from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import load_workbook

from dea.logging_utils import (
    action_step_to_log_record,
    write_entry_log_csv,
    write_entry_log_xlsx,
    write_validation_report_xlsx,
)
from dea.models import ActionStep, ValidationIssue


def _id_from_parts(*parts: str) -> str:
    return "".join(parts)


def _base_step(action: str, field: str, masked_value: str) -> ActionStep:
    return ActionStep(
        action=action,  # type: ignore[arg-type]
        screen="SCRN1",
        field=field,
        value="raw-not-logged",
        masked_value=masked_value,
        source_sheet="Clients",
        source_cell="D2",
        support_status="SUPPORTED",
    )


def test_planned_action_step_becomes_log_record_with_planned_status() -> None:
    rec = action_step_to_log_record(
        _base_step("ENTER_FIELD", "taxpayer.first_name", "Alex"),
        client_id="C-001",
        tax_year=2025,
        mode="dry_run",
        timestamp=datetime(2026, 1, 1, tzinfo=UTC),
    )
    assert rec.status == "PLANNED"


def test_skipped_manual_review_logs_status() -> None:
    rec = action_step_to_log_record(
        _base_step("SKIP_MANUAL_REVIEW", "w2.box_3_social_security_wages", "72000"),
        client_id="C-001",
        tax_year=2025,
        mode="dry_run",
    )
    assert rec.status == "SKIPPED_MANUAL_REVIEW"


def test_skipped_unsupported_logs_status() -> None:
    rec = action_step_to_log_record(
        _base_step("SKIP_UNSUPPORTED", "w2.box_4_social_security_tax", "4464"),
        client_id="C-001",
        tax_year=2025,
        mode="dry_run",
    )
    assert rec.status == "SKIPPED_UNSUPPORTED"


def test_csv_export_writes_expected_headers_and_rows(tmp_path) -> None:
    records = [
        action_step_to_log_record(
            _base_step("ENTER_FIELD", "taxpayer.first_name", "Alex"),
            client_id="C-001",
            tax_year=2025,
            mode="dry_run",
            timestamp=datetime(2026, 1, 1, tzinfo=UTC),
        )
    ]
    out = tmp_path / "entry_log.csv"
    write_entry_log_csv(records, out)

    content = out.read_text(encoding="utf-8")
    assert "timestamp,client_id,tax_year,mode,screen,field,source_sheet,source_cell,masked_value,action,status,error_message" in content
    assert "C-001" in content
    assert "taxpayer.first_name" in content


def test_xlsx_export_writes_expected_headers_and_rows(tmp_path) -> None:
    records = [
        action_step_to_log_record(
            _base_step("ENTER_FIELD", "taxpayer.first_name", "Alex"),
            client_id="C-001",
            tax_year=2025,
            mode="dry_run",
            timestamp=datetime(2026, 1, 1, tzinfo=UTC),
        )
    ]
    out = tmp_path / "entry_log.xlsx"
    write_entry_log_xlsx(records, out)

    wb = load_workbook(out)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]

    assert headers == [
        "timestamp",
        "client_id",
        "tax_year",
        "mode",
        "screen",
        "field",
        "source_sheet",
        "source_cell",
        "masked_value",
        "action",
        "status",
        "error_message",
    ]
    assert row[1] == "C-001"
    assert row[5] == "taxpayer.first_name"


def test_csv_and_xlsx_do_not_emit_full_ssn_or_ein(tmp_path) -> None:
    ssn = _id_from_parts("123", "45", "6789")
    ein = _id_from_parts("12", "345", "6789")
    records = [
        action_step_to_log_record(
            _base_step("ENTER_FIELD", "taxpayer.ssn", "***-**-6789"),
            client_id="C-001",
            tax_year=2025,
            mode="dry_run",
            timestamp=datetime(2026, 1, 1, tzinfo=UTC),
        ),
        action_step_to_log_record(
            _base_step("ENTER_FIELD", "w2.employer.ein", "**-***6789"),
            client_id="C-001",
            tax_year=2025,
            mode="dry_run",
            timestamp=datetime(2026, 1, 1, tzinfo=UTC),
        ),
    ]

    csv_path = tmp_path / "entry_log.csv"
    xlsx_path = tmp_path / "entry_log.xlsx"
    write_entry_log_csv(records, csv_path)
    write_entry_log_xlsx(records, xlsx_path)

    csv_content = csv_path.read_text(encoding="utf-8")
    assert ssn not in csv_content
    assert ein not in csv_content

    wb = load_workbook(xlsx_path)
    ws = wb.active
    joined = "\n".join("" if cell.value is None else str(cell.value) for row in ws.iter_rows() for cell in row)
    assert ssn not in joined
    assert ein not in joined


def test_validation_report_xlsx_headers_rows_and_masking(tmp_path) -> None:
    full_ssn = _id_from_parts("123", "45", "6789")
    issue = ValidationIssue(
        severity="ERROR",
        client_id="C-001",
        field="taxpayer.ssn",
        message="must contain exactly 9 digits",
        source_sheet="Clients",
        source_cell="F2",
    )
    out = tmp_path / "validation_report.xlsx"
    write_validation_report_xlsx([issue], out)

    wb = load_workbook(out)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]

    assert headers == ["severity", "client_id", "field", "message", "source_sheet", "source_cell"]
    assert row[0] == "ERROR"
    assert row[2] == "taxpayer.ssn"

    joined = "\n".join("" if cell.value is None else str(cell.value) for r in ws.iter_rows() for cell in r)
    assert full_ssn not in joined
