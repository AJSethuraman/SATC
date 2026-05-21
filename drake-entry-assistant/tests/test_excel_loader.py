from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import Workbook

from dea.excel_loader import ExcelLoadError, load_workbook_data


CLIENTS_HEADERS = [
    "ClientID", "TaxYear", "FilingStatus", "TP_First", "TP_Last", "TP_SSN",
    "TP_DOB", "TP_Occupation", "SP_First", "SP_Last", "SP_SSN", "SP_DOB",
    "SP_Occupation", "Address", "City", "State", "ZIP", "Phone", "Email",
]

W2_HEADERS = [
    "ClientID", "W2_ID", "Employee", "Employer_EIN", "Employer_Name", "Employer_Address",
    "Employer_City", "Employer_State", "Employer_ZIP", "Box1", "Box2", "Box3", "Box4",
    "Box5", "Box6", "Box12_Code_1", "Box12_Amount_1", "Box15_State",
    "Box15_Employer_State_ID", "Box16", "Box17",
]


def _id_from_parts(*parts: str) -> str:
    return "".join(parts)


def _create_base_workbook(tmp_path):
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)

    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)

    path = tmp_path / "input.xlsx"
    wb.save(path)
    return path


def test_loads_one_client_no_w2s(tmp_path) -> None:
    path = _create_base_workbook(tmp_path)
    wb = Workbook()
    # rebuild explicitly to keep sheets/headers deterministic
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "S", "Alex", "Rivera", _id_from_parts("123", "45", "6789"),
        "1988-04-14", "Engineer", "", "", "", "", "", "100 Main St",
        "Springfield", "IL", "62701", "", "",
    ])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    wb.save(path)

    loaded = load_workbook_data(path)
    assert len(loaded.client_batch.clients) == 1
    client = loaded.client_batch.clients[0]
    assert client.taxpayer.first_name == "Alex"
    assert client.address.zip == "62701"
    assert client.w2s == []


def test_loads_one_client_one_w2(tmp_path) -> None:
    path = _create_base_workbook(tmp_path)
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "MFJ", "Alex", "Rivera", _id_from_parts("123", "45", "6789"),
        "1988-04-14", "Engineer", "Jordan", "Rivera", _id_from_parts("111", "22", "3333"),
        "1990-11-30", "Teacher", "100 Main St", "Springfield", "IL", "62701", "", "",
    ])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    w2s.append([
        "C-001", "W2-001", "tp", _id_from_parts("12", "345", "6789"), "Acme", "500 Market Ave",
        "Springfield", "IL", "62702", "72000", "8500", "72000", "4464", "72000", "1044",
        "D", "3000", "IL", "IL-ID", "72000", "3200",
    ])
    wb.save(path)

    loaded = load_workbook_data(path)
    client = loaded.client_batch.clients[0]
    assert len(client.w2s) == 1
    w2 = client.w2s[0]
    assert w2.employer.ein.endswith("6789")
    assert w2.box_1_wages == Decimal("72000")
    assert len(w2.box_12_items) == 1
    assert len(w2.state_lines) == 1


def test_loads_multiple_w2s(tmp_path) -> None:
    path = _create_base_workbook(tmp_path)
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "MFJ", "Alex", "Rivera", _id_from_parts("123", "45", "6789"),
        "1988-04-14", "Engineer", "Jordan", "Rivera", _id_from_parts("111", "22", "3333"),
        "1990-11-30", "Teacher", "100 Main St", "Springfield", "IL", "62701", "", "",
    ])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    w2s.append(["C-001", "W2-001", "tp", _id_from_parts("12", "345", "6789"), "Acme", "500 Market Ave", "Springfield", "IL", "62702", "100", "10", "100", "6.2", "100", "1.45", "", "", "", "", "", ""])
    w2s.append(["C-001", "W2-002", "sp", _id_from_parts("98", "765", "4321"), "Beta", "900 State St", "Springfield", "IL", "62703", "200", "20", "200", "12.4", "200", "2.9", "", "", "", "", "", ""])
    wb.save(path)

    loaded = load_workbook_data(path)
    assert len(loaded.client_batch.clients[0].w2s) == 2


def test_source_cells_include_expected_refs(tmp_path) -> None:
    path = _create_base_workbook(tmp_path)
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "S", "Alex", "Rivera", _id_from_parts("123", "45", "6789"), "1988-04-14", "Engineer",
        "", "", "", "", "", "100 Main St", "Springfield", "IL", "62701", "", "",
    ])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    w2s.append([
        "C-001", "W2-001", "tp", _id_from_parts("12", "345", "6789"), "Acme", "500 Market Ave",
        "Springfield", "IL", "62702", "72000", "", "", "", "", "", "", "", "", "", "", "",
    ])
    wb.save(path)

    loaded = load_workbook_data(path)
    refs = loaded.source_cells
    assert refs["clients.C-001.taxpayer.first_name"].sheet == "Clients"
    assert refs["clients.C-001.taxpayer.first_name"].cell == "D2"
    assert refs["clients.C-001.taxpayer.ssn"].cell == "F2"
    assert refs["clients.C-001.address.zip"].cell == "Q2"
    assert refs["clients.C-001.w2s.W2-001.employer.ein"].sheet == "W2s"
    assert refs["clients.C-001.w2s.W2-001.box_1_wages"].cell == "J2"


def test_missing_clients_sheet_raises(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "W2s"
    ws.append(W2_HEADERS)
    path = tmp_path / "missing_clients.xlsx"
    wb.save(path)

    with pytest.raises(ExcelLoadError, match="Missing required sheet: Clients"):
        load_workbook_data(path)


def test_missing_w2s_sheet_raises(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(CLIENTS_HEADERS)
    path = tmp_path / "missing_w2s.xlsx"
    wb.save(path)

    with pytest.raises(ExcelLoadError, match="Missing required sheet: W2s"):
        load_workbook_data(path)


def test_missing_required_column_raises(tmp_path) -> None:
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append([col for col in CLIENTS_HEADERS if col != "TP_First"])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    path = tmp_path / "missing_col.xlsx"
    wb.save(path)

    with pytest.raises(ExcelLoadError, match="Missing required columns in Clients"):
        load_workbook_data(path)


def test_unknown_client_id_in_w2_raises(tmp_path) -> None:
    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "S", "Alex", "Rivera", _id_from_parts("123", "45", "6789"),
        "1988-04-14", "Engineer", "", "", "", "", "", "100 Main St", "Springfield", "IL", "62701", "", "",
    ])
    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    w2s.append([
        "C-999", "W2-001", "tp", _id_from_parts("12", "345", "6789"), "Acme", "500 Market Ave",
        "Springfield", "IL", "62702", "100", "10", "100", "6", "100", "1", "", "", "", "", "", "",
    ])
    path = tmp_path / "unknown_client.xlsx"
    wb.save(path)

    with pytest.raises(ExcelLoadError, match="unknown ClientID"):
        load_workbook_data(path)
