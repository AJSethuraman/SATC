from __future__ import annotations

from openpyxl import load_workbook

from dea.action_plan import generate_action_plan
from dea.adapters.fake import FakeDrakeAdapter
from dea.config_loader import load_screen_maps
from dea.demo import create_sample_workbook
from dea.excel_loader import load_workbook_data
from dea.logging_utils import write_entry_log_csv, write_entry_log_xlsx, write_validation_report_xlsx
from dea.output import write_action_plans_json
from dea.validation import validate_client_batch


def _id_from_parts(*parts: str) -> str:
    return "".join(parts)


def _xlsx_cells_as_text(path):
    wb = load_workbook(path)
    ws = wb.active
    return "\n".join("" if cell.value is None else str(cell.value) for row in ws.iter_rows() for cell in row)


def test_milestone_workflow_end_to_end(tmp_path) -> None:
    raw_ssn = _id_from_parts("123", "45", "6789")
    raw_ein = _id_from_parts("12", "345", "6789")

    workbook_path = tmp_path / "sample.xlsx"
    create_sample_workbook(workbook_path)

    loaded = load_workbook_data(workbook_path)
    issues = validate_client_batch(loaded.client_batch, loaded.source_cells)
    assert not [issue for issue in issues if issue.severity == "ERROR"]

    screen_maps = load_screen_maps("configs/drake/2025")
    plans = [
        generate_action_plan(client, screen_maps, source_cells=loaded.source_cells)
        for client in loaded.client_batch.clients
    ]

    validation_report = tmp_path / "validation_report.xlsx"
    action_plan_output = tmp_path / "action_plan.json"
    entry_csv = tmp_path / "entry_log.csv"
    entry_xlsx = tmp_path / "entry_log.xlsx"

    write_validation_report_xlsx(issues, validation_report)
    write_action_plans_json(plans, action_plan_output)

    adapter = FakeDrakeAdapter()
    all_records = []
    for plan in plans:
        result = adapter.execute_action_plan(plan, screen_maps)
        assert result.success is True
        all_records.extend(result.records)

    write_entry_log_csv(all_records, entry_csv)
    write_entry_log_xlsx(all_records, entry_xlsx)

    assert validation_report.exists()
    assert action_plan_output.exists()
    assert entry_csv.exists()
    assert entry_xlsx.exists()

    action_text = action_plan_output.read_text(encoding="utf-8")
    csv_text = entry_csv.read_text(encoding="utf-8")
    xlsx_text = _xlsx_cells_as_text(entry_xlsx)
    report_text = _xlsx_cells_as_text(validation_report)

    for text in [action_text, csv_text, xlsx_text, report_text]:
        assert raw_ssn not in text
        assert raw_ein not in text

    assert "***-**-6789" in action_text
    assert "**-***6789" in action_text
