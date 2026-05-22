from __future__ import annotations

from openpyxl import Workbook

from dea.action_plan import generate_action_plan
from dea.adapters.fake import FakeDrakeAdapter
from dea.config_loader import load_screen_maps
from dea.excel_loader import load_workbook_data
from dea.validation import validate_client_batch


CLIENTS_HEADERS = [
    "ClientID", "TaxYear", "FilingStatus", "TP_First", "TP_Last", "TP_SSN",
    "TP_DOB", "TP_Occupation", "SP_First", "SP_Last", "SP_SSN", "SP_DOB",
    "SP_Occupation", "Address", "City", "State", "ZIP", "Phone", "Email",
]

W2_HEADERS = [
    "ClientID", "W2_ID", "Employee", "Employer_EIN", "Employer_Name", "Employer_Address",
    "Employer_City", "Employer_State", "Employer_ZIP", "Box1", "Box2", "Box3", "Box4",
    "Box5", "Box6", "Box12_Code_1", "Box12_Amount_1", "Box15_State",
    "Box15_Employer_State_ID", "Box16", "Box17",
]


def _id_from_parts(*parts: str) -> str:
    return "".join(parts)


def test_integration_dry_run_then_fake_execution(tmp_path) -> None:
    ssn = _id_from_parts("123", "45", "6789")
    spouse_ssn = _id_from_parts("111", "22", "3333")
    ein = _id_from_parts("12", "345", "6789")

    wb = Workbook()
    clients = wb.active
    clients.title = "Clients"
    clients.append(CLIENTS_HEADERS)
    clients.append([
        "C-001", 2025, "MFJ", "Alex", "Rivera", ssn,
        "1988-04-14", "Engineer", "Jordan", "Rivera", spouse_ssn,
        "1990-11-30", "Teacher", "100 Main St", "Springfield", "IL", "62701", "", "",
    ])

    w2s = wb.create_sheet("W2s")
    w2s.append(W2_HEADERS)
    w2s.append([
        "C-001", "W2-001", "tp", ein, "Acme", "500 Market Ave",
        "Springfield", "IL", "62702", "72000", "8500", "72000", "4464", "72000", "1044",
        "D", "3000", "IL", "IL-ID", "72000", "3200",
    ])

    workbook_path = tmp_path / "sample.xlsx"
    wb.save(workbook_path)

    loaded = load_workbook_data(workbook_path)
    issues = validate_client_batch(loaded.client_batch, loaded.source_cells)

    assert not [i for i in issues if i.severity == "ERROR"]

    maps = load_screen_maps("configs/drake/2025")
    client = loaded.client_batch.clients[0]
    plan = generate_action_plan(client, maps, loaded.source_cells)

    adapter = FakeDrakeAdapter()
    result = adapter.execute_action_plan(plan, maps)

    assert result.success is True
    assert result.records
    assert any(r.screen == "SCRN1" for r in result.records)
    assert any(r.screen == "W2IN" for r in result.records)

    payload = "\n".join(f"{r.field}|{r.masked_value}|{r.error_message or ''}" for r in result.records)
    assert ssn not in payload
    assert ein not in payload
