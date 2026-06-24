"""Excel ingestion helpers for source workbooks.

This module loads normalized DEA models from a workbook that includes `Clients`
and `W2s` sheets while capturing source sheet/cell references.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from dea.models import (
    Address,
    Client,
    ClientBatch,
    Employer,
    LoadedWorkbook,
    SourceCellRef,
    Spouse,
    Taxpayer,
    W2,
    W2Box12Item,
    W2StateLine,
)


class ExcelLoadError(Exception):
    """Raised when the workbook structure or references are invalid."""


CLIENTS_SHEET = "Clients"
W2S_SHEET = "W2s"

REQUIRED_CLIENT_COLUMNS = [
    "ClientID",
    "TaxYear",
    "FilingStatus",
    "TP_First",
    "TP_Last",
    "TP_SSN",
    "TP_DOB",
    "TP_Occupation",
    "SP_First",
    "SP_Last",
    "SP_SSN",
    "SP_DOB",
    "SP_Occupation",
    "Address",
    "City",
    "State",
    "ZIP",
    "Phone",
    "Email",
]

REQUIRED_W2_COLUMNS = [
    "ClientID",
    "W2_ID",
    "Employee",
    "Employer_EIN",
    "Employer_Name",
    "Employer_Address",
    "Employer_City",
    "Employer_State",
    "Employer_ZIP",
    "Box1",
    "Box2",
    "Box3",
    "Box4",
    "Box5",
    "Box6",
    "Box12_Code_1",
    "Box12_Amount_1",
    "Box15_State",
    "Box15_Employer_State_ID",
    "Box16",
    "Box17",
]


def _header_index(sheet: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col_idx, value in enumerate(sheet.iter_rows(min_row=1, max_row=1, values_only=True).__next__(), start=1):
        if value is not None:
            header_map[str(value).strip()] = col_idx
    return header_map


def _require_columns(sheet_name: str, header_map: dict[str, int], required: list[str]) -> None:
    missing = [col for col in required if col not in header_map]
    if missing:
        raise ExcelLoadError(f"Missing required columns in {sheet_name}: {', '.join(missing)}")


def _cell_value(sheet: Worksheet, row: int, header_map: dict[str, int], column_name: str) -> object | None:
    return sheet.cell(row=row, column=header_map[column_name]).value


def _cell_ref(sheet: Worksheet, row: int, header_map: dict[str, int], column_name: str) -> SourceCellRef:
    cell = sheet.cell(row=row, column=header_map[column_name])
    return SourceCellRef(sheet=sheet.title, cell=cell.coordinate)


def _to_str(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value: object | None) -> Decimal:
    if value is None:
        return Decimal("0")
    text = str(value).strip()
    if text == "":
        return Decimal("0")
    try:
        return Decimal(text)
    except Exception:
        return Decimal("0")


def _to_int(value: object | None, *, context: str) -> int:
    """Coerce a cell to int, tolerating Excel's float form (2025.0).

    Raises ExcelLoadError with context rather than a bare ValueError so the
    loader's error contract holds for callers other than the CLI.
    """
    text = _to_str(value)
    if text == "":
        return 0
    try:
        return int(float(text)) if ("." in text or "e" in text.lower()) else int(text)
    except (TypeError, ValueError) as exc:
        raise ExcelLoadError(f"Invalid integer for {context}: {value!r}") from exc


def load_workbook_data(path: str | Path) -> LoadedWorkbook:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise ExcelLoadError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    if CLIENTS_SHEET not in wb.sheetnames:
        raise ExcelLoadError("Missing required sheet: Clients")
    if W2S_SHEET not in wb.sheetnames:
        raise ExcelLoadError("Missing required sheet: W2s")

    clients_sheet = wb[CLIENTS_SHEET]
    w2s_sheet = wb[W2S_SHEET]

    clients_header = _header_index(clients_sheet)
    w2_header = _header_index(w2s_sheet)
    _require_columns(CLIENTS_SHEET, clients_header, REQUIRED_CLIENT_COLUMNS)
    _require_columns(W2S_SHEET, w2_header, REQUIRED_W2_COLUMNS)

    source_cells: dict[str, SourceCellRef] = {}
    clients_by_id: dict[str, Client] = {}

    for row in range(2, clients_sheet.max_row + 1):
        client_id = _to_str(_cell_value(clients_sheet, row, clients_header, "ClientID"))
        if client_id == "":
            continue

        taxpayer = Taxpayer(
            first_name=_to_str(_cell_value(clients_sheet, row, clients_header, "TP_First")),
            last_name=_to_str(_cell_value(clients_sheet, row, clients_header, "TP_Last")),
            ssn=_to_str(_cell_value(clients_sheet, row, clients_header, "TP_SSN")),
            dob=_to_str(_cell_value(clients_sheet, row, clients_header, "TP_DOB")),
            occupation=_to_str(_cell_value(clients_sheet, row, clients_header, "TP_Occupation")) or None,
        )

        spouse_first = _to_str(_cell_value(clients_sheet, row, clients_header, "SP_First"))
        spouse_last = _to_str(_cell_value(clients_sheet, row, clients_header, "SP_Last"))
        spouse_ssn = _to_str(_cell_value(clients_sheet, row, clients_header, "SP_SSN"))
        spouse_dob = _to_str(_cell_value(clients_sheet, row, clients_header, "SP_DOB"))
        spouse_occ = _to_str(_cell_value(clients_sheet, row, clients_header, "SP_Occupation"))
        spouse: Spouse | None = None
        if any([spouse_first, spouse_last, spouse_ssn, spouse_dob, spouse_occ]):
            spouse = Spouse(
                first_name=spouse_first,
                last_name=spouse_last,
                ssn=spouse_ssn,
                dob=spouse_dob,
                occupation=spouse_occ or None,
            )

        address = Address(
            street=_to_str(_cell_value(clients_sheet, row, clients_header, "Address")),
            city=_to_str(_cell_value(clients_sheet, row, clients_header, "City")),
            state=_to_str(_cell_value(clients_sheet, row, clients_header, "State")),
            zip=_to_str(_cell_value(clients_sheet, row, clients_header, "ZIP")),
        )

        client = Client(
            client_id=client_id,
            tax_year=_to_int(
                _cell_value(clients_sheet, row, clients_header, "TaxYear"),
                context=f"TaxYear (client {client_id})",
            ),
            filing_status=_to_str(_cell_value(clients_sheet, row, clients_header, "FilingStatus")),
            taxpayer=taxpayer,
            spouse=spouse,
            address=address,
            w2s=[],
        )
        clients_by_id[client_id] = client

        base_key = f"clients.{client_id}"
        source_cells[f"{base_key}.taxpayer.first_name"] = _cell_ref(clients_sheet, row, clients_header, "TP_First")
        source_cells[f"{base_key}.taxpayer.last_name"] = _cell_ref(clients_sheet, row, clients_header, "TP_Last")
        source_cells[f"{base_key}.taxpayer.ssn"] = _cell_ref(clients_sheet, row, clients_header, "TP_SSN")
        source_cells[f"{base_key}.taxpayer.dob"] = _cell_ref(clients_sheet, row, clients_header, "TP_DOB")
        source_cells[f"{base_key}.filing_status"] = _cell_ref(clients_sheet, row, clients_header, "FilingStatus")
        source_cells[f"{base_key}.address.street"] = _cell_ref(clients_sheet, row, clients_header, "Address")
        source_cells[f"{base_key}.address.city"] = _cell_ref(clients_sheet, row, clients_header, "City")
        source_cells[f"{base_key}.address.state"] = _cell_ref(clients_sheet, row, clients_header, "State")
        source_cells[f"{base_key}.address.zip"] = _cell_ref(clients_sheet, row, clients_header, "ZIP")
        source_cells[f"{base_key}.spouse.first_name"] = _cell_ref(clients_sheet, row, clients_header, "SP_First")
        source_cells[f"{base_key}.spouse.last_name"] = _cell_ref(clients_sheet, row, clients_header, "SP_Last")
        source_cells[f"{base_key}.spouse.ssn"] = _cell_ref(clients_sheet, row, clients_header, "SP_SSN")
        source_cells[f"{base_key}.spouse.dob"] = _cell_ref(clients_sheet, row, clients_header, "SP_DOB")

    for row in range(2, w2s_sheet.max_row + 1):
        client_id = _to_str(_cell_value(w2s_sheet, row, w2_header, "ClientID"))
        if client_id == "":
            continue
        if client_id not in clients_by_id:
            raise ExcelLoadError(f"W2 row references unknown ClientID: {client_id}")

        client = clients_by_id[client_id]
        employee_token = _to_str(_cell_value(w2s_sheet, row, w2_header, "Employee")).lower()
        if employee_token in {"spouse", "sp"}:
            if client.spouse is None:
                # A W-2 assigned to a non-existent spouse is a data-integrity
                # error (wrong ClientID or a missing spouse row). Don't silently
                # reassign the income to the taxpayer.
                raise ExcelLoadError(
                    f"W2 row assigns income to spouse but client {client_id} has no spouse."
                )
            employee = client.spouse
        else:
            employee = client.taxpayer

        w2_id = _to_str(_cell_value(w2s_sheet, row, w2_header, "W2_ID"))
        employer = Employer(
            ein=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_EIN")),
            name=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_Name")),
            street=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_Address")),
            city=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_City")),
            state=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_State")),
            zip=_to_str(_cell_value(w2s_sheet, row, w2_header, "Employer_ZIP")),
        )

        box_12_items: list[W2Box12Item] = []
        code_1 = _to_str(_cell_value(w2s_sheet, row, w2_header, "Box12_Code_1"))
        amount_1_value = _cell_value(w2s_sheet, row, w2_header, "Box12_Amount_1")
        if code_1 or _to_str(amount_1_value):
            box_12_items.append(W2Box12Item(code=code_1, amount=_to_decimal(amount_1_value)))

        state_lines: list[W2StateLine] = []
        box15_state = _to_str(_cell_value(w2s_sheet, row, w2_header, "Box15_State"))
        box15_emp_state_id = _to_str(_cell_value(w2s_sheet, row, w2_header, "Box15_Employer_State_ID"))
        box16 = _cell_value(w2s_sheet, row, w2_header, "Box16")
        box17 = _cell_value(w2s_sheet, row, w2_header, "Box17")
        if box15_state or box15_emp_state_id or _to_str(box16) or _to_str(box17):
            state_lines.append(
                W2StateLine(
                    state=box15_state,
                    employer_state_id=box15_emp_state_id,
                    state_wages=_to_decimal(box16),
                    state_withholding=_to_decimal(box17),
                )
            )

        box1_val = _cell_value(w2s_sheet, row, w2_header, "Box1")
        box2_val = _cell_value(w2s_sheet, row, w2_header, "Box2")
        box3_val = _cell_value(w2s_sheet, row, w2_header, "Box3")
        box4_val = _cell_value(w2s_sheet, row, w2_header, "Box4")
        box5_val = _cell_value(w2s_sheet, row, w2_header, "Box5")
        box6_val = _cell_value(w2s_sheet, row, w2_header, "Box6")

        w2 = W2(
            w2_id=w2_id,
            client_id=client_id,
            employee=employee,
            employer=employer,
            box_1_wages=_to_decimal(box1_val),
            box_2_federal_withholding=_to_decimal(box2_val),
            box_3_social_security_wages=_to_decimal(box3_val),
            box_4_social_security_tax=_to_decimal(box4_val),
            box_5_medicare_wages=_to_decimal(box5_val),
            box_6_medicare_tax=_to_decimal(box6_val),
            box_1_raw=_to_str(box1_val),
            box_2_raw=_to_str(box2_val),
            box_3_raw=_to_str(box3_val),
            box_4_raw=_to_str(box4_val),
            box_5_raw=_to_str(box5_val),
            box_6_raw=_to_str(box6_val),
            box_12_items=box_12_items,
            state_lines=state_lines,
            manual_review_notes=[],
        )
        client.w2s.append(w2)

        w2_base = f"clients.{client_id}.w2s.{w2_id}"
        source_cells[f"{w2_base}.employer.ein"] = _cell_ref(w2s_sheet, row, w2_header, "Employer_EIN")
        source_cells[f"{w2_base}.employer.name"] = _cell_ref(w2s_sheet, row, w2_header, "Employer_Name")
        source_cells[f"{w2_base}.box_1_wages"] = _cell_ref(w2s_sheet, row, w2_header, "Box1")
        source_cells[f"{w2_base}.box_2_federal_withholding"] = _cell_ref(w2s_sheet, row, w2_header, "Box2")
        source_cells[f"{w2_base}.box_3_social_security_wages"] = _cell_ref(w2s_sheet, row, w2_header, "Box3")
        source_cells[f"{w2_base}.box_4_social_security_tax"] = _cell_ref(w2s_sheet, row, w2_header, "Box4")
        source_cells[f"{w2_base}.box_5_medicare_wages"] = _cell_ref(w2s_sheet, row, w2_header, "Box5")
        source_cells[f"{w2_base}.box_6_medicare_tax"] = _cell_ref(w2s_sheet, row, w2_header, "Box6")
        if box_12_items:
            source_cells[f"{w2_base}.box_12_items.0.code"] = _cell_ref(w2s_sheet, row, w2_header, "Box12_Code_1")
            source_cells[f"{w2_base}.box_12_items.0.amount"] = _cell_ref(w2s_sheet, row, w2_header, "Box12_Amount_1")
        if state_lines:
            source_cells[f"{w2_base}.state_lines.0.state"] = _cell_ref(w2s_sheet, row, w2_header, "Box15_State")
            source_cells[f"{w2_base}.state_lines.0.employer_state_id"] = _cell_ref(w2s_sheet, row, w2_header, "Box15_Employer_State_ID")
            source_cells[f"{w2_base}.state_lines.0.state_wages"] = _cell_ref(w2s_sheet, row, w2_header, "Box16")
            source_cells[f"{w2_base}.state_lines.0.state_withholding"] = _cell_ref(w2s_sheet, row, w2_header, "Box17")

    return LoadedWorkbook(client_batch=ClientBatch(clients=list(clients_by_id.values())), source_cells=source_cells)
