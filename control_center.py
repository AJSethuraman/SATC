#!/usr/bin/env python3
"""Credit-Risk Template Control Center -- one place to see and run everything.

Every template workbook carries its own data-path runner inside itself (the
_code_py tab -- the same bytes the in-Excel Extract button writes out). This
tool discovers those workbooks, shows their live state at a glance (last run,
alerts, staleness), and drives any of them by extracting the workbook's OWN
embedded runner. Nothing template-specific lives here: any workbook that
follows TEMPLATE_CONTRACT.md is supported automatically.

GUI (default):   python control_center.py
                 (Tkinter -- ships with standard Python on Windows.)
CLI (headless):  python control_center.py --list          # status board
                 python control_center.py --doctor        # will it work here?
                 python control_center.py --run <name> --demo
                 python control_center.py --run-all --demo
                 python control_center.py --tieout <name> <entity> [--demo]
                 python control_center.py --extract <name>

Needs only what the runners need: python 3.8+, openpyxl, pandas.
Close a workbook in Excel before refreshing it -- the runner writes the file.
"""
from __future__ import annotations

import argparse
import datetime
import os
import subprocess
import sys
import threading

APP_TITLE = "Credit-Risk Template Control Center"
SKIP_DIRS = {".git", "build", "__pycache__", ".pytest_cache", "node_modules"}
MAX_DEPTH = 2                    # repo root -> template dir -> workbook
STATUS_COL = 12                  # column L status panel (contract convention)

# Friendly one-liners keyed by workbook basename; fallback for unknowns.
PURPOSE = {
    "FRED_Credit_Risk_Dashboard": "National credit quality + geographic HPI watchlist (FRED)",
    "Consumer_Credit_Risk_Monitor": "Consumer credit cycle by product; watchlist gated (NY Fed HHDC)",
    "Macro_Early_Warning_Monitor": "Recession/credit-cycle signals + 50-state labor stress (FRED)",
    "Bank_Peer_Monitor": "Peer/competitor bank surveillance incl. loan-book + SVB metrics (FDIC)",
    "Mortgage_Delinquency_Monitor": "County mortgage delinquency in your footprint (CFPB)",
    "Crit_Class_Tracker": "Commercial criticized/classified per competitor + 8-K events (EDGAR)",
}

# The suite's ENTIRE credential surface (user choice: keys live in the
# workbook _config cells; every runner already reads them). --configure
# writes a value into every workbook that has the setting.
CONFIG_KEYS = {
    "fred_api_key": "FRED API key (free: fredaccount.stlouisfed.org/apikeys) "
                    "-- used by the FRED + Macro templates",
    "edgar_user_agent": "SEC fair-access identifier 'YourOrg you@bank.com' "
                        "-- required by the EDGAR tracker (not a secret)",
}

# Data hosts per provider, for --doctor (which template needs which host).
DOCTOR_HOSTS = [
    ("api.stlouisfed.org", "FRED + Macro templates (live FRED pulls)"),
    ("api.fdic.gov", "Bank Peer Monitor (live FDIC pulls; keyless)"),
    ("www.newyorkfed.org", "Consumer Credit Monitor (live HHDC download)"),
    ("files.consumerfinance.gov", "Mortgage Delinquency Monitor (live CFPB CSVs)"),
    ("www.sec.gov", "Crit/Class Tracker (EDGAR archives)"),
    ("data.sec.gov", "Crit/Class Tracker (EDGAR submissions/facts APIs)"),
]


class Template:
    def __init__(self, xlsm_path):
        self.xlsm_path = os.path.abspath(xlsm_path)
        self.folder = os.path.dirname(self.xlsm_path)
        self.name = os.path.splitext(os.path.basename(self.xlsm_path))[0]
        self.purpose = PURPOSE.get(self.name, "credit-risk template")

    def __repr__(self):
        return f"Template({self.name})"


# --------------------------------------------------------------------------
# Discovery + live status (read from the workbook's column-L status panel)
# --------------------------------------------------------------------------
def _has_code_tab(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ok = "_code_py" in wb.sheetnames
        wb.close()
        return ok
    except Exception:
        return False


def discover(root=".") -> list:
    root = os.path.abspath(root)
    found = []
    for dirpath, dirnames, filenames in os.walk(root):
        rel = os.path.relpath(dirpath, root)
        depth = 0 if rel == "." else rel.count(os.sep) + 1
        dirnames[:] = [d for d in dirnames
                       if d not in SKIP_DIRS and depth < MAX_DEPTH]
        for fn in filenames:
            # .xlsm is the finished product; a fallback .xlsx (bundle-built,
            # macro not yet pasted in) is equally runnable.
            if fn.lower().endswith((".xlsm", ".xlsx")) and not fn.startswith("~$"):
                p = os.path.join(dirpath, fn)
                if _has_code_tab(p):
                    found.append(Template(p))
    return sorted(found, key=lambda t: t.name.lower())


def read_status(t: Template) -> dict:
    """Last-run state straight from the workbook's status panel (column L of
    the first Dashboard_* tab) -- no run needed, just a read."""
    out = {"last_run": "never refreshed", "detail": "", "mtime": ""}
    try:
        out["mtime"] = datetime.datetime.fromtimestamp(
            os.path.getmtime(t.xlsm_path)).strftime("%Y-%m-%d %H:%M")
        import openpyxl
        wb = openpyxl.load_workbook(t.xlsm_path, read_only=True)
        dash = next((s for s in wb.sheetnames if s.startswith("Dashboard_")), None)
        if dash:
            ws = wb[dash]
            l1 = ws.cell(1, STATUS_COL).value
            l2 = ws.cell(2, STATUS_COL).value
            if l1:
                out["last_run"] = str(l1)
            if l2:
                out["detail"] = str(l2)
        wb.close()
    except Exception as exc:
        out["detail"] = f"(status unreadable: {exc})"
    return out


def read_setting(t: Template, key: str):
    """Read one [SETTINGS] value from the workbook's _config tab (or None if
    the workbook doesn't carry that setting)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(t.xlsm_path, read_only=True)
        if "_config" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["_config"]
        for row in ws.iter_rows(min_col=1, max_col=2):
            if str(row[0].value).strip() == key:
                val = "" if row[1].value is None else str(row[1].value).strip()
                wb.close()
                return val
        wb.close()
    except Exception:
        pass
    return None


def set_setting(t: Template, key: str, value: str) -> bool:
    """Write one [SETTINGS] value into the workbook's _config tab. Returns
    False if the workbook doesn't carry that setting. Close the workbook in
    Excel first -- this saves the file (keep_vba honored per L2)."""
    import openpyxl
    keep = t.xlsm_path.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(t.xlsm_path, keep_vba=keep)
    try:
        if "_config" not in wb.sheetnames:
            return False
        ws = wb["_config"]
        for row in ws.iter_rows(min_col=1, max_col=1):
            if str(row[0].value).strip() == key:
                ws.cell(row[0].row, 2).value = value
                wb.save(t.xlsm_path)
                return True
        return False
    finally:
        wb.close()


def configure_all(templates, key, value, log=print) -> int:
    """Write a setting into EVERY workbook that has it (configure once,
    applied suite-wide)."""
    if key not in CONFIG_KEYS:
        log(f"unknown setting '{key}' -- known: {', '.join(CONFIG_KEYS)}")
        return 2
    hits = 0
    for t in templates:
        try:
            if set_setting(t, key, value):
                log(f"[{t.name}] {key} set")
                hits += 1
        except Exception as exc:
            log(f"[{t.name}] FAILED ({exc}) -- is the workbook open in Excel?")
    if hits == 0:
        log(f"no discovered workbook carries '{key}'")
    return 0 if hits else 1


# --------------------------------------------------------------------------
# Extraction + run (the workbook's OWN embedded runner)
# --------------------------------------------------------------------------
def extract_runner(t: Template) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(t.xlsm_path, read_only=True)
    ws = wb["_code_py"]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    dest = os.path.join(t.folder, "runner.py")
    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")
    return dest


def _runner_source(runner_path):
    with open(runner_path, "r", encoding="utf-8") as fh:
        return fh.read()


def build_command(t: Template, runner_path: str, demo: bool) -> list:
    src = _runner_source(runner_path)
    cmd = [sys.executable, runner_path, "--workbook", t.xlsm_path]
    if demo:
        cmd.append("--demo")
    if "--backend" in src:               # grandfathered FRED runner
        cmd += ["--backend", "openpyxl"]
    return cmd


def run_refresh(t: Template, demo: bool, log=print) -> int:
    log(f"[{t.name}] extracting embedded runner ...")
    runner_path = extract_runner(t)
    cmd = build_command(t, runner_path, demo)
    log(f"[{t.name}] mode: {'DEMO (offline, deterministic)' if demo else 'LIVE'}")
    proc = subprocess.Popen(cmd, cwd=t.folder, stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True)
    for line in proc.stdout:
        log(line.rstrip())
    proc.wait()
    log(f"[{t.name}] {'DONE -- reopen the workbook to see the dashboards' if proc.returncode == 0 else f'FAILED (exit {proc.returncode})'}")
    return proc.returncode


def run_tieout(t: Template, entity: str, demo: bool, log=print) -> int:
    """Verification mode: print each value beside its official-document
    location (only templates whose runner supports --tieout)."""
    runner_path = extract_runner(t)
    if "--tieout" not in _runner_source(runner_path):
        log(f"[{t.name}] this template has no tie-out mode (its _provenance "
            "docs and source URLs are in the workbook instead).")
        return 2
    cmd = [sys.executable, runner_path, "--workbook", t.xlsm_path,
           "--tieout", entity]
    if demo:
        cmd.append("--demo")
    proc = subprocess.Popen(cmd, cwd=t.folder, stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True)
    for line in proc.stdout:
        log(line.rstrip())
    proc.wait()
    return proc.returncode


# --------------------------------------------------------------------------
# Doctor: "will this suite work on THIS machine?" in one command
# --------------------------------------------------------------------------
def doctor(root=".", log=print) -> int:
    log("=== Control Center doctor ===")
    ok = True
    # 1. python + deps
    log(f"python : {sys.version.split()[0]}")
    for mod in ("pandas", "openpyxl"):
        try:
            m = __import__(mod)
            log(f"{mod:<9}: OK ({getattr(m, '__version__', '?')})")
        except ImportError:
            log(f"{mod:<9}: MISSING -- run: python -m pip install pandas openpyxl")
            ok = False
    try:
        import tkinter  # noqa: F401
        log("tkinter  : OK (GUI available)")
    except ImportError:
        log("tkinter  : missing -- CLI mode still works (--list/--run)")
    # 2. workbooks
    ts = discover(root)
    log(f"workbooks: {len(ts)} discovered under {os.path.abspath(root)}")
    for t in ts:
        log(f"  - {t.name}")
    # 3. credentials (the whole surface: one free key + one identifier)
    log("credentials (live mode only; demo needs none):")
    for key, why in CONFIG_KEYS.items():
        holders = [(t, read_setting(t, key)) for t in ts]
        holders = [(t, v) for t, v in holders if v is not None]
        if not holders:
            continue
        env_note = ""
        if key == "fred_api_key" and os.environ.get("FRED_API_KEY"):
            env_note = " (FRED_API_KEY env var IS set -- cells optional)"
        for t, v in holders:
            state = "SET" if v else "blank"
            log(f"  {key:<18} {state:<6} in {t.name}{env_note}")
        if any(not v for _, v in holders) and not env_note:
            log(f"      -> fix once for all: python control_center.py "
                f"--configure {key} \"<value>\"   ({why})")
    log("  (FDIC, CFPB and NY Fed HHDC templates need NO credentials)")
    # 4. data-host reachability (LIVE pulls only; demo mode needs no network)
    log("data hosts (live mode; demo mode works regardless):")
    import urllib.request
    for host, needed_by in DOCTOR_HOSTS:
        try:
            req = urllib.request.Request(
                f"https://{host}/", method="HEAD",
                headers={"User-Agent": "control-center-doctor contact@example.com"})
            urllib.request.urlopen(req, timeout=6)
            log(f"  {host:<28} REACHABLE   ({needed_by})")
        except Exception as exc:
            code = getattr(exc, "code", None)
            if code is not None:         # an HTTP status means we REACHED it
                log(f"  {host:<28} REACHABLE   (HTTP {code}) ({needed_by})")
            else:
                log(f"  {host:<28} BLOCKED     ({needed_by})")
                log(f"      -> if live pulls fail, ask IT to allowlist {host}")
    log("verdict: " + ("environment OK -- demo mode always works; live mode "
                       "per host results above" if ok else
                       "fix the MISSING items above first"))
    return 0 if ok else 1


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def _resolve(arg, templates):
    if os.path.exists(arg):
        return Template(arg)
    matches = [t for t in templates if arg.lower() in t.name.lower()]
    if len(matches) == 1:
        return matches[0]
    sys.stderr.write(f"'{arg}' matches {len(matches)} templates; be more specific.\n")
    raise SystemExit(2)


def cli(argv):
    ap = argparse.ArgumentParser(description=APP_TITLE)
    ap.add_argument("--root", default=".", help="folder to scan for template workbooks")
    ap.add_argument("--list", action="store_true", help="status board of all templates")
    ap.add_argument("--doctor", action="store_true", help="check deps + data-host reachability")
    ap.add_argument("--run", metavar="NAME", help="refresh one template")
    ap.add_argument("--run-all", action="store_true", help="refresh every template, one at a time")
    ap.add_argument("--tieout", nargs=2, metavar=("NAME", "ENTITY"),
                    help="verification: values beside their official document locations")
    ap.add_argument("--extract", metavar="NAME", help="extract a template's runner.py and exit")
    ap.add_argument("--configure", nargs=2, metavar=("KEY", "VALUE"),
                    help="write a setting into EVERY workbook that has it "
                         "(keys: " + ", ".join(CONFIG_KEYS) + ")")
    ap.add_argument("--demo", action="store_true", help="offline deterministic demo data")
    args = ap.parse_args(argv)

    if args.doctor:
        return doctor(args.root)
    templates = discover(args.root)
    if args.list:
        if not templates:
            print("no template workbooks found (.xlsm/.xlsx with a _code_py tab)")
        for t in templates:
            st = read_status(t)
            print(f"{t.name}")
            print(f"    {t.purpose}")
            print(f"    {st['last_run']}" + (f" | {st['detail']}" if st['detail'] else "")
                  + f" | file: {st['mtime']}")
        return 0
    if args.extract:
        print(extract_runner(_resolve(args.extract, templates)))
        return 0
    if args.configure:
        return configure_all(templates, args.configure[0], args.configure[1])
    if args.tieout:
        t = _resolve(args.tieout[0], templates)
        return run_tieout(t, args.tieout[1], demo=args.demo)
    if args.run:
        return run_refresh(_resolve(args.run, templates), demo=args.demo)
    if args.run_all:
        worst = 0
        for t in templates:
            worst = max(worst, run_refresh(t, demo=args.demo))
        return worst
    return gui(templates, args.root)


# --------------------------------------------------------------------------
# GUI (Tkinter -- one template at a time; status board built in)
# --------------------------------------------------------------------------
def gui(templates, root):
    try:
        import tkinter as tk
        from tkinter import scrolledtext, messagebox, simpledialog
    except ImportError:
        sys.stderr.write("Tkinter unavailable -- use --list/--run/--doctor.\n")
        return 3

    # House palette (mirrors keybank_style tokens; this tool lives OUTSIDE
    # the workbooks and must stay stdlib-only, so tokens are inlined).
    INK, CANVAS, KEY_RED, SLATE = "#0A0908", "#F4F1EC", "#CC0000", "#57534B"

    win = tk.Tk()
    win.title(APP_TITLE)
    win.geometry("980x620")
    win.configure(bg=CANVAS)

    tk.Label(win, text=APP_TITLE, bg=INK, fg="white",
             font=("Arial", 14, "bold"), anchor="w", padx=14, pady=10
             ).pack(fill="x")
    tk.Frame(win, bg=KEY_RED, height=3).pack(fill="x")

    body = tk.Frame(win, bg=CANVAS)
    body.pack(fill="both", expand=True, padx=12, pady=10)

    left = tk.Frame(body, bg=CANVAS)
    left.pack(side="left", fill="y")
    tk.Label(left, text="Templates (pick one)", bg=CANVAS, fg=SLATE,
             font=("Arial", 9, "bold")).pack(anchor="w")
    lb = tk.Listbox(left, width=44, height=12, font=("Consolas", 10),
                    selectmode="browse", exportselection=False)
    lb.pack(fill="y", expand=True, pady=(2, 6))
    for t in templates:
        lb.insert("end", t.name)
    if templates:
        lb.selection_set(0)

    info = tk.Label(left, text="", bg=CANVAS, fg=SLATE, font=("Arial", 9),
                    justify="left", anchor="w", wraplength=330)
    info.pack(fill="x", pady=(0, 6))

    log_box = scrolledtext.ScrolledText(body, font=("Consolas", 9), bg="white",
                                        fg=INK, state="disabled", wrap="word")
    log_box.pack(side="right", fill="both", expand=True, padx=(12, 0))

    def log(msg):
        log_box.configure(state="normal")
        log_box.insert("end", msg + "\n")
        log_box.see("end")
        log_box.configure(state="disabled")

    def show_info(_evt=None):
        sel = lb.curselection()
        if not sel:
            return
        t = templates[sel[0]]
        st = read_status(t)
        info.configure(text=f"{t.purpose}\n{st['last_run']}"
                            + (f"\n{st['detail']}" if st['detail'] else "")
                            + f"\nfile saved: {st['mtime']}")
    lb.bind("<<ListboxSelect>>", show_info)

    buttons = []

    def set_busy(busy):
        for b in buttons:
            b.configure(state="disabled" if busy else "normal")

    def selected():
        sel = lb.curselection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Pick a template first.")
            return None
        return templates[sel[0]]

    def in_thread(work):
        set_busy(True)          # one job at a time

        def wrapped():
            try:
                work()
            except Exception as exc:
                win.after(0, log, f"ERROR: {exc}")
            finally:
                win.after(0, set_busy, False)
                win.after(0, show_info)
        threading.Thread(target=wrapped, daemon=True).start()

    def do_run(demo, everything=False):
        ts = templates if everything else ([selected()] if selected() else [])
        if not ts or ts == [None]:
            return
        in_thread(lambda: [run_refresh(t, demo=demo,
                                       log=lambda m: win.after(0, log, m))
                           for t in ts])

    def do_tieout():
        t = selected()
        if not t:
            return
        entity = simpledialog.askstring(
            APP_TITLE, f"Tie-out {t.name}:\nentity id (CERT / CIK / FIPS "
                       "-- see the workbook's _provenance tab):")
        if not entity:
            return
        demo = messagebox.askyesno(APP_TITLE, "Demo mode? (Yes = demo values "
                                   "with the real tie-out map; No = live)")
        in_thread(lambda: run_tieout(t, entity, demo,
                                     log=lambda m: win.after(0, log, m)))

    def do_doctor():
        in_thread(lambda: doctor(root, log=lambda m: win.after(0, log, m)))

    def do_configure():
        # The suite's whole credential surface is two values; ask for both,
        # write each into every workbook that carries it. Blank = skip.
        current_fred = next((read_setting(t, "fred_api_key") for t in templates
                             if read_setting(t, "fred_api_key") is not None), "")
        current_ua = next((read_setting(t, "edgar_user_agent") for t in templates
                           if read_setting(t, "edgar_user_agent") is not None), "")
        fred = simpledialog.askstring(
            APP_TITLE, "FRED API key (free at fredaccount.stlouisfed.org/"
            "apikeys)\nApplied to every workbook that uses FRED.\n"
            "Leave blank to keep current.", initialvalue=current_fred or "")
        ua = simpledialog.askstring(
            APP_TITLE, "EDGAR user agent -- 'YourOrg you@bank.com' (SEC "
            "fair-access rule, not a secret).\nLeave blank to keep current.",
            initialvalue=current_ua or "")

        def work():
            wrote = False
            if fred:
                configure_all(templates, "fred_api_key", fred,
                              log=lambda m: win.after(0, log, m))
                wrote = True
            if ua:
                configure_all(templates, "edgar_user_agent", ua,
                              log=lambda m: win.after(0, log, m))
                wrote = True
            win.after(0, log, "configure: done -- live refreshes are ready"
                      if wrote else "configure: nothing changed")
        in_thread(work)

    def do_open():
        t = selected()
        if not t:
            return
        if sys.platform == "win32":
            os.startfile(t.folder)                       # noqa: S606
        else:
            subprocess.Popen(["xdg-open" if sys.platform.startswith("linux")
                              else "open", t.folder])

    btns = tk.Frame(left, bg=CANVAS)
    btns.pack(fill="x")
    for label, cmd, accent in [
            ("Refresh (Demo)", lambda: do_run(True), INK),
            ("Refresh (Live)", lambda: do_run(False), KEY_RED),
            ("Refresh ALL (Demo)", lambda: do_run(True, everything=True), INK),
            ("Tie-out / verify...", do_tieout, SLATE),
            ("Configure keys...", do_configure, SLATE),
            ("Doctor (env check)", do_doctor, SLATE),
            ("Open folder", do_open, SLATE)]:
        b = tk.Button(btns, text=label, command=cmd, bg=accent, fg="white",
                      font=("Arial", 10, "bold"), relief="flat", padx=10, pady=5)
        b.pack(fill="x", pady=2)
        buttons.append(b)

    log(f"Scanned: {os.path.abspath(root)} -- {len(templates)} template(s).")
    log("Close a workbook in Excel before refreshing it. New here? Click "
        "'Doctor' first: it checks python deps and which data hosts this "
        "machine can reach (demo mode always works offline).")
    show_info()
    win.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(cli(sys.argv[1:]))
