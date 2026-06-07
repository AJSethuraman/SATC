#!/usr/bin/env python3
"""Tests for the year-by-year fee workbook. openpyxl-guarded."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import fee_workbook
import invoice_calc

try:
    from openpyxl import Workbook, load_workbook

    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover - depends on environment
    HAVE_OPENPYXL = False


class YearDetectTests(unittest.TestCase):
    def test_explicit_and_from_clients(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            folder = Path(d)
            self.assertEqual(fee_workbook.target_year(folder, 2031), 2031)
            (folder / "clients.json").write_text(
                json.dumps([{"client_name": "A", "tax_year": "2024"}, {"client_name": "B", "tax_year": "2025"}]),
                encoding="utf-8",
            )
            self.assertEqual(fee_workbook.target_year(folder, None), 2025)


@unittest.skipUnless(HAVE_OPENPYXL, "openpyxl not installed")
class WorkbookTests(unittest.TestCase):
    def test_round_trip_schedule(self) -> None:
        schedule = {
            "base_1040": {"description": "Form 1040", "price": 170.0, "additional": 170.0},
            "schedule_b": {"description": "Schedule B", "price": 5.0},
            invoice_calc.DISCOUNTS_KEY: {
                "express": {"description": "Express", "amount": -40.0},
                "friends_family": {"description": "F&F", "percent": 20.0},
            },
        }
        wb = Workbook()
        ws = wb.active
        fee_workbook.write_year_sheet(ws, schedule, 2025)
        parsed = fee_workbook.read_year_sheet(ws)
        self.assertEqual(parsed["base_1040"], {"description": "Form 1040", "price": 170.0, "additional": 170.0})
        self.assertEqual(parsed["schedule_b"], {"description": "Schedule B", "price": 5.0})
        self.assertEqual(parsed[invoice_calc.DISCOUNTS_KEY]["express"], {"description": "Express", "amount": -40.0})
        self.assertEqual(parsed[invoice_calc.DISCOUNTS_KEY]["friends_family"], {"description": "F&F", "percent": 20.0})

    def test_run_creates_workbook_next_year_and_applies_edits(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            folder = Path(d)
            (folder / "clients.json").write_text(
                json.dumps([{"client_name": "A", "tax_year": "2025"}]), encoding="utf-8"
            )
            result = fee_workbook.run_fee_workbook(folder)
            self.assertEqual(result["year"], 2025)
            self.assertEqual(result["sheets"], ["2025", "2026"])  # next year created, ordered
            self.assertTrue(result["next_year_created"])
            self.assertTrue((folder / "fee_schedule.json").exists())

            # Edit the 2025 base price in the workbook, then re-run -> JSON reflects it.
            wb = load_workbook(folder / "fee_schedule.xlsx")
            for row in wb["2025"].iter_rows():
                if row[0].value == "base_1040":
                    row[2].value = 199.0
            wb.save(folder / "fee_schedule.xlsx")
            fee_workbook.run_fee_workbook(folder)
            schedule = json.loads((folder / "fee_schedule.json").read_text())
            self.assertEqual(schedule["base_1040"]["price"], 199.0)


if __name__ == "__main__":
    unittest.main()
