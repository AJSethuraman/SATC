"""Excel export for the Tax Withholding Estimate audit tape.

Requires openpyxl (``pip install openpyxl``).  The import is deferred inside
:func:`render_excel` so the rest of the package stays zero-dependency.
"""

from __future__ import annotations

from decimal import Decimal
from typing import Any

from twe.models import ZERO, EstimateResult, EstimatorInput
# Shared rendering helpers live in report.py to keep the four renderers consistent.
from twe.report import _FREQ_LABELS, _STATUS_LABELS, _usd  # noqa: PLC2701

# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def render_excel(
    inp: EstimatorInput,
    result: EstimateResult,
    *,
    generated_at: str | None = None,
) -> bytes:
    """Return the bytes of an .xlsx audit tape for *inp* / *result*.

    Raises :class:`ImportError` if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export: pip install openpyxl"
        ) from exc

    from datetime import datetime as _dt
    from io import BytesIO

    if generated_at is None:
        generated_at = _dt.now().strftime("%B %d, %Y  %I:%M %p")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Withholding Tape"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.sheet_view.showGridLines = False

    # ---- colour palette (openpyxl uses hex without #) ----
    NAVY  = "0B1F3A"
    NAVY2 = "173361"
    GOLD  = "B08D57"
    CREAM = "F6F2EA"
    CREAM2 = "EFE9DC"
    PAPER = "FBF9F4"
    INK   = "0E1726"
    MUTED = "64748B"
    GREEN = "2F5D3A"
    RED   = "9B2226"

    def _fill(color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=color)

    def _font(*, bold: bool = False, color: str = INK, size: int = 9) -> Font:
        return Font(bold=bold, color=color, name="Calibri", size=size)

    def _align(h: str = "left", v: str = "center", indent: int = 1) -> Alignment:
        return Alignment(horizontal=h, vertical=v, indent=indent)

    CURR_FMT = '"$"#,##0.00'
    CURR_NEG_FMT = '"$"#,##0.00_);[Red]"($"#,##0.00)'

    # ---- row-counter state ----
    _row: list[int] = [0]

    def R() -> int:
        _row[0] += 1
        return _row[0]

    # ---- layout helpers ----

    def merged_row(
        text: str,
        bg: str = PAPER,
        fg: str = INK,
        *,
        bold: bool = False,
        size: int = 9,
        height: int = 15,
        indent: int = 1,
    ) -> None:
        r = R()
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = Font(bold=bold, color=fg, name="Calibri", size=size)
        c.fill = _fill(bg)
        c.alignment = _align("left", indent=indent)
        ws.row_dimensions[r].height = height

    def section_header(title: str) -> None:
        # small gap before each section
        gap = R()
        ws.row_dimensions[gap].height = 5

        r = R()
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = title.upper()
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        c.fill = _fill(NAVY)
        c.alignment = _align("left", indent=1)
        ws.row_dimensions[r].height = 18

    def sub_label(text: str) -> None:
        r = R()
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = _font(bold=True, color=INK, size=9)
        c.fill = _fill(CREAM2)
        c.alignment = _align("left", indent=1)
        ws.row_dimensions[r].height = 15

    def text_val_row(label: str, text: str, bg: str = PAPER) -> None:
        r = R()
        a, b = ws.cell(r, 1), ws.cell(r, 2)
        a.value = label
        b.value = text
        for c in (a, b):
            c.font = _font()
            c.fill = _fill(bg)
        a.alignment = _align("left", indent=1)
        b.alignment = _align("right", indent=0)
        ws.row_dimensions[r].height = 14

    def amt_row(
        label: str,
        value: Decimal,
        *,
        bg: str = PAPER,
        bold: bool = False,
        fg: str = INK,
        subtract: bool = False,
        indent: int = 1,
    ) -> None:
        r = R()
        a, b = ws.cell(r, 1), ws.cell(r, 2)
        a.value = label
        b.value = float(value)
        fnt = _font(bold=bold, color=fg)
        for c in (a, b):
            c.font = fnt
            c.fill = _fill(bg)
        a.alignment = _align("left", indent=indent)
        b.alignment = _align("right", indent=0)
        b.number_format = CURR_NEG_FMT if subtract else CURR_FMT
        if subtract:
            b.value = -float(value)
        ws.row_dimensions[r].height = 14

    def total_row(
        label: str,
        value: Decimal,
        *,
        fg: str = INK,
        double_bottom: bool = False,
    ) -> None:
        r = R()
        a, b = ws.cell(r, 1), ws.cell(r, 2)
        a.value = label
        b.value = float(value)
        gold_side = Side(style="thin", color=GOLD)
        double_side = Side(style="double", color=GOLD)
        bot = double_side if double_bottom else gold_side
        for c in (a, b):
            c.font = _font(bold=True, color=fg)
            c.fill = _fill(CREAM)
            c.border = Border(top=Side(style="thin", color=GOLD), bottom=bot)
        a.alignment = _align("left", indent=1)
        b.alignment = _align("right", indent=0)
        b.number_format = CURR_FMT
        ws.row_dimensions[r].height = 16

    def rates_row(marginal: str, effective: str) -> None:
        r = R()
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = f"Marginal rate: {marginal}   |   Effective rate: {effective}"
        c.font = _font(bold=True, color=NAVY2)
        c.fill = _fill(CREAM2)
        c.alignment = _align("left", indent=1)
        ws.row_dimensions[r].height = 14

    def note_row(text: str) -> None:
        r = R()
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = _font(color=MUTED, size=8)
        c.fill = _fill(PAPER)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 13

    # ====================================================================
    # CONTENT
    # ====================================================================

    b = result.breakdown
    rec = result.recommendation
    status = _STATUS_LABELS.get(result.filing_status, result.filing_status)

    def pct(v: Decimal) -> str:
        return f"{float(v) * 100:.2f}%"

    # ---- header ----
    merged_row(
        "SETHURAMAN ACCOUNTING  ·  TAX  ·  CONSULTING",
        bg=NAVY, fg="FFFFFF", bold=True, size=12, height=24, indent=1,
    )
    merged_row(
        "FEDERAL TAX WITHHOLDING ESTIMATE  —  AUDIT TAPE",
        bg=NAVY2, fg="C8D8F0", size=10, height=18, indent=1,
    )
    merged_row(
        f"Tax Year: {result.tax_year_used}   ·   Filing Status: {status}   ·   {generated_at}",
        bg=CREAM2, fg=INK, size=9, height=15, indent=1,
    )

    # ---- inputs ----
    section_header("Scenario Inputs")

    all_jobs = inp.jobs
    for i, job in enumerate(all_jobs):
        name = job.name or ("Paystub" if len(all_jobs) == 1 else f"Job {i + 1}")
        sub_label(f"  {name}")
        text_val_row(
            "    Frequency",
            _FREQ_LABELS.get(job.pay_frequency, job.pay_frequency),
        )
        if job.taxable_wages_per_period is not None:
            amt_row("    Taxable wages / period", job.taxable_wages_per_period, indent=2)
        if job.gross_pay_per_period != ZERO:
            amt_row("    Gross pay / period", job.gross_pay_per_period, indent=2)
        amt_row("    Federal withheld / period", job.federal_tax_withheld_per_period, indent=2)
        if job.retirement_pretax_per_period != ZERO:
            amt_row("    Retirement pre-tax / period", job.retirement_pretax_per_period, indent=2)
        if job.other_pretax_per_period != ZERO:
            amt_row("    Other pre-tax / period", job.other_pretax_per_period, indent=2)
        if job.ytd_taxable_wages is not None:
            amt_row("    YTD taxable wages", job.ytd_taxable_wages, indent=2)
        if job.ytd_federal_tax_withheld is not None:
            amt_row("    YTD federal withheld", job.ytd_federal_tax_withheld, indent=2)
        if job.pay_periods_remaining is not None:
            text_val_row("    Periods remaining", f"{job.pay_periods_remaining} of {job.periods_per_year}")
        else:
            text_val_row("    Periods / year", str(job.periods_per_year))

    oi = inp.other_income
    oi_items: list[tuple[str, Decimal]] = [
        ("Interest", oi.interest),
        ("Ordinary dividends", oi.ordinary_dividends),
        ("Qualified dividends", oi.qualified_dividends),
        ("IRA / retirement distributions", oi.taxable_retirement_distributions),
        ("Taxable Social Security", oi.taxable_social_security),
        ("Long-term capital gains", oi.long_term_capital_gains),
        ("Short-term capital gains", oi.short_term_capital_gains),
        ("Net self-employment income", oi.self_employment_net),
        ("Unemployment compensation", oi.unemployment),
        ("Other taxable income", oi.other_taxable_income),
        ("Spouse taxable wages", oi.spouse_taxable_wages),
        ("Spouse federal withheld", oi.spouse_federal_tax_withheld),
    ]
    nonzero_oi = [(lbl, v) for lbl, v in oi_items if v != ZERO]
    if nonzero_oi:
        sub_label("  Other Income")
        for lbl, v in nonzero_oi:
            amt_row("    " + lbl, v, indent=2)

    adj = inp.adjustments
    adj_items: list[tuple[str, Decimal]] = [
        ("Traditional IRA deduction", adj.traditional_ira_deduction),
        ("HSA deduction", adj.hsa_deduction),
        ("Student loan interest", adj.student_loan_interest),
        ("Other adjustments", adj.other_adjustments),
    ]
    nonzero_adj = [(lbl, v) for lbl, v in adj_items if v != ZERO]
    if nonzero_adj:
        sub_label("  Adjustments (above the line)")
        for lbl, v in nonzero_adj:
            amt_row("    " + lbl, v, indent=2)

    ded = inp.deductions
    sub_label("  Deductions")
    if ded.itemized_total is not None:
        amt_row("    Itemized deductions", ded.itemized_total, indent=2)
    else:
        text_val_row("    Type", "Standard deduction")
    if ded.extra_standard_deductions:
        text_val_row("    Additional standard deductions", str(ded.extra_standard_deductions))

    cr = inp.credits
    cr_items: list[tuple[str, Decimal]] = [
        ("Child tax credit", cr.child_tax_credit),
        ("Other nonrefundable credits", cr.other_nonrefundable_credits),
        ("Refundable credits", cr.refundable_credits),
    ]
    nonzero_cr = [(lbl, v) for lbl, v in cr_items if v != ZERO]
    if nonzero_cr:
        sub_label("  Credits")
        for lbl, v in nonzero_cr:
            amt_row("    " + lbl, v, indent=2)

    op = inp.other_payments
    op_items: list[tuple[str, Decimal]] = [
        ("Estimated tax payments", op.estimated_tax_payments),
        ("Other withholding", op.other_withholding),
    ]
    nonzero_op = [(lbl, v) for lbl, v in op_items if v != ZERO]
    if nonzero_op:
        sub_label("  Other Payments Already Made")
        for lbl, v in nonzero_op:
            amt_row("    " + lbl, v, indent=2)

    if inp.target_refund != ZERO or inp.prior_year_tax is not None or inp.prior_year_agi is not None:
        sub_label("  Options / Safe Harbor")
        if inp.target_refund != ZERO:
            amt_row("    Target refund", inp.target_refund, indent=2)
        if inp.prior_year_tax is not None:
            amt_row("    Prior-year total tax", inp.prior_year_tax, indent=2)
        if inp.prior_year_agi is not None:
            amt_row("    Prior-year AGI", inp.prior_year_agi, indent=2)

    # ---- income computation ----
    section_header("Income Computation")

    amt_row("Projected taxable wages", b.projected_taxable_wages)
    income_addends: list[tuple[str, Decimal]] = [
        ("+ Interest", oi.interest),
        ("+ Ordinary dividends", oi.ordinary_dividends),
        ("+ IRA / retirement distributions", oi.taxable_retirement_distributions),
        ("+ Taxable Social Security", oi.taxable_social_security),
        ("+ Long-term capital gains", oi.long_term_capital_gains),
        ("+ Short-term capital gains", oi.short_term_capital_gains),
        ("+ Net self-employment income", oi.self_employment_net),
        ("+ Unemployment compensation", oi.unemployment),
        ("+ Other taxable income", oi.other_taxable_income),
        ("+ Spouse taxable wages", oi.spouse_taxable_wages),
    ]
    for lbl, v in income_addends:
        if v != ZERO:
            amt_row(lbl, v, fg=MUTED)
    total_row("TOTAL INCOME", b.total_income)

    if b.adjustments_total != ZERO:
        for lbl, v in [
            ("– Traditional IRA deduction", adj.traditional_ira_deduction),
            ("– HSA deduction", adj.hsa_deduction),
            ("– Student loan interest", adj.student_loan_interest),
            ("– Other adjustments", adj.other_adjustments),
        ]:
            if v != ZERO:
                amt_row(lbl, v, subtract=True, fg=RED)

    total_row("ADJUSTED GROSS INCOME (AGI)", b.adjusted_gross_income)
    amt_row(f"– Deduction  ({b.deduction_kind})", b.deduction_used, subtract=True, fg=RED)
    total_row("TAXABLE INCOME", b.taxable_income, double_bottom=True)

    # ---- tax computation ----
    section_header("Tax Computation")

    amt_row("Ordinary income tax", b.ordinary_income_tax)
    if b.capital_gains_tax != ZERO:
        amt_row("+ Cap. gains / qual. dividend tax", b.capital_gains_tax, fg=MUTED)
    if b.self_employment_tax != ZERO:
        amt_row("+ Self-employment tax", b.self_employment_tax, fg=MUTED)
    if b.additional_medicare_tax != ZERO:
        amt_row("+ Additional Medicare Tax (0.9%)", b.additional_medicare_tax, fg=MUTED)
    if b.net_investment_income_tax != ZERO:
        amt_row("+ Net Investment Income Tax (3.8%)", b.net_investment_income_tax, fg=MUTED)
    total_row("Income tax before credits", b.income_tax_before_credits)
    if b.nonrefundable_credits != ZERO:
        amt_row("– Nonrefundable credits", b.nonrefundable_credits, subtract=True, fg=RED)
    if b.refundable_credits != ZERO:
        amt_row("– Refundable credits", b.refundable_credits, subtract=True, fg=RED)
    total_row("TOTAL TAX LIABILITY", b.total_tax_liability, double_bottom=True)
    rates_row(pct(b.marginal_rate), pct(b.effective_rate))

    # ---- withholding analysis ----
    section_header("Withholding Analysis")

    multi_job = len(rec.job_breakdown) > 1
    if multi_job:
        for job in rec.job_breakdown:
            sub_label(f"  {job.name}  ({job.pay_frequency})")
            amt_row(f"    YTD withheld  ({job.periods_elapsed} of {job.periods_per_year} periods)", job.ytd_withholding, indent=2)
            job_proj_remaining = job.projected_withholding - job.ytd_withholding
            if job_proj_remaining > ZERO:
                amt_row(f"    + Remaining  ({job.periods_remaining} periods, current rate)", job_proj_remaining, fg=MUTED, indent=2)
        amt_row("YTD withheld — all jobs", rec.ytd_withholding)
        proj_remaining = rec.projected_withholding_current_rate - rec.ytd_withholding
        if proj_remaining > ZERO:
            amt_row("+ Remaining — all jobs (current rate)", proj_remaining, fg=MUTED)
    else:
        proj_remaining = rec.projected_withholding_current_rate - rec.ytd_withholding
        amt_row(
            f"YTD withheld  ({rec.periods_elapsed} of {rec.periods_per_year} periods)",
            rec.ytd_withholding,
        )
        if proj_remaining > ZERO:
            amt_row(
                f"+ Remaining  ({rec.periods_remaining} periods, current rate)",
                proj_remaining,
                fg=MUTED,
            )
    if rec.other_payments_total != ZERO:
        amt_row("+ Other payments / spouse withheld", rec.other_payments_total, fg=MUTED)
    total_row("PROJECTED TOTAL PAYMENTS", rec.projected_total_payments)

    amt_row("Projected tax liability", b.total_tax_liability)
    amt_row("– Projected total payments", rec.projected_total_payments, subtract=True, fg=RED)

    balance = rec.projected_balance
    if balance >= ZERO:
        total_row("PROJECTED REFUND", balance, fg=GREEN, double_bottom=True)
    else:
        total_row("PROJECTED BALANCE DUE", -balance, fg=RED, double_bottom=True)

    # ---- recommendation ----
    section_header("Recommendation")

    if inp.target_refund != ZERO:
        amt_row("Target refund", inp.target_refund)

    if rec.is_over_withholding:
        reduction = rec.adjusted_job_withholding_per_period - rec.recommended_withholding_per_period
        step3_approx = reduction * rec.periods_per_year
        job_label = f"  ({rec.adjusted_job_name})" if multi_job else ""
        job_w4_owner = f"{rec.adjusted_job_name}'s" if multi_job else "your"
        merged_row(
            "You are on track to OVER-WITHHOLD for your target.",
            bg=CREAM, fg=GREEN, bold=True, size=9, height=15,
        )
        amt_row(f"Current withholding / paycheck{job_label}", rec.adjusted_job_withholding_per_period)
        amt_row(f"Recommended withholding / paycheck{job_label}", rec.recommended_withholding_per_period, bold=True, fg=GREEN)
        total_row(f"REDUCTION / PAYCHECK{job_label}", reduction, fg=GREEN, double_bottom=True)
        note_row(
            f"W-4 Step 3 — Enter ~{_usd(step3_approx)} in Step 3 of {job_w4_owner} W-4  "
            f"({_usd(reduction)} × {rec.periods_per_year} periods/yr = {_usd(step3_approx)})"
        )
    else:
        on_job = f'  (for "{rec.adjusted_job_name}" W-4)' if multi_job else ""
        merged_row(
            "Action recommended — currently UNDER-WITHHOLDING.",
            bg=CREAM, fg=RED, bold=True, size=9, height=15,
        )
        amt_row("Recommended withholding / paycheck" + on_job, rec.recommended_withholding_per_period, bold=True)
        amt_row("Current withholding / paycheck", rec.adjusted_job_withholding_per_period)
        total_row("EXTRA NEEDED / PAYCHECK  (W-4 Step 4c)", rec.additional_withholding_per_period, fg=RED, double_bottom=True)
        note_row(
            f"Enter {_usd(rec.additional_withholding_per_period)} as additional withholding{on_job} "
            f"on Form W-4, Step 4(c).  "
            f"({rec.periods_remaining} pay period{'s' if rec.periods_remaining != 1 else ''} remaining.)"
        )

    # ---- safe harbor ----
    if rec.safe_harbor_target is not None:
        section_header("Safe Harbor  (underpayment penalty threshold)")
        met = rec.projected_total_payments >= rec.safe_harbor_target
        amt_row("Minimum payments for safe harbor", rec.safe_harbor_target)
        amt_row("Projected total payments", rec.projected_total_payments)
        merged_row(
            "Status: SAFE HARBOR MET" if met else "Status: SAFE HARBOR NOT MET",
            bg=CREAM, fg=GREEN if met else RED, bold=True, size=9, height=15,
        )
        if rec.safe_harbor_additional_per_period is not None and rec.safe_harbor_additional_per_period > ZERO:
            amt_row("Extra / paycheck for safe harbor", rec.safe_harbor_additional_per_period)

    # ---- notes ----
    if result.notes:
        section_header("Notes")
        for note in result.notes:
            note_row(f"• {note}")

    # ---- footer ----
    gap = R()
    ws.row_dimensions[gap].height = 8
    merged_row(
        "For planning purposes only — not tax advice.  "
        "Federal income tax only.  Always verify with a qualified tax professional.  "
        "Sethuraman Accounting · Tax · Consulting",
        bg=NAVY, fg="9DB0CC", size=8, height=22, indent=1,
    )

    # Freeze the top 3 header rows so they stay visible while scrolling
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
