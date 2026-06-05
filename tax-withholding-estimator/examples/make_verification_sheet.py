"""Generate tax_verify_2025.xlsx — a step-by-step Excel verification of the
tax withholding estimator against 2025 IRS tables.

Run:
    python examples/make_verification_sheet.py

Opens (or creates) tax_verify_2025.xlsx in the current directory.
All yellow cells are user inputs. All blue cells are computed by Excel
formulas so you can change the inputs and watch everything update.

Compare the green "TWE Output" column against what `twe estimate` or
the web UI actually reports — they should match within $1 (rounding).
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "tax_verify_2025.xlsx")

# ── palette ──────────────────────────────────────────────────────────────────
HDR   = PatternFill("solid", fgColor="0F766E")   # teal header
INP   = PatternFill("solid", fgColor="FFFDE7")   # yellow input
CALC  = PatternFill("solid", fgColor="E3F2FD")   # blue calculated
RES   = PatternFill("solid", fgColor="E8F5E9")   # green result
WARN  = PatternFill("solid", fgColor="FFF3E0")   # orange info
EXPL  = PatternFill("solid", fgColor="F3E5F5")   # purple explanation
WHITE = PatternFill("solid", fgColor="FFFFFF")
NONE  = PatternFill("none")

HDR_F  = Font(bold=True, color="FFFFFF", size=10)
LBL_F  = Font(bold=True, size=9)
NORM_F = Font(size=9)
SMALL  = Font(size=8, italic=True, color="546E7A")
NOTE_F = Font(size=8, color="546E7A")

THIN = Side(style="thin", color="B0BEC5")
MED  = Side(style="medium", color="0F766E")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BOX = Border(left=MED, right=MED, top=MED, bottom=MED)

USD  = '#,##0.00'
PCT  = '0.00%'
INT_ = '#,##0'


def _hdr(ws, row, label, span=3):
    ws.cell(row, 1, label).font = HDR_F
    ws.cell(row, 1).fill = HDR
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row, 1).border = MED_BOX
    for c in range(2, span + 1):
        ws.cell(row, c).fill = HDR
        ws.cell(row, c).border = MED_BOX
    ws.row_dimensions[row].height = 16


def _row(ws, row, label, value=None, formula=None, fmt=USD,
         fill=None, note="", bold_label=False):
    lc = ws.cell(row, 1, label)
    lc.font = Font(bold=bold_label, size=9)
    lc.alignment = Alignment(indent=1, vertical="center")
    lc.border = BOX

    vc = ws.cell(row, 2)
    if formula:
        vc.value = formula
        vc.fill = fill or CALC
    elif value is not None:
        vc.value = value
        vc.fill = fill or INP
    else:
        vc.fill = fill or NONE
    vc.number_format = fmt
    vc.alignment = Alignment(horizontal="right", vertical="center")
    vc.border = BOX
    vc.font = NORM_F

    nc = ws.cell(row, 3, note)
    nc.font = NOTE_F
    nc.alignment = Alignment(vertical="center", wrap_text=True)
    nc.border = BOX

    ws.row_dimensions[row].height = 15
    return vc


EMPTY_BORDER = Border()

def _blank(ws, row):
    for c in range(1, 4):
        ws.cell(row, c).fill = WHITE
        ws.cell(row, c).border = EMPTY_BORDER
    ws.row_dimensions[row].height = 6


# ── named-cell helpers ───────────────────────────────────────────────────────
# We track cell addresses for formula references by name.
REF = {}

def _addr(cell):
    return f"$B${cell.row}"

def _reg(name, cell):
    REF[name] = _addr(cell)
    return cell


# ── "How It Works" explanation sheet ─────────────────────────────────────────
EXPLAIN = [
    ("OVERVIEW", None,
     "This sheet walks through how the Tax Withholding Estimator (TWE) calculates "
     "your recommended federal withholding per paycheck. Each section below matches "
     "a numbered section on the Verify 2025 sheet. Change the yellow input cells "
     "there and every formula updates automatically."),

    ("① INPUTS", None, None),

    ("  Filing status", None,
     "Controls which tax bracket table and standard deduction amount are used. "
     "The four options are: single, married_jointly, married_separately, head_of_household. "
     "The Verify 2025 sheet is pre-set to 'single' with 2025 single-filer constants."),

    ("  Standard deduction 2025", None,
     "Single: $15,000 | Married jointly: $30,000 | Head of household: $22,500 | "
     "Married separately: $15,000. Add $2,000 if single and age 65+ or blind "
     "(add $1,600 per qualifying person if married jointly)."),

    ("  Pay periods per year", None,
     "Weekly = 52 | Biweekly = 26 | Semimonthly = 24 | Monthly = 12. "
     "Your paystub should show pay period or check date — count how many times "
     "per year you receive a paycheck."),

    ("  Taxable wages this period", None,
     "The gross wages that appear in Box 1 of your W-2 equivalent. This is NOT "
     "your gross pay — it is gross pay MINUS pre-tax deductions (401k, HSA, "
     "medical premiums, FSA, etc.). Your paystub may label this 'Federal Taxable' "
     "or 'Current Taxable Wages'."),

    ("② INCOME PROJECTION — how YTD is inferred", None, None),

    ("  Elapsed periods", None,
     "= periods_per_year − remaining_periods.\n"
     "Example: biweekly (26) with 14 remaining → 26 − 14 = 12 elapsed."),

    ("  YTD taxable wages — if you leave it at 0", None,
     "The tool infers it as: elapsed_periods × taxable_wages_per_period.\n"
     "Example: 12 elapsed × $2,950/period = $35,400 inferred YTD wages.\n\n"
     "If you enter a non-zero value, the actual YTD figure from your paystub "
     "is used instead. Always prefer the real number from your paystub "
     "because it accounts for mid-year changes in pay rate or deductions."),

    ("  YTD withholding — if you leave it at 0", None,
     "Same inference logic: elapsed_periods × withholding_per_period.\n"
     "Again, if your paystub shows a YTD federal tax withheld amount, enter "
     "that real number — it is more accurate."),

    ("  Projected wages", None,
     "= YTD_wages_used + (remaining_periods × taxable_wages_per_period).\n"
     "This is the best estimate of what Box 1 of your W-2 will say."),

    ("③ ADJUSTMENTS → AGI", None, None),

    ("  What counts as an adjustment?", None,
     "Above-the-line deductions reduce income BEFORE the standard/itemized "
     "deduction is applied. Common ones:\n"
     "• Traditional IRA contribution (if deductible) — up to $7,000 ($8,000 age 50+)\n"
     "• HSA contribution (employee portion) — up to $4,300 single / $8,550 family\n"
     "• Student loan interest paid — up to $2,500\n"
     "• One-half of self-employment tax (computed automatically in Section ⑦)"),

    ("  One-half SE tax approximation in Section ③", None,
     "Section ③ uses a quick approximation so AGI can be estimated early. "
     "Section ⑦ computes the exact SE tax. If you have SE income, confirm "
     "the exact half-SE deduction matches between the two sections — "
     "they should be within a few dollars."),

    ("④ DEDUCTION → TAXABLE INCOME", None, None),

    ("  Standard vs. itemized", None,
     "The tool picks whichever is larger. If your itemized deductions "
     "(mortgage interest, state taxes up to $10K, charitable contributions, "
     "unreimbursed medical over 7.5% of AGI) exceed the standard deduction, "
     "enter your itemized total in the input cell. Otherwise leave it at 0 "
     "and the standard deduction applies automatically."),

    ("  Taxable income", None,
     "= AGI − deduction (standard or itemized, whichever is larger).\n"
     "This is the number the ordinary tax brackets are applied to, "
     "after the LTCG/qualified dividend portion is separated out."),

    ("⑤ ORDINARY INCOME TAX — how brackets work", None, None),

    ("  LTCG / qualified dividends are separated first", None,
     "Long-term capital gains and qualified dividends get their own preferential "
     "rate (0/15/20%). They are removed from ordinary taxable income so the "
     "ordinary brackets apply only to wages, interest, short-term gains, etc.\n\n"
     "Preferential income = MIN(LTCG + qual_dividends, taxable_income)\n"
     "Ordinary TI = taxable_income − preferential_income"),

    ("  Bracket math (single, 2025)", None,
     "The brackets are cumulative — each bracket only taxes the income WITHIN "
     "that range, not all income at the higher rate:\n"
     "  $0 – $11,925          → 10%  (max tax: $1,192.50)\n"
     "  $11,925 – $48,475      → 12%  (max additional: $4,386.00)\n"
     "  $48,475 – $103,350     → 22%  (max additional: $12,072.50)\n"
     "  $103,350 – $197,300    → 24%  (max additional: $22,548.00)\n"
     "  $197,300 – $250,525    → 32%  (max additional: $17,032.00)\n"
     "  $250,525 – $626,350    → 35%  (max additional: $131,538.75)\n"
     "  Over $626,350          → 37%\n\n"
     "The IFS formula on the Verify sheet encodes the cumulative amounts at "
     "each breakpoint so it only needs one lookup rather than stacking."),

    ("⑥ CAPITAL GAINS TAX — 'stacking' explained", None, None),

    ("  Why LTCG rates depend on ordinary income", None,
     "The 0%/15%/20% LTCG rate thresholds apply to TOTAL taxable income "
     "(ordinary + preferential combined), not just the gains themselves. "
     "So if your ordinary income already uses up the 0% room, your gains "
     "are taxed at 15% even if they would otherwise fall below the threshold.\n\n"
     "Example (single 2025):\n"
     "  Ordinary TI = $40,000 → uses $40,000 of the $48,350 zero-rate room\n"
     "  Room remaining = $48,350 − $40,000 = $8,350\n"
     "  LTCG = $5,000 → first $8,350 of LTCG taxed at 0%, so all $5,000 = 0%\n\n"
     "  Ordinary TI = $50,000 → already exceeds $48,350, zero-rate room = $0\n"
     "  LTCG = $5,000 → all taxed at 15%"),

    ("⑦ SELF-EMPLOYMENT TAX", None, None),

    ("  Why the 0.9235 factor?", None,
     "As an employee, your employer pays half of FICA (SS + Medicare). "
     "Self-employed people pay both halves, but the IRS lets you deduct the "
     "employer-equivalent half. The 0.9235 factor (= 1 − 0.0765) approximates "
     "this: it reduces your net earnings to the amount that would be left after "
     "a hypothetical employer's 7.65% share is removed."),

    ("  Social Security wage base", None,
     "2025 SS wage base = $176,100. Only earnings up to this amount are subject "
     "to the 12.4% SS portion. Medicare (2.9%) has no cap. If your projected "
     "wages already exceed $176,100, none of your SE income faces SS tax."),

    ("  One-half SE tax deduction", None,
     "Half of your total SE tax (SS + Medicare) is deductible above the line, "
     "reducing AGI. Section ③ estimates this; Section ⑦ computes it exactly. "
     "If SE income is large, update the half-SE input in Section ③ to the "
     "exact value from Section ⑦ for a more accurate AGI."),

    ("⑧ ADDITIONAL MEDICARE TAX & NIIT", None, None),

    ("  Additional Medicare Tax (0.9%)", None,
     "Applies to wages/SE earnings over $200,000 (single), $250,000 (married jointly), "
     "$125,000 (married separately). Your employer withholds this automatically once "
     "your wages exceed $200K in a calendar year, but they do NOT know about your "
     "spouse's wages or SE income — you may owe more at filing."),

    ("  Net Investment Income Tax — NIIT (3.8%)", None,
     "Applies to the LESSER of:\n"
     "  (a) net investment income (interest + dividends + capital gains), or\n"
     "  (b) AGI − $200,000 (single) / $250,000 (MFJ)\n\n"
     "If your AGI is below $200K, NIIT is $0 regardless of investment income."),

    ("⑨ TAX LIABILITY", None, None),

    ("  Order of credits", None,
     "Nonrefundable credits (Child Tax Credit, education credits, etc.) reduce "
     "tax but cannot create a refund — they stop at $0 tax. Refundable credits "
     "(EITC, Additional Child Tax Credit) can produce a refund even if tax = $0. "
     "The tool applies nonrefundable credits first, then refundable credits."),

    ("⑩ WITHHOLDING RECOMMENDATION", None, None),

    ("  How the per-paycheck number is calculated", None,
     "1. Compute total liability (Section ⑨)\n"
     "2. Add target refund (usually $0 for break-even)\n"
     "3. Subtract everything already secured:\n"
     "   • YTD withholding from your paystub\n"
     "   • Estimated tax payments already made\n"
     "   • Spouse/other withholding\n"
     "4. Result = amount still needed from remaining paychecks\n"
     "5. Divide by remaining_periods → recommended per-paycheck withholding\n"
     "6. If that exceeds your normal withholding, the difference is what to\n"
     "   enter on Form W-4 line 4(c) as extra withholding."),

    ("  Target refund", None,
     "Set to $0 to break even. A positive target refund (e.g. $500) means you "
     "want to over-withhold by that amount so you get a refund at filing. "
     "A negative target means you accept owing up to that amount at filing "
     "(only safe if within safe-harbor limits — see Section ⑪)."),

    ("⑪ SAFE HARBOR — penalty avoidance", None, None),

    ("  What is safe harbor?", None,
     "The IRS waives the underpayment penalty if you paid the lesser of:\n"
     "  (a) 90% of this year's tax liability, OR\n"
     "  (b) 100% of last year's tax (110% if last year's AGI exceeded $150,000)\n\n"
     "Safe harbor does NOT protect you from owing money — it just means no "
     "penalty on top of what you owe. The Section ⑪ row on the Verify sheet "
     "shows how much additional withholding per check is needed to hit this "
     "minimum, as an alternative to the full break-even recommendation."),

    ("  High-income rule (110%)", None,
     "If your prior-year AGI was over $150,000 (single or married jointly), "
     "you must pay 110% of last year's tax (not 100%) to qualify for the "
     "prior-year safe harbor. Married filing separately threshold: $75,000. "
     "This is captured in the IFS formula in Section ⑪."),
]


def _build_explain_sheet(wb):
    ws = wb.create_sheet("How It Works")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    ws.row_dimensions[1].height = 24
    t = ws.cell(1, 1, "How the Estimator Computes Your Withholding")
    t.font = Font(bold=True, size=13, color="0F766E")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:B1")

    ws.row_dimensions[2].height = 14
    sub = ws.cell(2, 1, "Plain-English explanation of each section on the 'Verify 2025' sheet.")
    sub.font = Font(size=9, italic=True, color="546E7A")
    ws.merge_cells("A2:B2")

    r = 3
    for topic, _unused, text in EXPLAIN:
        r += 1
        if text is None:
            # section header
            tc = ws.cell(r, 1, topic)
            tc.font = Font(bold=True, size=10, color="FFFFFF")
            tc.fill = HDR
            tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            tc.border = MED_BOX
            ws.cell(r, 2).fill = HDR
            ws.cell(r, 2).border = MED_BOX
            ws.row_dimensions[r].height = 16
        else:
            # label + explanation
            lc = ws.cell(r, 1, topic)
            lc.font = Font(bold=True, size=9, color="4A148C")
            lc.fill = EXPL
            lc.alignment = Alignment(indent=1, vertical="top", wrap_text=True)
            lc.border = BOX

            ec = ws.cell(r, 2, text)
            ec.font = Font(size=9)
            ec.alignment = Alignment(vertical="top", wrap_text=True)
            ec.border = BOX

            # auto-size row height based on approximate line count
            lines = text.count("\n") + max(1, len(text) // 80)
            ws.row_dimensions[r].height = max(15, lines * 14 + 6)

    ws.freeze_panes = "A3"


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Verify 2025"

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52

    # freeze header rows
    ws.freeze_panes = "A5"

    # ── title ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    t = ws.cell(1, 1, "Tax Withholding Estimator — 2025 Verification Sheet")
    t.font = Font(bold=True, size=13, color="0F766E")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:C1")

    ws.row_dimensions[2].height = 13
    d = ws.cell(2, 1,
        "Yellow = input you type.  Blue = Excel formula.  Green = key result.  "
        "See the 'How It Works' tab for plain-English logic explanations.")
    d.font = Font(size=8, italic=True, color="546E7A")
    ws.merge_cells("A2:C2")

    ws.row_dimensions[3].height = 6

    # column D header
    ws.column_dimensions["D"].width = 16
    dh = ws.cell(4, 4, "TWE Output\n(paste here)")
    dh.font = Font(bold=True, size=9, color="0F766E")
    dh.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    dh.fill = PatternFill("solid", fgColor="E0F7FA")
    ws.row_dimensions[4].height = 28

    r = 4  # current row counter

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1: INPUTS
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _hdr(ws, r, "① INPUTS  (pre-filled from examples/sample_input.json)", 4)

    r += 1; _hdr(ws, r, "  Filing & Year", 4)
    r += 1; _reg("filing",   _row(ws, r, "Filing status", "single",  fmt="@",
                                  note='single | married_jointly | married_separately | head_of_household'))
    r += 1; _reg("tax_year", _row(ws, r, "Tax year",      2025,       fmt=INT_))
    r += 1; _reg("std_ded",  _row(ws, r, "Standard deduction (2025 single)", 15000, fmt=USD,
                                  note="Single $15K | MFJ $30K | HoH $22.5K | MFS $15K  (+$2K if 65+/blind, single)"))

    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "  Paystub  (Job 1)", 4)
    r += 1; _reg("freq",      _row(ws, r, "Pay frequency",                   "biweekly", fmt="@",
                                   note="weekly=52  biweekly=26  semimonthly=24  monthly=12"))
    r += 1; _reg("ppy",       _row(ws, r, "Periods per year",                26,   fmt=INT_,
                                   note="biweekly = 26"))
    r += 1; _reg("taxable_pp",_row(ws, r, "Taxable wages this period ($)",   2950, fmt=USD,
                                   note="Gross pay MINUS pre-tax deductions (401k, HSA, etc.) — your 'Federal Taxable' line"))
    r += 1; _reg("wh_pp",     _row(ws, r, "Federal tax withheld per period ($)", 410, fmt=USD,
                                   note="'Federal Income Tax' on your paystub — NOT SS or Medicare"))
    r += 1; _reg("remaining", _row(ws, r, "Pay periods remaining",           14,   fmt=INT_,
                                   note="Count paychecks left this calendar year including the next one"))
    r += 1; _reg("ytd_wages", _row(ws, r, "YTD taxable wages ($)",           35100,fmt=USD,
                                   note="From paystub YTD column. Leave 0 → inferred as elapsed×per-period (see How It Works tab)"))
    r += 1; _reg("ytd_wh",    _row(ws, r, "YTD federal tax withheld ($)",    4920, fmt=USD,
                                   note="From paystub YTD column. Leave 0 → inferred as elapsed×per-period"))

    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "  Other Income (annual)", 4)
    r += 1; _reg("interest",  _row(ws, r, "Interest income ($)",                  350,  fmt=USD,
                                   note="1099-INT box 1"))
    r += 1; _reg("ord_div",   _row(ws, r, "Ordinary dividends ($)",               800,  fmt=USD,
                                   note="1099-DIV box 1a"))
    r += 1; _reg("qual_div",  _row(ws, r, "Qualified dividends ($)",              600,  fmt=USD,
                                   note="1099-DIV box 1b — must be ≤ ordinary dividends; taxed at 0/15/20% (see How It Works)"))
    r += 1; _reg("ira_dist",  _row(ws, r, "Taxable IRA / retirement distributions ($)", 5000, fmt=USD,
                                   note="1099-R box 2a taxable amount — Roth conversions, traditional withdrawals"))
    r += 1; _reg("ltcg",      _row(ws, r, "Long-term capital gains ($)",         2000,  fmt=USD,
                                   note="Held >1 year; taxed at 0/15/20% stacked on top of ordinary income (see How It Works)"))
    r += 1; _reg("stcg",      _row(ws, r, "Short-term capital gains ($)",        0,     fmt=USD,
                                   note="Held ≤1 year; taxed as ordinary income"))
    r += 1; _reg("se_net",    _row(ws, r, "Self-employment net income ($)",      0,     fmt=USD,
                                   note="Schedule C profit (after business expenses). SE tax computed in Section ⑦"))
    r += 1; _reg("unemp",     _row(ws, r, "Unemployment compensation ($)",       0,     fmt=USD))
    r += 1; _reg("other_inc", _row(ws, r, "Other taxable income ($)",            0,     fmt=USD))
    r += 1; _reg("sp_wages",  _row(ws, r, "Spouse taxable wages ($)",           0,     fmt=USD))
    r += 1; _reg("sp_wh",     _row(ws, r, "Spouse federal tax withheld ($)",    0,     fmt=USD))

    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "  Adjustments (above-the-line — reduce income before deduction)", 4)
    r += 1; _reg("ira_ded",   _row(ws, r, "Traditional IRA deduction ($)",  0,    fmt=USD,
                                   note="Deductible if no workplace plan, or income below phaseout; up to $7,000 ($8K age 50+)"))
    r += 1; _reg("hsa_ded",   _row(ws, r, "HSA deduction ($)",              2000, fmt=USD,
                                   note="Employee-paid HSA contributions not via payroll; 2025 limit $4,300 self / $8,550 family"))
    r += 1; _reg("sl_int",    _row(ws, r, "Student loan interest ($)",      1200, fmt=USD,
                                   note="Deductible up to $2,500; phases out at $85K–$100K AGI (single 2025)"))
    r += 1; _reg("other_adj", _row(ws, r, "Other adjustments ($)",          0,    fmt=USD,
                                   note="Educator expenses ($300), alimony (pre-2019 divorces), self-employed health ins., etc."))

    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "  Deductions & Credits", 4)
    r += 1; _reg("itemized",  _row(ws, r, "Itemized total ($, or 0 = use standard)", 0, fmt=USD,
                                   note="Mortgage int + state/local taxes (cap $10K) + charity + medical >7.5% AGI. 0 → standard wins"))
    r += 1; _reg("ctc",       _row(ws, r, "Child Tax Credit ($)",           0, fmt=USD,
                                   note="Up to $2,000/child under 17; phases out above $200K AGI (single). Enter expected credit amount."))
    r += 1; _reg("nr_cred",   _row(ws, r, "Other nonrefundable credits ($)",0, fmt=USD,
                                   note="Education credits, child/dependent care, savers credit, etc. Cannot exceed tax owed."))
    r += 1; _reg("ref_cred",  _row(ws, r, "Refundable credits ($)",         0, fmt=USD,
                                   note="EITC, Additional Child Tax Credit, American Opportunity Credit (40% refundable). Can produce refund."))

    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "  Other Payments & Target", 4)
    r += 1; _reg("est_pay",   _row(ws, r, "Estimated tax payments ($)",     0, fmt=USD,
                                   note="Quarterly Form 1040-ES payments already made this year"))
    r += 1; _reg("other_wh",  _row(ws, r, "Other withholding ($)",          0, fmt=USD,
                                   note="Backup withholding, pension withholding, etc."))
    r += 1; _reg("target_ref",_row(ws, r, "Target refund ($, 0 = break even)", 0, fmt=USD,
                                   note="Positive = over-withhold to get a refund. Negative = accept owing up to this amount at filing."))
    r += 1; _reg("py_tax",    _row(ws, r, "Prior-year tax ($)",           8200, fmt=USD,
                                   note="Your actual 2024 total tax (Form 1040 line 24). Used for safe-harbor calculation in Section ⑪"))
    r += 1; _reg("py_agi",    _row(ws, r, "Prior-year AGI ($)",          71000, fmt=USD,
                                   note="2024 AGI (Form 1040 line 11). If >$150K, safe harbor requires 110% of prior-year tax."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2: INCOME PROJECTION
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "② INCOME PROJECTION", 4)

    elapsed_formula = f"={REF['ppy']}-{REF['remaining']}"
    r += 1; _reg("elapsed", _row(ws, r, "Elapsed periods",
        formula=elapsed_formula, fmt=INT_,
        note="= periods_per_year − remaining. Example: biweekly 26 total, 14 remaining → 12 elapsed"))

    ytd_used = (f"=IF({REF['ytd_wages']}>0, {REF['ytd_wages']}, "
                f"{REF['taxable_pp']}*{REF['elapsed']})")
    r += 1; _reg("ytd_used", _row(ws, r, "YTD taxable wages used",
        formula=ytd_used, fmt=USD,
        note="Uses input if non-zero; otherwise infers as elapsed×per-period. Enter actual YTD from paystub for accuracy."))

    proj_wages = f"={REF['ytd_used']}+{REF['taxable_pp']}*{REF['remaining']}"
    r += 1; _reg("proj_wages", _row(ws, r, "Projected full-year taxable wages",
        formula=proj_wages, fmt=USD,
        note="= YTD used + (remaining × per-period). Best estimate of W-2 Box 1."))

    ytd_wh_used = (f"=IF({REF['ytd_wh']}>0, {REF['ytd_wh']}, "
                   f"{REF['wh_pp']}*{REF['elapsed']})")
    r += 1; _reg("ytd_wh_used", _row(ws, r, "YTD withholding used",
        formula=ytd_wh_used, fmt=USD,
        note="Uses input if non-zero; otherwise infers as elapsed×per-period"))

    proj_wh = f"={REF['ytd_wh_used']}+{REF['wh_pp']}*{REF['remaining']}"
    r += 1; _reg("proj_wh", _row(ws, r, "Projected total withholding (job 1 only)",
        formula=proj_wh, fmt=USD,
        note="Projected if nothing changes. Add spouse/other withholding in Section ⑩."))

    total_income = (f"={REF['proj_wages']}+{REF['interest']}+{REF['ord_div']}"
                    f"+{REF['ira_dist']}+{REF['stcg']}+{REF['ltcg']}"
                    f"+{REF['se_net']}+{REF['unemp']}+{REF['other_inc']}"
                    f"+{REF['sp_wages']}")
    r += 1; _reg("total_income", _row(ws, r, "Total income",
        formula=total_income, fmt=USD,
        note="Sum of all income sources before any deductions or adjustments"))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3: AGI
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "③ ADJUSTMENTS → AGI", 4)

    half_se_f = (f"=IF({REF['se_net']}<=0, 0, "
                 f"({REF['se_net']}*0.9235*(0.124+0.029))/2)")
    r += 1; _reg("half_se_approx", _row(ws, r, "One-half SE tax deduction (approx.)",
        formula=half_se_f, fmt=USD,
        note="Quick estimate used here. Compare to exact value in Section ⑦ — they should agree. Large SE income? Update AGI manually."))

    adj_total = (f"={REF['ira_ded']}+{REF['hsa_ded']}+{REF['sl_int']}"
                 f"+{REF['other_adj']}+{REF['half_se_approx']}")
    r += 1; _reg("adj_total", _row(ws, r, "Total adjustments (above-the-line)",
        formula=adj_total, fmt=USD,
        note="IRA + HSA + student loan interest + other + ½ SE tax"))

    agi_f = f"=MAX(0, {REF['total_income']}-{REF['adj_total']})"
    r += 1; _reg("agi", _row(ws, r, "Adjusted Gross Income (AGI)",
        formula=agi_f, fmt=USD, fill=RES, bold_label=True,
        note="Total income minus above-the-line adjustments. Many credits and phaseouts are based on AGI."))
    ws.cell(r, 4).number_format = USD

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4: DEDUCTION → TAXABLE INCOME
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "④ DEDUCTION → TAXABLE INCOME", 4)

    deduction_f = (f"=IF({REF['itemized']}>{REF['std_ded']}, "
                   f"{REF['itemized']}, {REF['std_ded']})")
    r += 1; _reg("deduction", _row(ws, r, "Deduction used (larger of standard / itemized)",
        formula=deduction_f, fmt=USD,
        note="Itemized wins only if it exceeds standard. Most filers use standard after 2017 TCJA."))

    ti_f = f"=MAX(0, {REF['agi']}-{REF['deduction']})"
    r += 1; _reg("taxable_income", _row(ws, r, "Taxable Income",
        formula=ti_f, fmt=USD, fill=RES, bold_label=True,
        note="= AGI − deduction. Brackets are applied to this number (after separating LTCG below)."))
    ws.cell(r, 4).number_format = USD

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 5: ORDINARY INCOME TAX (2025 Single brackets hardcoded)
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑤ ORDINARY INCOME TAX  (2025 Single brackets — update for other status)", 4)

    pref_f = (f"=MIN(MAX(0,{REF['qual_div']}+{REF['ltcg']}), {REF['taxable_income']})")
    r += 1; _reg("preferential", _row(ws, r, "Preferential income (LTCG + qual. dividends)",
        formula=pref_f, fmt=USD,
        note="Removed from ordinary TI and taxed at 0/15/20% in Section ⑥. Capped at taxable income."))

    ord_ti_f = f"=MAX(0, {REF['taxable_income']}-{REF['preferential']})"
    r += 1; _reg("ord_ti", _row(ws, r, "Ordinary taxable income",
        formula=ord_ti_f, fmt=USD,
        note="= Taxable income − preferential. This is what the 10–37% brackets apply to."))

    oti = REF['ord_ti']
    bracket_tax = (
        f"=IFS("
        f"{oti}<=0, 0, "
        f"{oti}<=11925, {oti}*0.10, "
        f"{oti}<=48475, 1192.50+({oti}-11925)*0.12, "
        f"{oti}<=103350, 5578.50+({oti}-48475)*0.22, "
        f"{oti}<=197300, 17651.00+({oti}-103350)*0.24, "
        f"{oti}<=250525, 40199.00+({oti}-197300)*0.32, "
        f"{oti}<=626350, 57231.00+({oti}-250525)*0.35, "
        f"TRUE, 188769.75+({oti}-626350)*0.37)"
    )
    r += 1; _reg("ord_tax", _row(ws, r, "Ordinary income tax",
        formula=bracket_tax, fmt=USD,
        note="Cumulative bracket amounts encoded: $1,192.50 at $11,925 / $5,578.50 at $48,475 / etc."))

    marg_f = (
        f"=IFS("
        f"{oti}<=0, 0.10, "
        f"{oti}<=11925, 0.10, "
        f"{oti}<=48475, 0.12, "
        f"{oti}<=103350, 0.22, "
        f"{oti}<=197300, 0.24, "
        f"{oti}<=250525, 0.32, "
        f"{oti}<=626350, 0.35, "
        f"TRUE, 0.37)"
    )
    r += 1; _reg("marginal", _row(ws, r, "Marginal rate",
        formula=marg_f, fmt=PCT,
        note="The rate that applies to your next dollar of ordinary income."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 6: CAPITAL GAINS TAX  (2025 single: 0% ≤ 48,350 / 15% ≤ 533,400 / 20%)
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑥ CAPITAL GAINS TAX  (2025 Single — 0/15/20% stacked on ordinary TI)", 4)

    zroom_f = f"=MAX(0, 48350-{REF['ord_ti']})"
    r += 1; _reg("zero_room", _row(ws, r, "Room remaining in 0% bracket",
        formula=zroom_f, fmt=USD,
        note="0% rate applies up to $48,350 of total TI (single 2025). Ordinary income fills this bucket first."))

    zamount_f = f"=MIN({REF['preferential']}, {REF['zero_room']})"
    r += 1; _reg("zero_amount", _row(ws, r, "Amount taxed at 0%",
        formula=zamount_f, fmt=USD,
        note="The slice of preferential income that fits in the remaining 0% room."))

    cg_rem1_f = f"=MAX(0, {REF['preferential']}-{REF['zero_amount']})"
    r += 1; _reg("cg_rem1", _row(ws, r, "Preferential income above 0% threshold",
        formula=cg_rem1_f, fmt=USD,
        note="Spills into 15% or 20% bracket."))

    froom_f = f"=MAX(0, 533400-MAX({REF['ord_ti']}, 48350))"
    r += 1; _reg("fifteen_room", _row(ws, r, "Room in 15% bracket",
        formula=froom_f, fmt=USD,
        note="15% rate applies from $48,350 to $533,400 of total TI (single 2025)."))

    famount_f = f"=MIN({REF['cg_rem1']}, {REF['fifteen_room']})"
    r += 1; _reg("fifteen_amount", _row(ws, r, "Amount taxed at 15%",
        formula=famount_f, fmt=USD))

    twenty_f = f"=MAX(0, {REF['cg_rem1']}-{REF['fifteen_amount']})"
    r += 1; _reg("twenty_amount", _row(ws, r, "Amount taxed at 20%",
        formula=twenty_f, fmt=USD,
        note="Any preferential income above $533,400 total TI."))

    cg_tax_f = f"={REF['fifteen_amount']}*0.15+{REF['twenty_amount']}*0.20"
    r += 1; _reg("cg_tax", _row(ws, r, "Capital gains tax",
        formula=cg_tax_f, fmt=USD,
        note="0% amount is free — no tax on that slice."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 7: SELF-EMPLOYMENT TAX
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑦ SELF-EMPLOYMENT TAX  (if applicable)", 4)

    se_base_f = f"=IF({REF['se_net']}<=0, 0, {REF['se_net']}*0.9235)"
    r += 1; _reg("se_base", _row(ws, r, "SE net earnings (× 0.9235 factor)",
        formula=se_base_f, fmt=USD,
        note="0.9235 = 1 − 0.0765 adjusts for the employer-half SS/Medicare that employees don't pay. See How It Works."))

    ss_room_f = f"=MAX(0, 176100-{REF['proj_wages']}-{REF['sp_wages']})"
    r += 1; _reg("ss_room", _row(ws, r, "Remaining SS wage base room (2025: $176,100)",
        formula=ss_room_f, fmt=USD,
        note="SS tax only applies to earnings up to $176,100. If wages already exceed this, SE income has no SS portion."))

    ss_taxable_f = f"=MIN({REF['se_base']}, {REF['ss_room']})"
    r += 1; _reg("ss_taxable_se", _row(ws, r, "SE net earnings subject to SS (12.4%)",
        formula=ss_taxable_f, fmt=USD))

    se_tax_f = (f"=IF({REF['se_net']}<=0, 0, "
                f"{REF['ss_taxable_se']}*0.124 + {REF['se_base']}*0.029)")
    r += 1; _reg("se_tax", _row(ws, r, "Self-employment tax (SS 12.4% + Medicare 2.9%)",
        formula=se_tax_f, fmt=USD,
        note="Both halves (employee + employer equivalent) are owed by self-employed filers."))

    half_se_exact_f = f"={REF['se_tax']}/2"
    r += 1; _reg("half_se_exact", _row(ws, r, "One-half SE tax (exact above-the-line deduction)",
        formula=half_se_exact_f, fmt=USD,
        note="Deduct this amount above-the-line. Compare to the approx. in Section ③ — if different, update AGI manually."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 8: ADDITIONAL MEDICARE TAX & NIIT
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑧ ADDITIONAL MEDICARE TAX & NIIT", 4)

    amt_base_f = f"={REF['proj_wages']}+{REF['sp_wages']}+{REF['se_base']}"
    r += 1; _reg("amt_base", _row(ws, r, "Medicare wages (wages + spouse wages + SE base)",
        formula=amt_base_f, fmt=USD,
        note="The base for Additional Medicare Tax. Your employer only sees their own wages — they miss spouse or SE income."))

    amt_f = f"=MAX(0, {REF['amt_base']}-200000)*0.009"
    r += 1; _reg("add_medicare", _row(ws, r, "Additional Medicare Tax (0.9% over $200K single)",
        formula=amt_f, fmt=USD,
        note="Threshold: $200K single | $250K MFJ | $125K MFS. No SS component — Medicare only."))

    inv_income_f = (f"=MAX(0, {REF['interest']}+{REF['ord_div']}"
                    f"+{REF['stcg']}+{REF['ltcg']})")
    r += 1; _reg("inv_income", _row(ws, r, "Net investment income",
        formula=inv_income_f, fmt=USD,
        note="Interest + dividends + capital gains (short & long). Does NOT include wages or SE income."))

    niit_f = f"=MIN({REF['inv_income']}, MAX(0,{REF['agi']}-200000))*0.038"
    r += 1; _reg("niit", _row(ws, r, "Net Investment Income Tax (3.8% over $200K single)",
        formula=niit_f, fmt=USD,
        note="= 3.8% × MIN(net investment income, MAX(0, AGI − $200K)). Zero if AGI ≤ $200K."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 9: TOTAL LIABILITY
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑨ TAX LIABILITY", 4)

    inc_tax_f = f"={REF['ord_tax']}+{REF['cg_tax']}"
    r += 1; _reg("inc_tax", _row(ws, r, "Income tax before credits",
        formula=inc_tax_f, fmt=USD,
        note="Ordinary income tax + capital gains tax"))

    tax_bc_f = f"={REF['inc_tax']}+{REF['se_tax']}+{REF['add_medicare']}+{REF['niit']}"
    r += 1; _reg("tax_before_credits", _row(ws, r, "Tax before credits",
        formula=tax_bc_f, fmt=USD,
        note="Income tax + SE tax + Additional Medicare Tax + NIIT"))

    nr_credits_f = f"=MIN({REF['ctc']}+{REF['nr_cred']}, {REF['tax_before_credits']})"
    r += 1; _reg("nr_credits", _row(ws, r, "Nonrefundable credits applied",
        formula=nr_credits_f, fmt=USD,
        note="Capped at tax before credits — cannot go below $0. Refundable credits applied separately below."))

    total_liability_f = (f"=MAX(0, {REF['tax_before_credits']}"
                         f"-{REF['nr_credits']}-{REF['ref_cred']})")
    r += 1; _reg("total_liability", _row(ws, r, "TOTAL TAX LIABILITY",
        formula=total_liability_f, fmt=USD, fill=RES, bold_label=True,
        note="= Tax before credits − nonrefundable credits − refundable credits. Cannot go below $0."))
    ws.cell(r, 4).number_format = USD

    eff_f = f"=IF({REF['agi']}>0, {REF['total_liability']}/{REF['agi']}, 0)"
    r += 1; _reg("eff_rate", _row(ws, r, "Effective tax rate",
        formula=eff_f, fmt=PCT,
        note="Total liability ÷ AGI. Different from marginal rate — shows average tax burden."))

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 10: WITHHOLDING RECOMMENDATION
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑩ WITHHOLDING RECOMMENDATION", 4)

    other_pay_f = f"={REF['est_pay']}+{REF['other_wh']}+{REF['sp_wh']}"
    r += 1; _reg("other_pay", _row(ws, r, "Other payments (estimated tax + withholding)",
        formula=other_pay_f, fmt=USD,
        note="Quarterly estimated tax payments + any other withholding sources + spouse withholding"))

    already_secured_f = f"={REF['ytd_wh_used']}+{REF['other_pay']}"
    r += 1; _reg("already_secured", _row(ws, r, "Already secured toward liability",
        formula=already_secured_f, fmt=USD,
        note="YTD withholding + all other payments. These are locked in — only future paychecks are adjustable."))

    others_future_f = "=0"
    r += 1; _reg("others_future", _row(ws, r, "Future withholding from OTHER jobs",
        formula=others_future_f, fmt=USD,
        note="If you have a second job, enter its expected remaining withholding here so we don't double-count."))

    req_remaining_f = (f"=MAX(0, ({REF['total_liability']}+{REF['target_ref']})"
                       f"-{REF['already_secured']}-{REF['others_future']})")
    r += 1; _reg("req_remaining", _row(ws, r, "Still needed from remaining paychecks",
        formula=req_remaining_f, fmt=USD,
        note="= (Liability + target refund) − already secured − other future. This is the total gap."))

    rec_pp_f = (f"=IF({REF['remaining']}>0, "
                f"{REF['req_remaining']}/{REF['remaining']}, 0)")
    r += 1; _reg("rec_pp", _row(ws, r, "Recommended withholding per paycheck",
        formula=rec_pp_f, fmt=USD, fill=RES, bold_label=True,
        note="= gap ÷ remaining periods. Set your W-4 so each check withholds at least this amount."))
    ws.cell(r, 4).number_format = USD

    add_pp_f = f"=MAX(0, {REF['rec_pp']}-{REF['wh_pp']})"
    r += 1; _reg("add_pp", _row(ws, r, "Extra per paycheck needed (W-4 line 4c)",
        formula=add_pp_f, fmt=USD, fill=RES, bold_label=True,
        note="= recommended − current withholding per period. Enter this on Form W-4 Step 4(c) as extra withholding."))
    ws.cell(r, 4).number_format = USD

    proj_balance_f = (f"={REF['proj_wh']}+{REF['other_pay']}-{REF['total_liability']}")
    r += 1; _reg("proj_balance", _row(ws, r, "Projected refund (+) / owe (−) at current rate",
        formula=proj_balance_f, fmt=USD, fill=RES,
        note="If you change NOTHING on your W-4. Positive = refund, negative = balance due at filing."))
    ws.cell(r, 4).number_format = USD

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 11: SAFE HARBOR
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1; _hdr(ws, r, "⑪ SAFE HARBOR  (minimum to avoid underpayment penalty)", 4)

    sh_current_f = f"={REF['total_liability']}*0.90"
    r += 1; _reg("sh_current", _row(ws, r, "90% of current-year liability",
        formula=sh_current_f, fmt=USD,
        note="Option A for safe harbor. Pay at least 90% of this year's tax to avoid penalty."))

    sh_prior_f = (f"=IF({REF['py_agi']}>150000, "
                  f"{REF['py_tax']}*1.10, {REF['py_tax']}*1.00)")
    r += 1; _reg("sh_prior", _row(ws, r, "100% (or 110%) of prior-year tax",
        formula=sh_prior_f, fmt=USD,
        note="Option B. If 2024 AGI >$150K, you need 110% of 2024 tax, not 100%. Prior-year safe harbor is often easier to hit."))

    sh_target_f = f"=MIN({REF['sh_current']}, {REF['sh_prior']})"
    r += 1; _reg("sh_target", _row(ws, r, "Safe-harbor target (lesser of the two)",
        formula=sh_target_f, fmt=USD,
        note="You only need to hit the LOWER of Option A or B. Usually Option B (prior year) is easier."))

    sh_add_pp_f = (f"=MAX(0, ({REF['sh_target']}-{REF['already_secured']}"
                   f"-{REF['others_future']})/{REF['remaining']})")
    r += 1; _reg("sh_add_pp", _row(ws, r, "Safe-harbor additional per paycheck",
        formula=sh_add_pp_f, fmt=USD, fill=RES,
        note="Minimum extra withholding per check to avoid underpayment penalty. May be less than full break-even amount."))
    ws.cell(r, 4).number_format = USD

    # ══════════════════════════════════════════════════════════════════════════
    # Footer
    # ══════════════════════════════════════════════════════════════════════════
    r += 1; _blank(ws, r)
    r += 1
    ws.row_dimensions[r].height = 40
    msg = ws.cell(r, 1,
        "COMPARISON: paste values from `twe estimate` (or the web UI) into column D on matching rows. "
        "Differences > $1 indicate a discrepancy to investigate. "
        "See the 'How It Works' tab for explanations of each calculation step.")
    msg.font = Font(size=9, italic=True, color="546E7A")
    msg.fill = WARN
    ws.merge_cells(f"A{r}:D{r}")
    msg.alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    # ── build the explanation sheet ───────────────────────────────────────────
    _build_explain_sheet(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    build()
