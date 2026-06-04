from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from collections import Counter
from .models import ExportResult
from .db import now
from .audit import append_audit_event, export_audit_log
from .review_engine import calculate_completion_status
from .template_engine import get_applicable_questions, load_template
from .dti_engine import load_dti_config, load_dti_inputs, compute_dti, block_lines, summarize_dti
from .cash_flow_engine import load_cash_flow_config, load_cash_flow_inputs, compute_cash_flow, source_lines, summarize_cash_flow
from .collateral_engine import (load_collateral_config, load_collateral_inputs, compute_collateral,
    collateral_lines, exposure_lines, summarize_collateral)
from .dscr_engine import (load_dscr_config, load_dscr_inputs, compute_dscr,
    cash_flow_lines as dscr_cash_flow_lines, debt_service_lines, loan_lines, summarize_dscr)
from .leverage_engine import (load_leverage_config, load_leverage_inputs, compute_leverage,
    input_lines as leverage_input_lines, summarize_leverage)
from .guarantor_engine import (load_guarantor_config, load_guarantor_inputs, compute_guarantor, summarize_guarantor)
from .global_engine import load_global_config, compute_global, summarize_global, carry_global
from .borrowing_base_engine import (load_borrowing_base_config, load_borrowing_base_inputs, compute_borrowing_base,
    collateral_lines as bb_collateral_lines, reserve_lines as bb_reserve_lines, line_lines as bb_line_lines,
    summarize_borrowing_base)
from .liquidity_engine import (load_liquidity_config, load_liquidity_inputs, compute_liquidity,
    input_lines as liquidity_input_lines, summarize_liquidity)

ROOT = Path(__file__).resolve().parents[1]

# --- Palette -----------------------------------------------------------------
NAVY   = "1F3A5F"   # primary brand
INK    = "13263D"   # darkest band
GOLD   = "C8A24B"   # accent rule
BAND   = "EEF2F8"   # zebra band
LINE   = "D7DEE8"   # hairline borders
WHITE  = "FFFFFF"
MUTED  = "5B6B7B"   # secondary text
# status / severity chips (fill, text)
GREEN_F, GREEN_T = "DCF1E5", "1E7F4F"
AMBER_F, AMBER_T = "FBE9C7", "8A5A00"
RED_F,   RED_T   = "F7D9D9", "9B1C1C"
SLATE_F, SLATE_T = "E7ECF3", "5B6B7B"

FONT = "Calibri"
_HAIR = Side(style="thin", color=LINE)
BOTTOM = Border(bottom=_HAIR)

CHIP = {
    "Complete": (GREEN_F, GREEN_T), "Ready": (GREEN_F, GREEN_T), "QC Approved": (GREEN_F, GREEN_T),
    "Ready for QC": (GREEN_F, GREEN_T), "Pass": (GREEN_F, GREEN_T), "Attached": (GREEN_F, GREEN_T),
    "Waived": (GREEN_F, GREEN_T),
    "Warning": (AMBER_F, AMBER_T), "Needs Review": (AMBER_F, AMBER_T), "Needed": (AMBER_F, AMBER_T),
    "Exception": (RED_F, RED_T), "Finding": (RED_F, RED_T), "Blocked": (RED_F, RED_T), "Open": (RED_F, RED_T),
    "Incomplete": (SLATE_F, SLATE_T), "Not Started": (SLATE_F, SLATE_T), "Not Required": (SLATE_F, SLATE_T),
}

def _fill(color): return PatternFill("solid", fgColor=color)

ALL_MODULES = ["cash_flow", "dti", "collateral", "dscr", "guarantor", "global", "leverage", "borrowing_base", "liquidity"]

def _template_modules(template):
    """Calc tabs to include for a template. Empty/absent = all. 'global' pulls
    in its dscr + guarantor feeders."""
    mods = list(getattr(template, "modules", []) or [])
    if not mods:
        return set(ALL_MODULES)
    s = set(mods)
    if "global" in s:
        s |= {"dscr", "guarantor"}
    return s

# fills for conditional formatting (need start/end colors)
_CF_GREEN = PatternFill(start_color=GREEN_F, end_color=GREEN_F, fill_type="solid")
_CF_AMBER = PatternFill(start_color=AMBER_F, end_color=AMBER_F, fill_type="solid")
_CF_RED = PatternFill(start_color=RED_F, end_color=RED_F, fill_type="solid")

def _ctx(conn, review_case_id):
    row = conn.execute("""SELECT rc.*, e.review_period,e.template_id,e.reviewer_name,e.qc_reviewer_name,c.client_name,lr.* FROM review_cases rc JOIN engagements e ON rc.engagement_id=e.engagement_id JOIN clients c ON e.client_id=c.client_id JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?""", (review_case_id,)).fetchone()
    return {k: row[k] for k in row.keys()}

def assert_export_allowed(conn, review_case_id, loan_record, template, override_reason=None):
    status = calculate_completion_status(conn, review_case_id, loan_record, template)
    if status["export_ready"] or override_reason: return status
    raise ValueError("Export blocked: " + "; ".join(status["blockers"] or ["Review status must be Ready for QC or QC Approved"]))

# --- Styling helpers ---------------------------------------------------------
def _short_date(v):
    s = "" if v is None else str(v)
    return s[:10] if len(s) >= 10 and s[4] == "-" and "T" in s else s

def _table_header(ws, headers, row=1):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = _fill(NAVY)
        c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        c.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[row].height = 24

def _chip(cell):
    style = CHIP.get(str(cell.value or "").strip())
    if style:
        fill, text = style
        cell.fill = _fill(fill)
        cell.font = Font(name=FONT, color=text, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _finish_table(ws, header_row, ncols, chip_cols=(), widths=None, landscape=False):
    """Zebra-band, border, freeze, autofilter and lay out a simple table."""
    last = ws.max_row
    for r in range(header_row + 1, last + 1):
        band = _fill(BAND) if (r - header_row) % 2 == 0 else _fill(WHITE)
        for cidx in range(1, ncols + 1):
            c = ws.cell(row=r, column=cidx)
            c.fill = band
            c.border = BOTTOM
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="center", wrap_text=cidx not in chip_cols)
        for cidx in chip_cols:
            _chip(ws.cell(row=r, column=cidx))
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if last >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last}"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if landscape:
        ws.page_setup.orientation = "landscape"

# --- Cover -------------------------------------------------------------------
def _build_cover(ws, ctx, template, metrics, dti=None, cf=None, coll=None, dscr=None, lev=None, guar=None, glob=None, bb=None, liq=None):
    ws.sheet_view.showGridLines = False
    widths = [2.5, 22, 30, 4, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # gold top rule
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 5
    ws["A1"].fill = _fill(GOLD)
    # banner
    ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 40
    t = ws["A2"]; t.value = "LINESHEET BUILDER"
    t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=22)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 22
    s = ws["A3"]; s.value = "Audit-Ready Commercial Loan Linesheet"
    s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=11)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for addr in ("A2", "A3"):
        for col in range(1, 7):
            ws.cell(row=int(addr[1:]), column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:F4"); ws.row_dimensions[4].height = 5
    ws["A4"].fill = _fill(GOLD)

    # KPI strip (row 6-8)
    kpis = [
        ("OVERALL STATUS", ctx["status"], True),
        ("COMPLETION", f"{metrics['completion_pct']}%", False),
        ("FINDINGS", str(metrics["findings"]), False),
        ("EVIDENCE OPEN", str(metrics["evidence_open"]), False),
    ]
    ws.row_dimensions[6].height = 8
    cells = ["B", "C", "D", "E"]
    for (label, value, is_status), col in zip(kpis, cells):
        lab = ws[f"{col}7"]; lab.value = label
        lab.font = Font(name=FONT, color=MUTED, bold=True, size=8)
        lab.alignment = Alignment(horizontal="center")
        lab.fill = _fill(BAND); lab.border = Border(top=_HAIR, left=_HAIR, right=_HAIR)
        val = ws[f"{col}8"]; val.value = value
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.fill = _fill(BAND); val.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR)
        if is_status:
            style = CHIP.get(str(value).strip(), (NAVY, WHITE))
            val.fill = _fill(style[0]); val.font = Font(name=FONT, color=style[1], bold=True, size=13)
        else:
            val.font = Font(name=FONT, color=NAVY, bold=True, size=16)
    ws.row_dimensions[8].height = 30

    # Engagement detail card (rows 10+)
    ws["B10"] = "ENGAGEMENT"
    ws["B10"].font = Font(name=FONT, color=GOLD, bold=True, size=10)
    rows = [
        ("Client", ctx["client_name"]),
        ("Review period", ctx["review_period"]),
        ("Template", f"{template.template_name}  ·  v{template.version}"),
        ("Loan ID", ctx["loan_id"]),
        ("Borrower", ctx["borrower_name"]),
        ("Product type", ctx.get("product_type")),
        ("Reviewer", ctx["reviewer_name"]),
        ("QC reviewer", ctx["qc_reviewer_name"]),
        ("Generated", _short_date(now()) + "  " + now()[11:19]),
    ]
    r0 = 11
    for i, (label, value) in enumerate(rows):
        r = r0 + i
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name=FONT, color=MUTED, bold=True, size=10)
        lc.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        vc = ws.cell(row=r, column=3, value=value)
        vc.font = Font(name=FONT, color=INK, size=10)
        vc.alignment = Alignment(vertical="center")
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = BOTTOM

    next_row = r0 + len(rows) + 1

    def _card(title, pairs):
        nonlocal next_row
        ws.cell(row=next_row, column=2, value=title).font = Font(name=FONT, color=GOLD, bold=True, size=10)
        next_row += 1
        for label, value, chip in pairs:
            lc = ws.cell(row=next_row, column=2, value=label)
            lc.font = Font(name=FONT, color=MUTED, bold=True, size=10); lc.alignment = Alignment(vertical="center")
            ws.merge_cells(start_row=next_row, start_column=3, end_row=next_row, end_column=6)
            vc = ws.cell(row=next_row, column=3, value=value)
            if chip:
                st = CHIP.get(chip, (GREEN_F, GREEN_T))
                vc.fill = _fill(st[0]); vc.font = Font(name=FONT, color=st[1], bold=True, size=10)
                vc.alignment = Alignment(vertical="center", indent=1)
            else:
                vc.font = Font(name=FONT, color=INK, size=10); vc.alignment = Alignment(vertical="center")
            for col in range(2, 7): ws.cell(row=next_row, column=col).border = BOTTOM
            next_row += 1
        next_row += 1

    # Cash flow / income summary — carried from the Cash Flow module
    if cf and cf.get("qualifying_monthly"):
        pairs = [("Qualifying income (mo.)", f"${cf['qualifying_monthly']:,.0f}", None),
                 ("Qualifying income (yr.)", f"${cf['qualifying_annual']:,.0f}", None)]
        if cf.get("business_income_reference_monthly"):
            pairs.append(("Business income (ref.)", f"${cf['business_income_reference_monthly']:,.0f}/mo", None))
        _card("CASH FLOW / INCOME", pairs)

    # Collateral / LTV summary — carried from the Collateral module
    if coll and coll.get("total_exposure"):
        _card("COLLATERAL / LTV", [
            ("LTV", f"{coll['ltv']:.1f}%   (guideline ≤ {coll['max_ltv']:.0f}%)", None),
            ("Net collateral value", f"${coll['net_collateral_value']:,.0f}", None),
            ("Excess / (shortfall)", f"${coll['excess']:,.0f}", None),
            ("Collateral assessment", coll["assessment"], coll["severity"]),
        ])

    # Debt service coverage — carried from the DSCR module (commercial)
    if dscr and dscr.get("total_debt_service"):
        pairs = [("DSCR", f"{dscr['dscr']:.2f}x   (min {dscr['min_dscr']:.2f}x)", None)]
        if dscr.get("loan_amount"):
            pairs.append(("Debt yield", f"{dscr['debt_yield']:.1f}%   (min {dscr['min_debt_yield']:.0f}%)", None))
        pairs.append(("DSCR assessment", dscr["assessment"], dscr["severity"]))
        _card("DEBT SERVICE (DSCR)", pairs)

    # Guarantor — carried from the Guarantor module (commercial)
    if guar and guar.get("assessment") not in (None, "Inputs required"):
        _card("GUARANTOR", [
            ("Net worth", f"${guar['net_worth']:,.0f}", None),
            ("Liquid assets", f"${guar['liquid_assets']:,.0f}", None),
            ("Personal DSCR", f"{guar['personal_dscr']:.2f}x", None),
            ("Guarantor assessment", guar["assessment"], guar["severity"]),
        ])

    # Global cash flow — carried from the Global DSCR capstone (commercial)
    if glob and glob.get("global_debt_service"):
        _card("GLOBAL CASH FLOW", [
            ("Global CFADS", f"${glob['global_cfads']:,.0f}", None),
            ("Global debt service", f"${glob['global_debt_service']:,.0f}", None),
            ("Global DSCR", f"{glob['global_dscr']:.2f}x   (min {glob['min_global_dscr']:.2f}x)", None),
            ("Global assessment", glob["assessment"], glob["severity"]),
        ])

    # Borrowing base — carried from the Borrowing Base module (ABL / lines)
    if bb and bb.get("assessment") not in (None, "Inputs required"):
        _card("BORROWING BASE", [
            ("Borrowing base", f"${bb['borrowing_base']:,.0f}", None),
            ("Current outstanding", f"${bb['current_outstanding']:,.0f}", None),
            ("Net availability", f"${bb['net_availability']:,.0f}", None),
            ("Borrowing base assessment", bb["assessment"], bb["severity"]),
        ])

    # Liquidity & reserves — carried from the Liquidity module
    if liq and liq.get("assessment") not in (None, "Inputs required"):
        _card("LIQUIDITY & RESERVES", [
            ("Liquid assets", f"${liq['liquid_assets']:,.0f}", None),
            ("Months of reserves", f"{liq['months_reserves']:.1f}   (min {liq['min_months_reserves']:.0f})", None),
            ("Liquidity assessment", liq["assessment"], liq["severity"]),
        ])

    # Leverage & liquidity — carried from the Leverage module (commercial)
    if lev and lev.get("assessment") not in (None, "Inputs required"):
        _card("LEVERAGE & LIQUIDITY", [
            ("Current ratio", f"{lev['current_ratio']:.2f}", None),
            ("Debt-to-worth", f"{lev['debt_to_worth']:.2f}x", None),
            ("Debt-to-EBITDA", f"{lev['debt_to_ebitda']:.2f}x", None),
            ("Leverage assessment", lev["assessment"], lev["severity"]),
        ])

    # Ability-to-Repay summary — carried from the DTI module (consumer reviews)
    if dti and dti.get("total_income"):
        ws.cell(row=next_row, column=2, value="ABILITY-TO-REPAY (DTI)").font = Font(name=FONT, color=GOLD, bold=True, size=10)
        next_row += 1
        _res = dti.get('net_residual_income') if dti.get('total_withholding') else dti['residual_income']
        atr = [
            ("Monthly gross income", f"${dti['total_income']:,.0f}"),
            ("Back-end DTI", f"{dti['back_end_dti']:.1f}%   (target ≤ {dti['back_end_target']:.0f}%, max {dti['back_end_max']:.0f}%)"),
            ("Front-end DTI", f"{dti['front_end_dti']:.1f}%   (target ≤ {dti['front_end_target']:.0f}%)"),
            ("Residual income (net)" if dti.get("total_withholding") else "Residual income", f"${_res:,.0f}"),
        ]
        for label, value in atr:
            lc = ws.cell(row=next_row, column=2, value=label)
            lc.font = Font(name=FONT, color=MUTED, bold=True, size=10); lc.alignment = Alignment(vertical="center")
            ws.merge_cells(start_row=next_row, start_column=3, end_row=next_row, end_column=6)
            vc = ws.cell(row=next_row, column=3, value=value)
            vc.font = Font(name=FONT, color=INK, size=10); vc.alignment = Alignment(vertical="center")
            for col in range(2, 7): ws.cell(row=next_row, column=col).border = BOTTOM
            next_row += 1
        # assessment chip
        ws.cell(row=next_row, column=2, value="ATR assessment").font = Font(name=FONT, color=MUTED, bold=True, size=10)
        ws.merge_cells(start_row=next_row, start_column=3, end_row=next_row, end_column=6)
        ac = ws.cell(row=next_row, column=3, value=dti["assessment"])
        chip = CHIP.get(dti["severity"] or "", (GREEN_F, GREEN_T))
        ac.fill = _fill(chip[0]); ac.font = Font(name=FONT, color=chip[1], bold=True, size=10)
        ac.alignment = Alignment(vertical="center", indent=1)
        next_row += 2

    foot = next_row
    ws.merge_cells(start_row=foot, start_column=2, end_row=foot, end_column=6)
    fc = ws.cell(row=foot, column=2, value="Confidential — prepared for internal credit review. Generated by Linesheet Builder.")
    fc.font = Font(name=FONT, color=MUTED, italic=True, size=8)
    ws.page_setup.orientation = "portrait"

# --- Ability-to-Repay (DTI) tab ----------------------------------------------
def _build_dti(ws, ctx, cfg, values, result, income_ref=None):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34

    # banner
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:C2"); ws.row_dimensions[2].height = 30
    t = ws["A2"]; t.value = "ABILITY-TO-REPAY  ·  DEBT-TO-INCOME WORKSHEET"
    t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("A3:C3"); ws.row_dimensions[3].height = 18
    s = ws["A3"]; s.value = f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  enter monthly amounts; ratios calculate automatically"
    s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in range(1, 4):
        ws.cell(row=2, column=col).fill = _fill(NAVY); ws.cell(row=3, column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:C4"); ws.row_dimensions[4].height = 5; ws["A4"].fill = _fill(GOLD)

    # column captions
    for col, cap in ((1, "Line item"), (2, "Monthly $"), (3, "Notes / source")):
        c = ws.cell(row=5, column=col, value=cap)
        c.font = Font(name=FONT, color=MUTED, bold=True, size=9)
        c.border = BOTTOM
        if col == 2: c.alignment = Alignment(horizontal="right")

    money = '$#,##0'
    row = 6
    totals = {}  # block -> subtotal row

    def section_bar(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label.upper())
        c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
        c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
        ws.row_dimensions[row].height = 18
        row += 1

    def input_line(key, label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10)
        amt = ws.cell(row=row, column=2)
        v = values.get(key)
        if v not in (None, "", 0, 0.0):
            amt.value = float(v)
        amt.number_format = money
        amt.fill = _fill("FCFCFD"); amt.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        amt.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3).border = BOTTOM
        for col in (1, 3): ws.cell(row=row, column=col).border = BOTTOM
        row += 1

    def subtotal(label, first, last, formula=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c = ws.cell(row=row, column=2, value=formula or f"=SUM(B{first}:B{last})")
        c.number_format = money; c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal="right"); c.fill = _fill(BAND)
        ws.cell(row=row, column=1).fill = _fill(BAND); ws.cell(row=row, column=3).fill = _fill(BAND)
        r = row; row += 2
        return r

    blocks = [("income", cfg["income"]["section_name"]),
              ("housing", cfg["housing"]["section_name"]),
              ("debts", cfg["debts"]["section_name"])]
    if cfg.get("deductions"):
        blocks.append(("deductions", cfg["deductions"]["section_name"]))
    for block, label in blocks:
        section_bar(label)
        first = row
        for key, lbl in block_lines(cfg, block):
            input_line(key, lbl)
        if block == "income" and income_ref:
            totals[block] = subtotal("Qualifying income (auto-fed from Cash Flow Analysis)", first, row - 1, formula=f"={income_ref}")
        else:
            totals[block] = subtotal(f"Total — {label}", first, row - 1)

    inc, hou, deb = totals["income"], totals["housing"], totals["debts"]
    ded = totals.get("deductions")
    th = result

    section_bar("Ability-to-Repay Results")
    def result_row(label, formula, fmt, bold=True):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=bold, color=INK)
        c = ws.cell(row=row, column=2, value=formula)
        c.number_format = fmt; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1
        return r

    oblig_row = result_row("Total monthly obligations (housing + debts)", f"=B{hou}+B{deb}", money)
    front_row = result_row("Front-end DTI  (housing ÷ income)", f'=IF(B{inc}=0,"",B{hou}/B{inc})', '0.0%')
    back_row = result_row("Back-end DTI  (obligations ÷ income)", f'=IF(B{inc}=0,"",(B{hou}+B{deb})/B{inc})', '0.0%')
    gross_res_label = "Monthly residual income (gross − obligations)" if ded else "Monthly residual income (income − obligations)"
    res_row = result_row(gross_res_label, f"=B{inc}-B{hou}-B{deb}", money)
    net_res_row = None
    if ded:
        result_row("Net monthly income (gross − payroll deductions)", f"=B{inc}-B{ded}", money)
        net_res_row = result_row("Net residual income (net of withholding)", f"=B{inc}-B{ded}-B{hou}-B{deb}", money)
    row += 1

    # guideline thresholds (static reference)
    fe, bt, bm, rmin = th["front_end_target"], th["back_end_target"], th["back_end_max"], th["residual_income_min"]
    for label, val in (("Guideline — front-end target", f"{fe:.0f}%"),
                       ("Guideline — back-end target", f"{bt:.0f}%"),
                       ("Guideline — back-end maximum", f"{bm:.0f}%"),
                       ("Guideline — minimum residual income", f"${rmin:,.0f}" if rmin else "n/a")):
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9, color=MUTED)
        gc = ws.cell(row=row, column=2, value=val); gc.font = Font(name=FONT, size=9, color=MUTED)
        gc.alignment = Alignment(horizontal="right")
        row += 1
    row += 1

    # assessment banner (live formula + conditional colour)
    ws.cell(row=row, column=1, value="ATR ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    a = ws.cell(row=row, column=2)
    a.value = (f'=IF(B{inc}=0,"Enter income to assess",'
               f'IF(B{back_row}>{bm/100},"Fails ATR — exceeds maximum DTI",'
               f'IF(OR(B{back_row}>{bt/100},B{front_row}>{fe/100}),'
               f'"Exceeds guidelines — documented exception required",'
               f'"Within ability-to-repay guidelines")))')
    a.font = Font(name=FONT, bold=True, size=11)
    a.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22
    assess_cell = f"B{row}"

    # --- conditional formatting (live colour as the user types) ---
    bmf, btf, fef = bm / 100, bt / 100, fe / 100
    cf = ws.conditional_formatting
    cf.add(f"B{back_row}", CellIsRule(operator="greaterThan", formula=[str(bmf)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{back_row}", CellIsRule(operator="greaterThan", formula=[str(btf)], fill=_CF_AMBER, stopIfTrue=True))
    cf.add(f"B{back_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"B{front_row}", CellIsRule(operator="greaterThan", formula=[str(fef)], fill=_CF_AMBER, stopIfTrue=True))
    cf.add(f"B{front_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    for rr in (res_row, net_res_row):
        if not rr:
            continue
        cf.add(f"B{rr}", CellIsRule(operator="lessThan", formula=["0"], fill=_CF_RED, stopIfTrue=True))
        if rmin:
            cf.add(f"B{rr}", CellIsRule(operator="lessThan", formula=[str(rmin)], fill=_CF_AMBER, stopIfTrue=True))
        cf.add(f"B{rr}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Fails",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Exceeds",{assess_cell}))'], fill=_CF_AMBER, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Within",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))

    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# --- Cash Flow / Income Analysis tab -----------------------------------------
def _build_cash_flow(ws, ctx, cfg, values, result):
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 40), ("B", 13), ("C", 13), ("D", 11), ("E", 12), ("F", 14)):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 30
    t = ws["A2"]; t.value = "CASH FLOW  ·  INCOME ANALYSIS"
    t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 18
    s = ws["A3"]; s.value = f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  gross / pre-tax; enter up to two periods, pick basis & method"
    s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in range(1, 7):
        ws.cell(row=2, column=col).fill = _fill(NAVY); ws.cell(row=3, column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:F4"); ws.row_dimensions[4].height = 5; ws["A4"].fill = _fill(GOLD)

    _table_header(ws, ["Income source", "Period 1", "Period 2", "Basis", "Method", "Monthly $"], row=5)
    for col in (2, 3, 6):
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    money = '$#,##0'
    db, dm = cfg["default_basis"], cfg["default_method"]
    dv_basis = DataValidation(type="list", formula1='"%s"' % ",".join(cfg["bases"]), allow_blank=True)
    dv_method = DataValidation(type="list", formula1='"%s"' % ",".join(cfg["methods"]), allow_blank=True)
    ws.add_data_validation(dv_basis); ws.add_data_validation(dv_method)

    row = 6
    qualifying_cells, reference_cells = [], []
    last_section = None
    for section, key, label, role in source_lines(cfg):
        if section != last_section:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1, value=section.upper())
            c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=9.5)
            c.alignment = Alignment(vertical="center", indent=1)
            for col in range(1, 7): ws.cell(row=row, column=col).fill = _fill(INK)
            ws.row_dimensions[row].height = 16
            last_section = section; row += 1
        v = values.get(key) or {}
        ref = role == "reference"
        lab = ws.cell(row=row, column=1, value=label)
        lab.font = Font(name=FONT, size=9.5, italic=ref, color=MUTED if ref else INK)
        for col, field in ((2, "period1"), (3, "period2")):
            cell = ws.cell(row=row, column=col)
            val = v.get(field)
            if val not in (None, "", 0, 0.0): cell.value = float(val)
            cell.number_format = money; cell.alignment = Alignment(horizontal="right")
            cell.fill = _fill("FCFCFD"); cell.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        bcell = ws.cell(row=row, column=4, value=v.get("basis") or db)
        mcell = ws.cell(row=row, column=5, value=v.get("method") or dm)
        for cc in (bcell, mcell):
            cc.font = Font(name=FONT, size=9); cc.alignment = Alignment(horizontal="center")
            cc.fill = _fill(BAND); cc.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        dv_basis.add(bcell); dv_method.add(mcell)
        mon = ws.cell(row=row, column=6)
        mon.value = (f'=ROUND(SWITCH($E{row},'
                     f'"Latest",IF(C{row}<>"",C{row},N(B{row})),'
                     f'"Lower of",IF(AND(B{row}<>"",C{row}<>""),MIN(B{row},C{row}),N(B{row})+N(C{row})),'
                     f'IF(AND(B{row}="",C{row}=""),0,AVERAGE(B{row},C{row})))'
                     f'/IF($D{row}="Monthly",1,12),2)')
        mon.number_format = money; mon.alignment = Alignment(horizontal="right")
        mon.font = Font(name=FONT, size=9.5, italic=ref, color=MUTED if ref else NAVY)
        for col in range(1, 7): ws.cell(row=row, column=col).border = BOTTOM
        (reference_cells if ref else qualifying_cells).append(f"F{row}")
        row += 1

    row += 1
    def total_row(label, cells, big=False, muted=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10.5 if big else 9.5, bold=True, color=MUTED if muted else NAVY)
        c = ws.cell(row=row, column=6, value=("=" + ("+".join(cells))) if cells else 0)
        c.number_format = money; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=12 if big else 10, bold=True, color=MUTED if muted else NAVY)
        c.fill = _fill(BAND); ws.cell(row=row, column=1).fill = _fill(BAND)
        for col in range(2, 6): ws.cell(row=row, column=col).fill = _fill(BAND)
        r = row; row += 1
        return r
    qm_row = total_row("TOTAL QUALIFYING MONTHLY INCOME", qualifying_cells, big=True)
    ws.cell(row=row, column=1, value="Total qualifying ANNUAL income").font = Font(name=FONT, size=9.5, bold=True, color=NAVY)
    ac = ws.cell(row=row, column=6, value=f"=F{qm_row}*12"); ac.number_format = money
    ac.alignment = Alignment(horizontal="right"); ac.font = Font(name=FONT, size=10, bold=True, color=NAVY); row += 1
    if reference_cells:
        total_row("Pro-rata business income (reference only — not counted)", reference_cells, muted=True)

    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return {"qualifying_monthly": f"F{qm_row}"}


# --- Collateral & LTV tab ----------------------------------------------------
def _build_collateral(ws, ctx, cfg, values, result):
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 38), ("B", 16), ("C", 12), ("D", 16)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:D2"); ws.row_dimensions[2].height = 30
    t = ws["A2"]; t.value = "COLLATERAL  ·  LTV ANALYSIS"
    t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("A3:D3"); ws.row_dimensions[3].height = 18
    s = ws["A3"]; s.value = f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  net value applies advance rates; LTV is market-value basis"
    s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in range(1, 5):
        ws.cell(row=2, column=col).fill = _fill(NAVY); ws.cell(row=3, column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:D4"); ws.row_dimensions[4].height = 5; ws["A4"].fill = _fill(GOLD)

    _table_header(ws, ["Collateral type", "Market value", "Advance %", "Eligible value"], row=5)
    for col in (2, 3, 4):
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    money = '$#,##0'
    row = 6
    coll_first = row
    for key, label, default_ar in collateral_lines(cfg):
        v = values.get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        mv = ws.cell(row=row, column=2)
        if _numlike(v.get("market_value")): mv.value = float(v["market_value"])
        mv.number_format = money; mv.alignment = Alignment(horizontal="right")
        mv.fill = _fill("FCFCFD"); mv.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ar = ws.cell(row=row, column=3)
        ar.value = float(v["advance_rate"]) if v.get("advance_rate") not in (None, "") else default_ar
        ar.number_format = '0"%"'; ar.alignment = Alignment(horizontal="right")
        ar.fill = _fill("FCFCFD"); ar.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        el = ws.cell(row=row, column=4, value=f"=ROUND(B{row}*C{row}/100,2)")
        el.number_format = money; el.alignment = Alignment(horizontal="right")
        el.font = Font(name=FONT, size=9.5, color=NAVY)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        row += 1
    coll_last = row - 1
    # collateral subtotals
    ws.cell(row=row, column=1, value="Total market value / net (eligible) value").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    mt = ws.cell(row=row, column=2, value=f"=SUM(B{coll_first}:B{coll_last})"); mt.number_format = money
    mt.font = Font(name=FONT, size=10, bold=True, color=NAVY); mt.alignment = Alignment(horizontal="right"); mt.fill = _fill(BAND)
    nt = ws.cell(row=row, column=4, value=f"=SUM(D{coll_first}:D{coll_last})"); nt.number_format = money
    nt.font = Font(name=FONT, size=10, bold=True, color=NAVY); nt.alignment = Alignment(horizontal="right"); nt.fill = _fill(BAND)
    ws.cell(row=row, column=1).fill = _fill(BAND); ws.cell(row=row, column=3).fill = _fill(BAND)
    market_row = row; net_row = row; row += 2

    # exposure
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="EXPOSURE"); c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1
    exp_first = row
    for key, label in exposure_lines(cfg):
        v = values.get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        amt = ws.cell(row=row, column=2)
        if _numlike(v.get("market_value")): amt.value = float(v["market_value"])
        amt.number_format = money; amt.alignment = Alignment(horizontal="right")
        amt.fill = _fill("FCFCFD"); amt.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        row += 1
    exp_last = row - 1
    ws.cell(row=row, column=1, value="Total exposure").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    et = ws.cell(row=row, column=2, value=f"=SUM(B{exp_first}:B{exp_last})"); et.number_format = money
    et.font = Font(name=FONT, size=10, bold=True, color=NAVY); et.alignment = Alignment(horizontal="right"); et.fill = _fill(BAND)
    ws.cell(row=row, column=1).fill = _fill(BAND); ws.cell(row=row, column=3).fill = _fill(BAND); ws.cell(row=row, column=4).fill = _fill(BAND)
    exp_row = row; row += 2

    # results
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="COLLATERAL RESULTS"); c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1
    mx, mc = result["max_ltv"], result["min_coverage"]
    def res(label, formula, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        cc = ws.cell(row=row, column=4, value=formula); cc.number_format = fmt
        cc.alignment = Alignment(horizontal="right"); cc.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r
    ltv_row = res(f"LTV  (exposure ÷ market value)   guideline ≤ {mx:.0f}%", f'=IF(B{market_row}=0,"",B{exp_row}/B{market_row})', '0.0%')
    cov_row = res(f"Collateral coverage  (net value ÷ exposure)   min {mc:.0f}%", f'=IF(B{exp_row}=0,"",D{net_row}/B{exp_row})', '0.0%')
    exc_row = res("Excess / (shortfall) of net collateral", f"=D{net_row}-B{exp_row}", money)
    row += 1
    ws.cell(row=row, column=1, value="COLLATERAL ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    a = ws.cell(row=row, column=2)
    a.value = (f'=IF(B{exp_row}=0,"Enter exposure to assess",'
               f'IF(D{cov_row}<{mc/100},"Undersecured — collateral shortfall",'
               f'IF(D{ltv_row}>{mx/100},"Exceeds LTV guideline","Adequately secured")))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    assess_cell = f"B{row}"

    cf = ws.conditional_formatting
    cf.add(f"D{ltv_row}", CellIsRule(operator="greaterThan", formula=[str(mx/100)], fill=_CF_AMBER, stopIfTrue=True))
    cf.add(f"D{ltv_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"D{cov_row}", CellIsRule(operator="lessThan", formula=[str(mc/100)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"D{cov_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(mc/100)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"D{exc_row}", CellIsRule(operator="lessThan", formula=["0"], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"D{exc_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Undersecured",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Exceeds",{assess_cell}))'], fill=_CF_AMBER, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Adequately",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))

    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True


def _numlike(v):
    return v not in (None, "", 0, 0.0)


def _calc_tab_header(ws, ctx, title, subtitle):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 48; ws.column_dimensions["B"].width = 18; ws.column_dimensions["C"].width = 14
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:C2"); ws.row_dimensions[2].height = 30
    t = ws["A2"]; t.value = title; t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=14)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells("A3:C3"); ws.row_dimensions[3].height = 18
    s = ws["A3"]; s.value = subtitle; s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=9)
    s.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 4): ws.cell(row=2, column=col).fill = _fill(NAVY); ws.cell(row=3, column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:C4"); ws.row_dimensions[4].height = 5; ws["A4"].fill = _fill(GOLD)
    _table_header(ws, ["Line item", "Amount", ""], row=5)
    ws.cell(row=5, column=2).alignment = Alignment(horizontal="right", vertical="center")


def _build_dscr(ws, ctx, cfg, values, result):
    _calc_tab_header(ws, ctx, "DEBT SERVICE COVERAGE (DSCR)",
                     f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  CFADS ÷ annual debt service; debt yield = NOI ÷ loan")
    money = '$#,##0'; rowmap = {}; row = 6

    def sec(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label.upper()); c.fill = _fill(INK)
        c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
        ws.row_dimensions[row].height = 16; row += 1

    def inp(key, label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        b = ws.cell(row=row, column=2)
        if _numlike(values.get(key)): b.value = float(values[key])
        b.number_format = money; b.alignment = Alignment(horizontal="right")
        b.fill = _fill("FCFCFD"); b.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ws.cell(row=row, column=1).border = BOTTOM
        rowmap[key] = row; row += 1

    def total(label, formula, fmt=money):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c = ws.cell(row=row, column=2, value=formula); c.number_format = fmt; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY); c.fill = _fill(BAND); ws.cell(row=row, column=1).fill = _fill(BAND)
        r = row; row += 1; return r

    sec(cfg["cash_flow"]["section_name"])
    for k, label, sign in dscr_cash_flow_lines(cfg): inp(k, label)
    parts = "".join(f"{'+' if sign > 0 else '-'}B{rowmap[k]}" for k, _, sign in dscr_cash_flow_lines(cfg))
    cfads_row = total("Cash flow available for debt service (CFADS)", f"=ROUND({parts},2)")
    row += 1
    sec(cfg["debt_service"]["section_name"]); ds_first = row
    for k, label in debt_service_lines(cfg): inp(k, label)
    ds_row = total("Total annual debt service", f"=SUM(B{ds_first}:B{row-1})")
    row += 1
    if list(loan_lines(cfg)):
        sec(cfg["loan"]["section_name"])
        for k, label in loan_lines(cfg): inp(k, label)
        row += 1
    sec("DSCR Results")
    noi_row = rowmap.get("net_operating_income"); loan_row = rowmap.get("loan_amount")
    md, mdy = result["min_dscr"], result["min_debt_yield"]

    def res(label, formula, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        c = ws.cell(row=row, column=2, value=formula); c.number_format = fmt; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    dscr_row = res(f"DSCR  (CFADS ÷ debt service)   min {md:.2f}x", f'=IF(B{ds_row}=0,"",B{cfads_row}/B{ds_row})', '0.00"x"')
    dy_row = res(f"Debt yield  (NOI ÷ loan amount)   min {mdy:.0f}%",
                 (f'=IF(B{loan_row}=0,"",B{noi_row}/B{loan_row})' if loan_row else '=""'), '0.0%')
    exc_row = res("Excess / (shortfall) cash flow", f"=B{cfads_row}-B{ds_row}", money)
    row += 1
    ws.cell(row=row, column=1, value="DSCR ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    dy_clause = f'IF(AND(B{loan_row}>0,B{dy_row}<{mdy/100}),"Below debt yield guideline","Meets coverage guidelines")' if loan_row else '"Meets coverage guidelines"'
    a = ws.cell(row=row, column=1, value=f'=IF(B{ds_row}=0,"Enter debt service",IF(B{dscr_row}<{md},"Below DSCR guideline",{dy_clause}))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"

    cf = ws.conditional_formatting
    cf.add(f"B{dscr_row}", CellIsRule(operator="lessThan", formula=[str(md)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{dscr_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(md)], fill=_CF_GREEN, stopIfTrue=True))
    if loan_row:
        cf.add(f"B{dy_row}", CellIsRule(operator="lessThan", formula=[str(mdy/100)], fill=_CF_AMBER, stopIfTrue=True))
        cf.add(f"B{dy_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(mdy/100)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"B{exc_row}", CellIsRule(operator="lessThan", formula=["0"], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{exc_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Below",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Meets",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True
    return {"cfads": f"B{cfads_row}", "ds": f"B{ds_row}"}


def _build_guarantor(ws, ctx, cfg, values, result):
    from .guarantor_engine import position_lines, cash_flow_lines as g_cf_lines, debt_service_lines as g_ds_lines, contingent_lines
    _calc_tab_header(ws, ctx, "GUARANTOR / GLOBAL FINANCIAL",
                     f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  personal net worth, liquidity and cash flow")
    money = '$#,##0'; rowmap = {}; row = 6

    def sec(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label.upper()); c.fill = _fill(INK)
        c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
        ws.row_dimensions[row].height = 16; row += 1

    def inp(key, label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        b = ws.cell(row=row, column=2)
        if _numlike(values.get(key)): b.value = float(values[key])
        b.number_format = money; b.alignment = Alignment(horizontal="right")
        b.fill = _fill("FCFCFD"); b.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ws.cell(row=row, column=1).border = BOTTOM; rowmap[key] = row; row += 1

    def total(label, formula, fmt=money):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c = ws.cell(row=row, column=2, value=formula); c.number_format = fmt; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY); c.fill = _fill(BAND); ws.cell(row=row, column=1).fill = _fill(BAND)
        r = row; row += 1; return r

    sec(cfg["financial_position"]["section_name"])
    asset_rows, liab_rows, liquid_rows = [], [], []
    for key, label, typ, liq in position_lines(cfg):
        inp(key, label)
        (liab_rows if typ == "liability" else asset_rows).append(rowmap[key])
        if liq: liquid_rows.append(rowmap[key])
    nw_formula = "+".join(f"B{r}" for r in asset_rows) + ("-" + "-".join(f"B{r}" for r in liab_rows) if liab_rows else "")
    nw_row = total("Net worth (assets − liabilities)", f"={nw_formula}")
    row += 1
    sec(cfg["cash_flow"]["section_name"])
    for k, label, sign in g_cf_lines(cfg): inp(k, label)
    pcf_parts = "".join(f"{'+' if sign > 0 else '-'}B{rowmap[k]}" for k, _, sign in g_cf_lines(cfg))
    pcf_row = total("Personal cash flow available", f"=ROUND({pcf_parts},2)")
    row += 1
    sec(cfg["debt_service"]["section_name"]); ds_first = row
    for k, label in g_ds_lines(cfg): inp(k, label)
    pds_row = total("Total personal debt service", f"=SUM(B{ds_first}:B{row-1})")
    row += 1
    if list(contingent_lines(cfg)):
        sec(cfg["contingent"]["section_name"])
        for k, label in contingent_lines(cfg): inp(k, label)
        row += 1
    sec("Guarantor Results")
    mpd = result["min_personal_dscr"]

    def res(label, formula, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        c = ws.cell(row=row, column=2, value=formula); c.number_format = fmt; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    res("Net worth", f"=B{nw_row}", money)
    res("Liquid assets", "=" + "+".join(f"B{r}" for r in liquid_rows) if liquid_rows else "=0", money)
    pdscr_row = res(f"Personal DSCR  (cash flow ÷ debt service)   min {mpd:.2f}x", f'=IF(B{pds_row}=0,"",B{pcf_row}/B{pds_row})', '0.00"x"')
    row += 1
    ws.cell(row=row, column=1, value="GUARANTOR ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    a = ws.cell(row=row, column=1, value=(f'=IF(B{pds_row}>0,IF(B{pdscr_row}<{mpd},"Below personal coverage",'
                                          f'IF(B{nw_row}<=0,"Negative net worth","Adequate guarantor support")),'
                                          f'IF(B{nw_row}<=0,"Negative net worth","Adequate guarantor support"))'))
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"
    cf = ws.conditional_formatting
    cf.add(f"B{pdscr_row}", CellIsRule(operator="lessThan", formula=[str(mpd)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{pdscr_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(mpd)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"B{nw_row}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{nw_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Below",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Negative",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Adequate",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True
    return {"pcf": f"B{pcf_row}", "pds": f"B{pds_row}"}


def _build_global(ws, ctx, cfg, result, dscr_sheet, dscr_cells, guar_sheet, guar_cells):
    _calc_tab_header(ws, ctx, "GLOBAL CASH FLOW / GLOBAL DSCR",
                     f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  business CFADS + guarantor cash flow ÷ total debt service")
    money = '$#,##0'; row = 6
    dq = f"'{dscr_sheet}'"; gq = f"'{guar_sheet}'"
    min_g = result["min_global_dscr"]

    def sec(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label.upper()); c.fill = _fill(INK)
        c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
        ws.row_dimensions[row].height = 16; row += 1

    def line(label, formula, bold=False, band=False):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=bold, color=NAVY if bold else INK)
        c = ws.cell(row=row, column=2, value=formula); c.number_format = money; c.alignment = Alignment(horizontal="right")
        c.font = Font(name=FONT, size=11 if bold else 10, bold=bold, color=NAVY)
        if band:
            c.fill = _fill(BAND); ws.cell(row=row, column=1).fill = _fill(BAND)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    sec("Global Cash Flow Available")
    bc = line("Business CFADS (from DSCR)", f"={dq}!{dscr_cells['cfads']}")
    pc = line("Guarantor personal cash flow (from Guarantor)", f"={gq}!{guar_cells['pcf']}")
    gc = line("Global CFADS", f"=B{bc}+B{pc}", bold=True, band=True)
    row += 1
    sec("Global Debt Service")
    bds = line("Business debt service (from DSCR)", f"={dq}!{dscr_cells['ds']}")
    pds = line("Personal debt service (from Guarantor)", f"={gq}!{guar_cells['pds']}")
    gds = line("Global debt service", f"=B{bds}+B{pds}", bold=True, band=True)
    row += 1
    sec("Global Coverage")
    ws.cell(row=row, column=1, value=f"Global DSCR  (global CFADS ÷ global debt service)   min {min_g:.2f}x").font = Font(name=FONT, size=10, bold=True, color=INK)
    gd = ws.cell(row=row, column=2, value=f'=IF(B{gds}=0,"",B{gc}/B{gds})'); gd.number_format = '0.00"x"'
    gd.alignment = Alignment(horizontal="right"); gd.font = Font(name=FONT, size=12, bold=True, color=NAVY)
    for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
    gd_row = row; row += 2
    ws.cell(row=row, column=1, value="GLOBAL ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    a = ws.cell(row=row, column=1, value=f'=IF(B{gds}=0,"Enter debt service",IF(B{gd_row}<{min_g},"Below global DSCR guideline","Meets global coverage"))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"
    cf = ws.conditional_formatting
    cf.add(f"B{gd_row}", CellIsRule(operator="lessThan", formula=[str(min_g)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{gd_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(min_g)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Below",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Meets",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True


def _build_leverage(ws, ctx, cfg, values, result):
    _calc_tab_header(ws, ctx, "LEVERAGE & LIQUIDITY",
                     f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  balance-sheet spreads (gross)")
    money = '$#,##0'; rowmap = {}; row = 6
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=cfg["inputs"]["section_name"].upper()); c.fill = _fill(INK)
    c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
    ws.row_dimensions[row].height = 16; row += 1
    for key, label in leverage_input_lines(cfg):
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        b = ws.cell(row=row, column=2)
        if _numlike(values.get(key)): b.value = float(values[key])
        b.number_format = money; b.alignment = Alignment(horizontal="right")
        b.fill = _fill("FCFCFD"); b.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ws.cell(row=row, column=1).border = BOTTOM; rowmap[key] = row; row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value="LEVERAGE & LIQUIDITY RESULTS"); c.fill = _fill(INK)
    c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1
    R = rowmap; mcr, mdtw, mdte = result["min_current_ratio"], result["max_debt_to_worth"], result["max_debt_to_ebitda"]

    def res(label, formula, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        cc = ws.cell(row=row, column=2, value=formula); cc.number_format = fmt; cc.alignment = Alignment(horizontal="right")
        cc.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    cr_row = res(f"Current ratio  (CA ÷ CL)   min {mcr:.2f}", f'=IF(B{R["current_liabilities"]}=0,"",B{R["current_assets"]}/B{R["current_liabilities"]})', '0.00')
    res("Working capital  (CA − CL)", f'=B{R["current_assets"]}-B{R["current_liabilities"]}', money)
    dtw_row = res(f"Debt-to-worth  (TL ÷ TNW)   max {mdtw:.2f}x", f'=IF(B{R["tangible_net_worth"]}=0,"",B{R["total_liabilities"]}/B{R["tangible_net_worth"]})', '0.00"x"')
    dte_row = res(f"Debt-to-EBITDA  (debt ÷ EBITDA)   max {mdte:.2f}x", f'=IF(B{R["ebitda"]}=0,"",B{R["total_debt"]}/B{R["ebitda"]})', '0.00"x"')
    row += 1
    ws.cell(row=row, column=1, value="LEVERAGE ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    breach = (f'OR(AND(B{R["current_liabilities"]}>0,B{cr_row}<{mcr}),'
              f'AND(B{R["tangible_net_worth"]}>0,B{dtw_row}>{mdtw}),'
              f'AND(B{R["ebitda"]}>0,B{dte_row}>{mdte}))')
    total_inputs = "+".join(f"B{r}" for r in R.values())
    a = ws.cell(row=row, column=1, value=f'=IF({total_inputs}=0,"Enter inputs",IF({breach},"Exceeds leverage / liquidity guidelines","Within leverage guidelines"))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"

    cf = ws.conditional_formatting
    cf.add(f"B{cr_row}", CellIsRule(operator="lessThan", formula=[str(mcr)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{cr_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(mcr)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"B{dtw_row}", CellIsRule(operator="greaterThan", formula=[str(mdtw)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{dtw_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"B{dte_row}", CellIsRule(operator="greaterThan", formula=[str(mdte)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{dte_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Exceeds",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Within",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True


# --- Borrowing Base tab ------------------------------------------------------
def _build_borrowing_base(ws, ctx, cfg, values, result):
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 38), ("B", 16), ("C", 12), ("D", 16)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:D2"); ws.row_dimensions[2].height = 30
    t = ws["A2"]; t.value = "BORROWING BASE"
    t.fill = _fill(NAVY); t.font = Font(name=FONT, color=WHITE, bold=True, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells("A3:D3"); ws.row_dimensions[3].height = 18
    s = ws["A3"]; s.value = f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  eligible collateral × advance rate − reserves vs. line"
    s.fill = _fill(NAVY); s.font = Font(name=FONT, color="C9D6E5", italic=True, size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in range(1, 5):
        ws.cell(row=2, column=col).fill = _fill(NAVY); ws.cell(row=3, column=col).fill = _fill(NAVY)
    ws.merge_cells("A4:D4"); ws.row_dimensions[4].height = 5; ws["A4"].fill = _fill(GOLD)
    _table_header(ws, ["Eligible collateral", "Value", "Advance %", "Eligible"], row=5)
    for col in (2, 3, 4):
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    money = '$#,##0'; row = 6; coll_first = row
    for key, label, default_ar in bb_collateral_lines(cfg):
        v = values.get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        val = ws.cell(row=row, column=2)
        if _numlike(v.get("value")): val.value = float(v["value"])
        val.number_format = money; val.alignment = Alignment(horizontal="right")
        val.fill = _fill("FCFCFD"); val.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ar = ws.cell(row=row, column=3)
        ar.value = float(v["advance_rate"]) if v.get("advance_rate") not in (None, "") else default_ar
        ar.number_format = '0"%"'; ar.alignment = Alignment(horizontal="right")
        ar.fill = _fill("FCFCFD"); ar.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        el = ws.cell(row=row, column=4, value=f"=ROUND(B{row}*C{row}/100,2)")
        el.number_format = money; el.alignment = Alignment(horizontal="right"); el.font = Font(name=FONT, size=9.5, color=NAVY)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        row += 1
    coll_last = row - 1
    ws.cell(row=row, column=1, value="Gross availability").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    gross = ws.cell(row=row, column=4, value=f"=SUM(D{coll_first}:D{coll_last})"); gross.number_format = money
    gross.font = Font(name=FONT, size=10, bold=True, color=NAVY); gross.alignment = Alignment(horizontal="right"); gross.fill = _fill(BAND)
    for col in (1, 2, 3): ws.cell(row=row, column=col).fill = _fill(BAND)
    gross_row = row; row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="RESERVES"); c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1; res_first = row
    res_keys = [k for k, _ in bb_reserve_lines(cfg)]
    for key, label in bb_reserve_lines(cfg):
        v = values.get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        amt = ws.cell(row=row, column=2)
        if _numlike(v.get("value")): amt.value = float(v["value"])
        amt.number_format = money; amt.alignment = Alignment(horizontal="right")
        amt.fill = _fill("FCFCFD"); amt.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        row += 1
    res_sum = f"SUM(B{res_first}:B{row-1})" if res_keys else "0"
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="LINE OF CREDIT"); c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1; line_rows = {}
    for key, label in bb_line_lines(cfg):
        v = values.get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        amt = ws.cell(row=row, column=2)
        if _numlike(v.get("value")): amt.value = float(v["value"])
        amt.number_format = money; amt.alignment = Alignment(horizontal="right")
        amt.fill = _fill("FCFCFD"); amt.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        line_rows[key] = row; row += 1
    com = f"B{line_rows['line_commitment']}"; out = f"B{line_rows['current_outstanding']}"
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="AVAILABILITY"); c.fill = _fill(INK); c.font = Font(name=FONT, color=WHITE, bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1

    def res_row(label, formula):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        cc = ws.cell(row=row, column=4, value=formula); cc.number_format = money; cc.alignment = Alignment(horizontal="right")
        cc.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 5): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    bb_row = res_row("Borrowing base (gross − reserves)", f"=D{gross_row}-{res_sum}")
    avail_row = res_row("Net availability", f"=MIN(D{bb_row},IF({com}>0,{com},D{bb_row}))-{out}")
    over_row = res_row("Overadvance", f"=MAX({out}-D{bb_row},0)")
    row += 1
    ws.cell(row=row, column=1, value="BORROWING BASE ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    a = ws.cell(row=row, column=1, value=f'=IF((D{gross_row}+{com}+{out})=0,"Enter inputs",IF({out}>D{bb_row},"Overadvance — borrowing base shortfall","Within borrowing base"))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"
    cf = ws.conditional_formatting
    cf.add(f"D{avail_row}", CellIsRule(operator="lessThan", formula=["0"], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"D{avail_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(f"D{over_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"D{over_row}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Overadvance",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Within",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True


def _build_liquidity(ws, ctx, cfg, values, result):
    _calc_tab_header(ws, ctx, "LIQUIDITY & RESERVES",
                     f"{ctx.get('borrower_name','')}  ·  Loan {ctx.get('loan_id','')}  ·  liquid assets ÷ monthly obligations = months of reserves")
    money = '$#,##0'; rowmap = {}; row = 6
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=cfg["inputs"]["section_name"].upper()); c.fill = _fill(INK)
    c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
    ws.row_dimensions[row].height = 16; row += 1
    for key, label in liquidity_input_lines(cfg):
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9.5)
        b = ws.cell(row=row, column=2)
        if _numlike(values.get(key)): b.value = float(values[key])
        b.number_format = money; b.alignment = Alignment(horizontal="right")
        b.fill = _fill("FCFCFD"); b.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR, top=_HAIR)
        ws.cell(row=row, column=1).border = BOTTOM; rowmap[key] = row; row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value="LIQUIDITY RESULTS"); c.fill = _fill(INK)
    c.font = Font(name=FONT, color=WHITE, bold=True, size=10); c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 4): ws.cell(row=row, column=col).fill = _fill(INK)
    row += 1
    R = rowmap; mm = result["min_months_reserves"]
    liquid_formula = f"B{R['cash_equivalents']}+B{R['marketable_securities']}+B{R['other_liquid_assets']}"

    def res(label, formula, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True, color=INK)
        cc = ws.cell(row=row, column=2, value=formula); cc.number_format = fmt; cc.alignment = Alignment(horizontal="right")
        cc.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        for col in range(1, 4): ws.cell(row=row, column=col).border = BOTTOM
        r = row; row += 1; return r

    res("Total liquid assets", f"={liquid_formula}", money)
    mo = R["monthly_obligations"]
    mr_row = res(f"Months of reserves  (liquid ÷ monthly obligations)   min {mm:.0f}", f'=IF(B{mo}=0,"",({liquid_formula})/B{mo})', '0.0')
    row += 1
    ws.cell(row=row, column=1, value="LIQUIDITY ASSESSMENT").font = Font(name=FONT, color=GOLD, bold=True, size=11); row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    a = ws.cell(row=row, column=1, value=f'=IF(B{mo}=0,"Enter obligations",IF(B{mr_row}<{mm},"Below reserve guideline","Adequate reserves"))')
    a.font = Font(name=FONT, bold=True, size=11); a.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22; assess_cell = f"A{row}"
    cf = ws.conditional_formatting
    cf.add(f"B{mr_row}", CellIsRule(operator="lessThan", formula=[str(mm)], fill=_CF_RED, stopIfTrue=True))
    cf.add(f"B{mr_row}", CellIsRule(operator="greaterThanOrEqual", formula=[str(mm)], fill=_CF_GREEN, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Below",{assess_cell}))'], fill=_CF_RED, stopIfTrue=True))
    cf.add(assess_cell, FormulaRule(formula=[f'ISNUMBER(SEARCH("Adequate",{assess_cell}))'], fill=_CF_GREEN, stopIfTrue=True))
    ws.freeze_panes = "A6"; ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.sheet_properties.pageSetUpPr.fitToPage = True


# --- Main export -------------------------------------------------------------
def generate_excel_linesheet(conn, review_case_id: int, template, output_dir: str | Path = ROOT / "outputs" / "excel", generated_by="system", override_reason=None):
    ctx = _ctx(conn, review_case_id); loan = ctx.copy()
    assert_export_allowed(conn, review_case_id, loan, template, override_reason)
    carry_global(conn, review_case_id)  # refresh derived global finding before reading exceptions
    output_dir = Path(output_dir); output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"linesheet_{ctx['loan_id']}_{now().replace(':','')}.xlsx"

    applicable = get_applicable_questions(loan, template)
    ans = {r["question_id"]: r for r in conn.execute("SELECT * FROM review_answers WHERE review_case_id=?", (review_case_id,)).fetchall()}
    exceptions = conn.execute("SELECT * FROM exceptions WHERE review_case_id=?", (review_case_id,)).fetchall()
    completion = calculate_completion_status(conn, review_case_id, loan, template)
    evidence_open = sum(1 for _, q in applicable if (a := ans.get(q.question_id)) and a["evidence_required"] and a["evidence_status"] not in ("Attached", "Waived"))
    metrics = {"completion_pct": completion["completion_pct"], "findings": len(exceptions), "evidence_open": evidence_open}

    mods = _template_modules(template)
    # For each module: load inputs once, summarize once (returns the full result
    # or None when empty); reuse the summary as the result, computing again only
    # for the empty/blank-tab case.
    dti_cfg = load_dti_config(); dti_values = load_dti_inputs(conn, review_case_id)
    # auto-feed DTI income from Cash Flow only when that tab is part of this template
    dti_summary = summarize_dti(conn, review_case_id, dti_cfg, auto_income=("cash_flow" in mods))
    dti_result = dti_summary if dti_summary is not None else compute_dti(dti_values, dti_cfg)
    cf_cfg = load_cash_flow_config(); cf_values = load_cash_flow_inputs(conn, review_case_id)
    cf_summary = summarize_cash_flow(conn, review_case_id, cf_cfg)
    cf_result = cf_summary if cf_summary is not None else compute_cash_flow(cf_values, cf_cfg)
    col_cfg = load_collateral_config(); col_values = load_collateral_inputs(conn, review_case_id)
    col_summary = summarize_collateral(conn, review_case_id, col_cfg)
    col_result = col_summary if col_summary is not None else compute_collateral(col_values, col_cfg)
    dscr_cfg = load_dscr_config(); dscr_values = load_dscr_inputs(conn, review_case_id)
    dscr_summary = summarize_dscr(conn, review_case_id, dscr_cfg)
    dscr_result = dscr_summary if dscr_summary is not None else compute_dscr(dscr_values, dscr_cfg)
    lev_cfg = load_leverage_config(); lev_values = load_leverage_inputs(conn, review_case_id)
    lev_summary = summarize_leverage(conn, review_case_id, lev_cfg)
    lev_result = lev_summary if lev_summary is not None else compute_leverage(lev_values, lev_cfg)
    guar_cfg = load_guarantor_config(); guar_values = load_guarantor_inputs(conn, review_case_id)
    guar_summary = summarize_guarantor(conn, review_case_id, guar_cfg)
    guar_result = guar_summary if guar_summary is not None else compute_guarantor(guar_values, guar_cfg)
    global_cfg = load_global_config()
    global_result = compute_global(dscr_summary, guar_summary, global_cfg)
    global_summary = global_result if (dscr_summary or guar_summary) else None
    bb_cfg = load_borrowing_base_config(); bb_values = load_borrowing_base_inputs(conn, review_case_id)
    bb_summary = summarize_borrowing_base(conn, review_case_id, bb_cfg)
    bb_result = bb_summary if bb_summary is not None else compute_borrowing_base(bb_values, bb_cfg)
    liq_cfg = load_liquidity_config(); liq_values = load_liquidity_inputs(conn, review_case_id)
    liq_summary = summarize_liquidity(conn, review_case_id, liq_cfg)
    liq_result = liq_summary if liq_summary is not None else compute_liquidity(liq_values, liq_cfg)

    wb = Workbook()
    wb.properties.title = f"Linesheet — {ctx['loan_id']} {ctx['borrower_name']}"
    wb.properties.creator = "Linesheet Builder"
    wb.calculation.fullCalcOnLoad = True  # recompute live DTI formulas on open

    # Cover (only carry modules selected for this template)
    ws = wb.active; ws.title = "Cover"
    _build_cover(ws, ctx, template, metrics,
                 dti_summary if "dti" in mods else None,
                 cf_summary if "cash_flow" in mods else None,
                 col_summary if "collateral" in mods else None,
                 dscr_summary if "dscr" in mods else None,
                 lev_summary if "leverage" in mods else None,
                 guar_summary if "guarantor" in mods else None,
                 global_summary if "global" in mods else None,
                 bb_summary if "borrowing_base" in mods else None,
                 liq_summary if "liquidity" in mods else None)

    # Loan Summary
    ws = wb.create_sheet("Loan Summary")
    _table_header(ws, ["Field", "Value"])
    summary = [
        ("Loan ID", ctx.get("loan_id"), None), ("Borrower", ctx.get("borrower_name"), None),
        ("Product type", ctx.get("product_type"), None),
        ("Commitment amount", ctx.get("commitment_amount"), "$#,##0"),
        ("Outstanding balance", ctx.get("outstanding_balance"), "$#,##0"),
        ("Origination date", _short_date(ctx.get("origination_date")), None),
        ("Maturity date", _short_date(ctx.get("maturity_date")), None),
        ("Risk rating", ctx.get("risk_rating"), None), ("Officer", ctx.get("officer"), None),
        ("Collateral type", ctx.get("collateral_type"), None), ("Guarantor", ctx.get("guarantor_name"), None),
        ("DSCR", ctx.get("dscr"), '0.00"x"'), ("LTV", ctx.get("ltv"), '0"%"'),
        ("Validation status", ctx.get("validation_status"), None),
    ]
    for label, value, fmt in summary:
        ws.append([label, value])
        if fmt and isinstance(value, (int, float)):
            ws.cell(row=ws.max_row, column=2).number_format = fmt
    _finish_table(ws, 1, 2, chip_cols=(), widths=[24, 34])
    # bold the field labels; chip the validation status value
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
    _chip(ws.cell(row=ws.max_row, column=2))

    # Calculation worksheets (only those selected for this template)
    cf_cells = None
    if "cash_flow" in mods:
        cf_cells = _build_cash_flow(wb.create_sheet("Cash Flow Analysis"), ctx, cf_cfg, cf_values, cf_result)
    if "dti" in mods:
        # auto-feed: when Cash Flow is filled, DTI income comes from it
        income_ref = f"'Cash Flow Analysis'!{cf_cells['qualifying_monthly']}" if (cf_cells and cf_summary) else None
        _build_dti(wb.create_sheet("Ability-to-Repay (DTI)"), ctx, dti_cfg, dti_values, dti_result, income_ref=income_ref)
    if "collateral" in mods:
        _build_collateral(wb.create_sheet("Collateral & LTV"), ctx, col_cfg, col_values, col_result)
    dscr_cells = guar_cells = None
    if "dscr" in mods:
        dscr_cells = _build_dscr(wb.create_sheet("Debt Service (DSCR)"), ctx, dscr_cfg, dscr_values, dscr_result)
    if "guarantor" in mods:
        guar_cells = _build_guarantor(wb.create_sheet("Guarantor"), ctx, guar_cfg, guar_values, guar_result)
    if "global" in mods and dscr_cells and guar_cells:
        _build_global(wb.create_sheet("Global Cash Flow"), ctx, global_cfg, global_result, "Debt Service (DSCR)", dscr_cells, "Guarantor", guar_cells)
    if "leverage" in mods:
        _build_leverage(wb.create_sheet("Leverage & Liquidity"), ctx, lev_cfg, lev_values, lev_result)
    if "borrowing_base" in mods:
        _build_borrowing_base(wb.create_sheet("Borrowing Base"), ctx, bb_cfg, bb_values, bb_result)
    if "liquidity" in mods:
        _build_liquidity(wb.create_sheet("Liquidity & Reserves"), ctx, liq_cfg, liq_values, liq_result)

    # Linesheet Questions  (header MUST remain at row 1, A1 == "Section")
    ws = wb.create_sheet("Linesheet Questions")
    _table_header(ws, ["Section", "ID", "Question", "Answer", "Source Value", "Status", "Severity", "Reviewer Comment", "Evidence"])
    prev_section = None
    for sec, q in applicable:
        a = ans.get(q.question_id)
        section_label = sec.section_name if sec.section_name != prev_section else ""
        prev_section = sec.section_name
        ws.append([
            section_label, q.question_id, q.question_text,
            a["answer_value"] if a else "",
            _short_date(a["source_value"]) if a else (_short_date(ctx.get(q.source_field)) if q.source_field else ""),
            a["answer_status"] if a else "Incomplete",
            a["severity"] if a else "",
            a["reviewer_comment"] if a else "",
            a["evidence_status"] if a else "Not Required",
        ])
    _finish_table(ws, 1, 9, chip_cols=(6, 7, 9), widths=[20, 6, 46, 12, 16, 13, 13, 34, 12], landscape=True)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws.cell(row=r, column=2).font = Font(name=FONT, size=9, color=MUTED)

    # Exceptions & Findings
    ws = wb.create_sheet("Exceptions & Findings")
    _table_header(ws, ["ID", "Section", "Question", "Issue", "Severity", "Status", "Reviewer Comment", "Evidence"])
    if exceptions:
        for e in exceptions:
            ws.append([e["exception_id"], e["section_id"], e["question_id"], e["issue_text"], e["severity"], e["status"], e["reviewer_comment"], e["evidence_status"]])
    else:
        ws.append(["", "", "", "No findings or exceptions recorded for this loan.", "", "", "", ""])
    _finish_table(ws, 1, 8, chip_cols=(5, 6, 8), widths=[6, 22, 12, 44, 13, 11, 30, 12], landscape=True)

    # Evidence Checklist
    ws = wb.create_sheet("Evidence Checklist")
    _table_header(ws, ["Section", "Question", "Evidence Required", "Evidence Status", "Comment"])
    for sec, q in applicable:
        a = ans.get(q.question_id); req = bool(a and a["evidence_required"])
        ws.append([sec.section_name, q.question_text, "Yes" if req else "No", a["evidence_status"] if a else "Not Required", a["reviewer_comment"] if a else ""])
    _finish_table(ws, 1, 5, chip_cols=(4,), widths=[22, 46, 16, 16, 36], landscape=True)

    # Audit Summary
    ws = wb.create_sheet("Audit Summary")
    _table_header(ws, ["ID", "Timestamp", "User", "Action", "Entity", "Reason"])
    for a in conn.execute("SELECT audit_id,timestamp,user,action_type,entity_type,reason FROM audit_log WHERE review_case_id=? OR loan_id=? ORDER BY audit_id", (review_case_id, ctx["loan_id"])).fetchall():
        ws.append([a["audit_id"], str(a["timestamp"]).replace("T", "  "), a["user"], a["action_type"], a["entity_type"], a["reason"]])
    _finish_table(ws, 1, 6, chip_cols=(), widths=[6, 22, 14, 22, 16, 28], landscape=True)

    wb.save(path)
    conn.execute("INSERT INTO exports (engagement_id, review_case_id, export_type, file_path, generated_by, generated_at, export_status) VALUES (?, ?, ?, ?, ?, ?, ?)", (ctx['engagement_id'], review_case_id, "excel", str(path), generated_by, now(), "Generated")); conn.commit()
    append_audit_event(conn, generated_by, "export_generated", "export", "excel", after_value=str(path), reason=override_reason, engagement_id=ctx['engagement_id'], review_case_id=review_case_id, loan_id=ctx['loan_id'], template_id=template.template_id, template_version=template.version)
    return ExportResult(export_type="excel", file_path=str(path), export_status="Generated")

def generate_data_mart_csv(conn, review_case_id: int, template, output_path: str | Path = ROOT / "outputs" / "data_mart" / "review_answers_export.csv", generated_by="system"):
    ctx=_ctx(conn, review_case_id); rows=[]
    def _row(qid, section, value, status, severity, comment):
        return {"client_name":ctx['client_name'],"review_period":ctx['review_period'],"template_id":template.template_id,"template_version":template.version,"review_case_id":review_case_id,"loan_id":ctx['loan_id'],"borrower_name":ctx['borrower_name'],"question_id":qid,"section":section,"answer_value":value,"status":status,"severity":severity,"exception_flag":bool(severity),"reviewer_comment":comment,"evidence_status":"","answered_by":generated_by,"answered_at":now(),"exported_at":now()}
    for a in conn.execute("SELECT * FROM review_answers WHERE review_case_id=?", (review_case_id,)).fetchall():
        r=_row(a['question_id'],a['section_id'],a['answer_value'],a['answer_status'],a['severity'],a['reviewer_comment'])
        r["evidence_status"]=a['evidence_status']; r["answered_by"]=a['answered_by']; r["answered_at"]=a['answered_at']
        rows.append(r)
    # Carry the cash-flow / income results into the data mart
    cf = summarize_cash_flow(conn, review_case_id)
    if cf:
        for qid, value in (("CF_QUALIFYING_MONTHLY", cf["qualifying_monthly"]),
                           ("CF_QUALIFYING_ANNUAL", cf["qualifying_annual"]),
                           ("CF_BUSINESS_INCOME_REF_MONTHLY", cf["business_income_reference_monthly"])):
            rows.append(_row(qid, "cash_flow", value, "Computed", None, "Carried from Cash Flow worksheet"))
    # Carry the collateral / LTV results into the data mart
    col = summarize_collateral(conn, review_case_id)
    if col:
        for qid, value in (("COLL_LTV_PCT", col["ltv"]), ("COLL_COVERAGE_PCT", col["coverage"]),
                           ("COLL_NET_VALUE", col["net_collateral_value"]), ("COLL_EXCESS", col["excess"]),
                           ("COLL_ASSESSMENT", col["assessment"])):
            rows.append(_row(qid, "collateral_ltv", value, col["assessment"], col["severity"], "Carried from Collateral worksheet"))
    # Carry the debt service coverage results into the data mart (commercial)
    dscr = summarize_dscr(conn, review_case_id)
    if dscr:
        for qid, value in (("DSCR_RATIO", dscr["dscr"]), ("DSCR_DEBT_YIELD_PCT", dscr["debt_yield"]),
                           ("DSCR_CFADS", dscr["cfads"]), ("DSCR_EXCESS_CASH_FLOW", dscr["excess_cash_flow"]),
                           ("DSCR_ASSESSMENT", dscr["assessment"])):
            rows.append(_row(qid, "debt_service", value, dscr["assessment"], dscr["severity"], "Carried from DSCR worksheet"))
    # Carry the leverage / liquidity results into the data mart (commercial)
    lev = summarize_leverage(conn, review_case_id)
    if lev:
        for qid, value in (("LEV_CURRENT_RATIO", lev["current_ratio"]), ("LEV_DEBT_TO_WORTH", lev["debt_to_worth"]),
                           ("LEV_DEBT_TO_EBITDA", lev["debt_to_ebitda"]), ("LEV_WORKING_CAPITAL", lev["working_capital"]),
                           ("LEV_ASSESSMENT", lev["assessment"])):
            rows.append(_row(qid, "leverage", value, lev["assessment"], lev["severity"], "Carried from Leverage worksheet"))
    # Carry guarantor results into the data mart (commercial)
    guar = summarize_guarantor(conn, review_case_id)
    if guar:
        for qid, value in (("GUAR_NET_WORTH", guar["net_worth"]), ("GUAR_LIQUID_ASSETS", guar["liquid_assets"]),
                           ("GUAR_PERSONAL_DSCR", guar["personal_dscr"]), ("GUAR_ASSESSMENT", guar["assessment"])):
            rows.append(_row(qid, "guarantor", value, guar["assessment"], guar["severity"], "Carried from Guarantor worksheet"))
    # Carry global cash flow / global DSCR into the data mart (commercial capstone)
    glob = summarize_global(conn, review_case_id)
    if glob:
        for qid, value in (("GLOBAL_CFADS", glob["global_cfads"]), ("GLOBAL_DEBT_SERVICE", glob["global_debt_service"]),
                           ("GLOBAL_DSCR_RATIO", glob["global_dscr"]), ("GLOBAL_ASSESSMENT", glob["assessment"])):
            rows.append(_row(qid, "global_cash_flow", value, glob["assessment"], glob["severity"], "Carried from Global worksheet"))
    # Carry borrowing base results into the data mart (ABL / lines)
    bb = summarize_borrowing_base(conn, review_case_id)
    if bb:
        for qid, value in (("BB_BORROWING_BASE", bb["borrowing_base"]), ("BB_NET_AVAILABILITY", bb["net_availability"]),
                           ("BB_OVERADVANCE", bb["overadvance"]), ("BB_ASSESSMENT", bb["assessment"])):
            rows.append(_row(qid, "borrowing_base", value, bb["assessment"], bb["severity"], "Carried from Borrowing Base worksheet"))
    # Carry liquidity & reserves results into the data mart
    liq = summarize_liquidity(conn, review_case_id)
    if liq:
        for qid, value in (("LIQ_LIQUID_ASSETS", liq["liquid_assets"]), ("LIQ_MONTHS_RESERVES", liq["months_reserves"]),
                           ("LIQ_ASSESSMENT", liq["assessment"])):
            rows.append(_row(qid, "liquidity", value, liq["assessment"], liq["severity"], "Carried from Liquidity worksheet"))
    # Carry the ability-to-repay results into the data mart (consumer reviews)
    dti = summarize_dti(conn, review_case_id, auto_income=("cash_flow" in _template_modules(template)))
    if dti:
        st, sev = dti["assessment"], dti["severity"]
        for qid, value in (("DTI_TOTAL_INCOME", dti["total_income"]), ("DTI_TOTAL_OBLIGATIONS", dti["total_obligations"]),
                           ("DTI_FRONT_END_PCT", dti["front_end_dti"]), ("DTI_BACK_END_PCT", dti["back_end_dti"]),
                           ("DTI_RESIDUAL_INCOME", dti["residual_income"]), ("DTI_NET_RESIDUAL_INCOME", dti["net_residual_income"]),
                           ("DTI_ASSESSMENT", dti["assessment"])):
            rows.append(_row(qid, "ability_to_repay", value, st, sev, "Carried from DTI worksheet"))
    Path(output_path).parent.mkdir(parents=True, exist_ok=True); pd.DataFrame(rows).to_csv(output_path, index=False)
    conn.execute("INSERT INTO exports (engagement_id, review_case_id, export_type, file_path, generated_by, generated_at, export_status) VALUES (?, ?, ?, ?, ?, ?, ?)", (ctx['engagement_id'], review_case_id, "data_mart", str(output_path), generated_by, now(), "Generated")); conn.commit()
    append_audit_event(conn, generated_by, "export_generated", "export", "data_mart", after_value=str(output_path), engagement_id=ctx['engagement_id'], review_case_id=review_case_id, loan_id=ctx['loan_id'], template_id=template.template_id, template_version=template.version)
    return ExportResult(export_type="data_mart", file_path=str(output_path), export_status="Generated")

def generate_exception_report_csv(conn, output_path: str | Path = ROOT / "outputs" / "exceptions" / "exceptions_report.csv"):
    df = pd.read_sql_query("""SELECT c.client_name,e.review_period,lr.loan_id,lr.borrower_name,x.section_id as section,x.question_id,x.issue_text,x.severity,x.status,x.reviewer_comment,x.evidence_status,x.created_at,x.updated_at FROM exceptions x JOIN review_cases rc ON x.review_case_id=rc.review_case_id JOIN loan_records lr ON x.loan_record_id=lr.loan_record_id JOIN engagements e ON rc.engagement_id=e.engagement_id JOIN clients c ON e.client_id=c.client_id ORDER BY x.exception_id""", conn)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True); df.to_csv(output_path, index=False)
    return ExportResult(export_type="exceptions", file_path=str(output_path), export_status="Generated")

def generate_audit_log_csv(conn, output_path: str | Path = ROOT / "outputs" / "audit" / "audit_log.csv"):
    return ExportResult(export_type="audit", file_path=export_audit_log(conn, output_path), export_status="Generated")


# --- Engagement-level Excel data mart ----------------------------------------
_DM_TABLES = {
    "Linesheets": ["review_case_id", "client_name", "review_period", "template_id", "template_name",
                   "template_version", "loan_id", "borrower_name", "product_type", "outstanding_balance",
                   "validation_status", "review_status", "completion_pct", "required_count", "answered_required",
                   "findings_count", "blockers_count", "dti_back_end_pct", "dti_front_end_pct",
                   "dti_residual_income", "dti_net_residual_income", "dti_assessment",
                   "cf_qualifying_monthly", "cf_qualifying_annual", "cf_business_income_ref_monthly",
                   "coll_ltv_pct", "coll_coverage_pct", "coll_net_value", "coll_excess", "coll_assessment",
                   "dscr_ratio", "dscr_debt_yield_pct", "dscr_assessment",
                   "lev_current_ratio", "lev_debt_to_worth", "lev_debt_to_ebitda", "lev_assessment",
                   "guar_net_worth", "guar_personal_dscr", "guar_assessment",
                   "global_dscr", "global_assessment",
                   "bb_borrowing_base", "bb_net_availability", "bb_assessment",
                   "liq_months_reserves", "liq_assessment"],
    "Answers": ["review_case_id", "loan_id", "borrower_name", "template_id", "section_id", "question_id",
                "answer_value", "answer_status", "severity", "exception_flag", "reviewer_comment",
                "evidence_required", "evidence_status", "answered_by", "answered_at"],
    "Findings": ["exception_id", "review_case_id", "loan_id", "borrower_name", "section_id", "question_id",
                 "issue_text", "severity", "status", "reviewer_comment", "evidence_status", "created_at", "updated_at"],
    "DTI": ["review_case_id", "loan_id", "borrower_name", "total_income", "total_housing", "total_other_debt",
            "total_obligations", "front_end_dti", "back_end_dti", "residual_income", "net_residual_income",
            "total_withholding", "assessment", "severity"],
    "CashFlow": ["review_case_id", "loan_id", "borrower_name", "qualifying_monthly", "qualifying_annual",
                 "business_income_reference_monthly"],
    "Collateral": ["review_case_id", "loan_id", "borrower_name", "total_market_value", "net_collateral_value",
                   "total_exposure", "ltv", "coverage", "excess", "assessment", "severity"],
    "DSCR": ["review_case_id", "loan_id", "borrower_name", "cfads", "net_operating_income", "total_debt_service",
             "loan_amount", "dscr", "debt_yield", "excess_cash_flow", "assessment", "severity"],
    "Leverage": ["review_case_id", "loan_id", "borrower_name", "current_ratio", "working_capital",
                 "debt_to_worth", "debt_to_ebitda", "assessment", "severity"],
    "Guarantor": ["review_case_id", "loan_id", "borrower_name", "total_assets", "total_liabilities", "liquid_assets",
                  "net_worth", "personal_cf_available", "personal_debt_service", "personal_dscr", "contingent_liabilities",
                  "assessment", "severity"],
    "Global": ["review_case_id", "loan_id", "borrower_name", "business_cfads", "personal_cf_available", "global_cfads",
               "business_debt_service", "personal_debt_service", "global_debt_service", "global_dscr", "assessment", "severity"],
    "BorrowingBase": ["review_case_id", "loan_id", "borrower_name", "gross_availability", "total_reserves", "borrowing_base",
                      "line_commitment", "current_outstanding", "net_availability", "overadvance", "assessment", "severity"],
    "Liquidity": ["review_case_id", "loan_id", "borrower_name", "liquid_assets", "monthly_obligations",
                  "months_reserves", "assessment", "severity"],
    "Audit": ["audit_id", "timestamp", "user", "action_type", "entity_type", "entity_id", "reason",
              "review_case_id", "loan_id"],
}

_DM_DICTIONARY = [
    ("Linesheets", "review_case_id", "Primary key — one row per loan / review case (linesheet)."),
    ("Linesheets", "completion_pct", "Percent of required questions answered."),
    ("Linesheets", "findings_count", "Open exceptions/findings on the case (incl. carried DTI finding)."),
    ("Linesheets", "dti_back_end_pct", "Back-end DTI carried from the Ability-to-Repay worksheet."),
    ("Linesheets", "cf_qualifying_monthly", "Total qualifying monthly income from the Cash Flow worksheet."),
    ("Answers", "review_case_id", "Foreign key to Linesheets."),
    ("Answers", "exception_flag", "TRUE when the answer raised a finding/exception."),
    ("Findings", "review_case_id", "Foreign key to Linesheets."),
    ("DTI", "review_case_id", "Foreign key to Linesheets — ability-to-repay results."),
    ("CashFlow", "review_case_id", "Foreign key to Linesheets — income analysis results."),
    ("Audit", "review_case_id", "Foreign key to Linesheets (nullable for engagement-level events)."),
]


def _build_portfolio(wb, ctx, lines, findings_rows):
    """Engagement rollup with aggregates and two bar charts."""
    ws = wb.create_sheet("Portfolio"); ws.sheet_view.showGridLines = False
    for col, w in (("A", 26), ("B", 10), ("C", 4), ("D", 26), ("E", 10)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:I1"); ws.row_dimensions[1].height = 5; ws["A1"].fill = _fill(GOLD)
    ws.merge_cells("A2:I2"); ws.row_dimensions[2].height = 30
    h = ws["A2"]; h.value = f"PORTFOLIO ROLLUP — {ctx.get('client_name','')} {ctx.get('review_period','')}"
    h.fill = _fill(NAVY); h.font = Font(name=FONT, color=WHITE, bold=True, size=14); h.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 10): ws.cell(row=2, column=col).fill = _fill(NAVY)
    ws.merge_cells("A3:I3"); ws.row_dimensions[3].height = 5; ws["A3"].fill = _fill(GOLD)

    def _avg(key):
        vals = [l[key] for l in lines if isinstance(l.get(key), (int, float))]
        return round(sum(vals) / len(vals), 2) if vals else None

    total = len(lines)
    findings = sum((l.get("findings_count") or 0) for l in lines)
    comps = [l["completion_pct"] for l in lines if isinstance(l.get("completion_pct"), (int, float))]
    avg_comp = round(sum(comps) / len(comps)) if comps else 0
    blocked = sum(1 for l in lines if (l.get("blockers_count") or 0))
    kpis = [("Linesheets", total), ("Findings", findings), ("Avg completion", f"{avg_comp}%"), ("With blockers", blocked)]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 2
        lc = ws.cell(row=5, column=col, value=label); lc.font = Font(name=FONT, color=MUTED, bold=True, size=8); lc.alignment = Alignment(horizontal="center")
        lc.fill = _fill(BAND); lc.border = Border(top=_HAIR, left=_HAIR, right=_HAIR)
        vc = ws.cell(row=6, column=col, value=value); vc.font = Font(name=FONT, color=NAVY, bold=True, size=15); vc.alignment = Alignment(horizontal="center")
        vc.fill = _fill(BAND); vc.border = Border(bottom=_HAIR, left=_HAIR, right=_HAIR)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
    ws.row_dimensions[6].height = 26

    def _table(top, title, pairs):
        ws.cell(row=top, column=1, value=title).font = Font(name=FONT, color=GOLD, bold=True, size=10)
        ws.cell(row=top + 1, column=1, value="Category").font = Font(name=FONT, color=WHITE, bold=True, size=9)
        ws.cell(row=top + 1, column=2, value="Count").font = Font(name=FONT, color=WHITE, bold=True, size=9)
        for cc in (1, 2):
            ws.cell(row=top + 1, column=cc).fill = _fill(NAVY)
        r = top + 2
        for k, v in pairs:
            ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=9)
            ws.cell(row=r, column=2, value=v).font = Font(name=FONT, size=9); ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            r += 1
        return top + 1, r - 1  # header row, last row

    val = sorted(Counter((l.get("validation_status") or "—") for l in lines).items())
    vh, vl = _table(9, "VALIDATION STATUS", val)
    ch = BarChart(); ch.type = "col"; ch.title = "Validation status"; ch.legend = None; ch.height = 6; ch.width = 11
    ch.add_data(Reference(ws, min_col=2, min_row=vh, max_row=vl), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=vh + 1, max_row=vl))
    ws.add_chart(ch, "D9")

    fbs = sorted(Counter((f.get("section_id") or "—") for f in findings_rows).items())
    next_top = vl + 3
    if fbs:
        fh, fl = _table(next_top, "FINDINGS BY SECTION", fbs)
        ch2 = BarChart(); ch2.type = "bar"; ch2.title = "Findings by section"; ch2.legend = None; ch2.height = 7; ch2.width = 11
        ch2.add_data(Reference(ws, min_col=2, min_row=fh, max_row=fl), titles_from_data=True)
        ch2.set_categories(Reference(ws, min_col=1, min_row=fh + 1, max_row=fl))
        ws.add_chart(ch2, f"D{next_top}")
        next_top = fl + 3

    ws.cell(row=next_top, column=1, value="AVERAGE METRICS").font = Font(name=FONT, color=GOLD, bold=True, size=10)
    r = next_top + 1
    for label, val in (("Avg back-end DTI %", _avg("dti_back_end_pct")), ("Avg DSCR (x)", _avg("dscr_ratio")),
                       ("Avg global DSCR (x)", _avg("global_dscr")), ("Avg LTV %", _avg("coll_ltv_pct")),
                       ("Avg months reserves", _avg("liq_months_reserves"))):
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, color=MUTED, bold=True, size=9)
        vc = ws.cell(row=r, column=2, value=val if val is not None else "—")
        vc.font = Font(name=FONT, color=INK, size=9); vc.alignment = Alignment(horizontal="right")
        for cc in (1, 2): ws.cell(row=r, column=cc).border = BOTTOM
        r += 1
    return ws


def _dm_table_sheet(wb, title, table_name, columns, rows):
    ws = wb.create_sheet(title)
    ws.append(columns)
    for r in (rows or [{}]):
        ws.append([r.get(c) for c in columns])
    ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"; ws.sheet_view.showGridLines = False
    for i, c in enumerate(columns, start=1):
        width = max([len(str(c))] + [len(str((r.get(c) if r else "") or "")) for r in (rows or [])])
        ws.column_dimensions[get_column_letter(i)].width = min(46, width + 2)
    return ws


def generate_data_mart_workbook(conn, engagement_id: int,
                                output_path: str | Path = ROOT / "outputs" / "data_mart" / "linesheet_data_mart.xlsx",
                                generated_by: str = "system"):
    """Consolidate every linesheet in an engagement into one pivot-ready Excel
    data mart: normalized Tables for Linesheets, Answers, Findings, DTI, Cash
    Flow and Audit, plus an Overview and Data Dictionary."""
    ctx = conn.execute("SELECT e.*, c.client_name FROM engagements e JOIN clients c ON e.client_id=c.client_id WHERE e.engagement_id=?",
                       (engagement_id,)).fetchone()
    ctx = dict(ctx) if ctx else {}
    try:
        template = load_template(ctx.get("template_id"))
    except Exception:
        template = None

    cases = [dict(r) for r in conn.execute(
        "SELECT rc.review_case_id, rc.status AS review_status, lr.* FROM review_cases rc "
        "JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.engagement_id=? ORDER BY rc.review_case_id",
        (engagement_id,)).fetchall()]
    case_ids = [c["review_case_id"] for c in cases]

    dti_auto = ("cash_flow" in _template_modules(template)) if template else True
    lines, dti_rows, cf_rows, col_rows, dscr_rows, lev_rows, guar_rows, global_rows, bb_rows, liq_rows = [], [], [], [], [], [], [], [], [], []
    for c in cases:
        rcid = c["review_case_id"]
        carry_global(conn, rcid)
        comp = calculate_completion_status(conn, rcid, c, template) if template else {}
        dti = summarize_dti(conn, rcid, auto_income=dti_auto)
        cf = summarize_cash_flow(conn, rcid)
        col = summarize_collateral(conn, rcid)
        dscr = summarize_dscr(conn, rcid)
        lev = summarize_leverage(conn, rcid)
        guar = summarize_guarantor(conn, rcid)
        glob = summarize_global(conn, rcid)
        bb = summarize_borrowing_base(conn, rcid)
        liq = summarize_liquidity(conn, rcid)
        findings = conn.execute("SELECT COUNT(*) FROM exceptions WHERE review_case_id=?", (rcid,)).fetchone()[0]
        lines.append({
            "review_case_id": rcid, "client_name": ctx.get("client_name"), "review_period": ctx.get("review_period"),
            "template_id": ctx.get("template_id"), "template_name": getattr(template, "template_name", None),
            "template_version": getattr(template, "version", None), "loan_id": c.get("loan_id"),
            "borrower_name": c.get("borrower_name"), "product_type": c.get("product_type"),
            "outstanding_balance": c.get("outstanding_balance"), "validation_status": c.get("validation_status"),
            "review_status": c.get("review_status"), "completion_pct": comp.get("completion_pct"),
            "required_count": comp.get("required_count"), "answered_required": comp.get("answered_required"),
            "findings_count": findings, "blockers_count": len(comp.get("blockers", [])),
            "dti_back_end_pct": dti["back_end_dti"] if dti else None,
            "dti_front_end_pct": dti["front_end_dti"] if dti else None,
            "dti_residual_income": dti["residual_income"] if dti else None,
            "dti_net_residual_income": dti["net_residual_income"] if dti else None,
            "dti_assessment": dti["assessment"] if dti else None,
            "cf_qualifying_monthly": cf["qualifying_monthly"] if cf else None,
            "cf_qualifying_annual": cf["qualifying_annual"] if cf else None,
            "cf_business_income_ref_monthly": cf["business_income_reference_monthly"] if cf else None,
            "coll_ltv_pct": col["ltv"] if col else None,
            "coll_coverage_pct": col["coverage"] if col else None,
            "coll_net_value": col["net_collateral_value"] if col else None,
            "coll_excess": col["excess"] if col else None,
            "coll_assessment": col["assessment"] if col else None,
            "dscr_ratio": dscr["dscr"] if dscr else None,
            "dscr_debt_yield_pct": dscr["debt_yield"] if dscr else None,
            "dscr_assessment": dscr["assessment"] if dscr else None,
            "lev_current_ratio": lev["current_ratio"] if lev else None,
            "lev_debt_to_worth": lev["debt_to_worth"] if lev else None,
            "lev_debt_to_ebitda": lev["debt_to_ebitda"] if lev else None,
            "lev_assessment": lev["assessment"] if lev else None,
            "guar_net_worth": guar["net_worth"] if guar else None,
            "guar_personal_dscr": guar["personal_dscr"] if guar else None,
            "guar_assessment": guar["assessment"] if guar else None,
            "global_dscr": glob["global_dscr"] if glob else None,
            "global_assessment": glob["assessment"] if glob else None,
            "bb_borrowing_base": bb["borrowing_base"] if bb else None,
            "bb_net_availability": bb["net_availability"] if bb else None,
            "bb_assessment": bb["assessment"] if bb else None,
            "liq_months_reserves": liq["months_reserves"] if liq else None,
            "liq_assessment": liq["assessment"] if liq else None,
        })
        if dti:
            dti_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                             **{k: dti.get(k) for k in ("total_income", "total_housing", "total_other_debt", "total_obligations",
                                "front_end_dti", "back_end_dti", "residual_income", "net_residual_income", "total_withholding",
                                "assessment", "severity")}})
        if cf:
            cf_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                            "qualifying_monthly": cf["qualifying_monthly"], "qualifying_annual": cf["qualifying_annual"],
                            "business_income_reference_monthly": cf["business_income_reference_monthly"]})
        if col:
            col_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                             **{k: col.get(k) for k in ("total_market_value", "net_collateral_value", "total_exposure",
                                "ltv", "coverage", "excess", "assessment", "severity")}})
        if dscr:
            dscr_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                              **{k: dscr.get(k) for k in ("cfads", "net_operating_income", "total_debt_service",
                                 "loan_amount", "dscr", "debt_yield", "excess_cash_flow", "assessment", "severity")}})
        if lev:
            lev_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                             **{k: lev.get(k) for k in ("current_ratio", "working_capital", "debt_to_worth",
                                "debt_to_ebitda", "assessment", "severity")}})
        if guar:
            guar_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                              **{k: guar.get(k) for k in ("total_assets", "total_liabilities", "liquid_assets", "net_worth",
                                 "personal_cf_available", "personal_debt_service", "personal_dscr", "contingent_liabilities",
                                 "assessment", "severity")}})
        if glob:
            global_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                                **{k: glob.get(k) for k in ("business_cfads", "personal_cf_available", "global_cfads",
                                   "business_debt_service", "personal_debt_service", "global_debt_service", "global_dscr",
                                   "assessment", "severity")}})
        if bb:
            bb_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                            **{k: bb.get(k) for k in ("gross_availability", "total_reserves", "borrowing_base",
                               "line_commitment", "current_outstanding", "net_availability", "overadvance",
                               "assessment", "severity")}})
        if liq:
            liq_rows.append({"review_case_id": rcid, "loan_id": c.get("loan_id"), "borrower_name": c.get("borrower_name"),
                             **{k: liq.get(k) for k in ("liquid_assets", "monthly_obligations", "months_reserves",
                                "assessment", "severity")}})

    answers = []
    for r in conn.execute(
        "SELECT ra.*, lr.loan_id, lr.borrower_name FROM review_answers ra "
        "JOIN review_cases rc ON ra.review_case_id=rc.review_case_id "
        "JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.engagement_id=? ORDER BY ra.review_case_id, ra.answer_id",
        (engagement_id,)).fetchall():
        d = dict(r); d["exception_flag"] = bool(d.get("severity")); d["template_id"] = ctx.get("template_id"); answers.append(d)
    findings_rows = [dict(r) for r in conn.execute(
        "SELECT x.*, lr.loan_id, lr.borrower_name FROM exceptions x "
        "JOIN review_cases rc ON x.review_case_id=rc.review_case_id "
        "JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.engagement_id=? ORDER BY x.exception_id",
        (engagement_id,)).fetchall()]
    ph = ",".join("?" * len(case_ids)) or "NULL"
    audit_rows = [dict(r) for r in conn.execute(
        f"SELECT audit_id,timestamp,user,action_type,entity_type,entity_id,reason,review_case_id,loan_id "
        f"FROM audit_log WHERE engagement_id=? OR review_case_id IN ({ph}) ORDER BY audit_id",
        [engagement_id] + case_ids).fetchall()]

    wb = Workbook()
    wb.properties.title = f"Linesheet Data Mart — {ctx.get('client_name','')} {ctx.get('review_period','')}"
    wb.properties.creator = "Linesheet Builder"

    ov = wb.active; ov.title = "Overview"; ov.sheet_view.showGridLines = False
    ov.column_dimensions["A"].width = 2.5; ov.column_dimensions["B"].width = 28; ov.column_dimensions["C"].width = 40
    ov.merge_cells("A1:D1"); ov.row_dimensions[1].height = 5; ov["A1"].fill = _fill(GOLD)
    ov.merge_cells("A2:D2"); ov.row_dimensions[2].height = 34
    h = ov["A2"]; h.value = "LINESHEET DATA MART"; h.fill = _fill(NAVY); h.font = Font(name=FONT, color=WHITE, bold=True, size=18)
    h.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5): ov.cell(row=2, column=col).fill = _fill(NAVY)
    ov.merge_cells("A3:D3"); ov.row_dimensions[3].height = 5; ov["A3"].fill = _fill(GOLD)
    meta = [("Client", ctx.get("client_name")), ("Review period", ctx.get("review_period")),
            ("Template", f"{getattr(template,'template_name', ctx.get('template_id'))}"),
            ("Linesheets (cases)", len(cases)), ("Answers", len(answers)), ("Findings", len(findings_rows)),
            ("Generated", now().replace("T", "  "))]
    r = 5
    for label, value in meta:
        ov.cell(row=r, column=2, value=label).font = Font(name=FONT, color=MUTED, bold=True, size=10)
        ov.cell(row=r, column=3, value=value).font = Font(name=FONT, color=INK, size=10)
        for col in (2, 3): ov.cell(row=r, column=col).border = BOTTOM
        r += 1
    r += 1
    ov.cell(row=r, column=2, value="TABLES").font = Font(name=FONT, color=GOLD, bold=True, size=10); r += 1
    for name, count in (("Linesheets", len(lines)), ("Answers", len(answers)), ("Findings", len(findings_rows)),
                        ("DTI", len(dti_rows)), ("CashFlow", len(cf_rows)), ("Collateral", len(col_rows)),
                        ("DSCR", len(dscr_rows)), ("Leverage", len(lev_rows)), ("Guarantor", len(guar_rows)),
                        ("Global", len(global_rows)), ("BorrowingBase", len(bb_rows)), ("Liquidity", len(liq_rows)),
                        ("Audit", len(audit_rows)), ("Data Dictionary", len(_DM_DICTIONARY))):
        ov.cell(row=r, column=2, value=name).font = Font(name=FONT, color=NAVY, bold=True, size=10)
        ov.cell(row=r, column=3, value=f"{count} rows").font = Font(name=FONT, color=MUTED, size=10)
        for col in (2, 3): ov.cell(row=r, column=col).border = BOTTOM
        r += 1

    _build_portfolio(wb, ctx, lines, findings_rows)
    _dm_table_sheet(wb, "Linesheets", "tbl_Linesheets", _DM_TABLES["Linesheets"], lines)
    _dm_table_sheet(wb, "Answers", "tbl_Answers", _DM_TABLES["Answers"], answers)
    _dm_table_sheet(wb, "Findings", "tbl_Findings", _DM_TABLES["Findings"], findings_rows)
    _dm_table_sheet(wb, "DTI", "tbl_DTI", _DM_TABLES["DTI"], dti_rows)
    _dm_table_sheet(wb, "CashFlow", "tbl_CashFlow", _DM_TABLES["CashFlow"], cf_rows)
    _dm_table_sheet(wb, "Collateral", "tbl_Collateral", _DM_TABLES["Collateral"], col_rows)
    _dm_table_sheet(wb, "DSCR", "tbl_DSCR", _DM_TABLES["DSCR"], dscr_rows)
    _dm_table_sheet(wb, "Leverage", "tbl_Leverage", _DM_TABLES["Leverage"], lev_rows)
    _dm_table_sheet(wb, "Guarantor", "tbl_Guarantor", _DM_TABLES["Guarantor"], guar_rows)
    _dm_table_sheet(wb, "Global", "tbl_Global", _DM_TABLES["Global"], global_rows)
    _dm_table_sheet(wb, "BorrowingBase", "tbl_BorrowingBase", _DM_TABLES["BorrowingBase"], bb_rows)
    _dm_table_sheet(wb, "Liquidity", "tbl_Liquidity", _DM_TABLES["Liquidity"], liq_rows)
    _dm_table_sheet(wb, "Audit", "tbl_Audit", _DM_TABLES["Audit"], audit_rows)
    _dm_table_sheet(wb, "Data Dictionary", "tbl_Dictionary", ["table", "column", "description"],
                    [{"table": t, "column": c, "description": d} for t, c, d in _DM_DICTIONARY])

    output_path = Path(output_path); output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    conn.execute("INSERT INTO exports (engagement_id, review_case_id, export_type, file_path, generated_by, generated_at, export_status) VALUES (?, ?, ?, ?, ?, ?, ?)",
                 (engagement_id, None, "data_mart_workbook", str(output_path), generated_by, now(), "Generated")); conn.commit()
    append_audit_event(conn, generated_by, "export_generated", "export", "data_mart_workbook",
                       after_value=str(output_path), engagement_id=engagement_id)
    return ExportResult(export_type="data_mart_workbook", file_path=str(output_path), export_status="Generated")
