from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _autosize(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build_workbook(records: list[dict[str, Any]], errors: list[dict[str, str]], scan_meta: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws_review = wb.active
    ws_review.title = "Review"
    ws_dups = wb.create_sheet("Duplicate Groups")
    ws_sum = wb.create_sheet("Summary")
    ws_err = wb.create_sheet("Errors")

    review_cols = [
        "Client Decision", "Recommended Action", "Duplicate Group ID", "Duplicate Confidence", "Full Path", "Folder Path",
        "File Name", "Extension", "Category", "Size Bytes", "Size MB", "Size GB", "Created Date", "Modified Date",
        "Last Accessed Date", "Top-Level Folder", "Parent Folder", "Likely Source / Machine", "Hash", "Reason Flagged",
        "Notes", "Scan Error",
    ]
    ws_review.append(review_cols)
    for r in records:
        ws_review.append([
            "", r["recommended_action"], r["duplicate_group_id"], r["duplicate_confidence"], r["full_path"], r["folder_path"], r["file_name"],
            r["extension"], r["category"], r["size_bytes"], r["size_mb"], r["size_gb"], _fmt_dt(r["created_date"]), _fmt_dt(r["modified_date"]),
            _fmt_dt(r["accessed_date"]), r["top_level_folder"], r["parent_folder"], r["likely_source"], r.get("hash", ""), r["reason_flagged"], "", r.get("scan_error", ""),
        ])
    dv = DataValidation(type="list", formula1='"Keep,Delete,Archive,Unsure,Needs Review"')
    ws_review.add_data_validation(dv)
    if len(records) > 0:
        dv.add(f"A2:A{len(records)+1}")

    dup_cols = ["Duplicate Group ID", "File Count", "Total Group Size GB", "Potential Recoverable Size GB", "File Name Examples", "Categories", "Oldest Modified Date", "Newest Modified Date", "Locations", "Hash"]
    ws_dups.append(dup_cols)

    dup_groups = defaultdict(list)
    for r in records:
        if r["duplicate_group_id"]:
            dup_groups[r["duplicate_group_id"]].append(r)

    for group_id in sorted(dup_groups):
        rows = dup_groups[group_id]
        total = sum(x["size_bytes"] for x in rows)
        recoverable = max(0, total - rows[0]["size_bytes"])
        ws_dups.append([
            group_id,
            len(rows),
            round(total / (1024**3), 2),
            round(recoverable / (1024**3), 2),
            ", ".join(sorted({x["file_name"] for x in rows})[:5]),
            ", ".join(sorted({x["category"] for x in rows})),
            _fmt_dt(min(x["modified_date"] for x in rows if x["modified_date"])),
            _fmt_dt(max(x["modified_date"] for x in rows if x["modified_date"])),
            " | ".join(sorted({x["folder_path"] for x in rows})[:5]),
            rows[0].get("hash", ""),
        ])

    category_counter = Counter(r["category"] for r in records)
    category_sizes = defaultdict(int)
    dup_count = defaultdict(int)
    dup_size = defaultdict(int)
    for r in records:
        category_sizes[r["category"]] += r["size_bytes"]
        if r["duplicate_group_id"]:
            dup_count[r["category"]] += 1
            dup_size[r["category"]] += r["size_bytes"]

    ws_sum.append(["Metric", "Value"])
    items = [
        ("Scan Root", scan_meta["scan_root"]),
        ("Output File", scan_meta["output_file"]),
        ("Scan Date", scan_meta["scan_date"]),
        ("File Type Mode", scan_meta["file_type_mode"]),
        ("Extension Filter", scan_meta["extension_filter"]),
        ("Duplicate Detection Enabled", str(scan_meta["duplicate_detection"])),
        ("Total Files Scanned", scan_meta["total_files"]),
        ("Total Size GB", scan_meta["total_size_gb"]),
        ("Total Duplicate Groups", scan_meta["total_duplicate_groups"]),
        ("Total Duplicate Files", scan_meta["total_duplicate_files"]),
        ("Potential Recoverable Duplicate Size GB", scan_meta["potential_recoverable_gb"]),
        ("Files Skipped Due To Error", scan_meta["files_skipped_error"]),
        ("Largest File Size GB", scan_meta["largest_file_gb"]),
        ("Oldest Modified Date", scan_meta["oldest_modified"]),
        ("Newest Modified Date", scan_meta["newest_modified"]),
    ]
    for k, v in items:
        ws_sum.append([k, v])
    ws_sum.append([])
    ws_sum.append(["Note", "This workbook is for review only. No files were deleted, moved, renamed, or modified by this tool."])
    ws_sum.append([])
    ws_sum.append(["Category", "File Count", "Total Size GB", "Duplicate File Count", "Duplicate Size GB"])
    for cat in sorted(category_counter):
        ws_sum.append([cat, category_counter[cat], round(category_sizes[cat]/(1024**3), 2), dup_count[cat], round(dup_size[cat]/(1024**3), 2)])

    ws_err.append(["Full Path", "Error Type", "Error Message", "Stage"])
    for err in errors:
        ws_err.append([err.get("full_path", ""), err.get("error_type", ""), err.get("error_message", ""), err.get("stage", "Other")])

    for ws in [ws_review, ws_dups, ws_err]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    ws_sum.auto_filter.ref = "A1:B16"

    bold = Font(bold=True)
    for ws in [ws_review, ws_dups, ws_sum, ws_err]:
        for c in ws[1]:
            c.font = bold

    dup_fill = PatternFill(start_color="FFF8D6", end_color="FFF8D6", fill_type="solid")
    err_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    ws_review.conditional_formatting.add(f"C2:C{len(records)+1}", FormulaRule(formula=["LEN($C2)>0"], fill=dup_fill))
    ws_review.conditional_formatting.add(f"V2:V{len(records)+1}", FormulaRule(formula=["LEN($V2)>0"], fill=err_fill))

    _autosize(ws_review)
    _autosize(ws_dups)
    _autosize(ws_sum)
    _autosize(ws_err)

    wb.save(output_path)
