from __future__ import annotations

from datetime import date, datetime
import importlib.util
from decimal import Decimal
from pathlib import Path
from typing import Any

from .simple_xlsx import read_xlsx
from .template_scanner import normalize_field_name


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _openpyxl_available() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


def read_workbook(path: str | Path) -> dict[str, list[dict]]:
    """Read workbook rows without mutating the source file.

    Uses openpyxl for normal Excel files when available. The lightweight
    simple_xlsx reader remains as a fallback for constrained environments.
    """
    path = Path(path)
    if not _openpyxl_available():
        return read_xlsx(path)

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    for worksheet in workbook.worksheets:
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            sheets[worksheet.title] = []
            continue
        headers = [_format_cell_value(value).strip() for value in raw_rows[0]]
        rows: list[dict] = []
        for raw_row in raw_rows[1:]:
            row: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = raw_row[index] if index < len(raw_row) else None
                row[header] = _format_cell_value(value)
            if any(value != "" for value in row.values()):
                rows.append(row)
        sheets[worksheet.title] = rows
    workbook.close()
    return sheets


class OccamWorkbook:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.sheets = read_workbook(self.path)

    def rows(self, sheet: str) -> list[dict]:
        return self.sheets.get(sheet, [])

    def settings_dict(self) -> dict:
        return {r.get("Setting", ""): r.get("Value", "") for r in self.rows("Settings")}

    def clients(self, sheet: str = "Clients") -> list[dict]:
        return self.rows(sheet)

    def get_client(self, match_value: str, match_field: str = "Client ID", sheet: str = "Clients") -> dict:
        wanted = str(match_value)
        for row in self.clients(sheet):
            if str(row.get(match_field, "")) == wanted:
                return row
        return {}

    def get_client_by_id(self, client_id: str, sheet: str = "Clients") -> dict:
        return self.get_client(client_id, "Client ID", sheet)

    def get_client_by_name(self, name: str, sheet: str = "Clients") -> dict:
        return self.get_client(name, "Client Name", sheet)

    def client_options(self, match_field: str = "Client ID", sheet: str = "Clients") -> list[dict]:
        options = []
        for row in self.clients(sheet):
            match_value = row.get(match_field, "")
            label = row.get("Client Name", "Unnamed Client")
            if match_field != "Client Name" and match_value:
                label = f"{label} ({match_field}: {match_value})"
            options.append({"label": label, "match_value": match_value, "client": row})
        return options

    def invoices_for_client(self, client_id: str) -> list[dict]:
        return [r for r in self.rows("Invoices") if str(r.get("Client ID", "")) == str(client_id)]

    def missing_items_for_client(self, client_id: str) -> list[dict]:
        return [r for r in self.rows("Missing Items") if str(r.get("Client ID", "")) == str(client_id)]

    def field_pool_for_client(self, client: dict, invoice: dict | None = None) -> dict[str, tuple[str, str]]:
        pool: dict[str, tuple[str, str]] = {}

        def add_map(row: dict, source: str):
            for key, value in row.items():
                if value not in (None, ""):
                    pool[normalize_field_name(key)] = (str(value), source)

        add_map(self.settings_dict(), "Settings sheet")
        add_map(client, "Clients table")
        if invoice:
            add_map(invoice, "Invoices table")
        missing = self.missing_items_for_client(client.get("Client ID", ""))
        if missing:
            items = [m.get("Missing Item", "") for m in missing if str(m.get("Status", "")).lower() != "received"]
            pool[normalize_field_name("Missing Items")] = ("; ".join(items), "Missing Items table")
        return pool
