from __future__ import annotations

import importlib.util
import json
import re
import uuid
from pathlib import Path
from datetime import datetime

from .audit import build_audit, write_json, now_iso
from .renderer import render_docx_template, render_email_template, scan_rendered_output_for_placeholders
from .simple_xlsx import write_xlsx

NAVY = "0B1F3A"
GOLD = "C9A227"
GRAY = "E5E7EB"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"


def sanitize_filename(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._ -]+", "", name).strip().replace(" ", "_")
    return clean[:80] or "Untitled"


def create_output_package_folder(output_root: str | Path, client_name: str, template_name: str) -> Path:
    root = Path(output_root)
    root.mkdir(parents=True, exist_ok=True)
    base = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{sanitize_filename(client_name)}_{sanitize_filename(Path(template_name).stem)}"
    for _ in range(10):
        folder = root / f"{base}_{uuid.uuid4().hex[:8]}"
        try:
            folder.mkdir(parents=True, exist_ok=False)
            return folder
        except FileExistsError:
            continue
    raise FileExistsError("Unable to create a unique output package folder after multiple attempts.")


def _report_rows(fields: list[dict], validation: dict, audit: dict) -> dict[str, list[dict]]:
    generated_files = audit.get("generated_files", [])
    outlook = audit.get("outlook_draft_status", {}) or {}
    rows_summary = [
        {"Item": "Run Status", "Value": validation.get("status", "")},
        {"Item": "Status", "Value": validation.get("status", "")},
        {"Item": "Client", "Value": audit.get("client_name", "")},
        {"Item": "Template", "Value": audit.get("template_name", "")},
        {"Item": "Template Type", "Value": audit.get("template_type", "")},
        {"Item": "Timestamp", "Value": audit.get("timestamp", "")},
        {"Item": "Run ID", "Value": audit.get("run_id", "")},
        {"Item": "Blocker Count", "Value": len(validation.get("blockers", []))},
        {"Item": "Warning Count", "Value": len(validation.get("warnings", []))},
        {"Item": "Unresolved Placeholder Count", "Value": len(validation.get("unresolved_placeholders", []))},
        {"Item": "Outlook Status", "Value": outlook.get("message", "Not attempted")},
        {"Item": "Generated Files", "Value": "\n".join(generated_files)},
        {"Item": "Next Action", "Value": "; ".join(validation.get("next_actions", []))},
    ]
    rows_validation = []
    for blocker in validation.get("blockers", []):
        rows_validation.append({"Type": "Blocked", "Message": blocker, "Next Action": "Resolve before generation or use."})
    for warning in validation.get("warnings", []):
        rows_validation.append({"Type": "Warning", "Message": warning, "Next Action": "Review before sending or filing."})
    rows_validation.append({"Type": "Template Diagnostics", "Message": f"Template reviewed: {audit.get('template_name', '')}", "Next Action": "Confirm template health check was acceptable."})
    if not validation.get("blockers") and not validation.get("warnings"):
        rows_validation.append({"Type": "Ready", "Message": "No blockers or warnings.", "Next Action": "Ready."})
    rows_audit = [{"Key": key, "Value": json.dumps(value) if isinstance(value, (list, dict)) else value} for key, value in audit.items()]
    field_rows = [
        {
            "Field": field.get("field", ""),
            "Required/Optional": "Required" if field.get("required", True) else "Optional",
            "Value": field.get("value", ""),
            "Source": field.get("source", ""),
            "Status": field.get("status", ""),
            "Overridden": "Yes" if field.get("overridden") else "No",
        }
        for field in fields
    ] or [{"Field": "", "Required/Optional": "", "Value": "", "Source": "", "Status": "", "Overridden": ""}]
    for token in validation.get("unresolved_placeholders", []):
        rows_validation.append({"Type": "Unresolved Placeholder", "Message": f"{{{{{token}}}}}", "Next Action": "Fix the template or provide a value, then regenerate."})
    return {
        "Summary": rows_summary,
        "Fields": field_rows,
        "Validation Results": rows_validation,
        "Audit Log": rows_audit,
    }


def _openpyxl_available() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


def _status_fill(status: str):
    from openpyxl.styles import PatternFill

    normalized = str(status).lower()
    if "ready" in normalized or "success" in normalized:
        return PatternFill("solid", fgColor=GREEN)
    if "blocked" in normalized or "error" in normalized or "do not send" in normalized:
        return PatternFill("solid", fgColor=RED)
    if "warning" in normalized or "needs review" in normalized:
        return PatternFill("solid", fgColor=AMBER)
    return PatternFill("solid", fgColor=GRAY)


def _write_openpyxl_report(path: Path, rows_by_sheet: dict[str, list[dict]], validation: dict) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    navy_fill = PatternFill("solid", fgColor=NAVY)
    gold_fill = PatternFill("solid", fgColor=GOLD)
    gray_fill = PatternFill("solid", fgColor=GRAY)
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color=NAVY, bold=True)

    for sheet_name, rows in rows_by_sheet.items():
        ws = workbook.create_sheet(sheet_name)
        headers = list(rows[0].keys()) if rows else ["Item", "Value"]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        title = ws.cell(row=1, column=1, value=f"Occam Template Desk - {sheet_name}")
        title.fill = navy_fill
        title.font = white_font
        title.alignment = Alignment(horizontal="center")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = gray_fill
            cell.font = dark_font
        for row_index, row in enumerate(rows, 3):
            for col, header in enumerate(headers, 1):
                value = row.get(header, "")
                cell = ws.cell(row=row_index, column=col, value=value)
                if header.lower() in {"status", "type"} or str(value) in {"Ready", "Needs Review", "Blocked", "Warning", "Error"}:
                    cell.fill = _status_fill(str(value))
                if any(token in header.lower() for token in ["amount", "fee"]):
                    cell.number_format = '$#,##0.00'
                if "date" in header.lower() or header.lower() == "timestamp":
                    cell.number_format = 'yyyy-mm-dd'
        if sheet_name == "Summary":
            ws.row_dimensions[3].height = 28
            status_cell = ws.cell(row=3, column=2)
            status_cell.fill = _status_fill(validation.get("status", ""))
            status_cell.font = Font(color=NAVY, bold=True, size=14)
            ws.cell(row=3, column=1).fill = gold_fill
            ws.cell(row=3, column=1).font = dark_font
        ws.freeze_panes = "A3"
        for col in range(1, len(headers) + 1):
            width = max(14, min(60, max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1)) + 3))
            ws.column_dimensions[get_column_letter(col)].width = width
    workbook.save(path)
    return str(path)


def write_validation_report(path: str | Path, fields: list[dict], validation: dict, audit: dict) -> str:
    path = Path(path)
    rows_by_sheet = _report_rows(fields, validation, audit)
    if _openpyxl_available():
        return _write_openpyxl_report(path, rows_by_sheet, validation)
    return str(write_xlsx(path, rows_by_sheet, styles=True))


def _post_render_placeholder_warnings(rendered: dict) -> tuple[list[str], list[str]]:
    warnings = []
    unresolved_tokens = []
    seen = set()
    for file_path in rendered.get("generated_files", []):
        path = Path(file_path)
        if path.name == "email_metadata.json":
            continue
        if path.suffix.lower() not in {".docx", ".txt", ".html", ".md"}:
            continue
        remaining = scan_rendered_output_for_placeholders(path)
        if remaining:
            unresolved_tokens.extend(remaining)
            token_list = ", ".join(f"{{{{{token}}}}}" for token in remaining)
            message = f"Rendered output still contains unreplaced placeholder(s) in {path.name}: {token_list}."
            if message not in seen:
                seen.add(message)
                warnings.append(message)
    return warnings, sorted(set(unresolved_tokens))


def build_output_package(template_path: str, template_type: str, values: dict, fields: list[dict], validation, client: dict, output_root: str | Path, outlook_status: dict | None = None, detected_placeholders: list[str] | None = None, user_overrides: dict | None = None, selected_invoice: dict | None = None, selected_attachments: list[str] | None = None) -> dict:
    if validation.blockers:
        raise ValueError("Blocked runs cannot generate output packages.")
    template = Path(template_path)
    package_dir = create_output_package_folder(output_root, client.get("Client Name", "Client"), template.name)
    if template_type == "email":
        rendered = render_email_template(template, values, package_dir, values.get("Client Email", ""), validation.status)
    else:
        rendered = render_docx_template(template, values, package_dir, f"{sanitize_filename(client.get('Client Name','Client'))}_{sanitize_filename(template.stem)}.docx")
    generated_files = rendered.get("generated_files", [])
    validation_dict = validation.to_dict()
    post_render_warnings, unresolved_placeholders = _post_render_placeholder_warnings(rendered)
    if post_render_warnings:
        validation_dict["warnings"] = validation_dict.get("warnings", []) + post_render_warnings
        validation_dict["status"] = "Blocked - Do Not Send"
        validation_dict["blockers"] = validation_dict.get("blockers", []) + ["Rendered output contains unreplaced placeholders and must not be sent."]
        validation_dict["next_actions"] = validation_dict.get("next_actions", []) + ["Fix unresolved placeholders before creating an Outlook draft or sending to a client."]
    validation_dict["unresolved_placeholders"] = unresolved_placeholders
    audit = build_audit(str(template), template_type, client, validation_dict, generated_files, outlook_status)
    audit["unresolved_placeholders"] = unresolved_placeholders
    audit["selected_invoice"] = selected_invoice or {}
    audit["selected_attachments"] = selected_attachments or []
    audit["user_overrides"] = user_overrides or {}
    snapshot = {"selected_template": str(template), "selected_client": client, "selected_invoice": selected_invoice or {}, "selected_attachments": selected_attachments or [], "detected_placeholders": detected_placeholders or [], "final_values": values, "field_sources": fields, "user_overrides": user_overrides or {}, "post_render_warnings": post_render_warnings, "unresolved_placeholders": unresolved_placeholders, "timestamp": now_iso()}
    write_json(package_dir / "input_snapshot.json", snapshot)
    write_json(package_dir / "rendered_values.json", values)
    write_json(package_dir / "audit_log.json", audit)
    if outlook_status:
        write_json(package_dir / "outlook_status.json", outlook_status)
    report_path = write_validation_report(package_dir / "validation_report.xlsx", fields, validation_dict, audit)
    return {"package_dir": str(package_dir), "rendered": rendered, "audit": audit, "validation": validation_dict, "validation_report": report_path, "audit_log": str(package_dir / "audit_log.json"), "unresolved_placeholders": unresolved_placeholders, "generated_files": generated_files + [str(package_dir / "input_snapshot.json"), str(package_dir / "audit_log.json"), str(package_dir / "rendered_values.json"), report_path]}


def update_outlook_status(package_dir: str | Path, status: dict) -> None:
    package_dir = Path(package_dir)
    write_json(package_dir / "outlook_status.json", status)
    audit_path = package_dir / "audit_log.json"
    if audit_path.exists():
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        audit["outlook_draft_status"] = status
        if "attachments" in status:
            audit["selected_attachments"] = status.get("attachments", [])
        write_json(audit_path, audit)
    snapshot_path = package_dir / "input_snapshot.json"
    if snapshot_path.exists() and "attachments" in status:
        snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
        snapshot["selected_attachments"] = status.get("attachments", [])
        write_json(snapshot_path, snapshot)


def list_recent_packages(output_root: str | Path, limit: int = 20) -> list[dict]:
    root = Path(output_root)
    if not root.exists():
        return []
    packages = []
    for folder in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.stat().st_mtime, reverse=True)[:limit]:
        audit_path = folder / "audit_log.json"
        data = {"folder": str(folder), "timestamp": "", "status": "", "client": "", "template": "", "generated_files": []}
        if audit_path.exists():
            try:
                audit = json.loads(audit_path.read_text(encoding="utf-8"))
                outlook = audit.get("outlook_draft_status", {}) or {}
                data.update({
                    "timestamp": audit.get("timestamp", ""),
                    "status": audit.get("validation_status", ""),
                    "client": audit.get("client_name", ""),
                    "template": audit.get("template_name", ""),
                    "blocker_count": len(audit.get("blockers", [])),
                    "warning_count": len(audit.get("warnings", [])),
                    "outlook_status": outlook.get("message", "Not attempted"),
                    "generated_files": audit.get("generated_files", []),
                })
            except (OSError, json.JSONDecodeError):
                data.update({"status": "Audit unavailable", "client": "Audit unavailable", "template": folder.name})
        packages.append(data)
    return packages
