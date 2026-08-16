"""Excel / CSV exporters for screens and single-company valuations (pure).

Everything here is in-memory (BytesIO) and returns ``bytes`` — no file paths, no
DB, no network — so the UI can hand the result straight to a Streamlit
``download_button`` and tests can assert on the bytes.

Honest-framing invariant: every workbook carries a README/Disclaimer tab with
the research-aid disclaimer, the as-of context, and a generated-at note. Missing
figures render as ``n/m`` / ``n/a`` (never a fabricated 0), mirroring the rest of
the app. Prices are non-canonical; forensic scores are screens, not verdicts.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- shared style tokens (match the UI navy/white aesthetic) ------------------
_NAVY = "FF1B2A4A"
_ACCENT = "FF2C4A7C"
_HEADER_FILL = PatternFill("solid", fgColor="FF1B2A4A")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=11)
_TITLE_FONT = Font(color=_NAVY, bold=True, size=14)
_LABEL_FONT = Font(color=_NAVY, bold=True)
_CAVEAT_FONT = Font(color="FF6B7280", italic=True, size=9)
_THIN = Side(style="thin", color="FFE3E7EE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MONEY_FMT = '_-$* #,##0_-;[Red]-$* #,##0_-;_-$* "-"_-'
_PCT_FMT = "0.0%;[Red]-0.0%"
_MULT_FMT = '0.00"x"'
_NUM_FMT = "0.00"

DISCLAIMER = (
    "Research aid — NOT investment advice and NOT a price target. Valuations are "
    "scenario estimates from explicit, editable assumptions. Prices are "
    "non-canonical (see source). Forensic / distress / manipulation figures are "
    "statistical SCREENS for further reading, never accusations or verdicts. "
    "Drake stays the system of record for filed returns and computations."
)


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #


def _now_note() -> str:
    return f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')} by stock-helper."


def _as_of_note(as_of: Any) -> str:
    if as_of in (None, ""):
        return "As-of: current view (latest available facts, filings, and prices)."
    return f"As-of: point-in-time view as of {as_of} (facts/filings after excluded)."


def _autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = None
        for cell in col_cells:
            letter = cell.column_letter
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        if letter is not None:
            ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _readme_sheet(wb: Workbook, title: str, meta: dict[str, Any] | None) -> None:
    ws = wb.create_sheet("README")
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    lines = [
        "",
        _as_of_note((meta or {}).get("as_of")),
        _now_note(),
    ]
    for extra_key in ("universe", "screen", "ticker", "bucket", "source_label"):
        val = (meta or {}).get(extra_key)
        if val:
            lines.append(f"{extra_key.replace('_', ' ').title()}: {val}")
    lines += ["", "Disclaimer:", DISCLAIMER]
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        if line in ("Disclaimer:",):
            cell.font = _LABEL_FONT
        elif line == DISCLAIMER:
            cell.font = _CAVEAT_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 90
    ws.column_dimensions["A"].width = 110


def _fmt_for(colname: str) -> str | None:
    """Best-effort number format from a column name."""
    low = colname.lower()
    if "gap" in low or "upside" in low or "yield" in low or low.endswith("%") or "margin" in low:
        return _PCT_FMT
    if low.startswith("ev_") or low in ("pe", "p_b", "p_fcf", "p_tbv") or "multiple" in low:
        return _MULT_FMT
    if "cap" in low or low in ("price", "fair_value", "net_debt"):
        return _MONEY_FMT
    if "score" in low or low.endswith("_z") or " z" in low:
        return _NUM_FMT
    return None


def _write_dataframe(ws, df: pd.DataFrame, *, start_row: int = 1) -> None:
    """Render a DataFrame with a styled header, borders, and per-column formats.
    None/NaN cells render as the honest 'n/m' placeholder."""
    headers = [str(c) for c in df.columns]
    _write_header(ws, headers, row=start_row)
    for r, (_, rowvals) in enumerate(df.iterrows(), start=start_row + 1):
        for c, colname in enumerate(df.columns, start=1):
            value = rowvals[colname]
            cell = ws.cell(row=r, column=c)
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = "n/m"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.value = float(value)
                fmt = _fmt_for(str(colname))
                if fmt:
                    cell.number_format = fmt
            else:
                cell.value = str(value)
            cell.border = _BORDER
    _autosize(ws)


# --------------------------------------------------------------------------- #
# Screen exports
# --------------------------------------------------------------------------- #


def screen_to_xlsx(df: pd.DataFrame, meta: dict[str, Any] | None = None) -> bytes:
    """A screen workbook: a styled ``Screen`` tab + a ``README`` disclaimer tab.

    ``df`` is the already-filtered/ranked screen table; ``meta`` may carry
    ``as_of``, ``screen``, ``universe``, ``source_label`` for the README."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Screen"
    if df is None or df.empty:
        ws["A1"] = "No candidates for this screen / filter set."
        ws["A1"].font = _LABEL_FONT
    else:
        _write_dataframe(ws, df)
    _readme_sheet(wb, "stock-helper — cross-sectional screen", meta)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def screen_to_csv(df: pd.DataFrame) -> bytes:
    """The screen table as UTF-8 CSV bytes (round-trippable with pandas)."""
    if df is None:
        df = pd.DataFrame()
    return df.to_csv(index=False).encode("utf-8")


# --------------------------------------------------------------------------- #
# Valuation exports
# --------------------------------------------------------------------------- #


def _g(obj: Any, name: str, default: Any = None) -> Any:
    """Defensive getattr: any sub-result on a ValuationResult may be None."""
    return getattr(obj, name, default) if obj is not None else default


def _rows_sheet(wb: Workbook, title: str, rows: list[tuple[str, Any]], fmts: dict[str, str] | None = None) -> None:
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    fmts = fmts or {}
    for i, (label, value) in enumerate(rows, start=3):
        lcell = ws.cell(row=i, column=1, value=label)
        lcell.font = _LABEL_FONT
        vcell = ws.cell(row=i, column=2)
        if value is None:
            vcell.value = "n/m"
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            vcell.value = float(value)
            if label in fmts:
                vcell.number_format = fmts[label]
        else:
            vcell.value = str(value)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 32


def valuation_to_xlsx(val: Any) -> bytes:
    """Full single-company valuation workbook.

    Tabs: Summary, DCF (assumptions + projection), Multiples, Quality &
    Forensics, README/Disclaimer. Every sub-result is read defensively so a
    partial ValuationResult (no price, no DCF, financial bucket, ...) still
    exports cleanly with n/m placeholders rather than crashing."""
    ticker = _g(val, "ticker", "")
    bucket = _g(val, "bucket", "")
    dcf = _g(val, "dcf")
    reverse = _g(val, "reverse")
    quality = _g(val, "quality")
    beneish = _g(val, "beneish")
    stress = _g(val, "stress")

    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary ---
    summary_rows: list[tuple[str, Any]] = [
        ("Ticker", ticker),
        ("Industry bucket", bucket),
        ("Price (non-canonical)", _g(val, "price")),
        ("Price date", _g(val, "price_date")),
        ("Fair value / share", _g(val, "fair_value_per_share")),
        ("Margin of safety", _g(val, "margin_of_safety")),
        ("Reverse-DCF implied growth", _g(val, "implied_growth")),
        ("Metric set", _g(val, "metric_set")),
        ("Flags", ", ".join(_g(val, "flags", []) or []) or "none"),
        ("Engine version", _g(val, "engine_version")),
    ]
    _rows_sheet(
        wb, "Summary", summary_rows,
        fmts={
            "Price (non-canonical)": _MONEY_FMT,
            "Fair value / share": _MONEY_FMT,
            "Margin of safety": _PCT_FMT,
            "Reverse-DCF implied growth": _PCT_FMT,
        },
    )

    # --- DCF assumptions + projection ---
    ws = wb.create_sheet("DCF")
    ws["A1"] = "DCF assumptions & projection"
    ws["A1"].font = _TITLE_FONT
    assum: list[tuple[str, Any]] = [
        ("Status", _g(dcf, "status")),
        ("Method", _g(dcf, "method")),
        ("Base FCF", _g(dcf, "base_fcf")),
        ("FCF basis", _g(dcf, "fcf_basis")),
        ("Stage-1 growth", _g(dcf, "stage1_growth")),
        ("Discount rate", _g(dcf, "discount_rate")),
        ("Terminal growth", _g(dcf, "terminal_growth")),
        ("Horizon (years)", _g(dcf, "stage1_years")),
        ("Net debt", _g(dcf, "net_debt")),
        ("Shares", _g(dcf, "shares")),
        ("PV explicit", _g(dcf, "pv_explicit")),
        ("PV terminal", _g(dcf, "pv_terminal")),
        ("Terminal weight", _g(dcf, "terminal_weight")),
        ("Equity value", _g(dcf, "equity_value")),
        ("Fair value / share", _g(dcf, "fair_value_per_share")),
    ]
    pct_labels = {"Stage-1 growth", "Discount rate", "Terminal growth", "Terminal weight"}
    money_labels = {"Base FCF", "Net debt", "PV explicit", "PV terminal", "Equity value",
                    "Fair value / share"}
    r = 3
    for label, value in assum:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        vcell = ws.cell(row=r, column=2)
        if value is None:
            vcell.value = "n/m"
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            vcell.value = float(value)
            if label in pct_labels:
                vcell.number_format = _PCT_FMT
            elif label in money_labels:
                vcell.number_format = _MONEY_FMT
        else:
            vcell.value = str(value)
        r += 1

    # projection table below the assumptions
    proj = _g(dcf, "projection", []) or []
    if proj:
        head_row = r + 1
        _write_header(ws, ["Year", "Projected FCF", "Discounted FCF"], row=head_row)
        ws.freeze_panes = None  # a mid-sheet projection shouldn't freeze the pane
        for i, tup in enumerate(proj, start=1):
            year, fcf_t, disc = tup
            ws.cell(row=head_row + i, column=1, value=int(year))
            c2 = ws.cell(row=head_row + i, column=2, value=float(fcf_t))
            c2.number_format = _MONEY_FMT
            c3 = ws.cell(row=head_row + i, column=3, value=float(disc))
            c3.number_format = _MONEY_FMT
    if reverse is not None:
        rr = (r + len(proj) + 4) if proj else (r + 2)
        ws.cell(row=rr, column=1, value="Reverse-DCF implied growth").font = _LABEL_FONT
        ig = _g(reverse, "implied_growth")
        rc = ws.cell(row=rr, column=2)
        if ig is None:
            rc.value = "n/m"
        else:
            rc.value = float(ig)
            rc.number_format = _PCT_FMT
        ws.cell(row=rr + 1, column=1, value="Historical FCF CAGR").font = _LABEL_FONT
        hc = _g(reverse, "historical_fcf_cagr")
        hcell = ws.cell(row=rr + 1, column=2)
        if hc is None:
            hcell.value = "n/m"
        else:
            hcell.value = float(hc)
            hcell.number_format = _PCT_FMT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # --- Multiples ---
    ws = wb.create_sheet("Multiples")
    _write_header(ws, ["Multiple", "Company", "Peer median", "Implied upside"], row=1)
    multiples = _g(val, "multiples", {}) or {}
    peer_rel = _g(val, "peer_relative", {}) or {}
    row_i = 2
    for key, m in multiples.items():
        label = _g(m, "label", key)
        mval = _g(m, "value")
        is_yield = str(key).endswith("_yield")
        ws.cell(row=row_i, column=1, value=label)
        c2 = ws.cell(row=row_i, column=2)
        if mval is None:
            c2.value = "n/m"
        else:
            c2.value = float(mval)
            c2.number_format = _PCT_FMT if is_yield else _MULT_FMT
        pr = peer_rel.get(key)
        pm = _g(pr, "peer_median")
        up = _g(pr, "upside_pct")
        c3 = ws.cell(row=row_i, column=3)
        c3.value = float(pm) if pm is not None else "n/m"
        if pm is not None:
            c3.number_format = _MULT_FMT
        c4 = ws.cell(row=row_i, column=4)
        if up is None:
            c4.value = "n/m"
        else:
            c4.value = float(up)
            c4.number_format = _PCT_FMT
        row_i += 1
    if not multiples:
        ws.cell(row=2, column=1, value="No market multiples (no price / intrinsic-only view).")
    _autosize(ws)

    # --- Quality & Forensics ---
    ws = wb.create_sheet("Quality & Forensics")
    _write_header(ws, ["Factor", "Value", "Zone / note"], row=1)
    row_i = 2
    factors = _g(quality, "factors", {}) or {}
    for _key, f in factors.items():
        if _g(f, "status") != "OK":
            continue
        ws.cell(row=row_i, column=1, value=_g(f, "label", _key))
        fv = _g(f, "value")
        ws.cell(row=row_i, column=2, value=float(fv) if fv is not None else "n/m").number_format = _NUM_FMT
        detail = _g(f, "detail", {}) or {}
        ws.cell(row=row_i, column=3, value=str(detail.get("zone", "")))
        row_i += 1
    # forensic panel
    row_i += 1
    ws.cell(row=row_i, column=1, value="Forensic screens (NOT accusations)").font = _LABEL_FONT
    row_i += 1
    if beneish is not None and _g(beneish, "m_score") is not None:
        ws.cell(row=row_i, column=1, value="Beneish M-score")
        ws.cell(row=row_i, column=2, value=float(_g(beneish, "m_score"))).number_format = _NUM_FMT
        ws.cell(row=row_i, column=3,
                value="possible manipulation (> -1.78)" if _g(beneish, "flag") else "below threshold")
        row_i += 1
    for key, reason, value in (_g(stress, "stress_flags", []) or []):
        ws.cell(row=row_i, column=1, value=str(key))
        ws.cell(row=row_i, column=2, value=float(value) if value is not None else "n/m")
        ws.cell(row=row_i, column=3, value=str(reason))
        row_i += 1
    _autosize(ws)

    _readme_sheet(wb, f"stock-helper valuation — {ticker}",
                  {"ticker": ticker, "bucket": bucket, "as_of": _g(val, "as_of"),
                   "source_label": "Stooq daily CSV (non-canonical)" if _g(val, "price") else None})
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def valuation_to_csv(val: Any) -> bytes:
    """A flat one-row CSV summary of a valuation (headline figures)."""
    row = {
        "ticker": _g(val, "ticker"),
        "bucket": _g(val, "bucket"),
        "price": _g(val, "price"),
        "fair_value_per_share": _g(val, "fair_value_per_share"),
        "margin_of_safety": _g(val, "margin_of_safety"),
        "implied_growth": _g(val, "implied_growth"),
        "flags": ", ".join(_g(val, "flags", []) or []),
    }
    return pd.DataFrame([row]).to_csv(index=False).encode("utf-8")
