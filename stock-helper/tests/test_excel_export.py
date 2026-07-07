"""Light checks for the Excel/CSV exporters: they return non-empty bytes with
the expected tabs, degrade gracefully on partial/None inputs, and the CSV
round-trips through pandas."""

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from stock_helper.reports import excel
from stock_helper.valuation.dcf import dcf_value
from stock_helper.valuation.types import ValuationResult


def _sheet_names(data: bytes) -> list[str]:
    return load_workbook(BytesIO(data)).sheetnames


def _screen_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Ticker": "AAA", "Sector": "industrial", "Market cap": 4.2e11,
             "FV gap": 0.18, "Value z": 1.4, "Quality z": 0.9, "EV/EBIT": 12.3,
             "Flags": "deep_value"},
            {"Ticker": "BBB", "Sector": "tech", "Market cap": None,
             "FV gap": None, "Value z": -0.3, "Quality z": 2.1, "EV/EBIT": None,
             "Flags": ""},
        ]
    )


def test_screen_to_xlsx_has_expected_tabs():
    data = excel.screen_to_xlsx(_screen_df(), {"as_of": None, "screen": "deep value"})
    assert isinstance(data, bytes) and len(data) > 0
    assert _sheet_names(data) == ["Screen", "README"]


def test_screen_to_xlsx_empty_df_is_friendly():
    data = excel.screen_to_xlsx(pd.DataFrame(), {})
    assert len(data) > 0
    wb = load_workbook(BytesIO(data))
    assert "Screen" in wb.sheetnames and "README" in wb.sheetnames


def test_screen_to_csv_round_trips():
    df = _screen_df()
    csv_bytes = excel.screen_to_csv(df)
    assert isinstance(csv_bytes, bytes) and csv_bytes
    back = pd.read_csv(BytesIO(csv_bytes))
    assert list(back["Ticker"]) == ["AAA", "BBB"]
    assert back.loc[0, "Value z"] == 1.4


def _valuation_result() -> ValuationResult:
    """A partial-but-real ValuationResult: a computed intrinsic DCF, no price,
    empty relative/quality — exercises graceful degradation."""
    dcf = dcf_value(
        fcf_0=1.0e10, stage1_growth=0.06, stage1_years=5, discount_rate=0.09,
        terminal_growth=0.025, shares=1.5e9,
    )
    return ValuationResult(
        ticker="AAA", bucket="industrial", as_of=None, price=None, price_date=None,
        dcf=dcf, fair_value_per_share=dcf.fair_value_per_share,
        flags=["beneish_manipulation_flag"],
        caveats=["Research aid — not advice."],
    )


def test_valuation_to_xlsx_has_expected_tabs():
    data = excel.valuation_to_xlsx(_valuation_result())
    assert isinstance(data, bytes) and len(data) > 0
    names = _sheet_names(data)
    for expected in ("Summary", "DCF", "Multiples", "Quality & Forensics", "README"):
        assert expected in names


def test_valuation_to_xlsx_survives_none():
    # Fully empty result (no sub-results at all) must still export.
    empty = ValuationResult(
        ticker="ZZZ", bucket="industrial", as_of=date(2020, 1, 1),
        price=None, price_date=None,
    )
    data = excel.valuation_to_xlsx(empty)
    assert len(data) > 0
    assert "Summary" in _sheet_names(data)


def test_valuation_to_csv_round_trips():
    csv_bytes = excel.valuation_to_csv(_valuation_result())
    back = pd.read_csv(BytesIO(csv_bytes))
    assert back.loc[0, "ticker"] == "AAA"


# --- fair-value bar (theme) ---------------------------------------------------

def test_fair_value_bar_none_safe():
    from stock_helper.ui import theme

    # No fair value -> a friendly placeholder figure, no crash.
    fig = theme.fair_value_bar(None, 100.0)
    assert fig is not None
    # Fair value but no price -> band + fair line, no marker, no crash.
    fig2 = theme.fair_value_bar(50.0, None, uncertainty="high")
    assert fig2 is not None


def test_uncertainty_from_dcf_tiers():
    from stock_helper.ui import theme

    class _D:
        def __init__(self, tw):
            self.terminal_weight = tw

    assert theme.uncertainty_from_dcf(_D(0.4)) == "low"
    assert theme.uncertainty_from_dcf(_D(0.6)) == "medium"
    assert theme.uncertainty_from_dcf(_D(0.7)) == "high"
    assert theme.uncertainty_from_dcf(_D(0.9)) == "very_high"
    assert theme.uncertainty_from_dcf(None) == "high"  # unknown -> conservative
    # Higher uncertainty demands a wider margin-of-safety band.
    assert theme.UNCERTAINTY_DISCOUNT["very_high"] > theme.UNCERTAINTY_DISCOUNT["low"]
