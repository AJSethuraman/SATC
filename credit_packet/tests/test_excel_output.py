import pytest
pytest.importorskip('openpyxl')
from openpyxl import load_workbook
from credit_packet.models import CompanyIdentity, ResearchPacket, FilingRecord, FinancialPeriod, CalculatedMetric, WatchlistFlag, Excerpt, FilingChange, SourceBoundBrief, BriefPoint, ReviewQuestionItem
from credit_packet.excel_render import render_excel
from credit_packet.render import render_markdown

PROHIBITED=['recommend approval','recommend decline','approved for credit','declined for credit','buy rating','sell rating','hold rating','investment recommendation','credit rating:','risk rating:','should lend','should not lend','safe investment','good credit risk','bad credit risk']


def sample_packet():
    return ResearchPacket(company=CompanyIdentity('AAPL','0000320193','Apple Inc.'),
        filings=[FilingRecord(accession_number='0001',form='10-K',filing_date='2025-10-31',report_date='2025-09-30',primary_document='a10k.htm',filing_url='https://example.com/10k',source='s'),FilingRecord(accession_number='0002',form='10-Q',filing_date='2025-05-01',report_date='2025-03-31',primary_document='a10q.htm',filing_url='https://example.com/10q',source='s')],
        financial_periods=[FinancialPeriod(fiscal_year=2024,period='FY',revenue=100,gross_profit=40,operating_income=30,net_income=20,cash_and_equivalents=50,current_assets=60,current_liabilities=30,total_assets=300,total_liabilities=180,stockholders_equity=120,long_term_debt=70,operating_cash_flow=-25,capex=10,tag_map={'revenue':'Revenues'})],
        calculated_metrics=[CalculatedMetric(fiscal_year=2024,revenue_growth=0.1,gross_margin=0.4,operating_margin=0.3,net_margin=0.2,cash_change_pct=-0.1,current_ratio=0.8,debt_change_pct=0.05,debt_to_equity=0.58,operating_cash_flow_margin=-0.25,free_cash_flow=-15,free_cash_flow_margin=-0.15)],
        watchlist_flags=[WatchlistFlag(code='NEGATIVE_FREE_CASH_FLOW',severity='high',description='desc',metric='m',threshold='<0',observed_value='-15',period='2024',source='metrics')],
        excerpts=[Excerpt(filing='10-K',section='Risk Factors',category='liquidity',text='x'*700,matched_keywords=['liquidity'],source_url='https://example.com/ex',accession_number='0001',filing_date='2025-10-31')],
        filing_changes=[FilingChange(section='Risk Factors',category='liquidity',old_excerpt='o'*500,new_excerpt='n'*500,change_type='modified',similarity_score=0.7,source_old='https://example.com/old',source_new='https://example.com/new')],
        review_questions=['What explains liquidity pressure?'],
        audit_log=['llm_provider:none','source_bound_validation:fallback'],
        source_bound_brief=SourceBoundBrief(summary_points=[BriefPoint(text='Flag detected',sources=['flag:1'])],review_questions=[ReviewQuestionItem(question='Q1',based_on=['flag:1'])],generation_mode='deterministic_fallback',validation_status='fallback')
    )


def test_workbook_renders_and_opens(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); wb=load_workbook(out); assert wb is not None

def test_visible_sheet_order(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); wb=load_workbook(out)
    assert wb.sheetnames==['Summary','Filing Activity','Financial Trends','Calculated Metrics','Watchlist Flags','Source-Bound Brief','Excerpts','Filing Changes','Review Questions','Memo Shell','Evidence Index','Sources & Audit']

def test_summary_and_memo_have_safety_phrase(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); wb=load_workbook(out)
    safety='Manual review required. No automated credit conclusion generated.'
    assert safety in ' '.join(str(c.value) for row in wb['Summary'].iter_rows(values_only=False) for c in row if c.value)
    assert safety in ' '.join(str(c.value) for row in wb['Memo Shell'].iter_rows(values_only=False) for c in row if c.value)

def test_tabular_sheets_have_tables_and_freeze(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); wb=load_workbook(out)
    tabs=['Filing Activity','Financial Trends','Calculated Metrics','Watchlist Flags','Source-Bound Brief','Excerpts','Filing Changes','Review Questions','Evidence Index']
    for t in tabs:
        ws=wb[t]; assert ws.freeze_panes=='A2'; assert len(ws.tables)>=1

def test_filing_activity_hyperlinks(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); ws=load_workbook(out)['Filing Activity']; assert ws['F2'].hyperlink is not None

def test_excerpts_preview_capped_and_full_preserved_hidden(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); ws=load_workbook(out)['Excerpts']
    assert len(ws['F2'].value)<=500 and len(ws['G2'].value)>=700 and ws.column_dimensions['G'].hidden

def test_changes_preview_capped_and_full_preserved_hidden(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); ws=load_workbook(out)['Filing Changes']
    assert len(ws['E2'].value)<=350 and len(ws['F2'].value)<=350 and len(ws['G2'].value)>=500 and len(ws['H2'].value)>=500
    assert ws.column_dimensions['G'].hidden and ws.column_dimensions['H'].hidden

def test_sources_audit_sections_and_field_audit(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); ws=load_workbook(out)['Sources & Audit']
    vals=' '.join(str(c.value) for row in ws.iter_rows(values_only=False) for c in row if c.value)
    for lab in ['Run Metadata','Source Documents','Field Audit','Data Quality']:
        assert lab in vals
    assert 'revenue' in vals and 'Revenues' in vals and 'source_bound_validation' in vals

def test_source_bound_brief_status_visible(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); ws=load_workbook(out)['Source-Bound Brief']
    assert ws['D2'].value=='deterministic_fallback' and ws['E2'].value=='fallback'

def test_visible_text_has_no_prohibited_phrases(tmp_path):
    out=tmp_path/'p.xlsx'; render_excel(sample_packet(),out); wb=load_workbook(out)
    all_text=[]
    for s in wb.sheetnames:
        for row in wb[s].iter_rows(values_only=True):
            for v in row:
                if isinstance(v,str): all_text.append(v.lower())
    text=' '.join(all_text)
    safety='manual review required. no automated credit conclusion generated.'
    for p in PROHIBITED:
        assert p not in text.replace(safety,'')

def test_markdown_output_still_works():
    assert 'Credit Research Packet' in render_markdown(sample_packet())
