from __future__ import annotations
from datetime import datetime, timezone
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from .excel_schema import *
from .excel_view_model import *

NAVY='1F3A5F'; WARN='FCE5CD'; DANGER='F4CCCC'; MISSING='EFEFEF'


def register_styles(_wb):
    return {
        'header_fill': PatternFill('solid', fgColor=NAVY),
        'header_font': Font(color='FFFFFF', bold=True),
        'thin': Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')),
    }

def apply_header_style(ws, styles):
    for c in ws[1]: c.fill=styles['header_fill']; c.font=styles['header_font']; c.border=styles['thin']

def apply_column_widths(ws, spec):
    for i,col in enumerate(spec.columns,1):
        letter=ws.cell(1,i).column_letter
        ws.column_dimensions[letter].width=col.width
        ws.column_dimensions[letter].hidden=col.hidden

def apply_number_formats(ws, spec):
    fmt={'money':'$#,##0','pct':'0.0%','ratio':'0.00x','decimal':'0.00','date':'yyyy-mm-dd','int':'0'}
    for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for i,col in enumerate(spec.columns,1):
            c=r[i-1]
            if col.kind in fmt and c.value not in (None,''): c.number_format=fmt[col.kind]
            if col.wrap: c.alignment=Alignment(wrap_text=True,vertical='top')

def apply_conditional_formatting(ws, sheet_name):
    if ws.max_row < 2:
        return
    if sheet_name=='Financial Trends':
        ws.conditional_formatting.add(f"M2:O{ws.max_row}", CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False, fill=PatternFill('solid', fgColor=DANGER)))
    if sheet_name=='Calculated Metrics':
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False, fill=PatternFill('solid', fgColor=DANGER)))
        ws.conditional_formatting.add(f"G2:G{ws.max_row}", CellIsRule(operator='lessThan', formula=['1'], stopIfTrue=False, fill=PatternFill('solid', fgColor=WARN)))
    if sheet_name=='Watchlist Flags':
        ws.conditional_formatting.add(f"A2:A{ws.max_row}", FormulaRule(formula=['$A2="high"'], fill=PatternFill('solid', fgColor=DANGER)))
        ws.conditional_formatting.add(f"A2:A{ws.max_row}", FormulaRule(formula=['$A2="medium"'], fill=PatternFill('solid', fgColor=WARN)))

def add_excel_table(ws, table_name, ref):
    tab=Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False)
    ws.add_table(tab)

def write_source_link(cell, label, target):
    cell.value=label
    if target: cell.hyperlink=target

def write_table_sheet(wb, spec, rows, styles):
    ws=wb.create_sheet(spec.name)
    ws.append([c.header for c in spec.columns])
    if not rows:
        rows=[{c.key:'' for c in spec.columns}]
    for row in rows:
        ws.append([row.get(c.key) for c in spec.columns])
    apply_header_style(ws, styles)
    apply_column_widths(ws, spec)
    apply_number_formats(ws, spec)
    for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for c in r: c.border=styles['thin']
    ws.freeze_panes='A2'
    if ws.max_row>=1 and spec.table_name:
        ref=f"A1:{ws.cell(1,len(spec.columns)).column_letter}{ws.max_row}"
        add_excel_table(ws,spec.table_name,ref)
    apply_conditional_formatting(ws, spec.name)
    return ws

def write_summary_sheet(wb, packet, model):
    ws=wb.create_sheet('Summary')
    latest_10k=next((f.filing_date for f in packet.filings if f.form=='10-K'),'Unavailable')
    latest_10q=next((f.filing_date for f in packet.filings if f.form=='10-Q'),'Unavailable')
    llm_mode=next((a.split(':',1)[1] for a in packet.audit_log if a.startswith('llm_provider:')),'none')
    sb_status=next((a.split(':',1)[1] for a in packet.audit_log if a.startswith('source_bound_validation:')),'unknown')
    high_count=sum(1 for f in packet.watchlist_flags if f.severity=='high')
    rows=[('Credit Research Packet',''),('Company',packet.company.name),('Ticker',packet.company.ticker),('CIK',packet.company.cik),('Generated',datetime.now(timezone.utc).isoformat()),('LLM Mode',llm_mode),('Source-Bound Validation',sb_status),('Latest 10-K',latest_10k),('Latest 10-Q',latest_10q),('Recent 8-K Count',sum(1 for f in packet.filings if f.form=='8-K')),('Watchlist Flag Count',len(packet.watchlist_flags)),('High Flag Count',high_count),('Excerpt Count',len(packet.excerpts)),('Filing Change Count',len(packet.filing_changes)),('Manual Credit Conclusion','Manual review required. No automated credit conclusion generated.')]
    for r in rows: ws.append(r)
    ws.append(('How to use this workbook','Review Filing Activity, Financial Trends, Calculated Metrics, Watchlist Flags, Excerpts, Filing Changes, Review Questions, Memo Shell, and Sources & Audit.'))
    ws.append(('Navigate','Filing Activity'))
    ws["B17"].hyperlink = "#'Filing Activity'!A1"
    ws["B17"].style = "Hyperlink"
    ws.append(('', 'Sources & Audit'))
    ws["B18"].hyperlink = "#'Sources & Audit'!A1"
    ws["B18"].style = "Hyperlink"
    dq=data_quality_rows(packet)
    ws.append(('Top data-quality notes','; '.join(x['issue'] for x in dq[:3]) if dq else 'None'))
    ws.column_dimensions['A'].width=36; ws.column_dimensions['B'].width=120
    for c in ws['A']: c.font=Font(bold=True)

def write_memo_shell(wb, packet):
    ws=wb.create_sheet('Memo Shell')
    for s in ['Company Overview','Recent Filing Activity','Financial Trend Summary','Rule-Based Watchlist Flags','Notable Filing Language','Open Questions','Manual Credit Conclusion']:
        ws.append([s]); ws.cell(ws.max_row,1).font=Font(bold=True); ws.append([''])
    ws.append(['Manual review required. No automated credit conclusion generated.'])
    ws.column_dimensions['A'].width=120

def write_sources_audit_sheet(wb, packet, generated_ts, styles):
    ws=wb.create_sheet('Sources & Audit')
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=90; ws.column_dimensions['C'].width=30

    def section(title, headers, rows):
        ws.append([title]); ws.cell(ws.max_row,1).font=Font(bold=True,color='FFFFFF'); ws.cell(ws.max_row,1).fill=PatternFill('solid',fgColor=NAVY)
        ws.append(headers)
        hdr=ws.max_row
        for c in ws[hdr]: c.font=Font(bold=True); c.fill=PatternFill('solid',fgColor='D9E1F2')
        if not rows: rows=[['','','']]
        for r in rows: ws.append(r)

    section('Run Metadata',['Key','Value',''],[[r['key'],r['value'],''] for r in run_metadata_rows(packet, generated_ts)] + [[a,'',''] for a in packet.audit_log])
    docs=[[r['label'],'Open filing',r['accession']] for r in source_document_rows(packet)]
    section('Source Documents',['Document','Source Link','Accession'],docs)
    # hyperlink docs
    start=ws.max_row-len(docs)+1 if docs else ws.max_row
    src_rows=source_document_rows(packet)
    for i,r in enumerate(src_rows, start): ws.cell(i,2).hyperlink=r['url']
    fa=[[r['fiscal_year'],r['field'],r['selected_tag']] for r in field_audit_rows(packet)]
    section('Field Audit',['Fiscal Year','Field','Selected Tag'],fa)
    dq=[[r['issue'],r['detail'],''] for r in data_quality_rows(packet)]
    section('Data Quality',['Issue','Detail',''],dq)


def render_excel(packet, output_path: Path, *, preview_chars: int = 500) -> None:
    from openpyxl import Workbook
    wb=Workbook(); wb.remove(wb.active)
    styles=register_styles(wb)

    write_summary_sheet(wb, packet, None)

    ws=write_table_sheet(wb,FILING_ACTIVITY_SPEC,filing_activity_rows(packet),styles)
    for r in range(2,ws.max_row+1): write_source_link(ws.cell(r,6),'Open filing',ws.cell(r,7).value)

    write_table_sheet(wb,FINANCIAL_TRENDS_SPEC,financial_trend_rows(packet),styles)
    write_table_sheet(wb,CALCULATED_METRICS_SPEC,calculated_metric_rows(packet),styles)
    write_table_sheet(wb,WATCHLIST_FLAGS_SPEC,watchlist_flag_rows(packet),styles)

    sb_rows=[]
    brief=packet.source_bound_brief
    if brief:
        for p in brief.summary_points: sb_rows.append({'type':'summary_point','text':p.text,'sources':', '.join(p.sources),'generation_mode':brief.generation_mode,'validation_status':brief.validation_status})
        for t in brief.review_themes: sb_rows.append({'type':'review_theme','text':f"{t.theme}: {t.why_it_matters}",'sources':', '.join(t.sources),'generation_mode':brief.generation_mode,'validation_status':brief.validation_status})
        for m in brief.missing_information: sb_rows.append({'type':'missing_information','text':f"{m.item}: {m.reason}",'sources':'','generation_mode':brief.generation_mode,'validation_status':brief.validation_status})
    write_table_sheet(wb,SOURCE_BOUND_BRIEF_SPEC,sb_rows or [{'type':'info','text':'No source-bound brief items available','sources':'','generation_mode':'none','validation_status':'unknown'}],styles)

    ws=write_table_sheet(wb,EXCERPTS_SPEC,excerpt_rows(packet,preview_chars=preview_chars),styles)
    for r in range(2,ws.max_row+1): write_source_link(ws.cell(r,8),ws.cell(r,8).value or 'Open source',ws.cell(r,9).value)

    ws=write_table_sheet(wb,FILING_CHANGES_SPEC,filing_change_rows(packet,preview_chars=350),styles)
    for r in range(2,ws.max_row+1):
        write_source_link(ws.cell(r,9),ws.cell(r,9).value or 'Open old source',ws.cell(r,11).value)
        write_source_link(ws.cell(r,10),ws.cell(r,10).value or 'Open new source',ws.cell(r,12).value)

    write_table_sheet(wb,REVIEW_QUESTIONS_SPEC,review_question_rows(packet),styles)
    write_memo_shell(wb,packet)
    write_table_sheet(wb,EVIDENCE_INDEX_SPEC,evidence_index_rows(packet),styles)
    generated_ts=datetime.now(timezone.utc).isoformat()
    write_sources_audit_sheet(wb,packet,generated_ts,styles)

    order=['Summary','Filing Activity','Financial Trends','Calculated Metrics','Watchlist Flags','Source-Bound Brief','Excerpts','Filing Changes','Review Questions','Memo Shell','Evidence Index','Sources & Audit']
    wb._sheets=[wb[s] for s in order]
    output_path.parent.mkdir(parents=True,exist_ok=True)
    wb.save(output_path)
