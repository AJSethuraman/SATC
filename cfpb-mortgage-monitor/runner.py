#!/usr/bin/env python3
"""County Mortgage Delinquency Monitor -- runner (the data path).

Single source of truth for the data path; embedded verbatim into the
workbook's _code_py tab. The VBA ExtractFiles macro writes this back out as
runner.py and the user runs it from PowerShell against the CLOSED workbook
(openpyxl backend).

Built to BUILD_SPEC_CFPB.md. THE SLOT MECHANISM (the FDIC template's [PEERS]
pattern, carried wholesale): the hand-picked _config [FOOTPRINT] list IS the
watchlist, keyed by 5-digit county FIPS, and [SERIES] carries the national
rows, the state dashboard rows (st_{FIPS2}_{measure}, slot-provisioned) and
the 2-measure county TEMPLATE rows expanded at run time into units
co_{slot:02d}_{measure}. Raw anchors depend ONLY on (slot, measure), so
editing WHICH county occupies a slot moves no formula -- add/remove =
edit a [FOOTPRINT] line + re-run, never a rebuild (within built capacity).

Clean seams (BUILD SPEC 0.3 / 1):
  * ADAPTER SEAM -- fetch_series(spec, secret) -> list[NormalizedRow], FIXED
    schema. Two Class A providers: CfpbProvider (live, plain urllib, KEYLESS,
    TWO-STAGE: scrape the download page for the CURRENT dated filenames --
    trap C2 -- then fetch 4 CSVs) + CfpbDemoProvider (offline deterministic,
    all tests). Wide CSVs are unpivoted by the ^\\d{4}-\\d{2}$ header pattern
    (trap C5), quoted FIPS stripped and kept as TEXT (trap C1), the National
    row filtered by RegionType (trap C7).
  * TRANSFORM REGISTRY -- pure functions (level, yoy_pct, dev_12m,
    rise_streak3) defined IDENTICALLY here and in the Excel formulas
    (parity-tested). Blank-tolerant: missing months blank, never zero.
  * THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT, direction-aware.
  * WATCHLIST VALIDATOR -- DEFAULT-DENY (BUILD SPEC 0.1): only rows carrying
    a genuine county key (county:NNNNN) are admitted; blank/malformed FIPS
    are refused BY NAME with a --lookup hint; states/metros/national are
    dashboard context, NEVER watchlist.
  * SUPPRESSION (BUILD SPEC 0.6, trap C4) -- a footprint county ABSENT from
    the fetched file is SUPPRESSED (below the ~1,000-sample-mortgage
    inclusion threshold), rendered explicitly, excluded from alert counts --
    never blank-and-silent, never zero.
  * VINTAGE / CONTINUITY (BUILD SPEC 0.6, traps C3/C6) -- every vintage
    revises the FULL history, so every run is a stateless FULL REPLACE; the
    vintage (thru-YYYY-MM) is recorded in the status panel, and a vintage
    older than continuity_months behind --asof raises a CONTINUITY WARNING
    (the CFPB publishes informally; it survived the 2025 turmoil once).

No AI/LLM anywhere; every computation is deterministic (BUILD SPEC 0.5).
Pure ASCII (L3).
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd  # noqa: F401  (contract dependency; openpyxl does the IO)

# Fixed raw-block geometry: ONE block per GEO SLOT in Raw_CFPB -- the
# national block, then 15 state slots, then the built county slots. Months
# as rows (newest first), the two measures as columns. Anchors depend only
# on (kind, slot, raw_slots) -- never on which county occupies the slot --
# so dashboard formulas survive any [FOOTPRINT] edit.
RAW_TAB = "Raw_CFPB"
RAW_SLOTS_DEFAULT = 72             # months kept per geo (6 years, newest-first)
FOOTPRINT_SLOTS_DEFAULT = 40       # BUILT county capacity (--footprint-slots)
STATE_SLOTS = 15                   # provisioned state dashboard slots (fixed)
RAW_HEADER_ROWS = 2
RAW_GAP_ROWS = 2
RAW_FIRST_ROW = 2
RAW_PERIOD_COL = 1                 # column A = month (YYYY-MM, text)

# Status panel column on each dashboard (free of any merged masthead cells).
STATUS_COL = 12                    # column L

DASH_TABS = ("Dashboard_State", "Dashboard_Trends")

MEASURES = ("d3089", "d90")
MEASURE_FILE = {"d3089": "30-89", "d90": "90-plus"}
MEASURE_COL = {"d3089": 2, "d90": 3}     # Raw_CFPB columns B / C

# The watchlist join-key whitelist (BUILD SPEC 0.1): default-deny. Only a
# literal 5-digit county FIPS key passes; states, metros, national and
# name-only rows are refused (metro/CBSA promotion is a spec change, not a
# config edit -- the metro file mixes non-FIPS keys).
ENTITY_KEY_PATTERN = r"^county:[0-9]{5}$"

MONTH_RE = re.compile(r"^\d{4}-\d{2}$")
VINTAGE_RE = re.compile(r"^(?:thru-)?(\d{4})-(\d{2})$")

# Stage-1 discovery (trap C2: filenames are DATED; hardcoded URLs die every
# release). The CFPB's own ecosystem scrapes the download page -- so do we.
FILENAME_RE = re.compile(
    r"(State|County)MortgagesPercent-(30-89|90-plus)DaysLate-"
    r"thru-(\d{4}-\d{2})\.csv")
DOWNLOAD_PAGE = ("https://www.consumerfinance.gov/data-research/"
                 "mortgage-performance-trends/download-the-data/")
FILES_BASE = ("https://files.consumerfinance.gov/data/"
              "mortgage-performance/downloads/")

THRESHOLD_IDS = ("d3089_dev12m", "d90_streak3", "d90_level")

SUPPRESSED_TEXT = "SUPPRESSED (below sample threshold) -- use the state row"


def slot_label(kind: str, slot: int) -> str:
    """Raw block label: 'nat', 'st01'..'st15', 'co01'..'coNN'."""
    if kind == "national":
        return "nat"
    return ("st" if kind == "state" else "co") + f"{slot:02d}"


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` tab; backend-agnostic)
# ---------------------------------------------------------------------------
SERIES_HEADER = [
    "id", "title", "category", "lane", "metric_type", "frequency", "sa_nsa",
    "units", "level_rate_index", "geo_segment", "source_class",
    "dashboard_capable", "watchlist_capable", "source_url", "table_id", "sheet",
    "series_label", "transform", "notes",
]

FOOTPRINT_HEADER = ["slot", "fips", "name", "state", "active"]


@dataclass
class SeriesSpec:
    id: str
    title: str
    category: str
    lane: str
    metric_type: str
    frequency: str
    sa_nsa: str
    units: str
    level_rate_index: str
    geo_segment: str
    source_class: str
    dashboard_capable: bool
    watchlist_capable: bool
    source_url: str
    table_id: str
    sheet: str
    series_label: str
    transform: str
    notes: str


@dataclass
class FootprintRow:
    """One [FOOTPRINT] line: the flexible county list. An empty row (slot
    number only) is provisioned headroom the user fills by hand."""
    slot: Optional[int]
    fips: str              # normalized 5-digit text; "" when blank
    name: str
    state: str
    active: bool

    @property
    def has_county(self) -> bool:
        return bool(self.fips or self.name)

    @property
    def entity_key(self) -> str:
        """The join key the gate whitelists: county:NNNNN from the fips cell."""
        return f"county:{self.fips}"


@dataclass
class Threshold:
    watch: Optional[float]
    alert: Optional[float]
    direction: str          # "above" | "below"


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, Threshold] = field(default_factory=dict)
    series: List[SeriesSpec] = field(default_factory=list)
    footprint: List[FootprintRow] = field(default_factory=list)

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))

    @property
    def footprint_slots(self) -> int:
        """The BUILT county slot capacity. Informational here -- the
        raw-layout check (block labels) is the hard guard against a stale
        value."""
        return int(float(self.settings.get("footprint_slots",
                                           FOOTPRINT_SLOTS_DEFAULT)))

    @property
    def continuity_months(self) -> int:
        v = _as_float(self.settings.get("continuity_months", "9"))
        return 9 if v is None else int(v)

    @property
    def vintage_override(self) -> str:
        return str(self.settings.get("vintage", "") or "").strip()


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _as_float(v):
    try:
        s = str(v).strip()
        return float(s) if s != "" else None
    except (TypeError, ValueError):
        return None


def _norm_fips(v, width: int = 5) -> str:
    """Normalize a [FOOTPRINT] fips cell to 5-digit TEXT (trap C1). Excel
    hands numeric cells over as int/float (1003 or 1003.0 for a cell the
    user typed without the leading zero) -- zero-pad those back. Text stays
    as typed (minus a stray quote wrapper), so a malformed value ('ABC12',
    '060370') reaches the gate and is refused, never silently coerced."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, int):
        s = str(v)
        return s.zfill(width) if len(s) <= width else s
    return str(v).strip().strip("'")


def _norm_slot(v) -> Optional[int]:
    try:
        s = str(v).strip()
        return int(float(s)) if s != "" else None
    except (TypeError, ValueError):
        return None


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists). Sections in col
    A: [SETTINGS], [THRESHOLDS], [FOOTPRINT], [SERIES]. The [FOOTPRINT]
    section is this template's addition (the FDIC [PEERS] pattern); the
    other three parse contract-identically."""
    cfg = Config()
    section = None
    series_header = None
    thr_header = None
    fp_header = None
    for raw in rows:
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            series_header = None
            thr_header = None
            fp_header = None
            continue
        if not a:
            continue
        if a.startswith("#"):          # in-sheet comment line, never data
            continue
        if section == "SETTINGS":
            if a.lower() in ("key", "name"):
                continue
            val = "" if len(raw) < 2 or raw[1] is None else raw[1]
            cfg.settings[a] = str(val).strip()
        elif section == "THRESHOLDS":
            if thr_header is None and a.lower() == "id":
                thr_header = [str(c).strip().lower() for c in raw]
                continue
            cells = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                     for i, h in enumerate(thr_header or
                                           ["id", "watch", "alert", "direction"])}
            cfg.thresholds[a] = Threshold(
                watch=_as_float(cells.get("watch")),
                alert=_as_float(cells.get("alert")),
                direction=str(cells.get("direction", "above")).strip().lower()
                or "above")
        elif section == "FOOTPRINT":
            if fp_header is None and a.lower() == "slot":
                fp_header = [str(c).strip().lower() for c in raw]
                continue
            v = {h: (None if i >= len(raw) else raw[i])
                 for i, h in enumerate(fp_header or FOOTPRINT_HEADER)}
            cfg.footprint.append(FootprintRow(
                slot=_norm_slot(v.get("slot")),
                fips=_norm_fips(v.get("fips")),
                name=("" if v.get("name") is None else str(v.get("name")).strip()),
                state=("" if v.get("state") is None
                       else str(v.get("state")).strip().upper()),
                active=_as_bool(v.get("active"))))
        elif section == "SERIES":
            if series_header is None:
                series_header = [str(c).strip() for c in raw]
                continue
            v = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                 for i, h in enumerate(series_header)}
            cfg.series.append(SeriesSpec(
                id=str(v.get("id", "")).strip(),
                title=str(v.get("title", "")).strip(),
                category=str(v.get("category", "")).strip(),
                lane=str(v.get("lane", "")).strip().lower(),
                metric_type=str(v.get("metric_type", "")).strip(),
                frequency=str(v.get("frequency", "")).strip(),
                sa_nsa=str(v.get("sa_nsa", "")).strip(),
                units=str(v.get("units", "")).strip(),
                level_rate_index=str(v.get("level_rate_index", "")).strip().lower(),
                geo_segment=str(v.get("geo_segment", "")).strip(),
                source_class=str(v.get("source_class", "")).strip().upper(),
                dashboard_capable=_as_bool(v.get("dashboard_capable", "")),
                watchlist_capable=_as_bool(v.get("watchlist_capable", "")),
                source_url=str(v.get("source_url", "")).strip(),
                table_id=str(v.get("table_id", "")).strip(),
                sheet=str(v.get("sheet", "")).strip(),
                series_label=str(v.get("series_label", "")).strip(),
                transform=str(v.get("transform", "level")).strip().lower(),
                notes=str(v.get("notes", "")).strip()))
    return cfg


def series_measure(spec: SeriesSpec) -> str:
    """Measure id from the series id suffix (nat_d90, st_06_d3089, co_d90)."""
    return spec.id.rsplit("_", 1)[-1]


def national_series(cfg: Config) -> List[SeriesSpec]:
    return [s for s in cfg.series if s.geo_segment == "national"]


def state_series(cfg: Config) -> List[SeriesSpec]:
    return [s for s in cfg.series if s.geo_segment.startswith("state:")]


def county_templates(cfg: Config) -> List[SeriesSpec]:
    return [s for s in cfg.series if s.geo_segment == "county"]


class SeriesError(Exception):
    pass


class StateCapacityError(Exception):
    """More distinct states in [SERIES] than the 15 provisioned state
    dashboard slots."""


def validate_series(series: Sequence[SeriesSpec]) -> None:
    """Every [SERIES] row must carry a known measure suffix, a registered
    transform and a recognized geo shape -- config drift refuses, not
    guesses."""
    for s in series:
        m = series_measure(s)
        if m not in MEASURES:
            raise SeriesError(
                f"Series '{s.id}' does not end in a known measure id "
                f"({', '.join(MEASURES)}).")
        if s.transform not in TRANSFORMS:
            raise SeriesError(
                f"Series '{s.id}' declares unregistered transform "
                f"'{s.transform}'; known transforms: "
                f"{', '.join(sorted(TRANSFORMS))}.")
        g = s.geo_segment
        if not (g == "national" or g == "county"
                or re.match(r"^state:[0-9]{2}$", g)):
            raise SeriesError(
                f"Series '{s.id}' has unrecognized geo_segment '{g}' "
                "(expected national, state:XX, or county).")


def state_slot_map(cfg: Config) -> Dict[str, int]:
    """fips2 -> state slot (1-based), in [SERIES] first-appearance order.
    The workbook is provisioned for STATE_SLOTS state rows; more distinct
    states than that is REFUSED, never truncated."""
    smap: Dict[str, int] = {}
    for s in state_series(cfg):
        fips2 = s.geo_segment.split(":", 1)[1]
        if fips2 not in smap:
            smap[fips2] = len(smap) + 1
    if len(smap) > STATE_SLOTS:
        raise StateCapacityError(
            f"[SERIES] lists {len(smap)} distinct states but the workbook "
            f"is provisioned for {STATE_SLOTS} state dashboard slots. Trim "
            f"the st_* rows to the states your footprint actually spans "
            f"(the county watchlist is unaffected).")
    return smap


# ---------------------------------------------------------------------------
# TRANSFORM REGISTRY -- pure functions over NEWEST-FIRST value lists
# (v[0] = latest month). Defined IDENTICALLY in the Excel formulas
# (build_workbook.py) -- parity is machine-tested, never assumed.
# ---------------------------------------------------------------------------
def t_level(v: Sequence[Optional[float]]) -> Optional[float]:
    """Latest value. Excel: =IF(v0="","",v0)."""
    return v[0] if v else None


def t_dev_12m(v: Sequence[Optional[float]]) -> Optional[float]:
    """v0 - mean(v1..v12): deviation from own trailing-12 average, in
    percentage points. Excel: =IF(OR(v0="",COUNT(v1:v12)=0),"",
    v0-AVERAGE(v1:v12)) -- AVERAGE ignores blanks, so the Python mean
    ignores missing months identically."""
    if not v or v[0] is None:
        return None
    window = [x for x in v[1:13] if x is not None]
    if not window:
        return None
    return v[0] - sum(window) / len(window)


def t_rise_streak3(v: Sequence[Optional[float]]) -> Optional[float]:
    """Consecutive-rise count CAPPED at 3. Excel (identical, capped):
    =IF(COUNT(v0:v3)<4,"",IF(v0>v1,1,0)+IF(AND(v0>v1,v1>v2),1,0)
    +IF(AND(v0>v1,v1>v2,v2>v3),1,0))."""
    if len(v) < 4 or any(v[i] is None for i in range(4)):
        return None
    s = 0
    if v[0] > v[1]:
        s += 1
    if v[0] > v[1] and v[1] > v[2]:
        s += 1
    if v[0] > v[1] and v[1] > v[2] and v[2] > v[3]:
        s += 1
    return float(s)


def t_yoy_pct(v: Sequence[Optional[float]]) -> Optional[float]:
    """Percent change vs 12 months earlier. Excel:
    =IF(OR(v0="",v12="",v12=0),"",(v0/v12-1)*100)."""
    if len(v) < 13 or v[0] is None or v[12] is None or v[12] == 0:
        return None
    return (v[0] / v[12] - 1.0) * 100.0


TRANSFORMS = {
    "level": t_level,
    "yoy_pct": t_yoy_pct,
    "dev_12m": t_dev_12m,
    "rise_streak3": t_rise_streak3,
}


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT, direction-aware
# ---------------------------------------------------------------------------
def status_for(value: Optional[float], thr: Optional[Threshold]) -> str:
    """direction='above': flag when value >= bound; 'below': value <= bound.
    Blank value -> OK (blanks are surfaced by the suppression/continuity
    paths, never fabricated into a flag)."""
    if value is None or thr is None or (isinstance(value, float)
                                        and math.isnan(value)):
        return "OK"
    above = thr.direction != "below"

    def hit(bound):
        if bound is None:
            return False
        return value >= bound if above else value <= bound

    if hit(thr.alert):
        return "ALERT"
    if hit(thr.watch):
        return "WATCH"
    return "OK"


SEVERITY = {"OK": 0, "WATCH": 1, "ALERT": 2}


# ---------------------------------------------------------------------------
# VINTAGE / CONTINUITY (BUILD SPEC 0.6 -- traps C3/C6)
# ---------------------------------------------------------------------------
def vintage_ym(vintage: Optional[str]) -> Optional[Tuple[int, int]]:
    """Parse 'thru-2025-09' (or bare '2025-09') -> (2025, 9); None if not
    parseable (e.g. the demo provider's fabricated label is still thru-*)."""
    if not vintage:
        return None
    m = VINTAGE_RE.match(str(vintage).strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def months_behind(asof: date, vintage: Optional[str]) -> Optional[int]:
    ym = vintage_ym(vintage)
    if ym is None:
        return None
    return (asof.year * 12 + asof.month) - (ym[0] * 12 + ym[1])


def continuity_warning(asof: date, vintage: Optional[str],
                       limit_months: int) -> Optional[str]:
    """The tripwire: the CFPB publishes semiannually-ish with a ~6-7 month
    lag and no formal cadence promise; a vintage more than limit_months
    behind asof means the program may have stalled (it survived the 2025
    turmoil once -- check the download page and the DataLumos mirror)."""
    behind = months_behind(asof, vintage)
    if behind is None:
        return (f"CONTINUITY WARNING: data vintage '{vintage}' is not "
                f"parseable as thru-YYYY-MM; cannot verify freshness.")
    if behind > limit_months:
        return (f"CONTINUITY WARNING: newest vintage {vintage} is {behind} "
                f"months behind asof {asof.isoformat()} (tripwire "
                f"{limit_months}). The CFPB publishes informally -- check "
                f"the download page; the DataLumos archive is the recovery "
                f"mirror if the program has stalled.")
    return None


# ---------------------------------------------------------------------------
# THE WATCHLIST VALIDATOR -- DEFAULT-DENY (BUILD SPEC 0.1 / sec 3).
# Gate 1: the footprint slot is active=TRUE (inactive = EXCLUDED, blanked).
# Gate 2: every watchlist-capable series row is the Class A county template
#         (a state/metro/national row marked watchlist-capable refuses the
#         whole run, series-named -- aggregates are context, never watchlist).
# Gate 3: the entity key built from the [FOOTPRINT] fips cell matches
#         ^county:[0-9]{5}$ (blank/malformed FIPS = refused BY NAME).
# ---------------------------------------------------------------------------
class WatchlistRefused(Exception):
    """Raised by the hard gates; carries the interpolated refusal message."""


class FootprintCapacityError(Exception):
    """More counties than BUILT slots: refuse with the rebuild command --
    the runner never truncates a footprint."""


def county_refusal_message(fp: FootprintRow, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: footprint slot {fp.slot} '
        f'"{fp.name or "(unnamed)"}" has fips="{fp.fips}", '
        f'active={"TRUE" if fp.active else "FALSE"}. '
        + " ".join(reasons) +
        " Only 5-digit county-FIPS-keyed rows can be monitored; states,"
        " metros and national are dashboard context, never watchlist."
        " Find the FIPS with: python runner.py --lookup"
        f' "{fp.name or "<county name>"}" (works offline with --demo).')


def series_refusal_message(spec: SeriesSpec, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: series "{spec.id}" (geo_segment='
        f'"{spec.geo_segment}", source_class="{spec.source_class}") is '
        f'marked watchlist_capable={"TRUE" if spec.watchlist_capable else "FALSE"}. '
        + " ".join(reasons) +
        " This template's watchlist lane admits ONLY the Class A county"
        " template rows; state, metro and national aggregates are dashboard"
        " context by design (BUILD SPEC 0.1).")


def gate_footprint_row(fp: FootprintRow) -> List[str]:
    """Failed-gate reasons for an ACTIVE footprint row carrying a county.
    Empty list == a genuinely county-FIPS-keyed row."""
    reasons = []
    if not re.match(ENTITY_KEY_PATTERN, fp.entity_key):
        reasons.append(
            f"Gate3: entity key '{fp.entity_key}' does not match the "
            f"join-key whitelist pattern {ENTITY_KEY_PATTERN} (default-deny; "
            "a blank or malformed FIPS cannot be joined).")
    return reasons


def gate_series_row(spec: SeriesSpec) -> List[str]:
    """Reasons a watchlist-capable series row fails the lane gates."""
    reasons = []
    if spec.geo_segment != "county":
        reasons.append(
            f"Gate2a: geo_segment '{spec.geo_segment}' is not the county "
            "template -- states/metros/national are never watchlist.")
    if spec.source_class != "A":
        reasons.append(
            "Gate2b: source_class is not A -- this template's only admitted "
            "class is the keyless public CFPB file adapter; any other class "
            "requires its own review before admission.")
    return reasons


def assert_series_admissible(series: Sequence[SeriesSpec]) -> None:
    """Hard gate (build-time AND run start): a fabricated watchlist-capable
    aggregate row or non-Class-A row refuses everything, series-named --
    defense in depth behind the runtime gates."""
    for s in series:
        if not s.watchlist_capable:
            continue
        reasons = gate_series_row(s)
        if reasons:
            raise WatchlistRefused(series_refusal_message(s, reasons))


def evaluate_footprint(cfg: Config):
    """Apply the default-deny gates to every [FOOTPRINT] row carrying a
    county. Returns (admitted, refusals, excluded): admitted rows are
    fetched and landed; refusals is [(row, message)] -- those slots are
    never fetched and stay blank; excluded is the active=FALSE list
    (blanked, not refused)."""
    admitted, refusals, excluded = [], [], []
    for fp in cfg.footprint:
        if not fp.has_county:
            continue                       # provisioned empty slot
        if not fp.active:
            excluded.append(fp)            # gate 1: exclusion, not refusal
            continue
        reasons = gate_footprint_row(fp)
        if reasons:
            refusals.append((fp, county_refusal_message(fp, reasons)))
        else:
            admitted.append(fp)
    return admitted, refusals, excluded


def assert_entity_gates(cfg: Config) -> None:
    """BUILD-TIME hard gate backing the runtime gates: a bad series row or a
    bad SEED county cannot even be built into the lane. (At run time the
    per-county gate refuses-and-continues so one typo does not kill the
    whole refresh; the refusal is rendered, the slot stays blank.)"""
    assert_series_admissible(cfg.series)
    _, refusals, _ = evaluate_footprint(cfg)
    if refusals:
        raise WatchlistRefused(refusals[0][1])


def validate_footprint_capacity(cfg: Config) -> None:
    """The flexible-footprint guard: every county-carrying row needs a slot
    inside the BUILT capacity, unique. Over capacity is REFUSED with the
    rebuild command -- never truncated."""
    cap = cfg.footprint_slots
    filled = [f for f in cfg.footprint if f.has_county]
    seen: Dict[int, FootprintRow] = {}
    bad = []
    for f in filled:
        if f.slot is None or f.slot < 1 or f.slot > cap:
            bad.append(f)
        elif f.slot in seen:
            raise FootprintCapacityError(
                f"[FOOTPRINT] slot {f.slot} is listed twice "
                f'("{seen[f.slot].name}" and "{f.name}"); slots must be '
                "unique -- fix the [FOOTPRINT] table.")
        else:
            seen[f.slot] = f
    if bad:
        need = max(len(filled), max((f.slot or 0) for f in filled))
        names = ", ".join(f'"{f.name or f.fips}" (slot {f.slot})' for f in bad)
        raise FootprintCapacityError(
            f"[FOOTPRINT] lists more counties than the workbook was built "
            f"for: {names} outside the built capacity of {cap} slots. The "
            f"runner never truncates a footprint -- rebuild with "
            f"--footprint-slots {max(need, cap + 1)} "
            f"(python make_workbook.py --footprint-slots {max(need, cap + 1)}).")


# ---------------------------------------------------------------------------
# THE ADAPTER SEAM -- the only provider-specific code
# ---------------------------------------------------------------------------
@dataclass
class NormalizedRow:
    id: str
    period: str            # month "YYYY-MM" (text)
    value: Optional[float]
    geo_segment: str       # national | state:XX | county:XXXXX
    source_class: str
    units: str = "pct"


@dataclass
class UnitSpec:
    """One expanded fetch/land unit: a (geo, measure) pair. County units are
    SLOT-keyed (co_{slot:02d}_{measure}) so anchors never move on a
    [FOOTPRINT] edit; state units carry the FIPS2 (st_{FIPS2}_{measure});
    national is nat_{measure}."""
    kind: str              # "national" | "state" | "county"
    slot: int              # 0 for national; state slot; footprint slot
    fips: str              # "" | 2-digit | 5-digit (text)
    measure: str
    id: str
    geo_segment: str
    source_class: str = "A"


def make_national_spec(measure: str) -> UnitSpec:
    return UnitSpec("national", 0, "", measure, f"nat_{measure}", "national")


def make_state_spec(slot: int, fips2: str, measure: str) -> UnitSpec:
    return UnitSpec("state", slot, fips2, measure,
                    f"st_{fips2}_{measure}", f"state:{fips2}")


def make_county_spec(fp: FootprintRow, measure: str) -> UnitSpec:
    return UnitSpec("county", fp.slot, fp.fips, measure,
                    f"co_{fp.slot:02d}_{measure}", fp.entity_key)


def parse_wide_csv(text: str) -> dict:
    """Parse one CFPB wide CSV (schema CONFIRMED from the inspected file +
    the CFPB's own generator source):

        RegionType,State,Name,FIPSCode,2008-01,...,2025-09
        National,,United States,-----,3.5,...
        County,AL,Baldwin County,'01003',2.8,...

    - month columns found by the ^\\d{4}-\\d{2}$ HEADER pattern, never by
      position (trap C5: new columns appear every release);
    - the literal single quotes around FIPS are STRIPPED, FIPS kept as TEXT
      (trap C1);
    - the National row (RegionType) is its own series, filtered out of the
      geo dict (trap C7); Metro/other region types are skipped;
    - values are percent floats; empty cells are None, never zero.
    Returns {"months": [asc], "national": {month: val} | None,
             "geos": {fips: {"name","state","values"}}}."""
    rdr = csv.reader(io.StringIO(text))
    try:
        header = next(rdr)
    except StopIteration:
        raise RuntimeError("CFPB CSV is empty (no header row).")
    header = [h.strip() for h in header]
    for req in ("RegionType", "State", "Name", "FIPSCode"):
        if req not in header:
            raise RuntimeError(
                f"CFPB CSV header is missing the '{req}' column -- schema "
                f"changed? Header starts: {header[:6]}")
    col = {name: header.index(name)
           for name in ("RegionType", "State", "Name", "FIPSCode")}
    month_cols = [(i, h) for i, h in enumerate(header) if MONTH_RE.match(h)]
    if not month_cols:
        raise RuntimeError("CFPB CSV header has no YYYY-MM month columns.")
    months = [h for _i, h in month_cols]
    out = {"months": months, "national": None, "geos": {}}

    def vals(row):
        d = {}
        for i, mon in month_cols:
            cell = row[i].strip() if i < len(row) else ""
            d[mon] = float(cell) if cell != "" else None
        return d

    for row in rdr:
        if not row or all(not c.strip() for c in row):
            continue
        rtype = row[col["RegionType"]].strip()
        if rtype == "National":
            out["national"] = vals(row)          # FIPSCode is '-----'
            continue
        if rtype not in ("State", "County"):
            continue                             # Metro etc: no FIPS join
        fips = row[col["FIPSCode"]].strip().strip("'")   # trap C1
        if not fips:
            continue
        out["geos"][fips] = {
            "name": row[col["Name"]].strip(),
            "state": row[col["State"]].strip(),
            "values": vals(row),
        }
    return out


class Provider:
    """Adapter interface: prime(asof) once per run (the 4 CSV fetches), then
    fetch_series(spec, secret) -> [NormalizedRow] slices per unit. A county
    ABSENT from the primed data returns [] -- the runner renders that as
    SUPPRESSED (trap C4), never zero."""

    source_class = "A"
    vintage: Optional[str] = None          # "thru-YYYY-MM"
    vintage_source: str = ""               # discovered | override | demo

    def prime(self, asof: date) -> None:  # pragma: no cover
        raise NotImplementedError

    def fetch_series(self, spec: UnitSpec, secret=None) -> List[NormalizedRow]:  # pragma: no cover
        raise NotImplementedError


# Embedded mini-index for the OFFLINE --lookup path: the 12 seed counties
# plus a dozen well-known large counties. Live mode greps the fetched county
# CSV (~470 counties) instead.
MINI_INDEX = {
    "01003": ("Baldwin County", "AL"), "01073": ("Jefferson County", "AL"),
    "06037": ("Los Angeles County", "CA"), "17031": ("Cook County", "IL"),
    "48201": ("Harris County", "TX"), "04013": ("Maricopa County", "AZ"),
    "06073": ("San Diego County", "CA"), "12086": ("Miami-Dade County", "FL"),
    "36047": ("Kings County", "NY"), "39035": ("Cuyahoga County", "OH"),
    "42101": ("Philadelphia County", "PA"), "53033": ("King County", "WA"),
    "06059": ("Orange County", "CA"), "48113": ("Dallas County", "TX"),
    "32003": ("Clark County", "NV"), "26163": ("Wayne County", "MI"),
    "36061": ("New York County", "NY"), "13121": ("Fulton County", "GA"),
    "25025": ("Suffolk County", "MA"), "08031": ("Denver County", "CO"),
    "27053": ("Hennepin County", "MN"), "37119": ("Mecklenburg County", "NC"),
    "51059": ("Fairfax County", "VA"), "12011": ("Broward County", "FL"),
}


class CfpbDemoProvider(Provider):
    """Deterministic offline stand-in (BUILD SPEC 0.7). Values are a pure
    function of (geo key, measure, vintage): ~1-6 pct with drift so a few
    footprint counties trip WATCH/ALERT, and histories CHANGE between
    vintages (the full-history revision semantics, trap C3). One seeded
    footprint county -- slot 12, King County WA 53033 -- is deliberately
    ABSENT from the demo data so the SUPPRESSED path (trap C4) is always
    exercised. NO network, NO key."""

    source_class = "A"
    DEFAULT_VINTAGE = "thru-2025-09"
    SUPPRESSED_FIPS = "53033"          # seed slot 12: King County WA

    # fixed stress tiers for seed counties (fiction assigned by hand, stated
    # in-sheet): d90 streak-3 + dev ALERTs, level per base below
    TIERS = {"12086": "alert", "48201": "alert",
             "04013": "watch", "39035": "watch"}

    def __init__(self, asof: Optional[date] = None,
                 raw_slots: int = RAW_SLOTS_DEFAULT,
                 vintage: Optional[str] = None):
        self.asof = asof or date(2026, 4, 30)
        self.raw_slots = int(raw_slots)
        v = (vintage or "").strip() or self.DEFAULT_VINTAGE
        if not VINTAGE_RE.match(v):
            raise RuntimeError(f"demo vintage '{v}' is not thru-YYYY-MM")
        self.vintage = v if v.startswith("thru-") else "thru-" + v
        self.vintage_source = "demo"

    def prime(self, asof: date) -> None:
        return None                     # nothing to fetch

    def _months(self) -> List[str]:
        y, m = vintage_ym(self.vintage)
        end = y * 12 + (m - 1)
        idx = range(end - self.raw_slots + 1, end + 1)
        return [f"{k // 12:04d}-{k % 12 + 1:02d}" for k in idx]

    @staticmethod
    def _seed(key: str) -> int:
        n = 0
        for i, c in enumerate(key):
            n = (n * 131 + (i + 1) * ord(c)) % (2 ** 32)
        return (n * 2654435761) % (2 ** 32)

    def _values(self, key: str, kind: str, measure: str) -> List[Optional[float]]:
        """Oldest-first value list for one geo x measure. Bases sit in the
        realistic 1-6 pct (d3089) / 0.4-3 pct (d90) ranges; tiered counties
        get deterministic rising tails (no wobble in the tail) so their
        dev_12m / rise_streak3 statuses are exact."""
        s = self._seed(key + "|" + measure)
        rev = self._seed(self.vintage) % 9          # vintage revision shift
        shift = (rev - 4) * 0.02
        n = self.raw_slots
        tier = self.TIERS.get(key, "calm") if kind == "county" else "calm"
        if measure == "d3089":
            base = 1.6 + (s % 27) * 0.1             # 1.6 - 4.2
        else:
            base = 0.45 + (s % 24) * 0.05           # 0.45 - 1.60
            if tier == "alert":
                base = 2.6 if key == "12086" else 2.0
        if kind == "national":
            base = 2.8 if measure == "d3089" else 1.0
        vals: List[Optional[float]] = []
        for i in range(n):
            wob = 0.12 * math.sin((i + s % 11) / 2.6)
            trend = ((s % 5) - 2) * 0.002 * i
            v = base + shift + trend + wob
            if tier != "calm" and i >= n - 14:
                # deterministic tail: freeze the wobble/trend at the window
                # edge so the transform arithmetic is exact
                v = base + shift + ((s % 5) - 2) * 0.002 * (n - 14) \
                    + 0.12 * math.sin((n - 14 + s % 11) / 2.6)
                k = i - (n - 14)                     # 0..13, 13 = newest
                if tier == "alert":
                    ramp3089 = [0, 0, 0, 0, 0, 0, 0, 0,
                                0.15, 0.32, 0.5, 0.7, 0.9, 1.1]
                    ramp90 = [0, 0, 0, 0, 0, 0,
                              0, 0, 0.05, 0.12, 0.22, 0.34, 0.45, 0.55]
                else:                                # watch tier
                    ramp3089 = [0, 0, 0, 0, 0, 0, 0, 0,
                                0.06, 0.13, 0.22, 0.33, 0.45, 0.45]
                    ramp90 = [0, 0, 0, 0, 0, 0,
                              0, 0, 0, 0, 0, 0, 0.10, 0.18]
                v += (ramp3089 if measure == "d3089" else ramp90)[k]
            vals.append(round(max(v, 0.05), 3))
        # one interior None per series (missing-month tolerance), well below
        # the 13-month transform window so headline statuses never see it
        if n >= 40:
            depth = 20 + (s % min(30, n - 21))       # 20..49 from newest
            vals[n - 1 - depth] = None
        return vals

    def fetch_series(self, spec: UnitSpec, secret=None) -> List[NormalizedRow]:
        if spec.kind == "county" and spec.fips == self.SUPPRESSED_FIPS:
            return []                    # ABSENT = suppressed (trap C4)
        key = ("nat" if spec.kind == "national"
               else ("st" + spec.fips if spec.kind == "state" else spec.fips))
        months = self._months()
        vals = self._values(key, spec.kind, spec.measure)
        return [NormalizedRow(id=spec.id, period=mon, value=val,
                              geo_segment=spec.geo_segment,
                              source_class=self.source_class)
                for mon, val in zip(months, vals)]


class CfpbProvider(Provider):
    """Live Class A provider: plain urllib against consumerfinance.gov
    (KEYLESS, public domain). TWO-STAGE (BUILD SPEC sec 1):

      1. GET the download page and regex the CURRENT dated filenames
         (trap C2). If the page scrape fails or the `vintage` setting is
         set (e.g. 'thru-2025-09'), construct the URLs directly.
      2. GET the 4 CSVs (State + County x both measures; the Metro file is
         SKIPPED -- mixed non-FIPS keys).

    Throttle 0.6s between requests, retry/backoff on 429/5xx, per-URL
    cache. The whole refresh is 5 small HTTP GETs."""

    source_class = "A"

    def __init__(self, min_interval: float = 0.6, max_retries: int = 4,
                 raw_slots: int = RAW_SLOTS_DEFAULT,
                 vintage_override: str = ""):
        self._min_interval = float(min_interval)
        self._max_retries = int(max_retries)
        self._raw_slots = int(raw_slots)
        self._override = str(vintage_override or "").strip()
        self._last = 0.0
        self._cache: Dict[str, bytes] = {}
        self._data: Optional[Dict[str, Dict[str, dict]]] = None

    def _throttle(self):
        gap = time.time() - self._last
        if gap < self._min_interval:
            time.sleep(self._min_interval - gap)
        self._last = time.time()

    def _http_get(self, url: str) -> bytes:
        """The one real network call -- isolated so tests can stub it."""
        import urllib.request
        req = urllib.request.Request(url, headers={
            "User-Agent": "cfpb-mortgage-monitor/1.0 (plain urllib)"})
        with urllib.request.urlopen(req, timeout=120) as resp:
            return resp.read()

    def _download(self, url: str, label: str) -> bytes:
        """Cache + throttle + backoff. 429/5xx retry with backoff; other
        HTTP codes fail fast (a 404 usually means the vintage moved -- the
        next discovery pass fixes it)."""
        if url in self._cache:                       # repeat fetch == cache hit
            return self._cache[url]
        import urllib.error
        for attempt in range(self._max_retries + 1):
            self._throttle()
            try:
                data = self._http_get(url)
                self._cache[url] = data
                return data
            except urllib.error.HTTPError as exc:
                if exc.code == 429 or exc.code >= 500:
                    if attempt < self._max_retries:
                        time.sleep(2.0 * (attempt + 1))
                        continue
                raise RuntimeError(f"CFPB fetch failed for {label}: "
                                   f"HTTP {exc.code}")
            except Exception as exc:
                if attempt < self._max_retries:
                    time.sleep(2.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"CFPB fetch failed for {label}: "
                                   f"{type(exc).__name__}")

    def discover(self) -> str:
        """Stage 1: the current vintage 'YYYY-MM'. Honors the `vintage`
        settings override (skip the scrape entirely); otherwise scrapes the
        download page for the dated filenames and takes the newest."""
        if self._override:
            ym = vintage_ym(self._override)
            if ym is None:
                raise RuntimeError(
                    f"_config vintage override '{self._override}' is not "
                    "thru-YYYY-MM (e.g. thru-2025-09).")
            self.vintage_source = "override"
            return f"{ym[0]:04d}-{ym[1]:02d}"
        html = self._download(DOWNLOAD_PAGE, "download page").decode(
            "utf-8", "replace")
        found = FILENAME_RE.findall(html)
        if not found:
            raise RuntimeError(
                "Could not discover any dated CFPB filenames on the download "
                "page (page layout changed, or blocked). Set the `vintage` "
                "key in _config [SETTINGS] (e.g. thru-2025-09) to construct "
                "the URLs directly.")
        self.vintage_source = "discovered"
        return max(ym for _geo, _meas, ym in found)

    def prime(self, asof: date) -> None:
        """Stage 2: fetch + parse the 4 CSVs for the discovered vintage."""
        ym = self.discover()
        self.vintage = "thru-" + ym
        self._data = {"State": {}, "County": {}}
        for geo in ("State", "County"):
            for m in MEASURES:
                fname = (f"{geo}MortgagesPercent-{MEASURE_FILE[m]}DaysLate-"
                         f"thru-{ym}.csv")
                raw = self._download(FILES_BASE + fname, fname)
                self._data[geo][m] = parse_wide_csv(
                    raw.decode("utf-8-sig", "replace"))

    def fetch_series(self, spec: UnitSpec, secret=None) -> List[NormalizedRow]:
        if self._data is None:
            raise RuntimeError("CfpbProvider.fetch_series before prime(): "
                               "the 4-CSV fetch must run once per refresh.")
        geo_file = "County" if spec.kind == "county" else "State"
        if spec.kind == "national":
            # the National row rides in every file (trap C7); read it from
            # the state file (smallest)
            parsed = self._data["State"][spec.measure]
            series = parsed["national"] or {}
        else:
            parsed = self._data[geo_file][spec.measure]
            geo = parsed["geos"].get(spec.fips)
            if geo is None:
                return []                # ABSENT = suppressed (trap C4)
            series = geo["values"]
        months = parsed["months"][-self._raw_slots:]
        return [NormalizedRow(id=spec.id, period=mon,
                              value=series.get(mon),
                              geo_segment=spec.geo_segment,
                              source_class=self.source_class)
                for mon in months]

    def lookup(self, name: str) -> List[dict]:
        """--lookup support: grep the fetched county CSV (one measure is
        enough) for name/state matches. LIVE path; the offline path searches
        the embedded MINI_INDEX instead."""
        ym = self.discover()
        fname = f"CountyMortgagesPercent-30-89DaysLate-thru-{ym}.csv"
        parsed = parse_wide_csv(
            self._download(FILES_BASE + fname, fname).decode("utf-8-sig",
                                                             "replace"))
        needle = name.strip().lower()
        hits = []
        for fips, geo in sorted(parsed["geos"].items()):
            hay = f"{geo['name']} {geo['state']}".lower()
            if needle in hay:
                hits.append({"fips": fips, "name": geo["name"],
                             "state": geo["state"]})
        return hits


def resolve_secret(cfg: Config) -> Optional[str]:
    """Contract plumbing: a licensed (Class C) secret would come from the
    env var NAMED in _config, never stored. Unused in v1 -- the CFPB files
    are public domain and keyless, and the gates admit only Class A."""
    name = str(cfg.setting("secret_env", "") or "").strip()
    if not name:
        return None
    return os.environ.get(name)


def make_provider(cfg: Config, demo: bool, asof: Optional[date]) -> Provider:
    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        return CfpbDemoProvider(asof=asof, raw_slots=cfg.raw_slots,
                                vintage=cfg.vintage_override or None)
    return CfpbProvider(
        min_interval=float(cfg.setting("http_min_interval", 0.6) or 0.6),
        max_retries=int(float(cfg.setting("cfpb_max_retries", 4) or 4)),
        raw_slots=cfg.raw_slots,
        vintage_override=cfg.vintage_override)


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder. One block per GEO
# SLOT: months as rows (newest-first), measures as columns B/C.
# ---------------------------------------------------------------------------
@dataclass
class GeoBlock:
    kind: str
    slot: int
    header_row: int
    label_row: int
    first_data_row: int
    slots: int                 # month rows (raw_slots)


def block_ordinal(kind: str, slot: int) -> int:
    if kind == "national":
        return 0
    if kind == "state":
        return slot                          # 1..STATE_SLOTS
    return STATE_SLOTS + slot                # county slots follow the states


def geo_block(kind: str, slot: int,
              raw_slots: int = RAW_SLOTS_DEFAULT) -> GeoBlock:
    """Deterministic block placement: anchors depend ONLY on (kind, slot,
    raw_slots) -- editing which county occupies the slot moves nothing."""
    stride = RAW_HEADER_ROWS + raw_slots + RAW_GAP_ROWS
    header_row = RAW_FIRST_ROW + block_ordinal(kind, slot) * stride
    return GeoBlock(kind, slot, header_row, header_row + 1,
                    header_row + RAW_HEADER_ROWS, raw_slots)


def assemble_months(measure_rows: Dict[str, List[NormalizedRow]],
                    raw_slots: int):
    """Per-measure NormalizedRows -> newest-first [(month, {measure: value})]
    rows ready to land. Months are unioned so a measure missing one month
    still aligns (blank cell, not a shifted row)."""
    months = sorted({r.period for rows in measure_rows.values() for r in rows},
                    reverse=True)[:raw_slots]
    per: Dict[str, Dict[str, Optional[float]]] = {p: {} for p in months}
    for m, rows in measure_rows.items():
        for r in rows:
            if r.period in per:
                per[r.period][m] = r.value
    return [(p, {m: per[p].get(m) for m in MEASURES}) for p in months]


# ---------------------------------------------------------------------------
# WRITE BACKEND -- openpyxl on the closed workbook (contract sec 4 / L2)
# ---------------------------------------------------------------------------
class OpenpyxlBackend:
    def __init__(self, path: str):
        self.path = path
        import openpyxl
        # L2: keep_vba=True ONLY for .xlsm; on .xlsx it injects a dangling
        # vbaProject relationship Excel rejects.
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        return parse_config([[c.value for c in row] for row in ws.iter_rows()])

    def _check_block(self, block: GeoBlock):
        """Refuse to write into a workbook whose raw layout doesn't match
        the computed one (raw_slots edited after build): every dashboard and
        Watchlist formula is anchored to the BUILT layout, so a silent
        mismatch would make the whole workbook quietly wrong. The block
        label is constant regardless of which county occupies the slot, so
        [FOOTPRINT] edits never trip this."""
        ws = self._wb[RAW_TAB]
        expected = slot_label(block.kind, block.slot)
        existing = ws.cell(block.header_row, 1).value
        if existing is not None and str(existing).strip() not in ("", expected):
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB} at row {block.header_row}: "
                f"expected block '{expected}' but found '{existing}'. The "
                f"workbook was built with a different raw_slots/"
                f"footprint_slots layout than _config now describes; "
                f"formulas are anchored to the built layout. Rebuild the "
                f"workbook (python make_workbook.py --footprint-slots N) "
                f"instead of editing the layout settings in place.")

    def clear_geo_block(self, block: GeoBlock):
        """Blank a block's identity AND data (stateless FULL REPLACE, BUILD
        SPEC 0.2 -- every CFPB vintage revises the ENTIRE history, and a
        removed/suppressed/failed geo must show empty, never last vintage's
        values masquerading as current)."""
        self._check_block(block)
        ws = self._wb[RAW_TAB]
        # L7: ws.cell(r, c, None) is a silent NO-OP in openpyxl (the value
        # arg is ignored when None) -- clearing MUST assign .value directly.
        for c in (2, 3, 4):                          # runtime identity cells
            ws.cell(block.header_row, c).value = None
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            for c in (1, 2, 3):
                ws.cell(r, c).value = None

    def write_block_identity(self, block: GeoBlock, unit_key: str,
                             name: str, note: str):
        self._check_block(block)
        ws = self._wb[RAW_TAB]
        ws.cell(block.header_row, 1, slot_label(block.kind, block.slot))
        ws.cell(block.header_row, 2, unit_key)
        ws.cell(block.header_row, 3, name)
        ws.cell(block.header_row, 4, note)

    def write_geo_block(self, block: GeoBlock, unit_key: str, name: str,
                        note: str, months):
        """months: newest-first [(month, {measure: value})]."""
        self.write_block_identity(block, unit_key, name, note)
        ws = self._wb[RAW_TAB]
        for i, (mon, mv) in enumerate(months[:block.slots]):
            r = block.first_data_row + i             # newest-first
            ws.cell(r, 1, str(mon))
            for m in MEASURES:
                v = mv.get(m)
                # explicit .value assignment: None must LAND as blank (L7)
                ws.cell(r, MEASURE_COL[m]).value = \
                    None if v is None else float(v)

    def write_status(self, status: dict):
        line2 = (f"Counties {status.get('counties_landed', 0)}/"
                 f"{status.get('counties_active', 0)} - "
                 f"{status.get('alert_counties', 0)} ALERT / "
                 f"{status.get('watch_counties', 0)} WATCH / "
                 f"{status.get('counties_suppressed', 0)} SUPPRESSED")
        n_ref = len(status.get("watchlist_refusals") or [])
        if n_ref:
            line2 += f" - {n_ref} REFUSED (see digest)"
        n_err = len(status.get("errors") or [])
        if n_err:
            line2 += f" - {n_err} FETCH ERRORS"
        line3 = (f"CFPB vintage: {status.get('vintage') or 'n/a'} "
                 f"({status.get('vintage_source') or 'n/a'}) - data lag "
                 f"~6-7 months by design")
        warn = status.get("continuity_warning")
        behind = status.get("months_behind")
        line4 = warn if warn else (
            f"Continuity OK: vintage {behind} months behind asof "
            f"(tripwire {status.get('continuity_months')})"
            if behind is not None else "Continuity: vintage unparsed")
        for tab in DASH_TABS:
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                # explicit .value writes: line 4 must OVERWRITE any previous
                # warning, never leave one stale (L7 discipline)
                ws.cell(1, STATUS_COL).value = ("Last run  "
                                                + status.get("timestamp", "")
                                                + f" ({status.get('mode', '')})")
                ws.cell(2, STATUS_COL).value = line2
                ws.cell(3, STATUS_COL).value = line3
                ws.cell(4, STATUS_COL).value = line4

    def finalize(self):
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def _measure_vals(months, measure) -> List[Optional[float]]:
    """Newest-first value list for one measure off assembled month rows."""
    return [mv.get(measure) for _mon, mv in months]


def compute_digest(cfg: Config, asof: date,
                   nat_months, state_landed, county_landed,
                   suppressed: List[FootprintRow],
                   smap: Dict[str, int]) -> dict:
    """Latest-month transform values + statuses + flag counts (the engine
    output consumed by the Watchlist parity tests and email_sim).
    Suppression (BUILD SPEC 0.6) is first-class: a suppressed county's
    STATUS is SUPPRESSED and it never counts toward the alert/watch KPIs."""
    thr_dev = cfg.thresholds.get("d3089_dev12m")
    thr_stk = cfg.thresholds.get("d90_streak3")
    thr_lvl = cfg.thresholds.get("d90_level")

    def geo_metrics(months):
        v3089 = _measure_vals(months, "d3089")
        v90 = _measure_vals(months, "d90")
        return {
            "d3089_latest": t_level(v3089),
            "d3089_dev12": t_dev_12m(v3089),
            "d90_latest": t_level(v90),
            "d90_dev12": t_dev_12m(v90),
            "d90_streak": t_rise_streak3(v90),
        }

    national = geo_metrics(nat_months) if nat_months else None

    states = []
    for fips2 in sorted(smap, key=lambda f: smap[f]):
        slot = smap[fips2]
        entry = {"slot": slot, "fips2": fips2,
                 "abbrev": (state_landed.get(fips2) or {}).get("abbrev", ""),
                 "landed": fips2 in state_landed}
        if fips2 in state_landed:
            entry.update(geo_metrics(state_landed[fips2]["months"]))
        states.append(entry)

    counties = []
    sup_slots = {f.slot for f in suppressed}
    for fp, months in sorted(
            list(county_landed.values()) + [(f, None) for f in suppressed],
            key=lambda t: t[0].slot):
        entry = {"slot": fp.slot, "fips": fp.fips, "name": fp.name,
                 "state": fp.state, "entity_key": fp.entity_key,
                 "suppressed": fp.slot in sup_slots}
        if entry["suppressed"]:
            entry.update({"d3089_latest": None, "d3089_dev12": None,
                          "d90_latest": None, "d90_dev12": None,
                          "d90_streak": None, "flags": {},
                          "status": "SUPPRESSED", "severity": None})
            counties.append(entry)
            continue
        entry.update(geo_metrics(months))
        flags = {
            "d3089_dev12m": status_for(entry["d3089_dev12"], thr_dev),
            "d90_streak3": status_for(entry["d90_streak"], thr_stk),
            "d90_level": status_for(entry["d90_latest"], thr_lvl),
        }
        sev = max(SEVERITY[f] for f in flags.values())
        entry["flags"] = flags
        entry["status"] = {0: "OK", 1: "WATCH", 2: "ALERT"}[sev]
        entry["severity"] = sev
        counties.append(entry)

    return {"national": national, "states": states, "counties": counties}


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None,
        provider: Optional[Provider] = None) -> dict:
    """One refresh against the CLOSED workbook. `provider` is injectable for
    tests (e.g. the two-vintage full-replace regression)."""
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any fetch (BUILD SPEC 0.1 / sec 3).
    validate_series(cfg.series)
    assert_series_admissible(cfg.series)
    validate_footprint_capacity(cfg)
    admitted, refusals, excluded = evaluate_footprint(cfg)
    smap = state_slot_map(cfg)

    if provider is None:
        provider = make_provider(cfg, demo, asof)
    mode = "demo" if isinstance(provider, CfpbDemoProvider) else "live"

    # Stateless FULL REPLACE (BUILD SPEC 0.2): blank the national block,
    # every state slot and every county slot 1..capacity FIRST -- every CFPB
    # vintage revises the ENTIRE history, and removed/refused/suppressed
    # slots must show blank under the fresh timestamp, never the previous
    # vintage. This also validates the workbook's raw layout before any
    # fetch.
    backend.clear_geo_block(geo_block("national", 0, cfg.raw_slots))
    for slot in range(1, STATE_SLOTS + 1):
        backend.clear_geo_block(geo_block("state", slot, cfg.raw_slots))
    for slot in range(1, cfg.footprint_slots + 1):
        backend.clear_geo_block(geo_block("county", slot, cfg.raw_slots))

    errors: List[str] = []
    nat_months = None
    state_landed: Dict[str, dict] = {}
    county_landed: Dict[int, Tuple[FootprintRow, list]] = {}
    suppressed: List[FootprintRow] = []
    try:
        provider.prime(asof)
    except Exception as exc:
        errors.append(f"fetch: {exc}")
    else:
        # ---- national ----
        try:
            rows = {m: provider.fetch_series(make_national_spec(m))
                    for m in MEASURES}
            nat_months = assemble_months(rows, cfg.raw_slots)
            backend.write_geo_block(
                geo_block("national", 0, cfg.raw_slots),
                "nat national", "UNITED STATES", "RegionType=National",
                nat_months)
        except Exception as exc:
            errors.append(f"national: {exc}")
        # ---- states (dashboard context) ----
        abbrevs = {}
        for s in state_series(cfg):
            fips2 = s.geo_segment.split(":", 1)[1]
            abbrevs.setdefault(fips2, s.title.split(" ", 1)[0])
        for fips2, slot in sorted(smap.items(), key=lambda kv: kv[1]):
            try:
                rows = {m: provider.fetch_series(
                    make_state_spec(slot, fips2, m)) for m in MEASURES}
                if not any(rows.values()):
                    errors.append(f"st{slot:02d} state:{fips2}: no data "
                                  "in the state file")
                    continue
                months = assemble_months(rows, cfg.raw_slots)
                ab = abbrevs.get(fips2, "")
                backend.write_geo_block(
                    geo_block("state", slot, cfg.raw_slots),
                    f"st{slot:02d} state:{fips2}", ab,
                    f"state fips {fips2}", months)
                state_landed[fips2] = {"months": months, "abbrev": ab,
                                       "slot": slot}
            except Exception as exc:
                errors.append(f"st{slot:02d} state:{fips2}: {exc}")
        # ---- footprint counties (the watchlist lane) ----
        for fp in admitted:
            try:
                rows = {m: provider.fetch_series(make_county_spec(fp, m))
                        for m in MEASURES}
            except Exception as exc:
                errors.append(f"co{fp.slot:02d} {fp.entity_key}: {exc}")
                continue
            if not any(rows.values()):
                # ABSENT from the vintage = SUPPRESSED (trap C4): identity
                # is written so the block is explicit, data stays blank, and
                # the Watchlist formula renders the SUPPRESSED text.
                backend.write_block_identity(
                    geo_block("county", fp.slot, cfg.raw_slots),
                    f"co{fp.slot:02d} {fp.entity_key}", fp.name,
                    "SUPPRESSED: absent from this vintage (below the "
                    "~1,000-mortgage sample threshold) -- use the state row")
                suppressed.append(fp)
                continue
            months = assemble_months(rows, cfg.raw_slots)
            backend.write_geo_block(
                geo_block("county", fp.slot, cfg.raw_slots),
                f"co{fp.slot:02d} {fp.entity_key}", fp.name,
                f"{fp.state}", months)
            county_landed[fp.slot] = (fp, months)

    digest = compute_digest(cfg, asof, nat_months, state_landed,
                            county_landed, suppressed, smap)
    warn = continuity_warning(asof, provider.vintage, cfg.continuity_months) \
        if provider.vintage else ("CONTINUITY WARNING: no vintage recorded "
                                  "(fetch failed?)")
    live = [c for c in digest["counties"] if not c["suppressed"]]
    # footprint states with no state dashboard row: honest note, not a gate
    missing_states = sorted({fp.state for fp in admitted if fp.state
                             and fp.state not in
                             {s.title.split(" ", 1)[0]
                              for s in state_series(cfg)}})
    notes = []
    if missing_states:
        notes.append("footprint states without a state dashboard row: "
                     + ", ".join(missing_states)
                     + " (add st_* rows to _config [SERIES] if wanted)")
    status = {
        "timestamp": asof.isoformat(),
        "mode": mode,
        "footprint_slots": cfg.footprint_slots,
        "counties_active": len(admitted) + len(refusals),
        "counties_landed": len(county_landed),
        "counties_suppressed": len(suppressed),
        "counties_excluded": len(excluded),
        # suppression excludes a county from every alert KPI (BUILD SPEC 0.6)
        "alert_counties": sum(1 for c in live if c["status"] == "ALERT"),
        "watch_counties": sum(1 for c in live if c["status"] == "WATCH"),
        "states_landed": len(state_landed),
        "vintage": provider.vintage,
        "vintage_source": provider.vintage_source,
        "months_behind": months_behind(asof, provider.vintage),
        "continuity_months": cfg.continuity_months,
        "continuity_warning": warn,
        "digest": digest,
        "watchlist_refusals": [m for _, m in refusals],
        "watchlist_admitted": [f"co{f.slot:02d} {f.entity_key}"
                               for f in admitted],
        "suppressed_counties": [f"co{f.slot:02d} {f.entity_key} {f.name}"
                                for f in suppressed],
        "notes": notes,
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


# ---------------------------------------------------------------------------
# CLI (contract sec 4) + --lookup (offline-friendly)
# ---------------------------------------------------------------------------
def do_lookup(name: str, demo: bool) -> int:
    needle = name.strip().lower()
    if demo:
        hits = [{"fips": f, "name": n, "state": st}
                for f, (n, st) in sorted(MINI_INDEX.items())
                if needle in f"{n} {st}".lower()]
        src = (f"offline mini-index ({len(MINI_INDEX)} well-known counties; "
               "run without --demo to grep the full CFPB county file, "
               "~470 counties)")
    else:
        prov = CfpbProvider()
        try:
            hits = prov.lookup(name)
        except Exception as exc:
            sys.stderr.write(
                f"lookup failed: {exc}\n"
                "Check network access to consumerfinance.gov (keyless; no "
                "account needed), or use the offline mini-index: "
                f'python runner.py --lookup "{name}" --demo\n')
            return 1
        src = f"CFPB county file {prov.vintage or ''}".strip()
    if not hits:
        print(f'No counties matched "{name}" ({src}).')
        print("NOTE: only counties above the ~1,000-sample-mortgage "
              "threshold appear in the live file; a missing rural county "
              "is SUPPRESSED by the source -- use its state row.")
        return 0
    print(f'{"FIPS":>6}  {"ST":<3} COUNTY        [{src}]')
    for d in hits:
        print(f'{d["fips"]:>6}  {d["state"]:<3} {d["name"]}')
    print("\nPut the FIPS into the _config [FOOTPRINT] table "
          "(slot | fips | name | state | active) and re-run the runner -- "
          "no rebuild needed within the built slot capacity. Keep the cell "
          "TEXT-formatted so the leading zero survives.")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="County Mortgage Delinquency Monitor runner "
                    "(CFPB Mortgage Performance Trends)")
    ap.add_argument("--workbook", "-w", default=None,
                    help="path to the .xlsm (required unless --lookup)")
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline CfpbDemoProvider "
                         "(no network; the live files are keyless anyway)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    ap.add_argument("--lookup", default=None, metavar="NAME",
                    help='find a county\'s FIPS by name: --lookup "Cook" '
                         "(offline with --demo, live otherwise)")
    args = ap.parse_args(argv)
    if args.lookup is not None:
        return do_lookup(args.lookup, args.demo)
    if not args.workbook:
        sys.stderr.write("--workbook is required (or use --lookup).\n")
        return 2
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, SeriesError, FootprintCapacityError,
            StateCapacityError) as exc:
        sys.stderr.write(f"GATE ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except SystemExit as exc:
        sys.stderr.write(f"{exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True,
                      **{k: v for k, v in status.items() if k != "digest"}}))
    warn = status.get("continuity_warning")
    sys.stderr.write(
        f"OK ({status['mode']}): {status['counties_landed']}/"
        f"{status['counties_active']} counties landed, "
        f"{status['alert_counties']} ALERT / {status['watch_counties']} "
        f"WATCH / {status['counties_suppressed']} SUPPRESSED, "
        f"states={status['states_landed']}, vintage={status['vintage']} "
        f"({status['vintage_source']})"
        + (f" -- {warn}" if warn else "") + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
