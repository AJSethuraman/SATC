#!/usr/bin/env python3
"""Email-simulate acceptance test (BUILD_SPEC_CFPB.md Phase 10, --demo).

Copies ONLY the .xlsm into a fresh, otherwise-empty folder, then reproduces
exactly what the VBA "Extract" button does:
  1. read the _code_py tab (one source line per cell, column A) and write
     runner.py next to the workbook -- byte-for-byte the macro's
     WriteTabToFile;
  2. shell Python to run the EXTRACTED runner.py against the CLOSED workbook
     in --demo mode (deterministic, offline -- the live files are keyless
     anyway).
Then it composes (does NOT send) the monitoring email from the run digest
and asserts the email contains the RANKED COUNTY WATCHLIST, the STATE
SUMMARY, the VINTAGE line, the SUPPRESSION section and the CONTINUITY
section, and that the workbook rebuilt itself with nothing but the workbook
+ extracted runner present (the workbook is the source of truth).
Deterministic from _config + CfpbDemoProvider at fixed --asof.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

import runner as R

XLSM_NAME = "Mortgage_Delinquency_Monitor.xlsm"
ASOF = "2026-04-30"


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write UTF-8."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def _fmt(v, spec="{:.2f}"):
    return "n/a" if v is None else spec.format(v)


def compose_email(status):
    """Deterministic monitoring email from the run status/digest. No send."""
    digest = status.get("digest", {})
    counties = digest.get("counties", [])
    live = [c for c in counties if not c.get("suppressed")]
    suppressed = [c for c in counties if c.get("suppressed")]
    lines = []
    lines.append("Subject: County Mortgage Delinquency Monitor -- "
                 f"{status['alert_counties']} ALERT / "
                 f"{status['watch_counties']} WATCH / "
                 f"{status['counties_suppressed']} SUPPRESSED counties "
                 f"({status['timestamp']})")
    lines.append("")
    lines.append(f"Run mode: {status['mode']}  |  counties landed: "
                 f"{status['counties_landed']}/{status['counties_active']}"
                 f"  |  refused: {len(status.get('watchlist_refusals', []))}")
    # the vintage line (trap C6: the lag is the lag -- label it)
    lines.append(f"Data vintage: {status.get('vintage') or 'n/a'} "
                 f"({status.get('vintage_source') or 'n/a'}) -- monthly "
                 f"series, ~6-7 months lagged by design; every vintage "
                 f"revises the FULL history (never diff vintages as "
                 f"movement)")
    lines.append("")
    lines.append("RANKED COUNTY WATCHLIST (by status severity, then 90+ "
                 "deviation vs own 12m avg)")
    lines.append("-" * 76)
    lines.append(f"  {'#':>2} {'County':<24} {'ST':<3} {'FIPS':>5} "
                 f"{'30-89':>6} {'dev12':>6} {'90+':>6} {'stk':>3}  Status")
    ranked = sorted(live, key=lambda c: (-(c["severity"] or 0),
                                         -(c["d90_dev12"] or 0.0)))
    for i, c in enumerate(ranked, start=1):
        lines.append(
            f"  {i:>2} {c['name']:<24} {c['state']:<3} {c['fips']:>5} "
            f"{_fmt(c['d3089_latest']):>6} {_fmt(c['d3089_dev12'], '{:+.2f}'):>6} "
            f"{_fmt(c['d90_latest']):>6} "
            f"{_fmt(c['d90_streak'], '{:.0f}'):>3}  {c['status']}")
    if not ranked:
        lines.append("  (no counties landed)")
    lines.append("")
    lines.append("STATE SUMMARY (dashboard context -- the fallback lane for "
                 "suppressed counties)")
    lines.append("-" * 76)
    nat = digest.get("national")
    if nat:
        lines.append(f"  {'NATIONAL':<10} 30-89 {_fmt(nat['d3089_latest'])} "
                     f"(dev12 {_fmt(nat['d3089_dev12'], '{:+.2f}')})  |  "
                     f"90+ {_fmt(nat['d90_latest'])} "
                     f"(dev12 {_fmt(nat['d90_dev12'], '{:+.2f}')})")
    for s in digest.get("states", []):
        if not s.get("landed"):
            continue
        lines.append(f"  {s['abbrev']:<10} 30-89 {_fmt(s['d3089_latest'])} "
                     f"(dev12 {_fmt(s['d3089_dev12'], '{:+.2f}')})  |  "
                     f"90+ {_fmt(s['d90_latest'])} "
                     f"(dev12 {_fmt(s['d90_dev12'], '{:+.2f}')})")
    lines.append("")
    lines.append("SUPPRESSED FOOTPRINT COUNTIES (absent from this vintage -- "
                 "below the ~1,000-sample threshold; excluded from alert "
                 "counts; absence is NOT zero)")
    lines.append("-" * 76)
    if suppressed:
        for c in sorted(suppressed, key=lambda x: x["slot"]):
            lines.append(f"  SUPPRESSED  co{c['slot']:02d} {c['name']} "
                         f"({c['state']} {c['fips']}) -- use the "
                         f"{c['state']} state row above")
    else:
        lines.append("  (none -- every active footprint county is in the "
                     "file)")
    lines.append("")
    lines.append("CONTINUITY")
    lines.append("-" * 10)
    warn = status.get("continuity_warning")
    if warn:
        lines.append("  " + warn)
    else:
        lines.append(f"  OK: vintage {status.get('vintage')} is "
                     f"{status.get('months_behind')} months behind asof "
                     f"(tripwire {status.get('continuity_months')}). The "
                     f"CFPB publishes informally -- this line flips to a "
                     f"WARNING if a release is skipped.")
    lines.append("")
    lines.append("WATCHLIST LANE")
    lines.append("--------------")
    refusals = status.get("watchlist_refusals", [])
    if refusals:
        for m in refusals:
            lines.append("  " + m)
    else:
        lines.append(f"  ADMITTED (county-FIPS-keyed Class A): "
                     f"{len(status.get('watchlist_admitted', []))} counties")
    for n in status.get("notes", []):
        lines.append("  NOTE: " + n)
    lines.append("")
    return "\n".join(lines)


def main():
    repo = os.path.dirname(os.path.abspath(__file__))
    xlsm = os.path.join(repo, XLSM_NAME)
    if not os.path.exists(xlsm):
        print("build the workbook first: python3 make_workbook.py",
              file=sys.stderr)
        return 1

    work = tempfile.mkdtemp(prefix="cfpb_emailsim_")
    try:
        # 1. The email contains ONLY the workbook.
        dst_xlsm = os.path.join(work, XLSM_NAME)
        shutil.copy(xlsm, dst_xlsm)
        before = sorted(os.listdir(work))
        assert before == [XLSM_NAME], before
        print(f"[email-sim] fresh folder contains only: {before}")

        # 2. Button step 1: extract runner.py from _code_py.
        runner_py = os.path.join(work, "runner.py")
        extract_code_tab(dst_xlsm, "_code_py", runner_py)
        print(f"[email-sim] extracted runner.py "
              f"({os.path.getsize(runner_py)} bytes) from _code_py")

        # 3. Button step 2: shell Python to run the EXTRACTED runner (demo).
        proc = subprocess.run(
            [sys.executable, runner_py, "--workbook", dst_xlsm, "--demo",
             "--asof", ASOF],
            capture_output=True, text=True, cwd=work)
        last = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() \
            else "(none)"
        print("[email-sim] runner stderr:", last)
        if proc.returncode != 0:
            print("[email-sim] FAILED: runner exited", proc.returncode)
            print(proc.stdout, proc.stderr)
            return 2

        # 4. Verify the rebuild: raw data present, macro intact, no extras.
        wb = openpyxl.load_workbook(dst_xlsm, keep_vba=True)
        cfg = R.parse_config([[c.value for c in row]
                              for row in wb["_config"].iter_rows()])
        b1 = R.geo_block("county", 1, cfg.raw_slots)
        raw_ok = wb[R.RAW_TAB].cell(b1.first_data_row, 2).value is not None
        macro_ok = wb.vba_archive is not None
        produced = sorted(os.listdir(work))
        wb.close()
        print(f"[email-sim] folder now: {produced}")
        print(f"[email-sim] raw populated={raw_ok}  macro intact={macro_ok}")

        # 5. Compose the monitoring email from the digest (deterministic).
        status = R.run(dst_xlsm, demo=True,
                       asof=R.datetime.strptime(ASOF, "%Y-%m-%d").date())
        email = compose_email(status)
        email_path = os.path.join(work, "monitoring_email.txt")
        with open(email_path, "w", encoding="utf-8") as fh:
            fh.write(email)
        print("\n" + "=" * 76)
        print(email)
        print("=" * 76 + "\n")

        digest = status["digest"]
        n_live = sum(1 for c in digest["counties"] if not c["suppressed"])
        n_states = sum(1 for s in digest["states"] if s["landed"])
        table_lines = [ln for ln in email.splitlines()
                       if ln.strip() and ln.strip()[0].isdigit()
                       and ("County" in ln or "Parish" in ln)]
        has_table = ("RANKED COUNTY WATCHLIST" in email
                     and len(table_lines) == n_live
                     and "ALERT" in email)
        state_lines = [ln for ln in email.splitlines()
                       if ln.startswith("  ") and "30-89 " in ln
                       and "|  90+ " in ln]
        has_states = ("STATE SUMMARY" in email
                      and len(state_lines) == n_states + 1  # + NATIONAL
                      and "NATIONAL" in email)
        has_vintage = ("Data vintage: thru-" in email
                       and "revises the FULL history" in email)
        has_suppression = ("SUPPRESSED FOOTPRINT COUNTIES" in email
                           and "King County" in email
                           and "use the WA state row" in email)
        has_continuity = ("CONTINUITY" in email
                          and ("tripwire" in email
                               or "CONTINUITY WARNING" in email))
        self_contained = (raw_ok and macro_ok
                          and set(produced) <= {XLSM_NAME, "runner.py"})

        ok = (has_table and has_states and has_vintage and has_suppression
              and has_continuity and self_contained)
        print(f"[email-sim] ranked county watchlist : {has_table} "
              f"({len(table_lines)} rows)")
        print(f"[email-sim] state summary           : {has_states} "
              f"({len(state_lines)} lines incl. national)")
        print(f"[email-sim] vintage line            : {has_vintage} "
              f"({status['vintage']}, {status['vintage_source']})")
        print(f"[email-sim] suppression section     : {has_suppression}")
        print(f"[email-sim] continuity section      : {has_continuity}")
        print(f"[email-sim] workbook self-contained : {self_contained}")
        print("[email-sim] RESULT:",
              "PASS -- workbook is the source of truth" if ok else "FAIL")
        return 0 if ok else 3
    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
