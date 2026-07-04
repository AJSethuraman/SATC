"""Headless test suite for the County Mortgage Delinquency Monitor.

Runs on Linux/CI with NO Excel and NO network, all in --demo mode. Mirrors
the named tests in BUILD_SPEC_CFPB.md Section 6 (Phases 1-9):

  test_config_parse                 (1 -- [FOOTPRINT] + over-capacity refusal)
  test_demo_provider_deterministic  (2)
  test_cfpb_provider_parse          (2 -- REAL-schema CSV fixture, quoted FIPS)
  test_filename_discovery           (2 -- download-page HTML fixture + override)
  test_provider_backoff             (2 -- stubbed 429/5xx, no network)
  test_transforms                   (3 -- hardcoded expected values, NaN tolerance)
  test_reload_headless              (4)
  test_watchlist_county_gates       (5 -- positive + every refusal shape)
  test_suppressed_county            (6 -- absent county -> SUPPRESSED, not zero)
  test_vintage_continuity           (7 -- stale vintage -> warning in status+email)
  test_full_replace                 (8 -- two demo vintages, no stale months)
  test_raw_landing_idempotent       (9)
  test_raw_layout_mismatch_refused  (9)
  test_clear_actually_blanks        (9 -- the L7 regression)
  test_vba_protection_keys_roundtrip(9)
  test_lookup_offline               (--lookup offline mini-index path)
"""
import json
import os
import re
import shutil
import struct
import sys
from dataclasses import replace
from datetime import date

import openpyxl
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import runner as R                     # noqa: E402
import series_seed as SEED             # noqa: E402
import build_workbook as BW            # noqa: E402
import assemble_xlsm                   # noqa: E402
import email_sim                       # noqa: E402
import vba_writer as V                 # noqa: E402

ASOF = date(2026, 4, 30)

TABS = ("Dashboard_State", "Dashboard_Trends", "Watchlist", "Raw_CFPB",
        "_config", "_code_py", "_code_vba", "_readme")


# --------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------
@pytest.fixture(scope="session")
def xlsm(tmp_path_factory):
    """Build the macro-enabled workbook once for the whole session."""
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    out = str(d / "Mortgage_Delinquency_Monitor.xlsm")
    BW.build(base)
    assemble_xlsm.assemble(base, out)
    return out


@pytest.fixture(scope="session")
def populated(xlsm, tmp_path_factory):
    """A fresh copy of the workbook with a --demo run applied
    (session-scoped: the read-only assertions share one copy)."""
    d = tmp_path_factory.mktemp("run")
    p = str(d / "run.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    return p


def _county(slot=13, fips="12345", name="Test County", state="TX",
            active=True):
    return R.FootprintRow(slot=slot, fips=fips, name=name, state=state,
                          active=active)


def _edit_setting(path, key, value):
    """Edit a [SETTINGS] value in a BUILT workbook exactly as a user would."""
    wb = openpyxl.load_workbook(path, keep_vba=path.endswith(".xlsm"))
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == key:
            target = row[0].row
            break
    assert target is not None, f"[SETTINGS] key {key} not found"
    ws.cell(target, 2, value)
    wb.save(path)
    wb.close()


# --------------------------------------------------------------------------
# Phase 1 -- config parse (incl. [FOOTPRINT] + capacity refusals)
# --------------------------------------------------------------------------
def test_config_parse():
    cfg = R.parse_config(BW.config_rows())
    # [SERIES]: 2 national + 20 state (10 seed states x 2 measures) + the
    # 2-measure county TEMPLATE rows
    assert len(cfg.series) == 24
    assert [s.id for s in R.national_series(cfg)] == ["nat_d3089", "nat_d90"]
    st = R.state_series(cfg)
    assert len(st) == 20
    assert [s.id for s in R.county_templates(cfg)] == ["co_d3089", "co_d90"]
    R.validate_series(cfg.series)                      # registry agreement
    # only the county templates are watchlist-capable (aggregates never are)
    assert [s.id for s in cfg.series if s.watchlist_capable] == \
        ["co_d3089", "co_d90"]
    for s in cfg.series:
        assert s.source_class == "A"
        assert R.series_measure(s) in R.MEASURES
    # thresholds: the three measure/transform bands, direction above
    assert set(cfg.thresholds) == set(R.THRESHOLD_IDS)
    assert cfg.thresholds["d3089_dev12m"].watch == pytest.approx(0.3)
    assert cfg.thresholds["d3089_dev12m"].alert == pytest.approx(0.6)
    assert cfg.thresholds["d90_streak3"].watch == pytest.approx(2)
    assert cfg.thresholds["d90_streak3"].alert == pytest.approx(3)
    assert cfg.thresholds["d90_level"].alert == pytest.approx(3.0)
    assert all(t.direction == "above" for t in cfg.thresholds.values())
    # [FOOTPRINT]: 40 provisioned slots, 12 seeded counties, FIPS 5-digit TEXT
    assert cfg.footprint_slots == 40 and len(cfg.footprint) == 40
    filled = [f for f in cfg.footprint if f.has_county]
    assert len(filled) == 12
    assert [f.slot for f in cfg.footprint] == list(range(1, 41))
    for f in filled:
        assert isinstance(f.fips, str) and len(f.fips) == 5 \
            and f.fips.isdigit(), f.fips
        assert re.match(R.ENTITY_KEY_PATTERN, f.entity_key)
    baldwin = next(f for f in filled if f.slot == 1)
    assert baldwin.fips == "01003" and baldwin.entity_key == "county:01003"
    assert baldwin.state == "AL" and baldwin.active is True
    for f in cfg.footprint[12:]:
        assert not f.has_county and not f.active       # provisioned headroom
    # settings the runner depends on
    assert cfg.raw_slots == 72
    assert cfg.continuity_months == 9
    assert cfg.vintage_override == ""
    # state slots derived from the seed footprint states, appearance order
    smap = R.state_slot_map(cfg)
    assert smap == {"01": 1, "06": 2, "17": 3, "48": 4, "04": 5, "12": 6,
                    "36": 7, "39": 8, "42": 9, "53": 10}
    # expansion footprint x measures -> the co_{slot:02d}_{measure} unit ids
    admitted, refusals, excluded = R.evaluate_footprint(cfg)
    assert len(admitted) == 12 and not refusals and not excluded
    spec = R.make_county_spec(baldwin, "d90")
    assert spec.id == "co_01_d90" and spec.geo_segment == "county:01003"
    unit_ids = {R.make_county_spec(f, m).id for f in admitted
                for m in R.MEASURES}
    assert len(unit_ids) == 24
    assert {"co_01_d3089", "co_12_d90", "co_08_d90"} <= unit_ids
    assert R.make_state_spec(1, "01", "d3089").id == "st_01_d3089"
    assert R.make_national_spec("d90").id == "nat_d90"
    # FIPS normalization (trap C1): numeric cells re-padded, text untouched
    assert R._norm_fips(1003) == "01003"
    assert R._norm_fips(1003.0) == "01003"
    assert R._norm_fips("01003") == "01003"
    assert R._norm_fips("'01003'") == "01003"
    assert R._norm_fips("ABC12") == "ABC12"            # reaches the gate
    # capacity is validated, never truncated: a 41st county refuses with the
    # exact rebuild command
    cfg.footprint.append(_county(slot=41, fips="48113",
                                 name="Overflow County"))
    with pytest.raises(R.FootprintCapacityError) as ei:
        R.validate_footprint_capacity(cfg)
    msg = str(ei.value)
    assert "Overflow County" in msg and "--footprint-slots 41" in msg
    assert "make_workbook.py" in msg and "never truncates" in msg
    # duplicate slots refuse too
    cfg2 = R.parse_config(BW.config_rows())
    cfg2.footprint.append(_county(slot=1, fips="48113", name="Dup County"))
    with pytest.raises(R.FootprintCapacityError, match="listed twice"):
        R.validate_footprint_capacity(cfg2)
    # more distinct states than the 15 provisioned slots refuses
    cfg3 = R.parse_config(BW.config_rows())
    tmpl = st[0]
    for i, f2 in enumerate(["08", "10", "13", "16", "20", "31"]):
        cfg3.series.append(replace(tmpl, id=f"st_{f2}_d3089",
                                   geo_segment=f"state:{f2}"))
    with pytest.raises(R.StateCapacityError, match="15 state"):
        R.state_slot_map(cfg3)


# --------------------------------------------------------------------------
# Phase 2 -- DemoProvider determinism + live-provider parse (no network)
# --------------------------------------------------------------------------
def test_demo_provider_deterministic():
    cfg = R.parse_config(BW.config_rows())
    filled = {f.slot: f for f in cfg.footprint if f.has_county}
    prov = R.CfpbDemoProvider(asof=ASOF, raw_slots=cfg.raw_slots)
    assert prov.vintage == "thru-2025-09" and prov.vintage_source == "demo"
    for slot, m in ((1, "d3089"), (5, "d90"), (8, "d90"), (11, "d3089")):
        spec = R.make_county_spec(filled[slot], m)
        a = prov.fetch_series(spec)
        b = R.CfpbDemoProvider(asof=ASOF,
                               raw_slots=cfg.raw_slots).fetch_series(spec)
        assert [(r.period, r.value) for r in a] == \
               [(r.period, r.value) for r in b]
        assert len(a) == cfg.raw_slots
        assert a[-1].period == "2025-09"     # months end AT the vintage
        assert a[0].period == "2019-10"
        assert a[0].geo_segment == filled[slot].entity_key
    # values live in the realistic percent ranges with one interior None
    vals = [r.value for r in prov.fetch_series(
        R.make_county_spec(filled[1], "d3089"))]
    nums = [v for v in vals if v is not None]
    assert all(0.05 <= v <= 8.0 for v in nums)
    assert vals.count(None) == 1
    assert None not in vals[-13:]            # never inside the 13m window
    # the deliberately-omitted county (seed slot 12, King WA) is ABSENT
    assert prov.fetch_series(
        R.make_county_spec(filled[12], "d90")) == []
    assert filled[12].fips == R.CfpbDemoProvider.SUPPRESSED_FIPS
    # national + state series exist
    assert len(prov.fetch_series(R.make_national_spec("d90"))) == 72
    assert len(prov.fetch_series(R.make_state_spec(1, "01", "d3089"))) == 72
    # a DIFFERENT vintage shifts the month grid AND revises overlapping
    # history (trap C3: full-history revisions are real)
    old = R.CfpbDemoProvider(asof=ASOF, raw_slots=cfg.raw_slots,
                             vintage="thru-2025-03")
    assert old.vintage == "thru-2025-03"
    rows_old = old.fetch_series(R.make_county_spec(filled[1], "d90"))
    rows_new = prov.fetch_series(R.make_county_spec(filled[1], "d90"))
    assert rows_old[-1].period == "2025-03"
    by_old = {r.period: r.value for r in rows_old}
    by_new = {r.period: r.value for r in rows_new}
    common = sorted(set(by_old) & set(by_new))
    assert len(common) >= 60
    assert any(by_old[p] != by_new[p] for p in common
               if by_old[p] is not None and by_new[p] is not None)
    # a bare vintage is normalized; a malformed one refuses
    assert R.CfpbDemoProvider(vintage="2025-06").vintage == "thru-2025-06"
    with pytest.raises(RuntimeError, match="thru-YYYY-MM"):
        R.CfpbDemoProvider(vintage="latest")


# A FIXED fixture built from the REAL inspected schema
# (COVERAGE_RESEARCH_CFPB.md): header RegionType,State,Name,FIPSCode plus
# YYYY-MM month columns; National row FIPSCode '-----'; QUOTED county FIPS;
# percent floats; a blank cell (missing month != 0); a MetroArea row that
# must be filtered out.
FIXTURE_MONTHS = "2008-01,2008-02,2025-06,2025-07,2025-08,2025-09"

COUNTY_CSV = (
    "RegionType,State,Name,FIPSCode," + FIXTURE_MONTHS + "\n"
    "National,,United States,-----,3.5,3.4,1.1,1.2,1.15,1.18\n"
    "County,AL,Baldwin County,'01003',2.8,2.7,1.4,1.5,,1.62\n"
    "County,AL,Jefferson County,'01073',3.0,3.1,1.8,1.7,1.75,1.9\n"
    "MetroArea,IL,Chicago-Naperville-Elgin,'16980',2.2,2.1,1.0,1.0,1.0,1.0\n"
)

STATE_CSV = (
    "RegionType,State,Name,FIPSCode," + FIXTURE_MONTHS + "\n"
    "National,,United States,-----,3.5,3.4,1.1,1.2,1.15,1.18\n"
    "State,AL,Alabama,'01',2.9,2.8,1.6,1.55,1.6,1.65\n"
    "State,CA,California,'06',3.2,3.0,0.9,0.95,0.9,0.88\n"
)

# The filename-discovery fixture: a small HTML snippet containing dated CSV
# hrefs (two vintages -- discovery must take the NEWEST).
PAGE_HTML = """<html><body><h1>Download the data</h1><ul>
<li><a href="https://files.consumerfinance.gov/data/mortgage-performance/downloads/CountyMortgagesPercent-30-89DaysLate-thru-2025-09.csv">County 30-89</a></li>
<li><a href="https://files.consumerfinance.gov/data/mortgage-performance/downloads/CountyMortgagesPercent-90-plusDaysLate-thru-2025-09.csv">County 90+</a></li>
<li><a href="https://files.consumerfinance.gov/data/mortgage-performance/downloads/StateMortgagesPercent-30-89DaysLate-thru-2025-09.csv">State 30-89</a></li>
<li><a href="https://files.consumerfinance.gov/data/mortgage-performance/downloads/StateMortgagesPercent-90-plusDaysLate-thru-2025-09.csv">State 90+</a></li>
<li>archived: <a href="/downloads/CountyMortgagesPercent-30-89DaysLate-thru-2025-03.csv">old vintage</a></li>
</ul></body></html>"""


def _fixture_get(calls):
    def fake_get(self, url):
        calls.append(url)
        if "download-the-data" in url:
            return PAGE_HTML.encode("utf-8")
        if "County" in url:
            return COUNTY_CSV.encode("utf-8")
        if "State" in url:
            return STATE_CSV.encode("utf-8")
        raise AssertionError(f"unexpected URL {url}")
    return fake_get


def test_cfpb_provider_parse(monkeypatch):
    """REAL-schema fixture: quoted FIPS stripped (kept TEXT), National row
    filtered by RegionType into its own series, wide->long by the YYYY-MM
    header pattern, percent floats, blank cell -> None, absent county -> []
    (the SUPPRESSED source, never zero), per-URL cache. NO network."""
    calls = []
    monkeypatch.setattr(R.CfpbProvider, "_http_get", _fixture_get(calls))
    prov = R.CfpbProvider(min_interval=0.0)
    prov.prime(ASOF)
    # stage 1 (page) + stage 2 (4 CSVs) = 5 GETs total
    assert len(calls) == 5
    assert "download-the-data" in calls[0]
    assert prov.vintage == "thru-2025-09"
    assert prov.vintage_source == "discovered"
    csv_urls = calls[1:]
    assert all(u.startswith(R.FILES_BASE) for u in csv_urls)
    assert sorted(u.rsplit("/", 1)[1] for u in csv_urls) == sorted([
        "StateMortgagesPercent-30-89DaysLate-thru-2025-09.csv",
        "StateMortgagesPercent-90-plusDaysLate-thru-2025-09.csv",
        "CountyMortgagesPercent-30-89DaysLate-thru-2025-09.csv",
        "CountyMortgagesPercent-90-plusDaysLate-thru-2025-09.csv"])
    # county series: quoted FIPS matched AFTER stripping, values by month
    fp = _county(slot=1, fips="01003", name="Baldwin County", state="AL")
    rows = prov.fetch_series(R.make_county_spec(fp, "d3089"))
    assert [r.period for r in rows] == \
        ["2008-01", "2008-02", "2025-06", "2025-07", "2025-08", "2025-09"]
    assert rows[0].value == pytest.approx(2.8)
    assert rows[-1].value == pytest.approx(1.62)
    assert rows[4].value is None                # blank cell -> None, NOT 0
    assert rows[0].geo_segment == "county:01003"
    assert rows[0].source_class == "A" and rows[0].units == "pct"
    # National row: RegionType filter, FIPSCode '-----' never a county
    nat = prov.fetch_series(R.make_national_spec("d90"))
    assert [r.value for r in nat][:2] == [3.5, 3.4]
    parsed = prov._data["County"]["d3089"]
    assert "-----" not in parsed["geos"]
    assert "16980" not in parsed["geos"]        # MetroArea filtered out
    assert set(parsed["geos"]) == {"01003", "01073"}
    assert parsed["geos"]["01003"]["name"] == "Baldwin County"
    # state series come from the state file
    st = prov.fetch_series(R.make_state_spec(2, "06", "d90"))
    assert st[-1].value == pytest.approx(0.88)
    # ABSENT county -> [] -- the suppression source (trap C4)
    ghost = _county(slot=2, fips="06037", name="Los Angeles County",
                    state="CA")
    assert prov.fetch_series(R.make_county_spec(ghost, "d90")) == []
    # raw_slots windows to the NEWEST months (header order, not position)
    small = R.CfpbProvider(min_interval=0.0, raw_slots=3)
    monkeypatch.setattr(R.CfpbProvider, "_http_get", _fixture_get(calls))
    small.prime(ASOF)
    rows3 = small.fetch_series(R.make_county_spec(fp, "d3089"))
    assert [r.period for r in rows3] == ["2025-07", "2025-08", "2025-09"]
    # per-URL cache: re-priming makes NO further calls
    n = len(calls)
    prov.prime(ASOF)
    assert len(calls) == n
    # fetch before prime is a programming-error refusal, not empty data
    prov3 = R.CfpbProvider(min_interval=0.0)
    with pytest.raises(RuntimeError, match="prime"):
        prov3.fetch_series(R.make_national_spec("d90"))
    # a schema drift (missing FIPSCode column) refuses loudly
    with pytest.raises(RuntimeError, match="FIPSCode"):
        R.parse_wide_csv("RegionType,State,Name,2020-01\nCounty,AL,X,1.0\n")
    with pytest.raises(RuntimeError, match="month"):
        R.parse_wide_csv("RegionType,State,Name,FIPSCode\nCounty,AL,X,'01'\n")


def test_filename_discovery(monkeypatch):
    """Stage-1 regex over the download-page fixture (two vintages present:
    take the newest) + the `vintage` override path (no scrape at all)."""
    # the regex itself, on the raw fixture
    found = R.FILENAME_RE.findall(PAGE_HTML)
    assert ("County", "30-89", "2025-09") in found
    assert ("State", "90-plus", "2025-09") in found
    assert ("County", "30-89", "2025-03") in found     # the older vintage
    calls = []
    monkeypatch.setattr(R.CfpbProvider, "_http_get", _fixture_get(calls))
    prov = R.CfpbProvider(min_interval=0.0)
    assert prov.discover() == "2025-09"                # newest wins
    assert prov.vintage_source == "discovered"
    # override: NO page call, URLs constructed directly
    calls2 = []
    monkeypatch.setattr(R.CfpbProvider, "_http_get", _fixture_get(calls2))
    prov2 = R.CfpbProvider(min_interval=0.0, vintage_override="thru-2025-09")
    assert prov2.discover() == "2025-09"
    assert prov2.vintage_source == "override"
    assert calls2 == []                                # scrape skipped
    prov2.prime(ASOF)
    assert len(calls2) == 4                            # only the CSVs
    assert all("download-the-data" not in u for u in calls2)
    # bare YYYY-MM override accepted; junk refused with the fix named
    assert R.CfpbProvider(vintage_override="2024-06").discover() == "2024-06"
    with pytest.raises(RuntimeError, match="thru-YYYY-MM"):
        R.CfpbProvider(vintage_override="latest").discover()
    # a page with no dated filenames refuses and points at the override knob
    def blank_page(self, url):
        return b"<html>nothing here</html>"
    monkeypatch.setattr(R.CfpbProvider, "_http_get", blank_page)
    with pytest.raises(RuntimeError, match="vintage"):
        R.CfpbProvider(min_interval=0.0).discover()


def test_provider_backoff(monkeypatch):
    """429 is retried with backoff and then succeeds -- stubbed, no network."""
    import urllib.error
    attempts = []

    def flaky_get(self, url):
        attempts.append(url)
        if len(attempts) < 3:
            raise urllib.error.HTTPError(url, 429, "Too Many Requests",
                                         None, None)
        return _fixture_get([])(self, url)

    sleeps = []
    monkeypatch.setattr(R.CfpbProvider, "_http_get", flaky_get)
    monkeypatch.setattr(R.time, "sleep", lambda s: sleeps.append(s))
    prov = R.CfpbProvider(min_interval=0.0, max_retries=4)
    prov.prime(ASOF)
    assert len(attempts) == 7              # 2 x 429 + page + 4 CSVs
    assert sleeps                          # backoff actually slept
    fp = _county(slot=1, fips="01003", name="Baldwin County", state="AL")
    assert len(prov.fetch_series(R.make_county_spec(fp, "d90"))) == 6
    # a persistent 503 fails with a labeled error, never an endless retry
    def dead_get(self, url):
        raise urllib.error.HTTPError(url, 503, "down", None, None)
    monkeypatch.setattr(R.CfpbProvider, "_http_get", dead_get)
    prov2 = R.CfpbProvider(min_interval=0.0, max_retries=1)
    with pytest.raises(RuntimeError, match="HTTP 503"):
        prov2.prime(ASOF)


def test_lookup_offline(capsys):
    """--lookup with --demo searches the embedded mini-index offline (the
    offline-FRIENDLY requirement) and prints County/State/FIPS."""
    rc = R.main(["--lookup", "King", "--demo"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "53033" in out and "WA" in out and "King County" in out
    assert "[FOOTPRINT]" in out and "TEXT" in out      # the how-to hint
    rc = R.main(["--lookup", "Xyzzyville", "--demo"])
    out = capsys.readouterr().out
    assert rc == 0 and "No counties matched" in out
    assert "SUPPRESSED" in out                         # the rural-county note
    # and the runner still requires a workbook when not looking up
    rc = R.main(["--demo"])
    err = capsys.readouterr().err
    assert rc == 2 and "--workbook" in err


# --------------------------------------------------------------------------
# Phase 3 -- transforms (hardcoded expectations, Excel parity semantics)
# --------------------------------------------------------------------------
def test_transforms():
    # dev_12m = v0 - AVERAGE(v1:v12), newest-first input
    assert R.t_dev_12m([2.0] + [1.0] * 12) == pytest.approx(1.0)
    assert R.t_dev_12m([1.5, 1.0, 2.0, 1.0, 2.0, 1.0, 2.0, 1.0, 2.0, 1.0,
                        2.0, 1.0, 2.0]) == pytest.approx(0.0)
    # AVERAGE ignores blanks: None inside the window drops out of the mean
    assert R.t_dev_12m([2.0, 1.0, None, 3.0]) == pytest.approx(0.0)
    # blank guards: v0 missing, or nothing numeric in the window
    assert R.t_dev_12m([None] + [1.0] * 12) is None
    assert R.t_dev_12m([2.0]) is None
    assert R.t_dev_12m([2.0, None, None]) is None
    assert R.t_dev_12m([]) is None
    # extra history beyond v12 is ignored (window is exactly v1..v12)
    assert R.t_dev_12m([2.0] + [1.0] * 12 + [99.0] * 10) == pytest.approx(1.0)

    # rise_streak3: capped at 3, strict rises, IF-chain identical to Excel
    assert R.t_rise_streak3([4.0, 3.0, 2.0, 1.0]) == 3
    assert R.t_rise_streak3([5.0, 4.0, 3.0, 2.0, 1.0, 0.5]) == 3   # the cap
    assert R.t_rise_streak3([4.0, 3.0, 3.5, 1.0]) == 1
    assert R.t_rise_streak3([4.0, 3.0, 2.0, 2.5]) == 2
    assert R.t_rise_streak3([3.0, 4.0, 2.0, 1.0]) == 0
    assert R.t_rise_streak3([2.0, 2.0, 1.0, 0.5]) == 0      # equal != rise
    # NaN/None tolerance: any blank inside v0..v3 blanks the streak
    assert R.t_rise_streak3([4.0, None, 2.0, 1.0]) is None
    assert R.t_rise_streak3([4.0, 3.0, 2.0]) is None        # too short
    # broken interior chain: v0>v1 false kills later terms too (AND chain)
    assert R.t_rise_streak3([4.0, 5.0, 3.0, 2.0]) == 0

    # level + yoy
    assert R.t_level([1.62, 1.5]) == pytest.approx(1.62)
    assert R.t_level([None, 1.5]) is None
    assert R.t_yoy_pct([2.2] + [1.0] * 11 + [2.0]) == pytest.approx(10.0)
    assert R.t_yoy_pct([2.2] + [1.0] * 11 + [0.0]) is None
    assert R.t_yoy_pct([2.2] * 5) is None

    # threshold engine, direction-aware (matches the Excel band rules)
    lvl = R.Threshold(watch=1.5, alert=3.0, direction="above")
    assert R.status_for(3.2, lvl) == "ALERT"
    assert R.status_for(3.0, lvl) == "ALERT"                # >= is inclusive
    assert R.status_for(1.6, lvl) == "WATCH"
    assert R.status_for(0.4, lvl) == "OK"
    assert R.status_for(None, lvl) == "OK"
    stk = R.Threshold(watch=2, alert=3, direction="above")
    assert R.status_for(3, stk) == "ALERT"
    assert R.status_for(2, stk) == "WATCH"
    assert R.status_for(1, stk) == "OK"
    below = R.Threshold(watch=1.0, alert=0.5, direction="below")
    assert R.status_for(0.4, below) == "ALERT"

    # registry drift refuses (unknown measure / transform / geo shape)
    cfg = R.parse_config(BW.config_rows())
    with pytest.raises(R.SeriesError, match="measure"):
        R.validate_series([replace(cfg.series[0], id="nat_dpd120")])
    with pytest.raises(R.SeriesError, match="transform"):
        R.validate_series([replace(cfg.series[0], transform="zscore_8q")])
    with pytest.raises(R.SeriesError, match="geo_segment"):
        R.validate_series([replace(cfg.series[0], geo_segment="metro:16980")])


# --------------------------------------------------------------------------
# Phase 4 -- headless reload (Excel proxy) + NO native charts
# --------------------------------------------------------------------------
def test_reload_headless(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    # macro survived the openpyxl round-trip
    assert wb.vba_archive is not None
    # tab set EXACT (contract sec 2)
    assert tuple(wb.sheetnames) == TABS
    # NO native chart objects anywhere (L4)
    for ws in wb.worksheets:
        assert getattr(ws, "_charts", []) == []
    cfg = R.parse_config(BW.config_rows())
    # Watchlist: every built slot row carries the formula set; the Status
    # formula renders REFUSED/OFF/SUPPRESSED + the --lookup hint by formula
    ws = wb["Watchlist"]
    for slot in (1, 12, 40):
        r = BW.wl_row(slot)
        for c in (1, 2, 3, 4, 5, 6, 7, 8, 9):
            v = ws.cell(r, c).value
            assert isinstance(v, str) and v.startswith("="), (slot, c)
        st = str(ws.cell(r, 8).value)
        assert "REFUSED" in st and "--lookup" in st
        assert "SUPPRESSED (below sample threshold)" in st and "OFF" in st
        assert "RANK" in str(ws.cell(r, 9).value)
        # identity flows BY FORMULA from the [FOOTPRINT] cells
        assert "_config" in str(ws.cell(r, 1).value)
        assert "_config" in str(ws.cell(r, 3).value)
        # dev_12m / streak formulas match the registry definitions
        assert "AVERAGE(" in str(ws.cell(r, 5).value)
        assert str(ws.cell(r, 7).value).count("IF(") >= 4
    # Dashboard_State: National pinned first; state slots read runtime
    # identity from the raw block headers
    ds = wb["Dashboard_State"]
    assert ds.cell(BW.DS_FIRST, 1).value == "NATIONAL (benchmark)"
    assert "Raw_CFPB" in str(ds.cell(BW.ds_row(1), 1).value)
    for c in (3, 4, 5, 6, 7, 8):
        assert str(ds.cell(BW.DS_FIRST, c).value).startswith("=")
    # Dashboard_Trends: month headers + INDEX readout cells + MATCH helper
    tr = wb["Dashboard_Trends"]
    assert "Raw_CFPB" in str(tr.cell(BW.TR_HDR, BW.TR_FIRST_MCOL).value)
    assert "LARGE(" in str(tr.cell(BW.TR_FIRST + 1, BW.TR_SLOT_COL).value)
    assert "INDEX(" in str(tr.cell(BW.TR_FIRST + 1, BW.TR_FIRST_MCOL).value)
    assert str(tr.cell(BW.TR_FIRST, BW.TR_FIRST_MCOL).value).startswith("=")
    # threshold cells are NUMERIC (L8: Excel's number>=text comparison is
    # silently FALSE and would kill every band rule)
    cws = wb["_config"]
    in_thr = False
    checked = 0
    for row in cws.iter_rows(min_col=1, max_col=3):
        v = row[0].value
        if v == "[THRESHOLDS]":
            in_thr = True
            continue
        if in_thr and isinstance(v, str) and v.startswith("["):
            break
        if in_thr and v in R.THRESHOLD_IDS:
            assert isinstance(row[1].value, (int, float)), v
            assert isinstance(row[2].value, (int, float)), v
            checked += 1
    assert checked == 3
    # seeded FIPS cells are TEXT (trap C1)
    in_fp = False
    for row in cws.iter_rows(min_col=1, max_col=2):
        v = row[0].value
        if v == "[FOOTPRINT]":
            in_fp = True
            continue
        if in_fp and isinstance(v, str) and v.startswith("["):
            break
        if in_fp and str(v).strip() == "1":
            assert row[1].value == "01003" and isinstance(row[1].value, str)
            break
    # the raw scaffold labels every geo block
    raw = wb["Raw_CFPB"]
    for kind, slot in (("national", 0), ("state", 7), ("county", 7),
                       ("county", 40)):
        b = R.geo_block(kind, slot, cfg.raw_slots)
        assert raw.cell(b.header_row, 1).value == R.slot_label(kind, slot)
        assert raw.cell(b.label_row, 1).value == "month"
        assert raw.cell(b.label_row, 2).value == "d3089_pct"
        assert raw.cell(b.label_row, 3).value == "d90_pct"
    wb.close()


# --------------------------------------------------------------------------
# Phase 5 -- county gates: positive admission + every refusal shape
# --------------------------------------------------------------------------
def test_watchlist_county_gates():
    cfg = R.parse_config(BW.config_rows())

    # POSITIVE: all 12 seeded FIPS-keyed counties are admitted
    admitted, refusals, excluded = R.evaluate_footprint(cfg)
    assert len(admitted) == 12 and refusals == [] and excluded == []
    assert all(re.match(R.ENTITY_KEY_PATTERN, f.entity_key)
               for f in admitted)

    # blank fips: refused BY NAME with the --lookup hint (never fetched)
    blank = _county(slot=13, fips="", name="Mystery County", state="ZZ")
    cfg.footprint[12] = blank
    admitted2, refusals2, _ = R.evaluate_footprint(cfg)
    assert len(admitted2) == 12                    # the good rows still pass
    assert len(refusals2) == 1
    msg = refusals2[0][1]
    assert 'footprint slot 13 "Mystery County"' in msg
    assert 'fips=""' in msg
    assert "county:[0-9]{5}" in msg                # the whitelist, named
    assert '--lookup "Mystery County"' in msg and "--demo" in msg
    assert "never watchlist" in msg                # aggregates stay out

    # inactive slot: EXCLUDED (blanked), not refused
    cfg3 = R.parse_config(BW.config_rows())
    idx = next(i for i, f in enumerate(cfg3.footprint) if f.slot == 2)
    cfg3.footprint[idx] = replace(cfg3.footprint[idx], active=False)
    admitted3, refusals3, excluded3 = R.evaluate_footprint(cfg3)
    assert len(admitted3) == 11 and refusals3 == []
    assert [f.slot for f in excluded3] == [2]

    # malformed keys are refused default-deny (text stays as typed -- only
    # NUMERIC cells are re-padded)
    for fips in ("1234", "123456", "ABC12", "12-34", "01 03", "county:01003"):
        reasons = R.gate_footprint_row(
            _county(slot=14, fips=fips, name="Bad Key"))
        assert reasons and "Gate3" in reasons[0], fips

    # defense in depth: a fabricated watchlist-capable AGGREGATE row refuses
    # the run, series-named (states/metros/national are context, never
    # watchlist -- BUILD SPEC 0.1)
    cfg4 = R.parse_config(BW.config_rows())
    state_row = next(s for s in cfg4.series if s.id == "st_01_d3089")
    with pytest.raises(R.WatchlistRefused) as ei:
        R.assert_series_admissible([replace(state_row,
                                            watchlist_capable=True)])
    assert 'series "st_01_d3089"' in str(ei.value)
    assert "never watchlist" in str(ei.value)
    # ... and a non-Class-A county template refuses too
    co = next(s for s in cfg4.series if s.id == "co_d90")
    with pytest.raises(R.WatchlistRefused) as ei2:
        R.assert_series_admissible([replace(co, source_class="C")])
    assert "its own review" in str(ei2.value)

    # the build-time hard gate backs the runtime gate
    cfg5 = R.parse_config(BW.config_rows())
    cfg5.footprint[13] = _county(slot=14, fips="not-a-fips",
                                 name="Aggregate Row")
    with pytest.raises(R.WatchlistRefused, match="Aggregate Row"):
        R.assert_entity_gates(cfg5)


# --------------------------------------------------------------------------
# Phase 6 -- suppression (trap C4: absent county != zero)
# --------------------------------------------------------------------------
def test_suppressed_county(xlsm, tmp_path):
    p = str(tmp_path / "sup.xlsm")
    shutil.copy(xlsm, p)
    status = R.run(p, demo=True, asof=ASOF)
    # the demo deliberately omits seed slot 12 (King County WA 53033)
    assert status["counties_suppressed"] == 1
    assert status["counties_landed"] == 11
    assert status["suppressed_counties"] == ["co12 county:53033 King County"]
    king = next(c for c in status["digest"]["counties"] if c["slot"] == 12)
    assert king["suppressed"] is True
    assert king["status"] == "SUPPRESSED"          # never blank, never zero
    assert king["d90_latest"] is None and king["severity"] is None
    # EXCLUDED from the alert KPIs: only live counties count
    live = [c for c in status["digest"]["counties"] if not c["suppressed"]]
    assert status["alert_counties"] == \
        sum(1 for c in live if c["status"] == "ALERT") == 2
    assert status["watch_counties"] == \
        sum(1 for c in live if c["status"] == "WATCH") == 2
    # the raw block is explicit: identity + SUPPRESSED note, data blank
    wb = openpyxl.load_workbook(p, keep_vba=True)
    cfg = R.parse_config([[c.value for c in row]
                          for row in wb["_config"].iter_rows()])
    b12 = R.geo_block("county", 12, cfg.raw_slots)
    raw = wb["Raw_CFPB"]
    assert raw.cell(b12.header_row, 2).value == "co12 county:53033"
    assert raw.cell(b12.header_row, 3).value == "King County"
    assert "SUPPRESSED" in str(raw.cell(b12.header_row, 4).value)
    assert "state row" in str(raw.cell(b12.header_row, 4).value)
    for r in range(b12.first_data_row, b12.first_data_row + 5):
        for c in (1, 2, 3):
            assert raw.cell(r, c).value is None
    # a landed neighbor block has data (the contrast)
    b11 = R.geo_block("county", 11, cfg.raw_slots)
    assert raw.cell(b11.first_data_row, 2).value is not None
    wb.close()
    # present in the email, explicitly, with the state-row fallback hint
    email = email_sim.compose_email(status)
    assert "SUPPRESSED FOOTPRINT COUNTIES" in email
    assert "King County" in email and "use the WA state row" in email
    assert "absence is NOT zero" in email


# --------------------------------------------------------------------------
# Phase 7 -- vintage recording + continuity tripwire (BUILD SPEC 0.6)
# --------------------------------------------------------------------------
def test_vintage_continuity(xlsm, tmp_path):
    # the primitives
    assert R.vintage_ym("thru-2025-09") == (2025, 9)
    assert R.vintage_ym("2025-09") == (2025, 9)
    assert R.vintage_ym("demo data") is None
    assert R.months_behind(date(2026, 4, 30), "thru-2025-09") == 7
    assert R.continuity_warning(date(2026, 4, 30), "thru-2025-09", 9) is None
    w = R.continuity_warning(date(2026, 9, 30), "thru-2025-09", 9)
    assert w and "CONTINUITY WARNING" in w and "12 months" in w
    assert "DataLumos" in w                       # the recovery-mirror hint

    # FRESH: vintage 7 months behind asof -> no warning, status panel OK
    p = str(tmp_path / "fresh.xlsm")
    shutil.copy(xlsm, p)
    st = R.run(p, demo=True, asof=ASOF)
    assert st["vintage"] == "thru-2025-09" and st["months_behind"] == 7
    assert st["continuity_warning"] is None
    wb = openpyxl.load_workbook(p, keep_vba=True)
    for tab in R.DASH_TABS:
        assert "thru-2025-09" in str(wb[tab].cell(3, R.STATUS_COL).value)
        assert "Continuity OK" in str(wb[tab].cell(4, R.STATUS_COL).value)
    wb.close()

    # STALE: pin the vintage via the _config `vintage` override (the demo
    # provider honors it too), then push asof 15 months past it -> WARNING
    # in the status dict, the status panel AND the email
    p2 = str(tmp_path / "stale.xlsm")
    shutil.copy(xlsm, p2)
    _edit_setting(p2, "vintage", "thru-2025-09")
    st2 = R.run(p2, demo=True, asof=date(2026, 12, 31))
    assert st2["vintage"] == "thru-2025-09"
    assert st2["months_behind"] == 15
    assert st2["continuity_warning"] and \
        "CONTINUITY WARNING" in st2["continuity_warning"]
    wb2 = openpyxl.load_workbook(p2, keep_vba=True)
    l4 = str(wb2["Dashboard_State"].cell(4, R.STATUS_COL).value)
    assert "CONTINUITY WARNING" in l4 and "tripwire 9" in l4
    wb2.close()
    email = email_sim.compose_email(st2)
    assert "CONTINUITY WARNING" in email
    # a fresh run over the stale workbook must OVERWRITE the warning (L7
    # discipline: explicit writes, no stale panel text)
    st3 = R.run(p2, demo=True, asof=ASOF)
    assert st3["continuity_warning"] is None
    wb3 = openpyxl.load_workbook(p2, keep_vba=True)
    assert "Continuity OK" in \
        str(wb3["Dashboard_State"].cell(4, R.STATUS_COL).value)
    wb3.close()


# --------------------------------------------------------------------------
# Phase 8 -- full replace (trap C3: every vintage revises the WHOLE history)
# --------------------------------------------------------------------------
def test_full_replace(xlsm, tmp_path):
    p = str(tmp_path / "replace.xlsm")
    shutil.copy(xlsm, p)
    cfg = R.parse_config(BW.config_rows())
    n = cfg.raw_slots

    def nat_months_landed():
        wb = openpyxl.load_workbook(p, keep_vba=True)
        b = R.geo_block("national", 0, n)
        ws = wb["Raw_CFPB"]
        months = [ws.cell(b.first_data_row + i, 1).value for i in range(n)]
        wb.close()
        return months

    def county1_value(month):
        wb = openpyxl.load_workbook(p, keep_vba=True)
        b = R.geo_block("county", 1, n)
        ws = wb["Raw_CFPB"]
        val = None
        for i in range(n):
            if ws.cell(b.first_data_row + i, 1).value == month:
                val = ws.cell(b.first_data_row + i, 3).value      # d90
        wb.close()
        return val

    # run 1: OLD vintage thru-2025-03
    old = R.CfpbDemoProvider(asof=date(2025, 11, 30), raw_slots=n,
                             vintage="thru-2025-03")
    R.run(p, demo=True, asof=date(2025, 11, 30), provider=old)
    months1 = nat_months_landed()
    assert months1[0] == "2025-03" and months1[-1] == "2019-04"
    v1 = county1_value("2024-12")
    assert v1 is not None

    # run 2: NEW vintage thru-2025-09 on the SAME workbook
    new = R.CfpbDemoProvider(asof=ASOF, raw_slots=n, vintage="thru-2025-09")
    st2 = R.run(p, demo=True, asof=ASOF, provider=new)
    assert st2["vintage"] == "thru-2025-09"
    months2 = nat_months_landed()
    assert months2[0] == "2025-09" and months2[-1] == "2019-10"
    # NO stale months survive: the old vintage's exclusive tail is gone
    old_only = set(months1) - set(months2)
    assert old_only == {"2019-04", "2019-05", "2019-06", "2019-07",
                        "2019-08", "2019-09"}
    assert not (set(months2) & old_only)
    # and the OVERLAPPING history was REPLACED, not diffed/kept: the cell
    # now holds the new vintage's revised value (trap C3)
    v2 = county1_value("2024-12")
    fresh = {r.period: r.value for r in new.fetch_series(
        R.make_county_spec(_county(slot=1, fips="01003",
                                   name="Baldwin County", state="AL"),
                           "d90"))}
    assert v2 == pytest.approx(fresh["2024-12"])
    assert v1 != v2                       # the revision actually shows


# --------------------------------------------------------------------------
# Phase 9 -- idempotence, layout guard, L7 clearing, VBA keys
# --------------------------------------------------------------------------
def test_raw_landing_idempotent(xlsm, tmp_path):
    def landed(path):
        wb = openpyxl.load_workbook(path, keep_vba=True)
        ws = wb["Raw_CFPB"]
        vals = [(c.row, c.column, c.value) for row in ws.iter_rows()
                for c in row if c.value is not None]
        wb.close()
        return vals

    # TRUE idempotence: a second run on the SAME already-populated workbook
    # lands identical values (clearing leaves no stale tail).
    p1 = str(tmp_path / "r1.xlsm")
    shutil.copy(xlsm, p1)
    R.run(p1, demo=True, asof=ASOF)
    first = landed(p1)
    R.run(p1, demo=True, asof=ASOF)
    assert landed(p1) == first
    # ... and determinism across fresh copies.
    p2 = str(tmp_path / "r2.xlsm")
    shutil.copy(xlsm, p2)
    R.run(p2, demo=True, asof=ASOF)
    assert landed(p2) == first
    # newest-first: every landed block's first data row is the vintage month
    wb = openpyxl.load_workbook(p1, keep_vba=True)
    cfg = R.parse_config([[c.value for c in row]
                          for row in wb["_config"].iter_rows()])
    for kind, slot in (("national", 0), ("state", 1), ("county", 1)):
        b = R.geo_block(kind, slot, cfg.raw_slots)
        assert wb["Raw_CFPB"].cell(b.first_data_row, 1).value == "2025-09"
        assert wb["Raw_CFPB"].cell(b.first_data_row, 3).value is not None
    wb.close()


def test_raw_layout_mismatch_refused(xlsm, tmp_path):
    """Editing raw_slots after build must be REFUSED (every formula is
    anchored to the built layout), not silently mis-written."""
    p = str(tmp_path / "mismatch.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)                # populate at raw_slots=72
    _edit_setting(p, "raw_slots", 24)
    with pytest.raises(RuntimeError, match="Raw layout mismatch"):
        R.run(p, demo=True, asof=ASOF)


def test_clear_actually_blanks(populated, tmp_path):
    """The L7 regression: openpyxl's ws.cell(r, c, None) silently IGNORES
    None -- clearing must assign .value explicitly. Clear a populated block
    directly and prove every cell actually blanked."""
    p = str(tmp_path / "clear.xlsm")
    shutil.copy(populated, p)
    backend = R.OpenpyxlBackend(p)
    cfg = backend.read_config()
    b1 = R.geo_block("county", 1, cfg.raw_slots)
    bn = R.geo_block("national", 0, cfg.raw_slots)
    ws = backend._wb["Raw_CFPB"]
    # populated before: identity + every data row carries values
    assert ws.cell(b1.header_row, 2).value == "co01 county:01003"
    n_before = sum(1 for r in range(b1.first_data_row,
                                    b1.first_data_row + cfg.raw_slots)
                   for c in (1, 2, 3) if ws.cell(r, c).value is not None)
    # full block minus the demo's one interior None per measure series
    assert n_before == cfg.raw_slots * 3 - 2
    backend.clear_geo_block(b1)
    backend.clear_geo_block(bn)
    backend.finalize()
    # reload from disk: the blanks must have SURVIVED the save
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws2 = wb["Raw_CFPB"]
    for b in (b1, bn):
        for c in (2, 3, 4):
            assert ws2.cell(b.header_row, c).value is None
        for r in range(b.first_data_row, b.first_data_row + cfg.raw_slots):
            for c in (1, 2, 3):
                assert ws2.cell(r, c).value is None
        # the build-time block label stays (the layout guard anchor)
        assert ws2.cell(b.header_row, 1).value == R.slot_label(b.kind, b.slot)
    wb.close()


def test_vba_protection_keys_roundtrip():
    """The PROJECT stream's CMG/DPB/GC must decrypt per [MS-OVBA] 2.4.3 to
    protection=none / password=none / visibility=visible -- exactly the way
    the VBE will read them (a malformed key triggers Excel's 'invalid key
    DPB' prompt)."""
    proj = V.build_project_stream(
        "X", [V.Module("X", "Sub A()\nEnd Sub")]).decode("latin-1")
    keys = {}
    for line in proj.splitlines():
        for k in ("CMG", "DPB", "GC"):
            if line.startswith(f'{k}="'):
                keys[k] = line.split('"')[1]
    assert V.ovba_decrypt(keys["CMG"]) == struct.pack("<I", 0)   # unprotected
    assert V.ovba_decrypt(keys["DPB"]) == b"\x00"               # no password
    assert V.ovba_decrypt(keys["GC"]) == b"\xff"                # visible
    # _VBA_PROJECT is the full 7-byte header ([MS-OVBA] 2.3.4.1)
    assert len(V.build_vba_project_stream()) == 7
    # the default VBA + Excel libraries are referenced in the dir stream
    d = V.build_dir_stream("X", [V.Module("X", "Sub A()\nEnd Sub")])
    assert b"Visual Basic For Applications" in d
    assert b"Microsoft Excel" in d
