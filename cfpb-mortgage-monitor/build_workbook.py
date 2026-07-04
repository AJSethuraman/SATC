#!/usr/bin/env python3
"""Build the County Mortgage Delinquency Monitor workbook (CFPB Mortgage
Performance Trends) -- KeyBank-styled.

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled
.xlsm. All visual style comes from keybank_style.py (tokens + helpers) --
never a hardcoded hex here. Built to BUILD_SPEC_CFPB.md (Sections 2-5).

THE SLOT MECHANISM (the FDIC [PEERS] pattern, wholesale): every Watchlist
row is anchored to a FOOTPRINT SLOT, not to a county. County identity
(name/FIPS/state) flows in BY FORMULA from the _config [FOOTPRINT] cells and
measure values BY FORMULA from the slot's fixed Raw_CFPB block -- so editing
WHICH county occupies a slot re-labels and re-points everything on recalc,
with zero rebuilt formulas. Capacity is a build knob: --footprint-slots
(default 40). State dashboard rows are provisioned to 15 fixed state slots;
their identity is runtime state the runner writes into the raw block headers.

Structure (Section 4):
  Dashboard_State  -- National (pinned first) + footprint states: both
                      measures, latest / dev_12m / 12m text-trend column
  Dashboard_Trends -- National + worst-5 footprint counties by 90+ level
                      (formula-ranked), 24-month readout as formula cells
  Watchlist        -- footprint counties ranked: County | State | FIPS |
                      30-89 latest | dev vs 12m | 90+ latest | rise streak |
                      Status | Rank (SUPPRESSED rendered explicitly)
  Raw_CFPB         -- one block per geo slot (nat, st01-15, co01-NN):
                      72 months x 2 measures, runner fills
  _config          -- [SETTINGS]/[THRESHOLDS]/[FOOTPRINT]/[SERIES]
  _code_py/_code_vba -- runner.py + macro.bas as flat ASCII text
  _readme          -- NMDB honesty, lag, suppression, revision + continuity

No native charts (L4). Embedded code is flat ASCII (L3). Every raw reference
is blank-guarded (empty cell renders blank, never 0). Heat direction: red =
rising delinquency (all three thresholds run above-is-bad).
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import runner as R
import series_seed as SEED

import keybank_style as KB
from keybank_style import (
    INK, KEY_RED, MONO_FONT, NOTE_FONT, SECT_FILL,
    brand_banner, kpi_tiles, header_row, watchlist_boundary,
    hide_gridlines, freeze_below,
)

HERE = os.path.dirname(os.path.abspath(__file__))
INK_BOLD = Font(name="Arial", bold=True, color=INK)
SMALL_BOLD = Font(name="Calibri", bold=True, size=9)

# ---- fixed geometry ------------------------------------------------------
DS_HDR = 7                      # Dashboard_State header row; National at 8
DS_FIRST = DS_HDR + 1
WL_HDR = 4                      # Watchlist header row; slot rows start at 5
WL_FIRST = WL_HDR + 1
WL_D90_COL = 6                  # F = 90+ latest (the trends ranking column)
WL_HELPER_DEV = 11              # K = d90 dev_12m helper (rank tiebreak)
WL_HELPER_SCORE = 12            # L = (severity, d90 dev) rank score
TR_HDR = 5                      # Dashboard_Trends header row; National at 6
TR_FIRST = TR_HDR + 1
TR_MONTHS = 24                  # 24-month readout, newest first
TR_FIRST_MCOL = 3               # months in C..Z
TR_SLOT_COL = 28                # AB = the MATCHed slot helper column

WL_COLS = ["County", "State", "FIPS", "30-89 latest", "30-89 dev 12m",
           "90+ latest", "90+ rise streak", "Status", "Rank"]
DS_COLS = ["Geo", "FIPS", "30-89 latest", "30-89 dev 12m", "30-89 trend 12m",
           "90+ latest", "90+ dev 12m", "90+ trend 12m"]


def ds_row(state_slot):
    """Dashboard_State row for a state slot (National sits at DS_FIRST)."""
    return DS_FIRST + state_slot


def wl_row(slot):
    return WL_FIRST + slot - 1


def stress_heat(ws, cell_range, direction="above"):
    """Direction-AWARE heat: red = stress. Every CFPB measure runs
    above-is-bad (rising delinquency = red), but the helper keeps the
    carried direction seam."""
    lo, hi = (KB.HEAT_GOOD, KB.HEAT_BAD) if direction != "below" \
        else (KB.HEAT_BAD, KB.HEAT_GOOD)
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="min", start_color=lo,
        mid_type="percentile", mid_value=50, mid_color=KB.HEAT_MID,
        end_type="max", end_color=hi))


def band_rules(ws, cell_range, first_cell, watch_ref, alert_ref, direction):
    """Per-cell WATCH/ALERT coloring straight off the _config [THRESHOLDS]
    cells (config-driven, direction-aware, blank-guarded via ISNUMBER)."""
    op = ">=" if direction != "below" else "<="
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{alert_ref})"],
        fill=KB.ALERT_FILL, font=KB.ALERT_FONT, stopIfTrue=True))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{watch_ref})"],
        fill=KB.TIGHTEN_FILL,
        font=Font(name="Calibri", bold=True, color=KB.SLATE),
        stopIfTrue=True))


def text_cell(ws, row, col, value):
    """Write literal text; openpyxl treats a leading '=' as a formula, so
    force it back to a string for source/doc lines."""
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# _config  (the knob panel: SETTINGS / THRESHOLDS / FOOTPRINT / SERIES)
# ==========================================================================
def config_rows(footprint_slots=R.FOOTPRINT_SLOTS_DEFAULT):
    rows = [["County Mortgage Delinquency Monitor (CFPB Mortgage Performance "
             "Trends) -- CONFIG (the knob panel). Edit values here; no code "
             "change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["demo_mode", "FALSE", "TRUE = offline deterministic "
         "CfpbDemoProvider (no network), for trying the button and tests. "
         "Live mode is KEYLESS too -- the CFPB files are public domain."],
        ["raw_slots", 72, "months kept per geo block (newest-first; 72 = 6 "
         "years). BUILT INTO every formula -- changing it requires "
         "rebuilding the workbook (make_workbook.py); the runner refuses a "
         "mismatched layout."],
        ["footprint_slots", footprint_slots, "the BUILT county capacity "
         "(informational; make_workbook.py --footprint-slots N sets it). "
         "More counties than this in [FOOTPRINT] = the runner refuses with "
         "the rebuild command -- never truncation."],
        ["http_min_interval", 0.6, "seconds between live CFPB requests "
         "(only ~5 per refresh; stay polite)."],
        ["cfpb_max_retries", 4, "retries with backoff on HTTP 429/5xx."],
        ["continuity_months", 9, "CONTINUITY tripwire: warn when the newest "
         "vintage is more than this many months behind --asof. The CFPB "
         "publishes semiannually-ish with a ~6-7 month lag and no formal "
         "promise; it survived the 2025 turmoil once."],
        ["vintage", "", "OPTIONAL override, e.g. thru-2025-09: skip the "
         "download-page scrape and fetch that vintage's files directly "
         "(useful if the page layout changes). Empty = discover the newest "
         "vintage from the page (filenames are DATED -- never hardcode)."],
        ["secret_env", "", "NAME of the env var holding a licensed (Class "
         "C) secret. Empty in v1 -- the CFPB files are keyless public "
         "domain, and the gates admit ONLY Class A."],
    ]
    rows += [[], ["[THRESHOLDS]"],
             ["id", "watch", "alert", "direction", "authority"]]
    for tid, watch, alert, direction, authority in SEED.THRESHOLDS:
        # watch/alert are NUMERIC cells (L8: Excel's number>=text comparison
        # is silently FALSE).
        rows.append([tid, watch, alert, direction, authority])
    rows += [[], ["[FOOTPRINT]"], list(SEED.FOOTPRINT_HEADER)]
    for frow in SEED.footprint_rows(footprint_slots):
        rows.append(frow)
    rows += [["# One county per row: slot | fips | name | state | active. "
              "Add a county = fill a free slot row + re-run the runner (no "
              "rebuild). Remove = active FALSE or clear the row (the slot "
              "blanks on the next run)."],
             ["# The seeded counties are ILLUSTRATIVE large counties -- "
              "replace them with your own collateral footprint. Find FIPS "
              "with: python runner.py --lookup \"<county name>\" (offline "
              "with --demo). Keep fips cells TEXT so leading zeros survive "
              "(trap C1)."],
             ["# Only ~470 large counties clear the CFPB's ~1,000-sample-"
              "mortgage threshold; a footprint county absent from the file "
              "shows SUPPRESSED -- use its state row (trap C4)."]]
    rows += [[], ["[SERIES]"], list(SEED.HEADER)]
    for r in SEED.all_series():
        rows.append([r[h] for h in SEED.HEADER])
    rows += [[], ["# nat_* and st_* rows are DASHBOARD CONTEXT (watchlist_"
                  "capable FALSE -- aggregates are never watchlist). The "
                  "st_* rows were derived from the seed footprint's states "
                  "at build time; if your footprint spans other states, add "
                  "matching st_{FIPS2}_{measure} rows (15 state slots are "
                  "provisioned)."],
             ["# co_d3089/co_d90 are the TEMPLATE rows the runner expands "
              "per active [FOOTPRINT] slot into units co_{slot}_{measure}. "
              "OUT of v1: the Metro file (mixed non-FIPS keys), FHFA NMDB "
              "fallback, any interpolation for suppressed counties."]]
    return rows


def write_config(wb, footprint_slots):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows(footprint_slots)
    thr_cells = {}                 # id -> (watch_ref, alert_ref, direction)
    set_cells = {}                 # settings key -> value-cell ref
    fp_cells = {}                  # slot -> dict(fips/name/state/active refs)
    in_thr = in_set = in_fp = False
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        head = row[0] if row else None
        if head == "[THRESHOLDS]":
            in_thr, in_set, in_fp = True, False, False
            continue
        if head == "[SETTINGS]":
            in_set, in_thr, in_fp = True, False, False
            continue
        if head == "[FOOTPRINT]":
            in_fp, in_thr, in_set = True, False, False
            continue
        if head is not None and isinstance(head, str) and head.startswith("["):
            in_thr = in_set = in_fp = False
        if not row or head is None or (isinstance(head, str)
                                       and str(head).startswith("#")):
            continue
        if in_set and head not in ("key",):
            set_cells[head] = f"_config!$B${i}"
        if in_thr and head not in ("id",):
            thr_cells[head] = (f"_config!$B${i}", f"_config!$C${i}",
                               str(row[3]).strip().lower())
        if in_fp and str(head).strip().lower() != "slot":
            slot = int(head)
            fp_cells[slot] = {
                "fips": f"_config!$B${i}", "name": f"_config!$C${i}",
                "state": f"_config!$D${i}", "active": f"_config!$E${i}"}
    ws["A1"].font = INK_BOLD
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("[") \
                and row[0].endswith("]"):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    return thr_cells, set_cells, fp_cells


# ==========================================================================
# Raw_CFPB scaffold  (one block per GEO slot; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, footprint_slots, raw_slots):
    ws = wb.create_sheet(R.RAW_TAB)
    hide_gridlines(ws)
    ws["A1"] = (f"{R.RAW_TAB} -- CFPB Mortgage Performance Trends values per "
                f"geo slot (or CfpbDemoProvider output in --demo), newest "
                f"month first. Blocks: nat, st01-st{R.STATE_SLOTS:02d}, "
                f"co01-co{footprint_slots:02d}; anchors never move when "
                f"[FOOTPRINT] changes. Values are percent shares; a missing "
                f"month is blank, never 0. FULL-REPLACED every run (every "
                f"vintage revises the whole history -- trap C3). Written by "
                f"runner.py -- do not edit.")
    ws["A1"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 12
    for col in ("B", "C"):
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["D"].width = 44
    blocks = [("national", 0)] + \
        [("state", s) for s in range(1, R.STATE_SLOTS + 1)] + \
        [("county", c) for c in range(1, footprint_slots + 1)]
    for kind, slot in blocks:
        b = R.geo_block(kind, slot, raw_slots)
        ws.cell(b.header_row, 1, R.slot_label(kind, slot)).font = INK_BOLD
        # cells B/C/D (unit key, name, note) are RUNTIME state the runner
        # writes/blanks per run -- empty at build.
        ws.cell(b.label_row, 1, "month").font = SMALL_BOLD
        ws.cell(b.label_row, 2, "d3089_pct").font = SMALL_BOLD
        ws.cell(b.label_row, 3, "d90_pct").font = SMALL_BOLD
    return ws


# ==========================================================================
# Formula builders  (the Excel side of the transform registry -- parity
# with runner.TRANSFORMS is machine-tested, never assumed)
# ==========================================================================
def _mref(kind, slot, measure, raw_slots, offset=0, absolute=False):
    b = R.geo_block(kind, slot, raw_slots)
    col = get_column_letter(R.MEASURE_COL[measure])
    dollar = "$" if absolute else ""
    return (f"{R.RAW_TAB}!{dollar}{col}{dollar}{b.first_data_row + offset}")


def latest_formula(kind, slot, measure, raw_slots):
    """t_level: =IF(v0="","",v0) -- blank-guarded, never a fabricated 0."""
    ref = _mref(kind, slot, measure, raw_slots)
    return f"=IF({ref}=\"\",\"\",{ref})"


def dev12_formula(kind, slot, measure, raw_slots):
    """t_dev_12m EXACT: v0-AVERAGE(v1:v12); AVERAGE ignores blanks just as
    the Python mean skips missing months."""
    b = R.geo_block(kind, slot, raw_slots)
    col = get_column_letter(R.MEASURE_COL[measure])
    v0 = f"{R.RAW_TAB}!{col}{b.first_data_row}"
    rng = (f"{R.RAW_TAB}!{col}{b.first_data_row + 1}:"
           f"{col}{b.first_data_row + 12}")
    return (f"=IF(OR({v0}=\"\",COUNT({rng})=0),\"\","
            f"{v0}-AVERAGE({rng}))")


def streak_formula(kind, slot, measure, raw_slots):
    """t_rise_streak3 EXACT (capped): IF(v0>v1,1,0)+IF(AND(v0>v1,v1>v2),1,0)
    +IF(AND(v0>v1,v1>v2,v2>v3),1,0), blank-guarded on COUNT(v0:v3)=4."""
    b = R.geo_block(kind, slot, raw_slots)
    col = get_column_letter(R.MEASURE_COL[measure])
    v = [f"{R.RAW_TAB}!{col}{b.first_data_row + k}" for k in range(4)]
    rng = f"{R.RAW_TAB}!{col}{b.first_data_row}:{col}{b.first_data_row + 3}"
    return (f"=IF(COUNT({rng})<4,\"\","
            f"IF({v[0]}>{v[1]},1,0)"
            f"+IF(AND({v[0]}>{v[1]},{v[1]}>{v[2]}),1,0)"
            f"+IF(AND({v[0]}>{v[1]},{v[1]}>{v[2]},{v[2]}>{v[3]}),1,0))")


def trend_formula(kind, slot, measure, raw_slots):
    """12-month ASCII text trend strip (newest on the left): '^' rise,
    'v' fall, '-' flat, month-over-month; blank until 13 months land. A
    formula-cell stand-in for a sparkline (no native charts, L4)."""
    b = R.geo_block(kind, slot, raw_slots)
    col = get_column_letter(R.MEASURE_COL[measure])

    def ref(k):
        return f"{R.RAW_TAB}!{col}{b.first_data_row + k}"

    rng = f"{R.RAW_TAB}!{col}{b.first_data_row}:{col}{b.first_data_row + 12}"
    parts = [f"IF({ref(k)}>{ref(k + 1)},\"^\",IF({ref(k)}<{ref(k + 1)},"
             f"\"v\",\"-\"))" for k in range(12)]
    return f"=IF(COUNT({rng})<13,\"\"," + "&".join(parts) + ")"


# ==========================================================================
# Dashboard_State  (National pinned first + 15 provisioned state slots)
# ==========================================================================
def write_dashboard_state(wb, footprint_slots, raw_slots, thr_cells,
                          n_active, n_states):
    ws = wb.create_sheet("Dashboard_State")
    hide_gridlines(ws)
    ncols = len(DS_COLS)

    wl_last = WL_FIRST + footprint_slots - 1
    brand_banner(ws, 1, ncols, "County Mortgage Delinquency Monitor",
                 "CFPB Mortgage Performance Trends (NMDB 5% sample) -- "
                 "monthly 30-89 and 90+ DPD shares, ~6-7 months lagged by "
                 "design (trap C6): the CONFIRMING outcome lane, not a "
                 "nowcast. National + footprint states below; the county "
                 "watchlist is its own tab.")
    tiles = [
        ("COUNTIES IN ALERT",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"ALERT\")",
         f"of {n_active} active footprint counties"),
        ("COUNTIES IN WATCH",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"WATCH\")",
         "elevated"),
        ("SUPPRESSED",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"SUPPRESSED*\")",
         "absent from vintage -- use state rows"),
    ]
    kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8
    header_row(ws, DS_HDR, DS_COLS, right_from=2)
    freeze_below(ws, DS_HDR)

    def geo_row(r, kind, slot, name_formula, fips_formula):
        ws.cell(r, 1, name_formula).font = KB.DATA_FONT
        ws.cell(r, 2, fips_formula).font = MONO_FONT
        cells = [
            (3, latest_formula(kind, slot, "d3089", raw_slots), "0.00"),
            (4, dev12_formula(kind, slot, "d3089", raw_slots), "+0.00;-0.00;0.00"),
            (5, trend_formula(kind, slot, "d3089", raw_slots), None),
            (6, latest_formula(kind, slot, "d90", raw_slots), "0.00"),
            (7, dev12_formula(kind, slot, "d90", raw_slots), "+0.00;-0.00;0.00"),
            (8, trend_formula(kind, slot, "d90", raw_slots), None),
        ]
        for c, f, fmt in cells:
            cell = ws.cell(r, c, f)
            if fmt:
                cell.number_format = fmt
                cell.alignment = KB.RIGHT
            else:
                cell.font = MONO_FONT
                cell.alignment = KB.CENTER
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    # National pinned first as the benchmark (identity is static -- the
    # National row rides in every CFPB file)
    geo_row(DS_FIRST, "national", 0, "NATIONAL (benchmark)", "-----")
    ws.cell(DS_FIRST, 1).font = INK_BOLD
    # 15 provisioned state slots: identity is RUNTIME state the runner
    # writes into the raw block header (name/abbrev in C, fips2 in D)
    for s in range(1, R.STATE_SLOTS + 1):
        b = R.geo_block("state", s, raw_slots)
        name_f = (f"=IF({R.RAW_TAB}!$C${b.header_row}=\"\",\"\","
                  f"{R.RAW_TAB}!$C${b.header_row})")
        fips_f = (f"=IF({R.RAW_TAB}!$D${b.header_row}=\"\",\"\","
                  f"{R.RAW_TAB}!$D${b.header_row})")
        geo_row(ds_row(s), "state", s, name_f, fips_f)

    first, last = DS_FIRST, ds_row(R.STATE_SLOTS)
    # threshold band rules where a threshold exists; heat everywhere
    w_ref, a_ref, direction = thr_cells["d3089_dev12m"]
    band_rules(ws, f"D{first}:D{last}", f"D{first}", w_ref, a_ref, direction)
    w_ref, a_ref, direction = thr_cells["d90_level"]
    band_rules(ws, f"F{first}:F{last}", f"F{first}", w_ref, a_ref, direction)
    for col in ("C", "D", "F", "G"):
        stress_heat(ws, f"{col}{first}:{col}{last}", "above")

    r = last + 2
    ws.cell(r, 1, f"State rows are dashboard CONTEXT ({n_states} states "
                  "derived from the seed footprint at build; 15 slots "
                  "provisioned) -- the fallback lane for SUPPRESSED "
                  "counties. dev 12m = latest minus own trailing-12 "
                  "average, in percentage points. Trend strip: '^' rise / "
                  "'v' fall / '-' flat, newest on the left.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 42
    ws.cell(r, 1).alignment = KB.WRAP

    widths = {"A": 24, "B": 8, "C": 11, "D": 12, "E": 14, "F": 11, "G": 12,
              "H": 14}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    # run-status readout in column L (free of the masthead merges)
    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 46
    for rr in (1, 2, 3, 4):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(5, R.STATUS_COL).font = NOTE_FONT          # macro writes here
    ws.cell(5, R.STATUS_COL).alignment = KB.LEFT
    return ws


# ==========================================================================
# Watchlist  (the ranked county list -- the county-FIPS-gated lane)
# ==========================================================================
BOUNDARY = ("COUNTY RULE. This lane admits ONLY rows carrying a genuine "
            "5-digit county FIPS join key (county:NNNNN from the "
            "[FOOTPRINT] fips cell, Class A) -- the default-deny validator "
            "refuses anything else BY NAME: states, metros and National are "
            "dashboard context, never watchlist. Blank/malformed FIPS shows "
            "REFUSED here; find FIPS with: python runner.py --lookup "
            "\"<county name>\" (offline with --demo). A county ABSENT from "
            "the fetched vintage is SUPPRESSED (below the ~1,000-sample "
            "threshold) -- use its state row; absence is NOT zero.")


def write_watchlist(wb, cfg, footprint_slots, raw_slots, thr_cells, fp_cells):
    ws = wb.create_sheet("Watchlist")
    hide_gridlines(ws)
    ncols = len(WL_COLS)
    brand_banner(ws, 1, ncols,
                 "Watchlist -- Ranked Footprint Counties",
                 "Ranked by (status severity, 90+ deviation vs own 12m "
                 "average) descending. Thresholds: 30-89 dev 12m "
                 "0.3/0.6pp; 90+ rise streak 2/3; 90+ level 1.5/3.0% -- "
                 "heuristic bands, honestly labeled in _config. Helper "
                 "columns K-L feed the rank.")
    watchlist_boundary(ws, 3, ncols, BOUNDARY)
    ws.row_dimensions[3].height = 68
    header_row(ws, WL_HDR, WL_COLS, right_from=3, center_cols=(6, 7, 8))
    ws.cell(WL_HDR, WL_HELPER_DEV, "d90 dev12 (helper)").font = SMALL_BOLD
    ws.cell(WL_HDR, WL_HELPER_SCORE, "rank score").font = SMALL_BOLD

    first, last = WL_FIRST, WL_FIRST + footprint_slots - 1
    nat_first = R.geo_block("national", 0, raw_slots).first_data_row
    nat_ref = f"{R.RAW_TAB}!$B${nat_first}"      # non-blank == a run landed
    thrE = thr_cells["d3089_dev12m"]
    thrG = thr_cells["d90_streak3"]
    thrF = thr_cells["d90_level"]

    for slot in range(1, footprint_slots + 1):
        r = wl_row(slot)
        pc = fp_cells[slot]
        # county identity BY FORMULA from [FOOTPRINT]: a swapped slot
        # re-labels everywhere on recalc (the flexible-footprint requirement)
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['state']}=\"\",\"\",{pc['state']})").font = KB.SECONDARY
        c3 = ws.cell(r, 3, f"=IF({pc['fips']}=\"\",\"\",{pc['fips']})")
        c3.font = MONO_FONT
        c3.number_format = "@"                   # FIPS stays TEXT (trap C1)
        ws.cell(r, 4, latest_formula("county", slot, "d3089", raw_slots)
                ).number_format = "0.00"
        ws.cell(r, 5, dev12_formula("county", slot, "d3089", raw_slots)
                ).number_format = "+0.00;-0.00;0.00"
        ws.cell(r, 6, latest_formula("county", slot, "d90", raw_slots)
                ).number_format = "0.00"
        ws.cell(r, 7, streak_formula("county", slot, "d90", raw_slots)
                ).number_format = "0"
        ws.cell(r, WL_HELPER_DEV,
                dev12_formula("county", slot, "d90", raw_slots)
                ).font = NOTE_FONT
        # Status: the formula-level gate rendering. Order matters: empty
        # slot -> blank; bad FIPS -> REFUSED (with the --lookup hint);
        # inactive -> OFF; pre-first-run -> (not yet run); no landed months
        # -> the explicit SUPPRESSED text (trap C4); else the three
        # threshold gates roll up to ALERT/WATCH/OK exactly like
        # runner.compute_digest.
        def gate(op_ref, cellref):
            w, a, _d = op_ref
            return (f"AND(ISNUMBER({cellref}),{cellref}>={{bound}})"
                    .format(bound=""), w, a)
        alert_cond = (f"OR(AND(ISNUMBER($E{r}),$E{r}>={thrE[1]}),"
                      f"AND(ISNUMBER($G{r}),$G{r}>={thrG[1]}),"
                      f"AND(ISNUMBER($F{r}),$F{r}>={thrF[1]}))")
        watch_cond = (f"OR(AND(ISNUMBER($E{r}),$E{r}>={thrE[0]}),"
                      f"AND(ISNUMBER($G{r}),$G{r}>={thrG[0]}),"
                      f"AND(ISNUMBER($F{r}),$F{r}>={thrF[0]}))")
        ws.cell(r, 8,
                f"=IF({pc['name']}=\"\",\"\","
                f"IF(NOT(AND(ISTEXT({pc['fips']}),LEN({pc['fips']})=5)),"
                f"\"REFUSED: blank/invalid FIPS -- run --lookup\","
                f"IF(UPPER({pc['active']}&\"\")<>\"TRUE\",\"OFF\","
                f"IF({nat_ref}=\"\",\"(not yet run)\","
                f"IF(AND($D{r}=\"\",$F{r}=\"\"),\"{R.SUPPRESSED_TEXT}\","
                f"IF({alert_cond},\"ALERT\","
                f"IF({watch_cond},\"WATCH\",\"OK\")))))))")
        # rank score = severity dominant, 90+ dev tiebreak (spec sec 4);
        # only live statuses rank -- SUPPRESSED/OFF/REFUSED stay out
        sc = get_column_letter(WL_HELPER_SCORE)
        ws.cell(r, WL_HELPER_SCORE,
                f"=IF(OR($H{r}=\"ALERT\",$H{r}=\"WATCH\",$H{r}=\"OK\"),"
                f"IF($H{r}=\"ALERT\",2,IF($H{r}=\"WATCH\",1,0))*1000"
                f"+IF(ISNUMBER(${get_column_letter(WL_HELPER_DEV)}{r}),"
                f"${get_column_letter(WL_HELPER_DEV)}{r},0),\"\")"
                ).font = NOTE_FONT
        ws.cell(r, 9, f"=IF(${sc}{r}=\"\",\"\","
                      f"RANK(${sc}{r},${sc}${first}:${sc}${last}))")
        for c in (4, 5, 6):
            ws.cell(r, c).alignment = KB.RIGHT
        for c in (7, 9):
            ws.cell(r, c).alignment = KB.CENTER
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    KB.alert_rule(ws, f"H{first}:H{last}")
    ws.conditional_formatting.add(
        f"H{first}:H{last}", FormulaRule(
            formula=[f'ISNUMBER(SEARCH("SUPPRESSED",H{first}))'],
            fill=KB.TIGHTEN_FILL,
            font=Font(name="Calibri", bold=True, color=KB.SLATE)))
    w_ref, a_ref, direction = thr_cells["d3089_dev12m"]
    band_rules(ws, f"E{first}:E{last}", f"E{first}", w_ref, a_ref, direction)
    w_ref, a_ref, direction = thr_cells["d90_level"]
    band_rules(ws, f"F{first}:F{last}", f"F{first}", w_ref, a_ref, direction)
    w_ref, a_ref, direction = thr_cells["d90_streak3"]
    band_rules(ws, f"G{first}:G{last}", f"G{first}", w_ref, a_ref, direction)
    stress_heat(ws, f"E{first}:E{last}", "above")

    r = last + 2
    ws.cell(r, 1, "Columns K-L are the rank helpers (90+ dev vs 12m avg; "
                  "severity*1000 + dev). SUPPRESSED counties carry no rank "
                  "and never count toward the alert KPIs -- absence from "
                  "the CFPB file is a sample-size fact, not a zero. Sort/"
                  "filter on Rank if desired.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    r += 1
    # build-time refusals (none under the seeded config; rendered
    # interpolated if a bad seed row is ever injected and the build-time
    # hard gate bypassed)
    _, refusals, _ = R.evaluate_footprint(cfg)
    for fp, message in refusals:
        ws.cell(r, 1, f"slot {fp.slot}").font = MONO_FONT
        ws.cell(r, 2, "REFUSED").font = Font(name="Calibri", bold=True,
                                             color=KB.CRIMSON)
        r += 1
        ws.cell(r, 1, message).font = NOTE_FONT
        ws.cell(r, 1).alignment = KB.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=ncols)
        ws.row_dimensions[r].height = 56
        r += 1

    widths = {"A": 24, "B": 7, "C": 8, "D": 12, "E": 13, "F": 11, "G": 14,
              "H": 44, "I": 6, "K": 10, "L": 10}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    freeze_below(ws, WL_HDR)
    return ws


# ==========================================================================
# Dashboard_Trends  (National + worst-5 footprint counties by 90+ level,
# 24-month formula readout -- no charts, L4)
# ==========================================================================
def write_dashboard_trends(wb, footprint_slots, raw_slots):
    ws = wb.create_sheet("Dashboard_Trends")
    hide_gridlines(ws)
    brand_banner(ws, 1, 8, "Trends -- 90+ DPD, National vs Worst Footprint "
                 "Counties",
                 "Worst-5 footprint counties by latest 90+ share "
                 "(formula-ranked off the Watchlist; ties keep the first "
                 "match). 24 months, newest on the left. Values are "
                 "percent of outstanding mortgages 90+ days late.")
    ws.cell(4, 1, "Every vintage revises the FULL history (trap C3) -- "
                  "these are the CURRENT vintage's readings, not a stitched "
                  "archive. ~6-7 month lag by design.").font = NOTE_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

    wl_last = WL_FIRST + footprint_slots - 1
    f_rng = f"Watchlist!$F${WL_FIRST}:$F${wl_last}"
    a_rng = f"Watchlist!$A${WL_FIRST}:$A${wl_last}"
    c_rng = f"Watchlist!$C${WL_FIRST}:$C${wl_last}"
    nat_b = R.geo_block("national", 0, raw_slots)
    co1 = R.geo_block("county", 1, raw_slots)
    stride = R.RAW_HEADER_ROWS + raw_slots + R.RAW_GAP_ROWS
    last_raw_row = R.geo_block("county", footprint_slots,
                               raw_slots).first_data_row + raw_slots
    d90_col_rng = f"{R.RAW_TAB}!$C$1:$C${last_raw_row}"

    # header row: month labels by formula off the national block (all geos
    # share the vintage's month grid)
    hdr = ws.cell(TR_HDR, 1, "Series")
    hdr.fill = KB.HDR_FILL
    hdr.font = KB.WHITE_BOLD
    hdr2 = ws.cell(TR_HDR, 2, "FIPS")
    hdr2.fill = KB.HDR_FILL
    hdr2.font = KB.WHITE_BOLD
    for m in range(TR_MONTHS):
        c = ws.cell(TR_HDR, TR_FIRST_MCOL + m,
                    f"=IF({R.RAW_TAB}!$A${nat_b.first_data_row + m}=\"\","
                    f"\"\",{R.RAW_TAB}!$A${nat_b.first_data_row + m})")
        c.fill = KB.HDR_FILL
        c.font = KB.WHITE_BOLD
        c.alignment = KB.CENTER
        ws.column_dimensions[get_column_letter(TR_FIRST_MCOL + m)].width = 9
    ws.row_dimensions[TR_HDR].height = 22

    # National row (pinned)
    ws.cell(TR_FIRST, 1, "NATIONAL (benchmark)").font = INK_BOLD
    ws.cell(TR_FIRST, 2, "-----").font = MONO_FONT
    for m in range(TR_MONTHS):
        ref = f"{R.RAW_TAB}!$C${nat_b.first_data_row + m}"
        c = ws.cell(TR_FIRST, TR_FIRST_MCOL + m,
                    f"=IF({ref}=\"\",\"\",{ref})")
        c.number_format = "0.00"
        c.alignment = KB.RIGHT

    # worst-5 county rows: slot helper in AB, identity + values by formula
    sc = get_column_letter(TR_SLOT_COL)
    for k in range(1, 6):
        r = TR_FIRST + k
        ws.cell(r, TR_SLOT_COL,
                f"=IFERROR(MATCH(LARGE({f_rng},{k}),{f_rng},0),\"\")"
                ).font = NOTE_FONT
        ws.cell(r, 1, f"=IF(${sc}{r}=\"\",\"\","
                      f"INDEX({a_rng},${sc}{r}))").font = KB.DATA_FONT
        ws.cell(r, 2, f"=IF(${sc}{r}=\"\",\"\","
                      f"INDEX({c_rng},${sc}{r}))").font = MONO_FONT
        ws.cell(r, 2).number_format = "@"
        for m in range(TR_MONTHS):
            idx = (f"INDEX({d90_col_rng},{co1.first_data_row}"
                   f"+(${sc}{r}-1)*{stride}+{m})")
            c = ws.cell(r, TR_FIRST_MCOL + m,
                        f"=IF(${sc}{r}=\"\",\"\","
                        f"IF({idx}=\"\",\"\",{idx}))")
            c.number_format = "0.00"
            c.alignment = KB.RIGHT
        for c in range(1, TR_FIRST_MCOL + TR_MONTHS):
            ws.cell(r, c).border = KB.HAIRLINE

    grid = (f"{get_column_letter(TR_FIRST_MCOL)}{TR_FIRST}:"
            f"{get_column_letter(TR_FIRST_MCOL + TR_MONTHS - 1)}"
            f"{TR_FIRST + 5}")
    stress_heat(ws, grid, "above")

    r = TR_FIRST + 7
    ws.cell(r, 1, "Column AB holds the MATCHed Watchlist row (== footprint "
                  "slot) for each ranked county; the 24 month cells INDEX "
                  "straight into that slot's fixed Raw_CFPB block. "
                  "SUPPRESSED and empty slots never rank (no 90+ value)."
             ).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8
    # run-status readout in column L rows 1-4 sits above the grid (the
    # banner merge stops at column H)
    for rr in (1, 2, 3, 4):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    freeze_below(ws, TR_HDR, first_col=3)
    return ws


# ==========================================================================
# Code-in-tab + readme
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
COUNTY MORTGAGE DELINQUENCY MONITOR (CFPB Mortgage Performance Trends)
======================================================================

WHAT THIS IS
  The leanest template in the series: county-level mortgage credit
  OUTCOMES for a collateral footprint. One workbook pulls the CFPB's
  Mortgage Performance Trends files (public domain, KEYLESS -- no API key,
  no account: five small CSV downloads), lands National + footprint-state
  + footprint-county 30-89 and 90+ DPD shares, and answers: "in the
  counties where my collateral actually sits, is mortgage delinquency
  rising?" It is the CONFIRMING counterpart to the macro template's
  leading indicators -- realized outcomes, ~6-7 months lagged by design.

THE FLEXIBLE FOOTPRINT (the point of this template)
  [FOOTPRINT] in _config: slot | fips | name | state | active.
  - ADD a county: type its 5-digit FIPS/name/state/active=TRUE on a free
    slot row, re-run the runner. No rebuild. Find FIPS from PowerShell:
        python runner.py --lookup "Cook"            (live: greps the file)
        python runner.py --lookup "Cook" --demo     (offline mini-index)
  - REMOVE a county: set active=FALSE (or clear the row); the slot blanks
    on the next run (stateless clearing -- never last vintage's values).
  - CAPACITY is built in (see footprint_slots in [SETTINGS]). More
    counties than slots = the runner REFUSES with the exact rebuild
    command (make_workbook.py --footprint-slots N). It never truncates.
  - KEEP FIPS CELLS TEXT-FORMATTED: the CFPB quote-wraps FIPS in its own
    files precisely so Excel cannot strip leading zeros (trap C1).
  County identity flows into every tab BY FORMULA from the [FOOTPRINT]
  cells, and raw anchors depend only on the SLOT -- a swap re-labels and
  re-points everything on recalc.

THE DATA (honest provenance)
  Source: CFPB Mortgage Performance Trends, built on the National Mortgage
  Database (NMDB, joint CFPB/FHFA) -- a nationally representative 5%
  (1-in-20) sample of outstanding closed-end first-lien 1-4-family
  mortgages, credit-repository sourced. These are credit-record buckets,
  not the MBA survey rates:
    30-89 DPD share = one-two missed payments / all outstanding loans
    90+  DPD share = three or more missed payments / all outstanding
  Monthly series from 2008-01, published ~SEMIANNUALLY with a ~6-7 MONTH
  LAG (trap C6) -- a confirming indicator, never a nowcast. Values are
  percent, one decimal in the source. U.S. government work: public domain
  (17 U.S.C. s105); credit "CFPB Mortgage Performance Trends (NMDB)" as
  good practice.

SUPPRESSION (trap C4 -- absence is NOT zero)
  Only counties averaging >= ~1,000 sample mortgages in the threshold year
  clear the CFPB's inclusion bar -- roughly 470 of ~3,100 counties, i.e. a
  big-county filter (in a 5% sample that's ~20,000 real mortgages).
  Suppressed counties are simply ABSENT from the file. This template
  renders them "SUPPRESSED (below sample threshold) -- use the state row",
  excludes them from every alert KPI, and lists them in the email digest.
  Never interpolated, never zero. Rural footprints: monitor the state row.

VINTAGES AND REVISIONS (trap C3 -- why every run is a FULL REPLACE)
  Every release revises the ENTIRE history ("rates may change between
  updates due to updated credit data"). The runner therefore blanks and
  rewrites every raw block on every run, and records the vintage
  (thru-YYYY-MM, from the dated filename) in the status panel. NEVER diff
  a stored history against a new vintage and read it as market movement.

FILENAME DISCOVERY (trap C2)
  The files are DATED (CountyMortgagesPercent-90-plusDaysLate-thru-
  YYYY-MM.csv), so hardcoded URLs die every release. The live provider
  scrapes the download page for the current names; if the page changes,
  set the `vintage` key in [SETTINGS] (e.g. thru-2025-09) to construct
  the URLs directly.

CONTINUITY TRIPWIRE (the 2025 lesson)
  The CFPB publishes this dataset informally -- no promised cadence -- and
  the program survived the 2025 turmoil once (the dataset was community-
  archived to DataLumos as a recovery mirror). If the newest vintage falls
  more than continuity_months (default 9) behind --asof, the run surfaces
  a CONTINUITY WARNING in the status panel and the email. If it trips:
  check the download page, then the DataLumos mirror; the FHFA NMDB
  aggregate statistics are a partial (state/CBSA, quarterly) fallback --
  there is NO like-for-like county-monthly substitute.

PROVIDERS (the adapter seam)
  Class A (public, KEYLESS):
    - CfpbProvider     -- LIVE plain-urllib, TWO-STAGE: scrape the download
                          page for dated filenames, then fetch 4 CSVs
                          (State + County x both measures; the Metro file
                          is SKIPPED -- mixed non-FIPS keys). Throttled
                          0.6s, backoff on 429/5xx, per-URL cache.
    - CfpbDemoProvider -- deterministic OFFLINE stand-in; used by every
                          test and the --demo flag. Demo stress counties
                          are FICTION assigned by hand; the demo also
                          omits one seeded county (King County WA) so the
                          SUPPRESSED path is always visible.
  Class C (licensed): NOT admitted anywhere in this template.

ONE-TIME SETUP
  1. Install Python 3 (add it to PATH) and the dependencies:
         pip install pandas openpyxl
  2. Enable macros when prompted (the Extract button needs them).
  That is all -- the live files need NO key.

RUN IT  (the workbook is the finished product; you only add data)
  1. In Excel press Alt+F8 -> run ExtractFiles. It writes runner.py (from
     _code_py), requirements.txt, and RUN.txt next to the workbook.
     NOTHING runs inside Excel.
  2. SAVE and CLOSE the workbook.
  3. From PowerShell (see RUN.txt):
       - Try-the-button (offline): python runner.py -w <book>.xlsm --demo
       - Live CFPB (keyless):      python runner.py -w <book>.xlsm
       - Find a FIPS:              python runner.py --lookup "<name>"
  4. Reopen the workbook -- the formulas recalc into the dashboards.

THE MEASURES AND BANDS (heuristic, honestly labeled)
  30-89 dev 12m: latest 30-89 share minus its own trailing-12 average
    (percentage points). WATCH +0.3 / ALERT +0.6 -- the early-stage
    pipeline turning against its own trend.
  90+ rise streak: consecutive monthly rises in the 90+ share, capped at
    3. WATCH 2 / ALERT 3 -- persistence beats level for a slow series.
  90+ level: WATCH 1.5 / ALERT 3.0 -- long-run non-crisis is under ~1%;
    GFC county peaks ran far higher.
  All bands are heuristic screens for portfolio review, not regulatory
  standards; they live in _config [THRESHOLDS] as numeric cells.

DATA-QUALITY TRAPS (baked in -- C1-C7)
  C1 quote-wrapped FIPS (stripped; kept TEXT end-to-end)
  C2 dated filenames (discovered each run; `vintage` override knob)
  C3 full-history revisions (stateless full replace; vintage labeled)
  C4 suppression by omission (absent != zero; explicit SUPPRESSED)
  C5 wide format (unpivoted by the YYYY-MM header pattern, never position)
  C6 ~6-7 month lag (labeled in the masthead and status panel)
  C7 National row inside every file (RegionType filter; own nat_* series)

NO AI IS INVOLVED IN THE DATA PATH. The gates, transforms, thresholds and
continuity guard are deterministic and config-driven: the same input
always yields the same workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() \
                and len(line) > 3 and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path, footprint_slots=R.FOOTPRINT_SLOTS_DEFAULT):
    rows = config_rows(footprint_slots)
    cfg = R.parse_config(rows)
    # BUILD-TIME hard gates (BUILD SPEC 0.1/sec 3): bad series rows, bad
    # seed counties, or an over-capacity seed refuse the BUILD itself.
    R.validate_series(cfg.series)
    R.validate_footprint_capacity(cfg)
    R.assert_entity_gates(cfg)
    smap = R.state_slot_map(cfg)
    admitted, refusals, excluded = R.evaluate_footprint(cfg)
    raw_slots = cfg.raw_slots

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, footprint_slots, raw_slots)
    # _config must be written before dashboards so threshold and
    # [FOOTPRINT] cell refs resolve.
    thr_cells, _set_cells, fp_cells = write_config(wb, footprint_slots)

    write_dashboard_state(wb, footprint_slots, raw_slots, thr_cells,
                          len(admitted), len(smap))
    write_dashboard_trends(wb, footprint_slots, raw_slots)
    write_watchlist(wb, cfg, footprint_slots, raw_slots, thr_cells, fp_cells)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    order = ["Dashboard_State", "Dashboard_Trends", "Watchlist", R.RAW_TAB,
             "_config", "_code_py", "_code_vba", "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order
                    else 99)
    wb.active = wb.sheetnames.index("Dashboard_State")
    wb.save(out_path)
    return out_path, len(cfg.series), len(admitted), footprint_slots


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "Mortgage_Delinquency_Monitor_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n_series, n_admitted, cap = build(out)
    print(f"built {path}: {n_series} series rows x {cap} county slots, "
          f"{n_admitted} seed counties admitted")
