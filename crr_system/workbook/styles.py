"""Shared styling for the CRR workbook.

Design language: report, not spreadsheet. No gridlines, no boxed cell grids -
tables use hairline row separators and banded fills; titles are dark text on
white with a navy accent rule; KPIs render as soft cards. Financial-model
text conventions still apply: blue = inputs, black = formulas, green =
cross-sheet links, yellow fill = key assumptions. Arial throughout.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

INK = "1F2937"          # near-black slate for headings
NAVY = "1F3864"
MID_BLUE = "2E5597"
LIGHT_BLUE = "D9E2F3"
PALE = "F2F6FC"
BAND = "F4F6FA"         # alternate-row band (on white panels)
HAIR = "DDE2EA"         # hairline separators
CANVAS = "E7EAEF"       # page background - light gray desk the panels sit on
AMBER = "FFF2CC"
GOLD = "D9A441"
RED = "C0504D"
ALERT_RED = "C00000"
PALE_RED = "FCE4E4"
PALE_GREEN = "E2EFDA"
GREY = "8A93A2"
LIGHT_GREY = "F2F2F2"

CHART_COLORS = ["2E5597", "D9A441", "8496B0", "1F3864", "C0504D", "6FA287"]

WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
CANVAS_FILL = PatternFill("solid", start_color=CANVAS)

# Information-cell treatments: every cell that carries data gets one of these
# so content reads as designed fields/values, not bare text on white.
FIELD = "EAF0F8"        # editable input fields (pairs with blue input text)
ICE = "F4F7FB"          # computed/derived values
GOLDTINT = "FAF3E3"     # values sourced from the crosswalk / reference layer
LABELTINT = "EDF1F7"    # table label columns / row headers
FIELD_FILL = PatternFill("solid", start_color=FIELD)
ICE_FILL = PatternFill("solid", start_color=ICE)
GOLDTINT_FILL = PatternFill("solid", start_color=GOLDTINT)
LABEL_FILL = PatternFill("solid", start_color=LABELTINT)

_FIELD_SIDE = Side(style="thin", color="C9D2E0")
FIELD_BOX = Border(left=_FIELD_SIDE, right=_FIELD_SIDE,
                   top=_FIELD_SIDE, bottom=_FIELD_SIDE)

NAVY_FONT = Font(name="Arial", size=10, bold=True, color="1F3864")
SEV_FONTS = {
    "High": Font(name="Arial", size=8.5, bold=True, color="C0504D"),
    "Medium": Font(name="Arial", size=8.5, bold=True, color="B07D2B"),
    "Low": Font(name="Arial", size=8.5, bold=True, color="8A93A2"),
}

INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
FORMULA_FONT = Font(name=FONT, size=10, color="000000")
LINK_FONT = Font(name=FONT, size=10, color="008000")
BODY_FONT = Font(name=FONT, size=10, color=INK)
BOLD_FONT = Font(name=FONT, size=10, bold=True, color=INK)
SMALL_FONT = Font(name=FONT, size=8, color=GREY, italic=True)
TITLE_FONT = Font(name=FONT, size=17, bold=True, color=INK)
SUBTITLE_FONT = Font(name=FONT, size=9.5, italic=True, color=GREY)
H2_FONT = Font(name=FONT, size=10.5, bold=True, color=NAVY)
H3_FONT = Font(name=FONT, size=10.5, bold=True, color=NAVY)
TH_FONT = Font(name=FONT, size=9, bold=True, color=NAVY)

TITLE_FILL = PatternFill("solid", start_color=NAVY)
SECTION_FILL = PatternFill("solid", start_color="F5EEDF")  # light gold tint
SUBHEAD_FILL = PatternFill("solid", start_color=PALE)
ASSUMPTION_FILL = PatternFill("solid", start_color="FFFF00")
BAND_FILL = PatternFill("solid", start_color=BAND)

_HAIR_SIDE = Side(style="thin", color=HAIR)
_RULE_SIDE = Side(style="medium", color=NAVY)
THIN = _HAIR_SIDE
BOX = Border(bottom=_HAIR_SIDE)            # hairline under each table row
RULE = Border(bottom=_RULE_SIDE)           # heavier rule under table headers
BOTTOM = Border(bottom=_HAIR_SIDE)

WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

FMT_USD = "$#,##0;($#,##0);\"-\""
FMT_USD_K = "$#,##0,;($#,##0,);\"-\""
FMT_X = "0.0x;(0.0x);\"-\""
FMT_X2 = "0.00x;(0.00x);\"-\""
FMT_PCT = "0.0%;(0.0%);\"-\""
FMT_NUM = "#,##0;(#,##0);\"-\""
FMT_DATE = "mm/dd/yyyy"


def title_bar(ws, text, last_col, row=1, subtitle=None):
    """Editorial title: large dark heading, grey subtitle, navy accent rule."""
    c = ws.cell(row=row, column=2 if last_col > 2 else 1, value=text)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[row].height = 30
    sub_row = row + 1
    if subtitle:
        s = ws.cell(row=sub_row, column=2 if last_col > 2 else 1, value=subtitle)
        s.font = SUBTITLE_FONT
        s.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        ws.row_dimensions[sub_row].height = 14
    rule_row = sub_row + 1 if subtitle else sub_row
    for col in range(2 if last_col > 2 else 1, last_col + 1):
        ws.cell(row=rule_row, column=col).fill = TITLE_FILL
    ws.cell(row=rule_row, column=last_col).fill = PatternFill("solid", start_color=GOLD)
    ws.row_dimensions[rule_row].height = 3


def section_bar(ws, row, text, first_col=1, last_col=8):
    """Slim section header: navy accent block + pale band + navy label."""
    a = ws.cell(row=row, column=first_col)
    a.fill = TITLE_FILL
    for col in range(first_col + 1, last_col + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    c = ws.cell(row=row, column=first_col + 1, value=text)
    c.font = H2_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 17


def col_headers(ws, row, headers, start_col=1, fill=None, height=22):
    """Report-style table header: bold navy text over a medium navy rule."""
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = TH_FONT
        if fill is not None:
            c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        c.border = RULE
    ws.row_dimensions[row].height = height


def set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_cell(c, *, font=None, fmt=None, fill=None, border=BOX, align=None):
    c.font = font or BODY_FONT
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if align:
        c.alignment = align
    return c


def band_rows(ws, first_row, last_row, first_col, last_col):
    """Alternate-row banding for data tables (replaces cell-grid borders)."""
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for col in range(first_col, last_col + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill is None or cell.fill.patternType is None:
                    cell.fill = BAND_FILL


def kpi_card(ws, row, col, label, formula, fmt=FMT_NUM, color=NAVY):
    """KPI card: white block on the gray canvas, gold rule under the figure."""
    lab = ws.cell(row=row, column=col, value=label.upper())
    lab.font = Font(name=FONT, size=7.5, bold=True, color=GREY)
    lab.alignment = Alignment(horizontal="left", vertical="bottom", indent=1, wrap_text=True)
    lab.fill = WHITE_FILL
    val = ws.cell(row=row + 1, column=col, value=formula)
    val.font = Font(name=FONT, size=18, bold=True, color=color)
    val.number_format = fmt
    val.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    val.fill = WHITE_FILL
    val.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row + 1].height = 26


def whiten(ws, r1, c1, r2, c2):
    """Paint a white panel behind a content block (only where unfilled)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill.patternType is None:
                cell.fill = WHITE_FILL


def canvas_pass(ws, last_col, last_row):
    """Fill every remaining unfilled cell with the gray canvas, so white
    panels and charts read as layered content on a page background."""
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill.patternType is None:
                cell.fill = CANVAS_FILL


def polish_chart(chart, *, legend_pos="b", colors=None):
    """Strip the default-Excel chart look: no frame, soft gridlines, palette.

    colors: per-series override. Default palette leads navy/gold; pass red
    (RED/ALERT_RED) explicitly for exception/alert series only.
    """
    from openpyxl.chart import LineChart, PieChart
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties as GraphicProperties
    from openpyxl.drawing.line import LineProperties

    palette = colors or CHART_COLORS
    chart.graphical_properties = GraphicProperties(
        solidFill="FFFFFF", ln=LineProperties(noFill=True))
    if chart.legend is not None:
        chart.legend.position = legend_pos
        chart.legend.overlay = False
    for axis in ("y_axis", "x_axis"):
        ax = getattr(chart, axis, None)
        if ax is None:
            continue
        ax.spPr = GraphicProperties(ln=LineProperties(solidFill=HAIR))
        if axis == "y_axis":
            ax.majorGridlines = ChartLines(
                spPr=GraphicProperties(ln=LineProperties(solidFill=HAIR)))
        else:
            ax.majorGridlines = None

    if isinstance(chart, PieChart):
        if colors:  # color slices individually (e.g. severity: red/gold/slate)
            series = chart.series[0]
            series.data_points = [
                DataPoint(idx=i, spPr=GraphicProperties(solidFill=c))
                for i, c in enumerate(colors)
            ]
        return chart
    for i, series in enumerate(getattr(chart, "series", [])):
        color = palette[i % len(palette)]
        if isinstance(chart, LineChart):
            series.graphicalProperties = GraphicProperties(
                ln=LineProperties(solidFill=color, w=22000))
            series.smooth = False
        else:
            series.graphicalProperties = GraphicProperties(
                solidFill=color, ln=LineProperties(noFill=True))
    return chart
