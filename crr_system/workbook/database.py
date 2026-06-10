"""Database (credit-record table) and Responses (normalized answers) sheets.

Database: one row per credit per review. Raw inputs are values; everything
derivable (quarter, category, exception counts, completion, rates) is a
formula over Responses, so dashboards recalculate when answers change.
Responses: one row per question per credit - the single source the
exception dashboard (No only), observations view (Obs only), and rate
calcs (N/A excluded) all read from.
"""

from __future__ import annotations

from openpyxl.worksheet.datavalidation import DataValidation

from . import styles as st

DB_HEADERS = [
    "Credit ID", "Seg", "Segment", "Borrower", "Commitment ($000)",
    "Review Date", "Quarter", "Reviewer", "LOB Grade", "CRR Grade",
    "Regulatory Category", "Grade Concurrence", "Leverage (x)", "DSCR (x)",
    "LTV (%)", "Exceptions (No)", "Observations (Obs)", "Applicable (excl. N/A)",
    "Answered", "Total Qs", "Completion %", "Exception Rate", "Status", "Notes",
]

RESP_HEADERS = [
    "Credit ID", "Seg", "QID", "Section", "Question", "Severity", "Answer",
    "Note", "Review Date", "Reviewer", "Days Open (No)", "Aging Bucket",
    "ObsIdx", "NoIdx",
]


def build_database(wb, credits):
    ws = wb.create_sheet("Database")
    st.set_widths(ws, [11, 6, 22, 34, 13, 11, 9, 11, 7, 7, 14, 13, 10, 9, 9,
                       11, 13, 13, 10, 9, 11, 11, 12, 30])
    st.col_headers(ws, 1, DB_HEADERS, fill=st.SECTION_FILL)
    for c in ws[1]:
        c.font = st.H2_FONT
    ws.freeze_panes = "A2"

    for i, cr in enumerate(credits):
        write_credit_row(ws, i + 2, cr)

    n = len(credits) + 1
    ws.auto_filter.ref = f"A1:X{n}"
    dv = DataValidation(type="list", formula1="=StatusList", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"W2:W{n}")
    return ws


def write_credit_row(ws, r, cr):
    vals = {
        "A": cr["credit_id"], "B": cr["segment"], "C": cr["segment_name"],
        "D": cr["borrower"], "E": cr["commitment"], "F": cr["review_date"],
        "H": cr["reviewer"], "I": cr["lob_grade"], "J": cr["crr_grade"],
        "M": cr["leverage"], "N": cr["dscr"], "O": cr["ltv"],
        "W": cr["status"],
    }
    for col, v in vals.items():
        cell = ws[f"{col}{r}"]
        cell.value = v
        cell.font = st.INPUT_FONT if col not in "ABCG" else st.BODY_FONT
        cell.border = st.BOX
    fml = {
        "G": f'=IF(F{r}="","",YEAR(F{r})&"-Q"&ROUNDUP(MONTH(F{r})/3,0))',
        "K": f'=IF(J{r}="","",VLOOKUP(J{r},Lists!$D$2:$E$9,2,0))',
        "L": f'=IF(OR(I{r}="",J{r}=""),"",IF(I{r}=J{r},"Concur","Grade Change"))',
        "P": f'=COUNTIFS(Responses!$A:$A,$A{r},Responses!$G:$G,"No")',
        "Q": f'=COUNTIFS(Responses!$A:$A,$A{r},Responses!$G:$G,"Obs")',
        "R": f'=COUNTIFS(Responses!$A:$A,$A{r},Responses!$G:$G,"Yes")+P{r}+Q{r}',
        "S": f'=R{r}+COUNTIFS(Responses!$A:$A,$A{r},Responses!$G:$G,"N/A")',
        "T": f'=COUNTIFS(Responses!$A:$A,$A{r})',
        "U": f'=IF(T{r}=0,"",S{r}/T{r})',
        "V": f'=IF(R{r}=0,"",P{r}/R{r})',
    }
    for col, f in fml.items():
        cell = ws[f"{col}{r}"]
        cell.value = f
        cell.font = st.FORMULA_FONT
        cell.border = st.BOX
    ws[f"E{r}"].number_format = st.FMT_USD
    ws[f"F{r}"].number_format = st.FMT_DATE
    ws[f"M{r}"].number_format = st.FMT_X
    ws[f"N{r}"].number_format = st.FMT_X2
    ws[f"O{r}"].number_format = st.FMT_PCT
    for col in "PQRST":
        ws[f"{col}{r}"].number_format = st.FMT_NUM
    ws[f"U{r}"].number_format = st.FMT_PCT
    ws[f"V{r}"].number_format = st.FMT_PCT


def build_responses(wb, responses):
    ws = wb.create_sheet("Responses")
    st.set_widths(ws, [11, 6, 9, 30, 60, 9, 8, 45, 11, 10, 10, 10, 7, 7])
    st.col_headers(ws, 1, RESP_HEADERS, fill=st.SECTION_FILL)
    for c in ws[1]:
        c.font = st.H2_FONT
    ws.freeze_panes = "A2"

    for i, rp in enumerate(responses):
        write_response_row(ws, i + 2, rp)

    ws.auto_filter.ref = f"A1:N{len(responses) + 1}"
    return ws


def write_response_row(ws, r, rp):
    vals = [rp["credit_id"], rp["segment"], rp["qid"], rp["section"],
            rp["question"], rp["severity"], rp["answer"], rp["note"] or None,
            rp["review_date"], rp["reviewer"]]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = st.INPUT_FONT if col in (7, 8) else st.BODY_FONT
        if col in (4, 5, 8):
            c.alignment = st.WRAP_TOP
    ws[f"I{r}"].number_format = st.FMT_DATE
    ws[f"K{r}"] = f'=IF(G{r}="No",MAX(0,AsOfDate-I{r}),"")'
    ws[f"K{r}"].font = st.FORMULA_FONT
    ws[f"K{r}"].number_format = st.FMT_NUM
    ws[f"L{r}"] = (f'=IF(K{r}="","",IF(K{r}<=90,"0-90",IF(K{r}<=180,"91-180",'
                   f'IF(K{r}<=270,"181-270",">270"))))')
    ws[f"L{r}"].font = st.FORMULA_FONT
    ws[f"M{r}"] = f'=IF(G{r}="Obs",COUNTIF(G$2:G{r},"Obs"),"")'
    ws[f"M{r}"].font = st.FORMULA_FONT
    ws[f"N{r}"] = f'=IF(G{r}="No",COUNTIF(G$2:G{r},"No"),"")'
    ws[f"N{r}"].font = st.FORMULA_FONT
