"""Premade dashboards - Excel-native formulas and charts over the Database
and Responses sheets. Exceptions count only "No"; the Observations view lists
only "Obs"; every rate excludes N/A from numerator and denominator.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

from . import styles as st
from .content import SEGMENTS, CORE_QUESTIONS

SEG_NAMES = [s["name"] for s in SEGMENTS.values()]
QUARTERS = ["2025-Q3", "2025-Q4", "2026-Q1", "2026-Q2"]


def _kpi(ws, row, col, label, formula, fmt=st.FMT_NUM):
    ws.cell(row=row, column=col, value=label).font = Font(name=st.FONT, size=9, color=st.GREY)
    c = ws.cell(row=row + 1, column=col, value=formula)
    c.font = Font(name=st.FONT, size=16, bold=True, color=st.NAVY)
    c.number_format = fmt


def _table(ws, top, left, title, headers, rows, widths=None):
    """rows: list of lists (values/formulas). Returns (first_data_row, last)."""
    ws.cell(row=top, column=left, value=title).font = st.H3_FONT
    st.col_headers(ws, top + 1, headers, start_col=left)
    for i, row_vals in enumerate(rows):
        r = top + 2 + i
        for j, v in enumerate(row_vals):
            c = ws.cell(row=r, column=left + j, value=v)
            c.font = st.FORMULA_FONT if isinstance(v, str) and v.startswith("=") else st.BODY_FONT
            c.border = st.BOX
    return top + 2, top + 1 + len(rows)


def build_dash_portfolio(wb, n_credits):
    ws = wb.create_sheet("Dash_Portfolio")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [2, 24, 13, 13, 13, 13, 3, 12] + [9] * 9)
    st.title_bar(ws, "Portfolio Overview Dashboard", 17,
                 subtitle="Recalculates from the Database; grade scale 1-8 "
                          "(5 = Special Mention, 6+ = Classified).")
    dbA = "Database!$A:$A"
    _kpi(ws, 4, 2, "Credits Reviewed", f'=COUNTA(Database!$A$2:$A${n_credits + 1})')
    _kpi(ws, 4, 3, "Complete", '=COUNTIF(Database!$W:$W,"Complete")')
    _kpi(ws, 4, 4, "Grade Changes", '=COUNTIF(Database!$L:$L,"Grade Change")')
    _kpi(ws, 4, 5, "Concurrence Rate",
         f'=IF(COUNTA(Database!$A$2:$A${n_credits + 1})=0,"",1-COUNTIF(Database!$L:$L,"Grade Change")/COUNTA(Database!$A$2:$A${n_credits + 1}))',
         st.FMT_PCT)
    _kpi(ws, 4, 6, "Criticized (CRR ≥ 5)", '=COUNTIF(Database!$J:$J,">=5")')

    rows = [[name,
             f'=COUNTIF(Database!$C:$C,B{9 + i})',
             f'=SUMIFS(Database!$P:$P,Database!$C:$C,B{9 + i})',
             f'=COUNTIFS(Database!$C:$C,B{9 + i},Database!$L:$L,"Grade Change")']
            for i, name in enumerate(SEG_NAMES)]
    first, last = _table(ws, 7, 2, "Reviews by Segment",
                         ["Segment", "Reviews", "Exceptions", "Downgrades"], rows)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Reviews & Exceptions by Segment"
    chart.style = 10
    chart.height, chart.width = 8, 16
    data = Reference(ws, min_col=3, max_col=4, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=2, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H7")

    top = last + 3
    rows = [[g, f'=COUNTIF(Database!$I:$I,B{top + 2 + i})',
             f'=COUNTIF(Database!$J:$J,B{top + 2 + i})']
            for i, g in enumerate(range(1, 9))]
    gfirst, glast = _table(ws, top, 2, "Grade Distribution (LOB vs CRR)",
                           ["Grade", "LOB Assigned", "CRR Concluded"], rows)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Grade Distribution: Assigned vs CRR-Concluded"
    chart2.height, chart2.width = 8, 16
    data = Reference(ws, min_col=3, max_col=4, min_row=gfirst - 1, max_row=glast)
    cats = Reference(ws, min_col=2, min_row=gfirst, max_row=glast)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, f"H{top}")

    # Migration matrix: LOB grade (rows) x CRR grade (cols)
    top = glast + 3
    ws.cell(row=top, column=2, value="Migration Matrix — LOB Assigned (rows) vs CRR Concluded (columns)").font = st.H3_FONT
    st.col_headers(ws, top + 1, ["LOB \\ CRR"] + [str(g) for g in range(1, 9)], start_col=2)
    for i, g in enumerate(range(1, 9)):
        r = top + 2 + i
        c = ws.cell(row=r, column=2, value=g)
        c.font = st.BOLD_FONT
        c.border = st.BOX
        c.alignment = st.CENTER
        for j, g2 in enumerate(range(1, 9)):
            cell = ws.cell(row=r, column=3 + j,
                           value=f'=IF(COUNTIFS(Database!$I:$I,$B{r},Database!$J:$J,{g2})=0,"",'
                                 f'COUNTIFS(Database!$I:$I,$B{r},Database!$J:$J,{g2}))')
            cell.font = st.FORMULA_FONT
            cell.border = st.BOX
            cell.alignment = st.CENTER
    mr = f"C{top + 2}:J{top + 9}"
    ws.conditional_formatting.add(mr, FormulaRule(
        formula=[f"COLUMN()-2>$B{top + 2}"], stopIfTrue=False,
        fill=PatternFill("solid", start_color=st.PALE_RED)))
    ws.cell(row=top + 11, column=2,
            value="Shaded cells right of the diagonal = CRR downgrades vs the LOB grade.").font = st.SMALL_FONT
    return ws


def build_dash_exceptions(wb, n_resp):
    ws = wb.create_sheet("Dash_Exceptions")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [2, 24, 12, 12, 12, 3, 6, 11, 7, 9, 10, 60, 45, 10])
    st.title_bar(ws, "Exceptions / Findings Dashboard", 14,
                 subtitle="Counts ONLY 'No' answers. Obs is never an exception; "
                          "N/A is excluded from every rate.")
    R = "Responses!"
    _kpi(ws, 4, 2, "Total Exceptions", f'=COUNTIF({R}$G:$G,"No")')
    _kpi(ws, 4, 3, "High Severity", f'=COUNTIFS({R}$G:$G,"No",{R}$F:$F,"High")')
    _kpi(ws, 4, 4, "Aged > 180 Days", f'=COUNTIFS({R}$G:$G,"No",{R}$K:$K,">180")')
    _kpi(ws, 4, 5, "Notes Missing",
         f'=COUNTIFS({R}$G:$G,"No",{R}$H:$H,"")+COUNTIFS({R}$G:$G,"Obs",{R}$H:$H,"")')

    codes = list(SEGMENTS.keys())
    rows = []
    for i, code in enumerate(codes):
        r = 9 + i
        app = (f'COUNTIFS({R}$B:$B,$B{r},{R}$G:$G,"Yes")'
               f'+COUNTIFS({R}$B:$B,$B{r},{R}$G:$G,"No")'
               f'+COUNTIFS({R}$B:$B,$B{r},{R}$G:$G,"Obs")')
        rows.append([code,
                     f'=COUNTIFS({R}$B:$B,$B{r},{R}$G:$G,"No")',
                     f'=IF({app}=0,"",COUNTIFS({R}$B:$B,$B{r},{R}$G:$G,"No")/({app}))'])
    first, last = _table(ws, 7, 2, "Exceptions by Segment",
                         ["Segment", "Exceptions", "Rate (excl. N/A)"], rows)
    for r in range(first, last + 1):
        ws.cell(row=r, column=4).number_format = st.FMT_PCT
    chart = BarChart()
    chart.type = "col"
    chart.title = "Exceptions by Segment"
    chart.height, chart.width = 7, 11
    data = Reference(ws, min_col=3, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=2, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G7")

    top = last + 3
    rows = [[sev, f'=COUNTIFS({R}$G:$G,"No",{R}$F:$F,B{top + 2 + i})']
            for i, sev in enumerate(["High", "Medium", "Low"])]
    sfirst, slast = _table(ws, top, 2, "Exceptions by Severity", ["Severity", "Count"], rows)
    pie = PieChart()
    pie.title = "By Severity"
    pie.height, pie.width = 6, 8
    data = Reference(ws, min_col=3, min_row=sfirst - 1, max_row=slast)
    cats = Reference(ws, min_col=2, min_row=sfirst, max_row=slast)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "O7")

    top = slast + 3
    rows = [[b, f'=COUNTIFS({R}$G:$G,"No",{R}$L:$L,B{top + 2 + i})']
            for i, b in enumerate(["0-90", "91-180", "181-270", ">270"])]
    afirst, alast = _table(ws, top, 2, "Exception Aging (days since review)",
                           ["Bucket", "Count"], rows)

    # Detail list driven by the NoIdx helper column - no array formulas.
    top = alast + 3
    ws.cell(row=top, column=2, value="Exception Detail (first 30)").font = st.H3_FONT
    hdr = top + 1
    st.col_headers(ws, hdr, ["#", "Credit", "Seg", "QID", "Severity",
                             "Question", "Finding Rationale", "Days Open"], start_col=7)
    for k in range(1, 31):
        r = hdr + k
        ws.cell(row=r, column=7, value=k).font = st.SMALL_FONT
        targets = [("A", 8), ("B", 9), ("C", 10), ("F", 11), ("E", 12), ("H", 13), ("K", 14)]
        for src_col, dest in targets:
            cell = ws.cell(row=r, column=dest,
                           value=f'=IFERROR(INDEX({R}${src_col}:${src_col},'
                                 f'MATCH($G{r},{R}$N:$N,0)),"")')
            cell.font = st.FORMULA_FONT
            if src_col in ("E", "H"):
                cell.alignment = st.WRAP_TOP
    return ws


def build_obs_view(wb):
    ws = wb.create_sheet("Obs_View")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [2, 38, 10, 3, 6, 11, 7, 9, 60, 45, 11])
    st.title_bar(ws, "Observations View", 11,
                 subtitle="Obs = pass with a required note. Never counted as an exception; "
                          "listed here so thematic patterns across files stay visible.")
    R = "Responses!"
    _kpi(ws, 4, 2, "Total Observations", f'=COUNTIF({R}$G:$G,"Obs")')

    # Thematic rollup: same Obs in many files is a pattern worth acting on.
    sections = []
    for s, _q, _v in CORE_QUESTIONS:
        if s not in sections:
            sections.append(s)
    for seg in SEGMENTS.values():
        for s, _qs in seg["sections_extra"]:
            if s not in sections:
                sections.append(s)
    rows = [[s, f'=COUNTIFS({R}$D:$D,B{9 + i},{R}$G:$G,"Obs")']
            for i, s in enumerate(sections)]
    first, last = _table(ws, 7, 2, "Observations by Section (thematic patterns)",
                         ["Section", "Obs Count"], rows)
    ws.conditional_formatting.add(
        f"B{first}:C{last}",
        FormulaRule(formula=[f"$C{first}>=2"],
                    fill=PatternFill("solid", start_color=st.AMBER)))
    ws.cell(row=last + 1, column=2,
            value="Amber = section with 2+ observations across files (potential thematic finding).").font = st.SMALL_FONT

    ws.cell(row=6, column=5, value="Observation Detail (first 30)").font = st.H3_FONT
    st.col_headers(ws, 7, ["#", "Credit", "Seg", "QID", "Question",
                           "Observation Note", "Review Date"], start_col=5)
    for k in range(1, 31):
        r = 7 + k
        ws.cell(row=r, column=5, value=k).font = st.SMALL_FONT
        for src_col, dest in [("A", 6), ("B", 7), ("C", 8), ("E", 9), ("H", 10), ("I", 11)]:
            cell = ws.cell(row=r, column=dest,
                           value=f'=IFERROR(INDEX({R}${src_col}:${src_col},'
                                 f'MATCH($E{r},{R}$M:$M,0)),"")')
            cell.font = st.FORMULA_FONT
            if src_col in ("E", "H"):
                cell.alignment = st.WRAP_TOP
            if src_col == "I":
                cell.number_format = st.FMT_DATE
    return ws


def build_dash_concentration(wb):
    ws = wb.create_sheet("Dash_Concentration")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [2, 42, 15, 15, 22, 3, 14, 14, 14, 16])
    st.title_bar(ws, "Concentration vs. Capital — Live Crosswalk Status", 10,
                 subtitle="Criteria pull from the confirmed Crosswalk for the bank's primary "
                          "agency, as of the Review As-Of Date. Supervisory criteria trigger "
                          "heightened scrutiny - they are not hard caps.")
    ws.cell(row=4, column=2, value="Bank inputs (edit on Settings sheet)").font = st.H3_FONT
    items = [
        ("Total Risk-Based Capital ($000)", "=TotalRBC", st.FMT_USD),
        ("Construction, Land & Development Loans ($000)", "=CLDLoans", st.FMT_USD),
        ("Total Non-Owner-Occupied CRE Loans ($000)", "=CRELoans", st.FMT_USD),
        ("Non-OO CRE Balance 36 Months Ago ($000)", "=CRELoans36", st.FMT_USD),
    ]
    for i, (label, f, fmt) in enumerate(items):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).font = st.BODY_FONT
        c = ws.cell(row=r, column=3, value=f)
        c.font = st.LINK_FONT
        c.number_format = fmt
        c.border = st.BOX

    top = 11
    ws.cell(row=top, column=2, value="Supervisory criteria status (2006 Interagency CRE Guidance, as confirmed in Crosswalk)").font = st.H3_FONT
    st.col_headers(ws, top + 1, ["Measure", "Bank Level", "Criterion (In Force)", "Status"], start_col=2)
    xl = ('SUMPRODUCT(MAX((Crosswalk!$O$2:$O$300="{m}|"&PrimaryAgency)'
          '*(Crosswalk!$P$2:$P$300=1)*Crosswalk!$C$2:$C$300))')
    cld_t = xl.format(m="CLD Concentration / Total Risk-Based Capital")
    cre_t = xl.format(m="Total CRE / Total Risk-Based Capital")
    gro_t = xl.format(m="CRE 36-Month Growth Trigger")
    r = top + 2
    ws.cell(row=r, column=2, value="CLD Loans / Total Risk-Based Capital").font = st.BODY_FONT
    ws.cell(row=r, column=3, value='=IF(TotalRBC=0,"",CLDLoans/TotalRBC)').number_format = st.FMT_PCT
    ws.cell(row=r, column=4, value=f'=IF({cld_t}=0,"Coverage gap",{cld_t}/100)').number_format = st.FMT_PCT
    ws.cell(row=r, column=5, value=f'=IF(OR(C{r}="",ISTEXT(D{r})),"n/a",IF(C{r}>=D{r},"Criterion met — heightened scrutiny","Below criterion"))')
    r += 1
    ws.cell(row=r, column=2, value="Total Non-OO CRE / Total Risk-Based Capital").font = st.BODY_FONT
    ws.cell(row=r, column=3, value='=IF(TotalRBC=0,"",CRELoans/TotalRBC)').number_format = st.FMT_PCT
    ws.cell(row=r, column=4, value=f'=IF({cre_t}=0,"Coverage gap",{cre_t}/100)').number_format = st.FMT_PCT
    ws.cell(row=r, column=5, value=f'=IF(OR(C{r}="",ISTEXT(D{r})),"n/a",IF(C{r}>=D{r},"Part 1 met — test growth","Below criterion"))')
    r += 1
    ws.cell(row=r, column=2, value="36-Month CRE Growth").font = st.BODY_FONT
    ws.cell(row=r, column=3, value='=IF(CRELoans36=0,"",(CRELoans-CRELoans36)/CRELoans36)').number_format = st.FMT_PCT
    ws.cell(row=r, column=4, value=f'=IF({gro_t}=0,"Coverage gap",{gro_t}/100)').number_format = st.FMT_PCT
    ws.cell(row=r, column=5, value=f'=IF(OR(C{r}="",ISTEXT(D{r})),"n/a",IF(AND(C{r}>=D{r},C{r-1}>=D{r-1}),"Combined criterion met — heightened scrutiny","Below combined criterion"))')
    for rr in range(top + 2, r + 1):
        for col in range(2, 6):
            ws.cell(row=rr, column=col).border = st.BOX
            if col >= 3:
                ws.cell(row=rr, column=col).font = st.FORMULA_FONT if col != 4 else st.LINK_FONT
    ws.conditional_formatting.add(f"B{top + 2}:E{r}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("met",$E{top + 2}))'],
        fill=PatternFill("solid", start_color=st.PALE_RED)))

    # Versioning proof: same threshold, three agencies, different status as-of.
    top = r + 3
    ws.cell(row=top, column=2,
            value="Leveraged-Lending Guidance status by agency (Total Debt / EBITDA criterion, as of Review As-Of Date)").font = st.H3_FONT
    st.col_headers(ws, top + 1, ["Agency", "Threshold", "Effective", "Rescinded", "Status As-Of"], start_col=2)
    for i, agency in enumerate(["OCC", "FDIC", "FRB"]):
        rr = top + 2 + i
        key = f'"Total Debt / EBITDA|{agency}"'
        ws.cell(row=rr, column=2, value=agency).font = st.BODY_FONT
        for dest, src in [(3, "C"), (4, "H"), (5, "I"), (6, "J")]:
            idx = f'INDEX(Crosswalk!${src}:${src},MATCH({key},Crosswalk!$O:$O,0))'
            c = ws.cell(row=rr, column=dest,
                        value=f'=IFERROR(IF({idx}=0,"—",{idx}),"not in crosswalk")')
            c.font = st.LINK_FONT
            c.border = st.BOX
            if dest == 3:
                c.number_format = st.FMT_X
            elif dest in (4, 5):
                c.number_format = st.FMT_DATE
        ws.cell(row=rr, column=2).border = st.BOX
    ws.conditional_formatting.add(f"B{top + 2}:F{top + 4}", FormulaRule(
        formula=[f'$F{top + 2}="Rescinded"'],
        fill=PatternFill("solid", start_color=st.LIGHT_GREY)))
    ws.cell(row=top + 6, column=2,
            value="A threshold can be Active for one agency and Rescinded for another on the same date; "
                  "reviews cite the rule in force at the review as-of date.").font = st.SMALL_FONT
    return ws


def build_dash_trends(wb):
    ws = wb.create_sheet("Dash_Trends")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [2, 14, 12, 14, 14, 14, 3] + [10] * 8)
    st.title_bar(ws, "Risk-Grade & Watch-List Trend", 15,
                 subtitle="By review quarter, from the Database. Watch = Special Mention (5); "
                          "Classified = Substandard or worse (6+).")
    D = "Database!"
    rows = []
    for i, q in enumerate(QUARTERS):
        r = 7 + i
        rows.append([
            q,
            f'=COUNTIF({D}$G:$G,B{r})',
            f'=COUNTIFS({D}$G:$G,B{r},{D}$J:$J,5)',
            f'=COUNTIFS({D}$G:$G,B{r},{D}$J:$J,">=6")',
            f'=IF(COUNTIF({D}$G:$G,B{r})=0,"",AVERAGEIFS({D}$J:$J,{D}$G:$G,B{r}))',
        ])
    first, last = _table(ws, 5, 2, "Quarterly trend",
                         ["Quarter", "Reviews", "Watch (SM)", "Classified (6+)", "Avg CRR Grade"], rows)
    for r in range(first, last + 1):
        ws.cell(row=r, column=6).number_format = "0.0"

    chart = LineChart()
    chart.title = "Watch & Classified Trend"
    chart.height, chart.width = 8, 14
    data = Reference(ws, min_col=4, max_col=5, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=2, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H5")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Reviews per Quarter"
    bar.height, bar.width = 8, 14
    data = Reference(ws, min_col=3, min_row=first - 1, max_row=last)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "H22")
    return ws
