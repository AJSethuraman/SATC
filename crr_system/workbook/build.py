"""Build the complete CRR Line Sheet System workbook.

    python -m workbook.build [output_path]

Order of operations:
  1. Static sheets (Home, Settings, Lists, Questions)
  2. Six segment line sheets (C&I is the reference template)
  3. Database + Responses populated with deterministic sample credits
  4. Dashboards
  5. Extraction engine runs on the fixture documents -> staging sheets
  6. Simulated reviewer confirms High-confidence rows (demo only - in
     production a human does this in Excel), engine promotes to Crosswalk /
     Assertions
  7. Sheet ordering/tab colors, then LibreOffice recalculation must report
     zero formula errors.
"""

from __future__ import annotations

import datetime as dt
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from engine import crosswalk as xw                      # noqa: E402
from engine import extract, ingest, staging             # noqa: E402
from workbook import dashboards, database, forms, sampledata  # noqa: E402
from workbook import styles as st                       # noqa: E402
from workbook.content import RATING_SCALE, SEGMENTS, question_rows  # noqa: E402

OUT_DEFAULT = ROOT / "output" / "CRR_Line_Sheet_System.xlsx"
FIX = ROOT / "fixtures"


# ---------------------------------------------------------------------------
# Static sheets
# ---------------------------------------------------------------------------

def build_home(wb):
    ws = wb.create_sheet("Home")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [3, 26, 70, 18, 30])
    st.title_bar(ws, "Credit Risk Review — Line Sheet System", 5,
                 subtitle="Independent review per the 2020 Interagency Guidance on Credit Risk "
                          "Review Systems: rely on LOB information only after critically "
                          "evaluating it; every conclusion traces to a source.")
    r = 4
    ws.cell(row=r, column=2, value="How the system works").font = st.H3_FONT
    flow = [
        ("1. Ingest", "Run the engine on source documents: regulatory PDFs, internal policy, "
                      "credit memos (python -m engine extract-thresholds / extract-cam)."),
        ("2. Stage", "Extracted rows land on Staging_TypeA / Staging_TypeB with the verbatim "
                     "source span, page/section, and a confidence flag. Nothing is live yet."),
        ("3. Confirm", "A reviewer marks each staged row Confirmed or Rejected. Low-confidence "
                       "and Coverage Gap rows are highlighted for resolution."),
        ("4. Promote", "python -m engine promote copies ONLY Confirmed rows to the Crosswalk "
                       "(thresholds) and Assertions (credit-memo facts) sheets."),
        ("5. Review", "Line sheets re-derive each asserted value independently, compare to the "
                      "in-force threshold for the review date and agency, and record "
                      "Yes / No / N/A / Obs answers."),
        ("6. Report", "Database and dashboards roll up exceptions (No only), observations "
                      "(Obs only), rates (N/A excluded), grades, and concentrations."),
    ]
    for i, (step, desc) in enumerate(flow):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=step).font = st.BOLD_FONT
        c = ws.cell(row=rr, column=3, value=desc)
        c.font = st.BODY_FONT
        c.alignment = st.WRAP_TOP
        ws.row_dimensions[rr].height = 26

    r += len(flow) + 2
    ws.cell(row=r, column=2, value="Answer convention (identical on every form)").font = st.H3_FONT
    legend = [
        ("Yes", "Clean pass. Note optional."),
        ("No", "Exception / finding. Drives the exception rollup and aging. Rationale note required."),
        ("N/A", "Not applicable. Excluded from numerator AND denominator of every rate."),
        ("Obs", "Pass with a note on record. Note required. Never counted as an exception; "
                "surfaces only in the Observations view for thematic patterns."),
    ]
    for i, (a, d) in enumerate(legend):
        rr = r + 1 + i
        c = ws.cell(row=rr, column=2, value=a)
        c.font = st.BOLD_FONT
        c.alignment = st.CENTER
        c.border = st.BOX
        ws.cell(row=rr, column=3, value=d).font = st.BODY_FONT

    r += len(legend) + 2
    ws.cell(row=r, column=2, value="Color conventions").font = st.H3_FONT
    colors = [
        ("Blue", st.INPUT_FONT, "Hardcoded inputs you can change"),
        ("Black", st.FORMULA_FONT, "Formulas and calculations"),
        ("Green", st.LINK_FONT, "Links pulling from other sheets (Crosswalk, Settings)"),
        ("Yellow fill", st.BODY_FONT, "Key assumptions needing attention"),
    ]
    for i, (name, font, desc) in enumerate(colors):
        rr = r + 1 + i
        c = ws.cell(row=rr, column=2, value=name)
        c.font = font
        if name == "Yellow fill":
            c.fill = st.ASSUMPTION_FILL
        ws.cell(row=rr, column=3, value=desc).font = st.BODY_FONT

    r += len(colors) + 2
    ws.cell(row=r, column=2, value="Navigation").font = st.H3_FONT
    nav = [
        ("Settings", "As-of date, primary agency, capital inputs"),
        ("LS_CI", "C&I / Commercial line sheet (reference template)"),
        ("LS_CRE", "Commercial Real Estate line sheet"),
        ("LS_LL", "Leveraged Lending line sheet"),
        ("LS_ABL", "Asset-Based Lending line sheet"),
        ("LS_ARG", "ARG / Workout line sheet"),
        ("LS_COMP", "General Compliance line sheet"),
        ("LS_IA", "Internal Audit test sheet - generic template for audit shops"),
        ("Database", "One row per credit per review"),
        ("Responses", "One row per question per credit"),
        ("Dash_Portfolio", "Portfolio overview & grade migration"),
        ("Dash_Exceptions", "Exceptions (No answers only), severity, aging"),
        ("Obs_View", "Observations only - thematic patterns"),
        ("Dash_Concentration", "Concentration vs capital, live crosswalk status"),
        ("Dash_Trends", "Watch-list / classified trend"),
        ("Crosswalk", "Confirmed thresholds with provenance & dating"),
        ("Assertions", "Confirmed credit-memo facts vs independent values"),
        ("Staging_TypeA", "Extracted thresholds awaiting confirmation"),
        ("Staging_TypeB", "Extracted credit-memo assertions awaiting confirmation"),
    ]
    for i, (sheet, desc) in enumerate(nav):
        rr = r + 1 + i
        c = ws.cell(row=rr, column=2, value=sheet)
        c.hyperlink = f"#'{sheet}'!A1"
        c.font = Font(name=st.FONT, size=10, color="0563C1", underline="single")
        ws.cell(row=rr, column=3, value=desc).font = st.BODY_FONT
    st.whiten(ws, 4, 2, r + len(nav) + 1, 5)
    st.canvas_pass(ws, 7, r + len(nav) + 8)
    return ws


def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [3, 44, 16, 56])
    st.title_bar(ws, "Settings & Bank Inputs", 4)
    ws.cell(row=3, column=2, value="Review controls").font = st.H3_FONT
    items = [
        (4, "Review As-Of Date", dt.date(2026, 6, 10), st.FMT_DATE,
         "Drives crosswalk Active/Rescinded status and exception aging."),
        (5, "Primary Supervisory Agency", "OCC", None,
         "Agency whose thresholds the dashboards cite. Forms use this plus internal policy."),
    ]
    for r, label, value, fmt, note in items:
        ws.cell(row=r, column=2, value=label).font = st.BOLD_FONT
        c = ws.cell(row=r, column=3, value=value)
        c.font = st.INPUT_FONT
        c.border = st.BOX
        if fmt:
            c.number_format = fmt
        ws.cell(row=r, column=4, value=note).font = st.SMALL_FONT
    dv = DataValidation(type="list", formula1="=AgencyList", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("C5")

    ws.cell(row=7, column=2, value="Capital & concentration inputs").font = st.H3_FONT
    cap = [
        (8, "Total Risk-Based Capital ($000)", 185_000,
         "Source: Call Report Schedule RC-R, 3/31/2026 filing."),
        (9, "Construction, Land & Development Loans ($000)", 142_000,
         "Source: internal MIS concentration report, 5/31/2026."),
        (10, "Total Non-Owner-Occupied CRE Loans ($000)", 612_000,
         "Source: internal MIS concentration report, 5/31/2026."),
        (11, "Non-OO CRE Balance 36 Months Ago ($000)", 425_000,
         "Source: internal MIS concentration report, 5/31/2023."),
    ]
    for r, label, value, src in cap:
        ws.cell(row=r, column=2, value=label).font = st.BOLD_FONT
        c = ws.cell(row=r, column=3, value=value)
        c.font = st.INPUT_FONT
        c.border = st.BOX
        c.number_format = st.FMT_USD
        ws.cell(row=r, column=4, value=src).font = st.SMALL_FONT

    names = {"AsOfDate": "$C$4", "PrimaryAgency": "$C$5", "TotalRBC": "$C$8",
             "CLDLoans": "$C$9", "CRELoans": "$C$10", "CRELoans36": "$C$11"}
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"Settings!{ref}"))
    st.whiten(ws, 3, 2, 11, 4)
    st.canvas_pass(ws, 6, 19)
    return ws


def build_lists(wb):
    ws = wb.create_sheet("Lists")
    st.set_widths(ws, [10, 12, 10, 7, 16, 12, 10, 11, 3, 7, 16, 70])
    cols = {
        "A": ("Answers", ["Yes", "No", "N/A", "Obs"]),
        "B": ("Status", ["In Progress", "Complete"]),
        "C": ("Agencies", ["OCC", "FRB", "FDIC", "CFPB", "Internal"]),
        "D": ("Grade", [g for g, _c, _d in RATING_SCALE]),
        "E": ("Category", [c for _g, c, _d in RATING_SCALE]),
        "F": ("PropType", ["Office", "Multifamily", "Retail", "Industrial", "Hospitality", "Other"]),
        "G": ("Severity", ["High", "Medium", "Low"]),
        "H": ("ReviewType", ["Periodic", "Targeted", "Continuous"]),
    }
    for col, (header, values) in cols.items():
        h = ws[f"{col}1"]
        h.value = header
        h.font = st.BOLD_FONT
        h.fill = st.SUBHEAD_FILL
        h.border = st.BOX
        for i, v in enumerate(values):
            c = ws[f"{col}{2 + i}"]
            c.value = v
            c.font = st.BODY_FONT
            c.border = st.BOX
    ws["J1"].value, ws["K1"].value, ws["L1"].value = "Grade", "Category", "Regulatory Definition"
    for cell in ("J1", "K1", "L1"):
        ws[cell].font = st.BOLD_FONT
        ws[cell].fill = st.SUBHEAD_FILL
        ws[cell].border = st.BOX
    for i, (g, cat, desc) in enumerate(RATING_SCALE):
        ws[f"J{2 + i}"].value = g
        ws[f"K{2 + i}"].value = cat
        ws[f"L{2 + i}"].value = desc
        for col in "JKL":
            ws[f"{col}{2 + i}"].font = st.BODY_FONT
            ws[f"{col}{2 + i}"].border = st.BOX

    ranges = {"AnswerList": "$A$2:$A$5", "StatusList": "$B$2:$B$3",
              "AgencyList": "$C$2:$C$6", "GradeList": "$D$2:$D$9",
              "PropTypeList": "$F$2:$F$7"}
    for name, ref in ranges.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"Lists!{ref}"))
    st.whiten(ws, 1, 1, 10, 12)
    st.canvas_pass(ws, 14, 16)
    return ws


def build_questions(wb):
    ws = wb.create_sheet("Questions")
    st.set_widths(ws, [10, 7, 38, 95, 10])
    st.col_headers(ws, 1, ["QID", "Seg", "Section (2020 CRR criteria)", "Question", "Severity"],
                   fill=st.SECTION_FILL)
    for c in ws[1]:
        c.font = st.H2_FONT
    r = 2
    for code in SEGMENTS:
        for qid, section, question, severity in question_rows(code):
            for col, v in enumerate([qid, code, section, question, severity], start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = st.BODY_FONT
                if col == 4:
                    cell.alignment = st.WRAP_TOP
            r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{r - 1}"
    st.whiten(ws, 1, 1, r - 1, 5)
    st.canvas_pass(ws, 7, r + 6)
    return ws


# ---------------------------------------------------------------------------
# Engine pipeline on fixtures (steps 5-6)
# ---------------------------------------------------------------------------

def run_engine_pipeline(out_path: Path):
    ll = extract.extract_thresholds(
        ingest.ingest(str(FIX / "fixture_2013_leveraged_lending.pdf")),
        citation="2013 Interagency Guidance on Leveraged Lending (OCC 2013-9 et al.)")
    notices = extract.extract_rescissions(
        ingest.ingest(str(FIX / "fixture_2025_rescission.pdf")))
    ll = xw.apply_rescissions(ll, notices)
    cre = extract.extract_thresholds(
        ingest.ingest(str(FIX / "fixture_2006_cre_guidance.pdf")),
        citation="2006 Interagency Guidance on CRE Concentration Risk Management")
    pol = extract.extract_thresholds(
        ingest.ingest(str(FIX / "fixture_internal_policy.docx")),
        citation="Commercial Credit Policy CP-100")
    cam = extract.extract_assertions(ingest.ingest(str(FIX / "fixture_sample_cam.docx")))
    staging.write_rows(str(out_path), ll + cre + pol + cam)


def simulate_reviewer_confirmation(out_path: Path):
    """Demo stand-in for the human confirmation step: confirm High-confidence
    rows, leave Medium/Low and Coverage Gap rows Pending so the gate is
    visible, and fill a few independent values to show the challenge columns.
    """
    wb = load_workbook(out_path)
    wsa = wb[staging.STAGING_A]
    for r in range(2, wsa.max_row + 1):
        if (wsa.cell(row=r, column=11).value == "High"
                and wsa.cell(row=r, column=10).value == "Staged"):
            wsa.cell(row=r, column=16).value = "Confirmed"
            wsa.cell(row=r, column=17).value = "Verified against source p./section cited."
    wsb = wb[staging.STAGING_B]
    independents = {
        "Total Debt / EBITDA (asserted)": (3.80, "Agree"),
        "Senior Debt / EBITDA (asserted)": (2.90, "Agree"),
        "DSCR (asserted)": (1.42, "Agree"),
        "Interest Coverage (asserted)": (5.03, "Disagree"),
        "Global DSCR (asserted)": (1.46, "Disagree"),
    }
    for r in range(2, wsb.max_row + 1):
        if wsb.cell(row=r, column=11).value == "High":
            wsb.cell(row=r, column=16).value = "Confirmed"
        metric = wsb.cell(row=r, column=5).value
        if metric in independents:
            val, verdict = independents[metric]
            wsb.cell(row=r, column=13).value = val
            wsb.cell(row=r, column=15).value = verdict
            if verdict == "Disagree":
                wsb.cell(row=r, column=17).value = (
                    "CRR re-derivation from spreads does not tie to memo; "
                    "see line-sheet ratio engine.")
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Finishing: order, tabs
# ---------------------------------------------------------------------------

SHEET_ORDER = ["Home", "Settings", "Lists", "Questions",
               "LS_CI", "LS_CRE", "LS_LL", "LS_ABL", "LS_ARG", "LS_COMP", "LS_IA",
               "Database", "Responses",
               "Dash_Portfolio", "Dash_Exceptions", "Obs_View",
               "Dash_Concentration", "Dash_Trends",
               "Crosswalk", "Assertions", "Staging_TypeA", "Staging_TypeB"]

TAB_COLORS = {
    "Home": st.NAVY, "Settings": st.NAVY, "Lists": "BF8F00", "Questions": "BF8F00",
    "LS_CI": st.MID_BLUE, "LS_CRE": st.MID_BLUE, "LS_LL": st.MID_BLUE,
    "LS_ABL": st.MID_BLUE, "LS_ARG": st.MID_BLUE, "LS_COMP": st.MID_BLUE,
    "LS_IA": st.MID_BLUE,
    "Database": "808080", "Responses": "808080",
    "Dash_Portfolio": "548235", "Dash_Exceptions": "548235", "Obs_View": "548235",
    "Dash_Concentration": "548235", "Dash_Trends": "548235",
    "Crosswalk": "BF8F00", "Assertions": "BF8F00",
    "Staging_TypeA": "C00000", "Staging_TypeB": "C00000",
}


def finish(out_path: Path):
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = load_workbook(out_path)
    order = [s for s in SHEET_ORDER if s in wb.sheetnames]
    wb._sheets = [wb[s] for s in order] + [ws for ws in wb._sheets if ws.title not in order]
    for name, color in TAB_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
    wb.active = 0
    wb.save(out_path)


def main(out_path: Path = OUT_DEFAULT) -> dict:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    build_home(wb)
    build_settings(wb)
    build_lists(wb)
    build_questions(wb)

    credits, responses = sampledata.generate()
    first_credit = {c["segment"]: c for c in reversed(credits)}
    for code in SEGMENTS:
        cr = first_credit[code]
        answers = {r["qid"]: (r["answer"], r["note"]) for r in responses
                   if r["credit_id"] == cr["credit_id"]}
        inputs = sampledata.MERIDIAN_INPUTS if code == "CI" else sampledata.FORM_INPUTS[code]
        asserted = sampledata.MERIDIAN_ASSERTED if code == "CI" else sampledata.FORM_ASSERTED[code]
        ws = forms.build_form(wb, code, inputs_data=inputs,
                              asserted_data=asserted, answers_by_qid=answers)
        hdr = {
            "credit_id": cr["credit_id"], "borrower": cr["borrower"],
            "commitment": cr["commitment"], "review_type": "Periodic",
            "review_date": cr["review_date"], "fin_date": cr["review_date"] - dt.timedelta(days=75),
            "reviewer": cr["reviewer"], "subtype": "Office" if code == "CRE" else None,
            "lob_grade": cr["lob_grade"], "crr_grade": cr["crr_grade"],
        }
        for key, value in hdr.items():
            ws[forms.FORM_CELLS[key]].value = value

    database.build_database(wb, credits)
    database.build_responses(wb, responses)

    dashboards.build_dash_portfolio(wb, len(credits))
    dashboards.build_dash_exceptions(wb, len(responses))
    dashboards.build_obs_view(wb)
    dashboards.build_dash_concentration(wb)
    dashboards.build_dash_trends(wb)

    wb.save(out_path)

    run_engine_pipeline(out_path)
    simulate_reviewer_confirmation(out_path)
    staging.promote_confirmed(str(out_path))
    finish(out_path)

    result = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "recalc.py"), str(out_path), "120"],
        capture_output=True, text=True)
    try:
        recalc = json.loads(result.stdout)
    except json.JSONDecodeError:
        recalc = {"error": result.stdout + result.stderr}
    return recalc


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else OUT_DEFAULT
    print(json.dumps(main(target), indent=2, default=str))
