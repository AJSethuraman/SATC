"""Line-sheet form builder. One builder, six segments (config-driven).

Form anatomy:
  A. Credit identification (inputs + grade-concurrence formula)
  B. Financial inputs - the CRR reviewer's independent spread (blue inputs)
  C. Ratio engine - independent (formula) vs. asserted (input) with variance,
     alignment flag, in-force threshold pulled from the confirmed Crosswalk
     (green cross-sheet link), and pass/exception flag
  D+ Question sections mapped to the 2020 CRR review criteria, answered
     Yes / No / N/A / Obs with note column and note-required checks
  Review summary - completion %, exceptions (No only), observations (Obs only),
     exception rate excluding N/A
"""

from __future__ import annotations

import re

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from . import styles as st
from .content import SEGMENTS, question_rows

LAST_COL = 10  # col J

FMTS = {
    "usd": st.FMT_USD, "pct": st.FMT_PCT, "x": st.FMT_X, "x2": st.FMT_X2,
    "num": st.FMT_NUM, "num1": "0.0;(0.0);\"-\"",
}

# Crosswalk metrics sourced from internal policy rather than a regulator.
THRESHOLD_AGENCY = {
    "Minimum DSCR": '"Internal"',
    "Maximum LTV": '"Internal"',
    "Maximum Advance Rate (Eligible AR)": '"Internal"',
    "Maximum Advance Rate (Eligible Inventory)": '"Internal"',
}

# Fixed header-cell map (same on every form; used by the save-form CLI).
FORM_CELLS = {
    "credit_id": "C5", "borrower": "C6", "commitment": "C7", "review_type": "C8",
    "review_date": "F5", "fin_date": "F6", "reviewer": "F7", "subtype": "F8",
    "lob_grade": "I5", "crr_grade": "I6",
}


def _label(ws, cell, text):
    c = ws[cell]
    c.value = text
    c.font = st.BOLD_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")


def _input(ws, cell, value=None, fmt=None, assumption=False):
    c = ws[cell]
    c.value = value
    c.font = st.INPUT_FONT
    c.border = st.FIELD_BOX
    c.fill = st.ASSUMPTION_FILL if assumption else st.FIELD_FILL
    if fmt:
        c.number_format = fmt
    return c


def _formula(ws, cell, formula, fmt=None, link=False, bold=False, fill=None):
    c = ws[cell]
    c.value = formula
    c.font = Font(name=st.FONT, size=10, bold=bold,
                  color="008000" if link else "000000")
    c.border = st.BOX
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


def _threshold_formula(metric: str, direction: str, fmt: str) -> str:
    """In-force threshold as of the form's Review Date ($F$5), agency-tagged.

    A row is in force when effective <= review date and (no rescinded date or
    rescinded date > review date) - so a credit reviewed before a rescission
    still cites the rule that governed it, and one reviewed after does not.
    MAX picks the binding (most conservative) level if several rows match.
    """
    agency = THRESHOLD_AGENCY.get(metric, "PrimaryAgency")
    k, c, h, i = ("Crosswalk!$O$2:$O$300", "Crosswalk!$C$2:$C$300",
                  "Crosswalk!$H$2:$H$300", "Crosswalk!$I$2:$I$300")
    arr = (f'SUMPRODUCT(MAX(({k}="{metric}|"&{agency})*({h}<>"")*({h}<=$F$5)'
           f'*(({i}="")+({i}>$F$5))*{c}))')
    scale = "/100" if fmt == "pct" else ""
    return f'=IF({arr}=0,"",{arr}{scale})'


def build_form(wb, seg_code, *, inputs_data, asserted_data, answers_by_qid):
    """Build one segment line sheet. answers_by_qid: qid -> (answer, note)."""
    seg = SEGMENTS[seg_code]
    ws = wb.create_sheet(seg["sheet"])
    ws.sheet_view.showGridLines = False
    st.set_widths(ws, [8, 52, 14, 14, 12, 14, 14, 16, 12, 30])

    title = seg.get("form_title") or f"{seg['name']} — Credit Risk Review Line Sheet"
    st.title_bar(ws, title, LAST_COL,
                 subtitle="Independent challenge function: re-derive, confirm, or dispute each "
                          "asserted value. Answers: Yes / No / N/A / Obs (Obs requires a note; "
                          "N/A is excluded from all rates).")

    # --- A. Credit identification -------------------------------------------
    labels = {
        "credit_id": "Credit ID", "borrower": "Borrower",
        "commitment": "Commitment ($000)", "review_type": "Review Type",
        "review_date": "Review Date", "fin_date": "Financials As-Of",
        "reviewer": "Reviewer", "subtype": "Subtype / Property Type",
        "lob_grade": "LOB Assigned Grade", "crr_grade": "CRR Concluded Grade",
        "category": "Regulatory Category", "concurrence": "Grade Concurrence",
        "section_a": "A.  Credit Identification",
    }
    labels.update(seg.get("header_labels", {}))
    st.section_bar(ws, 4, labels["section_a"], last_col=LAST_COL)
    pairs = [
        ("B5", labels["credit_id"]), ("B6", labels["borrower"]),
        ("B7", labels["commitment"]), ("B8", labels["review_type"]),
        ("E5", labels["review_date"]), ("E6", labels["fin_date"]),
        ("E7", labels["reviewer"]), ("E8", labels["subtype"]),
        ("H5", labels["lob_grade"]), ("H6", labels["crr_grade"]),
        ("H7", labels["category"]), ("H8", labels["concurrence"]),
    ]
    for cell, text in pairs:
        _label(ws, cell, text)
    for key in ("credit_id", "borrower", "commitment", "review_type",
                "review_date", "fin_date", "reviewer", "subtype",
                "lob_grade", "crr_grade"):
        _input(ws, FORM_CELLS[key])
    ws["C7"].number_format = FMTS[seg.get("commitment_fmt", "usd")]
    ws["F5"].number_format = st.FMT_DATE
    ws["F6"].number_format = st.FMT_DATE
    _formula(ws, "I7", '=IF($I$6="","",VLOOKUP($I$6,Lists!$D$2:$E$9,2,0))',
             link=True, fill=st.GOLDTINT_FILL)
    _formula(ws, "I8", '=IF(OR($I$5="",$I$6=""),"",IF($I$5=$I$6,"Concur","Grade Change"))',
             fill=st.ICE_FILL)

    dv_grade = DataValidation(type="list", formula1="=GradeList", allow_blank=True)
    ws.add_data_validation(dv_grade)
    dv_grade.add("I5:I6")
    dv_rt = DataValidation(type="list", formula1='"Periodic,Targeted,Continuous"', allow_blank=True)
    ws.add_data_validation(dv_rt)
    dv_rt.add("C8")
    dv_pt = DataValidation(type="list", formula1="=PropTypeList", allow_blank=True)
    ws.add_data_validation(dv_pt)
    dv_pt.add("F8")

    row = 10

    # --- B. Financial inputs -------------------------------------------------
    input_cells = {}
    if seg["inputs"]:
        st.section_bar(ws, row, "B.  Financial Inputs — CRR Independent Spread", last_col=LAST_COL)
        row += 1
        st.col_headers(ws, row, ["", "Input", "Value", "Source / Note"], height=14)
        row += 1
        for label, kind, note in seg["inputs"]:
            lab = ws.cell(row=row, column=2, value=label)
            lab.font = st.BODY_FONT
            lab.fill = st.LABEL_FILL
            lab.border = st.BOX
            cell = f"C{row}"
            assumption = "Key assumption" in note
            _input(ws, cell, inputs_data.get(label), FMTS[kind], assumption=assumption)
            n = ws.cell(row=row, column=4, value=note or None)
            n.font = st.SMALL_FONT
            input_cells[label] = f"${cell[0]}${cell[1:]}"
            row += 1
        row += 1

    # --- C. Ratio engine -------------------------------------------------------
    ratio_first = ratio_last = None
    if seg["ratios"]:
        ratio_title = ("C.  Ratio Engine — Independent vs. Asserted" if seg["inputs"]
                       else "B.  Ratio Engine — Independent vs. Asserted")
        st.section_bar(ws, row, ratio_title, last_col=LAST_COL)
        row += 1
        asserted = seg.get("asserted_label") or "Per CAM (Asserted)"
        independent = seg.get("independent_label") or "CRR Independent"
        st.col_headers(ws, row, [
            "ID", "Ratio", independent, asserted, "Variance",
            "Alignment", "Threshold (In Force)", "Vs Threshold", "", "Note",
        ])
        row += 1
        ratio_first = row
        for rid, label, template, fmt, metric, direction in seg["ratios"]:
            ws.cell(row=row, column=1, value=rid).font = st.SMALL_FONT
            rl = ws.cell(row=row, column=2, value=label)
            rl.font = st.BODY_FONT
            rl.fill = st.LABEL_FILL
            rl.border = st.BOX

            formula = re.sub(r"\[([^\]]+)\]", lambda m: input_cells[m.group(1)], template)
            _formula(ws, f"C{row}", formula, FMTS[fmt], fill=st.ICE_FILL)
            _input(ws, f"D{row}", asserted_data.get(rid), FMTS[fmt])
            _formula(ws, f"E{row}",
                     f'=IF(OR(ISTEXT(C{row}),C{row}="",D{row}=""),"",C{row}-D{row})',
                     FMTS[fmt], fill=st.ICE_FILL)
            _formula(ws, f"F{row}",
                     f'=IF(E{row}="","",IF(ABS(E{row})<=0.05*MAX(ABS(D{row}),0.0001),'
                     f'"Aligned","Variance"))', fill=st.ICE_FILL)
            if metric:
                _formula(ws, f"G{row}", _threshold_formula(metric, direction, fmt),
                         FMTS[fmt], link=True, fill=st.GOLDTINT_FILL)
                op = ">" if direction == "max" else "<"
                _formula(ws, f"H{row}",
                         f'=IF(OR(G{row}="",ISTEXT(C{row}),C{row}=""),"n/a",'
                         f'IF(C{row}{op}G{row},"Exception","Pass"))', fill=st.ICE_FILL)
                ws.cell(row=row, column=8).alignment = st.CENTER
            else:
                st.style_cell(ws.cell(row=row, column=7), border=st.BOX)
                c = ws.cell(row=row, column=8, value="n/a")
                c.font = st.SMALL_FONT
                c.border = st.BOX
                c.alignment = st.CENTER
            note_cell = ws.cell(row=row, column=10)
            note_cell.border = st.FIELD_BOX
            note_cell.fill = st.FIELD_FILL
            note_cell.font = st.INPUT_FONT
            row += 1
        ratio_last = row - 1
        rng = f"A{ratio_first}:J{ratio_last}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'$H{ratio_first}="Exception"'],
            fill=PatternFill("solid", start_color=st.PALE_RED)))
        ws.conditional_formatting.add(
            f"H{ratio_first}:H{ratio_last}",
            FormulaRule(formula=[f'H{ratio_first}="Pass"'],
                        font=Font(name=st.FONT, size=10, bold=True, color="2E7D32"),
                        fill=PatternFill("solid", start_color=st.PALE_GREEN)))
        ws.conditional_formatting.add(
            f"H{ratio_first}:H{ratio_last}",
            FormulaRule(formula=[f'H{ratio_first}="Exception"'],
                        font=Font(name=st.FONT, size=10, bold=True, color=st.ALERT_RED)))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'$F{ratio_first}="Variance"'],
            fill=PatternFill("solid", start_color=st.AMBER)))
        row += 1

    # --- D+. Question sections ------------------------------------------------
    sec_letter = ord("B") + (1 if seg["inputs"] else 0) + (1 if seg["ratios"] else 0)
    q_rows = []
    current_section = None
    questions = list(question_rows(seg_code))
    q_first = None
    for qid, section, question, severity in questions:
        if section != current_section:
            st.section_bar(ws, row, f"{chr(sec_letter)}.  {section}", last_col=LAST_COL)
            sec_letter += 1
            row += 1
            st.col_headers(ws, row, ["ID", "Review Question", "Severity", "Answer",
                                     "Reviewer Note / Finding Rationale", "", "", "", "", "Check"])
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
            row += 1
            current_section = section
        if q_first is None:
            q_first = row
        idc = ws.cell(row=row, column=1, value=qid)
        idc.font = st.SMALL_FONT
        idc.alignment = st.WRAP_TOP
        q = ws.cell(row=row, column=2, value=question)
        q.font = st.BODY_FONT
        q.alignment = st.WRAP_TOP
        sv = ws.cell(row=row, column=3, value=severity)
        sv.font = st.SEV_FONTS.get(severity, st.SMALL_FONT)
        sv.alignment = Alignment(horizontal="center", vertical="top")
        ans, note = answers_by_qid.get(qid, (None, None))
        a = _input(ws, f"D{row}", ans)
        a.alignment = st.CENTER
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
        n = ws.cell(row=row, column=5, value=note or None)
        n.font = st.INPUT_FONT
        n.alignment = st.WRAP_TOP
        n.border = st.FIELD_BOX
        n.fill = st.FIELD_FILL
        _formula(ws, f"J{row}",
                 f'=IF(AND(D{row}="Obs",E{row}=""),"Note required",'
                 f'IF(AND(D{row}="No",E{row}=""),"Rationale required",""))')
        ws.cell(row=row, column=10).font = Font(name=st.FONT, size=8, bold=True, color="C00000")
        for col in range(1, LAST_COL + 1):
            ws.cell(row=row, column=col).border = st.BOX
        ws.row_dimensions[row].height = 26
        q_rows.append(row)
        row += 1
    q_last = row - 1

    dv_ans = DataValidation(type="list", formula1="=AnswerList", allow_blank=True)
    ws.add_data_validation(dv_ans)
    for r in q_rows:
        dv_ans.add(f"D{r}")
    qrng = f"A{q_first}:J{q_last}"
    ws.conditional_formatting.add(qrng, FormulaRule(
        formula=[f'$D{q_first}="No"'], fill=PatternFill("solid", start_color=st.PALE_RED)))
    ws.conditional_formatting.add(
        f"D{q_first}:D{q_last}",
        FormulaRule(formula=[f'D{q_first}="Yes"'],
                    font=Font(name=st.FONT, size=10, bold=True, color="2E7D32")))
    ws.conditional_formatting.add(
        f"D{q_first}:D{q_last}",
        FormulaRule(formula=[f'D{q_first}="No"'],
                    font=Font(name=st.FONT, size=10, bold=True, color=st.ALERT_RED)))
    ws.conditional_formatting.add(
        f"D{q_first}:D{q_last}",
        FormulaRule(formula=[f'D{q_first}="Obs"'],
                    font=Font(name=st.FONT, size=10, bold=True, color=st.GOLD)))
    ws.conditional_formatting.add(qrng, FormulaRule(
        formula=[f'AND($D{q_first}="Obs",$E{q_first}="")'],
        fill=PatternFill("solid", start_color=st.AMBER)))

    # --- Review summary --------------------------------------------------------
    row += 1
    st.section_bar(ws, row, "Review Summary", last_col=LAST_COL)
    row += 1
    drange = f'D{q_first}:D{q_last}'
    summary = [
        ("Questions (total)", f'=ROWS({drange})-COUNTIF({drange},"")'
                              f'+COUNTIF({drange},"Yes")+COUNTIF({drange},"No")'
                              f'+COUNTIF({drange},"N/A")+COUNTIF({drange},"Obs")', st.FMT_NUM),
        ("Answered (Yes / No / N/A / Obs)",
         f'=COUNTIF({drange},"Yes")+COUNTIF({drange},"No")'
         f'+COUNTIF({drange},"N/A")+COUNTIF({drange},"Obs")', st.FMT_NUM),
        ("Completion %", None, st.FMT_PCT),
        ("Applicable (excl. N/A)", None, st.FMT_NUM),
        ("Exceptions (No)", f'=COUNTIF({drange},"No")', st.FMT_NUM),
        ("Observations (Obs)", f'=COUNTIF({drange},"Obs")', st.FMT_NUM),
        ("Exception Rate (No / Applicable)", None, st.FMT_PCT),
        ("Note Checks Outstanding", f'=COUNTIF(J{q_first}:J{q_last},"*required*")', st.FMT_NUM),
    ]
    base = row
    nq = len(questions)
    for i, (label, formula, fmt) in enumerate(summary):
        r = base + i
        ws.cell(row=r, column=2, value=label).font = st.BOLD_FONT
        if label == "Questions (total)":
            formula = f"={nq}"
        elif label == "Completion %":
            formula = f'=IF(C{base}=0,"",C{base+1}/C{base})'
        elif label == "Applicable (excl. N/A)":
            formula = f'=C{base+1}-COUNTIF({drange},"N/A")'
        elif label == "Exception Rate (No / Applicable)":
            formula = f'=IF(C{base+3}=0,"",C{base+4}/C{base+3})'
        c = _formula(ws, f"C{r}", formula, fmt, bold=True, fill=st.ICE_FILL)
        c.font = st.NAVY_FONT
    row = base + len(summary) + 1
    ws.cell(row=row, column=2, value="CRR Conclusion / Grade Rationale:").font = st.BOLD_FONT
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 3, end_column=9)
    box = ws.cell(row=row + 1, column=2)
    box.font = st.INPUT_FONT
    box.alignment = st.WRAP_TOP
    box.border = st.FIELD_BOX
    box.fill = st.FIELD_FILL

    # White document panel on the gray canvas.
    st.whiten(ws, 4, 1, row + 4, LAST_COL)
    st.canvas_pass(ws, LAST_COL + 4, row + 16)
    return ws
