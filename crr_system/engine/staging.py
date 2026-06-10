"""Staging layer: extracted rows land here, never directly in live logic.

Workflow:
  1. engine writes rows to Staging_TypeA / Staging_TypeB with provenance,
     confidence, and a Pending confirmation cell;
  2. a human reviewer marks each row Confirmed or Rejected in Excel;
  3. promote_confirmed() copies Confirmed Type A rows to the Crosswalk sheet
     and Confirmed Type B rows to the Assertions sheet, which is what the
     line sheets reference. Pending/Rejected rows never go live.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from .schema import ROW_TYPE_A, ROW_TYPE_B, ExtractedRow

FONT = "Arial"
NAVY = "1F3864"
LIGHT = "D9E2F3"
AMBER = "FFF2CC"
RED_FILL = "F8CBAD"

STAGING_A = "Staging_TypeA"
STAGING_B = "Staging_TypeB"
CROSSWALK = "Crosswalk"
ASSERTIONS = "Assertions"

HEADERS_A = [
    "Row ID", "Metric", "Proposed Value", "Unit", "Basis / Definition",
    "Agency", "Citation", "Effective Date", "Rescinded Date", "Status",
    "Confidence", "Source Document", "Page / Section", "Verbatim Source Span",
    "Extractor Notes", "Reviewer Confirmation", "Reviewer Notes",
]
HEADERS_B = [
    "Row ID", "Borrower", "Facility", "Category", "Field / Metric",
    "Asserted Value (per document)", "Unit", "Source Document",
    "Page / Section", "Verbatim Source Span", "Confidence", "Extractor Notes",
    "CRR Independent Value", "Variance", "Agree / Disagree",
    "Reviewer Confirmation", "Reviewer Notes",
]
HEADERS_XWALK = [
    "Row ID", "Metric", "Value", "Unit", "Basis / Definition", "Agency",
    "Citation", "Effective Date", "Rescinded Date", "Status As-Of Review Date",
    "Source Document", "Page / Section", "Verbatim Source Span", "Reviewer Notes",
]
HEADERS_ASSERT = [
    "Row ID", "Borrower", "Facility", "Category", "Field / Metric",
    "Asserted Value (per document)", "Unit", "CRR Independent Value",
    "Variance", "Agree / Disagree", "Source Document", "Page / Section",
    "Verbatim Source Span", "Reviewer Notes",
]


def _style_header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def ensure_staging_sheets(wb) -> None:
    if STAGING_A not in wb.sheetnames:
        ws = wb.create_sheet(STAGING_A)
        _style_header(ws, HEADERS_A, [13, 30, 12, 7, 24, 9, 22, 12, 12, 13, 11, 26, 16, 60, 30, 16, 26])
        dv = DataValidation(type="list", formula1='"Pending,Confirmed,Rejected"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add("P2:P500")
        ws.conditional_formatting.add(
            "A2:Q500",
            FormulaRule(formula=['$K2="Low"'], fill=PatternFill("solid", start_color=AMBER)),
        )
        ws.conditional_formatting.add(
            "A2:Q500",
            FormulaRule(formula=['$J2="Coverage Gap"'], fill=PatternFill("solid", start_color=RED_FILL)),
        )
    if STAGING_B not in wb.sheetnames:
        ws = wb.create_sheet(STAGING_B)
        _style_header(ws, HEADERS_B, [13, 22, 22, 12, 28, 34, 7, 24, 14, 60, 11, 28, 18, 10, 14, 16, 26])
        dv = DataValidation(type="list", formula1='"Pending,Confirmed,Rejected"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add("P2:P500")
        dv2 = DataValidation(type="list", formula1='"Agree,Disagree"', allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add("O2:O500")
        ws.conditional_formatting.add(
            "A2:Q500",
            FormulaRule(formula=['$K2="Low"'], fill=PatternFill("solid", start_color=AMBER)),
        )


def _existing_ids(ws) -> set:
    return {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}


def _maybe_number(value: str):
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def write_rows(workbook_path: str, rows: Iterable[ExtractedRow]) -> dict:
    """Append extracted rows to the staging sheets (dedup by Row ID)."""
    path = Path(workbook_path)
    wb = load_workbook(path) if path.exists() else Workbook()
    if not path.exists() and "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])
    ensure_staging_sheets(wb)
    wsa, wsb = wb[STAGING_A], wb[STAGING_B]
    ids_a, ids_b = _existing_ids(wsa), _existing_ids(wsb)
    body = Font(name=FONT, size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    counts = {"A": 0, "B": 0, "skipped": 0}

    for row in rows:
        if row.row_type == ROW_TYPE_A:
            if row.row_id in ids_a:
                counts["skipped"] += 1
                continue
            r = wsa.max_row + 1
            values = [
                row.row_id, row.metric, _maybe_number(row.proposed_value), row.unit,
                row.basis, row.agency, row.citation, row.effective_date,
                row.rescinded_date, row.status, row.confidence,
                row.anchor.document, row.anchor.locator(), row.source_span,
                row.notes, row.confirmation, row.reviewer_notes,
            ]
            ws = wsa
            ids_a.add(row.row_id)
            counts["A"] += 1
        elif row.row_type == ROW_TYPE_B:
            if row.row_id in ids_b:
                counts["skipped"] += 1
                continue
            r = wsb.max_row + 1
            asserted = _maybe_number(row.proposed_value)
            var_formula = (
                f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(M{r})),M{r}-F{r},"")'
            )
            values = [
                row.row_id, row.borrower, row.facility, row.category, row.metric,
                asserted, row.unit, row.anchor.document, row.anchor.locator(),
                row.source_span, row.confidence, row.notes,
                row.independent_value or None, var_formula, None,
                row.confirmation, row.reviewer_notes,
            ]
            ws = wsb
            ids_b.add(row.row_id)
            counts["B"] += 1
        else:
            continue
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = body
            c.alignment = wrap
            if isinstance(v, dt.date):
                c.number_format = "mm/dd/yyyy"

    wb.save(path)
    return counts


def promote_confirmed(workbook_path: str) -> dict:
    """Copy Confirmed staging rows to the live Crosswalk / Assertions sheets.

    Only rows whose Reviewer Confirmation cell reads 'Confirmed' move; the
    confirmation gate is the contamination barrier between extraction output
    and live line-sheet logic.
    """
    wb = load_workbook(workbook_path)
    counts = {"crosswalk": 0, "assertions": 0}
    if "AsOfDate" not in wb.defined_names:
        # Standalone staging workbook: provide the as-of input the crosswalk
        # status formulas reference (the full system defines it on Settings).
        from openpyxl.workbook.defined_name import DefinedName

        if "Settings" not in wb.sheetnames:
            ws = wb.create_sheet("Settings")
            ws["A1"] = "Review As-Of Date"
            ws["A1"].font = Font(name=FONT, size=10, bold=True)
            ws["B1"] = dt.date.today()
            ws["B1"].font = Font(name=FONT, size=10, color="0000FF")
            ws["B1"].number_format = "mm/dd/yyyy"
        wb.defined_names.add(DefinedName("AsOfDate", attr_text="Settings!$B$1"))
    body = Font(name=FONT, size=10)
    green = Font(name=FONT, size=10, color="008000")
    wrap = Alignment(vertical="top", wrap_text=True)

    if CROSSWALK not in wb.sheetnames:
        ws = wb.create_sheet(CROSSWALK)
        _style_header(ws, HEADERS_XWALK, [13, 30, 12, 7, 24, 9, 22, 12, 12, 18, 26, 16, 60, 26])
    if ASSERTIONS not in wb.sheetnames:
        ws = wb.create_sheet(ASSERTIONS)
        _style_header(ws, HEADERS_ASSERT, [13, 22, 22, 12, 28, 30, 7, 18, 10, 14, 24, 14, 60, 26])

    wsx, wsr = wb[CROSSWALK], wb[ASSERTIONS]
    live_x, live_r = _existing_ids(wsx), _existing_ids(wsr)

    wsa = wb[STAGING_A]
    for r in range(2, wsa.max_row + 1):
        if wsa.cell(row=r, column=16).value != "Confirmed":
            continue
        rid = wsa.cell(row=r, column=1).value
        if rid in live_x or rid is None:
            continue
        t = wsx.max_row + 1
        src = [wsa.cell(row=r, column=c).value for c in range(1, 18)]
        eff, resc = src[7], src[8]
        status_formula = (
            f'=IF(H{t}="","Unknown (Coverage Gap)",'
            f'IF(AsOfDate<H{t},"Not Yet Effective",'
            f'IF(AND(I{t}<>"",AsOfDate>=I{t}),"Rescinded","Active")))'
        )
        out = [src[0], src[1], src[2], src[3], src[4], src[5], src[6], eff,
               resc, status_formula, src[11], src[12], src[13], src[16]]
        for col, v in enumerate(out, start=1):
            c = wsx.cell(row=t, column=col, value=v)
            c.font = green if col == 10 else body
            c.alignment = wrap
            if isinstance(v, dt.date):
                c.number_format = "mm/dd/yyyy"
        live_x.add(rid)
        counts["crosswalk"] += 1

    wsb = wb[STAGING_B]
    for r in range(2, wsb.max_row + 1):
        if wsb.cell(row=r, column=16).value != "Confirmed":
            continue
        rid = wsb.cell(row=r, column=1).value
        if rid in live_r or rid is None:
            continue
        t = wsr.max_row + 1
        src = [wsb.cell(row=r, column=c).value for c in range(1, 18)]
        var_formula = f'=IF(AND(ISNUMBER(F{t}),ISNUMBER(H{t})),H{t}-F{t},"")'
        out = [src[0], src[1], src[2], src[3], src[4], src[5], src[6], src[12],
               var_formula, src[14], src[7], src[8], src[9], src[16]]
        for col, v in enumerate(out, start=1):
            c = wsr.cell(row=t, column=col, value=v)
            c.font = body
            c.alignment = wrap
        live_r.add(rid)
        counts["assertions"] += 1

    wb.save(workbook_path)
    return counts
