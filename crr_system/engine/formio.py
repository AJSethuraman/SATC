"""save-form: append a completed line-sheet form to Database + Responses.

    python -m engine save-form --workbook WB --sheet LS_CI

Reads the form's header cells, ratio engine, and answer rows; upserts one
Database row (matched on Credit ID) and replaces that credit's Responses
rows. Recalculate (or open in Excel) after running so cached ratio values
are current before the next save.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from workbook.content import SEGMENTS              # noqa: E402
from workbook.database import write_credit_row, write_response_row  # noqa: E402
from workbook.forms import FORM_CELLS              # noqa: E402

_QID = re.compile(r"^[A-Z]+-Q\d+$")

# Which ratio IDs feed the Database key-ratio columns, per segment.
_DB_RATIOS = {
    "CI": {"leverage": "CI-R1", "dscr": "CI-R3", "ltv": None},
    "CRE": {"leverage": None, "dscr": "CRE-R1", "ltv": "CRE-R2"},
    "LL": {"leverage": "LL-R3", "dscr": None, "ltv": None},
    "ABL": {"leverage": None, "dscr": None, "ltv": None},
    "ARG": {"leverage": None, "dscr": None, "ltv": None},
    "COMP": {"leverage": None, "dscr": None, "ltv": None},
    "IA": {"leverage": None, "dscr": None, "ltv": None},
}


def save_form(workbook_path: str, sheet: str, status: str = "In Progress") -> dict:
    seg_code = next((c for c, s in SEGMENTS.items() if s["sheet"] == sheet), None)
    if seg_code is None:
        raise ValueError(f"{sheet} is not a known line-sheet (expected one of "
                         f"{[s['sheet'] for s in SEGMENTS.values()]})")

    wb = load_workbook(workbook_path)
    wb_vals = load_workbook(workbook_path, data_only=True)
    form, form_vals = wb[sheet], wb_vals[sheet]

    credit_id = form[FORM_CELLS["credit_id"]].value
    if not credit_id:
        raise ValueError(f"{sheet}!{FORM_CELLS['credit_id']} (Credit ID) is empty")

    ratios = {}
    answers = []
    qmeta = {}
    qsheet = wb["Questions"]
    for r in range(2, qsheet.max_row + 1):
        qmeta[qsheet.cell(row=r, column=1).value] = (
            qsheet.cell(row=r, column=3).value, qsheet.cell(row=r, column=4).value,
            qsheet.cell(row=r, column=5).value)

    for r in range(1, form.max_row + 1):
        rid = form.cell(row=r, column=1).value
        if not rid:
            continue
        if _QID.match(str(rid)):
            answers.append((rid, form.cell(row=r, column=4).value,
                            form.cell(row=r, column=5).value))
        elif str(rid).startswith(f"{seg_code}-R"):
            v = form_vals.cell(row=r, column=3).value
            ratios[str(rid)] = v if isinstance(v, (int, float)) else None

    rmap = _DB_RATIOS[seg_code]
    review_date = form[FORM_CELLS["review_date"]].value
    reviewer = form[FORM_CELLS["reviewer"]].value
    credit = {
        "credit_id": credit_id, "segment": seg_code,
        "segment_name": SEGMENTS[seg_code]["name"],
        "borrower": form[FORM_CELLS["borrower"]].value,
        "commitment": form[FORM_CELLS["commitment"]].value,
        "review_date": review_date, "reviewer": reviewer,
        "lob_grade": form[FORM_CELLS["lob_grade"]].value,
        "crr_grade": form[FORM_CELLS["crr_grade"]].value,
        "leverage": ratios.get(rmap["leverage"]),
        "dscr": ratios.get(rmap["dscr"]),
        "ltv": ratios.get(rmap["ltv"]),
        "status": status,
    }

    db = wb["Database"]
    target = None
    for r in range(2, db.max_row + 2):
        if db.cell(row=r, column=1).value in (credit_id, None):
            target = r
            break
    write_credit_row(db, target, credit)

    resp = wb["Responses"]
    rows = [(resp.cell(row=r, column=1).value, r) for r in range(2, resp.max_row + 1)]
    keep = [r for cid, r in rows if cid not in (credit_id, None)]
    # Rebuild the sheet body without this credit's old rows, then append new.
    existing = []
    for r in keep:
        existing.append({
            "credit_id": resp.cell(row=r, column=1).value,
            "segment": resp.cell(row=r, column=2).value,
            "qid": resp.cell(row=r, column=3).value,
            "section": resp.cell(row=r, column=4).value,
            "question": resp.cell(row=r, column=5).value,
            "severity": resp.cell(row=r, column=6).value,
            "answer": resp.cell(row=r, column=7).value,
            "note": resp.cell(row=r, column=8).value,
            "review_date": resp.cell(row=r, column=9).value,
            "reviewer": resp.cell(row=r, column=10).value,
        })
    new_rows = []
    for qid, answer, note in answers:
        section, question, severity = qmeta.get(qid, ("", "", ""))
        new_rows.append({
            "credit_id": credit_id, "segment": seg_code, "qid": qid,
            "section": section, "question": question, "severity": severity,
            "answer": answer, "note": note, "review_date": review_date,
            "reviewer": reviewer,
        })
    if resp.max_row > 1:
        resp.delete_rows(2, resp.max_row - 1)
    for i, rp in enumerate(existing + new_rows):
        write_response_row(resp, i + 2, rp)
    resp.auto_filter.ref = f"A1:N{len(existing) + len(new_rows) + 1}"

    wb.save(workbook_path)
    return {"credit_id": credit_id, "database_row": target,
            "responses_written": len(new_rows),
            "note": "Run scripts/recalc.py (or open in Excel) to refresh calculations."}
