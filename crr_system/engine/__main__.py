"""CLI for the CRR extraction engine.

  python -m engine extract-thresholds DOC --workbook WB [--citation C]
         [--agency OCC ...] [--effective YYYY-MM-DD] [--unverified]
         [--rescission-doc DOC2]
  python -m engine extract-cam DOC --workbook WB [--borrower NAME]
  python -m engine apply-rescission DOC --workbook WB
  python -m engine promote --workbook WB

Everything lands in staging; `promote` moves only reviewer-Confirmed rows
into the live Crosswalk / Assertions sheets.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys

from . import crosswalk as xw
from . import extract, ingest, staging


def _date(s: str) -> dt.date:
    return dt.date.fromisoformat(s)


def cmd_extract_thresholds(args) -> dict:
    doc = ingest.ingest(args.document)
    rows = extract.extract_thresholds(
        doc,
        citation=args.citation or "",
        agencies=args.agency or None,
        effective_date=args.effective,
        status_verified=not args.unverified,
    )
    if args.rescission_doc:
        notices = extract.extract_rescissions(ingest.ingest(args.rescission_doc))
        rows = xw.apply_rescissions(rows, notices)
    counts = staging.write_rows(args.workbook, rows)
    return {"rows_extracted": len(rows), "staged": counts, "warnings": doc.warnings}


def cmd_extract_cam(args) -> dict:
    doc = ingest.ingest(args.document)
    rows = extract.extract_assertions(doc, borrower_hint=args.borrower or "")
    counts = staging.write_rows(args.workbook, rows)
    return {"rows_extracted": len(rows), "staged": counts, "warnings": doc.warnings}


def cmd_apply_rescission(args) -> dict:
    """Stamp rescinded dates onto already-staged Type A rows in the workbook."""
    from openpyxl import load_workbook

    notices = extract.extract_rescissions(ingest.ingest(args.document))
    wb = load_workbook(args.workbook)
    ws = wb[staging.STAGING_A]
    stamped = 0
    for notice in notices:
        target_kw = xw._keywords(notice.target_keywords)
        for r in range(2, ws.max_row + 1):
            agency = ws.cell(row=r, column=6).value
            if agency not in notice.agencies:
                continue
            cite = f"{ws.cell(row=r, column=7).value or ''} {ws.cell(row=r, column=12).value or ''}"
            if len(target_kw & xw._keywords(cite)) < 2:
                continue
            if ws.cell(row=r, column=9).value is None:
                ws.cell(row=r, column=9, value=notice.rescinded_date).number_format = "mm/dd/yyyy"
                note_cell = ws.cell(row=r, column=15)
                note_cell.value = (
                    (f"{note_cell.value} " if note_cell.value else "")
                    + f"Rescinded for {agency} per {notice.anchor.document} ({notice.anchor.locator()})."
                )
                stamped += 1
    wb.save(args.workbook)
    return {"notices_found": len(notices), "rows_stamped": stamped}


def cmd_promote(args) -> dict:
    return staging.promote_confirmed(args.workbook)


def main(argv=None) -> int:
    p = argparse.ArgumentParser(prog="engine", description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("extract-thresholds", help="Extract Type A regulatory/policy thresholds")
    a.add_argument("document")
    a.add_argument("--workbook", required=True)
    a.add_argument("--citation", default="")
    a.add_argument("--agency", action="append")
    a.add_argument("--effective", type=_date, default=None)
    a.add_argument("--unverified", action="store_true",
                   help="Current status not verified: record rows as Coverage Gap")
    a.add_argument("--rescission-doc", default=None)
    a.set_defaults(func=cmd_extract_thresholds)

    b = sub.add_parser("extract-cam", help="Extract Type B credit-memo assertions")
    b.add_argument("document")
    b.add_argument("--workbook", required=True)
    b.add_argument("--borrower", default="")
    b.set_defaults(func=cmd_extract_cam)

    c = sub.add_parser("apply-rescission", help="Stamp rescissions onto staged Type A rows")
    c.add_argument("document")
    c.add_argument("--workbook", required=True)
    c.set_defaults(func=cmd_apply_rescission)

    d = sub.add_parser("promote", help="Move reviewer-Confirmed rows to live sheets")
    d.add_argument("--workbook", required=True)
    d.set_defaults(func=cmd_promote)

    e = sub.add_parser("save-form", help="Append a completed line-sheet form to the Database")
    e.add_argument("--workbook", required=True)
    e.add_argument("--sheet", required=True)
    e.add_argument("--status", default="In Progress")
    e.set_defaults(func=lambda a: __import__("engine.formio", fromlist=["save_form"])
                   .save_form(a.workbook, a.sheet, a.status))

    args = p.parse_args(argv)
    result = args.func(args)
    json.dump(result, sys.stdout, indent=2, default=str)
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
