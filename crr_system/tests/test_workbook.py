"""Validation tests against the built workbook artifact.

Run `python -m workbook.build` first; these tests verify the delivered
workbook's calculated values (from the LibreOffice recalc cache), the
confirmation gating, and the save-form round trip.
"""

import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

OUT = ROOT / "output" / "CRR_Line_Sheet_System.xlsx"

pytestmark = pytest.mark.skipif(not OUT.exists(), reason="build the workbook first")

ERRS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


@pytest.fixture(scope="module")
def vals():
    return load_workbook(OUT, data_only=True)


@pytest.fixture(scope="module")
def raw():
    return load_workbook(OUT)


def test_zero_formula_errors_in_cache(vals):
    bad = []
    for ws in vals.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERRS:
                    bad.append(f"{ws.title}!{c.coordinate}={c.value}")
    assert not bad, bad[:20]


def test_all_sheets_present(raw):
    expected = {"Home", "Settings", "Lists", "Questions", "LS_CI", "LS_CRE",
                "LS_LL", "LS_ABL", "LS_ARG", "LS_COMP", "Database", "Responses",
                "Dash_Portfolio", "Dash_Exceptions", "Obs_View",
                "Dash_Concentration", "Dash_Trends", "Crosswalk", "Assertions",
                "Staging_TypeA", "Staging_TypeB"}
    assert expected <= set(raw.sheetnames)


def test_crosswalk_agency_versioning(vals):
    """OCC/FDIC rescinded vs FRB active for the same metric, as of 2026-06-10."""
    ws = vals["Crosswalk"]
    status = {}
    for r in range(2, ws.max_row + 1):
        metric, agency = ws.cell(row=r, column=2).value, ws.cell(row=r, column=6).value
        if metric == "Senior Debt / EBITDA":
            status[agency] = ws.cell(row=r, column=10).value
    assert status == {"OCC": "Rescinded", "FRB": "Active", "FDIC": "Rescinded"}


def test_crosswalk_provenance_complete(raw):
    ws = raw["Crosswalk"]
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=13).value, f"row {r} missing verbatim span"
        assert ws.cell(row=r, column=12).value, f"row {r} missing page/section"


def test_confirmation_gating(raw):
    """Pending/Rejected staged rows must not exist in the live sheets."""
    live = set()
    for sheet in ("Crosswalk", "Assertions"):
        ws = raw[sheet]
        for r in range(2, ws.max_row + 1):
            live.add(ws.cell(row=r, column=1).value)
    leaked = []
    for sheet in ("Staging_TypeA", "Staging_TypeB"):
        ws = raw[sheet]
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(row=r, column=1).value
            conf = ws.cell(row=r, column=16).value
            if rid in live and conf != "Confirmed":
                leaked.append((sheet, rid, conf))
    assert not leaked


def test_pending_rows_remain_in_staging(raw):
    ws = raw["Staging_TypeA"]
    confs = [ws.cell(row=r, column=16).value for r in range(2, ws.max_row + 1)]
    assert "Pending" in confs, "expected some rows still awaiting confirmation"


def test_ci_ratio_engine_values(vals):
    ws = vals["LS_CI"]
    got = {}
    for r in range(1, 80):
        rid = ws.cell(row=r, column=1).value
        if rid and str(rid).startswith("CI-R"):
            got[rid] = (ws.cell(row=r, column=3).value, ws.cell(row=r, column=8).value)
    assert got["CI-R1"][0] == pytest.approx(65000 / 17100)
    assert got["CI-R1"][1] == "Pass"          # 3.8x vs in-force 6.0x (pre-rescission review date)
    assert got["CI-R2"][0] == pytest.approx(49590 / 17100)
    assert got["CI-R3"][1] == "Pass"          # DSCR vs internal-policy 1.25x minimum
    assert got["CI-R5"][1] == "n/a"           # no threshold defined


def test_dated_threshold_lookup(vals):
    """C&I review dated 2025-08-14 cites OCC 6.0x; the dashboard table dated
    2026-06-10 shows the OCC row Rescinded - both from the same crosswalk."""
    ci = vals["LS_CI"]
    th = None
    for r in range(1, 80):
        if ci.cell(row=r, column=1).value == "CI-R1":
            th = ci.cell(row=r, column=7).value
    assert th == 6
    dc = vals["Dash_Concentration"]
    statuses = {}
    for r in range(1, 40):
        a = dc.cell(row=r, column=2).value
        if a in ("OCC", "FDIC", "FRB"):
            statuses[a] = dc.cell(row=r, column=6).value
    assert statuses == {"OCC": "Rescinded", "FDIC": "Rescinded", "FRB": "Active"}


def test_database_counts_match_responses(vals):
    db, resp = vals["Database"], vals["Responses"]
    answers = {}
    for r in range(2, resp.max_row + 1):
        cid = resp.cell(row=r, column=1).value
        ans = resp.cell(row=r, column=7).value
        answers.setdefault(cid, []).append(ans)
    for r in range(2, db.max_row + 1):
        cid = db.cell(row=r, column=1).value
        expect_no = sum(1 for a in answers[cid] if a == "No")
        expect_obs = sum(1 for a in answers[cid] if a == "Obs")
        assert db.cell(row=r, column=16).value == expect_no
        assert db.cell(row=r, column=17).value == expect_obs
        # N/A excluded from applicable
        applicable = sum(1 for a in answers[cid] if a in ("Yes", "No", "Obs"))
        assert db.cell(row=r, column=18).value == applicable


def test_exception_dashboard_counts_only_no(vals):
    resp = vals["Responses"]
    total_no = sum(1 for r in range(2, resp.max_row + 1)
                   if resp.cell(row=r, column=7).value == "No")
    total_obs = sum(1 for r in range(2, resp.max_row + 1)
                    if resp.cell(row=r, column=7).value == "Obs")
    de = vals["Dash_Exceptions"]
    assert de["B5"].value == total_no
    ov = vals["Obs_View"]
    assert ov["B5"].value == total_obs
    assert total_obs not in (None, 0)


def test_obs_requires_note_check_formula(raw):
    ws = raw["LS_CI"]
    found = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=10).value
        if isinstance(v, str) and "Note required" in v:
            found = True
    assert found


def test_save_form_roundtrip(tmp_path):
    from engine.formio import save_form

    wb_path = tmp_path / "wb.xlsx"
    shutil.copy(OUT, wb_path)
    wb = load_workbook(wb_path)
    ws = wb["LS_CI"]
    ws["C5"] = "CI-2026099"
    ws["C6"] = "Roundtrip Test Borrower"
    wb.save(wb_path)

    result = save_form(str(wb_path), "LS_CI", status="Complete")
    assert result["credit_id"] == "CI-2026099"
    assert result["responses_written"] == 27

    wb = load_workbook(wb_path)
    db = wb["Database"]
    ids = [db.cell(row=r, column=1).value for r in range(2, db.max_row + 1)]
    assert "CI-2026099" in ids
    resp = wb["Responses"]
    cnt = sum(1 for r in range(2, resp.max_row + 1)
              if resp.cell(row=r, column=1).value == "CI-2026099")
    assert cnt == 27
