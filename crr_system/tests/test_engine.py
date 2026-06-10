"""Engine tests against the known-value fixtures."""

import datetime as dt
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from engine import crosswalk as xw
from engine import extract, ingest, staging
from engine.schema import ROW_TYPE_A, ROW_TYPE_B, STATUS_COVERAGE_GAP

FIX = ROOT / "fixtures"
LL_PDF = str(FIX / "fixture_2013_leveraged_lending.pdf")
RESC_PDF = str(FIX / "fixture_2025_rescission.pdf")
CRE_PDF = str(FIX / "fixture_2006_cre_guidance.pdf")
CAM_DOCX = str(FIX / "fixture_sample_cam.docx")


@pytest.fixture(scope="module")
def ll_rows():
    return extract.extract_thresholds(
        ingest.ingest(LL_PDF), citation="2013 Interagency Guidance on Leveraged Lending"
    )


def _by_metric(rows, metric):
    return [r for r in rows if r.metric == metric]


class TestTypeAExtraction:
    def test_pulls_6x_total_leverage(self, ll_rows):
        rows = _by_metric(ll_rows, "Total Debt / EBITDA")
        values = {r.proposed_value for r in rows}
        assert "6.0" in values and "4.0" in values

    def test_pulls_3x_senior(self, ll_rows):
        assert {r.proposed_value for r in _by_metric(ll_rows, "Senior Debt / EBITDA")} == {"3.0"}

    def test_pulls_repayment_capacity(self, ll_rows):
        rows = _by_metric(ll_rows, "Repayment Capacity (base-case de-lever)")
        assert rows and rows[0].proposed_value == "50"

    def test_agency_tagging_one_row_per_agency(self, ll_rows):
        senior = _by_metric(ll_rows, "Senior Debt / EBITDA")
        assert {r.agency for r in senior} == {"OCC", "FRB", "FDIC"}

    def test_effective_date_detected(self, ll_rows):
        assert ll_rows[0].effective_date == dt.date(2013, 3, 21)

    def test_provenance_attached(self, ll_rows):
        for r in ll_rows:
            assert r.source_span, f"{r.metric} missing span"
            assert r.anchor.page is not None
            assert r.anchor.document.endswith(".pdf")

    def test_span_is_verbatim(self, ll_rows):
        doc = ingest.ingest(LL_PDF)
        for r in ll_rows[:6]:
            page_text = doc.pages[r.anchor.page - 1].text
            assert r.source_span[:60] in page_text

    def test_cre_fixture_values(self):
        rows = extract.extract_thresholds(
            ingest.ingest(CRE_PDF), citation="2006 Interagency CRE Concentration Guidance"
        )
        cld = _by_metric(rows, "CLD Concentration / Total Risk-Based Capital")
        total = _by_metric(rows, "Total CRE / Total Risk-Based Capital")
        assert "100" in {r.proposed_value for r in cld}
        assert "300" in {r.proposed_value for r in total}

    def test_unverified_status_records_coverage_gap(self):
        rows = extract.extract_thresholds(ingest.ingest(LL_PDF), status_verified=False)
        assert all(r.status == STATUS_COVERAGE_GAP for r in rows)


class TestVersioning:
    @pytest.fixture(scope="class")
    def stamped(self, ll_rows):
        notices = extract.extract_rescissions(ingest.ingest(RESC_PDF))
        assert notices, "rescission notice not detected"
        return xw.apply_rescissions([r for r in ll_rows], notices)

    def test_rescission_detects_agencies_and_date(self):
        notices = extract.extract_rescissions(ingest.ingest(RESC_PDF))
        n = notices[0]
        assert set(n.agencies) >= {"OCC", "FDIC"}
        assert "FRB" not in n.agencies
        assert n.rescinded_date == dt.date(2025, 12, 16)

    def test_occ_rescinded_frb_active_after_dec_2025(self, stamped):
        as_of = dt.date(2026, 1, 15)
        senior = [r for r in stamped if r.metric == "Senior Debt / EBITDA"]
        status = {r.agency: xw.status_as_of(r, as_of) for r in senior}
        assert status["OCC"] == xw.STATUS_RESCINDED
        assert status["FDIC"] == xw.STATUS_RESCINDED
        assert status["FRB"] == xw.STATUS_ACTIVE

    def test_all_active_before_rescission(self, stamped):
        as_of = dt.date(2025, 6, 30)
        senior = [r for r in stamped if r.metric == "Senior Debt / EBITDA"]
        assert all(xw.status_as_of(r, as_of) == xw.STATUS_ACTIVE for r in senior)

    def test_applicable_thresholds_filters_by_agency_and_date(self, stamped):
        post = xw.applicable_thresholds(stamped, as_of=dt.date(2026, 1, 15), metric="Senior Debt / EBITDA")
        assert {r.agency for r in post} == {"FRB"}


class TestTypeBExtraction:
    @pytest.fixture(scope="class")
    def cam_rows(self):
        return extract.extract_assertions(ingest.ingest(CAM_DOCX))

    def _value(self, rows, metric):
        matches = [r for r in rows if r.metric == metric]
        assert matches, f"missing {metric}"
        return matches[0].proposed_value

    def test_header_fields(self, cam_rows):
        assert "Meridian Fabrication" in self._value(cam_rows, "Borrower")
        assert "revolving" in self._value(cam_rows, "Facility")
        assert self._value(cam_rows, "Assigned Risk Grade").startswith("4")
        assert "Whitfield" in self._value(cam_rows, "Guarantor")
        assert "lien" in self._value(cam_rows, "Collateral").lower()

    def test_asserted_ratios(self, cam_rows):
        assert self._value(cam_rows, "Total Debt / EBITDA (asserted)") == "3.8"
        assert self._value(cam_rows, "Senior Debt / EBITDA (asserted)") == "2.9"
        assert self._value(cam_rows, "DSCR (asserted)") == "1.42"
        assert self._value(cam_rows, "Global DSCR (asserted)") == "1.55"

    def test_narratives_and_covenants(self, cam_rows):
        metrics = {r.metric for r in cam_rows}
        assert "Primary Repayment Source" in metrics
        assert "Secondary Repayment Source" in metrics
        assert "Key Assumption / Projection" in metrics
        covenants = [r for r in cam_rows if r.category == "covenant"]
        assert any("4.5x" in r.proposed_value for r in covenants)

    def test_borrower_attached_and_independent_blank(self, cam_rows):
        ratios = [r for r in cam_rows if r.category == "ratio"]
        assert all("Meridian" in r.borrower for r in ratios)
        assert all(r.independent_value == "" for r in cam_rows)
        assert all(r.confirmation == "Pending" for r in cam_rows)

    def test_provenance(self, cam_rows):
        for r in cam_rows:
            assert r.source_span and r.anchor.document.endswith(".docx")


class TestStagingWorkflow:
    def test_stage_and_promote_gating(self, tmp_path, ll_rows):
        wb_path = str(tmp_path / "staging.xlsx")
        cam_rows = extract.extract_assertions(ingest.ingest(CAM_DOCX))
        counts = staging.write_rows(wb_path, list(ll_rows) + cam_rows)
        assert counts["A"] > 0 and counts["B"] > 0

        # dedupe on second write
        counts2 = staging.write_rows(wb_path, list(ll_rows))
        assert counts2["A"] == 0 and counts2["skipped"] > 0

        # nothing promotes while Pending
        promoted = staging.promote_confirmed(wb_path)
        assert promoted == {"crosswalk": 0, "assertions": 0}

        # confirm two rows, promote, verify only those went live
        from openpyxl import load_workbook

        wb = load_workbook(wb_path)
        wb[staging.STAGING_A].cell(row=2, column=16).value = "Confirmed"
        wb[staging.STAGING_B].cell(row=2, column=16).value = "Confirmed"
        wb[staging.STAGING_B].cell(row=3, column=16).value = "Rejected"
        wb.save(wb_path)
        promoted = staging.promote_confirmed(wb_path)
        assert promoted == {"crosswalk": 1, "assertions": 1}

        wb = load_workbook(wb_path)
        assert wb[staging.CROSSWALK].max_row == 2
        assert wb[staging.ASSERTIONS].max_row == 2
        # provenance traveled with the promoted row
        assert wb[staging.CROSSWALK].cell(row=2, column=13).value  # verbatim span
