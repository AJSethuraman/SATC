#!/usr/bin/env python3
"""Build the Bank Counterparty & Peer Monitor workbook (FDIC BankFind) --
KeyBank-styled.

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled .xlsm.
All visual style comes from keybank_style.py (tokens + helpers) -- never a
hardcoded hex here. Built to BUILD_SPEC_FDIC.md (Sections 2-5).

THE INVERSION: banks are ROWS. Every dashboard/Watchlist row is anchored to a
PEER SLOT, not to a bank: bank identity (name/CERT/group) flows in BY FORMULA
from the _config [PEERS] cells and metric values BY FORMULA from the slot's
fixed Raw_FDIC block -- so editing WHICH bank occupies a slot re-labels and
re-points everything on recalc, with zero rebuilt formulas (the flexible-peers
USER REQUIREMENT). Capacity is a build knob: --peer-slots (default 40).

Structure (Section 4 + SPEC_COMPETITOR_PACK.md v1.1):
  Dashboard_AssetQuality         -- noncurrent, NCO(q), 30-89 PD, ALLL/loans,
                                    coverage, Texas; peer-median row
  Dashboard_Capital_Earnings     -- T1 leverage, total RBC (CBLR-blank
                                    tolerated), equity/assets, ROAQ, NIM,
                                    efficiency; peer-median row
  Dashboard_Funding_Concentration-- loans/deposits, brokered %, CRE
                                    concentration (proxy noted) + the SVB
                                    pack: uninsured %, unrealized/capital,
                                    FHLB/assets; peer-median
  Dashboard_LoanBook             -- v1.1 two-track: consumer DQ/NCO band +
                                    commercial Call-Report-floor band (honesty
                                    subtitle), per-class PD30-89/PD90+/NA/NCOq
  Watchlist       -- the ranked peer list: Bank | CERT | Group | Texas |
                     ALERT/WATCH flag counts | Rank | Status (+ per-metric
                     helper status columns feeding the COUNTIFs)
  Raw_FDIC        -- one block per slot: 16 quarters x 68 fields, runner fills
  _config         -- [SETTINGS]/[THRESHOLDS]/[PEERS]/[SERIES] (source of truth)
  _provenance     -- the tie-out map (contract sec 12): field/metric ->
                     schedule/line/MDRM + facsimile URL pattern; --tieout
                     reads it
  _code_py/_code_vba -- runner.py + macro.bas as flat ASCII text
  _readme         -- metric authorities, regulatory vintages, honest caveats

No native charts (L4). Embedded code is flat ASCII (L3). Every raw reference
is blank-guarded (empty cell renders blank, never 0). Heat direction agrees
with each metric's threshold direction (red = stress; below-is-bad metrics
get INVERTED heat -- the bureau lesson, extended per-metric).
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import runner as R
import series_seed as SEED
import provenance_seed as PROV

import keybank_style as KB
from keybank_style import (
    INK, KEY_RED, MONO_FONT, NOTE_FONT, SECT_FILL,
    brand_banner, kpi_tiles, header_row, watchlist_boundary,
    hide_gridlines, freeze_below,
)

HERE = os.path.dirname(os.path.abspath(__file__))
INK_BOLD = Font(name="Arial", bold=True, color=INK)
SMALL_BOLD = Font(name="Calibri", bold=True, size=9)


def stress_heat(ws, cell_range, direction="above"):
    """Direction-AWARE heat: red = stress always. direction='above' (high is
    bad: Texas, NCO, brokered %) shades low green -> high red; 'below' (low
    is bad: capital, coverage, ROA, NIM) is INVERTED -- low red -> high green.
    Matching the threshold engine's direction is the carried bureau lesson."""
    lo, hi = (KB.HEAT_GOOD, KB.HEAT_BAD) if direction != "below" \
        else (KB.HEAT_BAD, KB.HEAT_GOOD)
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="min", start_color=lo,
        mid_type="percentile", mid_value=50, mid_color=KB.HEAT_MID,
        end_type="max", end_color=hi))


def band_rules(ws, cell_range, first_cell, watch_ref, alert_ref, direction):
    """Per-cell WATCH/ALERT coloring straight off the _config [THRESHOLDS]
    cells (config-driven, direction-aware, blank-guarded via ISNUMBER)."""
    op = ">=" if direction != "below" else "<="
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{alert_ref})"],
        fill=KB.ALERT_FILL, font=KB.ALERT_FONT, stopIfTrue=True))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{watch_ref})"],
        fill=KB.TIGHTEN_FILL,
        font=Font(name="Calibri", bold=True, color=KB.SLATE),
        stopIfTrue=True))


def text_cell(ws, row, col, value):
    """Write literal text; openpyxl treats a leading '=' as a formula, so force
    it back to a string for source/doc lines."""
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# Lane layout (Section 4): which metric renders in which dashboard column.
# ==========================================================================
LANES = [
    ("Dashboard_AssetQuality",
     "Asset Quality Dashboard",
     "Noncurrent, quarterly NCO (seasonal: Q4 cleanup -- compare YoY, F7), "
     "30-89 day pipeline, allowance, coverage, Texas ratio (VARIANT). "
     "QBP Q1'26 industry reference: NCO 0.59%.",
     ["NCLNLSR", "NTLNLSQR", "PD3089R", "LNATRESR", "LNRESNCR", "TEXAS"]),
    ("Dashboard_Capital_Earnings",
     "Capital & Earnings Dashboard",
     "Tier 1 leverage (universal), total RBC (blank for CBLR electors -- "
     "null is not zero), equity/assets, quarterly ROA, NIM, efficiency. "
     "QBP Q1'26 industry reference: ROA 1.26%, NIM 3.31%.",
     ["RBC1AAJ", "RBCRWAJ", "EQV", "ROAQ", "NIMY", "EEFFR"]),
    ("Dashboard_Funding_Concentration",
     "Funding & Concentration Dashboard",
     "Loans/deposits, brokered-deposit share (FDI Act s29 salience), CRE "
     "concentration vs the 2006 interagency 300% screen (PROXY denominator "
     "EQ+LNATRES) -- plus the v1.1 SVB pack: uninsured-deposit share "
     "(DEPUNINS null renders BLANK + digest note, never 0), COMPUTED "
     "HTM+AFS unrealized loss / (EQ+LNATRES) (EQCCOMPI is an OCI FLOW -- "
     "not used), FHLB advances/assets. SVB '23: ~94% uninsured, unrealized "
     ">100% of capital.",
     ["LNDEPR", "BRODEPR", "CRECONR", "UNINSDEPR", "UNRLZCAPR", "FHLBASSR"]),
]

# --------------------------------------------------------------------------
# Dashboard_LoanBook (v1.1, SPEC_COMPETITOR_PACK.md sec 1): ONE tab, TWO
# slot-anchored bands. Band 1 = the CONSUMER track (retail classification is
# DPD-formula-driven, so DQ IS the classification story); band 2 = the
# COMMERCIAL floor (public Call-Report proxy -- the honesty subtitle below).
# --------------------------------------------------------------------------
LB_TAB = "Dashboard_LoanBook"
LB_CONSUMER = [
    "P3CRCDR", "P9CRCDR", "NACRCDR", "NTCRCDQR",
    "P3AUTOR", "P9AUTOR", "NAAUTOR", "NTAUTOQR",
    "P3CONOTHR", "P9CONOTHR", "NACONOTHR", "NTCONOTQR",
    "P3RERESR", "P9RERESR", "NARERESR", "NTRERESQR",
    "P3RELOCR", "P9RELOCR", "NARELOCR",
]
LB_COMMERCIAL = [
    "P3RECONSR", "P9RECONSR", "NARECONSR", "NTRECONQR",
    "P3RENRESR", "P9RENRESR", "NARENRESR", "NTRENREQR",
    "P3REMULTR", "P9REMULTR", "NAREMULTR", "NTREMULQR",
    "P3CIR", "P9CIR", "NACIR", "NTCIQR",
]
LB_CONSUMER_CAPTION = (
    "CONSUMER TRACK -- the DQ/NCO story IS the classification story: retail "
    "classification is DPD-formula-driven under the Uniform Retail Credit "
    "Classification policy. Card DQ/NCO runs structurally high vs other "
    "classes (QBP context) -- its bands are set accordingly. NCO rates are "
    "QUARTERLY flows annualized x4 (the FDIC's NTLNLSQR convention).")
# The spec sec 1 honesty subtitle, verbatim (ASCII dashes):
LB_COMMERCIAL_CAPTION = (
    "COMMERCIAL FLOOR -- Public Call-Report proxy -- commercial risk "
    "ratings lead delinquency; criticized/classified view via the EDGAR "
    "tracker.")

METRIC_LABEL = {
    "NCLNLSR": "Noncurrent %", "NTLNLSQR": "NCO % (q)", "PD3089R": "30-89 PD %",
    "LNATRESR": "ALLL/Loans %", "LNRESNCR": "Coverage %", "TEXAS": "Texas %",
    "RBC1AAJ": "T1 Leverage", "RBCRWAJ": "Total RBC", "EQV": "Equity/Assets",
    "ROAQ": "ROA (q)", "NIMY": "NIM", "EEFFR": "Efficiency",
    "LNDEPR": "Loans/Dep %", "BRODEPR": "Brokered %", "CRECONR": "CRE Conc %",
    # v1.1 SVB pack
    "UNINSDEPR": "Uninsured Dep %", "UNRLZCAPR": "Unrlz Loss/Cap %",
    "FHLBASSR": "FHLB/Assets %",
    # v1.1 loan-book pack (class prefix + rate kind)
    "P3CRCDR": "Card 30-89", "P9CRCDR": "Card 90+", "NACRCDR": "Card NA",
    "NTCRCDQR": "Card NCOq",
    "P3AUTOR": "Auto 30-89", "P9AUTOR": "Auto 90+", "NAAUTOR": "Auto NA",
    "NTAUTOQR": "Auto NCOq",
    "P3CONOTHR": "OthCons 30-89", "P9CONOTHR": "OthCons 90+",
    "NACONOTHR": "OthCons NA", "NTCONOTQR": "OthCons NCOq",
    "P3RERESR": "Resi 30-89", "P9RERESR": "Resi 90+", "NARERESR": "Resi NA",
    "NTRERESQR": "Resi NCOq",
    "P3RELOCR": "HELOC 30-89", "P9RELOCR": "HELOC 90+",
    "NARELOCR": "HELOC NA",
    "P3RECONSR": "Constr 30-89", "P9RECONSR": "Constr 90+",
    "NARECONSR": "Constr NA", "NTRECONQR": "Constr NCOq",
    "P3RENRESR": "CRE 30-89", "P9RENRESR": "CRE 90+", "NARENRESR": "CRE NA",
    "NTRENREQR": "CRE NCOq",
    "P3REMULTR": "Multif 30-89", "P9REMULTR": "Multif 90+",
    "NAREMULTR": "Multif NA", "NTREMULQR": "Multif NCOq",
    "P3CIR": "C&I 30-89", "P9CIR": "C&I 90+", "NACIR": "C&I NA",
    "NTCIQR": "C&I NCOq",
}

DASH_HDR = 7                    # header band row; slot rows start at 8
DASH_FIRST = DASH_HDR + 1
WL_HDR = 4                      # Watchlist header row; slot rows start at 5
WL_FIRST = WL_HDR + 1
WL_HELPER_COL0 = 10             # J.. = the per-metric status helpers
# rank-score column sits one past the last helper (53 metrics in v1.1)
WL_SCORE_COL = WL_HELPER_COL0 + len(SEED.METRIC_ROWS) + 1


def dash_row(slot):
    return DASH_FIRST + slot - 1


def wl_row(slot):
    return WL_FIRST + slot - 1


def lb_comm_hdr(peer_slots):
    """LoanBook band 2 (commercial floor) header row: below band 1's slot
    rows + median + a blank + the honesty-caption row."""
    return DASH_HDR + peer_slots + 4


def metric_home(metric_id, peer_slots=R.PEER_SLOTS_DEFAULT):
    """(tab, column_index, first_slot_row) where a metric's value cell
    lives; banks-as-rows means the (slot, metric) value exists exactly once.
    Row for a slot = first_slot_row + slot - 1 (the LoanBook commercial band
    starts below the consumer band, so first_slot_row is per-metric)."""
    for tab, _t, _s, mids in LANES:
        if metric_id in mids:
            return tab, 4 + mids.index(metric_id), DASH_FIRST
    if metric_id in LB_CONSUMER:
        return LB_TAB, 4 + LB_CONSUMER.index(metric_id), DASH_FIRST
    if metric_id in LB_COMMERCIAL:
        return (LB_TAB, 4 + LB_COMMERCIAL.index(metric_id),
                lb_comm_hdr(peer_slots) + 1)
    raise KeyError(metric_id)


# ==========================================================================
# _config  (the knob panel: SETTINGS / THRESHOLDS / PEERS / SERIES)
# ==========================================================================
def config_rows(peer_slots=R.PEER_SLOTS_DEFAULT):
    rows = [["Bank Counterparty & Peer Monitor (FDIC BankFind) -- CONFIG "
             "(the knob panel). Edit values here; no code change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["pack_version", R.PACK_VERSION, "metric-pack version BUILT into "
         "this workbook (v1.1 = the two-track loan-book + SVB competitor "
         "pack, _provenance tab and --tieout mode). The raw layout and every "
         "formula are anchored to this pack; the runner refuses a mismatched "
         "field layout -- rebuild, never edit in place."],
        ["demo_mode", "FALSE", "TRUE = offline deterministic FdicDemoProvider "
         "(no network), for trying the button and tests. Live mode is KEYLESS "
         "-- the FDIC API needs no key either."],
        ["raw_slots", 16, "quarters kept per bank (newest-first). BUILT INTO "
         "every formula -- changing it requires rebuilding the workbook "
         "(make_workbook.py); the runner refuses a mismatched layout."],
        ["peer_slots", peer_slots, "the BUILT peer capacity (informational; "
         "make_workbook.py --peer-slots N sets it). More banks than this in "
         "[PEERS] = the runner refuses with the rebuild command -- never "
         "truncation."],
        ["http_min_interval", 0.6, "seconds between live FDIC requests (rate "
         "limit undocumented -- stay polite)."],
        ["fdic_max_retries", 4, "retries with backoff on HTTP 429/5xx."],
        ["stale_multiplier", 2.0, "a bank is STALE when its latest REPDTE "
         "lags the peer-set max by more than this many quarters (merged-away "
         "banks keep returning old data with no error -- trap F2); stale "
         "banks are excluded from digest alert counts. NOTE: the Excel "
         "median rows still include them in v1 -- see _readme."],
        ["secret_env", "", "NAME of the env var holding a licensed (Class C) "
         "secret. Empty in v1 -- the FDIC API is keyless, and the gates "
         "admit ONLY Class A; a licensed swap requires its own review."],
    ]
    rows += [[], ["[THRESHOLDS]"],
             ["id", "watch", "alert", "direction", "authority"]]
    for tid, watch, alert, direction, authority in SEED.THRESHOLDS:
        # watch/alert are NUMERIC cells (the macro text-"0.5" lesson: Excel's
        # number>=text comparison is silently FALSE).
        rows.append([tid, watch, alert, direction, authority])
    rows += [[], ["[PEERS]"], list(SEED.PEER_HEADER)]
    for prow in SEED.peer_rows(peer_slots):
        rows.append(prow)
    rows += [["# One bank per row: slot | cert | name | group | active. Add "
              "a bank = fill a free slot row + re-run the runner (no "
              "rebuild). Remove = active FALSE or clear the row (the slot "
              "blanks on the next run)."],
             ["# CERT provenance: 628, 3511 and 639 are verified against the "
              "FDIC's own spec/captured response; ALL OTHER seeded CERTs are "
              "illustrative -- verify each with: python runner.py --lookup "
              "\"<bank name>\" before a live run."]]
    rows += [[], ["[SERIES]"], list(SEED.HEADER)]
    for r in SEED.all_series():
        rows.append([r[h] for h in SEED.HEADER])
    rows += [[], ["# [SERIES] is the bank-agnostic METRIC dictionary "
                  "(geo_segment=entity, expanded at run time to cert:NNN per "
                  "active peer). The gates admit only Class A CERT-keyed "
                  "rows; aggregates/name-only rows are refused by name."],
             ["# v1.1 pack IN (fields verified, RESEARCH_COMPETITOR_PACK.md "
              "93/93): per-class consumer DQ/NCO track + commercial floor "
              "(Dashboard_LoanBook), uninsured-deposit share, COMPUTED "
              "HTM+AFS unrealized loss / capital cushion, FHLB reliance."],
             ["# STILL OUT (never silently approximated): OREO in the Texas "
              "numerator, intangibles netting, AOCI as a stock (EQCCOMPI is "
              "a FLOW), category loan yields + UBPR percentiles (FFIEC "
              "bulk-ZIP job only), criticized/classified (EDGAR template)."]]
    return rows


def write_config(wb, peer_slots):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows(peer_slots)
    thr_cells = {}                 # id -> (watch_ref, alert_ref, direction)
    set_cells = {}                 # settings key -> value-cell ref
    peer_cells = {}                # slot -> dict(cert/name/group/active refs)
    in_thr = in_set = in_peers = False
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        head = row[0] if row else None
        if head == "[THRESHOLDS]":
            in_thr, in_set, in_peers = True, False, False
            continue
        if head == "[SETTINGS]":
            in_set, in_thr, in_peers = True, False, False
            continue
        if head == "[PEERS]":
            in_peers, in_thr, in_set = True, False, False
            continue
        if head is not None and isinstance(head, str) and head.startswith("["):
            in_thr = in_set = in_peers = False
        if not row or head is None or (isinstance(head, str)
                                       and str(head).startswith("#")):
            continue
        if in_set and head not in ("key",):
            set_cells[head] = f"_config!$B${i}"
        if in_thr and head not in ("id",):
            thr_cells[head] = (f"_config!$B${i}", f"_config!$C${i}",
                               str(row[3]).strip().lower())
        if in_peers and str(head).strip().lower() != "slot":
            slot = int(head)
            peer_cells[slot] = {
                "cert": f"_config!$B${i}", "name": f"_config!$C${i}",
                "group": f"_config!$D${i}", "active": f"_config!$E${i}"}
    ws["A1"].font = INK_BOLD
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("[") \
                and row[0].endswith("]"):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    return thr_cells, set_cells, peer_cells


# ==========================================================================
# Raw_FDIC scaffold  (one block per SLOT; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, peer_slots, raw_slots):
    ws = wb.create_sheet(R.RAW_TAB)
    hide_gridlines(ws)
    ws["A1"] = (f"{R.RAW_TAB} -- raw FDIC /financials fields per peer slot "
                f"(or FdicDemoProvider output in --demo), newest quarter "
                f"first. One block per SLOT: anchors never move when [PEERS] "
                f"changes. Dollar fields are $ thousands, ratio fields are "
                f"percent (trap F4). Written by runner.py -- do not edit.")
    ws["A1"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 12
    for j in range(len(R.RAW_FIELDS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12
    for slot in range(1, peer_slots + 1):
        b = R.slot_block(slot, raw_slots)
        ws.cell(b.header_row, 1, R.slot_label(slot)).font = INK_BOLD
        # cells B/C/D (unit key, bank name, group) are RUNTIME state the
        # runner writes/blanks per run -- empty at build.
        ws.cell(b.label_row, 1, "REPDTE").font = SMALL_BOLD
        for fname in R.RAW_FIELDS:
            c = ws.cell(b.label_row, R.field_col(fname), fname)
            c.font = SMALL_BOLD
    return ws


# ==========================================================================
# Formula builders  (the Excel side of the metric registry -- trap F5:
# definitions MUST equal runner.METRICS; parity is machine-tested)
# ==========================================================================
def _fref(slot, fname, raw_slots, offset=0):
    b = R.slot_block(slot, raw_slots)
    col = get_column_letter(R.field_col(fname))
    return f"{R.RAW_TAB}!{col}{b.first_data_row + offset}"


def _direct(ref):
    """Blank-guard every raw reference: an empty cell renders blank, never 0
    (IFERROR alone would fabricate 0.00 values and pre-run OK statuses)."""
    return f"=IF({ref}=\"\",\"\",{ref})"


def _ratio_formula(num_ref, den_ref, mult=100):
    """num/den*mult with blank guards AND a zero-denominator guard -- exactly
    runner._ratio / runner._pack_ratio_fn (None on missing input or zero
    denominator). mult=400 is the annualized-quarterly-flow convention."""
    return (f"=IF(OR({num_ref}=\"\",{den_ref}=\"\"),\"\","
            f"IF({den_ref}=0,\"\",{num_ref}/{den_ref}*{int(mult)}))")


def metric_formula(metric_id, slot, raw_slots):
    """The (slot, metric) headline cell: latest quarter (offset 0) off the
    slot's raw block. One branch per registered metric -- no silent
    fallback."""
    f = lambda name: _fref(slot, name, raw_slots)          # noqa: E731
    if metric_id == "PD3089R":
        return _ratio_formula(f("P3LNLS"), f("LNLSGR"))
    if metric_id == "TEXAS":
        n, e, a = f("NCLNLS"), f("EQ"), f("LNATRES")
        return (f"=IF(OR({n}=\"\",{e}=\"\",{a}=\"\"),\"\","
                f"IF({e}+{a}=0,\"\",{n}/({e}+{a})*100))")
    if metric_id == "LNDEPR":
        return _ratio_formula(f("LNLSNET"), f("DEP"))
    if metric_id == "BRODEPR":
        return _ratio_formula(f("BRO"), f("DEP"))
    if metric_id == "CRECONR":
        c, r2, m = f("LNRECONS"), f("LNRENRES"), f("LNREMULT")
        e, a = f("EQ"), f("LNATRES")
        return (f"=IF(OR({c}=\"\",{r2}=\"\",{m}=\"\",{e}=\"\",{a}=\"\"),\"\","
                f"IF({e}+{a}=0,\"\",({c}+{r2}+{m})/({e}+{a})*100))")
    if metric_id == "UNRLZCAPR":
        # ((SCHA-SCHF)+(SCAA-SCAF))/(EQ+LNATRES)*100 -- runner.d_unrlzcapr
        ha, hf = f("SCHA"), f("SCHF")
        aa, af = f("SCAA"), f("SCAF")
        e, a = f("EQ"), f("LNATRES")
        return (f"=IF(OR({ha}=\"\",{hf}=\"\",{aa}=\"\",{af}=\"\","
                f"{e}=\"\",{a}=\"\"),\"\",IF({e}+{a}=0,\"\","
                f"(({ha}-{hf})+({aa}-{af}))/({e}+{a})*100))")
    if metric_id in R.PACK_RATIOS:
        # the declarative pack table drives BOTH sides (trap F5): identical
        # num/den*mult here and in runner._pack_ratio_fn
        num, den, mult = R.PACK_RATIOS[metric_id]
        return _ratio_formula(f(num), f(den), mult)
    # direct API ratio: the metric id IS the field
    flds, fn = R.METRICS[metric_id]
    assert fn is None, f"unhandled derived metric {metric_id}"
    return _direct(f(flds[0]))


# ==========================================================================
# Dashboards  (banks as rows; per-cell threshold coloring; median row)
# ==========================================================================
def write_dashboard(wb, tab, title, subtitle, metric_ids, peer_slots,
                    raw_slots, thr_cells, peer_cells, n_active):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ncols = 3 + len(metric_ids)
    cols = ["Bank", "CERT", "Group"] + [METRIC_LABEL[m] for m in metric_ids]

    first, last = DASH_FIRST, DASH_FIRST + peer_slots - 1
    for slot in range(1, peer_slots + 1):
        r = dash_row(slot)
        pc = peer_cells[slot]
        # bank identity BY FORMULA from [PEERS]: a renamed/swapped slot
        # re-labels everywhere on recalc (flexible-peers requirement)
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cert']}=\"\",\"\",{pc['cert']})").font = MONO_FONT
        ws.cell(r, 3, f"=IF({pc['group']}=\"\",\"\",{pc['group']})").font = KB.SECONDARY
        for k, mid in enumerate(metric_ids):
            c = ws.cell(r, 4 + k, metric_formula(mid, slot, raw_slots))
            c.number_format = "0.00"
            c.alignment = KB.RIGHT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 2).alignment = KB.RIGHT
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    # PEER MEDIAN row: Excel MEDIAN ignores blanks/text. v1 caveat (spec test
    # 6): a STALE bank's landed values still enter these medians -- the
    # runner's digest medians exclude them; divergence documented in _readme.
    med = last + 1
    ws.cell(med, 1, "PEER MEDIAN (landed banks; includes stale -- see "
                    "_readme)").font = INK_BOLD
    for k, mid in enumerate(metric_ids):
        col = get_column_letter(4 + k)
        rng = f"{col}{first}:{col}{last}"
        c = ws.cell(med, 4 + k,
                    f"=IF(COUNT({rng})=0,\"\",MEDIAN({rng}))")
        c.number_format = "0.00"
        c.alignment = KB.RIGHT
        c.font = INK_BOLD
    ws.cell(med, 1).border = KB.HAIRLINE

    header_row(ws, DASH_HDR, cols, right_from=3)
    freeze_below(ws, DASH_HDR)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 13
    for k in range(len(metric_ids)):
        ws.column_dimensions[get_column_letter(4 + k)].width = 12

    brand_banner(ws, 1, ncols, title, subtitle)
    wl_last = WL_FIRST + peer_slots - 1
    tiles = [
        ("BANKS IN ALERT",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"ALERT\")",
         f"of {n_active} active"),
        ("BANKS IN WATCH",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"WATCH\")",
         "elevated"),
        ("BANKS TRACKED",
         f"=COUNTIF(Watchlist!$A${WL_FIRST}:$A${wl_last},\"?*\")",
         f"of {peer_slots} built slots"),
    ]
    kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8

    # run-status readout in column L (free of the masthead merges)
    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 40
    for rr in (1, 2, 3):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(4, R.STATUS_COL).font = NOTE_FONT          # macro writes here
    ws.cell(4, R.STATUS_COL).alignment = KB.LEFT

    # per-metric-column coloring: WATCH/ALERT band rules off the _config
    # threshold cells + direction-aware heat (red = stress, both directions)
    for k, mid in enumerate(metric_ids):
        col = get_column_letter(4 + k)
        rng = f"{col}{first}:{col}{last}"
        watch_ref, alert_ref, direction = thr_cells[mid]
        band_rules(ws, rng, f"{col}{first}", watch_ref, alert_ref, direction)
        stress_heat(ws, rng, direction)
    return ws


# ==========================================================================
# Dashboard_LoanBook  (v1.1: two slot-anchored bands on one tab)
# ==========================================================================
def _slot_band(ws, hdr, metric_ids, peer_slots, raw_slots, thr_cells,
               peer_cells):
    """One banks-as-rows band: header row + slot rows (identity by formula
    from [PEERS], values by formula from the slot's raw block) + PEER MEDIAN
    row + per-column threshold band rules and direction-aware heat. Same
    body as write_dashboard's -- shared so both LoanBook bands and future
    lanes stay identical in behavior."""
    ncols = 3 + len(metric_ids)
    cols = ["Bank", "CERT", "Group"] + [METRIC_LABEL[m] for m in metric_ids]
    first, last = hdr + 1, hdr + peer_slots
    for slot in range(1, peer_slots + 1):
        r = first + slot - 1
        pc = peer_cells[slot]
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cert']}=\"\",\"\",{pc['cert']})").font = MONO_FONT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 2).alignment = KB.RIGHT
        ws.cell(r, 3, f"=IF({pc['group']}=\"\",\"\",{pc['group']})").font = KB.SECONDARY
        for k, mid in enumerate(metric_ids):
            c = ws.cell(r, 4 + k, metric_formula(mid, slot, raw_slots))
            c.number_format = "0.00"
            c.alignment = KB.RIGHT
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE
    med = last + 1
    ws.cell(med, 1, "PEER MEDIAN (landed banks; includes stale -- see "
                    "_readme)").font = INK_BOLD
    for k, mid in enumerate(metric_ids):
        col = get_column_letter(4 + k)
        rng = f"{col}{first}:{col}{last}"
        c = ws.cell(med, 4 + k, f"=IF(COUNT({rng})=0,\"\",MEDIAN({rng}))")
        c.number_format = "0.00"
        c.alignment = KB.RIGHT
        c.font = INK_BOLD
    ws.cell(med, 1).border = KB.HAIRLINE
    header_row(ws, hdr, cols, right_from=3)
    for k, mid in enumerate(metric_ids):
        col = get_column_letter(4 + k)
        rng = f"{col}{first}:{col}{last}"
        watch_ref, alert_ref, direction = thr_cells[mid]
        band_rules(ws, rng, f"{col}{first}", watch_ref, alert_ref, direction)
        stress_heat(ws, rng, direction)
    return first, last, med


def write_loanbook(wb, peer_slots, raw_slots, thr_cells, peer_cells,
                   n_active):
    """The two-track loan-book tab (SPEC_COMPETITOR_PACK.md sec 1): consumer
    DQ/NCO band on top, commercial Call-Report FLOOR band below -- each with
    its own honest caption and PEER MEDIAN row per column."""
    ws = wb.create_sheet(LB_TAB)
    hide_gridlines(ws)
    ncols = 3 + len(LB_CONSUMER)                     # the wider band
    brand_banner(
        ws, 1, ncols,
        "Loan-Book Dashboard -- Two-Track Competitor Surveillance",
        "Per-class credit quality: 30-89 PD / 90+ PD (still accruing, "
        "disjoint from NA) / nonaccrual / QUARTERLY NCO annualized x4 (F1: "
        "never the YTD flow). Consumer classes carry the classification "
        "story; commercial classes are the public floor. Blank = null "
        "(never 0, F3); *AUTO exists from 2011, construction splits from "
        "~2007-08 (null-tolerant).")
    wl_last = WL_FIRST + peer_slots - 1
    tiles = [
        ("BANKS IN ALERT",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"ALERT\")",
         f"of {n_active} active"),
        ("BANKS IN WATCH",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"WATCH\")",
         "elevated"),
        ("BANKS TRACKED",
         f"=COUNTIF(Watchlist!$A${WL_FIRST}:$A${wl_last},\"?*\")",
         f"of {peer_slots} built slots"),
    ]
    kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])

    # band 1: consumer track, captioned on the spacer row above the header
    cap = ws.cell(6, 1, LB_CONSUMER_CAPTION)
    cap.font = SMALL_BOLD
    cap.alignment = KB.LEFT
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ncols)
    _slot_band(ws, DASH_HDR, LB_CONSUMER, peer_slots, raw_slots, thr_cells,
               peer_cells)

    # band 2: commercial floor, with the spec's honesty subtitle verbatim
    hdr2 = lb_comm_hdr(peer_slots)
    cap2 = ws.cell(hdr2 - 1, 1, LB_COMMERCIAL_CAPTION)
    cap2.font = Font(name="Calibri", bold=True, size=9, color=KB.CRIMSON)
    cap2.alignment = KB.LEFT
    ws.merge_cells(start_row=hdr2 - 1, start_column=1, end_row=hdr2 - 1,
                   end_column=3 + len(LB_COMMERCIAL))
    _slot_band(ws, hdr2, LB_COMMERCIAL, peer_slots, raw_slots, thr_cells,
               peer_cells)

    freeze_below(ws, DASH_HDR)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 13
    for k in range(len(LB_CONSUMER)):
        ws.column_dimensions[get_column_letter(4 + k)].width = 11

    # run-status readout: this tab is wider than the others, so the panel
    # lives in column X (runner.STATUS_COL_BY_TAB), clear of the banner merge
    scol_i = R.STATUS_COL_BY_TAB[LB_TAB]
    ws.column_dimensions[get_column_letter(scol_i)].width = 40
    for rr in (1, 2, 3):
        ws.cell(rr, scol_i).font = KB.SECONDARY
        ws.cell(rr, scol_i).alignment = KB.LEFT
    ws.cell(4, scol_i).font = NOTE_FONT               # macro writes here
    ws.cell(4, scol_i).alignment = KB.LEFT
    return ws


# ==========================================================================
# Watchlist  (the ranked peer list -- the entity-gated lane)
# ==========================================================================
BOUNDARY = ("ENTITY RULE. This lane admits ONLY rows carrying a genuine FDIC "
            "certificate join key (cert:NNN from the [PEERS] cert cell, "
            "Class A) -- the default-deny validator refuses anything else BY "
            "NAME: aggregates, holding companies and name-only rows cannot "
            "be monitored as counterparties. Blank/malformed CERT shows "
            "REFUSED here; find CERTs with: python runner.py --lookup "
            "\"<bank name>\". RSSD/holding-company promotion is a spec "
            "change, not a config edit.")

WL_COLS = ["Bank", "CERT", "Group", "Texas", "ALERT flags", "WATCH flags",
           "Rank", "Status"]


def write_watchlist(wb, cfg, peer_slots, thr_cells, peer_cells):
    metric_ids = [s.id for s in cfg.series]
    hj_name = get_column_letter(WL_HELPER_COL0)
    hx_name = get_column_letter(WL_HELPER_COL0 + len(metric_ids) - 1)
    ws = wb.create_sheet("Watchlist")
    hide_gridlines(ws)
    brand_banner(ws, 1, len(WL_COLS),
                 "Watchlist -- Ranked Peer / Counterparty Banks",
                 f"Ranked by (ALERT flag count, Texas ratio) descending. "
                 f"Flag counts aggregate all {len(metric_ids)} metric "
                 f"statuses (helper columns {hj_name}-{hx_name}); STALE "
                 f"banks are surfaced by the runner's status line and "
                 f"digest, not here (v1).")
    watchlist_boundary(ws, 3, len(WL_COLS), BOUNDARY)

    header_row(ws, WL_HDR, WL_COLS, right_from=3, center_cols=(6, 7))
    # helper mini-headers: one per metric status column + the rank score
    for k, mid in enumerate(metric_ids):
        c = ws.cell(WL_HDR, WL_HELPER_COL0 + k, mid)
        c.font = SMALL_BOLD
    ws.cell(WL_HDR, WL_SCORE_COL, "rank score").font = SMALL_BOLD

    first, last = WL_FIRST, WL_FIRST + peer_slots - 1
    tex_tab, tex_col, tex_first = metric_home("TEXAS", peer_slots)
    for slot in range(1, peer_slots + 1):
        r = wl_row(slot)
        pc = peer_cells[slot]
        dr = dash_row(slot)
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cert']}=\"\",\"\",{pc['cert']})").font = MONO_FONT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3, f"=IF({pc['group']}=\"\",\"\",{pc['group']})").font = KB.SECONDARY
        tex_ref = f"{tex_tab}!{get_column_letter(tex_col)}{tex_first + slot - 1}"
        ws.cell(r, 4, f"=IF({tex_ref}=\"\",\"\",{tex_ref})").number_format = "0.0"
        # per-metric status helpers (J..): same thresholds, same direction,
        # same blank-guard as runner.status_for -- parity-tested. The row on
        # the home tab is per-metric (the LoanBook commercial band sits
        # below the consumer band).
        for k, mid in enumerate(metric_ids):
            tab, ci, mfirst = metric_home(mid, peer_slots)
            ref = f"{tab}!{get_column_letter(ci)}{mfirst + slot - 1}"
            watch_ref, alert_ref, direction = thr_cells[mid]
            op = ">=" if direction != "below" else "<="
            ws.cell(r, WL_HELPER_COL0 + k,
                    f"=IF(NOT(ISNUMBER({ref})),\"\","
                    f"IF({ref}{op}{alert_ref},\"ALERT\","
                    f"IF({ref}{op}{watch_ref},\"WATCH\",\"OK\")))"
                    ).font = NOTE_FONT
        hj = get_column_letter(WL_HELPER_COL0)
        hx = get_column_letter(WL_HELPER_COL0 + len(metric_ids) - 1)
        ws.cell(r, 5, f"=IF($A{r}=\"\",\"\","
                      f"COUNTIF(${hj}{r}:${hx}{r},\"ALERT\"))")
        ws.cell(r, 6, f"=IF($A{r}=\"\",\"\","
                      f"COUNTIF(${hj}{r}:${hx}{r},\"WATCH\"))")
        # rank score = ALERT count dominant, Texas tiebreak (spec sec 4)
        zc = get_column_letter(WL_SCORE_COL)
        ws.cell(r, WL_SCORE_COL,
                f"=IF(NOT(ISNUMBER($E{r})),\"\","
                f"$E{r}*10000+IF(ISNUMBER($D{r}),$D{r},0))").font = NOTE_FONT
        ws.cell(r, 7, f"=IF(${zc}{r}=\"\",\"\","
                      f"RANK(${zc}{r},${zc}${first}:${zc}${last}))")
        # Status: the formula-level entity gate rendering -- a named REFUSED
        # for a blank/malformed CERT (with the --lookup hint), OFF for an
        # inactive slot, else the flag rollup. STALE is runtime state the
        # digest/status line carries (v1 -- see _readme).
        ws.cell(r, 8,
                f"=IF({pc['name']}=\"\",\"\","
                f"IF(NOT(ISNUMBER({pc['cert']})),"
                f"\"REFUSED: blank/invalid CERT -- run --lookup\","
                f"IF(UPPER({pc['active']}&\"\")<>\"TRUE\",\"OFF\","
                f"IF($E{r}>0,\"ALERT\",IF($F{r}>0,\"WATCH\",\"OK\")))))")
        for c in (4,):
            ws.cell(r, c).alignment = KB.RIGHT
        for c in (5, 6, 7, 8):
            ws.cell(r, c).alignment = KB.CENTER
        for c in range(1, len(WL_COLS) + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    KB.alert_rule(ws, f"H{first}:H{last}")
    stress_heat(ws, f"D{first}:D{last}", "above")   # red = high Texas
    stress_heat(ws, f"E{first}:E{last}", "above")   # red = many ALERT flags

    r = last + 2
    ws.cell(r, 1, f"Columns {hj_name}-{hx_name} are the per-metric status "
                  "helpers feeding the flag COUNTIFs; the rank score column "
                  "makes the (ALERT, Texas) ordering auditable. Sort/filter "
                  "on Rank if desired.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=len(WL_COLS))
    r += 1
    # build-time refusals (none under the seeded config; rendered
    # interpolated if a bad seed row is ever injected and the build-time
    # hard gate bypassed)
    _, refusals, _ = R.evaluate_peers(cfg)
    for peer, message in refusals:
        ws.cell(r, 1, f"slot {peer.slot}").font = MONO_FONT
        ws.cell(r, 2, "REFUSED").font = Font(name="Calibri", bold=True,
                                             color=KB.CRIMSON)
        r += 1
        ws.cell(r, 1, message).font = NOTE_FONT
        ws.cell(r, 1).alignment = KB.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=len(WL_COLS))
        ws.row_dimensions[r].height = 56
        r += 1

    widths = {"A": 26, "B": 9, "C": 13, "D": 10, "E": 11, "F": 11, "G": 7,
              "H": 34}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for k in range(len(metric_ids)):
        ws.column_dimensions[get_column_letter(WL_HELPER_COL0 + k)].width = 9
    freeze_below(ws, WL_HDR)
    return ws


# ==========================================================================
# _provenance  (TEMPLATE_CONTRACT.md sec 12 -- the tie-out map as a sheet)
# ==========================================================================
def write_provenance(wb):
    """Render provenance_seed (the PROVENANCE_MAP_FDIC.md transcription):
    citations-of-record header + the facsimile URL pattern, then one row per
    landed FIELD and per derived METRIC (field | schedule | line/caption |
    MDRM | flag | notes). The runner's --tieout mode reads THIS tab, so the
    workbook stays the source of truth."""
    ws = wb.create_sheet("_provenance")
    hide_gridlines(ws)
    widths = {"A": 12, "B": 34, "C": 46, "D": 34, "E": 10, "F": 60}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    r = 1
    for line in PROV.PREAMBLE:
        text_cell(ws, r, 1, line)
        ws.cell(r, 1).font = INK_BOLD if r == 1 else NOTE_FONT
        r += 1
    r += 1
    ws.cell(r, 1, "[FIELDS + METRICS]").font = KB.SECTION_FONT
    ws.cell(r, 1).fill = SECT_FILL
    r += 1
    hdr = r
    for j, name in enumerate(PROV.HEADER, start=1):
        c = ws.cell(hdr, j, name)
        c.font = SMALL_BOLD
        c.fill = KB.SUBHDR_FILL
    r += 1
    for row in PROV.ALL_ROWS:
        for j, val in enumerate(row, start=1):
            text_cell(ws, r, j, val)
            ws.cell(r, j).font = MONO_FONT if j in (1, 4) else NOTE_FONT
        if row[4].startswith("[~]") or row[4] == "UNVERIFIED":
            ws.cell(r, 5).font = Font(name="Calibri", bold=True, size=9,
                                      color=KB.CRIMSON)
        r += 1
    r += 1
    text_cell(ws, r, 1,
              "# Every [SERIES] metric has a row here: direct metrics under "
              "their field name, derived metrics under their metric id with "
              "the exact formula + component MDRMs. Sample-verify from "
              "PowerShell: python runner.py -w <book>.xlsm --tieout <CERT> "
              "[REPDTE] (--demo prints labeled demo values).")
    ws.cell(r, 1).font = NOTE_FONT
    freeze_below(ws, hdr)
    return ws


# ==========================================================================
# Code-in-tab + readme
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
BANK COUNTERPARTY & PEER MONITOR (FDIC BankFind) -- README
==========================================================

WHAT THIS IS
  A self-contained, reusable template for CREDIT-RISK REVIEW of BANKS: your
  hand-picked peer/counterparty list IS the watchlist, keyed by FDIC CERT.
  One workbook pulls quarterly Call-Report-derived metrics for every active
  bank in the _config [PEERS] table from the FDIC BankFind API (public
  domain, KEYLESS -- no API key, no account), lands every raw field as an
  audit trail, and presents banks-as-rows dashboards with per-metric
  threshold coloring, peer medians, and a ranked Watchlist.

THE FLEXIBLE PEER LIST (the point of this template)
  [PEERS] in _config: slot | cert | name | group | active.
  - ADD a bank: type its CERT/name/group/active=TRUE on a free slot row,
    re-run the runner. No rebuild. Find CERTs from PowerShell:
        python runner.py --lookup "Frost Bank"
  - REMOVE a bank: set active=FALSE (or clear the row); the slot blanks on
    the next run (stateless clearing -- never last run's values).
  - CAPACITY is built in (this workbook: see peer_slots in [SETTINGS]).
    More banks than slots = the runner REFUSES with the exact rebuild
    command (make_workbook.py --peer-slots N). It never truncates.
  Bank identity flows into every tab BY FORMULA from the [PEERS] cells, and
  raw-data anchors depend only on the SLOT -- so a swap re-labels and
  re-points everything on recalc.

PROVIDERS (the adapter seam -- one provider per class)
  Class A (public, KEYLESS):
    - FdicProvider     -- LIVE plain-urllib REST against api.fdic.gov/banks.
                          ONE bulk /financials request per refresh (whole
                          peer set), one /institutions roster call (merger
                          detection). Throttled 0.6s, backoff on 429/5xx.
                          NO API KEY NEEDED -- nothing to configure.
    - FdicDemoProvider -- deterministic OFFLINE stand-in; used by every test
                          and the --demo flag. Demo banks are illustrative:
                          the stressed ones are FICTION assigned by a hash,
                          not a statement about any real bank.
  Class C (licensed): NOT admitted anywhere in this template -- the gates
  admit only Class A, and a licensed swap requires its own review.

ONE-TIME SETUP
  1. Install Python 3 (add it to PATH) and the dependencies:
         pip install pandas openpyxl
  2. Enable macros when prompted (the Extract button needs them).
  That is all -- the live API needs NO key.

RUN IT  (the workbook is the finished product; you only add data)
  1. In Excel press Alt+F8 -> run ExtractFiles. It writes runner.py (from
     _code_py), requirements.txt, and RUN.txt next to the workbook.
     NOTHING runs inside Excel.
  2. SAVE and CLOSE the workbook.
  3. From PowerShell (see RUN.txt):
       - Try-the-button (offline): python runner.py -w <book>.xlsm --demo
       - Live FDIC (keyless):      python runner.py -w <book>.xlsm
       - Find a CERT:              python runner.py --lookup "<bank name>"
  4. Reopen the workbook -- the formulas recalc into the populated tabs.

THE METRICS (authorities named; "heuristic" means exactly that)
  Asset quality: NCLNLSR (noncurrent %), NTLNLSQR (QUARTERLY net charge-off
    % -- the YTD variant is deliberately excluded, trap F1; Q4 is seasonally
    heavy, compare YoY -- F7), PD3089R = P3LNLS/LNLSGR*100 (computed; no API
    ratio exists), LNATRESR (allowance/loans), LNRESNCR (coverage of
    noncurrents; below-is-bad).
  TEXAS RATIO -- VARIANT (documented): NCLNLS/(EQ+LNATRES)*100. The
    canonical Cassidy/RBC formulation adds OREO to the numerator and nets
    intangibles from equity; the ORE/INTAN field ids are UNVERIFIED, so v1
    computes the variant from verified fields only. Bands: 50 WATCH / 100
    ALERT (StL Fed 2025: 50-100 significant stress, >100 historic failure
    signal).
  Capital: RBC1AAJ (Tier 1 leverage -- universal incl. CBLR electors; PCA
    well-capitalized 5%; CBLR floor 8% EFFECTIVE JULY 1 2026 -- vintage
    dated, F8), RBCRWAJ (total risk-based; PCA well-cap 10%; JSON NULL for
    CBLR electors -- rendered blank, never zero, trap F3), EQV.
  Earnings: ROAQ (QUARTERLY ROA -- trap F1), NIMY, EEFFR (above-is-bad).
    QBP Q1'26 industry references: ROA 1.26%, NIM 3.31%, NCO 0.59%.
  Funding/concentration: LNDEPR = LNLSNET/DEP*100, BRODEPR = BRO/DEP*100
    (brokered deposits are restricted below well-capitalized -- FDI Act s29 /
    12 CFR 337.6; the healthy-bank bands are heuristic), CRECONR =
    (LNRECONS+LNRENRES+LNREMULT)/(EQ+LNATRES)*100 -- the 2006 interagency
    guidance screens total CRE >= 300% of TOTAL RISK-BASED CAPITAL; CBLR
    electors report none, so v1 uses the documented PROXY denominator
    EQ+LNATRES until the tier-1 dollar field is verified live. The
    guidance's 36-month growth leg is omitted (merger-distorted, F6).

THE v1.1 COMPETITOR PACK (SPEC_COMPETITOR_PACK.md; fields verified 93/93)
  TWO-TRACK LOAN BOOK (Dashboard_LoanBook, one tab, two bands):
  - CONSUMER track (cards CRCD, auto AUTO, other consumer CONOTH, 1-4
    family RERES + a HELOC RELOC drill-in): 30-89 PD / 90+ PD (still
    accruing -- DISJOINT from nonaccrual) / nonaccrual / QUARTERLY NCO
    annualized x4 (the FDIC's own NTLNLSQR convention; the YTD NT{c} flow
    is never fetched -- F1). Retail classification is DPD-formula-driven
    (Uniform Retail Credit Classification), so delinquency IS the
    classification story. Card bands sit high on purpose (QBP context).
  - COMMERCIAL floor (construction RECONS, nonfarm-nonres CRE RENRES,
    multifamily REMULT, C&I CI), same four rates -- honestly subtitled:
    a public Call-Report PROXY; commercial risk ratings lead delinquency;
    the criticized/classified view arrives via the EDGAR tracker (#6).
  - Naming honesty: verified R ratio twins consumed directly ONLY where
    the twin name fits the legacy 8-char field limit (P3CRCDR yes,
    P3CONOTHR no -- computed from the verified dollar triple + balance);
    quarterly NCO names are 8-char truncated (NTRECONQ, never NTRECONSQ).
  - HELOC has NO NCO column: LNRELOC is not in the verified balance list.
  SVB / FUNDING-STRESS metrics (on Dashboard_Funding_Concentration):
  - UNINSDEPR = DEPUNINS/DEP*100. RC-O Mem 2 is FILED by $1B+ reporters
    only -- a null renders BLANK plus a digest note, NEVER 0 (F3).
    Bands 40/60 heuristic (2023 failures: SVB ~94% uninsured).
  - UNRLZCAPR = ((SCHA-SCHF)+(SCAA-SCAF))/(EQ+LNATRES)*100 -- BOTH legs
    COMPUTED from the RC-B fair/amortized pairs (no named unrealized field
    exists). EQCCOMPI is the YTD OCI FLOW, not the AOCI stock -- it is
    deliberately NOT used. Bands 25/50 heuristic (SVB ran >100%).
  - FHLBASSR = OTHBFHLB/ASSET*100 (RC-M 5.a maturity lines summed; the
    structured-advance of-which line is NOT added). Bands 10/20 heuristic.
  Field vintages (null-tolerant, F-traps): *AUTO from 2011; construction
  splits from ~2007-08; DEPUNINS materially populated for larger reporters.

PROVENANCE & TIE-OUT (contract sec 12 -- verify against the OFFICIAL record)
  The _provenance tab maps EVERY landed field and EVERY metric to its Call
  Report schedule + line/caption + MDRM code, with per-row honesty flags
  ([V] verified against FFIEC captions; [~] match by caption). It carries
  the citations of record and the facsimile URL pattern:
    https://cdr.ffiec.gov/Public/ViewFacsimileDirect.aspx?ds=call
        &idType=fdiccert&id={CERT}&date={MMDDYYYY}
  Sample-verify any bank-quarter in minutes from PowerShell:
    python runner.py -w <book>.xlsm --tieout 628 2026-03-31
  prints each metric's value + schedule/line/MDRM + the facsimile and
  BankFind URLs. Works with --demo too (values labeled as demo fiction;
  the provenance columns are the real map either way).

REGULATORY VINTAGES (dated -- trap F8)
  - CBLR floor lowered 9% -> 8% effective July 1, 2026 (OCC NR 2026-30 /
    FDIC FIL). The RBC1AAJ WATCH band of 8 reflects the NEW floor.
  - UBPR peer groups changed Feb 26, 2026: commercial banks now 8 groups by
    asset size only. Any future peer-band logic must record this vintage.
  - PCA thresholds per 12 CFR 324.403 (well-capitalized: total RBC >= 10,
    tier 1 >= 8, CET1 >= 6.5, leverage >= 5).

STALENESS IS A FIRST-CLASS FAILURE (trap F2 -- the merged-bank lesson)
  A merged-away or closed bank keeps returning its FINAL REPDTE with no
  error. Every run, the runner compares each bank's latest REPDTE to the
  peer-set max; a bank more than stale_multiplier quarters behind is marked
  STALE in the digest and status line, EXCLUDED from the alert counts, and
  annotated "possible merger/closure -- check /institutions ACTIVE,
  /history". The roster call surfaces ACTIVE=0/ENDEFYMD directly.
  KNOWN v1 DIVERGENCE (deliberate, documented): the PEER MEDIAN rows in the
  dashboards are Excel MEDIAN() formulas over the landed slot rows -- Excel
  cannot see runtime staleness, so a STALE bank's last-landed values still
  enter the WORKBOOK medians, while the runner's DIGEST medians exclude
  them. The digest is authoritative for alerting; treat workbook medians as
  approximate whenever the status line shows STALE > 0. (v1.1 candidate:
  blank stale slots' latest-quarter cells at land time.)

THE WATCHLIST BOUNDARY (default-deny, defense in depth)
  A [PEERS] row is admitted ONLY if: the slot is active=TRUE, every metric
  row is Class A, and the entity key built from the cert cell matches
  cert:NNNNNNN. Blank/malformed CERTs are refused BY NAME with the --lookup
  hint (and the Watchlist Status column shows REFUSED by formula).
  Aggregates, holding companies and name-only rows cannot be monitored as
  counterparties. A build-time hard gate backs the runtime gate.

NOT AVAILABLE (honestly out of scope -- verified)
  - CAMELS/UFIRS ratings: CONFIDENTIAL, never public. The quantitative
    lanes mirror C-A-E-L + concentration; M/S are not computable.
  - Problem-bank NAMES: the QBP publishes only the count (54 in Q1 2026).
  - Holding-company (FR Y-9C) consolidated data: FFIEC/Chicago Fed, not
    this API. This template monitors the BANK (CERT) level only (the EDGAR
    template covers the HC view).
  - Category-level loan YIELDS and UBPR peer PERCENTILES: verified ABSENT
    from the FDIC API (an FFIEC bulk-ZIP ingest job if ever needed).
  - AOCI as a STOCK (EQCCOMPI is the YTD flow); OREO/intangibles for the
    canonical Texas ratio -- still unverified, never approximated.
  - Commercial criticized/classified loans: not public in Call Reports;
    the EDGAR tracker (#6) carries that view.
  - Intraquarter data: none exists publicly; call reports land ~35-60 days
    after quarter-end.

DATA-QUALITY TRAPS (baked in -- F1-F8)
  F1 YTD-vs-quarterly variants (ROAQ/NTLNLSQR used, never ROA/NTLNLSR)
  F2 merged-bank stale REPDTEs (staleness guard + roster ACTIVE check)
  F3 JSON null != 0 (CBLR ratios, BRO) -- blank cells, never zeros
  F4 mixed units in one record ($000 vs %) -- units carried per field
  F5 computed-ratio parity (Python == Excel, machine-tested)
  F6 merger survivorship (>25% single-quarter asset jump flagged in digest)
  F7 NCO Q4 seasonality (compare YoY -- subtitle note)
  F8 regulatory vintages dated (CBLR 8% 7/2026; UBPR 8-band 2/2026)

NO AI IS INVOLVED IN THE DATA PATH. The gates, metric functions, thresholds
and staleness guard are deterministic and config-driven: the same input
always yields the same workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() \
                and len(line) > 3 and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path, peer_slots=R.PEER_SLOTS_DEFAULT):
    rows = config_rows(peer_slots)
    cfg = R.parse_config(rows)
    # BUILD-TIME hard gates (BUILD SPEC 0.1/sec 3): bad metric rows, bad seed
    # peers, or an over-capacity seed refuse the BUILD itself.
    R.validate_metrics(cfg.series)
    R.validate_peer_capacity(cfg)
    R.assert_entity_gates(cfg)
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    raw_slots = cfg.raw_slots

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, peer_slots, raw_slots)
    # _config must be written before dashboards so threshold and [PEERS]
    # cell refs resolve.
    thr_cells, _set_cells, peer_cells = write_config(wb, peer_slots)

    for tab, title, subtitle, metric_ids in LANES:
        write_dashboard(wb, tab, title, subtitle, metric_ids, peer_slots,
                        raw_slots, thr_cells, peer_cells, len(admitted))
    write_loanbook(wb, peer_slots, raw_slots, thr_cells, peer_cells,
                   len(admitted))
    write_watchlist(wb, cfg, peer_slots, thr_cells, peer_cells)
    write_provenance(wb)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    # v1.1 tab order: LoanBook after Funding_Concentration, _provenance
    # after _config (SPEC_COMPETITOR_PACK.md sec 3 / contract sec 12)
    order = ["Dashboard_AssetQuality", "Dashboard_Capital_Earnings",
             "Dashboard_Funding_Concentration", LB_TAB, "Watchlist",
             R.RAW_TAB, "_config", "_provenance", "_code_py", "_code_vba",
             "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order
                    else 99)
    wb.active = wb.sheetnames.index("Dashboard_AssetQuality")
    wb.save(out_path)
    return out_path, len(cfg.series), len(admitted), peer_slots


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "Bank_Peer_Monitor_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n_metrics, n_admitted, cap = build(out)
    print(f"built {path}: {n_metrics} metrics x {cap} slots, "
          f"{n_admitted} seed peers admitted")
