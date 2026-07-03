#!/usr/bin/env python3
"""Bank Counterparty & Peer Monitor -- runner (the data path).

Single source of truth for the data path; embedded verbatim into the workbook's
_code_py tab. The VBA ExtractFiles macro writes this back out as runner.py and
the user runs it from PowerShell against the CLOSED workbook (openpyxl backend).

Built to BUILD_SPEC_FDIC.md. THE INVERSION: rows are BANKS, not series types --
the hand-picked [PEERS] list IS the watchlist, keyed by FDIC CERT, and [SERIES]
is a bank-agnostic 15-metric dictionary expanded at run time into units
s{slot:02d}_{METRIC}. Raw anchors depend ONLY on (slot, field), so editing
WHICH bank occupies a slot moves no formula (the flexible-peers USER
REQUIREMENT: add/remove = edit a [PEERS] line + re-run, never a rebuild).

Clean seams (BUILD SPEC 0.3 / 1a):
  * ADAPTER SEAM -- fetch_series(spec, secret) -> list[NormalizedRow], FIXED
    schema. Two Class A providers: FdicProvider (live, plain urllib, KEYLESS
    -- the FDIC BankFind API needs no API key) + FdicDemoProvider (offline
    deterministic, all tests). One in-process Class C stub proves the seam
    contract; note the gates admit ONLY class "A".
  * METRIC REGISTRY -- named PURE derived-ratio functions (PD3089R, TEXAS,
    LNDEPR, BRODEPR, CRECONR) defined IDENTICALLY here and in the Excel
    formulas (trap F5, parity-tested). None-tolerant: a missing field yields
    blank, never zero (trap F3: JSON null != 0).
  * THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT, direction-aware.
  * WATCHLIST VALIDATOR -- DEFAULT-DENY (BUILD SPEC 0.1 / sec 3): only rows
    carrying a genuine entity key (cert:NNNNNNN) are admitted; blank/malformed
    CERTs are refused BY NAME with a --lookup hint; non-Class-A metric rows
    refuse the run.
  * STALENESS GUARD (BUILD SPEC 0.6) -- merged-away banks return stale
    REPDTEs with no error (trap F2): a bank whose latest REPDTE lags the
    peer-set max by more than stale_multiplier quarters is STALE, excluded
    from digest alert counts, and surfaced for a merger/closure check.

No AI/LLM anywhere; every computation is deterministic (BUILD SPEC 0.5).
Pure ASCII (L3).
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd

# Fixed raw-block geometry: ONE block per PEER SLOT in Raw_FDIC, quarters as
# rows (newest first), fields as columns. Anchors depend only on (slot, field)
# -- never on which bank occupies the slot -- so dashboard formulas survive
# any [PEERS] edit.
RAW_TAB = "Raw_FDIC"
RAW_SLOTS_DEFAULT = 16         # quarters kept per bank (newest-first)
PEER_SLOTS_DEFAULT = 40        # BUILT peer capacity (make_workbook.py --peer-slots)
RAW_HEADER_ROWS = 2
RAW_GAP_ROWS = 2
RAW_FIRST_ROW = 2
RAW_PERIOD_COL = 1             # column A = REPDTE (ISO)

# Status panel column on each dashboard (free of any merged masthead cells).
# Dashboard_LoanBook is wider than the others (its banner merge spans past
# column L), so its status panel lives further right.
STATUS_COL = 12                # column L
STATUS_COL_BY_TAB = {"Dashboard_LoanBook": 24}   # column X

DASH_TABS = ("Dashboard_AssetQuality", "Dashboard_Capital_Earnings",
             "Dashboard_Funding_Concentration", "Dashboard_LoanBook")

# The pack version BUILT into this runner's raw layout / metric registry.
# A workbook built by a different pack is refused (rebuild, never mis-write).
PACK_VERSION = "1.1"

# Every raw field landed per bank (the audit trail: each derived metric's
# inputs are visible in Raw_FDIC). Verified names from risview_properties.yaml
# / the captured response -- COVERAGE_RESEARCH_FDIC.md; v1.1 competitor-pack
# fields verified in RESEARCH_COMPETITOR_PACK.md (93/93). Mixed units per
# record (trap F4): dollar fields are $ THOUSANDS, ratio fields are PERCENT.
# NAMING RULES (research-verified): quarterly NCO fields truncate the base to
# 8 chars before the Q (NTRECONQ, never NTRECONSQ); P3/P9/NA ratio twins are
# used ONLY where the un-truncated twin name fits the same 8-char limit
# (P3CRCDR yes, P3CONOTHR no -- those rates are computed from the verified
# dollar triple + balance instead).
RAW_FIELDS = [
    "ASSET", "DEP", "LNLSGR", "LNLSNET", "BRO", "EQ",           # $000
    "NCLNLS", "LNATRES", "P3LNLS",                              # $000
    "LNRECONS", "LNRENRES", "LNREMULT",                         # $000
    "NCLNLSR", "NTLNLSQR", "LNATRESR", "LNRESNCR",              # pct
    "RBC1AAJ", "RBCRWAJ", "EQV", "ROAQ", "NIMY", "EEFFR",       # pct
    # --- v1.1 consumer track: verified R ratio twins (pct) ---
    "P3CRCDR", "P9CRCDR", "NACRCDR",                            # pct
    "P3AUTOR", "P9AUTOR", "NAAUTOR",                            # pct
    "P3RERESR", "P9RERESR", "NARERESR",                         # pct
    "P3RELOCR", "P9RELOCR", "NARELOCR",                         # pct (HELOC)
    # other-consumer PD/NA dollar triple (twin name would exceed 8 chars)
    "P3CONOTH", "P9CONOTH", "NACONOTH",                         # $000
    # consumer balances + quarterly NCO dollars (F1: never the YTD NT{c})
    "LNCRCD", "LNAUTO", "LNCONOTH", "LNRERES",                  # $000
    "NTCRCDQ", "NTAUTOQ", "NTCONOTQ", "NTRERESQ",               # $000
    # --- v1.1 commercial floor ---
    "P3CIR", "P9CIR", "NACIR",                                  # pct twins
    "P3RECONS", "P9RECONS", "NARECONS",                         # $000
    "P3RENRES", "P9RENRES", "NARENRES",                         # $000
    "P3REMULT", "P9REMULT", "NAREMULT",                         # $000
    "LNCI",                                                     # $000
    "NTRECONQ", "NTRENREQ", "NTREMULQ", "NTCIQ",                # $000
    # --- v1.1 SVB / funding-stress pack ---
    "DEPUNINS",                                                 # $000 (nullable)
    "SCHA", "SCHF", "SCAA", "SCAF",                             # $000 HTM/AFS
    "OTHBFHLB",                                                 # $000 FHLB adv
]
PCT_FIELDS = {
    "NCLNLSR", "NTLNLSQR", "LNATRESR", "LNRESNCR",
    "RBC1AAJ", "RBCRWAJ", "EQV", "ROAQ", "NIMY", "EEFFR",
    "P3CRCDR", "P9CRCDR", "NACRCDR", "P3AUTOR", "P9AUTOR", "NAAUTOR",
    "P3RERESR", "P9RERESR", "NARERESR", "P3RELOCR", "P9RELOCR", "NARELOCR",
    "P3CIR", "P9CIR", "NACIR",
}
DOLLAR_FIELDS = {f for f in RAW_FIELDS if f not in PCT_FIELDS}
FIELD_UNITS = {f: ("USD_thousands" if f in DOLLAR_FIELDS else "pct")
               for f in RAW_FIELDS}

# The FDIC bulk endpoint caps fields= at 250; the ONE bulk request must stay
# far under it (spec sec 3). Guarded here AND at request build time.
MAX_REQUEST_FIELDS = 250
assert len(RAW_FIELDS) + 2 < MAX_REQUEST_FIELDS, "fields= list over the cap"

# The watchlist join-key whitelist (BUILD SPEC 0.1 / sec 3): default-deny.
# Only a literal FDIC certificate key passes; aggregates, name-only rows and
# unanticipated key forms are refused (holding-company/RSSD promotion is a
# spec change, not a config edit).
ENTITY_KEY_PATTERN = r"^cert:[0-9]{1,7}$"

QUARTER_DAYS = 92              # generous quarter length for the staleness math

FDIC_FIN_URL = "https://api.fdic.gov/banks/financials"
FDIC_INST_URL = "https://api.fdic.gov/banks/institutions"


def field_col(fname: str) -> int:
    """Raw_FDIC column for a field: B onward, in RAW_FIELDS order (fixed)."""
    return 2 + RAW_FIELDS.index(fname)


def slot_label(slot: int) -> str:
    return f"slot{slot:02d}"


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` tab; backend-agnostic)
# ---------------------------------------------------------------------------
SERIES_HEADER = [
    "id", "title", "category", "lane", "metric_type", "frequency", "sa_nsa",
    "units", "level_rate_index", "geo_segment", "source_class",
    "dashboard_capable", "watchlist_capable", "source_url", "table_id", "sheet",
    "series_label", "transform", "notes",
]

PEER_HEADER = ["slot", "cert", "name", "group", "active"]


@dataclass
class SeriesSpec:
    id: str
    title: str
    category: str
    lane: str
    metric_type: str
    frequency: str
    sa_nsa: str
    units: str
    level_rate_index: str
    geo_segment: str
    source_class: str
    dashboard_capable: bool
    watchlist_capable: bool
    source_url: str
    table_id: str
    sheet: str
    series_label: str
    transform: str
    notes: str


@dataclass
class PeerRow:
    """One [PEERS] line: the flexible peer list (USER REQUIREMENT). An empty
    row (slot number only) is provisioned headroom the user fills by hand."""
    slot: Optional[int]
    cert: str              # normalized digit string; "" when blank
    name: str
    group: str             # peer | counterparty | self (free text, rendered)
    active: bool

    @property
    def has_bank(self) -> bool:
        return bool(self.cert or self.name)

    @property
    def entity_key(self) -> str:
        """The join key the gate whitelists: cert:NNN built from the cert cell."""
        return f"cert:{self.cert}"


@dataclass
class Threshold:
    watch: Optional[float]
    alert: Optional[float]
    direction: str          # "above" | "below"


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, Threshold] = field(default_factory=dict)
    series: List[SeriesSpec] = field(default_factory=list)
    peers: List[PeerRow] = field(default_factory=list)

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))

    @property
    def peer_slots(self) -> int:
        """The BUILT slot capacity. Informational here -- the raw-layout check
        (slot header labels) is the hard guard against a stale value."""
        return int(float(self.settings.get("peer_slots", PEER_SLOTS_DEFAULT)))

    @property
    def stale_multiplier(self) -> float:
        v = _as_float(self.settings.get("stale_multiplier", "2.0"))
        return 2.0 if v is None else v


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _as_float(v):
    try:
        s = str(v).strip()
        return float(s) if s != "" else None
    except (TypeError, ValueError):
        return None


def _norm_cert(v) -> str:
    """Normalize a [PEERS] cert cell to a plain digit string. Excel hands
    numeric cells over as int/float (628 or 628.0); text stays as typed so a
    malformed value ('ABC', '12-34') reaches the gate and is refused, never
    silently coerced."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _norm_slot(v) -> Optional[int]:
    try:
        s = str(v).strip()
        return int(float(s)) if s != "" else None
    except (TypeError, ValueError):
        return None


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists). Sections in col A:
    [SETTINGS], [THRESHOLDS], [PEERS], [SERIES]. The [PEERS] section is this
    template's addition; the other three parse contract-identically."""
    cfg = Config()
    section = None
    series_header = None
    thr_header = None
    peer_header = None
    for raw in rows:
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            series_header = None
            thr_header = None
            peer_header = None
            continue
        if not a:
            continue
        if a.startswith("#"):          # in-sheet comment line, never data
            continue
        if section == "SETTINGS":
            if a.lower() in ("key", "name"):
                continue
            val = "" if len(raw) < 2 or raw[1] is None else raw[1]
            cfg.settings[a] = str(val).strip()
        elif section == "THRESHOLDS":
            if thr_header is None and a.lower() == "id":
                thr_header = [str(c).strip().lower() for c in raw]
                continue
            cells = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                     for i, h in enumerate(thr_header or
                                           ["id", "watch", "alert", "direction"])}
            cfg.thresholds[a] = Threshold(
                watch=_as_float(cells.get("watch")),
                alert=_as_float(cells.get("alert")),
                direction=str(cells.get("direction", "above")).strip().lower()
                or "above")
        elif section == "PEERS":
            if peer_header is None and a.lower() == "slot":
                peer_header = [str(c).strip().lower() for c in raw]
                continue
            v = {h: (None if i >= len(raw) else raw[i])
                 for i, h in enumerate(peer_header or PEER_HEADER)}
            cfg.peers.append(PeerRow(
                slot=_norm_slot(v.get("slot")),
                cert=_norm_cert(v.get("cert")),
                name=("" if v.get("name") is None else str(v.get("name")).strip()),
                group=("" if v.get("group") is None
                       else str(v.get("group")).strip().lower()),
                active=_as_bool(v.get("active"))))
        elif section == "SERIES":
            if series_header is None:
                series_header = [str(c).strip() for c in raw]
                continue
            v = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                 for i, h in enumerate(series_header)}
            cfg.series.append(SeriesSpec(
                id=str(v.get("id", "")).strip(),
                title=str(v.get("title", "")).strip(),
                category=str(v.get("category", "")).strip(),
                lane=str(v.get("lane", "")).strip().lower(),
                metric_type=str(v.get("metric_type", "")).strip(),
                frequency=str(v.get("frequency", "")).strip(),
                sa_nsa=str(v.get("sa_nsa", "")).strip(),
                units=str(v.get("units", "")).strip(),
                level_rate_index=str(v.get("level_rate_index", "")).strip().lower(),
                geo_segment=str(v.get("geo_segment", "")).strip(),
                source_class=str(v.get("source_class", "")).strip().upper(),
                dashboard_capable=_as_bool(v.get("dashboard_capable", "")),
                watchlist_capable=_as_bool(v.get("watchlist_capable", "")),
                source_url=str(v.get("source_url", "")).strip(),
                table_id=str(v.get("table_id", "")).strip(),
                sheet=str(v.get("sheet", "")).strip(),
                series_label=str(v.get("series_label", "")).strip(),
                transform=str(v.get("transform", "direct")).strip().lower(),
                notes=str(v.get("notes", "")).strip()))
    return cfg


# ---------------------------------------------------------------------------
# METRIC REGISTRY -- named PURE functions of the landed fields (BUILD SPEC 0.5
# / sec 2). Defined IDENTICALLY in the Excel formulas (build_workbook.py) --
# parity is tested, never assumed (trap F5). None-tolerant everywhere: a
# missing/null field yields None (blank), NEVER zero (trap F3).
# ---------------------------------------------------------------------------
def _ratio(num: Optional[float], den: Optional[float]) -> Optional[float]:
    """num/den*100 -- None on any missing input or a zero denominator, exactly
    like the Excel form IF(OR(n="",d=""),"",IF(d=0,"",n/d*100))."""
    if num is None or den is None or den == 0:
        return None
    return num / den * 100.0


def _sum(*vals) -> Optional[float]:
    """None-propagating sum: a null component (e.g. LNREMULT) blanks the
    composite rather than understating it as zero (trap F3)."""
    total = 0.0
    for v in vals:
        if v is None:
            return None
        total += v
    return total


def d_pd3089r(f):
    """30-89 day past due %: P3LNLS/LNLSGR*100 (no API ratio field exists)."""
    return _ratio(f.get("P3LNLS"), f.get("LNLSGR"))


def d_texas(f):
    """Texas ratio VARIANT: NCLNLS/(EQ+LNATRES)*100. Canonical adds OREO and
    nets intangibles -- ORE/INTAN field ids are UNVERIFIED (Open Q), so v1
    computes the documented variant from verified fields only."""
    return _ratio(f.get("NCLNLS"), _sum(f.get("EQ"), f.get("LNATRES")))


def d_lndepr(f):
    """Loans/deposits %: LNLSNET/DEP*100."""
    return _ratio(f.get("LNLSNET"), f.get("DEP"))


def d_brodepr(f):
    """Brokered deposit share %: BRO/DEP*100 (BRO null -> blank, never 0)."""
    return _ratio(f.get("BRO"), f.get("DEP"))


def d_creconr(f):
    """CRE concentration PROXY: (LNRECONS+LNRENRES+LNREMULT)/(EQ+LNATRES)*100.
    Guidance uses total risk-based capital; CBLR electors lack it, so the
    documented proxy denominator stands until the tier-1 dollar field is
    verified live."""
    return _ratio(_sum(f.get("LNRECONS"), f.get("LNRENRES"), f.get("LNREMULT")),
                  _sum(f.get("EQ"), f.get("LNATRES")))


# --------------------------------------------------------------------------
# v1.1 COMPETITOR PACK ratios -- ONE declarative table drives BOTH the Python
# functions (below) and the Excel formulas (build_workbook reads this table),
# so the two definitions cannot drift (trap F5).
#   metric id -> (numerator field, denominator field, multiplier)
# Multiplier 100 = plain percent; 400 = QUARTERLY flow ANNUALIZED (x4), the
# same convention as the FDIC's own NTLNLSQR (PROVENANCE_MAP_FDIC.md sec F).
# All None-tolerant: null numerator OR denominator -> blank, never 0 (F3) --
# DEPUNINS especially is genuinely null below the $1B reporting threshold.
# --------------------------------------------------------------------------
PACK_RATIOS = {
    # consumer PD/NA rates computed from the verified dollar triple + balance
    # (the R-twin name would exceed the 8-char field-name limit -- unverified)
    "P3CONOTHR": ("P3CONOTH", "LNCONOTH", 100),
    "P9CONOTHR": ("P9CONOTH", "LNCONOTH", 100),
    "NACONOTHR": ("NACONOTH", "LNCONOTH", 100),
    # consumer quarterly NCO rates (F1: the Q dollar flow, never YTD NT{c})
    "NTCRCDQR": ("NTCRCDQ", "LNCRCD", 400),
    "NTAUTOQR": ("NTAUTOQ", "LNAUTO", 400),
    "NTCONOTQR": ("NTCONOTQ", "LNCONOTH", 400),
    "NTRERESQR": ("NTRERESQ", "LNRERES", 400),
    # commercial-floor PD/NA rates (twin names exceed 8 chars -- computed)
    "P3RECONSR": ("P3RECONS", "LNRECONS", 100),
    "P9RECONSR": ("P9RECONS", "LNRECONS", 100),
    "NARECONSR": ("NARECONS", "LNRECONS", 100),
    "P3RENRESR": ("P3RENRES", "LNRENRES", 100),
    "P9RENRESR": ("P9RENRES", "LNRENRES", 100),
    "NARENRESR": ("NARENRES", "LNRENRES", 100),
    "P3REMULTR": ("P3REMULT", "LNREMULT", 100),
    "P9REMULTR": ("P9REMULT", "LNREMULT", 100),
    "NAREMULTR": ("NAREMULT", "LNREMULT", 100),
    # commercial quarterly NCO rates (8-char truncation: NTRECONQ etc.)
    "NTRECONQR": ("NTRECONQ", "LNRECONS", 400),
    "NTRENREQR": ("NTRENREQ", "LNRENRES", 400),
    "NTREMULQR": ("NTREMULQ", "LNREMULT", 400),
    "NTCIQR": ("NTCIQ", "LNCI", 400),
    # SVB pack simple ratios
    "UNINSDEPR": ("DEPUNINS", "DEP", 100),   # null -> BLANK + digest note
    "FHLBASSR": ("OTHBFHLB", "ASSET", 100),
}

# verified R-twin ratio fields consumed DIRECTLY (metric id IS the field)
PACK_DIRECT = [
    "P3CRCDR", "P9CRCDR", "NACRCDR",
    "P3AUTOR", "P9AUTOR", "NAAUTOR",
    "P3RERESR", "P9RERESR", "NARERESR",
    "P3RELOCR", "P9RELOCR", "NARELOCR",
    "P3CIR", "P9CIR", "NACIR",
]


def _pack_ratio_fn(num, den, mult):
    """Factory for the PACK_RATIOS functions: num/den*mult, None-tolerant,
    zero-denominator-safe -- exactly the Excel form
    IF(OR(n="",d=""),"",IF(d=0,"",n/d*mult))."""
    def fn(f):
        n, d = f.get(num), f.get(den)
        if n is None or d is None or d == 0:
            return None
        return n / d * float(mult)
    return fn


def d_unrlzcapr(f):
    """Unrealized securities loss / capital cushion (SVB metric):
    ((SCHA-SCHF)+(SCAA-SCAF))/(EQ+LNATRES)*100. BOTH legs COMPUTED from the
    verified fair/amortized pairs -- no named unrealized field exists, and
    EQCCOMPI is the YTD OCI FLOW, not the AOCI stock (never used). Positive =
    unrealized LOSS eroding the cushion. Any null leg blanks the metric."""
    vals = [f.get(k) for k in ("SCHA", "SCHF", "SCAA", "SCAF")]
    if any(v is None for v in vals):
        return None
    scha, schf, scaa, scaf = vals
    return _ratio((scha - schf) + (scaa - scaf),
                  _sum(f.get("EQ"), f.get("LNATRES")))


# Loan class per LoanBook metric (drives the email per-class alert section
# and the tieout grouping). Consumer classes are the DQ/NCO track; commercial
# classes are the public Call-Report FLOOR (criticized/classified via EDGAR).
LOANBOOK_CLASS = {}
for _mid, _cls in (
        (("P3CRCDR", "P9CRCDR", "NACRCDR", "NTCRCDQR"), "credit card"),
        (("P3AUTOR", "P9AUTOR", "NAAUTOR", "NTAUTOQR"), "auto"),
        (("P3CONOTHR", "P9CONOTHR", "NACONOTHR", "NTCONOTQR"),
         "other consumer"),
        (("P3RERESR", "P9RERESR", "NARERESR", "NTRERESQR"), "resi 1-4 fam"),
        (("P3RELOCR", "P9RELOCR", "NARELOCR"), "HELOC"),
        (("P3RECONSR", "P9RECONSR", "NARECONSR", "NTRECONQR"),
         "construction"),
        (("P3RENRESR", "P9RENRESR", "NARENRESR", "NTRENREQR"),
         "CRE nonfarm"),
        (("P3REMULTR", "P9REMULTR", "NAREMULTR", "NTREMULQR"),
         "multifamily"),
        (("P3CIR", "P9CIR", "NACIR", "NTCIQR"), "C&I")):
    for _m in _mid:
        LOANBOOK_CLASS[_m] = _cls
CONSUMER_CLASSES = ("credit card", "auto", "other consumer", "resi 1-4 fam",
                    "HELOC")
COMMERCIAL_CLASSES = ("construction", "CRE nonfarm", "multifamily", "C&I")


# metric id -> (fields consumed, derived fn or None for a direct API ratio)
METRICS = {
    "NCLNLSR": (("NCLNLSR",), None),
    "NTLNLSQR": (("NTLNLSQR",), None),
    "PD3089R": (("P3LNLS", "LNLSGR"), d_pd3089r),
    "LNATRESR": (("LNATRESR",), None),
    "LNRESNCR": (("LNRESNCR",), None),
    "TEXAS": (("NCLNLS", "EQ", "LNATRES"), d_texas),
    "RBC1AAJ": (("RBC1AAJ",), None),
    "RBCRWAJ": (("RBCRWAJ",), None),
    "EQV": (("EQV",), None),
    "ROAQ": (("ROAQ",), None),
    "NIMY": (("NIMY",), None),
    "EEFFR": (("EEFFR",), None),
    "LNDEPR": (("LNLSNET", "DEP"), d_lndepr),
    "BRODEPR": (("BRO", "DEP"), d_brodepr),
    "CRECONR": (("LNRECONS", "LNRENRES", "LNREMULT", "EQ", "LNATRES"),
                d_creconr),
}
# v1.1 pack: direct R-twins, the declarative simple ratios, and the one
# multi-leg SVB derivation.
for _f in PACK_DIRECT:
    METRICS[_f] = ((_f,), None)
for _mid, (_num, _den, _mult) in PACK_RATIOS.items():
    METRICS[_mid] = ((_num, _den), _pack_ratio_fn(_num, _den, _mult))
METRICS["UNRLZCAPR"] = (("SCHA", "SCHF", "SCAA", "SCAF", "EQ", "LNATRES"),
                        d_unrlzcapr)


class MetricError(Exception):
    pass


def validate_metrics(series: Sequence[SeriesSpec]) -> None:
    """Every [SERIES] row must be a registered metric whose transform column
    agrees with the registry (direct field vs named derived function), and
    must consume only landed RAW_FIELDS -- config drift refuses, not guesses."""
    for s in series:
        if s.id not in METRICS:
            raise MetricError(f"Series '{s.id}' is not a registered metric; "
                              f"known metrics: {', '.join(sorted(METRICS))}.")
        flds, fn = METRICS[s.id]
        expected = "direct" if fn is None else "derived"
        if s.transform != expected:
            raise MetricError(f"Series '{s.id}' declares transform "
                              f"'{s.transform}' but the registry defines it as "
                              f"'{expected}'.")
        for f in flds:
            if f not in RAW_FIELDS:
                raise MetricError(f"Series '{s.id}' consumes unlanded field "
                                  f"'{f}'.")


def metric_value(metric_id: str, fields: Dict[str, Optional[float]]):
    """The single headline computation: latest-quarter fields -> metric value.
    Identical to the Excel formula off the same Raw_FDIC cells."""
    flds, fn = METRICS[metric_id]
    if fn is None:
        return fields.get(flds[0])
    return fn(fields)


# ---------------------------------------------------------------------------
# THE WATCHLIST VALIDATOR -- DEFAULT-DENY (BUILD SPEC 0.1 / sec 3).
# Gate 1: the peer slot is active=TRUE (inactive = EXCLUDED, slot blanked).
# Gate 2: every metric row is Class A (a non-A row refuses the whole run --
#         it would poison every bank's counts).
# Gate 3: the entity key built from the [PEERS] cert cell matches
#         ^cert:[0-9]{1,7}$ (blank/malformed cert = refused BY NAME).
# ---------------------------------------------------------------------------
class WatchlistRefused(Exception):
    """Raised by the hard gates; carries the interpolated refusal message."""


class PeerCapacityError(Exception):
    """More peers than BUILT slots: refuse with the rebuild command -- the
    runner never truncates a peer list (USER REQUIREMENT)."""


def peer_refusal_message(peer: PeerRow, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: peer slot {peer.slot} "{peer.name or "(unnamed)"}" '
        f'has cert="{peer.cert}", active={"TRUE" if peer.active else "FALSE"}. '
        + " ".join(reasons) +
        " Only CERT-keyed FDIC institutions can be monitored as counterparties;"
        " aggregates, holding companies and name-only rows cannot. Find the"
        f' CERT with: python runner.py --lookup "{peer.name or "<bank name>"}".')


def metric_refusal_message(spec: SeriesSpec, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: series "{spec.id}" has '
        f'source_class="{spec.source_class}", '
        f'watchlist_capable={"TRUE" if spec.watchlist_capable else "FALSE"}. '
        + " ".join(reasons) +
        " This template admits only the public FDIC BankFind adapter"
        " (Class A); any other class requires its own review before"
        " admission.")


def gate_peer_row(peer: PeerRow) -> List[str]:
    """Failed-gate reasons for an ACTIVE peer row carrying a bank. Empty list
    == a genuinely CERT-keyed row."""
    reasons = []
    if not re.match(ENTITY_KEY_PATTERN, peer.entity_key):
        reasons.append(
            f"Gate3: entity key '{peer.entity_key}' does not match the "
            f"join-key whitelist pattern {ENTITY_KEY_PATTERN} (default-deny; "
            "a blank or malformed CERT cannot be fetched).")
    return reasons


def gate_metric_row(spec: SeriesSpec) -> List[str]:
    reasons = []
    if not spec.watchlist_capable:
        reasons.append("Gate1: watchlist_capable is not TRUE.")
    if spec.source_class != "A":
        reasons.append("Gate2: source_class is not A -- this template's only "
                       "admitted class is the keyless public FDIC adapter; "
                       "any other class (including a future licensed swap) "
                       "requires its own review before admission.")
    return reasons


def assert_metrics_admissible(series: Sequence[SeriesSpec]) -> None:
    """Hard gate (build-time AND run start): a fabricated non-Class-A metric
    row refuses everything, series-named -- defense in depth behind gate 2."""
    for s in series:
        reasons = gate_metric_row(s)
        if reasons:
            raise WatchlistRefused(metric_refusal_message(s, reasons))


def evaluate_peers(cfg: Config):
    """Apply the default-deny gates to every [PEERS] row carrying a bank.
    Returns (admitted, refusals, excluded): admitted peers are fetched and
    landed; refusals is [(peer, message)] -- those slots are never fetched and
    are blanked; excluded is the active=FALSE list (blanked, not refused)."""
    admitted, refusals, excluded = [], [], []
    for peer in cfg.peers:
        if not peer.has_bank:
            continue                       # provisioned empty slot
        if not peer.active:
            excluded.append(peer)          # gate 1: exclusion, not refusal
            continue
        reasons = gate_peer_row(peer)
        if reasons:
            refusals.append((peer, peer_refusal_message(peer, reasons)))
        else:
            admitted.append(peer)
    return admitted, refusals, excluded


def assert_entity_gates(cfg: Config) -> None:
    """BUILD-TIME hard gate backing the runtime gates (BUILD SPEC 0.1/sec 3):
    a bad metric row or a bad SEED peer cannot even be built into the lane.
    (At run time the per-peer gate refuses-and-continues so one typo does not
    kill the whole refresh; the refusal is rendered, the slot blanked.)"""
    assert_metrics_admissible(cfg.series)
    _, refusals, _ = evaluate_peers(cfg)
    if refusals:
        raise WatchlistRefused(refusals[0][1])


def validate_peer_capacity(cfg: Config) -> None:
    """The flexible-peers guard: every bank-carrying row needs a slot inside
    the BUILT capacity, unique. Over capacity is REFUSED with the rebuild
    command -- never truncated (USER REQUIREMENT)."""
    cap = cfg.peer_slots
    banked = [p for p in cfg.peers if p.has_bank]
    seen: Dict[int, PeerRow] = {}
    bad = []
    for p in banked:
        if p.slot is None or p.slot < 1 or p.slot > cap:
            bad.append(p)
        elif p.slot in seen:
            raise PeerCapacityError(
                f"[PEERS] slot {p.slot} is listed twice "
                f'("{seen[p.slot].name}" and "{p.name}"); slots must be '
                "unique -- fix the [PEERS] table.")
        else:
            seen[p.slot] = p
    if bad:
        need = max(len(banked), max((p.slot or 0) for p in banked))
        names = ", ".join(f'"{p.name or p.cert}" (slot {p.slot})' for p in bad)
        raise PeerCapacityError(
            f"[PEERS] lists more peers than the workbook was built for: "
            f"{names} outside the built capacity of {cap} slots. The runner "
            f"never truncates a peer list -- rebuild with "
            f"--peer-slots {max(need, cap + 1)} "
            f"(python make_workbook.py --peer-slots {max(need, cap + 1)}).")


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT, direction-aware
# ---------------------------------------------------------------------------
def status_for(value: Optional[float], thr: Optional[Threshold]) -> str:
    """direction='above': flag when value >= bound; 'below': value <= bound
    (capital/coverage/earnings run below-is-bad). Blank value -> OK (blank is
    surfaced by the staleness/error paths, never fabricated into a flag)."""
    if value is None or thr is None or (isinstance(value, float)
                                        and math.isnan(value)):
        return "OK"
    above = thr.direction != "below"

    def hit(bound):
        if bound is None:
            return False
        return value >= bound if above else value <= bound

    if hit(thr.alert):
        return "ALERT"
    if hit(thr.watch):
        return "WATCH"
    return "OK"


# ---------------------------------------------------------------------------
# STALENESS GUARD (BUILD SPEC 0.6 -- trap F2, the merged-bank lesson)
# ---------------------------------------------------------------------------
def is_stale_bank(last_period: Optional[str], set_max_period: Optional[str],
                  stale_multiplier: float) -> bool:
    """Fetch success is NOT data currency: a merged-away bank keeps returning
    its final REPDTE with no error. A bank is STALE when its latest REPDTE
    lags the PEER-SET max by more than stale_multiplier quarters (generous
    92-day quarters). Valueless bank -> stale."""
    if last_period is None:
        return True
    if set_max_period is None:
        return False                      # nothing landed anywhere: no baseline
    try:
        last = date.fromisoformat(str(last_period)[:10])
        mx = date.fromisoformat(str(set_max_period)[:10])
    except ValueError:
        return True
    return (mx - last).days > stale_multiplier * QUARTER_DAYS


# ---------------------------------------------------------------------------
# THE ADAPTER SEAM (BUILD SPEC sec 1a) -- the only provider-specific code
# ---------------------------------------------------------------------------
@dataclass
class NormalizedRow:
    id: str
    period: str            # ISO date string (quarter-end REPDTE)
    value: Optional[float]
    geo_segment: str       # the ENTITY key: cert:NNN
    source_class: str
    units: str = ""


@dataclass
class FieldSpec:
    """One expanded fetch/land unit: (peer slot, cert) x raw field. The id
    scheme s{slot:02d}_{FIELD} mirrors the digest's s{slot:02d}_{METRIC}
    units; anchors derive from (slot, field) alone."""
    slot: int
    cert: str
    fname: str
    id: str
    geo_segment: str
    units: str
    source_class: str = "A"


def make_field_spec(peer: PeerRow, fname: str) -> FieldSpec:
    return FieldSpec(slot=peer.slot, cert=peer.cert, fname=fname,
                     id=f"s{peer.slot:02d}_{fname}",
                     geo_segment=peer.entity_key, units=FIELD_UNITS[fname])


class Provider:
    """Adapter interface: prime(certs, asof) once per run (the ONE bulk call),
    then fetch_series(spec, secret) -> [NormalizedRow] slices per unit."""

    source_class = "A"
    vintage: Optional[str] = None      # meta.index.createTimestamp (data age)

    def __init__(self):
        self.roster: Dict[str, dict] = {}

    def prime(self, certs: Sequence[str], asof: date,
              names: Optional[Dict[str, str]] = None) -> None:  # pragma: no cover
        raise NotImplementedError

    def fetch_series(self, spec: FieldSpec, secret=None) -> List[NormalizedRow]:  # pragma: no cover
        raise NotImplementedError


class FdicDemoProvider(Provider):
    """Deterministic offline stand-in (BUILD SPEC 0.7). Seeded per (cert,
    field) -- the SAME bank yields the SAME history in any slot, so a [PEERS]
    swap visibly moves data between slots (the flexibility test). Realistic
    ranges per metric; the seed profile puts one illustrative bank in a
    Texas-ALERT tier and one in a Texas-WATCH tier; one bank's BRO is null
    for the whole series (trap F3); every series carries one interior None.
    NO network, NO key; never stale at its own asof."""

    source_class = "A"

    def __init__(self, asof: Optional[date] = None,
                 raw_slots: int = RAW_SLOTS_DEFAULT):
        super().__init__()
        self.asof = asof or date(2026, 3, 31)
        self.raw_slots = raw_slots
        self.vintage = "demo data (FdicDemoProvider -- no live vintage)"
        self._profiles: Dict[str, List[Tuple[str, Dict[str, Optional[float]]]]] = {}

    def prime(self, certs, asof=None, names=None):
        names = names or {}
        for c in certs:
            self.roster[str(c)] = {
                "CERT": str(c), "NAME": names.get(str(c), f"Bank {c}"),
                "ACTIVE": 1, "BKCLASS": "N", "ENDEFYMD": None,
                "FED_RSSD": None}

    @staticmethod
    def _seed(cert: str) -> int:
        try:
            n = int(cert)
        except ValueError:
            n = sum((i + 1) * ord(c) for i, c in enumerate(cert))
        return (n * 2654435761) % (2 ** 32)

    def _periods(self) -> List[str]:
        idx = pd.period_range(end=pd.Period(self.asof, freq="Q"),
                              periods=self.raw_slots, freq="Q")
        return [p.to_timestamp(how="end").date().isoformat() for p in idx]

    def _profile(self, cert: str):
        """Full consistent quarter-by-quarter field history for one bank:
        dollar fields derive from the SAME wobbled ratios the ratio fields
        report, so every derived metric (Texas, coverage, CRE) is internally
        consistent -- the parity tests lean on that."""
        if cert in self._profiles:
            return self._profiles[cert]
        s = self._seed(cert)
        n = self.raw_slots
        tier = s % 13                       # 0 = ALERT tier, 1 = WATCH tier
        # newest-quarter targets per tier (ratios in percent)
        if tier == 0:      # deep-stress illustrative bank: Texas > 100
            eq_r, ncl, res_r, p3, nco = 0.035, 13.0, 6.5, 3.4, 2.3
            roaq, nimy, eeffr = -0.6, 1.9, 88.0
        elif tier == 1:    # stressed illustrative bank: Texas 50-100
            eq_r, ncl, res_r, p3, nco = 0.055, 8.6, 6.4, 2.1, 1.3
            roaq, nimy, eeffr = 0.3, 2.35, 74.0
        else:              # healthy ranges
            eq_r = 0.092 + (s % 25) * 0.001
            ncl = 0.35 + (s % 14) * 0.09
            res_r = 1.05 + (s % 9) * 0.09
            p3 = 0.5 + (s % 9) * 0.1
            nco = 0.12 + (s % 10) * 0.05
            roaq = 0.85 + (s % 8) * 0.09
            nimy = 2.65 + (s % 14) * 0.11
            eeffr = 52.0 + (s % 17)
        asset0 = 2.0e6 + (s % 400) * 1.0e6          # $000: $2B - $402B
        loans_r = 0.55 + (s % 18) * 0.01
        dep_r = 0.72 + (s % 9) * 0.01
        bro_r = ((s >> 4) % 29) / 200.0     # 0-14%: a couple of WATCHes, no ALERT
        bro_none = (s % 17) == 3                     # whole-series null BRO
        cre_c = 0.02 + (s % 6) * 0.005
        cre_n = 0.10 + (s % 12) * 0.01
        cre_m = 0.03 + (s % 7) * 0.01
        # ---- v1.1 pack seeding (per class; newest-quarter targets, pct) ----
        # The stress tiers trip the new bands (spec sec 3: 1-2 demo banks);
        # healthy ranges sit under the ALERT bands (a few WATCHes by design).
        # Class rates: {class: (PD30-89, PD90+, nonaccrual, NCOq annualized)}
        if tier == 0:
            rate = {"crcd": (3.8, 2.9, 0.8, 4.8), "auto": (4.4, 1.2, 1.7, 2.3),
                    "conoth": (3.6, 2.2, 1.9, 3.9), "reres": (4.2, 2.1, 2.4, 0.9),
                    "reloc": (3.1, 1.6, 2.2, 0.0), "recons": (3.4, 1.7, 3.3, 1.7),
                    "renres": (2.2, 1.1, 2.5, 1.2), "remult": (2.0, 0.9, 2.2, 1.1),
                    "ci": (1.9, 0.9, 2.1, 1.3)}
            unins, unrlz, fhlb = 69.0, 56.0, 22.0
        elif tier == 1:
            rate = {"crcd": (2.6, 1.7, 0.4, 3.1), "auto": (2.9, 0.6, 0.9, 1.2),
                    "conoth": (2.2, 1.2, 1.1, 2.2), "reres": (2.4, 1.1, 1.3, 0.3),
                    "reloc": (1.8, 0.9, 1.2, 0.0), "recons": (1.8, 0.8, 1.7, 0.7),
                    "renres": (1.2, 0.6, 1.3, 0.6), "remult": (1.1, 0.55, 1.2, 0.55),
                    "ci": (1.1, 0.5, 1.2, 0.7)}
            unins, unrlz, fhlb = 55.0, 33.0, 12.0
        else:
            rate = {
                "crcd": (0.9 + (s % 7) * 0.13, 0.7 + (s % 5) * 0.11,
                         0.1 + (s % 4) * 0.09, 1.2 + (s % 10) * 0.18),
                "auto": (1.0 + (s % 8) * 0.14, 0.15 + (s % 4) * 0.07,
                         0.2 + (s % 5) * 0.09, 0.3 + (s % 6) * 0.09),
                "conoth": (1.0 + (s % 7) * 0.12, 0.3 + (s % 5) * 0.09,
                           0.3 + (s % 4) * 0.11, 1.0 + (s % 8) * 0.11),
                "reres": (0.8 + (s % 6) * 0.13, 0.3 + (s % 4) * 0.1,
                          0.4 + (s % 5) * 0.09, 0.02 + (s % 5) * 0.02),
                "reloc": (0.5 + (s % 5) * 0.12, 0.2 + (s % 4) * 0.08,
                          0.3 + (s % 4) * 0.09, 0.0),
                "recons": (0.4 + (s % 6) * 0.12, 0.2 + (s % 4) * 0.09,
                           0.4 + (s % 6) * 0.12, 0.05 + (s % 5) * 0.06),
                "renres": (0.3 + (s % 5) * 0.11, 0.15 + (s % 4) * 0.07,
                           0.3 + (s % 5) * 0.11, 0.1 + (s % 4) * 0.07),
                "remult": (0.3 + (s % 4) * 0.11, 0.15 + (s % 3) * 0.07,
                           0.3 + (s % 4) * 0.1, 0.1 + (s % 3) * 0.07),
                "ci": (0.3 + (s % 5) * 0.1, 0.15 + (s % 4) * 0.06,
                       0.3 + (s % 5) * 0.1, 0.1 + (s % 5) * 0.07)}
            unins = 22.0 + (s % 11) * 2.0            # 22-42: WATCH edge only
            unrlz = 6.0 + (s % 10) * 2.0             # 6-24: under the 25 WATCH
            fhlb = 1.0 + (s % 9)                     # 1-9: under the 10 WATCH
        # class balance shares of total loans / assets
        cc_sh = 0.01 + (s % 6) * 0.01                # cards 1-6% of loans
        au_sh = 0.02 + (s % 5) * 0.01                # auto 2-6%
        oc_sh = 0.02 + (s % 4) * 0.01                # other consumer 2-5%
        rr_sh = 0.18 + (s % 8) * 0.02                # 1-4 family 18-32%
        ci_sh = 0.18 + (s % 9) * 0.02                # C&I 18-34%
        htm_sh = 0.10 + (s % 6) * 0.02               # HTM securities / assets
        afs_sh = 0.12 + (s % 7) * 0.02               # AFS securities / assets
        # F-trap vintages/coverage: one bank reports NO auto book (fields null
        # from 2011 only / not a lender) and one bank's DEPUNINS is null for
        # the whole series (RC-O Mem 2 is filed by $1B+ reporters only) --
        # blanks, never zeros (F3).
        auto_none = (s % 19) == 9
        depunins_none = (s % 23) == 7
        periods = self._periods()
        quarters = []
        for i, iso in enumerate(periods):            # i=0 oldest .. n-1 newest
            ramp = 0.7 + 0.3 * (i / max(1, n - 1))   # deterioration narrative
            wob = 1.0 + 0.04 * math.sin((i + s % 7) / 2.1)
            a = asset0 * ((1.009) ** i) * (1.0 + 0.008 * math.sin((i + s % 9) / 2.7))
            loans = loans_r * a
            ncl_i = ncl * ramp * wob
            res_i = res_r * (0.9 + 0.1 * ramp) * wob
            f: Dict[str, Optional[float]] = {}
            f["ASSET"] = round(a)
            f["DEP"] = round(dep_r * a)
            f["LNLSGR"] = round(loans)
            f["LNLSNET"] = round(loans * (1.0 - res_i / 100.0))
            f["BRO"] = None if bro_none else round(bro_r * dep_r * a)
            f["EQ"] = round(eq_r * a)
            f["NCLNLS"] = round(ncl_i / 100.0 * loans)
            f["LNATRES"] = round(res_i / 100.0 * loans)
            f["P3LNLS"] = round(p3 * ramp / 100.0 * loans)
            f["LNRECONS"] = round(cre_c * loans)
            f["LNRENRES"] = round(cre_n * loans)
            f["LNREMULT"] = round(cre_m * loans)
            f["NCLNLSR"] = round(ncl_i, 4)
            f["NTLNLSQR"] = round(nco * ramp * wob, 4)
            f["LNATRESR"] = round(res_i, 4)
            f["LNRESNCR"] = round(res_i / ncl_i * 100.0, 4)
            f["RBC1AAJ"] = round(eq_r * 100.0 - 0.3 + 0.2 * math.sin(i + s), 4)
            # CBLR-elector shape: risk-based ratio can be null (trap F3)
            f["RBCRWAJ"] = round(eq_r * 100.0 + 5.2 + 0.3 * math.sin(i + s), 4)
            f["EQV"] = round(eq_r * 100.0, 4)
            f["ROAQ"] = round(roaq * (0.85 + 0.15 * ramp) * wob, 4)
            f["NIMY"] = round(nimy * wob, 4)
            f["EEFFR"] = round(eeffr * wob, 4)
            # ---- v1.1 pack fields (internally consistent: dollar triples
            # derive from the SAME wobbled class rates the R-twins report,
            # so Python/Excel parity holds on the landed values) ----
            bal = {"crcd": round(cc_sh * loans), "auto": round(au_sh * loans),
                   "conoth": round(oc_sh * loans), "reres": round(rr_sh * loans),
                   "ci": round(ci_sh * loans),
                   "recons": f["LNRECONS"], "renres": f["LNRENRES"],
                   "remult": f["LNREMULT"]}
            f["LNCRCD"], f["LNAUTO"] = bal["crcd"], bal["auto"]
            f["LNCONOTH"], f["LNRERES"] = bal["conoth"], bal["reres"]
            f["LNCI"] = bal["ci"]

            def rw(cls, k):                          # ramped+wobbled class rate
                return rate[cls][k] * ramp * wob
            # verified R-twin rate fields land as percents directly
            for cls, pre in (("crcd", "CRCD"), ("auto", "AUTO"),
                             ("reres", "RERES"), ("reloc", "RELOC"),
                             ("ci", "CI")):
                f["P3" + pre + "R"] = round(rw(cls, 0), 4)
                f["P9" + pre + "R"] = round(rw(cls, 1), 4)
                f["NA" + pre + "R"] = round(rw(cls, 2), 4)
            # computed classes land the dollar triple off the same rates
            for cls, suf in (("conoth", "CONOTH"), ("recons", "RECONS"),
                             ("renres", "RENRES"), ("remult", "REMULT")):
                f["P3" + suf] = round(rw(cls, 0) / 100.0 * bal[cls])
                f["P9" + suf] = round(rw(cls, 1) / 100.0 * bal[cls])
                f["NA" + suf] = round(rw(cls, 2) / 100.0 * bal[cls])
            # quarterly NCO dollars: rate is ANNUALIZED, flow = rate/400 * bal
            for cls, fld in (("crcd", "NTCRCDQ"), ("auto", "NTAUTOQ"),
                             ("conoth", "NTCONOTQ"), ("reres", "NTRERESQ"),
                             ("recons", "NTRECONQ"), ("renres", "NTRENREQ"),
                             ("remult", "NTREMULQ"), ("ci", "NTCIQ")):
                f[fld] = round(rw(cls, 3) / 400.0 * bal[cls])
            if auto_none:                            # not an auto lender
                for fld in ("LNAUTO", "P3AUTOR", "P9AUTOR", "NAAUTOR",
                            "NTAUTOQ"):
                    f[fld] = None
            # SVB pack: uninsured share, HTM/AFS unrealized (60/40 split so
            # ((SCHA-SCHF)+(SCAA-SCAF))/(EQ+LNATRES) == the seeded target),
            # FHLB advances
            f["DEPUNINS"] = (None if depunins_none
                             else round(unins * ramp * wob / 100.0 * f["DEP"]))
            cushion = f["EQ"] + f["LNATRES"]
            loss = unrlz * ramp * wob / 100.0 * cushion
            f["SCHA"] = round(htm_sh * a)
            f["SCHF"] = round(htm_sh * a - 0.6 * loss)
            f["SCAA"] = round(afs_sh * a)
            f["SCAF"] = round(afs_sh * a - 0.4 * loss)
            f["OTHBFHLB"] = round(fhlb * ramp * wob / 100.0 * a)
            quarters.append((iso, f))
        # one interior None per (cert, field) series -- exercises the
        # missing-value path deterministically, at offsets >= 2 from newest so
        # the latest-quarter headline formulas never see it.
        n_q = len(quarters)
        if n_q >= 6:
            for j, fname in enumerate(RAW_FIELDS):
                k_from_new = 2 + (s + 7 * j) % (n_q - 3)
                quarters[n_q - 1 - k_from_new][1][fname] = None
        self._profiles[cert] = quarters
        return quarters

    def fetch_series(self, spec: FieldSpec, secret=None) -> List[NormalizedRow]:
        quarters = self._profile(spec.cert)          # oldest-first
        return [NormalizedRow(id=spec.id, period=iso, value=f[spec.fname],
                              geo_segment=spec.geo_segment,
                              source_class=self.source_class, units=spec.units)
                for iso, f in quarters]


class FdicProvider(Provider):
    """Live Class A provider: plain urllib REST against api.fdic.gov (BUILD
    SPEC 0.3 -- KEYLESS: the FDIC BankFind API needs no API key; swagger marks
    api_key optional and every observed client is keyless). ONE bulk
    /financials request per refresh covers the whole peer set (30 banks x
    ~24 fields x 16 quarters << the 10,000-row cap), plus one /institutions
    roster call for identity + merger detection (trap F2). Rate limit is
    UNDOCUMENTED -- be polite: min_interval throttle, 429/5xx backoff,
    per-URL cache."""

    source_class = "A"

    def __init__(self, min_interval: float = 0.6, max_retries: int = 4,
                 raw_slots: int = RAW_SLOTS_DEFAULT, limit: int = 10000):
        super().__init__()
        self._min_interval = float(min_interval)
        self._max_retries = int(max_retries)
        self._raw_slots = int(raw_slots)
        self._limit = int(limit)
        self._last = 0.0
        self._cache: Dict[str, bytes] = {}
        self._bulk: Optional[Dict[str, Dict[str, Dict[str, Optional[float]]]]] = None

    def _throttle(self):
        gap = time.time() - self._last
        if gap < self._min_interval:
            time.sleep(self._min_interval - gap)
        self._last = time.time()

    def _http_get(self, url: str) -> bytes:
        """The one real network call -- isolated so tests can stub it."""
        import urllib.request
        with urllib.request.urlopen(url, timeout=60) as resp:
            return resp.read()

    def _download(self, url: str, label: str) -> bytes:
        """Cache + throttle + backoff. 429/5xx retry with backoff; other HTTP
        codes fail fast (bad filter syntax / bad field name)."""
        if url in self._cache:                       # repeat fetch == cache hit
            return self._cache[url]
        import urllib.error
        for attempt in range(self._max_retries + 1):
            self._throttle()
            try:
                data = self._http_get(url)
                self._cache[url] = data
                return data
            except urllib.error.HTTPError as exc:
                if exc.code == 429 or exc.code >= 500:
                    if attempt < self._max_retries:
                        time.sleep(2.0 * (attempt + 1))
                        continue
                raise RuntimeError(f"FDIC fetch failed for {label}: "
                                   f"HTTP {exc.code}")
            except Exception as exc:
                if attempt < self._max_retries:
                    time.sleep(2.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"FDIC fetch failed for {label}: "
                                   f"{type(exc).__name__}")

    @staticmethod
    def _iso(repdte: str) -> str:
        r = str(repdte).strip()
        if len(r) == 8 and r.isdigit():              # "20260331" -> ISO
            return f"{r[0:4]}-{r[4:6]}-{r[6:8]}"
        return r

    @staticmethod
    def _oldest_repdte(asof: date, raw_slots: int) -> str:
        """Lower REPDTE bound: raw_slots quarters back from asof (fetch only
        what lands)."""
        months = raw_slots * 3
        y = asof.year - (months // 12)
        m = asof.month - (months % 12)
        if m < 1:
            m += 12
            y -= 1
        return f"{y:04d}{m:02d}01"

    def prime(self, certs, asof, names=None):
        """The ONE bulk /financials call + the /institutions roster call.
        Guard: meta.total > limit errors CLEARLY -- silent truncation would
        quietly drop whole banks (BUILD SPEC sec 1)."""
        certs = [str(c) for c in certs if str(c)]
        self._bulk = {}
        if not certs:
            return
        from urllib.parse import urlencode
        filt = ("CERT:(" + " OR ".join(certs) + ") AND REPDTE:["
                + self._oldest_repdte(asof, self._raw_slots) + " TO *]")
        field_list = ["CERT", "REPDTE"] + list(RAW_FIELDS)
        if len(field_list) >= MAX_REQUEST_FIELDS:    # spec sec 3 guard
            raise RuntimeError(
                f"fields= list has {len(field_list)} entries -- the FDIC "
                f"bulk endpoint caps at {MAX_REQUEST_FIELDS}.")
        url = FDIC_FIN_URL + "?" + urlencode({
            "filters": filt,
            "fields": ",".join(field_list),
            "sort_by": "REPDTE", "sort_order": "DESC",
            "limit": str(self._limit), "format": "json"})
        payload = json.loads(self._download(url, "financials bulk").decode("utf-8"))
        meta = payload.get("meta") or {}
        total = meta.get("total")
        if isinstance(total, dict):                  # ES-style nested total
            total = total.get("value")
        if total is not None and int(total) > self._limit:
            raise RuntimeError(
                f"FDIC bulk response reports {total} rows but the request "
                f"limit is {self._limit}: the peer set x quarter window "
                f"exceeds one page. Narrow raw_slots or the peer list -- "
                f"refusing silent truncation.")
        self.vintage = str((meta.get("index") or {}).get("createTimestamp")
                           or "")
        for rec in payload.get("data", []):
            d = rec.get("data") or {}                # DOUBLE WRAP: data[i].data
            cert = _norm_cert(d.get("CERT"))
            iso = self._iso(d.get("REPDTE", ""))
            if not cert or not iso:
                continue
            entry = {}
            for f in RAW_FIELDS:
                v = d.get(f)
                # JSON null -> None, NEVER zero (trap F3: CBLR risk-based
                # ratios and non-applicable items are genuinely null)
                entry[f] = None if v is None else float(v)
            self._bulk.setdefault(cert, {})[iso] = entry
        # roster: identity + merger detection (ACTIVE/ENDEFYMD -- trap F2)
        rurl = FDIC_INST_URL + "?" + urlencode({
            "filters": "CERT:(" + " OR ".join(certs) + ")",
            "fields": "CERT,NAME,FED_RSSD,ACTIVE,BKCLASS,ENDEFYMD",
            "limit": str(max(len(certs), 10)), "format": "json"})
        try:
            rp = json.loads(self._download(rurl, "institutions roster").decode("utf-8"))
            for rec in rp.get("data", []):
                d = rec.get("data") or {}
                c = _norm_cert(d.get("CERT"))
                if c:
                    self.roster[c] = d
        except Exception:
            # roster is advisory (merger notes); the financials landing and
            # the staleness guard still protect the run without it
            self.roster = self.roster or {}

    def fetch_series(self, spec: FieldSpec, secret=None) -> List[NormalizedRow]:
        if self._bulk is None:
            raise RuntimeError("FdicProvider.fetch_series before prime(): the "
                               "bulk request must run once per refresh.")
        quarters = self._bulk.get(spec.cert, {})
        rows = []
        for iso in sorted(quarters):                 # oldest-first, like demo
            rows.append(NormalizedRow(
                id=spec.id, period=iso, value=quarters[iso].get(spec.fname),
                geo_segment=spec.geo_segment, source_class=self.source_class,
                units=spec.units))
        return rows

    def lookup(self, name: str) -> List[dict]:
        """--lookup support (USER REQUIREMENT): fuzzy /institutions search so
        the user can find CERTs from PowerShell. LIVE-ONLY."""
        from urllib.parse import urlencode
        url = FDIC_INST_URL + "?" + urlencode({
            "search": name,
            "fields": "NAME,CERT,CITY,STALP,ASSET,ACTIVE",
            "limit": "15", "format": "json"})
        payload = json.loads(self._download(url, f'lookup "{name}"').decode("utf-8"))
        return [rec.get("data") or {} for rec in payload.get("data", [])]


class ClassCStubProvider(Provider):
    """In-process OAuth client_credentials STUB carried from the contract's
    seam rehearsal. Makes NO live request. NOTE: even a real Class C adapter
    is REFUSED by this template's gate 2 (only class "A" is admitted); a
    licensed swap requires its own review, not a config edit. secret_env
    stays wired per contract for a hypothetical authenticated v2."""

    source_class = "C"

    def __init__(self, secret_env: str):
        super().__init__()
        self.secret_env = secret_env
        self._token = None

    def _authenticate(self, secret):
        if not secret:
            raise PermissionError("HTTP 401: missing/invalid access token "
                                  "(client_credentials not supplied).")
        self._token = "stub-token"
        return self._token

    def prime(self, certs, asof, names=None):
        return None

    def fetch_series(self, spec: FieldSpec, secret=None) -> List[NormalizedRow]:
        self._authenticate(secret)
        # schema-shaped, VALUELESS row: proves the seam contract without
        # fabricating licensed data.
        return [NormalizedRow(id=spec.id, period="1900-01-01", value=None,
                              geo_segment=spec.geo_segment, source_class="C",
                              units=spec.units)]


def resolve_secret(cfg: Config) -> Optional[str]:
    """Class C secret from the env var whose NAME is recorded in _config;
    never hardcoded. Unused in v1 (the FDIC API is keyless)."""
    name = str(cfg.setting("secret_env", "") or "").strip()
    if not name:
        return None
    return os.environ.get(name)


def make_provider(cfg: Config, demo: bool, asof: Optional[date]) -> Provider:
    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        return FdicDemoProvider(asof=asof, raw_slots=cfg.raw_slots)
    # KEYLESS live provider -- no env var, no secret, nothing to fail fast on.
    return FdicProvider(
        min_interval=float(cfg.setting("http_min_interval", 0.6) or 0.6),
        max_retries=int(float(cfg.setting("fdic_max_retries", 4) or 4)),
        raw_slots=cfg.raw_slots)


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder. One block per SLOT:
# quarters as rows (newest-first), fields as columns.
# ---------------------------------------------------------------------------
@dataclass
class SlotBlock:
    slot: int
    header_row: int
    label_row: int
    first_data_row: int
    slots: int                 # quarter rows (raw_slots)


def slot_block(slot: int, raw_slots: int = RAW_SLOTS_DEFAULT) -> SlotBlock:
    """Deterministic block placement: anchors depend ONLY on (slot, raw_slots)
    -- editing which bank occupies the slot moves nothing."""
    stride = RAW_HEADER_ROWS + raw_slots + RAW_GAP_ROWS
    header_row = RAW_FIRST_ROW + (slot - 1) * stride
    return SlotBlock(slot, header_row, header_row + 1,
                     header_row + RAW_HEADER_ROWS, raw_slots)


def assemble_quarters(field_rows: Dict[str, List[NormalizedRow]],
                      raw_slots: int):
    """Per-field NormalizedRows -> newest-first [(iso, {field: value})] rows
    ready to land. Periods are unioned so a field missing one quarter still
    aligns (blank cell, not a shifted row)."""
    periods = sorted({r.period for rows in field_rows.values() for r in rows},
                     reverse=True)[:raw_slots]
    per: Dict[str, Dict[str, Optional[float]]] = {p: {} for p in periods}
    for fname, rows in field_rows.items():
        for r in rows:
            if r.period in per:
                per[r.period][fname] = r.value
    return [(p, {f: per[p].get(f) for f in RAW_FIELDS}) for p in periods]


# ---------------------------------------------------------------------------
# WRITE BACKEND -- openpyxl on the closed workbook (contract sec 4 / L2)
# ---------------------------------------------------------------------------
class OpenpyxlBackend:
    def __init__(self, path: str):
        self.path = path
        import openpyxl
        # L2: keep_vba=True ONLY for .xlsm; on .xlsx it injects a dangling
        # vbaProject relationship Excel rejects as "format/extension not valid".
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        return parse_config([[c.value for c in row] for row in ws.iter_rows()])

    def _check_slot(self, block: SlotBlock):
        """Refuse to write into a workbook whose raw layout doesn't match the
        computed one (raw_slots/peer_slots edited after build): every
        dashboard and Watchlist formula is anchored to the BUILT layout, so a
        silent mismatch would make the whole workbook quietly wrong. The slot
        label is constant regardless of which bank occupies the slot, so
        [PEERS] edits never trip this."""
        ws = self._wb[RAW_TAB]
        existing = ws.cell(block.header_row, 1).value
        if existing is not None and str(existing).strip() not in (
                "", slot_label(block.slot)):
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB} at row {block.header_row}: "
                f"expected block '{slot_label(block.slot)}' but found "
                f"'{existing}'. The workbook was built with a different "
                f"raw_slots/peer_slots layout than _config now describes; "
                f"formulas are anchored to the built layout. Rebuild the "
                f"workbook (python make_workbook.py --peer-slots N) instead "
                f"of editing the layout settings in place.")
        # v1.1 pack guard: the FIELD layout must match too -- a workbook built
        # by an earlier pack has fewer raw columns, and writing this runner's
        # field set into it would land data under no label with stale
        # formulas. The last field's label column is the cheap sentinel.
        lbl = ws.cell(block.label_row, field_col(RAW_FIELDS[-1])).value
        if str(lbl or "").strip() != RAW_FIELDS[-1]:
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB}: block "
                f"'{slot_label(block.slot)}' does not label field "
                f"'{RAW_FIELDS[-1]}' at column {field_col(RAW_FIELDS[-1])} "
                f"(found '{lbl}'). The workbook was built by a different "
                f"metric-pack version than this runner (pack {PACK_VERSION}); "
                f"rebuild it (python make_workbook.py) or extract the "
                f"matching runner from ITS OWN _code_py tab.")

    def clear_slot_block(self, block: SlotBlock):
        """Blank a slot's bank identity AND data (stateless rebuild, BUILD
        SPEC 0.2): a deactivated/removed/failed bank must show empty, never
        last run's values masquerading as current."""
        self._check_slot(block)
        ws = self._wb[RAW_TAB]
        # NOTE: ws.cell(r, c, None) is a silent NO-OP in openpyxl (the value
        # arg is ignored when None) -- clearing MUST assign .value directly,
        # or a deactivated bank's data would survive as if current.
        for c in (2, 3, 4):                          # runtime identity cells
            ws.cell(block.header_row, c).value = None
        last_col = 1 + len(RAW_FIELDS)
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            for c in range(1, last_col + 1):
                ws.cell(r, c).value = None

    def write_slot_block(self, block: SlotBlock, peer: PeerRow, quarters):
        self._check_slot(block)
        ws = self._wb[RAW_TAB]
        ws.cell(block.header_row, 1, slot_label(block.slot))
        ws.cell(block.header_row, 2,
                f"s{block.slot:02d} {peer.entity_key}")   # the unit key
        ws.cell(block.header_row, 3, peer.name)
        ws.cell(block.header_row, 4, f"group={peer.group}")
        for i, (iso, fvals) in enumerate(quarters[:block.slots]):
            r = block.first_data_row + i                 # newest-first
            ws.cell(r, 1, iso)
            for fname in RAW_FIELDS:
                v = fvals.get(fname)
                # explicit .value assignment: None must LAND as blank (the
                # ws.cell value arg silently drops None -- see clear above)
                ws.cell(r, field_col(fname)).value = \
                    None if v is None else float(v)

    def write_status(self, status: dict):
        line2 = (f"Banks {status.get('banks_landed', 0)}/"
                 f"{status.get('banks_active', 0)} - "
                 f"{status.get('alert_banks', 0)} ALERT / "
                 f"{status.get('watch_banks', 0)} WATCH / "
                 f"{status.get('stale_banks', 0)} STALE")
        n_ref = len(status.get("watchlist_refusals") or [])
        if n_ref:
            line2 += f" - {n_ref} REFUSED (see digest)"
        n_err = len(status.get("errors") or [])
        if n_err:
            line2 += f" - {n_err} FETCH ERRORS (slots blanked)"
        line3 = f"FDIC data vintage: {status.get('vintage') or 'n/a'}"
        for tab in DASH_TABS:
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                col = STATUS_COL_BY_TAB.get(tab, STATUS_COL)
                ws.cell(1, col, "Last run  " + status.get("timestamp", "")
                        + f" ({status.get('mode', '')})")
                ws.cell(2, col, line2)
                ws.cell(3, col, line3)

    def finalize(self):
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def compute_digest(cfg: Config, landed: Dict[int, Tuple[PeerRow, list]],
                   roster: Dict[str, dict]) -> dict:
    """Per-bank latest-quarter metric values + statuses + flag counts (the
    engine output consumed by the Watchlist parity tests and email-sim).
    Staleness (BUILD SPEC 0.6) is first-class: a stale bank's STATUS is
    forced to STALE and it never counts toward the alert/watch KPIs or the
    digest medians. NOTE the documented v1 divergence: the EXCEL median rows
    include stale banks' landed values (MEDIAN cannot see runtime staleness);
    the digest median excludes them -- see _readme."""
    per_bank_latest: Dict[int, Optional[str]] = {}
    for slot, (peer, quarters) in landed.items():
        last_p = next((p for p, f in quarters
                       if any(v is not None for v in f.values())), None)
        per_bank_latest[slot] = last_p
    set_max = max((p for p in per_bank_latest.values() if p), default=None)

    banks = []
    for slot in sorted(landed):
        peer, quarters = landed[slot]
        latest_fields = quarters[0][1] if quarters else {}
        last_p = per_bank_latest[slot]
        stale = is_stale_bank(last_p, set_max, cfg.stale_multiplier)
        metrics = {}
        alert_n = watch_n = 0
        for spec in cfg.series:
            v = metric_value(spec.id, latest_fields)
            st = status_for(v, cfg.thresholds.get(spec.id))
            if v is None:
                st = ""                       # blank cell -> blank status (Excel parity)
            metrics[spec.id] = {"value": v, "status": st,
                                "dimension": spec.category}
            # flag counts mirror the Watchlist COUNTIF columns exactly --
            # computed for every landed bank (incl. stale; the top-level KPIs
            # below are where staleness excludes)
            if st == "ALERT":
                alert_n += 1
            elif st == "WATCH":
                watch_n += 1
        notes = []
        # trap F6: merger survivorship -- flag >25% single-quarter asset jump
        if len(quarters) >= 2:
            a0 = quarters[0][1].get("ASSET")
            a1 = quarters[1][1].get("ASSET")
            if a0 is not None and a1 not in (None, 0):
                jump = (a0 / a1 - 1.0) * 100.0
                if jump > 25.0:
                    notes.append(f"ASSET +{jump:.0f}% in one quarter -- "
                                 "possible merger survivorship distortion "
                                 "(trap F6); ratios and CRE growth screens "
                                 "unreliable this quarter.")
        # SVB pack: DEPUNINS null -> BLANK + note, never zero (trap F3; RC-O
        # Mem 2 is filed only by $1B+ reporters -- below that the API carries
        # estimates or nulls). The note makes the blank auditable.
        if (latest_fields and latest_fields.get("DEPUNINS") is None
                and any(v is not None for v in latest_fields.values())):
            notes.append("DEPUNINS is null this quarter -- uninsured-deposit "
                         "share left BLANK, never 0 (RC-O Mem 2 is filed by "
                         "$1B+ reporters; smaller banks carry FDIC estimates "
                         "or nulls).")
        rr = roster.get(peer.cert) or {}
        if rr and str(rr.get("ACTIVE", "1")).strip() not in ("1", "True", "true"):
            notes.append(f"roster reports ACTIVE={rr.get('ACTIVE')} "
                         f"(ENDEFYMD={rr.get('ENDEFYMD')}) -- possible "
                         "merger/closure; check /institutions and /history.")
        if stale:
            notes.append("latest REPDTE lags the peer-set max by more than "
                         f"{cfg.stale_multiplier:g} quarters -- possible "
                         "merger/closure -- check /institutions ACTIVE, "
                         "/history (trap F2).")
        status = ("STALE" if stale else
                  "ALERT" if alert_n > 0 else
                  "WATCH" if watch_n > 0 else "OK")
        banks.append({
            "slot": slot, "id_prefix": f"s{slot:02d}", "cert": peer.cert,
            "name": peer.name, "group": peer.group,
            "entity_key": peer.entity_key, "asof_period": last_p,
            "stale": stale, "status": status,
            "alert_count": alert_n, "watch_count": watch_n,
            "texas": metrics.get("TEXAS", {}).get("value"),
            "metrics": metrics, "notes": notes,
        })

    # digest peer medians EXCLUDE stale banks (spec 0.6); Excel medians can't
    # -- the honest v1 divergence, documented in _readme.
    medians = {}
    for spec in cfg.series:
        vals = sorted(b["metrics"][spec.id]["value"] for b in banks
                      if not b["stale"]
                      and b["metrics"][spec.id]["value"] is not None)
        if not vals:
            medians[spec.id] = None
        else:
            k = len(vals)
            medians[spec.id] = (vals[k // 2] if k % 2 else
                                (vals[k // 2 - 1] + vals[k // 2]) / 2.0)
    return {"banks": banks, "medians": medians, "set_max_repdte": set_max}


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None,
        provider: Optional[Provider] = None) -> dict:
    """One refresh against the CLOSED workbook. `provider` is injectable for
    tests (e.g. the frozen-bank staleness regression)."""
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any fetch (BUILD SPEC 0.1 / sec 3).
    validate_metrics(cfg.series)
    assert_metrics_admissible(cfg.series)
    validate_peer_capacity(cfg)
    admitted, refusals, excluded = evaluate_peers(cfg)

    if provider is None:
        provider = make_provider(cfg, demo, asof)
    mode = "demo" if isinstance(provider, FdicDemoProvider) else "live"

    # Stateless rebuild (BUILD SPEC 0.2): blank EVERY slot 1..capacity first
    # -- deactivated/refused/emptied slots show blank under the fresh
    # timestamp, never last run's bank. This also validates the workbook's
    # raw layout against _config before any fetch.
    for slot in range(1, cfg.peer_slots + 1):
        backend.clear_slot_block(slot_block(slot, cfg.raw_slots))

    errors: List[str] = []
    landed: Dict[int, Tuple[PeerRow, list]] = {}
    try:
        provider.prime([p.cert for p in admitted], asof,
                       names={p.cert: p.name for p in admitted})
    except Exception as exc:
        errors.append(f"bulk fetch: {exc}")
    else:
        for peer in admitted:
            try:
                field_rows = {f: provider.fetch_series(make_field_spec(peer, f))
                              for f in RAW_FIELDS}
                quarters = assemble_quarters(field_rows, cfg.raw_slots)
            except Exception as exc:
                errors.append(f"s{peer.slot:02d} cert:{peer.cert}: {exc}")
                continue
            backend.write_slot_block(slot_block(peer.slot, cfg.raw_slots),
                                     peer, quarters)
            landed[peer.slot] = (peer, quarters)

    digest = compute_digest(cfg, landed, provider.roster)
    fresh = [b for b in digest["banks"] if not b["stale"]]
    status = {
        "timestamp": asof.isoformat(),
        "mode": mode,
        "pack_version": PACK_VERSION,
        "peer_slots": cfg.peer_slots,
        "banks_active": len(admitted) + len(refusals),
        "banks_landed": len(landed),
        "banks_excluded": len(excluded),
        # staleness excludes a bank from every alert KPI (BUILD SPEC 0.6)
        "alert_banks": sum(1 for b in fresh if b["status"] == "ALERT"),
        "watch_banks": sum(1 for b in fresh if b["status"] == "WATCH"),
        "stale_banks": sum(1 for b in digest["banks"] if b["stale"]),
        "alert_flags": sum(b["alert_count"] for b in fresh),
        "watch_flags": sum(b["watch_count"] for b in fresh),
        "vintage": provider.vintage,
        "digest": digest,
        "watchlist_refusals": [m for _, m in refusals],
        "watchlist_admitted": [f"s{p.slot:02d} {p.entity_key}" for p in admitted],
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


# ---------------------------------------------------------------------------
# CLI (contract sec 4) + --lookup (USER REQUIREMENT, live-only)
# ---------------------------------------------------------------------------
def do_lookup(name: str, demo: bool) -> int:
    if demo:
        sys.stderr.write(
            "--lookup is LIVE-ONLY: it queries the FDIC BankFind API "
            "(https://api.fdic.gov/banks/institutions) and cannot run in "
            "--demo/offline mode. Run it without --demo -- the API is "
            "keyless, no API key or account is needed:\n"
            '    python runner.py --lookup "' + name + '"\n')
        return 2
    prov = FdicProvider()
    try:
        hits = prov.lookup(name)
    except Exception as exc:
        sys.stderr.write(f"lookup failed: {exc}\n"
                         "Check network/proxy access to api.fdic.gov "
                         "(keyless; no API key needed).\n")
        return 1
    if not hits:
        print(f'No institutions matched "{name}".')
        return 0
    print(f'{"CERT":>8}  {"ACTIVE":>6}  {"ASSET($000)":>14}  '
          f'{"CITY":<20} {"ST":<3} NAME')
    for d in hits:
        asset = d.get("ASSET")
        asset_s = "" if asset is None else f"{float(asset):,.0f}"
        print(f'{str(d.get("CERT", "")):>8}  {str(d.get("ACTIVE", "")):>6}  '
              f'{asset_s:>14}  {str(d.get("CITY", "")):<20} '
              f'{str(d.get("STALP", "")):<3} {d.get("NAME", "")}')
    print("\nPut the CERT into the _config [PEERS] table "
          "(slot | cert | name | group | active) and re-run the runner -- "
          "no rebuild needed within the built slot capacity.")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Bank Counterparty & Peer Monitor runner (FDIC BankFind)")
    ap.add_argument("--workbook", "-w", default=None,
                    help="path to the .xlsm (required unless --lookup)")
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline FdicDemoProvider "
                         "(no network; the live API is keyless anyway)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    ap.add_argument("--lookup", default=None, metavar="NAME",
                    help='find a bank\'s CERT by name (live-only): '
                         '--lookup "Wells Fargo"')
    args = ap.parse_args(argv)
    if args.lookup is not None:
        return do_lookup(args.lookup, args.demo)
    if not args.workbook:
        sys.stderr.write("--workbook is required (or use --lookup).\n")
        return 2
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, MetricError, PeerCapacityError) as exc:
        sys.stderr.write(f"GATE ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except SystemExit as exc:
        sys.stderr.write(f"{exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True,
                      **{k: v for k, v in status.items() if k != "digest"}}))
    sys.stderr.write(
        f"OK ({status['mode']}): {status['banks_landed']}/"
        f"{status['banks_active']} banks landed, "
        f"{status['alert_banks']} ALERT / {status['watch_banks']} WATCH / "
        f"{status['stale_banks']} STALE banks "
        f"({status['alert_flags']} ALERT flags), "
        f"refused={len(status['watchlist_refusals'])}, "
        f"vintage={status['vintage']}.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
