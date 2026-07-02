#!/usr/bin/env python3
"""Email-simulate acceptance test (BUILD_SPEC_FDIC.md Phase 9, --demo).

Copies ONLY the .xlsm into a fresh, otherwise-empty folder, then reproduces
exactly what the VBA "Extract" button does:
  1. read the _code_py tab (one source line per cell, column A) and write
     runner.py next to the workbook -- byte-for-byte the macro's WriteTabToFile;
  2. shell Python to run the EXTRACTED runner.py against the CLOSED workbook in
     --demo mode (deterministic, offline -- the live API is keyless anyway).
Then it composes (does NOT send) the monitoring email from the run digest and
asserts the email contains the RANKED PEER TABLE (bank names + Texas + flag
counts), the PER-DIMENSION alert lines, the STALENESS section, and the
DATA-VINTAGE line, and that the workbook rebuilt itself with nothing but the
workbook + extracted runner present (the workbook is the source of truth).
Deterministic from _config + FdicDemoProvider at fixed --asof.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

import runner as R

XLSM_NAME = "Bank_Peer_Monitor.xlsm"
ASOF = "2026-03-31"

DIMENSIONS = ["asset_quality", "composite", "capital", "earnings", "funding",
              "concentration"]


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write UTF-8."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def compose_email(status):
    """Deterministic monitoring email from the run status/digest. No send."""
    digest = status.get("digest", {})
    banks = digest.get("banks", [])
    fresh = [b for b in banks if not b.get("stale")]
    stale = [b for b in banks if b.get("stale")]
    lines = []
    lines.append("Subject: Bank Peer Monitor -- "
                 f"{status['alert_banks']} ALERT / {status['watch_banks']} "
                 f"WATCH / {status['stale_banks']} STALE banks "
                 f"({status['timestamp']})")
    lines.append("")
    lines.append(f"Run mode: {status['mode']}  |  banks landed: "
                 f"{status['banks_landed']}/{status['banks_active']}  |  "
                 f"refused: {len(status.get('watchlist_refusals', []))}")
    # the data-vintage line (meta.index.createTimestamp on live runs)
    lines.append(f"Data vintage: {status.get('vintage') or 'n/a'}  |  "
                 f"peer-set latest REPDTE: {digest.get('set_max_repdte')}")
    lines.append("")
    lines.append("RANKED PEER TABLE (by ALERT flags, then Texas; "
                 "stale banks excluded)")
    lines.append("-" * 68)
    lines.append(f"  {'#':>2} {'Bank':<28} {'CERT':>7} {'Group':<13} "
                 f"{'Texas':>7} {'A':>2} {'W':>2}  Status")
    ranked = sorted(fresh, key=lambda b: (-(b["alert_count"]),
                                          -(b["texas"] or 0.0)))
    for i, b in enumerate(ranked, start=1):
        tex = "n/a" if b["texas"] is None else f"{b['texas']:.1f}"
        lines.append(f"  {i:>2} {b['name']:<28} {b['cert']:>7} "
                     f"{b['group']:<13} {tex:>7} {b['alert_count']:>2} "
                     f"{b['watch_count']:>2}  {b['status']}")
    if not ranked:
        lines.append("  (no banks landed)")
    lines.append("")
    lines.append("PER-DIMENSION ALERTS (metric-level, non-stale banks)")
    lines.append("-" * 52)
    any_dim = False
    for dim in DIMENSIONS:
        hits = []
        for b in fresh:
            for mid, m in b["metrics"].items():
                if m["dimension"] == dim and m["status"] == "ALERT":
                    hits.append(f"{b['name']} {mid}={m['value']:.2f}")
        if hits:
            any_dim = True
            lines.append(f"  {dim:<14} " + "; ".join(sorted(hits)))
    if not any_dim:
        lines.append("  (no metric-level alerts)")
    lines.append("")
    lines.append("STALENESS FLAGS (excluded from alert counts -- possible "
                 "merger/closure)")
    lines.append("-" * 68)
    if stale:
        for b in sorted(stale, key=lambda x: x["slot"]):
            lines.append(f"  STALE  s{b['slot']:02d} {b['name']:<28} "
                         f"last REPDTE {b['asof_period']}")
            for n in b.get("notes", []):
                lines.append(f"         {n}")
    else:
        lines.append("  (no stale banks -- every bank reports the peer-set "
                     "latest quarter)")
    lines.append("")
    lines.append("WATCHLIST LANE")
    lines.append("--------------")
    refusals = status.get("watchlist_refusals", [])
    if refusals:
        for m in refusals:
            lines.append("  " + m)
    else:
        lines.append(f"  ADMITTED (CERT-keyed Class A): "
                     f"{len(status.get('watchlist_admitted', []))} banks")
    notes = [n for b in fresh for n in b.get("notes", [])]
    if notes:
        lines.append("")
        lines.append("NOTES (survivorship / roster)")
        for n in notes:
            lines.append("  " + n)
    lines.append("")
    return "\n".join(lines)


def main():
    repo = os.path.dirname(os.path.abspath(__file__))
    xlsm = os.path.join(repo, XLSM_NAME)
    if not os.path.exists(xlsm):
        print("build the workbook first: python3 make_workbook.py", file=sys.stderr)
        return 1

    work = tempfile.mkdtemp(prefix="fdic_emailsim_")
    try:
        # 1. The email contains ONLY the workbook.
        dst_xlsm = os.path.join(work, XLSM_NAME)
        shutil.copy(xlsm, dst_xlsm)
        before = sorted(os.listdir(work))
        assert before == [XLSM_NAME], before
        print(f"[email-sim] fresh folder contains only: {before}")

        # 2. Button step 1: extract runner.py from _code_py.
        runner_py = os.path.join(work, "runner.py")
        extract_code_tab(dst_xlsm, "_code_py", runner_py)
        print(f"[email-sim] extracted runner.py ({os.path.getsize(runner_py)} "
              f"bytes) from _code_py")

        # 3. Button step 2: shell Python to run the EXTRACTED runner (demo).
        proc = subprocess.run(
            [sys.executable, runner_py, "--workbook", dst_xlsm, "--demo",
             "--asof", ASOF],
            capture_output=True, text=True, cwd=work)
        last = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() else "(none)"
        print("[email-sim] runner stderr:", last)
        if proc.returncode != 0:
            print("[email-sim] FAILED: runner exited", proc.returncode)
            print(proc.stdout, proc.stderr)
            return 2

        # 4. Verify the rebuild: raw data present, macro intact, nothing extra.
        wb = openpyxl.load_workbook(dst_xlsm, keep_vba=True)
        b1 = R.slot_block(1, 16)
        raw_ok = wb[R.RAW_TAB].cell(b1.first_data_row, 2).value is not None
        macro_ok = wb.vba_archive is not None
        produced = sorted(os.listdir(work))
        print(f"[email-sim] folder now: {produced}")
        print(f"[email-sim] raw populated={raw_ok}  macro intact={macro_ok}")

        # 5. Compose the monitoring email from the digest (deterministic).
        status = R.run(dst_xlsm, demo=True,
                       asof=R.datetime.strptime(ASOF, "%Y-%m-%d").date())
        email = compose_email(status)
        email_path = os.path.join(work, "monitoring_email.txt")
        with open(email_path, "w", encoding="utf-8") as fh:
            fh.write(email)
        print("\n" + "=" * 72)
        print(email)
        print("=" * 72 + "\n")

        n_banks = status["banks_landed"]
        table_lines = [ln for ln in email.splitlines()
                       if ln.strip() and ln.strip()[0].isdigit()
                       and ("peer" in ln or "counterparty" in ln or "self" in ln)]
        has_table = ("RANKED PEER TABLE" in email
                     and len(table_lines) == n_banks
                     and "ALERT" in email and "Texas" in email)
        has_dimension = ("PER-DIMENSION ALERTS" in email
                         and any(d in email for d in ("composite", "capital")))
        has_staleness = "STALENESS FLAGS" in email
        has_vintage = "Data vintage:" in email
        self_contained = (raw_ok and macro_ok
                          and set(produced) <= {XLSM_NAME, "runner.py"})

        ok = (has_table and has_dimension and has_staleness and has_vintage
              and self_contained)
        print(f"[email-sim] ranked peer table      : {has_table} "
              f"({len(table_lines)} rows)")
        print(f"[email-sim] per-dimension alerts   : {has_dimension}")
        print(f"[email-sim] staleness section      : {has_staleness}")
        print(f"[email-sim] data-vintage line      : {has_vintage}")
        print(f"[email-sim] workbook self-contained: {self_contained}")
        print("[email-sim] RESULT:",
              "PASS -- workbook is the source of truth" if ok else "FAIL")
        return 0 if ok else 3
    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
