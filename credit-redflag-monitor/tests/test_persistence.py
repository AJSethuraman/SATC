"""Disposition persistence -- the highest-risk piece (spec section 7.3).

Acceptance: run twice, enter dummy dispositions between runs, confirm they
survive the second run.
"""

from __future__ import annotations

from openpyxl import load_workbook

from redflag_monitor.excel_writer import (
    FLAGS_COLUMNS,
    SHEET_DICT,
    SHEET_FLAGS,
    SHEET_INTERNAL,
    SHEET_NEWS,
    write_workbook,
)
from redflag_monitor.seed import seed_signals


def _data_row(series_id="DGS10", as_of="2024-06-05", current=4.22, flag="Y"):
    return {
        "Series ID": series_id,
        "Signal": "10-Yr Treasury",
        "Category": "Rate",
        "Current": current,
        "As-Of": as_of,
        "Prior": 4.10,
        "Δ Abs": round(current - 4.10, 4),
        "Δ %": 2.9,
        "Direction": "both",
        "Threshold": "abs_change 0.5 (both)",
        "Auto-Flag": flag,
    }


def _col_index(name):
    return FLAGS_COLUMNS.index(name) + 1


def _find_row(ws, series_id, as_of):
    sid_c = _col_index("Series ID")
    asof_c = _col_index("As-Of")
    for r in range(2, ws.max_row + 1):
        if (
            str(ws.cell(row=r, column=sid_c).value) == series_id
            and str(ws.cell(row=r, column=asof_c).value) == as_of
        ):
            return r
    return None


def _set_disposition(path, series_id, as_of, matters, note, owner="Pat", reviewed="2024-06-10"):
    wb = load_workbook(path)
    ws = wb[SHEET_FLAGS]
    r = _find_row(ws, series_id, as_of)
    assert r is not None
    ws.cell(row=r, column=_col_index("Matters? (Y/N)"), value=matters)
    ws.cell(row=r, column=_col_index("Disposition / Notes"), value=note)
    ws.cell(row=r, column=_col_index("Owner"), value=owner)
    ws.cell(row=r, column=_col_index("Reviewed"), value=reviewed)
    wb.save(path)


def test_dispositions_survive_second_run(tmp_path):
    path = tmp_path / "wb.xlsx"
    signals = seed_signals()

    write_workbook(path, flag_rows=[_data_row()], history_rows=[], signals=signals)
    _set_disposition(path, "DGS10", "2024-06-05", "Y", "watch closely")

    # Second run: same period, data column changes (Current moved).
    write_workbook(path, flag_rows=[_data_row(current=4.30)], history_rows=[], signals=signals)

    wb = load_workbook(path)
    ws = wb[SHEET_FLAGS]
    r = _find_row(ws, "DGS10", "2024-06-05")
    assert ws.cell(row=r, column=_col_index("Matters? (Y/N)")).value == "Y"
    assert ws.cell(row=r, column=_col_index("Disposition / Notes")).value == "watch closely"
    assert ws.cell(row=r, column=_col_index("Owner")).value == "Pat"
    # Data column updated, human columns preserved.
    assert ws.cell(row=r, column=_col_index("Current")).value == 4.30


def test_new_period_appends_and_keeps_old_disposition(tmp_path):
    path = tmp_path / "wb.xlsx"
    signals = seed_signals()

    write_workbook(path, flag_rows=[_data_row(as_of="2024-06-05")], history_rows=[], signals=signals)
    _set_disposition(path, "DGS10", "2024-06-05", "Y", "june view")

    # Next month: a new observation period for the same signal.
    write_workbook(path, flag_rows=[_data_row(as_of="2024-07-05", current=4.5)], history_rows=[], signals=signals)

    wb = load_workbook(path)
    ws = wb[SHEET_FLAGS]
    june = _find_row(ws, "DGS10", "2024-06-05")
    july = _find_row(ws, "DGS10", "2024-07-05")
    assert june is not None and july is not None
    # Old disposition retained on its own row.
    assert ws.cell(row=june, column=_col_index("Disposition / Notes")).value == "june view"
    # New period starts blank.
    assert ws.cell(row=july, column=_col_index("Matters? (Y/N)")).value in (None, "")


def test_dictionary_sheet_not_clobbered(tmp_path):
    path = tmp_path / "wb.xlsx"
    signals = seed_signals()
    write_workbook(path, flag_rows=[_data_row()], history_rows=[], signals=signals)

    # Owner tunes a threshold in the dictionary sheet.
    wb = load_workbook(path)
    ws = wb[SHEET_DICT]
    ws.cell(row=2, column=1, value="TUNED")  # mutate first cell as a sentinel
    wb.save(path)

    write_workbook(path, flag_rows=[_data_row()], history_rows=[], signals=signals)
    wb = load_workbook(path)
    assert wb[SHEET_DICT].cell(row=2, column=1).value == "TUNED"


def test_all_sheets_present(tmp_path):
    path = tmp_path / "wb.xlsx"
    write_workbook(path, flag_rows=[_data_row()], history_rows=[], signals=seed_signals())
    wb = load_workbook(path)
    for name in (SHEET_FLAGS, "All Signals (History)", SHEET_DICT, SHEET_INTERNAL, SHEET_NEWS):
        assert name in wb.sheetnames
    # Flags is first.
    assert wb.sheetnames[0] == SHEET_FLAGS


def test_matters_dropdown_validation_present(tmp_path):
    path = tmp_path / "wb.xlsx"
    write_workbook(path, flag_rows=[_data_row()], history_rows=[], signals=seed_signals())
    wb = load_workbook(path)
    ws = wb[SHEET_FLAGS]
    validations = list(ws.data_validations.dataValidation)
    assert any('"Y,N"' in (dv.formula1 or "") for dv in validations)
