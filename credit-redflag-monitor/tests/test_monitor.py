"""End-to-end run with the deterministic demo client."""

from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from redflag_monitor.demo import DemoFredClient
from redflag_monitor.excel_writer import SHEET_FLAGS, read_dictionary_signals
from redflag_monitor.history import read_history
from redflag_monitor.monitor import resolve_signals, run
from redflag_monitor.seed import seed_signals


def _run(tmp_path, run_date="2026-06-19"):
    wb_path = tmp_path / "wb.xlsx"
    hist_path = tmp_path / "hist.csv"
    signals = resolve_signals(wb_path)
    fetcher = DemoFredClient(signals, today=date.fromisoformat(run_date))
    summary = run(
        fetcher,
        workbook_path=wb_path,
        history_path=hist_path,
        run_date=run_date,
    )
    return summary, wb_path, hist_path


def test_end_to_end_produces_workbook_and_flags(tmp_path):
    summary, wb_path, hist_path = _run(tmp_path)
    assert wb_path.exists()
    assert summary.n_signals == 14
    # Demo engineers DGS10 and UNRATE to breach; T10Y2Y is inverted (level_below).
    assert summary.n_flagged >= 2
    assert "10-Yr Treasury" in summary.flagged_labels


def test_history_is_appended_each_run(tmp_path):
    _run(tmp_path, run_date="2026-05-15")
    _, _, hist_path = _run(tmp_path, run_date="2026-06-19")
    rows = read_history(hist_path)
    run_dates = {r["run_date"] for r in rows}
    assert run_dates == {"2026-05-15", "2026-06-19"}
    # Two runs * 14 signals (TERMCBCCALLNS has valid quarterly data).
    assert len(rows) >= 24


def test_workbook_dictionary_round_trips(tmp_path):
    _, wb_path, _ = _run(tmp_path)
    # After the first run the dictionary sheet exists and resolve_signals reads it.
    loaded = read_dictionary_signals(wb_path)
    assert {s.series_id for s in loaded} == {s.series_id for s in seed_signals()}


def test_termcbccallns_handled_without_blank_current(tmp_path):
    _, wb_path, _ = _run(tmp_path)
    wb = load_workbook(wb_path)
    ws = wb[SHEET_FLAGS]
    # Find TERMCBCCALLNS row; Current must be a real number despite "." gaps.
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "TERMCBCCALLNS":
            current = ws.cell(row=r, column=4).value
            assert isinstance(current, (int, float))
            break
    else:
        raise AssertionError("TERMCBCCALLNS row not found")
