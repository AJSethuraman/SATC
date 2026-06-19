"""Excel writer with disposition persistence (spec sections 7.3 and 8).

The workbook is the disposition surface. The single hardest requirement is that
when the monthly run regenerates the workbook, the team's prior
``Matters? / Disposition / Owner / Reviewed`` entries carry forward. We key each
disposition row by ``(series_id, observation_period)``: on re-run we match
existing rows, preserve the human columns, update only the data columns, and
append new rows for new periods. We never clobber a human cell.

Sheets:
    Flags                  -- one row per (signal, observation period); dispositioned
    All Signals (History)  -- the append-only time-series log
    Signal Dictionary      -- the editable config (owner tunes here)
    Internal (paste)       -- empty, same shape as Flags
    News (manual)          -- optional manual-paste stub
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from redflag_monitor.config import (
    DICTIONARY_COLUMNS,
    Signal,
    signals_from_rows,
)
from redflag_monitor.history import HISTORY_COLUMNS

SHEET_FLAGS = "Flags"
SHEET_HISTORY = "All Signals (History)"
SHEET_DICT = "Signal Dictionary"
SHEET_INTERNAL = "Internal (paste)"
SHEET_NEWS = "News (manual)"

# Column layout for the Flags / Internal sheets.
DATA_COLUMNS = [
    "Series ID",
    "Signal",
    "Category",
    "Current",
    "As-Of",
    "Prior",
    "Δ Abs",  # Δ Abs
    "Δ %",  # Δ %
    "Direction",
    "Threshold",
    "Auto-Flag",
]
HUMAN_COLUMNS = [
    "Matters? (Y/N)",
    "Disposition / Notes",
    "Owner",
    "Reviewed",
]
FLAGS_COLUMNS = DATA_COLUMNS + HUMAN_COLUMNS

# Keying for disposition persistence.
KEY_COLUMNS = ("Series ID", "As-Of")

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_FLAG_FILL = PatternFill("solid", fgColor="F8CBAD")  # soft red for Auto-Flag = Y
_NUMERIC_HEADERS = {"Current", "Prior", "Δ Abs", "Δ %", "threshold_value"}


def has_dictionary(path: str | Path) -> bool:
    """Whether the workbook already carries an owner-editable dictionary sheet."""
    path = Path(path)
    if not path.exists():
        return False
    wb = load_workbook(path, read_only=True)
    try:
        return SHEET_DICT in wb.sheetnames
    finally:
        wb.close()


def read_dictionary_signals(path: str | Path) -> list[Signal]:
    """Load the Signal Dictionary from the workbook sheet (owner's source of truth)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_DICT]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            return []
        records = [dict(zip(header, raw)) for raw in rows]
    finally:
        wb.close()
    return signals_from_rows(records)


def _row_key(row: dict[str, Any]) -> tuple[str, str]:
    return (str(row.get(KEY_COLUMNS[0], "")), str(row.get(KEY_COLUMNS[1], "")))


def _style_header(ws: Worksheet, columns: list[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, columns: list[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(40, len(name) + 6))


def _read_existing_flags(ws: Worksheet) -> dict[tuple[str, str], dict[str, Any]]:
    """Read an existing Flags sheet into a key -> full-row-dict map.

    Matching is by header name (not position) so a hand-reordered sheet still
    resolves. This is what lets human dispositions survive a re-run.
    """
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return {}
    name_to_idx = {str(h): i for i, h in enumerate(header) if h is not None}

    existing: dict[tuple[str, str], dict[str, Any]] = {}
    for raw in rows:
        record: dict[str, Any] = {}
        for name, i in name_to_idx.items():
            record[name] = raw[i] if i < len(raw) else None
        key = _row_key(record)
        if key == ("", ""):
            continue
        existing[key] = record
    return existing


def _merge_flags(
    existing: dict[tuple[str, str], dict[str, Any]],
    current_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Merge current-run data rows with existing rows, preserving human columns.

    - Existing periods stay (their human + data columns are retained).
    - For a current row whose key already exists, refresh the data columns but
      keep the human columns untouched.
    - New keys are appended with blank human columns.
    """
    merged: dict[tuple[str, str], dict[str, Any]] = {}

    # Seed with everything already in the sheet (prior months keep their data).
    for key, record in existing.items():
        merged[key] = {col: record.get(col) for col in FLAGS_COLUMNS}

    for row in current_rows:
        key = _row_key(row)
        if key in merged:
            preserved = {col: merged[key].get(col) for col in HUMAN_COLUMNS}
            updated = {col: row.get(col) for col in DATA_COLUMNS}
            updated.update(preserved)
            merged[key] = updated
        else:
            new_row = {col: row.get(col) for col in DATA_COLUMNS}
            for col in HUMAN_COLUMNS:
                new_row[col] = None
            merged[key] = new_row

    def sort_key(item: tuple[tuple[str, str], dict[str, Any]]):
        record = item[1]
        return (
            str(record.get("Category", "")),
            str(record.get("Signal", "")),
            str(record.get("As-Of", "")),
        )

    return [record for _, record in sorted(merged.items(), key=sort_key)]


def _write_flags_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """(Re)write the Flags sheet body, then apply validation + formatting."""
    if SHEET_FLAGS in wb.sheetnames:
        del wb[SHEET_FLAGS]
    ws = wb.create_sheet(SHEET_FLAGS, index=0)
    _style_header(ws, FLAGS_COLUMNS)

    for r, record in enumerate(rows, start=2):
        for c, name in enumerate(FLAGS_COLUMNS, start=1):
            value = record.get(name)
            cell = ws.cell(row=r, column=c, value=value)
            if name in _NUMERIC_HEADERS and isinstance(value, (int, float)):
                cell.number_format = "0.00"

    _autosize(ws, FLAGS_COLUMNS)
    _apply_flag_styling(ws, n_rows=len(rows))


def _apply_flag_styling(ws: Worksheet, n_rows: int) -> None:
    """Y/N dropdown on Matters? and red fill where Auto-Flag = Y."""
    last_row = max(2, n_rows + 1)

    matters_col = FLAGS_COLUMNS.index("Matters? (Y/N)") + 1
    matters_letter = get_column_letter(matters_col)
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = "Enter Y or N"
    dv.prompt = "Does this flagged move matter to our book?"
    ws.add_data_validation(dv)
    dv.add(f"{matters_letter}2:{matters_letter}{last_row}")

    if n_rows:
        flag_col = FLAGS_COLUMNS.index("Auto-Flag") + 1
        flag_letter = get_column_letter(flag_col)
        last_letter = get_column_letter(len(FLAGS_COLUMNS))
        rule = FormulaRule(formula=[f'${flag_letter}2="Y"'], fill=_FLAG_FILL)
        ws.conditional_formatting.add(f"A2:{last_letter}{last_row}", rule)


def _write_history_sheet(wb: Workbook, history_rows: list[dict[str, Any]]) -> None:
    """Mirror the append-only history CSV into the workbook (full rewrite)."""
    if SHEET_HISTORY in wb.sheetnames:
        del wb[SHEET_HISTORY]
    ws = wb.create_sheet(SHEET_HISTORY)
    _style_header(ws, HISTORY_COLUMNS)
    for r, record in enumerate(history_rows, start=2):
        for c, name in enumerate(HISTORY_COLUMNS, start=1):
            ws.cell(row=r, column=c, value=record.get(name))
    _autosize(ws, HISTORY_COLUMNS)


def _ensure_dictionary_sheet(wb: Workbook, signals: list[Signal]) -> None:
    """Create the Signal Dictionary sheet from ``signals`` only if it's absent.

    Once it exists, it is the owner-editable source of truth -- we never
    overwrite it, so threshold tuning survives every run.
    """
    if SHEET_DICT in wb.sheetnames:
        return
    ws = wb.create_sheet(SHEET_DICT)
    _style_header(ws, DICTIONARY_COLUMNS)
    for r, signal in enumerate(signals, start=2):
        row = signal.as_row()
        for c, name in enumerate(DICTIONARY_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(name))
            if name == "threshold_value" and isinstance(row.get(name), (int, float)):
                cell.number_format = "0.00"
    _autosize(ws, DICTIONARY_COLUMNS)


def _ensure_stub_sheet(wb: Workbook, name: str, columns: list[str], note: str | None) -> None:
    """Create a headers-only stub sheet if absent (never clobber a paste)."""
    if name in wb.sheetnames:
        return
    ws = wb.create_sheet(name)
    _style_header(ws, columns)
    _autosize(ws, columns)
    if note:
        ws.cell(row=2, column=1, value=note).font = Font(italic=True, color="808080")


def write_workbook(
    path: str | Path,
    *,
    flag_rows: list[dict[str, Any]],
    history_rows: list[dict[str, Any]],
    signals: list[Signal],
    include_news: bool = True,
) -> None:
    """Produce or refresh the workbook, preserving prior human dispositions.

    ``flag_rows`` carry only the data columns (see :data:`DATA_COLUMNS`); the
    human columns are merged in from the existing sheet.
    """
    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
        existing = (
            _read_existing_flags(wb[SHEET_FLAGS]) if SHEET_FLAGS in wb.sheetnames else {}
        )
    else:
        wb = Workbook()
        # Drop the default empty sheet openpyxl creates.
        default = wb.active
        if default is not None:
            wb.remove(default)
        existing = {}

    merged = _merge_flags(existing, flag_rows)
    _write_flags_sheet(wb, merged)
    _write_history_sheet(wb, history_rows)
    _ensure_dictionary_sheet(wb, signals)
    _ensure_stub_sheet(
        wb,
        SHEET_INTERNAL,
        FLAGS_COLUMNS,
        "Paste internal flags (segment growth, rising book losses) here so "
        "internal + external get dispositioned together.",
    )
    if include_news:
        _ensure_stub_sheet(
            wb,
            SHEET_NEWS,
            ["Date", "Headline", "Source", "Relevance / Notes"],
            "Paste headlines for human skim (manual, deferred feature).",
        )

    # Order: Flags, History, Dictionary, Internal, News.
    desired = [SHEET_FLAGS, SHEET_HISTORY, SHEET_DICT, SHEET_INTERNAL]
    if include_news:
        desired.append(SHEET_NEWS)
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
