"""Headless test suite for the bureau Consumer Credit-Risk Monitor.

Runs on Linux/CI with NO Excel and NO network, all in --demo mode. Mirrors the
named tests in BUILD_SPEC_BUREAU.md Section 6 (Phases 1-7):

  test_config_parse                     (Phase 1)
  test_demo_provider_deterministic      (Phase 2)
  test_transforms                       (Phase 3)
  test_reload_headless                  (Phase 4)
  test_watchlist_refusal                (Phase 5)
  test_watchlist_gate_defense_in_depth  (Phase 5 negative)
  test_class_c_stub                     (Phase 6)
"""
import math
import os
import shutil
import sys
from dataclasses import replace
from datetime import date

import openpyxl
import pandas as pd
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import runner as R                     # noqa: E402
import series_seed as SEED             # noqa: E402
import build_workbook as BW            # noqa: E402
import assemble_xlsm                   # noqa: E402
import vba_writer as V                 # noqa: E402
import struct                          # noqa: E402

ASOF = date(2026, 3, 31)


# --------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------
@pytest.fixture(scope="session")
def xlsm(tmp_path_factory):
    """Build the macro-enabled workbook once for the whole session."""
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    out = str(d / "Consumer_Credit_Risk_Monitor.xlsm")
    BW.build(base)
    assemble_xlsm.assemble(base, out)
    return out


@pytest.fixture
def populated(xlsm, tmp_path):
    """A fresh copy of the workbook with a --demo run applied."""
    p = str(tmp_path / "run.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    return p


# --------------------------------------------------------------------------
# Phase 1 -- config parse
# --------------------------------------------------------------------------
def test_config_parse():
    cfg = R.parse_config(BW.config_rows())
    assert len(cfg.series) == 18
    ids = {s.id for s in cfg.series}
    assert "hhdc_total_balance" in ids
    assert "WATCHLIST_MSA_PLACEHOLDER" in ids
    # every source-locator / capability column loads
    for s in cfg.series:
        assert s.source_class in ("A", "C")
        assert s.lane in ("dashboard", "watchlist")
        assert isinstance(s.dashboard_capable, bool)
        assert isinstance(s.watchlist_capable, bool)
        assert s.transform in R.TRANSFORMS
    # public stand-in invariant: NO row is watchlist-promotable
    promotable = [s for s in cfg.series
                  if s.lane == "watchlist" and s.watchlist_capable and s.source_class == "C"]
    assert promotable == []
    # and every seeded row is Class A / dashboard-capable / not watchlist-capable
    for s in cfg.series:
        assert s.source_class == "A"
        assert s.watchlist_capable is False


# --------------------------------------------------------------------------
# Phase 2 -- DemoProvider determinism + idempotent raw landing
# --------------------------------------------------------------------------
def test_demo_provider_deterministic():
    cfg = R.parse_config(BW.config_rows())
    prov = R.HhdcDemoProvider(asof=ASOF, slots=cfg.raw_slots)
    spec = next(s for s in cfg.series if s.id == "hhdc_card_balance")
    a = prov.fetch_series(spec)
    b = R.HhdcDemoProvider(asof=ASOF, slots=cfg.raw_slots).fetch_series(spec)
    assert [(r.period, r.value) for r in a] == [(r.period, r.value) for r in b]
    assert len(a) == cfg.raw_slots


def test_raw_landing_idempotent(xlsm, tmp_path):
    def landed(path):
        wb = openpyxl.load_workbook(path, keep_vba=True)
        ws = wb["Raw_HHDC"]
        vals = [(c.row, c.value) for col in ws.iter_cols(min_col=2, max_col=2)
                for c in col if c.value is not None]
        wb.close()
        return vals

    # TRUE idempotence: a second run on the SAME already-populated workbook
    # lands byte-identical values (slot-clearing leaves no stale tail).
    p1 = str(tmp_path / "r1.xlsm")
    shutil.copy(xlsm, p1)
    R.run(p1, demo=True, asof=ASOF)
    first = landed(p1)
    R.run(p1, demo=True, asof=ASOF)
    assert landed(p1) == first
    # ... and determinism across fresh copies.
    p2 = str(tmp_path / "r2.xlsm")
    shutil.copy(xlsm, p2)
    R.run(p2, demo=True, asof=ASOF)
    assert landed(p2) == first
    # newest-first: B4 (first data row of the first block) is populated
    wb = openpyxl.load_workbook(p1, keep_vba=True)
    assert wb["Raw_HHDC"]["B4"].value is not None
    wb.close()


def test_raw_layout_mismatch_refused(xlsm, tmp_path):
    """Editing raw_slots after build must be REFUSED (dashboard formulas are
    anchored to the built layout), not silently mis-written."""
    p = str(tmp_path / "mismatch.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)                     # populate at slots=60
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == "raw_slots":
            target = row[0].row
            break
    assert target is not None
    ws.cell(target, 2, "30")
    wb.save(p)
    wb.close()
    with pytest.raises(RuntimeError, match="Raw layout mismatch"):
        R.run(p, demo=True, asof=ASOF)


# --------------------------------------------------------------------------
# Phase 3 -- transform registry + threshold engine (hardcoded expected)
# --------------------------------------------------------------------------
def test_transforms():
    s = pd.Series([100.0, 102.0, 104.0, 106.0, 110.0, 112.0, 114.0, 116.0])

    # level == identity
    lvl = R.t_level(s, "quarterly")
    assert list(lvl) == list(s)

    # yoy_pct (quarterly -> 4-period lag): index 4 = (110-100)/100*100 = 10.0
    yoy = R.t_yoy_pct(s, "quarterly")
    assert math.isnan(yoy.iloc[0])
    assert yoy.iloc[4] == pytest.approx(10.0)
    assert yoy.iloc[5] == pytest.approx((112 - 102) / 102 * 100)

    # zscore_8q: mean=108, std(ddof=1)=sqrt(240/7); z(last)=(116-108)/std
    z = R.t_zscore_8q(s, "quarterly")
    std = math.sqrt(240.0 / 7.0)
    assert z.iloc[7] == pytest.approx((116 - 108) / std)
    assert math.isnan(z.iloc[6])           # min_periods=8

    # index_to_pct: percent vs the BASE (first valid obs), not a YoY alias.
    # Pin a point where the two genuinely differ: at i=5, index-to-base is
    # (112/100-1)*100 = 12.0 while yoy is (112-102)/102*100 ~ 9.80.
    idx = R.t_index_to_pct(s, "quarterly")
    assert idx.iloc[0] == pytest.approx(0.0)
    assert idx.iloc[7] == pytest.approx(16.0)          # (116/100-1)*100
    assert idx.iloc[5] == pytest.approx(12.0)
    assert R.t_yoy_pct(s, "quarterly").iloc[5] != pytest.approx(12.0)

    # NaN propagates through pct changes (no forward-fill, any pandas version)
    s_nan = pd.Series([100.0, math.nan, 104.0, 106.0, 110.0, 112.0])
    yoy_nan = R.t_yoy_pct(s_nan, "quarterly")
    assert math.isnan(yoy_nan.iloc[5])                 # lag hits the NaN slot


def test_threshold_engine():
    above = R.Threshold(watch=2.0, alert=3.0, direction="above")
    spec = R.SeriesSpec(**_spec_kwargs())
    assert R.status_for(spec, 3.5, above) == "ALERT"
    assert R.status_for(spec, 2.5, above) == "WATCH"
    assert R.status_for(spec, 1.0, above) == "OK"
    assert R.status_for(spec, None, above) == "OK"
    assert R.status_for(spec, float("nan"), above) == "OK"
    below = R.Threshold(watch=2.0, alert=1.0, direction="below")
    assert R.status_for(spec, 0.5, below) == "ALERT"
    assert R.status_for(spec, 1.5, below) == "WATCH"
    assert R.status_for(spec, 3.0, below) == "OK"


def _spec_kwargs():
    return dict(id="x", title="x", category="c", lane="dashboard",
                metric_type="delinq_rate", frequency="quarterly", sa_nsa="NSA",
                units="pct", level_rate_index="rate", geo_segment="national",
                source_class="A", dashboard_capable=True, watchlist_capable=False,
                source_url="", table_id="", sheet="", series_label="",
                transform="level", notes="")


# --------------------------------------------------------------------------
# Phase 4 -- headless reload (Excel proxy) + NO native charts
# --------------------------------------------------------------------------
def test_reload_headless(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    # macro survived the openpyxl round-trip
    assert wb.vba_archive is not None
    for tab in ("Dashboard_Balances", "Dashboard_Delinquency",
                "Dashboard_Originations", "Watchlist", "Raw_HHDC",
                "_config", "_code_py", "_code_vba", "_readme"):
        assert tab in wb.sheetnames
    # NO native chart objects anywhere (the top "recovered content" trigger, L4)
    for ws in wb.worksheets:
        assert getattr(ws, "_charts", []) == []
    # a dashboard headline formula cell is present
    bal = wb["Dashboard_Balances"]
    assert any(isinstance(bal.cell(r, 6).value, str) and str(bal.cell(r, 6).value).startswith("=")
               for r in range(8, 30))
    wb.close()


# --------------------------------------------------------------------------
# Phase 5 -- watchlist refusal + defense in depth
# --------------------------------------------------------------------------
def test_watchlist_refusal(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    ws = wb["Watchlist"]
    blob = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    wb.close()
    # series-named refusal with interpolated id / geo_segment / source_class
    assert "WATCHLIST REFUSED" in blob
    assert 'series "WATCHLIST_MSA_PLACEHOLDER"' in blob
    assert 'geo_segment="msa"' in blob
    assert 'source_class="A"' in blob
    assert "licensed (Class C)" in blob
    # NO public/national series leaked into the lane as data
    assert "hhdc_total_balance" not in blob
    assert "hhdc_card_90plus" not in blob


def test_watchlist_gate_defense_in_depth():
    cfg = R.parse_config(BW.config_rows())

    # (a) flip the MSA placeholder watchlist_capable=TRUE -> still refused by the
    #     source_class="A" gate (build-time hard gate raises).
    flipped = [replace(s, watchlist_capable=True) if s.id == "WATCHLIST_MSA_PLACEHOLDER"
               else s for s in cfg.series]
    with pytest.raises(R.WatchlistRefused):
        R.assert_no_public_in_watchlist(flipped)

    # the runtime evaluator also refuses it, naming Gate2 (source_class)
    msa = next(s for s in flipped if s.id == "WATCHLIST_MSA_PLACEHOLDER")
    reasons = R.gate_watchlist_row(msa)
    assert any("Gate2" in r for r in reasons)     # source_class still A

    # (b) flip a national public row into the watchlist lane -> refused by BOTH
    #     the source_class gate and the geo-whitelist gate.
    nat = next(s for s in cfg.series if s.id == "hhdc_total_balance")
    promoted = replace(nat, lane="watchlist", watchlist_capable=True)
    reasons = R.gate_watchlist_row(promoted)
    assert any("Gate2" in r for r in reasons)     # source_class A
    assert any("Gate3" in r for r in reasons)     # geo national not in whitelist
    with pytest.raises(R.WatchlistRefused):
        R.assert_no_public_in_watchlist([promoted])

    # a genuine licensed MSA row passes all three gates
    licensed = replace(promoted, source_class="C", geo_segment="msa")
    assert R.gate_watchlist_row(licensed) == []


# --------------------------------------------------------------------------
# Phase 6 -- Class C OAuth stub (fully in-process, no live call)
# --------------------------------------------------------------------------
def test_class_c_stub():
    spec = R.SeriesSpec(**{**_spec_kwargs(), "source_class": "C",
                           "geo_segment": "msa", "lane": "watchlist"})
    prov = R.ClassCStubProvider(secret_env="BUREAU_CLASS_C_SECRET")

    # (a) fail-fast: no secret -> HTTP 401 / missing-token path, NO network
    with pytest.raises(PermissionError) as ei:
        prov.fetch_series(spec, secret=None)
    assert "401" in str(ei.value)

    # (b) with a secret, the stub returns the normalized schema (Section 1a):
    # real NormalizedRow instances, schema-shaped but VALUELESS.
    rows = prov.fetch_series(spec, secret="dummy-client-credentials")
    assert isinstance(rows, list) and len(rows) == 1
    nr = rows[0]
    assert isinstance(nr, R.NormalizedRow)
    assert nr.id == spec.id
    assert nr.geo_segment == spec.geo_segment
    assert nr.source_class == "C"
    assert nr.value is None                  # the stub never fabricates data

    # (c) the seam type is the same contract used everywhere else
    assert issubclass(R.ClassCStubProvider, R.Provider)
    assert prov.source_class == "C"


def test_hhdc_provider_cache_and_unbound_parse():
    """Live-provider seam behavior WITHOUT any network: the URL cache serves
    repeated fetches (conditional-fetch/idempotency, L5), and _parse_table
    refuses with a clear message until bound to the real HHDC schema."""
    prov = R.HhdcProvider()
    url = "https://example.invalid/hhdc.xlsx"
    prov._cache[url] = b"fixed-upstream-snapshot"
    assert prov._download(url) == b"fixed-upstream-snapshot"
    assert prov._download(url) == b"fixed-upstream-snapshot"   # served from cache
    spec = R.SeriesSpec(**{**_spec_kwargs(), "source_url": url})
    with pytest.raises(NotImplementedError, match="demo"):
        prov.fetch_series(spec)


def test_vba_protection_keys_roundtrip():
    """The PROJECT stream's CMG/DPB/GC must decrypt per [MS-OVBA] 2.4.3 to
    protection=none / password=none / visibility=visible -- exactly the way
    the VBE will read them (a malformed key triggers Excel's 'invalid key
    DPB' prompt)."""
    proj = V.build_project_stream(
        "X", [V.Module("X", "Sub A()\nEnd Sub")]).decode("latin-1")
    keys = {}
    for line in proj.splitlines():
        for k in ("CMG", "DPB", "GC"):
            if line.startswith(f'{k}="'):
                keys[k] = line.split('"')[1]
    assert V.ovba_decrypt(keys["CMG"]) == struct.pack("<I", 0)   # unprotected
    assert V.ovba_decrypt(keys["DPB"]) == b"\x00"               # no password
    assert V.ovba_decrypt(keys["GC"]) == b"\xff"                # visible
    # _VBA_PROJECT is the full 7-byte header ([MS-OVBA] 2.3.4.1)
    assert len(V.build_vba_project_stream()) == 7
    # the default VBA + Excel libraries are referenced in the dir stream
    d = V.build_dir_stream("X", [V.Module("X", "Sub A()\nEnd Sub")])
    assert b"Visual Basic For Applications" in d
    assert b"Microsoft Excel" in d


def test_resolve_secret_fail_fast(monkeypatch, xlsm, tmp_path):
    """A live (non-demo) run with a named secret_env that is unset fails fast."""
    p = str(tmp_path / "live.xlsm")
    shutil.copy(xlsm, p)
    # inject secret_env into the workbook's _config
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == "secret_env":
            target = row[0].row
            break
    assert target is not None
    ws.cell(target, 2, "BUREAU_CLASS_C_SECRET_UNSET")
    wb.save(p)
    wb.close()
    monkeypatch.delenv("BUREAU_CLASS_C_SECRET_UNSET", raising=False)
    with pytest.raises(SystemExit):
        R.run(p, demo=False, asof=ASOF)
