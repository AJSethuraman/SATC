#!/usr/bin/env python3
"""Consumer Credit-Risk Monitor -- runner (the data path).

Single source of truth for the data path; embedded verbatim into the workbook's
_code_py tab. The VBA ExtractFiles macro writes this back out as runner.py and
the user runs it from PowerShell against the CLOSED workbook (openpyxl backend).

Built to BUILD_SPEC_BUREAU.md. Clean seams (BUILD SPEC 0.3 / 1a / L6):

  * ADAPTER SEAM -- the only provider-specific code. Every adapter implements
    fetch_series(spec, secret) -> list[NormalizedRow] with a FIXED normalized
    schema; nothing downstream calls provider-specific code. v1 ships two Class A
    providers (HhdcDemoProvider offline + HhdcProvider live) and an in-process
    Class C OAuth stub for the swap rehearsal. Licensed feeds are a v2 module swap.
  * TRANSFORM REGISTRY -- deterministic named transforms (BUILD SPEC sec 3).
  * THRESHOLD ENGINE  -- config-driven OK/WATCH/ALERT (BUILD SPEC sec 3).
  * WATCHLIST VALIDATOR -- DEFAULT-DENY WHITELIST (BUILD SPEC 0.1 / sec 3): the
    watchlist lane is refused for any public/national/annual-aggregate series;
    it opens only for a licensed Class C MSA/account join key.

No AI/LLM anywhere; transforms are pure (BUILD SPEC 0.5). Pure ASCII (L3).
"""
from __future__ import annotations

import argparse
import json
import math
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Sequence

import pandas as pd

# Fixed raw-block geometry (newest observation first) so dashboard formulas have
# stable anchors that never shift across refreshes.
RAW_TAB = "Raw_HHDC"
RAW_SLOTS_DEFAULT = 60          # observation rows kept per series (newest-first)
RAW_HEADER_ROWS = 2
RAW_GAP_ROWS = 2
RAW_FIRST_ROW = 2
RAW_PERIOD_COL = 1             # column A
RAW_VALUE_COL = 2             # column B

# Status panel column on each dashboard (free of any merged masthead cells).
STATUS_COL = 12               # column L

FREQ_PERIODS = {"quarterly": 4, "annual": 1, "monthly": 12}

# The watchlist join-key whitelist (BUILD SPEC sec 3, gate 3): default-deny.
WATCHLIST_GEO_WHITELIST = {"msa", "account"}
# Geo tags explicitly NOT joinable (documented; refused by default-deny anyway).
NON_JOINABLE_GEO = {"national", "state_annual", "region", "census_division", ""}


def _freq(frequency: str) -> str:
    f = (frequency or "").strip().lower()
    if f.startswith("q"):
        return "quarterly"
    if f.startswith("a") or f.startswith("y"):
        return "annual"
    if f.startswith("m"):
        return "monthly"
    return "quarterly"


def periods_per_year(frequency: str) -> int:
    return FREQ_PERIODS[_freq(frequency)]


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` tab; backend-agnostic)
# ---------------------------------------------------------------------------
SERIES_HEADER = [
    "id", "title", "category", "lane", "metric_type", "frequency", "sa_nsa",
    "units", "level_rate_index", "geo_segment", "source_class",
    "dashboard_capable", "watchlist_capable", "source_url", "table_id", "sheet",
    "series_label", "transform", "notes",
]


@dataclass
class SeriesSpec:
    id: str
    title: str
    category: str
    lane: str
    metric_type: str
    frequency: str
    sa_nsa: str
    units: str
    level_rate_index: str
    geo_segment: str
    source_class: str
    dashboard_capable: bool
    watchlist_capable: bool
    source_url: str
    table_id: str
    sheet: str
    series_label: str
    transform: str
    notes: str


@dataclass
class Threshold:
    watch: Optional[float]
    alert: Optional[float]
    direction: str          # "above" | "below"


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, Threshold] = field(default_factory=dict)
    series: List[SeriesSpec] = field(default_factory=list)

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _as_float(v):
    try:
        s = str(v).strip()
        return float(s) if s != "" else None
    except (TypeError, ValueError):
        return None


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists). Sections in col A:
    [SETTINGS], [THRESHOLDS], [SERIES]."""
    cfg = Config()
    section = None
    series_header = None
    thr_header = None
    for raw in rows:
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            series_header = None
            thr_header = None
            continue
        if not a:
            continue
        if a.startswith("#"):          # in-sheet comment line, never data
            continue
        if section == "SETTINGS":
            if a.lower() in ("key", "name"):
                continue
            val = "" if len(raw) < 2 or raw[1] is None else raw[1]
            cfg.settings[a] = str(val).strip()
        elif section == "THRESHOLDS":
            if thr_header is None and a.lower() == "id":
                thr_header = [str(c).strip().lower() for c in raw]
                continue
            cells = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                     for i, h in enumerate(thr_header or ["id", "watch", "alert", "direction"])}
            cfg.thresholds[a] = Threshold(
                watch=_as_float(cells.get("watch")),
                alert=_as_float(cells.get("alert")),
                direction=str(cells.get("direction", "above")).strip().lower() or "above")
        elif section == "SERIES":
            if series_header is None:
                series_header = [str(c).strip() for c in raw]
                continue
            v = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                 for i, h in enumerate(series_header)}
            cfg.series.append(SeriesSpec(
                id=str(v.get("id", "")).strip(),
                title=str(v.get("title", "")).strip(),
                category=str(v.get("category", "")).strip(),
                lane=str(v.get("lane", "")).strip().lower(),
                metric_type=str(v.get("metric_type", "")).strip(),
                frequency=str(v.get("frequency", "")).strip(),
                sa_nsa=str(v.get("sa_nsa", "")).strip(),
                units=str(v.get("units", "")).strip(),
                level_rate_index=str(v.get("level_rate_index", "")).strip().lower(),
                geo_segment=str(v.get("geo_segment", "")).strip().lower(),
                source_class=str(v.get("source_class", "")).strip().upper(),
                dashboard_capable=_as_bool(v.get("dashboard_capable", "")),
                watchlist_capable=_as_bool(v.get("watchlist_capable", "")),
                source_url=str(v.get("source_url", "")).strip(),
                table_id=str(v.get("table_id", "")).strip(),
                sheet=str(v.get("sheet", "")).strip(),
                series_label=str(v.get("series_label", "")).strip(),
                transform=str(v.get("transform", "level")).strip().lower(),
                notes=str(v.get("notes", "")).strip()))
    return cfg


# ---------------------------------------------------------------------------
# THE WATCHLIST VALIDATOR -- DEFAULT-DENY WHITELIST (BUILD SPEC 0.1 / sec 3)
# ---------------------------------------------------------------------------
class WatchlistRefused(Exception):
    """Raised by the gate; carries the series-named refusal message."""


def watchlist_refusal_message(spec: SeriesSpec, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: series "{spec.id}" has geo_segment="{spec.geo_segment}", '
        f'watchlist_capable={"TRUE" if spec.watchlist_capable else "FALSE"}, '
        f'source_class="{spec.source_class}". ' + " ".join(reasons) +
        " The watchlist lane requires a licensed (Class C) MSA feed (TransUnion "
        "Prama Benchmarking) or an account-level feed (Experian Risk & Retention "
        "Triggers / TransUnion TruVision). No public/national/annual-aggregate "
        "series may feed this lane.")


def gate_watchlist_row(spec: SeriesSpec) -> List[str]:
    """Return the list of failed-gate reasons for a lane='watchlist' row.
    Empty list == passes all gates (a true licensed MSA/account join key)."""
    reasons = []
    if not spec.watchlist_capable:
        reasons.append("Gate1: watchlist_capable is not TRUE.")
    if spec.source_class != "C":
        reasons.append("Gate2: source_class is not C (licensed) -- public/free "
                       "sources can never feed the watchlist.")
    if spec.geo_segment not in WATCHLIST_GEO_WHITELIST:
        reasons.append(f"Gate3: geo_segment '{spec.geo_segment}' is not in the "
                       f"allowed join-key whitelist {sorted(WATCHLIST_GEO_WHITELIST)}.")
    return reasons


def evaluate_watchlist(series: Sequence[SeriesSpec]):
    """Apply the default-deny whitelist to every lane='watchlist' row.
    Returns (admitted, refusals) -- admitted rows pass ALL gates; refusals is a
    list of (spec, message). Under the v1 public stand-in, admitted is empty."""
    admitted, refusals = [], []
    for s in series:
        if s.lane != "watchlist":
            continue
        reasons = gate_watchlist_row(s)
        if reasons:
            refusals.append((s, watchlist_refusal_message(s, reasons)))
        else:
            admitted.append(s)
    return admitted, refusals


def assert_no_public_in_watchlist(series: Sequence[SeriesSpec]) -> None:
    """Build-time hard gate: a Class A row that is watchlist_capable must never
    exist -- refuse the build, naming the series (BUILD SPEC 0.1)."""
    for s in series:
        if s.watchlist_capable and s.source_class != "C":
            raise WatchlistRefused(watchlist_refusal_message(
                s, ["Gate2: a non-Class-C (public) series is flagged "
                    "watchlist_capable, which is forbidden."]))


# ---------------------------------------------------------------------------
# TRANSFORM REGISTRY (pure, deterministic -- BUILD SPEC sec 3)
# ---------------------------------------------------------------------------
def t_level(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64")


def t_yoy_pct(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64").pct_change(periods=periods_per_year(frequency)) * 100.0


def t_qoq_pct(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64").pct_change(periods=1) * 100.0


def t_mom_pct(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64").pct_change(periods=1) * 100.0


def t_zscore_8q(s: pd.Series, frequency: str) -> pd.Series:
    x = s.astype("float64")
    mean = x.rolling(window=8, min_periods=8).mean()
    std = x.rolling(window=8, min_periods=8).std(ddof=1)
    return (x - mean) / std.replace(0.0, math.nan)


def t_index_to_pct(s: pd.Series, frequency: str) -> pd.Series:
    return t_yoy_pct(s, frequency)


TRANSFORMS = {
    "level": t_level, "yoy_pct": t_yoy_pct, "qoq_pct": t_qoq_pct,
    "mom_pct": t_mom_pct, "zscore_8q": t_zscore_8q, "index_to_pct": t_index_to_pct,
}


class TransformError(Exception):
    pass


def validate_transforms(series: Sequence[SeriesSpec]) -> None:
    for s in series:
        if s.transform not in TRANSFORMS:
            raise TransformError(f"Series '{s.id}' references unknown transform "
                                 f"'{s.transform}'.")


def apply_transform(spec: SeriesSpec, s: pd.Series) -> pd.Series:
    return TRANSFORMS[spec.transform](s, spec.frequency)


def latest_valid(s: pd.Series):
    s = s.dropna()
    return None if s.empty else float(s.iloc[-1])


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT (BUILD SPEC sec 3)
# ---------------------------------------------------------------------------
def status_for(spec: SeriesSpec, value: Optional[float], thr: Optional[Threshold]) -> str:
    if value is None or thr is None or (isinstance(value, float) and math.isnan(value)):
        return "OK"
    above = thr.direction != "below"
    def hit(bound):
        if bound is None:
            return False
        return value >= bound if above else value <= bound
    if hit(thr.alert):
        return "ALERT"
    if hit(thr.watch):
        return "WATCH"
    return "OK"


# ---------------------------------------------------------------------------
# THE ADAPTER SEAM (BUILD SPEC sec 1a) -- the only provider-specific code
# ---------------------------------------------------------------------------
@dataclass
class NormalizedRow:
    id: str
    period: str            # ISO date string (period end)
    value: Optional[float]
    geo_segment: str
    source_class: str
    units: str = ""


class Provider:
    """Adapter interface contract: fetch_series(spec, secret) -> [NormalizedRow]."""

    source_class = "A"

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:  # pragma: no cover
        raise NotImplementedError

    def last_observation_period(self, rows: List[NormalizedRow]) -> Optional[str]:
        """Derived from already-fetched data -- no extra call (L5)."""
        ps = [r.period for r in rows if r.value is not None]
        return max(ps) if ps else None


class HhdcDemoProvider(Provider):
    """Deterministic offline stand-in (BUILD SPEC 0.7 / Phase 2). Seeded
    pseudo-walk per id, fixed asof, NO network, NO key -- used by ALL tests."""

    source_class = "A"

    def __init__(self, asof: Optional[date] = None, slots: int = RAW_SLOTS_DEFAULT):
        self.asof = asof or date(2026, 3, 31)
        self.slots = slots

    def _seed(self, sid: str) -> int:
        return sum((i + 1) * ord(c) for i, c in enumerate(sid)) % 997

    def _base(self, spec: SeriesSpec, seed: int) -> float:
        mt = spec.metric_type
        if mt == "delinq_rate":
            return 1.5 + (seed % 30) / 10.0          # ~1.5-4.5 %
        if mt == "delinq_flow":
            return 1.0 + (seed % 20) / 10.0          # ~1-3 %
        if mt == "origination":
            return 5.0 + (seed % 200) / 10.0         # millions / $bn
        if "tn" in spec.units:
            return 1.0 + (seed % 120) / 10.0         # $ trillions
        return 100.0 + (seed % 800)                  # $ billions

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        seed = self._seed(spec.id)
        freq = _freq(spec.frequency)
        n = self.slots
        pcode = {"quarterly": "Q", "annual": "Y", "monthly": "M"}[freq]
        idx = pd.period_range(end=pd.Period(self.asof, freq=pcode[0]), periods=n, freq=pcode[0])
        idx = idx.to_timestamp(how="end").normalize()
        v = self._base(spec, seed)
        rows = []
        for i in range(n):
            wobble = math.sin((i + seed) / 5.0) * 0.04
            drift = 0.015 if i > n - 8 else 0.0       # gentle recent rise -> trips some thresholds
            v = max(0.05, v * (1.0 + wobble + drift))
            rows.append(NormalizedRow(
                id=spec.id, period=idx[i].date().isoformat(),
                value=round(v, 4), geo_segment=spec.geo_segment,
                source_class=spec.source_class, units=spec.units))
        # exercise the missing-value path deterministically
        rows[seed % n] = NormalizedRow(spec.id, rows[seed % n].period, None,
                                       spec.geo_segment, spec.source_class, spec.units)
        return rows


class HhdcProvider(Provider):
    """Live Class A provider: downloads the NY Fed Household Debt & Credit public
    tables and maps each id via its source-locator. HHDC is a BULK published
    table, NOT a per-series API, and its literal column schema is UNKNOWN at spec
    time (Open Question #5) -- so _parse_table must be bound to the real columns
    before live use. Tests never exercise this path (they use HhdcDemoProvider);
    live-source idempotence is tested only against cached bytes, never the
    network (HHDC revises)."""

    source_class = "A"

    def __init__(self, min_interval: float = 0.0, max_retries: int = 3):
        self._min_interval = float(min_interval)
        self._max_retries = int(max_retries)
        self._last = 0.0
        self._cache: Dict[str, bytes] = {}

    def _throttle(self):
        gap = time.time() - self._last
        if gap < self._min_interval:
            time.sleep(self._min_interval - gap)
        self._last = time.time()

    def _download(self, url: str) -> bytes:
        if url in self._cache:                        # conditional-fetch/idempotency (L5)
            return self._cache[url]
        import urllib.request
        for attempt in range(self._max_retries + 1):
            self._throttle()
            try:
                with urllib.request.urlopen(url, timeout=30) as resp:
                    data = resp.read()
                self._cache[url] = data
                return data
            except Exception:
                if attempt < self._max_retries:
                    time.sleep(2.0 * (attempt + 1))
                    continue
                raise

    def _parse_table(self, data: bytes, spec: SeriesSpec) -> List[NormalizedRow]:
        # SCHEMA BINDING REQUIRED (Open Question #5): the literal HHDC table
        # columns are not established by research. Bind table_id/sheet/series_label
        # to the actual published layout here before any live run.
        raise NotImplementedError(
            "HhdcProvider._parse_table is unbound: the literal NY Fed HHDC table "
            "schema is an open question. Use --demo for build/test; bind this to "
            "the real published columns before a live pull.")

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        if not spec.source_url:
            raise ValueError(f"Series '{spec.id}' has no source_url for live HHDC fetch.")
        return self._parse_table(self._download(spec.source_url), spec)


class ClassCStubProvider(Provider):
    """In-process OAuth client_credentials STUB for the Phase 6 swap rehearsal.
    Makes NO live request -- it asserts the fail-fast + 401/missing-token code
    path without a token request. Live Class C calls are forbidden in v1."""

    source_class = "C"

    def __init__(self, secret_env: str):
        self.secret_env = secret_env
        self._token = None

    def _authenticate(self, secret):
        # Equifax pattern: token from client_id + client_secret + scope; a request
        # without a valid token returns HTTP 401. Stubbed -- never hits a network.
        if not secret:
            raise PermissionError("HTTP 401: missing/invalid access token "
                                  "(client_credentials not supplied).")
        self._token = "stub-token"
        return self._token

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        self._authenticate(secret)
        # A real adapter would call the licensed endpoint here and normalize.
        return []


def resolve_secret(cfg: Config) -> Optional[str]:
    """Class C secret from the env var whose NAME is recorded in _config; never
    hardcoded. Absent env var -> fail fast (handled by the caller)."""
    name = str(cfg.setting("secret_env", "") or "").strip()
    if not name:
        return None
    return os.environ.get(name)


def make_provider(cfg: Config, demo: bool, asof: Optional[date]) -> Provider:
    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        return HhdcDemoProvider(asof=asof, slots=cfg.raw_slots)
    return HhdcProvider(min_interval=float(cfg.setting("http_min_interval", 0.0) or 0.0))


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder
# ---------------------------------------------------------------------------
@dataclass
class RawBlock:
    id: str
    header_row: int
    label_row: int
    first_data_row: int
    slots: int


def _col(idx: int) -> str:
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def raw_layout(series: Sequence[SeriesSpec], slots: int = RAW_SLOTS_DEFAULT) -> Dict[str, RawBlock]:
    """Deterministic block placement in Raw_HHDC: same input -> same anchors."""
    stride = RAW_HEADER_ROWS + slots + RAW_GAP_ROWS
    blocks = {}
    for i, s in enumerate(series):
        header_row = RAW_FIRST_ROW + i * stride
        blocks[s.id] = RawBlock(s.id, header_row, header_row + 1,
                                header_row + RAW_HEADER_ROWS, slots)
    return blocks


# ---------------------------------------------------------------------------
# WRITE BACKEND -- openpyxl on the closed workbook (BUILD SPEC sec 5 / L2)
# ---------------------------------------------------------------------------
class OpenpyxlBackend:
    def __init__(self, path: str):
        self.path = path
        import openpyxl
        # L2: keep_vba=True ONLY for .xlsm; on .xlsx it injects a dangling
        # vbaProject relationship Excel rejects as "format/extension not valid".
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        return parse_config([[c.value for c in row] for row in ws.iter_rows()])

    def write_raw_block(self, block: RawBlock, spec: SeriesSpec, rows: List[NormalizedRow]):
        ws = self._wb[RAW_TAB]
        ws.cell(block.header_row, 1, spec.id)
        ws.cell(block.header_row, 2, spec.title)
        ws.cell(block.header_row, 3, f"freq={spec.frequency}; transform={spec.transform}; "
                                     f"class={spec.source_class}; geo={spec.geo_segment}")
        ws.cell(block.label_row, 1, "period")
        ws.cell(block.label_row, 2, "value")
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            ws.cell(r, 1, None)
            ws.cell(r, 2, None)
        # newest-first
        tail = rows[-block.slots:] if len(rows) > block.slots else rows
        for i, nr in enumerate(reversed(tail)):
            rr = block.first_data_row + i
            ws.cell(rr, 1, nr.period)
            ws.cell(rr, 2, None if nr.value is None else float(nr.value))

    def write_status(self, status: dict):
        for tab in ("Dashboard_Balances", "Dashboard_Delinquency", "Dashboard_Originations"):
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                ws.cell(1, STATUS_COL, "Last run  " + status.get("timestamp", ""))
                ws.cell(2, STATUS_COL, f"Series {status.get('series_pulled', 0)}/"
                                       f"{status.get('series_in_dict', 0)} - "
                                       f"{status.get('alert_count', 0)} ALERT / "
                                       f"{status.get('watch_count', 0)} WATCH")

    def finalize(self):
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def compute_digest(cfg: Config, series_rows: Dict[str, List[NormalizedRow]]) -> List[dict]:
    """Per-series headline transform value + OK/WATCH/ALERT status (the engine
    output consumed by dashboards + email-sim)."""
    digest = []
    for spec in cfg.series:
        if not spec.dashboard_capable or spec.id not in series_rows:
            continue
        vals = [r.value for r in series_rows[spec.id]]
        s = pd.Series(vals, dtype="float64")
        headline = latest_valid(apply_transform(spec, s))
        thr = cfg.thresholds.get(spec.id)
        digest.append({"id": spec.id, "title": spec.title, "metric": spec.metric_type,
                       "transform": spec.transform, "value": headline,
                       "status": status_for(spec, headline, thr)})
    return digest


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None) -> dict:
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any fetch (BUILD SPEC 0.1, sec 3).
    validate_transforms(cfg.series)
    assert_no_public_in_watchlist(cfg.series)
    admitted, refusals = evaluate_watchlist(cfg.series)

    if not (demo or _as_bool(cfg.setting("demo_mode", "false"))):
        # live HHDC needs no secret; a Class C swap would fail fast here.
        secret_name = str(cfg.setting("secret_env", "") or "").strip()
        if secret_name and resolve_secret(cfg) is None:
            raise SystemExit(f"Secret env var '{secret_name}' is not set. "
                             "Set it before a licensed (Class C) run; never hardcode it.")
    provider = make_provider(cfg, demo, asof)
    mode = "demo" if isinstance(provider, HhdcDemoProvider) else "live"

    blocks = raw_layout(cfg.series, slots=cfg.raw_slots)
    series_rows: Dict[str, List[NormalizedRow]] = {}
    pulled, errors = 0, []
    for spec in cfg.series:
        if spec.lane == "watchlist":
            continue                                  # gated lane is never fetched in v1
        try:
            rows = provider.fetch_series(spec)
        except Exception as exc:
            errors.append(f"{spec.id}: {exc}")
            continue
        series_rows[spec.id] = rows
        backend.write_raw_block(blocks[spec.id], spec, rows)
        pulled += 1

    digest = compute_digest(cfg, series_rows)
    status = {
        "timestamp": asof.isoformat(),
        "mode": mode,
        "series_in_dict": len(cfg.series),
        "series_pulled": pulled,
        "alert_count": sum(1 for d in digest if d["status"] == "ALERT"),
        "watch_count": sum(1 for d in digest if d["status"] == "WATCH"),
        "digest": digest,
        "watchlist_refusals": [m for _, m in refusals],
        "watchlist_admitted": [s.id for s in admitted],
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Consumer Credit-Risk Monitor runner")
    ap.add_argument("--workbook", "-w", required=True)
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline HhdcDemoProvider (no network/key)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    args = ap.parse_args(argv)
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, TransformError) as exc:
        sys.stderr.write(f"GATE ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except SystemExit as exc:
        sys.stderr.write(f"{exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True, **{k: v for k, v in status.items() if k != "digest"}}))
    sys.stderr.write(
        f"OK ({status['mode']}): {status['series_pulled']}/{status['series_in_dict']} series, "
        f"{status['alert_count']} ALERT, {status['watch_count']} WATCH, "
        f"watchlist admitted={len(status['watchlist_admitted'])} "
        f"refused={len(status['watchlist_refusals'])}.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
