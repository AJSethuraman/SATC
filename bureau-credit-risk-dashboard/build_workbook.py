#!/usr/bin/env python3
"""Build the Consumer Credit-Risk Monitor workbook (bureau feed) -- KeyBank-styled.

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled .xlsm.
All visual style comes from keybank_style.py (tokens + helpers) -- never a
hardcoded hex here. Built to BUILD_SPEC_BUREAU.md (Sections 2-6).

Structure (Section 4):
  Dashboard_Balances / _Delinquency / _Originations  -- formula-driven panels
  Watchlist        -- the GATED lane: shows the series-named refusal, not data
  Raw_HHDC         -- one raw tab, fixed-anchor blocks, runner fills the values
  _config          -- the Section 2 dictionary + [THRESHOLDS] (source of truth)
  _code_py/_code_vba -- runner.py + macro.bas as flat ASCII text
  _readme          -- provider notes, score-scale + compliance, run steps

No native charts (L4). Embedded code is flat ASCII (L3).
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import runner as R
import series_seed as SEED

import keybank_style as KB
from keybank_style import (
    INK, KEY_RED, MONO_FONT, NOTE_FONT, SECT_FILL,
    brand_banner, kpi_tiles, header_row, watchlist_boundary,
    yoy_heat, hide_gridlines, freeze_below,
)

HERE = os.path.dirname(os.path.abspath(__file__))
INK_BOLD = Font(name="Arial", bold=True, color=INK)


def text_cell(ws, row, col, value):
    """Write literal text; openpyxl treats a leading '=' as a formula, so force
    it back to a string for source/doc lines."""
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# _config  (the knob panel: SETTINGS / THRESHOLDS / SERIES)
# ==========================================================================
# Per-id thresholds. Keyed to seed ids; demo data trips a few of these so the
# email-sim has color. direction "above" = flag when value >= bound.
THRESHOLDS = [
    # id, watch, alert, direction
    ("hhdc_card_90plus", 2.0, 3.0, "above"),
    ("hhdc_auto_90plus", 2.0, 3.0, "above"),
    ("hhdc_mortgage_90plus", 1.5, 2.5, "above"),
    ("hhdc_flow_to_30", 1.5, 2.5, "above"),
    ("hhdc_flow_to_90", 1.0, 2.0, "above"),
    ("hhdc_total_balance", 5.0, 10.0, "above"),
    ("hhdc_card_balance", 6.0, 12.0, "above"),
]


def config_rows():
    rows = [["Consumer Credit-Risk Monitor (bureau feed) -- CONFIG (the knob "
             "panel). Edit values here; no code change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["demo_mode", "FALSE", "TRUE = offline deterministic HhdcDemoProvider "
         "(no network/key), for trying the button and tests."],
        ["raw_slots", "60", "observations kept per series (newest-first)."],
        ["http_min_interval", "0.0", "seconds between live HHDC downloads "
         "(conditional-fetch; HHDC is a static published table, no API quota)."],
        ["secret_env", "", "NAME of the env var holding a licensed (Class C) "
         "secret. Empty under the public stand-in. The secret is NEVER hardcoded "
         "here -- only the variable name; the value lives in the environment."],
    ]
    rows += [[], ["[THRESHOLDS]"], ["id", "watch", "alert", "direction"]]
    for tid, watch, alert, direction in THRESHOLDS:
        rows.append([tid, watch, alert, direction])
    rows += [[], ["[SERIES]"], SEED.HEADER]
    for r in SEED.all_series():
        rows.append([r[h] for h in SEED.HEADER])
    rows += [[], ["# To add a series: append a [SERIES] row. To open the "
                  "watchlist lane you need a licensed Class C MSA/account feed "
                  "(source_class=C, watchlist_capable=TRUE, geo_segment in "
                  "{msa,account}); a public row can never be promoted into it."]]
    return rows


def write_config(wb):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows()
    thr_cells = {}                 # id -> (watch_ref, alert_ref, direction)
    in_thr = False
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if row and row[0] == "[THRESHOLDS]":
            in_thr = True
            continue
        if row and isinstance(row[0], str) and row[0].startswith("[") and row[0] != "[THRESHOLDS]":
            in_thr = False
        if in_thr and row and row[0] not in ("[THRESHOLDS]", "id"):
            tid = row[0]
            thr_cells[tid] = (f"_config!$B${i}", f"_config!$C${i}",
                              str(row[3]).strip().lower())
    ws["A1"].font = INK_BOLD
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("[") and row[0].endswith("]"):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    return thr_cells


# ==========================================================================
# Raw_HHDC scaffold  (single tab, fixed anchors; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, specs, blocks):
    ws = wb.create_sheet(R.RAW_TAB)
    hide_gridlines(ws)
    ws["A1"] = (f"{R.RAW_TAB} -- raw HHDC observations (or DemoProvider output in "
                f"--demo), newest-first. Written by runner.py; dashboards read "
                f"these by formula. Do not edit by hand.")
    ws["A1"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    for s in specs:
        if s.lane == "watchlist":
            continue               # gated lane is never fetched / never lands raw
        b = blocks[s.id]
        ws.cell(b.header_row, 1, s.id).font = INK_BOLD
        ws.cell(b.header_row, 2, s.title)
        ws.cell(b.header_row, 3, f"freq={s.frequency}; transform={s.transform}; "
                                 f"class={s.source_class}; geo={s.geo_segment}")
        ws.cell(b.header_row, 3).font = NOTE_FONT
        ws.cell(b.label_row, 1, "period").font = Font(name="Calibri", bold=True, size=9)
        ws.cell(b.label_row, 2, "value").font = Font(name="Calibri", bold=True, size=9)
    return ws


# ==========================================================================
# Dashboards  (formula-driven; conditional-format heat; no native charts)
# ==========================================================================
def _ref(b, offset=0):
    return f"{R.RAW_TAB}!{get_column_letter(R.RAW_VALUE_COL)}{b.first_data_row + offset}"


DASH_COLS = ["Category", "Series ID", "Title", "Latest", "Prior", "Headline",
             "Trend (8)", "Status"]
COL = {name: get_column_letter(i + 1) for i, name in enumerate(DASH_COLS)}
NCOLS = len(DASH_COLS)
HEADLINE_COL = 6
STATUS_C = 8


def _headline_formula(spec, b):
    """Headline = the named transform's latest value, by formula off the raw
    block. yoy_pct -> YoY %; level -> latest level."""
    latest = _ref(b, 0)
    if spec.transform in ("yoy_pct", "index_to_pct"):
        ppy = R.periods_per_year(spec.frequency)
        yago = _ref(b, ppy)
        return f"=IFERROR(({latest}-{yago})/{yago}*100,\"\")"
    return f"=IFERROR({latest},\"\")"


def _status_formula(r, thr_cells, spec):
    """OK/WATCH/ALERT by formula, referencing the _config [THRESHOLDS] cells
    (config-driven, not baked into the dashboard). Empty if no threshold."""
    cells = thr_cells.get(spec.id)
    if not cells:
        return ""
    watch_ref, alert_ref, direction = cells
    hc = f"${COL['Headline']}{r}"
    op = ">=" if direction != "below" else "<="
    return (f"=IF(NOT(ISNUMBER({hc})),\"\","
            f"IF({hc}{op}{alert_ref},\"ALERT\","
            f"IF({hc}{op}{watch_ref},\"WATCH\",\"OK\")))")


def write_dashboard(wb, tab, specs, blocks, thr_cells, title, subtitle,
                    metric_types, headline_label):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    rows = sorted([s for s in specs if s.lane == "dashboard"
                   and s.dashboard_capable and s.metric_type in metric_types],
                  key=lambda s: (s.category, s.id))

    hdr = 7
    first = hdr + 1
    r = first
    cols = list(DASH_COLS)
    cols[HEADLINE_COL - 1] = headline_label
    for s in rows:
        b = blocks[s.id]
        latest, prior = _ref(b, 0), _ref(b, 1)
        ws.cell(r, 1, s.category).font = KB.SECONDARY
        ws.cell(r, 2, s.id).font = MONO_FONT
        ws.cell(r, 3, s.title).font = KB.DATA_FONT
        ws.cell(r, 4, f"=IFERROR({latest},\"\")")
        ws.cell(r, 5, f"=IFERROR({prior},\"\")")
        ws.cell(r, HEADLINE_COL, _headline_formula(s, b))
        # col 7 (Trend) left blank -- sparklines painted by the macro.
        status = _status_formula(r, thr_cells, s)
        if status:
            ws.cell(r, STATUS_C, status)
        for c in (4, 5):
            ws.cell(r, c).number_format = KB.FMT_RATE
        ws.cell(r, HEADLINE_COL).number_format = KB.FMT_YOY
        for c in (4, 5, HEADLINE_COL):
            ws.cell(r, c).alignment = KB.RIGHT
        ws.cell(r, STATUS_C).alignment = KB.CENTER
        for c in range(1, NCOLS + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        r += 1
    last = r - 1

    header_row(ws, hdr, cols, right_from=4, center_cols=(7, 8))
    freeze_below(ws, hdr)
    _dash_widths(ws)

    brand_banner(ws, 1, NCOLS, title, subtitle)
    sc = COL["Status"]
    if last >= first:
        tiles = [
            ("SERIES IN ALERT", f"=COUNTIF({sc}{first}:{sc}{last},\"ALERT\")",
             f"of {len(rows)}"),
            ("SERIES IN WATCH", f"=COUNTIF({sc}{first}:{sc}{last},\"WATCH\")",
             "elevated"),
            ("SERIES TRACKED", f"={len(rows)}", "this lane"),
        ]
        kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8

    # run-status readout in column L (free of the masthead merges)
    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 34
    for rr in (1, 2):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(3, R.STATUS_COL, "Anonymized 5% sample -- aggregate only").font = NOTE_FONT
    ws.cell(3, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(4, R.STATUS_COL).font = NOTE_FONT          # macro writes status here
    ws.cell(4, R.STATUS_COL).alignment = KB.LEFT

    if last >= first:
        KB.alert_rule(ws, f"{sc}{first}:{sc}{last}")
        # heat the headline column (red = stress on rates/growth)
        yoy_heat(ws, f"{COL['Headline']}{first}:{COL['Headline']}{last}")
    return ws


def _dash_widths(ws):
    widths = {"A": 14, "B": 24, "C": 40, "D": 11, "E": 11, "F": 12, "G": 14, "H": 12}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


# ==========================================================================
# Watchlist  (the GATED lane -- shows the refusal, not data)
# ==========================================================================
BOUNDARY = ("GATED LANE. The watchlist requires a licensed (Class C) MSA or "
            "account-level feed. Under the public stand-in every series is "
            "Class A / national / aggregate and is refused by the default-deny "
            "validator -- by design. No public/national/annual-aggregate series "
            "may feed this lane.")

WL_COLS = ["Series ID", "geo_segment", "source_class", "Validator decision"]


def write_watchlist(wb, specs):
    # Build-time hard gate + the per-row refusal messages (runner sec 3).
    R.assert_no_public_in_watchlist(specs)
    admitted, refusals = R.evaluate_watchlist(specs)

    ws = wb.create_sheet("Watchlist")
    hide_gridlines(ws)
    brand_banner(ws, 1, len(WL_COLS), "Watchlist -- Gated Lane",
                 "Licensed MSA / account-level feed required (TransUnion Prama / "
                 "Experian Triggers / TransUnion TruVision).")
    watchlist_boundary(ws, 3, len(WL_COLS), BOUNDARY)

    hdr = 4
    header_row(ws, hdr, WL_COLS, right_from=None)
    r = hdr + 1
    if not refusals and not admitted:
        ws.cell(r, 1, "No lane=\"watchlist\" rows in _config.").font = KB.SECONDARY
        r += 1
    for spec, message in refusals:
        ws.cell(r, 1, spec.id).font = MONO_FONT
        ws.cell(r, 2, spec.geo_segment).font = KB.SECONDARY
        ws.cell(r, 3, spec.source_class).font = KB.SECONDARY
        ws.cell(r, 4, "REFUSED").font = Font(name="Calibri", bold=True, color=KB.CRIMSON)
        for c in range(1, len(WL_COLS) + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        r += 1
        # the full series-named refusal message, wrapped under the row
        ws.cell(r, 1, message).font = NOTE_FONT
        ws.cell(r, 1).alignment = KB.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(WL_COLS))
        ws.row_dimensions[r].height = 56
        r += 1
    for spec in admitted:                  # empty under the public stand-in
        ws.cell(r, 1, spec.id).font = MONO_FONT
        ws.cell(r, 2, spec.geo_segment).font = KB.SECONDARY
        ws.cell(r, 3, spec.source_class).font = KB.SECONDARY
        ws.cell(r, 4, "ADMITTED (licensed)").font = KB.DATA_FONT
        r += 1
    widths = {"A": 30, "B": 16, "C": 14, "D": 22}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    freeze_below(ws, hdr)
    return ws


# ==========================================================================
# Code-in-tab + readme
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
CONSUMER CREDIT-RISK MONITOR (BUREAU FEED) -- README
====================================================

WHAT THIS IS
  A self-contained, reusable template. One workbook pulls aggregate consumer
  credit-risk metrics from a credit-bureau-sourced public feed (the NY Fed
  Household Debt & Credit report, built on an anonymized 5% Equifax sample),
  lands them raw, and presents them as formula-driven dashboards. A gated
  "watchlist" lane is reserved for a future licensed (Class C) MSA / account
  feed and is refused under the public stand-in by design. All code, config and
  docs live inside this workbook, so it can be emailed, opened elsewhere, and
  re-run.

PROVIDERS (the adapter seam -- one provider per class)
  Class A (free, no contract):
    - HhdcProvider     -- LIVE download of the NY Fed HHDC public tables.
    - HhdcDemoProvider -- deterministic OFFLINE stand-in (no network/key); used
                          by every test and the --demo button.
  Class C (licensed, v2 swap only): a module swap behind the same seam. v1 ships
  only an in-process OAuth client_credentials STUB; no live licensed call is made.

ONE-TIME SETUP
  1. Install Python 3 (add it to PATH) and the dependencies:
         pip install openpyxl pandas
  2. Enable macros when prompted (the Extract button needs them).

RUN IT  (the workbook is the finished product; you only add data)
  1. In Excel press Alt+F8 -> run ExtractFiles. It writes three files next to the
     workbook: runner.py (from _code_py), requirements.txt, and RUN.txt (the
     exact PowerShell commands). NOTHING runs inside Excel.
  2. SAVE and CLOSE the workbook.
  3. Open RUN.txt and follow it from PowerShell: install deps, then run
     runner.py against the CLOSED workbook. It writes the data into the file.
       - Try-the-button (offline): runner.py -w <book>.xlsm --demo
       - Live HHDC: runner.py -w <book>.xlsm   (downloads the public tables)
  4. Reopen the workbook -- the formulas recalc into the populated dashboards.
  Optional: Alt+F8 -> PaintSparklines draws the Trend column.

THE TABS
  Dashboard_Balances / _Delinquency / _Originations -- formula panels: latest,
      prior, a named-transform headline (YoY% for balances/originations; the
      level for delinquency rates/flows), an optional trend sparkline, and a
      config-driven OK/WATCH/ALERT status. Heat shading marks stress.
  Watchlist -- the GATED lane. Shows the series-named refusal and the licensed-
      feed requirement, not data, under the public stand-in.
  Raw_HHDC -- raw observations, newest-first (the audit trail).
  _config  -- THE KNOB PANEL: series dictionary + thresholds (source of truth).
  _code_py / _code_vba -- the runner and macro as plain text.

THE WATCHLIST BOUNDARY (default-deny, defense in depth)
  A row only feeds the watchlist if ALL THREE hold: watchlist_capable=TRUE AND
  source_class="C" (licensed) AND geo_segment in {msa, account}. Every public
  stand-in series is Class A / national / aggregate, so the lane is refused --
  and stays refused even if someone flips a single capability flag, because the
  source_class="A" gate still catches it. No public/national/annual-aggregate
  series can ever localize a portfolio subset.

SCORE-SCALE HONESTY
  Risk scores differ by scale and are never cross-compared without explicit
  normalization: Equifax Risk Score 280-850 (prime >= 660); VantageScore and
  FICO use different ranges. v1 records the scale and refuses silent cross-scale
  comparison; it does not invent a crosswalk.

COMPLIANCE (design intent; see Open Questions -- several items UNKNOWN)
  - Public HHDC is an anonymized 5% sample (NOT a census), aggregate-only and
    NOT account-joinable; microdata is restricted to Fed researchers.
  - Secrets (licensed Class C) come from an env var named in _config (secret_env)
    ONLY -- never hardcoded, never written to disk by the bootstrap, never echoed.
    Absent env var -> the run fails fast.
  - FCRA permissible purpose ("account review") is tied specifically to the
    account-level licensed products; whether/how it applies to aggregate feeds is
    UNKNOWN and must be confirmed at contract time. Redistribution terms for any
    licensed feed are UNKNOWN at spec time. v1 touches no licensed data.

DATA-QUALITY TRAPS (baked into the adapter / notes)
  - 5% anonymized sample, not a census -- sampling noise grows at finer cuts.
  - One-quarter origination lag is standard.
  - State data is ANNUAL Q4 only; quarterly all-state granularity is withheld by
    Equifax contract -- so state data is NOT a usable watchlist geo key.
  - Vendor reference figures are point-in-time and change each release.

NO AI IS INVOLVED IN THE DATA PATH. Transforms, thresholds and the watchlist
gate are deterministic and config-driven: the same input always yields the same
workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() and len(line) > 3 \
                and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path):
    rows = config_rows()
    cfg = R.parse_config(rows)
    specs = cfg.series
    R.assert_no_public_in_watchlist(specs)
    R.validate_transforms(specs)
    blocks = R.raw_layout(specs, slots=cfg.raw_slots)

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, specs, blocks)
    thr_cells = {}
    # _config must be written before dashboards so threshold cell refs resolve.
    # We write _config into a temp pass to capture thr_cells, then real sheets.
    thr_cells = write_config(wb)

    write_dashboard(wb, "Dashboard_Balances", specs, blocks, thr_cells,
                    "Balances Dashboard",
                    "Household debt balances by product (HHDC, 5% Equifax sample). "
                    "Headline = YoY %.",
                    metric_types={"balance"}, headline_label="YoY %")
    write_dashboard(wb, "Dashboard_Delinquency", specs, blocks, thr_cells,
                    "Delinquency Dashboard",
                    "90+ DPD rates and transition flows by DPD bucket. "
                    "Headline = current rate (level).",
                    metric_types={"delinq_rate", "delinq_flow"},
                    headline_label="Rate %")
    write_dashboard(wb, "Dashboard_Originations", specs, blocks, thr_cells,
                    "Originations Dashboard",
                    "New originations by product. NOTE: a one-quarter origination "
                    "lag is standard. Headline = YoY %.",
                    metric_types={"origination"}, headline_label="YoY %")
    write_watchlist(wb, specs)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    order = ["Dashboard_Balances", "Dashboard_Delinquency", "Dashboard_Originations",
             "Watchlist", "Raw_HHDC", "_config", "_code_py", "_code_vba", "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.active = wb.sheetnames.index("Dashboard_Balances")
    wb.save(out_path)
    admitted, refusals = R.evaluate_watchlist(specs)
    return out_path, len(specs), len(refusals), len(admitted)


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "Consumer_Credit_Risk_Monitor_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n, nref, nadm = build(out)
    print(f"built {path}: {n} series, {nref} watchlist-refused, {nadm} admitted")
