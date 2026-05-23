from __future__ import annotations

import argparse
import html
import json
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SATC = {"navy": "0B1F3A", "gold": "B08D57", "cream": "F6F2EA", "paper": "FBF9F4", "hairline": "D9CFB8"}


@dataclass
class InventoryItem:
    original_file_path: str
    filename: str
    parent_folder: str
    inferred_document_type: str
    inferred_tax_year: str
    inferred_entity_name: str
    confidence: float
    evidence: str
    duplicate_flag: bool
    notes: str = ""


def load_config(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def infer_year(text: str) -> str:
    m = re.search(r"(?<!\d)(20\d{2})(?!\d)", text)
    return m.group(1) if m else ""


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def _contains_phrase(text: str, phrase: str) -> bool:
    return re.search(rf"\b{re.escape(_norm(phrase))}\b", _norm(text)) is not None


def infer_entity_name(filename: str) -> str:
    parts = [p for p in re.split(r"[_\- ]+", Path(filename).stem) if p]
    flt = [p for p in parts if not re.match(r"^(20\d{2}|w2|w-2|1098|1099|nec|div|int|b|t)$", p.lower())]
    return flt[0] if flt else ""


def infer_document_type(path_text: str) -> tuple[str, float, str]:
    t = _norm(path_text)
    if _contains_phrase(t, "1098 t") or _contains_phrase(t, "tuition statement"):
        return "1098-t education", 0.9, "matched education phrases"
    if (_contains_phrase(t, "w2") or _contains_phrase(t, "w 2") or _contains_phrase(t, "w-2")):
        return "w-2", 0.9, "matched w-2 phrase"
    if _contains_phrase(t, "1099 nec"):
        return "1099-nec", 0.9, "matched 1099-nec phrase"
    if _contains_phrase(t, "1099 int"):
        return "1099-int", 0.9, "matched 1099-int phrase"
    if _contains_phrase(t, "1099 div"):
        return "1099-div", 0.9, "matched 1099-div phrase"
    if _contains_phrase(t, "1099 b") or _contains_phrase(t, "brokerage") or _contains_phrase(t, "consolidated 1099"):
        return "1099-b / brokerage", 0.9, "matched brokerage phrase"
    if (_contains_phrase(t, "1098 mortgage") or _contains_phrase(t, "mortgage interest statement") or _contains_phrase(t, "form 1098 mortgage")):
        return "1098 mortgage", 0.92, "matched mortgage-specific phrase"
    if _contains_phrase(t, "rent ledger") or _contains_phrase(t, "rental income summary"):
        return "rental income support", 0.88, "matched rental income phrase"
    if _contains_phrase(t, "property tax"):
        if _contains_phrase(t, "rental"):
            return "rental property tax", 0.82, "matched property tax with rental context"
        return "property tax", 0.7, "matched property tax phrase without rental context"
    return "unknown", 0.35, "no confident phrase match"


def build_inventory(input_folder: Path) -> list[InventoryItem]:
    files = [p for p in input_folder.rglob("*") if p.is_file()]
    seen = Counter([p.name.lower() for p in files])
    items = []
    for f in files:
        rel = str(f.relative_to(input_folder)).replace("\\", "/")
        dt, conf, ev = infer_document_type(rel)
        items.append(InventoryItem(str(f.resolve()), f.name, f.parent.name, dt, infer_year(rel), infer_entity_name(f.name), conf, ev, seen[f.name.lower()] > 1))
    return items


def detect_modules(inventory: list[InventoryItem], modules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    detected = []
    search_rows = [_norm(f"{i.filename} {i.parent_folder} {i.inferred_document_type}") for i in inventory]
    for m in modules:
        trigger_hits = []
        strong_hits = 0
        for idx, row in enumerate(search_rows):
            matched = [t for t in m.get("trigger_docs", []) if _contains_phrase(row, t)]
            if matched:
                trigger_hits.append({"file": inventory[idx].filename, "matches": matched})
            strong_hits += sum(1 for t in m.get("strong_trigger_docs", []) if _contains_phrase(row, t))
        if trigger_hits:
            if strong_hits >= 2 or len(trigger_hits) >= 2:
                strength, conf = "Strong", 0.9
            elif strong_hits == 1:
                strength, conf = "Medium", 0.75
            else:
                strength, conf = "Weak", 0.55
            detected.append({"module_id": m["module_id"], "display_name": m["display_name"], "confidence": conf, "strength": strength, "triggering_documents": trigger_hits, "evidence": f"{strength} signal from trigger document patterns.", "notes": "Conservative module inference for preparer review."})
    return detected


def _evaluate_item(item: dict[str, Any], detected: dict[str, Any], module_name: str, inventory: list[InventoryItem], item_kind: str) -> dict[str, Any]:
    aliases = item.get("aliases", [])
    matches_files, matched_aliases = [], []
    for inv in inventory:
        row = _norm(f"{inv.filename} {inv.parent_folder} {inv.inferred_document_type}")
        hits = [a for a in aliases if _contains_phrase(row, a)]
        if hits:
            matches_files.append(inv.filename)
            matched_aliases.extend(hits)
    matched_aliases = sorted(set(matched_aliases))
    cond = bool(item.get("conditional", False))
    found = bool(matches_files)
    if found:
        status = "Found"
    elif item_kind == "optional":
        status = "Optional"
    elif cond or detected["strength"] == "Weak":
        status = "Needs Review"
    elif detected["strength"] == "Strong" and item_kind == "critical":
        status = "Missing"
    else:
        status = "Needs Review"
    return {
        "module": module_name,
        "item_id": item.get("id", ""),
        "item": item.get("label", ""),
        "status": status,
        "reason": item.get("condition_note", "Checklist item evaluation"),
        "matched_files": matches_files,
        "matched_aliases": matched_aliases,
        "evidence": f"Module strength={detected['strength']}; aliases checked={', '.join(aliases)}",
        "confidence": detected["confidence"],
    }


def generate_missing_items(detected_modules: list[dict[str, Any]], modules: list[dict[str, Any]], inventory: list[InventoryItem]) -> list[dict[str, Any]]:
    by_id = {m["module_id"]: m for m in modules}
    rows = []
    for d in detected_modules:
        mod = by_id[d["module_id"]]
        for item in mod.get("expected_critical_docs", []):
            rows.append(_evaluate_item(item, d, mod["display_name"], inventory, "critical"))
        for item in mod.get("review_needed_docs", []):
            rows.append(_evaluate_item(item, d, mod["display_name"], inventory, "review"))
        for item in mod.get("optional_support_docs", []):
            rows.append(_evaluate_item(item, d, mod["display_name"], inventory, "optional"))
    return rows


def _style_sheet(ws) -> None:
    head_fill = PatternFill("solid", fgColor=SATC["navy"])
    head_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or "")) + 2, 16), 42)


def write_outputs(output_dir: Path, inventory: list[InventoryItem], detected: list[dict[str, Any]], missing: list[dict[str, Any]], modules: list[dict[str, Any]]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    inv_records = [asdict(i) for i in inventory]
    (output_dir / "inventory.json").write_text(json.dumps(inv_records, indent=2), encoding="utf-8")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Inventory"
    inv_headers = list(inv_records[0].keys()) if inv_records else []
    ws1.append(inv_headers)
    for r in inv_records:
        ws1.append([r[h] for h in inv_headers])
    _style_sheet(ws1)

    ws2 = wb.create_sheet("Detected Modules")
    ws2.append(["module_id", "display_name", "strength", "confidence", "evidence", "triggering_documents"])
    for d in detected:
        ws2.append([d["module_id"], d["display_name"], d["strength"], d["confidence"], d["evidence"], json.dumps(d["triggering_documents"])])
    _style_sheet(ws2)

    ws3 = wb.create_sheet("Missing and Review Items")
    headers = ["module", "item_id", "item", "status", "reason", "matched_files", "matched_aliases", "evidence", "confidence"]
    ws3.append(headers)
    for row in missing:
        ws3.append([row[h] if h not in {"matched_files", "matched_aliases"} else ", ".join(row[h]) for h in headers])
    _style_sheet(ws3)
    wb.save(output_dir / "inventory.xlsx")

    questions = defaultdict(list)
    mod_map = {m["display_name"]: m for m in modules}
    for r in missing:
        if r["status"] in {"Missing", "Needs Review"}:
            questions[r["module"]].extend(mod_map.get(r["module"], {}).get("client_questions", []))
    lines = ["Additional Items / Questions Needed", ""]
    for mod, qs in questions.items():
        lines.append(mod)
        for i, q in enumerate(dict.fromkeys(qs), 1):
            lines.append(f"{i}. {q}")
        lines.append("")
    (output_dir / "client_follow_up_questions.txt").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")

    prep = ["Tax Packet QA - Preparer Notes", "", "Detected Modules:"]
    for d in detected:
        prep.append(f"- {d['display_name']} [{d['strength']}] ({d['confidence']}): {d['evidence']}")
    prep.append("\nMissing / Needs Review:")
    for r in missing:
        if r["status"] in {"Missing", "Needs Review"}:
            prep.append(f"- [{r['status']}] {r['module']} | {r['item']} | matched files: {', '.join(r['matched_files']) if r['matched_files'] else 'none'}")
    (output_dir / "preparer_notes.txt").write_text("\n".join(prep) + "\n", encoding="utf-8")


def _render_template(path: Path, replacements: dict[str, str]) -> str:
    content = path.read_text(encoding="utf-8")
    for k, v in replacements.items():
        content = content.replace("{{ " + k + " }}", v).replace("{{" + k + "}}", v)
    return content


def render_reports(output_dir: Path, inventory: list[InventoryItem], detected: list[dict[str, Any]], missing: list[dict[str, Any]], client_name: str, modules: list[dict[str, Any]]) -> None:
    ts = datetime.now(UTC).isoformat()
    counts = Counter(i.inferred_document_type for i in inventory)
    inv_rows = "".join([f"<tr><td>{html.escape(i.filename)}</td><td>{html.escape(i.parent_folder)}</td><td>{html.escape(i.inferred_document_type)}</td><td>{i.inferred_tax_year}</td><td>{html.escape(i.inferred_entity_name)}</td><td>{i.confidence}</td><td>{html.escape(i.evidence)}</td><td>{i.duplicate_flag}</td></tr>" for i in inventory])
    doc_counts = "".join([f"<li>{html.escape(k)}: {v}</li>" for k, v in counts.items()])

    inv_html = _render_template(Path(__file__).parent / "templates" / "inventory_report.html", {
        "client_name": html.escape(client_name), "run_ts": ts, "file_count": str(len(inventory)), "doc_counts": doc_counts, "inventory_rows": inv_rows
    })
    (output_dir / "inventory_report.html").write_text(inv_html, encoding="utf-8")

    detected_rows = "".join([f"<li>{html.escape(d['display_name'])} ({d['strength']}, {d['confidence']}): {html.escape(d['evidence'])}</li>" for d in detected])
    miss_rows = "".join([f"<tr><td>{html.escape(r['module'])}</td><td>{html.escape(r['item'])}</td><td>{r['status']}</td><td>{html.escape(r['reason'])}</td><td>{html.escape(', '.join(r['matched_aliases']))}</td><td>{html.escape(', '.join(r['matched_files']))}</td><td>{r['confidence']}</td></tr>" for r in missing if r['status'] in {'Missing','Needs Review'}])
    prep_notes = "".join([f"<li>{html.escape(n)}</li>" for m in modules for n in m.get('preparer_notes', [])])
    miss_html = _render_template(Path(__file__).parent / "templates" / "missing_items_report.html", {
        "client_name": html.escape(client_name), "run_ts": ts, "detected_modules": detected_rows, "missing_rows": miss_rows, "preparer_notes": prep_notes
    })
    (output_dir / "missing_items_report.html").write_text(miss_html, encoding="utf-8")


def run(input_folder: Path, config_path: Path, output_dir: Path, metadata_path: Path | None = None) -> None:
    cfg = load_config(config_path)
    meta = load_config(metadata_path) if metadata_path else {}
    inv = build_inventory(input_folder)
    modules = cfg.get("modules", [])
    det = detect_modules(inv, modules)
    miss = generate_missing_items(det, modules, inv)
    write_outputs(output_dir, inv, det, miss, modules)
    render_reports(output_dir, inv, det, miss, meta.get("client_name", input_folder.name), modules)


def main() -> None:
    p = argparse.ArgumentParser(description="Tax Packet QA v0.1")
    p.add_argument("input_folder")
    p.add_argument("--client-metadata", default=None)
    p.add_argument("--config", default=str(Path(__file__).parent / "config" / "tax_modules.yaml"))
    p.add_argument("--output", default=str(Path(__file__).parent / "outputs" / "run_output"))
    a = p.parse_args()
    run(Path(a.input_folder), Path(a.config), Path(a.output), Path(a.client_metadata) if a.client_metadata else None)


if __name__ == "__main__":
    main()
