from pathlib import Path
from openpyxl import load_workbook
from tax_packet_qa import build_inventory, detect_modules, load_config, generate_missing_items, write_outputs, run


def test_missing_review_and_outputs_written(tmp_path):
    base = Path(__file__).resolve().parents[1]
    inventory = build_inventory(base / "sample_client" / "sorted_documents")
    config = load_config(base / "config" / "tax_modules.yaml")
    detected = detect_modules(inventory, config["modules"])
    rows = generate_missing_items(detected, config["modules"], inventory)
    assert any(r["status"] in {"Missing", "Needs Review"} for r in rows)
    assert any(r["matched_aliases"] for r in rows if r["status"] == "Found")

    write_outputs(tmp_path, inventory, detected, rows, config["modules"])
    assert (tmp_path / "inventory.json").exists()
    assert (tmp_path / "inventory.xlsx").exists()
    assert (tmp_path / "client_follow_up_questions.txt").exists()
    assert (tmp_path / "preparer_notes.txt").exists()

    wb = load_workbook(tmp_path / "inventory.xlsx")
    assert {"Inventory", "Detected Modules", "Missing and Review Items"}.issubset(set(wb.sheetnames))


def test_weak_module_gives_needs_review_not_missing(tmp_path):
    weak_root = tmp_path / "input" / "Rental"
    weak_root.mkdir(parents=True)
    (weak_root / "Lease_2025.pdf").write_text("x", encoding="utf-8")
    inventory = build_inventory(tmp_path / "input")
    base = Path(__file__).resolve().parents[1]
    config = load_config(base / "config" / "tax_modules.yaml")
    detected = detect_modules(inventory, config["modules"])
    rows = generate_missing_items(detected, config["modules"], inventory)
    rental_rows = [r for r in rows if r["module"] == "Rental Property"]
    assert rental_rows
    assert any(r["status"] == "Needs Review" for r in rental_rows)
    assert all(not (r["status"] == "Missing" and "if" in r["reason"].lower()) for r in rental_rows)


def test_source_files_not_modified(tmp_path):
    src = tmp_path / "input" / "01_W2"
    src.mkdir(parents=True)
    file_path = src / "W2_Test_2025.pdf"
    file_path.write_text("placeholder", encoding="utf-8")
    before = file_path.stat().st_mtime_ns

    base = Path(__file__).resolve().parents[1]
    run(tmp_path / "input", base / "config" / "tax_modules.yaml", tmp_path / "outputs")

    assert before == file_path.stat().st_mtime_ns
