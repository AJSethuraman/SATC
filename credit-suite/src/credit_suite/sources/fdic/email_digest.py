"""The monitoring email, composed from a run digest.

Half of the shipped `email_sim.py` acceptance (contract section 9). The other
half -- copy ONLY the .xlsm into an empty folder, extract `runner.py` from
`_code_py`, and have it rebuild the workbook with nothing else present -- is
what the build-time inliner delivers, so it lives with issue #167. Splitting
them here rather than quietly dropping the harder half.

`extract_code_tab` reproduces the VBA button byte for byte, and is kept beside
the composer because both are about the workbook being the source of truth.

Deterministic from `_config` + FdicDemoProvider at a fixed `--asof`.
"""

import json
import os
import shutil
import sys

import openpyxl

from credit_suite.sources.fdic import engine_api as R

XLSM_NAME = "Bank_Peer_Monitor.xlsm"
ASOF = "2026-03-31"

DIMENSIONS = ["asset_quality", "composite", "capital", "earnings", "funding",
              "concentration",
              # v1.1 competitor pack dimensions
              "consumer_credit", "commercial_credit", "funding_stress"]

# v1.1: the per-CLASS loan-book alert section groups by loan class (the
# runner's LOANBOOK_CLASS map), consumer track first, commercial floor after.
LOAN_CLASSES = list(R.CONSUMER_CLASSES) + list(R.COMMERCIAL_CLASSES)


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write UTF-8."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def compose_email(status):
    """Deterministic monitoring email from the run status/digest. No send."""
    digest = status.get("digest", {})
    banks = digest.get("banks", [])
    fresh = [b for b in banks if not b.get("stale")]
    stale = [b for b in banks if b.get("stale")]
    lines = []
    lines.append("Subject: Bank Peer Monitor -- "
                 f"{status['alert_banks']} ALERT / {status['watch_banks']} "
                 f"WATCH / {status['stale_banks']} STALE banks "
                 f"({status['timestamp']})")
    lines.append("")
    lines.append(f"Run mode: {status['mode']}  |  banks landed: "
                 f"{status['banks_landed']}/{status['banks_active']}  |  "
                 f"refused: {len(status.get('watchlist_refusals', []))}")
    # the data-vintage line (meta.index.createTimestamp on live runs)
    lines.append(f"Data vintage: {status.get('vintage') or 'n/a'}  |  "
                 f"peer-set latest REPDTE: {digest.get('set_max_repdte')}")
    lines.append("")
    lines.append("RANKED PEER TABLE (by ALERT flags, then Texas; "
                 "stale banks excluded)")
    lines.append("-" * 68)
    lines.append(f"  {'#':>2} {'Bank':<28} {'CERT':>7} {'Group':<13} "
                 f"{'Texas':>7} {'A':>2} {'W':>2}  Status")
    ranked = sorted(fresh, key=lambda b: (-(b["alert_count"]),
                                          -(b["texas"] or 0.0)))
    for i, b in enumerate(ranked, start=1):
        tex = "n/a" if b["texas"] is None else f"{b['texas']:.1f}"
        lines.append(f"  {i:>2} {b['name']:<28} {b['cert']:>7} "
                     f"{b['group']:<13} {tex:>7} {b['alert_count']:>2} "
                     f"{b['watch_count']:>2}  {b['status']}")
    if not ranked:
        lines.append("  (no banks landed)")
    lines.append("")
    lines.append("PER-DIMENSION ALERTS (metric-level, non-stale banks)")
    lines.append("-" * 52)
    any_dim = False
    for dim in DIMENSIONS:
        hits = []
        for b in fresh:
            for mid, m in b["metrics"].items():
                if m["dimension"] == dim and m["status"] == "ALERT":
                    hits.append(f"{b['name']} {mid}={m['value']:.2f}")
        if hits:
            any_dim = True
            lines.append(f"  {dim:<14} " + "; ".join(sorted(hits)))
    if not any_dim:
        lines.append("  (no metric-level alerts)")
    lines.append("")
    lines.append("PER-CLASS LOAN-BOOK ALERTS (v1.1 pack; consumer = DQ/NCO "
                 "track, commercial = Call-Report floor)")
    lines.append("-" * 68)
    any_cls = False
    for cls in LOAN_CLASSES:
        hits = []
        for b in fresh:
            for mid, m in b["metrics"].items():
                if (R.LOANBOOK_CLASS.get(mid) == cls
                        and m["status"] == "ALERT"):
                    hits.append(f"{b['name']} {mid}={m['value']:.2f}")
        if hits:
            any_cls = True
            lines.append(f"  {cls:<15} " + "; ".join(sorted(hits)))
    if not any_cls:
        lines.append("  (no per-class loan-book alerts)")
    lines.append("")
    lines.append("STALENESS FLAGS (excluded from alert counts -- possible "
                 "merger/closure)")
    lines.append("-" * 68)
    if stale:
        for b in sorted(stale, key=lambda x: x["slot"]):
            lines.append(f"  STALE  s{b['slot']:02d} {b['name']:<28} "
                         f"last REPDTE {b['asof_period']}")
            for n in b.get("notes", []):
                lines.append(f"         {n}")
    else:
        lines.append("  (no stale banks -- every bank reports the peer-set "
                     "latest quarter)")
    lines.append("")
    lines.append("WATCHLIST LANE")
    lines.append("--------------")
    refusals = status.get("watchlist_refusals", [])
    if refusals:
        for m in refusals:
            lines.append("  " + m)
    else:
        lines.append(f"  ADMITTED (CERT-keyed Class A): "
                     f"{len(status.get('watchlist_admitted', []))} banks")
    notes = [n for b in fresh for n in b.get("notes", [])]
    if notes:
        lines.append("")
        lines.append("NOTES (survivorship / roster)")
        for n in notes:
            lines.append("  " + n)
    lines.append("")
    return "\n".join(lines)
