#!/usr/bin/env python3
"""Bank Peer Monitor -- the FDIC runner, on the shared credit-suite engine.

    python runner.py --workbook Bank_Peer_Monitor.xlsm [--demo] [--asof YYYY-MM-DD]
    python runner.py --lookup "KeyBank"            resolve a name to a CERT
    python runner.py -w BOOK.xlsm --tieout 628 [20260331]

Exit codes (TEMPLATE_CONTRACT section 4): 0 OK, 1 run error, 2 gate error,
3 missing secret. JSON status on stdout, human summary on stderr.

Everything generic lives in ``credit_suite.engine``. What is here is FDIC's:
the status wording, the three digest annotations (merger survivorship, the
uninsured-deposit null, the roster ACTIVE flag), the tie-out rendering against
the Call Report facsimile, and the CERT lookup.

The status dict keeps FDIC's published key names (``banks``, ``cert``,
``texas``, ``set_max_repdte``) rather than the engine's generic ones. That is a
deliberate compatibility surface: the monitoring email and the acceptance
harness read those keys, and renaming them would be an output change in a change
whose whole purpose is that outputs do not change.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from typing import Dict, List, Optional

from credit_suite.engine import runtime
from credit_suite.engine.config import Config
from credit_suite.engine.digest import EntityContext
from credit_suite.engine.gates import EntityCapacityError, WatchlistRefused
from credit_suite.engine.metrics import MetricError
from credit_suite.engine.provider import MissingSecret, resolve_secret
from credit_suite.engine.workbook import OpenpyxlBackend, RawLayoutMismatch
from credit_suite.sources.fdic import adapter, fields
from credit_suite.sources.fdic.engine_api import (DASH_TABS, STATUS_COL,
                                                  STATUS_COL_BY_TAB)
from credit_suite.sources.fdic.spec import FDIC

CDR_FACSIMILE_URL = ("https://cdr.ffiec.gov/Public/ViewFacsimileDirect.aspx"
                     "?ds=call&idType=fdiccert&id={cert}&date={mmddyyyy}")
BANKFIND_URL = ("https://banks.data.fdic.gov/bankfind-suite/bankfind/"
                "details/{cert}")


# --------------------------------------------------------------------------
# FDIC's digest annotations -- what makes a blank or an odd figure auditable
# --------------------------------------------------------------------------

def annotate_merger_survivorship(ctx: EntityContext) -> List[str]:
    """Trap F6: a >25% single-quarter asset jump usually means an acquisition,
    which makes every ratio and growth screen for that quarter unreliable."""
    if len(ctx.periods) < 2:
        return []
    newest = ctx.periods[0][1].get("ASSET")
    prior = ctx.periods[1][1].get("ASSET")
    if newest is None or prior in (None, 0):
        return []
    jump = (newest / prior - 1.0) * 100.0
    if jump <= 25.0:
        return []
    return ["ASSET +%.0f%% in one quarter -- possible merger survivorship "
            "distortion (trap F6); ratios and CRE growth screens unreliable "
            "this quarter." % jump]


def annotate_uninsured_deposit_null(ctx: EntityContext) -> List[str]:
    """Trap F3, made auditable: DEPUNINS is genuinely null below the $1B
    reporting threshold, so the blank is explained rather than shown as 0."""
    if not ctx.latest_fields:
        return []
    if ctx.latest_fields.get("DEPUNINS") is not None:
        return []
    if not any(v is not None for v in ctx.latest_fields.values()):
        return []
    return ["DEPUNINS is null this quarter -- uninsured-deposit share left "
            "BLANK, never 0 (RC-O Mem 2 is filed by $1B+ reporters; smaller "
            "banks carry FDIC estimates or nulls)."]


def annotate_roster_inactive(ctx: EntityContext) -> List[str]:
    """The institutions roster is the independent check on a bank still being
    a bank -- a merged one keeps answering the financials endpoint."""
    roster = ctx.roster_row
    if not roster:
        return []
    active = str(roster.get("ACTIVE", "1")).strip()
    if active in ("1", "True", "true"):
        return []
    return ["roster reports ACTIVE=%s (ENDEFYMD=%s) -- possible "
            "merger/closure; check /institutions and /history."
            % (roster.get("ACTIVE"), roster.get("ENDEFYMD"))]


ANNOTATORS = (annotate_merger_survivorship, annotate_uninsured_deposit_null,
              annotate_roster_inactive)


# --------------------------------------------------------------------------
# the run
# --------------------------------------------------------------------------

def _to_published_status(status: dict) -> dict:
    """The engine's generic keys -> FDIC's published ones.

    Not cosmetic: the monitoring email and the acceptance harness read these
    names, and a rename would be an output change inside a change whose whole
    purpose is that outputs do not change.
    """
    digest = status.pop("digest")
    banks = []
    for entity in digest["entities"]:
        bank = dict(entity)
        bank["cert"] = bank.pop("key")
        bank["texas"] = bank.pop("headline")
        banks.append(bank)

    out = dict(status)
    out["peer_slots"] = out.pop("entity_slots")
    out["banks_active"] = out.pop("entities_active")
    out["banks_landed"] = out.pop("entities_landed")
    out["banks_excluded"] = out.pop("entities_excluded")
    out["alert_banks"] = out.pop("alert_entities")
    out["watch_banks"] = out.pop("watch_entities")
    out["stale_banks"] = out.pop("stale_entities")
    out["digest"] = {"banks": banks, "medians": digest["medians"],
                     "set_max_repdte": digest["set_max_period"]}
    return out


def status_lines(status: dict) -> List[str]:
    """The three status-panel lines, in FDIC's vocabulary."""
    line2 = ("Banks %s/%s - %s ALERT / %s WATCH / %s STALE"
             % (status.get("banks_landed", 0), status.get("banks_active", 0),
                status.get("alert_banks", 0), status.get("watch_banks", 0),
                status.get("stale_banks", 0)))
    refused = len(status.get("watchlist_refusals") or [])
    if refused:
        line2 += " - %d REFUSED (see digest)" % refused
    errors = len(status.get("errors") or [])
    if errors:
        line2 += " - %d FETCH ERRORS (slots blanked)" % errors
    return [
        "Last run  %s (%s)" % (status.get("timestamp", ""), status.get("mode", "")),
        line2,
        "FDIC data vintage: %s" % (status.get("vintage") or "n/a"),
    ]


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None,
        provider=None) -> dict:
    """One refresh against the CLOSED workbook.

    ``provider`` is injectable so a test can drive a frozen or failing source
    without the runner knowing how one is built.
    """
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path, FDIC, fields.RAW_FIELDS)
    cfg = backend.read_config()

    if provider is None:
        provider = adapter.make_provider(cfg, demo, asof)
    mode = "demo" if isinstance(provider, adapter.FdicDemoProvider) else "live"

    status = runtime.run_refresh(
        backend=backend, cfg=cfg, provider=provider,
        registry=fields.REGISTRY, fields=fields.RAW_FIELDS,
        field_units=fields.FIELD_UNITS, asof=asof, mode=mode,
        annotators=ANNOTATORS, headline_metric="TEXAS",
        secret=resolve_secret(cfg))

    published = _to_published_status(status)
    backend.write_status_lines(status_lines(published), DASH_TABS,
                               STATUS_COL_BY_TAB, STATUS_COL)
    backend.finalize()
    return published


# --------------------------------------------------------------------------
# tie-out (contract section 12) and lookup (section 13)
# --------------------------------------------------------------------------

def facsimile_url(cert: str, iso_repdte: str) -> str:
    """The filed Call Report's public facsimile, keyed by CERT + MMDDYYYY.

    This URL is the whole point of provenance: it opens the actual filing a
    reviewer reads to check a flagged number by eye.
    """
    d = str(iso_repdte)[:10]
    return CDR_FACSIMILE_URL.format(
        cert=cert, mmddyyyy="%s%s%s" % (d[5:7], d[8:10], d[0:4]))


def bankfind_url(cert: str) -> str:
    return BANKFIND_URL.format(cert=cert)


def read_provenance_rows(wb) -> Dict[str, dict]:
    """Parse the `_provenance` tab into ``{field: {...}}``. Empty when absent."""
    if "_provenance" not in wb.sheetnames:
        return {}
    ws = wb["_provenance"]
    rows: Dict[str, dict] = {}
    started = False
    for row in ws.iter_rows(values_only=True):
        first = "" if not row or row[0] is None else str(row[0]).strip()
        if not started:
            started = first.lower() == "field"
            continue
        if not first:
            continue
        cells = list(row) + [None] * 6
        rows[first] = {
            "field": first, "schedule": cells[1] or "", "caption": cells[2] or "",
            "mdrm": cells[3] or "", "flag": cells[4] or "", "notes": cells[5] or "",
        }
    return rows


def _cmd_lookup(name: str, demo: bool) -> int:
    """Resolve an institution name to its CERT.

    Live-only, and the refusal says so in the words the shipped monitor used:
    it names the endpoint, says the API is keyless (so nobody goes hunting for
    a key that does not exist), and prints the exact command to run.
    """
    if demo:
        sys.stderr.write(
            "--lookup is LIVE-ONLY: it queries the FDIC BankFind API "
            "(https://api.fdic.gov/banks/institutions) and cannot run in "
            "--demo/offline mode. Run it without --demo -- the API is "
            "keyless, no API key or account is needed:\n"
            '    python runner.py --lookup "' + name + '"\n')
        return runtime.EXIT_GATE_ERROR
    try:
        hits = adapter.FdicProvider().lookup(name)
    except Exception as exc:                       # noqa: BLE001
        sys.stderr.write("lookup failed: %s\n"
                         "Check network/proxy access to api.fdic.gov "
                         "(keyless; no API key needed).\n" % exc)
        return runtime.EXIT_RUN_ERROR
    if not hits:
        print('No institutions matched "%s".' % name)
        return runtime.EXIT_OK
    print("%8s  %6s  %14s  %-20s %-3s NAME"
          % ("CERT", "ACTIVE", "ASSET($000)", "CITY", "ST"))
    for row in hits:
        asset = row.get("ASSET")
        asset_s = "" if asset is None else "{:,.0f}".format(float(asset))
        print("%8s  %6s  %14s  %-20s %-3s %s"
              % (str(row.get("CERT", "")), str(row.get("ACTIVE", "")), asset_s,
                 str(row.get("CITY", "")), str(row.get("STALP", "")),
                 row.get("NAME", "")))
    print("\nPut the CERT into the _config [PEERS] table "
          "(slot | cert | name | group | active) and re-run the runner -- "
          "no rebuild needed within the built slot capacity.")
    return runtime.EXIT_OK



def _wrap(text: str, width: int) -> List[str]:
    """Wrap without importing textwrap for one call, and never mid-code:
    an MDRM code split across two lines stops being searchable, which is the
    one property it exists for."""
    words = str(text).split()
    lines, current = [], ""
    for word in words:
        if current and len(current) + 1 + len(word) > width:
            lines.append(current)
            current = word
        else:
            current = (current + " " + word) if current else word
    if current:
        lines.append(current)
    return lines or [""]


def _describe_source(schedule: str, caption: str) -> List[tuple]:
    """Label the two provenance columns by WHAT THEY HOLD, not by position.

    The `_provenance` tab uses three conventions across the same column pair and
    a fixed label is wrong for some row under every choice:

      NCLNLSR   schedule="FDIC-computed ratio"        caption="100 x (1407+1403)/2122"
      PD3089R   schedule="derived: P3LNLS/LNLSGR*100" caption="RC-N 9 col A over RC-C Pt I 12"
      RBC1AAJ   schedule="FILED: RC-R Pt I 31"        caption="Leverage ratio"

    So each value is classified on its own: a form location, a calculation, or
    the form's own wording. Getting this wrong is not cosmetic -- printing
    "Where on form: FDIC-computed ratio" tells a reviewer to go and look for a
    page that does not exist.
    """
    def classify(text: str) -> str:
        t = (text or "").strip()
        if not t:
            return "empty"
        low = t.lower()
        # Arithmetic wins over a mention. "100 x annualized qtr (4635-4605) /
        # RC-K avg loans" names a schedule but IS a calculation, and calling it
        # a location sends a reviewer looking for a page of arithmetic.
        if (low.startswith(("derived", "fdic-computed", "filed but"))
                or any(ch in t for ch in "*/+")
                or " x " in low):
            return "calc"
        if (t.startswith("FILED:") or t.startswith("RC") or t.startswith("RI")
                or " RC-" in t):
            return "where"
        return "wording"

    LABELS = {"where": "Where on form", "calc": "How built",
              "wording": "Filed as"}
    out = []
    for text in (schedule, caption):
        kind = classify(text)
        if kind == "empty":
            continue
        label = LABELS[kind]
        # Two of the same kind: the second is the detail of the first.
        if out and out[-1][0] == label:
            label = "  ... also"
        out.append((label, str(text).strip()))
    return out


def do_tieout(workbook: str, args: List[str], demo: bool,
              brief: bool = False, filing: bool = False) -> int:
    """``--tieout CERT [REPDTE]``: sample-verify the feed against the filing.

    Contract section 12. The point is that a flagged number can be defended in
    minutes: every value is printed beside the Call Report schedule, line and
    MDRM code it came from, plus the URL of the official facsimile a reviewer
    can open and read.

    Works in ``--demo`` too, loudly labelled. The values are fiction there but
    the schedule/line/MDRM columns are the real tie-out map either way, which is
    the half worth checking offline.
    """
    import re

    import openpyxl

    from credit_suite.engine.config import EntityRow, norm_key
    from credit_suite.sources.fdic import plain
    from credit_suite.sources.fdic.engine_api import (assemble_quarters,
                                                      make_field_spec,
                                                      metric_value,
                                                      validate_metrics)

    cert = norm_key(args[0])
    repdte = args[1] if len(args) > 1 else None
    if not FDIC.entity.admits("cert:%s" % cert):
        sys.stderr.write("--tieout: '%s' is not a valid FDIC CERT (digits "
                         'only). Find one with: python runner.py --lookup '
                         '"<bank name>".\n' % args[0])
        return runtime.EXIT_GATE_ERROR

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=False)
    try:
        rows = [list(row) for row in wb["_config"].iter_rows(values_only=True)]
        from credit_suite.engine.config import parse_config
        cfg = parse_config(rows, FDIC)
        prov_rows = read_provenance_rows(wb)
    finally:
        wb.close()

    if not prov_rows:
        sys.stderr.write("--tieout: this workbook has no _provenance tab -- it "
                         "predates pack v1.1; rebuild it.\n")
        return runtime.EXIT_RUN_ERROR

    validate_metrics(cfg.series)
    provider = adapter.make_provider(cfg, demo, None)
    mode = "demo" if isinstance(provider, adapter.FdicDemoProvider) else "live"
    try:
        provider.prime([cert], date.today(), names={cert: "cert:%s" % cert})
        entity = EntityRow(slot=1, key=cert, name="cert:%s" % cert,
                           group="tieout", active=True, key_prefix="cert")
        field_rows = {f: provider.fetch_series(make_field_spec(entity, f))
                      for f in fields.RAW_FIELDS}
    except Exception as exc:                       # noqa: BLE001
        sys.stderr.write("--tieout fetch failed for cert %s: %s\n" % (cert, exc))
        return runtime.EXIT_RUN_ERROR

    quarters = assemble_quarters(field_rows, cfg.raw_slots)
    if not quarters:
        sys.stderr.write("--tieout: no quarters returned for cert %s.\n" % cert)
        return runtime.EXIT_RUN_ERROR

    if repdte:
        wanted = str(repdte).strip()
        if len(wanted) == 8 and wanted.isdigit():
            wanted = "%s-%s-%s" % (wanted[0:4], wanted[4:6], wanted[6:8])
        match = [(p, f) for p, f in quarters if p == wanted]
        if not match:
            sys.stderr.write("--tieout: REPDTE %s not in the fetched window "
                             "(%s .. %s).\n"
                             % (repdte, quarters[-1][0], quarters[0][0]))
            return runtime.EXIT_RUN_ERROR
        iso, values = match[0]
    else:
        iso, values = quarters[0]

    print("TIE-OUT  cert %s  REPDTE %s  (pack %s, %s mode)"
          % (cert, iso, FDIC.pack_version, mode))
    if mode == "demo":
        print("  *** DEMO VALUES -- deterministic FdicDemoProvider fiction, "
              "NOT this bank's filing; the schedule/line/MDRM columns are "
              "the real tie-out map. Re-run without --demo for live "
              "values. ***")
    print("  Call Report facsimile : %s" % facsimile_url(cert, iso))
    print("  BankFind (pull check) : %s" % bankfind_url(cert))
    print("  Data vintage          : %s" % (provider.vintage or "n/a"))
    print("  Tie-out is two-step: API value <-> BankFind page (pull check) "
          "<-> CDR facsimile schedule/line (filing check).")
    print()
    if not brief:
        print("  WHAT THE TERMS MEAN")
        for term, meaning in plain.GLOSSARY:
            first = True
            for line in _wrap(meaning, 88):
                print("    %-16s %s" % (("%s -" % term) if first else "", line))
                first = False
        print()

    def _confidence(flag: str) -> str:
        """The flag column in words. `[V]` is not self-explanatory and its
        absence is the part that actually matters to a reviewer."""
        f = (flag or "").strip()
        if "[V]" in f and "comp" in f:
            return "VERIFIED (computed from verified lines)"
        if "[V]" in f:
            return "VERIFIED against the FFIEC form caption"
        if "[~]" in f:
            return "PARTIAL - matched by caption, not by code"
        return "NOT VERIFIED" if not f else f

    missing = []
    for series in cfg.series:
        value = metric_value(series.id, values)
        shown = "(blank)" if value is None else "{:,.2f}".format(value)
        row = prov_rows.get(series.id)
        if row is None:
            missing.append(series.id)
            row = {"schedule": "(no provenance row)", "caption": "",
                   "mdrm": "", "flag": "", "notes": ""}

        print("  %s = %s" % (series.id, shown))
        meaning = plain.describe(series.id)
        if meaning:
            for i, line in enumerate(_wrap(meaning, 84)):
                print("      %-14s %s" % ("What it is" if i == 0 else "", line))
        for label, text in _describe_source(row["schedule"], row["caption"]):
            for i, line in enumerate(_wrap(text, 84)):
                print("      %-14s %s" % (label if i == 0 else "", line))
        if row["mdrm"]:
            print("      %-14s %s" % ("Code", row["mdrm"]))
        print("      %-14s %s" % ("Checked", _confidence(row["flag"])))
        note = row["notes"] or ""
        loan_class = fields.LOANBOOK_CLASS.get(series.id)
        if loan_class:
            note = ("[%s] %s" % (loan_class, note)).strip()
        if note:
            for i, line in enumerate(_wrap(note, 84)):
                print("      %-14s %s" % ("Caveat" if i == 0 else "", line))
        print()

    if missing:
        sys.stderr.write("WARNING: no _provenance row for: %s\n"
                         % ", ".join(missing))
    print()
    print("  Per-FIELD rows (every raw input's schedule/line/MDRM) are on "
          "the workbook's _provenance tab.")

    if filing:
        _tieout_to_filing(cert, iso, values, prov_rows, mode)
    return runtime.EXIT_OK


def _tieout_to_filing(cert: str, iso: str, landed: dict, prov_rows: dict,
                      mode: str) -> None:
    """The first link of the chain: the bank's own filed XBRL, from the FFIEC.

    Compares every RAW dollar field the workbook landed against the line(s)
    the provenance map says it comes from, evaluated in the filing itself.
    Ratios are skipped -- they are FDIC-computed, and the arithmetic leg above
    already recomputes them from these same raw lines.
    """
    from credit_suite.sources.fdic import filing as F

    print()
    print("  FILING CHECK -- the bank's own XBRL from cdr.ffiec.gov, not the "
          "FDIC's republication")
    if mode == "demo":
        print("  *** demo values cannot tie to a real filing; this compares the "
              "filing against fiction and is shown only to exercise the path. ***")
    try:
        facts = F.parse_facts(F.fetch_xbrl(cert, iso), iso)
    except Exception as exc:                          # noqa: BLE001
        print("  could not fetch the filing: %s" % exc)
        print("  -> the chain stops at the FDIC API for this bank-quarter; "
              "that is stated, not hidden.")
        return
    expressions = {f: prov_rows[f]["mdrm"] for f in fields.RAW_FIELDS
                   if f in prov_rows and prov_rows[f]["mdrm"]}
    rows = F.tie(facts, landed, expressions, units=fields.FIELD_UNITS)
    checked = [r for r in rows if not r.note]
    ties = sum(r.verdict == "TIES" for r in checked)
    differs = [r for r in checked if r.verdict.startswith("DIFFERS")]
    absent = [r for r in checked if r.verdict == "NOT IN FILING"]
    print("  %d facts in the filing for %s" % (len(facts), iso))
    print("  %d raw dollar lines compared: %d tie, %d differ, %d not in the filing; "
          "%d skipped with a stated reason"
          % (len(checked), ties, len(differs), len(absent), len(rows) - len(checked)))
    print("  filed values are DOLLARS; landed values are THOUSANDS -- shown in thousands")
    print("  %-10s %18s %18s  %-22s %s" % ("field", "filed (k$)", "landed (k$)",
                                            "filed as", "verdict"))
    for r in rows:
        filed = "-" if r.filed_thousands is None else "{:,}".format(r.filed_thousands)
        got = "-" if r.landed_thousands is None else "{:,.0f}".format(r.landed_thousands)
        print("  %-10s %18s %18s  %-22s %s" % (r.field, filed, got, r.used or "-", r.verdict))
    if differs:
        print("  DIFFERENCES ARE FINDINGS. Each one is either a line the map cites "
              "wrongly for this filer, or a real disagreement between the FDIC's "
              "republication and the filing -- and only reading the form says which.")
    if any(r.used.startswith("RCFD") for r in rows):
        print("  RCFD = consolidated (form 031, foreign offices); RCON = domestic only. "
              "The FDIC's totals are the consolidated ones -- following a bare RCON "
              "code to the facsimile for this bank lands on the wrong line.")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Bank Peer Monitor refresh")
    ap.add_argument("-w", "--workbook", default=None,
                    help="path to the .xlsm (required unless --lookup)")
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline provider; no key, no network")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    ap.add_argument("--lookup", metavar="NAME", default=None,
                    help="resolve an institution name to its CERT (live only)")
    ap.add_argument("--filing", action="store_true",
                    help="with --tieout: also fetch the bank's filed XBRL from "
                         "cdr.ffiec.gov and tie every raw field to it (network)")
    ap.add_argument("--tieout", nargs="+", metavar="ARG", default=None,
                    help="CERT [REPDTE] -- print each value with its source")
    return ap


def main(argv: Optional[List[str]] = None) -> int:
    args = build_parser().parse_args(argv)

    if args.lookup:
        return _cmd_lookup(args.lookup, args.demo)
    if not args.workbook:
        print("--workbook is required (or use --lookup).", file=sys.stderr)
        return runtime.EXIT_GATE_ERROR
    if args.tieout is not None and len(args.tieout) > 2:
        print("--tieout takes CERT [REPDTE].", file=sys.stderr)
        return runtime.EXIT_GATE_ERROR
    if args.tieout:
        try:
            return do_tieout(args.workbook, args.tieout, args.demo,
                             filing=args.filing)
        except Exception as exc:                   # noqa: BLE001
            print("TIE-OUT ERROR: %s" % exc, file=sys.stderr)
            return runtime.EXIT_RUN_ERROR

    asof = (datetime.strptime(args.asof, "%Y-%m-%d").date()
            if args.asof else None)
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, MetricError, EntityCapacityError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        print("GATE ERROR: %s" % exc, file=sys.stderr)
        return runtime.EXIT_GATE_ERROR
    except MissingSecret as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        print("MISSING SECRET: %s" % exc, file=sys.stderr)
        return runtime.EXIT_MISSING_SECRET
    except (RawLayoutMismatch, Exception) as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}))
        print("RUN ERROR: %s" % exc, file=sys.stderr)
        return runtime.EXIT_RUN_ERROR

    ok = runtime.run_succeeded({
        "entities_active": status["banks_active"],
        "entities_landed": status["banks_landed"]})
    payload = {"ok": ok, **{k: v for k, v in status.items() if k != "digest"}}
    print(json.dumps(payload))

    print("%s: %s/%s banks landed, %s ALERT / %s WATCH / %s STALE banks "
          "(%s ALERT flags), refused=%s, vintage=%s"
          % ("OK (demo)" if status["mode"] == "demo" else "OK",
             status["banks_landed"], status["banks_active"],
             status["alert_banks"], status["watch_banks"],
             status["stale_banks"], status["alert_flags"],
             len(status["watchlist_refusals"]), status["vintage"]),
          file=sys.stderr)
    return runtime.EXIT_OK if ok else runtime.EXIT_NOTHING_PULLED


if __name__ == "__main__":
    raise SystemExit(main())
