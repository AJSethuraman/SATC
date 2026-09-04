"""Writing into the closed workbook, via openpyxl.

Two carried lessons live in this file, and both are the kind that pass every
test while being wrong in production.

**L2 -- ``keep_vba=True`` only for ``.xlsm``.** On an ``.xlsx`` it injects a
dangling vbaProject relationship that Excel rejects outright as "format or
extension not valid". The workbook still saves; it just will not open.

**L7 -- ``ws.cell(r, c, None)`` silently does nothing.** openpyxl ignores the
value argument when it is ``None``, so the obvious way to blank a cell is a
no-op. Cells are blanked ONLY by assigning ``.value`` explicitly. The bug this
caused was invisible: clear-blocks appeared to work because a successful run
rewrote the same shape over the top, and only a *failed* fetch after a
successful one would have left last quarter's figures sitting under this
quarter's timestamp. Every write path below assigns ``.value``.

The raw-layout check is defence against a different silent failure: every
dashboard and Watchlist formula is anchored to the layout the workbook was
*built* with, so a ``raw_slots`` edited in ``_config`` afterwards would make the
runner write correct data into cells no formula reads. That is refused with the
rebuild command rather than written.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Sequence, Tuple

from credit_suite.engine import rawlayout
from credit_suite.engine.config import (Config, EntityRow, MonitorSpec,
                                        parse_config)
from credit_suite.engine.rawlayout import SlotBlock


class RawLayoutMismatch(RuntimeError):
    """The workbook's built layout is not the one `_config` now describes.

    Refused rather than written: the data would land correctly and be read by
    nothing, which looks exactly like a workbook full of blanks.
    """


class OpenpyxlBackend:
    """The closed-workbook writer. One instance per run."""

    def __init__(self, path: str, spec: MonitorSpec, fields: Sequence[str]):
        import openpyxl

        self.path = path
        self.spec = spec
        self.fields = list(fields)
        # L2: keep_vba only for .xlsm.
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    # -- reading ---------------------------------------------------------

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        return parse_config(rows, self.spec)

    @property
    def sheetnames(self) -> List[str]:
        return list(self._wb.sheetnames)

    # -- the layout guard ------------------------------------------------

    def _check_slot(self, block: SlotBlock) -> None:
        ws = self._wb[self.spec.raw_tab]
        label = rawlayout.slot_label(block.slot)

        existing = ws.cell(block.header_row, 1).value
        if existing is not None and str(existing).strip() not in ("", label):
            raise RawLayoutMismatch(
                "Raw layout mismatch in %s at row %d: expected block '%s' but "
                "found '%s'. The workbook was built with a different "
                "raw_slots/entity_slots layout than _config now describes; "
                "formulas are anchored to the built layout. Rebuild the "
                "workbook (%s N) instead of editing the layout settings in "
                "place." % (self.spec.raw_tab, block.header_row, label,
                            existing, self.spec.rebuild_command))

        # Pack guard: the FIELD layout must match too. A workbook built by an
        # earlier pack has fewer raw columns, and writing this field set into
        # it would land data under no label, read by stale formulas. The last
        # field's label column is the cheap sentinel.
        last = self.fields[-1]
        col = rawlayout.field_col(last, self.fields)
        found = ws.cell(block.label_row, col).value
        if str(found or "").strip() != last:
            raise RawLayoutMismatch(
                "Raw layout mismatch in %s: block '%s' does not label field "
                "'%s' at column %d (found '%s'). The workbook was built by a "
                "different metric-pack version than this runner (pack %s); "
                "rebuild it or extract the matching runner from ITS OWN "
                "_code_py tab." % (self.spec.raw_tab, label, last, col, found,
                                   self.spec.pack_version))

    # -- writing ---------------------------------------------------------

    def clear_slot_block(self, block: SlotBlock) -> None:
        """Blank a slot's identity AND data -- the stateless-rebuild property.

        A deactivated, removed or failed entity must show empty, never last
        run's values masquerading as current. Column 1 of the header row keeps
        its slot label: that is the layout sentinel, not data.
        """
        self._check_slot(block)
        ws = self._wb[self.spec.raw_tab]
        for col in (2, 3, 4):                        # runtime identity cells
            ws.cell(block.header_row, col).value = None   # L7: assign .value
        last_col = 1 + len(self.fields)
        for row in range(block.first_data_row, block.first_data_row + block.slots):
            for col in range(1, last_col + 1):
                ws.cell(row, col).value = None            # L7: assign .value

    def write_slot_block(self, block: SlotBlock, entity: EntityRow,
                         periods: Sequence[Tuple[str, Dict[str, Optional[float]]]]
                         ) -> None:
        self._check_slot(block)
        ws = self._wb[self.spec.raw_tab]
        ws.cell(block.header_row, 1, rawlayout.slot_label(block.slot))
        ws.cell(block.header_row, 2, "s%02d %s" % (block.slot, entity.entity_key))
        ws.cell(block.header_row, 3, entity.name)
        ws.cell(block.header_row, 4, "group=%s" % entity.group)
        for offset, (period, values) in enumerate(periods[:block.slots]):
            row = block.first_data_row + offset          # newest-first
            ws.cell(row, 1, period)
            for fname in self.fields:
                value = values.get(fname)
                # L7 again: None must LAND as blank, so assign .value.
                ws.cell(row, rawlayout.field_col(fname, self.fields)).value = \
                    None if value is None else float(value)

    def write_status_lines(self, lines: Sequence[str],
                           tabs: Sequence[str],
                           column_by_tab: Dict[str, int],
                           default_column: int) -> None:
        """Write the status panel. Which lines to write is the source's call.

        The panel's *placement* is shared (rows 1..n of a status column, with a
        per-tab override where a banner merge runs past the default column);
        its wording is not, because it names the source's own vocabulary.
        """
        for tab in tabs:
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                col = column_by_tab.get(tab, default_column)
                for offset, line in enumerate(lines):
                    ws.cell(1 + offset, col, line)

    def finalize(self) -> None:
        self._wb.save(self.path)
