#!/usr/bin/env python3
"""Build a SECOND workbook: the same numbers, as charts.

    python tools/chartbook.py --workbook Bank_Peer_Monitor.xlsm
    python tools/chartbook.py --workbook Bank_Peer_Monitor.xlsm --out Peers_Charts.xlsx

The monitor answers "who is bad now". This answers "who is getting worse, and
how does that compare", which is the question a credit review actually asks. It
reads the history already sitting in `Raw_FDIC` -- sixteen quarters, twelve
banks -- and writes native Excel line charts over it.

WHY A SECOND WORKBOOK RATHER THAN A TAB IN THE FIRST

Carried lesson L4 says no native charts, for two reasons, and they turned out to
have different lifespans. **Both were tested on the build PC rather than
inherited**, because a lesson that blocks something valuable is worth
re-checking:

* *"the top unreadable-content / recovered trigger"* -- **did not reproduce.**
  A native openpyxl LineChart opened cleanly in Excel 16.0, twice: once in a
  bare `.xlsx`, and once added to the real `Bank_Peer_Monitor.xlsm` beside its
  VBA project, where the ExtractFiles macro still ran afterwards. Zero dialogs
  in both.
* *"they re-emit on every refresh"* -- **still true, and it is the binding
  one.** The runner rewrites the monitor on every run, so a chart tab there
  would be regenerated each time and any analyst's customisation lost.

So charts live here instead. Regenerated wholesale, nothing to re-emit. And
because this workbook carries no macros it opens with **no security banner** --
the monitor's own `.xlsm` is blocked by Excel's Mark of the Web when it arrives
by email, and a chart nobody can see is not a chart.

WHAT THE FIRST VERSION GOT WRONG, AND HOW IT WAS FOUND

The first build "opened with zero dialogs" and a harness read a cell out of it,
and on that basis it was called done. Exporting a chart to PNG and *looking*
showed two defects no harness could see: **no axis numbers at all** (openpyxl
3.1 hides axes unless told otherwise), and the peer sheet was twelve coloured
lines with the legend sitting on the data -- unreadable, and past three or four
hues indistinguishable to a colour-blind reader. Opening the artifact means
looking at it. Now:

* axes are drawn, with tick labels;
* the peer overview draws every bank in grey with the **peer median** in
  colour, so the shape of the group reads at a glance and any one bank is a
  click away (Excel highlights a series when you click it);
* below it, one small chart per bank -- that bank in colour against the peer
  median in grey -- which is the "where do we sit" view without twelve hues.

The numbers are not recomputed here. They come from the monitor's own engine, so
a chart cannot disagree with the dashboard it was drawn from.
"""

from __future__ import annotations

import argparse
import statistics
import sys
from pathlib import Path
from typing import List, Optional, Sequence

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl                                        # noqa: E402
from openpyxl.chart import LineChart, Reference        # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter           # noqa: E402

import trend                                           # noqa: E402
from trend import MEANS, WORSE_WHEN, read_panel        # noqa: E402

try:
    from credit_suite.sources.fdic import plain
except ImportError:                                    # pragma: no cover
    plain = None

#: Peer-comparison sheets: one metric, every bank.
PEER_METRICS = ["NCLNLSR", "NTLNLSQR", "PD3089R", "LNRESNCR", "TEXAS", "EQV"]

#: The four stages of a loan going wrong, for the one-bank view. Same book at
#: each stage, so the lines are comparable and the lead-lag is the point.
STAGES = [("P3CONOTH_BOOK", "30-89 days late"),
          ("P9CONOTH_BOOK", "90+ days, still accruing"),
          ("NACONOTH_BOOK", "Nonaccrual"),
          ("NTCONOTQ_BOOK", "Net charge-offs (annualised)")]

#: Colours. Four categorical hues validated by the data-viz palette checker
#: (light surface); grey for context. Hex without '#', as openpyxl wants.
HUES = ["2A78D6", "EB6834", "1BAF7A", "EDA100"]
CONTEXT = "B9B8B3"
ACCENT = HUES[0]
PT = 12700                                             # EMU per point

HEAD = PatternFill("solid", fgColor="1E3D63")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(size=9, color="667286")
MEDIAN_LABEL = "Peer median"


def _quarter(iso: str) -> str:
    return "%sQ%d" % (iso[2:4], (int(iso[5:7]) + 2) // 3)


def _sheet_name(prefix: str, name: str) -> str:
    """Excel caps sheet names at 31 chars and forbids []:*?/\\ ."""
    clean = "".join(c for c in name if c not in "[]:*?/\\")
    return ("%s_%s" % (prefix, clean))[:31]


def _median_row(rows: List[tuple]) -> List[Optional[float]]:
    """Per-quarter median across the banks, None where nobody reported."""
    if not rows:
        return []
    width = max(len(v) for _, v in rows)
    out: List[Optional[float]] = []
    for i in range(width):
        got = [v[i] for _, v in rows if i < len(v) and v[i] is not None]
        out.append(statistics.median(got) if got else None)
    return out


def _write_grid(ws, row0: int, periods: List[str], rows: List[tuple]) -> int:
    """Header + one row per series. Returns the last row written."""
    ws.cell(row0, 1, "Series").fill = HEAD
    ws.cell(row0, 1).font = HEAD_FONT
    for i, iso in enumerate(periods):
        cell = ws.cell(row0, 2 + i, _quarter(iso))
        cell.fill, cell.font = HEAD, HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    for j, (label, values) in enumerate(rows, start=1):
        ws.cell(row0 + j, 1, label)
        for i, v in enumerate(values):
            c = ws.cell(row0 + j, 2 + i)
            if v is not None:
                c.value = round(float(v), 4)
                c.number_format = "0.00"
    ws.column_dimensions["A"].width = 30
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 8
    return row0 + len(rows)


def _style(series, colour: str, width_pt: float) -> None:
    series.smooth = False
    series.marker.symbol = "none"
    series.graphicalProperties.line.solidFill = colour
    series.graphicalProperties.line.width = int(width_pt * PT)


def _base_chart(title: str, y_title: str, height: float, width: float) -> LineChart:
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height, chart.width = height, width
    chart.y_axis.title = y_title
    chart.y_axis.number_format = "0.00"
    chart.y_axis.majorGridlines = None
    # openpyxl >= 3.1 hides both axes unless told not to. The first build of
    # this file shipped charts with no numbers on either axis and a harness
    # called them fine; a human would not have.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # Legend below the plot, and NOT overlaid: without `overlay = False` Excel
    # draws it on top of the category labels, which is what the second export
    # showed. A y-axis title overlaps the tick numbers under Excel's auto
    # layout for these charts, so the measure's meaning lives in the sheet
    # header (A2) and the chart title instead.
    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.y_axis.title = None
    return chart


def _series_refs(ws, row0: int, n_rows: int, n_cols: int):
    data = Reference(ws, min_col=1, max_col=1 + n_cols,
                     min_row=row0 + 1, max_row=row0 + n_rows)
    cats = Reference(ws, min_col=2, max_col=1 + n_cols, min_row=row0, max_row=row0)
    return data, cats


def _add_stages_chart(ws, title: str, y_title: str, row0: int, n_rows: int,
                      n_cols: int, anchor: str) -> None:
    """Up to four series, each its own validated hue, direct legend."""
    chart = _base_chart(title, y_title, 9.5, 26)
    data, cats = _series_refs(ws, row0, n_rows, n_cols)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    for i, series in enumerate(chart.series):
        _style(series, HUES[i % len(HUES)], 2.0)
    ws.add_chart(chart, anchor)


def _add_peer_overview(ws, metric: str, meaning: str, row0: int, n_banks: int,
                       n_cols: int, anchor: str) -> None:
    """Every bank in grey, the peer median in colour. The last grid row is the
    median, written by the caller after the banks."""
    chart = _base_chart("%s (%s) -- all peers, with the peer median"
                        % (MEANS.get(metric, metric).capitalize(), metric),
                        meaning[:40], 9.5, 26)
    data, cats = _series_refs(ws, row0, n_banks + 1, n_cols)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)
    for i, series in enumerate(chart.series):
        if i == n_banks:                               # the median row
            _style(series, ACCENT, 2.75)
        else:
            _style(series, CONTEXT, 1.0)
    ws.add_chart(chart, anchor)


def _add_small_multiple(ws, bank: str, metric: str, row0: int, bank_row: int,
                        median_row: int, n_cols: int, anchor: str) -> None:
    """One bank in colour against the peer median in grey."""
    chart = _base_chart(bank, "", 5.2, 8.6)
    chart.legend = None
    chart.y_axis.title = None
    cats = Reference(ws, min_col=2, max_col=1 + n_cols, min_row=row0, max_row=row0)
    for r, colour, width in ((median_row, CONTEXT, 1.25), (bank_row, ACCENT, 2.0)):
        ref = Reference(ws, min_col=1, max_col=1 + n_cols, min_row=r, max_row=r)
        chart.add_data(ref, titles_from_data=True, from_rows=True)
        _style(chart.series[-1], colour, width)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def build(source: Path, out: Path, banks: Optional[List[str]] = None) -> Path:
    panels = read_panel(source)
    if "NCLNLSR" not in panels:
        raise SystemExit("no trendable data in %s" % source.name)
    periods = list(reversed(panels["NCLNLSR"].series[
        next(iter(panels["NCLNLSR"].series))].periods))
    all_banks = sorted(panels["NCLNLSR"].series)
    chosen = [b for b in all_banks if not banks or b in banks]
    n_cols = len(periods)

    wb = openpyxl.Workbook()
    about = wb.active
    about.title = "About"
    about["A1"] = "%s -- charts" % source.stem
    about["A1"].font = Font(bold=True, size=15)
    lines = [
        "",
        "What this is",
        "  The same numbers as the monitor, drawn over time. The monitor shows the latest",
        "  quarter against a threshold; this shows the direction, the pace, and how each",
        "  bank compares with its peers.",
        "",
        "How to read the peer sheets",
        "  The big chart draws every bank in grey and the peer median in blue, so the shape",
        "  of the group reads at a glance. Click any grey line to highlight that bank.",
        "  Below the numbers, one small chart per bank: that bank in blue against the",
        "  peer median in grey. That is the 'where do we sit' view.",
        "",
        "Where the numbers come from",
        "  Read from the monitor's own Raw_FDIC block and computed by its own engine, so a",
        "  chart here cannot disagree with a dashboard there. Every measure traces to its",
        "  Call Report schedule, line and MDRM code -- run the monitor with --tieout to see",
        "  the full map, in plain English, with a link to the filed document.",
        "",
        "This workbook has no macros",
        "  So it opens without the security banner that blocks the monitor's own .xlsm when",
        "  that arrives by email. Nothing here needs to be enabled or unblocked.",
        "",
        "What was left blank, and why",
        "  A class rate is left blank for any quarter where that bank's book in the class",
        "  was under $100M. A rate on a near-empty book is arithmetic, not information:",
        "  a $560k recovery on a $2M book is -28% annualised, drawn as faithfully as a",
        "  real number. The sheet says how many quarters went, and how big the largest",
        "  blanked book was, so the reason can be checked rather than taken on trust.",
        "",
        "  A charge-off rate is left blank for any quarter in which that bank absorbed",
        "  another bank. A Call Report reports charge-offs as a running total from 1",
        "  January, and a quarter's figure is that total less the previous quarter's --",
        "  so across a merger the subtraction spans two banks and the quarter is not a",
        "  quarter of anything. Capital One's other-consumer rate read 670% for the",
        "  quarter it absorbed its card bank, which was true arithmetic about nothing.",
        "  The quarters come from the FDIC's own merger record on the monitor's",
        "  _mergers tab, never from the shape of the numbers. Balances and 30-89 / 90+ /",
        "  nonaccrual rates are NOT blanked: they are correct as at the date and simply",
        "  describe a larger bank.",
        "",
        "It is generated, not edited",
        "  Re-running the tool replaces it. Keep your own notes somewhere else.",
        "",
        "Coverage",
        "  %d banks, %d quarters (%s to %s), %d peer-comparison sheets and one"
        % (len(chosen), len(periods), _quarter(periods[0]),
           _quarter(periods[-1]), len(PEER_METRICS)),
        "  stage-comparison sheet per bank.",
    ]
    for i, text in enumerate(lines, start=2):
        about.cell(i, 1, text)
        if text and not text.startswith("  "):
            about.cell(i, 1).font = Font(bold=True, size=11)
    about.column_dimensions["A"].width = 95

    # --- one metric, every bank ------------------------------------------
    for metric in PEER_METRICS:
        panel = panels.get(metric)
        if panel is None:
            continue
        ws = wb.create_sheet(_sheet_name("Peers", metric))
        meaning = (plain.describe(metric) if plain else None) or MEANS.get(metric, "")
        # the code and the words, always both: "i dont know what RCON2200
        # means without looking it up" (the firm, 5 Sep 2026)
        ws["A1"] = "%s -- %s" % (MEANS.get(metric, metric).capitalize(), metric)
        ws["A1"].font = TITLE_FONT
        ws["A2"] = meaning
        ws["A2"].font = NOTE_FONT
        direction = WORSE_WHEN.get(metric)
        ws["A3"] = ("Higher is worse." if direction == "up"
                    else "Lower is worse." if direction == "down"
                    else "Direction is not a verdict for this measure.")
        ws["A3"].font = NOTE_FONT

        rows = []
        for bank in chosen:
            series = panel.series.get(bank)
            if series:
                rows.append((bank, list(reversed(series.values))))
        rows.append((MEDIAN_LABEL, _median_row(rows)))
        grid_top = 20
        _write_grid(ws, grid_top, periods, rows)
        ws.cell(grid_top + len(rows), 1).font = Font(bold=True)
        _add_peer_overview(ws, metric, meaning, grid_top, len(rows) - 1, n_cols, "A5")

        # small multiples, three across, below the grid
        median_row = grid_top + len(rows)
        first = grid_top + len(rows) + 3
        ws.cell(first - 1, 1, "Each bank against the peer median").font = Font(bold=True)
        cols = ("A", "G", "M")
        for k, (bank, _values) in enumerate(rows[:-1]):
            anchor = "%s%d" % (cols[k % 3], first + (k // 3) * 11)
            _add_small_multiple(ws, bank, metric, grid_top, grid_top + 1 + k,
                                median_row, n_cols, anchor)
        ws.freeze_panes = "B21"

    # --- one bank, the four stages ---------------------------------------
    for bank in chosen:
        ws = wb.create_sheet(_sheet_name("Stages", bank))
        ws["A1"] = "%s -- other consumer loans, four stages" % bank
        ws["A1"].font = TITLE_FONT
        ws["A2"] = ("The same loan book at each stage of going wrong. If arrears "
                    "lead write-offs, these move in sequence rather than together.")
        ws["A2"].font = NOTE_FONT
        ws["A3"] = ("Charge-offs are annualised (a quarter's flow x4); the others "
                    "are point-in-time balances. Compare shapes, not heights.")
        ws["A3"].font = NOTE_FONT
        stage_ids = {m for m, _ in STAGES}
        gaps = [b for b in trend.LAST_MERGER_BLANKS if b[0] == bank
                and b[1] in stage_ids]
        small = [b for b in trend.LAST_MATERIALITY_BLANKS if b[0] == bank
                 and b[1] in stage_ids]
        row = 4
        if small:
            quarters = sorted({b[2][:10] for b in small})
            biggest = max((b[3] or 0) for b in small)
            ws["A%d" % row] = (
                "%d quarter%s left blank because the book was under $%dM: %s. "
                "The largest book blanked was $%.1fM. A rate on a book that "
                "size is arithmetic, not information."
                % (len(quarters), "" if len(quarters) == 1 else "s",
                   trend.MATERIALITY_FLOOR_K // 1000, ", ".join(quarters),
                   biggest / 1000.0))
            ws["A%d" % row].font = NOTE_FONT
            row += 1
        if gaps:
            ws["A%d" % row] = ("Charge-offs are left blank for %s: %s"
                               % (", ".join(sorted({g[2][:10] for g in gaps})),
                                  gaps[0][3]))
            ws["A%d" % row].font = NOTE_FONT
        elif trend.LAST_MERGERS_UNKNOWN:
            ws["A%d" % row] = (
                "No merger record was available for this workbook, so no "
                "quarter is marked. A quarter that spans a merger mixes two "
                "banks and its charge-off rate is not comparable -- run the "
                "monitor live to fetch the record.")
            ws["A%d" % row].font = NOTE_FONT
        rows = []
        for metric, label in STAGES:
            panel = panels.get(metric)
            series = panel.series.get(bank) if panel else None
            if series:
                rows.append((label, list(reversed(series.values))))
        if not rows:
            continue
        _write_grid(ws, 20, periods, rows)
        _add_stages_chart(ws, "%s -- delinquency to charge-off" % bank,
                          "% of that loan book", 20, len(rows), n_cols, "A5")
        ws.freeze_panes = "B21"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("-w", "--workbook", required=True)
    ap.add_argument("-o", "--out")
    ap.add_argument("--bank", action="append",
                    help="limit to these banks (repeatable)")
    args = ap.parse_args(argv)
    source = Path(args.workbook)
    out = Path(args.out) if args.out else source.with_name(source.stem + "_Charts.xlsx")
    written = build(source, out, args.bank)
    book = openpyxl.load_workbook(written)
    print("wrote %s" % written)
    print("  %d sheets, %d charts"
          % (len(book.sheetnames),
             sum(len(book[s]._charts) for s in book.sheetnames)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
