#!/usr/bin/env python3
"""Peer trend analysis over the history the workbook already holds.

    python tools/trend.py --workbook Bank_Peer_Monitor.xlsm
    python tools/trend.py --workbook Bank_Peer_Monitor.xlsm --metric NCLNLSR

The dashboards are a snapshot: one column, the latest quarter, each bank against
a threshold. That answers "who is bad now" and cannot answer "who is getting
worse", which is the question a credit review actually asks. The data to answer
it is already in the workbook -- ``Raw_FDIC`` carries 16 quarters per bank -- and
nothing reads it.

WHAT THIS MEASURES, AND WHAT IT DOES NOT

Three things a level cannot tell you, computed per metric per bank:

* **Direction and pace.** The change over the last 4 quarters and the last 8, so
  a bank that doubled from a low base is distinguishable from one that drifted.
* **Persistence.** How many of the last 4 quarters moved the same way. A metric
  that rose 4 quarters running is a different object from one that rose once and
  is noisy, and a level chart shows them identically.
* **Divergence from the peer set.** The bank's move against the median move of
  its peers. A whole sector deteriorating together is a cycle; one bank
  deteriorating while its peers do not is a finding.

It does NOT decide anything. It has no thresholds and lights no flags -- those
live in `_config` and belong to the desk. Every figure here is arithmetic on
filed Call Report data, and every metric traces to its schedule and MDRM code
through ``--tieout``. Where a metric is a documented variant or proxy rather
than a filed line, that is the workbook's own note and it applies here too.

**Direction is not goodness.** Rising equity and rising noncurrent loans both
"go up"; only the metric's own polarity says which is deterioration. That is
declared per metric below rather than guessed from the name.
"""

from __future__ import annotations

import argparse
import statistics
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl  # noqa: E402

#: Metrics worth trending, and which direction is deterioration.
#: "up" means a rising number is worse. Anything not listed is still computed
#: when asked for by name, but is reported without a deterioration verdict --
#: guessing polarity from a field name is how a capital ratio gets read as a
#: delinquency rate.
WORSE_WHEN = {
    "NCLNLSR": "up",     "NTLNLSQR": "up",   "PD3089R": "up",
    "LNRESNCR": "down",  "LNATRESR": None,   "TEXAS": "up",
    "EQV": "down",       "RBC1AAJ": "down",  "RBCRWAJ": "down",
    "ROAQ": "down",      "NIMY": "down",     "EEFFR": "up",
    "LNDEPR": "up",      "BRODEPR": "up",    "CRECONR": "up",
    "UNINSDEPR": "up",   "UNRLZCAPR": "up",  "FHLBASSR": "up",
    "P3CRCDR": "up",     "P9CRCDR": "up",    "NACRCDR": "up",  "NTCRCDQ_BOOK": "up",
    "P3AUTOR": "up",     "P9AUTOR": "up",    "NAAUTOR": "up",  "NTAUTOQ_BOOK": "up",
    "P3CONOTH_BOOK": "up",   "P9CONOTH_BOOK": "up",  "NACONOTH_BOOK": "up", "NTCONOTQ_BOOK": "up",
    "P3RERESR": "up",    "P9RERESR": "up",   "NARERESR": "up",
    "P3RENRES_BOOK": "up",   "P9RENRES_BOOK": "up",  "NARENRES_BOOK": "up",
    "P3CIR": "up",       "P9CIR": "up",      "NACIR": "up",    "NTCIQ_BOOK": "up",
}

#: What each headline metric is, in words, so a reader who does not live in
#: MDRM codes can still act on the output.
MEANS = {
    "NCLNLSR":  "noncurrent loans as % of loans (90+ days past due, plus nonaccrual)",
    "NTLNLSQR": "net charge-offs as % of loans, annualised",
    "PD3089R":  "loans 30-89 days past due as % of loans -- the early-warning bucket",
    "LNRESNCR": "allowance as % of noncurrent loans -- reserve coverage",
    "TEXAS":    "Texas ratio variant: noncurrent loans / (equity + allowance)",
    "EQV":      "equity as % of assets",
    "RBC1AAJ":  "tier 1 leverage ratio",
    "ROAQ":     "return on assets, annualised",
    "CRECONR":  "commercial real estate concentration as % of capital (proxy)",
    "BRODEPR":  "brokered deposits as % of deposits",
    "UNINSDEPR": "estimated uninsured deposits as % of deposits",
}


@dataclass
class Series:
    bank: str
    cert: str
    periods: List[str]
    values: List[Optional[float]]

    def clean(self) -> List[tuple]:
        return [(p, v) for p, v in zip(self.periods, self.values) if v is not None]

    @property
    def latest(self) -> Optional[float]:
        c = self.clean()
        return c[0][1] if c else None

    def change(self, quarters: int) -> Optional[float]:
        c = self.clean()
        if len(c) <= quarters:
            return None
        return c[0][1] - c[quarters][1]

    def run_length(self) -> int:
        """Consecutive quarters (most recent first) moving the same direction.

        Signed: +3 means three straight rises. Returns 0 when the latest move is
        flat or there is too little data.
        """
        c = [v for _, v in self.clean()]
        if len(c) < 3:
            return 0
        diffs = [c[i] - c[i + 1] for i in range(len(c) - 1)]
        if not diffs or diffs[0] == 0:
            return 0
        sign = 1 if diffs[0] > 0 else -1
        run = 0
        for d in diffs:
            if (d > 0 and sign > 0) or (d < 0 and sign < 0):
                run += 1
            else:
                break
        return run * sign


@dataclass
class Panel:
    metric: str
    series: Dict[str, Series] = field(default_factory=dict)

    def peer_median_change(self, quarters: int) -> Optional[float]:
        moves = [s.change(quarters) for s in self.series.values()]
        moves = [m for m in moves if m is not None]
        return statistics.median(moves) if moves else None


def read_panel(path: Path, raw_tab: str = "Raw_FDIC",
               derive: bool = True) -> Dict[str, Panel]:
    """Every metric, every bank, every quarter, out of the raw block.

    Twenty-eight of the fifty-three dashboard metrics are not landed as raw
    fields -- they are computed by formula in the workbook from raw components
    (`PD3089R` is `P3LNLS / LNLSGR`, and so on). Reading only the raw block
    therefore silently covers less than half the metrics a reader can see on a
    dashboard, and a tool that quietly trends 25 of 53 while looking like it
    trends all of them is worse than one that refuses.

    So the derived ones are computed here through the ENGINE'S OWN
    ``metric_value`` -- the same function the workbook's own runner and
    ``--tieout`` use. Re-implementing the formulas in this tool would create a
    second definition of every ratio, and the first quarter they disagreed the
    trend would contradict the dashboard with nothing to say which was right.
    """
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    ws = wb[raw_tab]
    panels: Dict[str, Panel] = {}
    bank = cert = None
    fields: List[str] = []
    periods: List[str] = []
    rows: List[List] = []

    def flush():
        if not bank or not fields or not rows:
            return
        for col, name in enumerate(fields[1:], start=1):
            if not name:
                continue
            panel = panels.setdefault(name, Panel(metric=name))
            vals = []
            for row in rows:
                v = row[col] if col < len(row) else None
                vals.append(float(v) if isinstance(v, (int, float)) else None)
            panel.series[bank] = Series(bank=bank, cert=cert or "",
                                        periods=list(periods), values=vals)

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        text = str(a) if a is not None else ""
        if text.startswith("slot"):
            flush()
            bank = str(ws.cell(r, 3).value or "").strip() or None
            tag = str(ws.cell(r, 2).value or "")
            cert = tag.split("cert:")[-1].strip() if "cert:" in tag else ""
            fields, periods, rows = [], [], []
        elif text == "REPDTE":
            fields = [str(ws.cell(r, c).value or "").strip()
                      for c in range(1, ws.max_column + 1)]
        elif text[:2] in ("19", "20") and bank:
            periods.append(text[:10])
            rows.append([ws.cell(r, c).value for c in range(1, len(fields) + 1)])
    flush()
    global LAST_MERGERS_UNKNOWN
    mergers = read_mergers(wb)
    LAST_MERGERS_UNKNOWN = mergers is None
    if derive:
        _add_derived(panels, _metric_ids(wb))
        apply_merger_flags(panels, mergers)
    return panels


#: The quarters a merger makes uncomparable, and the metrics it touches.
#:
#: 4 September 2026: a chart drew Capital One's other-consumer charge-off rate
#: at 670% for the quarter ending 31 December 2022. The firm did not believe
#: it and asked for the cause rather than a threshold. The cause was a merger
#: -- Capital One Bank (USA), N.A. into Capital One, N.A. on 3 October 2022 --
#: and the FDIC derives a quarterly flow by subtracting the previous quarter's
#: year-to-date total, which across a merger spans two banks.
#:
#: A size floor was built first and then removed, because it would have hidden
#: this by luck: on a $500M book the same $5.3M draws a plausible 4.3% that
#: nobody questions. The guard is the merger record, from the FDIC's own
#: history, read out of the workbook's `_mergers` tab. This tool never infers
#: a merger from the shape of the numbers.
MERGERS_TAB = "_mergers"


def read_mergers(wb) -> Optional[Dict[str, List[dict]]]:
    """`{cert: [{quarter, effective, out_name, why}, ...]}` from the workbook.

    Returns None when the workbook cannot answer -- no `_mergers` tab (an
    older build), or a tab whose block says the run never fetched one. None is
    NOT the same as `{}`: one means "nobody asked", the other "asked, none
    found", and a caller that collapses them is back to the 670.
    """
    if MERGERS_TAB not in wb.sheetnames:
        return None
    ws = wb[MERGERS_TAB]
    header = None
    out: Dict[str, List[dict]] = {}
    asked = False
    for row in ws.iter_rows(values_only=True):
        first = "" if not row or row[0] is None else str(row[0]).strip()
        if header is None:
            if first == "bank":
                header = [str(c).strip() if c else "" for c in row]
            continue
        if not first:
            continue
        if first.startswith("(this run did not fetch"):
            return None
        if first.startswith("("):
            asked = True                       # "asked, and none found"
            continue
        cells = list(row) + [None] * len(header)
        cert = str(cells[1] or "").strip()
        if not cert:
            continue
        asked = True
        out.setdefault(cert, []).append({
            "bank": first,
            "effective": str(cells[2] or "")[:10],
            "quarter": str(cells[3] or "")[:10],
            "out_name": str(cells[4] or ""),
            "why": str(cells[8] or ""),
        })
    return out if (out or asked) else None


#: What the last `apply_merger_flags` blanked -- (bank, metric, period, why) --
#: so a sheet or a report can say why the gap is there. Module-level on
#: purpose: `read_panel` returns a plain dict of panels, and hiding a
#: pseudo-panel inside it would miscount every "N metrics" line downstream.
LAST_MERGER_BLANKS: List[tuple] = []

#: True when the last read could not establish a merger record at all.
LAST_MERGERS_UNKNOWN = False


def flow_metrics() -> frozenset:
    """The metrics built from a quarterly flow -- the engine's own list."""
    from credit_suite.sources.fdic import engine_api as R
    return R.QUARTERLY_FLOW_METRICS


def apply_merger_flags(panels: Dict[str, Panel],
                       mergers: Optional[Dict[str, List[dict]]]
                       ) -> List[tuple]:
    """Blank every quarterly-flow rate for a quarter that spans a merger.

    Balances and point-in-time rates are left alone on purpose: they are
    correct as at the date and simply describe a larger bank. It is the FLOW
    that mixes two banks.
    """
    blanked: List[tuple] = []
    if not mergers:
        LAST_MERGER_BLANKS[:] = blanked
        return blanked
    by_cert = {str(c): v for c, v in mergers.items()}
    flows = flow_metrics()
    for metric, panel in panels.items():
        if metric not in flows:
            continue
        for bank, series in panel.series.items():
            events = by_cert.get(str(series.cert))
            if not events:
                continue
            for event in events:
                quarter = event.get("quarter", "")
                for i, period in enumerate(series.periods):
                    if period[:10] == quarter and series.values[i] is not None:
                        series.values[i] = None
                        blanked.append((bank, metric, period,
                                        event.get("why") or
                                        ("quarter spans a merger effective %s"
                                         % event.get("effective", "?"))))
    LAST_MERGER_BLANKS[:] = blanked
    return blanked


def _metric_ids(wb) -> List[str]:
    """The metric ids the workbook itself declares, from its `_config` tab.

    Read from the workbook rather than hardcoded here: `_config` is the source
    of truth for what this monitor computes (contract section 3), and a second
    list in a tool is a list that goes stale the first time somebody adds a
    metric."""
    try:
        from credit_suite.engine.config import parse_config
        from credit_suite.sources.fdic.spec import FDIC
        rows = [list(r) for r in wb["_config"].iter_rows(values_only=True)]
        return [s.id for s in parse_config(rows, FDIC).series]
    except Exception:                                  # noqa: BLE001
        return []


def _add_derived(panels: Dict[str, Panel], metric_ids: List[str]) -> None:
    """Fill in the formula-driven metrics, using the engine's definitions."""
    try:
        from credit_suite.sources.fdic.engine_api import metric_value
    except ImportError:                                # pragma: no cover
        return

    if not panels:
        return
    any_panel = next(iter(panels.values()))
    banks = list(any_panel.series)
    # Which metrics are ALREADY landed raw, decided once. Deciding it inside
    # the quarter loop meant every metric looked "already present" from the
    # second quarter onwards, so each derived series held exactly one value
    # and every trend over it was silently empty.
    already_raw = {m for m in metric_ids if m in panels}

    # Rebuild a per-bank, per-quarter field dict from what was read.
    for bank in banks:
        sample = any_panel.series[bank]
        periods = sample.periods
        for index, period in enumerate(periods):
            values = {}
            for name, panel in panels.items():
                series = panel.series.get(bank)
                if series and index < len(series.values):
                    values[name] = series.values[index]
            for metric in metric_ids:
                if metric in already_raw:
                    continue
                try:
                    computed = metric_value(metric, values)
                except Exception:                      # noqa: BLE001
                    computed = None
                panel = panels.setdefault(metric, Panel(metric=metric))
                target = panel.series.get(bank)
                if target is None:
                    target = Series(bank=bank, cert=sample.cert,
                                    periods=list(periods),
                                    values=[None] * len(periods))
                    panel.series[bank] = target
                target.values[index] = computed


def spark(values: Sequence[Optional[float]]) -> str:
    """Oldest-to-newest bar sparkline. ASCII-safe fallback is not attempted --
    the caller reconfigures stdout, because a chart that crashes the report is
    worse than no chart (that was a real defect in check_parity)."""
    vals = [v for v in reversed(values) if v is not None]
    if len(vals) < 2:
        return ""
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return "-" * len(vals)
    blocks = "▁▂▃▄▅▆▇█"
    return "".join(blocks[min(7, int((v - lo) / (hi - lo) * 7.999))] for v in vals)


def deteriorating(panel: Panel, quarters: int = 4) -> List[tuple]:
    """Banks whose move over `quarters` is in the metric's bad direction,
    worst first, with their divergence from the peer median move."""
    polarity = WORSE_WHEN.get(panel.metric)
    if polarity is None:
        return []
    median = panel.peer_median_change(quarters)
    out = []
    for s in panel.series.values():
        change = s.change(quarters)
        if change is None or s.latest is None:
            continue
        bad = change if polarity == "up" else -change
        if bad <= 0:
            continue
        divergence = None
        if median is not None:
            divergence = (change - median) if polarity == "up" else (median - change)
        out.append((bad, s, change, divergence))
    return sorted(out, key=lambda t: -t[0])


def report(path: Path, only: Optional[str] = None, quarters: int = 4,
           top: int = 5) -> int:
    panels = read_panel(path)
    banks = max((len(p.series) for p in panels.values()), default=0)
    depth = max((len(s.periods) for p in panels.values()
                 for s in p.series.values()), default=0)
    print("PEER TREND  %s" % path.name)
    print("  %d metrics x %d banks x %d quarters read from Raw_FDIC"
          % (len(panels), banks, depth))
    span = None
    for p in panels.values():
        for s in p.series.values():
            if s.periods:
                span = (s.periods[-1], s.periods[0]); break
        if span: break
    if span:
        print("  window %s .. %s   (change measured over %d quarters)"
              % (span[0], span[1], quarters))
    print()

    names = [only] if only else [m for m in MEANS if m in panels]
    for metric in names:
        panel = panels.get(metric)
        if panel is None:
            print("  %s: not in this workbook" % metric)
            continue
        rows = deteriorating(panel, quarters)
        median = panel.peer_median_change(quarters)
        print("== %s -- %s" % (metric, MEANS.get(metric, "")))
        print("   peer median move over %dq: %s   "
              "(\"worse by\" = this bank's move minus the peer median, "
              "signed so bigger is always worse)"
              % (quarters, "n/a" if median is None else "%+.2f" % median))
        if not rows:
            print("   no bank deteriorated on this metric over the window.\n")
            continue
        print("   %-26s %8s %9s %9s %6s  %s"
              % ("bank", "latest", "%dq chg" % quarters, "worse by", "run", "16q trend"))
        for _, s, change, divergence in rows[:top]:
            run = s.run_length()
            runtxt = ("%+d" % run) if run else "."
            print("   %-26s %8.2f %+9.2f %9s %6s  %s"
                  % (s.bank[:26], s.latest, change,
                     "n/a" if divergence is None else "%+.2f" % divergence,
                     runtxt, spark(s.values)))
        print()
    return 0


def main(argv: Optional[List[str]] = None) -> int:
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("-w", "--workbook", required=True)
    ap.add_argument("--metric", help="one metric instead of the headline set")
    ap.add_argument("--quarters", type=int, default=4)
    ap.add_argument("--top", type=int, default=5)
    args = ap.parse_args(argv)
    return report(Path(args.workbook), args.metric, args.quarters, args.top)


if __name__ == "__main__":
    raise SystemExit(main())
