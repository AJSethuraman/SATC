"""Verify EVERY observation of every FRED series, not just the latest one.

The first pass proved 142 figures -- the newest observation of each series. The
firm's instruction is all of the raw data, verified. Most of the agency
documents already downloaded carry the whole history, so most of the 13,841
observations can be checked against the body that computed them rather than
against nothing.

Where a source only publishes the current period -- S&P's Case-Shiller release
is the case -- that is stated per series as unverified history, never quietly
counted as done.

Nothing here derives, adjusts or recomputes a value. It compares the number in
the workbook with the number in the agency's own file, and records which.
"""
import csv
import io
import json
import pathlib
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

SB = pathlib.Path(r"C:\Users\ajish\AppData\Local\Temp\claude"
                  r"\C--Users-ajish-SATC\261f7248-3cbc-4aa2-aacf-e4ff9181778a\scratchpad")
C = SB / "sources"
sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

ours = json.loads((SB / "fred_ours.json").read_text())
series_meta = {r["series_id"]: r for r in json.loads((SB / "fred_series.json").read_text())}


def q_of(iso):
    y, m = int(iso[:4]), int(iso[5:7])
    return str(y), str((m - 1) // 3 + 1)


# ---------------------------------------------------------------- FHFA ------
def fhfa_tables():
    state, metro, us = {}, {}, {}
    for line in csv.reader(io.StringIO((C / "fhfa_state.csv").read_text(errors="replace"))):
        if len(line) >= 4:
            state.setdefault(line[0].strip(), {})[(line[1], line[2])] = line[3]
    for line in csv.reader(io.StringIO((C / "fhfa_metro.csv").read_text(errors="replace"))):
        if len(line) >= 5 and line[1].strip().isdigit():
            metro.setdefault(line[1].strip(), {})[(line[2], line[3])] = line[4]
    for line in csv.reader(io.StringIO((C / "fhfa_us.csv").read_text(errors="replace"))):
        if len(line) >= 4:
            us.setdefault(line[0].strip(), {})[(line[1], line[2])] = line[3]
    return state, metro, us


FHFA_STATE, FHFA_METRO, FHFA_US = fhfa_tables()


def fhfa_lookup(sid, iso):
    cat = series_meta.get(sid, {}).get("category", "")
    y, q = q_of(iso)
    if cat == "hpi_state":
        row = FHFA_STATE.get(sid[:2])
    elif cat == "hpi_metro":
        code = re.search(r"(\d{5})", sid)
        row = FHFA_METRO.get(code.group(1)) if code else None
    else:
        row = FHFA_US.get({"USSTHPI": "USA"}.get(sid))
    if not row:
        return None
    v = row.get((y, q))
    return None if v in (None, "", "-") else float(v)


# --------------------------------------------------- FHFA monthly (PO) ------
def fhfa_monthly():
    """USA, seasonally adjusted, from FHFA's own monthly purchase-only file."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO((C / "fhfa_po_monthly.csv").read_bytes()),
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [str(c.value).replace("\n", " ").strip() if c.value else "" for c in ws[4]]
    col = next(i for i, h in enumerate(hdr) if h.startswith("USA") and "(SA)" in h)
    out = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        stamp = r[0]
        if stamp is None or r[col] is None:
            continue
        y, m = ((stamp.year, stamp.month) if hasattr(stamp, "year")
                else (int(str(stamp)[:4]), int(str(stamp)[5:7])))
        out["%04d-%02d" % (y, m)] = float(r[col])
    wb.close()
    return out


FHFA_MONTHLY = fhfa_monthly()


def fhfa_monthly_lookup(sid, iso):
    return FHFA_MONTHLY.get(iso[:7])


# ------------------------------------------------------------- Fed tables ---
FED_PAGES = ("chgallsa", "delallsa", "chgtop100sa", "deltop100sa",
             "chgothersa", "delothersa")
ROW = re.compile(r"(\d{4}):(\d)((?:\s*-?\d+\.\d+&nbsp;)+)")


def fed_tables():
    out = {}
    for name in FED_PAGES:
        html = (C / ("fed_%s.html" % name)).read_text(encoding="utf-8", errors="replace")
        flat = re.sub(r"<[^>]+>", " ", html)
        flat = re.sub(r"[ \t]+", " ", flat)
        rows = {}
        for m in ROW.finditer(flat):
            yy, qq, body = m.groups()
            rows[(yy, qq)] = [float(v) for v in re.findall(r"-?\d+\.\d+", body)]
        out[name] = rows
    return out


FED = fed_tables()
FED_LAYOUT = json.loads((SB / "fred_fed_layout.json").read_text()) \
    if (SB / "fred_fed_layout.json").exists() else None
if FED_LAYOUT is None:
    import importlib.util
    spec = importlib.util.spec_from_file_location("ff", SB / "fred_fed.py")
    # fred_fed.py executes on import; read its LAYOUT literal instead
    text = (SB / "fred_fed.py").read_text(encoding="utf-8")
    body = text[text.index("LAYOUT = {"):]
    body = body[:body.index("\n}") + 2]
    FED_LAYOUT = eval(body.split("=", 1)[1].strip())          # noqa: S307
    (SB / "fred_fed_layout.json").write_text(json.dumps(FED_LAYOUT), encoding="utf-8")


def fed_lookup(sid, iso):
    spec = FED_LAYOUT.get(sid)
    if not spec:
        return None
    table, col = spec
    row = FED[table].get(q_of(iso))
    if not row or col >= len(row):
        return None
    return row[col]


# ------------------------------------------------------------------ G.19 ----
MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def g19_hist():
    html = (C / "g19_hist_sa.csv").read_text(encoding="utf-8", errors="replace")
    out = {}
    for mon, year, body in re.findall(
            r'<th[^>]*>\s*([A-Z][a-z]{2})\s+(\d{4})\s*</th>'
            r'((?:\s*<td[^>]*>[^<]*</td>)+)', html):
        vals = []
        for cell in re.findall(r"<td[^>]*>([^<]*)</td>", body):
            cell = cell.strip().replace(",", "")
            vals.append(float(cell) if re.fullmatch(r"-?\d+(\.\d+)?", cell) else None)
        out[(year, mon)] = vals
    return out


G19 = g19_hist()
G19_COL = {"TOTALSL": 0, "REVOLSL": 1, "NONREVSL": 2}


def g19_lookup(sid, iso):
    col = G19_COL.get(sid)
    if col is None:
        return None
    row = G19.get((iso[:4], MONTHS[int(iso[5:7]) - 1]))
    return None if not row or col >= len(row) else row[col]


# ------------------------------------------------------------------- DSR ----
def dsr_table():
    html = (C / "dsr_new.htm").read_text(encoding="utf-8", errors="replace")
    flat = re.sub(r"<[^>]+>", " ", html)
    flat = re.sub(r"[ \t]+", " ", flat)
    return {(y, q): [float(v) for v in body.split()]
            for y, q, body in re.findall(r"(\d{4}):(\d)((?:\s+-?\d+\.\d+){3})", flat)}


DSR = dsr_table()
DSR_COL = {"TDSP": 0, "MDSP": 1, "CDSP": 2}


def dsr_lookup(sid, iso):
    col = DSR_COL.get(sid)
    row = DSR.get(q_of(iso))
    return None if col is None or not row else row[col]


# ----------------------------------------------------------------- SLOOS ----
def sloos_tables():
    html = (C / "sloos_chartdata.htm").read_text(encoding="utf-8", errors="replace")
    def txt(x):
        return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", x)).replace("&nbsp;", " ").strip()
    out = []
    for tb in re.findall(r"<table.*?</table>", html, re.S):
        rows = {}
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", tb, re.S):
            cells = [txt(c) for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.S)]
            if cells and re.fullmatch(r"\d{4}:\d", cells[0]):
                rows[tuple(cells[0].split(":"))] = cells
        out.append(rows)
    return out


SLOOS = sloos_tables()
SLOOS_LAYOUT = {"DRTSCILM": (0, 1), "DRTSCIS": (0, 2), "DRSDCILM": (0, 5),
                "DRSDCIS": (0, 6), "SUBLPDRCSC": (1, 2), "SUBLPDRCSN": (1, 3),
                "SUBLPDRCSM": (1, 4), "DRTSSP": (2, 4), "SUBLPDHMSENQ": (2, 5),
                "DRTSCLCC": (3, 1), "STDSAUTO": (3, 3), "STDSOTHCONS": (3, 4)}


def sloos_lookup(sid, iso):
    spec = SLOOS_LAYOUT.get(sid)
    if not spec:
        return None
    ti, ci = spec
    row = SLOOS[ti].get(q_of(iso))
    if not row or ci >= len(row) or not re.fullmatch(r"-?\d+(\.\d+)?", row[ci]):
        return None
    return float(row[ci])


# ------------------------------------------------------------------- Z.1 ----
def z1_series():
    z = zipfile.ZipFile(C / "z1_pkg_zip.bin")
    want = {"FL075035403.Q": "BOGZ1FL075035403Q", "FL075035503.Q": "BOGZ1FL075035503Q"}
    found = {v: {} for v in want.values()}
    cur = None
    with z.open("Z1_data.xml") as fh:
        for event, elem in ET.iterparse(fh, events=("start", "end")):
            tag = elem.tag.rsplit("}", 1)[-1]
            if event == "start" and tag == "Series":
                name = elem.get("SERIES_NAME")
                cur = want.get(name)
            elif event == "end":
                if tag == "Obs" and cur:
                    p, v = elem.get("TIME_PERIOD"), elem.get("OBS_VALUE")
                    if p and v not in (None, "", "ND"):
                        found[cur][p] = float(v)
                elif tag == "Series":
                    cur = None
                    elem.clear()
    return found


Z1 = z1_series()
QEND = ((3, 31), (6, 30), (9, 30), (12, 31))


def z1_lookup(sid, iso):
    obs = Z1.get(sid)
    if not obs:
        return None
    y, m = int(iso[:4]), int(iso[5:7])
    em, ed = QEND[(m - 1) // 3]
    return obs.get("%04d-%02d-%02d" % (y, em, ed))


LOOKUPS = [
    ("FHFA monthly purchase-only house price index", fhfa_monthly_lookup,
     lambda s: s == "HPIPONM226S"),
    ("FHFA All-Transactions house price index", fhfa_lookup,
     lambda s: series_meta.get(s, {}).get("category", "").startswith("hpi_")
     and series_meta.get(s, {}).get("category") != "hpi_caseshiller"
     and s not in ("CSUSHPINSA", "CSUSHPISA", "HPIPONM226S")),
    ("Federal Reserve Board charge-off / delinquency table", fed_lookup,
     lambda s: s in FED_LAYOUT),
    ("Federal Reserve Board G.19 historical table", g19_lookup,
     lambda s: s in G19_COL),
    ("Federal Reserve Board Household Debt Service Ratio release", dsr_lookup,
     lambda s: s in DSR_COL),
    ("Federal Reserve Board Senior Loan Officer Opinion Survey chart data",
     sloos_lookup, lambda s: s in SLOOS_LAYOUT),
    ("Federal Reserve Board Z.1 complete data package", z1_lookup,
     lambda s: s in Z1),
]

rows = []
per_series = {}
for sid, block in sorted(ours.items()):
    src_name, fn = None, None
    for name, f, applies in LOOKUPS:
        if applies(sid):
            src_name, fn = name, f
            break
    n = tied = differ = nosrc = 0
    for obs in block["observations"]:
        iso, val = obs["date"], obs["value"]
        if val is None:
            continue
        n += 1
        theirs = fn(sid, iso) if fn else None
        if theirs is None:
            nosrc += 1
            verdict = "NO SOURCE FOR THIS PERIOD"
        else:
            ok = abs(float(val) - float(theirs)) < 0.005
            verdict = "TIED" if ok else "DIFFERS"
            tied += ok
            differ += (not ok)
        rows.append({"series": sid, "date": iso, "ours": val, "theirs": theirs,
                     "verdict": verdict, "source": src_name or "no full-history source",
                     "tab": block["tab"], "row": obs["row"]})
    per_series[sid] = {"n": n, "tied": tied, "differs": differ, "no_source": nosrc,
                       "source": src_name or "no full-history source"}

(SB / "fred_history_rows.json").write_text(json.dumps(rows), encoding="utf-8")
(SB / "fred_history_summary.json").write_text(json.dumps(per_series, indent=1), encoding="utf-8")

tot = sum(v["n"] for v in per_series.values())
tied = sum(v["tied"] for v in per_series.values())
differ = sum(v["differs"] for v in per_series.values())
nosrc = sum(v["no_source"] for v in per_series.values())
print("observations in the workbook : %d" % tot)
print("  tied to the agency's file  : %d" % tied)
print("  DIFFER                     : %d" % differ)
print("  no source for that period  : %d" % nosrc)
print()
from collections import Counter
bysrc = Counter()
for sid, v in per_series.items():
    bysrc[v["source"]] += v["n"]
for k, n in bysrc.most_common():
    t = sum(v["tied"] for v in per_series.values() if v["source"] == k)
    print("  %-58s %6d obs, %6d tied" % (k[:58], n, t))
print("\nseries with any DIFFERS:")
for sid, v in sorted(per_series.items()):
    if v["differs"]:
        print("   %-16s %d of %d differ" % (sid, v["differs"], v["n"]))
