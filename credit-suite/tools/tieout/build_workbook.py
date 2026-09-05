"""One workbook: the data, and the tabs that say what it is and is not.

The firm asked for the limits noted "on a tab that explains this and other
similar factors", for the unverifiable parts highlighted so they can see what is
being referred to, and for a tab describing ratios without building any.

Nothing on the data tabs is calculated here.
"""
import csv
import json
import pathlib
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CS = pathlib.Path(r"C:\Users\ajish\SATC-cs\credit-suite")
SB = pathlib.Path(r"C:\Users\ajish\AppData\Local\Temp\claude"
                  r"\C--Users-ajish-SATC\261f7248-3cbc-4aa2-aacf-e4ff9181778a\scratchpad")
D = CS / "verified-data"
sys.path.insert(0, str(SB))
sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

from workbook_tabs import START_HERE, LIMITS, SOURCES, proven_tab   # noqa: E402

INK = "FF151515"
RED = "FFB00020"
GREEN = "FF0A6B3D"
HEAD = PatternFill("solid", fgColor="FF151515")
WARN = PatternFill("solid", fgColor="FFFFF3F3")
CAUTION = PatternFill("solid", fgColor="FFFFFBE6")
OKFILL = PatternFill("solid", fgColor="FFF2FBF6")
THIN = Side(style="thin", color="FFD8D8D8")
BOX = Border(bottom=THIN)

audit = {}
p = SB / "evidence_audit.json"
if p.exists():
    audit = json.loads(p.read_text())

wb = openpyxl.Workbook()
wb.remove(wb.active)


def prose(name, blocks, width=118):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = width
    ws.sheet_view.showGridLines = False
    for r, (kind, text) in enumerate(blocks, start=1):
        c = ws.cell(r, 1, text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "h1":
            c.font = Font(bold=True, size=17, color=INK)
            ws.row_dimensions[r].height = 28
        elif kind == "h2":
            c.font = Font(bold=True, size=13, color=INK)
            ws.row_dimensions[r].height = 24
        elif kind in ("warn", "caution", "ok"):
            c.font = Font(bold=(kind == "warn"), size=11,
                          color=RED if kind == "warn" else
                          (GREEN if kind == "ok" else INK))
            c.fill = WARN if kind == "warn" else (OKFILL if kind == "ok" else CAUTION)
            ws.row_dimensions[r].height = max(30, 14 * (len(text) // 100 + 1))
        elif kind == "":
            ws.row_dimensions[r].height = 8
        else:
            c.font = Font(size=11, color=INK)
            ws.row_dimensions[r].height = max(15, 14 * (len(text) // 104 + 1))
    return ws


def table(name, path, shade=None):
    ws = wb.create_sheet(name)
    with open(path, encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    header = rows[0]
    for i, row in enumerate(rows, start=1):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j)
            if i == 1:
                cell.value = v
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
                cell.fill = HEAD
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                num = None
                if v not in ("", None):
                    try:
                        num = float(v)
                    except ValueError:
                        num = None
                cell.value = num if num is not None else v
                cell.border = BOX
                cell.alignment = Alignment(vertical="top",
                                           wrap_text=len(str(v)) > 60)
        if i > 1 and shade:
            fill = shade(header, row)
            if fill:
                for j in range(1, len(header) + 1):
                    ws.cell(i, j).fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(header)), len(rows))
    WIDE = {"bank", "title", "verified_meaning", "note", "why_not_verified",
            "plain_english", "the_trap", "why_it_makes_sense",
            "what_this_means", "question_it_answers", "ratio"}
    for j, n in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(j)].width = (
            48 if n in WIDE else (36 if n.endswith("url") else
                                  min(max(len(n) + 3, 12), 26)))
    ws.row_dimensions[1].height = 32
    return len(rows) - 1


def shade_unverified(header, row):
    idx = {n: k for k, n in enumerate(header)}
    if idx.get("verified") is None or row[idx["verified"]] != "no":
        return None
    why = ""
    for k in ("why_not_verified", "verified_meaning"):
        if k in idx and row[idx[k]]:
            why = row[idx[k]]
            break
    return CAUTION if "FDIC calculates" in why else WARN


prose("START HERE", START_HERE)
prose("WHAT WAS PROVEN", proven_tab(audit))
prose("LIMITS", LIMITS)
prose("THE SOURCES", SOURCES)
n1 = table("BANK DATA", D / "bank-values.csv", shade_unverified)
n2 = table("MACRO DATA", D / "macro-observations.csv", shade_unverified)
n3 = table("FIELD DICTIONARY", D / "field-dictionary.csv")
n4 = table("RATIOS", D / "ratios-worth-building.csv")
n5 = table("NOT COMPARABLE", D / "not-comparable-periods.csv")

out = D / "SATC-verified-credit-data.xlsx"
wb.save(out)
print("tabs   : %s" % ", ".join(wb.sheetnames))
print("rows   : bank %d, macro %d, fields %d, ratios %d, not-comparable %d"
      % (n1, n2, n3, n4, n5))
print("written: %s  (%.1f MB)" % (out.name, out.stat().st_size / 1e6))
