"""Three more publishers behind the FRED workbook: G.19, the DSR release, FHFA monthly.

Nine series. FRED redistributes all nine; the Board computes G.19 and the debt
service ratios, and FHFA computes the monthly purchase-only house price index.
So each comparison is workbook cell against the computing agency's own document.

Where an agency publishes fewer decimals than FRED carries, that is said out
loud per line and the tolerance is the agency's own precision -- not a bar
lowered to make a number agree.
"""
import io
import json
import pathlib
import re
import sys
import zipfile

import openpyxl

SB = pathlib.Path(r"C:\Users\ajish\AppData\Local\Temp\claude"
                  r"\C--Users-ajish-SATC\261f7248-3cbc-4aa2-aacf-e4ff9181778a\scratchpad")
C = SB / "sources"
sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

ours = json.loads((SB / "fred_ours.json").read_text())
results = []

MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def add(sid, source, where, theirs, tol, note=None, units=None):
    """One comparison, recorded whatever it says."""
    block = ours.get(sid)
    entry = {"series": sid, "source": source, "source_where": where,
             "tolerance": tol, "note": note, "units": units}
    if block is None:
        entry.update(verdict="COULD NOT", why="not in the workbook")
        results.append(entry)
        return entry
    latest = block.get("latest")
    entry.update(title=block["title"], tab=block["tab"])
    if latest is None:
        entry.update(verdict="COULD NOT", why="the workbook landed no observations")
        results.append(entry)
        return entry
    entry.update(cell="B%d" % latest["row"], ours=latest["value"], date=latest["date"])
    if theirs is None:
        entry.update(verdict="COULD NOT",
                     why="the agency's document has no value at that period")
    else:
        entry["theirs"] = theirs
        entry["diff"] = round(entry["ours"] - theirs, 8)
        entry["verdict"] = "TIED" if abs(entry["diff"]) <= tol else "DIFFERS"
    results.append(entry)
    return entry


# ---------------------------------------------------------------- G.19 ------
# Two documents from the same release. The historical table carries the levels
# to the cent; the current release carries every one of the five but rounds to
# a tenth of a billion. Read both -- two renderings of one source is not the
# mirror, and where they disagree that is a finding, not a tie.

def g19_hist(path):
    """{('Jun', '2026'): [total, revolving, nonrevolving]} in millions."""
    html = (C / path).read_text(encoding="utf-8", errors="replace")
    out = {}
    for mon, year, body in re.findall(
            r'<th[^>]*>\s*([A-Z][a-z]{2})\s+(\d{4})\s*</th>'
            r'((?:\s*<td[^>]*>[^<]*</td>)+)', html):
        vals = []
        for cell in re.findall(r"<td[^>]*>([^<]*)</td>", body):
            cell = cell.strip().replace(",", "")
            vals.append(float(cell) if re.fullmatch(r"-?\d+(\.\d+)?", cell) else None)
        out[(mon, year)] = vals
    return out


def g19_release_row(label):
    """The last (most recent month) cell of a labelled row in the current release."""
    html = (C / "g19_current_default.htm").read_text(encoding="utf-8", errors="replace")
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S):
        text = re.sub(r"<[^>]+>", " ", row)
        text = re.sub(r"\s+", " ", text).strip()
        if not text.startswith(label):
            continue
        nums = [c.strip().replace(",", "")
                for c in re.findall(r"<td[^>]*>(.*?)</td>", row, re.S)]
        nums = [re.sub(r"<[^>]+>|&nbsp;", "", n).strip() for n in nums]
        nums = [float(n) for n in nums if re.fullmatch(r"-?\d+(\.\d+)?", n)]
        if nums:
            return nums[-1], text[:70]
    return None, None


sa = g19_hist("g19_hist_sa.csv")
latest_month = max(sa, key=lambda k: (int(k[1]), MONTHS.index(k[0])))
tot, rev, nonrev = sa[latest_month][:3]
mon, year = latest_month
G19H = ("Federal Reserve Board, G.19 Consumer Credit, historical table "
        "'Consumer credit outstanding, seasonally adjusted, levels'")
add("TOTALSL", G19H, "%s %s row, column 'Total'" % (mon, year), tot, 0.005,
    units="millions of dollars")
add("REVOLSL", G19H, "%s %s row, column 'Revolving'" % (mon, year), rev, 0.005,
    units="millions of dollars")
add("NONREVSL", G19H, "%s %s row, column 'Nonrevolving'" % (mon, year), nonrev, 0.005,
    units="millions of dollars")

G19C = "Federal Reserve Board, G.19 Consumer Credit, current release"
nsa_total, _ = g19_release_row("Total 4,512.7")   # NSA levels table, 'Total' row
if nsa_total is None:
    # the NSA levels row is labelled just 'Total'; take the second such row
    html = (C / "g19_current_default.htm").read_text(encoding="utf-8", errors="replace")
    rows = []
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S):
        text = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", row)).strip()
        if re.match(r"^Total\b", text) and "percent change" not in text.lower() \
           and "flow" not in text.lower():
            nums = [re.sub(r"<[^>]+>|&nbsp;", "", n).strip().replace(",", "")
                    for n in re.findall(r"<td[^>]*>(.*?)</td>", row, re.S)]
            nums = [float(n) for n in nums if re.fullmatch(r"-?\d+(\.\d+)?", n)]
            if nums:
                rows.append(nums[-1])
    nsa_total = rows[-1] if rows else None

pct, _ = g19_release_row("Total percent change")
add("TOTALNS", G19C,
    "table 'Consumer Credit Outstanding (Levels), not seasonally adjusted, "
    "billions of dollars', row 'Total', column '%s %s p'" % (mon, year),
    nsa_total * 1000.0 if nsa_total is not None else None, 50.0,
    note="the release rounds to a tenth of a billion; the tolerance is "
         "$50 million, which is that rounding and nothing wider",
    units="millions of dollars (release prints billions)")
add("TOTALSLAR", G19C,
    "table 'Consumer Credit Outstanding, seasonally adjusted', row "
    "'Total percent change (annual rate)', column '%s %s p'" % (mon, year),
    pct, 0.05,
    note="the release prints one decimal; the tolerance is that rounding",
    units="percent, annual rate")

# ----------------------------------------------------------------- DSR ------
# Columns after the quarter label: FOR, then DSR total, mortgage, consumer.
# The page at /releases/housedebt/ is an ARCHIVE, frozen at 2024:1: from the
# 2024:Q2 publication the Board moved the DSR to a credit-bureau methodology
# and to a new address. Reading the archive would have compared a 2026 figure
# against a discontinued series computed a different way -- which is exactly
# the "same basis" check, and it is the reason this file names the release it
# read rather than saying "the Fed".
dsr_html = (C / "dsr_new.htm").read_text(encoding="utf-8", errors="replace")
flat = re.sub(r"<[^>]+>", " ", dsr_html)
flat = re.sub(r"[ \t]+", " ", flat)
dsr = {}
for year, q, body in re.findall(r"(\d{4}):(\d)((?:\s+-?\d+\.\d+){3})", flat):
    dsr[(year, q)] = [float(v) for v in body.split()]
dsr_latest = max(dsr) if dsr else None
DSRSRC = ("Federal Reserve Board, Household Debt Service Ratio release "
          "(credit-bureau methodology, in force from the 2024:Q2 publication)")
for sid, col, name in (("TDSP", 0, "DSR"), ("MDSP", 1, "Mortgage DSR"),
                       ("CDSP", 2, "Consumer DSR")):
    row = dsr.get(dsr_latest)
    add(sid, DSRSRC,
        "row '%s:%s', column '%s'" % (dsr_latest[0], dsr_latest[1], name)
        if dsr_latest else "no rows parsed",
        row[col] if row else None, 0.005,
        note="the Board publishes two decimals; FRED carries six. The "
             "tolerance is the Board's own printed precision.",
        units="percent of disposable personal income")

# --------------------------------------------------------- FHFA monthly -----
# hpi_po_monthly_hist is served as a workbook, whatever its extension says.
wb = openpyxl.load_workbook(io.BytesIO((C / "fhfa_po_monthly.csv").read_bytes()),
                            data_only=True)
ws = wb[wb.sheetnames[0]]
# Three title rows sit above the header; row 4 carries the column names.
HEADER_ROW = 4
header = [str(c.value).replace("\n", " ").strip() if c.value is not None else ""
          for c in ws[HEADER_ROW]]
rows = [[c.value for c in r] for r in ws.iter_rows(min_row=HEADER_ROW + 1)]
print("FHFA monthly sheet %r, header: %s" % (wb.sheetnames[0], header[:8]))

hp = ours.get("HPIPONM226S", {}).get("latest")
theirs = None
where = "hpi_po_monthly_hist, USA (SA) row for the workbook's latest month"
if hp:
    want_y, want_m = int(hp["date"][:4]), int(hp["date"][5:7])
    def col(*names):
        for i, h in enumerate(header):
            if any(n.lower() in h.lower() for n in names):
                return i
        return None
    # One row per month; one column per census division plus USA, each in a
    # seasonally adjusted and an unadjusted flavour. FRED's HPIPONM226S is the
    # seasonally adjusted national index, so it is the USA (SA) column and not
    # the USA (NSA) one sitting immediately beside it.
    c_sa = col("USA  (SA)", "USA (SA)")
    for r in rows:
        stamp = r[0]
        if stamp is None:
            continue
        y, m = getattr(stamp, "year", None), getattr(stamp, "month", None)
        if y is None:
            text = str(stamp)
            if len(text) < 7:
                continue
            y, m = int(text[:4]), int(text[5:7])
        if y == want_y and m == want_m:
            theirs = float(r[c_sa]) if c_sa is not None and r[c_sa] is not None else None
            where = ("hpi_po_monthly_hist sheet %r, row %04d-%02d, column %r "
                     "(column %d of %d)"
                     % (wb.sheetnames[0], y, m, header[c_sa], c_sa + 1, len(header)))
            break
add("HPIPONM226S", "FHFA, monthly purchase-only house price index",
    where, theirs, 0.005, units="index, Jan 1991 = 100")

(SB / "fred_other_results.json").write_text(json.dumps(results, indent=1), encoding="utf-8")
from collections import Counter
print("\nseries compared: %d  %s"
      % (len(results), Counter(r["verdict"] for r in results).most_common()))
for r in results:
    mark = " " if r["verdict"] == "TIED" else "*"
    print("%s %-12s ours %14s  theirs %14s  %s"
          % (mark, r["series"], r.get("ours"), r.get("theirs"),
             r.get("why") or r["source_where"][:66]))
