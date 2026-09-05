"""The ours side of the FRED set: what the workbook holds, read from the workbook.

Not re-fetched from FRED. The artifact is the .xlsm a person opens, and the
raw tabs are where its numbers live: a header row per series (id, title, meta),
then date/value rows newest-first.
"""
import json
import pathlib
import sys

import openpyxl

CS = pathlib.Path(r"C:\Users\ajish\SATC-cs\credit-suite")
SB = pathlib.Path(r"C:\Users\ajish\AppData\Local\Temp\claude"
                  r"\C--Users-ajish-SATC\261f7248-3cbc-4aa2-aacf-e4ff9181778a\scratchpad")
WB = CS / "example-output" / "FRED_Credit_Risk_Dashboard.xlsm"
sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

wb = openpyxl.load_workbook(WB, keep_vba=True, data_only=True)
out = {}
for tab in ("Raw_Consumer", "Raw_Commercial", "Raw_Price"):
    ws = wb[tab]
    row = 1
    current = None
    while row <= ws.max_row:
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        if isinstance(a, str) and isinstance(b, str) and b and a.isupper() and " " not in a:
            current = {"series": a, "title": b, "meta": ws.cell(row, 3).value,
                       "tab": tab, "header_row": row, "observations": []}
            out[a] = current
        elif current is not None and isinstance(a, str) and a[:2] in ("19", "20"):
            value = ws.cell(row, 2).value
            current["observations"].append(
                {"date": str(a)[:10], "value": value, "row": row})
        row += 1
wb.close()

for s in out.values():
    s["latest"] = s["observations"][0] if s["observations"] else None
    s["n"] = len(s["observations"])

(SB / "fred_ours.json").write_text(json.dumps(out, indent=1), encoding="utf-8")
empty = [k for k, v in out.items() if not v["observations"]]
print("series blocks read from the workbook: %d" % len(out))
print("with no observations: %d %s" % (len(empty), empty[:8]))
print("total observations: %d" % sum(v["n"] for v in out.values()))
for k in list(out)[:3]:
    v = out[k]
    print("  %-14s %-46s latest %s = %s (cell B%d)"
          % (k, v["title"][:46], v["latest"]["date"], v["latest"]["value"], v["latest"]["row"]))
