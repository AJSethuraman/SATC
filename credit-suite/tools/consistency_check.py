#!/usr/bin/env python3
"""Run the data-consistency checks against a workbook that already exists.

    python tools/consistency_check.py example-output/FRED_Credit_Risk_Dashboard.xlsm
    python tools/consistency_check.py Bank_Peer_Monitor.xlsm

Exit 0 when nothing failed, 2 when something did -- drift is a gate error, not
a crash, the same convention ``tools/conformance.py`` and
``tools/check_parity.py`` already use. An UNKNOWN never sets the exit code; it
is printed, because a gate that refuses on "could not establish" trains the
operator to bypass it.

**Why a tool and not only a test.** The builder already derives ``_config``
from ``series_seed.py``, so a freshly built workbook agrees with the seed by
construction and a test over the builder can only ever confirm that. The file
on somebody's desk is a different question. The shipped FRED workbook declares
``billions $`` for four G.19 series that the seed corrected to ``millions $``
on 5 September 2026 -- the source was fixed, the artifact was not, and
``example-output/`` is in ``.gitignore``, so nothing in version control can
tell you. Point this at the artifact.

Every line carries its denominator. A check with nothing to look at prints
NONE, never ok.
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl                                            # noqa: E402

import trend                                               # noqa: E402
from credit_suite.engine import consistency as K           # noqa: E402
from credit_suite.sources.fdic import consistency as FDIC  # noqa: E402
from credit_suite.sources.fred import consistency as FRED  # noqa: E402
from credit_suite.sources.fred import runner as FRED_RUNNER  # noqa: E402

FRED_RAW_TABS = ("Raw_Consumer", "Raw_Commercial", "Raw_Price")


# --------------------------------------------------------------------------
# reading what is already in the file
# --------------------------------------------------------------------------

def fred_checks(path: str) -> list:
    """S1, C3 and C5 over a FRED workbook on disk."""
    wb = openpyxl.load_workbook(path, keep_vba=path.lower().endswith(".xlsm"),
                                data_only=True)
    try:
        if "_config" not in wb.sheetnames:
            return [K.undetermined("S1", "%s has no _config tab" % path)]
        rows = [list(row) for row in wb["_config"].iter_rows(values_only=True)]
        results = [FRED.config_matches_seed(rows)]

        cfg = FRED_RUNNER.parse_config(rows)
        blocks = FRED_RUNNER.raw_layout(cfg.series, cfg.raw_slots)
        vintages = {}
        grids = {}
        for spec in cfg.series:
            block = blocks[spec.series_id]
            if block.tab not in wb.sheetnames:
                continue
            ws = wb[block.tab]
            vintages[spec.series_id] = FRED.vintage_of(
                ws.cell(block.header_row, FRED_RUNNER.RAW_VALUE_COL + 1).value)
            dates = []
            for offset in range(block.slots):
                value = ws.cell(block.first_data_row + offset,
                                FRED_RUNNER.RAW_DATE_COL).value
                if value in (None, ""):
                    break
                dates.append(value)
            grids[spec.series_id] = (dates, spec.frequency, True)
        # No per-run manifest exists yet, so "did the vintage move forward"
        # cannot be answered from one file. That is UNKNOWN, and saying so is
        # the point (design item C2 is the manifest that would settle it).
        results.append(FRED.vintage_check(vintages, previous=None))
        results.append(FRED.date_grid_all(grids))
        return results
    finally:
        wb.close()


def fdic_panel(path: str) -> dict:
    """`{(cert, period): {field: value}}` from ``Raw_FDIC``.

    Read through ``trend.read_panel`` rather than a second parser of the same
    tab: two readers of one layout is two definitions of what a bank-quarter
    is, and the first time they disagree neither is the answer.
    """
    panels = trend.read_panel(path, raw_tab="Raw_FDIC", derive=False)
    out = defaultdict(dict)
    for field, panel in panels.items():
        for series in panel.series.values():
            for period, value in zip(series.periods, series.values):
                if value is not None:
                    out[(str(series.cert), str(period)[:10])][field] = value
    return dict(out)


#: What the demo build's own status panel says about itself.
DEMO_MARKER = "(demo)"

#: Why the identity set does not run against a demo build. Measured on the
#: workbook ``monitorbuild`` produces, 5 September 2026: the synthetic provider
#: rounds each published ratio to four decimal places independently of its
#: components (``NCLNLSR = 0.8706`` against ``NCLNLS/LNLSGR = 0.87063304``),
#: rounds net loans independently of gross less reserve (off by one thousand
#: dollars in 44 of 156 bank-quarters), and draws each loan class independently
#: of the noncurrent total (a constant 1.5769 for Wells Fargo in every
#: quarter). None of that is a defect in the monitor -- these identities are
#: assertions about the PUBLISHER's numbers, and a demo build has no publisher.
#: Running them anyway would refuse every demo build and teach whoever sees it
#: that this tool cries wolf.
DEMO_REASON = (
    "this workbook was built from the synthetic demo provider, whose values "
    "do not satisfy the publisher's own identities (it rounds each ratio "
    "independently of its components). The identity set asserts things about "
    "the FDIC's numbers; run it against a live build")


def _is_demo(path: str) -> bool:
    from credit_suite.sources.fdic.engine_api import (DASH_TABS, STATUS_COL,
                                                      STATUS_COL_BY_TAB)
    wb = openpyxl.load_workbook(path, read_only=True,
                                keep_vba=path.lower().endswith(".xlsm"))
    try:
        for tab in DASH_TABS:
            if tab not in wb.sheetnames:
                continue
            col = STATUS_COL_BY_TAB.get(tab, STATUS_COL)
            value = wb[tab].cell(1, col).value
            if value and DEMO_MARKER in str(value):
                return True
    finally:
        wb.close()
    return False


def fdic_checks(path: str) -> list:
    panel = fdic_panel(path)
    if _is_demo(path):
        # The denominator survives even when nothing could be settled: "0 of
        # 192 bank-quarters, 192 unknown" is a different report from "nothing
        # examined", and only one of them says how big the hole is.
        results = [K.decide(check, len(panel), unknowns=[DEMO_REASON],
                            unknown=len(panel), unit="bank-quarters")
                   for check in ("I2", "I3", "I4", "I5")]
    else:
        results = list(FDIC.identity_set(panel))
    wb = openpyxl.load_workbook(path, keep_vba=path.lower().endswith(".xlsm"),
                                data_only=True)
    try:
        record = trend.read_mergers(wb)
    finally:
        wb.close()
    results.append(FDIC.comparability_check(sorted(panel), record))
    return results


# --------------------------------------------------------------------------

def which(path: str) -> str:
    wb = openpyxl.load_workbook(path, read_only=True,
                                keep_vba=path.lower().endswith(".xlsm"))
    try:
        tabs = set(wb.sheetnames)
    finally:
        wb.close()
    if "Raw_FDIC" in tabs:
        return "fdic"
    if tabs & set(FRED_RAW_TABS):
        return "fred"
    raise SystemExit(
        "%s carries neither Raw_FDIC nor a FRED raw tab, so there is nothing "
        "here this knows how to check. Refusing rather than reporting a clean "
        "run over a file it did not understand." % path)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("--monitor", choices=("fdic", "fred"),
                    help="skip the tab sniff and say which monitor this is")
    args = ap.parse_args(argv)

    path = args.workbook
    if not Path(path).exists():
        raise SystemExit("%s does not exist" % path)
    monitor = args.monitor or which(path)

    results = fdic_checks(path) if monitor == "fdic" else fred_checks(path)
    print("%s -- %s monitor" % (path, monitor))
    for result in results:
        print("  " + result.summary())

    failed = [r for r in results if r.blocking]
    unknown = [r for r in results if r.verdict == K.UNKNOWN]
    nothing = [r for r in results if r.verdict == K.NONE]
    print("\n%d checks: %d failed, %d unknown, %d examined nothing"
          % (len(results), len(failed), len(unknown), len(nothing)))
    if nothing:
        print("A check that examined nothing is not a check that passed: %s"
              % ", ".join(r.check for r in nothing))
    return 2 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
