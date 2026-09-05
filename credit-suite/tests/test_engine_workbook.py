"""Metrics, the provider seam, the workbook writer and the run guards.

The two tests that earn their place most are ``test_clearing_actually_blanks``
(carried lesson L7) and the raw-layout refusals. Both cover failures that are
invisible: the workbook saves, opens, and shows numbers -- just the wrong ones,
under the right timestamp.
"""

from __future__ import annotations

import itertools

import openpyxl
import pytest

from credit_suite.engine import metrics, provider, rawlayout, runtime, workbook
from credit_suite.engine.config import Config, EntityRow
from credit_suite.sources.fdic.spec import FDIC


FIELDS = ["ASSET", "DEP", "LNLSGR"]
UNITS = {f: "USD_thousands" for f in FIELDS}


def entity(slot=1, key="628", name="Test Bank", group="peer", active=True):
    return EntityRow(slot=slot, key=key, name=name, group=group, active=active,
                     key_prefix="cert")


def make_workbook(tmp_path, fields=FIELDS, slots=3, raw_slots=4, name="wb.xlsx"):
    """A minimal workbook carrying the raw scaffold a runner writes into."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FDIC.raw_tab
    ws.cell(1, 1, "Raw block note")
    for slot in range(1, slots + 1):
        block = rawlayout.slot_block(slot, raw_slots)
        ws.cell(block.header_row, 1, rawlayout.slot_label(slot))
        ws.cell(block.label_row, 1, "REPDTE")
        for fname in fields:
            ws.cell(block.label_row, rawlayout.field_col(fname, fields), fname)
    cfg = wb.create_sheet("_config")
    cfg.cell(1, 1, "[SETTINGS]")
    cfg.cell(2, 1, "raw_slots")
    cfg.cell(2, 2, raw_slots)
    cfg.cell(3, 1, "peer_slots")
    cfg.cell(3, 2, slots)
    path = tmp_path / name
    wb.save(path)
    return path


def backend_for(tmp_path, **kw):
    path = make_workbook(tmp_path, **kw)
    return workbook.OpenpyxlBackend(str(path), FDIC, kw.get("fields", FIELDS)), path


# --------------------------------------------------------------------------
# metrics
# --------------------------------------------------------------------------

def test_a_missing_input_blanks_the_metric_and_never_reads_as_zero():
    """A null uninsured-deposit figure shown as 0% would read as a bank with no
    uninsured deposits -- the opposite of unknown."""
    assert metrics.ratio(None, 100.0) is None
    assert metrics.ratio(5.0, None) is None
    assert metrics.total(1.0, None, 2.0) is None
    assert metrics.ratio(0.0, 100.0) == 0.0, "a real zero is still a zero"


def test_a_zero_denominator_blanks_rather_than_raising():
    assert metrics.ratio(5.0, 0.0) is None


def test_the_multiplier_annualises_a_quarterly_flow():
    assert metrics.ratio(1.0, 100.0, 400) == 4.0
    assert metrics.ratio(1.0, 100.0, 100) == 1.0


def test_validate_metrics_refuses_config_drift_rather_than_guessing():
    registry = metrics.build_registry(direct=["ASSET"], ratios={}, derived={})

    class Row:
        def __init__(self, id, transform):
            self.id, self.transform = id, transform

    with pytest.raises(metrics.MetricError, match="not a registered metric"):
        metrics.validate_metrics([Row("NOPE", "direct")], registry, FIELDS)
    with pytest.raises(metrics.MetricError, match="declares transform"):
        metrics.validate_metrics([Row("ASSET", "derived")], registry, FIELDS)

    registry2 = metrics.build_registry(
        direct=[], ratios={"R": ("MISSING", "DEP", 100)}, derived={})
    with pytest.raises(metrics.MetricError, match="unlanded field"):
        metrics.validate_metrics([Row("R", "derived")], registry2, FIELDS)


# --------------------------------------------------------------------------
# the provider seam
# --------------------------------------------------------------------------

def test_the_secret_is_read_by_name_from_the_environment(monkeypatch):
    """The env var's NAME is config; its value is never stored anywhere."""
    cfg = Config(spec=FDIC, settings={"secret_env": "CREDIT_SUITE_TEST_KEY"})
    monkeypatch.delenv("CREDIT_SUITE_TEST_KEY", raising=False)
    assert provider.resolve_secret(cfg) is None
    monkeypatch.setenv("CREDIT_SUITE_TEST_KEY", "s3cret")
    assert provider.resolve_secret(cfg) == "s3cret"


def test_a_keyless_monitor_resolves_no_secret():
    assert provider.resolve_secret(Config(spec=FDIC, settings={})) is None
    assert provider.resolve_secret(
        Config(spec=FDIC, settings={"secret_env": "  "})) is None


def test_a_licensed_adapter_refuses_to_call_without_its_secret():
    """Class C stays gated until a contract exists.

    The refusal is a 401-shaped PermissionError, which is the shipped monitor's
    contract -- the secret arrives per call, so nothing licensed is ever held on
    the instance.
    """
    stub = provider.ClassCStubProvider(secret_env="FDIC_CLASS_C_SECRET")
    spec = provider.make_field_spec(entity(), "ASSET", UNITS)
    with pytest.raises(PermissionError, match="401"):
        stub.fetch_series(spec, secret=None)

    rows = stub.fetch_series(spec, secret="dummy-client-credentials")
    assert len(rows) == 1
    assert rows[0].value is None, "the stub must never fabricate licensed data"
    assert rows[0].source_class == "C"


def test_a_missing_secret_maps_to_the_contracts_exit_code_3():
    """``MissingSecret`` is deliberately unraised by any adapter today: FDIC's
    source is keyless, so there is no secret to be missing.

    It is kept because exit code 3 is in the contract and a keyed source needs
    somewhere to raise from -- FRED, whose live pull takes an API key, is the
    first. Asserting the mapping here stops it being untested code that only
    gets exercised the day it matters.
    """
    assert issubclass(provider.MissingSecret, SystemExit)
    assert runtime.EXIT_MISSING_SECRET == 3


def test_a_licensed_adapter_is_refused_by_the_gate_not_by_missing_code():
    """The stub exists so the Class C path is rehearsed. If it were deleted the
    gate would be untestable, and an untested gate is the one that opens."""
    from credit_suite.engine import gates
    from credit_suite.engine.config import SeriesSpec

    row = SeriesSpec(id="X", title="", category="", lane="", metric_type="",
                     frequency="", sa_nsa="", units="", level_rate_index="",
                     geo_segment="", source_class=provider.ClassCStubProvider.source_class,
                     dashboard_capable=True, watchlist_capable=True,
                     source_url="", table_id="", sheet="", series_label="",
                     transform="direct", notes="")
    assert gates.gate_metric_row(row, FDIC), "Class C was admitted"


def test_field_specs_carry_the_entity_key_and_the_slot_scheme():
    spec = provider.make_field_spec(entity(slot=7, key="628"), "ASSET", UNITS)
    assert spec.id == "s07_ASSET"
    assert spec.geo_segment == "cert:628"
    assert spec.units == "USD_thousands"


# --------------------------------------------------------------------------
# the workbook writer -- L7 and the layout guard
# --------------------------------------------------------------------------

def test_clearing_actually_blanks(tmp_path):
    """Carried lesson L7, and the reason it exists.

    ``ws.cell(r, c, None)`` is a silent no-op in openpyxl. The bug it caused was
    invisible: a successful run rewrote the same shape over the top, so clearing
    only mattered after a FAILED fetch -- at which point last quarter's figures
    sat under this quarter's timestamp, looking current.
    """
    backend, path = backend_for(tmp_path)
    block = rawlayout.slot_block(1, 4)
    backend.write_slot_block(block, entity(), [
        ("2026-03-31", {"ASSET": 100.0, "DEP": 50.0, "LNLSGR": 25.0}),
        ("2025-12-31", {"ASSET": 90.0, "DEP": 45.0, "LNLSGR": 20.0}),
    ])
    backend.finalize()

    wb = openpyxl.load_workbook(path)
    ws = wb[FDIC.raw_tab]
    assert ws.cell(block.first_data_row, 2).value == 100.0, "nothing was written"
    assert ws.cell(block.header_row, 3).value == "Test Bank"
    wb.close()

    backend2 = workbook.OpenpyxlBackend(str(path), FDIC, FIELDS)
    backend2.clear_slot_block(block)
    backend2.finalize()

    wb = openpyxl.load_workbook(path)
    ws = wb[FDIC.raw_tab]
    for row in range(block.first_data_row, block.first_data_row + block.slots):
        for col in range(1, 1 + len(FIELDS) + 1):
            assert ws.cell(row, col).value is None, \
                "stale data survived at r%dc%d" % (row, col)
    for col in (2, 3, 4):
        assert ws.cell(block.header_row, col).value is None, "identity survived"
    assert ws.cell(block.header_row, 1).value == "slot01", \
        "the layout sentinel must survive -- it is not data"
    wb.close()


def test_a_null_value_lands_as_a_blank_not_a_zero(tmp_path):
    """Trap F3 through the write path: a null must not become 0.0 in the sheet."""
    backend, path = backend_for(tmp_path)
    block = rawlayout.slot_block(1, 4)
    backend.write_slot_block(block, entity(), [
        ("2026-03-31", {"ASSET": 100.0, "DEP": None, "LNLSGR": 0.0}),
    ])
    backend.finalize()

    ws = openpyxl.load_workbook(path)[FDIC.raw_tab]
    assert ws.cell(block.first_data_row, 2).value == 100.0
    assert ws.cell(block.first_data_row, 3).value is None, "null became a number"
    assert ws.cell(block.first_data_row, 4).value == 0.0, "a real zero was lost"


def test_a_relabelled_slot_is_refused_with_the_rebuild_command(tmp_path):
    """Formulas are anchored to the BUILT layout, so writing into a workbook
    whose layout moved would land correct data where nothing reads it."""
    path = make_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    block = rawlayout.slot_block(2, 4)
    wb[FDIC.raw_tab].cell(block.header_row, 1, "slot99")
    wb.save(path)

    backend = workbook.OpenpyxlBackend(str(path), FDIC, FIELDS)
    with pytest.raises(workbook.RawLayoutMismatch) as exc:
        backend.clear_slot_block(block)
    assert "slot02" in str(exc.value) and "slot99" in str(exc.value)
    assert FDIC.rebuild_command in str(exc.value), "no rebuild command to act on"


def test_a_workbook_built_by_another_pack_is_refused(tmp_path):
    """The last field's label is the cheap sentinel: an older pack has fewer
    raw columns, and this field set would land under no label at all."""
    path = make_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    block = rawlayout.slot_block(1, 4)
    col = rawlayout.field_col(FIELDS[-1], FIELDS)
    wb[FDIC.raw_tab].cell(block.label_row, col, "SOMETHINGELSE")
    wb.save(path)

    backend = workbook.OpenpyxlBackend(str(path), FDIC, FIELDS)
    with pytest.raises(workbook.RawLayoutMismatch, match=FDIC.pack_version):
        backend.clear_slot_block(block)


def test_an_editable_peer_swap_never_trips_the_layout_guard(tmp_path):
    """The property that makes a peer list a config edit: the slot label does
    not depend on who occupies the slot."""
    backend, path = backend_for(tmp_path)
    block = rawlayout.slot_block(1, 4)
    backend.write_slot_block(block, entity(key="628", name="Bank A"),
                             [("2026-03-31", {"ASSET": 1.0, "DEP": 2.0,
                                              "LNLSGR": 3.0})])
    backend.clear_slot_block(block)
    backend.write_slot_block(block, entity(key="3511", name="Bank B"),
                             [("2026-03-31", {"ASSET": 9.0, "DEP": 8.0,
                                              "LNLSGR": 7.0})])
    backend.finalize()
    ws = openpyxl.load_workbook(path)[FDIC.raw_tab]
    assert ws.cell(block.header_row, 2).value == "s01 cert:3511"
    assert ws.cell(block.first_data_row, 2).value == 9.0


def test_keep_vba_is_never_used_on_an_xlsx(tmp_path):
    """L2: keep_vba on an .xlsx injects a dangling relationship Excel rejects
    outright as 'format or extension not valid'."""
    backend, path = backend_for(tmp_path)
    backend.finalize()
    assert path.suffix == ".xlsx"
    reopened = openpyxl.load_workbook(path)          # would raise if corrupt
    assert FDIC.raw_tab in reopened.sheetnames


# --------------------------------------------------------------------------
# the run guards
# --------------------------------------------------------------------------

def test_zero_pulls_where_pulls_were_expected_is_a_failure():
    """Without this a total outage exits 0 over a workbook of blanks under a
    fresh timestamp, which reads as 'checked, nothing wrong'."""
    assert runtime.run_succeeded({"entities_active": 12, "entities_landed": 0}) \
        is False
    assert runtime.run_succeeded({"entities_active": 12, "entities_landed": 1}) \
        is True


def test_a_monitor_with_nothing_to_pull_is_not_a_failure():
    """An empty peer list is a configuration, not an outage."""
    assert runtime.run_succeeded({"entities_active": 0, "entities_landed": 0}) \
        is True


def test_the_exit_codes_are_the_ones_the_contract_names():
    assert (runtime.EXIT_OK, runtime.EXIT_RUN_ERROR, runtime.EXIT_GATE_ERROR,
            runtime.EXIT_MISSING_SECRET) == (0, 1, 2, 3)
