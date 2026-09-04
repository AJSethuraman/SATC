"""The email-sim acceptance (contract section 9), as far as it can go at #165.

The shipped acceptance has two halves:

1. **The monitoring email composes from the run digest** and carries the ranked
   peer table, the per-dimension alerts, the per-class loan-book alerts, the
   staleness section and the data-vintage line. That is asserted here.

2. **The workbook rebuilds itself in an empty folder** -- copy only the `.xlsm`,
   extract `runner.py` from `_code_py`, run it, and the workbook repopulates
   with nothing else present. That cannot pass yet: `_code_py` currently holds a
   runner that imports `credit_suite`, and making it self-contained is precisely
   what the build-time inliner does. It is issue #167's acceptance criterion,
   and `test_the_self_contained_half_is_not_claimed_yet` records that the gap is
   known rather than letting silence imply it passes.
"""

from __future__ import annotations

from datetime import date

import openpyxl
import pytest

import monitorbuild
from credit_suite.sources.fdic import email_digest, fields, runner

ASOF = date(2026, 3, 31)


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    """A demo-run workbook plus its status, built through the engine."""
    keep = tmp_path_factory.mktemp("emailsim")
    with monitorbuild.built_monitor("fdic", run_demo=False) as (workbook, _):
        target = keep / workbook.name
        target.write_bytes(workbook.read_bytes())
    status = runner.run(str(target), demo=True, asof=ASOF)
    return target, status


def test_the_email_carries_the_ranked_peer_table(built):
    _workbook, status = built
    email = email_digest.compose_email(status)
    rows = [ln for ln in email.splitlines()
            if ln.strip() and ln.strip()[0].isdigit()
            and any(g in ln for g in ("peer", "counterparty", "self"))]
    assert "RANKED PEER TABLE" in email
    assert len(rows) == status["banks_landed"] == 12
    assert "ALERT" in email and "Texas" in email


def test_the_email_carries_every_section_a_reviewer_acts_on(built):
    _workbook, status = built
    email = email_digest.compose_email(status)
    assert "PER-DIMENSION ALERTS" in email
    assert any(d in email for d in ("composite", "capital"))
    assert "PER-CLASS LOAN-BOOK ALERTS" in email
    assert any(c in email for c in fields.CONSUMER_CLASSES)
    assert any(c in email for c in fields.COMMERCIAL_CLASSES)
    assert "STALENESS FLAGS" in email
    assert "Data vintage:" in email


def test_the_email_is_deterministic_for_a_fixed_asof(built):
    _workbook, status = built
    assert email_digest.compose_email(status) == \
        email_digest.compose_email(status)


def test_the_macro_survives_a_refresh(built):
    """L2 through the whole path: the VBA project must still be there after the
    runner has opened, written and saved the closed workbook."""
    workbook, _status = built
    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        assert wb.vba_archive is not None, "the refresh dropped the macro"
    finally:
        wb.close()


def test_the_raw_block_is_populated_after_a_refresh(built):
    from credit_suite.sources.fdic.engine_api import RAW_TAB, slot_block

    workbook, _status = built
    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        block = slot_block(1, 16)
        assert wb[RAW_TAB].cell(block.first_data_row, 2).value is not None
    finally:
        wb.close()


def test_the_workbook_rebuilds_itself_in_an_empty_folder(built, tmp_path):
    """The second half of the shipped acceptance, closed by the inliner (#167).

    Copy ONLY the .xlsm into an empty folder. Reproduce what the VBA button
    does -- read `_code_py` and write `runner.py` beside the workbook. Then run
    it with an isolated interpreter and no `credit_suite` anywhere, and the
    workbook must repopulate itself with nothing else present.

    At #165 this could not pass, and a test asserted the gap rather than letting
    silence imply a pass. This is that test, inverted, now that it holds.
    """
    import json
    import os
    import subprocess
    import sys

    workbook, _status = built
    folder = tmp_path / "email"
    folder.mkdir()
    target = folder / workbook.name
    target.write_bytes(workbook.read_bytes())
    assert sorted(p.name for p in folder.iterdir()) == [workbook.name]

    runner_py = folder / "runner.py"
    email_digest.extract_code_tab(str(target), "_code_py", str(runner_py))
    assert runner_py.stat().st_size > 1000

    env = {"PATH": os.environ.get("PATH", ""),
           "SYSTEMROOT": os.environ.get("SYSTEMROOT", ""),
           "TEMP": str(folder), "TMP": str(folder)}
    result = subprocess.run(
        [sys.executable, "-I", "runner.py", "--workbook", workbook.name,
         "--demo", "--asof", email_digest.ASOF],
        cwd=str(folder), capture_output=True, text=True, env=env, timeout=900)
    assert result.returncode == 0, result.stdout + result.stderr

    payload = json.loads([ln for ln in result.stdout.splitlines()
                          if ln.startswith("{")][-1])
    assert payload["ok"] is True and payload["banks_landed"] == 12

    # Nothing but the workbook and the extracted runner: the workbook is the
    # source of truth, and it needed nothing else to rebuild.
    assert sorted(p.name for p in folder.iterdir()) == \
        sorted([workbook.name, "runner.py"])

    wb = openpyxl.load_workbook(target, keep_vba=True)
    try:
        assert wb.vba_archive is not None, "the self-refresh dropped the macro"
    finally:
        wb.close()
