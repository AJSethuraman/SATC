"""tools/chartbook.py: the shape of the chart workbook, and the two defects that
"opened with zero dialogs" could not see.

A harness read a cell out of the first build and it was called done. Exporting
a chart to PNG and looking showed no axis numbers and a legend drawn over the
data. These tests pin the properties that fix produced, in the chart XML
openpyxl writes -- which is not the same as looking, and does not claim to be.
The look lives in the acceptance runbook.
"""
from __future__ import annotations

import statistics
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

import chartbook as C                                  # noqa: E402
import monitorbuild                                    # noqa: E402


def line_colour(series) -> str:
    """After a save/load round-trip openpyxl hands the fill back as a bare hex
    string; before one it is a ColorChoice. Accept either."""
    fill = series.graphicalProperties.line.solidFill
    if isinstance(fill, str):
        return fill
    rgb = fill.srgbClr
    return rgb if isinstance(rgb, str) else rgb.val


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    out = tmp_path_factory.mktemp("charts") / "Charts.xlsx"
    with monitorbuild.built_monitor("fdic") as (workbook, _stdout):
        C.build(Path(workbook), out)
    return openpyxl.load_workbook(out)


def test_one_sheet_per_peer_metric_and_per_bank_plus_about(built):
    names = built.sheetnames
    assert names[0] == "About"
    assert sum(n.startswith("Peers_") for n in names) == len(C.PEER_METRICS)
    assert sum(n.startswith("Stages_") for n in names) == 12


def test_every_chart_draws_its_axes(built):
    """openpyxl 3.1 hides axes unless told not to. The first build shipped
    charts with no numbers on either axis."""
    charts = [ch for s in built.sheetnames for ch in built[s]._charts]
    assert charts, "no charts at all"
    for chart in charts:
        assert chart.x_axis.delete is False and chart.y_axis.delete is False


def test_no_legend_sits_on_top_of_the_plot(built):
    for s in built.sheetnames:
        for chart in built[s]._charts:
            if chart.legend is not None:
                assert chart.legend.position == "b"
                assert chart.legend.overlay is False


def test_the_peer_overview_is_grey_context_with_one_coloured_median(built):
    """Twelve hues would be unreadable and, past three or four, indistinguishable
    to a colour-blind reader. The overview is every bank in grey and the peer
    median in colour; each bank gets its own small chart below."""
    ws = built["Peers_NCLNLSR"]
    overview = ws._charts[0]
    colours = [line_colour(s) for s in overview.series]
    assert colours.count(C.ACCENT) == 1 and colours.count(C.CONTEXT) == len(colours) - 1
    assert len(ws._charts) == 1 + 12                   # overview + one per bank


def test_the_median_row_is_the_median_of_the_banks(built):
    ws = built["Peers_NCLNLSR"]
    grid = list(ws.iter_rows(min_row=20, max_row=33, values_only=True))
    header, *body = grid
    banks = [r for r in body if r[0] != C.MEDIAN_LABEL]
    median = next(r for r in body if r[0] == C.MEDIAN_LABEL)
    for col in range(1, len(header)):
        got = [r[col] for r in banks if r[col] is not None]
        if got:
            assert median[col] == pytest.approx(round(statistics.median(got), 4))


def test_a_stage_chart_uses_the_four_validated_hues(built):
    ws = next(built[s] for s in built.sheetnames if s.startswith("Stages_"))
    colours = [line_colour(s) for s in ws._charts[0].series]
    assert colours == C.HUES[:len(colours)]


def test_sheet_names_survive_excels_limits():
    assert len(C._sheet_name("Stages", "Bank of New York Mellon Trust Co")) <= 31
    assert "/" not in C._sheet_name("Peers", "A/B:C*D?")
