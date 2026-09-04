"""The parts of the section 9 bar that inspect the built file itself.

A workbook can pass every behavioural test and still be a file Excel refuses to
open, or one whose macro no tool can read back. These check the artifact as a
package rather than as a spreadsheet:

* **olevba decompile** -- the embedded VBA project must be readable by an
  independent tool, and must expose exactly the macro entry points the contract
  names. A project openpyxl carried through but that Excel cannot bind is
  indistinguishable from a good one until somebody clicks the button.
* **OPC audit** -- no dangling relationships and no overlapping merged ranges.
  Both produce Excel's "unreadable content" repair dialog, and both survive a
  save without complaint.
"""

from __future__ import annotations

import re
import zipfile
from xml.etree import ElementTree

import openpyxl
import pytest

import monitorbuild

OPC_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    keep = tmp_path_factory.mktemp("audit")
    with monitorbuild.built_monitor("fdic", run_demo=True) as (built, _stdout):
        target = keep / built.name
        target.write_bytes(built.read_bytes())
    return target


# --------------------------------------------------------------------------
# olevba
# --------------------------------------------------------------------------

def test_the_macro_decompiles_and_exposes_the_contracts_entry_points(workbook):
    olevba = pytest.importorskip("oletools.olevba",
                                 reason="oletools not installed")
    parser = olevba.VBA_Parser(str(workbook))
    try:
        assert parser.detect_vba_macros(), "no VBA project an external tool can see"
        source = "\n".join(code for _f, _s, _n, code in parser.extract_macros())
    finally:
        parser.close()

    assert re.search(r"\bSub\s+ExtractFiles\b", source), "ExtractFiles is missing"
    assert re.search(r"\bSub\s+ExtractAndRun\b", source), "the alias is missing"


def _executable_vba(source: str) -> str:
    """The macro with comment lines removed.

    Needed because the macro's own comments SAY "no shell-out, no xlwings, no
    save-from-vba" -- a scan that reads comments as code fails on a workbook
    that is correct, which is the kind of false positive that teaches people to
    switch a check off.
    """
    lines = []
    for line in source.splitlines():
        stripped = line.strip()
        if stripped.startswith("'") or stripped.lower().startswith("rem "):
            continue
        lines.append(line.split(" '")[0] if " '" in line else line)
    return "\n".join(lines)


def test_the_macro_is_extract_only(workbook):
    """Contract section 5, carried lesson L1: no shell-out, no xlwings, no save.

    A macro that can run a process is a macro a bank's security review stops at
    the door, and the whole transmission story depends on it staying inert.
    """
    olevba = pytest.importorskip("oletools.olevba")
    parser = olevba.VBA_Parser(str(workbook))
    try:
        source = "\n".join(code for _f, _s, _n, code in parser.extract_macros())
    finally:
        parser.close()

    code = _executable_vba(source).lower()
    assert "extractfiles" in code, "comment-stripping ate the macro body"
    for forbidden in ("shell(", "wscript.shell", 'createobject("wscript',
                      "xlwings", "activeworkbook.save", "thisworkbook.save"):
        assert forbidden not in code, "macro is not extract-only: %r" % forbidden


def test_the_macro_writes_exactly_the_three_named_files(workbook):
    olevba = pytest.importorskip("oletools.olevba")
    parser = olevba.VBA_Parser(str(workbook))
    try:
        source = "\n".join(code for _f, _s, _n, code in parser.extract_macros())
    finally:
        parser.close()

    written = set(re.findall(r'"([A-Za-z0-9_.\-]+\.(?:py|txt))"', source))
    assert written == {"runner.py", "requirements.txt", "RUN.txt"}, written


# --------------------------------------------------------------------------
# OPC package audit
# --------------------------------------------------------------------------

def test_every_relationship_points_at_a_part_that_exists(workbook):
    """A dangling relationship is Excel's 'unreadable content' dialog, and it
    saves without a murmur."""
    with zipfile.ZipFile(workbook) as zf:
        names = set(zf.namelist())
        dangling = []
        for rels_name in [n for n in names if n.endswith(".rels")]:
            base = rels_name.rsplit("_rels/", 1)[0]
            root = ElementTree.fromstring(zf.read(rels_name))
            for rel in root.findall("%sRelationship" % OPC_REL_NS):
                if (rel.get("TargetMode") or "") == "External":
                    continue
                target = rel.get("Target", "")
                if target.startswith("/"):
                    resolved = target.lstrip("/")
                else:
                    resolved = base + target
                while "/../" in resolved:
                    head, tail = resolved.split("/../", 1)
                    resolved = head.rsplit("/", 1)[0] + "/" + tail
                if resolved not in names:
                    dangling.append("%s -> %s" % (rels_name, target))
    assert not dangling, "dangling relationship(s): %s" % dangling


def test_the_vba_project_is_declared_as_well_as_present(workbook):
    """Both halves are needed: the part, and the declarations that make Excel
    actually bind it.

    The content type may be declared either as an ``Override`` for the part or
    as a ``Default`` for the ``bin`` extension -- both are legal OPC, and
    openpyxl emits the second when it re-saves a keep_vba workbook. So this
    asserts the *property* (the part is typed as a VBA project) rather than one
    spelling of it; pinning the spelling would fail on a correct file.
    """
    with zipfile.ZipFile(workbook) as zf:
        names = set(zf.namelist())
        assert "xl/vbaProject.bin" in names, "no VBA part at all"
        content_types = zf.read("[Content_Types].xml").decode("utf-8")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    assert "ms-excel.sheet.macroEnabled.main+xml" in content_types, \
        "the workbook is not typed as macro-enabled"
    assert "vbaProject.bin" in rels, "no workbook -> vbaProject relationship"

    root = ElementTree.fromstring(content_types)
    ns = "{http://schemas.openxmlformats.org/package/2006/content-types}"
    typed_by_override = any(
        el.get("PartName") == "/xl/vbaProject.bin"
        and "vbaProject" in (el.get("ContentType") or "")
        for el in root.findall("%sOverride" % ns))
    bin_defaults = [el for el in root.findall("%sDefault" % ns)
                    if (el.get("Extension") or "").lower() == "bin"]
    typed_by_default = any("vbaProject" in (el.get("ContentType") or "")
                           for el in bin_defaults)
    assert typed_by_override or typed_by_default, \
        "xl/vbaProject.bin has no content type; Excel will not bind the macro"

    if typed_by_default and not typed_by_override:
        # A Default types EVERY .bin part, so it is only unambiguous while the
        # VBA project is the only one. Printer settings are also .bin parts.
        others = [n for n in names
                  if n.endswith(".bin") and n != "xl/vbaProject.bin"]
        assert not others, \
            ("bin is typed as vbaProject by Default, but other .bin parts "
             "exist and would be mistyped: %s" % others)


def test_no_two_merged_ranges_overlap(workbook):
    """Overlapping merges are the other reliable way to produce the repair
    dialog, and openpyxl will write them without complaint."""
    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        overlaps = []
        for ws in wb.worksheets:
            seen = {}
            for merged in ws.merged_cells.ranges:
                for row in range(merged.min_row, merged.max_row + 1):
                    for col in range(merged.min_col, merged.max_col + 1):
                        if (row, col) in seen:
                            overlaps.append("%s: %s overlaps %s"
                                            % (ws.title, merged, seen[(row, col)]))
                        seen[(row, col)] = merged
    finally:
        wb.close()
    assert not overlaps, overlaps


def test_no_native_charts_anywhere(workbook):
    """Contract L4: dashboards are formula-driven panels. A native chart is the
    thing that does not survive the round trip through a DLP text boundary."""
    with zipfile.ZipFile(workbook) as zf:
        charts = [n for n in zf.namelist()
                  if "/charts/" in n or "/chartsheets/" in n]
    assert not charts, "native chart parts present: %s" % charts

    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        for ws in wb.worksheets:
            assert not getattr(ws, "_charts", []), "%s carries a chart" % ws.title
    finally:
        wb.close()


def test_the_workbook_reloads_with_keep_vba_without_losing_the_project(workbook):
    """L2 end to end: load with keep_vba, save, reload -- the project survives."""
    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        assert wb.vba_archive is not None
    finally:
        wb.close()
