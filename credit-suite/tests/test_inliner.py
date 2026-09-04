"""The transmission format survives consolidation (contract section 11, issue #167).

The property being defended is blunt: a colleague receives ONE plain-text file
through corporate email and ends up with a working workbook. Sharing an engine
between monitors threatened that, because a monitor no longer carries its own
runner to embed.

The tests that matter here are the slow ones. `test_the_bundle_builds_and_runs_in_an_empty_folder`
runs the emitted script in a scratch directory with an isolated interpreter and
no `credit_suite` on the path at all -- if inlining were incomplete, that is
where it shows, and nothing cheaper would catch it.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from credit_suite import parity
from credit_suite.bundles import SPECS
from credit_suite.engine import inline

#: Run bundles with an isolated interpreter and an empty PYTHONPATH, so a stray
#: development path cannot make an incomplete bundle look complete.
ISOLATED = [sys.executable, "-I"]


def _run(argv, cwd):
    env = {"PATH": __import__("os").environ.get("PATH", ""),
           "SYSTEMROOT": __import__("os").environ.get("SYSTEMROOT", ""),
           "TEMP": str(cwd), "TMP": str(cwd)}
    return subprocess.run(argv, cwd=str(cwd), capture_output=True, text=True,
                          env=env, timeout=900)


@pytest.fixture(scope="module", params=sorted(SPECS))
def bundle(request, tmp_path_factory):
    """One emitted bundle, and a fresh empty folder holding only it."""
    name = request.param
    spec = SPECS[name]
    text = inline.render_bundle(spec)
    folder = tmp_path_factory.mktemp("bundle_%s" % name)
    path = folder / ("build_%s.py" % spec.name)
    with open(path, "w", encoding="ascii", newline="\n") as handle:
        handle.write(text)
    return name, spec, folder, path


# --------------------------------------------------------------------------
# the format
# --------------------------------------------------------------------------

@pytest.mark.parametrize("name", sorted(SPECS))
def test_the_bundle_is_pure_ascii(name):
    """Anything non-ASCII may not survive the email boundary intact."""
    text = inline.render_bundle(SPECS[name])
    text.encode("ascii")                          # raises if not
    assert "\r" not in text, "a CR would change on the far side of a mail gateway"


@pytest.mark.parametrize("name", sorted(SPECS))
def test_the_bundle_is_byte_identical_run_to_run(name):
    """A bundle that differs each time cannot be diffed or checksummed, and
    'did this change?' stops having an answer.

    gzip stamps the clock into its header by default, which is exactly the kind
    of drift that looks like nothing and ruins the property.
    """
    first = inline.render_bundle(SPECS[name])
    second = inline.render_bundle(SPECS[name])
    assert first == second


def test_the_encoder_carries_no_clock():
    """Assert the MECHANISM, not the symptom.

    Comparing two encodes only proves the clock did not tick between them,
    which it usually will not -- the mutation check caught exactly that by
    surviving a removal of ``mtime=0``. gzip stores its timestamp in header
    bytes 4..7, so read them: zero, or the bundle carries a clock.
    """
    import base64

    header = base64.b64decode(inline.encode(b"payload"))[:10]
    assert header[:2] == b"\x1f\x8b", "not a gzip stream"
    assert header[4:8] == b"\x00\x00\x00\x00", \
        "gzip header carries a timestamp: %r" % header[4:8]

    assert inline.encode(b"same") == inline.encode(b"same")
    assert inline.encode(b"different") != inline.encode(b"same")


def test_a_module_reachable_only_through_a_function_level_import_is_inlined(tmp_path):
    """Discovery must look inside functions, not just at module scope.

    A lazily-imported module is still needed at runtime, and a bundle missing it
    fails in the empty folder it exists for. The real graph happens not to have
    a lazy-ONLY import today, so this builds one -- otherwise the guard is a
    branch nothing can exercise, which is how it survived a mutation.
    """
    root = tmp_path
    pkg = root / inline.PACKAGE
    (pkg / "sub").mkdir(parents=True)
    (pkg / "__init__.py").write_text("", encoding="utf-8")
    (pkg / "sub" / "__init__.py").write_text("", encoding="utf-8")
    (pkg / "sub" / "lazy.py").write_text("VALUE = 1\n", encoding="utf-8")
    (pkg / "sub" / "entry.py").write_text(
        "def later():\n"
        "    from credit_suite.sub.lazy import VALUE\n"
        "    return VALUE\n", encoding="utf-8")

    found = inline.discover(["credit_suite.sub.entry"], root=root)
    assert "credit_suite.sub.lazy" in found, \
        "a lazily-imported module was left out of the bundle"
    # ...and it is NOT an ordering constraint, because the import has not run
    # when the module executes. That is how a deliberate cycle is broken.
    assert found.index("credit_suite.sub.entry") < len(found)


@pytest.mark.parametrize("name", sorted(SPECS))
def test_every_module_the_source_needs_is_inlined(name):
    modules = inline.discover(list(SPECS[name].roots))
    assert "credit_suite.engine.package" in modules, "cannot assemble the .xlsm"
    assert any(m.endswith(".layout") for m in modules)
    assert any(m.endswith(".runner") for m in modules)
    assert "credit_suite" in modules, "the package itself must exist first"


def test_modules_are_ordered_so_dependencies_execute_first():
    """`from x import y` needs x already executed, or y will not exist."""
    modules = inline.discover(list(SPECS["fdic"].roots))
    position = {m: i for i, m in enumerate(modules)}
    for parent, child in [("credit_suite", "credit_suite.engine"),
                          ("credit_suite.engine.config", "credit_suite.engine.gates"),
                          ("credit_suite.engine.provider",
                           "credit_suite.sources.fdic.adapter"),
                          ("credit_suite.sources.fdic.fields",
                           "credit_suite.sources.fdic.engine_api")]:
        assert position[parent] < position[child], "%s after %s" % (parent, child)


def test_a_non_ascii_payload_is_refused_rather_than_shipped(tmp_path):
    with pytest.raises(inline.InlineError, match="ASCII"):
        inline._assert_ascii("ok\nnot ok: ⚠\n", "bundle")


# --------------------------------------------------------------------------
# the acceptance: an empty folder, an isolated interpreter
# --------------------------------------------------------------------------

def test_the_bundle_builds_and_runs_in_an_empty_folder(bundle):
    """The whole point, end to end.

    No `credit_suite` installed, no PYTHONPATH, nothing in the folder but the
    one script. If any module were missed, this is where it surfaces.
    """
    name, spec, folder, path = bundle

    assert sorted(p.name for p in folder.iterdir()) == [path.name], \
        "the folder must start with only the bundle in it"

    result = _run(ISOLATED + [path.name], folder)
    assert result.returncode == 0, result.stdout + result.stderr

    produced = sorted(p.name for p in folder.iterdir())
    for expected in ("%s.xlsm" % spec.workbook,
                     "%s_fallback.xlsx" % spec.workbook,
                     "runner.py", "macro.bas", "requirements.txt"):
        assert expected in produced, "%s missing; got %s" % (expected, produced)

    wb = openpyxl.load_workbook(folder / ("%s.xlsm" % spec.workbook), keep_vba=True)
    try:
        assert wb.vba_archive is not None, "the built .xlsm carries no macro"
    finally:
        wb.close()


def test_the_extracted_runner_refreshes_the_workbook_alone(bundle):
    """Step two of the button: the runner the macro writes out must work with
    nothing else present either."""
    name, spec, folder, path = bundle
    if not (folder / "runner.py").exists():
        _run(ISOLATED + [path.name], folder)

    asof = parity.SPINE_BASELINES[name]["asof"]
    result = _run(ISOLATED + ["runner.py", "--workbook",
                              "%s.xlsm" % spec.workbook, "--demo",
                              "--asof", asof], folder)
    assert result.returncode == 0, result.stdout + result.stderr
    payload = json.loads([ln for ln in result.stdout.splitlines()
                          if ln.startswith("{")][-1])
    assert payload["ok"] is True
    assert payload["mode"] == "demo"


def test_parity_survives_inlining(bundle):
    """The bundle-built workbook must still match the pre-consolidation golden.

    Inlining is a change to how the code travels, not to what it computes, and
    this is the assertion that keeps those two things separate.
    """
    name, spec, folder, path = bundle
    workbook = folder / ("%s.xlsm" % spec.workbook)
    if not workbook.exists():
        result = _run(ISOLATED + [path.name], folder)
        assert result.returncode == 0, result.stdout + result.stderr

    baseline = parity.SPINE_BASELINES[name]
    golden = parity.read_golden(parity.repo_root() / baseline["demo_golden"])
    current = parity.snapshot_workbook(workbook, source=baseline["workbook"])
    diffs = parity.diff_snapshots(golden, current, ignore=parity.MIGRATION_IGNORE)

    compared = len(set(golden["cells"]) | set(current["cells"]))
    assert compared > 20000, "only %d cells compared" % compared
    assert not diffs, "inlining moved something:\n%s" % parity.describe(diffs)


def test_the_code_py_tab_is_now_self_contained(bundle):
    """The half of the email-sim acceptance that #165 could not claim.

    `test_the_self_contained_half_is_not_claimed_yet` in test_fdic_email_sim.py
    asserted the gap; this asserts it is closed.
    """
    name, spec, folder, path = bundle
    workbook = folder / ("%s.xlsm" % spec.workbook)
    if not workbook.exists():
        _run(ISOLATED + [path.name], folder)

    wb = openpyxl.load_workbook(workbook, keep_vba=True)
    try:
        ws = wb["_code_py"]
        source = "\n".join(str(ws.cell(r, 1).value or "")
                           for r in range(1, ws.max_row + 1))
    finally:
        wb.close()

    assert "_install()" in source, "_code_py does not carry the inlined engine"
    assert "_B64" in source
    for line in source.splitlines():
        stripped = line.strip()
        assert not stripped.startswith("from credit_suite"), \
            "_code_py still imports credit_suite at module scope: %s" % stripped
