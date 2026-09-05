"""TEMPLATE_CONTRACT section 9 -- the verification bar, run against the ENGINE.

Ported from `fdic-peer-monitor/tests/test_runner.py` and pointed at
`credit_suite`. Kept as close to the original as possible: these tests are the
bar the monitor shipped against, and rewriting them while migrating would mean
the migration was checked by tests written after the fact.

Runs on Linux/CI with NO Excel and NO network, all in --demo mode. Mirrors the
named tests in BUILD_SPEC_FDIC.md Section 6 (Phases 1-8):

  test_config_parse                     (Phase 1 -- incl. over-capacity refusal)
  test_demo_provider_deterministic      (Phase 2)
  test_fdic_provider_parse_and_cache    (Phase 2 -- captured-envelope fixture)
  test_fdic_provider_backoff            (Phase 2 -- stubbed 429, no network)
  test_lookup_offline_message           (Phase 2)
  test_transforms_and_derived           (Phase 3 -- hardcoded expected values)
  test_reload_headless                  (Phase 4)
  test_watchlist_entity_gates           (Phase 5 -- positive + refusals)
  test_stale_bank_flag                  (Phase 6 -- the merged-bank regression)
  test_peer_flexibility                 (Phase 7 -- the USER REQUIREMENT)
  test_raw_landing_idempotent           (Phase 8)
  test_raw_layout_mismatch_refused      (Phase 8)
  test_vba_protection_keys_roundtrip    (carried)
"""
import json
import os
import re
import shutil
import struct
import sys
from dataclasses import replace
from datetime import date

import openpyxl
import pytest

from credit_suite.sources.fdic import engine_api as R      # noqa: E402
from credit_suite.sources.fdic import runner as FDIC_RUNNER  # noqa: E402
from credit_suite.sources.fdic import series_seed as SEED    # noqa: E402
from credit_suite.sources.fdic import layout as BW           # noqa: E402
from credit_suite.sources.fdic import adapter                # noqa: E402
from credit_suite.sources.fdic import engine_api as assemble_xlsm  # noqa: E402
from credit_suite.engine import vba as V                     # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

ASOF = date(2026, 3, 31)

#: The FDIC history endpoint's shape, with one real merger in it (Capital One
#: absorbed its card bank on 3 October 2022 -- the 670% quarter). The provider
#: asks for this on every prime; a fixture that did not answer would leave the
#: merger record UNKNOWN and quietly stop testing the path.
HISTORY_FIXTURE = json.dumps({
    "meta": {"total": 1},
    "data": [{"data": {"CERT": 33954, "EFFDATE": "2022-10-03T00:00:00",
                       "CHANGECODE": 223,
                       "CHANGECODE_DESC": "Merger -Without Assistance",
                       "ACQ_CERT": 4297, "OUT_CERT": 33954,
                       "OUT_NAME": "Capital One Bank (USA), NA"}}],
}).encode("utf-8")

# v1.1 tab order: LoanBook after Funding_Concentration, _provenance after
# _config (SPEC_COMPETITOR_PACK.md sec 3 / contract sec 12)
TABS = ("Dashboard_AssetQuality", "Dashboard_Capital_Earnings",
        "Dashboard_Funding_Concentration", "Dashboard_LoanBook", "Watchlist",
        "Raw_FDIC", "_config", "_provenance", "_mergers", "_code_py",
        "_code_vba", "_readme")

N_METRICS = 53          # 15 core + 38 v1.1 pack


# --------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------
@pytest.fixture(scope="session")
def xlsm(tmp_path_factory):
    """Build the macro-enabled workbook once for the whole session."""
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    out = str(d / "Bank_Peer_Monitor.xlsm")
    BW.build(base)
    assemble_xlsm.assemble(base, out)
    return out


@pytest.fixture(scope="session")
def populated(xlsm, tmp_path_factory):
    """A fresh copy of the workbook with a --demo run applied (session-scoped:
    the read-only assertions share one copy)."""
    d = tmp_path_factory.mktemp("run")
    p = str(d / "run.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    return p


def _peer(slot=13, cert="12345", name="Test Bank", group="peer", active=True):
    return R.PeerRow(slot=slot, key=cert, name=name, group=group,
                     active=active)


def _edit_peer_row(path, slot, cert=None, name=None, group=None, active=None):
    """Edit a [PEERS] line in a BUILT workbook exactly as a user would (type
    into the _config cells) -- returns after saving."""
    wb = openpyxl.load_workbook(path, keep_vba=path.endswith(".xlsm"))
    ws = wb["_config"]
    in_peers = False
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v == "[PEERS]":
            in_peers = True
            continue
        if in_peers and isinstance(v, str) and v.startswith("["):
            break
        if in_peers and str(v).strip() == str(slot):
            target = row[0].row
            break
    assert target is not None, f"[PEERS] slot {slot} row not found"
    if cert is not None:
        ws.cell(target, 2, cert)
    if name is not None:
        ws.cell(target, 3, name)
    if group is not None:
        ws.cell(target, 4, group)
    if active is not None:
        ws.cell(target, 5, active)
    wb.save(path)
    wb.close()


# --------------------------------------------------------------------------
# Phase 1 -- config parse (incl. [PEERS] + over-capacity refusal)
# --------------------------------------------------------------------------
def test_config_parse():
    cfg = R.parse_config(BW.config_rows())
    # the 53-metric dictionary (15 core + 38 pack), bank-agnostic,
    # entity-keyed placeholder
    assert len(cfg.series) == N_METRICS
    assert {s.id for s in cfg.series} == set(R.METRICS)
    for s in cfg.series:
        assert s.geo_segment == "entity" and s.source_class == "A"
        assert s.watchlist_capable is True
        assert s.transform in ("direct", "derived")
    R.validate_metrics(cfg.series)                     # registry agreement
    # thresholds: every metric banded, directions sane
    assert set(cfg.thresholds) == set(R.METRICS)
    # bands load from the seed (values are calibrated data, not hardcoded here,
    # so a recalibration never breaks this test) -- just check they round-trip
    seed_thr = {t[0]: t for t in SEED.THRESHOLDS}
    assert cfg.thresholds["TEXAS"].watch == pytest.approx(seed_thr["TEXAS"][1])
    assert cfg.thresholds["TEXAS"].alert == pytest.approx(seed_thr["TEXAS"][2])
    assert cfg.thresholds["TEXAS"].watch < cfg.thresholds["TEXAS"].alert  # above-dir
    assert cfg.thresholds["RBC1AAJ"].direction == "below"
    assert cfg.thresholds["LNRESNCR"].direction == "below"
    # [PEERS]: 40 provisioned slots, 12 seeded banks, certs normalized
    assert cfg.entity_slots == 40 and len(cfg.entities) == 40
    banked = [p for p in cfg.entities if p.has_entity]
    assert len(banked) == 12
    assert [p.slot for p in cfg.entities] == list(range(1, 41))
    jpm = next(p for p in banked if p.slot == 1)
    assert jpm.key == "628" and jpm.entity_key == "cert:628"
    assert jpm.active is True and jpm.group == "peer"
    assert next(p for p in banked if p.slot == 9).group == "self"
    for p in cfg.entities[12:]:
        assert not p.has_entity and not p.active        # provisioned headroom
    # settings the runner depends on
    assert cfg.raw_slots == 16
    assert cfg.stale_multiplier == pytest.approx(2.0)
    # expansion peers x metrics -> the s{slot:02d}_{ID} unit ids
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    assert len(admitted) == 12 and not refusals and not excluded
    fspec = R.make_field_spec(jpm, "NCLNLSR")
    assert fspec.id == "s01_NCLNLSR" and fspec.geo_segment == "cert:628"
    unit_ids = {f"s{p.slot:02d}_{s.id}" for p in admitted for s in cfg.series}
    assert len(unit_ids) == 12 * N_METRICS
    assert {"s01_TEXAS", "s12_CRECONR", "s05_RBCRWAJ",
            "s03_NTCRCDQ_BOOK", "s07_UNRLZCAPR"} <= unit_ids
    # capacity is validated, never truncated: a 41st bank refuses with the
    # exact rebuild command (the flexible-peers USER REQUIREMENT)
    cfg.entities.append(_peer(slot=41, cert="99999", name="Overflow Bank"))
    with pytest.raises(R.PeerCapacityError) as ei:
        R.validate_peer_capacity(cfg)
    msg = str(ei.value)
    assert "Overflow Bank" in msg and "--peer-slots" in msg
    assert "make_workbook.py" in msg and "never truncates" in msg
    # duplicate slots refuse too
    cfg2 = R.parse_config(BW.config_rows())
    cfg2.entities.append(_peer(slot=1, cert="55555", name="Dup Bank"))
    with pytest.raises(R.PeerCapacityError, match="listed twice"):
        R.validate_peer_capacity(cfg2)


# --------------------------------------------------------------------------
# Phase 2 -- DemoProvider determinism + live-provider parse/cache (no network)
# --------------------------------------------------------------------------
def test_demo_provider_deterministic():
    cfg = R.parse_config(BW.config_rows())
    peers = {p.slot: p for p in cfg.entities if p.has_entity}
    prov = R.FdicDemoProvider(asof=ASOF, raw_slots=cfg.raw_slots)
    for slot, fname in ((1, "NCLNLSR"), (5, "EQ"),
                        (6, "ASSET"), (12, "RBCRWAJ")):
        spec = R.make_field_spec(peers[slot], fname)
        a = prov.fetch_series(spec)
        b = R.FdicDemoProvider(asof=ASOF,
                               raw_slots=cfg.raw_slots).fetch_series(spec)
        assert [(r.period, r.value) for r in a] == \
               [(r.period, r.value) for r in b]
        assert len(a) == cfg.raw_slots
        assert a[-1].period == "2026-03-31"            # never stale at asof
        assert a[0].geo_segment == peers[slot].entity_key
    # seeding is per (cert, field): the same bank produces the same history
    # in ANY slot (a [PEERS] swap moves the data with the bank)
    p_alt = replace(peers[1], slot=27)
    a = prov.fetch_series(R.make_field_spec(peers[1], "ASSET"))
    b = prov.fetch_series(R.make_field_spec(p_alt, "ASSET"))
    assert [r.value for r in a] == [r.value for r in b]
    # one interior None per series (missing-value path), never the newest
    vals = [r.value for r in prov.fetch_series(
        R.make_field_spec(peers[1], "NCLNLSR"))]
    assert vals[-1] is not None and None in vals[1:-1]
    # whole-series null BRO for one seeded bank (trap F3 shape)
    bro = [r.value for r in prov.fetch_series(
        R.make_field_spec(peers[5], "BRO"))]
    assert all(v is None for v in bro)


# A FIXED fixture mimicking the captured API envelope from
# COVERAGE_RESEARCH_FDIC.md: double-wrapped records (data[i].data), meta.total,
# meta.index.createTimestamp, totals.count; JSON null for BRO; REPDTE yyyymmdd.
def _fin_fixture(total=4):
    def rec(cert, repdte, asset, nclnlsr, bro):
        d = {"CERT": cert, "REPDTE": repdte, "ASSET": asset,
             "NCLNLSR": nclnlsr, "BRO": bro, "EQ": asset // 10,
             "LNATRES": asset // 100, "NCLNLS": asset // 50,
             "RBCRWAJ": None}                     # CBLR-elector null
        for f in R.RAW_FIELDS:
            d.setdefault(f, 1.0)
        return {"data": d, "score": 1.0}
    return json.dumps({
        "meta": {"total": total,
                 "parameters": {"limit": 10000},
                 "index": {"name": "risview",
                           "createTimestamp": "2026-06-15T04:00:00Z"}},
        "data": [
            rec(628, "20260331", 242664, 0.84, None),
            rec(628, "20251231", 240000, 0.80, None),
            rec(3511, "20260331", 380997000, 0.55, 12345),
            rec(3511, "20251231", 379000000, 0.51, 11111),
        ],
        "totals": {"count": total},
    }).encode("utf-8")


ROSTER_FIXTURE = json.dumps({
    "meta": {"total": 2},
    "data": [
        {"data": {"CERT": 628, "NAME": "JPMorgan Chase Bank NA",
                  "FED_RSSD": 852218, "ACTIVE": 1, "BKCLASS": "N",
                  "ENDEFYMD": None}, "score": 1.0},
        {"data": {"CERT": 3511, "NAME": "Wells Fargo Bank NA",
                  "FED_RSSD": 451965, "ACTIVE": 1, "BKCLASS": "N",
                  "ENDEFYMD": None}, "score": 1.0},
    ],
    "totals": {"count": 2},
}).encode("utf-8")


def test_fdic_provider_parse_and_cache(monkeypatch):
    """Captured-envelope fixture: double-wrap parse, null->None, per-series
    slicing off ONE bulk call, per-URL cache, meta.total>limit guard, vintage
    recorded. NO network (the one real call is stubbed out)."""
    calls = []

    def fake_get(self, url):
        calls.append(url)
        if "history" in url:
            return HISTORY_FIXTURE
        return ROSTER_FIXTURE if "institutions" in url else _fin_fixture()

    monkeypatch.setattr(R.FdicProvider, "_http_get", fake_get)
    prov = R.FdicProvider(min_interval=0.0)
    prov.prime(["628", "3511"], ASOF, names={})
    # ONE bulk financials call + one roster call + one merger-history call --
    # never per-series and never per-bank requests
    assert len(calls) == 3
    assert sum("history" in c for c in calls) == 1
    fin_url = calls[0]
    assert "financials" in fin_url and "format=json" in fin_url
    assert "CERT%3A%28628+OR+3511%29" in fin_url          # CERT:(628 OR 3511)
    assert "REPDTE%3A%5B" in fin_url                      # REPDTE:[... TO *]
    assert "sort_by=REPDTE" in fin_url and "sort_order=DESC" in fin_url
    assert "limit=10000" in fin_url and "NCLNLSR" in fin_url
    # data vintage (meta.index.createTimestamp) recorded for the status line
    assert prov.vintage == "2026-06-15T04:00:00Z"
    # per-series slicing: (cert, field) -> ordered rows, REPDTE -> ISO
    p1 = _peer(slot=1, cert="628", name="JPM")
    rows = prov.fetch_series(R.make_field_spec(p1, "NCLNLSR"))
    assert [r.period for r in rows] == ["2025-12-31", "2026-03-31"]
    assert [r.value for r in rows] == [0.80, 0.84]
    assert rows[0].geo_segment == "cert:628" and rows[0].source_class == "A"
    # JSON null -> None, NEVER zero (trap F3): BRO for 628, RBCRWAJ for all
    bro = prov.fetch_series(R.make_field_spec(p1, "BRO"))
    assert [r.value for r in bro] == [None, None]
    rbc = prov.fetch_series(R.make_field_spec(p1, "RBCRWAJ"))
    assert all(r.value is None for r in rbc)
    # $000 dollar fields land as floats untouched (units carried per field)
    asset = prov.fetch_series(
        R.make_field_spec(_peer(slot=2, cert="3511", name="WF"), "ASSET"))
    assert asset[-1].value == pytest.approx(380997000.0)
    assert asset[-1].units == "USD_thousands"
    assert prov.fetch_series(
        R.make_field_spec(p1, "NCLNLSR"))[0].units == "pct"
    # roster landed (merger detection inputs)
    assert prov.roster["628"]["ACTIVE"] == 1
    # cache hit: re-priming the same request makes NO further call
    prov.prime(["628", "3511"], ASOF, names={})
    assert len(calls) == 3
    # meta.total > limit refuses CLEARLY -- never silent truncation
    def fat_get(self, url):
        return _fin_fixture(total=20000)
    monkeypatch.setattr(R.FdicProvider, "_http_get", fat_get)
    prov2 = R.FdicProvider(min_interval=0.0)
    with pytest.raises(RuntimeError, match="refusing silent truncation"):
        prov2.prime(["628"], ASOF)
    # fetch before prime is a programming-error refusal, not empty data
    prov3 = R.FdicProvider(min_interval=0.0)
    with pytest.raises(RuntimeError, match="prime"):
        prov3.fetch_series(R.make_field_spec(p1, "ASSET"))


def test_fdic_provider_backoff(monkeypatch):
    """429 is retried with backoff and then succeeds -- unit-level, stubbed."""
    import urllib.error
    attempts = []

    def flaky_get(self, url):
        attempts.append(url)
        if len(attempts) < 3:
            raise urllib.error.HTTPError(url, 429, "Too Many Requests",
                                         None, None)
        if "history" in url:
            return HISTORY_FIXTURE
        return ROSTER_FIXTURE if "institutions" in url else _fin_fixture()

    sleeps = []
    monkeypatch.setattr(R.FdicProvider, "_http_get", flaky_get)
    monkeypatch.setattr(R.time, "sleep", lambda s: sleeps.append(s))
    prov = R.FdicProvider(min_interval=0.0, max_retries=4)
    prov.prime(["628", "3511"], ASOF)
    assert len(attempts) == 5              # 2 x 429 + financials + roster + history
    assert sleeps                          # backoff actually slept
    rows = prov.fetch_series(
        R.make_field_spec(_peer(slot=1, cert="628", name="JPM"), "ASSET"))
    assert len(rows) == 2
    # a persistent 500 fails with a labeled error, never a stack of retries
    def dead_get(self, url):
        raise urllib.error.HTTPError(url, 503, "down", None, None)
    monkeypatch.setattr(R.FdicProvider, "_http_get", dead_get)
    prov2 = R.FdicProvider(min_interval=0.0, max_retries=1)
    with pytest.raises(RuntimeError, match="HTTP 503"):
        prov2.prime(["628"], ASOF)


def test_lookup_offline_message(capsys):
    """--lookup is live-only: in --demo it prints a CLEAR offline message
    (naming the live endpoint and the keyless run form) and exits nonzero
    without touching any network."""
    rc = R.main(["--lookup", "Frost Bank", "--demo"])
    err = capsys.readouterr().err
    assert rc != 0
    assert "LIVE-ONLY" in err and "--demo" in err
    assert "keyless" in err and "--lookup \"Frost Bank\"" in err
    # and the runner still requires a workbook when not looking up
    rc = R.main(["--demo"])
    err = capsys.readouterr().err
    assert rc == 2 and "--workbook" in err


# --------------------------------------------------------------------------
# Phase 3 -- derived metrics + threshold engine (hardcoded expected values)
# --------------------------------------------------------------------------
def test_transforms_and_derived():
    # TEXAS = NCLNLS/(EQ+LNATRES)*100 on fixed inputs
    assert R.d_texas({"NCLNLS": 50.0, "EQ": 80.0, "LNATRES": 20.0}) \
        == pytest.approx(50.0)
    assert R.d_texas({"NCLNLS": 120.0, "EQ": 80.0, "LNATRES": 20.0}) \
        == pytest.approx(120.0)
    # None-field tolerance: missing input -> None (blank), NEVER 0 (trap F3)
    assert R.d_texas({"NCLNLS": 50.0, "EQ": None, "LNATRES": 20.0}) is None
    assert R.d_texas({"NCLNLS": None, "EQ": 80.0, "LNATRES": 20.0}) is None
    assert R.d_texas({"NCLNLS": 50.0, "EQ": 0.0, "LNATRES": 0.0}) is None
    # BRODEPR with BRO=None -> None; zero denominator -> None
    assert R.d_brodepr({"BRO": None, "DEP": 1000.0}) is None
    assert R.d_brodepr({"BRO": 25.0, "DEP": 250.0}) == pytest.approx(10.0)
    assert R.d_brodepr({"BRO": 25.0, "DEP": 0.0}) is None
    # PD3089R = P3LNLS/LNLSGR*100
    assert R.d_pd3089r({"P3LNLS": 15.0, "LNLSGR": 1000.0}) == pytest.approx(1.5)
    assert R.d_pd3089r({"P3LNLS": 15.0, "LNLSGR": None}) is None
    # LNDEPR = LNLSNET/DEP*100
    assert R.d_lndepr({"LNLSNET": 800.0, "DEP": 1000.0}) == pytest.approx(80.0)
    # CRECONR: any null CRE component blanks the composite (null != 0)
    f = {"LNRECONS": 30.0, "LNRENRES": 80.0, "LNREMULT": 20.0,
         "EQ": 80.0, "LNATRES": 20.0}
    assert R.d_creconr(f) == pytest.approx(130.0)
    assert R.d_creconr({**f, "LNREMULT": None}) is None
    assert R.d_creconr({**f, "EQ": None}) is None
    # direct metrics pass the field through; absent field -> None
    assert R.metric_value("NCLNLSR", {"NCLNLSR": 0.84}) == pytest.approx(0.84)
    assert R.metric_value("RBCRWAJ", {"RBCRWAJ": None}) is None
    assert R.metric_value("ROAQ", {}) is None
    # derived metrics route through the registry identically
    assert R.metric_value("TEXAS", {"NCLNLS": 50.0, "EQ": 80.0,
                                    "LNATRES": 20.0}) == pytest.approx(50.0)
    # threshold engine, direction-aware (matches the Excel band rules)
    above = R.Threshold(watch=50.0, alert=100.0, direction="above")
    assert R.status_for(108.0, above) == "ALERT"
    assert R.status_for(52.0, above) == "WATCH"
    assert R.status_for(10.0, above) == "OK"
    assert R.status_for(None, above) == "OK"
    below = R.Threshold(watch=0.5, alert=0.0, direction="below")   # ROAQ
    assert R.status_for(-0.1, below) == "ALERT"
    assert R.status_for(0.4, below) == "WATCH"
    assert R.status_for(1.2, below) == "OK"
    cov = R.Threshold(watch=100.0, alert=60.0, direction="below")  # LNRESNCR
    assert R.status_for(55.0, cov) == "ALERT"
    assert R.status_for(75.0, cov) == "WATCH"
    assert R.status_for(180.0, cov) == "OK"
    # registry drift refuses (unknown metric / transform mismatch)
    cfg = R.parse_config(BW.config_rows())
    with pytest.raises(R.MetricError, match="not a registered metric"):
        R.validate_metrics([replace(cfg.series[0], id="UNKNOWNX")])
    with pytest.raises(R.MetricError, match="transform"):
        R.validate_metrics([replace(
            next(s for s in cfg.series if s.id == "TEXAS"),
            transform="direct")])


# --------------------------------------------------------------------------
# Phase 4 -- headless reload (Excel proxy) + NO native charts
# --------------------------------------------------------------------------
def test_reload_headless(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    # macro survived the openpyxl round-trip
    assert wb.vba_archive is not None
    # tab set EXACT (contract sec 2)
    assert tuple(wb.sheetnames) == TABS
    # NO native chart objects anywhere (the top "recovered content" trigger, L4)
    for ws in wb.worksheets:
        assert getattr(ws, "_charts", []) == []
    cfg = R.parse_config(BW.config_rows())
    # every (slot, metric) cell carries a formula on its lane tab -- all 40
    # built slot rows, plus the MEDIAN row under them
    for tab, _t, _s, mids in BW.LANES:
        ws = wb[tab]
        for k in range(len(mids)):
            col = 4 + k
            for slot in (1, 12, 40):                     # seeded, last, empty
                v = ws.cell(BW.dash_row(slot), col).value
                assert isinstance(v, str) and v.startswith("="), (tab, slot)
            med = ws.cell(BW.dash_row(40) + 1, col).value
            assert isinstance(med, str) and "MEDIAN(" in med, tab
        # bank identity flows BY FORMULA from the [PEERS] cells
        assert "_config" in str(ws.cell(BW.dash_row(1), 1).value)
    # Watchlist: flag counts, rank, status (with the formula-level REFUSED
    # rendering + --lookup hint) on every slot row
    ws = wb["Watchlist"]
    for slot in (1, 12, 40):
        r = BW.wl_row(slot)
        assert "COUNTIF" in str(ws.cell(r, 5).value)
        assert "RANK" in str(ws.cell(r, 7).value)
        st = str(ws.cell(r, 8).value)
        assert "REFUSED" in st and "--lookup" in st and "OFF" in st
    # threshold cells are NUMERIC (the text-"0.5" lesson: Excel's number>=text
    # comparison is silently FALSE and would kill every band rule)
    cws = wb["_config"]
    in_thr = False
    checked = 0
    for row in cws.iter_rows(min_col=1, max_col=3):
        v = row[0].value
        if v == "[THRESHOLDS]":
            in_thr = True
            continue
        if in_thr and isinstance(v, str) and v.startswith("["):
            break
        if in_thr and v in R.METRICS:
            assert isinstance(row[1].value, (int, float)), v
            assert isinstance(row[2].value, (int, float)), v
            checked += 1
    assert checked == N_METRICS
    # the raw scaffold labels every field column per slot block
    raw = wb["Raw_FDIC"]
    b = R.slot_block(7, cfg.raw_slots)
    assert raw.cell(b.header_row, 1).value == "slot07"
    assert raw.cell(b.label_row, R.field_col("EQ")).value == "EQ"
    assert raw.cell(b.label_row, R.field_col("EEFFR")).value == "EEFFR"
    wb.close()


# --------------------------------------------------------------------------
# Phase 5 -- entity gates: positive admission + every refusal shape
# --------------------------------------------------------------------------
def test_watchlist_entity_gates():
    cfg = R.parse_config(BW.config_rows())

    # POSITIVE: all 12 seeded CERT-keyed banks are admitted; nothing refused
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    assert len(admitted) == 12 and refusals == [] and excluded == []
    assert all(re.match(R.ENTITY_KEY_PATTERN, p.entity_key) for p in admitted)

    # blank cert: refused BY NAME with the --lookup hint (never fetched)
    blank = _peer(slot=13, cert="", name="Mystery Trust Co")
    cfg.entities[12] = blank
    admitted2, refusals2, _ = R.evaluate_peers(cfg)
    assert len(admitted2) == 12                    # the good rows still pass
    assert len(refusals2) == 1
    msg = refusals2[0][1]
    assert 'peer slot 13 "Mystery Trust Co"' in msg
    assert 'cert=""' in msg
    assert "Only CERT-keyed FDIC institutions" in msg
    assert '--lookup "Mystery Trust Co"' in msg
    assert "cert:[0-9]{1,7}" in msg                # the whitelist, named

    # inactive slot: EXCLUDED (blanked), not refused
    cfg3 = R.parse_config(BW.config_rows())
    idx = next(i for i, p in enumerate(cfg3.entities) if p.slot == 2)
    cfg3.entities[idx] = replace(cfg3.entities[idx], active=False)
    admitted3, refusals3, excluded3 = R.evaluate_peers(cfg3)
    assert len(admitted3) == 11 and refusals3 == []
    assert [p.slot for p in excluded3] == [2]

    # defense in depth: a fabricated non-"A" class metric row refuses the
    # run, series-named -- a good entity key cannot carry a bad class
    cfg4 = R.parse_config(BW.config_rows())
    bad = replace(next(s for s in cfg4.series if s.id == "TEXAS"),
                  source_class="C")
    with pytest.raises(R.WatchlistRefused) as ei:
        R.assert_metrics_admissible([bad])
    assert 'series "TEXAS"' in str(ei.value)
    assert 'source_class="C"' in str(ei.value)
    assert "its own review" in str(ei.value)

    # malformed entity keys are refused default-deny
    for cert in ("12345678", "abc", "12-34", "62 8", "cert:628"):
        reasons = R.gate_peer_row(_peer(slot=14, cert=cert, name="Bad Key"))
        assert reasons and "Gate3" in reasons[0], cert
    # and the build-time hard gate backs the runtime gate
    cfg5 = R.parse_config(BW.config_rows())
    cfg5.entities[13] = _peer(slot=14, cert="not-a-cert", name="Aggregate Row")
    with pytest.raises(R.WatchlistRefused, match="Aggregate Row"):
        R.assert_entity_gates(cfg5)


# --------------------------------------------------------------------------
# Phase 6 -- staleness (the merged-bank regression, trap F2)
# --------------------------------------------------------------------------
class FrozenPeerProvider(R.FdicDemoProvider):
    """DemoProvider variant that freezes ONE bank three quarters behind the
    peer-set max -- the exact shape of a merged-away bank (data returned, no
    error, REPDTE just stops advancing)."""

    def __init__(self, frozen_cert, **kw):
        super().__init__(**kw)
        self.frozen_cert = str(frozen_cert)
        y, m = self.asof.year, self.asof.month - 9
        if m < 1:
            m += 12
            y -= 1
        self._old = R.FdicDemoProvider(asof=date(y, m, 28),
                                       raw_slots=self.raw_slots)

    def fetch_series(self, spec, secret=None):
        if spec.key == self.frozen_cert:
            return self._old.fetch_series(spec, secret)
        return super().fetch_series(spec, secret)


def test_stale_bank_flag(xlsm, tmp_path):
    # freeze the demo ALERT-tier bank (cert 6548, slot 5): fresh it carries
    # 11 ALERT flags; frozen it must carry NONE of them into the KPIs.
    p = str(tmp_path / "stale.xlsm")
    shutil.copy(xlsm, p)
    frozen = FrozenPeerProvider("6548", asof=ASOF, raw_slots=16)
    frozen_status = R.run(p, demo=True, asof=ASOF, provider=frozen)
    p2 = str(tmp_path / "fresh.xlsm")
    shutil.copy(xlsm, p2)
    fresh_status = R.run(p2, demo=True, asof=ASOF)

    fb = next(b for b in frozen_status["digest"]["banks"] if b["cert"] == "6548")
    assert fb["stale"] is True
    assert fb["status"] == "STALE"                 # visible in the digest
    assert fb["asof_period"] < frozen_status["digest"]["set_max_repdte"]
    assert any("merger/closure" in n for n in fb["notes"])
    # excluded from the alert KPIs: the bank ALERTs when fresh
    fresh_b = next(b for b in fresh_status["digest"]["banks"]
                   if b["cert"] == "6548")
    assert fresh_b["status"] == "ALERT" and fresh_b["stale"] is False
    assert frozen_status["stale_banks"] == 1 and fresh_status["stale_banks"] == 0
    assert frozen_status["alert_banks"] == fresh_status["alert_banks"] - 1
    assert frozen_status["alert_flags"] == \
        fresh_status["alert_flags"] - fresh_b["alert_count"]
    # digest medians exclude the stale bank (its extreme Texas drops out);
    # the EXCEL median rows still include it -- the documented v1 divergence
    # (spec test 6) recorded in _readme, asserted there below.
    assert frozen_status["digest"]["medians"]["TEXAS"] < \
        fresh_status["digest"]["medians"]["TEXAS"]
    wb = openpyxl.load_workbook(p, keep_vba=True)
    readme = "\n".join(str(c.value) for row in wb["_readme"].iter_rows()
                       for c in row if c.value)
    assert "KNOWN v1 DIVERGENCE" in readme
    assert "MEDIAN" in readme and "STALE" in readme
    # the frozen slot's data IS still landed (audit trail; Excel reads it)
    b5 = R.slot_block(5, 16)
    assert wb["Raw_FDIC"].cell(b5.first_data_row, 2).value is not None
    wb.close()

    # the primitive itself
    assert R.is_stale_bank("2025-06-30", "2026-03-31", 2.0) is True   # 3 qtrs
    assert R.is_stale_bank("2025-12-31", "2026-03-31", 2.0) is False  # 1 qtr
    assert R.is_stale_bank(None, "2026-03-31", 2.0) is True           # valueless
    assert R.is_stale_bank("2026-03-31", None, 2.0) is False          # no baseline


# --------------------------------------------------------------------------
# Phase 7 -- peer flexibility (the USER REQUIREMENT: edit + re-run, NO rebuild)
# --------------------------------------------------------------------------
def test_peer_flexibility(populated, tmp_path):
    p = str(tmp_path / "flex.xlsm")
    shutil.copy(populated, p)                     # BUILT + populated workbook
    # capture the pre-edit dashboard formulas: the edit must move DATA, not
    # formulas (anchors depend only on the slot)
    wb = openpyxl.load_workbook(p, keep_vba=True)
    pre_formula = wb["Dashboard_AssetQuality"].cell(BW.dash_row(13), 4).value
    b13 = R.slot_block(13, 16)
    b2 = R.slot_block(2, 16)
    assert wb["Raw_FDIC"].cell(b13.first_data_row, 2).value is None  # empty slot
    assert wb["Raw_FDIC"].cell(b2.first_data_row, 2).value is not None
    wb.close()

    # the user edit: swap a NEW bank into free slot 13, deactivate slot 2
    _edit_peer_row(p, 13, cert=12345, name="Swapped In Bank", group="peer",
                   active="TRUE")
    _edit_peer_row(p, 2, active="FALSE")
    status = R.run(p, demo=True, asof=ASOF)       # re-run -- NO rebuild

    assert status["banks_landed"] == 12           # 12 seed - 1 off + 1 new
    assert status["banks_excluded"] == 1
    slots = {b["slot"] for b in status["digest"]["banks"]}
    assert 13 in slots and 2 not in slots

    wb = openpyxl.load_workbook(p, keep_vba=True)
    raw = wb["Raw_FDIC"]
    # slot 13's raw block now carries the new unit id + data
    assert raw.cell(b13.header_row, 2).value == "s13 cert:12345"
    assert raw.cell(b13.header_row, 3).value == "Swapped In Bank"
    assert raw.cell(b13.first_data_row, 2).value is not None
    assert raw.cell(b13.first_data_row, 1).value == "2026-03-31"
    # the deactivated slot is BLANKED (stateless clearing), identity included
    assert raw.cell(b2.header_row, 2).value is None
    for c in range(1, 2 + len(R.RAW_FIELDS)):
        assert raw.cell(b2.first_data_row, c).value is None
    # anchors/formulas untouched: same formula text as before the edit
    assert wb["Dashboard_AssetQuality"].cell(BW.dash_row(13), 4).value \
        == pre_formula
    wb.close()


# --------------------------------------------------------------------------
# Phase 8 -- idempotence + layout guard
# --------------------------------------------------------------------------
def test_raw_landing_idempotent(xlsm, tmp_path):
    def landed(path):
        wb = openpyxl.load_workbook(path, keep_vba=True)
        ws = wb["Raw_FDIC"]
        vals = [(c.row, c.column, c.value) for row in ws.iter_rows()
                for c in row if c.value is not None]
        wb.close()
        return vals

    # TRUE idempotence: a second run on the SAME already-populated workbook
    # lands identical values (slot clearing leaves no stale tail).
    p1 = str(tmp_path / "r1.xlsm")
    shutil.copy(xlsm, p1)
    R.run(p1, demo=True, asof=ASOF)
    first = landed(p1)
    R.run(p1, demo=True, asof=ASOF)
    assert landed(p1) == first
    # ... and determinism across fresh copies.
    p2 = str(tmp_path / "r2.xlsm")
    shutil.copy(xlsm, p2)
    R.run(p2, demo=True, asof=ASOF)
    assert landed(p2) == first
    # newest-first: slot 1's first data row is the asof quarter
    wb = openpyxl.load_workbook(p1, keep_vba=True)
    b1 = R.slot_block(1, 16)
    assert wb["Raw_FDIC"].cell(b1.first_data_row, 1).value == "2026-03-31"
    assert wb["Raw_FDIC"].cell(b1.first_data_row, 2).value is not None
    wb.close()


def test_raw_layout_mismatch_refused(xlsm, tmp_path):
    """Editing raw_slots after build must be REFUSED (every formula is
    anchored to the built layout), not silently mis-written."""
    p = str(tmp_path / "mismatch.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)                # populate at raw_slots=16
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == "raw_slots":
            target = row[0].row
            break
    assert target is not None
    ws.cell(target, 2, 8)
    wb.save(p)
    wb.close()
    with pytest.raises(RuntimeError, match="Raw layout mismatch"):
        R.run(p, demo=True, asof=ASOF)


# --------------------------------------------------------------------------
# Seam extras -- Class C stub (contract) + VBA project keys
# --------------------------------------------------------------------------
def test_class_c_stub():
    """The seam-contract stub. NOTE: class C is REFUSED by this template's
    gate 2 -- the stub only proves the fetch_series seam shape."""
    spec = R.FieldSpec(slot=1, key="628", fname="ASSET", id="s01_ASSET",
                       geo_segment="cert:628", units="USD_thousands",
                       source_class="C")
    prov = R.ClassCStubProvider(secret_env="FDIC_CLASS_C_SECRET")
    with pytest.raises(PermissionError) as ei:
        prov.fetch_series(spec, secret=None)
    assert "401" in str(ei.value)
    rows = prov.fetch_series(spec, secret="dummy-client-credentials")
    assert isinstance(rows, list) and len(rows) == 1
    assert isinstance(rows[0], R.NormalizedRow)
    assert rows[0].value is None                 # the stub never fabricates data
    # and even a well-formed Class C metric row is refused by gate 2
    cfg = R.parse_config(BW.config_rows())
    licensed = replace(cfg.series[0], source_class="C")
    assert any("Gate2" in r for r in R.gate_metric_row(licensed))


# --------------------------------------------------------------------------
# v1.1 competitor pack (SPEC_COMPETITOR_PACK.md sec 4)
# --------------------------------------------------------------------------
def test_pack_fields_in_bulk_request(monkeypatch):
    """Every pack field rides the ONE bulk /financials request's fields=
    list, and the list stays far under the 250-field cap."""
    from urllib.parse import urlparse, parse_qs
    calls = []

    def fake_get(self, url):
        calls.append(url)
        if "history" in url:
            return HISTORY_FIXTURE
        return ROSTER_FIXTURE if "institutions" in url else _fin_fixture()

    monkeypatch.setattr(R.FdicProvider, "_http_get", fake_get)
    prov = R.FdicProvider(min_interval=0.0)
    prov.prime(["628", "3511"], ASOF)
    assert len(calls) == 3                        # STILL one bulk + roster + history
    qs = parse_qs(urlparse(calls[0]).query)
    fields = qs["fields"][0].split(",")
    assert fields[:2] == ["CERT", "REPDTE"]
    assert len(fields) == 2 + len(R.RAW_FIELDS)
    assert len(fields) < 250                      # the FDIC fields= cap
    # spot the pack families: consumer twins, dollar triples, truncated NCO
    # names, balances, SVB fields
    for f in ("P3CRCD", "NAAUTO", "P3RELOC", "LNRELOC", "P3CONOTH", "LNCRCD",
              "NTCRCDQ", "NTCONOTQ", "NTRECONQ", "NTRENREQ", "NTREMULQ",
              "NTCIQ", "P3RECONS", "NAREMULT", "LNCI", "DEPUNINS", "SCHA",
              "SCHF", "SCAA", "SCAF", "OTHBFHLB"):
        assert f in fields, f
    # the YTD NCO fields are NEVER requested (trap F1) and no un-truncated
    # 9-char name slipped in
    # the YTD NCO fields, the FDIC's over-total-assets ratio twins (#268),
    # and any metric id that is not itself a landed field
    for banned in ("NTCRCD", "NTAUTO", "NTCI", "NTRECONS", "NTRECONSQ",
                   "P3CRCDR", "NAAUTOR", "P3RELOCR", "NACIR",
                   "P3CONOTH_BOOK", "P3CRCD_BOOK", "EQCCOMPI"):
        assert banned not in fields, banned
    # every [SERIES] metric's inputs are landed fields (audit trail intact)
    cfg = R.parse_config(BW.config_rows())
    R.validate_metrics(cfg.series)


def test_consumer_track_rates():
    """Hardcoded expectations for the consumer-track transforms: every class
    rate is num/den*100 over its OWN BOOK, since #268 -- the FDIC's ratio
    twins divide by total assets and are no longer landed. Quarterly NCO
    rates are the Q dollar flow ANNUALIZED x4 over balances. None-tolerant
    throughout."""
    # the card triple, over the card book
    assert R.metric_value("P3CRCD_BOOK", {"P3CRCD": 25.1, "LNCRCD": 1000.0}) \
        == pytest.approx(2.51)
    assert R.metric_value("NARELOC_BOOK", {"NARELOC": 7.0, "LNRELOC": 1000.0}) \
        == pytest.approx(0.7)
    # no book, no rate -- and never a zero (trap F3)
    assert R.metric_value("P3CRCD_BOOK", {"P3CRCD": 25.1, "LNCRCD": 0}) is None
    assert R.metric_value("P9AUTO_BOOK", {}) is None

    # card NCOq: 10/1000 * 400 = 4.0 (annualized, NTLNLSQR convention)
    assert R.metric_value("NTCRCDQ_BOOK", {"NTCRCDQ": 10.0, "LNCRCD": 1000.0}) \
        == pytest.approx(4.0)
    assert R.metric_value("NTAUTOQ_BOOK", {"NTAUTOQ": 2.5, "LNAUTO": 1000.0}) \
        == pytest.approx(1.0)
    # computed PD rates: dollar triple over balance x100
    assert R.metric_value("P3CONOTH_BOOK",
                          {"P3CONOTH": 15.0, "LNCONOTH": 1000.0}) \
        == pytest.approx(1.5)
    assert R.metric_value("NACONOTH_BOOK",
                          {"NACONOTH": 8.0, "LNCONOTH": 400.0}) \
        == pytest.approx(2.0)
    # commercial floor: truncated-name NCO flows annualize identically
    assert R.metric_value("NTRECONQ_BOOK",
                          {"NTRECONQ": 5.0, "LNRECONS": 1000.0}) \
        == pytest.approx(2.0)
    assert R.metric_value("NARENRES_BOOK",
                          {"NARENRES": 12.0, "LNRENRES": 600.0}) \
        == pytest.approx(2.0)
    # None-tolerance: null flow, null balance, zero balance -> blank never 0
    assert R.metric_value("NTCRCDQ_BOOK", {"NTCRCDQ": None, "LNCRCD": 1000.0}) is None
    assert R.metric_value("NTCRCDQ_BOOK", {"NTCRCDQ": 10.0, "LNCRCD": None}) is None
    assert R.metric_value("NTCRCDQ_BOOK", {"NTCRCDQ": 10.0, "LNCRCD": 0.0}) is None
    # Excel side is generated from the SAME declarative table (trap F5)
    fx = BW.metric_formula("NTCRCDQ_BOOK", 1, 16)
    assert "*400))" in fx
    n_ref = BW._fref(1, "NTCRCDQ", 16)
    d_ref = BW._fref(1, "LNCRCD", 16)
    assert fx == (f"=IF(OR({n_ref}=\"\",{d_ref}=\"\"),\"\","
                  f"IF({d_ref}=0,\"\",{n_ref}/{d_ref}*400))")
    # demo determinism: the stress bank trips the consumer bands (spec sec 3)
    prov = R.FdicDemoProvider(asof=ASOF, raw_slots=16)
    f6548 = prov._profile("6548")[-1][1]
    assert R.metric_value("NTCRCDQ_BOOK", f6548) > 4.0        # card NCO ALERT
    assert R.metric_value("P3AUTO_BOOK", f6548) > 4.0     # auto PD ALERT
    assert R.metric_value("NARERES_BOOK", f6548) > 2.0    # resi NA ALERT


def test_svb_derived_metrics():
    """Hardcoded SVB-pack expectations, incl. the binding null shape:
    DEPUNINS=None -> BLANK, never zero."""
    # uninsured share
    assert R.metric_value("UNINSDEPR", {"DEPUNINS": 400.0, "DEP": 1000.0}) \
        == pytest.approx(40.0)
    v = R.metric_value("UNINSDEPR", {"DEPUNINS": None, "DEP": 1000.0})
    assert v is None and v != 0                   # blank, NOT zero
    assert R.metric_value("UNINSDEPR", {"DEPUNINS": 400.0, "DEP": 0.0}) is None
    # unrealized loss / capital cushion: ((1000-900)+(2000-1850))/(400+100)
    f = {"SCHA": 1000.0, "SCHF": 900.0, "SCAA": 2000.0, "SCAF": 1850.0,
         "EQ": 400.0, "LNATRES": 100.0}
    assert R.metric_value("UNRLZCAPR", f) == pytest.approx(50.0)
    # unrealized GAIN runs negative (fair above amortized), never clipped
    assert R.metric_value("UNRLZCAPR", {**f, "SCHF": 1100.0, "SCAF": 2100.0}) \
        == pytest.approx(-40.0)
    # any null leg blanks the composite (null != 0)
    for leg in ("SCHA", "SCHF", "SCAA", "SCAF", "EQ", "LNATRES"):
        assert R.metric_value("UNRLZCAPR", {**f, leg: None}) is None
    assert R.metric_value("UNRLZCAPR", {**f, "EQ": 0.0, "LNATRES": 0.0}) is None
    # FHLB reliance
    assert R.metric_value("FHLBASSR", {"OTHBFHLB": 30.0, "ASSET": 300.0}) \
        == pytest.approx(10.0)
    assert R.metric_value("FHLBASSR", {"OTHBFHLB": None, "ASSET": 300.0}) is None
    # demo shape: cert 7213 reports DEPUNINS null for the WHOLE series ->
    # blank + a digest note; the stress banks trip the SVB bands
    prov = R.FdicDemoProvider(asof=ASOF, raw_slots=16)
    f7213 = prov._profile("7213")[-1][1]
    assert f7213["DEPUNINS"] is None
    assert R.metric_value("UNINSDEPR", f7213) is None
    f6548 = prov._profile("6548")[-1][1]
    assert R.metric_value("UNINSDEPR", f6548) > 60.0
    assert R.metric_value("UNRLZCAPR", f6548) > 50.0
    assert R.metric_value("FHLBASSR", f6548) > 20.0
    f6384 = prov._profile("6384")[-1][1]
    assert 60.0 < R.metric_value("UNINSDEPR", f6384) < 75.0   # calibrated WATCH band
    # Excel parity: the UNRLZCAPR formula computes BOTH legs identically
    fx = BW.metric_formula("UNRLZCAPR", 3, 16)
    ha, hf = BW._fref(3, "SCHA", 16), BW._fref(3, "SCHF", 16)
    aa, af = BW._fref(3, "SCAA", 16), BW._fref(3, "SCAF", 16)
    e, a = BW._fref(3, "EQ", 16), BW._fref(3, "LNATRES", 16)
    assert f"(({ha}-{hf})+({aa}-{af}))/({e}+{a})*100" in fx
    assert fx.startswith("=IF(OR(") and f"IF({e}+{a}=0" in fx


def test_loanbook_tab(populated):
    """Reload the populated workbook: both LoanBook bands present (consumer
    caption + the spec's commercial honesty subtitle), slot rows carry
    formulas, PEER MEDIAN rows per band, statuses flow into the Watchlist
    helpers, and the digest statuses match the workbook thresholds."""
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    assert tuple(wb.sheetnames) == TABS           # inserted in ORDER
    ws = wb["Dashboard_LoanBook"]
    # band captions: consumer track + the commercial-floor honesty subtitle
    cap1 = str(ws.cell(6, 1).value)
    assert "CONSUMER TRACK" in cap1 and "DPD-formula-driven" in cap1
    hdr2 = BW.lb_comm_hdr(40)
    cap2 = str(ws.cell(hdr2 - 1, 1).value)
    assert "COMMERCIAL FLOOR" in cap2
    assert "Public Call-Report proxy" in cap2
    assert "criticized/classified view via the EDGAR tracker" in cap2
    # headers label every class column in both bands
    assert ws.cell(BW.DASH_HDR, 4).value == "Card 30-89"
    assert ws.cell(BW.DASH_HDR, 3 + len(BW.LB_CONSUMER)).value == "HELOC NA"
    assert ws.cell(hdr2, 4).value == "Constr 30-89"
    assert ws.cell(hdr2, 3 + len(BW.LB_COMMERCIAL)).value == "C&I NCOq"
    # every (slot, metric) cell is a formula in BOTH bands; median row under
    for band_hdr, mids in ((BW.DASH_HDR, BW.LB_CONSUMER),
                           (hdr2, BW.LB_COMMERCIAL)):
        for k in range(len(mids)):
            for slot in (1, 12, 40):
                v = ws.cell(band_hdr + slot, 4 + k).value
                assert isinstance(v, str) and v.startswith("="), (band_hdr, slot)
            med = ws.cell(band_hdr + 41, 4 + k).value
            assert isinstance(med, str) and "MEDIAN(" in med
        assert "PEER MEDIAN" in str(ws.cell(band_hdr + 41, 1).value)
        # identity flows by formula from [PEERS]
        assert "_config" in str(ws.cell(band_hdr + 1, 1).value)
    # NCO columns annualize x4; twin columns are blank-guarded directs
    assert "*400))" in str(ws.cell(BW.DASH_HDR + 1, 7).value)   # NTCRCDQ_BOOK
    # Watchlist helpers cover all 53 metrics and point pack metrics at the
    # LoanBook tab (commercial band rows offset below the consumer band)
    wl = wb["Watchlist"]
    cfg = R.parse_config(BW.config_rows())
    mids = [s.id for s in cfg.series]
    for k, mid in enumerate(mids):
        assert wl.cell(BW.WL_HDR, BW.WL_HELPER_COL0 + k).value == mid
    k_cons = mids.index("NTCRCDQ_BOOK")
    h = str(wl.cell(BW.wl_row(1), BW.WL_HELPER_COL0 + k_cons).value)
    assert "Dashboard_LoanBook!" in h and "ALERT" in h
    k_comm = mids.index("NACI_BOOK")
    h2 = str(wl.cell(BW.wl_row(1), BW.WL_HELPER_COL0 + k_comm).value)
    tab, col, first = BW.metric_home("NACI_BOOK", 40)
    from openpyxl.utils import get_column_letter
    assert f"Dashboard_LoanBook!{get_column_letter(col)}{first}" in h2
    assert first == hdr2 + 1                     # the commercial band offset
    wb.close()
    # digest statuses agree with the seeded thresholds (values -> statuses)
    status = R.run(populated, demo=True, asof=ASOF)
    b5 = next(b for b in status["digest"]["banks"] if b["cert"] == "6548")
    assert b5["metrics"]["NTCRCDQ_BOOK"]["status"] == "ALERT"
    assert b5["metrics"]["NARENRES_BOOK"]["status"] == "ALERT"
    b6 = next(b for b in status["digest"]["banks"] if b["cert"] == "6384")
    assert b6["metrics"]["NTCRCDQ_BOOK"]["status"] == "WATCH"
    b4 = next(b for b in status["digest"]["banks"] if b["cert"] == "7213")
    assert b4["metrics"]["UNINSDEPR"]["value"] is None
    assert b4["metrics"]["UNINSDEPR"]["status"] == ""     # blank, not OK/0
    assert any("DEPUNINS" in n for n in b4["notes"])      # the blank+note
    # per-band medians exist in the digest too (peer-median per column)
    assert status["digest"]["medians"]["NTCRCDQ_BOOK"] is not None
    assert status["digest"]["medians"]["NACI_BOOK"] is not None


def test_provenance_tab(populated):
    """Contract sec 12: the _provenance tab renders the tie-out map -- header
    block with the citations of record + a WELL-FORMED facsimile URL pattern,
    and one row per landed field AND per metric (every pack metric
    included)."""
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    assert wb.sheetnames.index("_provenance") == \
        wb.sheetnames.index("_config") + 1
    ws = wb["_provenance"]
    text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                     if c.value is not None)
    # citations of record + the two-step tie-out + the URL patterns
    assert "CITATIONS OF RECORD" in text
    assert "Fed MDRM dictionary" in text
    assert "TWO-STEP TIE-OUT" in text
    m = re.search(r"https://cdr\.ffiec\.gov/Public/ViewFacsimileDirect\.aspx"
                  r"\?ds=call&idType=fdiccert&id=\{CERT\}"
                  r"&date=\{MMDDYYYY\}", text)
    assert m, "facsimile URL pattern missing or malformed"
    assert "banks.data.fdic.gov" in text
    # structured rows: field | schedule | line/caption | MDRM | flag | notes
    rows = R.read_provenance_rows(wb)
    wb.close()
    cfg = R.parse_config(BW.config_rows())
    missing = [s.id for s in cfg.series if s.id not in rows]
    assert not missing, f"metrics without a provenance row: {missing}"
    for f in R.RAW_FIELDS:
        assert f in rows, f"landed field without a provenance row: {f}"
    # spot-verify content against PROVENANCE_MAP_FDIC.md
    assert rows["DEPUNINS"]["mdrm"] == "RCON5597"
    assert rows["DEPUNINS"]["schedule"] == "RC-O Mem 2"
    assert rows["SCHA"]["mdrm"] == "RCON1754 (RCFD1754 031)"
    assert rows["NTRECONQ"]["flag"] == "[V]"
    assert "NTRECONSQ" in rows["NTRECONQ"]["notes"]       # truncation quirk
    # honesty flags carried: unmapped MDRMs say so instead of inventing
    assert rows["NTRERESQ"]["flag"] == "[~]"
    assert "not in tie-out map" in rows["NTRERESQ"]["mdrm"]
    assert rows["UNRLZCAPR"]["mdrm"] == \
        "((1754-1771)+(1772-1773)) / (3210+3123)"
    assert "RCOA7204" in rows["RBC1AAJ"]["mdrm"]


def _tieout_blocks(out):
    """Split the tie-out into one block per metric.

    The output became a block per metric on 4 Sep 2026 -- a single dense line
    per metric was correct and unreadable, and the columns truncated mid-word.
    This test moved with it rather than being loosened: every metric must still
    carry its value AND its MDRM code, they must simply now be found in the
    same block rather than on the same line.
    """
    blocks, current, name = {}, [], None
    for line in out.splitlines():
        stripped = line.strip()
        head = re.match(r"^([A-Z0-9_]+) = (-?[\d,]+\.\d\d|\(blank\))$", stripped)
        if head:
            if name:
                blocks[name] = current
            name, current = head.group(1), [stripped]
        elif name:
            if stripped:
                current.append(stripped)
            elif current:
                blocks[name] = current
                name, current = None, []
    if name:
        blocks[name] = current
    return blocks


def test_tieout_mode(populated, capsys):
    """`--tieout CERT [REPDTE]` in demo mode: clearly labeled demo values, a
    block per metric carrying value + MDRM + what it means in plain English,
    and the facsimile + BankFind URLs for the bank-quarter."""
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "6548"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "DEMO VALUES" in out                   # clearly labeled
    assert ("https://cdr.ffiec.gov/Public/ViewFacsimileDirect.aspx?ds=call"
            "&idType=fdiccert&id=6548&date=03312026") in out
    assert R.bankfind_url("6548") in out
    cfg = R.parse_config(BW.config_rows())
    blocks = _tieout_blocks(out)
    for spec in cfg.series:                       # value + MDRM per metric
        block = blocks.get(spec.id)
        assert block is not None, spec.id
        assert re.search(r"(-?[\d,]+\.\d\d|\(blank\))", block[0]), spec.id
        assert any(l.startswith("Checked") for l in block), spec.id

    # THE CODES SURVIVE THE PLAIN ENGLISH. The whole risk of rewriting this for
    # readability was quietly dropping the identifiers that make a number
    # checkable, so that is asserted rather than assumed.
    prov = R.FdicDemoProvider(asof=ASOF, raw_slots=16)
    fields = prov._profile("6548")[-1][1]
    texas = blocks["TEXAS"]
    assert f"{R.metric_value('TEXAS', fields):,.2f}" in texas[0]
    assert any("(1407+1403) / (3210+3123)" in l for l in texas)
    assert any("RCON5597" in l for l in blocks["UNINSDEPR"])

    # ... and the plain English is actually there, for every metric.
    from credit_suite.sources.fdic import plain
    assert plain.describe("NCLNLSR"), "headline metric lost its description"
    undescribed = [s.id for s in cfg.series if not plain.describe(s.id)]
    assert not undescribed, (
        "no plain-English description for %s -- a reader who does not know "
        "MDRM codes cannot act on these rows" % undescribed)
    for term, _meaning in plain.GLOSSARY[:3]:
        assert term in out, "glossary term %r missing from the tie-out" % term
    # explicit REPDTE selects the quarter; a bad one refuses clearly
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "6548",
                 "2025-12-31"])
    out2 = capsys.readouterr().out
    assert rc == 0 and "date=12312025" in out2
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "6548",
                 "1999-12-31"])
    err = capsys.readouterr().err
    assert rc == 1 and "not in the fetched window" in err
    # gate: a malformed CERT is refused with the --lookup hint
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "not-a-cert"])
    err = capsys.readouterr().err
    assert rc == 2 and "--lookup" in err


def test_vba_protection_keys_roundtrip():
    """The PROJECT stream's CMG/DPB/GC must decrypt per [MS-OVBA] 2.4.3 to
    protection=none / password=none / visibility=visible -- exactly the way
    the VBE will read them (a malformed key triggers Excel's 'invalid key
    DPB' prompt)."""
    proj = V.build_project_stream(
        "X", [V.Module("X", "Sub A()\nEnd Sub")]).decode("latin-1")
    keys = {}
    for line in proj.splitlines():
        for k in ("CMG", "DPB", "GC"):
            if line.startswith(f'{k}="'):
                keys[k] = line.split('"')[1]
    assert V.ovba_decrypt(keys["CMG"]) == struct.pack("<I", 0)   # unprotected
    assert V.ovba_decrypt(keys["DPB"]) == b"\x00"               # no password
    assert V.ovba_decrypt(keys["GC"]) == b"\xff"                # visible
    # _VBA_PROJECT is the full 7-byte header ([MS-OVBA] 2.3.4.1)
    assert len(V.build_vba_project_stream()) == 7
    # the default VBA + Excel libraries are referenced in the dir stream
    d = V.build_dir_stream("X", [V.Module("X", "Sub A()\nEnd Sub")])
    assert b"Visual Basic For Applications" in d
    assert b"Microsoft Excel" in d
