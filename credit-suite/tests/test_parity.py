"""The parity safety net (issue #164): does the golden harness actually catch drift?

Two kinds of test live here.

*Harness tests* build a tiny synthetic workbook, plant a change in it, and
assert the diff names that change. They are fast and they are what the
mutation check in ``tools/mutation_check.py`` is pointed at.

*Baseline tests* assert the committed goldens still describe the real shipped
workbooks, and that the demo goldens are not vacuous -- a golden full of blanks
would "pass" every future parity check while proving nothing.
"""

from __future__ import annotations

import json

import openpyxl
import pytest

from credit_suite import parity

ASOF_FREE = "the harness must not depend on the clock"


# --------------------------------------------------------------------------
# a tiny workbook whose status is genuinely formula-driven
# --------------------------------------------------------------------------

def write_workbook(path, trigger: float = 9.0, formula: str | None = None):
    """A 1-sheet workbook: an input, a status formula over it, and a count.

    The status is computed, never stored -- which is the whole point. A harness
    that read cells and stopped would record ``=IF(...)`` here and would be
    blind to the status moving from ALERT to OK.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    ws["A1"] = "Series"
    ws["A2"] = "DRCCLACBS"
    ws["B1"] = "Value"
    ws["B2"] = trigger
    ws["C1"] = "Status"
    ws["C2"] = formula or '=IF(B2>5,"ALERT","OK")'
    ws["D1"] = "Alerts"
    ws["D2"] = '=COUNTIF(C2:C2,"ALERT")'

    second = wb.create_sheet("_config")
    second["A1"] = "[THRESHOLDS]"
    second["B1"] = 5.0
    wb.save(path)
    return path


@pytest.fixture
def base_snapshot(tmp_path):
    return parity.snapshot_workbook(write_workbook(tmp_path / "base.xlsx"),
                                    source="base.xlsx")


# --------------------------------------------------------------------------
# capture: deterministic, and status-aware
# --------------------------------------------------------------------------

def test_snapshot_of_an_unchanged_workbook_is_byte_identical_twice(tmp_path):
    path = write_workbook(tmp_path / "twice.xlsx")
    first = parity.dumps(parity.snapshot_workbook(path, source="twice.xlsx"))
    second = parity.dumps(parity.snapshot_workbook(path, source="twice.xlsx"))
    assert first == second
    assert "202" not in first.split('"cells"')[0].replace('"twice.xlsx"', ""), \
        "a golden header must carry no date -- " + ASOF_FREE


def test_cells_are_ordered_by_sheet_then_row_then_column(tmp_path):
    snapshot = parity.snapshot_workbook(write_workbook(tmp_path / "order.xlsx"),
                                        source="order.xlsx")
    keys = list(snapshot["cells"])
    assert keys == ["Watchlist!A1", "Watchlist!B1", "Watchlist!C1", "Watchlist!D1",
                    "Watchlist!A2", "Watchlist!B2", "Watchlist!C2", "Watchlist!D2",
                    "_config!A1", "_config!B1"]


def test_snapshot_stores_the_computed_status_not_the_formula_text(base_snapshot):
    # B2 = 9 > 5, so the status resolves to ALERT and the count to 1.
    assert base_snapshot["cells"]["Watchlist!C2"] == ["ALERT", '=IF(B2>5,"ALERT","OK")']
    assert base_snapshot["cells"]["Watchlist!D2"][0] == 1
    # ...and the literal alongside it is stored as a literal.
    assert base_snapshot["cells"]["Watchlist!B2"] == [9.0]


def test_a_formula_the_engine_cannot_run_is_still_pinned_by_its_source(tmp_path):
    """HYPERLINK is not implemented by the recalc engine (see UNSUPPORTED_FUNCTIONS).

    Its value is unpinnable, so nothing may be claimed for it -- but the formula
    text is compared regardless, so a provenance URL that changed is still
    caught. That is the property this guarantees.
    """
    path = tmp_path / "link.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Raw_FRED"
    wb["Raw_FRED"]["A1"] = '=HYPERLINK("https://fred.stlouisfed.org/series/X","FRED: X")'
    wb.save(path)
    before = parity.snapshot_workbook(path, source="link.xlsx")
    assert list(parity.unevaluated(before)) == ["Raw_FRED!A1"]

    wb["Raw_FRED"]["A1"] = '=HYPERLINK("https://example.invalid/series/X","FRED: X")'
    wb.save(path)
    diffs = parity.diff_snapshots(before,
                                  parity.snapshot_workbook(path, source="link.xlsx"))
    assert [(d.kind, d.key) for d in diffs] == [("formula", "Raw_FRED!A1")]


def test_float_noise_normalises_away_but_a_real_move_does_not():
    # 12 significant digits: the recalc engine and openpyxl must agree on a
    # value, but a percentage that actually moved must not be rounded away.
    assert parity.normalise(0.87060000000001) == parity.normalise(0.8706)
    assert parity.normalise(0.8706) != parity.normalise(0.8707)
    assert parity.normalise(1234567890.12345) == parity.normalise(1234567890.123451)


# --------------------------------------------------------------------------
# diff: a planted change is caught and named
# --------------------------------------------------------------------------

def test_clean_rebuild_shows_no_differences(base_snapshot, tmp_path):
    again = parity.snapshot_workbook(write_workbook(tmp_path / "again.xlsx"),
                                     source="base.xlsx")
    assert parity.diff_snapshots(base_snapshot, again) == []


def test_a_planted_value_change_is_detected_and_named(base_snapshot, tmp_path):
    path = write_workbook(tmp_path / "moved.xlsx")
    wb = openpyxl.load_workbook(path)
    wb["Watchlist"]["A2"] = "CORCCACBS"     # a label moved
    wb.save(path)

    diffs = parity.diff_snapshots(
        base_snapshot, parity.snapshot_workbook(path, source="base.xlsx"))

    assert [(d.kind, d.key) for d in diffs] == [("value", "Watchlist!A2")]
    assert diffs[0].expected == "DRCCLACBS"
    assert diffs[0].actual == "CORCCACBS"
    assert "Watchlist!A2" in parity.describe(diffs)


def test_a_planted_status_change_is_detected_and_named(base_snapshot, tmp_path):
    # Move only the INPUT. Every formula's text is untouched, so a harness that
    # snapshotted cells without recalculating would see nothing wrong on C2/D2
    # -- and a status headed for KeyBank would move silently.
    path = write_workbook(tmp_path / "flipped.xlsx", trigger=1.0)
    diffs = parity.diff_snapshots(
        base_snapshot, parity.snapshot_workbook(path, source="base.xlsx"))

    by_key = {d.key: d for d in diffs}
    assert set(by_key) == {"Watchlist!B2", "Watchlist!C2", "Watchlist!D2"}
    assert (by_key["Watchlist!C2"].expected, by_key["Watchlist!C2"].actual) == \
        ("ALERT", "OK")
    assert (by_key["Watchlist!D2"].expected, by_key["Watchlist!D2"].actual) == (1, 0)
    assert all(d.kind == "value" for d in diffs), "no formula text changed"


def test_a_rewritten_formula_is_detected_even_when_the_value_holds(base_snapshot,
                                                                  tmp_path):
    # Same answer (ALERT), different rule. Parity cares about both.
    path = write_workbook(tmp_path / "rewritten.xlsx",
                          formula='=IF(B2>=5,"ALERT","OK")')
    diffs = parity.diff_snapshots(
        base_snapshot, parity.snapshot_workbook(path, source="base.xlsx"))

    assert [(d.kind, d.key) for d in diffs] == [("formula", "Watchlist!C2")]
    assert diffs[0].actual == '=IF(B2>=5,"ALERT","OK")'


def test_added_and_removed_cells_are_named(base_snapshot, tmp_path):
    path = write_workbook(tmp_path / "shifted.xlsx")
    wb = openpyxl.load_workbook(path)
    wb["Watchlist"]["A2"].value = None      # L7: blanking needs .value
    wb["Watchlist"]["A3"] = "NEW"
    wb.save(path)

    diffs = parity.diff_snapshots(
        base_snapshot, parity.snapshot_workbook(path, source="base.xlsx"))
    assert [(d.kind, d.key) for d in diffs] == [
        ("cell_removed", "Watchlist!A2"), ("cell_added", "Watchlist!A3")]


def test_a_dropped_tab_is_named(base_snapshot, tmp_path):
    path = write_workbook(tmp_path / "notab.xlsx")
    wb = openpyxl.load_workbook(path)
    del wb["_config"]
    wb.save(path)

    diffs = parity.diff_snapshots(
        base_snapshot, parity.snapshot_workbook(path, source="base.xlsx"))
    assert ("sheet_removed", "_config") in [(d.kind, d.key) for d in diffs]


def test_ignore_forgives_the_named_cell_and_nothing_else(base_snapshot, tmp_path):
    path = write_workbook(tmp_path / "ignored.xlsx")
    wb = openpyxl.load_workbook(path)
    wb["Watchlist"]["A2"] = "CORCCACBS"
    wb["_config"]["A1"] = "[SETTINGS]"
    wb.save(path)
    current = parity.snapshot_workbook(path, source="base.xlsx")

    assert [d.key for d in parity.diff_snapshots(base_snapshot, current)] == \
        ["Watchlist!A2", "_config!A1"]
    assert [d.key for d in parity.diff_snapshots(base_snapshot, current,
                                                 ignore=["Watchlist!A2"])] == \
        ["_config!A1"]
    assert parity.diff_snapshots(base_snapshot, current,
                                 ignore=["Watchlist!*", "_config!*"]) == []


def test_differences_are_reported_in_workbook_reading_order(tmp_path):
    """Sheet order, then row, then column -- not alphabetical.

    ``Raw`` comes before ``_config`` in the workbook but after it alphabetically,
    and row 9 comes before row 10 in the workbook but after it as text. A diff a
    reviewer reads top-to-bottom has to follow the workbook.
    """
    def build(path, marker):
        wb = openpyxl.Workbook()
        first = wb.active
        first.title = "Raw"
        first["A9"] = "nine" + marker
        first["A10"] = "ten" + marker
        second = wb.create_sheet("_config")
        second["A1"] = "cfg" + marker
        wb.save(path)
        return path

    before = parity.snapshot_workbook(build(tmp_path / "o1.xlsx", ""), source="o")
    after = parity.snapshot_workbook(build(tmp_path / "o2.xlsx", "!"), source="o")
    assert [d.key for d in parity.diff_snapshots(before, after)] == \
        ["Raw!A9", "Raw!A10", "_config!A1"]


def test_a_band_that_turns_from_number_to_text_is_caught_as_a_status_move(tmp_path):
    """Carried lesson L8, seen through the parity harness.

    Excel's ``number >= text`` is silently FALSE, so a threshold cell written as
    text downgrades every ALERT to OK with no error anywhere. The formula text is
    unchanged; only the recomputed status moves -- which is precisely what this
    harness exists to see.
    """
    def build(path, band):
        wb = openpyxl.Workbook()
        cfg = wb.active
        cfg.title = "_config"
        cfg["B1"] = band
        ws = wb.create_sheet("Watchlist")
        ws["A1"] = 0.9
        ws["B1"] = '=IF(A1>=zscore_band,"ALERT","OK")'
        wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
            "zscore_band", attr_text="_config!$B$1"))
        wb.save(path)
        return path

    numeric = parity.snapshot_workbook(build(tmp_path / "num.xlsx", 0.5), source="b")
    text = parity.snapshot_workbook(build(tmp_path / "txt.xlsx", "0.5"), source="b")

    assert numeric["cells"]["Watchlist!B1"][0] == "ALERT"
    assert text["cells"]["Watchlist!B1"][0] == "OK", "L8 does not reproduce"

    diffs = parity.diff_snapshots(numeric, text)
    assert [(d.kind, d.key) for d in diffs] == [
        ("value", "_config!B1"), ("value", "Watchlist!B1")]
    assert (diffs[1].expected, diffs[1].actual) == ("ALERT", "OK")


def test_a_moved_defined_name_is_named(tmp_path):
    def build(path, target):
        wb = openpyxl.Workbook()
        wb.active.title = "_config"
        wb["_config"]["B1"] = 0.5
        wb["_config"]["B2"] = 1.5
        wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
            "zscore_band", attr_text=target))
        wb.save(path)
        return path

    before = parity.snapshot_workbook(build(tmp_path / "d1.xlsx", "_config!$B$1"),
                                      source="d")
    after = parity.snapshot_workbook(build(tmp_path / "d2.xlsx", "_config!$B$2"),
                                     source="d")
    diffs = parity.diff_snapshots(before, after)
    assert [(d.kind, d.key) for d in diffs] == [("defined_name", "zscore_band")]
    assert (diffs[0].expected, diffs[0].actual) == ("_config!$B$1", "_config!$B$2")


def test_dumps_is_ascii_and_one_cell_per_line_even_for_unicode_content(tmp_path):
    """Contract section 11: everything that travels must be pure ASCII.

    FRED's flags really are ``[warning] ALERT`` with a warning sign and its trend
    column really is made of arrows, so this is not hypothetical.
    """
    path = tmp_path / "uni.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Watchlist"
    wb["Watchlist"]["A1"] = "⚠ ALERT"
    wb["Watchlist"]["A2"] = "▲▼→"
    wb.save(path)

    text = parity.dumps(parity.snapshot_workbook(path, source="uni.xlsx"))
    text.encode("ascii")                                    # raises if not
    assert "\\u26a0" in text
    assert len([ln for ln in text.splitlines() if ln.startswith('    "')]) == 2
    assert parity.loads(text)["cells"]["Watchlist!A1"] == ["⚠ ALERT"]


def test_assert_parity_raises_naming_the_cell(base_snapshot, tmp_path):
    golden = parity.write_golden(tmp_path / "g.json", base_snapshot)
    moved = write_workbook(tmp_path / "moved2.xlsx", trigger=1.0)
    with pytest.raises(AssertionError, match=r"Watchlist!C2"):
        parity.assert_parity(golden, moved)
    parity.assert_parity(golden, write_workbook(tmp_path / "same.xlsx"))


# --------------------------------------------------------------------------
# the golden file itself
# --------------------------------------------------------------------------

def test_golden_round_trips_through_the_file(base_snapshot, tmp_path):
    path = parity.write_golden(tmp_path / "g.json", base_snapshot)
    assert parity.read_golden(path) == base_snapshot


def test_a_foreign_file_is_refused_rather_than_half_read(tmp_path):
    path = tmp_path / "foreign.json"
    path.write_text(json.dumps({"schema": "something-else", "cells": {}}))
    with pytest.raises(ValueError, match="parity-golden@1"):
        parity.read_golden(path)


@pytest.mark.parametrize("name", sorted(parity.SPINE_BASELINES))
@pytest.mark.parametrize("kind", ["shipped_golden", "demo_golden"])
def test_committed_goldens_are_pure_ascii_one_cell_per_line(name, kind):
    path = parity.repo_root() / parity.SPINE_BASELINES[name][kind]
    raw = path.read_bytes()
    assert raw.decode("ascii")                     # contract section 11: ASCII only
    assert b"\r\n" not in raw                      # stable across platforms
    text = raw.decode("ascii")
    cell_lines = [ln for ln in text.splitlines() if ln.startswith('    "')]
    assert len(cell_lines) == len(parity.read_golden(path)["cells"])


# --------------------------------------------------------------------------
# baselines: the committed goldens still describe the real workbooks
# --------------------------------------------------------------------------

@pytest.mark.parametrize("name", sorted(parity.SPINE_BASELINES))
def test_shipped_golden_matches_the_committed_workbook(name):
    root = parity.repo_root()
    spec = parity.SPINE_BASELINES[name]
    parity.assert_parity(root / spec["shipped_golden"], root / spec["workbook"])


@pytest.mark.parametrize("name", sorted(parity.SPINE_BASELINES))
def test_demo_golden_is_populated_and_its_flags_discriminate(name):
    """A vacuous baseline would pass every future parity check and prove nothing.

    The shipped ``.xlsm`` is an unpopulated template; the demo golden is the one
    with data in it, so it is the one that must be shown to have data in it --
    and its flag columns must be shown to be *doing* something: at least one
    cell lit and at least one not, off the same formula. (The two monitors spell
    a lit flag differently -- FDIC ``ALERT``, FRED ``[warning] ALERT`` -- so the
    check keys off the formula, not off a hardcoded vocabulary.)
    """
    golden = parity.read_golden(parity.repo_root()
                                / parity.SPINE_BASELINES[name]["demo_golden"])
    payloads = [p for p in golden["cells"].values() if len(p) > 1]

    numbers = [p[0] for p in payloads if isinstance(p[0], float) and p[0] != 0]
    assert len(numbers) > 200, "raw values are not reaching the formula panels"

    flags = [p for p in payloads if "ALERT" in p[1]]
    assert flags, "no flag formula in the workbook at all"
    lit = [p[0] for p in flags if isinstance(p[0], str) and "ALERT" in p[0]]
    unlit = [p[0] for p in flags if not (isinstance(p[0], str) and "ALERT" in p[0])]
    assert lit, "no flag is lit, so no flag can be shown to move"
    assert unlit, "every flag is lit -- the threshold is not discriminating"


@pytest.mark.parametrize("name", sorted(parity.SPINE_BASELINES))
@pytest.mark.parametrize("kind", ["shipped_golden", "demo_golden"])
def test_every_unpinned_formula_is_one_the_engine_documents_it_cannot_run(name, kind):
    """No silent holes: a formula that stopped evaluating must be named, not shrugged at."""
    golden = parity.read_golden(parity.repo_root()
                                / parity.SPINE_BASELINES[name][kind])
    for key, formula in parity.unevaluated(golden).items():
        assert any(fn + "(" in formula for fn in parity.UNSUPPORTED_FUNCTIONS), \
            "%s did not resolve and is not a documented gap: %s" % (key, formula)
