"""The merger record: the guard the 670% chart needed.

Capital One's other-consumer charge-off rate read 670% for the quarter ending
31 December 2022. It was arithmetically right and described nothing: the bank
absorbed Capital One Bank (USA), N.A. on 3 October 2022, and the FDIC derives
a quarterly flow by subtracting the previous quarter's year-to-date total,
which across a merger spans two banks.

The firm's instruction was to recognise the cause, not to hide the number
behind a size threshold, so everything here is driven by the FDIC's own
history record. Offline throughout: ``mergers.fetch`` takes its downloader,
and the end-to-end test injects a provider rather than opening a socket.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

import monitorbuild                                        # noqa: E402
from credit_suite.sources.fdic import layout, mergers, runner  # noqa: E402

#: Real rows, as the FDIC returned them on 5 September 2026 for the twelve-bank
#: peer set (banks.data.fdic.gov/api/history, filters ACQ_CERT:(...)).
LIVE_ROWS = [
    {"CERT": 33954, "EFFDATE": "2022-10-03T00:00:00", "CHANGECODE": 223,
     "CHANGECODE_DESC": "Merger -Without Assistance", "ACQ_CERT": 4297,
     "OUT_CERT": 33954, "OUT_NAME": "Capital One Bank (USA), National Association"},
    {"CERT": 33954, "EFFDATE": "2022-10-03T00:00:00", "CHANGECODE": 810,
     "CHANGECODE_DESC": "Participated in Absorbtion/Consolidation/Merger",
     "ACQ_CERT": 4297, "OUT_CERT": 33954, "OUT_NAME": ""},
    {"CERT": 59017, "EFFDATE": "2023-05-01T00:00:00", "CHANGECODE": 211,
     "CHANGECODE_DESC": "Failure - Whole Institution", "ACQ_CERT": 628,
     "OUT_CERT": 59017, "OUT_NAME": "First Republic Bank"},
    {"CERT": 5649, "EFFDATE": "2025-05-18T00:00:00", "CHANGECODE": 223,
     "CHANGECODE_DESC": "Merger -Without Assistance", "ACQ_CERT": 4297,
     "OUT_CERT": 5649, "OUT_NAME": "Discover Bank"},
    {"CERT": 12345, "EFFDATE": "2024-02-11T00:00:00", "CHANGECODE": 712,
     "CHANGECODE_DESC": "Branch Purchased", "ACQ_CERT": 628,
     "OUT_CERT": 12345, "OUT_NAME": "Somebody Else NA"},
]


def payload(rows, total=None):
    return json.dumps({"meta": {"total": total if total is not None else len(rows)},
                       "data": [{"data": r} for r in rows]}).encode("utf-8")


# --------------------------------------------------------------------------
# classifying the FDIC's history
# --------------------------------------------------------------------------

def test_the_capital_one_merger_is_found_with_its_quarter():
    found, unclassified = mergers.classify(LIVE_ROWS)
    assert unclassified == []
    capone = [m for m in found if m.cert == "4297"]
    assert [(m.effective, m.quarter, m.out_cert) for m in capone] == [
        ("2022-10-03", "2022-12-31", "33954"),
        ("2025-05-18", "2025-06-30", "5649")]
    assert capone[0].code == 223


def test_a_branch_purchase_is_not_a_merger():
    """A branch purchase names an acquirer and moves some offices. It does not
    put another bank's year-to-date into this bank's filing."""
    found, unclassified = mergers.classify(LIVE_ROWS)
    assert unclassified == []
    assert not any(m.code == 712 for m in found)
    assert [m.out_cert for m in found if m.cert == "628"] == ["59017"]


def test_the_mirror_row_does_not_double_count():
    """The FDIC writes a second row (code 810) against the acquirer for the
    same event. Two rows, one merger."""
    found, _ = mergers.classify(LIVE_ROWS)
    assert len([m for m in found
                if (m.cert, m.effective) == ("4297", "2022-10-03")]) == 1


def test_an_unrecognised_change_code_is_reported_not_dropped():
    """Unknown is a third answer. A code in neither list is handed back, so a
    merger this template has never seen cannot vanish silently."""
    row = dict(LIVE_ROWS[0], CHANGECODE=999, CHANGECODE_DESC="Something New")
    found, unclassified = mergers.classify([row])
    assert found == []
    assert [r["CHANGECODE"] for r in unclassified] == [999]


def test_a_row_with_no_acquirer_or_date_is_reported_not_guessed():
    found, unclassified = mergers.classify([
        dict(LIVE_ROWS[0], ACQ_CERT=None),
        dict(LIVE_ROWS[0], EFFDATE=""),
    ])
    assert found == [] and len(unclassified) == 2


@pytest.mark.parametrize("effective,quarter", [
    ("2022-10-03", "2022-12-31"), ("2022-07-01", "2022-09-30"),
    ("2026-06-18", "2026-06-30"), ("2025-05-18", "2025-06-30"),
    ("2023-01-01", "2023-03-31"), ("2023-12-31", "2023-12-31"),
])
def test_the_effective_date_maps_to_the_quarter_that_contains_it(effective, quarter):
    """The first filing that covers the combined bank is the one whose flow is
    contaminated -- the quarter the merger happened in, not the next."""
    assert mergers.quarter_end(effective) == quarter


def test_the_sentence_carries_the_code_and_what_it_means():
    """Behaviour 15: show the jargon and say what it means, never one instead
    of the other."""
    found, _ = mergers.classify(LIVE_ROWS)
    line = found[0].sentence("Capital One NA")
    assert "223" in line and "Merger -Without Assistance" in line
    assert "merged another into itself" in line
    assert "year-to-date" in line


# --------------------------------------------------------------------------
# fetching, without a socket
# --------------------------------------------------------------------------

def test_one_request_covers_the_whole_peer_set():
    url = mergers.request_url(["628", "4297", "17534"])
    import urllib.parse
    query = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert query["filters"] == ["ACQ_CERT:(17534 OR 4297 OR 628)"]
    assert "history" in url


def test_fetch_classifies_what_it_downloads():
    calls = []

    def download(url, label):
        calls.append(label)
        return payload(LIVE_ROWS)

    found, unclassified = mergers.fetch(["4297", "628"], download)
    assert calls == ["history"]
    assert {m.cert for m in found} == {"4297", "628"}
    assert unclassified == []


def test_a_truncated_history_page_is_refused():
    """A merger nobody sees is a quarter nobody marks, so a short page is an
    error rather than a smaller answer."""
    def download(url, label):
        return payload(LIVE_ROWS, total=5000)

    with pytest.raises(RuntimeError) as ei:
        mergers.fetch(["4297"], download, limit=1000)
    assert "truncated" in str(ei.value)


def test_no_certs_asks_nobody():
    def download(url, label):        # pragma: no cover - must not run
        raise AssertionError("asked with no certs")

    assert mergers.fetch([], download) == ([], [])


# --------------------------------------------------------------------------
# the workbook: written by the runner, read by the trend tool
# --------------------------------------------------------------------------

class FakeProvider:
    """A provider carrying a real merger record, so the offline bar exercises
    the path the live run takes."""

    def __init__(self, record):
        self.mergers = record


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    with monitorbuild.built_monitor("fdic") as (workbook, _stdout):
        path = Path(tmp_path_factory.mktemp("mergers") / "Bank_Peer_Monitor.xlsm")
        path.write_bytes(Path(workbook).read_bytes())
    return path


def test_the_runner_writes_the_merger_record_onto_the_tab(built, tmp_path):
    import shutil

    import trend as T

    book = tmp_path / "with_mergers.xlsm"
    shutil.copyfile(built, book)
    found, _ = mergers.classify(LIVE_ROWS)
    backend = runner.OpenpyxlBackend(str(book), runner.FDIC, runner.fields.RAW_FIELDS)
    cfg = backend.read_config()
    written = runner._write_merger_block(backend, cfg,
                                         FakeProvider(mergers.by_cert(found)))
    backend.finalize()
    assert written == 3

    wb = openpyxl.load_workbook(book, keep_vba=True)
    text = "\n".join(str(c.value) for row in wb[layout.MERGERS_TAB].iter_rows()
                     for c in row if c.value)
    assert "Capital One NA" in text and "2022-12-31" in text
    assert "223 -- Merger -Without Assistance" in text

    record = T.read_mergers(wb)
    wb.close()
    assert record is not None
    assert [e["quarter"] for e in record["4297"]] == ["2022-12-31", "2025-06-30"]


def test_a_run_that_never_asked_says_so_rather_than_showing_an_empty_tab(built, tmp_path):
    """The demo provider does not fetch history. An empty tab that reads as
    'no mergers' is the 670 all over again, so it says UNKNOWN instead."""
    import shutil

    import trend as T

    book = tmp_path / "no_record.xlsm"
    shutil.copyfile(built, book)
    backend = runner.OpenpyxlBackend(str(book), runner.FDIC, runner.fields.RAW_FIELDS)
    cfg = backend.read_config()
    assert runner._write_merger_block(backend, cfg, FakeProvider(None)) == 0
    backend.finalize()

    wb = openpyxl.load_workbook(book, keep_vba=True)
    text = "\n".join(str(c.value) for row in wb[layout.MERGERS_TAB].iter_rows()
                     for c in row if c.value)
    assert "UNKNOWN" in text
    assert T.read_mergers(wb) is None
    wb.close()


def test_asked_and_none_found_is_not_the_same_as_never_asked(built, tmp_path):
    import shutil

    import trend as T

    book = tmp_path / "none_found.xlsm"
    shutil.copyfile(built, book)
    backend = runner.OpenpyxlBackend(str(book), runner.FDIC, runner.fields.RAW_FIELDS)
    cfg = backend.read_config()
    assert runner._write_merger_block(backend, cfg, FakeProvider({})) == 0
    backend.finalize()

    wb = openpyxl.load_workbook(book, keep_vba=True)
    assert T.read_mergers(wb) == {}          # asked; none found
    wb.close()


def test_a_workbook_with_no_merger_tab_reports_unknown(built, tmp_path):
    """An older workbook cannot answer. That is UNKNOWN, not 'no mergers'."""
    import shutil

    import trend as T

    book = tmp_path / "old.xlsm"
    shutil.copyfile(built, book)
    wb = openpyxl.load_workbook(book, keep_vba=True)
    del wb[layout.MERGERS_TAB]
    assert T.read_mergers(wb) is None
    wb.close()


def test_the_tab_explains_itself_to_a_reader(built):
    wb = openpyxl.load_workbook(built, keep_vba=True)
    text = "\n".join(str(c.value) for row in wb[layout.MERGERS_TAB].iter_rows()
                     for c in row if c.value)
    wb.close()
    assert "running total from 1 January" in text
    assert "Never inferred from the numbers." in text
    assert "quarter affected" in text


def test_only_the_charted_window_is_asked_for():
    """The endpoint returns a bank's whole life -- Wells Fargo back to 1972.
    Only a quarter we chart can be contaminated for a quarter we chart, and a
    tab of 250 rows from the 1970s is a tab nobody reads."""
    import urllib.parse

    url = mergers.request_url(["4297"], since="2022-09-01")
    query = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert query["filters"] == ["ACQ_CERT:(4297) AND EFFDATE:[2022-09-01 TO *]"]


@pytest.mark.parametrize("day,start", [
    ("2022-09-01", "2022-07-01"), ("2022-09-30", "2022-07-01"),
    ("2023-01-15", "2023-01-01"), ("2026-06-30", "2026-04-01"),
])
def test_the_window_starts_at_the_first_day_of_the_oldest_charted_quarter(day, start):
    """Citibank merged on 1 July 2022 -- inside the quarter ending 30
    September 2022, which this monitor charts. The first live run asked from
    the oldest REPDTE (2022-09-01) and missed it by two months."""
    assert mergers.quarter_start(day) == start
