"""#259: a ratio on a book that does not exist is "N/A", never "OK".

Bank of New York Mellon has no credit-card book. The FDIC publishes 0.00 for
its card delinquency rate, 0.00 is a number, and the Watchlist read OK --
"checked and clean" where the truth was "nothing to check". Behaviour 6:
unknown is a third answer, drawn differently from both.

The shape of the fix changed the next day. #259 first added a *guarded direct*
metric -- a landed rate that reads None when its book is zero -- because the
FDIC's fifteen class rates were landed as rates. #268 replaced all fifteen
with ratios computed over the book, and a ratio already returns None on a zero
denominator, so the guarded-direct machinery had no users left and went. What
is pinned here is the behaviour, which did not change: three blanks, told
apart.

Three places have to agree, and each is pinned: the engine's value, the
digest's status, and the workbook's formulas.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

from credit_suite.engine import digest, thresholds       # noqa: E402
from credit_suite.engine.config import Threshold         # noqa: E402
from credit_suite.sources.fdic import engine_api as R    # noqa: E402
from credit_suite.sources.fdic import layout             # noqa: E402

ABOVE = Threshold(1.0, 2.0, "above")


# --------------------------------------------------------------------------
# the engine's value
# --------------------------------------------------------------------------

def test_a_class_ratio_on_a_zero_book_reads_none_not_zero():
    """What the FDIC publishes (0.00) versus what is true (no book)."""
    assert R.metric_value("P3CRCD_BOOK", {"P3CRCD": 0.0, "LNCRCD": 0}) is None
    assert R.metric_value("P3CRCD_BOOK", {"P3CRCD": 0.0, "LNCRCD": None}) is None
    assert R.metric_value("P3CRCD_BOOK", {"P3CRCD": 0.0, "LNCRCD": 5_000}) == 0.0
    assert R.metric_value("P3CRCD_BOOK",
                          {"P3CRCD": 85.0, "LNCRCD": 5_000}) == pytest.approx(1.7)


def test_a_whole_bank_metric_stands_on_no_book_and_is_untouched():
    assert R.metric_value("NCLNLSR", {"NCLNLSR": 0.0}) == 0.0
    assert R.balance_field("NCLNLSR") is None


@pytest.mark.parametrize("metric,balance", [
    ("P3CRCD_BOOK", "LNCRCD"), ("NACRCD_BOOK", "LNCRCD"),
    ("P3RELOC_BOOK", "LNRELOC"),          # HELOC got its book in #268
    ("NTCONOTQ_BOOK", "LNCONOTH"), ("NTCIQ_BOOK", "LNCI"),
    ("P3REMULT_BOOK", "LNREMULT"), ("UNINSDEPR", "DEP"),
    ("TEXAS", None), ("CRECONR", None), ("NCLNLSR", None), ("EQV", None),
])
def test_every_metric_knows_the_book_it_stands_on(metric, balance):
    assert R.balance_field(metric) == balance


def test_every_loan_class_rate_stands_on_a_landed_book():
    """Report the denominator: every class rate, over its own class, with the
    balance actually landed rather than assumed."""
    rates = sorted(m for m in R.LOANBOOK_CLASS if m in R.METRICS)
    assert len(rates) == 35
    for metric in rates:
        balance = R.balance_field(metric)
        assert balance in R.RAW_FIELDS, metric
        assert balance.startswith("LN"), (metric, balance)


# --------------------------------------------------------------------------
# the digest's status: three blanks, three different answers
# --------------------------------------------------------------------------

def test_no_book_is_not_applicable_and_no_number_is_blank():
    reg = R.METRICS
    assert digest.metric_status(reg, "P3CRCD_BOOK",
                                {"P3CRCD": 0.0, "LNCRCD": 0}, ABOVE) == "N/A"
    assert digest.metric_status(reg, "P3CRCD_BOOK",
                                {"P3CRCD": None, "LNCRCD": 5_000}, ABOVE) == ""
    assert digest.metric_status(reg, "P3CRCD_BOOK",
                                {"P3CRCD": 20.0, "LNCRCD": 5_000}, ABOVE) == "OK"
    assert digest.metric_status(reg, "P3CRCD_BOOK",
                                {"P3CRCD": 150.0, "LNCRCD": 5_000}, ABOVE) == "ALERT"


def test_a_flow_ratio_gets_the_same_split():
    reg = R.METRICS
    assert digest.metric_status(reg, "NTCONOTQ_BOOK",
                                {"NTCONOTQ": 5_318, "LNCONOTH": 0}, ABOVE) == "N/A"
    assert digest.metric_status(reg, "NTCONOTQ_BOOK",
                                {"NTCONOTQ": None, "LNCONOTH": 3_173}, ABOVE) == ""


def test_the_third_answer_is_neither_ok_nor_blank():
    assert thresholds.NOT_APPLICABLE not in (thresholds.OK, "")
    # the split lives above status_for, which still reads a blank as OK
    assert thresholds.status_for(None, ABOVE) == thresholds.OK


# --------------------------------------------------------------------------
# the workbook's formulas say the same thing
# --------------------------------------------------------------------------

def test_the_value_cell_blanks_a_ratio_on_a_zero_book():
    formula = layout.metric_formula("P3CRCD_BOOK", 1, 16)
    num = layout._fref(1, "P3CRCD", 16)
    den = layout._fref(1, "LNCRCD", 16)
    assert formula == ('=IF(OR(%s="",%s=""),"",IF(%s=0,"",%s/%s*100))'
                       % (num, den, den, num, den))


def test_a_landed_metric_keeps_the_plain_blank_guard():
    ref = layout._fref(1, "NCLNLSR", 16)
    assert layout.metric_formula("NCLNLSR", 1, 16) == '=IF(%s="","",%s)' % (ref, ref)


@pytest.fixture(scope="module")
def watchlist_helpers(tmp_path_factory):
    """The Watchlist helper formulas of a freshly built (unpopulated) workbook,
    keyed by metric id, for peer slot 1."""
    import monitorbuild
    with monitorbuild.built_monitor("fdic") as (workbook, _stdout):
        wb = openpyxl.load_workbook(workbook, keep_vba=True)
        ws = wb["Watchlist"]
        cfg = R.parse_config(list(wb["_config"].iter_rows(values_only=True)))
        r = layout.wl_row(1)
        out = {s.id: ws.cell(r, layout.WL_HELPER_COL0 + k).value
               for k, s in enumerate(cfg.series)}
        wb.close()
    return out


def test_the_watchlist_helper_says_na_when_the_book_is_zero(watchlist_helpers):
    f = watchlist_helpers["P3CRCD_BOOK"]
    bal = layout._fref(1, "LNCRCD", 16)
    assert 'IF(OR(%s="",%s=0),"N/A","")' % (bal, bal) in f
    assert re.search(r'"ALERT".*"WATCH".*"OK"', f)


def test_a_metric_with_no_book_keeps_the_plain_blank(watchlist_helpers):
    f = watchlist_helpers["TEXAS"]
    assert "N/A" not in f
    assert f.startswith('=IF(NOT(ISNUMBER(')
