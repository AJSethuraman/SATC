#!/usr/bin/env python3
"""Import an existing client list (CSV or Excel) into clients.json.

A standalone onboarding adapter: point it at a spreadsheet of clients and it maps
the columns to clients.json fields and appends any clients not already present
(deduped by email then name, so re-running is safe). Common headers (Name, Email,
Phone, Tax Year, Filing Status, ...) are recognized automatically; an optional
``import_map.json`` (header -> field) overrides or extends the mapping, and any
unrecognized column is kept under a snake_cased key so no data is lost.

Source file: ``client_list.csv`` or ``client_list.xlsx`` in the folder. Built as
pure mapping functions plus a thin run_import I/O layer. Excel support uses the
openpyxl dependency the suite already has (imported lazily).
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import core
import generate_documents
import sort_tax_docs

IMPORT_MAP_FILENAME = "import_map.json"
SOURCE_BASENAMES = ("client_list", "clients_import", "client_import")
SOURCE_SUFFIXES = (".csv", ".xlsx")

# Normalized header -> clients.json field for the headers people commonly use.
DEFAULT_HEADER_MAP: dict[str, str] = {
    "name": "client_name",
    "client": "client_name",
    "client name": "client_name",
    "full name": "client_name",
    "taxpayer": "client_name",
    "email": "email",
    "e mail": "email",
    "email address": "email",
    "phone": "phone",
    "telephone": "phone",
    "phone number": "phone",
    "tax year": "tax_year",
    "year": "tax_year",
    "filing status": "filing_status",
    "status": "filing_status",
    "address": "address",
    "mailing address": "address",
}


def normalize_header(header: str) -> str:
    """Lowercase and treat underscores/hyphens as spaces, collapsing whitespace."""

    return " ".join(str(header).lower().replace("_", " ").replace("-", " ").split())


def header_to_field(header: str, mapping: dict) -> str:
    """Resolve a spreadsheet header to a clients.json field name."""

    normalized = normalize_header(header)
    if normalized in mapping:
        return mapping[normalized]
    return normalized.replace(" ", "_")  # keep unknown columns, snake_cased


def map_row(row: dict, mapping: dict) -> dict:
    """Map one spreadsheet row to a client record (blank cells dropped)."""

    client: dict = {}
    for header, value in row.items():
        if header is None:
            continue
        text = "" if value is None else str(value).strip()
        if text == "":
            continue
        client[header_to_field(header, mapping)] = text
    return client


def build_clients(rows: list[dict], mapping: dict) -> tuple[list[dict], list[str]]:
    """Map rows to client records, skipping rows with no client_name."""

    clients: list[dict] = []
    warnings: list[str] = []
    for index, row in enumerate(rows, start=1):
        client = map_row(row, mapping)
        if not client.get("client_name"):
            if client:  # a non-empty row with no resolvable name
                warnings.append(f"Row {index}: no client name; skipped.")
            continue
        clients.append(client)
    return clients, warnings


def load_mapping(input_folder: Path) -> dict:
    """Default header map, overlaid with any import_map.json in the folder."""

    mapping = dict(DEFAULT_HEADER_MAP)
    path = input_folder / IMPORT_MAP_FILENAME
    if path.exists():
        try:
            override = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(override, dict):
                mapping.update({normalize_header(k): v for k, v in override.items()})
        except json.JSONDecodeError:
            pass
    return mapping


def find_source_file(input_folder: Path) -> Path | None:
    for basename in SOURCE_BASENAMES:
        for suffix in SOURCE_SUFFIXES:
            candidate = input_folder / f"{basename}{suffix}"
            if candidate.is_file():
                return candidate
    return None


def read_rows(source: Path) -> list[dict]:
    """Read a CSV or Excel source into a list of header->value dicts."""

    if source.suffix.lower() == ".csv":
        with source.open(newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))

    from openpyxl import load_workbook  # lazy: only needed for Excel sources

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = [dict(zip(headers, values)) for values in rows_iter]
    workbook.close()
    return rows


def run_import(input_folder, status_callback=None) -> dict:
    """Import the client list source into clients.json (non-destructive append)."""

    input_folder = Path(input_folder)
    output_folder = sort_tax_docs.setup_output_folders(input_folder)

    base_result = {
        "tool": "import",
        "output_folder": output_folder,
        "clients_file": None,
        "imported": 0,
        "added": 0,
        "skipped": 0,
        "warnings": [],
    }

    source = find_source_file(input_folder)
    if source is None:
        names = ", ".join(f"{b}.csv/.xlsx" for b in SOURCE_BASENAMES[:1])
        return {**base_result, "summary": f"No client list found (looked for {names}); nothing imported."}

    if status_callback:
        status_callback(f"Importing clients from {source.name}")
    mapping = load_mapping(input_folder)
    try:
        rows = read_rows(source)
    except Exception as exc:
        return {**base_result, "summary": f"Could not read {source.name} ({exc}); nothing imported."}

    clients, warnings = build_clients(rows, mapping)
    clients_file = input_folder / "clients.json"
    existing: list[dict] = []
    if clients_file.exists():
        try:
            loaded = json.loads(clients_file.read_text(encoding="utf-8"))
            existing = loaded if isinstance(loaded, list) else [loaded]
        except json.JSONDecodeError:
            existing = []
    merged, added, skipped = core.append_new_clients(existing, clients)
    if added:
        clients_file.write_text(json.dumps(merged, indent=2), encoding="utf-8")

    return {
        **base_result,
        "clients_file": clients_file,
        "imported": len(clients),
        "added": added,
        "skipped": skipped,
        "warnings": warnings,
        "summary": (
            f"Imported {len(clients)} client(s) from {source.name}: "
            f"added {added} new" + (f", {skipped} already present." if skipped else ".")
        ),
    }


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Import a CSV/Excel client list into clients.json.")
    parser.add_argument("input_folder", help="Folder containing client_list.csv or client_list.xlsx.")
    args = parser.parse_args()

    folder = Path(args.input_folder).expanduser().resolve()
    if not folder.is_dir():
        print(f"Input folder does not exist or is not a directory: {folder}")
        return 1

    result = run_import(folder, status_callback=print)
    print(result["summary"])
    for warning in result["warnings"]:
        print(f"  NOTE: {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
