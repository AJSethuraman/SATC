#!/usr/bin/env python3
"""EDGAR Criticized/Classified Tracker -- runner (the data path).

Single source of truth for the data path; embedded verbatim into the workbook's
_code_py tab. The VBA ExtractFiles macro writes this back out as runner.py and
the user runs it from PowerShell against the CLOSED workbook (openpyxl backend).

Built to BUILD_SPEC_EDGAR.md (grounded in COVERAGE_RESEARCH_EDGAR.md), governed
by TEMPLATE_CONTRACT.md. THE MISSION: whose commercial book is being risk-rated
down -- before delinquency shows it? Quarterly criticized/classified per
competitor HOLDING COMPANY from its own filed 10-Q/10-K credit-quality tables
(ASC 326-20-50-5 credit quality indicators), plus an 8-K credit-event lane.

The [PEERS] CIK list IS the watchlist (contract sec 13): slot | cik | name |
ticker | active, expandable by config edit within built slot capacity. Raw
anchors depend ONLY on the slot, so editing WHICH bank occupies a slot moves
no formula.

Clean seams (BUILD SPEC 0.3 / contract sec 6):
  * PROVIDER SEAM -- one provider (EDGAR), isolated, plain urllib, KEYLESS,
    but a declared User-Agent "{org} {email}" is REQUIRED on every live
    request (absent/generic UA = silent 403 + ~10-min IP block). Two
    implementations: EdgarProvider (live, two-stage) + EdgarDemoProvider
    (offline deterministic -- every test uses it).
  * DISCLOSURE-DIALECT HONESTY (BUILD SPEC 0.6) -- each bank carries a
    `family` flag (grades_full | criticized_only | ig_nig | unmapped)
    derived from its [MEMBER_MAP] rows. Metrics a family cannot support
    render "N/A (family)" -- never approximated. Newly observed members
    bootstrap into [MEMBER_MAP] as grade="unmapped" and are flagged --
    never silently guessed (contract sec 13).
  * METRIC ENGINE -- canonical rollups computed here for the digest AND
    landed into per-slot canonical metric blocks so the dashboard formulas
    stay simple references; the member-level facts land in per-slot fact
    blocks as the audit trail (accession per fact -- contract sec 12).
  * WATCHLIST VALIDATOR -- DEFAULT-DENY: only ^cik:[0-9]{1,10}$ keys are
    admitted; blank/malformed CIKs are refused BY NAME with a --lookup hint.
  * STALENESS GUARD -- a bank whose latest CQI quarter lags the peer-set max
    by more than stale_multiplier quarters is STALE (late filer/delisted?),
    excluded from alert counts.

Traps carried (COVERAGE_RESEARCH_EDGAR.md E1-E10): E1 convenience APIs are
non-dimensional (extracted instances only for CQI) - E2 accrued-interest twin
tags (per-bank preference) - E3 vintage line items IGNORED (grade totals are
single facts; summing vintages double-counts) - E4 extension members
(bootstrap + drift flag) - E5 class taxonomies differ (coarse [CLASS_MAP]) -
E6 no Q4 10-Q (Q4 = the 10-K, 60-90d lag) - E7 amendment dedupe by latest
acceptanceDateTime - E8 CIK padding split (10-digit on data.sec.gov, unpadded
in /Archives) - E9 uniform definitions, bank-specific application timing -
E10 corporate proxies block sec.gov (--selftest + allowlist note).

No AI/LLM anywhere; every computation is deterministic (BUILD SPEC 0.5).
Pure ASCII (L3).
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Fixed raw-block geometry: ONE block per PEER SLOT in Raw_EDGAR. Each block:
#   header row   (slot label + runtime identity: unit key, name, family,
#                 latest accession, tag preference)
#   label row    (metric column labels)
#   raw_slots    quarterly CANONICAL METRIC rows, newest-first (dashboards
#                read THESE -- rollups land as values, formulas stay simple)
#   facts label + FACT_ROWS latest-quarter MEMBER-LEVEL fact rows (the audit
#                trail: class x grade member amounts + accession per fact)
#   events label + EVENT_ROWS 8-K event rows (trailing 4 quarters)
# Anchors depend only on (slot, raw_slots) -- never on which bank occupies
# the slot -- so dashboard formulas survive any [PEERS] edit.
# ---------------------------------------------------------------------------
RAW_TAB = "Raw_EDGAR"
RAW_SLOTS_DEFAULT = 12         # quarters kept per bank (newest-first)
PEER_SLOTS_DEFAULT = 20        # BUILT peer capacity (make_workbook.py --peer-slots)
FACT_ROWS = 36                 # latest-quarter member-fact audit rows per slot
EVENT_ROWS = 8                 # 8-K event rows per slot (trailing 4 quarters)
RAW_FIRST_ROW = 2

# Canonical metric block columns (B onward). ACCESSION/TAG are per-quarter
# provenance text columns (contract sec 12: accession recorded per value).
RAW_METRIC_COLS = [
    "CRIT_RATIO", "CLASS_RATIO", "SM_RATIO",
    "CRIT_QOQ", "CRIT_YOY", "CRIT_GROWTH",
    "CRIT_USD", "CLASS_USD", "SM_USD", "TOTAL_COMM_USD",
    "MIX_CI", "MIX_CRE", "MIX_CONSTR", "MIX_OTHER",
    "ACCESSION", "TAG",
]
FACT_COLS = ["period", "class_qname", "class", "member_qname", "grade",
             "value_usd", "tag", "accession"]
EVENT_COLS = ["date", "item", "form", "accession", "url"]

STATUS_COL = 12                # column L (Dashboard_Criticized is 11 wide)
STATUS_COL_BY_TAB = {"Dashboard_Mix_Events": 16}   # wider tab -> column P
DASH_TABS = ("Dashboard_Criticized", "Dashboard_Mix_Events")

PACK_VERSION = "1.0"

# The watchlist join-key whitelist (BUILD SPEC 0.1): default-deny. Only a
# literal SEC CIK key passes; name-only rows, tickers-as-keys and
# unanticipated key forms are refused.
ENTITY_KEY_PATTERN = r"^cik:[0-9]{1,10}$"

QUARTER_DAYS = 92
EVENT_WINDOW_DAYS = 366        # the 8-K lane keeps trailing ~4 quarters

# --- taxonomy constants (verified in COVERAGE_RESEARCH_EDGAR.md) -----------
TAG_PRIMARY = "FinancingReceivableBeforeAllowanceForCreditLoss"
TAG_EXCL = ("FinancingReceivableExcludingAccruedInterest"
            "BeforeAllowanceForCreditLoss")
CQI_TAGS = (TAG_PRIMARY, TAG_EXCL)
CQI_AXIS = "InternalCreditAssessmentAxis"
# The class axis keeps its legacy name; many filers dimension the same table
# by the portfolio-segment axis instead -- both accepted, legacy preferred.
CLASS_AXES = ("FinancingReceivableRecordedInvestmentByClassOf"
              "FinancingReceivableAxis",
              "FinancingReceivablePortfolioSegmentAxis")
US_GAAP_URI_HINT = "fasb.org/us-gaap"

# --- canonical grades / classes / families ---------------------------------
GRADES = ("pass", "special_mention", "substandard", "doubtful", "loss",
          "criticized_accruing", "criticized_nonaccruing",
          "investment_grade", "noninvestment_grade", "ignore", "unmapped")
# NOTE: "ignore" is for filer-tagged SUBTOTAL members (us-gaap:CriticizedMember
# parents SM/Substandard/Doubtful/Loss -- counting it alongside its leaves
# would double-count); "investment_grade"/"noninvestment_grade" carry the
# ig_nig dialect without flooding the unmapped lane.
CRITICIZED_GRADES = {"special_mention", "substandard", "doubtful", "loss",
                     "criticized_accruing", "criticized_nonaccruing"}
CLASSIFIED_GRADES = {"substandard", "doubtful", "loss"}
CLASSES = ("ci", "cre", "construction", "other_commercial")
CLASS_VALUES = CLASSES + ("excluded", "unmapped")
FAMILIES = ("grades_full", "criticized_only", "ig_nig", "unmapped")

# --- 8-K item lane (harvested from the submissions `items` array) ----------
HARVEST_ITEMS = ("1.03", "2.04", "2.06", "4.02", "2.02")
WATCH_ITEMS = ("2.04", "4.02", "1.03")     # automatic WATCH flag on the bank
ITEM_LABEL = {
    "1.03": "bankruptcy/receivership",
    "2.04": "triggering events / debt acceleration",
    "2.06": "material impairment",
    "4.02": "non-reliance on prior financials",
    "2.02": "results of operations (informational)",
}

# --- SEC hosts (E8: CIK padding split; E10: proxy allowlist) ----------------
SUBMISSIONS_URL = "https://data.sec.gov/submissions/CIK{cik10}.json"
ARCHIVES_DIR_URL = ("https://www.sec.gov/Archives/edgar/data/"
                    "{cik}/{accession_nodash}/")
TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
FTS_URL = "https://efts.sec.gov/LATEST/search-index?q=%22bank%22"
BROWSE_URL = ("https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany"
              "&CIK={cik}&type=10-Q&dateb=&owner=include&count=40")

SEC_HOSTS = ("data.sec.gov", "www.sec.gov", "efts.sec.gov")


def cik10(cik) -> str:
    """data.sec.gov wants the 10-digit zero-padded CIK (trap E8)."""
    return f"{int(str(cik)):010d}"


def cik_unpadded(cik) -> str:
    """/Archives paths want the UNPADDED CIK (trap E8)."""
    return str(int(str(cik)))


def accession_nodash(acc: str) -> str:
    return str(acc).replace("-", "")


def filing_dir_url(cik, accession) -> str:
    return ARCHIVES_DIR_URL.format(cik=cik_unpadded(cik),
                                   accession_nodash=accession_nodash(accession))


def metric_col(name: str) -> int:
    """Raw_EDGAR column for a canonical metric: B onward, fixed order."""
    return 2 + RAW_METRIC_COLS.index(name)


def fact_col(name: str) -> int:
    return 1 + FACT_COLS.index(name)


def event_col(name: str) -> int:
    return 1 + EVENT_COLS.index(name)


def slot_label(slot: int) -> str:
    return f"slot{slot:02d}"


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` tab; the workbook is the source of
# truth -- BUILD SPEC 0.4, INCLUDING [PEERS], [MEMBER_MAP] and [CLASS_MAP])
# ---------------------------------------------------------------------------
PEER_HEADER = ["slot", "cik", "name", "ticker", "active"]
MEMBER_MAP_HEADER = ["cik", "family", "member_qname", "grade"]
CLASS_MAP_HEADER = ["cik", "class_member_qname", "class"]


@dataclass
class PeerRow:
    """One [PEERS] line. An empty row (slot number only) is provisioned
    headroom the user fills by hand (add a bank = edit + re-run)."""
    slot: Optional[int]
    cik: str               # normalized digit string; "" when blank
    name: str
    ticker: str
    active: bool

    @property
    def has_bank(self) -> bool:
        return bool(self.cik or self.name)

    @property
    def entity_key(self) -> str:
        return f"cik:{self.cik}"


@dataclass
class MemberMapRow:
    cik: str
    family: str
    member_qname: str
    grade: str


@dataclass
class ClassMapRow:
    cik: str               # "*" = standard us-gaap member, applies to all
    class_qname: str
    klass: str


@dataclass
class Threshold:
    watch: Optional[float]
    alert: Optional[float]
    direction: str


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, Threshold] = field(default_factory=dict)
    peers: List[PeerRow] = field(default_factory=list)
    member_map: List[MemberMapRow] = field(default_factory=list)
    class_map: List[ClassMapRow] = field(default_factory=list)

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))

    @property
    def peer_slots(self) -> int:
        return int(float(self.settings.get("peer_slots", PEER_SLOTS_DEFAULT)))

    @property
    def stale_multiplier(self) -> float:
        v = _as_float(self.settings.get("stale_multiplier", "2.0"))
        return 2.0 if v is None else v

    @property
    def user_agent(self) -> str:
        return str(self.settings.get("edgar_user_agent", "") or "").strip()

    # -- derived lookup structures -----------------------------------------
    def grade_of(self, cik: str, qname: str) -> Optional[str]:
        for r in self.member_map:
            if r.cik == str(cik) and r.member_qname == qname:
                return r.grade
        return None

    def family_of(self, cik: str) -> str:
        fams = [r.family for r in self.member_map
                if r.cik == str(cik) and r.family in FAMILIES]
        return fams[0] if fams else "unmapped"

    def class_of(self, cik: str, qname: str) -> Optional[str]:
        best = None
        for r in self.class_map:
            if r.class_qname != qname:
                continue
            if r.cik == str(cik):
                return r.klass          # per-bank row wins over "*"
            if r.cik == "*":
                best = r.klass
        return best


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _as_float(v):
    try:
        s = str(v).strip()
        return float(s) if s != "" else None
    except (TypeError, ValueError):
        return None


def _norm_cik(v) -> str:
    """Normalize a [PEERS] cik cell to a plain digit string. Excel hands
    numeric cells over as int/float; text stays as typed so a malformed value
    reaches the gate and is refused, never silently coerced."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _norm_slot(v) -> Optional[int]:
    try:
        s = str(v).strip()
        return int(float(s)) if s != "" else None
    except (TypeError, ValueError):
        return None


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists). Sections in col A:
    [SETTINGS], [THRESHOLDS], [PEERS], [MEMBER_MAP], [CLASS_MAP]. Lines
    starting `#` are comments; blank col-A rows (the bootstrap provisioning
    inside MEMBER_MAP/CLASS_MAP) are skipped."""
    cfg = Config()
    section = None
    hdr: Dict[str, List[str]] = {}
    for raw in rows:
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            continue
        if not a or a.startswith("#"):
            continue
        cells = list(raw) + [None] * 8
        if section == "SETTINGS":
            if a.lower() in ("key", "name"):
                continue
            cfg.settings[a] = ("" if cells[1] is None else str(cells[1]).strip())
        elif section == "THRESHOLDS":
            if a.lower() == "id":
                continue
            cfg.thresholds[a] = Threshold(
                watch=_as_float(cells[1]), alert=_as_float(cells[2]),
                direction=(str(cells[3] or "above")).strip().lower() or "above")
        elif section == "PEERS":
            if a.lower() == "slot":
                hdr["peers"] = [str(c).strip().lower() for c in raw]
                continue
            v = {h: (None if i >= len(raw) else raw[i])
                 for i, h in enumerate(hdr.get("peers", PEER_HEADER))}
            cfg.peers.append(PeerRow(
                slot=_norm_slot(v.get("slot")),
                cik=_norm_cik(v.get("cik")),
                name=("" if v.get("name") is None else str(v.get("name")).strip()),
                ticker=("" if v.get("ticker") is None
                        else str(v.get("ticker")).strip().upper()),
                active=_as_bool(v.get("active"))))
        elif section == "MEMBER_MAP":
            if a.lower() == "cik":
                continue
            cfg.member_map.append(MemberMapRow(
                cik=_norm_cik(cells[0]),
                family=str(cells[1] or "").strip().lower(),
                member_qname=str(cells[2] or "").strip(),
                grade=str(cells[3] or "").strip().lower()))
        elif section == "CLASS_MAP":
            if a.lower() == "cik":
                continue
            cfg.class_map.append(ClassMapRow(
                cik=("*" if str(cells[0]).strip() == "*"
                     else _norm_cik(cells[0])),
                class_qname=str(cells[1] or "").strip(),
                klass=str(cells[2] or "").strip().lower()))
    return cfg


# ---------------------------------------------------------------------------
# METRIC REGISTRY -- the canonical rollups (BUILD SPEC sec 3). Values are
# COMPUTED here and LANDED into the per-slot canonical metric blocks; the
# dashboards read those blocks by simple reference, so Python and Excel
# cannot drift on the arithmetic. Family gating (Tier 2) is enforced here
# AND rendered in-formula off the landed family cell.
# ---------------------------------------------------------------------------
@dataclass
class MetricSpec:
    id: str
    title: str
    tier: int                  # 1 = all crit-capable families, 2 = grades_full
    raw_col: str               # canonical metric block column
    source_class: str = "A"
    watchlist_capable: bool = True


METRIC_SPECS = [
    MetricSpec("criticized_ratio", "Criticized / total commercial %", 1,
               "CRIT_RATIO"),
    MetricSpec("classified_ratio", "Classified / total commercial %", 2,
               "CLASS_RATIO"),
    MetricSpec("sm_ratio", "Special mention / total commercial %", 2,
               "SM_RATIO"),
    MetricSpec("crit_qoq_delta", "Criticized ratio QoQ delta (pp)", 1,
               "CRIT_QOQ"),
    MetricSpec("crit_yoy_delta", "Criticized ratio YoY delta (pp)", 1,
               "CRIT_YOY"),
    MetricSpec("crit_growth", "Criticized $ QoQ growth %", 1, "CRIT_GROWTH"),
    MetricSpec("mix_ci", "Criticized $ share: C&I %", 1, "MIX_CI"),
    MetricSpec("mix_cre", "Criticized $ share: CRE %", 1, "MIX_CRE"),
    MetricSpec("mix_construction", "Criticized $ share: construction %", 1,
               "MIX_CONSTR"),
    MetricSpec("mix_other", "Criticized $ share: other commercial %", 1,
               "MIX_OTHER"),
]
METRICS = {m.id: m for m in METRIC_SPECS}
BANDED_METRICS = ("criticized_ratio", "crit_qoq_delta", "sm_ratio")


class MetricError(Exception):
    pass


def _ratio(num: Optional[float], den: Optional[float]) -> Optional[float]:
    if num is None or den is None or den == 0:
        return None
    return num / den * 100.0


# ---------------------------------------------------------------------------
# WATCHLIST VALIDATOR -- DEFAULT-DENY (BUILD SPEC 0.1 / sec 3)
# ---------------------------------------------------------------------------
class WatchlistRefused(Exception):
    """Raised by the hard gates; carries the interpolated refusal message."""


class PeerCapacityError(Exception):
    """More peers than BUILT slots: refuse with the rebuild command -- the
    runner never truncates a peer list."""


class UserAgentMissing(Exception):
    """Live mode with a blank edgar_user_agent (BUILD SPEC 0.3): fail fast --
    an absent/generic UA earns silent 403s and a ~10-minute IP block."""


def peer_refusal_message(peer: PeerRow, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: peer slot {peer.slot} "{peer.name or "(unnamed)"}" '
        f'has cik="{peer.cik}", active={"TRUE" if peer.active else "FALSE"}. '
        + " ".join(reasons) +
        " Only CIK-keyed SEC registrants (the holding company) can be"
        " monitored; tickers, series names and name-only rows cannot. Find"
        f' the CIK with: python runner.py --lookup "{peer.name or "<name>"}"'
        " (resolves names via company_tickers.json).")


def metric_refusal_message(spec: MetricSpec, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: metric "{spec.id}" has '
        f'source_class="{spec.source_class}", '
        f'watchlist_capable={"TRUE" if spec.watchlist_capable else "FALSE"}. '
        + " ".join(reasons) +
        " This template admits only the public keyless EDGAR adapter"
        " (Class A); any other class requires its own review before"
        " admission.")


def gate_peer_row(peer: PeerRow) -> List[str]:
    reasons = []
    if not re.match(ENTITY_KEY_PATTERN, peer.entity_key):
        reasons.append(
            f"Gate3: entity key '{peer.entity_key}' does not match the "
            f"join-key whitelist pattern {ENTITY_KEY_PATTERN} (default-deny; "
            "a blank or malformed CIK cannot be fetched).")
    return reasons


def gate_metric_spec(spec: MetricSpec) -> List[str]:
    reasons = []
    if not spec.watchlist_capable:
        reasons.append("Gate1: watchlist_capable is not TRUE.")
    if spec.source_class != "A":
        reasons.append("Gate2: source_class is not A -- the only admitted "
                       "class is the keyless public EDGAR adapter; any other "
                       "class (including a future licensed swap) requires "
                       "its own review before admission.")
    return reasons


def assert_metrics_admissible(specs: Sequence[MetricSpec]) -> None:
    for s in specs:
        reasons = gate_metric_spec(s)
        if reasons:
            raise WatchlistRefused(metric_refusal_message(s, reasons))


def evaluate_peers(cfg: Config):
    """(admitted, refusals, excluded): admitted peers are fetched and landed;
    refusals is [(peer, message)] -- those slots are never fetched and are
    blanked; excluded is the active=FALSE list (blanked, not refused)."""
    admitted, refusals, excluded = [], [], []
    for peer in cfg.peers:
        if not peer.has_bank:
            continue
        if not peer.active:
            excluded.append(peer)
            continue
        reasons = gate_peer_row(peer)
        if reasons:
            refusals.append((peer, peer_refusal_message(peer, reasons)))
        else:
            admitted.append(peer)
    return admitted, refusals, excluded


def assert_entity_gates(cfg: Config) -> None:
    """BUILD-TIME hard gate backing the runtime gates (defense in depth)."""
    assert_metrics_admissible(METRIC_SPECS)
    _, refusals, _ = evaluate_peers(cfg)
    if refusals:
        raise WatchlistRefused(refusals[0][1])


def validate_peer_capacity(cfg: Config) -> None:
    cap = cfg.peer_slots
    banked = [p for p in cfg.peers if p.has_bank]
    seen: Dict[int, PeerRow] = {}
    bad = []
    for p in banked:
        if p.slot is None or p.slot < 1 or p.slot > cap:
            bad.append(p)
        elif p.slot in seen:
            raise PeerCapacityError(
                f"[PEERS] slot {p.slot} is listed twice "
                f'("{seen[p.slot].name}" and "{p.name}"); slots must be '
                "unique -- fix the [PEERS] table.")
        else:
            seen[p.slot] = p
    if bad:
        need = max(len(banked), max((p.slot or 0) for p in banked))
        names = ", ".join(f'"{p.name or p.cik}" (slot {p.slot})' for p in bad)
        raise PeerCapacityError(
            f"[PEERS] lists more peers than the workbook was built for: "
            f"{names} outside the built capacity of {cap} slots. The runner "
            f"never truncates a peer list -- rebuild with "
            f"--peer-slots {max(need, cap + 1)} "
            f"(python make_workbook.py --peer-slots {max(need, cap + 1)}).")


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT, direction-aware.
# Thresholds are HEURISTIC and labeled so in _config (BUILD SPEC sec 3).
# ---------------------------------------------------------------------------
def status_for(value: Optional[float], thr: Optional[Threshold]) -> str:
    if value is None or thr is None or (isinstance(value, float)
                                        and math.isnan(value)):
        return "OK"
    above = thr.direction != "below"

    def hit(bound):
        if bound is None:
            return False
        return value >= bound if above else value <= bound

    if hit(thr.alert):
        return "ALERT"
    if hit(thr.watch):
        return "WATCH"
    return "OK"


# ---------------------------------------------------------------------------
# STALENESS GUARD (BUILD SPEC sec 3 -- late filer / delisted?)
# ---------------------------------------------------------------------------
def is_stale_bank(last_period: Optional[str], set_max_period: Optional[str],
                  stale_multiplier: float) -> bool:
    if last_period is None:
        return True
    if set_max_period is None:
        return False
    try:
        last = date.fromisoformat(str(last_period)[:10])
        mx = date.fromisoformat(str(set_max_period)[:10])
    except ValueError:
        return True
    return (mx - last).days > stale_multiplier * QUARTER_DAYS


# ---------------------------------------------------------------------------
# THE PROVIDER SEAM -- normalized facts + events
# ---------------------------------------------------------------------------
@dataclass
class Fact:
    """One dimensional CQI fact: amortized cost for (grade member x class
    member) at a quarter end, with its accession + acceptance timestamp
    (provenance, contract sec 12). Member qnames stay NAMESPACE-QUALIFIED
    (extensions matter -- trap E4)."""
    cik: str
    period: str            # ISO quarter-end (the filing's reportDate)
    tag: str               # TAG_PRIMARY or TAG_EXCL (trap E2)
    class_qname: str       # "" = grade total across classes
    grade_qname: str
    value: float           # USD
    accession: str
    acceptance: str
    form: str
    source_class: str = "A"


@dataclass
class Event:
    """One harvested 8-K credit-event item."""
    cik: str
    date: str              # filingDate ISO
    item: str
    form: str
    accession: str
    source_class: str = "A"


@dataclass
class NormalizedRow:
    """The contract sec 6 seam shape (id/period/value/geo_segment/
    source_class/units); canonical-metric rows are emitted through it by the
    engine for the digest."""
    id: str
    period: str
    value: Optional[float]
    geo_segment: str
    source_class: str
    units: str = "pct"


class Provider:
    source_class = "A"
    vintage: Optional[str] = None

    def prime(self, ciks: Sequence[str], asof: date,
              names: Optional[Dict[str, str]] = None) -> None:  # pragma: no cover
        raise NotImplementedError

    def fetch_facts(self, cik: str) -> List[Fact]:  # pragma: no cover
        raise NotImplementedError

    def fetch_events(self, cik: str) -> List[Event]:  # pragma: no cover
        raise NotImplementedError


# ---------------------------------------------------------------------------
# EdgarDemoProvider -- deterministic offline stand-in (BUILD SPEC 0.7).
# Values are FICTION seeded per (cik, class, grade, quarter): the same bank
# yields the same history in any slot. Families are honored: grades_full
# banks emit pass/SM/substandard/doubtful facts (Huntington emits its OLEM
# extension member instead of the standard SM member -- exercises the
# extension mapping); criticized_only banks emit pass + criticized-accruing/
# nonaccruing extension members; ig_nig banks emit ONLY IG/NIG members so
# Tier 1 AND Tier 2 render N/A. One bank (Comerica) trips the
# criticized_ratio ALERT band; one (Zions) carries a 2.04 8-K (auto-WATCH);
# one unknown extension member appears for exactly one bank (Fifth Third) so
# the bootstrap/unmapped path is exercised. Accessions are deterministic
# ("0000035527-26-000001" style). NO network, NO key.
# ---------------------------------------------------------------------------
DEMO_FAMILY = {
    "35527": "grades_full", "1281761": "grades_full", "109380": "grades_full",
    "28412": "grades_full", "49196": "grades_full", "92230": "grades_full",
    "875357": "grades_full", "91576": "criticized_only",
    "36270": "criticized_only", "36104": "ig_nig",
}
DEMO_ALERT_CIK = "28412"       # Comerica: criticized_ratio ALERT tier
DEMO_WATCH_CIK = "109380"      # Zions: WATCH tier + the 2.04 8-K
DEMO_EXTENSION_CIK = "35527"   # Fifth Third: emits one UNSEEDED ext member
DEMO_EXTENSION_MEMBER = "fitb:CriticizedRestructuredMember"

DEMO_CLASS_MEMBERS = [
    ("us-gaap:CommercialPortfolioSegmentMember", 0.42),
    ("us-gaap:CommercialRealEstatePortfolioSegmentMember", 0.34),
    ("us-gaap:ConstructionLoansMember", 0.10),
    ("us-gaap:CommercialLoanMember", 0.14),
]
DEMO_CO_MEMBERS = {            # criticized_only extension qnames per bank
    "91576": ("key:CriticizedAccruingMember", "key:CriticizedNonaccrualMember"),
    "36270": ("mtb:CriticizedAccruingMember", "mtb:CriticizedNonaccrualMember"),
}


class EdgarDemoProvider(Provider):
    source_class = "A"

    def __init__(self, asof: Optional[date] = None,
                 raw_slots: int = RAW_SLOTS_DEFAULT):
        self.asof = asof or date(2026, 3, 31)
        self.raw_slots = int(raw_slots)
        self.vintage = "demo data (EdgarDemoProvider -- no live vintage)"
        self.names: Dict[str, str] = {}
        self._facts: Dict[str, List[Fact]] = {}
        self._events: Dict[str, List[Event]] = {}

    def prime(self, ciks, asof=None, names=None):
        self.names.update(names or {})
        for c in ciks:
            self._profile(str(c))

    @staticmethod
    def _seed(cik: str) -> int:
        try:
            n = int(cik)
        except ValueError:
            n = sum((i + 1) * ord(c) for i, c in enumerate(cik))
        return (n * 2654435761) % (2 ** 32)

    def _periods(self) -> List[str]:
        idx = pd.period_range(end=pd.Period(self.asof, freq="Q"),
                              periods=self.raw_slots, freq="Q")
        return [p.to_timestamp(how="end").date().isoformat() for p in idx]

    @staticmethod
    def family_for(cik: str) -> str:
        c = str(int(cik)) if str(cik).isdigit() else str(cik)
        if c in DEMO_FAMILY:
            return DEMO_FAMILY[c]
        return FAMILIES[EdgarDemoProvider._seed(c) % 3]

    def _grade_members(self, cik: str):
        """[(grade_qname, criticized_share_weight)] emitted per family.
        Weights split the bank's criticized total across its members."""
        fam = self.family_for(cik)
        if fam == "criticized_only":
            acc, nac = DEMO_CO_MEMBERS.get(
                cik, ("ext:CriticizedAccruingMember",
                      "ext:CriticizedNonaccrualMember"))
            return fam, [("us-gaap:PassMember", None),
                         (acc, 0.7), (nac, 0.3)]
        if fam == "ig_nig":
            return fam, [("us-gaap:InternalInvestmentGradeMember", None),
                         ("us-gaap:InternalNoninvestmentGradeMember", None)]
        sm_q = ("hban:OlemMember" if cik == "49196"
                else "us-gaap:SpecialMentionMember")
        members = [("us-gaap:PassMember", None),
                   (sm_q, 0.45),
                   ("us-gaap:SubstandardMember", 0.45),
                   ("us-gaap:DoubtfulMember", 0.10)]
        return fam, members

    def _crit_target(self, cik: str, qi: int, n: int) -> float:
        """Criticized ratio (pct) at quarter index qi (0=oldest)."""
        s = self._seed(cik)
        ramp = 0.75 + 0.25 * (qi / max(1, n - 1))
        wob = 1.0 + 0.05 * math.sin((qi + s % 7) / 1.9)
        if cik == DEMO_ALERT_CIK:
            base = 7.2
        elif cik == DEMO_WATCH_CIK:
            base = 4.6
        else:
            base = 1.6 + (s % 9) * 0.22          # 1.6 - 3.4, under WATCH
        return base * ramp * wob

    def _profile(self, cik: str):
        if cik in self._facts:
            return
        s = self._seed(cik)
        fam, members = self._grade_members(cik)
        periods = self._periods()
        n = len(periods)
        total0 = 20.0e9 + (s % 60) * 1.0e9       # $20B - $80B commercial book
        facts: List[Fact] = []
        for qi, iso in enumerate(periods):       # qi=0 oldest .. n-1 newest
            q_cal = (int(iso[5:7]) + 2) // 3
            form = "10-K" if q_cal == 4 else "10-Q"   # Q4 via the 10-K (E6)
            acc = f"{int(cik):010d}-26-{qi + 1:06d}"
            acceptance = (f"{iso[:4]}-{int(iso[5:7]):02d}-15"
                          f"T16:0{qi % 6}:00.000Z")
            total = total0 * (1.012 ** qi)
            crit_r = self._crit_target(cik, qi, n) / 100.0
            for ci_, (cq, share) in enumerate(DEMO_CLASS_MEMBERS):
                cls_total = total * share
                # per-class criticized intensity varies by class (CRE hotter)
                mult = (0.8, 1.4, 1.3, 0.7)[ci_]
                cls_crit = cls_total * crit_r * mult
                if fam == "ig_nig":
                    nig = cls_total * (0.12 + (s % 5) * 0.01)
                    facts.append(Fact(cik, iso, TAG_EXCL, cq,
                                      "us-gaap:InternalInvestmentGradeMember",
                                      round(cls_total - nig, -3), acc,
                                      acceptance, form))
                    facts.append(Fact(cik, iso, TAG_EXCL, cq,
                                      "us-gaap:InternalNoninvestmentGradeMember",
                                      round(nig, -3), acc, acceptance, form))
                    continue
                crit_left = cls_crit
                for gq, w in members:
                    if w is None:
                        continue
                    v = cls_crit * w
                    crit_left -= v
                    facts.append(Fact(cik, iso, TAG_EXCL, cq, gq,
                                      round(v, -3), acc, acceptance, form))
                facts.append(Fact(cik, iso, TAG_EXCL, cq,
                                  "us-gaap:PassMember",
                                  round(cls_total - cls_crit, -3), acc,
                                  acceptance, form))
            # the one unknown extension member (bootstrap path, trap E4):
            # a small extra criticized bucket on the C&I class, all quarters
            if cik == DEMO_EXTENSION_CIK:
                facts.append(Fact(cik, iso, TAG_EXCL,
                                  DEMO_CLASS_MEMBERS[0][0],
                                  DEMO_EXTENSION_MEMBER,
                                  round(total * 0.0011, -3), acc,
                                  acceptance, form))
        self._facts[cik] = facts
        # events: every bank files a 2.02 (informational); Zions adds the
        # 2.04 acceleration (auto-WATCH lane); the ALERT bank adds a 2.06.
        latest = periods[-1]
        yr = latest[:4]
        events = [Event(cik, f"{yr}-01-21", "2.02", "8-K",
                        f"{int(cik):010d}-26-000101")]
        if cik == DEMO_WATCH_CIK:
            events.append(Event(cik, f"{yr}-02-11", "2.04", "8-K",
                                f"{int(cik):010d}-26-000102"))
        if cik == DEMO_ALERT_CIK:
            events.append(Event(cik, f"{yr}-02-25", "2.06", "8-K",
                                f"{int(cik):010d}-26-000103"))
        self._events[cik] = sorted(events, key=lambda e: e.date, reverse=True)

    def fetch_facts(self, cik: str) -> List[Fact]:
        self._profile(str(cik))
        return list(self._facts[str(cik)])

    def fetch_events(self, cik: str) -> List[Event]:
        self._profile(str(cik))
        return list(self._events[str(cik)])


# ---------------------------------------------------------------------------
# EdgarProvider -- LIVE, two-stage (BUILD SPEC sec 1).
#   Stage 1: data.sec.gov/submissions/CIK{10d}.json -- COLUMNAR arrays;
#            latest 10-Q/10-K per calendar quarter (Q4 = the 10-K); /A
#            amendments deduped by acceptanceDateTime; the 8-K `items` array
#            harvested from the SAME response (zero extra requests).
#   Stage 2: filing directory index.json -> the EDGAR-generated EXTRACTED
#            instance (*_htm.xml) -> stdlib XML parse: contexts
#            (xbrldi:explicitMember) + facts on the two CQI tags dimensioned
#            by InternalCreditAssessmentAxis x the class axis. Vintage line
#            items are IGNORED by exact-tag matching (trap E3); member qnames
#            stay namespace-qualified (trap E4).
# Keyless, but User-Agent "{org} {email}" REQUIRED (fail fast if blank).
# Throttle >= 0.15s/request; per-URL cache; retry/backoff on 429/5xx.
# ---------------------------------------------------------------------------
class EdgarProvider(Provider):
    source_class = "A"

    def __init__(self, user_agent: str, min_interval: float = 0.2,
                 max_retries: int = 4, raw_slots: int = RAW_SLOTS_DEFAULT,
                 asof: Optional[date] = None):
        ua = str(user_agent or "").strip()
        if not ua:
            raise UserAgentMissing(
                "edgar_user_agent is blank. SEC fair-access policy REQUIRES "
                "a declared User-Agent of the form \"{org} {email}\" on "
                "every request (an absent/generic UA earns silent HTTP 403s "
                "and a ~10-minute IP block). Open the workbook's _config tab "
                "and set edgar_user_agent, e.g. "
                "\"ExampleBank CreditRisk jane.doe@example.com\". "
                "No API key or account is needed.")
        self.user_agent = ua
        self._min_interval = max(0.15, float(min_interval))
        self._max_retries = int(max_retries)
        self._raw_slots = int(raw_slots)
        self.asof = asof or date.today()
        self._last = 0.0
        self._cache: Dict[str, bytes] = {}
        self._facts: Dict[str, List[Fact]] = {}
        self._events: Dict[str, List[Event]] = {}
        self.names: Dict[str, str] = {}
        self.errors: List[str] = []

    # -- transport ----------------------------------------------------------
    def _throttle(self):
        gap = time.time() - self._last
        if gap < self._min_interval:
            time.sleep(self._min_interval - gap)
        self._last = time.time()

    def _http_get(self, url: str) -> bytes:
        """The one real network call -- isolated so tests can stub it."""
        import urllib.request
        req = urllib.request.Request(url, headers={
            "User-Agent": self.user_agent,
            "Accept-Encoding": "identity"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.read()

    def _download(self, url: str, label: str) -> bytes:
        if url in self._cache:
            return self._cache[url]
        import urllib.error
        for attempt in range(self._max_retries + 1):
            self._throttle()
            try:
                data = self._http_get(url)
                self._cache[url] = data
                return data
            except urllib.error.HTTPError as exc:
                if exc.code == 403:
                    raise RuntimeError(
                        f"EDGAR fetch failed for {label}: HTTP 403. SEC "
                        "serves silent 403s for missing/generic User-Agent "
                        "headers -- check edgar_user_agent in _config "
                        "(\"{org} {email}\") and run --selftest. Corporate "
                        "proxies also block sec.gov (allowlist "
                        + ", ".join(SEC_HOSTS) + ").")
                if exc.code == 429 or exc.code >= 500:
                    if attempt < self._max_retries:
                        time.sleep(2.0 * (attempt + 1))
                        continue
                raise RuntimeError(f"EDGAR fetch failed for {label}: "
                                   f"HTTP {exc.code}")
            except Exception as exc:
                if attempt < self._max_retries:
                    time.sleep(2.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"EDGAR fetch failed for {label}: "
                                   f"{type(exc).__name__}: {exc}")

    # -- stage 1: submissions -------------------------------------------------
    @staticmethod
    def parse_submissions(payload: dict) -> List[dict]:
        """COLUMNAR parallel arrays -> row dicts (trap: the arrays are
        columns, not records). Accepts either the top-level payload (uses
        filings.recent) or an older filings.files page (arrays at top level)."""
        block = payload.get("filings", {}).get("recent", payload)
        forms = block.get("form", [])
        keys = ("accessionNumber", "filingDate", "reportDate",
                "acceptanceDateTime", "primaryDocument", "items")
        cols = {k: block.get(k, []) for k in keys}
        rows = []
        for i, form in enumerate(forms):
            row = {"form": form}
            for k in keys:
                v = cols[k][i] if i < len(cols[k]) else ""
                row[k] = "" if v is None else str(v)
            rows.append(row)
        return rows

    @staticmethod
    def select_quarterly(rows: List[dict], raw_slots: int,
                         asof: date) -> List[dict]:
        """Latest 10-Q/10-K per calendar quarter (Q4 = the 10-K -- trap E6);
        /A amendments dedupe by (period) keeping the LATEST
        acceptanceDateTime (trap E7). Newest-first, raw_slots deep."""
        per: Dict[str, dict] = {}
        for r in rows:
            if r.get("form") not in ("10-Q", "10-Q/A", "10-K", "10-K/A"):
                continue
            period = str(r.get("reportDate") or "")[:10]
            if not period or period > asof.isoformat():
                continue
            prev = per.get(period)
            if prev is None or (str(r.get("acceptanceDateTime") or "")
                                > str(prev.get("acceptanceDateTime") or "")):
                per[period] = r
        chosen = [dict(per[p], reportDate=p)
                  for p in sorted(per, reverse=True)[:raw_slots]]
        return chosen

    @staticmethod
    def harvest_events(rows: List[dict], cik: str, asof: date) -> List[Event]:
        """The 8-K credit-event lane, from the SAME submissions response
        (zero extra requests): `items` filtered to {1.03, 2.04, 2.06, 4.02}
        (+2.02 informational), trailing ~4 quarters."""
        floor = (asof - timedelta(days=EVENT_WINDOW_DAYS)).isoformat()
        events = []
        for r in rows:
            if r.get("form") not in ("8-K", "8-K/A"):
                continue
            d = str(r.get("filingDate") or "")[:10]
            if not d or d < floor or d > asof.isoformat():
                continue
            items = [i.strip() for i in str(r.get("items") or "").split(",")]
            for it in items:
                if it in HARVEST_ITEMS:
                    events.append(Event(str(cik), d, it, r["form"],
                                        r.get("accessionNumber", "")))
        return sorted(events, key=lambda e: (e.date, e.item), reverse=True)

    # -- stage 2: extracted-instance XML -------------------------------------
    @staticmethod
    def parse_instance(data: bytes, cik: str, period: str, accession: str,
                       acceptance: str, form: str) -> List[Fact]:
        """Parse an EDGAR-generated extracted XBRL instance (*_htm.xml) with
        stdlib XML. Keeps facts on the two CQI tags whose context (a) has an
        xbrldi:explicitMember on InternalCreditAssessmentAxis, (b) optionally
        one on the class axis, (c) NO other axes (geography etc. would
        double-count), and (d) an instant equal to the filing's reportDate
        (current period only -- comparative columns dropped). Vintage
        disclosures are LINE ITEMS, not an axis (research-verified): exact
        tag matching ignores them entirely (trap E3). Member qnames are kept
        namespace-qualified with a canonical us-gaap prefix; extension
        prefixes survive as-published (trap E4)."""
        import xml.etree.ElementTree as ET
        from io import BytesIO

        ns: Dict[str, str] = {}
        for _ev, (prefix, uri) in ET.iterparse(BytesIO(data),
                                               events=("start-ns",)):
            ns.setdefault(prefix, uri)
        root = ET.fromstring(data)

        def canon(qname_text: str) -> Tuple[str, str]:
            """document QName text -> (canonical_prefix, localname)."""
            t = (qname_text or "").strip()
            if ":" in t:
                pfx, local = t.split(":", 1)
            else:
                pfx, local = "", t
            uri = ns.get(pfx, "")
            if US_GAAP_URI_HINT in uri:
                return "us-gaap", local
            return (pfx or "ext"), local

        def local(tag: str) -> str:
            return tag.rsplit("}", 1)[-1]

        def tag_uri(tag: str) -> str:
            return tag[1:].split("}", 1)[0] if tag.startswith("{") else ""

        # contexts: id -> (instant, {axis_local: member_qname}) or None when
        # the context carries axes we do not accept.
        contexts: Dict[str, Optional[Tuple[str, Dict[str, str]]]] = {}
        for el in root:
            if local(el.tag) != "context":
                continue
            cid = el.get("id", "")
            instant = ""
            dims: Dict[str, str] = {}
            ok = True
            for sub in el.iter():
                l = local(sub.tag)
                if l == "instant":
                    instant = (sub.text or "").strip()
                elif l == "explicitMember":
                    ax_pfx, ax_local = canon(sub.get("dimension", ""))
                    m_pfx, m_local = canon((sub.text or ""))
                    member = f"{m_pfx}:{m_local}"
                    if ax_local == CQI_AXIS:
                        dims[CQI_AXIS] = member
                    elif ax_local in CLASS_AXES:
                        dims["CLASS"] = member
                    else:
                        ok = False        # unexpected extra axis -> drop
                elif l == "typedMember":
                    ok = False
            contexts[cid] = (instant, dims) if ok else None

        facts: List[Fact] = []
        for el in root:
            l = local(el.tag)
            if l not in CQI_TAGS:
                continue
            if US_GAAP_URI_HINT not in tag_uri(el.tag):
                continue                   # extension re-declaration -- skip
            ctx = contexts.get(el.get("contextRef", ""))
            if not ctx:
                continue
            instant, dims = ctx
            if CQI_AXIS not in dims:
                continue                   # undimensioned totals: not CQI
            if instant[:10] != str(period)[:10]:
                continue                   # comparative prior column
            txt = (el.text or "").strip()
            if not txt or el.get("{http://www.w3.org/2001/XMLSchema-instance}"
                                 "nil") == "true":
                continue
            try:
                v = float(txt)
            except ValueError:
                continue
            facts.append(Fact(str(cik), str(period)[:10], l,
                              dims.get("CLASS", ""), dims[CQI_AXIS], v,
                              accession, acceptance, form))
        return facts

    # -- orchestration --------------------------------------------------------
    def _submission_rows(self, cik: str) -> List[dict]:
        url = SUBMISSIONS_URL.format(cik10=cik10(cik))
        payload = json.loads(self._download(url, f"submissions cik:{cik}")
                             .decode("utf-8"))
        self.names[str(cik)] = payload.get("name", "") or ""
        rows = self.parse_submissions(payload)
        # older history via filings.files[] when the recent window is shallow
        oldest_needed = (self.asof
                         - timedelta(days=(self._raw_slots + 2) * QUARTER_DAYS))
        have_oldest = min((r.get("filingDate") or "9999" for r in rows),
                          default="9999")
        for extra in payload.get("filings", {}).get("files", []):
            if have_oldest <= oldest_needed.isoformat():
                break
            name = extra.get("name")
            if not name:
                continue
            more = json.loads(self._download(
                "https://data.sec.gov/submissions/" + name,
                f"submissions page {name}").decode("utf-8"))
            page_rows = self.parse_submissions(more)
            rows.extend(page_rows)
            have_oldest = min((r.get("filingDate") or "9999"
                               for r in page_rows), default=have_oldest)
        return rows

    def _instance_url(self, cik: str, accession: str,
                      primary_doc: str) -> Optional[str]:
        base = filing_dir_url(cik, accession)
        idx = json.loads(self._download(base + "index.json",
                                        f"index cik:{cik} {accession}")
                         .decode("utf-8"))
        items = (idx.get("directory") or {}).get("item", [])
        names = [it.get("name", "") for it in items]
        stem = os.path.splitext(str(primary_doc or "").split("/")[-1])[0]
        cands = [n for n in names if n.endswith("_htm.xml")]
        if not cands:
            return None
        for n in cands:
            if stem and n.startswith(stem):
                return base + n
        return base + cands[0]

    def prime(self, ciks, asof=None, names=None):
        if names:
            for c, n in names.items():
                self.names.setdefault(str(c), n)
        for cik in [str(c) for c in ciks if str(c)]:
            rows = self._submission_rows(cik)
            self._events[cik] = self.harvest_events(rows, cik, self.asof)
            facts: List[Fact] = []
            for filing in self.select_quarterly(rows, self._raw_slots,
                                                self.asof):
                acc = filing.get("accessionNumber", "")
                try:
                    url = self._instance_url(cik, acc,
                                             filing.get("primaryDocument", ""))
                    if url is None:
                        self.errors.append(
                            f"cik:{cik} {acc}: no extracted instance "
                            "(*_htm.xml) in the filing directory -- pre-iXBRL "
                            "filing? (EX-101.INS backfill is v1.1)")
                        continue
                    data = self._download(url, f"instance cik:{cik} {acc}")
                except RuntimeError as exc:
                    self.errors.append(str(exc))
                    continue
                facts.extend(self.parse_instance(
                    data, cik, filing.get("reportDate", ""), acc,
                    filing.get("acceptanceDateTime", ""),
                    filing.get("form", "")))
            self._facts[cik] = facts
        self.vintage = (f"EDGAR submissions as of {self.asof.isoformat()} "
                        f"(accessions recorded per fact)")

    def fetch_facts(self, cik: str) -> List[Fact]:
        return list(self._facts.get(str(cik), []))

    def fetch_events(self, cik: str) -> List[Event]:
        return list(self._events.get(str(cik), []))

    # -- helpers exposed to the CLI -------------------------------------------
    def lookup(self, name: str) -> List[dict]:
        payload = json.loads(self._download(TICKERS_URL, "company_tickers")
                             .decode("utf-8"))
        needle = name.strip().lower()
        hits = []
        for rec in payload.values():
            title = str(rec.get("title", ""))
            ticker = str(rec.get("ticker", ""))
            if needle in title.lower() or needle == ticker.lower():
                hits.append({"cik": str(rec.get("cik_str", "")),
                             "ticker": ticker, "title": title})
        return hits[:15]


def resolve_secret(cfg: Config) -> Optional[str]:
    """Contract plumbing: the NAME of a licensed (Class C) secret env var.
    Unused in v1 -- EDGAR is keyless and the gates admit only Class A."""
    name = str(cfg.setting("secret_env", "") or "").strip()
    if not name:
        return None
    return os.environ.get(name)


def make_provider(cfg: Config, demo: bool, asof: Optional[date]) -> Provider:
    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        return EdgarDemoProvider(asof=asof, raw_slots=cfg.raw_slots)
    return EdgarProvider(
        user_agent=cfg.user_agent,
        min_interval=float(cfg.setting("http_min_interval", 0.2) or 0.2),
        max_retries=int(float(cfg.setting("edgar_max_retries", 4) or 4)),
        raw_slots=cfg.raw_slots, asof=asof)


# ---------------------------------------------------------------------------
# THE METRIC ENGINE -- facts + [MEMBER_MAP]/[CLASS_MAP] -> canonical quarterly
# rollups (BUILD SPEC sec 3). Family-gated Tier 2; unmapped members surface,
# never guessed; the values landed here ARE what the dashboards reference.
# ---------------------------------------------------------------------------
@dataclass
class QuarterRollup:
    period: str
    tag: str
    accession: str
    values: Dict[str, Optional[float]]      # keyed by RAW_METRIC_COLS ids
    facts: List[Tuple[Fact, str, str]]      # (fact, class, grade) annotated


def _pick_tag(qfacts: List[Fact]) -> str:
    """Per-bank accrued-interest twin preference (trap E2): the tag the bank
    actually populates wins; a tie prefers the ExcludingAccruedInterest twin
    (the post-2022 disclosure style)."""
    n_ex = sum(1 for f in qfacts if f.tag == TAG_EXCL)
    n_pr = sum(1 for f in qfacts if f.tag == TAG_PRIMARY)
    return TAG_EXCL if n_ex >= n_pr else TAG_PRIMARY


def rollup_bank(cik: str, facts: List[Fact], cfg: Config
                ) -> Tuple[List[QuarterRollup], Set[str], Set[str]]:
    """All quarterly rollups for one bank, newest-first, plus the sets of
    UNMAPPED grade member qnames and UNMAPPED class member qnames observed
    (the bootstrap feed -- contract sec 13)."""
    fam = cfg.family_of(cik)
    unmapped_members: Set[str] = set()
    unmapped_classes: Set[str] = set()
    by_period: Dict[str, List[Fact]] = {}
    for f in facts:
        by_period.setdefault(f.period, []).append(f)

    quarters: List[QuarterRollup] = []
    for period in sorted(by_period, reverse=True):
        qfacts = by_period[period]
        tag = _pick_tag(qfacts)
        qfacts = [f for f in qfacts if f.tag == tag]
        per_class = [f for f in qfacts if f.class_qname]
        use = per_class if per_class else qfacts   # totals-only fallback
        annotated: List[Tuple[Fact, str, str]] = []
        total = crit = classified = sm = 0.0
        crit_by_class: Dict[str, float] = {}
        for f in use:
            grade = cfg.grade_of(cik, f.grade_qname)
            if grade is None:
                grade = "unmapped"
            if grade == "unmapped":
                # both never-seen members AND members already bootstrapped
                # as unmapped stay flagged until manually mapped
                unmapped_members.add(f.grade_qname)
            if f.class_qname:
                klass = cfg.class_of(cik, f.class_qname)
                if klass is None:
                    klass = "unmapped"
                if klass == "unmapped":
                    unmapped_classes.add(f.class_qname)
            else:
                klass = "total"
            annotated.append((f, klass, grade))
            if klass in ("excluded", "unmapped"):
                continue                 # consumer books / unknowable classes
            if grade == "ignore":
                continue                 # filer-tagged subtotals (E4 note)
            # unmapped grade members ARE part of the amortized-cost total
            # (a fact, not a guess) but NEVER enter a numerator; the bank is
            # flagged "needs manual mapping" in the digest.
            total += f.value
            if grade in CRITICIZED_GRADES:
                crit += f.value
                crit_by_class[klass] = crit_by_class.get(klass, 0.0) + f.value
            if grade in CLASSIFIED_GRADES:
                classified += f.value
            if grade == "special_mention":
                sm += f.value

        v: Dict[str, Optional[float]] = {c: None for c in RAW_METRIC_COLS}
        v["TOTAL_COMM_USD"] = total if total else None
        crit_capable = fam in ("grades_full", "criticized_only")
        if crit_capable:
            v["CRIT_RATIO"] = _ratio(crit, total)
            v["CRIT_USD"] = crit if total else None
            # class mix only when the bank actually dimensions by class --
            # a totals-only quarter leaves the mix BLANK, never 0
            if crit > 0 and per_class:
                v["MIX_CI"] = _ratio(crit_by_class.get("ci", 0.0), crit)
                v["MIX_CRE"] = _ratio(crit_by_class.get("cre", 0.0), crit)
                v["MIX_CONSTR"] = _ratio(
                    crit_by_class.get("construction", 0.0), crit)
                v["MIX_OTHER"] = _ratio(
                    crit_by_class.get("other_commercial", 0.0), crit)
        if fam == "grades_full":         # Tier 2, flag-gated (BUILD SPEC 0.6)
            v["CLASS_RATIO"] = _ratio(classified, total)
            v["CLASS_USD"] = classified if total else None
            v["SM_RATIO"] = _ratio(sm, total)
            v["SM_USD"] = sm if total else None
        quarters.append(QuarterRollup(period, tag,
                                      use[0].accession if use else "",
                                      v, annotated))

    # deltas off the newest-first series (pp; growth in %)
    for i, q in enumerate(quarters):
        r0 = q.values["CRIT_RATIO"]
        r1 = (quarters[i + 1].values["CRIT_RATIO"]
              if i + 1 < len(quarters) else None)
        r4 = (quarters[i + 4].values["CRIT_RATIO"]
              if i + 4 < len(quarters) else None)
        q.values["CRIT_QOQ"] = (None if r0 is None or r1 is None
                                else r0 - r1)
        q.values["CRIT_YOY"] = (None if r0 is None or r4 is None
                                else r0 - r4)
        c0 = q.values["CRIT_USD"]
        c1 = (quarters[i + 1].values["CRIT_USD"]
              if i + 1 < len(quarters) else None)
        q.values["CRIT_GROWTH"] = (None if not c0 or not c1
                                   else (c0 / c1 - 1.0) * 100.0)
    return quarters, unmapped_members, unmapped_classes


def metric_rows_for_digest(slot: int, cik: str,
                           quarters: List[QuarterRollup]
                           ) -> List[NormalizedRow]:
    """Contract sec 6 seam shape: canonical metric NormalizedRows."""
    rows = []
    for q in quarters:
        for spec in METRIC_SPECS:
            rows.append(NormalizedRow(
                id=f"s{slot:02d}_{spec.id}", period=q.period,
                value=q.values[spec.raw_col], geo_segment=f"cik:{cik}",
                source_class="A"))
    return rows


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder.
# ---------------------------------------------------------------------------
@dataclass
class SlotBlock:
    slot: int
    header_row: int
    label_row: int
    first_data_row: int        # first canonical-metric quarter row
    facts_label_row: int
    first_fact_row: int
    events_label_row: int
    first_event_row: int
    slots: int                 # quarter rows (raw_slots)


def slot_stride(raw_slots: int) -> int:
    return 2 + raw_slots + 1 + FACT_ROWS + 1 + EVENT_ROWS + 2


def slot_block(slot: int, raw_slots: int = RAW_SLOTS_DEFAULT) -> SlotBlock:
    header_row = RAW_FIRST_ROW + (slot - 1) * slot_stride(raw_slots)
    label_row = header_row + 1
    first_data_row = label_row + 1
    facts_label_row = first_data_row + raw_slots
    first_fact_row = facts_label_row + 1
    events_label_row = first_fact_row + FACT_ROWS
    first_event_row = events_label_row + 1
    return SlotBlock(slot, header_row, label_row, first_data_row,
                     facts_label_row, first_fact_row, events_label_row,
                     first_event_row, raw_slots)


# ---------------------------------------------------------------------------
# WRITE BACKEND -- openpyxl on the closed workbook (contract sec 4 / L2)
# ---------------------------------------------------------------------------
class OpenpyxlBackend:
    def __init__(self, path: str):
        self.path = path
        import openpyxl
        # L2: keep_vba=True ONLY for .xlsm.
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        return parse_config([[c.value for c in row] for row in ws.iter_rows()])

    def _check_slot(self, block: SlotBlock):
        """Refuse to write into a workbook whose raw layout doesn't match the
        computed one (raw_slots/peer_slots edited after build; formulas are
        anchored to the BUILT layout). The slot label is constant regardless
        of which bank occupies the slot, so [PEERS] edits never trip this."""
        ws = self._wb[RAW_TAB]
        existing = ws.cell(block.header_row, 1).value
        if existing is not None and str(existing).strip() not in (
                "", slot_label(block.slot)):
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB} at row {block.header_row}: "
                f"expected block '{slot_label(block.slot)}' but found "
                f"'{existing}'. The workbook was built with a different "
                f"raw_slots/peer_slots layout than _config now describes; "
                f"formulas are anchored to the built layout. Rebuild the "
                f"workbook (python make_workbook.py --peer-slots N) instead "
                f"of editing the layout settings in place.")
        lbl = ws.cell(block.label_row, metric_col(RAW_METRIC_COLS[-1])).value
        if str(lbl or "").strip() != RAW_METRIC_COLS[-1]:
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB}: block "
                f"'{slot_label(block.slot)}' does not label metric column "
                f"'{RAW_METRIC_COLS[-1]}' at column "
                f"{metric_col(RAW_METRIC_COLS[-1])} (found '{lbl}'). The "
                f"workbook was built by a different pack than this runner "
                f"(pack {PACK_VERSION}); rebuild it (python make_workbook.py) "
                f"or extract the matching runner from ITS OWN _code_py tab.")

    def clear_slot_block(self, block: SlotBlock):
        """Blank a slot's bank identity AND data (stateless full-replace,
        BUILD SPEC 0.2). NOTE (L7): openpyxl's ws.cell(r, c, None) silently
        IGNORES None -- cells are blanked ONLY by assigning .value."""
        self._check_slot(block)
        ws = self._wb[RAW_TAB]
        for c in (2, 3, 4, 5, 6):                # runtime identity cells
            ws.cell(block.header_row, c).value = None
        last_metric_col = 1 + len(RAW_METRIC_COLS)
        for r in range(block.first_data_row,
                       block.first_data_row + block.slots):
            for c in range(1, last_metric_col + 1):
                ws.cell(r, c).value = None
        for r in range(block.first_fact_row,
                       block.first_fact_row + FACT_ROWS):
            for c in range(1, len(FACT_COLS) + 1):
                ws.cell(r, c).value = None
        for r in range(block.first_event_row,
                       block.first_event_row + EVENT_ROWS):
            for c in range(1, len(EVENT_COLS) + 1):
                ws.cell(r, c).value = None

    def write_slot_block(self, block: SlotBlock, peer: PeerRow, family: str,
                         quarters: List[QuarterRollup],
                         events: List[Event]):
        self._check_slot(block)
        ws = self._wb[RAW_TAB]
        ws.cell(block.header_row, 1, slot_label(block.slot))
        ws.cell(block.header_row, 2, f"s{block.slot:02d} {peer.entity_key}")
        ws.cell(block.header_row, 3, peer.name)
        ws.cell(block.header_row, 4, family)
        latest_acc = quarters[0].accession if quarters else ""
        ws.cell(block.header_row, 5, latest_acc)
        ws.cell(block.header_row, 6, quarters[0].tag if quarters else "")
        # canonical metric rows, newest-first (dashboards read these)
        for i, q in enumerate(quarters[:block.slots]):
            r = block.first_data_row + i
            ws.cell(r, 1, q.period)
            for cname in RAW_METRIC_COLS:
                cell = ws.cell(r, metric_col(cname))
                if cname == "ACCESSION":
                    cell.value = q.accession
                elif cname == "TAG":
                    cell.value = q.tag
                else:
                    v = q.values.get(cname)
                    # L7: explicit .value assignment so None LANDS as blank
                    cell.value = None if v is None else float(v)
        # latest-quarter member facts (the audit trail)
        latest_facts = quarters[0].facts if quarters else []
        shown = latest_facts[:FACT_ROWS]
        for i, (f, klass, grade) in enumerate(shown):
            r = block.first_fact_row + i
            ws.cell(r, fact_col("period")).value = f.period
            ws.cell(r, fact_col("class_qname")).value = f.class_qname
            ws.cell(r, fact_col("class")).value = klass
            ws.cell(r, fact_col("member_qname")).value = f.grade_qname
            ws.cell(r, fact_col("grade")).value = grade
            ws.cell(r, fact_col("value_usd")).value = float(f.value)
            ws.cell(r, fact_col("tag")).value = f.tag
            ws.cell(r, fact_col("accession")).value = f.accession
        if len(latest_facts) > FACT_ROWS:
            ws.cell(block.first_fact_row + FACT_ROWS - 1, 1,
                    f"(+{len(latest_facts) - FACT_ROWS + 1} more facts -- "
                    f"see --tieout)")
        # 8-K events, newest-first (already windowed to trailing 4 quarters)
        for i, e in enumerate(events[:EVENT_ROWS]):
            r = block.first_event_row + i
            ws.cell(r, event_col("date")).value = e.date
            ws.cell(r, event_col("item")).value = e.item
            ws.cell(r, event_col("form")).value = e.form
            ws.cell(r, event_col("accession")).value = e.accession
            ws.cell(r, event_col("url")).value = filing_dir_url(peer.cik,
                                                                e.accession)

    # -- [MEMBER_MAP]/[CLASS_MAP] bootstrap (contract sec 13) ----------------
    def _append_section_rows(self, section: str, rows: List[List],
                             ncols: int) -> List[List]:
        """Write rows into the PROVISIONED blank rows of a _config section
        (never inserts -- inserting would shift every threshold/peer cell
        the dashboards reference). Returns rows that did NOT fit."""
        ws = self._wb["_config"]
        start = None
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == section:
                start = r + 1
                break
        if start is None:
            return rows
        blanks = []
        for r in range(start, ws.max_row + 1):
            a = ws.cell(r, 1).value
            s = str(a or "").strip()
            if s.startswith("[") and s.endswith("]"):
                break
            if a is None or s == "":
                blanks.append(r)
        existing = set()
        for r in range(start, ws.max_row + 1):
            s = str(ws.cell(r, 1).value or "").strip()
            if s.startswith("[") and s.endswith("]"):
                break
            key = tuple(str(ws.cell(r, c).value or "").strip()
                        for c in range(1, ncols + 1))
            existing.add(key)
        leftover = []
        for row in rows:
            key = tuple(str(v) for v in row)
            if key in existing:
                continue
            if not blanks:
                leftover.append(row)
                continue
            r = blanks.pop(0)
            for c, v in enumerate(row, start=1):
                ws.cell(r, c).value = v
            existing.add(key)
        return leftover

    def append_member_map(self, rows: List[List]) -> List[List]:
        return self._append_section_rows("[MEMBER_MAP]", rows,
                                         len(MEMBER_MAP_HEADER))

    def append_class_map(self, rows: List[List]) -> List[List]:
        return self._append_section_rows("[CLASS_MAP]", rows,
                                         len(CLASS_MAP_HEADER))

    def write_status(self, status: dict):
        line2 = (f"Banks {status.get('banks_landed', 0)}/"
                 f"{status.get('banks_active', 0)} - "
                 f"{status.get('alert_banks', 0)} ALERT / "
                 f"{status.get('watch_banks', 0)} WATCH / "
                 f"{status.get('stale_banks', 0)} STALE")
        n_ref = len(status.get("watchlist_refusals") or [])
        if n_ref:
            line2 += f" - {n_ref} REFUSED (see digest)"
        n_un = len(status.get("unmapped_members") or [])
        if n_un:
            line2 += f" - {n_un} UNMAPPED MEMBERS (see [MEMBER_MAP])"
        n_err = len(status.get("errors") or [])
        if n_err:
            line2 += f" - {n_err} FETCH ERRORS"
        line3 = (f"Vintage: {status.get('vintage') or 'n/a'} | latest "
                 f"accession: {status.get('latest_accession') or 'n/a'}")
        for tab in DASH_TABS:
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                col = STATUS_COL_BY_TAB.get(tab, STATUS_COL)
                ws.cell(1, col, "Last run  " + status.get("timestamp", "")
                        + f" ({status.get('mode', '')})")
                ws.cell(2, col, line2)
                ws.cell(3, col, line3)

    def finalize(self):
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def _events_summary(events: List[Event]) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    for e in events:
        d = out.setdefault(e.item, {"count": 0, "latest": "", "accessions": []})
        d["count"] += 1
        d["latest"] = max(d["latest"], e.date)
        d["accessions"].append(e.accession)
    return out


def compute_digest(cfg: Config,
                   landed: Dict[int, Tuple[PeerRow, str,
                                           List[QuarterRollup],
                                           List[Event], Set[str], Set[str]]]
                   ) -> dict:
    """Per-bank latest-quarter metrics + statuses + the 8-K lane + staleness
    + the unmapped-member bootstrap feed. Statuses mirror the Watchlist
    formulas EXACTLY (parity-tested). NOTE the documented v1 divergence: the
    EXCEL median row includes stale banks' landed values; the digest median
    excludes them -- see _readme."""
    per_bank_latest: Dict[int, Optional[str]] = {}
    for slot, (peer, fam, quarters, events, um, uc) in landed.items():
        per_bank_latest[slot] = quarters[0].period if quarters else None
    set_max = max((p for p in per_bank_latest.values() if p), default=None)

    banks = []
    for slot in sorted(landed):
        peer, fam, quarters, events, um, uc = landed[slot]
        latest = quarters[0] if quarters else None
        stale = is_stale_bank(per_bank_latest[slot], set_max,
                              cfg.stale_multiplier)
        metrics = {}
        alert_n = watch_n = 0
        for spec in METRIC_SPECS:
            v = latest.values.get(spec.raw_col) if latest else None
            st = status_for(v, cfg.thresholds.get(spec.id)) \
                if spec.id in BANDED_METRICS else "OK"
            if v is None:
                st = ""                    # blank cell -> blank status
            metrics[spec.id] = {"value": v, "status": st, "tier": spec.tier}
            if spec.id in BANDED_METRICS:
                if st == "ALERT":
                    alert_n += 1
                elif st == "WATCH":
                    watch_n += 1
        ev_sum = _events_summary(events)
        watch_8k = sum(ev_sum.get(i, {}).get("count", 0) for i in WATCH_ITEMS)
        notes = []
        if um:
            notes.append("UNMAPPED MEMBERS (needs manual mapping in "
                         "[MEMBER_MAP] -- never guessed): " + ", ".join(sorted(um)))
        if uc:
            notes.append("UNMAPPED CLASS MEMBERS (excluded from rollups "
                         "until mapped in [CLASS_MAP]): " + ", ".join(sorted(uc)))
        if fam == "criticized_only":
            notes.append("family=criticized_only: classified/SM ratios are "
                         "N/A (disclosure family) -- never approximated.")
        if fam == "ig_nig":
            notes.append("family=ig_nig: IG/NIG disclosure only -- "
                         "criticized metrics N/A; MD&A text fallback is "
                         "v1.1 (see _readme).")
        if watch_8k:
            worst = [i for i in WATCH_ITEMS if i in ev_sum]
            notes.append("8-K credit event(s) " + ", ".join(
                f"item {i} ({ITEM_LABEL[i]}) latest {ev_sum[i]['latest']}"
                for i in worst) + " -- automatic WATCH flag.")
        if stale:
            notes.append(f"latest CQI quarter lags the peer-set max by more "
                         f"than {cfg.stale_multiplier:g} quarters -- check "
                         "submissions: late filer or delisted?")
        crit_capable = fam in ("grades_full", "criticized_only")
        if stale:
            status = "STALE"
        elif not crit_capable:
            status = "WATCH" if watch_8k > 0 else "N/A"
        elif alert_n > 0:
            status = "ALERT"
        elif watch_n > 0 or watch_8k > 0:
            status = "WATCH"
        else:
            status = "OK"
        banks.append({
            "slot": slot, "id_prefix": f"s{slot:02d}", "cik": peer.cik,
            "name": peer.name, "ticker": peer.ticker, "family": fam,
            "entity_key": peer.entity_key,
            "asof_period": per_bank_latest[slot],
            "latest_accession": latest.accession if latest else None,
            "tag": latest.tag if latest else None,
            "stale": stale, "status": status,
            "alert_count": alert_n, "watch_count": watch_n,
            "watch_8k": watch_8k, "events": ev_sum,
            "criticized_ratio": metrics["criticized_ratio"]["value"],
            "metrics": metrics,
            "unmapped_members": sorted(um),
            "unmapped_classes": sorted(uc),
            "notes": notes,
        })

    # digest peer medians: criticized-capable AND non-stale banks only
    fresh_capable = [b for b in banks if not b["stale"]
                     and b["family"] in ("grades_full", "criticized_only")]
    medians = {}
    for spec in METRIC_SPECS:
        vals = sorted(b["metrics"][spec.id]["value"] for b in fresh_capable
                      if b["metrics"][spec.id]["value"] is not None)
        if not vals:
            medians[spec.id] = None
        else:
            k = len(vals)
            medians[spec.id] = (vals[k // 2] if k % 2 else
                                (vals[k // 2 - 1] + vals[k // 2]) / 2.0)
    return {"banks": banks, "medians": medians, "set_max_period": set_max}


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None,
        provider: Optional[Provider] = None) -> dict:
    """One refresh against the CLOSED workbook. `provider` is injectable for
    tests (e.g. the frozen-bank staleness regression)."""
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any fetch (BUILD SPEC 0.1 / sec 3).
    assert_metrics_admissible(METRIC_SPECS)
    validate_peer_capacity(cfg)
    admitted, refusals, excluded = evaluate_peers(cfg)

    if provider is None:
        provider = make_provider(cfg, demo, asof)   # fails fast on blank UA
    mode = "demo" if isinstance(provider, EdgarDemoProvider) else "live"

    # Stateless full-replace (BUILD SPEC 0.2): blank EVERY slot first --
    # deactivated/refused/emptied slots show blank under the fresh timestamp,
    # never last run's bank (L7). Also validates the raw layout up front.
    for slot in range(1, cfg.peer_slots + 1):
        backend.clear_slot_block(slot_block(slot, cfg.raw_slots))

    errors: List[str] = []
    landed: Dict[int, Tuple] = {}
    try:
        provider.prime([p.cik for p in admitted], asof,
                       names={p.cik: p.name for p in admitted})
    except Exception as exc:
        errors.append(f"prime: {exc}")
    else:
        errors.extend(getattr(provider, "errors", []) or [])
        for peer in admitted:
            try:
                facts = provider.fetch_facts(peer.cik)
                events = provider.fetch_events(peer.cik)
                quarters, um, uc = rollup_bank(peer.cik, facts, cfg)
            except Exception as exc:
                errors.append(f"s{peer.slot:02d} cik:{peer.cik}: {exc}")
                continue
            fam = cfg.family_of(peer.cik)
            backend.write_slot_block(slot_block(peer.slot, cfg.raw_slots),
                                     peer, fam, quarters, events)
            landed[peer.slot] = (peer, fam, quarters, events, um, uc)

    # bootstrap: append newly observed members/classes into the visible
    # _config sections as unmapped (never guessed -- contract sec 13)
    member_rows, class_rows = [], []
    for slot, (peer, fam, quarters, events, um, uc) in landed.items():
        for q in sorted(um):
            member_rows.append([peer.cik, fam, q, "unmapped"])
        for q in sorted(uc):
            class_rows.append([peer.cik, q, "unmapped"])
    left_m = backend.append_member_map(member_rows)
    left_c = backend.append_class_map(class_rows)
    if left_m or left_c:
        errors.append(f"[MEMBER_MAP]/[CLASS_MAP] bootstrap area full: "
                      f"{len(left_m) + len(left_c)} row(s) not written -- "
                      "rebuild the workbook (python make_workbook.py) for "
                      "more provisioned rows.")

    digest = compute_digest(cfg, landed)
    fresh = [b for b in digest["banks"] if not b["stale"]]
    all_unmapped = sorted({q for b in digest["banks"]
                           for q in b["unmapped_members"]})
    latest_acc = next((b["latest_accession"] for b in digest["banks"]
                       if b["latest_accession"]), None)
    status = {
        "timestamp": asof.isoformat(),
        "mode": mode,
        "pack_version": PACK_VERSION,
        "peer_slots": cfg.peer_slots,
        "banks_active": len(admitted) + len(refusals),
        "banks_landed": len(landed),
        "banks_excluded": len(excluded),
        "alert_banks": sum(1 for b in fresh if b["status"] == "ALERT"),
        "watch_banks": sum(1 for b in fresh if b["status"] == "WATCH"),
        "stale_banks": sum(1 for b in digest["banks"] if b["stale"]),
        "na_banks": sum(1 for b in fresh if b["status"] == "N/A"),
        "alert_flags": sum(b["alert_count"] for b in fresh),
        "watch_flags": sum(b["watch_count"] for b in fresh),
        "watch_8k_events": sum(b["watch_8k"] for b in fresh),
        "unmapped_members": all_unmapped,
        "latest_accession": latest_acc,
        "vintage": provider.vintage,
        "digest": digest,
        "watchlist_refusals": [m for _, m in refusals],
        "watchlist_admitted": [f"s{p.slot:02d} {p.entity_key}"
                               for p in admitted],
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


# ---------------------------------------------------------------------------
# TIE-OUT MODE (contract sec 12) -- verify the feed against the OFFICIAL
# record: for one bank(-quarter), print each metric's value + tag/axis + the
# member-level breakdown + accession + the EDGAR viewer URLs.
# ---------------------------------------------------------------------------
def read_provenance_rows(wb) -> Dict[str, dict]:
    if "_provenance" not in wb.sheetnames:
        return {}
    ws = wb["_provenance"]
    rows: Dict[str, dict] = {}
    in_table = False
    for row in ws.iter_rows(values_only=True):
        a = ("" if not row or row[0] is None else str(row[0])).strip()
        if a.lower() == "metric":
            in_table = True
            continue
        if not in_table or not a or a.startswith("#") or a.startswith("["):
            continue
        cells = [("" if i >= len(row) or row[i] is None else str(row[i]))
                 for i in range(6)]
        rows[a] = {"metric": cells[0], "source": cells[1], "tag": cells[2],
                   "accession": cells[3], "flag": cells[4], "notes": cells[5]}
    return rows


def _parse_quarter_arg(arg: str) -> str:
    """Accept 2026-03-31 or 2026Q1 -> ISO quarter end."""
    a = str(arg).strip().upper()
    m = re.match(r"^(\d{4})Q([1-4])$", a)
    if m:
        p = pd.Period(f"{m.group(1)}Q{m.group(2)}", freq="Q")
        return p.to_timestamp(how="end").date().isoformat()
    return a[:10]


def do_tieout(workbook: str, args: List[str], demo: bool) -> int:
    """`--tieout CIK [quarter]`: value + tag + member breakdown + accession +
    the https://www.sec.gov/Archives/edgar/data/{cik}/{accession}/ URL.
    Works with --demo (values labeled loudly as demo fiction)."""
    import openpyxl
    cik = _norm_cik(args[0])
    quarter = _parse_quarter_arg(args[1]) if len(args) > 1 else None
    if not re.match(ENTITY_KEY_PATTERN, f"cik:{cik}"):
        sys.stderr.write(f"--tieout: '{args[0]}' is not a valid SEC CIK "
                         "(digits only). Find one with: python runner.py "
                         "--lookup \"<company name>\".\n")
        return 2
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=False)
    try:
        ws = wb["_config"]
        cfg = parse_config([[c for c in row]
                            for row in ws.iter_rows(values_only=True)])
        prov_rows = read_provenance_rows(wb)
    finally:
        wb.close()
    if not prov_rows:
        sys.stderr.write("--tieout: this workbook has no _provenance tab -- "
                         "rebuild it (python make_workbook.py).\n")
        return 1
    try:
        provider = make_provider(cfg, demo, None)
    except UserAgentMissing as exc:
        sys.stderr.write(f"--tieout: {exc}\n")
        return 3
    mode = "demo" if isinstance(provider, EdgarDemoProvider) else "live"
    try:
        provider.prime([cik], date.today(), names={cik: f"cik:{cik}"})
        facts = provider.fetch_facts(cik)
        events = provider.fetch_events(cik)
    except Exception as exc:
        sys.stderr.write(f"--tieout fetch failed for cik {cik}: {exc}\n")
        return 1
    quarters, um, uc = rollup_bank(cik, facts, cfg)
    if not quarters:
        sys.stderr.write(f"--tieout: no CQI quarters returned for cik "
                         f"{cik}.\n")
        return 1
    if quarter:
        match = [q for q in quarters if q.period == quarter]
        if not match:
            sys.stderr.write(
                f"--tieout: quarter {quarter} not in the fetched window "
                f"({quarters[-1].period} .. {quarters[0].period}).\n")
            return 1
        q = match[0]
    else:
        q = quarters[0]
    fam = cfg.family_of(cik)

    print(f"TIE-OUT  cik {cik}  quarter {q.period}  family {fam}  "
          f"(pack {PACK_VERSION}, {mode} mode)")
    if mode == "demo":
        print("  *** DEMO VALUES -- deterministic EdgarDemoProvider FICTION, "
              "NOT this company's filing; the note/tag/axis columns are the "
              "real tie-out map. Re-run without --demo (edgar_user_agent "
              "set) for live values. ***")
    print(f"  Filing directory : {filing_dir_url(cik, q.accession)}")
    print(f"  EDGAR filings    : {BROWSE_URL.format(cik=cik10(cik))}")
    print(f"  Accession        : {q.accession}   tag: {q.tag}")
    print(f"  Data vintage     : {provider.vintage or 'n/a'}")
    print("  Tie-out: open the filing directory, open the 10-Q/10-K R-file "
          "or the primary document's Credit Quality Indicators note, and "
          "match the member amounts below.")
    print()
    hdr = (f"  {'metric':<18} {'value':>14}  {'source document / note':<44} "
           f"flag")
    print(hdr)
    print("  " + "-" * (len(hdr) + 12))
    missing = []
    for spec in METRIC_SPECS:
        v = q.values.get(spec.raw_col)
        vs = "(N/A)" if v is None else f"{v:,.2f}"
        pr = prov_rows.get(spec.id)
        if pr is None:
            missing.append(spec.id)
            pr = {"source": "(no provenance row)", "flag": "", "notes": ""}
        print(f"  {spec.id:<18} {vs:>14}  {pr['source'][:44]:<44} "
              f"{pr['flag']}")
    if missing:
        sys.stderr.write("WARNING: no _provenance row for: "
                         + ", ".join(missing) + "\n")
    print()
    print(f"  MEMBER BREAKDOWN ({len(q.facts)} facts, accession per fact):")
    print(f"  {'class':<16} {'grade':<22} {'member (qname)':<44} "
          f"{'USD':>16}  accession")
    for f, klass, grade in q.facts:
        print(f"  {klass:<16} {grade:<22} {f.grade_qname:<44} "
              f"{f.value:>16,.0f}  {f.accession}")
    if um or uc:
        print()
        print("  UNMAPPED (needs manual mapping -- never guessed): "
              + ", ".join(sorted(um | uc)))
    if events:
        print()
        print("  8-K EVENTS (trailing 4 quarters):")
        for e in events:
            print(f"    {e.date}  item {e.item:<5} {ITEM_LABEL.get(e.item, '')}"
                  f"  {e.accession}  {filing_dir_url(cik, e.accession)}")
    return 0


# ---------------------------------------------------------------------------
# --lookup (name -> CIK via company_tickers.json; offline: seeded index)
# ---------------------------------------------------------------------------
EMBEDDED_INDEX = [
    ("35527", "FITB", "Fifth Third Bancorp"),
    ("1281761", "RF", "Regions Financial Corp"),
    ("109380", "ZION", "Zions Bancorporation NA"),
    ("28412", "CMA", "Comerica Inc"),
    ("49196", "HBAN", "Huntington Bancshares Inc"),
    ("92230", "TFC", "Truist Financial Corp"),
    ("91576", "KEY", "KeyCorp"),
    ("36270", "MTB", "M&T Bank Corp"),
    ("36104", "USB", "US Bancorp"),
    ("875357", "BOKF", "BOK Financial Corp"),
    ("19617", "JPM", "JPMorgan Chase & Co"),
]


def do_lookup(name: str, demo: bool, workbook: Optional[str]) -> int:
    needle = name.strip().lower()
    if demo:
        hits = [(c, t, n) for c, t, n in EMBEDDED_INDEX
                if needle in n.lower() or needle == t.lower()]
        if workbook and os.path.exists(workbook):
            import openpyxl
            wb = openpyxl.load_workbook(workbook, read_only=True)
            cfg = parse_config([[c for c in row] for row in
                                wb["_config"].iter_rows(values_only=True)])
            wb.close()
            for p in cfg.peers:
                if p.has_bank and (needle in p.name.lower()
                                   or needle == p.ticker.lower()):
                    hits.append((p.cik, p.ticker, p.name + " ([PEERS] row)"))
        seen = set()
        hits = [h for h in hits if not (h[0] in seen or seen.add(h[0]))]
        print("OFFLINE LOOKUP (--demo): searching the seeded [PEERS] rows + "
              "a small embedded index only. Live --lookup resolves ANY "
              "registrant name/ticker via the SEC's company_tickers.json "
              "(keyless; needs edgar_user_agent set).")
        if not hits:
            print(f'No offline match for "{name}". Run live: '
                  f'python runner.py -w <book>.xlsm --lookup "{name}"')
            return 0
        for c, t, n in hits:
            print(f"  CIK {c:>10}  {t:<6} {n}")
        return 0
    ua = os.environ.get("EDGAR_USER_AGENT", "")
    if workbook and os.path.exists(workbook):
        import openpyxl
        wb = openpyxl.load_workbook(workbook, read_only=True)
        cfg = parse_config([[c for c in row] for row in
                            wb["_config"].iter_rows(values_only=True)])
        wb.close()
        ua = cfg.user_agent or ua
    try:
        prov = EdgarProvider(user_agent=ua)
    except UserAgentMissing as exc:
        sys.stderr.write(f"lookup: {exc}\n(Or set the EDGAR_USER_AGENT "
                         "environment variable when running without "
                         "--workbook.)\n")
        return 3
    try:
        hits = prov.lookup(name)
    except Exception as exc:
        sys.stderr.write(f"lookup failed: {exc}\nCheck network/proxy access "
                         "to www.sec.gov (run --selftest; allowlist "
                         + ", ".join(SEC_HOSTS) + ").\n")
        return 1
    if not hits:
        print(f'No registrants matched "{name}".')
        return 0
    for h in hits:
        print(f'  CIK {h["cik"]:>10}  {h["ticker"]:<6} {h["title"]}')
    print("\nPut the CIK into the _config [PEERS] table "
          "(slot | cik | name | ticker | active) and re-run the runner -- "
          "no rebuild needed within the built slot capacity.")
    return 0


# ---------------------------------------------------------------------------
# --selftest (trap E10): probe the three SEC hosts; print allowlist guidance.
# ---------------------------------------------------------------------------
def do_selftest(workbook: Optional[str]) -> int:
    ua = os.environ.get("EDGAR_USER_AGENT", "")
    if workbook and os.path.exists(workbook):
        import openpyxl
        wb = openpyxl.load_workbook(workbook, read_only=True)
        cfg = parse_config([[c for c in row] for row in
                            wb["_config"].iter_rows(values_only=True)])
        wb.close()
        ua = cfg.user_agent or ua
    if not ua:
        print("NOTE: no edgar_user_agent configured -- probing with a "
              "placeholder UA; SEC may 403 it. Set edgar_user_agent in "
              "_config (\"{org} {email}\") before a live run.")
        ua = "selftest placeholder@example.com"
    probes = [
        ("data.sec.gov (submissions JSON)",
         SUBMISSIONS_URL.format(cik10=cik10("35527"))),
        ("www.sec.gov (/Archives + company_tickers.json)", TICKERS_URL),
        ("efts.sec.gov (full-text search fallback lane)", FTS_URL),
    ]
    import urllib.request
    import urllib.error
    ok = True
    for label, url in probes:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": ua})
            with urllib.request.urlopen(req, timeout=20) as resp:
                print(f"  OK    {label}  (HTTP {resp.status})")
        except urllib.error.HTTPError as exc:
            # ANY HTTP response proves the host is reachable
            note = (" -- 403 usually means the User-Agent was rejected: set "
                    "edgar_user_agent to \"{org} {email}\""
                    if exc.code == 403 else "")
            print(f"  WARN  {label}  (HTTP {exc.code}{note})")
        except Exception as exc:
            ok = False
            print(f"  FAIL  {label}  ({type(exc).__name__}: {exc})")
    if not ok:
        print()
        print("Connectivity failed. Corporate proxies commonly block "
              "sec.gov (observed fact -- trap E10). Ask your network team "
              "to allowlist HTTPS to: " + ", ".join(SEC_HOSTS) + ". "
              "All requests are keyless GETs with a declared User-Agent, "
              "throttled well under SEC's 10 req/s fair-access cap.")
    return 0 if ok else 1


# ---------------------------------------------------------------------------
# CLI (contract sec 4)
# ---------------------------------------------------------------------------
def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="EDGAR Criticized/Classified Tracker runner (SEC EDGAR)")
    ap.add_argument("--workbook", "-w", default=None,
                    help="path to the .xlsm (required unless "
                         "--lookup/--selftest)")
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline EdgarDemoProvider (no "
                         "network, no User-Agent needed)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    ap.add_argument("--lookup", default=None, metavar="NAME",
                    help='find a registrant CIK by name/ticker: '
                         '--lookup "Frost" (live via company_tickers.json; '
                         "offline searches the seeded index)")
    ap.add_argument("--tieout", default=None, nargs="+", metavar="ARG",
                    help="CIK [quarter]: print value + tag + member "
                         "breakdown + accession + EDGAR viewer URL for one "
                         "bank(-quarter); works with --demo (labeled)")
    ap.add_argument("--selftest", action="store_true",
                    help="probe the three SEC hosts and print proxy "
                         "allowlist guidance")
    args = ap.parse_args(argv)
    if args.selftest:
        return do_selftest(args.workbook)
    if args.lookup is not None:
        return do_lookup(args.lookup, args.demo, args.workbook)
    if not args.workbook:
        sys.stderr.write("--workbook is required (or use --lookup / "
                         "--selftest).\n")
        return 2
    if args.tieout is not None:
        if len(args.tieout) > 2:
            sys.stderr.write("--tieout takes CIK [quarter].\n")
            return 2
        try:
            return do_tieout(args.workbook, args.tieout, args.demo)
        except Exception as exc:
            sys.stderr.write(f"TIEOUT ERROR: {exc}\n")
            return 1
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, MetricError, PeerCapacityError) as exc:
        sys.stderr.write(f"GATE ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except UserAgentMissing as exc:
        sys.stderr.write(f"MISSING SETTING: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True,
                      **{k: v for k, v in status.items() if k != "digest"}}))
    sys.stderr.write(
        f"OK ({status['mode']}): {status['banks_landed']}/"
        f"{status['banks_active']} banks landed, "
        f"{status['alert_banks']} ALERT / {status['watch_banks']} WATCH / "
        f"{status['na_banks']} N-A / {status['stale_banks']} STALE, "
        f"8-K watch events={status['watch_8k_events']}, "
        f"unmapped members={len(status['unmapped_members'])}, "
        f"refused={len(status['watchlist_refusals'])}.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
