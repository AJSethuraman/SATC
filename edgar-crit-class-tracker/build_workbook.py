#!/usr/bin/env python3
"""Build the EDGAR Criticized/Classified Tracker workbook -- KeyBank-styled.

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled .xlsm.
All visual style comes from keybank_style.py (tokens + helpers) -- never a
hardcoded hex here. Built to BUILD_SPEC_EDGAR.md (Sections 2-5).

Banks are ROWS. Every dashboard/Watchlist row is anchored to a PEER SLOT:
bank identity (name/CIK/ticker) flows in BY FORMULA from the _config [PEERS]
cells, the FAMILY flag and metric values BY FORMULA from the slot's fixed
Raw_EDGAR block -- so editing WHICH bank occupies a slot re-labels and
re-points everything on recalc, with zero rebuilt formulas. Capacity is a
build knob: --peer-slots (default 20).

Structure (Section 4):
  Dashboard_Criticized -- criticized_ratio + QoQ/YoY deltas + sm/classified
                          (family-N/A-gated IN FORMULA), family flag, status;
                          peer-median row (criticized-capable subset labeled)
  Dashboard_Mix_Events -- criticized-$ mix by class; the 8-K event lane
                          (item counts + latest event + accession link)
  Watchlist            -- ranked: Bank | CIK | Family | Criticized % | dQoQ |
                          8-K flags | Rank | Status (+ helper status columns)
  Raw_EDGAR            -- one block per slot: canonical metric rows (12
                          quarters, newest-first) + latest-quarter member
                          FACT rows (audit trail) + 8-K EVENT rows
  _config              -- [SETTINGS]/[THRESHOLDS]/[PEERS]/[MEMBER_MAP]/
                          [CLASS_MAP] (source of truth; the map sections
                          carry PROVISIONED blank rows for the bootstrap)
  _provenance          -- the tie-out map + QUOTED uniform interagency
                          definitions (contract sec 12); --tieout reads it
  _code_py/_code_vba   -- runner.py + macro.bas as flat ASCII text
  _readme              -- family limitations, honest caveats, run steps

No native charts (L4). Embedded code is flat ASCII (L3). Every raw reference
is blank-guarded (empty cell renders blank, never 0). Threshold cells are
NUMERIC (L8). Heat direction: red = RISING criticized (stress).
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import runner as R
import member_seed as SEED
import provenance_seed as PROV

import keybank_style as KB
from keybank_style import (
    INK, KEY_RED, MONO_FONT, NOTE_FONT, SECT_FILL,
    brand_banner, kpi_tiles, header_row, watchlist_boundary,
    hide_gridlines, freeze_below,
)

HERE = os.path.dirname(os.path.abspath(__file__))
INK_BOLD = Font(name="Arial", bold=True, color=INK)
SMALL_BOLD = Font(name="Calibri", bold=True, size=9)

DASH_HDR = 7                    # header band row; slot rows start at 8
DASH_FIRST = DASH_HDR + 1
WL_HDR = 4
WL_FIRST = WL_HDR + 1
WL_HELPER_COL0 = 10             # J..L = per-metric status helpers
WL_SCORE_COL = WL_HELPER_COL0 + len(R.BANDED_METRICS)   # M = rank score

CRIT_TAB = "Dashboard_Criticized"
MIX_TAB = "Dashboard_Mix_Events"

CRIT_COLS = ["Bank", "CIK", "Ticker", "Family", "Criticized %", "dQoQ pp",
             "dYoY pp", "Crit growth %", "SM %", "Classified %", "Status"]
MIX_COLS = ["Bank", "CIK", "Family", "Criticized $M", "Mix C&I %",
            "Mix CRE %", "Mix Constr %", "Mix Other %", "8-K 2.04",
            "8-K 2.06", "8-K 4.02", "8-K 1.03", "8-K 2.02", "Latest 8-K"]
WL_COLS = ["Bank", "CIK", "Family", "Criticized %", "dQoQ pp", "8-K flags",
           "Rank", "Status"]


def dash_row(slot):
    return DASH_FIRST + slot - 1


def wl_row(slot):
    return WL_FIRST + slot - 1


def stress_heat(ws, cell_range, direction="above"):
    """Red = stress. All this template's bands run above-is-bad (rising
    criticized), but the helper stays direction-aware for reuse."""
    lo, hi = (KB.HEAT_GOOD, KB.HEAT_BAD) if direction != "below" \
        else (KB.HEAT_BAD, KB.HEAT_GOOD)
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="min", start_color=lo,
        mid_type="percentile", mid_value=50, mid_color=KB.HEAT_MID,
        end_type="max", end_color=hi))


def band_rules(ws, cell_range, first_cell, watch_ref, alert_ref, direction):
    """Per-cell WATCH/ALERT coloring straight off the _config [THRESHOLDS]
    cells (config-driven, blank-guarded via ISNUMBER -- text 'N/A (family)'
    never colors)."""
    op = ">=" if direction != "below" else "<="
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{alert_ref})"],
        fill=KB.ALERT_FILL, font=KB.ALERT_FONT, stopIfTrue=True))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f"AND(ISNUMBER({first_cell}),{first_cell}{op}{watch_ref})"],
        fill=KB.TIGHTEN_FILL,
        font=Font(name="Calibri", bold=True, color=KB.SLATE),
        stopIfTrue=True))


def text_cell(ws, row, col, value):
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# _config  ([SETTINGS] / [THRESHOLDS] / [PEERS] / [MEMBER_MAP] / [CLASS_MAP])
# ==========================================================================
def config_rows(peer_slots=R.PEER_SLOTS_DEFAULT):
    rows = [["EDGAR Criticized/Classified Tracker -- CONFIG (the knob "
             "panel). Edit values here; no code change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["pack_version", R.PACK_VERSION, "pack version BUILT into this "
         "workbook. The raw layout and every formula are anchored to it; "
         "the runner refuses a mismatched layout -- rebuild, never edit in "
         "place."],
        ["demo_mode", "FALSE", "TRUE = offline deterministic "
         "EdgarDemoProvider (no network, no User-Agent needed), for trying "
         "the button and tests."],
        ["raw_slots", 12, "quarters kept per bank (newest-first). BUILT "
         "INTO every formula -- changing it requires rebuilding the "
         "workbook (make_workbook.py); the runner refuses a mismatched "
         "layout."],
        ["peer_slots", peer_slots, "the BUILT peer capacity "
         "(make_workbook.py --peer-slots N sets it). More banks than this "
         "in [PEERS] = the runner refuses with the rebuild command -- "
         "never truncation."],
        ["edgar_user_agent", "", "REQUIRED for live runs: SEC fair-access "
         "policy demands a declared User-Agent \"{org} {email}\" on every "
         "request (blank/generic = silent 403 + ~10-min IP block). Example: "
         "ExampleBank CreditRisk jane.doe@example.com. Still KEYLESS -- no "
         "API key or account."],
        ["http_min_interval", 0.2, "seconds between live SEC requests "
         "(fair-access cap is 10/s -- stay well under; floor 0.15s "
         "enforced)."],
        ["edgar_max_retries", 4, "retries with backoff on HTTP 429/5xx "
         "(403 fails fast -- it means the User-Agent was rejected)."],
        ["stale_multiplier", 2.0, "a bank is STALE when its latest CQI "
         "quarter lags the peer-set max by more than this many quarters "
         "(late filer or delisted? -- check submissions); stale banks are "
         "excluded from digest alert counts. NOTE: the Excel median row "
         "still includes them -- see _readme."],
        ["secret_env", "", "NAME of the env var holding a licensed "
         "(Class C) secret. Empty in v1 -- EDGAR is keyless, and the gates "
         "admit ONLY Class A; a licensed swap requires its own review."],
    ]
    rows += [[], ["[THRESHOLDS]"],
             ["id", "watch", "alert", "direction", "authority"]]
    for tid, watch, alert, direction, authority in SEED.THRESHOLDS:
        # watch/alert are NUMERIC cells (L8: Excel's number>=text comparison
        # is silently FALSE and would downgrade every ALERT).
        rows.append([tid, watch, alert, direction, authority])
    rows += [["# 8-K lane rule (not banded): ANY item 2.04/4.02/1.03 filing "
              "in the trailing 4 quarters = automatic WATCH flag on the "
              "bank row."]]
    rows += [[], ["[PEERS]"], list(SEED.PEER_HEADER)]
    for prow in SEED.peer_rows(peer_slots):
        rows.append(prow)
    rows += [["# One bank per row: slot | cik | name | ticker | active. Add "
              "a bank = fill a free slot row + re-run (no rebuild). Remove "
              "= active FALSE or clear the row (the slot blanks on the next "
              "run). The registrant is the HOLDING COMPANY (CIK), not the "
              "bank charter (CERT) -- see _readme."],
             ["# CIK provenance: all 10 seeded CIKs and their disclosure "
              "families are research-verified (COVERAGE_RESEARCH_EDGAR.md). "
              "Find more with: python runner.py --lookup \"<name>\"."]]
    rows += [[], ["[MEMBER_MAP]"], list(SEED.MEMBER_MAP_HEADER)]
    for mrow in SEED.member_map_rows():
        rows.append(mrow)
    rows += [[] for _ in range(SEED.MEMBER_MAP_SPARE_ROWS)]
    rows += [["# cik | family | member_qname | grade. Grades: pass, "
              "special_mention, substandard, doubtful, loss, "
              "criticized_accruing, criticized_nonaccruing, "
              "investment_grade, noninvestment_grade, ignore, unmapped."],
             ["# The runner APPENDS newly observed members into the blank "
              "rows above as grade=\"unmapped\" and flags them -- map them "
              "by hand (edit the grade cell) and re-run; NEVER guessed. "
              "us-gaap:CriticizedMember is a SUBTOTAL (parents SM/Sub/"
              "Doubtful/Loss) -- seeded grade=\"ignore\" so it never "
              "double-counts."],
             ["# Extension qnames (key:/mtb:/hban:) are ILLUSTRATIVE seeds "
              "for the researched disclosure shapes; the first live fetch "
              "resolves the real ones via this same bootstrap."]]
    rows += [[], ["[CLASS_MAP]"], list(SEED.CLASS_MAP_HEADER)]
    for crow in SEED.class_map_rows():
        rows.append(crow)
    rows += [[] for _ in range(SEED.CLASS_MAP_SPARE_ROWS)]
    rows += [["# cik | class_member_qname | class. Classes: ci, cre, "
              "construction, other_commercial, excluded (consumer/resi -- "
              "dropped from the COMMERCIAL denominators), unmapped. "
              "cik=\"*\" rows are standard us-gaap members applying to "
              "every bank; per-bank extension class members bootstrap into "
              "the blank rows above as \"unmapped\" (excluded from rollups "
              "until you map them -- trap E5)."]]
    return rows


def write_config(wb, peer_slots):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows(peer_slots)
    thr_cells = {}                 # id -> (watch_ref, alert_ref, direction)
    peer_cells = {}                # slot -> dict(cik/name/ticker/active refs)
    section = None
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        head = row[0] if row else None
        if isinstance(head, str) and head.startswith("[") \
                and head.endswith("]"):
            section = head
            continue
        if not row or head is None or (isinstance(head, str)
                                       and str(head).startswith("#")):
            continue
        if section == "[THRESHOLDS]" and head != "id":
            thr_cells[head] = (f"_config!$B${i}", f"_config!$C${i}",
                               str(row[3]).strip().lower())
        if section == "[PEERS]" and str(head).strip().lower() != "slot":
            slot = int(head)
            peer_cells[slot] = {
                "cik": f"_config!$B${i}", "name": f"_config!$C${i}",
                "ticker": f"_config!$D${i}", "active": f"_config!$E${i}"}
    ws["A1"].font = INK_BOLD
    widths = {"A": 26, "B": 18, "C": 46, "D": 14, "E": 10}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("[") \
                and row[0].endswith("]"):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    return thr_cells, peer_cells


# ==========================================================================
# Raw_EDGAR scaffold  (one block per SLOT; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, peer_slots, raw_slots):
    ws = wb.create_sheet(R.RAW_TAB)
    hide_gridlines(ws)
    ws["A1"] = (f"{R.RAW_TAB} -- canonical quarterly rollups + member-level "
                f"CQI facts + 8-K events per peer slot (EdgarDemoProvider "
                f"output in --demo), newest first. One block per SLOT: "
                f"anchors never move when [PEERS] changes. Ratios in "
                f"percent, amounts in USD. Written by runner.py -- do not "
                f"edit. Accession + tag recorded per quarter and per fact "
                f"(contract sec 12).")
    ws["A1"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 12
    for j in range(len(R.RAW_METRIC_COLS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13
    for slot in range(1, peer_slots + 1):
        b = R.slot_block(slot, raw_slots)
        ws.cell(b.header_row, 1, R.slot_label(slot)).font = INK_BOLD
        # cells B..F (unit key, name, family, latest accession, tag) are
        # RUNTIME state the runner writes/blanks per run -- empty at build.
        ws.cell(b.label_row, 1, "period").font = SMALL_BOLD
        for cname in R.RAW_METRIC_COLS:
            c = ws.cell(b.label_row, R.metric_col(cname), cname)
            c.font = SMALL_BOLD
        fl = ws.cell(b.facts_label_row, 1, "facts: period")
        fl.font = SMALL_BOLD
        for cname in R.FACT_COLS[1:]:
            ws.cell(b.facts_label_row, R.fact_col(cname), cname).font = \
                SMALL_BOLD
        el = ws.cell(b.events_label_row, 1, "events: date")
        el.font = SMALL_BOLD
        for cname in R.EVENT_COLS[1:]:
            ws.cell(b.events_label_row, R.event_col(cname), cname).font = \
                SMALL_BOLD
    return ws


# ==========================================================================
# Shared formula fragments
# ==========================================================================
def _fam_ref(slot, raw_slots):
    b = R.slot_block(slot, raw_slots)
    return f"{R.RAW_TAB}!$D${b.header_row}"


def _acc_ref(slot, raw_slots):
    b = R.slot_block(slot, raw_slots)
    return f"{R.RAW_TAB}!$E${b.header_row}"


def _mref(slot, cname, raw_slots):
    """Latest-quarter canonical metric cell for a slot."""
    b = R.slot_block(slot, raw_slots)
    col = get_column_letter(R.metric_col(cname))
    return f"{R.RAW_TAB}!{col}{b.first_data_row}"


def _item_range(slot, raw_slots):
    b = R.slot_block(slot, raw_slots)
    col = get_column_letter(R.event_col("item"))
    return (f"{R.RAW_TAB}!${col}${b.first_event_row}:"
            f"${col}${b.first_event_row + R.EVENT_ROWS - 1}")


def _guard(ref):
    """Blank-guard: an empty cell renders blank, never 0."""
    return f"=IF({ref}=\"\",\"\",{ref})"


def _tier1_formula(slot, cname, raw_slots):
    """Tier 1 value with the family honesty gate rendered IN FORMULA:
    ig_nig/unmapped families show 'N/A (family)', never a number."""
    fam = _fam_ref(slot, raw_slots)
    ref = _mref(slot, cname, raw_slots)
    return (f"=IF({fam}=\"\",\"\","
            f"IF(OR({fam}=\"ig_nig\",{fam}=\"unmapped\"),\"N/A (family)\","
            f"IF({ref}=\"\",\"\",{ref})))")


def _tier2_formula(slot, cname, raw_slots):
    """Tier 2 (grades_full only -- BUILD SPEC 0.6): other families render
    'N/A (family)', never approximated."""
    fam = _fam_ref(slot, raw_slots)
    ref = _mref(slot, cname, raw_slots)
    return (f"=IF({fam}=\"\",\"\","
            f"IF({fam}<>\"grades_full\",\"N/A (family)\","
            f"IF({ref}=\"\",\"\",{ref})))")


def _watch8k_formula(slot, raw_slots, name_ref):
    itm = _item_range(slot, raw_slots)
    return (f"=IF({name_ref}=\"\",\"\","
            + "+".join(f"COUNTIF({itm},\"{i}\")" for i in R.WATCH_ITEMS)
            + ")")


# ==========================================================================
# Dashboard_Criticized
# ==========================================================================
def write_dashboard_criticized(wb, peer_slots, raw_slots, thr_cells,
                               peer_cells, n_active):
    ws = wb.create_sheet(CRIT_TAB)
    hide_gridlines(ws)
    ncols = len(CRIT_COLS)
    first, last = DASH_FIRST, DASH_FIRST + peer_slots - 1
    for slot in range(1, peer_slots + 1):
        r = dash_row(slot)
        pc = peer_cells[slot]
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = \
            KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cik']}=\"\",\"\",{pc['cik']})").font = \
            MONO_FONT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3, f"=IF({pc['ticker']}=\"\",\"\",{pc['ticker']})").font \
            = KB.SECONDARY
        ws.cell(r, 4, _guard(_fam_ref(slot, raw_slots))).font = KB.SECONDARY
        ws.cell(r, 5, _tier1_formula(slot, "CRIT_RATIO", raw_slots))
        ws.cell(r, 6, _guard(_mref(slot, "CRIT_QOQ", raw_slots)))
        ws.cell(r, 7, _guard(_mref(slot, "CRIT_YOY", raw_slots)))
        ws.cell(r, 8, _guard(_mref(slot, "CRIT_GROWTH", raw_slots)))
        ws.cell(r, 9, _tier2_formula(slot, "SM_RATIO", raw_slots))
        ws.cell(r, 10, _tier2_formula(slot, "CLASS_RATIO", raw_slots))
        ws.cell(r, 11, f"=IF(Watchlist!$H${wl_row(slot)}=\"\",\"\","
                       f"Watchlist!$H${wl_row(slot)})")
        for c in range(5, 11):
            ws.cell(r, c).number_format = "0.00"
            ws.cell(r, c).alignment = KB.RIGHT
        ws.cell(r, 11).alignment = KB.CENTER
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    # PEER MEDIAN row: Excel MEDIAN ignores blanks AND the N/A text, so the
    # median is automatically the criticized-capable subset. v1 caveat: a
    # STALE bank's landed values still enter it (the digest median excludes
    # them) -- divergence documented in _readme.
    med = last + 1
    ws.cell(med, 1, "PEER MEDIAN (criticized-capable families; includes "
                    "stale -- see _readme)").font = INK_BOLD
    for c in range(5, 11):
        col = get_column_letter(c)
        rng = f"{col}{first}:{col}{last}"
        cell = ws.cell(med, c, f"=IF(COUNT({rng})=0,\"\",MEDIAN({rng}))")
        cell.number_format = "0.00"
        cell.alignment = KB.RIGHT
        cell.font = INK_BOLD
    ws.cell(med, 1).border = KB.HAIRLINE

    header_row(ws, DASH_HDR, CRIT_COLS, right_from=4, center_cols=(10,))
    freeze_below(ws, DASH_HDR)
    widths = {"A": 26, "B": 10, "C": 8, "D": 15, "E": 12, "F": 10, "G": 10,
              "H": 12, "I": 10, "J": 12, "K": 14}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    brand_banner(
        ws, 1, ncols,
        "Criticized / Classified Dashboard -- Competitor Commercial Books",
        "Quarterly criticized/classified from each holding company's own "
        "filed 10-Q/10-K Credit Quality Indicators note (ASC 326-20-50-5). "
        "Tier 2 (SM/classified) is grades_full families only -- other "
        "families render N/A, never approximated (BUILD SPEC 0.6). "
        "Definitions uniform (2004 Interagency Agreement); application "
        "timing is bank judgment (E9). Q4 arrives via the 10-K, 60-90 days "
        "after year-end (E6).")
    wl_last = WL_FIRST + peer_slots - 1
    tiles = [
        ("BANKS IN ALERT",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"ALERT\")",
         f"of {n_active} active"),
        ("BANKS IN WATCH",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"WATCH\")",
         "incl. 8-K auto-watch"),
        ("BANKS TRACKED",
         f"=COUNTIF(Watchlist!$A${WL_FIRST}:$A${wl_last},\"?*\")",
         f"of {peer_slots} built slots"),
    ]
    kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8

    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 44
    for rr in (1, 2, 3):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(4, R.STATUS_COL).font = NOTE_FONT          # macro writes here
    ws.cell(4, R.STATUS_COL).alignment = KB.LEFT

    # coloring: band rules off the numeric _config threshold cells; heat red
    # = RISING criticized (all above-is-bad)
    for c, mid in ((5, "criticized_ratio"), (6, "crit_qoq_delta"),
                   (9, "sm_ratio")):
        col = get_column_letter(c)
        rng = f"{col}{first}:{col}{last}"
        watch_ref, alert_ref, direction = thr_cells[mid]
        band_rules(ws, rng, f"{col}{first}", watch_ref, alert_ref, direction)
        stress_heat(ws, rng, direction)
    for c in (7, 8, 10):
        col = get_column_letter(c)
        stress_heat(ws, f"{col}{first}:{col}{last}", "above")
    KB.alert_rule(ws, f"K{first}:K{last}")
    return ws


# ==========================================================================
# Dashboard_Mix_Events
# ==========================================================================
def write_dashboard_mix(wb, peer_slots, raw_slots, peer_cells, n_active):
    ws = wb.create_sheet(MIX_TAB)
    hide_gridlines(ws)
    ncols = len(MIX_COLS)
    first, last = DASH_FIRST, DASH_FIRST + peer_slots - 1
    for slot in range(1, peer_slots + 1):
        r = dash_row(slot)
        pc = peer_cells[slot]
        b = R.slot_block(slot, raw_slots)
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = \
            KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cik']}=\"\",\"\",{pc['cik']})").font = \
            MONO_FONT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3, _guard(_fam_ref(slot, raw_slots))).font = KB.SECONDARY
        cu = _mref(slot, "CRIT_USD", raw_slots)
        ws.cell(r, 4, f"=IF({cu}=\"\",\"\",{cu}/1000000)")
        for k, cname in enumerate(("MIX_CI", "MIX_CRE", "MIX_CONSTR",
                                   "MIX_OTHER")):
            ws.cell(r, 5 + k, _guard(_mref(slot, cname, raw_slots)))
        itm = _item_range(slot, raw_slots)
        for k, item in enumerate(("2.04", "2.06", "4.02", "1.03", "2.02")):
            ws.cell(r, 9 + k,
                    f"=IF($A{r}=\"\",\"\",COUNTIF({itm},\"{item}\"))")
        ev_date = f"{R.RAW_TAB}!$A${b.first_event_row}"
        ev_item = (f"{R.RAW_TAB}!${get_column_letter(R.event_col('item'))}"
                   f"${b.first_event_row}")
        ws.cell(r, 14, f"=IF({ev_date}=\"\",\"\","
                       f"{ev_date}&\"  item \"&{ev_item})")
        for c in range(4, 9):
            ws.cell(r, c).number_format = "0.00"
            ws.cell(r, c).alignment = KB.RIGHT
        for c in range(9, 14):
            ws.cell(r, c).alignment = KB.CENTER
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    header_row(ws, DASH_HDR, MIX_COLS, right_from=3,
               center_cols=(8, 9, 10, 11, 12))
    freeze_below(ws, DASH_HDR)
    widths = {"A": 26, "B": 10, "C": 15, "D": 13, "E": 10, "F": 10, "G": 11,
              "H": 11, "I": 9, "J": 9, "K": 9, "L": 9, "M": 9, "N": 22}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    brand_banner(
        ws, 1, ncols,
        "Criticized Mix & 8-K Credit Events",
        "Left: criticized-$ share by coarse commercial class ([CLASS_MAP]: "
        "C&I / CRE / construction / other). Right: the 8-K credit-event "
        "lane harvested from the submissions items[] array -- 2.04 debt "
        "acceleration, 2.06 impairment, 4.02 non-reliance, 1.03 "
        "receivership (any 2.04/4.02/1.03 = automatic WATCH), 2.02 "
        "informational. Trailing 4 quarters; accession links in Raw_EDGAR "
        "event blocks.")
    wl_last = WL_FIRST + peer_slots - 1
    tiles = [
        ("8-K WATCH EVENTS",
         f"=SUM(Watchlist!$F${WL_FIRST}:$F${wl_last})",
         "2.04 / 4.02 / 1.03 filings, trailing 4Q"),
        ("BANKS IN ALERT",
         f"=COUNTIF(Watchlist!$H${WL_FIRST}:$H${wl_last},\"ALERT\")",
         f"of {n_active} active"),
        ("BANKS TRACKED",
         f"=COUNTIF(Watchlist!$A${WL_FIRST}:$A${wl_last},\"?*\")",
         f"of {peer_slots} built slots"),
    ]
    kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8

    scol_i = R.STATUS_COL_BY_TAB[MIX_TAB]
    ws.column_dimensions[get_column_letter(scol_i)].width = 44
    for rr in (1, 2, 3):
        ws.cell(rr, scol_i).font = KB.SECONDARY
        ws.cell(rr, scol_i).alignment = KB.LEFT
    ws.cell(4, scol_i).font = NOTE_FONT
    ws.cell(4, scol_i).alignment = KB.LEFT

    for c in (4, 5, 6, 7, 8):
        col = get_column_letter(c)
        stress_heat(ws, f"{col}{first}:{col}{last}", "above")
    for c in (9, 11, 12):                       # the auto-WATCH item counts
        col = get_column_letter(c)
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            FormulaRule(formula=[f"AND(ISNUMBER({col}{first}),"
                                 f"{col}{first}>0)"],
                        fill=KB.ALERT_FILL, font=KB.ALERT_FONT))
    return ws


# ==========================================================================
# Watchlist  (the ranked entity-gated lane)
# ==========================================================================
BOUNDARY = ("ENTITY RULE. This lane admits ONLY rows carrying a genuine SEC "
            "CIK join key (cik:NNN from the [PEERS] cik cell, Class A) -- "
            "the default-deny validator refuses anything else BY NAME: "
            "tickers-as-keys, series names and name-only rows cannot be "
            "monitored. Blank/malformed CIK shows REFUSED here; find CIKs "
            "with: python runner.py --lookup \"<name>\" "
            "(company_tickers.json). The registrant is the HOLDING COMPANY "
            "-- bank-charter (CERT) promotion is a spec change, not a "
            "config edit.")


def write_watchlist(wb, cfg, peer_slots, raw_slots, thr_cells, peer_cells):
    ws = wb.create_sheet("Watchlist")
    hide_gridlines(ws)
    hj = get_column_letter(WL_HELPER_COL0)
    hx = get_column_letter(WL_HELPER_COL0 + len(R.BANDED_METRICS) - 1)
    brand_banner(ws, 1, len(WL_COLS),
                 "Watchlist -- Ranked Competitor Banks (Criticized Lane)",
                 f"Ranked by (ALERT flag count, criticized ratio) "
                 f"descending. Flag counts aggregate the banded metric "
                 f"statuses (helper columns {hj}-{hx}); any 2.04/4.02/1.03 "
                 f"8-K = automatic WATCH. ig_nig/unmapped families rank as "
                 f"N/A -- never approximated. STALE banks are surfaced by "
                 f"the runner's status line and digest, not here (v1).")
    watchlist_boundary(ws, 3, len(WL_COLS), BOUNDARY)

    header_row(ws, WL_HDR, WL_COLS, right_from=3, center_cols=(5, 6, 7))
    for k, mid in enumerate(R.BANDED_METRICS):
        ws.cell(WL_HDR, WL_HELPER_COL0 + k, mid).font = SMALL_BOLD
    ws.cell(WL_HDR, WL_SCORE_COL, "rank score").font = SMALL_BOLD

    first, last = WL_FIRST, WL_FIRST + peer_slots - 1
    zc = get_column_letter(WL_SCORE_COL)
    for slot in range(1, peer_slots + 1):
        r = wl_row(slot)
        pc = peer_cells[slot]
        fam = _fam_ref(slot, raw_slots)
        ws.cell(r, 1, f"=IF({pc['name']}=\"\",\"\",{pc['name']})").font = \
            KB.DATA_FONT
        ws.cell(r, 2, f"=IF({pc['cik']}=\"\",\"\",{pc['cik']})").font = \
            MONO_FONT
        ws.cell(r, 2).number_format = "0"
        ws.cell(r, 3, _guard(fam)).font = KB.SECONDARY
        ws.cell(r, 4, _tier1_formula(slot, "CRIT_RATIO", raw_slots))
        ws.cell(r, 4).number_format = "0.00"
        ws.cell(r, 5, _guard(_mref(slot, "CRIT_QOQ", raw_slots)))
        ws.cell(r, 5).number_format = "0.00"
        ws.cell(r, 6, _watch8k_formula(slot, raw_slots, f"$A{r}"))
        # per-metric status helpers: same thresholds, same direction, same
        # blank-guard as runner.status_for -- parity-tested (all bands run
        # above-is-bad in this template).
        for k, mid in enumerate(R.BANDED_METRICS):
            cname = R.METRICS[mid].raw_col
            ref = _mref(slot, cname, raw_slots)
            watch_ref, alert_ref, _direction = thr_cells[mid]
            ws.cell(r, WL_HELPER_COL0 + k,
                    f"=IF(NOT(ISNUMBER({ref})),\"\","
                    f"IF({ref}>={alert_ref},\"ALERT\","
                    f"IF({ref}>={watch_ref},\"WATCH\",\"OK\")))"
                    ).font = NOTE_FONT
        # rank score: ALERT count dominant, criticized ratio tiebreak;
        # N/A families carry no score (they rank last, status N/A).
        ws.cell(r, WL_SCORE_COL,
                f"=IF($A{r}=\"\",\"\","
                f"IF(OR({fam}=\"ig_nig\",{fam}=\"unmapped\"),\"\","
                f"COUNTIF(${hj}{r}:${hx}{r},\"ALERT\")*10000"
                f"+IF(ISNUMBER($D{r}),$D{r},0)))").font = NOTE_FONT
        ws.cell(r, 7, f"=IF(${zc}{r}=\"\",\"\","
                      f"RANK(${zc}{r},${zc}${first}:${zc}${last}))")
        # Status: the formula-level entity gate rendering -- REFUSED for a
        # blank/malformed CIK (with the --lookup hint), OFF for an inactive
        # slot, N/A for non-criticized-capable families (WATCH if an 8-K
        # credit event fired), else the flag rollup. Mirrors compute_digest
        # EXACTLY (parity-tested). STALE is runtime state (digest/status
        # line -- v1, see _readme).
        ws.cell(r, 8,
                f"=IF({pc['name']}=\"\",\"\","
                f"IF(NOT(ISNUMBER({pc['cik']})),"
                f"\"REFUSED: blank/invalid CIK -- run --lookup\","
                f"IF(UPPER({pc['active']}&\"\")<>\"TRUE\",\"OFF\","
                f"IF({fam}=\"\",\"\","
                f"IF(OR({fam}=\"ig_nig\",{fam}=\"unmapped\"),"
                f"IF($F{r}>0,\"WATCH\",\"N/A (family)\"),"
                f"IF(COUNTIF(${hj}{r}:${hx}{r},\"ALERT\")>0,\"ALERT\","
                f"IF(OR(COUNTIF(${hj}{r}:${hx}{r},\"WATCH\")>0,"
                f"$F{r}>0),\"WATCH\",\"OK\")))))))")
        ws.cell(r, 4).alignment = KB.RIGHT
        ws.cell(r, 5).alignment = KB.RIGHT
        for c in (6, 7, 8):
            ws.cell(r, c).alignment = KB.CENTER
        for c in range(1, len(WL_COLS) + 1):
            ws.cell(r, c).border = KB.HAIRLINE

    KB.alert_rule(ws, f"H{first}:H{last}")
    stress_heat(ws, f"D{first}:D{last}", "above")   # red = high criticized
    stress_heat(ws, f"F{first}:F{last}", "above")   # red = 8-K credit events

    r = last + 2
    ws.cell(r, 1, f"Columns {hj}-{hx} are the per-metric status helpers "
                  f"feeding the flag COUNTIFs; column {zc} makes the (ALERT "
                  "flags, criticized) ordering auditable. Unmapped-member "
                  "notices and refusals also land in the run digest / "
                  "email.").font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=len(WL_COLS))
    r += 1
    _, refusals, _ = R.evaluate_peers(cfg)
    for peer, message in refusals:
        ws.cell(r, 1, f"slot {peer.slot}").font = MONO_FONT
        ws.cell(r, 2, "REFUSED").font = Font(name="Calibri", bold=True,
                                             color=KB.CRIMSON)
        r += 1
        ws.cell(r, 1, message).font = NOTE_FONT
        ws.cell(r, 1).alignment = KB.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=len(WL_COLS))
        ws.row_dimensions[r].height = 56
        r += 1

    widths = {"A": 26, "B": 10, "C": 15, "D": 12, "E": 10, "F": 10, "G": 7,
              "H": 34}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for k in range(len(R.BANDED_METRICS) + 1):
        ws.column_dimensions[
            get_column_letter(WL_HELPER_COL0 + k)].width = 12
    freeze_below(ws, WL_HDR)
    return ws


# ==========================================================================
# _provenance  (contract sec 12 -- the tie-out map as a sheet)
# ==========================================================================
def write_provenance(wb):
    ws = wb.create_sheet("_provenance")
    hide_gridlines(ws)
    widths = {"A": 22, "B": 42, "C": 52, "D": 40, "E": 8, "F": 62}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    r = 1
    for line in PROV.PREAMBLE:
        text_cell(ws, r, 1, line)
        ws.cell(r, 1).font = INK_BOLD if r == 1 or line.startswith("[") \
            else NOTE_FONT
        r += 1
    r += 1
    ws.cell(r, 1, "[INPUTS + METRICS]").font = KB.SECTION_FONT
    ws.cell(r, 1).fill = SECT_FILL
    r += 1
    hdr = r
    for j, name in enumerate(PROV.HEADER, start=1):
        c = ws.cell(hdr, j, name)
        c.font = SMALL_BOLD
        c.fill = KB.SUBHDR_FILL
    r += 1
    for row in PROV.ALL_ROWS:
        for j, val in enumerate(row, start=1):
            text_cell(ws, r, j, val)
            ws.cell(r, j).font = MONO_FONT if j in (1, 3) else NOTE_FONT
        if "SEEDED" in row[4]:
            ws.cell(r, 5).font = Font(name="Calibri", bold=True, size=9,
                                      color=KB.CRIMSON)
        r += 1
    r += 1
    text_cell(ws, r, 1,
              "# Every canonical metric has a row here. Sample-verify from "
              "PowerShell: python runner.py -w <book>.xlsm --tieout <CIK> "
              "[quarter] (--demo prints labeled demo values).")
    ws.cell(r, 1).font = NOTE_FONT
    freeze_below(ws, hdr)
    return ws


# ==========================================================================
# Code-in-tab + readme
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
EDGAR CRITICIZED/CLASSIFIED TRACKER -- README
=============================================

WHAT THIS IS
  A self-contained, reusable template answering: "Whose COMMERCIAL book is
  being risk-rated down -- before delinquency shows it?" One workbook pulls
  each competitor holding company's own filed 10-Q/10-K Credit Quality
  Indicators note (ASC 326-20-50-5: amortized cost by internal credit grade,
  by class, quarterly) straight from SEC EDGAR -- keyless, public domain --
  plus an 8-K credit-event lane, and presents banks-as-rows dashboards with
  criticized/classified ratios, migration deltas, class mix, and a ranked
  Watchlist.

THE TWO-TRACK DESIGN (why this template exists)
  Commercial credit deteriorates through RISK RATINGS long before it shows
  in delinquency; consumer credit shows in DQ/NCO directly. The FDIC Bank
  Peer Monitor carries the consumer DQ/NCO track and the commercial
  Call-Report floor; THIS tracker carries the commercial criticized/
  classified view from the banks' own SEC filings. Same [PEERS] slot
  mechanism, same contract, complementary lanes.

DISCLOSURE FAMILIES (the honesty core -- BUILD SPEC 0.6)
  Banks disclose commercial credit quality in DIALECTS. Each bank carries a
  family flag derived from its [MEMBER_MAP] rows:
    grades_full     Pass/Special Mention/Substandard/Doubtful(/Loss) --
                    Tier 1 AND Tier 2 metrics (criticized, classified, SM).
    criticized_only Pass/Criticized-accruing/Criticized-nonaccruing (e.g.
                    KeyCorp, M&T) -- Tier 1 only; classified/SM render
                    "N/A (family)", NEVER approximated.
    ig_nig          Investment/Non-investment grade only (e.g. USB; JPM) --
                    criticized is NOT computable from XBRL; every ratio
                    renders "N/A (family)". Their criticized numbers live
                    in MD&A text -- the full-text-search fallback is v1.1.
    unmapped        No [MEMBER_MAP] rows yet -- N/A until mapped.
  Research estimate (UNVERIFIED -- validate live): ~60-75% of $10B+
  regionals disclose SM/Substandard separately; >90% support at least
  pass-vs-criticized.

THE MEMBER BOOTSTRAP (contract sec 13 -- never guess)
  Filers use EXTENSION members for non-canonical buckets (criticized-
  accruing, OLEM, ...). On every run, members not in [MEMBER_MAP] are
  APPENDED there as grade="unmapped" (into the provisioned blank rows),
  flagged in the digest/email, EXCLUDED from every numerator, and included
  in the amortized-cost denominator (they are part of the total -- a fact,
  not a guess). Map them by hand (edit the grade cell) and re-run. Same
  bootstrap for class members in [CLASS_MAP] (unmapped classes are excluded
  from rollups entirely). Grade "ignore" exists for filer-tagged SUBTOTALS
  (us-gaap:CriticizedMember parents the adverse leaves -- summing both
  would double-count).

ONE-TIME SETUP
  1. Install Python 3 (add to PATH) and the dependencies:
         pip install pandas openpyxl
  2. Enable macros when prompted (the Extract button needs them).
  3. FOR LIVE RUNS ONLY: set edgar_user_agent in _config to
     "{your org} {your email}" -- SEC fair-access policy REQUIRES a
     declared User-Agent (blank/generic = silent 403 + ~10-min IP block).
     Still KEYLESS: no API key, no account.

RUN IT  (the workbook is the finished product; you only add data)
  1. Alt+F8 -> ExtractFiles: writes runner.py (from _code_py),
     requirements.txt and RUN.txt next to the workbook. NOTHING runs
     inside Excel.
  2. SAVE and CLOSE the workbook.
  3. From PowerShell (see RUN.txt):
       - Offline demo:      python runner.py -w <book>.xlsm --demo
       - Live EDGAR:        python runner.py -w <book>.xlsm
       - Find a CIK:        python runner.py -w <book>.xlsm --lookup "Frost"
       - Verify a value:    python runner.py -w <book>.xlsm --tieout 35527
       - Connectivity:      python runner.py -w <book>.xlsm --selftest
  4. Reopen the workbook -- the formulas recalc into the dashboards.

THE DATA PATH (research-verified -- COVERAGE_RESEARCH_EDGAR.md)
  Stage 1: data.sec.gov/submissions/CIK{10-digit}.json -- COLUMNAR arrays;
    latest 10-Q/10-K per calendar quarter (Q4 = the 10-K; no Q4 10-Q
    exists -- expect the Q4 point 60-90 days after year-end, trap E6);
    10-Q/A amendments deduped by latest acceptanceDateTime (E7). The 8-K
    lane is harvested from the SAME response's items[] array -- zero extra
    requests: 2.04 debt acceleration, 2.06 impairment, 4.02 non-reliance,
    1.03 receivership (any of 2.04/4.02/1.03 = automatic WATCH), 2.02
    informational.
  Stage 2: each filing directory's index.json -> the EDGAR-generated
    EXTRACTED XBRL INSTANCE (*_htm.xml) -> stdlib XML parse. THE DECISIVE
    RESEARCH FINDING: SEC's convenience JSON APIs (companyfacts/frames) are
    structurally NON-DIMENSIONAL and cannot carry grade x class tables --
    the instance bypass is the only clean keyless path (E1).
  Facts: FinancingReceivableBeforeAllowanceForCreditLoss and its
    ExcludingAccruedInterest twin (per-bank preference recorded in the TAG
    column, E2), dimensioned InternalCreditAssessmentAxis x class axis.
    VINTAGE disclosures are LINE ITEMS, not an axis: exact-tag matching
    ignores them -- summing vintage rows on top of grade totals would
    double-count (E3). Member qnames stay namespace-qualified (E4). CIK is
    10-digit padded on data.sec.gov, UNPADDED in /Archives (E8).

METRICS + THRESHOLDS (heuristic, labeled -- edit in _config)
  criticized_ratio  (SM+Substandard+Doubtful+Loss | native criticized rows)
                    / total commercial amortized cost. WATCH 4 / ALERT 6.
  classified_ratio  Substandard+Doubtful+Loss / total (grades_full only).
  sm_ratio          Special Mention / total (grades_full only). W 2 / A 3.
  crit_qoq_delta    pp change vs prior quarter. WATCH +0.5 / ALERT +1.0.
  crit_yoy_delta, crit_growth, class mix (C&I/CRE/construction/other).
  8-K lane          counts + latest date per item, trailing 4 quarters,
                    accession links; 2.04/4.02/1.03 = auto-WATCH.
  Uniform definitions QUOTED with citations on the _provenance tab (2004
  Interagency Agreement; SR 93-30; OCC Rating Credit Risk handbook).
  Classified = Substandard+Doubtful+Loss; Criticized = SM + Classified.

HONEST CAVEATS (read before trusting a quarter)
  - APPLICATION TIMING (E9): definitions are uniform; WHEN a bank downgrades
    is bank judgment and exam-cycle driven. QoQ jumps can be exam artifacts.
  - Q4 LAG (E6): the Q4 point arrives with the 10-K (60/75/90 days), later
    than 10-Qs (40/45 days). A Q4-missing bank is often just the calendar.
    Fiscal-year filers off calendar quarters land on their fiscal ends.
  - HC vs CERT: this tracker monitors the HOLDING COMPANY consolidation
    (SEC registrant, CIK). The FDIC template monitors the BANK charter
    (CERT). The two consolidations differ (nonbank subs, parent-only debt)
    -- reconcile approximately, never silently; automated HC<->CERT
    reconciliation is out of v1.
  - STALE banks (late filer/delisted) are excluded from digest alert counts
    but their landed values still enter the EXCEL median row (Excel cannot
    see runtime staleness) -- the digest median is authoritative.
  - Unmapped grade members sit in the DENOMINATOR only (part of the total)
    and are flagged until you map them; unmapped CLASS members are excluded
    from rollups entirely until mapped.
  - PROXY/ALLOWLIST (E10): corporate proxies commonly block sec.gov. Run
    --selftest; allowlist data.sec.gov, www.sec.gov, efts.sec.gov.

NOT AVAILABLE (honestly out of scope, v1)
  - MD&A full-text fallback for ig_nig banks (v1.1; efts.sec.gov lane).
  - FSNDS bulk-ZIP backfill (documented path for deep history).
  - Vintage-level stress metrics; paid data; HC<->CERT auto-reconciliation.

NO AI IS INVOLVED IN THE DATA PATH. The gates, rollups, thresholds and
staleness guard are deterministic and config-driven: the same input always
yields the same workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() \
                and len(line) > 3 and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path, peer_slots=R.PEER_SLOTS_DEFAULT):
    rows = config_rows(peer_slots)
    cfg = R.parse_config(rows)
    # BUILD-TIME hard gates (BUILD SPEC 0.1 / sec 3): bad seed peers or an
    # over-capacity seed refuse the BUILD itself (defense in depth).
    R.assert_metrics_admissible(R.METRIC_SPECS)
    R.validate_peer_capacity(cfg)
    R.assert_entity_gates(cfg)
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    raw_slots = cfg.raw_slots

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, peer_slots, raw_slots)
    thr_cells, peer_cells = write_config(wb, peer_slots)
    write_dashboard_criticized(wb, peer_slots, raw_slots, thr_cells,
                               peer_cells, len(admitted))
    write_dashboard_mix(wb, peer_slots, raw_slots, peer_cells, len(admitted))
    write_watchlist(wb, cfg, peer_slots, raw_slots, thr_cells, peer_cells)
    write_provenance(wb)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    order = [CRIT_TAB, MIX_TAB, "Watchlist", R.RAW_TAB, "_config",
             "_provenance", "_code_py", "_code_vba", "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order
                    else 99)
    wb.active = wb.sheetnames.index(CRIT_TAB)
    wb.save(out_path)
    return out_path, len(R.METRIC_SPECS), len(admitted), peer_slots


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "Crit_Class_Tracker_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n_metrics, n_admitted, cap = build(out)
    print(f"built {path}: {n_metrics} metrics x {cap} slots, "
          f"{n_admitted} seed peers admitted")
