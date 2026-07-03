#!/usr/bin/env python3
"""Email-simulate acceptance test (BUILD_SPEC_EDGAR.md Phase 9, --demo).

Copies ONLY the .xlsm into a fresh, otherwise-empty folder, then reproduces
exactly what the VBA "Extract" button does:
  1. read the _code_py tab (one source line per cell, column A) and write
     runner.py next to the workbook -- byte-for-byte the macro's
     WriteTabToFile;
  2. shell Python to run the EXTRACTED runner.py against the CLOSED workbook
     in --demo mode (deterministic, offline -- no User-Agent needed).
Then it composes (does NOT send) the monitoring email from the run digest and
asserts the email contains the RANKED CRITICIZED TABLE (with family flags and
the Tier-2 "n/a" rendering for non-grades_full families), the 8-K CREDIT
EVENTS section, the STALENESS section, and the UNMAPPED-MEMBER section, and
that the workbook rebuilt itself with nothing but the workbook + extracted
runner present (the workbook is the source of truth). Deterministic from
_config + EdgarDemoProvider at fixed --asof.
"""
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

import runner as R

XLSM_NAME = "Crit_Class_Tracker.xlsm"
ASOF = "2026-03-31"


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write UTF-8."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def _fmt(v, na="n/a"):
    return na if v is None else f"{v:.2f}"


def compose_email(status):
    """Deterministic monitoring email from the run status/digest. No send."""
    digest = status.get("digest", {})
    banks = digest.get("banks", [])
    fresh = [b for b in banks if not b.get("stale")]
    stale = [b for b in banks if b.get("stale")]
    lines = []
    lines.append("Subject: EDGAR Criticized/Classified Tracker -- "
                 f"{status['alert_banks']} ALERT / {status['watch_banks']} "
                 f"WATCH / {status['stale_banks']} STALE banks "
                 f"({status['timestamp']})")
    lines.append("")
    lines.append(f"Run mode: {status['mode']}  |  banks landed: "
                 f"{status['banks_landed']}/{status['banks_active']}  |  "
                 f"refused: {len(status.get('watchlist_refusals', []))}  |  "
                 f"N/A families: {status.get('na_banks', 0)}")
    lines.append(f"Data vintage: {status.get('vintage') or 'n/a'}  |  "
                 f"peer-set latest quarter: {digest.get('set_max_period')}  |"
                 f"  latest accession: {status.get('latest_accession')}")
    lines.append("")
    lines.append("RANKED CRITICIZED TABLE (by ALERT flags, then criticized "
                 "ratio; stale excluded; Tier-2 n/a = disclosure family, "
                 "never approximated)")
    lines.append("-" * 76)
    lines.append(f"  {'#':>2} {'Bank':<26} {'CIK':>8} {'Family':<16} "
                 f"{'Crit%':>6} {'dQoQ':>6} {'SM%':>6} {'Clsfd%':>6} "
                 f"{'8K':>2}  Status")
    ranked = sorted(fresh, key=lambda b: (-(b["alert_count"]),
                                          -(b["criticized_ratio"] or 0.0)))
    for i, b in enumerate(ranked, start=1):
        m = b["metrics"]
        lines.append(
            f"  {i:>2} {b['name']:<26} {b['cik']:>8} {b['family']:<16} "
            f"{_fmt(b['criticized_ratio']):>6} "
            f"{_fmt(m['crit_qoq_delta']['value']):>6} "
            f"{_fmt(m['sm_ratio']['value']):>6} "
            f"{_fmt(m['classified_ratio']['value']):>6} "
            f"{b['watch_8k']:>2}  {b['status']}")
    if not ranked:
        lines.append("  (no banks landed)")
    lines.append("")
    lines.append("FAMILY LIMITS (BUILD SPEC 0.6 -- what n/a means)")
    lines.append("-" * 52)
    fams = {}
    for b in fresh:
        fams.setdefault(b["family"], []).append(b["name"])
    for fam in ("grades_full", "criticized_only", "ig_nig", "unmapped"):
        if fam in fams:
            note = {"grades_full": "Tier 1 + Tier 2",
                    "criticized_only": "Tier 1 only; SM/classified n/a",
                    "ig_nig": "IG/NIG only; criticized n/a (MD&A text "
                              "fallback is v1.1)",
                    "unmapped": "no [MEMBER_MAP] rows -- map, then re-run"}
            lines.append(f"  {fam:<16} {note[fam]:<48} "
                         + "; ".join(sorted(fams[fam])))
    lines.append("")
    lines.append("8-K CREDIT EVENTS (trailing 4 quarters; 2.04/4.02/1.03 = "
                 "automatic WATCH)")
    lines.append("-" * 72)
    any_ev = False
    for b in sorted(banks, key=lambda x: x["slot"]):
        for item in R.HARVEST_ITEMS:
            ev = b["events"].get(item)
            if not ev:
                continue
            any_ev = True
            flag = "WATCH" if item in R.WATCH_ITEMS else "info"
            lines.append(f"  {b['name']:<26} item {item} "
                         f"({R.ITEM_LABEL[item]}) x{ev['count']} latest "
                         f"{ev['latest']} [{flag}] acc "
                         f"{ev['accessions'][0]}")
    if not any_ev:
        lines.append("  (no 8-K credit events in the window)")
    lines.append("")
    lines.append("STALENESS FLAGS (excluded from alert counts -- late filer "
                 "or delisted?)")
    lines.append("-" * 72)
    if stale:
        for b in sorted(stale, key=lambda x: x["slot"]):
            lines.append(f"  STALE  s{b['slot']:02d} {b['name']:<26} "
                         f"last CQI quarter {b['asof_period']}")
            for n in b.get("notes", []):
                lines.append(f"         {n}")
    else:
        lines.append("  (no stale banks -- every bank reports the peer-set "
                     "latest quarter)")
    lines.append("")
    lines.append("UNMAPPED MEMBERS (needs manual mapping in [MEMBER_MAP] -- "
                 "never guessed)")
    lines.append("-" * 72)
    any_um = False
    for b in sorted(banks, key=lambda x: x["slot"]):
        for q in b.get("unmapped_members", []):
            any_um = True
            lines.append(f"  {b['name']:<26} {q}  (in denominator only; "
                         "map the grade cell + re-run)")
        for q in b.get("unmapped_classes", []):
            any_um = True
            lines.append(f"  {b['name']:<26} {q}  (class member -- excluded "
                         "from rollups until mapped)")
    if not any_um:
        lines.append("  (none -- every observed member is mapped)")
    lines.append("")
    lines.append("WATCHLIST LANE")
    lines.append("--------------")
    refusals = status.get("watchlist_refusals", [])
    if refusals:
        for m in refusals:
            lines.append("  " + m)
    else:
        lines.append(f"  ADMITTED (CIK-keyed Class A): "
                     f"{len(status.get('watchlist_admitted', []))} banks")
    lines.append("")
    return "\n".join(lines)


def main():
    repo = os.path.dirname(os.path.abspath(__file__))
    xlsm = os.path.join(repo, XLSM_NAME)
    if not os.path.exists(xlsm):
        print("build the workbook first: python3 make_workbook.py",
              file=sys.stderr)
        return 1

    work = tempfile.mkdtemp(prefix="edgar_emailsim_")
    try:
        # 1. The email contains ONLY the workbook.
        dst_xlsm = os.path.join(work, XLSM_NAME)
        shutil.copy(xlsm, dst_xlsm)
        before = sorted(os.listdir(work))
        assert before == [XLSM_NAME], before
        print(f"[email-sim] fresh folder contains only: {before}")

        # 2. Button step 1: extract runner.py from _code_py.
        runner_py = os.path.join(work, "runner.py")
        extract_code_tab(dst_xlsm, "_code_py", runner_py)
        print(f"[email-sim] extracted runner.py "
              f"({os.path.getsize(runner_py)} bytes) from _code_py")

        # 3. Button step 2: shell Python to run the EXTRACTED runner (demo).
        proc = subprocess.run(
            [sys.executable, runner_py, "--workbook", dst_xlsm, "--demo",
             "--asof", ASOF],
            capture_output=True, text=True, cwd=work)
        last = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() \
            else "(none)"
        print("[email-sim] runner stderr:", last)
        if proc.returncode != 0:
            print("[email-sim] FAILED: runner exited", proc.returncode)
            print(proc.stdout, proc.stderr)
            return 2

        # 4. Verify the rebuild: raw data present, macro intact, no extras.
        wb = openpyxl.load_workbook(dst_xlsm, keep_vba=True)
        b1 = R.slot_block(1, R.RAW_SLOTS_DEFAULT)
        raw_ok = wb[R.RAW_TAB].cell(b1.first_data_row, 2).value is not None
        macro_ok = wb.vba_archive is not None
        produced = sorted(os.listdir(work))
        wb.close()
        print(f"[email-sim] folder now: {produced}")
        print(f"[email-sim] raw populated={raw_ok}  macro intact={macro_ok}")

        # 5. Compose the monitoring email from the digest (deterministic).
        status = R.run(dst_xlsm, demo=True,
                       asof=R.datetime.strptime(ASOF, "%Y-%m-%d").date())
        email = compose_email(status)
        email_path = os.path.join(work, "monitoring_email.txt")
        with open(email_path, "w", encoding="utf-8") as fh:
            fh.write(email)
        print("\n" + "=" * 76)
        print(email)
        print("=" * 76 + "\n")

        n_banks = len([b for b in status["digest"]["banks"]
                       if not b["stale"]])
        table_lines = [ln for ln in email.splitlines()
                       if ln.strip() and ln.strip()[0].isdigit()
                       and ("grades_full" in ln or "criticized_only" in ln
                            or "ig_nig" in ln or "unmapped" in ln)]
        has_table = ("RANKED CRITICIZED TABLE" in email
                     and len(table_lines) == n_banks
                     and "ALERT" in email)
        # Tier-2 N/A rendering: the criticized_only and ig_nig banks show n/a
        has_na = any("criticized_only" in ln and "n/a" in ln
                     for ln in table_lines) \
            and any("ig_nig" in ln and "n/a" in ln for ln in table_lines)
        has_events = ("8-K CREDIT EVENTS" in email and "item 2.04" in email
                      and "WATCH" in email)
        has_staleness = "STALENESS FLAGS" in email
        has_unmapped = ("UNMAPPED MEMBERS" in email
                        and R.DEMO_EXTENSION_MEMBER in email)
        has_vintage = "Data vintage:" in email
        self_contained = (raw_ok and macro_ok
                          and set(produced) <= {XLSM_NAME, "runner.py"})

        ok = (has_table and has_na and has_events and has_staleness
              and has_unmapped and has_vintage and self_contained)
        print(f"[email-sim] ranked criticized table : {has_table} "
              f"({len(table_lines)} rows)")
        print(f"[email-sim] family/Tier-2 n-a       : {has_na}")
        print(f"[email-sim] 8-K events section      : {has_events}")
        print(f"[email-sim] staleness section       : {has_staleness}")
        print(f"[email-sim] unmapped-member section : {has_unmapped}")
        print(f"[email-sim] data-vintage line       : {has_vintage}")
        print(f"[email-sim] workbook self-contained : {self_contained}")
        print("[email-sim] RESULT:",
              "PASS -- workbook is the source of truth" if ok else "FAIL")
        return 0 if ok else 3
    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
