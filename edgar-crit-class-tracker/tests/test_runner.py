"""Headless test suite for the EDGAR Criticized/Classified Tracker.

Runs on Linux/CI with NO Excel and NO network, all in --demo mode (any live
network call is stubbed). Mirrors the named tests in BUILD_SPEC_EDGAR.md
Section 6 (Phases 1-8):

  test_config_parse                  (Phase 1 -- maps + slots + over-capacity)
  test_demo_provider_deterministic   (Phase 2 -- families honored)
  test_submissions_parse             (Phase 2 -- columnar fixture + items +
                                      amendment dedupe + Q4-is-10-K)
  test_instance_parse                (Phase 2 -- fixture XML: standard +
                                      extension members, vintage IGNORED,
                                      hardcoded expected rollups)
  test_backoff                       (Phase 2 -- stubbed 429/5xx/403 + UA gate)
  test_member_bootstrap              (Phase 3 -- unmapped append, never guessed)
  test_metrics                       (Phase 4 -- hardcoded ratios/deltas +
                                      family Tier gating)
  test_reload_headless               (Phase 5)
  test_watchlist_entity_gates        (Phase 6 -- + defense in depth)
  test_stale_bank                    (Phase 6)
  test_8k_event_lane                 (Phase 6 -- items filter + auto-WATCH)
  test_provenance_and_tieout         (Phase 7)
  test_raw_landing_idempotent        (Phase 8)
  test_raw_layout_mismatch_refused   (Phase 8)
  test_clear_actually_blanks         (L7 regression)
  test_vba_protection_keys_roundtrip (carried)
"""
import json
import os
import re
import shutil
import struct
import sys
from dataclasses import replace
from datetime import date

import openpyxl
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import runner as R                     # noqa: E402
import member_seed as SEED             # noqa: E402
import build_workbook as BW            # noqa: E402
import assemble_xlsm                   # noqa: E402
import vba_writer as V                 # noqa: E402

ASOF = date(2026, 3, 31)

TABS = ("Dashboard_Criticized", "Dashboard_Mix_Events", "Watchlist",
        "Raw_EDGAR", "_config", "_provenance", "_code_py", "_code_vba",
        "_readme")

UA = "TestOrg test@example.com"


# --------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------
@pytest.fixture(scope="session")
def xlsm(tmp_path_factory):
    """Build the macro-enabled workbook once for the whole session."""
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    out = str(d / "Crit_Class_Tracker.xlsm")
    BW.build(base)
    assemble_xlsm.assemble(base, out)
    return out


@pytest.fixture(scope="session")
def populated(xlsm, tmp_path_factory):
    """A fresh copy of the workbook with a --demo run applied."""
    d = tmp_path_factory.mktemp("run")
    p = str(d / "run.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    return p


def _cfg():
    return R.parse_config(BW.config_rows())


def _peer(slot=11, cik="12345", name="Test Bank", ticker="TB", active=True):
    return R.PeerRow(slot=slot, cik=cik, name=name, ticker=ticker,
                     active=active)


def _edit_peer_row(path, slot, cik=None, name=None, ticker=None, active=None):
    """Edit a [PEERS] line in a BUILT workbook exactly as a user would."""
    wb = openpyxl.load_workbook(path, keep_vba=path.endswith(".xlsm"))
    ws = wb["_config"]
    in_peers = False
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v == "[PEERS]":
            in_peers = True
            continue
        if in_peers and isinstance(v, str) and v.startswith("["):
            break
        if in_peers and str(v).strip() == str(slot):
            target = row[0].row
            break
    assert target is not None, f"[PEERS] slot {slot} row not found"
    if cik is not None:
        ws.cell(target, 2, cik)
    if name is not None:
        ws.cell(target, 3, name)
    if ticker is not None:
        ws.cell(target, 4, ticker)
    if active is not None:
        ws.cell(target, 5, active)
    wb.save(path)
    wb.close()


# --------------------------------------------------------------------------
# Phase 1 -- config parse ([PEERS]/[MEMBER_MAP]/[CLASS_MAP] + capacity)
# --------------------------------------------------------------------------
def test_config_parse():
    cfg = _cfg()
    # [PEERS]: 20 provisioned slots, the 10 research-verified seed banks
    assert cfg.peer_slots == 20 and len(cfg.peers) == 20
    banked = [p for p in cfg.peers if p.has_bank]
    assert len(banked) == 10
    assert [p.slot for p in cfg.peers] == list(range(1, 21))
    fitb = next(p for p in banked if p.slot == 1)
    assert fitb.cik == "35527" and fitb.entity_key == "cik:35527"
    assert fitb.ticker == "FITB" and fitb.active is True
    for p in cfg.peers[10:]:
        assert not p.has_bank and not p.active     # provisioned headroom
    # [MEMBER_MAP]: families per research; standard + extension seeds
    assert cfg.family_of("35527") == "grades_full"
    assert cfg.family_of("91576") == "criticized_only"
    assert cfg.family_of("36270") == "criticized_only"
    assert cfg.family_of("36104") == "ig_nig"
    assert cfg.family_of("999999") == "unmapped"   # no rows -> unmapped
    assert cfg.grade_of("35527", "us-gaap:SpecialMentionMember") \
        == "special_mention"
    assert cfg.grade_of("35527", "us-gaap:CriticizedMember") == "ignore"
    assert cfg.grade_of("49196", "hban:OlemMember") == "special_mention"
    assert cfg.grade_of("91576", "key:CriticizedAccruingMember") \
        == "criticized_accruing"
    assert cfg.grade_of("36104", "us-gaap:InternalInvestmentGradeMember") \
        == "investment_grade"
    assert cfg.grade_of("35527", "fitb:NeverSeenMember") is None
    # [CLASS_MAP]: "*" wildcard standard members + excluded consumer books
    assert cfg.class_of("35527",
                        "us-gaap:CommercialPortfolioSegmentMember") == "ci"
    assert cfg.class_of("91576",
                        "us-gaap:ConstructionLoansMember") == "construction"
    assert cfg.class_of("35527",
                        "us-gaap:ResidentialPortfolioSegmentMember") \
        == "excluded"
    assert cfg.class_of("35527", "fitb:WeirdClassMember") is None
    # settings + thresholds (numeric, L8)
    assert cfg.raw_slots == 12
    assert cfg.user_agent == ""
    assert cfg.stale_multiplier == pytest.approx(2.0)
    assert set(cfg.thresholds) == set(R.BANDED_METRICS)
    assert cfg.thresholds["criticized_ratio"].watch == pytest.approx(4.0)
    assert cfg.thresholds["criticized_ratio"].alert == pytest.approx(6.0)
    assert cfg.thresholds["crit_qoq_delta"].alert == pytest.approx(1.0)
    assert cfg.thresholds["sm_ratio"].direction == "above"
    # slot expansion -> the s{slot:02d}_{metric} unit ids (seam shape)
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    assert len(admitted) == 10 and not refusals and not excluded
    q = R.QuarterRollup("2026-03-31", R.TAG_EXCL, "acc",
                        {c: 1.0 for c in R.RAW_METRIC_COLS}, [])
    rows = R.metric_rows_for_digest(1, "35527", [q])
    assert {r.id for r in rows} >= {"s01_criticized_ratio", "s01_sm_ratio"}
    assert all(r.geo_segment == "cik:35527" for r in rows)
    # over-capacity refused with the exact rebuild command, never truncated
    cfg.peers.append(_peer(slot=21, cik="99999", name="Overflow Bank"))
    with pytest.raises(R.PeerCapacityError) as ei:
        R.validate_peer_capacity(cfg)
    msg = str(ei.value)
    assert "Overflow Bank" in msg and "--peer-slots" in msg
    assert "make_workbook.py" in msg and "never truncates" in msg
    # duplicate slots refuse too
    cfg2 = _cfg()
    cfg2.peers.append(_peer(slot=1, cik="55555", name="Dup Bank"))
    with pytest.raises(R.PeerCapacityError, match="listed twice"):
        R.validate_peer_capacity(cfg2)


# --------------------------------------------------------------------------
# Phase 2 -- demo provider determinism + families honored
# --------------------------------------------------------------------------
def test_demo_provider_deterministic():
    a = R.EdgarDemoProvider(asof=ASOF, raw_slots=12)
    b = R.EdgarDemoProvider(asof=ASOF, raw_slots=12)
    for cik in ("35527", "91576", "36104", "28412"):
        fa = [(f.period, f.class_qname, f.grade_qname, f.value, f.accession)
              for f in a.fetch_facts(cik)]
        fb = [(f.period, f.class_qname, f.grade_qname, f.value, f.accession)
              for f in b.fetch_facts(cik)]
        assert fa == fb and fa                       # deterministic per key
    facts = a.fetch_facts("35527")
    periods = sorted({f.period for f in facts})
    assert len(periods) == 12 and periods[-1] == "2026-03-31"
    # deterministic accession shape "0000035527-26-000001"
    accs = sorted({f.accession for f in facts})
    assert accs[0] == "0000035527-26-000001"
    assert all(re.match(r"^\d{10}-26-\d{6}$", acc) for acc in accs)
    # Q4 arrives via the 10-K (trap E6)
    q4 = [f for f in facts if f.period == "2025-12-31"]
    assert q4 and all(f.form == "10-K" for f in q4)
    assert all(f.form == "10-Q" for f in facts if f.period == "2026-03-31")
    # families honored: grades_full emits pass/SM/substandard/doubtful
    g35 = {f.grade_qname for f in facts}
    assert {"us-gaap:PassMember", "us-gaap:SpecialMentionMember",
            "us-gaap:SubstandardMember", "us-gaap:DoubtfulMember"} <= g35
    # ... plus the ONE unseeded extension member, for exactly this bank
    assert R.DEMO_EXTENSION_MEMBER in g35
    for cik in ("1281761", "109380", "28412", "49196", "92230", "875357",
                "91576", "36270", "36104"):
        assert R.DEMO_EXTENSION_MEMBER not in \
            {f.grade_qname for f in a.fetch_facts(cik)}
    # Huntington emits its OLEM extension label instead of the standard SM
    g49 = {f.grade_qname for f in a.fetch_facts("49196")}
    assert "hban:OlemMember" in g49
    assert "us-gaap:SpecialMentionMember" not in g49
    # criticized_only banks emit pass + the two criticized buckets, no grades
    gk = {f.grade_qname for f in a.fetch_facts("91576")}
    assert gk == {"us-gaap:PassMember", "key:CriticizedAccruingMember",
                  "key:CriticizedNonaccrualMember"}
    # ig_nig banks emit ONLY IG/NIG members (Tier 1 AND 2 render N/A)
    gu = {f.grade_qname for f in a.fetch_facts("36104")}
    assert gu == {"us-gaap:InternalInvestmentGradeMember",
                  "us-gaap:InternalNoninvestmentGradeMember"}
    # events: one 2.02 everywhere; the 2.04 only on the designated bank
    ev_z = a.fetch_events(R.DEMO_WATCH_CIK)
    assert [e.item for e in ev_z] == ["2.04", "2.02"]     # newest-first
    assert all(e.item != "2.04" for e in a.fetch_events("35527"))


# --------------------------------------------------------------------------
# Phase 2 -- submissions parse (columnar fixture, offline)
# --------------------------------------------------------------------------
SUBMISSIONS_FIXTURE = {
    "cik": "35527",
    "name": "FIFTH THIRD BANCORP",
    "filings": {
        "recent": {
            # COLUMNAR parallel arrays -- the shape that bites (research)
            "form": ["10-Q", "8-K", "10-K", "10-Q/A", "10-Q", "8-K", "8-K"],
            "accessionNumber": [
                "0000035527-26-000010", "0000035527-26-000009",
                "0000035527-26-000008", "0000035527-25-000007",
                "0000035527-25-000006", "0000035527-25-000005",
                "0000035527-24-000004"],
            "filingDate": ["2026-05-05", "2026-02-11", "2026-02-25",
                           "2025-11-20", "2025-11-04", "2025-10-17",
                           "2024-07-01"],
            "reportDate": ["2026-03-31", "2026-02-11", "2025-12-31",
                           "2025-09-30", "2025-09-30", "2025-10-17",
                           "2024-06-30"],
            "acceptanceDateTime": [
                "2026-05-05T16:01:00.000Z", "2026-02-11T09:00:00.000Z",
                "2026-02-25T17:00:00.000Z", "2025-11-20T12:00:00.000Z",
                "2025-11-04T16:02:00.000Z", "2025-10-17T08:30:00.000Z",
                "2024-07-01T10:00:00.000Z"],
            "primaryDocument": ["fitb-20260331.htm", "fitb-8k.htm",
                                "fitb-20251231.htm", "fitb-20250930a.htm",
                                "fitb-20250930.htm", "fitb-8k2.htm",
                                "fitb-8k3.htm"],
            "items": ["", "2.02,2.04", "", "", "", "2.06,9.01",
                      "2.04"],
        },
        "files": [],
    },
}


def test_submissions_parse(monkeypatch):
    rows = R.EdgarProvider.parse_submissions(SUBMISSIONS_FIXTURE)
    assert len(rows) == 7
    assert rows[0]["form"] == "10-Q"
    assert rows[0]["accessionNumber"] == "0000035527-26-000010"
    # quarter selection: latest per calendar quarter, Q4 = the 10-K (E6),
    # the 10-Q/A supersedes its 10-Q by acceptanceDateTime (E7)
    asof = date(2026, 6, 30)
    chosen = R.EdgarProvider.select_quarterly(rows, 4, asof)
    assert [c["reportDate"] for c in chosen] == \
        ["2026-03-31", "2025-12-31", "2025-09-30"]
    assert chosen[1]["form"] == "10-K"                  # Q4-is-10-K
    assert chosen[2]["form"] == "10-Q/A"                # amendment wins
    assert chosen[2]["accessionNumber"] == "0000035527-25-000007"
    # 8-K items harvest: {1.03, 2.04, 2.06, 4.02, 2.02}, trailing window --
    # 9.01 is never harvested; the 2024 8-K falls outside the window
    events = R.EdgarProvider.harvest_events(rows, "35527", asof)
    assert [(e.date, e.item) for e in events] == \
        [("2026-02-11", "2.04"), ("2026-02-11", "2.02"),
         ("2025-10-17", "2.06")]
    assert all(e.cik == "35527" for e in events)
    # full prime() through stubbed transport: submissions -> index.json ->
    # extracted instance; ONE instance fetch per selected filing, cached
    calls = []

    def fake_get(self, url):
        calls.append(url)
        if "data.sec.gov/submissions" in url:
            return json.dumps(SUBMISSIONS_FIXTURE).encode()
        if url.endswith("index.json"):
            acc = url.rstrip("/").split("/")[-2]
            return json.dumps({"directory": {"item": [
                {"name": "ignore.htm"},
                {"name": f"doc-{acc}_htm.xml"}]}}).encode()
        assert url.endswith("_htm.xml")
        accnd = url.split("doc-")[1].split("_")[0]
        period = {"000003552726000010": "2026-03-31",
                  "000003552726000008": "2025-12-31",
                  "000003552725000007": "2025-09-30"}[accnd]
        return _instance_xml(period).encode()

    monkeypatch.setattr(R.EdgarProvider, "_http_get", fake_get)
    prov = R.EdgarProvider(user_agent=UA, min_interval=0.0, raw_slots=4)
    prov.asof = asof
    prov.prime(["35527"], asof)
    # CIK padding split (E8): 10-digit on data.sec.gov, unpadded /Archives
    assert any("data.sec.gov/submissions/CIK0000035527.json" in u
               for u in calls)
    assert any("www.sec.gov/Archives/edgar/data/35527/" in u for u in calls)
    assert not any("/edgar/data/0000035527/" in u for u in calls)
    facts = prov.fetch_facts("35527")
    assert {f.period for f in facts} == \
        {"2026-03-31", "2025-12-31", "2025-09-30"}
    assert {f.accession for f in facts if f.period == "2025-09-30"} == \
        {"0000035527-25-000007"}                       # the amendment's facts
    assert [e.item for e in prov.fetch_events("35527")] == \
        ["2.04", "2.02", "2.06"]
    n = len(calls)
    prov.prime(["35527"], asof)                        # per-URL cache
    assert len(calls) == n


# --------------------------------------------------------------------------
# Phase 2 -- extracted-instance XML parse (fixture; hardcoded rollups)
# --------------------------------------------------------------------------
def _instance_xml(period="2026-03-31"):
    """A miniature EDGAR extracted instance: standard members AND one
    extension member AND vintage line items that MUST be ignored, plus
    trap contexts (prior period, extra axis, class-only, subtotal)."""
    prior = "2025-12-31" if period != "2025-12-31" else "2025-09-30"
    CLASS_AXIS = ("us-gaap:FinancingReceivableRecordedInvestmentByClassOf"
                  "FinancingReceivableAxis")

    def ctx(cid, members, instant):
        dims = "".join(
            f'<xbrldi:explicitMember dimension="{ax}">{m}'
            f'</xbrldi:explicitMember>' for ax, m in members)
        return (f'<xbrli:context id="{cid}"><xbrli:entity>'
                f'<xbrli:identifier scheme="http://www.sec.gov/CIK">'
                f'0000035527</xbrli:identifier>'
                f'<xbrli:segment>{dims}</xbrli:segment></xbrli:entity>'
                f'<xbrli:period><xbrli:instant>{instant}</xbrli:instant>'
                f'</xbrli:period></xbrli:context>')

    CQI = "us-gaap:InternalCreditAssessmentAxis"
    ci = "us-gaap:CommercialPortfolioSegmentMember"
    cre = "us-gaap:CommercialRealEstatePortfolioSegmentMember"
    xs = []
    for cid, grade, cls in (
            ("c_pass_ci", "us-gaap:PassMember", ci),
            ("c_sm_ci", "us-gaap:SpecialMentionMember", ci),
            ("c_sub_ci", "us-gaap:SubstandardMember", ci),
            ("c_ext_ci", "fitb:CriticizedRestructuredMember", ci),
            ("c_crit_ci", "us-gaap:CriticizedMember", ci),
            ("c_pass_cre", "us-gaap:PassMember", cre),
            ("c_sm_cre", "us-gaap:SpecialMentionMember", cre),
            ("c_sub_cre", "us-gaap:SubstandardMember", cre)):
        xs.append(ctx(cid, [(CQI, grade), (CLASS_AXIS, cls)], period))
    xs.append(ctx("c_prior", [(CQI, "us-gaap:SpecialMentionMember"),
                              (CLASS_AXIS, ci)], prior))
    xs.append(ctx("c_geo", [(CQI, "us-gaap:SpecialMentionMember"),
                            (CLASS_AXIS, ci),
                            ("srt:StatementGeographicalAxis",
                             "country:US")], period))
    xs.append(ctx("c_classonly", [(CLASS_AXIS, ci)], period))

    T = ("us-gaap:FinancingReceivableExcludingAccruedInterest"
         "BeforeAllowanceForCreditLoss")
    TP = "us-gaap:FinancingReceivableBeforeAllowanceForCreditLoss"
    VINTAGE = ("us-gaap:FinancingReceivableExcludingAccruedInterest"
               "YearOneOriginatedCurrentFiscalYear")

    def fact(tag, cid, val):
        return (f'<{tag} contextRef="{cid}" unitRef="usd" decimals="-3">'
                f'{val}</{tag.split(" ")[0]}>')

    fx = [
        fact(T, "c_pass_ci", 90000), fact(T, "c_sm_ci", 4000),
        fact(T, "c_sub_ci", 5000), fact(T, "c_ext_ci", 1000),
        fact(T, "c_crit_ci", 9000),          # subtotal -> grade "ignore"
        fact(T, "c_pass_cre", 45000), fact(T, "c_sm_cre", 2000),
        fact(T, "c_sub_cre", 3000),
        fact(TP, "c_pass_ci", 88888),        # twin-tag minority (E2)
        fact(T, "c_prior", 3500),            # comparative column -> dropped
        fact(T, "c_geo", 7777),              # extra axis -> dropped
        fact(T, "c_classonly", 55555),       # no CQI axis -> dropped
        fact(VINTAGE, "c_sm_ci", 999999),    # vintage LINE ITEM -> IGNORED
        fact(VINTAGE, "c_sub_ci", 888888),
    ]
    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<xbrli:xbrl xmlns:xbrli="http://www.xbrl.org/2003/instance" '
        'xmlns:xbrldi="http://xbrl.org/2006/xbrldi" '
        'xmlns:us-gaap="http://fasb.org/us-gaap/2025" '
        'xmlns:fitb="http://www.53.com/20260331" '
        'xmlns:srt="http://fasb.org/srt/2025" '
        'xmlns:country="http://xbrl.sec.gov/country/2025">'
        + "".join(xs) + "".join(fx) + "</xbrli:xbrl>")


def test_instance_parse():
    data = _instance_xml().encode()
    facts = R.EdgarProvider.parse_instance(
        data, "35527", "2026-03-31", "0000035527-26-000010",
        "2026-05-05T16:01:00.000Z", "10-Q")
    # kept: 8 EXCL-tag facts + 1 primary-tag twin; dropped: prior period,
    # extra axis, class-only, and BOTH vintage line items (trap E3)
    assert len(facts) == 9
    assert all(f.accession == "0000035527-26-000010" for f in facts)
    by = {(f.tag, f.grade_qname, f.class_qname): f.value for f in facts}
    T, TP = R.TAG_EXCL, R.TAG_PRIMARY
    ci = "us-gaap:CommercialPortfolioSegmentMember"
    assert by[(T, "us-gaap:PassMember", ci)] == 90000.0
    assert by[(T, "fitb:CriticizedRestructuredMember", ci)] == 1000.0
    assert by[(TP, "us-gaap:PassMember", ci)] == 88888.0
    assert not any(v in (999999.0, 888888.0, 3500.0, 7777.0, 55555.0)
                   for v in by.values())
    # namespace-qualified member names survive (extension prefix intact)
    assert any(f.grade_qname.startswith("fitb:") for f in facts)

    # HARDCODED expected rollups through the seeded config (grades_full):
    # tag preference picks EXCL (8 facts vs 1 -- trap E2); the subtotal
    # CriticizedMember is grade "ignore" (never summed); the unmapped
    # extension member sits in the DENOMINATOR only.
    cfg = _cfg()
    quarters, um, uc = R.rollup_bank("35527", facts, cfg)
    assert um == {"fitb:CriticizedRestructuredMember"} and uc == set()
    q = quarters[0]
    assert q.tag == R.TAG_EXCL and q.period == "2026-03-31"
    v = q.values
    assert v["TOTAL_COMM_USD"] == pytest.approx(150000.0)   # incl. ext 1000
    assert v["CRIT_USD"] == pytest.approx(14000.0)          # excl. ext + subtotal
    assert v["CRIT_RATIO"] == pytest.approx(14000.0 / 150000.0 * 100)
    assert v["CLASS_USD"] == pytest.approx(8000.0)
    assert v["CLASS_RATIO"] == pytest.approx(8000.0 / 150000.0 * 100)
    assert v["SM_USD"] == pytest.approx(6000.0)
    assert v["SM_RATIO"] == pytest.approx(4.0)
    assert v["MIX_CI"] == pytest.approx(9000.0 / 14000.0 * 100)
    assert v["MIX_CRE"] == pytest.approx(5000.0 / 14000.0 * 100)
    assert v["MIX_CONSTR"] == pytest.approx(0.0)
    # deltas need history: single quarter -> None (blank), never zero
    assert v["CRIT_QOQ"] is None and v["CRIT_YOY"] is None


# --------------------------------------------------------------------------
# Phase 2 -- transport: backoff, 403-UA message, UA fail-fast
# --------------------------------------------------------------------------
def test_backoff(monkeypatch, xlsm, tmp_path):
    import urllib.error
    attempts = []

    def flaky_get(self, url):
        attempts.append(url)
        if len(attempts) < 3:
            raise urllib.error.HTTPError(url, 429, "Too Many Requests",
                                         None, None)
        return json.dumps(SUBMISSIONS_FIXTURE).encode()

    sleeps = []
    monkeypatch.setattr(R.EdgarProvider, "_http_get", flaky_get)
    monkeypatch.setattr(R.time, "sleep", lambda s: sleeps.append(s))
    prov = R.EdgarProvider(user_agent=UA, min_interval=0.0, max_retries=4)
    rows = prov._submission_rows("35527")
    assert len(attempts) == 3 and sleeps           # 2 x 429, then success
    assert rows and rows[0]["form"] == "10-Q"

    # persistent 5xx fails with a labeled error, never an endless retry
    def dead_get(self, url):
        raise urllib.error.HTTPError(url, 503, "down", None, None)
    monkeypatch.setattr(R.EdgarProvider, "_http_get", dead_get)
    prov2 = R.EdgarProvider(user_agent=UA, min_interval=0.0, max_retries=1)
    with pytest.raises(RuntimeError, match="HTTP 503"):
        prov2._submission_rows("35527")

    # 403 fails FAST with the User-Agent + allowlist guidance (no retries)
    calls_403 = []

    def forbidden_get(self, url):
        calls_403.append(url)
        raise urllib.error.HTTPError(url, 403, "Forbidden", None, None)
    monkeypatch.setattr(R.EdgarProvider, "_http_get", forbidden_get)
    prov3 = R.EdgarProvider(user_agent=UA, min_interval=0.0, max_retries=4)
    with pytest.raises(RuntimeError) as ei:
        prov3._submission_rows("35527")
    assert "User-Agent" in str(ei.value) and "--selftest" in str(ei.value)
    assert len(calls_403) == 1

    # UA REQUIRED live: blank edgar_user_agent fails fast (BUILD SPEC 0.3)
    with pytest.raises(R.UserAgentMissing, match="edgar_user_agent"):
        R.EdgarProvider(user_agent="")
    cfg = _cfg()
    assert cfg.user_agent == ""
    with pytest.raises(R.UserAgentMissing):
        R.make_provider(cfg, demo=False, asof=None)
    # ... and a live CLI run exits 3 (missing prerequisite), workbook intact
    p = str(tmp_path / "ua.xlsm")
    shutil.copy(xlsm, p)
    rc = R.main(["--workbook", p])
    assert rc == 3
    # the demo path never needs a UA
    assert isinstance(R.make_provider(cfg, demo=True, asof=None),
                      R.EdgarDemoProvider)
    # throttle floor: 0.15s minimum is enforced even if configured lower
    prov4 = R.EdgarProvider(user_agent=UA, min_interval=0.01)
    assert prov4._min_interval >= 0.15


# --------------------------------------------------------------------------
# Phase 3 -- member bootstrap (contract sec 13: visible, never guessed)
# --------------------------------------------------------------------------
def test_member_bootstrap(xlsm, tmp_path):
    p = str(tmp_path / "boot.xlsm")
    shutil.copy(xlsm, p)
    status = R.run(p, demo=True, asof=ASOF)
    # the demo's one unknown extension member surfaced, bank-attributed
    assert R.DEMO_EXTENSION_MEMBER in status["unmapped_members"]
    b = next(x for x in status["digest"]["banks"]
             if x["cik"] == R.DEMO_EXTENSION_CIK)
    assert b["unmapped_members"] == [R.DEMO_EXTENSION_MEMBER]
    assert any("needs manual mapping" in n.lower() or
               "UNMAPPED MEMBERS" in n for n in b["notes"])
    # ... APPENDED into [MEMBER_MAP] as grade="unmapped" (provisioned rows)
    def map_rows(path):
        wb = openpyxl.load_workbook(path, read_only=True)
        cfg2 = R.parse_config([[c for c in row] for row in
                               wb["_config"].iter_rows(values_only=True)])
        wb.close()
        return [(m.cik, m.family, m.member_qname, m.grade)
                for m in cfg2.member_map]
    rows1 = map_rows(p)
    hits = [r for r in rows1 if r[2] == R.DEMO_EXTENSION_MEMBER]
    assert hits == [(R.DEMO_EXTENSION_CIK, "grades_full",
                     R.DEMO_EXTENSION_MEMBER, "unmapped")]
    # NEVER guessed into a numerator: the unmapped member's USD sits in the
    # denominator only. Verify against the provider's own facts.
    prov = R.EdgarDemoProvider(asof=ASOF, raw_slots=12)
    facts = [f for f in prov.fetch_facts(R.DEMO_EXTENSION_CIK)
             if f.period == "2026-03-31"]
    ext = sum(f.value for f in facts
              if f.grade_qname == R.DEMO_EXTENSION_MEMBER)
    total = sum(f.value for f in facts)
    crit = sum(f.value for f in facts if f.grade_qname in (
        "us-gaap:SpecialMentionMember", "us-gaap:SubstandardMember",
        "us-gaap:DoubtfulMember"))
    assert ext > 0
    assert b["metrics"]["criticized_ratio"]["value"] == \
        pytest.approx(crit / total * 100)
    assert b["metrics"]["criticized_ratio"]["value"] < \
        (crit + ext) / total * 100
    # idempotent: a second run appends NOTHING new, still flags it
    status2 = R.run(p, demo=True, asof=ASOF)
    assert map_rows(p) == rows1
    assert R.DEMO_EXTENSION_MEMBER in status2["unmapped_members"]
    # family N/A gating survives the bootstrap: ig_nig stays all-N/A
    usb = next(x for x in status2["digest"]["banks"] if x["cik"] == "36104")
    assert usb["family"] == "ig_nig" and usb["status"] == "N/A"
    assert usb["metrics"]["criticized_ratio"]["value"] is None


# --------------------------------------------------------------------------
# Phase 4 -- metrics: hardcoded ratios/deltas + family Tier gating
# --------------------------------------------------------------------------
def _mini_cfg():
    cfg = R.Config()
    cfg.thresholds = dict(_cfg().thresholds)
    for q, g in (("us-gaap:PassMember", "pass"),
                 ("us-gaap:SpecialMentionMember", "special_mention"),
                 ("us-gaap:SubstandardMember", "substandard")):
        cfg.member_map.append(R.MemberMapRow("999", "grades_full", q, g))
    for q, g in (("us-gaap:PassMember", "pass"),
                 ("x:CritAccrMember", "criticized_accruing"),
                 ("x:CritNonaccrMember", "criticized_nonaccruing")):
        cfg.member_map.append(R.MemberMapRow("998", "criticized_only", q, g))
    for q, g in (("us-gaap:InternalInvestmentGradeMember",
                  "investment_grade"),
                 ("us-gaap:InternalNoninvestmentGradeMember",
                  "noninvestment_grade")):
        cfg.member_map.append(R.MemberMapRow("997", "ig_nig", q, g))
    cfg.class_map.append(R.ClassMapRow(
        "*", "us-gaap:CommercialPortfolioSegmentMember", "ci"))
    cfg.class_map.append(R.ClassMapRow(
        "*", "us-gaap:ResidentialPortfolioSegmentMember", "excluded"))
    return cfg


def _facts_for(cik, ratios, members, total=100000.0):
    """One class, fixed total; criticized split across `members` weights."""
    periods = ["2026-03-31", "2025-12-31", "2025-09-30", "2025-06-30",
               "2025-03-31"][:len(ratios)]
    ci = "us-gaap:CommercialPortfolioSegmentMember"
    out = []
    for i, (p, r) in enumerate(zip(periods, ratios)):
        crit = total * r / 100.0
        acc = f"acc-{i}"
        for q, w in members:
            out.append(R.Fact(cik, p, R.TAG_EXCL, ci, q, crit * w, acc,
                              "2026-01-01T00:00:00Z", "10-Q"))
        out.append(R.Fact(cik, p, R.TAG_EXCL, ci, "us-gaap:PassMember",
                          total - crit, acc, "2026-01-01T00:00:00Z", "10-Q"))
    return out


def test_metrics():
    cfg = _mini_cfg()
    # grades_full: ratios newest-first 10, 8, 7, 6, 5 (SM 40% / sub 60%)
    facts = _facts_for("999", [10.0, 8.0, 7.0, 6.0, 5.0],
                       [("us-gaap:SpecialMentionMember", 0.4),
                        ("us-gaap:SubstandardMember", 0.6)])
    quarters, um, uc = R.rollup_bank("999", facts, cfg)
    assert not um and not uc
    v = quarters[0].values
    assert v["CRIT_RATIO"] == pytest.approx(10.0)
    assert v["SM_RATIO"] == pytest.approx(4.0)
    assert v["CLASS_RATIO"] == pytest.approx(6.0)
    assert v["CRIT_QOQ"] == pytest.approx(2.0)          # 10 - 8 pp
    assert v["CRIT_YOY"] == pytest.approx(5.0)          # 10 - 5 pp
    assert v["CRIT_GROWTH"] == pytest.approx(25.0)      # 10000/8000 - 1
    assert v["MIX_CI"] == pytest.approx(100.0)
    assert quarters[-1].values["CRIT_QOQ"] is None      # no prior -> blank
    # criticized_only: Tier 1 computes, Tier 2 renders None (N/A -- 0.6)
    f2 = _facts_for("998", [6.5, 6.0], [("x:CritAccrMember", 0.7),
                                        ("x:CritNonaccrMember", 0.3)])
    q2, _, _ = R.rollup_bank("998", f2, cfg)
    v2 = q2[0].values
    assert v2["CRIT_RATIO"] == pytest.approx(6.5)
    assert v2["CRIT_QOQ"] == pytest.approx(0.5)
    assert v2["SM_RATIO"] is None and v2["CLASS_RATIO"] is None
    assert v2["SM_USD"] is None and v2["CLASS_USD"] is None
    # ig_nig: Tier 1 AND Tier 2 N/A -- never approximated
    ig = "us-gaap:InternalInvestmentGradeMember"
    nig = "us-gaap:InternalNoninvestmentGradeMember"
    f3 = [R.Fact("997", "2026-03-31", R.TAG_EXCL,
                 "us-gaap:CommercialPortfolioSegmentMember", m, val,
                 "acc", "t", "10-Q")
          for m, val in ((ig, 80000.0), (nig, 20000.0))]
    q3, _, _ = R.rollup_bank("997", f3, cfg)
    v3 = q3[0].values
    assert v3["CRIT_RATIO"] is None and v3["SM_RATIO"] is None
    assert v3["TOTAL_COMM_USD"] == pytest.approx(100000.0)  # total still lands
    # excluded classes stay out of the denominators
    resi = R.Fact("999", "2026-03-31", R.TAG_EXCL,
                  "us-gaap:ResidentialPortfolioSegmentMember",
                  "us-gaap:PassMember", 5.0e6, "acc-0", "t", "10-Q")
    q4, _, _ = R.rollup_bank("999", facts + [resi], cfg)
    assert q4[0].values["TOTAL_COMM_USD"] == pytest.approx(100000.0)
    assert q4[0].values["CRIT_RATIO"] == pytest.approx(10.0)
    # totals-only fallback (no class axis): ratio computes, mix stays blank
    f5 = [R.Fact("999", "2026-03-31", R.TAG_EXCL, "", g, val, "a", "t",
                 "10-Q")
          for g, val in (("us-gaap:PassMember", 92000.0),
                         ("us-gaap:SubstandardMember", 8000.0))]
    q5, _, _ = R.rollup_bank("999", f5, cfg)
    assert q5[0].values["CRIT_RATIO"] == pytest.approx(8.0)
    assert q5[0].values["MIX_CI"] is None
    # threshold engine (direction-aware, blank-tolerant)
    thr = cfg.thresholds["criticized_ratio"]
    assert R.status_for(6.5, thr) == "ALERT"
    assert R.status_for(4.2, thr) == "WATCH"
    assert R.status_for(2.0, thr) == "OK"
    assert R.status_for(None, thr) == "OK"
    qoq = cfg.thresholds["crit_qoq_delta"]
    assert R.status_for(1.2, qoq) == "ALERT"
    assert R.status_for(0.6, qoq) == "WATCH"


# --------------------------------------------------------------------------
# Phase 5 -- headless reload (Excel proxy) + NO native charts
# --------------------------------------------------------------------------
def test_reload_headless(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    assert wb.vba_archive is not None              # macro survived
    assert tuple(wb.sheetnames) == TABS            # tab set EXACT
    for ws in wb.worksheets:
        assert getattr(ws, "_charts", []) == []    # zero native charts (L4)
    # every (slot, metric) dashboard cell carries a formula; family-gated
    # Tier 2 columns carry the N/A rendering IN FORMULA
    ws = wb["Dashboard_Criticized"]
    for slot in (1, 10, 20):
        r = BW.dash_row(slot)
        for c in range(5, 11):
            v = ws.cell(r, c).value
            assert isinstance(v, str) and v.startswith("="), (slot, c)
        assert "_config" in str(ws.cell(r, 1).value)       # identity by formula
        assert "N/A (family)" in str(ws.cell(r, 5).value)  # Tier 1 gate
        assert "grades_full" in str(ws.cell(r, 9).value)   # Tier 2 gate
    med = ws.cell(BW.dash_row(20) + 1, 5).value
    assert isinstance(med, str) and "MEDIAN(" in med
    # Watchlist: helpers, 8-K counts, rank, gated status with --lookup hint
    wl = wb["Watchlist"]
    for slot in (1, 10, 20):
        r = BW.wl_row(slot)
        assert "COUNTIF" in str(wl.cell(r, 6).value)       # 8-K flags
        assert "RANK" in str(wl.cell(r, 7).value)
        st = str(wl.cell(r, 8).value)
        assert "REFUSED" in st and "--lookup" in st and "OFF" in st
        assert "N/A (family)" in st
    for k, mid in enumerate(R.BANDED_METRICS):
        assert wl.cell(BW.WL_HDR, BW.WL_HELPER_COL0 + k).value == mid
    # Mix/Events tab: per-item COUNTIFs off the event blocks
    mx = wb["Dashboard_Mix_Events"]
    assert "COUNTIF" in str(mx.cell(BW.dash_row(1), 9).value)
    assert "2.04" in str(mx.cell(BW.dash_row(1), 9).value)
    # threshold cells are NUMERIC (L8)
    cws = wb["_config"]
    in_thr = False
    checked = 0
    for row in cws.iter_rows(min_col=1, max_col=3):
        v = row[0].value
        if v == "[THRESHOLDS]":
            in_thr = True
            continue
        if in_thr and isinstance(v, str) and v.startswith("["):
            break
        if in_thr and v in R.METRICS:
            assert isinstance(row[1].value, (int, float)), v
            assert isinstance(row[2].value, (int, float)), v
            checked += 1
    assert checked == len(R.BANDED_METRICS)
    # raw scaffold labels every block (metric sentinel, facts, events)
    raw = wb["Raw_EDGAR"]
    b = R.slot_block(7, 12)
    assert raw.cell(b.header_row, 1).value == "slot07"
    assert raw.cell(b.label_row, R.metric_col("CRIT_RATIO")).value \
        == "CRIT_RATIO"
    assert raw.cell(b.label_row, R.metric_col("TAG")).value == "TAG"
    assert str(raw.cell(b.facts_label_row, 1).value).startswith("facts")
    assert str(raw.cell(b.events_label_row, 1).value).startswith("events")
    wb.close()


# --------------------------------------------------------------------------
# Phase 6 -- entity gates: positive admission + every refusal shape
# --------------------------------------------------------------------------
def test_watchlist_entity_gates():
    cfg = _cfg()
    admitted, refusals, excluded = R.evaluate_peers(cfg)
    assert len(admitted) == 10 and refusals == [] and excluded == []
    assert all(re.match(R.ENTITY_KEY_PATTERN, p.entity_key)
               for p in admitted)

    # blank cik: refused BY NAME with the --lookup hint (never fetched)
    cfg.peers[10] = _peer(slot=11, cik="", name="Mystery Trust Co")
    admitted2, refusals2, _ = R.evaluate_peers(cfg)
    assert len(admitted2) == 10 and len(refusals2) == 1
    msg = refusals2[0][1]
    assert 'peer slot 11 "Mystery Trust Co"' in msg
    assert 'cik=""' in msg
    assert "cik:[0-9]{1,10}" in msg                # the whitelist, named
    assert '--lookup "Mystery Trust Co"' in msg
    assert "company_tickers.json" in msg

    # inactive slot: EXCLUDED (blanked), not refused
    cfg3 = _cfg()
    idx = next(i for i, p in enumerate(cfg3.peers) if p.slot == 2)
    cfg3.peers[idx] = replace(cfg3.peers[idx], active=False)
    admitted3, refusals3, excluded3 = R.evaluate_peers(cfg3)
    assert len(admitted3) == 9 and refusals3 == []
    assert [p.slot for p in excluded3] == [2]

    # malformed entity keys are refused default-deny (series-named)
    for cik in ("12345678901", "abc", "12-34", "62 8", "cik:628", "FITB"):
        reasons = R.gate_peer_row(_peer(slot=12, cik=cik, name="Bad Key"))
        assert reasons and "Gate3" in reasons[0], cik

    # defense in depth: a fabricated non-"A" metric spec refuses the run
    bad = replace(R.METRICS["criticized_ratio"], source_class="C")
    with pytest.raises(R.WatchlistRefused) as ei:
        R.assert_metrics_admissible([bad])
    assert 'metric "criticized_ratio"' in str(ei.value)
    assert 'source_class="C"' in str(ei.value)
    assert "its own review" in str(ei.value)
    bad2 = replace(R.METRICS["sm_ratio"], watchlist_capable=False)
    assert any("Gate1" in r for r in R.gate_metric_spec(bad2))

    # the build-time hard gate backs the runtime gate
    cfg5 = _cfg()
    cfg5.peers[11] = _peer(slot=12, cik="not-a-cik", name="Aggregate Row")
    with pytest.raises(R.WatchlistRefused, match="Aggregate Row"):
        R.assert_entity_gates(cfg5)


# --------------------------------------------------------------------------
# Phase 6 -- staleness (late filer / delisted)
# --------------------------------------------------------------------------
class FrozenPeerProvider(R.EdgarDemoProvider):
    """Demo variant freezing ONE bank three quarters behind the peer-set max
    -- the shape of a delisted/late filer (data returned, no error, the
    latest CQI quarter just stops advancing)."""

    def __init__(self, frozen_cik, **kw):
        super().__init__(**kw)
        self.frozen_cik = str(frozen_cik)
        y, m = self.asof.year, self.asof.month - 9
        if m < 1:
            m += 12
            y -= 1
        self._old = R.EdgarDemoProvider(asof=date(y, m, 28),
                                        raw_slots=self.raw_slots)

    def fetch_facts(self, cik):
        if str(cik) == self.frozen_cik:
            return self._old.fetch_facts(cik)
        return super().fetch_facts(cik)

    def fetch_events(self, cik):
        if str(cik) == self.frozen_cik:
            return self._old.fetch_events(cik)
        return super().fetch_events(cik)


def test_stale_bank(xlsm, tmp_path):
    # freeze the demo ALERT-tier bank (Comerica): fresh it ALERTs; frozen it
    # must carry NONE of those flags into the KPIs.
    p = str(tmp_path / "stale.xlsm")
    shutil.copy(xlsm, p)
    frozen = FrozenPeerProvider(R.DEMO_ALERT_CIK, asof=ASOF, raw_slots=12)
    fs = R.run(p, demo=True, asof=ASOF, provider=frozen)
    p2 = str(tmp_path / "fresh.xlsm")
    shutil.copy(xlsm, p2)
    fr = R.run(p2, demo=True, asof=ASOF)

    fb = next(b for b in fs["digest"]["banks"]
              if b["cik"] == R.DEMO_ALERT_CIK)
    assert fb["stale"] is True and fb["status"] == "STALE"
    assert fb["asof_period"] < fs["digest"]["set_max_period"]
    assert any("late filer or delisted" in n for n in fb["notes"])
    fresh_b = next(b for b in fr["digest"]["banks"]
                   if b["cik"] == R.DEMO_ALERT_CIK)
    assert fresh_b["status"] == "ALERT" and fresh_b["stale"] is False
    assert fs["stale_banks"] == 1 and fr["stale_banks"] == 0
    # STALE excluded from alert counts (spec sec 3)
    assert fs["alert_banks"] == fr["alert_banks"] - 1
    assert fs["alert_flags"] == fr["alert_flags"] - fresh_b["alert_count"]
    # digest medians exclude the stale bank (its high ratio drops out); the
    # EXCEL median row still includes it -- documented in _readme
    assert fs["digest"]["medians"]["criticized_ratio"] < \
        fr["digest"]["medians"]["criticized_ratio"]
    wb = openpyxl.load_workbook(p, keep_vba=True)
    readme = "\n".join(str(c.value) for row in wb["_readme"].iter_rows()
                       for c in row if c.value)
    assert "median" in readme.lower() and "STALE" in readme
    # the frozen slot's data IS still landed (audit trail)
    b4 = R.slot_block(4, 12)
    assert wb["Raw_EDGAR"].cell(b4.first_data_row, 2).value is not None
    wb.close()
    # the primitive itself
    assert R.is_stale_bank("2025-06-30", "2026-03-31", 2.0) is True
    assert R.is_stale_bank("2025-12-31", "2026-03-31", 2.0) is False
    assert R.is_stale_bank(None, "2026-03-31", 2.0) is True
    assert R.is_stale_bank("2026-03-31", None, 2.0) is False


# --------------------------------------------------------------------------
# Phase 6 -- the 8-K event lane (items filter + automatic WATCH)
# --------------------------------------------------------------------------
def test_8k_event_lane(populated):
    # items filtering is covered on the fixture in test_submissions_parse;
    # here: the lane end-to-end (digest + landed event blocks + auto-WATCH).
    status = R.run(populated, demo=True, asof=ASOF)
    dig = status["digest"]
    z = next(b for b in dig["banks"] if b["cik"] == R.DEMO_WATCH_CIK)
    assert z["events"]["2.04"]["count"] == 1
    assert z["watch_8k"] == 1
    assert any("automatic WATCH" in n for n in z["notes"])
    assert z["status"] == "WATCH"
    # 2.06 is harvested + displayed but does NOT auto-WATCH
    cma = next(b for b in dig["banks"] if b["cik"] == R.DEMO_ALERT_CIK)
    assert cma["events"]["2.06"]["count"] == 1 and cma["watch_8k"] == 0
    assert cma["status"] == "ALERT"
    # landed event rows: newest-first, item + accession + /Archives URL
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    raw = wb["Raw_EDGAR"]
    b3 = R.slot_block(3, 12)                    # Zions is slot 3
    assert raw.cell(b3.first_event_row, R.event_col("item")).value == "2.04"
    acc = raw.cell(b3.first_event_row, R.event_col("accession")).value
    assert acc == "0000109380-26-000102"
    url = raw.cell(b3.first_event_row, R.event_col("url")).value
    assert url == ("https://www.sec.gov/Archives/edgar/data/109380/"
                   "000010938026000102/")
    wb.close()
    # auto-WATCH in isolation: an otherwise-OK bank with a 2.04 goes WATCH
    cfg = _mini_cfg()
    peer = R.PeerRow(1, "999", "OK Bank", "OKB", True)
    q = R.QuarterRollup("2026-03-31", R.TAG_EXCL, "acc-1",
                        {**{c: None for c in R.RAW_METRIC_COLS},
                         "CRIT_RATIO": 2.0}, [])
    ev = R.Event("999", "2026-02-01", "2.04", "8-K", "acc-9")
    d_ok = R.compute_digest(cfg, {1: (peer, "grades_full", [q], [], set(),
                                      set())})
    d_ev = R.compute_digest(cfg, {1: (peer, "grades_full", [q], [ev], set(),
                                      set())})
    assert d_ok["banks"][0]["status"] == "OK"
    assert d_ev["banks"][0]["status"] == "WATCH"
    # ... and an 8-K credit event on an ig_nig bank still surfaces as WATCH
    peer2 = R.PeerRow(1, "997", "IG Bank", "IGB", True)
    qn = R.QuarterRollup("2026-03-31", R.TAG_EXCL, "acc-1",
                         {c: None for c in R.RAW_METRIC_COLS}, [])
    d_ig = R.compute_digest(cfg, {1: (peer2, "ig_nig", [qn], [ev], set(),
                                      set())})
    assert d_ig["banks"][0]["status"] == "WATCH"
    d_ig2 = R.compute_digest(cfg, {1: (peer2, "ig_nig", [qn], [], set(),
                                       set())})
    assert d_ig2["banks"][0]["status"] == "N/A"


# --------------------------------------------------------------------------
# Phase 7 -- provenance + tie-out (contract sec 12)
# --------------------------------------------------------------------------
def test_provenance_and_tieout(populated, capsys):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    assert wb.sheetnames.index("_provenance") == \
        wb.sheetnames.index("_config") + 1
    ws = wb["_provenance"]
    text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                     if c.value is not None)
    # the QUOTED uniform interagency definitions, with citations
    assert "2004" in text and "Uniform Agreement" in text
    assert "SR 93-30" in text and "Rating Credit Risk" in text
    assert "not adversely classified" in text          # SM quote
    assert "well-defined weakness" in text             # Substandard quote
    assert "highly questionable and improbable" in text  # Doubtful quote
    assert "basically worthless asset" in text         # Loss quote
    assert "CRITICIZED = Special Mention + Classified" in text
    # viewer URL patterns (browse + /Archives filing directory)
    assert "cgi-bin/browse-edgar?action=getcompany" in text
    assert re.search(r"https://www\.sec\.gov/Archives/edgar/data/"
                     r"\{CIK\}/\{ACCESSION-NODASH\}/", text)
    assert "ASC 326-20-50-5" in text
    # every canonical metric has a structured provenance row
    rows = R.read_provenance_rows(wb)
    wb.close()
    missing = [m for m in R.METRICS if m not in rows]
    assert not missing, f"metrics without a provenance row: {missing}"
    for m in ("criticized_ratio", "classified_ratio", "sm_ratio"):
        assert "Credit Quality Indicators" in rows[m]["source"]
        assert "accession" in rows[m]["accession"]
    assert "InternalCreditAssessmentAxis" in \
        rows["criticized_ratio"]["tag"]

    # --tieout in demo mode: LOUD demo label, value + tag + member breakdown
    # + accession + the /Archives viewer URL
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "28412"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "DEMO VALUES" in out and "FICTION" in out
    assert "family grades_full" in out
    assert ("https://www.sec.gov/Archives/edgar/data/28412/"
            "000002841226000012/") in out
    assert "0000028412-26-000012" in out               # newest accession
    assert R.TAG_EXCL in out
    assert "us-gaap:SubstandardMember" in out          # member breakdown
    assert "MEMBER BREAKDOWN" in out
    for m in R.METRICS:
        assert m in out, m
    # engine parity: the tieout criticized value equals the digest's
    status = R.run(populated, demo=True, asof=ASOF)
    cma = next(b for b in status["digest"]["banks"] if b["cik"] == "28412")
    assert f"{cma['criticized_ratio']:,.2f}" in out
    # quarter selection works, both ISO and YYYYQn; bad quarter refused
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "28412",
                 "2025Q4"])
    out2 = capsys.readouterr().out
    assert rc == 0 and "quarter 2025-12-31" in out2
    assert "0000028412-26-000011" in out2
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "28412",
                 "1999-12-31"])
    err = capsys.readouterr().err
    assert rc == 1 and "not in the fetched window" in err
    # malformed CIK refused with the --lookup hint
    rc = R.main(["--workbook", populated, "--demo", "--tieout", "not-a-cik"])
    err = capsys.readouterr().err
    assert rc == 2 and "--lookup" in err
    # offline --lookup: clear message + the seeded index answers
    rc = R.main(["--lookup", "KeyCorp", "--demo"])
    out3 = capsys.readouterr().out
    assert rc == 0 and "OFFLINE LOOKUP" in out3
    assert "91576" in out3 and "company_tickers.json" in out3


# --------------------------------------------------------------------------
# Phase 8 -- idempotence + layout guard + L7 clearing
# --------------------------------------------------------------------------
def _landed(path, tab="Raw_EDGAR"):
    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws = wb[tab]
    vals = [(c.row, c.column, c.value) for row in ws.iter_rows()
            for c in row if c.value is not None]
    wb.close()
    return vals


def test_raw_landing_idempotent(xlsm, tmp_path):
    # TRUE idempotence: a second run on the SAME already-populated workbook
    # lands identical values (slot clearing leaves no stale tail), and the
    # member bootstrap appends nothing new.
    p1 = str(tmp_path / "r1.xlsm")
    shutil.copy(xlsm, p1)
    R.run(p1, demo=True, asof=ASOF)
    first = _landed(p1)
    cfg_first = _landed(p1, "_config")
    R.run(p1, demo=True, asof=ASOF)
    assert _landed(p1) == first
    assert _landed(p1, "_config") == cfg_first
    # ... and determinism across fresh copies
    p2 = str(tmp_path / "r2.xlsm")
    shutil.copy(xlsm, p2)
    R.run(p2, demo=True, asof=ASOF)
    assert _landed(p2) == first
    # newest-first: slot 1's first metric row is the asof quarter, with the
    # per-quarter accession + tag provenance columns filled
    wb = openpyxl.load_workbook(p1, keep_vba=True)
    b1 = R.slot_block(1, 12)
    raw = wb["Raw_EDGAR"]
    assert raw.cell(b1.first_data_row, 1).value == "2026-03-31"
    assert raw.cell(b1.first_data_row,
                    R.metric_col("ACCESSION")).value == \
        "0000035527-26-000012"
    assert raw.cell(b1.first_data_row, R.metric_col("TAG")).value == \
        R.TAG_EXCL
    assert raw.cell(b1.header_row, 4).value == "grades_full"
    # fact block carries the member-level audit trail
    assert raw.cell(b1.first_fact_row, R.fact_col("member_qname")).value
    assert raw.cell(b1.first_fact_row, R.fact_col("accession")).value
    wb.close()


def test_raw_layout_mismatch_refused(xlsm, tmp_path):
    """Editing raw_slots after build must be REFUSED (every formula is
    anchored to the built layout), not silently mis-written."""
    p = str(tmp_path / "mismatch.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == "raw_slots":
            target = row[0].row
            break
    assert target is not None
    ws.cell(target, 2, 8)
    wb.save(p)
    wb.close()
    with pytest.raises(RuntimeError, match="Raw layout mismatch"):
        R.run(p, demo=True, asof=ASOF)


def test_clear_actually_blanks(populated, tmp_path):
    """L7 regression: openpyxl's ws.cell(r, c, None) silently IGNORES None.
    A deactivated bank's slot must ACTUALLY blank -- identity, metric rows,
    fact rows AND event rows -- never last run's values under a fresh
    timestamp."""
    p = str(tmp_path / "clear.xlsm")
    shutil.copy(populated, p)
    b3 = R.slot_block(3, 12)                       # Zions: has events too
    wb = openpyxl.load_workbook(p, keep_vba=True)
    raw = wb["Raw_EDGAR"]
    assert raw.cell(b3.header_row, 2).value is not None       # pre: populated
    assert raw.cell(b3.first_data_row, 2).value is not None
    assert raw.cell(b3.first_fact_row, 1).value is not None
    assert raw.cell(b3.first_event_row, 1).value is not None
    wb.close()
    _edit_peer_row(p, 3, active="FALSE")
    status = R.run(p, demo=True, asof=ASOF)
    assert status["banks_excluded"] == 1
    wb = openpyxl.load_workbook(p, keep_vba=True)
    raw = wb["Raw_EDGAR"]
    for c in (2, 3, 4, 5, 6):                      # identity cells
        assert raw.cell(b3.header_row, c).value is None
    for r in range(b3.first_data_row, b3.first_data_row + 12):
        for c in range(1, 2 + len(R.RAW_METRIC_COLS)):
            assert raw.cell(r, c).value is None
    for r in range(b3.first_fact_row, b3.first_fact_row + R.FACT_ROWS):
        for c in range(1, len(R.FACT_COLS) + 1):
            assert raw.cell(r, c).value is None
    for r in range(b3.first_event_row, b3.first_event_row + R.EVENT_ROWS):
        for c in range(1, len(R.EVENT_COLS) + 1):
            assert raw.cell(r, c).value is None
    # the label scaffold survives (only runtime state is cleared)
    assert raw.cell(b3.header_row, 1).value == "slot03"
    assert raw.cell(b3.label_row, R.metric_col("TAG")).value == "TAG"
    wb.close()


# --------------------------------------------------------------------------
# Carried -- VBA project protection keys ([MS-OVBA] 2.4.3)
# --------------------------------------------------------------------------
def test_vba_protection_keys_roundtrip():
    proj = V.build_project_stream(
        "X", [V.Module("X", "Sub A()\nEnd Sub")]).decode("latin-1")
    keys = {}
    for line in proj.splitlines():
        for k in ("CMG", "DPB", "GC"):
            if line.startswith(f'{k}="'):
                keys[k] = line.split('"')[1]
    assert V.ovba_decrypt(keys["CMG"]) == struct.pack("<I", 0)  # unprotected
    assert V.ovba_decrypt(keys["DPB"]) == b"\x00"               # no password
    assert V.ovba_decrypt(keys["GC"]) == b"\xff"                # visible
    assert len(V.build_vba_project_stream()) == 7
    d = V.build_dir_stream("X", [V.Module("X", "Sub A()\nEnd Sub")])
    assert b"Visual Basic For Applications" in d
    assert b"Microsoft Excel" in d
