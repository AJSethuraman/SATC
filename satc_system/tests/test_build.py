"""Build-level tests: the demo workbook assembles and wires formulas correctly.

These do not require LibreOffice (formula *evaluation* is checked separately via
scripts/recalc.py). They verify structure, prefilled inputs, and that every
formula token resolved to a real reference (no leftover ``{...}`` / ``[XW ...]``).
"""

from __future__ import annotations

import re

from openpyxl import load_workbook

from satc.build import build_demo_workbook


def test_demo_workbook_builds(tmp_path):
    out = build_demo_workbook(tmp_path / "wb.xlsx", tax_year=2024)
    assert out.exists()
    wb = load_workbook(out)
    names = wb.sheetnames
    for sheet in ("Cover", "Staging", "Data Mart", "Prior-vs-Current", "Proforma",
                  "Client Delivery", "Document Repository", "Dashboards"):
        assert sheet in names, f"missing {sheet} sheet"
    assert any(n.startswith("Tax Law US") for n in names)
    # All four return types are present in the demo book.
    for rt in ("1040", "1120S", "1065", "1120"):
        assert any(n.startswith(rt + " ") for n in names), f"missing {rt} sheet"


def test_no_unresolved_formula_tokens(tmp_path):
    out = build_demo_workbook(tmp_path / "wb.xlsx", tax_year=2024)
    wb = load_workbook(out)
    leftover = re.compile(r"\{[a-zA-Z0-9_]+\}|\[XW|\[XWFS|\[CF")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    assert not leftover.search(cell.value), f"Unresolved token in {ws.title}!{cell.coordinate}: {cell.value}"


def test_1040_prefilled_inputs_and_links(tmp_path):
    out = build_demo_workbook(tmp_path / "wb.xlsx", tax_year=2024)
    wb = load_workbook(out)
    ws = next(ws for ws in wb.worksheets if ws.title.startswith("1040"))
    # Find the wages input by its label and confirm the prefilled value.
    found_wages = False
    for row in ws.iter_rows():
        if isinstance(row[1].value, str) and row[1].value.startswith("Wages (sum"):
            assert row[2].value == 145000
            found_wages = True
    assert found_wages

    # A crosswalk link formula should reference the Tax Law sheet.
    has_xw_link = any(
        isinstance(c.value, str) and c.value.startswith("=") and "Tax Law US" in c.value
        for r in ws.iter_rows() for c in r
    )
    assert has_xw_link


def test_the_workbook_and_the_app_agree_on_a_due_date():
    """A workbook that computes its own deadlines answers "when is this due"
    differently from the app, forever, and nothing ever notices.

    The 1120-S case is the tell: March 15, 2026 is a Sunday, so the real date is
    the 16th. The old hardcoded table said the 15th.
    """
    from satc.obligations.due_dates import compute, tax_year
    from satc.obligations.rules import rule
    from satc.workbook.dashboards import _RULE_FOR_TYPE, _due_dates

    for return_type, rule_key in _RULE_FOR_TYPE.items():
        book_due, book_ext = _due_dates(return_type, 2025)
        app = compute(rule(rule_key), tax_year(2025))
        assert book_due == app.due, return_type
        assert book_ext == app.extended_due, return_type

    assert _due_dates("1120S", 2025)[0].day == 16, "the weekend shift was lost"
