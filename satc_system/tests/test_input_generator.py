"""The Drake input generator emits a workbook in the expected Clients / W2s shape."""

from __future__ import annotations

from openpyxl import load_workbook

from satc.drake.input_generator import generate_drake_intake_workbook
from satc.fixtures import synthetic_drake_intake


def test_generated_intake_has_expected_shape(tmp_path):
    out = generate_drake_intake_workbook(tmp_path / "intake.xlsx", synthetic_drake_intake())
    wb = load_workbook(out)
    assert wb.sheetnames == ["Clients", "W2s"]
    assert wb["Clients"].max_row == 2          # header + 1 client
    assert wb["W2s"].max_row == 3              # header + 2 W-2s
