"""Stage 8 validation — the hard guarantees the whole system rests on.

Most important: NO PII ever lands in the workbook. Also checks provenance is
intact and that year-to-year carryforwards tie. Formula-error-freeness is
verified separately by scripts/recalc.py (requires LibreOffice).
"""

from __future__ import annotations

import re

from openpyxl import load_workbook

from satc.build import build_demo_workbook
from satc.drake import parse_preparer_set
from satc.fixtures import synthetic_identities, synthetic_mart, synthetic_preparer_set_text
from satc.proforma import roll_forward


def _all_strings(path) -> list[str]:
    wb = load_workbook(path)
    out: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    out.append(cell.value)
    return out


def test_no_client_legal_names_or_full_tins_in_workbook(tmp_path):
    out = build_demo_workbook(tmp_path / "wb.xlsx", tax_year=2024)
    blob = "\n".join(_all_strings(out))

    for rec in synthetic_identities():
        # The client's legal name must never appear (vault-only).
        assert rec.legal_name not in blob, f"legal name leaked: {rec.legal_name}"
        # The full TIN (any separators removed) must never appear.
        full_tin = "".join(ch for ch in rec.tin if ch.isdigit())
        digits_only = re.sub(r"\D", "", blob)
        assert full_tin not in digits_only, f"full TIN leaked: {rec.client_id}"

    # No full SSN-format token anywhere; only masked ***-**-#### is allowed.
    assert not re.search(r"\b\d{3}-\d{2}-\d{4}\b", blob), "an unmasked SSN appears in the workbook"


def test_workbook_references_clients_only_by_id(tmp_path):
    out = build_demo_workbook(tmp_path / "wb.xlsx", tax_year=2024)
    blob = "\n".join(_all_strings(out))
    # The opaque client_id handles are how clients are referenced.
    assert "SATC-001000" in blob


def test_drake_parsed_values_keep_worksheet_provenance():
    ps = parse_preparer_set(synthetic_preparer_set_text(), client_id="SATC-001000", tax_year=2024)
    assert ps.staged is not None and ps.staged.fields
    for f in ps.staged.fields:
        assert f.provenance.source_kind == "DRAKE_OUTPUT"
        assert f.provenance.source_ref.worksheet_title


def test_carryforwards_tie_year_to_year():
    mart = synthetic_mart()
    seeds_25 = roll_forward(mart, from_year=2024, to_year=2025)
    # Every open carryforward as of 2024 advances into 2025 with its amount intact.
    open_2024 = [c for c in mart.carryforwards
                 if c.applied_to_year is None and c.tax_year_generated <= 2024 and c.amount != 0]
    carried = [cf for s in seeds_25.values() for cf in s.carryforwards]
    assert len(carried) == len(open_2024)
    assert sum(c.amount for c in carried) == sum(c.amount for c in open_2024)
