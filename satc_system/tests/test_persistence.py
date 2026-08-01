"""Tests for the SQLite store of record (durability + vault/mart separation)."""

from __future__ import annotations

from openpyxl import load_workbook

from satc.fixtures import synthetic_identities
from satc.persistence import SATCStore, export_mart_to_excel


def test_closing_a_request_survives_a_restart(tmp_path):
    """And the REASON survives with it — that reason is the record."""
    from satc.models.evidence import mark_not_applicable

    store = SATCStore(tmp_path)
    store.seed_if_empty()
    item = next(i for i in store.load_requested_items() if i.is_open)
    store.save_requested_items([mark_not_applicable(item, "account closed in March")])
    store.close()

    # Reopen the same directory — a fresh process would see exactly this.
    store2 = SATCStore(tmp_path)
    reloaded = {i.request_id: i for i in store2.load_requested_items()}
    assert reloaded[item.request_id].status == "not_applicable"
    assert reloaded[item.request_id].not_applicable_reason == "account closed in March"
    store2.close()


def test_pii_lives_only_in_the_vault_file_not_the_mart(tmp_path):
    store = SATCStore(tmp_path)
    store.seed_if_empty()
    store.close()

    mart_bytes = (tmp_path / "satc_mart.db").read_bytes().decode("latin-1")
    vault_bytes = (tmp_path / "satc_vault.db").read_bytes().decode("latin-1")
    for rec in synthetic_identities():
        # The legal name and full TIN must NOT appear in the mart database…
        assert rec.legal_name not in mart_bytes
        assert rec.tin.replace("-", "") not in mart_bytes.replace("-", "")
        # …and are NOT plaintext in the vault file either (encrypted at rest)…
        assert rec.legal_name not in vault_bytes
        assert rec.tin.replace("-", "") not in vault_bytes.replace("-", "")

    # …but they ARE retrievable through the store API (decrypted on read).
    names = set(SATCStore(tmp_path).names().values())
    assert {rec.legal_name for rec in synthetic_identities()} <= names


def test_export_produces_an_excel_mirror(tmp_path):
    store = SATCStore(tmp_path)
    store.seed_if_empty()
    out = export_mart_to_excel(store, tmp_path / "export.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    assert "Data Mart" in wb.sheetnames
    assert "Dashboards" in wb.sheetnames


def test_loaded_mart_matches_fixture_shape(tmp_path):
    store = SATCStore(tmp_path)
    store.seed_if_empty()
    mart = store.load_mart()
    assert len(mart.returns) == 6
    assert len(mart.requested_items) == 4
    assert len(mart.received_documents) == 6
    assert any(c.kind == "NOL" for c in mart.carryforwards)


def test_every_client_keyed_table_is_registered_for_deletion():
    """A table missing from the delete registry survives delete_client.

    Not tidiness: §10.28 requires returning a client's records, and a "deleted"
    client whose evidence rows are still on disk is a compliance defect. This
    test fails the moment a new client_id-keyed table is added without
    registering it, which is exactly when it is cheap to fix.
    """
    import tempfile

    from satc.persistence.store import (
        _MART_TABLES_BY_CLIENT,
        _VAULT_TABLES_BY_CLIENT,
        SATCStore,
    )

    with tempfile.TemporaryDirectory() as tmp, SATCStore(tmp) as store:
        for conn, registered in ((store.mart, _MART_TABLES_BY_CLIENT),
                                 (store.vault, _VAULT_TABLES_BY_CLIENT)):
            for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"):
                table = row["name"]
                cols = {c["name"] for c in conn.execute(f"PRAGMA table_info({table})")}
                if "client_id" in cols:
                    assert table in registered, (
                        f"{table} is keyed by client_id but is not in the delete "
                        f"registry — deleting a client would leave its rows behind")


def test_deleting_a_client_leaves_nothing_behind():
    """The property the registry exists to guarantee, checked directly."""
    import tempfile

    from satc.persistence.store import (
        _MART_TABLES_BY_CLIENT,
        _VAULT_TABLES_BY_CLIENT,
        SATCStore,
    )

    with tempfile.TemporaryDirectory() as tmp, SATCStore(tmp) as store:
        store.seed_if_empty()
        victim = next(iter(store.names()))
        store.delete_client(victim)
        for conn, tables in ((store.mart, _MART_TABLES_BY_CLIENT),
                             (store.vault, _VAULT_TABLES_BY_CLIENT)):
            for table in tables:
                left = conn.execute(
                    f"SELECT COUNT(*) AS n FROM {table} WHERE client_id=?",
                    (victim,)).fetchone()["n"]
                assert left == 0, f"{table} still holds rows for the deleted client"
