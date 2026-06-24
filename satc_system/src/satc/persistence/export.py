"""Excel export of the data mart (Excel is the export, not the store of record)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from satc.models.mart import DataMart
from satc.persistence.store import SATCStore
from satc.workbook.dashboards import build_dashboards_sheet
from satc.workbook.mart_sheets import build_data_mart_sheet
from satc.workbook.repository_sheet import build_repository_sheet


def export_mart_to_excel(source: SATCStore | DataMart, out_path: str | Path,
                         tax_year: int = 2024) -> Path:
    """Write the current data mart to a branded Excel workbook. Returns the path."""
    mart = source.load_mart() if isinstance(source, SATCStore) else source
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    ranges = build_data_mart_sheet(wb.create_sheet("Data Mart"), mart)
    repo = build_repository_sheet(wb.create_sheet("Document Repository"), mart)
    build_dashboards_sheet(wb.create_sheet("Dashboards"),
                           mart_ranges=ranges, repo_ranges=repo, tax_year=tax_year)
    wb.save(out)
    return out
