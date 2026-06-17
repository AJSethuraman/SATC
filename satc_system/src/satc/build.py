"""Workbook build orchestrator.

Assembles the SATC workbook from its parts: a branded cover, the dated tax-law
reference sheets (which expose the crosswalk parameters as cells), and the
config-driven line sheets prefilled with a client's confirmed values. Output is a
single .xlsx; run ``scripts/recalc.py`` afterward to evaluate formulas and confirm
zero errors.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from satc.config import load_extraction_map, load_line_sheet, templatize
from satc.crosswalk import CrosswalkLibrary
from satc.fixtures import (
    synthetic_1040_values,
    synthetic_documents,
    synthetic_entity_values,
    synthetic_identities,
    synthetic_mart,
)
from satc.ingest import MAPPING_1040, MapExtractor, StagingGate
from satc.proforma import compare_years, roll_forward
from satc.workbook.cover import build_cover
from satc.workbook.line_sheet import BuildContext, LineSheetBuilder
from satc.workbook.mart_sheets import (
    build_comparison_sheet,
    build_data_mart_sheet,
    build_proforma_sheet,
)
from satc.workbook.reference import build_reference_sheet
from satc.workbook.staging_sheet import build_staging_sheet

DEFAULT_OUT = Path(__file__).resolve().parents[2] / "build" / "SATC_Workbook.xlsx"


def _reference_sheets(wb: Workbook, lib: CrosswalkLibrary, tax_year: int,
                      jurisdictions: list[str]) -> dict[str, str]:
    """Build one reference sheet per jurisdiction; return the merged cell registry."""
    registry: dict[str, str] = {}
    for juris in jurisdictions:
        xw = lib.resolve_or_none(tax_year, juris)
        if xw is None:
            continue
        ws = wb.create_sheet(f"Tax Law {juris} {tax_year}")
        registry.update(build_reference_sheet(ws, xw))
    return registry


def build_demo_workbook(out_path: str | Path = DEFAULT_OUT, tax_year: int = 2024) -> Path:
    """Build the demo workbook for synthetic client SATC-001000 (1040, OH)."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    lib = CrosswalkLibrary().load()
    identities = {r.client_id: r for r in synthetic_identities()}
    client = identities["SATC-001000"].to_public()
    state = client.home_state or "OH"

    # Stage 1: extract synthetic source docs and run them through the gate.
    gate = StagingGate()
    for doc in synthetic_documents():
        cfg = load_extraction_map(doc["doc_key"])
        gate.add(MapExtractor(cfg).extract(
            document_id=doc["document_id"], client_id=client.client_id,
            tax_year=tax_year, labeled_fields=doc["labeled"]))
    gate.auto_confirm_high()
    confirmed_values = gate.to_line_values(MAPPING_1040)

    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover")
    staging = wb.create_sheet("Staging")
    build_staging_sheet(staging, gate)
    # Reference sheets for every jurisdiction in the book of synthetic clients.
    registry = _reference_sheets(wb, lib, tax_year, ["US", "OH", "MI", "MA"])

    contents = [
        ("Staging", "Document extraction + confirmation gate (intake control)"),
        (f"Tax Law US {tax_year}", "Federal parameters in force, with citations"),
        (f"Tax Law OH/MI/MA {tax_year}", "State parameters in force, with citations"),
    ]

    # 1040 — fed by the confirmed gate output overlaid on the synthetic workpaper.
    values_1040 = {**synthetic_1040_values(tax_year), **confirmed_values}
    _add_line_sheet(wb, "1040", client, state, values_1040, registry, tax_year)
    contents.append((f"1040 — {client.client_id}",
                     "Individual workpaper: intake, schedules, state, §8867, reconciliation"))

    # Entity returns (1120-S / 1065 / 1120) for the other synthetic clients.
    for cid, return_type, label in (
        ("SATC-002000", "1120S", "S-corp: M-1, K-1 tie-out, basis & AAA, Sch L"),
        ("SATC-003000", "1065", "Partnership: capital accounts, guaranteed pmts, M-1"),
        ("SATC-004000", "1120", "C-corp: NOL, charitable 10% limit, M-1, tax tie"),
    ):
        ent = identities[cid].to_public()
        _add_line_sheet(wb, return_type, ent, ent.home_state,
                        synthetic_entity_values(return_type), registry, tax_year)
        contents.append((f"{return_type} — {cid}", label))

    # Stage 5: data mart + prior-vs-current comparison + proforma seed.
    mart = synthetic_mart()
    build_data_mart_sheet(wb.create_sheet("Data Mart"), mart)
    comparison_rows = compare_years(
        mart, client_id="SATC-001000", return_type="1040", jurisdiction="US",
        prior_year=tax_year - 1, current_year=tax_year)
    build_comparison_sheet(
        wb.create_sheet("Prior-vs-Current"), comparison_rows,
        title=f"SATC-001000 · 1040 · US · {tax_year - 1} vs {tax_year}")
    seeds = roll_forward(mart, from_year=tax_year, to_year=tax_year + 1)
    build_proforma_sheet(wb.create_sheet("Proforma"), seeds, to_year=tax_year + 1)
    contents += [
        ("Data Mart", "Normalized year-over-year client record store (SQL-portable)"),
        ("Prior-vs-Current", "Variance flags: swings, dropped 1099s, dependent changes"),
        ("Proforma", f"Carryforwards & basis seeded into {tax_year + 1}"),
    ]

    build_cover(cover, tax_year=tax_year, contents=contents)
    wb.save(out)
    return out


def _add_line_sheet(wb: Workbook, return_type: str, client, state: str,
                    values: dict, registry: dict[str, str], tax_year: int) -> None:
    """Build one line sheet for a client/return type and add it to the workbook."""
    config = templatize(load_line_sheet(return_type), {"STATE": state})
    config["meta"]["subtitle"] = (
        f"Client {client.client_id} ({client.entity_type})  ·  TY{tax_year}  ·  "
        f"Federal + {state}   —   Drake is the system of record."
    )
    ctx = BuildContext(xw_registry=registry, values=values, default_jurisdiction="US")
    ws = wb.create_sheet(f"{return_type} — {client.client_id}")
    LineSheetBuilder(ws, config, ctx).build()


def main() -> None:
    out = build_demo_workbook()
    print(f"Built {out}")


if __name__ == "__main__":
    main()
