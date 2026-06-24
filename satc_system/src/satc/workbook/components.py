"""Reusable workbook building blocks (sit on top of :mod:`satc.workbook.styles`).

These helpers keep every sheet consistent and enforce the house rules:
  * data never lives in merged cells (merges are used only for titles / section
    headers / free-text notes — never for a value a formula reads);
  * the review dropdown (Done / Exception / N/A / Note) is one shared data
    validation;
  * branded page header/footer on every sheet for PDF export.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from satc.workbook import styles as S
from satc.workbook.styles import CellStyle
from satc.models.review import REVIEW_CHOICES


def write(ws: Worksheet, row: int, col: int, value, style: CellStyle | None = None,
          number_format: str | None = None):
    """Write a value into a cell and apply a style bundle."""
    cell = ws.cell(row=row, column=col, value=value)
    if style is not None:
        style.apply(cell)
    if number_format is not None:
        cell.number_format = number_format
    return cell


def merge_text(ws: Worksheet, row: int, first_col: int, last_col: int, text,
               style: CellStyle) -> None:
    """Write a *non-data* text label across merged columns (titles/headers/notes)."""
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    write(ws, row, first_col, text, style)


def section_header(ws: Worksheet, row: int, text: str, last_col: int,
                   first_col: int = 1) -> int:
    """Navy section banner spanning the content width. Returns the next free row."""
    ws.row_dimensions[row].height = 22
    merge_text(ws, row, first_col, last_col, f"  {text}", S.SECTION)
    return row + 1


def subsection(ws: Worksheet, row: int, text: str, last_col: int, first_col: int = 1) -> int:
    ws.row_dimensions[row].height = 18
    merge_text(ws, row, first_col, last_col, f"  {text}", S.SUBSECTION)
    return row + 1


def note_row(ws: Worksheet, row: int, text: str, last_col: int, first_col: int = 1) -> int:
    merge_text(ws, row, first_col, last_col, text, S.NOTE)
    ws.row_dimensions[row].height = 14
    return row + 1


def column_headers(ws: Worksheet, row: int, headers: list[str], start_col: int = 1) -> int:
    for i, text in enumerate(headers):
        write(ws, row, start_col + i, text, S.COLHEAD)
    return row + 1


def set_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def review_validation(ws: Worksheet) -> DataValidation:
    """Create and register the shared review dropdown for this sheet."""
    formula = '"' + ",".join(c for c in REVIEW_CHOICES if c != "") + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = "Choose Done, Exception, N/A, or Note."
    dv.errorTitle = "Review status"
    ws.add_data_validation(dv)
    return dv


def text_list_validation(ws: Worksheet, choices: list[str]) -> DataValidation:
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    return dv


def page_setup(ws: Worksheet, title: str, orientation: str = "portrait",
               fit_width: int = 1) -> None:
    """Apply branded header/footer + print fit for clean PDF export."""
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddHeader.left.text = "SETHURAMAN  ·  Accounting · Tax · Consulting"
    ws.oddHeader.left.font = "Garamond,Italic"
    ws.oddHeader.left.size = 9
    ws.oddHeader.right.text = title
    ws.oddHeader.right.size = 9
    ws.oddFooter.center.text = S.brand_footer_text()
    ws.oddFooter.center.size = 7
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 7


def freeze(ws: Worksheet, cell: str) -> None:
    ws.freeze_panes = cell


def col_letter(idx: int) -> str:
    return get_column_letter(idx)
