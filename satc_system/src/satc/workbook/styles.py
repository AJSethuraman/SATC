"""SATC visual identity for the workbook (re-skin of the CRR styling system).

Firm: Sethuraman Accounting, Tax & Consulting (SATC).
Tagline: "Complex work, made clear." — Occam's razor: the simplest correct answer.

This module centralizes the palette, fonts, number formats, and reusable cell
styling so every sheet shares one coherent look: a warm cream/paper canvas with
white content panels, navy section headers, and gold rules/accents. Red is
reserved for exceptions only.

It also layers the financial-model color convention on top of the brand:
  * INPUT (confirmed source value)      -> navy text on a soft input tint
  * COMPUTED (in-sheet formula)         -> ink/black text
  * LINK (pulls from another sheet)     -> gold-deep text (cross-sheet reference)
  * CARRYFORWARD (from the data mart)   -> green text
  * EXCEPTION / pending                 -> red text
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Palette (hex, no leading '#')
# ---------------------------------------------------------------------------
class C:
    NAVY = "0B1F3A"          # primary
    NAVY_DEEP = "061A35"
    NAVY_SOFT = "173361"
    GOLD = "B08D57"          # accent
    GOLD_LIGHT = "D4B97E"
    GOLD_DEEP = "8A6F44"
    CREAM = "F6F2EA"
    CREAM_2 = "EFE9DC"
    PAPER = "FBF9F4"         # page canvas (warm, not gray)
    HAIRLINE = "D9CFB8"
    INK = "0E1726"
    CHARCOAL = "1F2733"
    WHITE = "FFFFFF"
    RED = "9B2226"           # exceptions only
    INPUT_TINT = "F3EEE3"    # soft tint behind preparer/source inputs
    GREEN = "2F5D3A"         # carryforward links


# Fonts: Cormorant Garamond for titles (Excel fallback Garamond -> Times New Roman);
# Hanken Grotesk for body (Excel fallback Calibri -> Arial). openpyxl writes one
# family name; we use the safe Excel fallbacks so the file renders everywhere.
TITLE_FONT = "Garamond"
TITLE_FALLBACK = "Times New Roman"
BODY_FONT = "Calibri"
BODY_FALLBACK = "Arial"


# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------
class NF:
    # Currency: parentheses for negatives, dash for zero.
    USD = '$#,##0;($#,##0);"-"'
    USD_CENTS = '$#,##0.00;($#,##0.00);"-"'
    NUM = '#,##0;(#,##0);"-"'
    PCT1 = '0.0%'
    PCT2 = '0.00%'
    RATE_PER_MILE = '$0.00'
    YEAR = '0'               # years as plain integers, never 2,024
    DATE = 'mm/dd/yyyy'
    TEXT = '@'


# ---------------------------------------------------------------------------
# Reusable side/border primitives
# ---------------------------------------------------------------------------
def _side(color: str = C.HAIRLINE, style: str = "thin") -> Side:
    return Side(style=style, color=color)


HAIRLINE_BORDER = Border(
    left=_side(), right=_side(), top=_side(), bottom=_side()
)
BOTTOM_HAIRLINE = Border(bottom=_side())
GOLD_BOTTOM = Border(bottom=_side(C.GOLD, "medium"))
NAVY_BOX = Border(left=_side(C.NAVY), right=_side(C.NAVY), top=_side(C.NAVY), bottom=_side(C.NAVY))


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


@dataclass(frozen=True)
class CellStyle:
    """A bundle of openpyxl style attributes applied together to a cell."""

    font: Font | None = None
    fill: PatternFill | None = None
    border: Border | None = None
    alignment: Alignment | None = None
    number_format: str | None = None

    def apply(self, cell) -> None:
        if self.font is not None:
            cell.font = self.font
        if self.fill is not None:
            cell.fill = self.fill
        if self.border is not None:
            cell.border = self.border
        if self.alignment is not None:
            cell.alignment = self.alignment
        if self.number_format is not None:
            cell.number_format = self.number_format


# ---------------------------------------------------------------------------
# Named style bundles (the vocabulary every builder uses)
# ---------------------------------------------------------------------------
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

TITLE = CellStyle(
    font=Font(name=TITLE_FONT, size=26, bold=True, color=C.NAVY),
    alignment=Alignment(horizontal="left", vertical="center"),
)
WORDMARK = CellStyle(
    font=Font(name=TITLE_FONT, size=15, bold=True, color=C.NAVY, italic=False),
    alignment=Alignment(horizontal="left", vertical="center"),
)
SUBTITLE = CellStyle(
    font=Font(name=BODY_FONT, size=11, italic=True, color=C.GOLD_DEEP),
    alignment=Alignment(horizontal="left", vertical="center"),
)
SECTION = CellStyle(
    font=Font(name=BODY_FONT, size=12, bold=True, color=C.CREAM),
    fill=fill(C.NAVY),
    alignment=Alignment(horizontal="left", vertical="center"),
    border=Border(bottom=_side(C.GOLD, "thin")),
)
SUBSECTION = CellStyle(
    font=Font(name=BODY_FONT, size=10.5, bold=True, color=C.WHITE),
    fill=fill(C.NAVY_SOFT),
    alignment=Alignment(horizontal="left", vertical="center"),
)
COLHEAD = CellStyle(
    font=Font(name=BODY_FONT, size=9.5, bold=True, color=C.NAVY),
    fill=fill(C.CREAM_2),
    alignment=CENTER,
    border=BOTTOM_HAIRLINE,
)
LABEL = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.INK),
    alignment=LEFT,
)
LABEL_MUTED = CellStyle(
    font=Font(name=BODY_FONT, size=9, italic=True, color=C.CHARCOAL),
    alignment=LEFT_WRAP,
)
INPUT = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.NAVY),
    fill=fill(C.INPUT_TINT),
    border=HAIRLINE_BORDER,
    alignment=RIGHT,
    number_format=NF.USD,
)
INPUT_TEXT = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.NAVY),
    fill=fill(C.INPUT_TINT),
    border=HAIRLINE_BORDER,
    alignment=LEFT,
)
COMPUTED = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.INK),
    border=HAIRLINE_BORDER,
    alignment=RIGHT,
    number_format=NF.USD,
)
COMPUTED_BOLD = CellStyle(
    font=Font(name=BODY_FONT, size=10, bold=True, color=C.INK),
    fill=fill(C.CREAM),
    border=HAIRLINE_BORDER,
    alignment=RIGHT,
    number_format=NF.USD,
)
LINK = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.GOLD_DEEP),
    border=HAIRLINE_BORDER,
    alignment=RIGHT,
    number_format=NF.USD,
)
CARRYFORWARD = CellStyle(
    font=Font(name=BODY_FONT, size=10, color=C.GREEN),
    border=HAIRLINE_BORDER,
    alignment=RIGHT,
    number_format=NF.USD,
)
EXCEPTION = CellStyle(
    font=Font(name=BODY_FONT, size=10, bold=True, color=C.RED),
    alignment=LEFT,
)
NOTE = CellStyle(
    font=Font(name=BODY_FONT, size=9, italic=True, color=C.CHARCOAL),
    alignment=LEFT_WRAP,
)
PANEL = CellStyle(fill=fill(C.WHITE), border=HAIRLINE_BORDER)
FOOTER = CellStyle(
    font=Font(name=BODY_FONT, size=8, italic=True, color=C.GOLD_DEEP),
    alignment=Alignment(horizontal="center", vertical="center"),
)


def paper_canvas(ws: Worksheet, max_col: int = 14, max_row: int = 200) -> None:
    """Fill the visible grid with the warm paper canvas and hide gridlines."""
    ws.sheet_view.showGridLines = False
    canvas = fill(C.PAPER)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = canvas


def brand_footer_text() -> str:
    """The footer line stamped on PDF export."""
    return ("Sethuraman Accounting, Tax & Consulting  ·  Complex work, made clear.  "
            "·  Workpaper — preparer review required  ·  Drake is the system of record.")
