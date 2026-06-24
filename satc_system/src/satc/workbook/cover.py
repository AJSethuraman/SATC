"""Branded SATC workbook cover sheet.

Renders the monogram seal (serif "S" in gold on navy), the SETHURAMAN wordmark,
the sub-tag and tagline, an index of the workbook's contents, the color-coding
legend, and the standing confidentiality / "no PII in the workbook" notice.
"""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from satc.workbook import components as X
from satc.workbook import styles as S
from satc.workbook.styles import C, CellStyle, fill


def _seal(ws: Worksheet) -> None:
    """A simple monogram seal: serif 'S' in gold inside a navy box with a gold ring."""
    for r in range(2, 6):
        for c in range(2, 5):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(C.NAVY)
            cell.border = Border(*(Side(style="medium", color=C.GOLD),) * 4)
    ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=4)
    seal = ws.cell(row=2, column=2, value="S")
    seal.font = Font(name=S.TITLE_FONT, size=44, bold=True, color=C.GOLD)
    seal.alignment = Alignment(horizontal="center", vertical="center")


def build_cover(ws: Worksheet, *, tax_year: int, contents: list[tuple[str, str]],
                prepared_on: date | None = None) -> None:
    S.paper_canvas(ws, max_col=12, max_row=60)
    X.set_widths(ws, {"A": 3, "B": 12, "C": 14, "D": 14, "E": 16, "F": 16, "G": 16, "H": 10})
    prepared_on = prepared_on or date.today()

    _seal(ws)

    X.merge_text(ws, 2, 6, 11, "SETHURAMAN", CellStyle(
        font=Font(name=S.TITLE_FONT, size=34, bold=True, color=C.NAVY),
        alignment=Alignment(horizontal="left", vertical="center")))
    X.merge_text(ws, 3, 6, 11, "A C C O U N T I N G   ·   T A X   ·   C O N S U L T I N G",
                 CellStyle(font=Font(name=S.BODY_FONT, size=10, color=C.GOLD_DEEP),
                           alignment=Alignment(horizontal="left", vertical="center")))
    # Hairline rule with center dot
    ws.row_dimensions[5].height = 6
    X.merge_text(ws, 5, 6, 11, "—————————————————  ·  —————————————————",
                 CellStyle(font=Font(name=S.BODY_FONT, size=9, color=C.GOLD),
                           alignment=Alignment(horizontal="center", vertical="center")))

    X.merge_text(ws, 7, 2, 11, "Tax Workpapers & Client Data Mart", CellStyle(
        font=Font(name=S.TITLE_FONT, size=20, bold=True, color=C.NAVY),
        alignment=Alignment(horizontal="left", vertical="center")))
    X.merge_text(ws, 8, 2, 11, '"Complex work, made clear."  —  Occam\'s razor: the simplest correct answer, never oversimplified.',
                 S.SUBTITLE)

    row = 10
    meta = [
        ("Tax year", str(tax_year)),
        ("Prepared", prepared_on.strftime("%B %d, %Y")),
        ("Firm", "Sethuraman Accounting, Tax & Consulting (SATC)"),
        ("System of record", "Drake Software — this workbook is a workpaper layer, not a tax engine"),
    ]
    for label, value in meta:
        X.write(ws, row, 2, label, S.LABEL_MUTED)
        X.merge_text(ws, row, 3, 11, value, S.LABEL)
        row += 1

    row += 1
    row = X.section_header(ws, row, "Contents", last_col=11, first_col=2)
    for name, desc in contents:
        X.write(ws, row, 2, "›", CellStyle(font=Font(name=S.BODY_FONT, color=C.GOLD, bold=True)))
        X.write(ws, row, 3, name, S.LABEL)
        X.merge_text(ws, row, 5, 11, desc, S.LABEL_MUTED)
        row += 1

    row += 1
    row = X.section_header(ws, row, "How to read this workbook — color legend", last_col=11, first_col=2)
    legend = [
        (C.NAVY, "Input — a value confirmed from a source document or keyed by the preparer."),
        (C.INK, "Computed — an in-sheet formula (a cross-check or subtotal)."),
        (C.GOLD_DEEP, "Link — pulls a tax-law parameter from the Tax Law reference sheet."),
        (C.GREEN, "Carryforward — pulled from the client data mart (prior-year proforma)."),
        (C.RED, "Exception / pending — needs attention; pending tax-law guidance."),
    ]
    for color, text in legend:
        X.write(ws, row, 3, "■", CellStyle(font=Font(name=S.BODY_FONT, size=12, color=color)))
        X.merge_text(ws, row, 5, 11, text, S.LABEL)
        row += 1

    row += 1
    X.merge_text(ws, row, 2, 11,
                 "CONFIDENTIAL — De-identified workpaper. Full SSN/EIN and source documents live in the "
                 "firm's secure store (Teams/SharePoint), never in this file. Clients are referenced by "
                 "client_id; only masked / last-4 values appear here.",
                 CellStyle(font=Font(name=S.BODY_FONT, size=9, italic=True, color=C.GOLD_DEEP),
                           fill=fill(C.CREAM_2),
                           alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)))
    ws.row_dimensions[row].height = 40
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)

    X.page_setup(ws, "Cover", orientation="portrait")
    X.freeze(ws, "A1")
