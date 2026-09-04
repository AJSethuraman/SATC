"""One lead shape, wherever the lead came from.

The firm, 26 August 2026:

    "the interview has to be set up - so like the questions we ask are based on
     the info we got in the intake lead which is this workbook
     it is possible that a lead has to be input manually though, they may just
     give us contact info"

Two doors, one shape. A lead arrives either as a row in `SATC_leads.xlsx` --
which is where the website's submissions land -- or as a person typing what
somebody said on the phone. Both produce the same dict, so everything
downstream reads one thing.

WHAT THE WORKBOOK ACTUALLY HOLDS, from the firm's own export:

    Received  Name  Email  Phone  Location  Preferred  Services
    Individual complexity  Business structure  Tax status  Bookkeeping status
    Urgency  Deadline  Notes  Raw JSON  Lead Number

`Raw JSON` is the submission verbatim and every other column is derived from
it, so that column is what is read. The flat columns are the fallback for a row
somebody typed in by hand.

WHAT THE INTERVIEW WAS DOING WITH IT. Three of the nine keys. `services`,
`business_structure`, `business_complexity`, `revenue_band`, `urgency` and
`notes` were collected by the website, carried into the workbook, and read by
nothing -- including the prospect's own note, which on the sample lead says
"Sold a rental in March and have a K-1 from a partnership I joined last year.
Prior accountant retired." That is three interview answers, in the client's
own words, that nobody was shown.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent

# Workbook column -> where it lives in the lead. Only used when `Raw JSON` is
# missing or unreadable; the JSON is authoritative when it is there.
COLUMNS = {
    "name": ("contact", "name"),
    "email": ("contact", "email"),
    "phone": ("contact", "phone"),
    "location": ("contact", "location"),
    "preferred": ("contact", "preferred"),
    "notes": (None, "notes"),
    "urgency": (None, "urgency"),
    "tax status": (None, "tax_status"),
}
_LISTS = {"services", "individual_complexity", "business_structure",
          "business_complexity"}
_COLUMN_LISTS = {
    "services": "services",
    "individual complexity": "individual_complexity",
    "business structure": "business_structure",
}


class LeadError(RuntimeError):
    pass


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def normalise(lead: dict) -> dict:
    """A lead in the one shape, with the list fields always lists.

    The website sends lists; a hand-typed row sends "individual_tax,
    business_advisory". Downstream code should not have to know which.
    """
    out = dict(lead or {})
    out["contact"] = dict(out.get("contact") or {})
    for key in _LISTS:
        v = out.get(key)
        if v is None:
            continue
        if isinstance(v, str):
            v = [p.strip() for p in v.split(",") if p.strip()]
        out[key] = list(v)
    return out


def by_hand(*, name: str = "", email: str = "", phone: str = "",
            location: str = "", notes: str = "", **rest) -> dict:
    """A lead for somebody who just gave us contact details.

    This is a REAL lead, not a lesser one. The firm's point: "they may just
    give us contact info". So it carries the same keys, with the answers the
    website would have collected simply absent -- and absent is honest. It is
    not the same as answered "none", and the interview asks either way.
    """
    if not (_clean(name) or _clean(email) or _clean(phone)):
        raise LeadError(
            "a lead needs at least a name, an email or a phone number. "
            "Without one of the three there is nobody to come back to."
        )
    lead = {
        "contact": {"name": _clean(name), "email": _clean(email),
                    "phone": _clean(phone), "location": _clean(location),
                    "preferred": _clean(rest.get("preferred")),
                    "consent": bool(rest.get("consent", False))},
        "notes": _clean(notes),
        # NOT a claim of "none". Nobody asked, so nothing is recorded, and the
        # interview asks every one of these from scratch.
        "_by_hand": True,
    }
    for key in ("services", "individual_complexity", "business_structure"):
        if rest.get(key):
            lead[key] = rest[key]
    return normalise(lead)


def from_row(header: list[str], row: list) -> dict:
    """One workbook row -> a lead. `Raw JSON` wins where it is present."""
    cells = {_clean(h).lower(): v for h, v in zip(header, row)}
    raw = _clean(cells.get("raw json"))
    if raw:
        try:
            lead = normalise(json.loads(raw))
        except json.JSONDecodeError as exc:
            raise LeadError(
                f"the Raw JSON column on this row will not parse ({exc}). "
                f"That column is the submission verbatim; a row whose JSON is "
                f"broken has lost something the flat columns do not carry."
            ) from None
    else:
        lead = {"contact": {}}
        for col, (group, key) in COLUMNS.items():
            v = _clean(cells.get(col))
            if not v:
                continue
            if group:
                lead[group][key] = v
            else:
                lead[key] = v
        for col, key in _COLUMN_LISTS.items():
            v = _clean(cells.get(col))
            if v:
                lead[key] = v
        lead["_by_hand"] = True
        lead = normalise(lead)

    received = _clean(cells.get("received"))
    if received:
        lead["_received"] = received
    number = _clean(cells.get("lead number"))
    if number:
        # The workbook writes "2026 - 0001" and an engagement ref is
        # "2026-0001". Same number, two spellings, and the ref is byte-compared
        # across every document -- so it is normalised here rather than
        # discovered later on a letter.
        lead["_lead_number"] = re.sub(r"\s*-\s*", "-", number)
    return lead


def from_workbook(path: Path | str, *, sheet: str | None = None) -> list[dict]:
    """Every row of the leads workbook, newest first.

    openpyxl is imported here rather than at module scope: the pipeline runs
    without it, and a workbook nobody opened should not be a dependency of
    rendering a letter.
    """
    try:
        import openpyxl
    except ImportError:
        raise LeadError(
            "reading the leads workbook needs openpyxl — `pip install "
            "openpyxl`. Everything else in this folder runs without it."
        ) from None

    wb = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise LeadError(f"{Path(path).name} has no rows at all")
    header = [_clean(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(_clean(c) for c in row):
            continue
        out.append(from_row(header, list(row)))
    out.sort(key=lambda l: l.get("_received", ""), reverse=True)
    return out


# The intake form's own labels, read out of it. `individual_tax` and
# `100k_500k` are what the site POSTS; "Individual tax preparation" and
# "$100k – $500k" are what the prospect actually clicked, and a preparer
# reading the lead should see the words the client saw.
SITE_CONFIG = ROOT.parent / "website" / "intake-config.js"


def site_labels(path: Path | str | None = None) -> dict:
    """value -> label, across every option the intake form offers.

    Read rather than copied. A duplicate list here would drift from the site
    the way the old `prefill_map` did, and nobody would notice until a lead
    read half in English and half in code.
    """
    try:
        js = Path(path or SITE_CONFIG).read_text(encoding="utf-8")
    except OSError:
        return {}
    out = {}
    for value, label in re.findall(r"value:\s*'([^']+)'\s*,\s*label:\s*'([^']*)'", js):
        out.setdefault(value, label)
    for value, label in re.findall(r'value:\s*"([^"]+)"\s*,\s*label:\s*"([^"]*)"', js):
        out.setdefault(value, label)
    return out


def _english(values, labels) -> str:
    if isinstance(values, str):
        values = [values]
    return ", ".join(labels.get(v, v) for v in values or [])


def summary(lead: dict) -> list[tuple[str, str]]:
    """What this prospect told us, in the order a preparer wants it.

    Everything, not the three keys the interview happens to prefill from.
    Shown before the sitting starts: a preparer who has read the note does not
    have to discover halfway down that there is a rental.
    """
    lead = normalise(lead)
    labels = site_labels()
    c = lead.get("contact") or {}
    rows = []

    def add(label, value):
        if value:
            rows.append((label, value))

    add("Name they gave", c.get("name"))
    add("Email", c.get("email"))
    add("Phone", c.get("phone"))
    add("Where they are", c.get("location"))
    add("How to reach them", c.get("preferred"))
    add("What they asked for", _english(lead.get("services"), labels))
    add("What applies to them", _english(lead.get("individual_complexity"), labels))
    add("How the business is set up", _english(lead.get("business_structure"), labels))
    add("About the business", _english(lead.get("business_complexity"), labels))
    add("Revenue", _english(lead.get("revenue_band"), labels))
    add("Where their filings stand", _english(lead.get("tax_status"), labels))
    add("How soon", _english(lead.get("urgency"), labels))
    add("Their deadline", lead.get("deadline"))
    add("What they wrote", lead.get("notes"))
    return rows
