"""Formatted Excel workbook output -- the primary client deliverable.

Tabs:
* ``Dashboard``       -- tier assessment, headline metrics, flagged threshold
                         exceptions, and the data-gap (data-maturity) findings.
* ``Card Detail``     -- monthly portfolio time series with a delinquency
                         trend chart and a charge-off rate chart.
* ``Migration Matrix``-- account- and dollar-weighted roll-rate matrices
                         with heat shading.
* ``Vintage Curves``  -- cumulative-loss table by cohort x months-on-book
                         with a line chart, plus the cohort summary.
* ``Concentration``   -- score band / vintage / line-size tables with charts.
* ``Utilization``     -- utilization distribution with chart.

Blocked metrics render as an explicit "not computable" note pointing at the
data-gap findings rather than silently disappearing.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ucpa.metrics.results import MetricResult, ReviewResult

NAVY = "1F3864"
LIGHT = "D9E2F3"
RED = "C00000"
AMBER = "BF8F00"
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
AMBER_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color=NAVY)
SECTION_FONT = Font(size=11, bold=True, color=NAVY)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT = "0.00%"
MONEY = "#,##0"
NUM2 = "0.00"


def _title(ws: Worksheet, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    return row + 2


def _section(ws: Worksheet, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    return row + 1


def _write_table(
    ws: Worksheet,
    row: int,
    df: pd.DataFrame,
    formats: Optional[dict[str, str]] = None,
    max_rows: Optional[int] = None,
) -> int:
    """Write a DataFrame with styled headers; return the next free row."""
    formats = formats or {}
    data = df if max_rows is None else df.head(max_rows)
    for j, col in enumerate(data.columns, start=1):
        cell = ws.cell(row=row, column=j, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    for i, (_, rec) in enumerate(data.iterrows(), start=1):
        for j, col in enumerate(data.columns, start=1):
            value = rec[col]
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.strftime("%Y-%m")
            elif hasattr(value, "item"):
                value = value.item()
            cell = ws.cell(row=row + i, column=j, value=value)
            cell.border = BORDER
            if col in formats:
                cell.number_format = formats[col]
    return row + len(data) + 2


def _autosize(ws: Worksheet, widths: Sequence[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_value(value: object) -> object:
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m")
    if isinstance(value, float):
        return round(value, 6)
    return value


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
_HEADLINES: tuple[tuple[str, str, str, str], ...] = (
    # (metric, summary_key, label, number_format)
    ("delinquency_distribution", "dpd30plus_balance_rate", "30+ DPD balance rate", PCT),
    ("delinquency_distribution", "dpd90plus_balance_rate", "90+ DPD balance rate", PCT),
    ("portfolio_time_series", "dpd30plus_yoy_delta", "30+ DPD rate YoY change (pp)", PCT),
    ("portfolio_time_series", "gross_co_rate_yoy_delta", "Gross CO rate YoY change (pp)", PCT),
    ("portfolio_time_series", "balance_growth_12m", "Open-balance growth (12m)", PCT),
    ("migration_matrix", "current_to_dpd30", "Current -> 30DPD monthly roll ($)", PCT),
    ("migration_matrix", "dpd30_cure_rate", "30DPD cure rate ($)", PCT),
    ("charge_off_rates", "gross_co_rate_t12", "Gross charge-off rate (T12 ann.)", PCT),
    ("charge_off_rates", "net_co_rate_t12", "Net charge-off rate (T12 ann.)", PCT),
    ("recovery_trends", "recovery_rate_t12", "Recovery rate (T12)", PCT),
    ("vintage_curves", "max_cum_loss_mob12", "Worst-cohort cum loss @ MOB 12", PCT),
    ("concentration", "subprime_balance_share", "Subprime balance share", PCT),
    ("utilization_distribution", "portfolio_utilization", "Portfolio utilization", PCT),
    ("utilization_distribution", "high_util_balance_share", "Balance share >90% utilized", PCT),
    ("line_management", "exposure_added_total", "Line exposure added (panel)", MONEY),
)


def _dashboard(ws: Worksheet, review: ReviewResult) -> None:
    det = review.tier_detection
    row = _title(ws, 1, f"Asset-Quality Review Dashboard -- {review.product_type}")

    row = _section(ws, row, "Portfolio")
    for label, key in (
        ("As of", "as_of"),
        ("Open accounts", "open_accounts"),
        ("Open balance ($)", "open_balance"),
        ("Panel months", "panel_months"),
        ("Tape rows", "rows"),
    ):
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=_fmt_value(review.portfolio_summary.get(key)))
        if key == "open_balance":
            cell.number_format = MONEY
        row += 1
    row += 1

    row = _section(ws, row, "Data-tier assessment")
    ws.cell(row=row, column=1, value="Detected data tier")
    ws.cell(row=row, column=2, value=f"Tier {det.detected_tier}")
    row += 1
    ws.cell(row=row, column=1, value="Longitudinal monthly panel")
    ws.cell(row=row, column=2, value="Yes" if det.is_panel else "No (snapshot)")
    row += 1
    if det.missing_for_tier0:
        ws.cell(row=row, column=1, value="MISSING TIER 0 (minimum standard) FIELDS")
        ws.cell(row=row, column=2, value=", ".join(det.missing_for_tier0)).font = Font(color=RED)
        row += 1
    ws.cell(row=row, column=1, value="Missing for next tier")
    ws.cell(
        row=row,
        column=2,
        value=", ".join(det.missing_for_next_tier) if det.missing_for_next_tier else "None - highest tier reached",
    )
    row += 2

    row = _section(ws, row, "Headline metrics")
    for metric, key, label, fmt in _HEADLINES:
        result = review.result_for(metric)
        value = result.summary.get(key) if result is not None else None
        ws.cell(row=row, column=1, value=label)
        if value is None:
            cell = ws.cell(row=row, column=2, value="n/a - see data gaps")
            cell.font = Font(italic=True, color="808080")
        else:
            cell = ws.cell(row=row, column=2, value=float(value))
            cell.number_format = fmt
        row += 1
    row += 1

    row = _section(ws, row, f"Automated observations -- rule-based, deterministic ({len(review.observations)})")
    if review.observations:
        obs_df = pd.DataFrame(
            [
                {"severity": o.severity, "rule": o.rule_id, "observation": o.text}
                for o in review.observations
            ]
        )
        start = row
        row = _write_table(ws, row, obs_df)
        for i, o in enumerate(review.observations, start=1):
            if o.severity == "ELEVATED":
                fill = RED_FILL
            elif o.severity == "NOTABLE":
                fill = AMBER_FILL
            else:
                continue
            for j in range(1, len(obs_df.columns) + 1):
                ws.cell(row=start + i, column=j).fill = fill
    else:
        ws.cell(row=row, column=1, value="No observations produced.")
        row += 2

    row = _section(ws, row, f"Threshold exceptions ({len(review.exceptions)})")
    if review.exceptions:
        exc = pd.DataFrame(
            [
                {
                    "severity": e.severity,
                    "metric": e.metric,
                    "check": e.check,
                    "observed": e.observed,
                    "limit": e.limit,
                    "direction": e.direction,
                }
                for e in review.exceptions
            ]
        )
        start = row
        row = _write_table(ws, row, exc, formats={"observed": NUM2, "limit": NUM2})
        for i, e in enumerate(review.exceptions, start=1):
            fill = RED_FILL if e.severity == "EXCEPTION" else AMBER_FILL
            for j in range(1, len(exc.columns) + 1):
                ws.cell(row=start + i, column=j).fill = fill
    else:
        ws.cell(row=row, column=1, value="No thresholds breached.")
        row += 2

    row = _section(ws, row, f"Data-gap findings -- data-maturity assessment ({len(review.gaps)})")
    if review.gaps:
        gaps = pd.DataFrame(
            [
                {
                    "metric": g.metric,
                    "scope": g.scope,
                    "tier required": g.tier_required,
                    "missing fields": ", ".join(g.missing_fields),
                    "impact": g.description,
                }
                for g in review.gaps
            ]
        )
        row = _write_table(ws, row, gaps)
    else:
        ws.cell(row=row, column=1, value="None - the tape supports the full metric battery.")
        row += 2

    _autosize(ws, (40, 28, 34, 12, 12, 12, 90))


# ---------------------------------------------------------------------------
# Detail sheets
# ---------------------------------------------------------------------------
def _blocked_note(ws: Worksheet, row: int, result: Optional[MetricResult], name: str) -> bool:
    """Write a not-computable note when the metric is missing/blocked."""
    if result is not None and result.status != "blocked":
        return False
    note = (
        result.gaps[0].description
        if result is not None and result.gaps
        else f"{name}: not computable on this tape - see Dashboard data-gap findings."
    )
    ws.cell(row=row, column=1, value=note).font = Font(italic=True, color=RED)
    return True


def _line_chart(
    ws: Worksheet,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    anchor: str,
    y_pct: bool = True,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.height = 8
    chart.width = 20
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    if y_pct:
        chart.y_axis.numFmt = "0.0%"
    ws.add_chart(chart, anchor)


_TS_FORMATS = {
    "open_balance": MONEY,
    "balance_current": MONEY,
    "balance_dpd30": MONEY,
    "balance_dpd60": MONEY,
    "balance_dpd90": MONEY,
    "balance_dpd120": MONEY,
    "gross_charge_offs": MONEY,
    "recoveries": MONEY,
    "dpd30plus_rate": PCT,
    "dpd90plus_rate": PCT,
    "gross_co_rate_ann": PCT,
    "portfolio_utilization": PCT,
    "high_util_balance_share": PCT,
}


def _card_detail(ws: Worksheet, review: ReviewResult) -> None:
    row = _title(ws, 1, "Card Portfolio Detail -- monthly time series")
    ts = review.result_for("portfolio_time_series")
    if _blocked_note(ws, row, ts, "portfolio_time_series"):
        ws.cell(
            row=row + 2,
            column=1,
            value="Trend analytics require a monthly longitudinal panel (Tier 1).",
        ).font = Font(italic=True)
        return
    assert ts is not None

    monthly = ts.tables["monthly"]
    cols = list(monthly.columns)
    n = len(monthly)
    row = _section(ws, row, "Consolidated monthly portfolio panel")
    start = row
    row = _write_table(ws, row, monthly, formats=_TS_FORMATS)

    def col_ref(*names: str) -> tuple[int, int]:
        idx = [cols.index(c) + 1 for c in names if c in cols]
        return min(idx), max(idx)

    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + n)
    anchor_col = get_column_letter(len(cols) + 2)
    charts: list[tuple[str, tuple[int, int], bool]] = [
        ("30+/90+ DPD balance rate by month", col_ref("dpd30plus_rate", "dpd90plus_rate"), True),
        ("Gross charge-off rate (annualized) by month", col_ref("gross_co_rate_ann"), True),
        ("Open balance by month ($)", col_ref("open_balance"), False),
    ]
    if "portfolio_utilization" in cols:
        charts.append(("Portfolio utilization by month", col_ref("portfolio_utilization"), True))
    anchor_row = 3
    for title, (lo, hi), pct in charts:
        _line_chart(
            ws,
            title,
            Reference(ws, min_col=lo, max_col=hi, min_row=start, max_row=start + n),
            cats,
            f"{anchor_col}{anchor_row}",
            y_pct=pct,
        )
        anchor_row += 17

    _autosize(ws, tuple([14] + [13] * (len(cols) - 1)))


def _migration_sheet(ws: Worksheet, review: ReviewResult) -> None:
    row = _title(ws, 1, "Month-over-Month Migration / Roll-Rate Matrix")
    result = review.result_for("migration_matrix")
    if _blocked_note(ws, row, result, "migration_matrix"):
        return
    assert result is not None

    for name, label, fmt in (
        ("row_pct", "Average monthly migration matrix -- % of accounts", PCT),
        ("balance_row_pct", "Average monthly roll-rate matrix -- % of dollars", PCT),
        ("counts", "Transition counts (all month pairs)", MONEY),
    ):
        table = result.tables[name]
        row = _section(ws, row, label)
        start = row
        formats = {c: fmt for c in table.columns if c != "from_bucket"}
        row = _write_table(ws, row, table, formats=formats)
        if fmt == PCT:  # heat-shade probability cells
            rng = (
                f"B{start + 1}:{get_column_letter(len(table.columns))}{start + len(table)}"
            )
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="FFFFFF",
                    end_type="num", end_value=1, end_color="C00000",
                ),
            )

    trend = result.tables["trend"]
    row = _section(ws, row, "Front-end roll-rate trend (dollar-weighted)")
    start = row
    row = _write_table(
        ws, row, trend, formats={"current_to_dpd30": PCT, "dpd30_to_dpd60": PCT}
    )
    _line_chart(
        ws,
        "Current->30 and 30->60 monthly roll rates",
        Reference(ws, min_col=2, max_col=3, min_row=start, max_row=start + len(trend)),
        Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(trend)),
        f"F{start}",
    )
    _autosize(ws, (16, 12, 12, 12, 12, 12, 12))


def _vintage_sheet(ws: Worksheet, review: ReviewResult) -> None:
    row = _title(ws, 1, "Vintage Cumulative Gross-Loss Curves (loss / original credit line)")
    result = review.result_for("vintage_curves")
    if _blocked_note(ws, row, result, "vintage_curves"):
        return
    assert result is not None

    summary = result.tables["cohort_summary"]
    row = _section(ws, row, "Cohort summary")
    row = _write_table(
        ws,
        row,
        summary,
        formats={
            "orig_credit_line": MONEY,
            "cum_loss_mob12": PCT,
            "cum_loss_latest": PCT,
        },
    )

    curves = result.tables["curves"]
    row = _section(ws, row, "Cumulative loss by months-on-book (columns = origination quarter)")
    start = row
    formats = {c: PCT for c in curves.columns if c != "months_on_book"}
    row = _write_table(ws, row, curves, formats=formats)
    n_rows = len(curves)
    n_cols = len(curves.columns)
    _line_chart(
        ws,
        "Vintage cumulative loss curves",
        Reference(ws, min_col=2, max_col=n_cols, min_row=start, max_row=start + n_rows),
        Reference(ws, min_col=1, min_row=start + 1, max_row=start + n_rows),
        f"{get_column_letter(n_cols + 2)}3",
    )
    _autosize(ws, tuple([18] + [11] * (n_cols - 1)))


def _concentration_sheet(ws: Worksheet, review: ReviewResult) -> None:
    row = _title(ws, 1, "Concentration Analysis (latest month, open accounts)")
    result = review.result_for("concentration")
    if _blocked_note(ws, row, result, "concentration"):
        return
    assert result is not None

    anchor_col = "G"
    for name, label in (
        ("by_score_band", "By score band"),
        ("by_vintage_year", "By origination vintage year"),
        ("by_line_size", "By credit-line size"),
    ):
        if name not in result.tables:
            ws.cell(row=row, column=1, value=f"{label}: blocked - see data-gap findings.").font = Font(italic=True, color=RED)
            row += 2
            continue
        table = result.tables[name]
        row = _section(ws, row, label)
        start = row
        row = _write_table(
            ws, row, table, formats={"balance": MONEY, "balance_share": PCT}
        )
        chart = BarChart()
        chart.title = f"Balance share {label.lower()}"
        chart.height = 6
        chart.width = 14
        share_col = list(table.columns).index("balance_share") + 1
        chart.add_data(
            Reference(ws, min_col=share_col, max_col=share_col, min_row=start, max_row=start + len(table)),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(table)))
        chart.y_axis.numFmt = "0%"
        ws.add_chart(chart, f"{anchor_col}{start}")
    _autosize(ws, (18, 12, 16, 14))


def _utilization_sheet(ws: Worksheet, review: ReviewResult) -> None:
    row = _title(ws, 1, "Utilization & Line Management")
    result = review.result_for("utilization_distribution")
    if not _blocked_note(ws, row, result, "utilization_distribution"):
        assert result is not None
        table = result.tables["distribution"]
        row = _section(ws, row, "Utilization distribution (latest month, open accounts)")
        start = row
        row = _write_table(ws, row, table, formats={"balance": MONEY, "balance_share": PCT})
        chart = BarChart()
        chart.title = "Balance share by utilization bucket"
        chart.height = 7
        chart.width = 16
        chart.add_data(
            Reference(ws, min_col=4, max_col=4, min_row=start, max_row=start + len(table)),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(table)))
        chart.y_axis.numFmt = "0%"
        ws.add_chart(chart, f"G{start}")
    else:
        row += 2

    lm = review.result_for("line_management")
    if not _blocked_note(ws, row, lm, "line_management"):
        assert lm is not None
        row = _section(ws, row, "Line-increase activity by score band (full panel)")
        row = _write_table(ws, row, lm.tables["increases_by_band"], formats={"amount_added": MONEY})
        for label, key, fmt in (
            ("Line-increase events", "line_increase_events", MONEY),
            ("Exposure added ($)", "exposure_added_total", MONEY),
            ("Increase $ share to below-prime", "increase_share_below_prime", PCT),
            ("Increases 30+ DPD within 6 months", "increases_gone_bad_rate", PCT),
        ):
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=float(lm.summary[key]))
            cell.number_format = fmt
            row += 1
    _autosize(ws, (34, 14, 16, 14))


def write_workbook(review: ReviewResult, path: str | Path) -> Path:
    """Write the formatted review workbook for ``review`` to ``path``.

    Returns:
        The path written.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    assert ws is not None
    ws.title = "Dashboard"
    _dashboard(ws, review)
    _card_detail(wb.create_sheet("Card Detail"), review)
    _migration_sheet(wb.create_sheet("Migration Matrix"), review)
    _vintage_sheet(wb.create_sheet("Vintage Curves"), review)
    _concentration_sheet(wb.create_sheet("Concentration"), review)
    _utilization_sheet(wb.create_sheet("Utilization"), review)

    wb.save(path)
    return path
