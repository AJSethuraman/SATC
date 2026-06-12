"""End-to-end tests: generate -> degrade -> detect tier -> compute -> export."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from ucpa.engine import run_review
from ucpa.excel_report import write_workbook
from ucpa.findings_template import write_findings_template
from ucpa.generator import degrade_to_tier
from ucpa.products import CreditCardModule, PersonalLoanModule, StudentLoanModule

EXPECTED_SHEETS = [
    "Dashboard",
    "Card Detail",
    "Migration Matrix",
    "Vintage Curves",
    "Concentration",
    "Utilization",
]


def test_end_to_end_tier2(small_tape: pd.DataFrame, tmp_path) -> None:
    review = run_review(small_tape, CreditCardModule())
    assert review.tier_detection.detected_tier == 2
    assert all(r.status != "blocked" for r in review.metric_results)

    xlsx = write_workbook(review, tmp_path / "review.xlsx")
    wb = load_workbook(xlsx)
    assert wb.sheetnames == EXPECTED_SHEETS
    ws = wb["Dashboard"]
    cells = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, 60)}
    assert cells["Detected data tier"] == "Tier 2"
    assert cells["Open accounts"] == 490
    assert cells["30+ DPD balance rate"] == pytest.approx(0.010977873267800774)

    findings = write_findings_template(review, tmp_path / "findings.md")
    text = findings.read_text()
    assert "[ANALYST TO COMPLETE]" in text  # interpretation left to humans
    assert "Tier 2" in text
    assert "dpd30plus_balance_rate" in text
    # Rule-based observations appear, clearly labeled as non-conclusions.
    assert "Automated observations (deterministic, rule-based)" in text
    assert "(OBS-DQ-01)" in text
    assert "NOT analytical conclusions" in text


def test_end_to_end_degraded_tier1_reports_gaps(small_tape: pd.DataFrame, tmp_path) -> None:
    tape = degrade_to_tier(small_tape, 1)
    review = run_review(tape, CreditCardModule())
    assert review.tier_detection.detected_tier == 1

    blocked = {r.metric for r in review.metric_results if r.status == "blocked"}
    assert blocked == {"recovery_trends", "line_management"}
    gap_metrics = {g.metric for g in review.gaps}
    assert "charge_off_rates" in gap_metrics  # net CO blocked, gross computed
    assert "recovery_amount" in {f for g in review.gaps for f in g.missing_fields}

    xlsx = write_workbook(review, tmp_path / "review_t1.xlsx")
    assert load_workbook(xlsx).sheetnames == EXPECTED_SHEETS
    findings = write_findings_template(review, tmp_path / "findings_t1.md")
    assert "Data-maturity gap assessment" in findings.read_text()


def test_end_to_end_degraded_tier0(small_tape: pd.DataFrame, tmp_path) -> None:
    tape = degrade_to_tier(small_tape, 0)
    review = run_review(tape, CreditCardModule())
    assert review.tier_detection.detected_tier == 0

    computed = {r.metric for r in review.metric_results if r.status != "blocked"}
    assert computed == {"delinquency_distribution", "concentration"}
    # Snapshot tapes get no trend analytics: time series must be blocked.
    blocked_names = {r.metric for r in review.metric_results if r.status == "blocked"}
    assert "portfolio_time_series" in blocked_names
    # Every blocked metric leaves a structured data-gap finding.
    blocked = [r for r in review.metric_results if r.status == "blocked"]
    assert all(r.gaps for r in blocked)

    xlsx = write_workbook(review, tmp_path / "review_t0.xlsx")
    assert load_workbook(xlsx).sheetnames == EXPECTED_SHEETS


def test_phase2_products_define_interface_only() -> None:
    for cls in (PersonalLoanModule, StudentLoanModule):
        module = cls()
        with pytest.raises(NotImplementedError):
            module.metric_specs()
