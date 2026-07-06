"""Seam 3 — workbook build/reload + byte-identical determinism."""

import io

import openpyxl

from popbench import demo, mapping
from popbench.engine import build


def _demo_bytes():
    pop = demo.demo_population()
    props = mapping.propose(list(pop.columns))
    m = mapping.confirm([mapping.ColumnMap(p.header, p.field_id)
                         for p in props if p.field_id])
    return pop, m, build(pop, m)


def test_build_is_byte_identical():
    pop, m, first = _demo_bytes()
    second = build(pop, m)
    assert first == second  # same input -> identical bytes (no wall-clock leak)


def test_workbook_has_expected_tabs_and_reloads():
    _, _, data = _demo_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # Minimal demo carries FICO -> a CFPB band stratification tab; no product,
    # no delinquency -> no Strat_product_type / Delinquency / URCCP tabs.
    assert set(wb.sheetnames) == {
        "Dashboard", "Raw_Input", "_map", "_cleaning", "_config", "Population",
        "Strat_cfpb", "_code_py", "_readme"}
    assert wb.sheetnames[0] == "Dashboard"   # styled front page


def test_population_tab_carries_computed_values():
    _, _, data = _demo_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Population"]
    assert ws.cell(3, 2).value == 4              # loan count
    assert ws.cell(4, 2).value == 100000.0       # total balance
    # WA row for fico_orig: dollar 716.0, count 715.0
    header = [c.value for c in ws[6]]
    assert header[0] == "attribute"
    fico = [c.value for c in ws[7]]
    assert fico[0] == "fico_orig"
    assert abs(fico[1] - 716.0) < 1e-9
    assert abs(fico[2] - 715.0) < 1e-9


def test_code_py_tab_is_pure_ascii_for_discovery():
    _, _, data = _demo_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "_code_py" in wb.sheetnames          # control_center discovery key
    ws = wb["_code_py"]
    text = "\n".join(str(c.value or "") for c in ws["A"])
    text.encode("ascii")                        # raises if any non-ASCII slipped in
    assert "build_demo_bytes" in text


def test_map_tab_records_confirmed_mapping():
    _, _, data = _demo_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["_map"]
    rows = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(4, 9)}
    assert rows.get("loan_id") == "Loan No"
    assert rows.get("current_balance") == "Curr Bal"
