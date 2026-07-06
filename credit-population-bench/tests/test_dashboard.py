"""Seam 3 — the styled Excel Dashboard front page (house style + heat)."""

import io

import openpyxl

from popbench import demo, mapping, contract
from popbench.engine import build


def _demo_wb():
    pop = demo.demo_population_full()
    props = mapping.propose(list(pop.columns))
    m = mapping.confirm([mapping.ColumnMap(p.header, p.field_id,
                         contract.UNIT_FRACTION if p.needs_unit else None)
                         for p in props if p.field_id])
    data = build(pop, m, cohort_specs=demo.demo_cohorts())
    return openpyxl.load_workbook(io.BytesIO(data))


def test_dashboard_is_first_tab_and_styled():
    wb = _demo_wb()
    assert wb.sheetnames[0] == "Dashboard"
    ws = wb["Dashboard"]
    assert ws.sheet_view.showGridLines is False
    # KeyBank INK header band with CANVAS text + KEY_RED accent rule
    assert str(ws["A2"].fill.fgColor.rgb).endswith("0A0908")
    assert str(ws["A2"].font.color.rgb).endswith("F4F1EC")
    assert str(ws["A1"].fill.fgColor.rgb).endswith("CC0000")


def test_dashboard_has_heat_color_scales():
    ws = _demo_wb()["Dashboard"]
    # FICO risk gradient + vintage loss triangle each get a color-scale rule
    assert len(ws.conditional_formatting) >= 2


def test_dashboard_carries_no_tin_pattern():
    from satc_credit_core import pii
    wb = _demo_wb()
    text = "\n".join(str(c.value) for row in wb["Dashboard"].iter_rows() for c in row if c.value)
    pii.assert_no_tin(text, "dashboard")
