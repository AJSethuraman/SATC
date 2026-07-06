"""The Excel ``Dashboard`` tab — the deliverable's front page.

Brings the HTML examination report's *quality* into the workbook itself
(KeyBank house style: INK header band, KEY_RED accent, CANVAS KPI tiles,
conditional-format heat on the FICO risk gradient and the vintage loss
triangle) so the reviewer opens a polished dashboard, not raw value dumps.
Rendered from an :class:`~popbench.engine.AnalysisResult`; values only, styled.
"""

from __future__ import annotations

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from popbench import metrics, strata
from popbench.delinquency import BUCKET_LABEL

# --- KeyBank palette (ARGB, no '#') ---
INK, CANVAS, MIST, PAPER = "0A0908", "F4F1EC", "E4DFD5", "FFFFFF"
INK_TEXT, SLATE = "16130F", "57534B"
KEY_RED, CRIMSON, RED_SOFT, WARN_BG = "CC0000", "960019", "F7DEDE", "EFD7CF"

_TITLE = Font(name="Arial", bold=True, size=18, color=CANVAS)
_SUB = Font(name="Arial", size=10, color="CFC7BA")
_BANDF = Font(name="Arial", bold=True, size=11, color=CANVAS)
_KLABEL = Font(name="Arial", bold=True, size=9, color=SLATE)
_KVAL = Font(name="Arial", bold=True, size=18, color=INK_TEXT)
_HDR = Font(name="Arial", bold=True, size=9, color=SLATE)
_DATA = Font(name="Calibri", size=10, color=INK_TEXT)
_LBL = Font(name="Calibri", bold=True, size=10, color=INK_TEXT)
_RED = Font(name="Calibri", bold=True, size=10, color=KEY_RED)
_CRIT = Font(name="Arial", bold=True, size=10, color=CRIMSON)

_BAND_FILL = PatternFill("solid", fgColor=INK)
_TILE_FILL = PatternFill("solid", fgColor=CANVAS)
_RED_ACCENT = PatternFill("solid", fgColor=KEY_RED)
_SOFT_FILL = PatternFill("solid", fgColor=RED_SOFT)
_HDR_BOTTOM = Border(bottom=Side(style="medium", color=INK))
_HAIR = Border(bottom=Side(style="thin", color=MIST))
_L = Alignment("left", "center")
_R = Alignment("right", "center")
_C = Alignment("center", "center")

USD, PCT, PCT2, FIC = "#,##0", "0.0%", "0.00%", "0"
NCOLS = 8


def _band(ws, row, text, sub=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row, 1, text); c.font = _BANDF; c.fill = _BAND_FILL; c.alignment = _L
    return row + 1


def _hdr(ws, row, labels, aligns=None):
    for i, t in enumerate(labels, 1):
        c = ws.cell(row, i, t); c.font = _HDR; c.border = _HDR_BOTTOM
        c.alignment = _R if (aligns and aligns[i-1] == "r") else _L
    return row + 1


def build_dashboard(ws: Worksheet, result, sampling_doc=None) -> None:
    clean = result.clean
    ws.sheet_view.showGridLines = False
    widths = [26, 13, 13, 13, 13, 13, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- header band -------------------------------------------------------
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2"); ws.merge_cells("A3:H3")
    ws.row_dimensions[1].height = 8
    t = ws.cell(2, 1, "Consumer Portfolio Review — Dashboard"); t.font = _TITLE
    t.fill = _BAND_FILL; t.alignment = _L
    ws.cell(1, 1).fill = _RED_ACCENT
    pop_n = int(len(clean)); pop_b = float(clean["current_balance"].astype("float64").sum())
    sub = ws.cell(3, 1, f"{pop_n:,} loans  ·  ${pop_b:,.0f} UPB  ·  every rate on both a "
                        f"dollar and a count basis  ·  URCCP from the shared credit-core")
    sub.font = _SUB; sub.fill = _BAND_FILL; sub.alignment = _L
    row = 5

    # --- KPI tiles ---------------------------------------------------------
    wa = {w.attribute: w for w in metrics.weighted_average_table(clean, ["fico_orig", "dti", "rate"])}
    has_dpd = "dpd_min" in clean.columns
    d90b = float(clean.loc[clean["dpd_min"] >= 90, "current_balance"].astype("float64").sum()) if has_dpd else 0
    d90c = int((clean["dpd_min"] >= 90).sum()) if has_dpd else 0
    tiles = [("LOANS", f"{pop_n:,}"), ("TOTAL UPB", f"${pop_b/1e6:.1f}M")]
    if "fico_orig" in wa: tiles.append(("WA FICO", f"{wa['fico_orig'].dollar_weighted:.0f}"))
    if "dti" in wa: tiles.append(("WA DTI", f"{wa['dti'].dollar_weighted*100:.1f}%"))
    if has_dpd and pop_b:
        tiles.append(("90+ DPD  $ / #", f"{d90b/pop_b*100:.1f}% / {d90c/pop_n*100:.1f}%"))
    if result.charge_off is not None:
        tiles.append(("GROSS CHARGE-OFF $", f"{result.charge_off.dollar_rate*100:.2f}%"))
    # two tiles per 2-col block, up to 4 across; wrap to a second row
    col = 1
    for label, value in tiles:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        lc = ws.cell(row, col, label); lc.font = _KLABEL; lc.fill = _TILE_FILL; lc.alignment = _L
        ws.cell(row, col+1).fill = _TILE_FILL
        vc = ws.cell(row+1, col, value); vc.font = _KVAL; vc.fill = _TILE_FILL; vc.alignment = _L
        ws.cell(row+1, col+1).fill = _TILE_FILL
        col += 2
        if col > NCOLS:
            col = 1; row += 2
    row = row + (2 if col > 1 else 0) + 1

    # --- delinquency (both bases) -----------------------------------------
    if result.delinquency is not None:
        row = _band(ws, row, "Delinquency distribution  —  dollar AND count basis")
        row = _hdr(ws, row, ["Bucket", "Count", "Count %", "Balance $", "Dollar %"],
                   ["l", "r", "r", "r", "r"])
        for d in result.delinquency["per_bucket"] + result.delinquency["cumulative"]:
            ws.cell(row, 1, d.label).font = _LBL
            ws.cell(row, 2).value = d.count; ws.cell(row, 2).font = _DATA
            ws.cell(row, 3).value = d.count_rate; ws.cell(row, 3).number_format = PCT; ws.cell(row, 3).font = _DATA
            ws.cell(row, 4).value = d.dollars; ws.cell(row, 4).number_format = USD; ws.cell(row, 4).font = _DATA
            ws.cell(row, 5).value = d.dollar_rate; ws.cell(row, 5).number_format = PCT; ws.cell(row, 5).font = _RED
            for c in range(1, 6): ws.cell(row, c).border = _HAIR
            row += 1
        row += 1

    # --- FICO risk gradient (heat) ----------------------------------------
    grad = strata.fico_risk_gradient(clean)
    if grad:
        row = _band(ws, row, "FICO band  —  risk gradient (90+ dollar rate heat-shaded)")
        row = _hdr(ws, row, ["Band", "Loans", "UPB $", "% of $", "WA DTI", "90+ $"],
                   ["l", "r", "r", "r", "r", "r"])
        first = row
        for g in grad:
            ws.cell(row, 1, g["band"]).font = _LBL
            ws.cell(row, 2).value = g["n"]; ws.cell(row, 2).font = _DATA
            ws.cell(row, 3).value = g["balance"]; ws.cell(row, 3).number_format = USD; ws.cell(row, 3).font = _DATA
            ws.cell(row, 4).value = g["share"]; ws.cell(row, 4).number_format = PCT; ws.cell(row, 4).font = _DATA
            ws.cell(row, 5).value = g["wa_dti"]; ws.cell(row, 5).number_format = PCT; ws.cell(row, 5).font = _DATA
            ws.cell(row, 6).value = g["rate90"]; ws.cell(row, 6).number_format = PCT; ws.cell(row, 6).font = _LBL
            for c in range(1, 7): ws.cell(row, c).border = _HAIR
            row += 1
        ws.conditional_formatting.add(
            f"F{first}:F{row-1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFF4F1EC",
                           end_type="num", end_value=0.35, end_color="FFCC0000"))
        row += 1

    # --- URCCP -------------------------------------------------------------
    if result.urccp:
        row = _band(ws, row, "URCCP classification  —  shared floors (65 FR 36903)")
        row = _hdr(ws, row, ["Structure", "Loans", "UPB $", "Sub #", "Sub $", "Loss #", "Loss $"],
                   ["l", "r", "r", "r", "r", "r", "r"])
        for u in result.urccp:
            ws.cell(row, 1, u.structure.replace("_", " ")).font = _LBL
            ws.cell(row, 2).value = u.count; ws.cell(row, 2).font = _DATA
            ws.cell(row, 3).value = u.balance; ws.cell(row, 3).number_format = USD; ws.cell(row, 3).font = _DATA
            ws.cell(row, 4).value = u.substandard_count; ws.cell(row, 4).font = _DATA
            ws.cell(row, 5).value = u.substandard_dollars; ws.cell(row, 5).number_format = USD; ws.cell(row, 5).font = _DATA
            ws.cell(row, 6).value = u.loss_count; ws.cell(row, 6).font = _RED
            ws.cell(row, 7).value = u.loss_dollars; ws.cell(row, 7).number_format = USD; ws.cell(row, 7).font = _RED
            for c in range(1, 8): ws.cell(row, c).border = _HAIR
            row += 1
        row += 1

    # --- vintage triangle (heat) ------------------------------------------
    if result.vintage_rows and result.triangle is not None:
        tri = result.triangle
        row = _band(ws, row, "Vintage loss triangle  —  cumulative gross-loss rate by MOB")
        row = _hdr(ws, row, ["Vintage"] + [f"MOB {m}" for m in tri.mob_bounds],
                   ["l"] + ["r"] * len(tri.mob_bounds))
        first = row
        last_col = 1 + len(tri.mob_bounds)
        for v in sorted(tri.rows):
            ws.cell(row, 1, v).font = _LBL
            for j, m in enumerate(tri.mob_bounds, 2):
                c = ws.cell(row, j); c.value = tri.rows[v][m]; c.number_format = PCT2; c.font = _DATA
                c.border = _HAIR
            row += 1
        rng = f"B{first}:{get_column_letter(last_col)}{row-1}"
        ws.conditional_formatting.add(
            rng, ColorScaleRule(start_type="min", start_color="FFF4F1EC",
                                end_type="max", end_color="FFCC0000"))
        row += 1

    # --- sample coverage + OCC + banner -----------------------------------
    occ = sampling_doc if sampling_doc is not None else result.sampling_doc
    if occ is not None:
        row = _band(ws, row, "Judgmental sample  —  coverage per segment (mark picks on the Selection tab)")
        row = _hdr(ws, row, ["Segment", "Selected", "Count cov", "Dollar cov"], ["l", "r", "r", "r"])
        for c in occ.coverage:
            ws.cell(row, 1, c.value).font = _LBL
            ws.cell(row, 2, f"{c.sel_count}/{c.seg_count}").font = _DATA; ws.cell(row, 2).alignment = _R
            ws.cell(row, 3).value = c.count_coverage; ws.cell(row, 3).number_format = PCT; ws.cell(row, 3).font = _DATA
            ws.cell(row, 4).value = c.dollar_coverage; ws.cell(row, 4).number_format = PCT; ws.cell(row, 4).font = _RED
            for cc in range(1, 5): ws.cell(row, cc).border = _HAIR
            row += 1
        ws.cell(row, 1, "Sample size").font = _KLABEL
        ws.cell(row, 2, f"{occ.sample_count} loans / ${occ.sample_dollars:,.0f}  "
                        f"({occ.overall_dollar_coverage*100:.1f}% $)").font = _LBL
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        b = ws.cell(row, 1, occ.non_extrapolable); b.font = _CRIT; b.fill = _SOFT_FILL; b.alignment = _L
        ws.row_dimensions[row].height = 28
