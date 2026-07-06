"""Workbook round-trip — read the reviewer's edits back and recompute.

This is what makes judgmental sampling *operable* in the deliverable: the
reviewer marks ``SELECT = Y`` on the Selection tab (and may edit the ``_cohorts``
rules), presses the button, and :func:`run_workbook` reads Raw_Input + the
persisted mapping (``_map``) + the marked selection + cohort rules, recomputes
every lens, and writes the workbook back — coverage and the OCC documentation
now reflect the reviewer's actual pick. No web app; the pick surface is Excel.
"""

from __future__ import annotations

import openpyxl
import pandas as pd

from popbench import cohorts as coh_mod, engine, mapping as mapping_mod

_TRUE = {"y", "yes", "true", "1", "x", "t", "✓"}


def read_raw_input(wb) -> pd.DataFrame:
    """The loaded population from the Raw_Input tab (header row 1)."""
    data = list(wb["Raw_Input"].iter_rows(values_only=True))
    if not data:
        return pd.DataFrame()
    headers = list(data[0])
    df = pd.DataFrame(list(data[1:]), columns=headers)
    return df.dropna(how="all").reset_index(drop=True)


def reconstruct_mapping(wb):
    """Rebuild the confirmed Mapping from the ``_map`` tab (data from row 4)."""
    rows = []
    for row in wb["_map"].iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        vals = list(row) + [None, None, None]
        rows.append((vals[0], vals[1], vals[2] or ""))
    return mapping_mod.from_rows(rows)


def read_selection(wb) -> set:
    """Loan ids the reviewer marked SELECT=Y on the Selection tab (data row 5)."""
    if "Selection" not in wb.sheetnames:
        return set()
    sel = set()
    for row in wb["Selection"].iter_rows(min_row=5, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        mark = row[1] if len(row) > 1 else None
        if mark is not None and str(mark).strip().lower() in _TRUE:
            sel.add(str(row[0]))
    return sel


def read_cohorts(wb):
    """Cohort definitions from the ``_cohorts`` tab (data from row 5)."""
    if "_cohorts" not in wb.sheetnames:
        return []
    grouped: dict[str, dict] = {}
    order: list[str] = []
    for row in wb["_cohorts"].iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        vals = list(row) + [None] * 6
        cid, label, group, field, op, value = vals[:6]
        if cid not in grouped:
            grouped[cid] = {"label": label, "rules": []}
            order.append(cid)
        g = None if group in (None, "") else group
        grouped[cid]["rules"].append(coh_mod.Rule(field, op, value, g))
    return [coh_mod.Cohort(cid, grouped[cid]["label"] or cid, tuple(grouped[cid]["rules"]))
            for cid in order]


def run_workbook(path: str, sample_dimension: str = "product_type") -> bytes:
    """The button: recompute from the workbook's own Raw_Input + reviewer edits,
    honoring the marked selection and cohort rules, and write it back."""
    wb = openpyxl.load_workbook(path)
    raw = read_raw_input(wb)
    m = reconstruct_mapping(wb)
    selection = read_selection(wb)          # exactly what the reviewer marked
    cohort_specs = read_cohorts(wb) or None
    data = engine.build(raw, m, cohort_specs=cohort_specs,
                        sample_selection=selection, sample_dimension=sample_dimension)
    with open(path, "wb") as fh:
        fh.write(data)
    return data
