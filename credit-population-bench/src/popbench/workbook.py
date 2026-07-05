"""Deterministic workbook assembly — the output surface.

Excel is where the reviewer loads the file and reads the results; the math
already happened in pandas. This module lays computed *values* into tabs and
serializes them **byte-identically** for the same inputs (fixed document
properties + a zip repack that pins every member timestamp — the same trick
``credit-review-os`` uses so a rebuild diffs clean).

Slice 1 tabs: ``Raw_Input`` (the loaded population), ``_map`` (the confirmed
column mapping + units), ``_cleaning`` (the audit trail), ``_config`` (the knob
panel), ``Population`` (headline metrics), ``_readme``. Later slices add the
stratification, cohort, vintage, sampling, and packaging tabs; the ``.xlsm``
macro + ``_code_py`` runner land with the packaging slice.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# House-style-adjacent minimal palette (warm neutrals; red = stress later).
_INK = "16130F"
_BAND = "0A0908"
_CANVAS = "F4F1EC"
_PAPER = "FFFFFF"

_HDR_FONT = Font(name="Arial", bold=True, size=11, color=_PAPER)
_HDR_FILL = PatternFill("solid", fgColor=_BAND)
_LABEL_FONT = Font(name="Arial", bold=True, size=10, color=_INK)
_DATA_FONT = Font(name="Calibri", size=11, color=_INK)
_KPI_FILL = PatternFill("solid", fgColor=_CANVAS)
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

# Fixed epoch — the build must never depend on the wall clock.
_EPOCH = datetime(2026, 1, 1)

FMT_USD = '#,##0'
FMT_NUM = '#,##0.0'


def _safe_tab(name: str, existing: list[str]) -> str:
    """Excel tab names are <=31 chars and unique; truncate and de-dup."""
    base = name[:31]
    candidate = base
    i = 2
    while candidate in existing:
        suffix = f"_{i}"
        candidate = base[:31 - len(suffix)] + suffix
        i += 1
    return candidate


def _band(ws: Worksheet, row: int, cols: list[str]) -> int:
    for c, text in enumerate(cols, start=1):
        cell = ws.cell(row, c, text)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _LEFT
    return row + 1


def _write_frame(ws: Worksheet, df: pd.DataFrame, start_row: int = 1) -> int:
    row = _band(ws, start_row, list(df.columns))
    for _, rec in df.iterrows():
        for c, col in enumerate(df.columns, start=1):
            val = rec[col]
            if pd.isna(val):
                val = None
            elif hasattr(val, "item"):
                val = val.item()
            ws.cell(row, c).value = val
        row += 1
    return row


def build_workbook(
    raw: pd.DataFrame,
    map_rows: list[tuple[str, str, str]],
    cleaning_rows: list[tuple[str, str, int, str]],
    population: dict,
    wa_rows: list[tuple],
    config_rows: list[tuple[str, str]] | None = None,
    delinquency_result: dict | None = None,
    urccp_rows: list | None = None,
    stratifications: list | None = None,
    cohort_comparison: object | None = None,
) -> Workbook:
    """Assemble the in-memory workbook from already-computed values.

    ``delinquency_result`` / ``urccp_rows`` add the Delinquency and URCCP tabs
    when the population carried the fields to compute them (else those tabs are
    simply absent — feature-gated output)."""
    wb = Workbook()
    wb.properties.creator = "Consumer Credit Population Bench"
    wb.properties.created = _EPOCH
    wb.properties.modified = _EPOCH

    # Raw_Input
    ws = wb.active
    ws.title = "Raw_Input"
    _write_frame(ws, raw)

    # _map
    ws = wb.create_sheet("_map")
    ws.cell(1, 1, "Confirmed column mapping (canonical <- header; units)").font = _LABEL_FONT
    r = _band(ws, 3, ["canonical_field", "source_header", "unit"])
    for fid, header, unit in map_rows:
        ws.cell(r, 1, fid); ws.cell(r, 2, header); ws.cell(r, 3, unit)
        r += 1

    # _cleaning
    ws = wb.create_sheet("_cleaning")
    ws.cell(1, 1, "Cleaning audit — every automatic normalization applied").font = _LABEL_FONT
    r = _band(ws, 3, ["rule", "column", "rows_affected", "detail"])
    for rule, col, n, detail in cleaning_rows:
        ws.cell(r, 1, rule); ws.cell(r, 2, col)
        ws.cell(r, 3).value = int(n); ws.cell(r, 4, detail)
        r += 1

    # Population
    ws = wb.create_sheet("Population")
    ws.cell(1, 1, "Population summary").font = Font(name="Arial", bold=True, size=14, color=_INK)
    ws.cell(3, 1, "Loan count").font = _LABEL_FONT
    ws.cell(3, 2).value = int(population["count"])
    ws.cell(4, 1, "Total balance (UPB)").font = _LABEL_FONT
    ws.cell(4, 2).value = float(population["total_balance"])
    ws.cell(4, 2).number_format = FMT_USD
    for cc in (ws.cell(3, 1), ws.cell(4, 1), ws.cell(3, 2), ws.cell(4, 2)):
        cc.fill = _KPI_FILL
    r = _band(ws, 6, ["attribute", "dollar_weighted", "count_weighted",
                      "n", "coverage_$", "note"])
    for attr, dollar, count, n, cov, note in wa_rows:
        ws.cell(r, 1, attr)
        ws.cell(r, 2).value = None if dollar is None else float(dollar)
        ws.cell(r, 3).value = None if count is None else float(count)
        ws.cell(r, 4).value = int(n)
        ws.cell(r, 5).value = float(cov); ws.cell(r, 5).number_format = FMT_USD
        ws.cell(r, 6, note)
        r += 1

    # Delinquency (both bases, always labeled) — only when computed.
    if delinquency_result is not None:
        ws = wb.create_sheet("Delinquency")
        ws.cell(1, 1, "Delinquency rates — dollar AND count basis").font = \
            Font(name="Arial", bold=True, size=14, color=_INK)
        ws.cell(2, 1, delinquency_result["basis_note"]).font = _DATA_FONT
        r = _band(ws, 4, ["bucket", "count", "balance_$",
                          "count_rate (n/N)", "dollar_rate ($/$)"])
        def _rate_rows(rows):
            nonlocal r
            for row in rows:
                ws.cell(r, 1, row.label)
                ws.cell(r, 2).value = int(row.count)
                ws.cell(r, 3).value = float(row.dollars); ws.cell(r, 3).number_format = FMT_USD
                ws.cell(r, 4).value = float(row.count_rate); ws.cell(r, 4).number_format = "0.0%"
                ws.cell(r, 5).value = float(row.dollar_rate); ws.cell(r, 5).number_format = "0.0%"
                r += 1
        _rate_rows(delinquency_result["per_bucket"])
        r = _band(ws, r + 1, ["cumulative", "count", "balance_$",
                              "count_rate", "dollar_rate"])
        _rate_rows(delinquency_result["cumulative"])

    # URCCP classification (shared floors) — only when computed.
    if urccp_rows:
        ws = wb.create_sheet("URCCP")
        ws.cell(1, 1, "URCCP pool classification (65 FR 36903; shared floors)").font = \
            Font(name="Arial", bold=True, size=14, color=_INK)
        r = _band(ws, 3, ["structure", "count", "balance_$",
                          "Substandard_#", "Substandard_$",
                          "Loss_#", "Loss_$", "note"])
        for row in urccp_rows:
            ws.cell(r, 1, row.structure)
            ws.cell(r, 2).value = int(row.count)
            ws.cell(r, 3).value = float(row.balance); ws.cell(r, 3).number_format = FMT_USD
            ws.cell(r, 4).value = int(row.substandard_count)
            ws.cell(r, 5).value = float(row.substandard_dollars); ws.cell(r, 5).number_format = FMT_USD
            ws.cell(r, 6).value = None if row.loss_count is None else int(row.loss_count)
            ws.cell(r, 7).value = None if row.loss_dollars is None else float(row.loss_dollars)
            if row.loss_dollars is not None:
                ws.cell(r, 7).number_format = FMT_USD
            ws.cell(r, 8, row.note)
            r += 1

    # Stratification tabs (group-by partitions) — one per requested dimension.
    for strat in (stratifications or []):
        title = _safe_tab(f"Strat_{strat.dimension}", wb.sheetnames)
        ws = wb.create_sheet(title)
        ws.cell(1, 1, f"Group-by: {strat.dimension}").font = \
            Font(name="Arial", bold=True, size=14, color=_INK)
        recon = "reconciles to population" if strat.reconciles else "DOES NOT reconcile"
        ws.cell(2, 1, f"partition ({recon}); WA shown dollar-weighted "
                      f"(count basis on Population)").font = _DATA_FONT
        # collect attributes present across cells for stable columns
        attrs: list[str] = []
        for c in strat.cells:
            for w in c.wa:
                if w.attribute not in attrs:
                    attrs.append(w.attribute)
        header = ["cell", "count", "balance_$"] + [f"WA_{a}" for a in attrs]
        r = _band(ws, 4, header)
        for c in strat.cells:
            ws.cell(r, 1, c.value)
            ws.cell(r, 2).value = int(c.count)
            ws.cell(r, 3).value = float(c.balance); ws.cell(r, 3).number_format = FMT_USD
            wa_by = {w.attribute: w for w in c.wa}
            for j, a in enumerate(attrs, start=4):
                w = wa_by.get(a)
                ws.cell(r, j).value = None if (w is None or w.dollar_weighted is None) \
                    else float(w.dollar_weighted)
            r += 1
        # total row for the partition
        ws.cell(r, 1, "TOTAL").font = _LABEL_FONT
        ws.cell(r, 2).value = int(strat.total_count)
        ws.cell(r, 3).value = float(strat.total_balance); ws.cell(r, 3).number_format = FMT_USD

    # Cohorts — independent reports + overlap matrix + residual (never summed).
    if cohort_comparison is not None:
        cc = cohort_comparison
        ws = wb.create_sheet("Cohorts")
        ws.cell(1, 1, "Derived cohorts — reported independently").font = \
            Font(name="Arial", bold=True, size=14, color=_INK)
        ws.cell(2, 1, "Cohorts OVERLAP and leave gaps; never sum them as a "
                      "partition. See the overlap matrix + residual below.").font = _DATA_FONT
        r = _band(ws, 4, ["cohort", "count", "balance_$", "% of pop $"])
        for rep in cc.reports:
            ws.cell(r, 1, rep.label)
            ws.cell(r, 2).value = int(rep.count)
            ws.cell(r, 3).value = float(rep.balance); ws.cell(r, 3).number_format = FMT_USD
            share = (rep.balance / cc.population_dollars) if cc.population_dollars else 0.0
            ws.cell(r, 4).value = float(share); ws.cell(r, 4).number_format = "0.0%"
            r += 1
        # overlap matrix
        r = _band(ws, r + 1, ["overlap (pairwise)", "count", "balance_$"])
        ids = [rep.id for rep in cc.reports]
        label_by = {rep.id: rep.label for rep in cc.reports}
        for (a, b), n in cc.overlap_count.items():
            ws.cell(r, 1, f"{label_by[a]}  x  {label_by[b]}")
            ws.cell(r, 2).value = int(n)
            ws.cell(r, 3).value = float(cc.overlap_dollars[(a, b)]); ws.cell(r, 3).number_format = FMT_USD
            r += 1
        # residual + population
        ws.cell(r, 1, "In NO cohort (residual)").font = _LABEL_FONT
        ws.cell(r, 2).value = int(cc.residual_count)
        ws.cell(r, 3).value = float(cc.residual_dollars); ws.cell(r, 3).number_format = FMT_USD
        r += 1
        ws.cell(r, 1, "Population").font = _LABEL_FONT
        ws.cell(r, 2).value = int(cc.population_count)
        ws.cell(r, 3).value = float(cc.population_dollars); ws.cell(r, 3).number_format = FMT_USD

    # _config (knob panel — populated further by later slices)
    ws = wb.create_sheet("_config")
    r = _band(ws, 1, ["[SETTINGS]", "", ""])
    for key, val in (config_rows or [("weight_field", "current_balance")]):
        ws.cell(r, 1, key); ws.cell(r, 2, val)
        r += 1

    # _readme
    ws = wb.create_sheet("_readme")
    for i, line in enumerate(_README_LINES, start=1):
        ws.cell(i, 1, line).font = _DATA_FONT
    return wb


_README_LINES = (
    "Consumer Credit Population Analysis Bench",
    "",
    "Load a loan-level consumer-loan flat file on Raw_Input; confirm the column",
    "mapping on _map; the engine validates/cleans (audit on _cleaning) and writes",
    "population metrics here. Every rate ships on BOTH a dollar and a count basis",
    "(labeled). Weighted averages are balance-weighted by current UPB by default.",
    "",
    "Nothing is computed on a dirty population: hard errors refuse; only safe,",
    "declared normalizations run automatically and are recorded on _cleaning.",
    "",
    "PII boundary: this runtime workbook may hold real borrower data and is for",
    "use only on bank equipment; the shipped tool/repo/demo fixtures carry none.",
)


def workbook_bytes(wb: Workbook) -> bytes:
    """Serialize deterministically: identical content -> identical bytes."""
    raw = io.BytesIO()
    wb.save(raw)
    fixed = (b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
             + _EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ").encode() + b"</dcterms:modified>")
    out = io.BytesIO()
    with zipfile.ZipFile(raw) as src, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for name in src.namelist():
            data = src.read(name)
            if name == "docProps/core.xml":
                data = re.sub(rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                              fixed, data)
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(info, data)
    return out.getvalue()
