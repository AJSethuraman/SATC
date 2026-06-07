#!/usr/bin/env python3
"""Year-by-year fee schedule in one workbook.

Keeps a ``fee_schedule.xlsx`` where each tax year is its own sheet -- so you can see
prices and discounts side by side across years, edit them in Excel, and build next
year from this one. Running the tool:

  * creates the workbook (and the target year's sheet) from the current fee schedule
    if it does not exist yet;
  * reads the target year's sheet and writes it to ``fee_schedule.json`` -- the file
    the Calculate Invoices tool uses -- so your workbook edits take effect;
  * copies the target year to a next-year sheet (ready to adjust) unless told not to.

Each sheet has a FORMS table (Key / Description / Price / Additional) and a DISCOUNTS
table (Key / Description / Amount / Percent). Needs openpyxl (imported lazily).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import generate_documents
import invoice_calc
import sort_tax_docs

WORKBOOK_FILENAME = "fee_schedule.xlsx"
FORMS_MARKER = "FORMS"
DISCOUNTS_MARKER = "DISCOUNTS"


def target_year(input_folder: Path, explicit=None) -> int:
    """Year to operate on: explicit, else the latest client tax_year, else this year."""

    if explicit:
        return int(str(explicit)[:4])
    data_file = generate_documents.find_client_data_file(input_folder)
    years: list[int] = []
    if data_file is not None:
        for client in generate_documents.load_clients(data_file):
            try:
                years.append(int(str(client.get("tax_year", "")).strip()[:4]))
            except (ValueError, TypeError):
                continue
    return max(years) if years else date.today().year


def _num(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def write_year_sheet(worksheet, schedule: dict, year: int) -> None:
    """Lay a fee schedule onto a worksheet as FORMS and DISCOUNTS tables."""

    worksheet.delete_rows(1, worksheet.max_row or 1)
    worksheet["A1"] = "Fee Schedule"
    worksheet["B1"] = int(year)
    worksheet.append([])
    worksheet.append([FORMS_MARKER])
    worksheet.append(["Key", "Description", "Price", "Additional"])
    for key, entry in schedule.items():
        if key == invoice_calc.DISCOUNTS_KEY or not isinstance(entry, dict) or "price" not in entry:
            continue
        worksheet.append([key, entry.get("description", ""), entry.get("price"), entry.get("additional", "")])
    worksheet.append([])
    worksheet.append([DISCOUNTS_MARKER])
    worksheet.append(["Key", "Description", "Amount", "Percent"])
    for key, config in (schedule.get(invoice_calc.DISCOUNTS_KEY) or {}).items():
        worksheet.append([key, config.get("description", ""), config.get("amount", ""), config.get("percent", "")])


def read_year_sheet(worksheet) -> dict:
    """Parse a worksheet's FORMS and DISCOUNTS tables back into a fee schedule dict."""

    schedule: dict = {}
    discounts: dict = {}
    mode = None
    for row in worksheet.iter_rows(values_only=True):
        first = row[0] if row else None
        if first == FORMS_MARKER:
            mode = "forms_header"
            continue
        if first == DISCOUNTS_MARKER:
            mode = "discounts_header"
            continue
        if mode == "forms_header":
            mode = "forms"
            continue
        if mode == "discounts_header":
            mode = "discounts"
            continue
        if not first:
            continue
        if mode == "forms":
            entry = {"description": row[1] or "", "price": _num(row[2]) or 0.0}
            additional = _num(row[3]) if len(row) > 3 else None
            if additional is not None:
                entry["additional"] = additional
            schedule[str(first)] = entry
        elif mode == "discounts":
            config = {"description": row[1] or ""}
            percent = _num(row[3]) if len(row) > 3 else None
            amount = _num(row[2]) if len(row) > 2 else None
            if percent is not None:
                config["percent"] = percent
            elif amount is not None:
                config["amount"] = amount
            discounts[str(first)] = config
    if discounts:
        schedule[invoice_calc.DISCOUNTS_KEY] = discounts
    return schedule


def run_fee_workbook(input_folder, year=None, make_next=True, status_callback=None) -> dict:
    """Sync the per-year fee workbook with fee_schedule.json and prep next year."""

    input_folder = Path(input_folder)
    sort_tax_docs.setup_output_folders(input_folder)

    base_result = {
        "tool": "feeworkbook",
        "workbook_path": None,
        "year": None,
        "sheets": [],
        "next_year_created": False,
        "warnings": [],
    }

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return {**base_result, "summary": "The Fee Workbook tool needs openpyxl (pip install openpyxl)."}

    import json

    year = target_year(input_folder, year)
    workbook_path = input_folder / WORKBOOK_FILENAME
    if status_callback:
        status_callback(f"Updating fee workbook for {year}")

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the default empty sheet

    sheet_name = str(year)
    if sheet_name in workbook.sheetnames:
        schedule = read_year_sheet(workbook[sheet_name])
    else:
        # Seed this year's sheet from the current fee_schedule.json (or the default).
        schedule, _ = invoice_calc.load_fee_schedule(input_folder)
        write_year_sheet(workbook.create_sheet(sheet_name), schedule, year)

    # Apply the workbook year to the JSON the invoice tool reads.
    (input_folder / invoice_calc.FEE_SCHEDULE_FILENAME).write_text(
        json.dumps(schedule, indent=2), encoding="utf-8"
    )

    next_created = False
    if make_next:
        next_name = str(year + 1)
        if next_name not in workbook.sheetnames:
            copied = workbook.copy_worksheet(workbook[sheet_name])
            copied.title = next_name
            copied["B1"] = year + 1
            next_created = True

    # Keep sheets ordered by year for easy year-over-year viewing.
    workbook._sheets.sort(key=lambda ws: ws.title)  # noqa: SLF001 - stable order by year
    workbook.save(workbook_path)

    return {
        **base_result,
        "workbook_path": workbook_path,
        "year": year,
        "sheets": list(workbook.sheetnames),
        "next_year_created": next_created,
        "summary": (
            f"Fee workbook synced for {year} (sheets: {', '.join(workbook.sheetnames)})"
            + (f"; created {year + 1} for next year." if next_created else ".")
        ),
    }


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Manage a year-by-year fee schedule workbook.")
    parser.add_argument("input_folder", help="Folder containing the fee workbook / clients.")
    parser.add_argument("--year", default="", help="Tax year to apply (default: latest client year).")
    parser.add_argument("--no-next", action="store_true", help="Do not create the next-year sheet.")
    args = parser.parse_args()

    folder = Path(args.input_folder).expanduser().resolve()
    if not folder.is_dir():
        print(f"Input folder does not exist or is not a directory: {folder}")
        return 1

    result = run_fee_workbook(folder, year=args.year or None, make_next=not args.no_next, status_callback=print)
    print(result["summary"])
    if result["workbook_path"]:
        print(f"Workbook: {result['workbook_path']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
