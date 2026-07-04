"""Headless test suite for the Macro Early-Warning Monitor.

Runs on Linux/CI with NO Excel and NO network, all in --demo mode. Mirrors the
named tests in BUILD_SPEC_MACRO.md Section 6 (Phases 1-7):

  test_config_parse                     (Phase 1)
  test_demo_provider_deterministic      (Phase 2)
  test_fred_provider_parse_and_cache    (Phase 2 -- fixture JSON, no network)
  test_transforms                       (Phase 3)
  test_reload_headless                  (Phase 4)
  test_watchlist_admits_states          (Phase 5 POSITIVE)
  test_watchlist_refusal                (Phase 5)
  test_watchlist_gate_defense_in_depth  (Phase 5 negative)
  test_stale_flag                       (Phase 6 -- the SLIND regression)
  test_raw_landing_idempotent           (Phase 7)
  test_raw_layout_mismatch_refused      (Phase 7)
"""
import json
import math
import os
import re
import shutil
import struct
import sys
from dataclasses import replace
from datetime import date

import openpyxl
import pandas as pd
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import runner as R                     # noqa: E402
import series_seed as SEED             # noqa: E402
import build_workbook as BW            # noqa: E402
import assemble_xlsm                   # noqa: E402
import vba_writer as V                 # noqa: E402

ASOF = date(2026, 3, 31)

TABS = ("Dashboard_Conditions", "Dashboard_Labor", "Dashboard_Housing_Sentiment",
        "Watchlist", "Raw_FRED", "_config", "_code_py", "_code_vba", "_readme")


# --------------------------------------------------------------------------
# Fixtures
# --------------------------------------------------------------------------
@pytest.fixture(scope="session")
def xlsm(tmp_path_factory):
    """Build the macro-enabled workbook once for the whole session."""
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    out = str(d / "Macro_Early_Warning_Monitor.xlsm")
    BW.build(base)
    assemble_xlsm.assemble(base, out)
    return out


@pytest.fixture(scope="session")
def populated(xlsm, tmp_path_factory):
    """A fresh copy of the workbook with a --demo run applied (session-scoped:
    the run lands 172 series; the read-only assertions share one copy)."""
    d = tmp_path_factory.mktemp("run")
    p = str(d / "run.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)
    return p


def _spec_kwargs():
    return dict(id="x", title="x", category="labor", lane="dashboard",
                metric_type="claims", frequency="monthly", sa_nsa="SA",
                units="pct", level_rate_index="rate", geo_segment="national",
                source_class="A", dashboard_capable=True, watchlist_capable=False,
                source_url="", table_id="", sheet="", series_label="",
                transform="level", notes="")


# --------------------------------------------------------------------------
# Phase 1 -- config parse
# --------------------------------------------------------------------------
def test_config_parse():
    cfg = R.parse_config(BW.config_rows())
    # ~171 rows: 21 national + 51 UR + 50 ICLAIMS + 50 PHCI
    assert len(cfg.series) == 172
    ids = {s.id for s in cfg.series}
    assert {"T10Y3M", "SAHMREALTIME", "ICSA", "BAMLH0A0HYM2", "DRTSCILM",
            "CAUR", "TXICLAIMS", "NYPHCI", "DCUR"} <= ids
    # forbidden dead/frozen ids must NOT be fetchable rows (the SLIND lesson)
    for dead in ("TEDRATE", "CFSI", "USSLIND"):
        assert dead not in ids
    assert not any("SLIND" in i for i in ids)
    # weekly yoy_ma4 needs >= 56 observations; raw_slots is build-bound
    assert cfg.raw_slots >= 56
    # every watchlist row is genuinely state-keyed; every national row is not
    # watchlist-promotable
    for s in cfg.series:
        assert s.source_class == "A"
        assert s.lane in ("dashboard", "watchlist")
        assert s.transform in R.TRANSFORMS
        if s.lane == "watchlist":
            assert re.match(r"^state:[A-Z]{2}$", s.geo_segment), s.id
            assert s.watchlist_capable is True
        else:
            assert s.geo_segment == "national"
            assert s.watchlist_capable is False
    assert sum(1 for s in cfg.series if s.lane == "watchlist") == 151
    # settings the runner depends on
    assert cfg.sahm_state_band == pytest.approx(0.5)
    assert cfg.stale_multiplier == pytest.approx(2.0)


# --------------------------------------------------------------------------
# Phase 2 -- DemoProvider determinism + live-provider parse/cache (no network)
# --------------------------------------------------------------------------
def test_demo_provider_deterministic():
    cfg = R.parse_config(BW.config_rows())
    prov = R.FredDemoProvider(asof=ASOF, slots=cfg.raw_slots)
    for sid in ("ICSA", "CAUR", "TXICLAIMS"):     # weekly + monthly cadences
        spec = next(s for s in cfg.series if s.id == sid)
        a = prov.fetch_series(spec)
        b = R.FredDemoProvider(asof=ASOF, slots=cfg.raw_slots).fetch_series(spec)
        assert [(r.period, r.value) for r in a] == [(r.period, r.value) for r in b]
        assert len(a) == cfg.raw_slots


FIXTURE_JSON = json.dumps({
    "observations": [
        {"date": "2026-03-01", "value": "4.2"},
        {"date": "2026-02-01", "value": "."},        # FRED missing marker
        {"date": "2026-01-01", "value": "4.0"},
    ]
}).encode("utf-8")


def test_fred_provider_parse_and_cache(monkeypatch):
    """FIXED in-test JSON fixture incl. '.' -> None; the URL cache serves
    repeat calls; NO network (the one real call is stubbed out)."""
    calls = []

    def fake_get(self, url):
        calls.append(url)
        return FIXTURE_JSON

    monkeypatch.setattr(R.FredProvider, "_http_get", fake_get)
    prov = R.FredProvider(api_key="test-key", min_interval=0.0, slots=60)
    spec = R.SeriesSpec(**{**_spec_kwargs(), "id": "CAUR"})

    rows = prov.fetch_series(spec)
    # desc from the API -> normalized oldest-first
    assert [r.period for r in rows] == ["2026-01-01", "2026-02-01", "2026-03-01"]
    assert [r.value for r in rows] == [4.0, None, 4.2]
    assert rows[0].id == "CAUR" and rows[0].source_class == "A"
    # the request carried the contract params, never a hardcoded key elsewhere
    assert "series_id=CAUR" in calls[0] and "file_type=json" in calls[0]
    assert "sort_order=desc" in calls[0] and "limit=60" in calls[0]

    # cache hit: a second fetch makes NO further call
    prov.fetch_series(spec)
    assert len(calls) == 1

    # missing key fails fast, no request
    prov2 = R.FredProvider(api_key="", min_interval=0.0)
    with pytest.raises(ValueError, match="FRED API key"):
        prov2.fetch_series(spec)
    assert len(calls) == 1


def test_fred_provider_429_backoff(monkeypatch):
    """429 is retried with backoff and then succeeds -- unit-level, stubbed."""
    import urllib.error
    attempts = []

    def flaky_get(self, url):
        attempts.append(url)
        if len(attempts) < 3:
            raise urllib.error.HTTPError(url, 429, "Too Many Requests", None, None)
        return FIXTURE_JSON

    sleeps = []
    monkeypatch.setattr(R.FredProvider, "_http_get", flaky_get)
    monkeypatch.setattr(R.time, "sleep", lambda s: sleeps.append(s))
    prov = R.FredProvider(api_key="k", min_interval=0.0, max_retries=4)
    rows = prov.fetch_series(R.SeriesSpec(**_spec_kwargs()))
    assert len(attempts) == 3 and len(rows) == 3
    assert sleeps                                  # backoff actually slept


# --------------------------------------------------------------------------
# Phase 3 -- transform registry + threshold engine (hardcoded expected)
# --------------------------------------------------------------------------
def test_transforms():
    s = pd.Series([100.0, 102.0, 104.0, 106.0, 110.0, 112.0, 114.0, 116.0])

    # level == identity
    assert list(R.t_level(s, "quarterly")) == list(s)

    # yoy_pct (quarterly -> 4-period lag): index 4 = (110-100)/100*100 = 10.0
    yoy = R.t_yoy_pct(s, "quarterly")
    assert math.isnan(yoy.iloc[0])
    assert yoy.iloc[4] == pytest.approx(10.0)
    assert yoy.iloc[5] == pytest.approx((112 - 102) / 102 * 100)

    # qoq/mom = 1-period change
    assert R.t_qoq_pct(s, "quarterly").iloc[1] == pytest.approx(2.0)
    assert R.t_mom_pct(s, "monthly").iloc[1] == pytest.approx(2.0)

    # zscore_8q: mean=108, std(ddof=1)=sqrt(240/7); z(last)=(116-108)/std
    z = R.t_zscore_8q(s, "quarterly")
    std = math.sqrt(240.0 / 7.0)
    assert z.iloc[7] == pytest.approx((116 - 108) / std)
    assert math.isnan(z.iloc[6])           # min_periods=8

    # index_to_pct: percent vs the BASE (first valid obs), not a YoY alias
    idx = R.t_index_to_pct(s, "quarterly")
    assert idx.iloc[0] == pytest.approx(0.0)
    assert idx.iloc[7] == pytest.approx(16.0)          # (116/100-1)*100
    assert idx.iloc[5] == pytest.approx(12.0)
    assert R.t_yoy_pct(s, "quarterly").iloc[5] != pytest.approx(12.0)

    # ma4: 4-period MA, NaN until 4 obs
    m = R.t_ma4(s, "weekly")
    assert math.isnan(m.iloc[2])
    assert m.iloc[3] == pytest.approx(103.0)           # mean(100,102,104,106)
    assert m.iloc[7] == pytest.approx(113.0)           # mean(110,112,114,116)

    # yoy_ma4 on a linear weekly series 1..60 (P=52):
    # ma4(59)=mean(57..60)=58.5; ma4(7)=mean(5..8)=6.5 -> (58.5-6.5)/6.5*100
    lin = pd.Series([float(i) for i in range(1, 61)])
    y4 = R.t_yoy_ma4(lin, "weekly")
    assert y4.iloc[59] == pytest.approx((58.5 - 6.5) / 6.5 * 100)
    assert math.isnan(y4.iloc[54])                     # needs 52+4 history

    # yoy_ma3 monthly (P=12) on 1..20: ma3(19)=19, ma3(7)=7 -> (19-7)/7*100
    lin20 = pd.Series([float(i) for i in range(1, 21)])
    y3 = R.t_yoy_ma3(lin20, "monthly")
    assert y3.iloc[19] == pytest.approx((19.0 - 7.0) / 7.0 * 100)

    # chg_3p: arithmetic 3-period change
    c = R.t_chg_3p(s, "monthly")
    assert c.iloc[3] == pytest.approx(6.0)
    assert math.isnan(c.iloc[2])

    # sahm_gap on 1..18: mean(latest 3)=mean(16,17,18)=17; the 12 observations
    # before those 3 are 4..15 -> min 4; gap = 13. Matches the Excel formula
    # AVERAGE(v0:v2)-MIN(v3:v14) exactly (spec sec 3 parity requirement).
    g = R.t_sahm_gap(pd.Series([float(i) for i in range(1, 19)]), "monthly")
    assert g.iloc[17] == pytest.approx(13.0)
    assert g.iloc[14] == pytest.approx(13.0)           # first valid: 15 obs
    assert math.isnan(g.iloc[13])
    # a rise in the latest 3 vs a flat trailing window is measured directly
    flat = [5.0] * 15 + [5.2, 5.4, 5.6]
    g2 = R.t_sahm_gap(pd.Series(flat), "monthly")
    assert g2.iloc[-1] == pytest.approx(0.4)           # mean(5.2,5.4,5.6)-5.0

    # NaN propagates (no forward-fill, any pandas version; Oct-2025 hole M3)
    s_nan = pd.Series([100.0, math.nan, 104.0, 106.0, 110.0, 112.0])
    assert math.isnan(R.t_yoy_pct(s_nan, "quarterly").iloc[5])
    assert math.isnan(R.t_ma4(s_nan, "weekly").iloc[4])       # window hits NaN
    g_nan = R.t_sahm_gap(pd.Series([math.nan] + [1.0] * 17), "monthly")
    assert math.isnan(g_nan.iloc[14])                  # min-window hits the NaN
    assert g_nan.iloc[15] == pytest.approx(0.0)        # NaN rolled out of range


def test_threshold_engine():
    above = R.Threshold(watch=2.0, alert=3.0, direction="above")
    spec = R.SeriesSpec(**_spec_kwargs())
    assert R.status_for(spec, 3.5, above) == "ALERT"
    assert R.status_for(spec, 2.5, above) == "WATCH"
    assert R.status_for(spec, 1.0, above) == "OK"
    assert R.status_for(spec, None, above) == "OK"
    assert R.status_for(spec, float("nan"), above) == "OK"
    # the curve thresholds run direction="below" (inversion)
    below = R.Threshold(watch=0.0, alert=-0.5, direction="below")
    assert R.status_for(spec, -0.7, below) == "ALERT"
    assert R.status_for(spec, -0.2, below) == "WATCH"
    assert R.status_for(spec, 1.0, below) == "OK"
    # the state band engine matches the Watchlist formula exactly
    wl = R.SeriesSpec(**{**_spec_kwargs(), "transform": "sahm_gap",
                         "lane": "watchlist", "geo_segment": "state:CA"})
    assert R.watchlist_status(wl, 0.55, 0.5) == "ALERT"
    assert R.watchlist_status(wl, 0.35, 0.5) == "WATCH"
    assert R.watchlist_status(wl, 0.10, 0.5) == "OK"
    assert R.watchlist_status(wl, None, 0.5) == "OK"
    ctx = R.SeriesSpec(**{**_spec_kwargs(), "transform": "yoy_ma4",
                          "lane": "watchlist", "geo_segment": "state:CA"})
    assert R.watchlist_status(ctx, 99.0, 0.5) == "OK"  # context lanes never band-alert


# --------------------------------------------------------------------------
# Phase 4 -- headless reload (Excel proxy) + NO native charts
# --------------------------------------------------------------------------
def test_reload_headless(populated):
    wb = openpyxl.load_workbook(populated, keep_vba=True)
    # macro survived the openpyxl round-trip
    assert wb.vba_archive is not None
    # tab set EXACT (contract sec 2)
    assert tuple(wb.sheetnames) == TABS
    # NO native chart objects anywhere (the top "recovered content" trigger, L4)
    for ws in wb.worksheets:
        assert getattr(ws, "_charts", []) == []
    # a headline formula exists on every dashboard row, for EVERY transform
    cfg = R.parse_config(BW.config_rows())
    by_tab = {
        "Dashboard_Conditions": {"curve", "spreads", "conditions", "sloos"},
        "Dashboard_Labor": {"labor"},
        "Dashboard_Housing_Sentiment": {"housing", "sentiment", "confirm"},
    }
    for tab, cats in by_tab.items():
        ws = wb[tab]
        n_rows = sum(1 for s in cfg.series
                     if s.lane == "dashboard" and s.category in cats)
        formulas = [ws.cell(r, 6).value for r in range(8, 8 + n_rows)]
        assert all(isinstance(f, str) and f.startswith("=") for f in formulas), tab
    # the sahm_state_band defined name resolves into _config, and the cell it
    # points at is NUMERIC -- Excel's number>=text comparison is always FALSE,
    # so a text "0.5" would silently downgrade every state ALERT to WATCH.
    dn = wb.defined_names["sahm_state_band"]
    assert "_config" in dn.attr_text
    ref = dn.attr_text.split("!")[1].replace("$", "")
    band = wb["_config"][ref].value
    assert isinstance(band, (int, float)) and band == pytest.approx(0.5)
    wb.close()


# --------------------------------------------------------------------------
# Phase 5 -- watchlist admits states (POSITIVE) + refusal + defense in depth
# --------------------------------------------------------------------------
def test_watchlist_admits_states(populated):
    cfg = R.parse_config(BW.config_rows())
    admitted, refusals = R.evaluate_watchlist(cfg.series)
    # POSITIVE: every state row passes all three gates; nothing is refused
    assert len(admitted) == 151
    assert refusals == []
    assert {s.geo_segment.split(":")[1] for s in admitted} == set(SEED.STATES)

    wb = openpyxl.load_workbook(populated, keep_vba=True)
    ws = wb["Watchlist"]
    # 51 ranked state rows, one per state, starting under the header band
    states = [ws.cell(r, 1).value for r in range(5, 5 + 51)]
    assert states == sorted(SEED.STATES)
    for r in range(5, 5 + 51):
        st = ws.cell(r, 1).value
        assert str(ws.cell(r, 2).value).startswith("=")        # UR gap formula
        assert str(ws.cell(r, 5).value).startswith("=IF")      # Rank formula
        assert "sahm_state_band" in str(ws.cell(r, 6).value)   # Status via defined name
        if st == "DC":   # DC has no ICLAIMS/PHCI family -- cells stay blank
            assert ws.cell(r, 3).value is None
            assert ws.cell(r, 4).value is None
        else:
            assert str(ws.cell(r, 3).value).startswith("=")
            assert str(ws.cell(r, 4).value).startswith("=")
    blob = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "WATCHLIST REFUSED" not in blob             # no refusal for state rows
    # admitted rows were FETCHED: their raw blocks carry demo values
    raw = wb["Raw_FRED"]
    blocks = R.raw_layout(cfg.series, slots=cfg.raw_slots)
    for sid in ("CAUR", "TXICLAIMS", "NYPHCI", "DCUR"):
        b = blocks[sid]
        assert raw.cell(b.first_data_row, 2).value is not None, sid
    wb.close()


def test_watchlist_refusal():
    """A national row injected into the watchlist lane is refused,
    series-named, with the config row's fields interpolated."""
    cfg = R.parse_config(BW.config_rows())
    nat = next(s for s in cfg.series if s.id == "SAHMREALTIME")
    injected = replace(nat, lane="watchlist")
    with pytest.raises(R.WatchlistRefused) as ei:
        R.assert_no_national_in_watchlist([injected])
    msg = str(ei.value)
    assert 'series "SAHMREALTIME"' in msg
    assert 'geo_segment="national"' in msg
    assert "cannot localize a portfolio footprint" in msg
    assert "state-keyed labor/coincident series" in msg
    # the runtime evaluator refuses it too (never fetched, never landed)
    admitted, refusals = R.evaluate_watchlist([injected])
    assert admitted == [] and len(refusals) == 1
    assert "SAHMREALTIME" in refusals[0][1]


def test_watchlist_gate_defense_in_depth():
    cfg = R.parse_config(BW.config_rows())

    # (a) flip a national row's capability flag -> the geo gate (3) still
    #     refuses: a flipped flag cannot admit a national aggregate.
    nat = next(s for s in cfg.series if s.id == "ICSA")
    flipped = replace(nat, lane="watchlist", watchlist_capable=True)
    reasons = R.gate_watchlist_row(flipped)
    assert any("Gate3" in r for r in reasons)
    assert not any("Gate1" in r for r in reasons)      # flag passed; geo caught it
    with pytest.raises(R.WatchlistRefused):
        R.assert_no_national_in_watchlist([flipped])

    # (b) fabricate a Class C state row -> the class gate (2) refuses: this
    #     template admits ONLY the public FRED adapter; an unanticipated class
    #     must not slip through on a good geo key.
    ca = next(s for s in cfg.series if s.id == "CAUR")
    licensed = replace(ca, source_class="C")
    reasons = R.gate_watchlist_row(licensed)
    assert any("Gate2" in r for r in reasons)
    with pytest.raises(R.WatchlistRefused):
        R.assert_no_national_in_watchlist([licensed])

    # (c) a genuine state Class A row passes all three gates
    assert R.gate_watchlist_row(ca) == []

    # (d) non-state geo keys are refused default-deny (msa/county/us/blank)
    for geo in ("msa:12345", "county:06037", "us", "", "state:cal", "state:ca"):
        bad = replace(ca, geo_segment=geo)
        assert any("Gate3" in r for r in R.gate_watchlist_row(bad)), geo


# --------------------------------------------------------------------------
# Phase 6 -- staleness (the SLIND regression test)
# --------------------------------------------------------------------------
class FrozenProvider(R.FredDemoProvider):
    """DemoProvider variant that freezes ONE series two years before asof --
    the exact shape of the frozen {ST}SLIND family (returns data, no error)."""

    def __init__(self, frozen_id, **kw):
        super().__init__(**kw)
        self.frozen_id = frozen_id

    def fetch_series(self, spec, secret=None):
        if spec.id == self.frozen_id:
            old = R.FredDemoProvider(
                asof=date(self.asof.year - 2, self.asof.month, 28),
                slots=self.slots)
            return old.fetch_series(spec, secret)
        return super().fetch_series(spec, secret)


def test_stale_flag():
    cfg = R.parse_config(BW.config_rows())
    prov = FrozenProvider("NFCI", asof=ASOF, slots=cfg.raw_slots)
    admitted, _ = R.evaluate_watchlist(cfg.series)
    admitted_ids = {s.id for s in admitted}
    series_rows = {}
    for spec in cfg.series:
        if spec.lane == "watchlist" and spec.id not in admitted_ids:
            continue
        series_rows[spec.id] = prov.fetch_series(spec)
    digest = R.compute_digest(cfg, series_rows, ASOF, admitted_ids)

    frozen = next(d for d in digest if d["id"] == "NFCI")
    assert frozen["stale"] is True
    assert frozen["status"] == "STALE"                 # visible in the digest
    assert frozen["asof_period"] < "2025-01-01"        # fetched, but frozen
    # excluded from alert counts: NFCI trips ALERT when fresh (see email-sim);
    # frozen it must count as neither ALERT nor WATCH.
    assert all(d["status"] not in ("ALERT", "WATCH")
               for d in digest if d["stale"])
    fresh_digest = R.compute_digest(
        cfg, {**series_rows, "NFCI": R.FredDemoProvider(
            asof=ASOF, slots=cfg.raw_slots).fetch_series(
                next(s for s in cfg.series if s.id == "NFCI"))},
        ASOF, admitted_ids)
    fresh_nfci = next(d for d in fresh_digest if d["id"] == "NFCI")
    assert fresh_nfci["stale"] is False and fresh_nfci["status"] == "ALERT"
    n_alerts = sum(1 for d in digest if d["status"] == "ALERT")
    n_alerts_fresh = sum(1 for d in fresh_digest if d["status"] == "ALERT")
    assert n_alerts == n_alerts_fresh - 1              # the stale row dropped out

    # the primitive itself, across cadences (generous month math)
    spec_m = R.SeriesSpec(**_spec_kwargs())
    assert R.is_stale(spec_m, "2026-03-31", ASOF, 2.0) is False
    assert R.is_stale(spec_m, "2026-01-15", ASOF, 2.0) is True    # > 62 days
    assert R.is_stale(spec_m, None, ASOF, 2.0) is True            # valueless
    spec_w = R.SeriesSpec(**{**_spec_kwargs(), "frequency": "weekly"})
    assert R.is_stale(spec_w, "2026-03-25", ASOF, 2.0) is False
    assert R.is_stale(spec_w, "2026-03-01", ASOF, 2.0) is True    # > 14 days
    spec_q = R.SeriesSpec(**{**_spec_kwargs(), "frequency": "quarterly"})
    assert R.is_stale(spec_q, "2025-12-31", ASOF, 2.0) is False
    assert R.is_stale(spec_q, "2025-06-30", ASOF, 2.0) is True


# --------------------------------------------------------------------------
# Phase 7 -- idempotence + layout guard
# --------------------------------------------------------------------------
def test_raw_landing_idempotent(xlsm, tmp_path):
    def landed(path):
        wb = openpyxl.load_workbook(path, keep_vba=True)
        ws = wb["Raw_FRED"]
        vals = [(c.row, c.value) for col in ws.iter_cols(min_col=2, max_col=2)
                for c in col if c.value is not None]
        wb.close()
        return vals

    # TRUE idempotence: a second run on the SAME already-populated workbook
    # lands byte-identical values (slot-clearing leaves no stale tail).
    p1 = str(tmp_path / "r1.xlsm")
    shutil.copy(xlsm, p1)
    R.run(p1, demo=True, asof=ASOF)
    first = landed(p1)
    R.run(p1, demo=True, asof=ASOF)
    assert landed(p1) == first
    # ... and determinism across fresh copies.
    p2 = str(tmp_path / "r2.xlsm")
    shutil.copy(xlsm, p2)
    R.run(p2, demo=True, asof=ASOF)
    assert landed(p2) == first
    # newest-first: B4 (first data row of the first block) is populated
    wb = openpyxl.load_workbook(p1, keep_vba=True)
    assert wb["Raw_FRED"]["B4"].value is not None
    wb.close()


def test_raw_layout_mismatch_refused(xlsm, tmp_path):
    """Editing raw_slots after build must be REFUSED (dashboard and Watchlist
    formulas are anchored to the built layout), not silently mis-written."""
    p = str(tmp_path / "mismatch.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)                     # populate at slots=60
    wb = openpyxl.load_workbook(p, keep_vba=True)
    ws = wb["_config"]
    target = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == "raw_slots":
            target = row[0].row
            break
    assert target is not None
    ws.cell(target, 2, "30")
    wb.save(p)
    wb.close()
    with pytest.raises(RuntimeError, match="Raw layout mismatch"):
        R.run(p, demo=True, asof=ASOF)


# --------------------------------------------------------------------------
# Seam extras -- Class C stub (contract), VBA project keys, key fail-fast
# --------------------------------------------------------------------------
def test_class_c_stub():
    """The seam-contract stub. NOTE: unlike the bureau template, class C is
    REFUSED by this template's watchlist gate 2 -- the stub only proves the
    fetch_series seam shape."""
    spec = R.SeriesSpec(**{**_spec_kwargs(), "source_class": "C"})
    prov = R.ClassCStubProvider(secret_env="MACRO_CLASS_C_SECRET")
    with pytest.raises(PermissionError) as ei:
        prov.fetch_series(spec, secret=None)
    assert "401" in str(ei.value)
    rows = prov.fetch_series(spec, secret="dummy-client-credentials")
    assert isinstance(rows, list) and len(rows) == 1
    assert isinstance(rows[0], R.NormalizedRow)
    assert rows[0].value is None                 # the stub never fabricates data
    assert issubclass(R.ClassCStubProvider, R.Provider)
    # and even a well-formed Class C state row is refused by gate 2
    wl = R.SeriesSpec(**{**_spec_kwargs(), "source_class": "C",
                         "lane": "watchlist", "watchlist_capable": True,
                         "geo_segment": "state:CA"})
    assert any("Gate2" in r for r in R.gate_watchlist_row(wl))


def test_live_run_requires_key(monkeypatch, xlsm, tmp_path):
    """A live (non-demo) run without FRED_API_KEY and with a blank _config
    fred_api_key cell fails fast (exit path 3) before any fetch."""
    p = str(tmp_path / "live.xlsm")
    shutil.copy(xlsm, p)
    monkeypatch.delenv("FRED_API_KEY", raising=False)
    with pytest.raises(SystemExit, match="FRED API key missing"):
        R.run(p, demo=False, asof=ASOF)


def test_vba_protection_keys_roundtrip():
    """The PROJECT stream's CMG/DPB/GC must decrypt per [MS-OVBA] 2.4.3 to
    protection=none / password=none / visibility=visible -- exactly the way
    the VBE will read them (a malformed key triggers Excel's 'invalid key
    DPB' prompt)."""
    proj = V.build_project_stream(
        "X", [V.Module("X", "Sub A()\nEnd Sub")]).decode("latin-1")
    keys = {}
    for line in proj.splitlines():
        for k in ("CMG", "DPB", "GC"):
            if line.startswith(f'{k}="'):
                keys[k] = line.split('"')[1]
    assert V.ovba_decrypt(keys["CMG"]) == struct.pack("<I", 0)   # unprotected
    assert V.ovba_decrypt(keys["DPB"]) == b"\x00"               # no password
    assert V.ovba_decrypt(keys["GC"]) == b"\xff"                # visible
    # _VBA_PROJECT is the full 7-byte header ([MS-OVBA] 2.3.4.1)
    assert len(V.build_vba_project_stream()) == 7
    # the default VBA + Excel libraries are referenced in the dir stream
    d = V.build_dir_stream("X", [V.Module("X", "Sub A()\nEnd Sub")])
    assert b"Visual Basic For Applications" in d
    assert b"Microsoft Excel" in d


def test_clear_actually_blanks(xlsm, tmp_path):
    """openpyxl's ws.cell(r, c, None) silently IGNORES None -- clearing must
    assign .value explicitly, or a failed fetch would leave last run's stale
    values under a fresh timestamp (regression for the silent no-op)."""
    p = str(tmp_path / "clear.xlsm")
    shutil.copy(xlsm, p)
    R.run(p, demo=True, asof=ASOF)                      # populate
    backend = R.OpenpyxlBackend(p)
    cfg = backend.read_config()
    blocks = R.raw_layout(cfg.series, slots=cfg.raw_slots)
    spec = next(s for s in cfg.series if s.id in blocks and s.lane != "watchlist")
    b = blocks[spec.id]
    ws = backend._wb[R.RAW_TAB]
    assert ws.cell(b.first_data_row, 2).value is not None   # was populated
    backend.clear_raw_block(b, spec)
    for rr in range(b.first_data_row, b.first_data_row + b.slots):
        assert ws.cell(rr, 1).value is None
        assert ws.cell(rr, 2).value is None
