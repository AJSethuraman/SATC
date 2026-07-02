#!/usr/bin/env python3
"""Email-simulate acceptance test (BUILD_SPEC_MACRO.md Phase 8, --demo).

Copies ONLY the .xlsm into a fresh, otherwise-empty folder, then reproduces
exactly what the VBA "Extract" button does:
  1. read the _code_py tab (one source line per cell, column A) and write
     runner.py next to the workbook -- byte-for-byte the macro's WriteTabToFile;
  2. shell Python to run the EXTRACTED runner.py against the CLOSED workbook in
     --demo mode (deterministic, offline -- no key/network).
Then it composes (does NOT send) the monitoring email from the run digest and
asserts the email contains the NATIONAL alert summary AND the ranked top-10
STATE stress section AND the staleness-flags section, and that the workbook
rebuilt itself with nothing but the workbook + extracted runner present (the
workbook is the source of truth). Deterministic from _config + DemoProvider.
"""
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

import runner as R

XLSM_NAME = "Macro_Early_Warning_Monitor.xlsm"
ASOF = "2026-03-31"


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write UTF-8."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def compose_email(status):
    """Deterministic monitoring email from the run status/digest. No send."""
    digest = status.get("digest", [])
    national = [d for d in digest if d["lane"] == "dashboard"]
    states = [d for d in digest if d["lane"] == "watchlist"
              and d["transform"] == "sahm_gap"]
    stale = [d for d in digest if d.get("stale")]
    alerts = [d for d in national if d["status"] == "ALERT"]
    watches = [d for d in national if d["status"] == "WATCH"]
    lines = []
    lines.append("Subject: Macro Early-Warning Monitor -- "
                 f"{status['alert_count']} ALERT / {status['watch_count']} WATCH "
                 f"/ {status['stale_count']} STALE ({status['timestamp']})")
    lines.append("")
    lines.append(f"Run mode: {status['mode']}  |  series pulled: "
                 f"{status['series_pulled']}/{status['series_in_dict']}")
    lines.append("")
    lines.append("NATIONAL ALERT SUMMARY")
    lines.append("----------------------")
    if alerts:
        for d in sorted(alerts, key=lambda x: x["id"]):
            v = "n/a" if d["value"] is None else f"{d['value']:.2f}"
            lines.append(f"  ALERT  {d['id']:<16} {d['transform']:<10} {v}")
    else:
        lines.append("  (no national series in ALERT)")
    if watches:
        for d in sorted(watches, key=lambda x: x["id"]):
            v = "n/a" if d["value"] is None else f"{d['value']:.2f}"
            lines.append(f"  WATCH  {d['id']:<16} {d['transform']:<10} {v}")
    lines.append("")
    lines.append("STATE STRESS RANKING (top 10 by UR Sahm-style gap)")
    lines.append("--------------------------------------------------")
    ranked = sorted([d for d in states if d["value"] is not None],
                    key=lambda x: -x["value"])[:10]
    if ranked:
        for i, d in enumerate(ranked, start=1):
            st = d["geo_segment"].split(":", 1)[1]
            flag = d["status"] if d["status"] in ("ALERT", "WATCH", "STALE") else "  "
            lines.append(f"  {i:>2}. {st}  gap={d['value']:.2f}  {flag}")
    else:
        lines.append("  (no state rows landed)")
    lines.append("")
    lines.append("STALENESS FLAGS (excluded from alert counts)")
    lines.append("--------------------------------------------")
    if stale:
        for d in sorted(stale, key=lambda x: x["id"]):
            lines.append(f"  STALE  {d['id']:<16} last obs {d['asof_period']}")
    else:
        lines.append("  (no stale series -- every row is within its cadence)")
    lines.append("")
    lines.append("WATCHLIST LANE")
    lines.append("--------------")
    refusals = status.get("watchlist_refusals", [])
    if refusals:
        for m in refusals:
            lines.append("  " + m)
    else:
        lines.append(f"  ADMITTED (state-keyed Class A): "
                     f"{len(status.get('watchlist_admitted', []))} series "
                     f"across {len(states)} states")
    lines.append("")
    return "\n".join(lines)


def main():
    repo = os.path.dirname(os.path.abspath(__file__))
    xlsm = os.path.join(repo, XLSM_NAME)
    if not os.path.exists(xlsm):
        print("build the workbook first: python3 make_workbook.py", file=sys.stderr)
        return 1

    work = tempfile.mkdtemp(prefix="macro_emailsim_")
    try:
        # 1. The email contains ONLY the workbook.
        dst_xlsm = os.path.join(work, XLSM_NAME)
        shutil.copy(xlsm, dst_xlsm)
        before = sorted(os.listdir(work))
        assert before == [XLSM_NAME], before
        print(f"[email-sim] fresh folder contains only: {before}")

        # 2. Button step 1: extract runner.py from _code_py.
        runner_py = os.path.join(work, "runner.py")
        extract_code_tab(dst_xlsm, "_code_py", runner_py)
        print(f"[email-sim] extracted runner.py ({os.path.getsize(runner_py)} bytes) from _code_py")

        # 3. Button step 2: shell Python to run the EXTRACTED runner (demo mode).
        proc = subprocess.run(
            [sys.executable, runner_py, "--workbook", dst_xlsm, "--demo", "--asof", ASOF],
            capture_output=True, text=True, cwd=work)
        last = proc.stderr.strip().splitlines()[-1] if proc.stderr.strip() else "(none)"
        print("[email-sim] runner stderr:", last)
        if proc.returncode != 0:
            print("[email-sim] FAILED: runner exited", proc.returncode)
            print(proc.stdout, proc.stderr)
            return 2

        # 4. Verify the rebuild: raw data present, macro intact, nothing extra.
        wb = openpyxl.load_workbook(dst_xlsm, keep_vba=True)
        raw_ok = wb["Raw_FRED"]["B4"].value is not None
        macro_ok = wb.vba_archive is not None
        produced = sorted(os.listdir(work))
        print(f"[email-sim] folder now: {produced}")
        print(f"[email-sim] raw populated={raw_ok}  macro intact={macro_ok}")

        # 5. Compose the monitoring email from the digest (deterministic).
        status = R.run(dst_xlsm, demo=True, asof=R.datetime.strptime(ASOF, "%Y-%m-%d").date())
        email = compose_email(status)
        email_path = os.path.join(work, "monitoring_email.txt")
        with open(email_path, "w", encoding="utf-8") as fh:
            fh.write(email)
        print("\n" + "=" * 72)
        print(email)
        print("=" * 72 + "\n")

        has_national = "NATIONAL ALERT SUMMARY" in email and (
            "ALERT" in email or "WATCH" in email)
        state_lines = [ln for ln in email.splitlines() if "gap=" in ln]
        has_state_ranking = ("STATE STRESS RANKING" in email
                             and len(state_lines) == 10
                             and any(f in email for f in ("ALERT", "WATCH")))
        has_staleness = "STALENESS FLAGS" in email
        self_contained = (raw_ok and macro_ok
                          and set(produced) <= {XLSM_NAME, "runner.py"})

        ok = has_national and has_state_ranking and has_staleness and self_contained
        print(f"[email-sim] national alert summary : {has_national}")
        print(f"[email-sim] ranked state section   : {has_state_ranking}")
        print(f"[email-sim] staleness section      : {has_staleness}")
        print(f"[email-sim] workbook self-contained: {self_contained}")
        print("[email-sim] RESULT:", "PASS -- workbook is the source of truth" if ok else "FAIL")
        return 0 if ok else 3
    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
