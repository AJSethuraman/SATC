#!/usr/bin/env python3
"""Build the Macro Early-Warning Monitor workbook (FRED provider) -- KeyBank-styled.

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled .xlsm.
All visual style comes from keybank_style.py (tokens + helpers) -- never a
hardcoded hex here. Built to BUILD_SPEC_MACRO.md (Sections 2-6).

Structure (Section 4):
  Dashboard_Conditions        -- curve, credit spreads (display-only lane),
                                 NFCI/ANFCI/STLFSI4/KCFSI, SLOOS
  Dashboard_Labor             -- Sahm pair, claims family, AWHMAN
  Dashboard_Housing_Sentiment -- PERMIT/HOUST (3m-smoothed), UMCSENT, NEWORDER,
                                 RECPROUSM156N (confirmatory)
  Watchlist        -- the ADMITTED state lane: 51 states ranked by UR sahm_gap,
                      with claims yoy_ma4 + PHCI chg_3p context columns
  Raw_FRED         -- one raw tab, fixed-anchor blocks, runner fills the values
  _config          -- the Section 2 dictionary + [THRESHOLDS] (source of truth)
  _code_py/_code_vba -- runner.py + macro.bas as flat ASCII text
  _readme          -- provider notes, licensing flags, run steps

No native charts (L4). Embedded code is flat ASCII (L3).
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import runner as R
import series_seed as SEED

import keybank_style as KB
from keybank_style import (
    INK, KEY_RED, MONO_FONT, NOTE_FONT, SECT_FILL,
    brand_banner, kpi_tiles, header_row, watchlist_boundary,
    hide_gridlines, freeze_below,
)


def stress_heat(ws, cell_range):
    """Heat where HIGH = stress (spreads, stress indexes, UR gaps): low renders
    calm green, high renders red -- matching the threshold engine's
    direction="above". (KB.yoy_heat runs the OTHER way -- high=good -- which is
    inverted for these lanes; the bureau lesson carried.)"""
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="min", start_color=KB.HEAT_GOOD,
        mid_type="percentile", mid_value=50, mid_color=KB.HEAT_MID,
        end_type="max", end_color=KB.HEAT_BAD))

HERE = os.path.dirname(os.path.abspath(__file__))
INK_BOLD = Font(name="Arial", bold=True, color=INK)


def text_cell(ws, row, col, value):
    """Write literal text; openpyxl treats a leading '=' as a formula, so force
    it back to a string for source/doc lines."""
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# _config  (the knob panel: SETTINGS / THRESHOLDS / SERIES)
# ==========================================================================
# Per-id national thresholds (BUILD SPEC sec 2/3). State watchlist alerting
# uses the single sahm_state_band setting instead -- 151 threshold rows would
# be noise. direction "above" = flag when value >= bound.
THRESHOLDS = [
    # id, watch, alert, direction
    ("T10Y3M", 0.0, -0.5, "below"),          # inversion: watch at 0, alert at -0.5
    ("SAHMREALTIME", 0.30, 0.50, "above"),   # the canonical trigger
    ("ICSA", 10.0, 25.0, "above"),           # YoY% of the 4-week MA
    ("BAMLH0A0HYM2", 5.0, 8.0, "above"),     # static reference, NOT a computed pctile
    ("NFCI", 0.0, 0.5, "above"),
    ("STLFSI4", 0.0, 1.0, "above"),
    ("DRTSCILM", 20.0, 40.0, "above"),       # underwriting cycle
    ("RECPROUSM156N", 20.0, 50.0, "above"),  # confirmatory only
]


def config_rows():
    rows = [["Macro Early-Warning Monitor (FRED) -- CONFIG (the knob panel). "
             "Edit values here; no code change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["demo_mode", "FALSE", "TRUE = offline deterministic FredDemoProvider "
         "(no network/key), for trying the button and tests."],
        ["raw_slots", "60", "observations kept per series (newest-first). BUILT "
         "INTO the dashboard/Watchlist formulas -- changing it requires "
         "rebuilding the workbook (make_workbook.py); the runner refuses a "
         "mismatched layout. Weekly yoy_ma4 needs >= 56."],
        ["fred_api_key", "", "fallback cell for the FRED API key. The "
         "FRED_API_KEY env var is preferred; the key is never hardcoded in "
         "code and never echoed. Demo mode needs no key."],
        ["http_min_interval", "0.6", "seconds between live FRED requests "
         "(documented limit 120/min -> 0.6s is compliant)."],
        ["fred_max_retries", "4", "retries with backoff on HTTP 429 / "
         "transient errors."],
        ["stale_multiplier", "2.0", "a series is STALE when its last "
         "observation is older than this x its cadence (the SLIND lesson); "
         "stale rows are excluded from alert counts."],
        # NUMERIC cell (not text): the Watchlist Status formula compares
        # against it via the defined name, and Excel's number>=text is FALSE.
        ["sahm_state_band", 0.5, "state watchlist alert band: UR sahm_gap "
         ">= band = ALERT, >= 0.6 x band = WATCH. Referenced by defined name "
         "from the Watchlist formulas AND read by the runner's digest."],
        ["secret_env", "", "NAME of the env var holding a licensed (Class C) "
         "secret. Empty in v1 -- and note this template's watchlist admits "
         "ONLY Class A; a Class C swap requires its own review."],
    ]
    rows += [[], ["[THRESHOLDS]"], ["id", "watch", "alert", "direction"]]
    for tid, watch, alert, direction in THRESHOLDS:
        rows.append([tid, watch, alert, direction])
    rows += [[], ["[SERIES]"], SEED.HEADER]
    for r in SEED.all_series():
        rows.append([r[h] for h in SEED.HEADER])
    rows += [[], ["# To add a series: append a [SERIES] row. The watchlist "
                  "lane admits ONLY genuinely state-keyed Class A rows "
                  "(watchlist_capable=TRUE, source_class=A, geo_segment "
                  "matching state:XX); national aggregates are refused, "
                  "series-named. MSA/county promotion is a spec change."],
             ["# EXCLUDED (dead/frozen -- never re-add as fetchable rows): "
              "TEDRATE, CFSI, {ST}SLIND/USSLIND (frozen at Feb 2020: the "
              "staleness lesson), Conference Board LEI (not on FRED)."]]
    return rows


def write_config(wb):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows()
    thr_cells = {}                 # id -> (watch_ref, alert_ref, direction)
    set_cells = {}                 # settings key -> value-cell ref
    in_thr = in_set = False
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if row and row[0] == "[THRESHOLDS]":
            in_thr, in_set = True, False
            continue
        if row and row[0] == "[SETTINGS]":
            in_set, in_thr = True, False
            continue
        if row and isinstance(row[0], str) and row[0].startswith("["):
            in_thr = in_set = False
        if in_set and row and row[0] not in ("key",):
            set_cells[row[0]] = f"_config!$B${i}"
        if in_thr and row and row[0] not in ("id",):
            tid = row[0]
            thr_cells[tid] = (f"_config!$B${i}", f"_config!$C${i}",
                              str(row[3]).strip().lower())
    ws["A1"].font = INK_BOLD
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("[") and row[0].endswith("]"):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    # the single state alert band, addressable by name from Watchlist formulas
    wb.defined_names.add(DefinedName("sahm_state_band",
                                     attr_text=set_cells["sahm_state_band"]))
    return thr_cells, set_cells


# ==========================================================================
# Raw_FRED scaffold  (single tab, fixed anchors; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, specs, blocks):
    ws = wb.create_sheet(R.RAW_TAB)
    hide_gridlines(ws)
    ws["A1"] = (f"{R.RAW_TAB} -- raw FRED observations (or DemoProvider output in "
                f"--demo), newest-first. Written by runner.py; dashboards and the "
                f"Watchlist read these by formula. Do not edit by hand.")
    ws["A1"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    for s in specs:
        # UNLIKE the bureau template, watchlist rows get scaffolded too:
        # admitted state rows ARE fetched and land here (spec sec 2/4).
        b = blocks[s.id]
        ws.cell(b.header_row, 1, s.id).font = INK_BOLD
        ws.cell(b.header_row, 2, s.title)
        ws.cell(b.header_row, 3, f"freq={s.frequency}; transform={s.transform}; "
                                 f"class={s.source_class}; geo={s.geo_segment}")
        ws.cell(b.header_row, 3).font = NOTE_FONT
        ws.cell(b.label_row, 1, "period").font = Font(name="Calibri", bold=True, size=9)
        ws.cell(b.label_row, 2, "value").font = Font(name="Calibri", bold=True, size=9)
    return ws


# ==========================================================================
# Dashboards  (formula-driven; conditional-format heat; no native charts)
# ==========================================================================
def _ref(b, offset=0):
    return f"{R.RAW_TAB}!{get_column_letter(R.RAW_VALUE_COL)}{b.first_data_row + offset}"


def _rng(b, first_offset, last_offset):
    col = get_column_letter(R.RAW_VALUE_COL)
    return (f"{R.RAW_TAB}!{col}{b.first_data_row + first_offset}:"
            f"{col}{b.first_data_row + last_offset}")


DASH_COLS = ["Category", "Series ID", "Title", "Latest", "Prior", "Signal",
             "Trend (8)", "Status"]
COL = {name: get_column_letter(i + 1) for i, name in enumerate(DASH_COLS)}
NCOLS = len(DASH_COLS)
HEADLINE_COL = 6
STATUS_C = 8


def _blank(ref):
    """Empty raw cells are NOT Excel errors: IFERROR(ref,"") coerces them to 0,
    which fabricates values (0.00 'Prior') and pre-run 'OK' statuses. Guard
    every raw reference so a missing observation renders blank instead."""
    return f"IF({ref}=\"\",\"\",{ref})"


def _pct_change_formula(latest, ago):
    return (f"=IF(OR({latest}=\"\",{ago}=\"\"),\"\","
            f"IFERROR(({latest}-{ago})/{ago}*100,\"\"))")


def _ma_yoy_formula(b, window, ppy):
    """(AVERAGE(v0:v{w-1}) - AVERAGE(vP:vP+{w-1})) / AVERAGE(vP:...) * 100,
    blank-guarded on the window anchors (v0 and vP)."""
    now = _rng(b, 0, window - 1)
    ago = _rng(b, ppy, ppy + window - 1)
    v0, vp = _ref(b, 0), _ref(b, ppy)
    return (f"=IF(OR({v0}=\"\",{vp}=\"\"),\"\","
            f"IFERROR((AVERAGE({now})-AVERAGE({ago}))/AVERAGE({ago})*100,\"\"))")


def _headline_formula(spec, b):
    """Headline = the named transform's latest value, by formula off the raw
    block -- covering EVERY registered transform so the workbook and the
    runner's digest can never disagree on the same series (no silent level
    fallback; the bureau review lesson)."""
    latest = _ref(b, 0)
    t = spec.transform
    ppy = R.periods_per_year(spec.frequency)
    if t == "yoy_pct":
        return _pct_change_formula(latest, _ref(b, ppy))
    if t in ("qoq_pct", "mom_pct"):
        return _pct_change_formula(latest, _ref(b, 1))
    if t == "zscore_8q":
        win = _rng(b, 0, 7)
        return (f"=IF({latest}=\"\",\"\","
                f"IFERROR(({latest}-AVERAGE({win}))/STDEV({win}),\"\"))")
    if t == "index_to_pct":
        # base = the series' first (oldest) observation = the LAST non-empty
        # cell of the newest-first block; LOOKUP(2,1/(rng<>"")) finds it.
        rng = _rng(b, 0, b.slots - 1)
        return (f"=IF({latest}=\"\",\"\","
                f"IFERROR(({latest}/LOOKUP(2,1/({rng}<>\"\"),{rng})-1)*100,\"\"))")
    if t == "ma4":
        v3 = _ref(b, 3)
        return (f"=IF(OR({latest}=\"\",{v3}=\"\"),\"\","
                f"AVERAGE({_rng(b, 0, 3)}))")
    if t == "yoy_ma4":
        return _ma_yoy_formula(b, 4, ppy)
    if t == "yoy_ma3":
        return _ma_yoy_formula(b, 3, ppy)
    if t == "chg_3p":
        v3 = _ref(b, 3)
        return f"=IF(OR({latest}=\"\",{v3}=\"\"),\"\",{latest}-{v3})"
    if t == "sahm_gap":
        # EXACTLY the Python t_sahm_gap: mean(latest 3) - min(12 before)
        v14 = _ref(b, 14)
        return (f"=IF(OR({latest}=\"\",{v14}=\"\"),\"\","
                f"AVERAGE({_rng(b, 0, 2)})-MIN({_rng(b, 3, 14)}))")
    return f"={_blank(latest)}"          # level: pass-through, blank-guarded


def _status_formula(r, thr_cells, spec):
    """OK/WATCH/ALERT by formula, referencing the _config [THRESHOLDS] cells
    (config-driven, not baked into the dashboard). Empty if no threshold."""
    cells = thr_cells.get(spec.id)
    if not cells:
        return ""
    watch_ref, alert_ref, direction = cells
    hc = f"${COL['Signal']}{r}"
    op = ">=" if direction != "below" else "<="
    return (f"=IF(NOT(ISNUMBER({hc})),\"\","
            f"IF({hc}{op}{alert_ref},\"ALERT\","
            f"IF({hc}{op}{watch_ref},\"WATCH\",\"OK\")))")


def write_dashboard(wb, tab, specs, blocks, thr_cells, title, subtitle,
                    categories, heat=True):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    rows = sorted([s for s in specs if s.lane == "dashboard"
                   and s.dashboard_capable and s.category in categories],
                  key=lambda s: (s.category, s.id))

    hdr = 7
    first = hdr + 1
    r = first
    for s in rows:
        b = blocks[s.id]
        latest, prior = _ref(b, 0), _ref(b, 1)
        ws.cell(r, 1, s.category).font = KB.SECONDARY
        ws.cell(r, 2, s.id).font = MONO_FONT
        ws.cell(r, 3, s.title).font = KB.DATA_FONT
        ws.cell(r, 4, f"={_blank(latest)}")
        ws.cell(r, 5, f"={_blank(prior)}")
        ws.cell(r, HEADLINE_COL, _headline_formula(s, b))
        # col 7 (Trend) left blank -- sparklines painted by the macro.
        status = _status_formula(r, thr_cells, s)
        if status:
            ws.cell(r, STATUS_C, status)
        for c in (4, 5):
            ws.cell(r, c).number_format = KB.FMT_RATE
        ws.cell(r, HEADLINE_COL).number_format = KB.FMT_YOY
        for c in (4, 5, HEADLINE_COL):
            ws.cell(r, c).alignment = KB.RIGHT
        ws.cell(r, STATUS_C).alignment = KB.CENTER
        for c in range(1, NCOLS + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        r += 1
    last = r - 1

    header_row(ws, hdr, DASH_COLS, right_from=3, center_cols=(6, 7))
    freeze_below(ws, hdr)
    _dash_widths(ws)

    brand_banner(ws, 1, NCOLS, title, subtitle)
    sc = COL["Status"]
    if last >= first:
        tiles = [
            ("SERIES IN ALERT", f"=COUNTIF({sc}{first}:{sc}{last},\"ALERT\")",
             f"of {len(rows)}"),
            ("SERIES IN WATCH", f"=COUNTIF({sc}{first}:{sc}{last},\"WATCH\")",
             "elevated"),
            ("SERIES TRACKED", f"={len(rows)}", "this lane"),
        ]
        kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8

    # run-status readout in column L (free of the masthead merges)
    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 34
    for rr in (1, 2):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(3, R.STATUS_COL, "Leading indicators of credit deterioration").font = NOTE_FONT
    ws.cell(3, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(4, R.STATUS_COL).font = NOTE_FONT          # macro writes status here
    ws.cell(4, R.STATUS_COL).alignment = KB.LEFT

    if last >= first:
        KB.alert_rule(ws, f"{sc}{first}:{sc}{last}")
        if heat:
            # red = stress: hot spreads / rising claims / tightening shade red,
            # agreeing with the ALERT/WATCH direction of the thresholds here.
            stress_heat(ws, f"{COL['Signal']}{first}:{COL['Signal']}{last}")
    return ws


def _dash_widths(ws):
    widths = {"A": 12, "B": 18, "C": 40, "D": 11, "E": 11, "F": 12, "G": 14, "H": 12}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


# ==========================================================================
# Watchlist  (the ADMITTED state lane -- ranked footprint stress)
# ==========================================================================
BOUNDARY = ("GEOGRAPHIC RULE. This lane admits ONLY genuinely state-keyed "
            "Class A series (geo_segment matching state:XX) -- the default-"
            "deny validator refuses anything else, series-named. National/"
            "aggregate series cannot localize a portfolio footprint; join "
            "your portfolio's state exposure on the State column. MSA/county "
            "promotion is a spec change, not a config edit.")

WL_COLS = ["State", "UR gap", "Claims YoY 4wk-MA", "Coincident 3m chg",
           "Rank", "Status"]


def _wl_formulas(st, blocks, band_name="sahm_state_band"):
    """The three family formulas for one state, blank-guarded, matching the
    Python transforms exactly (spec sec 3). Missing family (DC claims/PHCI)
    -> None -> the cell stays blank and the guards elsewhere tolerate it."""
    out = {}
    ur = blocks.get(f"{st}UR")
    if ur:
        v0, v14 = _ref(ur, 0), _ref(ur, 14)
        out["ur_gap"] = (f"=IF(OR({v0}=\"\",{v14}=\"\"),\"\","
                         f"AVERAGE({_rng(ur, 0, 2)})-MIN({_rng(ur, 3, 14)}))")
    cl = blocks.get(f"{st}ICLAIMS")
    if cl:
        out["claims"] = _ma_yoy_formula(cl, 4, R.FREQ_PERIODS["weekly"])
    ph = blocks.get(f"{st}PHCI")
    if ph:
        v0, v3 = _ref(ph, 0), _ref(ph, 3)
        out["phci"] = f"=IF(OR({v0}=\"\",{v3}=\"\"),\"\",{v0}-{v3})"
    return out


def write_watchlist(wb, specs, blocks):
    # Build-time hard gate + the per-row refusal messages (runner sec 3).
    R.assert_no_national_in_watchlist(specs)
    admitted, refusals = R.evaluate_watchlist(specs)
    states = sorted({s.geo_segment.split(":", 1)[1] for s in admitted})

    ws = wb.create_sheet("Watchlist")
    hide_gridlines(ws)
    brand_banner(ws, 1, len(WL_COLS), "Watchlist -- State Footprint Stress",
                 "Macro stress at the state join key: UR Sahm-style gap ranks "
                 "the lane; claims and coincident-index columns corroborate. "
                 "Join portfolio state exposure here.")
    watchlist_boundary(ws, 3, len(WL_COLS), BOUNDARY)

    hdr = 4
    header_row(ws, hdr, WL_COLS, right_from=1, center_cols=(4, 5))
    first = hdr + 1
    r = first
    last = first + len(states) - 1
    for st in states:
        f = _wl_formulas(st, blocks)
        ws.cell(r, 1, st).font = MONO_FONT
        if "ur_gap" in f:
            ws.cell(r, 2, f["ur_gap"]).number_format = KB.FMT_Z
        if "claims" in f:            # DC has no ICLAIMS family -- cell stays blank
            ws.cell(r, 3, f["claims"]).number_format = KB.FMT_YOY
        if "phci" in f:              # DC has no PHCI family -- cell stays blank
            ws.cell(r, 4, f["phci"]).number_format = KB.FMT_Z
        # rank on the UR gap, highest stress = 1 (descending)
        ws.cell(r, 5, (f"=IF(NOT(ISNUMBER($B{r})),\"\","
                       f"RANK($B{r},$B${first}:$B${last}))"))
        # status off the sahm_state_band DEFINED NAME (-> _config), matching
        # runner.watchlist_status exactly: >= band ALERT, >= 0.6*band WATCH.
        ws.cell(r, 6, (f"=IF(NOT(ISNUMBER($B{r})),\"\","
                       f"IF($B{r}>=sahm_state_band,\"ALERT\","
                       f"IF($B{r}>=0.6*sahm_state_band,\"WATCH\",\"\")))"))
        for c in (2, 3, 4):
            ws.cell(r, c).alignment = KB.RIGHT
        for c in (5, 6):
            ws.cell(r, c).alignment = KB.CENTER
        for c in range(1, len(WL_COLS) + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        r += 1

    if states:
        KB.alert_rule(ws, f"F{first}:F{last}")
        stress_heat(ws, f"B{first}:B{last}")   # red = high UR gap = stress

    # refusals (none under the seeded config -- every state row is admitted;
    # rendered series-named if a bad row is ever injected and the build gate
    # is bypassed)
    for spec, message in refusals:
        ws.cell(r, 1, spec.id).font = MONO_FONT
        ws.cell(r, 2, "REFUSED").font = Font(name="Calibri", bold=True, color=KB.CRIMSON)
        r += 1
        ws.cell(r, 1, message).font = NOTE_FONT
        ws.cell(r, 1).alignment = KB.LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(WL_COLS))
        ws.row_dimensions[r].height = 56
        r += 1

    widths = {"A": 10, "B": 12, "C": 18, "D": 18, "E": 8, "F": 12}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    freeze_below(ws, hdr)
    return ws


# ==========================================================================
# Code-in-tab + readme
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
MACRO EARLY-WARNING MONITOR (FRED) -- README
============================================

WHAT THIS IS
  A self-contained, reusable template for CREDIT-RISK REVIEW. One workbook
  pulls macro/credit-cycle early-warning series from FRED, lands them raw, and
  presents them as formula-driven dashboards plus a state-keyed watchlist.
  Every lane is a leading indicator of credit deterioration: labor signals
  lead consumer delinquency; the curve and recession probabilities time the
  cycle for loss forecasting; spreads price wholesale credit; SLOOS tracks
  the underwriting cycle; the state watchlist turns macro stress into
  footprint-level early warning a portfolio's state exposure can join on.
  All code, config and docs live inside this workbook, so it can be emailed,
  opened elsewhere, and re-run.

PROVIDERS (the adapter seam -- one provider per class)
  Class A (public, keyed):
    - FredProvider     -- LIVE plain-urllib REST against api.stlouisfed.org.
                          Key from the FRED_API_KEY env var (preferred) or the
                          _config fred_api_key cell; never hardcoded/echoed.
                          Throttled 0.6s/request (120/min limit), backoff on 429.
    - FredDemoProvider -- deterministic OFFLINE stand-in (no network/key); used
                          by every test and the --demo button.
  Class C (licensed): NOT admitted anywhere in this template -- the watchlist
  gate admits only Class A, and a licensed swap requires its own review.

ONE-TIME SETUP
  1. Install Python 3 (add it to PATH) and the dependencies:
         pip install pandas openpyxl
  2. Enable macros when prompted (the Extract button needs them).
  3. For live pulls: get a free FRED API key (fred.stlouisfed.org) and set it:
         PowerShell:  $env:FRED_API_KEY = "yourkey"

RUN IT  (the workbook is the finished product; you only add data)
  1. In Excel press Alt+F8 -> run ExtractFiles. It writes three files next to the
     workbook: runner.py (from _code_py), requirements.txt, and RUN.txt (the
     exact PowerShell commands). NOTHING runs inside Excel.
  2. SAVE and CLOSE the workbook.
  3. Open RUN.txt and follow it from PowerShell: install deps, then run
     runner.py against the CLOSED workbook. It writes the data into the file.
       - Try-the-button (offline): runner.py -w <book>.xlsm --demo
       - Live FRED: $env:FRED_API_KEY = "..." then runner.py -w <book>.xlsm
  4. Reopen the workbook -- the formulas recalc into the populated dashboards.
  Optional: Alt+F8 -> PaintSparklines draws the Trend column.

THE TABS (each lane in credit-risk language)
  Dashboard_Conditions -- cycle position & wholesale credit pricing: the yield
      curve (inversion leads recessions), ICE BofA credit spreads (display-only
      lane -- see licensing), NFCI/ANFCI/STLFSI4/KCFSI stress gauges, and SLOOS
      C&I tightening (the underwriting cycle that precedes charge-offs).
  Dashboard_Labor -- labor leads consumer delinquency: the Sahm pair (the
      canonical recession trigger), the weekly claims family (the timeliest
      tripwire), and manufacturing hours (cut before heads).
  Dashboard_Housing_Sentiment -- demand, sentiment, and the confirmatory lane:
      permits/starts (3m-smoothed -- preliminary prints), UMich sentiment
      (1-month delay label), core capex orders, and the Chauvet-Piger recession
      probability (2-3 months behind: confirmation, never early warning).
  Watchlist -- ADMITTED state rows, one per state, ranked by the UR Sahm-style
      gap with claims-YoY and coincident-index corroboration. Join portfolio
      state exposure on the State column. Joining exposure weights stays a
      manual step in v1 -- the workbook shows footprint stress only.
  Raw_FRED -- raw observations, newest-first (the audit trail).
  _config  -- THE KNOB PANEL: series dictionary + thresholds (source of truth).
  _code_py / _code_vba -- the runner and macro as plain text.

THE WATCHLIST BOUNDARY (default-deny, defense in depth)
  A row only feeds the watchlist if ALL THREE hold: watchlist_capable=TRUE AND
  source_class="A" (this template's only admitted class -- anything else,
  including a future licensed feed, requires its own review) AND geo_segment
  matches the literal pattern state:XX. National/aggregate series cannot
  localize a portfolio footprint; the watchlist admits only state-keyed
  labor/coincident series. A build-time hard gate backs the runtime gate.

STALENESS IS A FIRST-CLASS FAILURE (the SLIND lesson)
  The Philly Fed state LEADING indexes ({ST}SLIND/USSLIND) still return data
  frozen at Feb 2020 -- fetch success is not data currency. The runner compares
  every series' last observation to its cadence (stale_multiplier x period
  length), marks laggards STALE in the digest/status line, and EXCLUDES them
  from alert counts. The frozen family is excluded from the dictionary; its
  leading signal is replicated with permits + claims + the national curve.

COMPLIANCE / LICENSING (binding -- from the coverage research)
  - ICE BofA OAS series (BAMLH0A0HYM2, BAMLC0A0CM): (c) Ice Data Indices --
    display-only with attribution; FRED serves only a ~3-year rolling window
    from Apr 2026, so long-run thresholds (HY OAS >= 8.0) are STATIC REFERENCES,
    never computed percentiles. Embedding cached ICE values in a shared
    workbook is a FLAG-FOR-LEGAL item (exact note text: UNKNOWN -- re-verify).
  - UMCSENT: source-mandated 1-month delay -- the tile is "as of prior month";
    (c) University of Michigan attribution. Never diff against same-month series.
  - FRED API terms (June 2024): no redistribution of third-party proprietary
    content, no AI/ML training use, API key mandatory. Fed/Treasury/BLS/Census
    series are unproblematic for an internal workbook.
  - EQFXSUBPRIME* (county subprime shares): licensing note UNVERIFIED -- kept
    OUT of v1; promotion candidate after legal review.

DATA-QUALITY TRAPS (baked into the transforms / notes)
  - Oct-2025 hole: the BLS household survey was never collected, so {ST}UR has
    a gap and {ST}PHCI skipped the month -- transforms tolerate interior NaN.
  - Annual benchmark revisions (Jan-Mar): LAUS re-benchmarks and all 50 PHCI
    models re-estimate; state histories revise every spring. Alerts are
    point-in-time, not replayable without ALFRED vintages.
  - NSA weekly noise: state claims alert only on the 4-week-MA YoY transform,
    never raw deltas (California holiday/backlog spikes).
  - Release calendars float (2026 BLS dates move between Friday and Tuesday);
    never weekday heuristics -- poll fred/release/dates if scheduling is added.
  - Series-ID churn (TEDRATE, CFSI, STLFSI->4): live fetches surface
    "(DISCONTINUED)" titles rather than serving stale data.

NO AI IS INVOLVED IN THE DATA PATH. Transforms, thresholds, the staleness
guard and the watchlist gate are deterministic and config-driven: the same
input always yields the same workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() and len(line) > 3 \
                and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path):
    rows = config_rows()
    cfg = R.parse_config(rows)
    specs = cfg.series
    R.assert_no_national_in_watchlist(specs)
    R.validate_transforms(specs)
    blocks = R.raw_layout(specs, slots=cfg.raw_slots)

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, specs, blocks)
    # _config must be written before dashboards so threshold cell refs and the
    # sahm_state_band defined name resolve.
    thr_cells, _set_cells = write_config(wb)

    write_dashboard(wb, "Dashboard_Conditions", specs, blocks, thr_cells,
                    "Conditions Dashboard",
                    "Cycle position & wholesale credit pricing: curve inversion, "
                    "ICE BofA spreads (display-only lane), stress gauges, SLOOS "
                    "underwriting cycle.",
                    categories={"curve", "spreads", "conditions", "sloos"})
    write_dashboard(wb, "Dashboard_Labor", specs, blocks, thr_cells,
                    "Labor Dashboard",
                    "Labor leads consumer delinquency: the Sahm pair, the weekly "
                    "claims family (timeliest tripwire), manufacturing hours.",
                    categories={"labor"})
    write_dashboard(wb, "Dashboard_Housing_Sentiment", specs, blocks, thr_cells,
                    "Housing & Sentiment Dashboard",
                    "Demand, sentiment, and the confirmatory lane: 3m-smoothed "
                    "permits/starts, UMich (1-mo delay), core capex orders, "
                    "recession probability (confirmation only).",
                    categories={"housing", "sentiment", "confirm"})
    write_watchlist(wb, specs, blocks)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    order = ["Dashboard_Conditions", "Dashboard_Labor",
             "Dashboard_Housing_Sentiment", "Watchlist", "Raw_FRED",
             "_config", "_code_py", "_code_vba", "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.active = wb.sheetnames.index("Dashboard_Conditions")
    wb.save(out_path)
    admitted, refusals = R.evaluate_watchlist(specs)
    return out_path, len(specs), len(refusals), len(admitted)


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "Macro_Early_Warning_Monitor_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n, nref, nadm = build(out)
    print(f"built {path}: {n} series, {nref} watchlist-refused, {nadm} admitted")
