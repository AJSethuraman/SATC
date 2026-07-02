#!/usr/bin/env python3
"""Macro Early-Warning Monitor -- runner (the data path).

Single source of truth for the data path; embedded verbatim into the workbook's
_code_py tab. The VBA ExtractFiles macro writes this back out as runner.py and
the user runs it from PowerShell against the CLOSED workbook (openpyxl backend).

Built to BUILD_SPEC_MACRO.md. Clean seams (BUILD SPEC 0.3 / 1a):

  * ADAPTER SEAM -- the only provider-specific code. Every adapter implements
    fetch_series(spec, secret) -> list[NormalizedRow] with a FIXED normalized
    schema; nothing downstream calls provider-specific code. v1 ships two Class A
    providers (FredDemoProvider offline + FredProvider live urllib REST) and an
    in-process Class C OAuth stub for the seam contract; note this template's
    watchlist gate admits ONLY class "A" (spec sec 3 gate 2), so a Class C swap
    requires its own review before admission.
  * TRANSFORM REGISTRY -- deterministic named transforms (BUILD SPEC sec 3).
  * THRESHOLD ENGINE  -- config-driven OK/WATCH/ALERT (BUILD SPEC sec 3).
  * WATCHLIST VALIDATOR -- DEFAULT-DENY WHITELIST (BUILD SPEC 0.1 / sec 3):
    only genuinely state-keyed (geo_segment="state:XX") Class A series may feed
    the watchlist lane; national aggregates are refused, series-named.
  * STALENESS GUARD (BUILD SPEC 0.6, the SLIND lesson) -- fetch success is not
    data currency: a series whose last observation is older than
    stale_multiplier x its cadence is marked STALE and excluded from alert KPIs.

No AI/LLM anywhere; transforms are pure (BUILD SPEC 0.5). Pure ASCII (L3).
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Sequence

import pandas as pd

# Fixed raw-block geometry (newest observation first) so dashboard formulas have
# stable anchors that never shift across refreshes.
RAW_TAB = "Raw_FRED"
RAW_SLOTS_DEFAULT = 60          # observation rows kept per series (newest-first)
RAW_HEADER_ROWS = 2
RAW_GAP_ROWS = 2
RAW_FIRST_ROW = 2
RAW_PERIOD_COL = 1             # column A
RAW_VALUE_COL = 2             # column B

# Status panel column on each dashboard (free of any merged masthead cells).
STATUS_COL = 12               # column L

FREQ_PERIODS = {"quarterly": 4, "annual": 1, "monthly": 12, "weekly": 52,
                "daily": 261}

# Staleness cadence in days (BUILD SPEC 0.6): generous month math -- a series
# is stale when (asof - last_obs) exceeds stale_multiplier x these.
CADENCE_DAYS = {"daily": 1, "weekly": 7, "monthly": 31, "quarterly": 92,
                "annual": 366}

# The watchlist join-key whitelist (BUILD SPEC sec 3, gate 3): default-deny.
# Only a literal two-letter state key passes; national/us/msa/county/anything
# else is refused (MSA/county promotion is a spec change, not a config edit).
WATCHLIST_GEO_PATTERN = r"^state:[A-Z]{2}$"

FRED_OBS_URL = "https://api.stlouisfed.org/fred/series/observations"


def _freq(frequency: str) -> str:
    f = (frequency or "").strip().lower()
    if f.startswith("q"):
        return "quarterly"
    if f.startswith("a") or f.startswith("y"):
        return "annual"
    if f.startswith("m"):
        return "monthly"
    if f.startswith("w"):
        return "weekly"
    if f.startswith("d"):
        return "daily"
    return "monthly"


def periods_per_year(frequency: str) -> int:
    return FREQ_PERIODS[_freq(frequency)]


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` tab; backend-agnostic)
# ---------------------------------------------------------------------------
SERIES_HEADER = [
    "id", "title", "category", "lane", "metric_type", "frequency", "sa_nsa",
    "units", "level_rate_index", "geo_segment", "source_class",
    "dashboard_capable", "watchlist_capable", "source_url", "table_id", "sheet",
    "series_label", "transform", "notes",
]


@dataclass
class SeriesSpec:
    id: str
    title: str
    category: str
    lane: str
    metric_type: str
    frequency: str
    sa_nsa: str
    units: str
    level_rate_index: str
    geo_segment: str
    source_class: str
    dashboard_capable: bool
    watchlist_capable: bool
    source_url: str
    table_id: str
    sheet: str
    series_label: str
    transform: str
    notes: str


@dataclass
class Threshold:
    watch: Optional[float]
    alert: Optional[float]
    direction: str          # "above" | "below"


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, Threshold] = field(default_factory=dict)
    series: List[SeriesSpec] = field(default_factory=list)

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))

    @property
    def sahm_state_band(self) -> float:
        """The single state-watchlist alert band (spec sec 3): a per-id
        threshold row for each of 151 state series would be noise. Referenced
        by defined name from the Watchlist formulas; documented default 0.50."""
        v = _as_float(self.settings.get("sahm_state_band", "0.5"))
        return 0.5 if v is None else v

    @property
    def stale_multiplier(self) -> float:
        v = _as_float(self.settings.get("stale_multiplier", "2.0"))
        return 2.0 if v is None else v


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _as_float(v):
    try:
        s = str(v).strip()
        return float(s) if s != "" else None
    except (TypeError, ValueError):
        return None


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists). Sections in col A:
    [SETTINGS], [THRESHOLDS], [SERIES]."""
    cfg = Config()
    section = None
    series_header = None
    thr_header = None
    for raw in rows:
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            series_header = None
            thr_header = None
            continue
        if not a:
            continue
        if a.startswith("#"):          # in-sheet comment line, never data
            continue
        if section == "SETTINGS":
            if a.lower() in ("key", "name"):
                continue
            val = "" if len(raw) < 2 or raw[1] is None else raw[1]
            cfg.settings[a] = str(val).strip()
        elif section == "THRESHOLDS":
            if thr_header is None and a.lower() == "id":
                thr_header = [str(c).strip().lower() for c in raw]
                continue
            cells = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                     for i, h in enumerate(thr_header or ["id", "watch", "alert", "direction"])}
            cfg.thresholds[a] = Threshold(
                watch=_as_float(cells.get("watch")),
                alert=_as_float(cells.get("alert")),
                direction=str(cells.get("direction", "above")).strip().lower() or "above")
        elif section == "SERIES":
            if series_header is None:
                series_header = [str(c).strip() for c in raw]
                continue
            v = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                 for i, h in enumerate(series_header)}
            cfg.series.append(SeriesSpec(
                id=str(v.get("id", "")).strip(),
                title=str(v.get("title", "")).strip(),
                category=str(v.get("category", "")).strip(),
                lane=str(v.get("lane", "")).strip().lower(),
                metric_type=str(v.get("metric_type", "")).strip(),
                frequency=str(v.get("frequency", "")).strip(),
                sa_nsa=str(v.get("sa_nsa", "")).strip(),
                units=str(v.get("units", "")).strip(),
                level_rate_index=str(v.get("level_rate_index", "")).strip().lower(),
                # geo_segment keeps its case: the gate-3 whitelist is the
                # LITERAL pattern ^state:[A-Z]{2}$ (spec sec 3).
                geo_segment=str(v.get("geo_segment", "")).strip(),
                source_class=str(v.get("source_class", "")).strip().upper(),
                dashboard_capable=_as_bool(v.get("dashboard_capable", "")),
                watchlist_capable=_as_bool(v.get("watchlist_capable", "")),
                source_url=str(v.get("source_url", "")).strip(),
                table_id=str(v.get("table_id", "")).strip(),
                sheet=str(v.get("sheet", "")).strip(),
                series_label=str(v.get("series_label", "")).strip(),
                transform=str(v.get("transform", "level")).strip().lower(),
                notes=str(v.get("notes", "")).strip()))
    return cfg


# ---------------------------------------------------------------------------
# THE WATCHLIST VALIDATOR -- DEFAULT-DENY WHITELIST (BUILD SPEC 0.1 / sec 3)
# ---------------------------------------------------------------------------
class WatchlistRefused(Exception):
    """Raised by the gate; carries the series-named refusal message."""


def watchlist_refusal_message(spec: SeriesSpec, reasons: List[str]) -> str:
    return (
        f'WATCHLIST REFUSED: series "{spec.id}" has geo_segment="{spec.geo_segment}", '
        f'watchlist_capable={"TRUE" if spec.watchlist_capable else "FALSE"}, '
        f'source_class="{spec.source_class}". ' + " ".join(reasons) +
        " National/aggregate series cannot localize a portfolio footprint; "
        "the watchlist admits only state-keyed labor/coincident series "
        "(geo_segment matching state:XX, Class A public FRED).")


def gate_watchlist_row(spec: SeriesSpec) -> List[str]:
    """Return the list of failed-gate reasons for a lane='watchlist' row.
    Empty list == passes all gates (a genuinely state-keyed Class A series)."""
    reasons = []
    if not spec.watchlist_capable:
        reasons.append("Gate1: watchlist_capable is not TRUE.")
    if spec.source_class != "A":
        reasons.append("Gate2: source_class is not A -- this template's only "
                       "admitted class is the public FRED adapter; any other "
                       "class (including a future Class C swap) requires its "
                       "own review before admission.")
    if not re.match(WATCHLIST_GEO_PATTERN, spec.geo_segment):
        reasons.append(f"Gate3: geo_segment '{spec.geo_segment}' does not match "
                       f"the join-key whitelist pattern {WATCHLIST_GEO_PATTERN} "
                       "(default-deny; MSA/county promotion is a spec change, "
                       "not a config edit).")
    return reasons


def evaluate_watchlist(series: Sequence[SeriesSpec]):
    """Apply the default-deny whitelist to every lane='watchlist' row.
    Returns (admitted, refusals) -- admitted rows pass ALL gates and ARE
    fetched/landed in Raw_FRED; refusals is a list of (spec, message) and
    those rows are never fetched."""
    admitted, refusals = [], []
    for s in series:
        if s.lane != "watchlist":
            continue
        reasons = gate_watchlist_row(s)
        if reasons:
            refusals.append((s, watchlist_refusal_message(s, reasons)))
        else:
            admitted.append(s)
    return admitted, refusals


def assert_no_national_in_watchlist(series: Sequence[SeriesSpec]) -> None:
    """Build-time hard gate backing gates 1-3 (BUILD SPEC 0.1 / sec 3): any
    lane='watchlist' row that fails a gate refuses the BUILD, naming the
    series -- a national row cannot even be built into the lane."""
    for s in series:
        if s.lane == "watchlist":
            reasons = gate_watchlist_row(s)
            if reasons:
                raise WatchlistRefused(watchlist_refusal_message(s, reasons))


# ---------------------------------------------------------------------------
# TRANSFORM REGISTRY (pure, deterministic -- BUILD SPEC sec 3)
# ---------------------------------------------------------------------------
def t_level(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64")


def _pct_change_shift(s: pd.Series, periods: int) -> pd.Series:
    """Shift-based percent change: NaN-propagating and identical across pandas
    versions (pct_change's fill_method default changed between 1.x/2.x/3.x)."""
    x = s.astype("float64")
    return (x / x.shift(periods) - 1.0) * 100.0


def t_yoy_pct(s: pd.Series, frequency: str) -> pd.Series:
    return _pct_change_shift(s, periods_per_year(frequency))


def t_qoq_pct(s: pd.Series, frequency: str) -> pd.Series:
    return _pct_change_shift(s, 1)


def t_mom_pct(s: pd.Series, frequency: str) -> pd.Series:
    return _pct_change_shift(s, 1)


def t_zscore_8q(s: pd.Series, frequency: str) -> pd.Series:
    x = s.astype("float64")
    mean = x.rolling(window=8, min_periods=8).mean()
    std = x.rolling(window=8, min_periods=8).std(ddof=1)
    return (x - mean) / std.replace(0.0, math.nan)


def t_index_to_pct(s: pd.Series, frequency: str) -> pd.Series:
    """Percent relative to the series' base = its first valid observation
    (BUILD SPEC sec 3: base-relative -- the bureau fix carried; NOT a YoY
    alias)."""
    x = s.astype("float64")
    valid = x.dropna()
    if valid.empty or valid.iloc[0] == 0.0:
        return x * math.nan
    return (x / valid.iloc[0] - 1.0) * 100.0


def _ma(s: pd.Series, window: int) -> pd.Series:
    return s.astype("float64").rolling(window=window, min_periods=window).mean()


def t_ma4(s: pd.Series, frequency: str) -> pd.Series:
    """4-period moving average (weekly claims smoothing)."""
    return _ma(s, 4)


def _yoy_of_ma(s: pd.Series, frequency: str, window: int) -> pd.Series:
    """Percent change of the window-period MA vs its value periods_per_year
    earlier (52 weekly / 12 monthly / 4 quarterly). NaN-propagating."""
    ma = _ma(s, window)
    return (ma / ma.shift(periods_per_year(frequency)) - 1.0) * 100.0


def t_yoy_ma4(s: pd.Series, frequency: str) -> pd.Series:
    return _yoy_of_ma(s, frequency, 4)


def t_yoy_ma3(s: pd.Series, frequency: str) -> pd.Series:
    return _yoy_of_ma(s, frequency, 3)


def t_chg_3p(s: pd.Series, frequency: str) -> pd.Series:
    """Arithmetic change over 3 periods (PHCI 3-month change)."""
    x = s.astype("float64")
    return x - x.shift(3)


def t_sahm_gap(s: pd.Series, frequency: str) -> pd.Series:
    """Sahm-STYLE state gap: mean(latest 3 observations) - min(the 12
    observations before those 3). Defined EXACTLY this way in BOTH Python and
    the Excel formula (AVERAGE(v0:v2) - MIN(v3:v14)) so workbook and digest
    can never disagree (spec sec 3; the canonical Sahm rule uses the min of
    the trailing 3m-MA -- this variant is chosen for exact Excel parity).
    At position i: mean(x[i-2:i+1]) - min(x[i-14:i-2]); NaN when history or
    any window member is missing."""
    x = s.astype("float64")
    m3 = x.rolling(window=3, min_periods=3).mean()
    min12 = x.shift(3).rolling(window=12, min_periods=12).min()
    return m3 - min12


TRANSFORMS = {
    "level": t_level, "yoy_pct": t_yoy_pct, "qoq_pct": t_qoq_pct,
    "mom_pct": t_mom_pct, "zscore_8q": t_zscore_8q, "index_to_pct": t_index_to_pct,
    "ma4": t_ma4, "yoy_ma4": t_yoy_ma4, "yoy_ma3": t_yoy_ma3,
    "chg_3p": t_chg_3p, "sahm_gap": t_sahm_gap,
}


class TransformError(Exception):
    pass


def validate_transforms(series: Sequence[SeriesSpec]) -> None:
    for s in series:
        if s.transform not in TRANSFORMS:
            raise TransformError(f"Series '{s.id}' references unknown transform "
                                 f"'{s.transform}'.")


def apply_transform(spec: SeriesSpec, s: pd.Series) -> pd.Series:
    return TRANSFORMS[spec.transform](s, spec.frequency)


def latest_valid(s: pd.Series):
    s = s.dropna()
    return None if s.empty else float(s.iloc[-1])


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven OK/WATCH/ALERT (BUILD SPEC sec 3)
# ---------------------------------------------------------------------------
def status_for(spec: SeriesSpec, value: Optional[float], thr: Optional[Threshold]) -> str:
    if value is None or thr is None or (isinstance(value, float) and math.isnan(value)):
        return "OK"
    above = thr.direction != "below"
    def hit(bound):
        if bound is None:
            return False
        return value >= bound if above else value <= bound
    if hit(thr.alert):
        return "ALERT"
    if hit(thr.watch):
        return "WATCH"
    return "OK"


def watchlist_status(spec: SeriesSpec, value: Optional[float], band: float) -> str:
    """State-watchlist alerting off the single sahm_state_band setting (spec
    sec 3): the UR sahm_gap trips ALERT at >= band and WATCH at >= 0.6 x band
    -- identical to the Watchlist tab's formula off the sahm_state_band
    defined name. Non-sahm_gap watchlist rows (claims/coincident) render as
    context and stay OK here."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "OK"
    if spec.transform != "sahm_gap":
        return "OK"
    if value >= band:
        return "ALERT"
    if value >= band * 0.6:
        return "WATCH"
    return "OK"


# ---------------------------------------------------------------------------
# STALENESS GUARD (BUILD SPEC 0.6 -- the SLIND lesson)
# ---------------------------------------------------------------------------
def is_stale(spec: SeriesSpec, last_period: Optional[str], asof: date,
             stale_multiplier: float) -> bool:
    """Fetch success is NOT data currency: the frozen Philly Fed SLIND family
    still returns data ending Feb 2020 without erroring. A series is STALE
    when its newest real observation is older than stale_multiplier x its
    cadence (generous per-period day counts in CADENCE_DAYS)."""
    if last_period is None:
        return True                     # fetched but valueless == frozen/empty
    try:
        last = date.fromisoformat(str(last_period)[:10])
    except ValueError:
        return True
    cadence = CADENCE_DAYS[_freq(spec.frequency)]
    return (asof - last).days > stale_multiplier * cadence


# ---------------------------------------------------------------------------
# THE ADAPTER SEAM (BUILD SPEC sec 1a) -- the only provider-specific code
# ---------------------------------------------------------------------------
@dataclass
class NormalizedRow:
    id: str
    period: str            # ISO date string (period end)
    value: Optional[float]
    geo_segment: str
    source_class: str
    units: str = ""


def last_observation_period(rows: List[NormalizedRow]) -> Optional[str]:
    """Newest period with a real value, derived from already-fetched data --
    no extra call (L5). Used by the digest (per-series as-of + staleness)."""
    ps = [r.period for r in rows if r.value is not None]
    return max(ps) if ps else None


class Provider:
    """Adapter interface contract: fetch_series(spec, secret) -> [NormalizedRow]."""

    source_class = "A"

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:  # pragma: no cover
        raise NotImplementedError

    def last_observation_period(self, rows: List[NormalizedRow]) -> Optional[str]:
        return last_observation_period(rows)


class FredDemoProvider(Provider):
    """Deterministic offline stand-in (BUILD SPEC 0.7 / Phase 2). Seeded
    pseudo-walk per id, fixed asof, NO network, NO key -- used by ALL tests.
    Never stale at its own asof: every series ends at the current period."""

    source_class = "A"

    def __init__(self, asof: Optional[date] = None, slots: int = RAW_SLOTS_DEFAULT):
        self.asof = asof or date(2026, 3, 31)
        self.slots = slots

    def _seed(self, sid: str) -> int:
        return sum((i + 1) * ord(c) for i, c in enumerate(sid)) % 997

    def _base(self, spec: SeriesSpec, seed: int) -> float:
        mt = spec.metric_type
        if mt == "unemp_rate":
            return 3.0 + (seed % 35) / 10.0             # ~3.0-6.5 %
        if mt == "sahm":
            return 0.05 + (seed % 30) / 100.0           # ~0.05-0.35 pp
        if mt == "claims":
            return 8.0 + (seed % 300)                   # thousands / count
        if mt == "coincident":
            return 100.0 + (seed % 300)                 # index level
        if mt == "curve_spread":
            return 0.2 + (seed % 15) / 10.0             # ~0.2-1.7 pp
        if mt == "oas":
            return 3.0 + (seed % 30) / 10.0             # ~3-6 %
        if mt == "fin_conditions":
            return 0.1 + (seed % 8) / 10.0              # index around 0
        if mt == "sloos":
            return 5.0 + (seed % 30)                    # net pct tightening
        if mt == "housing":
            return 1200.0 + (seed % 400)                # thousands SAAR
        if mt == "sentiment":
            return 60.0 + (seed % 40)                   # index
        if mt == "recession_prob":
            return 2.0 + (seed % 30)                    # pct
        if mt == "hours":
            return 40.0 + (seed % 20) / 10.0
        return 100.0 + (seed % 800)

    def _periods(self, spec: SeriesSpec, n: int):
        pcode = {"quarterly": "Q", "annual": "Y", "monthly": "M",
                 "weekly": "W", "daily": "D"}[_freq(spec.frequency)]
        idx = pd.period_range(end=pd.Period(self.asof, freq=pcode), periods=n,
                              freq=pcode)
        return idx.to_timestamp(how="end").normalize()

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        seed = self._seed(spec.id)
        n = self.slots
        idx = self._periods(spec, n)
        rows = []
        if spec.metric_type == "unemp_rate":
            # Additive path so sahm_gap produces a SPREAD of state values:
            # seed-dependent drift over the last 8 observations lifts
            # mean(latest 3) above the trailing 12-obs min by ~7 x drift, so a
            # handful of states trip the 0.5 band, more trip 0.3 (illustrative
            # demo -- the email-sim's ranked state section has color).
            base = self._base(spec, seed)
            # per-step recent drift: nonzero for ~1 in 4 seeds, sized so a
            # handful of states clear 0.5 and a few more clear 0.3
            d = (seed % 11 - 7) * 0.03 if seed % 11 >= 8 else 0.0
            for i in range(n):
                wob = math.sin((i + seed) / 4.0) * 0.15
                drift = max(0, i - (n - 9)) * d
                rows.append(NormalizedRow(
                    id=spec.id, period=idx[i].date().isoformat(),
                    value=round(base + wob + drift, 4),
                    geo_segment=spec.geo_segment,
                    source_class=spec.source_class, units=spec.units))
        else:
            v = self._base(spec, seed)
            for i in range(n):
                wobble = math.sin((i + seed) / 5.0) * 0.04
                drift = 0.015 if i > n - 8 else 0.0      # gentle recent rise -> trips some thresholds
                v = max(0.05, v * (1.0 + wobble + drift))
                rows.append(NormalizedRow(
                    id=spec.id, period=idx[i].date().isoformat(),
                    value=round(v, 4), geo_segment=spec.geo_segment,
                    source_class=spec.source_class, units=spec.units))
        # exercise the missing-value path deterministically -- at an interior
        # slot OUTSIDE every headline-formula window (newest offsets 0..14 for
        # sahm_gap, 52..55 for weekly yoy_ma4), so the Excel formulas (which
        # skip blanks) and the NaN-propagating Python transforms agree on the
        # headline value (the bureau parity lesson).
        k = min(n - 1, 20 + seed % 20)
        rows[k] = NormalizedRow(spec.id, rows[k].period, None,
                                spec.geo_segment, spec.source_class, spec.units)
        return rows


class FredProvider(Provider):
    """Live Class A provider: plain urllib REST against api.stlouisfed.org
    (BUILD SPEC 0.3 -- no fredapi dependency; requirements stay
    pandas+openpyxl). Throttled to min_interval seconds/request (documented
    limit 120/min -> 0.6s default) with retry/backoff on HTTP 429 and
    transient errors. Fetches only what lands: sort_order=desc & limit=slots
    (spec sec 1). The API key is NEVER hardcoded, logged or echoed."""

    source_class = "A"

    def __init__(self, api_key: str, min_interval: float = 0.6,
                 max_retries: int = 4, slots: int = RAW_SLOTS_DEFAULT):
        self._api_key = api_key
        self._min_interval = float(min_interval)
        self._max_retries = int(max_retries)
        self._slots = int(slots)
        self._last = 0.0
        self._cache: Dict[str, bytes] = {}

    def _throttle(self):
        gap = time.time() - self._last
        if gap < self._min_interval:
            time.sleep(self._min_interval - gap)
        self._last = time.time()

    def _http_get(self, url: str) -> bytes:
        """The one real network call -- isolated so tests can stub it."""
        import urllib.request
        with urllib.request.urlopen(url, timeout=30) as resp:
            return resp.read()

    def _download(self, url: str, label: str) -> bytes:
        """Cache + throttle + backoff. Error messages carry the series label,
        NEVER the url (it embeds the api key)."""
        if url in self._cache:                        # repeat fetch == cache hit (L5)
            return self._cache[url]
        import urllib.error
        for attempt in range(self._max_retries + 1):
            self._throttle()
            try:
                data = self._http_get(url)
                self._cache[url] = data
                return data
            except urllib.error.HTTPError as exc:
                # 429 (rate limit) and 5xx are retryable with backoff; other
                # HTTP codes fail fast (bad series id / bad key).
                if exc.code == 429 or exc.code >= 500:
                    if attempt < self._max_retries:
                        time.sleep(2.0 * (attempt + 1))
                        continue
                raise RuntimeError(f"FRED fetch failed for {label}: HTTP {exc.code}")
            except Exception as exc:
                if attempt < self._max_retries:
                    time.sleep(2.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"FRED fetch failed for {label}: {type(exc).__name__}")

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        if not self._api_key:
            raise ValueError("FRED API key missing: set the FRED_API_KEY env "
                             "var or the _config [SETTINGS] fred_api_key cell.")
        from urllib.parse import urlencode
        url = FRED_OBS_URL + "?" + urlencode({
            "series_id": spec.id, "api_key": self._api_key,
            "file_type": "json", "sort_order": "desc",
            "limit": str(self._slots)})
        payload = json.loads(self._download(url, spec.id).decode("utf-8"))
        obs = payload.get("observations", [])
        rows: List[NormalizedRow] = []
        for o in reversed(obs):                       # desc from API -> oldest-first
            raw = o.get("value")
            val = None if raw is None or str(raw).strip() == "." else _as_float(raw)
            rows.append(NormalizedRow(
                id=spec.id, period=str(o.get("date", "")).strip(), value=val,
                geo_segment=spec.geo_segment, source_class=spec.source_class,
                units=spec.units))
        return rows


class ClassCStubProvider(Provider):
    """In-process OAuth client_credentials STUB carried from the contract's
    seam rehearsal. Makes NO live request. NOTE: even a real Class C adapter
    is REFUSED by this template's watchlist gate 2 (only class "A" is
    admitted); a licensed swap requires its own review, not a config edit."""

    source_class = "C"

    def __init__(self, secret_env: str):
        self.secret_env = secret_env
        self._token = None

    def _authenticate(self, secret):
        # token from client_id + client_secret + scope; a request without a
        # valid token returns HTTP 401. Stubbed -- never hits a network.
        if not secret:
            raise PermissionError("HTTP 401: missing/invalid access token "
                                  "(client_credentials not supplied).")
        self._token = "stub-token"
        return self._token

    def fetch_series(self, spec: SeriesSpec, secret=None) -> List[NormalizedRow]:
        self._authenticate(secret)
        # A real adapter would call the licensed endpoint and normalize. The
        # stub returns one schema-shaped, VALUELESS row: it proves the seam
        # contract (Section 1a) without fabricating licensed data.
        return [NormalizedRow(id=spec.id, period="1900-01-01", value=None,
                              geo_segment=spec.geo_segment, source_class="C",
                              units=spec.units)]


def resolve_fred_key(cfg: Config) -> str:
    """Key resolution (spec sec 1): env FRED_API_KEY preferred, else the
    _config [SETTINGS] fred_api_key cell. Never hardcoded, never echoed."""
    return (os.environ.get("FRED_API_KEY", "").strip()
            or str(cfg.setting("fred_api_key", "") or "").strip())


def resolve_secret(cfg: Config) -> Optional[str]:
    """Class C secret from the env var whose NAME is recorded in _config; never
    hardcoded. Absent env var -> fail fast (handled by the caller). v2 lane."""
    name = str(cfg.setting("secret_env", "") or "").strip()
    if not name:
        return None
    return os.environ.get(name)


def make_provider(cfg: Config, demo: bool, asof: Optional[date]) -> Provider:
    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        return FredDemoProvider(asof=asof, slots=cfg.raw_slots)
    key = resolve_fred_key(cfg)
    if not key:
        raise SystemExit(
            "FRED API key missing for a live run: set the FRED_API_KEY env var "
            "(PowerShell: $env:FRED_API_KEY = \"...\") or put the key in the "
            "_config [SETTINGS] fred_api_key cell. Use --demo to run offline.")
    return FredProvider(
        api_key=key,
        min_interval=float(cfg.setting("http_min_interval", 0.6) or 0.6),
        max_retries=int(float(cfg.setting("fred_max_retries", 4) or 4)),
        slots=cfg.raw_slots)


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder
# ---------------------------------------------------------------------------
@dataclass
class RawBlock:
    id: str
    header_row: int
    label_row: int
    first_data_row: int
    slots: int


def _col(idx: int) -> str:
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def raw_layout(series: Sequence[SeriesSpec], slots: int = RAW_SLOTS_DEFAULT) -> Dict[str, RawBlock]:
    """Deterministic block placement in Raw_FRED: same input -> same anchors.
    Every series gets a block -- unlike the bureau template, ADMITTED
    watchlist rows ARE fetched here and the Watchlist tab formulas read them."""
    stride = RAW_HEADER_ROWS + slots + RAW_GAP_ROWS
    blocks = {}
    for i, s in enumerate(series):
        header_row = RAW_FIRST_ROW + i * stride
        blocks[s.id] = RawBlock(s.id, header_row, header_row + 1,
                                header_row + RAW_HEADER_ROWS, slots)
    return blocks


# ---------------------------------------------------------------------------
# WRITE BACKEND -- openpyxl on the closed workbook (contract sec 4 / L2)
# ---------------------------------------------------------------------------
class OpenpyxlBackend:
    def __init__(self, path: str):
        self.path = path
        import openpyxl
        # L2: keep_vba=True ONLY for .xlsm; on .xlsx it injects a dangling
        # vbaProject relationship Excel rejects as "format/extension not valid".
        keep_vba = path.lower().endswith(".xlsm")
        self._wb = openpyxl.load_workbook(path, keep_vba=keep_vba)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        return parse_config([[c.value for c in row] for row in ws.iter_rows()])

    def _check_block(self, block: RawBlock, spec: SeriesSpec):
        """Refuse to write into a workbook whose raw layout doesn't match the
        computed one (e.g. raw_slots edited after build): every dashboard and
        Watchlist formula is anchored to the BUILT layout, so a silent
        mismatch would make the whole workbook quietly wrong. Rebuild instead."""
        ws = self._wb[RAW_TAB]
        existing = ws.cell(block.header_row, 1).value
        if existing is not None and str(existing).strip() not in ("", spec.id):
            raise RuntimeError(
                f"Raw layout mismatch in {RAW_TAB} at row {block.header_row}: "
                f"expected block '{spec.id}' but found '{existing}'. The workbook "
                f"was built with a different raw_slots/series layout than _config "
                f"now describes; dashboard formulas are anchored to the built "
                f"layout. Rebuild the workbook (make_workbook.py) instead of "
                f"editing raw_slots in place.")

    def clear_raw_block(self, block: RawBlock, spec: SeriesSpec):
        """Blank a block's data slots (stateless rebuild: a series that fails
        to fetch must show empty, never last run's stale values).
        NOTE: openpyxl's ws.cell(r, c, None) silently IGNORES None -- cells
        must be blanked by assigning .value explicitly."""
        self._check_block(block, spec)
        ws = self._wb[RAW_TAB]
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            ws.cell(r, 1).value = None
            ws.cell(r, 2).value = None

    def write_raw_block(self, block: RawBlock, spec: SeriesSpec, rows: List[NormalizedRow]):
        self._check_block(block, spec)
        ws = self._wb[RAW_TAB]
        ws.cell(block.header_row, 1, spec.id)
        ws.cell(block.header_row, 2, spec.title)
        ws.cell(block.header_row, 3, f"freq={spec.frequency}; transform={spec.transform}; "
                                     f"class={spec.source_class}; geo={spec.geo_segment}")
        ws.cell(block.label_row, 1, "period")
        ws.cell(block.label_row, 2, "value")
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            ws.cell(r, 1).value = None
            ws.cell(r, 2).value = None
        # newest-first
        tail = rows[-block.slots:] if len(rows) > block.slots else rows
        for i, nr in enumerate(reversed(tail)):
            rr = block.first_data_row + i
            ws.cell(rr, 1, nr.period)
            ws.cell(rr, 2).value = None if nr.value is None else float(nr.value)

    def write_status(self, status: dict):
        line2 = (f"Series {status.get('series_pulled', 0)}/"
                 f"{status.get('series_in_dict', 0)} - "
                 f"{status.get('alert_count', 0)} ALERT / "
                 f"{status.get('watch_count', 0)} WATCH - "
                 f"{status.get('stale_count', 0)} STALE")
        n_err = len(status.get("errors") or [])
        if n_err:
            line2 += f" - {n_err} FETCH ERRORS (blocks blanked)"
        for tab in ("Dashboard_Conditions", "Dashboard_Labor",
                    "Dashboard_Housing_Sentiment"):
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                ws.cell(1, STATUS_COL, "Last run  " + status.get("timestamp", ""))
                ws.cell(2, STATUS_COL, line2)

    def finalize(self):
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def compute_digest(cfg: Config, series_rows: Dict[str, List[NormalizedRow]],
                   asof: date, admitted_ids: Sequence[str] = ()) -> List[dict]:
    """Per-series headline transform value + status (the engine output consumed
    by dashboards + email-sim). Covers dashboard rows AND admitted watchlist
    rows. Staleness (BUILD SPEC 0.6) is first-class: a stale row's status is
    forced to STALE and it never counts toward alert/watch KPIs."""
    admitted_ids = set(admitted_ids)
    digest = []
    for spec in cfg.series:
        if spec.id not in series_rows:
            continue
        if spec.lane == "watchlist":
            if spec.id not in admitted_ids:
                continue                # refused rows never reach the digest
        elif not spec.dashboard_capable:
            continue
        vals = [r.value for r in series_rows[spec.id]]
        s = pd.Series(vals, dtype="float64")
        headline = latest_valid(apply_transform(spec, s))
        last_p = last_observation_period(series_rows[spec.id])
        stale = is_stale(spec, last_p, asof, cfg.stale_multiplier)
        if stale:
            status = "STALE"            # a frozen series never satisfies a cell
        elif spec.lane == "watchlist":
            status = watchlist_status(spec, headline, cfg.sahm_state_band)
        else:
            status = status_for(spec, headline, cfg.thresholds.get(spec.id))
        digest.append({"id": spec.id, "title": spec.title, "lane": spec.lane,
                       "geo_segment": spec.geo_segment,
                       "metric": spec.metric_type, "transform": spec.transform,
                       "value": headline, "asof_period": last_p,
                       "stale": stale, "status": status})
    return digest


def run(workbook_path: str, demo: bool = False, asof: Optional[date] = None) -> dict:
    asof = asof or date.today()
    backend = OpenpyxlBackend(workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any fetch (BUILD SPEC 0.1, sec 3).
    validate_transforms(cfg.series)
    assert_no_national_in_watchlist(cfg.series)
    admitted, refusals = evaluate_watchlist(cfg.series)
    admitted_ids = {s.id for s in admitted}

    provider = make_provider(cfg, demo, asof)   # live mode fails fast w/o key
    mode = "demo" if isinstance(provider, FredDemoProvider) else "live"

    blocks = raw_layout(cfg.series, slots=cfg.raw_slots)
    series_rows: Dict[str, List[NormalizedRow]] = {}
    pulled, errors = 0, []
    # Stateless rebuild (BUILD SPEC 0.2): blank EVERY block first -- including
    # the watchlist-lane blocks, which in THIS template are fetched (admitted
    # state rows land in Raw_FRED so the Watchlist tab formulas read them) --
    # so a series whose fetch fails shows empty under the fresh timestamp,
    # never last run's stale values masquerading as current. This also
    # validates the workbook's raw layout against _config before any fetch.
    for spec in cfg.series:
        backend.clear_raw_block(blocks[spec.id], spec)
    for spec in cfg.series:
        if spec.lane == "watchlist" and spec.id not in admitted_ids:
            continue                    # refused rows are NEVER fetched
        try:
            rows = provider.fetch_series(spec)
        except Exception as exc:
            errors.append(f"{spec.id}: {exc}")
            continue
        series_rows[spec.id] = rows
        backend.write_raw_block(blocks[spec.id], spec, rows)
        pulled += 1

    digest = compute_digest(cfg, series_rows, asof, admitted_ids)
    status = {
        "timestamp": asof.isoformat(),
        "mode": mode,
        "series_in_dict": len(cfg.series),
        "series_pulled": pulled,
        "alert_count": sum(1 for d in digest if d["status"] == "ALERT"),
        "watch_count": sum(1 for d in digest if d["status"] == "WATCH"),
        "stale_count": sum(1 for d in digest if d["stale"]),
        "digest": digest,
        "watchlist_refusals": [m for _, m in refusals],
        "watchlist_admitted": sorted(admitted_ids),
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Macro Early-Warning Monitor runner")
    ap.add_argument("--workbook", "-w", required=True)
    ap.add_argument("--demo", action="store_true",
                    help="deterministic offline FredDemoProvider (no network/key)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD (testing)")
    args = ap.parse_args(argv)
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, demo=args.demo, asof=asof)
    except (WatchlistRefused, TransformError) as exc:
        sys.stderr.write(f"GATE ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except SystemExit as exc:
        sys.stderr.write(f"{exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True, **{k: v for k, v in status.items() if k != "digest"}}))
    sys.stderr.write(
        f"OK ({status['mode']}): {status['series_pulled']}/{status['series_in_dict']} series, "
        f"{status['alert_count']} ALERT, {status['watch_count']} WATCH, "
        f"{status['stale_count']} STALE, "
        f"watchlist admitted={len(status['watchlist_admitted'])} "
        f"refused={len(status['watchlist_refusals'])}.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
