import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from queue import Queue
import threading
import time

from openpyxl import load_workbook

from create_demo_data import make_demo_tree
from scanner import ScanConfig, run_scan


def _collect_events(q: Queue) -> list[dict]:
    out = []
    while not q.empty():
        out.append(q.get_nowait())
    return out


def _summary_metrics(path: Path) -> dict:
    summary = load_workbook(path)["Summary"]
    return {r[0].value: r[1].value for r in summary.iter_rows(min_row=2, max_col=2) if r[0].value}


def test_end_to_end(tmp_path: Path) -> None:
    root = make_demo_tree(tmp_path / "demo")
    out = tmp_path / "report.xlsx"
    q = Queue()
    run_scan(ScanConfig(root_path=root, output_file=out, filter_mode="all", extension_text="", detect_duplicates=True), q)
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Review", "Duplicate Groups", "Summary", "Errors"]
    rows = list(wb["Review"].iter_rows(min_row=2, values_only=True))
    by_name = [r for r in rows if r[6] == "samefile.txt"]
    assert len(by_name) == 2
    assert all((r[2] or "") == "" for r in by_name)
    assert _summary_metrics(out)["Scan Status"] == "Completed"


def test_extension_filters(tmp_path: Path) -> None:
    root = make_demo_tree(tmp_path / "demo")
    out_inc = tmp_path / "inc.xlsx"
    run_scan(ScanConfig(root_path=root, output_file=out_inc, filter_mode="include", extension_text="txt", detect_duplicates=False), Queue())
    exts = {r[7].value for r in load_workbook(out_inc)["Review"].iter_rows(min_row=2)}
    assert exts == {".txt"}

    out_exc = tmp_path / "exc.xlsx"
    run_scan(ScanConfig(root_path=root, output_file=out_exc, filter_mode="exclude", extension_text="txt", detect_duplicates=False), Queue())
    exts2 = {r[7].value for r in load_workbook(out_exc)["Review"].iter_rows(min_row=2)}
    assert ".txt" not in exts2


def test_empty_scan_exports_workbook_and_progress(tmp_path: Path) -> None:
    root = tmp_path / "empty"
    root.mkdir()
    out = tmp_path / "empty.xlsx"
    q = Queue()
    run_scan(ScanConfig(root_path=root, output_file=out, filter_mode="all", extension_text="", detect_duplicates=True), q)
    assert out.exists()
    events = _collect_events(q)
    assert any(e.get("type") == "progress" and e.get("phase") == "Metadata Scan" for e in events)


def test_cancel_before_start(tmp_path: Path) -> None:
    root = make_demo_tree(tmp_path / "demo")
    out = tmp_path / "cancel.xlsx"
    cancel = threading.Event(); cancel.set()
    run_scan(ScanConfig(root_path=root, output_file=out, filter_mode="all", extension_text="", detect_duplicates=True), Queue(), cancel)
    assert out.exists()
    metrics = _summary_metrics(out)
    assert metrics["Scan Status"] == "Cancelled"
    assert metrics.get("Cancelled At")


def test_cancel_during_scanning(tmp_path: Path) -> None:
    root = tmp_path / "many"
    for i in range(5000):
        d = root / f"a{i//100}"
        d.mkdir(parents=True, exist_ok=True)
        (d / f"f{i}.txt").write_text("x")
    out = tmp_path / "cancel_scan.xlsx"
    cancel = threading.Event()
    t = threading.Timer(0.01, cancel.set)
    t.start()
    run_scan(ScanConfig(root_path=root, output_file=out, filter_mode="all", extension_text="", detect_duplicates=False), Queue(), cancel)
    t.cancel()
    assert out.exists()
    assert _summary_metrics(out)["Scan Status"] == "Cancelled"


def test_cancel_during_hashing(tmp_path: Path, monkeypatch) -> None:
    root = tmp_path / "hashcancel"
    (root / "a").mkdir(parents=True)
    (root / "b").mkdir(parents=True)
    big = b"z" * (2 * 1024 * 1024)
    (root / "a" / "dup.bin").write_bytes(big)
    (root / "b" / "dup2.bin").write_bytes(big)
    out = tmp_path / "cancel_hash.xlsx"

    import scanner
    orig = scanner._hash_file

    def slow_hash(path, cancel_event, chunk_cb=None):
        time.sleep(0.03)
        return orig(path, cancel_event, chunk_cb)

    monkeypatch.setattr(scanner, "_hash_file", slow_hash)
    cancel = threading.Event()
    t = threading.Timer(0.02, cancel.set)
    t.start()
    q = Queue()
    run_scan(ScanConfig(root_path=root, output_file=out, filter_mode="all", extension_text="", detect_duplicates=True), q, cancel)
    t.cancel()
    assert out.exists()
    metrics = _summary_metrics(out)
    assert metrics["Scan Status"] == "Cancelled"
    assert metrics.get("Cancelled At")
    events = _collect_events(q)
    assert any(e.get("type") == "progress" and e.get("phase") == "Hashing" for e in events)
