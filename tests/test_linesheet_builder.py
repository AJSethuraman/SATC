import ast
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import load_workbook

from linesheet_builder.db import init_db, get_connection, create_or_get_client, create_engagement
from linesheet_builder.sample_data import create_demo_loan_tape
from linesheet_builder.import_engine import load_loan_tape, save_raw_import, create_import_batch
from linesheet_builder.mapping_engine import load_mapping_profile, apply_mapping, persist_loan_records, save_mapping_profile
from linesheet_builder.validation_engine import validate_loan_records, persist_validation_issues
from linesheet_builder.template_engine import load_template_yaml
from linesheet_builder.review_engine import create_review_cases, save_answer, calculate_completion_status, set_review_status
from linesheet_builder.export_engine import generate_excel_linesheet, generate_data_mart_csv, generate_audit_log_csv, generate_data_mart_workbook
from linesheet_builder.audit import append_audit_event
from linesheet_builder.rules_engine import evaluate_rule, UnsafeRuleError, determine_question_status
from linesheet_builder.dti_engine import load_dti_config, compute_dti, save_dti_inputs, load_dti_inputs
from linesheet_builder.cash_flow_engine import (load_cash_flow_config, compute_cash_flow, normalize_line,
    save_cash_flow_inputs, load_cash_flow_inputs, summarize_cash_flow)

ROOT=Path(__file__).resolve().parents[1]
TEMPLATE_PATH=ROOT/"configs"/"templates"/"commercial_linesheet_v1.yaml"
MAPPING_PATH=ROOT/"configs"/"mappings"/"demo_bank_mapping.yaml"

@pytest.fixture()
def workflow(tmp_path):
    db=tmp_path/"app.db"; init_db(db); conn=get_connection(db)
    cid=create_or_get_client(conn,"Demo Bank"); eid=create_engagement(conn,cid,"Q4 2025","Commercial Loan Review","commercial_linesheet_v1","Reviewer","QC")
    tape=create_demo_loan_tape(tmp_path/"demo.xlsx"); df=load_loan_tape(tape)
    raw=save_raw_import(tape, tmp_path/"raw_imports")
    batch=create_import_batch(conn,eid,"demo.xlsx",raw,df,"Reviewer")
    profile=load_mapping_profile(MAPPING_PATH); save_mapping_profile(profile, tmp_path/"mapping.yaml", conn, cid, "commercial_linesheet_v1", "Reviewer"); mapped=apply_mapping(df, profile)
    ids=persist_loan_records(conn,eid,batch,mapped,"Reviewer")
    template=load_template_yaml(TEMPLATE_PATH)
    rows=pd.read_sql_query("SELECT * FROM loan_records WHERE engagement_id=?", conn, params=(eid,)).to_dict("records")
    results=validate_loan_records(rows, template); persist_validation_issues(conn,results,eid,"Reviewer")
    cases=create_review_cases(conn,eid,"Reviewer","QC")
    return {"conn":conn,"eid":eid,"df":df,"raw":raw,"mapped":mapped,"template":template,"results":results,"cases":cases,"tmp":tmp_path}

def test_demo_loan_tape_loads_successfully(tmp_path):
    path=create_demo_loan_tape(tmp_path/"demo.xlsx")
    df=load_loan_tape(path)
    assert len(df) >= 10
    assert {"Loan Number","Borrower","DSCR","LTV"}.issubset(df.columns)

def test_raw_imported_data_is_preserved(tmp_path):
    path=create_demo_loan_tape(tmp_path/"demo.xlsx")
    raw=save_raw_import(path, tmp_path/"raw")
    assert Path(raw).exists()
    assert Path(raw).read_bytes() == Path(path).read_bytes()

def test_mapping_profile_maps_columns_to_standard_schema(workflow):
    mapped=workflow["mapped"]
    assert "loan_id" in mapped.columns and "borrower_name" in mapped.columns
    assert mapped.loc[0,"loan_id"] == "L1001"
    assert "raw_payload_json" in mapped.columns and "Loan Number" in mapped.loc[0,"raw_payload_json"]

def test_validation_blocks_duplicate_missing_borrower_and_date_relationship(workflow):
    issues=[i for r in workflow["results"] for i in r["issues"]]
    codes={i.issue_code for i in issues}
    assert "DUPLICATE_LOAN_ID" in codes
    assert "MISSING_BORROWER_NAME" in codes
    assert "MATURITY_BEFORE_ORIGINATION" in codes
    blocked=[r for r in workflow["results"] if r["status"]=="Blocked"]
    assert len(blocked) >= 2

def test_dscr_below_120_creates_validation_warning_and_template_finding(workflow):
    issues=[i for r in workflow["results"] for i in r["issues"]]
    assert any(i.issue_code=="DSCR_BELOW_120" and i.severity=="Warning" for i in issues)
    q=[q for s in workflow["template"].sections for q in s.questions if q.question_id=="FA2"][0]
    status=determine_question_status(q, answer=1.10, source={"dscr":1.10})
    assert status["exception_flag"] and status["severity"] == "Finding"

def test_template_loader_rejects_duplicate_question_ids(tmp_path):
    text=TEMPLATE_PATH.read_text().replace("question_id: SO2", "question_id: SO1")
    p=tmp_path/"bad.yaml"; p.write_text(text)
    with pytest.raises(ValueError, match="Duplicate question_id"):
        load_template_yaml(p)

def test_required_unanswered_questions_block_final_export(workflow):
    conn=workflow["conn"]; rcid=workflow["cases"][0]
    row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(rcid,)).fetchone()
    loan={k:row[k] for k in row.keys()}
    status=calculate_completion_status(conn,rcid,loan,workflow["template"])
    assert not status["export_ready"]
    assert any("Required unanswered" in b for b in status["blockers"])

def test_finding_questions_require_comments_and_evidence(workflow):
    conn=workflow["conn"]; template=workflow["template"]
    rcid=next(c for c in workflow["cases"] if conn.execute("SELECT validation_status FROM loan_records lr JOIN review_cases rc ON lr.loan_record_id=rc.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone()[0] != "Blocked")
    row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(rcid,)).fetchone(); loan={k:row[k] for k in row.keys()}
    sec=next(s for s in template.sections if s.section_id=="loan_terms"); q=next(q for q in sec.questions if q.question_id=="LT1")
    save_answer(conn,rcid,loan,sec,q,"No","","Needed","Reviewer",template.template_id,template.version)
    status=calculate_completion_status(conn,rcid,loan,template)
    assert any("Reviewer comment required: LT1" in b for b in status["blockers"])
    assert any("Evidence unresolved: LT1" in b for b in status["blockers"])

def complete_case(conn, rcid, loan, template):
    for sec in template.sections:
        for q in sec.questions:
            if q.applies_if and not evaluate_rule(q.applies_if, source=loan): continue
            if q.answer_type=="yes_no_na": ans="Yes"
            elif q.answer_type in ("number","currency","percent"): ans=loan.get(q.source_field) or 1.25
            elif q.answer_type=="date": ans=loan.get(q.source_field) or "2026-01-01"
            elif q.answer_type=="select": ans=q.options[0]
            elif q.answer_type=="multi_select": ans=q.options[0]
            else: ans="Reviewed and supported."
            save_answer(conn,rcid,loan,sec,q,ans,"Supported by workpaper.","Attached","Reviewer",template.template_id,template.version)
    set_review_status(conn,rcid,"Ready for QC")

def test_export_engine_generates_excel_and_data_mart_columns_and_audit(workflow):
    conn=workflow["conn"]; template=workflow["template"]
    ready_case=None
    for c in workflow["cases"]:
        row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone(); loan={k:row[k] for k in row.keys()}
        if loan["validation_status"] == "Ready": ready_case=(c,loan); break
    rcid, loan=ready_case; complete_case(conn,rcid,loan,template)
    excel=generate_excel_linesheet(conn,rcid,template,workflow["tmp"] / "excel", "Reviewer")
    assert Path(excel.file_path).exists()
    wb=load_workbook(excel.file_path)
    assert ["Cover","Loan Summary","Cash Flow Analysis","Ability-to-Repay (DTI)","Collateral & LTV","Linesheet Questions","Exceptions & Findings","Evidence Checklist","Audit Summary"] == wb.sheetnames
    assert wb["Linesheet Questions"][1][0].value == "Section"
    csv=generate_data_mart_csv(conn,rcid,template,workflow["tmp"]/"data_mart"/"review_answers_export.csv","Reviewer")
    df=pd.read_csv(csv.file_path)
    required={"client_name","review_period","template_id","template_version","review_case_id","loan_id","borrower_name","question_id","section","answer_value","status","severity","exception_flag","reviewer_comment","evidence_status","answered_by","answered_at","exported_at"}
    assert required.issubset(df.columns)
    audit=generate_audit_log_csv(conn, workflow["tmp"]/"audit"/"audit.csv")
    adf=pd.read_csv(audit.file_path)
    actions=set(adf.action_type)
    assert {"import_created","mapping_saved","loan_records_normalized","validation_run","answer_changed","export_generated"}.issubset(actions)

def test_audit_log_records_exception_creation(workflow):
    conn=workflow["conn"]; template=workflow["template"]; rcid=workflow["cases"][0]
    row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(rcid,)).fetchone(); loan={k:row[k] for k in row.keys()}
    sec=next(s for s in template.sections if s.section_id=="borrower_relationship"); q=next(q for q in sec.questions if q.question_id=="BR1")
    save_answer(conn,rcid,loan,sec,q,"No","Issue documented","Attached","Reviewer",template.template_id,template.version)
    actions=[r[0] for r in conn.execute("SELECT action_type FROM audit_log").fetchall()]
    assert "exception_created" in actions

def test_dti_results_carry_into_findings_and_data_mart(workflow):
    conn=workflow["conn"]; template=workflow["template"]
    rcid=None
    for c in workflow["cases"]:
        row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone(); loan={k:row[k] for k in row.keys()}
        if loan["validation_status"]=="Ready": rcid=c; break
    # over-guideline DTI (back-end ~45.7%) should carry as a finding
    save_dti_inputs(conn, rcid, {"base_income":9000,"principal_interest":2510,"auto":620,"credit_cards":360,"student_loans":320,"installment":180,"other_debt":120}, "Reviewer", loan_id=loan["loan_id"])
    dti_exc=conn.execute("SELECT severity,issue_text FROM exceptions WHERE review_case_id=? AND question_id='DTI_ATR'",(rcid,)).fetchone()
    assert dti_exc is not None and dti_exc["severity"]=="Finding" and "back-end DTI" in dti_exc["issue_text"]
    complete_case(conn, rcid, loan, template)
    csv=generate_data_mart_csv(conn, rcid, template, workflow["tmp"]/"dm_dti.csv", "Reviewer")
    df=pd.read_csv(csv.file_path)
    back=df[df.question_id=="DTI_BACK_END_PCT"]
    assert len(back)==1 and back.iloc[0]["section"]=="ability_to_repay" and back.iloc[0]["severity"]=="Finding"
    assert "DTI_ASSESSMENT" in set(df.question_id)
    # clearing the worksheet (within guidelines) resolves the carried finding
    save_dti_inputs(conn, rcid, {"base_income":12000,"principal_interest":1500,"auto":200}, "Reviewer", loan_id=loan["loan_id"])
    assert conn.execute("SELECT COUNT(*) FROM exceptions WHERE review_case_id=? AND question_id='DTI_ATR'",(rcid,)).fetchone()[0]==0

CF_CONFIG_PATH=ROOT/"configs"/"cash_flow_v1.yaml"

def test_cash_flow_normalize_methods_and_basis():
    assert normalize_line(80000, 120000, "Annual", "Latest") == round(120000/12,2)
    assert normalize_line(80000, 120000, "Annual", "Average") == round(100000/12,2)
    assert normalize_line(80000, 120000, "Annual", "Lower of") == round(80000/12,2)
    assert normalize_line(0, 5000, "Monthly", "Latest") == 5000
    assert normalize_line(0, 90000, "Annual", "Average") == round(90000/12,2)  # one period present

def test_cash_flow_distributions_qualify_business_income_is_reference():
    cfg=load_cash_flow_config(CF_CONFIG_PATH)
    vals={"w2_wages":{"period2":96000,"basis":"Annual","method":"Latest"},
          "k1_distributions":{"period1":24000,"period2":36000,"basis":"Annual","method":"Average"},
          "k1_ordinary_income":{"period2":120000,"basis":"Annual","method":"Latest"}}
    r=compute_cash_flow(vals, cfg)
    assert r["qualifying_monthly"]==10500.0          # 8000 wages + 2500 distributions
    assert r["qualifying_annual"]==126000.0
    assert r["business_income_reference_monthly"]==10000.0  # ordinary income excluded from qualifying

def test_cash_flow_persist_summarize_and_audit(workflow):
    conn=workflow["conn"]; rcid=workflow["cases"][0]
    save_cash_flow_inputs(conn, rcid, {"w2_wages":{"period2":90000,"basis":"Annual","method":"Latest"}}, "Reviewer", loan_id="L1001")
    got=load_cash_flow_inputs(conn, rcid)
    assert got["w2_wages"]["period2"]==90000 and got["w2_wages"]["method"]=="Latest"
    s=summarize_cash_flow(conn, rcid)
    assert s and s["qualifying_monthly"]==round(90000/12,2)
    actions=[r[0] for r in conn.execute("SELECT action_type FROM audit_log WHERE review_case_id=?",(rcid,)).fetchall()]
    assert "cash_flow_updated" in actions
    assert summarize_cash_flow(conn, workflow["cases"][1]) is None  # empty -> None

def test_cash_flow_tab_and_data_mart_carry(workflow):
    conn=workflow["conn"]; template=workflow["template"]
    for c in workflow["cases"]:
        row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone(); loan={k:row[k] for k in row.keys()}
        if loan["validation_status"]=="Ready": rcid=c; break
    save_cash_flow_inputs(conn, rcid, {
        "w2_wages":{"period2":96000,"basis":"Annual","method":"Latest"},
        "k1_distributions":{"period1":24000,"period2":36000,"basis":"Annual","method":"Average"},
        "k1_ordinary_income":{"period2":120000,"basis":"Annual","method":"Latest"}}, "Reviewer", loan_id=loan["loan_id"])
    complete_case(conn, rcid, loan, template)
    excel=generate_excel_linesheet(conn, rcid, template, workflow["tmp"]/"excel_cf", "Reviewer")
    ws=load_workbook(excel.file_path)["Cash Flow Analysis"]
    assert any(isinstance(c.value,str) and c.value.startswith("=ROUND(SWITCH(") for col in ws.iter_cols(min_col=6,max_col=6) for c in col if c.value)
    csv=generate_data_mart_csv(conn, rcid, template, workflow["tmp"]/"dm_cf.csv", "Reviewer")
    df=pd.read_csv(csv.file_path)
    qm=df[df.question_id=="CF_QUALIFYING_MONTHLY"]
    assert len(qm)==1 and qm.iloc[0]["section"]=="cash_flow" and float(qm.iloc[0]["answer_value"])==10500.0

def test_template_builder_roundtrips_and_validates(tmp_path):
    from linesheet_builder.template_builder import q, section, build_template, write_template_yaml
    t=build_template("custom_demo_v1","Custom Demo",[
        section("s1","Borrower",[q("Q1","Identity verified?",required=True,exception_if='answer == "No"',severity="Finding")]),
        section("s2","Income",[q("Q2","Qualifying income","currency",required=True)])])
    p=write_template_yaml(t, tmp_path/"custom_demo_v1.yaml")
    loaded=load_template_yaml(p)  # parse back through the real loader -> validates structure
    assert loaded.template_id=="custom_demo_v1" and len(loaded.sections)==2
    q1=loaded.sections[0].questions[0]
    assert q1.question_id=="Q1" and q1.required and q1.exception_if=='answer == "No"' and q1.data_mart_field=="q1"

def test_template_builder_generates_catalog_en_masse(tmp_path):
    from linesheet_builder.template_builder import catalog, generate_templates
    from linesheet_builder.template_engine import discover_templates
    templates=catalog()
    assert len(templates) >= 4
    paths=generate_templates(templates, tmp_path)
    assert len(paths)==len(templates)
    reg=discover_templates(tmp_path)
    for t in templates:
        assert t.template_id in reg
        loaded=load_template_yaml(reg[t.template_id]["path"])  # each one is valid
        assert loaded.sections and all(s.questions for s in loaded.sections)

def test_template_builder_rejects_duplicate_question_ids():
    from linesheet_builder.template_builder import q, section, build_template
    with pytest.raises(ValueError, match="Duplicate question_id"):
        build_template("dup_v1","Dup",[section("s1","S",[q("X1","a"),q("X1","b")])])

def test_template_builder_options_and_severity_roundtrip(tmp_path):
    from linesheet_builder.template_builder import q, section, build_template, write_template_yaml
    t=build_template("opts_v1","Opts",[section("s1","S",[
        q("Q1","Pick one","select",required=True,options=["A","B","C"]),
        q("Q2","Concern?",required=True,warning_if='answer == "Yes"',severity="Needs Review")])])
    loaded=load_template_yaml(write_template_yaml(t, tmp_path/"opts_v1.yaml"))
    q1,q2=loaded.sections[0].questions
    assert q1.options==["A","B","C"] and q1.answer_type=="select"
    assert q2.warning_if=='answer == "Yes"' and q2.severity=="Needs Review"

def test_template_builder_ui_page_saves(tmp_path):
    pytest.importorskip("streamlit")
    import os
    from streamlit.testing.v1 import AppTest
    from linesheet_builder.template_builder import TEMPLATES_DIR
    tid="ui_test_tmp_v1"; path=TEMPLATES_DIR/f"{tid}.yaml"
    try:
        at=AppTest.from_file(str(ROOT/"app.py"), default_timeout=90).run()
        at.sidebar.radio[0].set_value("Template Builder").run()
        def click(label):
            for b in at.button:
                if b.label==label: b.click().run(); return True
            return False
        assert click("Add preset"); assert not at.exception
        at.text_input[0].set_value(tid).run()
        assert click("💾 Save template"); assert not at.exception
        assert path.exists() and load_template_yaml(str(path)).sections
    finally:
        if path.exists(): os.remove(str(path))

def test_data_mart_workbook_consolidates_engagement(workflow):
    conn=workflow["conn"]; template=workflow["template"]; eid=workflow["eid"]
    # complete one Ready case and attach DTI + cash flow so results carry into the mart
    for c in workflow["cases"]:
        row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone(); loan={k:row[k] for k in row.keys()}
        if loan["validation_status"]=="Ready": rcid=c; break
    save_dti_inputs(conn, rcid, {"base_income":9000,"principal_interest":2510,"auto":620,"credit_cards":360,"student_loans":320,"installment":180,"other_debt":120}, "Reviewer")
    save_cash_flow_inputs(conn, rcid, {"w2_wages":{"period2":96000,"basis":"Annual","method":"Latest"}}, "Reviewer")
    complete_case(conn, rcid, loan, template)
    res=generate_data_mart_workbook(conn, eid, workflow["tmp"]/"mart.xlsx", "Reviewer")
    wb=load_workbook(res.file_path)
    for s in ("Overview","Linesheets","Answers","Findings","DTI","CashFlow","Collateral","Audit","Data Dictionary"):
        assert s in wb.sheetnames
    ws=wb["Linesheets"]
    header=[c.value for c in ws[1]]
    assert "dti_back_end_pct" in header and "cf_qualifying_monthly" in header
    ncases=conn.execute("SELECT COUNT(*) FROM review_cases WHERE engagement_id=?",(eid,)).fetchone()[0]
    assert ws.max_row-1 == ncases  # one row per linesheet
    assert len(wb["Linesheets"].tables)==1  # registered as an Excel Table for pivots
    # the completed case carries its DTI back-end into the mart
    idx=header.index("review_case_id"); bidx=header.index("dti_back_end_pct")
    vals={ws.cell(row=r,column=idx+1).value: ws.cell(row=r,column=bidx+1).value for r in range(2,ws.max_row+1)}
    assert vals.get(rcid)==45.67

COLL_CONFIG_PATH=ROOT/"configs"/"collateral_v1.yaml"

def test_collateral_compute_secured_overltv_and_shortfall():
    from linesheet_builder.collateral_engine import load_collateral_config, compute_collateral
    cfg=load_collateral_config(COLL_CONFIG_PATH)
    # $1,000,000 RE @ 80% advance = 800k eligible; exposure 600k -> secured, LTV 60%
    sec=compute_collateral({"real_estate":{"market_value":1000000},"loan_balance":{"market_value":600000}}, cfg)
    assert sec["net_collateral_value"]==800000 and sec["ltv"]==60.0 and sec["severity"] is None and sec["assessment"]=="Adequately secured"
    # exposure 900k > market-LTV guideline (90% > 80%) but still secured by net value (800k<900k -> shortfall first)
    short=compute_collateral({"real_estate":{"market_value":1000000},"loan_balance":{"market_value":900000}}, cfg)
    assert short["net_collateral_value"] < short["total_exposure"] and short["severity"]=="Finding" and "shortfall" in short["assessment"]
    # high advance asset clears coverage but market LTV exceeds guideline
    over=compute_collateral({"cash_deposits":{"market_value":1000000},"loan_balance":{"market_value":850000}}, cfg)
    assert over["coverage"]>=100 and over["ltv"]==85.0 and over["severity"]=="Finding" and "LTV" in over["assessment"]

def test_collateral_persist_summarize_carry_and_audit(workflow):
    from linesheet_builder.collateral_engine import save_collateral_inputs, load_collateral_inputs, summarize_collateral
    conn=workflow["conn"]; rcid=workflow["cases"][0]
    save_collateral_inputs(conn, rcid, {"real_estate":{"market_value":1000000,"advance_rate":80},"loan_balance":{"market_value":900000}}, "Reviewer", loan_id="L1001")
    got=load_collateral_inputs(conn, rcid)
    assert got["real_estate"]["market_value"]==1000000 and got["real_estate"]["advance_rate"]==80
    s=summarize_collateral(conn, rcid); assert s and s["severity"]=="Finding"
    exc=conn.execute("SELECT severity FROM exceptions WHERE review_case_id=? AND question_id='COLL_LTV'",(rcid,)).fetchone()
    assert exc and exc["severity"]=="Finding"
    actions=[r[0] for r in conn.execute("SELECT action_type FROM audit_log WHERE review_case_id=?",(rcid,)).fetchall()]
    assert "collateral_updated" in actions
    # clearing to adequately secured resolves the carried finding
    save_collateral_inputs(conn, rcid, {"real_estate":{"market_value":1000000,"advance_rate":80},"loan_balance":{"market_value":500000}}, "Reviewer", loan_id="L1001")
    assert conn.execute("SELECT COUNT(*) FROM exceptions WHERE review_case_id=? AND question_id='COLL_LTV'",(rcid,)).fetchone()[0]==0

def test_rules_engine_safe_supported_and_no_raw_eval():
    assert evaluate_rule('answer == "No"', answer="No")
    assert evaluate_rule('value < 1.20', value=1.1)
    assert evaluate_rule('source.dscr < 1.20', source={"dscr":1.1})
    assert evaluate_rule('source.ltv > 80', source={"ltv":85})
    assert evaluate_rule('source.nonaccrual_flag == true', source={"nonaccrual_flag":True})
    assert evaluate_rule('answer in ["No", "N/A"]', answer="N/A")
    assert evaluate_rule('required answer is blank', answer="", required=True)
    with pytest.raises(UnsafeRuleError): evaluate_rule('__import__("os").system("echo bad")')
    source=Path(ROOT/"linesheet_builder"/"rules_engine.py").read_text()
    tree=ast.parse(source)
    assert not any(isinstance(n, ast.Call) and getattr(n.func, 'id', '') == 'eval' for n in ast.walk(tree))

DTI_CONFIG_PATH=ROOT/"configs"/"dti_worksheet_v1.yaml"

def test_dti_config_loads_three_blocks():
    cfg=load_dti_config(DTI_CONFIG_PATH)
    assert {"income","housing","debts"}.issubset(cfg)
    assert any(k=="base_income" for k,_ in [(l["key"],l["label"]) for l in cfg["income"]["lines"]])

def test_dti_compute_ratios_and_within_guidelines():
    cfg=load_dti_config(DTI_CONFIG_PATH)
    r=compute_dti({"base_income":8000,"principal_interest":1600,"property_taxes":300,"auto":400,"credit_cards":200}, cfg)
    assert r["total_income"]==8000 and r["total_housing"]==1900 and r["total_other_debt"]==600
    assert r["total_obligations"]==2500 and r["residual_income"]==5500
    assert r["front_end_dti"]==round(1900/8000*100,2)
    assert r["back_end_dti"]==round(2500/8000*100,2)
    assert r["severity"] is None and r["assessment"].startswith("Within")

def test_dti_compute_exception_and_fail_bands():
    cfg=load_dti_config(DTI_CONFIG_PATH)
    exc=compute_dti({"base_income":9000,"principal_interest":2510,"auto":620,"credit_cards":360,"student_loans":320,"installment":180,"other_debt":120}, cfg)
    assert exc["back_end_dti"]>43 and exc["back_end_dti"]<=50
    assert exc["severity"]=="Finding" and "Exceeds" in exc["assessment"]
    fail=compute_dti({"base_income":5000,"principal_interest":2000,"auto":700,"credit_cards":500}, cfg)
    assert fail["back_end_dti"]>50 and fail["severity"]=="Blocked" and "Fails" in fail["assessment"]

def test_dti_net_residual_with_payroll_withholding():
    cfg=load_dti_config(DTI_CONFIG_PATH)
    r=compute_dti({"base_income":8000,"principal_interest":1600,"property_taxes":300,"auto":400,"credit_cards":200,
                   "payroll_withholding":1500}, cfg)
    assert r["total_withholding"]==1500
    assert r["net_income"]==6500            # 8000 - 1500
    assert r["net_residual_income"]==4000   # 6500 - 2500 obligations
    # gross ratios are unaffected by withholding
    assert r["back_end_dti"]==round(2500/8000*100,2)

def test_dti_net_equals_gross_when_no_withholding():
    cfg=load_dti_config(DTI_CONFIG_PATH)
    r=compute_dti({"base_income":8000,"principal_interest":1900,"auto":600}, cfg)
    assert r["total_withholding"]==0
    assert r["net_income"]==r["total_income"]
    assert r["net_residual_income"]==r["residual_income"]

def test_dti_inputs_persist_and_reload_with_audit(workflow):
    conn=workflow["conn"]; rcid=workflow["cases"][0]
    save_dti_inputs(conn, rcid, {"base_income":7000,"auto":450}, "Reviewer", loan_id="L1001")
    save_dti_inputs(conn, rcid, {"base_income":7200}, "Reviewer", loan_id="L1001")  # upsert
    vals=load_dti_inputs(conn, rcid)
    assert vals["base_income"]==7200 and vals["auto"]==450
    actions=[r[0] for r in conn.execute("SELECT action_type FROM audit_log WHERE review_case_id=?",(rcid,)).fetchall()]
    assert "dti_updated" in actions

def test_dti_tab_appears_in_workbook(workflow):
    conn=workflow["conn"]; template=workflow["template"]
    for c in workflow["cases"]:
        row=conn.execute("SELECT lr.* FROM review_cases rc JOIN loan_records lr ON rc.loan_record_id=lr.loan_record_id WHERE rc.review_case_id=?",(c,)).fetchone(); loan={k:row[k] for k in row.keys()}
        if loan["validation_status"]=="Ready":
            rcid=c; break
    save_dti_inputs(conn, rcid, {"base_income":9000,"principal_interest":2510,"auto":620,"credit_cards":360}, "Reviewer")
    complete_case(conn, rcid, loan, template)
    excel=generate_excel_linesheet(conn, rcid, template, workflow["tmp"]/"excel2", "Reviewer")
    wb=load_workbook(excel.file_path)
    assert "Ability-to-Repay (DTI)" in wb.sheetnames
    ws=wb["Ability-to-Repay (DTI)"]
    assert any(str(cell.value).startswith("=SUM(") for col in ws.iter_cols(min_col=2,max_col=2) for cell in col if cell.value)
