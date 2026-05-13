from __future__ import annotations

import hashlib
import importlib.util
from datetime import date
from pathlib import Path

import pytest

from occam_template_desk.core.data_source import OccamWorkbook
from occam_template_desk.core.field_mapping import build_field_records, records_to_values
from occam_template_desk.core.invoice import invoice_selection_state, template_requires_invoice
from occam_template_desk.core.outlook_workflow import create_outlook_draft_if_allowed
from occam_template_desk.core.package_builder import build_output_package
from occam_template_desk.core.renderer import detect_unresolved_placeholders, render_email_template, scan_rendered_output_for_placeholders
from occam_template_desk.core.simple_xlsx import read_xlsx
from occam_template_desk.core.validation import validate_run


def _sha(path: str | Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


@pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None, reason="openpyxl is not installed in this environment")
def test_openpyxl_workbook_loading_with_blanks_dates_and_numbers(tmp_path):
    from openpyxl import Workbook

    workbook_path = tmp_path / "real.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(["Client ID", "Client Name", "Client Email", "Tax Year", "Fee Amount", "Start Date", "Blank Field"])
    ws.append(["C-1", "Blank Date Client", None, 2025, 1234.5, date(2026, 5, 13), None])
    ws.append(["C-2", "Sparse Client"])
    inv = wb.create_sheet("Invoices")
    inv.append(["Invoice Number", "Client ID", "Invoice Amount", "Due Date"])
    inv.append(["INV-1", "C-1", 250, date(2026, 5, 20)])
    wb.create_sheet("Settings").append(["Setting", "Value"])
    wb.save(workbook_path)

    source_hash = _sha(workbook_path)
    occam = OccamWorkbook(workbook_path)

    client = occam.get_client("C-1", "Client ID")
    assert client["Client Email"] == ""
    assert client["Fee Amount"] == "1234.5"
    assert client["Start Date"] == "2026-05-13"
    assert client["Blank Field"] == ""
    assert occam.get_client("C-2", "Client ID")["Tax Year"] == ""
    assert _sha(workbook_path) == source_hash


def test_client_match_field_supports_client_id_and_client_name():
    wb = OccamWorkbook("occam_template_desk/sample_data/Occam_Data.xlsx")
    by_id = wb.get_client("C-1001", "Client ID")
    by_name = wb.get_client("Arbor & Finch LLC", "Client Name")
    assert by_id["Client Name"] == "Arbor & Finch LLC"
    assert by_name["Client ID"] == "C-1001"
    options = wb.client_options("Client ID")
    assert any(option["match_value"] == "C-1001" for option in options)


def test_invoice_selection_logic_zero_one_and_multiple():
    assert template_requires_invoice(["Invoice Number"], "Email.html")
    assert invoice_selection_state([], True)["status"] == "none"
    one = [{"Invoice Number": "INV-1"}]
    assert invoice_selection_state(one, True)["status"] == "auto_selected"
    assert invoice_selection_state(one, True)["selected_invoice"] == one[0]
    many = [{"Invoice Number": "INV-1"}, {"Invoice Number": "INV-2"}]
    state = invoice_selection_state(many, True)
    assert state["status"] == "multiple"
    assert state["requires_user_selection"] is True


def test_no_invoice_for_invoice_template_blocks(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    values = {"Client Name": "No Invoice", "Client Email": "client@example.com"}
    result = validate_run(template, "email", values, ["Invoice Number", "Invoice Amount"], tmp_path, subject="Invoice")
    assert result.status == "Blocked"
    assert any("Invoice Number" in blocker for blocker in result.blockers)
    assert any("Invoice Amount" in blocker for blocker in result.blockers)


class MockOutlookService:
    def __init__(self, available=True):
        self.available = available
        self.calls = []
        self.sent = False

    def is_available(self):
        return self.available

    def create_draft(self, **kwargs):
        self.calls.append(kwargs)
        return {"attempted": True, "created": True, "mode": "local_outlook", "message": "mock draft"}

    def fallback_status(self, output_dir, reason):
        return {"attempted": True, "created": False, "mode": "fallback_files", "message": reason}


def test_blocked_validation_prevents_package_generation_and_outlook(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    validation = validate_run(template, "email", {"Client Name": "", "Client Email": "bad"}, ["Client Name"], tmp_path, subject="")
    with pytest.raises(ValueError):
        build_output_package(template, "email", {}, [], validation, {"Client Name": ""}, tmp_path)
    service = MockOutlookService()
    status = create_outlook_draft_if_allowed(validation, "local_outlook", {"subject": "x", "body": "y", "body_path": "draft.html"}, {"Client Email": "a@example.com"}, service=service)
    assert status["attempted"] is False
    assert service.calls == []
    assert service.sent is False


def test_outlook_draft_creation_uses_mock_and_never_sends(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    validation = validate_run(template, "email", {"Client Name": "A", "Client Email": "a@example.com", "Invoice Number": "INV", "Invoice Amount": "10"}, [], tmp_path, subject="Invoice")
    service = MockOutlookService()
    status = create_outlook_draft_if_allowed(validation, "local_outlook", {"subject": "Invoice", "body": "Body", "body_path": "draft.html"}, {"Client Email": "a@example.com"}, attachments=["attachment.pdf"], service=service, output_dir=tmp_path)
    assert status["created"] is True
    assert service.calls[0]["to"] == "a@example.com"
    assert service.calls[0]["subject"] == "Invoice"
    assert service.calls[0]["html_body"] is True
    assert service.calls[0]["attachments"] == ["attachment.pdf"]
    assert service.sent is False


def test_unreplaced_placeholders_are_detected_in_text_email(tmp_path):
    template = tmp_path / "email.txt"
    template.write_text("Subject: Hello {{Client Name}}\nBody {{Missing Field}}", encoding="utf-8")
    rendered = render_email_template(template, {"Client Name": "Alex", "Client Email": "a@example.com"}, tmp_path)
    assert detect_unresolved_placeholders(rendered["body"]) == ["Missing Field"]
    assert scan_rendered_output_for_placeholders(rendered["body_path"]) == ["Missing Field"]


def test_source_workbook_and_template_are_not_modified(tmp_path):
    workbook = Path("occam_template_desk/sample_data/Occam_Data.xlsx")
    template = Path("occam_template_desk/sample_templates/Documents/Individual Tax Engagement Letter.docx")
    workbook_hash = _sha(workbook)
    template_hash = _sha(template)
    values = {"Client Name": "Arbor", "Tax Year": "2025", "Contact First Name": "Alex", "Firm Name": "Occam Advisors", "Fee Amount": "100", "Billing Terms": "Due", "Partner": "Dana", "Manager": "Mia", "Payment Link": "https://pay"}
    fields = [{"field": key, "value": value, "source": "test", "status": "Filled"} for key, value in values.items()]
    validation = validate_run(template, "document", values, [], tmp_path)
    build_output_package(str(template), "document", values, fields, validation, {"Client Name": "Arbor", "Client ID": "C-1"}, tmp_path)
    assert _sha(workbook) == workbook_hash
    assert _sha(template) == template_hash


def test_validation_report_has_expected_sheets_and_labels(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    values = {"Client Name":"Arbor", "Client Email":"alex@example.com", "Invoice Number":"INV-1", "Invoice Amount":"100", "Firm Name":"Occam Advisors", "Contact First Name":"Alex", "Service Description":"Tax", "Due Date":"2026-05-20", "Payment Link":"https://pay", "Manager":"Mia"}
    fields = [{"field": key, "value": value, "source": "test", "status": "Filled"} for key, value in values.items()]
    validation = validate_run(template, "email", values, [], tmp_path, subject="Invoice INV-1")
    package = build_output_package(template, "email", values, fields, validation, {"Client Name":"Arbor", "Client ID":"C-1"}, tmp_path)
    report = Path(package["validation_report"])
    assert report.exists()
    if importlib.util.find_spec("openpyxl") is not None:
        from openpyxl import load_workbook
        workbook = load_workbook(report, read_only=True)
        assert workbook.sheetnames == ["Summary", "Fields", "Validation Results", "Audit Log"]
        assert workbook["Summary"]["A1"].value.startswith("Occam Template Desk")
        summary_labels = [cell.value for cell in workbook["Summary"]["A"]]
        assert "Status" in summary_labels
        assert "Blocker Count" in summary_labels
        assert "Warning Count" in summary_labels
        workbook.close()
    else:
        sheets = read_xlsx(report)
        assert set(sheets) == {"Summary", "Fields", "Validation Results", "Audit Log"}
        assert any(row.get("Item") == "Status" for row in sheets["Summary"])
