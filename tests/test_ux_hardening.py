from __future__ import annotations

import importlib.util
import json
from pathlib import Path

from occam_template_desk.core.data_source import OccamWorkbook
from occam_template_desk.core.field_mapping import build_field_records, records_to_values
from occam_template_desk.core.outlook_workflow import create_outlook_draft_if_allowed
from occam_template_desk.core.package_builder import build_output_package, update_outlook_status
from occam_template_desk.core.renderer import render_email_template
from occam_template_desk.core.session_state import clear_generation_state, get_generation_state, save_generation_state
from occam_template_desk.core.simple_xlsx import read_xlsx
from occam_template_desk.core.validation import validate_run


def _summary_labels_from_report(report_path: str | Path) -> dict:
    if importlib.util.find_spec("openpyxl") is not None:
        from openpyxl import load_workbook

        workbook = load_workbook(report_path, read_only=True, data_only=True)
        try:
            rows = list(workbook["Summary"].iter_rows(values_only=True))
            return {str(row[0]): "" if len(row) < 2 or row[1] is None else str(row[1]) for row in rows if row and row[0] and row[0] != "Item" and not str(row[0]).startswith("Occam Template Desk")}
        finally:
            workbook.close()
    summary = read_xlsx(report_path)["Summary"]
    return {row.get("Item"): row.get("Value") for row in summary}


class MockOutlookService:
    def __init__(self, available=True):
        self.available = available
        self.calls = []

    def is_available(self):
        return self.available

    def create_draft(self, **kwargs):
        self.calls.append(kwargs)
        return {"attempted": True, "created": True, "mode": "local_outlook", "message": "mock draft"}

    def fallback_status(self, output_dir, reason):
        return {"attempted": True, "created": False, "mode": "fallback_files", "message": reason}


def _email_values():
    return {"Client Name":"Arbor", "Client Email":"alex@example.com", "Invoice Number":"INV-1", "Invoice Amount":"100", "Firm Name":"Occam Advisors", "Contact First Name":"Alex", "Service Description":"Tax", "Due Date":"2026-05-20", "Payment Link":"https://pay", "Manager":"Mia"}


def test_streamlit_safe_generation_state_helpers():
    store = {}
    package = {"package_dir": "out", "validation": {"status": "Ready"}}
    state = save_generation_state(store, package, {"subject": "Hi"}, {"status": "Ready"}, "template", {"Client Name": "A"}, {"Client Email": "a@example.com"})
    assert get_generation_state(store) == state
    assert state["selected_template"] == "template"
    clear_generation_state(store)
    assert get_generation_state(store) is None


def test_outlook_draft_blocked_when_unresolved_placeholders_remain(tmp_path):
    service = MockOutlookService()
    validation = {"status": "Blocked - Do Not Send", "blockers": ["Rendered output contains unreplaced placeholders"], "unresolved_placeholders": ["Missing Field"]}
    status = create_outlook_draft_if_allowed(validation, "local_outlook", {"subject": "Hi", "body": "Body", "body_path": "draft.html"}, {"Client Email": "a@example.com"}, service=service, output_dir=tmp_path)
    assert status["attempted"] is False
    assert service.calls == []


def test_outlook_draft_blocked_when_body_has_placeholder_even_without_validation_flag(tmp_path):
    service = MockOutlookService()
    validation = {"status": "Ready", "blockers": [], "unresolved_placeholders": []}
    status = create_outlook_draft_if_allowed(validation, "local_outlook", {"subject": "Hi", "body": "Hello {{Missing Field}}", "body_path": "draft.html"}, {"Client Email": "a@example.com"}, service=service, output_dir=tmp_path)
    assert status["attempted"] is False
    assert "Missing Field" in status["unresolved_placeholders"]
    assert service.calls == []


def test_duplicate_package_generation_does_not_collide(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    values = _email_values()
    fields = [{"field": key, "value": value, "source": "test", "status": "Filled"} for key, value in values.items()]
    validation = validate_run(template, "email", values, [], tmp_path, subject="Invoice INV-1")
    first = build_output_package(template, "email", values, fields, validation, {"Client Name":"Arbor", "Client ID":"C-1"}, tmp_path)
    second = build_output_package(template, "email", values, fields, validation, {"Client Name":"Arbor", "Client ID":"C-1"}, tmp_path)
    assert first["package_dir"] != second["package_dir"]
    assert Path(first["package_dir"]).exists()
    assert Path(second["package_dir"]).exists()


def test_missing_items_render_as_readable_list(tmp_path):
    wb = OccamWorkbook("occam_template_desk/sample_data/Occam_Data.xlsx")
    client = wb.get_client("C-1001", "Client ID")
    pool = wb.field_pool_for_client(client)
    assert pool["missing items"][0] == "- Signed engagement letter\n- December bank statement"
    template = tmp_path / "missing.txt"
    template.write_text("Subject: Missing\n{{Missing Items}}", encoding="utf-8")
    fields = build_field_records(["Missing Items"], pool)
    rendered = render_email_template(template, records_to_values(fields), tmp_path)
    assert "- Signed engagement letter" in rendered["body"]
    assert "\n- December bank statement" in rendered["body"]


def test_validation_report_summary_includes_counts(tmp_path):
    template = tmp_path / "email.txt"
    template.write_text("Subject: Hi {{Client Name}}\nHello {{Missing Field}}", encoding="utf-8")
    values = {"Client Name": "A", "Client Email": "a@example.com"}
    fields = [{"field": "Client Name", "value": "A", "source": "test", "status": "Filled"}]
    validation = validate_run(template, "email", values, ["Client Name"], tmp_path, subject="Hi A")
    package = build_output_package(str(template), "email", values, fields, validation, {"Client Name":"A", "Client ID":"C-1"}, tmp_path)
    labels = _summary_labels_from_report(package["validation_report"])
    assert labels["Status"] == "Blocked - Do Not Send"
    assert labels["Blocker Count"] == "1"
    assert labels["Warning Count"] == "1"
    assert "generated_email.txt" in labels["Generated Files"]


def test_audit_log_updates_after_outlook_attempt(tmp_path):
    template = "occam_template_desk/sample_templates/Emails/Invoice Delivery Email.html"
    values = _email_values()
    fields = [{"field": key, "value": value, "source": "test", "status": "Filled"} for key, value in values.items()]
    validation = validate_run(template, "email", values, [], tmp_path, subject="Invoice INV-1")
    package = build_output_package(template, "email", values, fields, validation, {"Client Name":"Arbor", "Client ID":"C-1"}, tmp_path)
    status = {"attempted": True, "created": False, "mode": "fallback_files", "message": "Unavailable"}
    update_outlook_status(package["package_dir"], status)
    assert json.loads((Path(package["package_dir"]) / "outlook_status.json").read_text())["message"] == "Unavailable"
    audit = json.loads(Path(package["audit_log"]).read_text())
    assert audit["outlook_draft_status"] == status
