#!/usr/bin/env python3
"""Build the FRED Credit-Risk Dashboard workbook — KeyBank-styled.

PATCHED to use the house design system. All visual style now comes from
``keybank_style`` (tokens + helpers) and ``keybank_charts`` (native Excel line
charts); the data path, formulas, fixed-anchor raw layout and watchlist gate are
unchanged. To re-skin or extend, change a token in keybank_style.py — never a
hardcoded hex here.

Drop these three files into fred-credit-risk-dashboard/ next to make_workbook.py:
    build_workbook.py   (this file)
    keybank_style.py
    keybank_charts.py

Produces a base .xlsx; assemble_xlsm.py wraps it into the macro-enabled .xlsm.
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import runner as R
import series_seed as SEED

# ---- the house design system (single source of style) --------------------
import keybank_style as KB
from keybank_style import (
    INK, ONYX, KEY_RED, CRIMSON, CANVAS, MIST, SLATE,
    HDR_FILL, SECT_FILL, MONO_FONT, NOTE_FONT,
    brand_banner, kpi_tiles, header_row, section_band, watchlist_boundary,
    zscore_heat, yoy_heat, alert_rule, tighten_rule, hide_gridlines, freeze_below,
)
# NOTE: native openpyxl LineCharts were removed -- they are the most common
# trigger for Excel's "unreadable content / recovered" repair and are rebuilt on
# every openpyxl refresh. The dashboards stay fully intact without them. Charts
# can be re-added later in Excel (or via VBA, like the sparklines).

HERE = os.path.dirname(os.path.abspath(__file__))

# A couple of local text styles built from tokens (kept here, not inlined ad-hoc).
INK_BOLD = Font(name="Arial", bold=True, color=INK)
THIN = Side(style="thin", color="E7E2D8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def text_cell(ws, row, col, value):
    """Write a literal-text cell. openpyxl treats a string starting with '='
    as a formula; force it back to a string so source and doc lines round-trip."""
    cell = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


# ==========================================================================
# _config  (unchanged content; restyled section bands)
# ==========================================================================
def config_rows():
    rows = [["FRED Credit-Risk Dashboard -- CONFIG (the knob panel). "
             "Edit values here; no code change needed."], []]
    rows.append(["[SETTINGS]"])
    rows.append(["key", "value", "help"])
    rows += [
        ["fred_api_key", "", "Optional: paste your FRED key here if you don't set "
         "the FRED_API_KEY environment variable. Get one free at "
         "https://fredaccount.stlouisfed.org/apikeys"],
        ["write_backend", "auto", "auto|xlwings|openpyxl. xlwings writes into the open book."],
        ["demo_mode", "FALSE", "TRUE = offline synthetic data (no key/network), for trying the button."],
        ["raw_slots", "100", "observations kept per series (newest-first)."],
        ["stale_multiplier", "2.0", "flag a series stale if older than this x its cadence."],
        ["fred_min_interval", "0.6", "seconds between FRED requests (FRED allows ~120/min; keep >=0.5)."],
        ["fred_max_retries", "4", "retries with backoff if FRED returns a rate-limit error."],
    ]
    rows += [[], ["[THRESHOLDS]"], ["key", "value", "help"]]
    rows.append(["zscore_band", 1.0, "flag a loss/delinquency series when its 8-period z-score >= this."])
    rows.append(["sloos_band", 20.0, "flag a SLOOS series when net-tightening >= this (percent)."])
    rows += [[], ["[SERIES]"], SEED.HEADER]
    for r in SEED.all_series():
        rows.append([r[h] for h in SEED.HEADER])
    rows += [[], ["[CBSA_EXTENSIONS]"], ["cbsa", "name", "series_id"]]
    for r in SEED.cbsa_extension_rows():
        rows.append([r["cbsa"], r["name"], r["series_id"]])
    rows += [[], ["# To add a metro: add an FHFA series row to [SERIES] using "
                  "series_id ATNHPIUS<CBSA>Q, category hpi_metro, lane price, "
                  "geo_segment cbsa:<CBSA>, watchlist_capable TRUE, transform yoy_pct."]]
    return rows


def write_config(wb):
    ws = wb.create_sheet("_config")
    hide_gridlines(ws)
    rows = config_rows()
    zrow = srow = None
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if row and row[0] == "zscore_band":
            zrow = i
        if row and row[0] == "sloos_band":
            srow = i
    ws["A1"].font = INK_BOLD
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["O"].width = 50
    wb.defined_names["zscore_band"] = DefinedName("zscore_band", attr_text=f"_config!$B${zrow}")
    wb.defined_names["sloos_band"] = DefinedName("sloos_band", attr_text=f"_config!$B${srow}")
    # Onyx section bands for the [SECTION] markers — quiet system-tab styling.
    for i, row in enumerate(rows, start=1):
        if row and isinstance(row[0], str) and row[0].startswith("["):
            ws.cell(i, 1).font = KB.SECTION_FONT
            ws.cell(i, 1).fill = SECT_FILL
    return ws


# ==========================================================================
# Raw_* scaffolds  (fixed anchors; runner fills the values)
# ==========================================================================
def write_raw_scaffold(wb, specs, blocks):
    by_tab = {}
    for s in specs:
        by_tab.setdefault(blocks[s.series_id].tab, []).append(s)
    for tab, tab_specs in by_tab.items():
        ws = wb.create_sheet(tab)
        hide_gridlines(ws)
        ws["A1"] = (f"{tab} -- raw FRED observations, newest-first. Written by runner.py; "
                    f"dashboards read these by formula. Do not edit by hand.")
        ws["A1"].font = NOTE_FONT
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 38
        for s in tab_specs:
            b = blocks[s.series_id]
            ws.cell(b.header_row, 1, s.series_id).font = INK_BOLD
            ws.cell(b.header_row, 2, s.title)
            ws.cell(b.header_row, 3, f"freq={s.frequency}; transform={s.transform}; "
                                     f"{'WATCHLIST' if s.watchlist_capable else 'dashboard'}")
            ws.cell(b.header_row, 3).font = NOTE_FONT
            ws.cell(b.label_row, 1, "date").font = Font(name="Calibri", bold=True, size=9)
            ws.cell(b.label_row, 2, "value").font = Font(name="Calibri", bold=True, size=9)
    return by_tab


# ==========================================================================
# Dashboards
# ==========================================================================
def _ref(b, offset=0):
    return f"{b.tab}!{get_column_letter(R.RAW_VALUE_COL)}{b.first_data_row + offset}"


def _tier(series_id):
    if series_id.endswith("ACBS"):
        return "All banks"
    if series_id.endswith("T100S"):
        return "Top 100"
    if series_id.endswith("OBS"):
        return "Not top 100"
    return "-"


# Trend (8q) sits before Flag; the macro paints sparklines into it after data lands.
DASH_COLS = ["Tier", "Category", "Series ID", "Title", "Latest", "Prior",
             "YoY %", "Z-score (8)", "Trend (8q)", "Flag"]
COL = {name: get_column_letter(i + 1) for i, name in enumerate(DASH_COLS)}
NCOLS = len(DASH_COLS)


def write_dashboard(wb, tab, specs, blocks, title, subtitle, lane):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)

    rows = sorted([s for s in specs if s.dashboard_capable and not s.is_dead
                   and s.lane == lane],
                  key=lambda s: (s.category, s.series_id))

    hdr = 7                       # banner(1-2) + KPI tiles(3-5) + status(6) sit above
    first = hdr + 1
    r = first
    for s in rows:
        b = blocks[s.series_id]
        ppy = R.periods_per_year(s.frequency)
        latest, prior, yago = _ref(b, 0), _ref(b, 1), _ref(b, ppy)
        win = f"{b.tab}!{get_column_letter(R.RAW_VALUE_COL)}{b.first_data_row}:" \
              f"{get_column_letter(R.RAW_VALUE_COL)}{b.first_data_row + 7}"
        ws.cell(r, 1, _tier(s.series_id)).font = KB.SECONDARY
        ws.cell(r, 2, s.category).font = KB.DATA_FONT
        ws.cell(r, 3, s.series_id).font = MONO_FONT
        ws.cell(r, 4, s.title).font = KB.DATA_FONT
        ws.cell(r, 5, f"=IFERROR({latest},\"\")")
        ws.cell(r, 6, f"=IFERROR({prior},\"\")")
        ws.cell(r, 7, f"=IFERROR(({latest}-{yago})/{yago}*100,\"\")")
        ws.cell(r, 8, f"=IFERROR(({latest}-AVERAGE({win}))/STDEV({win}),\"\")")
        # col 9 (Trend) intentionally left blank — sparklines painted by the macro.
        zc, lc = f"${COL['Z-score (8)']}{r}", f"${COL['Latest']}{r}"
        if s.alert_rule == "zscore":
            flag = f"=IF(AND(ISNUMBER({zc}),{zc}>=zscore_band),\"⚠ ALERT\",\"\")"
        elif s.alert_rule == "sloos_level":
            flag = f"=IF(AND(ISNUMBER({lc}),{lc}>=sloos_band),\"⚠ TIGHTENING\",\"\")"
        else:
            flag = ""
        ws.cell(r, 10, flag)
        for c in (5, 6):
            ws.cell(r, c).number_format = KB.FMT_RATE
        ws.cell(r, 7).number_format = KB.FMT_YOY
        ws.cell(r, 8).number_format = KB.FMT_Z
        for c in (5, 6, 7, 8):
            ws.cell(r, c).alignment = KB.RIGHT
        for c in range(1, NCOLS + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        ws.cell(r, 10).alignment = KB.LEFT if flag else KB.CENTER
        r += 1
    last = r - 1

    # --- header band + freeze ---
    header_row(ws, hdr, DASH_COLS, right_from=4, center_cols=(8, 9))
    freeze_below(ws, hdr)
    _dash_widths(ws)

    # --- banner + KPI tiles (written after we know `last`) ---
    brand_banner(ws, 1, NCOLS, title, subtitle)
    zc, fc = COL["Z-score (8)"], COL["Flag"]
    if last >= first:
        tiles = [
            ("PORTFOLIO STRESS INDEX",
             f"=IFERROR(TEXT(AVERAGE({zc}{first}:{zc}{last}),\"0.00\"),\"—\")",
             "8-qtr composite z-score"),
            ("SERIES IN ALERT",
             f"=COUNTIF({fc}{first}:{fc}{last},\"*ALERT*\")",
             f"of {len(rows)} {lane}"),
            ("TIGHTENING SIGNALS",
             f"=COUNTIF({fc}{first}:{fc}{last},\"*TIGHTENING*\")",
             "SLOOS net tightening"),
        ]
        kpi_tiles(ws, 3, tiles, span=3, accents=[KEY_RED, KEY_RED, INK])
    ws.row_dimensions[6].height = 8     # thin spacer before the header band

    # Run-status readout, right of the merged masthead. The runner fills L1/L2
    # and the macro fills L4 — all in column L, free of the banner/KPI merges.
    scol = get_column_letter(R.STATUS_COL)
    ws.column_dimensions[scol].width = 34
    for rr in (1, 2):
        ws.cell(rr, R.STATUS_COL).font = KB.SECONDARY
        ws.cell(rr, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(3, R.STATUS_COL, "Member FDIC").font = NOTE_FONT
    ws.cell(3, R.STATUS_COL).alignment = KB.LEFT
    ws.cell(4, R.STATUS_COL).font = NOTE_FONT      # macro writes its status here
    ws.cell(4, R.STATUS_COL).alignment = KB.LEFT

    # --- heat + flag rules ---
    if last >= first:
        zscore_heat(ws, f"{zc}{first}:{zc}{last}")
        alert_rule(ws, f"{fc}{first}:{fc}{last}")
        tighten_rule(ws, f"{fc}{first}:{fc}{last}")

    return ws


def _dash_widths(ws):
    widths = {"A": 12, "B": 16, "C": 16, "D": 50, "E": 10, "F": 10, "G": 9,
              "H": 12, "I": 18, "J": 14}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


# ==========================================================================
# Watchlist_Geo  (ranked, validator-gated, geographic only)
# ==========================================================================
BOUNDARY = ("Geographic stress watchlist -- apply against portfolio collateral "
            "location manually. National credit-quality series are excluded by "
            "design; they cannot localize a portfolio subset.")

WL_COLS = ["Geography", "Series ID", "Source", "Latest", "YoY %", "Rank",
           "Trend (recent->older)"]


def _geo_label(spec):
    title = spec.title
    if " -- " in title:
        return title.split(" -- ", 1)[1]
    return spec.geo_segment


def _source(category):
    return {"hpi_state": "FHFA state", "hpi_metro": "FHFA metro (CBSA)",
            "hpi_caseshiller": "Case-Shiller metro"}.get(category, category)


def write_watchlist(wb, specs, blocks):
    R.validate_watchlist(specs)
    wl = sorted(R.watchlist_series(specs), key=lambda s: (s.category, _geo_label(s)))

    ws = wb.create_sheet("Watchlist_Geo")
    hide_gridlines(ws)
    brand_banner(ws, 1, len(WL_COLS),
                 "Watchlist_Geo — Geographic Stress Watchlist",
                 "States & metros ranked by house-price deterioration (FHFA / Case-Shiller).")
    watchlist_boundary(ws, 3, len(WL_COLS), BOUNDARY)   # the red gate banner

    hdr = 4
    header_row(ws, hdr, WL_COLS, right_from=3, center_cols=(6,))
    first = hdr + 1
    last = hdr + len(wl)
    yoy_col = get_column_letter(5)
    yoy_range = f"${yoy_col}${first}:${yoy_col}${last}"
    r = first
    for s in wl:
        b = blocks[s.series_id]
        ppy = R.periods_per_year(s.frequency)
        latest, yago = _ref(b, 0), _ref(b, ppy)
        v0, v1, v2, v3 = _ref(b, 0), _ref(b, 1), _ref(b, 2), _ref(b, 3)
        ws.cell(r, 1, _geo_label(s)).font = KB.DATA_FONT
        ws.cell(r, 2, s.series_id).font = MONO_FONT
        ws.cell(r, 3, _source(s.category)).font = KB.SECONDARY
        ws.cell(r, 4, f"=IFERROR({latest},\"\")")
        ws.cell(r, 5, f"=IFERROR(({latest}-{yago})/{yago}*100,\"\")")
        ws.cell(r, 6, f"=IFERROR(RANK({yoy_col}{r},{yoy_range},1),\"\")")
        arrow = (f"=IFERROR(IF({v0}>{v1},\"▲\",IF({v0}<{v1},\"▼\",\"→\"))"
                 f"&IF({v1}>{v2},\"▲\",IF({v1}<{v2},\"▼\",\"→\"))"
                 f"&IF({v2}>{v3},\"▲\",IF({v2}<{v3},\"▼\",\"→\")),\"\")")
        ws.cell(r, 7, arrow)
        ws.cell(r, 4).number_format = KB.FMT_YOY
        ws.cell(r, 5).number_format = KB.FMT_YOY
        for c in (4, 5, 6):
            ws.cell(r, c).alignment = KB.RIGHT
        ws.cell(r, 7).alignment = KB.CENTER
        for c in range(1, len(WL_COLS) + 1):
            ws.cell(r, c).border = KB.HAIRLINE
        r += 1
    widths = {"A": 40, "B": 16, "C": 18, "D": 10, "E": 9, "F": 8, "G": 16}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    freeze_below(ws, hdr)
    if last >= first:
        yoy_heat(ws, f"{yoy_col}{first}:{yoy_col}{last}")     # red = deterioration
        rank_col = get_column_letter(6)
        ws.conditional_formatting.add(
            f"{rank_col}{first}:{rank_col}{last}",
            ColorScaleRule(start_type="min", start_color=KB.HEAT_BAD,
                           end_type="max", end_color=KB.HEAT_GOOD))
    return ws


# ==========================================================================
# Code-in-tab + readme  (quiet system tabs)
# ==========================================================================
def write_code_tab(wb, tab, source_path, language):
    ws = wb.create_sheet(tab)
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 120
    with open(source_path, "r", encoding="utf-8") as fh:
        lines = fh.read().split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    if language == "vba":
        # Drop the leading "Attribute VB_Name = ..." line(s): not valid to paste
        # into a module code pane, so this tab copies straight into the editor.
        while lines and lines[0].startswith("Attribute "):
            lines = lines[1:]
    for i, line in enumerate(lines, start=1):
        text_cell(ws, i, 1, line)
        ws.cell(i, 1).font = MONO_FONT
    return len(lines)


README = """\
FRED CREDIT-RISK DASHBOARD -- README
====================================

WHAT THIS IS
  A self-contained, reusable template. One workbook pulls credit-risk data from
  FRED (Federal Reserve Economic Data), lands it raw, and presents it as
  formula-driven dashboards plus a geographic stress watchlist for commercial
  loan-portfolio targeting. All code, config and docs live inside this workbook,
  so it can be emailed, opened on another machine, and re-run with one click.

ONE-TIME SETUP
  1. Install Python 3 (add it to PATH) and the dependencies:
         pip install fredapi xlwings openpyxl pandas
  2. Get a free FRED API key: https://fredaccount.stlouisfed.org/apikeys
  3. Provide the key ONE of two ways (never hardcoded in code):
       - set an environment variable  FRED_API_KEY=your_key   (preferred), or
       - paste it into the _config tab: [SETTINGS] -> fred_api_key cell.
  4. Enable macros when prompted (the button needs them).

RUN IT  (the workbook is already the finished product; you only add data)
  1. In Excel press Alt+F8 -> run ExtractFiles. It writes three files next to the
     workbook: runner.py (from the _code_py tab), requirements.txt, and RUN.txt
     (the exact PowerShell commands). Nothing runs inside Excel.
  2. SAVE and CLOSE the workbook.
  3. Open RUN.txt and follow it from PowerShell: install the deps, then run
     runner.py --backend openpyxl against the closed workbook. It pulls FRED per
     _config and writes the data into the file.
  4. Reopen the workbook -- the formulas recalc into the populated dashboards.
  No key yet? Set _config demo_mode = TRUE before step 1 and use the --demo
  command in RUN.txt (offline synthetic data). For live data, set the
  FRED_API_KEY environment variable or the _config cell. Optional: after
  reopening, Alt+F8 -> PaintSparklines draws the Trend column (native sparklines
  are the one thing the data step can't write).

THE LOOK (house style: KeyBank)
  Every tab is styled from keybank_style.py (tokens + helpers) -- so this
  workbook and any future one built from the template look like one family.
  Black grounds, red leads sparingly, neutrals breathe. Dashboards: black banner
  + red rule, KPI tiles, brand-muted heat on the z-score column, and an optional
  8-quarter sparkline per row. Watchlist: the red geographic-boundary gate.
  System tabs: quiet Onyx bands.

THE TABS
  Dashboard_Consumer / _Commercial / _Price  -- formula-driven panels: latest,
      prior, YoY, an 8-period z-score, a Flag column, and an (optional) trend
      sparkline. Heat shading marks stress.
  Watchlist_Geo  -- states & metros ranked by house-price deterioration (YoY).
  Raw_Consumer / _Commercial / _Price  -- raw observations, newest-first.
  _config  -- THE KNOB PANEL: series dictionary, threshold bands, key fallback.
  _code_py / _code_vba  -- the runner and macro as plain text.

THE WATCHLIST BOUNDARY
  Only house-price indices with a geographic key a loan portfolio can join on
  (FHFA state/metro, Case-Shiller metros) may feed Watchlist_Geo. National
  credit-quality series are refused by the runner -- they cannot localize a
  portfolio subset.

NO AI IS INVOLVED IN THE DATA PATH. Thresholds, transforms and rankings are
deterministic and config-driven: the same FRED data always yields the same
workbook.
"""


def write_readme(wb):
    ws = wb.create_sheet("_readme")
    hide_gridlines(ws)
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README.split("\n"), start=1):
        text_cell(ws, i, 1, line)
        if line and not line.startswith(" ") and line == line.upper() and len(line) > 3 \
                and not line.startswith("="):
            ws.cell(i, 1).font = INK_BOLD
    return ws


# ==========================================================================
# Orchestration
# ==========================================================================
def build(out_path):
    rows = config_rows()
    cfg = R.parse_config(rows)
    specs = cfg.series
    R.validate_watchlist(specs)
    R.validate_transforms(specs)
    blocks = R.raw_layout(specs, slots=cfg.raw_slots)

    wb = Workbook()
    wb.remove(wb.active)

    write_raw_scaffold(wb, specs, blocks)         # raw tabs first (charts/formulas read them)
    write_dashboard(wb, "Dashboard_Consumer", specs, blocks,
                    "Consumer Credit-Risk Dashboard",
                    "Charge-offs, delinquencies, G.19, debt-service, SLOOS -- national, bank-tier where available.",
                    lane="consumer")
    write_dashboard(wb, "Dashboard_Commercial", specs, blocks,
                    "Commercial Credit-Risk Dashboard",
                    "C&I, CRE, all-loans charge-offs/delinquencies + SLOOS diffusion -- national.",
                    lane="commercial")
    write_dashboard(wb, "Dashboard_Price", specs, blocks,
                    "Price Dashboard",
                    "National house-price indices + commercial-real-estate price context.",
                    lane="price")
    write_watchlist(wb, specs, blocks)
    write_config(wb)
    write_code_tab(wb, "_code_py", os.path.join(HERE, "runner.py"), "python")
    write_code_tab(wb, "_code_vba", os.path.join(HERE, "macro.bas"), "vba")
    write_readme(wb)

    order = ["Dashboard_Consumer", "Dashboard_Commercial", "Dashboard_Price",
             "Watchlist_Geo", "Raw_Consumer", "Raw_Commercial", "Raw_Price",
             "_config", "_code_py", "_code_vba", "_readme"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.active = wb.sheetnames.index("Dashboard_Consumer")
    wb.save(out_path)
    return out_path, len(specs), len(R.watchlist_series(specs))


if __name__ == "__main__":
    out = os.path.join(HERE, "build", "FRED_Credit_Risk_Dashboard_base.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    path, n, nwl = build(out)
    print(f"built {path}: {n} series, {nwl} watchlist-capable")
