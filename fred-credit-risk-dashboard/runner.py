#!/usr/bin/env python3
"""FRED Credit-Risk Dashboard -- runner (the data path).

This module is the SINGLE SOURCE OF TRUTH for the data path and is embedded
verbatim into the workbook's ``_code_py`` tab. The VBA "Extract & Run" button
writes it back out as ``runner.py`` and executes it.

It owns, in clearly isolated sections so a future template can swap pieces
without a rebuild (BUILD SPEC sec 0.3):

  * PROVIDER ADAPTER -- the only FRED-specific code. Swap this class for v2.
  * TRANSFORM REGISTRY -- deterministic, pure, named transforms (sec 3).
  * WATCHLIST VALIDATOR -- the hard sample/watchlist boundary gate (sec 0.1).
  * RAW LAYOUT -- fixed-anchor block geometry shared with the builder so
    dashboard formulas never break across refreshes.
  * WRITE BACKENDS -- xlwings (primary, writes into the open book) and
    openpyxl (fallback, writes the closed file). Presentation logic never
    sees the backend.

No AI/LLM is involved anywhere in this path. Every transform is pure; a given
FRED input produces exactly one output (BUILD SPEC sec 0.5).
"""
from __future__ import annotations

import argparse
import json
import math
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Callable, Dict, List, Optional, Sequence

import pandas as pd

# Lane -> raw tab name. Presentation tabs read from these.
LANE_RAW_TAB = {
    "consumer": "Raw_Consumer",
    "commercial": "Raw_Commercial",
    "price": "Raw_Price",
}

# Fixed raw-block geometry (newest observation first). Both the builder and the
# runner derive anchors from this, so refilling raw never shifts a formula.
RAW_SLOTS_DEFAULT = 100      # observation rows kept per series (newest-first)
RAW_HEADER_ROWS = 2          # row 0: id/title/meta; row 1: "date"/"value"
RAW_GAP_ROWS = 2             # blank rows between series blocks
RAW_FIRST_ROW = 2            # blocks start at sheet row 2 (row 1 is a banner)
RAW_DATE_COL = 1             # column A
RAW_VALUE_COL = 2            # column B

# The dashboard run-status readout lands in a free column to the right of the
# table. The branded layout merges A:J for the banner/KPI strip, so columns >= L
# are the only collision-free home; two compact lines mirror the design masthead
# ("Last run …" / "Pulled n/total · s stale · a alerts"). The builder pre-styles
# these cells; the runner only writes values, and the macro writes row 4.
STATUS_COL = 12              # column L

# FRED missing-observation sentinel.
FRED_MISSING = "."

# How many periods make a year / one step, by frequency.
FREQ_PERIODS = {
    "quarterly": {"year": 4, "step": 1},
    "monthly": {"year": 12, "step": 1},
    "annual": {"year": 1, "step": 1},
    "weekly": {"year": 52, "step": 1},
}


def _freq_key(frequency: str) -> str:
    f = (frequency or "").strip().lower()
    if f.startswith("q"):
        return "quarterly"
    if f.startswith("m"):
        return "monthly"
    if f.startswith("a") or f.startswith("y"):
        return "annual"
    if f.startswith("w"):
        return "weekly"
    return "quarterly"


def periods_per_year(frequency: str) -> int:
    return FREQ_PERIODS[_freq_key(frequency)]["year"]


# ---------------------------------------------------------------------------
# CONFIG MODEL (parsed from the `_config` sheet; backend-agnostic)
# ---------------------------------------------------------------------------
@dataclass
class SeriesSpec:
    series_id: str
    title: str
    category: str
    lane: str
    metric_type: str
    frequency: str
    sa_nsa: str
    units: str
    level_rate_index: str
    geo_segment: str
    dashboard_capable: bool
    watchlist_capable: bool
    transform: str
    alert_rule: str          # zscore | sloos_level | none  (extends the dict)
    notes: str

    @property
    def is_dead(self) -> bool:
        """Documented-dead / discontinued series are kept in the dictionary
        for the record but never pulled live (BUILD SPEC sec 2)."""
        n = (self.notes or "").lower()
        return ("discontinued" in n) or ("documented-dead" in n) or ("do not pull" in n)


@dataclass
class Config:
    settings: Dict[str, str] = field(default_factory=dict)
    thresholds: Dict[str, float] = field(default_factory=dict)
    series: List[SeriesSpec] = field(default_factory=list)
    cbsa_extensions: List[Dict[str, str]] = field(default_factory=list)

    def setting(self, key: str, default=None):
        return self.settings.get(key, default)

    @property
    def zscore_band(self) -> float:
        return float(self.thresholds.get("zscore_band", 1.0))

    @property
    def sloos_band(self) -> float:
        return float(self.thresholds.get("sloos_band", 20.0))

    @property
    def raw_slots(self) -> int:
        return int(float(self.settings.get("raw_slots", RAW_SLOTS_DEFAULT)))

    @property
    def stale_multiplier(self) -> float:
        return float(self.settings.get("stale_multiplier", 2.0))


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


_SERIES_HEADER = [
    "series_id", "title", "category", "lane", "metric_type", "frequency",
    "sa_nsa", "units", "level_rate_index", "geo_segment", "dashboard_capable",
    "watchlist_capable", "transform", "alert_rule", "notes",
]


def parse_config(rows: Sequence[Sequence]) -> Config:
    """Parse the `_config` sheet (list of row value-lists) into a Config.

    Sections are delimited by a marker in column A: ``[SETTINGS]``,
    ``[THRESHOLDS]``, ``[SERIES]``, ``[CBSA_EXTENSIONS]``. Editing the sheet
    changes behaviour with no code edit (BUILD SPEC sec 4, the knob panel).
    """
    cfg = Config()
    section = None
    series_header: Optional[List[str]] = None
    cbsa_header: Optional[List[str]] = None
    for raw in rows:
        cells = list(raw) + [""] * (len(_SERIES_HEADER) - len(raw)) if raw else []
        a = ("" if not raw or raw[0] is None else str(raw[0])).strip()
        if a.startswith("[") and a.endswith("]"):
            section = a.strip("[]").strip().upper()
            series_header = None
            cbsa_header = None
            continue
        if not a:
            continue
        if section in ("SETTINGS", "THRESHOLDS"):
            if a.lower() in ("key", "name"):     # header row
                continue
            val = "" if len(raw) < 2 or raw[1] is None else raw[1]
            if section == "SETTINGS":
                cfg.settings[a] = str(val).strip()
            else:
                try:
                    cfg.thresholds[a] = float(val)
                except (TypeError, ValueError):
                    cfg.thresholds[a] = 0.0
        elif section == "SERIES":
            if series_header is None:
                series_header = [str(c).strip() for c in raw]
                continue
            vals = {h: ("" if i >= len(raw) or raw[i] is None else raw[i])
                    for i, h in enumerate(series_header)}
            cfg.series.append(SeriesSpec(
                series_id=str(vals.get("series_id", "")).strip(),
                title=str(vals.get("title", "")).strip(),
                category=str(vals.get("category", "")).strip(),
                lane=str(vals.get("lane", "")).strip().lower(),
                metric_type=str(vals.get("metric_type", "")).strip(),
                frequency=str(vals.get("frequency", "")).strip(),
                sa_nsa=str(vals.get("sa_nsa", "")).strip(),
                units=str(vals.get("units", "")).strip(),
                level_rate_index=str(vals.get("level_rate_index", "")).strip().lower(),
                geo_segment=str(vals.get("geo_segment", "")).strip(),
                dashboard_capable=_as_bool(vals.get("dashboard_capable", "")),
                watchlist_capable=_as_bool(vals.get("watchlist_capable", "")),
                transform=str(vals.get("transform", "level")).strip().lower(),
                alert_rule=str(vals.get("alert_rule", "none")).strip().lower(),
                notes=str(vals.get("notes", "")).strip(),
            ))
        elif section == "CBSA_EXTENSIONS":
            if cbsa_header is None:
                cbsa_header = [str(c).strip() for c in raw]
                continue
            cfg.cbsa_extensions.append(
                {h: ("" if i >= len(raw) else str(raw[i]).strip())
                 for i, h in enumerate(cbsa_header)})
    return cfg


# ---------------------------------------------------------------------------
# WATCHLIST VALIDATOR -- the hard gate (BUILD SPEC sec 0.1)
# ---------------------------------------------------------------------------
class WatchlistBoundaryError(Exception):
    """Raised when config tries to route a non-geographic series into the
    commercial watchlist lane. The danger is presenting a national trend as if
    it can point at loans -- this is the analog of a 'never plug' rule."""


# Geo keys that are NOT a portfolio-joinable location (national aggregates).
_NATIONAL_GEO = {"", "national", "us", "usa", "nation", "aggregate", "n/a", "none"}
# Categories whose series may legitimately be watchlist-capable.
_HPI_CATEGORIES = {"hpi_state", "hpi_metro", "hpi_caseshiller"}


def validate_watchlist(series: Sequence[SeriesSpec]) -> None:
    """Enforce: a series may be watchlist_capable ONLY if it carries a real
    geographic key a loan portfolio can join on -- i.e. a house-price index
    (FHFA state/metro, Case-Shiller metro). Charge-off, delinquency, G.19,
    DSR, SLOOS and CRE-price series are dashboard-only and must never reach
    the watchlist. Refuse with an error naming the offending series."""
    for s in series:
        if not s.watchlist_capable:
            continue
        geo = (s.geo_segment or "").strip().lower()
        reasons = []
        if geo in _NATIONAL_GEO:
            reasons.append("no geographic key (geo_segment is national/blank)")
        if s.lane != "price":
            reasons.append(f"lane is '{s.lane}', not 'price'")
        if s.category not in _HPI_CATEGORIES:
            reasons.append(
                f"category '{s.category}' is not a house-price index "
                f"({'/'.join(sorted(_HPI_CATEGORIES))})")
        if reasons:
            raise WatchlistBoundaryError(
                f"Series '{s.series_id}' ({s.title}) is marked watchlist_capable "
                f"but cannot localize a portfolio subset: " + "; ".join(reasons) +
                ". National credit-quality series are excluded by design.")


def watchlist_series(series: Sequence[SeriesSpec]) -> List[SeriesSpec]:
    """Series eligible to feed Watchlist_Geo (after validation)."""
    return [s for s in series if s.watchlist_capable and not s.is_dead]


# ---------------------------------------------------------------------------
# TRANSFORM REGISTRY -- pure, deterministic (BUILD SPEC sec 3)
# Each transform takes a date-indexed Series + its frequency and returns a
# derived Series of the same index. No transform invents data; missing stays
# missing. The builder maps each name to the matching Excel formula; the
# runner uses them for the Python-side alert cross-check.
# ---------------------------------------------------------------------------
def t_level(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64")


def t_yoy_pct(s: pd.Series, frequency: str) -> pd.Series:
    n = periods_per_year(frequency)
    return s.astype("float64").pct_change(periods=n) * 100.0


def t_qoq_pct(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64").pct_change(periods=1) * 100.0


def t_mom_pct(s: pd.Series, frequency: str) -> pd.Series:
    return s.astype("float64").pct_change(periods=1) * 100.0


def t_zscore_8q(s: pd.Series, frequency: str) -> pd.Series:
    """Z-score of each point vs. its trailing 8-period mean/std -- the dashboard
    alert primitive. Window is 8 by name; std uses sample (ddof=1)."""
    x = s.astype("float64")
    mean = x.rolling(window=8, min_periods=8).mean()
    std = x.rolling(window=8, min_periods=8).std(ddof=1)
    return (x - mean) / std.replace(0.0, math.nan)


def t_index_to_pct(s: pd.Series, frequency: str) -> pd.Series:
    """Convert an index series to YoY % change. Must NEVER be applied to a
    dollar-level series (e.g. the Z.1 CRE dollar levels) -- the validator and
    config guard against that."""
    return t_yoy_pct(s, frequency)


TRANSFORMS: Dict[str, Callable[[pd.Series, str], pd.Series]] = {
    "level": t_level,
    "yoy_pct": t_yoy_pct,
    "qoq_pct": t_qoq_pct,
    "mom_pct": t_mom_pct,
    "zscore_8q": t_zscore_8q,
    "index_to_pct": t_index_to_pct,
}

# Transforms that are only meaningful on an index (% of an index), never on a
# dollar level. Used by validate_transforms().
_INDEX_ONLY_TRANSFORMS = {"index_to_pct"}
_DOLLAR_LEVELS = {"level", "dollar", "dollars", "$"}


class TransformMisuseError(Exception):
    pass


def validate_transforms(series: Sequence[SeriesSpec]) -> None:
    """Guard: index_to_pct must not be wired onto a dollar-level series."""
    for s in series:
        if s.transform in _INDEX_ONLY_TRANSFORMS and s.level_rate_index in _DOLLAR_LEVELS:
            raise TransformMisuseError(
                f"Series '{s.series_id}' is a {s.level_rate_index} series but is "
                f"assigned transform '{s.transform}', which is index-only. Use "
                f"'level' or 'yoy_pct' for dollar-level series.")
    for s in series:
        if s.transform not in TRANSFORMS:
            raise TransformMisuseError(
                f"Series '{s.series_id}' references unknown transform '{s.transform}'.")


def apply_transform(spec: SeriesSpec, s: pd.Series) -> pd.Series:
    return TRANSFORMS[spec.transform](s, spec.frequency)


# ---------------------------------------------------------------------------
# THRESHOLD ENGINE -- config-driven (BUILD SPEC sec 3)
# ---------------------------------------------------------------------------
def latest_valid(s: pd.Series):
    s = s.dropna()
    return None if s.empty else float(s.iloc[-1])


def evaluate_alert(spec: SeriesSpec, s: pd.Series, cfg: Config) -> Optional[dict]:
    """Return an alert dict if the series breaches its configured band, else
    None. Mirrors the Excel flag so the runner can report a count."""
    if spec.alert_rule == "zscore":
        z = latest_valid(t_zscore_8q(s, spec.frequency))
        if z is not None and z >= cfg.zscore_band:
            return {"series_id": spec.series_id, "rule": "zscore", "value": round(z, 2),
                    "band": cfg.zscore_band}
    elif spec.alert_rule == "sloos_level":
        v = latest_valid(t_level(s, spec.frequency))
        if v is not None and v >= cfg.sloos_band:
            return {"series_id": spec.series_id, "rule": "sloos_level", "value": round(v, 1),
                    "band": cfg.sloos_band}
    return None


# ---------------------------------------------------------------------------
# PROVIDER ADAPTER -- the ONLY FRED-specific code (BUILD SPEC sec 0.3, sec 1)
# Swap this class to point the template at another source. Nothing downstream
# (transforms, validator, layout, backends) knows it is FRED.
# ---------------------------------------------------------------------------
def coerce_series(raw: pd.Series) -> pd.Series:
    """FRED returns '.' for missing observations -- coerce to NaN, NEVER to 0
    (BUILD SPEC sec 2). Keep the DatetimeIndex; do not resample."""
    s = raw.copy()
    s = s.replace(FRED_MISSING, pd.NA)
    s = pd.to_numeric(s, errors="coerce")
    try:
        s.index = pd.to_datetime(s.index)
    except (TypeError, ValueError):
        pass
    return s.sort_index()


class Provider:
    """Provider protocol. Implementations must return a tidy, date-indexed,
    NaN-for-missing float Series and a last-observation date."""

    def fetch(self, series_id: str) -> pd.Series:  # pragma: no cover - interface
        raise NotImplementedError

    def last_observation_date(self, series_id: str) -> Optional[date]:  # pragma: no cover
        raise NotImplementedError


class FredProvider(Provider):
    """FRED adapter via the `fredapi` library (BUILD SPEC sec 1)."""

    def __init__(self, api_key: str):
        from fredapi import Fred          # imported lazily so the module loads w/o it
        self._fred = Fred(api_key=api_key)

    def fetch(self, series_id: str) -> pd.Series:
        return coerce_series(self._fred.get_series(series_id))

    def last_observation_date(self, series_id: str) -> Optional[date]:
        try:
            info = self._fred.get_series_info(series_id)
            return pd.to_datetime(info["observation_end"]).date()
        except Exception:
            s = self.fetch(series_id).dropna()
            return None if s.empty else s.index.max().date()


class DemoProvider(Provider):
    """Deterministic synthetic provider for offline demos and the email-sim
    acceptance test -- NO network, NO key. Produces plausible, reproducible
    shapes per series so the workbook populates and dashboards light up.

    This is a DEV/DEMO convenience, not part of the data path: enable with
    --demo or demo_mode=TRUE in _config. It never touches FRED.
    """

    def __init__(self, asof: Optional[date] = None):
        self.asof = asof or date(2026, 3, 1)

    def _seed(self, series_id: str) -> int:
        return sum((i + 1) * ord(c) for i, c in enumerate(series_id)) % 997

    def fetch(self, series_id: str) -> pd.Series:
        seed = self._seed(series_id)
        # Deterministic pseudo-walk (no randomness -> reproducible).
        quarterly = not series_id.endswith(("SL", "NS", "SLAR"))  # G.19 monthly-ish
        n = 60
        freq = "Q" if quarterly else "M"
        idx = pd.period_range(end=pd.Period(self.asof, freq=freq[0]), periods=n, freq=freq[0])
        idx = idx.to_timestamp(how="end").normalize()
        base = 2.0 + (seed % 50) / 10.0
        vals = []
        v = base
        for i in range(n):
            wobble = math.sin((i + seed) / 5.0) * 0.15 + ((seed >> (i % 5)) & 1) * 0.05
            drift = 0.02 if i > n - 8 else 0.0     # gentle recent rise to trip some flags
            v = max(0.05, v + wobble * 0.2 + drift)
            vals.append(round(v, 3))
        s = pd.Series(vals, index=idx)
        s.iloc[seed % n] = math.nan               # exercise the missing-value path
        return s

    def last_observation_date(self, series_id: str) -> Optional[date]:
        return self.asof


def resolve_api_key(cfg: Config) -> Optional[str]:
    """Env var first, then the designated `_config` fallback cell. Never a
    hardcoded key (BUILD SPEC sec 1)."""
    key = os.environ.get("FRED_API_KEY", "").strip()
    if key:
        return key
    cell = str(cfg.setting("fred_api_key", "") or "").strip()
    if cell and cell.lower() not in ("", "<paste-your-fred-api-key-here>", "none"):
        return cell
    return None


# ---------------------------------------------------------------------------
# RAW LAYOUT -- fixed anchors shared with the builder
# ---------------------------------------------------------------------------
@dataclass
class RawBlock:
    series_id: str
    tab: str
    header_row: int          # 1-based row of the id/title banner
    label_row: int           # "date"/"value" header
    first_data_row: int      # newest observation row
    slots: int

    def data_cell(self, offset: int) -> str:
        """A1-style address of the value `offset` periods back from newest."""
        return f"{_col_letter(RAW_VALUE_COL)}{self.first_data_row + offset}"

    def date_cell(self, offset: int) -> str:
        return f"{_col_letter(RAW_DATE_COL)}{self.first_data_row + offset}"

    def value_range(self, count: int) -> str:
        c = _col_letter(RAW_VALUE_COL)
        return f"{c}{self.first_data_row}:{c}{self.first_data_row + count - 1}"


def _col_letter(idx: int) -> str:
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def raw_layout(series: Sequence[SeriesSpec], slots: int = RAW_SLOTS_DEFAULT) -> Dict[str, RawBlock]:
    """Deterministic block placement: same input -> same anchors, every run."""
    stride = RAW_HEADER_ROWS + slots + RAW_GAP_ROWS
    per_tab_index: Dict[str, int] = {}
    blocks: Dict[str, RawBlock] = {}
    for s in series:
        tab = LANE_RAW_TAB.get(s.lane, "Raw_Price")
        i = per_tab_index.get(tab, 0)
        per_tab_index[tab] = i + 1
        header_row = RAW_FIRST_ROW + i * stride
        blocks[s.series_id] = RawBlock(
            series_id=s.series_id, tab=tab, header_row=header_row,
            label_row=header_row + 1, first_data_row=header_row + RAW_HEADER_ROWS,
            slots=slots)
    return blocks


# ---------------------------------------------------------------------------
# WRITE BACKENDS -- presentation never sees these (BUILD SPEC sec 5/6)
# ---------------------------------------------------------------------------
class Backend:
    def read_config(self) -> Config:                      # pragma: no cover
        raise NotImplementedError

    def write_raw_block(self, block: RawBlock, spec: SeriesSpec, s: pd.Series):  # pragma: no cover
        raise NotImplementedError

    def write_status(self, status: dict):                 # pragma: no cover
        raise NotImplementedError

    def finalize(self):                                   # pragma: no cover
        pass


def _series_newest_first(s: pd.Series, slots: int):
    s = s.sort_index()
    tail = s.iloc[-slots:] if len(s) > slots else s
    pairs = list(tail.items())[::-1]      # newest first
    return pairs


def _status_lines(status: dict):
    """Two compact masthead lines for the dashboard status readout."""
    ts = status.get("timestamp", "")
    pulled = status.get("series_pulled", 0)
    total = status.get("series_in_dict", 0)
    stale = len(status.get("stale", []))
    alerts = status.get("alert_count", 0)
    line1 = f"Last run  {ts}"
    line2 = (f"Pulled {pulled}/{total} · {stale} stale · "
             f"{alerts} alert{'' if alerts == 1 else 's'}")
    return line1, line2


class OpenpyxlBackend(Backend):
    """Writes the closed workbook file in place (portable, fewer deps).
    Preserves the embedded VBA project via keep_vba=True."""

    def __init__(self, path: str):
        self.path = path
        import openpyxl
        self._wb = openpyxl.load_workbook(path, keep_vba=True)

    def read_config(self) -> Config:
        ws = self._wb["_config"]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        return parse_config(rows)

    def write_raw_block(self, block: RawBlock, spec: SeriesSpec, s: pd.Series):
        ws = self._wb[block.tab]
        ws.cell(block.header_row, RAW_DATE_COL, spec.series_id)
        ws.cell(block.header_row, RAW_VALUE_COL, spec.title)
        ws.cell(block.header_row, RAW_VALUE_COL + 1, f"freq={spec.frequency}; transform={spec.transform}")
        ws.cell(block.label_row, RAW_DATE_COL, "date")
        ws.cell(block.label_row, RAW_VALUE_COL, "value")
        # Clear the slot region first (stateless refresh).
        for r in range(block.first_data_row, block.first_data_row + block.slots):
            ws.cell(r, RAW_DATE_COL, None)
            ws.cell(r, RAW_VALUE_COL, None)
        for i, (idx, val) in enumerate(_series_newest_first(s, block.slots)):
            r = block.first_data_row + i
            ws.cell(r, RAW_DATE_COL, pd.Timestamp(idx).date().isoformat())
            ws.cell(r, RAW_VALUE_COL, None if pd.isna(val) else float(val))

    def write_status(self, status: dict):
        line1, line2 = _status_lines(status)
        for tab in ("Dashboard_Consumer", "Dashboard_Commercial", "Dashboard_Price"):
            if tab in self._wb.sheetnames:
                ws = self._wb[tab]
                ws.cell(1, STATUS_COL, line1)
                ws.cell(2, STATUS_COL, line2)

    def finalize(self):
        self._wb.save(self.path)


class XlwingsBackend(Backend):
    """Writes into the ALREADY-OPEN workbook (smoothest one-click feel) --
    the recommended path on the work machine (BUILD SPEC sec 5/6)."""

    def __init__(self, path: str):
        import xlwings as xw
        self.path = path
        self._app = None
        target = os.path.basename(path)
        for bk in xw.books:
            if bk.name.lower() == target.lower():
                self._book = bk
                break
        else:
            self._book = xw.Book(path)

    def read_config(self) -> Config:
        sht = self._book.sheets["_config"]
        rows = sht.used_range.value or []
        if rows and not isinstance(rows[0], list):
            rows = [rows]
        return parse_config(rows)

    def write_raw_block(self, block: RawBlock, spec: SeriesSpec, s: pd.Series):
        sht = self._book.sheets[block.tab]
        sht.range((block.header_row, RAW_DATE_COL)).value = spec.series_id
        sht.range((block.header_row, RAW_VALUE_COL)).value = spec.title
        sht.range((block.header_row, RAW_VALUE_COL + 1)).value = (
            f"freq={spec.frequency}; transform={spec.transform}")
        sht.range((block.label_row, RAW_DATE_COL)).value = "date"
        sht.range((block.label_row, RAW_VALUE_COL)).value = "value"
        clear_rng = sht.range((block.first_data_row, RAW_DATE_COL),
                              (block.first_data_row + block.slots - 1, RAW_VALUE_COL))
        clear_rng.clear_contents()
        data = []
        for idx, val in _series_newest_first(s, block.slots):
            data.append([pd.Timestamp(idx).date().isoformat(),
                         None if pd.isna(val) else float(val)])
        if data:
            sht.range((block.first_data_row, RAW_DATE_COL)).value = data

    def write_status(self, status: dict):
        line1, line2 = _status_lines(status)
        for tab in ("Dashboard_Consumer", "Dashboard_Commercial", "Dashboard_Price"):
            try:
                sht = self._book.sheets[tab]
            except Exception:
                continue
            sht.range((1, STATUS_COL)).value = line1
            sht.range((2, STATUS_COL)).value = line2

    def finalize(self):
        self._book.save()


def make_backend(name: str, path: str) -> Backend:
    name = (name or "auto").lower()
    if name in ("auto", "xlwings"):
        try:
            return XlwingsBackend(path)
        except Exception as exc:
            if name == "xlwings":
                raise
            sys.stderr.write(f"[info] xlwings unavailable ({exc}); using openpyxl backend.\n")
    return OpenpyxlBackend(path)


# ---------------------------------------------------------------------------
# STALE-SERIES CHECK (BUILD SPEC sec 2)
# ---------------------------------------------------------------------------
def is_stale(last_obs: Optional[date], frequency: str, asof: date, multiplier: float) -> bool:
    if last_obs is None:
        return True
    days_per = {"quarterly": 92, "monthly": 31, "annual": 366, "weekly": 7}[_freq_key(frequency)]
    return (asof - last_obs).days > days_per * multiplier


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------
def run(workbook_path: str, backend_name: str = "auto", demo: bool = False,
        asof: Optional[date] = None) -> dict:
    asof = asof or date.today()
    backend = make_backend(backend_name, workbook_path)
    cfg = backend.read_config()

    # Hard gates BEFORE any pull (BUILD SPEC sec 0.1, sec 3).
    validate_watchlist(cfg.series)
    validate_transforms(cfg.series)

    if demo or _as_bool(cfg.setting("demo_mode", "false")):
        provider: Provider = DemoProvider(asof=asof)
        mode = "demo"
    else:
        key = resolve_api_key(cfg)
        if not key:
            raise SystemExit(
                "FRED API key not found. Set the FRED_API_KEY environment "
                "variable or paste your key into the _config tab "
                "(SETTINGS -> fred_api_key). Get a free key at "
                "https://fredaccount.stlouisfed.org/apikeys")
        provider = FredProvider(key)
        mode = "live"

    blocks = raw_layout(cfg.series, slots=cfg.raw_slots)
    pulled, stale, alerts, errors = 0, [], [], []
    pullable = [s for s in cfg.series if not s.is_dead]
    for spec in pullable:
        try:
            s = provider.fetch(spec.series_id)
        except Exception as exc:
            errors.append(f"{spec.series_id}: {exc}")
            continue
        block = blocks[spec.series_id]
        backend.write_raw_block(block, spec, s)
        pulled += 1
        last = provider.last_observation_date(spec.series_id)
        if is_stale(last, spec.frequency, asof, cfg.stale_multiplier):
            stale.append(spec.series_id)
        if spec.alert_rule in ("zscore", "sloos_level"):
            a = evaluate_alert(spec, s, cfg)
            if a:
                alerts.append(a)

    status = {
        "timestamp": asof.isoformat() if asof else "",
        "mode": mode,
        "series_in_dict": len(cfg.series),
        "series_pulled": pulled,
        "alert_count": len(alerts),
        "alerts": alerts[:25],
        "stale": stale,
        "errors": errors[:25],
    }
    backend.write_status(status)
    backend.finalize()
    return status


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="FRED Credit-Risk Dashboard runner")
    ap.add_argument("--workbook", "-w", required=True, help="path to the .xlsm")
    ap.add_argument("--backend", default="auto", choices=["auto", "xlwings", "openpyxl"])
    ap.add_argument("--demo", action="store_true",
                    help="use the offline synthetic provider (no FRED key/network)")
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD as-of date (testing)")
    args = ap.parse_args(argv)
    asof = datetime.strptime(args.asof, "%Y-%m-%d").date() if args.asof else None
    try:
        status = run(args.workbook, backend_name=args.backend, demo=args.demo, asof=asof)
    except (WatchlistBoundaryError, TransformMisuseError) as exc:
        sys.stderr.write(f"BOUNDARY ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 2
    except SystemExit as exc:
        sys.stderr.write(f"{exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 3
    except Exception as exc:
        sys.stderr.write(f"RUN ERROR: {exc}\n")
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1
    print(json.dumps({"ok": True, **status}))
    sys.stderr.write(
        f"OK: {status['series_pulled']}/{status['series_in_dict']} series, "
        f"{status['alert_count']} alerts, {len(status['stale'])} stale.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
