"""Integration tests: build pipeline, VBA embedding, refresh round-trip.

Run: python3 -m pytest tests/ -q
"""
import os
import sys
import zipfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import assemble_xlsm
import build_workbook
import runner as R
import vba_writer
from oletools.olevba import VBA_Parser, decompress_stream

EXPECTED_TABS = {
    "Dashboard_Consumer", "Dashboard_Commercial", "Dashboard_Price",
    "Watchlist_Geo", "Raw_Consumer", "Raw_Commercial", "Raw_Price",
    "_config", "_code_py", "_code_vba", "_readme",
}


# --------------------------------------------------------------------------
# MS-OVBA compression round-trips through the real decompressor at all sizes,
# including the 3641-4095 remainder edge case that broke the raw-only encoder.
# --------------------------------------------------------------------------
@pytest.mark.parametrize("size", [0, 1, 8, 100, 4096, 4097, 3640, 3641, 4095,
                                  7796, 8192, 12000])
def test_compress_roundtrip(size):
    data = bytes((i * 37 + 11) % 256 for i in range(size))
    comp = vba_writer.compress(data)
    assert comp[0] == 0x01
    back = bytes(decompress_stream(bytearray(comp)))
    assert back == data


# --------------------------------------------------------------------------
# Build -> assemble produces a valid, macro-enabled, self-contained .xlsm.
# --------------------------------------------------------------------------
@pytest.fixture(scope="module")
def built(tmp_path_factory):
    d = tmp_path_factory.mktemp("wb")
    base = str(d / "base.xlsx")
    xlsm = str(d / "FRED.xlsm")
    _, n, nwl = build_workbook.build(base)
    assemble_xlsm.assemble(base, xlsm)
    return {"base": base, "xlsm": xlsm, "n": n, "nwl": nwl}


def test_base_has_all_tabs(built):
    wb = openpyxl.load_workbook(built["base"])
    assert EXPECTED_TABS.issubset(set(wb.sheetnames))


def test_counts(built):
    assert built["n"] == 147
    assert built["nwl"] == 89


def test_xlsm_is_macro_enabled(built):
    z = zipfile.ZipFile(built["xlsm"])
    ct = z.read("[Content_Types].xml").decode()
    assert "sheet.macroEnabled.main+xml" in ct
    assert "vbaProject" in ct
    assert "xl/vbaProject.bin" in z.namelist()
    rels = z.read("xl/_rels/workbook.xml.rels").decode()
    assert "vbaProject.bin" in rels


def test_xlsm_macro_extracts(built):
    vp = VBA_Parser(built["xlsm"])
    assert vp.detect_vba_macros()
    mods = [(vn, code) for (_, _, vn, code) in vp.extract_macros()]
    vp.close()
    assert any("FREDDashboard" in vn for vn, _ in mods)
    assert any("ExtractAndRun" in code for _, code in mods)


def test_code_tab_roundtrips_to_runner(built):
    """The _code_py tab must reproduce runner.py faithfully (workbook = source
    of truth)."""
    wb = openpyxl.load_workbook(built["xlsm"])
    ws = wb["_code_py"]
    lines = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    rebuilt = "\n".join("" if v is None else str(v) for v in lines)
    original = open(os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                 "runner.py"), encoding="utf-8").read().rstrip("\n")
    assert rebuilt.rstrip("\n") == original


# --------------------------------------------------------------------------
# A demo refresh fills raw data AND preserves the macro (stateless rebuild,
# self-contained across runs).
# --------------------------------------------------------------------------
def test_refresh_preserves_macro_and_fills_data(built, tmp_path):
    import shutil
    work = str(tmp_path / "work.xlsm")
    shutil.copy(built["xlsm"], work)
    status = R.run(work, backend_name="openpyxl", demo=True,
                   asof=__import__("datetime").date(2026, 3, 1))
    assert status["series_pulled"] >= 140
    assert status["alert_count"] >= 1
    wb = openpyxl.load_workbook(work, keep_vba=True)
    assert wb.vba_archive is not None                      # macro survived
    assert wb["Raw_Consumer"]["B4"].value is not None      # data landed
    vp = VBA_Parser(work)
    assert vp.detect_vba_macros()
    vp.close()


def test_validator_blocks_bad_config_at_build():
    """If config marks a delinquency series watchlist-capable, build must refuse
    (the hard gate, enforced at build time too)."""
    specs = R.parse_config([["[SERIES]"], __import__("series_seed").HEADER]
                           + [[r[h] for h in __import__("series_seed").HEADER]
                              for r in __import__("series_seed").all_series()]).series
    # corrupt one consumer series
    for s in specs:
        if s.series_id == "DRCCLACBS":
            s.watchlist_capable = True
    with pytest.raises(R.WatchlistBoundaryError):
        R.validate_watchlist(specs)


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))
