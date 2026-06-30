#!/usr/bin/env python3
"""Email-simulate the acceptance test (BUILD SPEC phase 5).

Copies ONLY the .xlsm into a fresh, otherwise-empty folder, then reproduces
exactly what the VBA "Extract & Run" button does:
  1. read the _code_py tab (one source line per cell, column A) and write
     runner.py next to the workbook -- byte-for-byte the macro's WriteTabToFile;
  2. shell Python to run runner.py against the workbook (demo mode, so no key /
     network is needed for the simulation).
Then it asserts the workbook rebuilt itself with no external files present --
proving the workbook is the source of truth.
"""
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl


def extract_code_tab(xlsm_path, tab, out_path):
    """Mirror the VBA extractor: join column-A cells with LF, write to file."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True)
    ws = wb[tab]
    last = ws.max_row
    lines = []
    for r in range(1, last + 1):
        v = ws.cell(r, 1).value
        lines.append("" if v is None else str(v))
    wb.close()
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def main():
    repo = os.path.dirname(os.path.abspath(__file__))
    xlsm = os.path.join(repo, "FRED_Credit_Risk_Dashboard.xlsm")
    if not os.path.exists(xlsm):
        print("build the workbook first: python3 make_workbook.py", file=sys.stderr)
        return 1

    work = tempfile.mkdtemp(prefix="fred_emailsim_")
    try:
        # 1. The email contains ONLY the workbook.
        dst_xlsm = os.path.join(work, "FRED_Credit_Risk_Dashboard.xlsm")
        shutil.copy(xlsm, dst_xlsm)
        before = sorted(os.listdir(work))
        assert before == ["FRED_Credit_Risk_Dashboard.xlsm"], before
        print(f"[email-sim] fresh folder contains only: {before}")

        # 2. Button step 1: extract runner.py from _code_py.
        runner_py = os.path.join(work, "runner.py")
        extract_code_tab(dst_xlsm, "_code_py", runner_py)
        print(f"[email-sim] extracted runner.py ({os.path.getsize(runner_py)} bytes) from _code_py")

        # 3. Button step 2: shell Python to run it (demo mode for the sim).
        proc = subprocess.run(
            [sys.executable, runner_py, "--workbook", dst_xlsm,
             "--backend", "openpyxl", "--demo", "--asof", "2026-03-01"],
            capture_output=True, text=True, cwd=work)
        print("[email-sim] runner stderr:", proc.stderr.strip().splitlines()[-1]
              if proc.stderr.strip() else "(none)")
        if proc.returncode != 0:
            print("[email-sim] FAILED: runner exited", proc.returncode)
            print(proc.stdout, proc.stderr)
            return 2

        # 4. Verify the rebuild: raw data present, formulas/macro intact, and
        #    nothing but the workbook + extracted runner exists.
        wb = openpyxl.load_workbook(dst_xlsm, keep_vba=True)
        raw_ok = wb["Raw_Consumer"]["B4"].value is not None
        wl_ok = wb["Watchlist_Geo"]["B5"].value is not None
        macro_ok = wb.vba_archive is not None
        produced = sorted(os.listdir(work))
        print(f"[email-sim] folder now: {produced}")
        print(f"[email-sim] raw populated={raw_ok}  watchlist={wl_ok}  macro intact={macro_ok}")

        ok = (raw_ok and wl_ok and macro_ok
              and set(produced) == {"FRED_Credit_Risk_Dashboard.xlsm", "runner.py"})
        print("[email-sim] RESULT:", "PASS -- workbook is the source of truth" if ok else "FAIL")
        return 0 if ok else 3
    finally:
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
