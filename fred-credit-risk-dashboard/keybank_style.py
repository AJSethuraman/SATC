"""KeyBank Excel design system — the single source of style for our workbooks.

Drop-in for openpyxl builders (e.g. build_workbook.py). Import the tokens and
helpers from here instead of hand-coding fills/fonts, so every workbook we ship
— this FRED dashboard and any future template — looks like one brand.

    from keybank_style import (
        INK, ONYX, KEY_RED, CRIMSON, CANVAS, MIST, SLATE,
        HDR_FILL, SECT_FILL, WHITE_BOLD, TITLE_FONT, DATA_FONT,
        brand_banner, kpi_tiles, section_band, header_row,
        zscore_heat, alert_rule, tighten_rule, watchlist_boundary,
        hide_gridlines, freeze_below,
    )

Design rule of thumb: BLACK grounds, RED leads sparingly, NEUTRALS breathe.
Only one red accent rule per banner; red elsewhere means "look here" (alert).
Fonts are the universally-installed Excel set — the brand lives in colour and
structure, never in a font we'd have to ship.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# ===========================================================================
# TOKENS  (hex WITHOUT the leading '#', the way openpyxl wants them)
# ===========================================================================
# -- brand core
INK      = "0A0908"   # header bands, primary surfaces on dark
ONYX     = "16130F"   # system-tab (_config/_code/_readme) section bands
KEY_RED  = "CC0000"   # the accent rule, alerts, KPI accent — leads sparingly
CRIMSON  = "960019"   # alert TEXT / pressed red
# -- surfaces
PAPER    = "FFFFFF"   # sheet background
CANVAS   = "F4F1EC"   # KPI tiles, sub-header fills
MIST     = "E4DFD5"   # caution fill, dividers
STONE    = "B9B4AC"   # disabled, faint rules
# -- text
SLATE    = "57534B"   # secondary text / labels
INK_TEXT = "16130F"   # primary data text (warm black, never pure 000)
# -- semantic + heat (brand-muted — NOT openpyxl's default 63BE7B / F8696B)
ALERT_FG   = "F7DEDE"  # deterioration fill
POSITIVE   = "1E7A47"  # improvement text
HEAT_BAD   = "E0A6A6"  # z high  -> stress
HEAT_WARM  = "EFD7CF"  # z mid-high
HEAT_MID   = "F4F1EC"  # z ~ 0   (== CANVAS)
HEAT_GOOD  = "BBD3BD"  # z low   -> calm

PALETTE = {  # handy for tests / docs / a legend tab
    "INK": INK, "ONYX": ONYX, "KEY_RED": KEY_RED, "CRIMSON": CRIMSON,
    "PAPER": PAPER, "CANVAS": CANVAS, "MIST": MIST, "STONE": STONE,
    "SLATE": SLATE, "ALERT_FG": ALERT_FG, "POSITIVE": POSITIVE,
    "HEAT_BAD": HEAT_BAD, "HEAT_WARM": HEAT_WARM, "HEAT_MID": HEAT_MID,
    "HEAT_GOOD": HEAT_GOOD,
}

# ===========================================================================
# FONTS  (Excel-installed only:  Arial = display,  Calibri = data,  Consolas = code)
# ===========================================================================
TITLE_FONT   = Font(name="Arial",   bold=True,  size=16, color=PAPER)    # banner title
SUBTITLE_FONT= Font(name="Calibri", size=11,    color="B9B4AC")          # banner subtitle (on dark)
WHITE_BOLD   = Font(name="Arial",   bold=True,  size=11, color=PAPER)    # header row
KPI_LABEL    = Font(name="Arial",   bold=True,  size=9,  color=SLATE)
KPI_NUMBER   = Font(name="Arial",   bold=True,  size=20, color=INK)
DATA_FONT    = Font(name="Calibri", size=11,    color=INK_TEXT)
SECONDARY    = Font(name="Calibri", size=11,    color=SLATE)
NOTE_FONT    = Font(name="Calibri", italic=True, size=9, color=SLATE)
MONO_FONT    = Font(name="Consolas", size=9,    color=INK_TEXT)
SECTION_FONT = Font(name="Arial",   bold=True,  size=11, color=PAPER)    # _config [SECTION] bands
ALERT_FONT   = Font(name="Arial",   bold=True,  size=11, color=CRIMSON)

# ===========================================================================
# FILLS / BORDERS
# ===========================================================================
HDR_FILL     = PatternFill("solid", fgColor=INK)      # dashboard/watchlist header band
SECT_FILL    = PatternFill("solid", fgColor=ONYX)     # _config section bands
KPI_FILL     = PatternFill("solid", fgColor=CANVAS)
SUBHDR_FILL  = PatternFill("solid", fgColor=CANVAS)
ALERT_FILL   = PatternFill("solid", fgColor=ALERT_FG)
TIGHTEN_FILL = PatternFill("solid", fgColor=MIST)
POS_FILL     = PatternFill("solid", fgColor="E4EFE7")

_thin = Side(style="thin", color="E7E2D8")
HAIRLINE  = Border(bottom=_thin)
RED_TOP   = Border(top=Side(style="medium", color=KEY_RED))   # KPI tile accent
RED_RULE  = Side(style="medium", color=KEY_RED)               # under banners

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
WRAP   = Alignment(wrap_text=True, vertical="top")

# Number formats used across the dashboards.
FMT_RATE = "0.00"      # Latest / Prior
FMT_YOY  = "0.0"       # YoY %, z-score uses 0.00
FMT_Z    = "0.00"


# ===========================================================================
# COMPOSED HELPERS  — call these from the builder; they encode the rules.
# ===========================================================================
def hide_gridlines(ws):
    """Every styled tab hides interior gridlines; borders only where designed."""
    ws.sheet_view.showGridLines = False


def freeze_below(ws, header_row_idx, first_col=1):
    """Freeze panes just under the column-header row so it stays on scroll."""
    ws.freeze_panes = ws.cell(header_row_idx + 1, first_col)


def brand_banner(ws, row, ncols, title, subtitle, height=46):
    """Black banner with the red accent rule: the signature of every dashboard.

    Insert the key-mark PNG over column A separately (ws.add_image) if desired.
    """
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row, 1, title)
    c.fill = HDR_FILL
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # subtitle on the next row, still on the black band
    ws.merge_cells(f"A{row+1}:{last}{row+1}")
    s = ws.cell(row + 1, 1, subtitle)
    s.fill = HDR_FILL
    s.font = SUBTITLE_FONT
    s.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    # the single red accent rule lives on the bottom edge of the band
    for col in range(1, ncols + 1):
        cell = ws.cell(row + 1, col)
        cell.border = Border(bottom=RED_RULE)
    ws.row_dimensions[row].height = height * 0.6
    ws.row_dimensions[row + 1].height = height * 0.4
    return row + 2


def kpi_tiles(ws, row, tiles, span=3, accents=None):
    """A strip of merged-cell KPI tiles: cream fill + 3pt KEY_RED top border.

    tiles : list of (LABEL, big_value, sub_caption)
    span  : columns each tile occupies
    accents: optional list of hex (defaults to KEY_RED for all)
    Returns the next free row.
    """
    accents = accents or [KEY_RED] * len(tiles)
    col = 1
    for (label, value, sub), accent in zip(tiles, accents):
        last = get_column_letter(col + span - 1)
        # three stacked merged ranges: label / value / caption
        for r, text, font in ((row, label, KPI_LABEL),
                              (row + 1, value, Font(name="Arial", bold=True, size=20, color=INK)),
                              (row + 2, sub, SECONDARY)):
            ws.merge_cells(f"{get_column_letter(col)}{r}:{last}{r}")
            cell = ws.cell(r, col, text)
            cell.fill = KPI_FILL
            cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # red accent on the tile's top edge
        for cc in range(col, col + span):
            top = ws.cell(row, cc)
            top.border = Border(top=Side(style="medium", color=accent))
        col += span
    ws.row_dimensions[row + 1].height = 26
    return row + 3


def header_row(ws, row, columns, right_from=None, center_cols=()):
    """The Ink (black) column-header band with white Arial-bold labels.

    columns    : list of header strings
    right_from : index (0-based) at/after which numeric headers are right-aligned
    center_cols: indices to centre (e.g. Flag, Trend)
    """
    for i, name in enumerate(columns):
        cell = ws.cell(row, i + 1, name)
        cell.fill = HDR_FILL
        cell.font = WHITE_BOLD
        if i in center_cols:
            cell.alignment = CENTER
        elif right_from is not None and i >= right_from:
            cell.alignment = RIGHT
        else:
            cell.alignment = LEFT
        cell.border = Border(bottom=Side(style="thin", color=ONYX))
    ws.row_dimensions[row].height = 22
    return row + 1


def section_band(ws, row, label, ncols=3):
    """Onyx section band for the quiet system tabs (_config, _code, _readme)."""
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    cell = ws.cell(row, 1, label)
    cell.fill = SECT_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return row + 1


def watchlist_boundary(ws, row, ncols, text):
    """The red geographic-boundary gate banner — the one place red text leads."""
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    cell = ws.cell(row, 1, text)
    cell.fill = ALERT_FILL
    cell.font = ALERT_FONT
    cell.alignment = WRAP
    ws.row_dimensions[row].height = 42
    return row + 1


# --- conditional formatting -------------------------------------------------
def zscore_heat(ws, cell_range):
    """Brand-muted diverging heat for a z-score column (calm -> stress)."""
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="num", start_value=-2, start_color=HEAT_GOOD,
        mid_type="num",   mid_value=0,   mid_color=HEAT_MID,
        end_type="num",   end_value=2,   end_color=HEAT_BAD))


def yoy_heat(ws, cell_range):
    """Diverging heat for a YoY % column: red = deterioration, green = growth."""
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="num", start_value=-10, start_color=HEAT_BAD,
        mid_type="num",   mid_value=0,    mid_color=PAPER,
        end_type="num",   end_value=10,   end_color=HEAT_GOOD))


def alert_rule(ws, flag_range):
    """Highlight any flag cell whose text contains ALERT."""
    first = flag_range.split(":")[0]
    ws.conditional_formatting.add(flag_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("ALERT",{first}))'],
        fill=ALERT_FILL, font=ALERT_FONT))


def tighten_rule(ws, flag_range):
    """Quieter neutral highlight for TIGHTENING flags (no amber — off-brand)."""
    first = flag_range.split(":")[0]
    ws.conditional_formatting.add(flag_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("TIGHTENING",{first}))'],
        fill=TIGHTEN_FILL, font=Font(name="Arial", bold=True, color=SLATE)))
