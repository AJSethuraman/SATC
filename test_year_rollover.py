#!/usr/bin/env python3
"""Tests for the year-rollover tool. Standard library only."""

from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date
from pathlib import Path

import year_rollover as yr

try:
    import fee_workbook
    from openpyxl import load_workbook

    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover - depends on environment
    HAVE_OPENPYXL = False


class RollForwardTests(unittest.TestCase):
    def test_static_fields_carry_per_year_fields_reset(self) -> None:
        client = {
            "client_name": "Jo Sample", "email": "jo@x.com", "filing_status": "Single",
            "expected_documents": ["W-2"], "services": ["state_return"], "spouse_name": "Pat",
            "tax_year": "2024", "total": "440.00", "line_items": [{"x": 1}], "amount_paid": "440.00",
            "form_8879_signed": True, "engagement_letter_signed": "2025-02-01", "return_filed": "2025-03-01",
        }
        client["subtotal"] = "440.00"
        client["discount"] = "-40.00"
        client["discount_lines"] = [{"description": "Express", "amount": "-40.00"}]
        client["express_applied"] = True
        client["returns"] = [{"return_type": "Federal Income Tax"}]
        client["efiled_returns"] = [{"name": "Federal Income Tax"}]
        carried = yr.roll_forward(client, "2025")
        # static carries
        self.assertEqual(carried["client_name"], "Jo Sample")
        self.assertEqual(carried["expected_documents"], ["W-2"])
        self.assertEqual(carried["spouse_name"], "Pat")  # custom field preserved
        # year bumped
        self.assertEqual(carried["tax_year"], "2025")
        # all per-year status/results wiped (incl. discount + filed returns)
        for gone in ("total", "line_items", "subtotal", "discount", "discount_lines",
                     "express_applied", "amount_paid", "form_8879_signed",
                     "engagement_letter_signed", "return_filed", "returns", "efiled_returns"):
            self.assertNotIn(gone, carried)

    def test_next_year_from_clients(self) -> None:
        self.assertEqual(yr.next_year([{"tax_year": "2024"}, {"tax_year": "2023"}]), "2025")
        self.assertEqual(yr.next_year([{"tax_year": "2024"}], explicit="2030"), "2030")
        self.assertEqual(yr.next_year([{}]), str(date.today().year))


class RunTests(unittest.TestCase):
    def test_no_data_file(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            self.assertIn("No clients", yr.run_rollover(Path(d))["summary"])

    def test_creates_next_year_folder_with_config(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            folder = Path(d)
            (folder / "clients.json").write_text(json.dumps([
                {"client_name": "Jo Sample", "tax_year": "2024", "total": "100.00", "paid": True},
                {"client_name": "Riley Carter", "tax_year": "2024"},
            ]), encoding="utf-8")
            (folder / "firm.json").write_text(json.dumps({"firm_name": "Acme"}), encoding="utf-8")

            result = yr.run_rollover(folder)
            self.assertEqual(result["new_year"], "2025")
            self.assertEqual(result["client_count"], 2)

            target = folder / "2025"
            self.assertTrue((target / "firm.json").exists())  # config copied
            carried = json.loads((target / "clients.json").read_text())
            self.assertEqual({c["tax_year"] for c in carried}, {"2025"})
            self.assertTrue(all("total" not in c and "paid" not in c for c in carried))

            # Re-running does not duplicate clients.
            again = yr.run_rollover(folder)
            self.assertEqual(again["client_count"], 0)

    @unittest.skipUnless(HAVE_OPENPYXL, "openpyxl not installed")
    def test_rollover_applies_workbook_year_prices(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            folder = Path(d)
            (folder / "clients.json").write_text(
                json.dumps([{"client_name": "A", "tax_year": "2024"}]), encoding="utf-8"
            )
            fee_workbook.run_fee_workbook(folder, year=2024)  # creates 2024 + 2025 sheets
            workbook = load_workbook(folder / "fee_schedule.xlsx")
            for row in workbook["2025"].iter_rows():
                if row[0].value == "base_1040":
                    row[2].value = 185.0  # reprice 2025
            workbook.save(folder / "fee_schedule.xlsx")

            result = yr.run_rollover(folder, new_year="2025")
            self.assertTrue(result["workbook_applied"])
            schedule = json.loads((folder / "2025" / "fee_schedule.json").read_text())
            self.assertEqual(schedule["base_1040"]["price"], 185.0)  # 2025 prices, not 2024's


if __name__ == "__main__":
    unittest.main()
