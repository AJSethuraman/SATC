"""Mode B workbook assembly: Cover, PS_ sheets, Products roll-up, Data Mart,
Findings, ``_config`` — reusing the Mode A machinery wherever the shapes align
(Findings register + aggregates, knob-panel style, mart flat grid, ``_map``).
"""

from __future__ import annotations

import json
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import Engagement, Program
from credit_review.config_mode_b import ProductSample
from credit_review.findings import FindingsRefs, build_findings
from credit_review.linesheet import ExceptionRef, FMT_NUM, FMT_USD
from credit_review.mart import mart_id
from credit_review.methodology import (
    PII_NOTES_MODE_B,
    build_methodology_mode_b,
    build_readme,
)
from credit_review.product_sheet import FMT_PCT1, ProductRefs, build_product_sheet

_INK_BOLD = Font(name="Arial", bold=True, size=11, color=KB.INK)


@dataclass(frozen=True)
class _ProductHandle:
    """Duck-types config.Loan for the Findings register (id + sheet name)."""

    loan_id: str
    sheet_name: str


# ---------------------------------------------------------------------------
# _config (Mode B knob panel): thresholds + per-class tolerances
# ---------------------------------------------------------------------------
def write_config_sheet_mode_b(ws: Worksheet, program: Program,
                              engagement: Engagement) -> dict[str, str]:
    KB.hide_gridlines(ws)
    rows: list[list] = [
        ["Credit Review OS -- CONFIG (the knob panel). Policy limits, fringe "
         "bands, and tolerances here drive the PS_ sheet formulas."],
        [],
        ["[ENGAGEMENT]"],
        ["key", "value", "help"],
        ["client_name", engagement.client_name, "Client bank under review."],
        ["engagement_id", engagement.engagement_id, "Unique engagement identifier."],
        ["review_as_of", engagement.review_as_of, "Review as-of date."],
        ["lob", program.lob, "Line of business this program covers."],
        ["review_mode", program.review_mode,
         "product_conformance: sampled conformance testing + URCCP pool "
         "classification (Mode B)."],
        [],
        ["[THRESHOLDS]"],
        ["key", "value", "help"],
    ]
    registry_row: dict[str, int] = {}
    for key in sorted(engagement.thresholds):
        rows.append([key, engagement.thresholds[key],
                     "Policy limit / fringe band; [POL] formulas point here."])
        registry_row[key] = len(rows)
    rows += [[], ["[TOLERANCES]"], ["key", "value", "help"]]
    for cls in sorted(engagement.tolerances):
        key = f"tol_{cls}"
        rows.append([key, engagement.tolerances[cls],
                     "Max acceptable exception rate for this test class "
                     "(compliance flags per-occurrence regardless)."])
        registry_row[key] = len(rows)

    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if row and row[0] == "review_as_of":
            ws.cell(i, 2).number_format = "mm/dd/yyyy"
        if row and isinstance(row[0], str) and row[0].startswith("["):
            band = ws.cell(i, 1)
            band.font = KB.SECTION_FONT
            band.fill = KB.SECT_FILL
    ws["A1"].font = _INK_BOLD
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72
    return {key: f"_config!$B${r}" for key, r in registry_row.items()}


# ---------------------------------------------------------------------------
# Cover (Mode B): engagement identity + sample plan evidence
# ---------------------------------------------------------------------------
def build_cover_mode_b(ws: Worksheet, program: Program, engagement: Engagement,
                       samples: dict[str, ProductSample]) -> None:
    KB.hide_gridlines(ws)
    for letter, width in {"A": 2, "B": 36, "C": 64, "D": 10}.items():
        ws.column_dimensions[letter].width = width

    row = KB.brand_banner(
        ws, 1, 4, f"Retail Review — {engagement.client_name}",
        f"{program.title}  ·  Engagement {engagement.engagement_id}")
    row += 1

    def section(label: str) -> None:
        nonlocal row
        cell = ws.cell(row, 2, label)
        cell.font = Font(name="Arial", bold=True, size=11, color=KB.INK)
        ws.row_dimensions[row].height = 20
        row += 1

    def kv(label: str, value) -> None:
        nonlocal row
        ws.cell(row, 2, label).font = KB.SECONDARY
        c = ws.cell(row, 3, value)
        c.font = KB.DATA_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    section("Engagement")
    kv("Client bank", engagement.client_name)
    kv("Engagement ID", engagement.engagement_id)
    kv("Line of business", program.lob)
    kv("Review mode", "product conformance (Mode B)")
    row += 1
    section("Sample plan")
    for pid, block in engagement.overlay_products.items():
        pop = block["population"]
        kv(f"{pid} — population", f"{pop['count']:,} loans / ${pop['dollars']:,}")
        for seg in block["sample_plan"]["segments"]:
            stratum = ", ".join(f"{k}: {v}" for k, v in (seg.get("stratum") or {}).items())
            kv(f"    {seg['id']} ({seg['method']}"
               f"{', ' + stratum if stratum else ''})",
               f"n={seg['size']} — {seg['basis']}")
        kv(f"{pid} — files in this workbook", samples[pid].sample_n
           if hasattr(samples[pid], "sample_n") else len(samples[pid].files))
    row += 1
    section("Reviewer & independence")
    reviewer = engagement.reviewer
    kv("Reviewer", reviewer.get("name"))
    kv("Independent", "Yes" if reviewer.get("independent") else "No")
    if reviewer.get("independence_note"):
        kv("Independence basis", reviewer["independence_note"])
    row += 1
    note = ws.cell(row, 2,
                   "Mode B files are identified by loan/account number only — no "
                   "individual's name appears anywhere in this workpaper, and no "
                   "TIN in any form.")
    note.font = KB.NOTE_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    note.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Products roll-up (Mode B's Master)
# ---------------------------------------------------------------------------
def build_products_rollup(ws: Worksheet, engagement: Engagement,
                          refs: list[ProductRefs],
                          findings: FindingsRefs) -> dict[str, int]:
    KB.hide_gridlines(ws)
    widths = {"A": 16, "B": 12, "C": 14, "D": 10, "E": 11, "F": 11,
              "G": 13, "H": 11, "I": 13}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    row = KB.brand_banner(ws, 1, 9, "Products — Retail Roll-up",
                          f"{engagement.client_name}  ·  {engagement.engagement_id}")
    row = KB.header_row(ws, row, ["Product", "Pop count", "Pop $", "Sample",
                                  "Coverage", "Open find.", "Substd $", "Loss $",
                                  "Substd+Loss %"], right_from=1)
    product_row: dict[str, int] = {}
    for r in refs:
        pop = engagement.overlay_products[r.product_id]["population"]
        sheet = f"'{r.sheet_name}'"
        values = [
            (1, r.product_id, None),
            (2, pop["count"], FMT_NUM),
            (3, pop["dollars"], FMT_USD),
            (4, r.sample_n, FMT_NUM),
            (5, f"=$D{row}/$B{row}", FMT_PCT1),
            (6, findings.open_count_formula(("A", r.product_id)), FMT_NUM),
            (7, f"={sheet}!{r.substandard_dollars}", FMT_USD),
            (8, f"={sheet}!{r.loss_dollars}", FMT_USD),
            (9, f"=IFERROR(($G{row}+$H{row})/$C{row},\"\")", FMT_PCT1),
        ]
        for col, value, fmt in values:
            c = ws.cell(row, col, value)
            c.font = KB.DATA_FONT
            c.alignment = Alignment(horizontal="right" if col > 1 else "left",
                                    vertical="center")
            if fmt:
                c.number_format = fmt
        product_row[r.product_id] = row
        row += 1
    return product_row


def add_urccp_totals(ws: Worksheet, findings_refs: FindingsRefs,
                     refs: list[ProductRefs]) -> None:
    """URCCP asset-quality block on Findings (Mode B's analogue of the
    criticized/classified summary — retail has no Special Mention)."""
    row = findings_refs.next_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    band = ws.cell(row, 1, "Asset quality — URCCP (Substandard >=90 DPD; "
                           "Loss per product type)")
    band.fill = KB.HDR_FILL
    band.font = KB.WHITE_BOLD
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1
    sub = "+".join(f"'{r.sheet_name}'!{r.substandard_dollars.replace('C', '$C$', 1)}"
                   for r in refs)
    loss = "+".join(f"'{r.sheet_name}'!{r.loss_dollars.replace('C', '$C$', 1)}"
                    for r in refs)
    for label, formula in (("Substandard $ (portfolio)", f"={sub}"),
                           ("Loss $ (portfolio)", f"={loss}")):
        ws.cell(row, 1, label).font = KB.SECONDARY
        c = ws.cell(row, 2, formula)
        c.font = _INK_BOLD
        c.number_format = '$#,##0'
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1


# ---------------------------------------------------------------------------
# Data Mart (Mode B: product-level rows — de-identified by construction)
# ---------------------------------------------------------------------------
MART_COLUMNS_B = ["mart_id", "product", "population_count", "population_dollars",
                  "sample_n", "coverage", "open_findings", "substandard_dollars",
                  "loss_dollars"]


def build_mart_mode_b(ws: Worksheet, engagement: Engagement,
                      refs: list[ProductRefs], product_row: dict[str, int]) -> None:
    KB.hide_gridlines(ws)
    ws.column_dimensions["A"].width = 22
    for letter in "BCDEFGHI":
        ws.column_dimensions[letter].width = 16
    row = KB.header_row(ws, 1, list(MART_COLUMNS_B), right_from=2)
    src = {"population_count": "B", "population_dollars": "C", "sample_n": "D",
           "coverage": "E", "open_findings": "F", "substandard_dollars": "G",
           "loss_dollars": "H"}
    for i, r in enumerate(refs, start=1):
        master_row = product_row[r.product_id]
        ws.cell(row, 1, mart_id(engagement, i).replace("-L", "-P")).font = KB.DATA_FONT
        ws.cell(row, 2, r.product_id).font = KB.DATA_FONT
        for col, name in enumerate(MART_COLUMNS_B[2:], start=3):
            c = ws.cell(row, col, f"=Products!${src[name]}${master_row}")
            c.font = KB.DATA_FONT
            c.alignment = Alignment(horizontal="right", vertical="center")
            if name.endswith("dollars"):
                c.number_format = '$#,##0'
        row += 1


# ---------------------------------------------------------------------------
# Assembly (called from build_engagement_workbook)
# ---------------------------------------------------------------------------
def assemble_mode_b(wb: Workbook, program: Program, engagement: Engagement,
                    samples: dict[str, ProductSample]) -> None:
    cover = wb.active
    cover.title = "Cover"
    build_cover_mode_b(cover, program, engagement, samples)

    pol = write_config_sheet_mode_b(wb.create_sheet("_config"), program, engagement)

    product_refs: list[ProductRefs] = []
    per_product: list[tuple[_ProductHandle, list[ExceptionRef]]] = []
    for product in program.products:
        ws = wb.create_sheet(f"PS_{product['id']}")
        refs = build_product_sheet(ws, product, engagement,
                                   samples[product["id"]], pol)
        product_refs.append(refs)
        handle = _ProductHandle(loan_id=refs.product_id, sheet_name=refs.sheet_name)
        exceptions = [ExceptionRef(
            row_id=t.test_id, label=t.label, exc_class=t.test_class, subtype=None,
            severity=t.severity, flag_cell=t.flag_cell, status_cell=t.status_cell,
            owner_cell=t.owner_cell, due_cell=t.due_cell, cleared_cell=t.cleared_cell)
            for t in refs.tests]
        per_product.append((handle, exceptions))

    ws_products = wb.create_sheet("Products")
    ws_mart = wb.create_sheet("Data Mart")
    ws_findings = wb.create_sheet("Findings")
    findings_refs = build_findings(ws_findings, engagement, per_product)
    product_row = build_products_rollup(ws_products, engagement, product_refs,
                                        findings_refs)
    add_urccp_totals(ws_findings, findings_refs, product_refs)
    build_mart_mode_b(ws_mart, engagement, product_refs, product_row)
    build_methodology_mode_b(wb.create_sheet("_methodology"), program,
                             engagement, product_row)
    build_readme(wb.create_sheet("_readme"), program, engagement,
                 pii_notes=PII_NOTES_MODE_B)

    structure = {
        "version": 1,
        "mode": "product_conformance",
        "engagement_id": engagement.engagement_id,
        "lob": program.lob,
        "products": [
            {"id": r.product_id, "sheet": r.sheet_name,
             "products_row": product_row[r.product_id],
             "grid": {"first": r.grid_first, "last": r.grid_last,
                      "segment_col": r.segment_col, "fails_col": r.fails_col,
                      "fringe_col": r.fringe_col},
             "norms": {"fringe_rate": r.fringe_rate, "core_rate": r.core_rate,
                       "gap": r.fringe_gap, "flag": r.fringe_gap_flag},
             "segments": r.segment_rows,
             "panel": {"substandard_dollars": r.substandard_dollars,
                       "substandard_count": r.substandard_count,
                       "loss_dollars": r.loss_dollars,
                       "loss_count": r.loss_count},
             "tests": [{"id": t.test_id, "label": t.label, "class": t.test_class,
                        "severity": t.severity, "flag": t.flag_cell,
                        "status": t.status_cell, "rate": t.rate_cell,
                        "fails": t.fails_cell}
                       for t in r.tests]}
            for r in product_refs],
    }
    ws_map = wb.create_sheet("_map")
    ws_map["A1"] = json.dumps(structure, separators=(",", ":"), sort_keys=True)
    ws_map.sheet_state = "hidden"

    wb.move_sheet("_config", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_config"))
    wb.move_sheet("_readme", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_readme"))
    wb.move_sheet("_map", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_map"))
