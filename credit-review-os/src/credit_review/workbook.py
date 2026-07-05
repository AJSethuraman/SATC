"""Assemble the engagement workbook (slice 1: Cover, one ``LS_`` per loan, ``_config``).

The build is pure and deterministic — no network, no clock: identical inputs
must produce byte-identical output (use :func:`workbook_bytes`, which also
normalizes zip-container timestamps that plain ``Workbook.save`` would stamp
with the wall clock).
"""

from __future__ import annotations

import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from credit_review.config import (
    ConfigError,
    Engagement,
    Loan,
    Program,
    check_overlay_fits_program,
    load_engagement,
    load_loans,
    load_program,
)
from credit_review.config_sheet import write_config_sheet
from credit_review.cover import build_cover
from credit_review.findings import add_asset_quality, build_findings
from credit_review.linesheet import BuildContext, LinesheetBuilder
from credit_review.master import build_master
from credit_review.mart import build_mart, mart_id
from credit_review.methodology import build_methodology, build_readme

PACKAGE_DIR = Path(__file__).parent
PROGRAMS_DIR = PACKAGE_DIR / "programs"
ENGAGEMENTS_DIR = PACKAGE_DIR / "engagements"

# Fixed document properties: the build must not depend on the wall clock.
_EPOCH = datetime(2026, 1, 1)


def build_engagement_workbook(program: Program, engagement: Engagement,
                              loans: list[Loan]) -> Workbook:
    """Build the in-memory engagement workbook: ``Cover``, one ``LS_<loan_id>``
    linesheet per loan, and the ``_config`` knob panel, in KeyBank house style."""
    if program.review_mode != "loan_level":
        raise ConfigError(
            f"review_mode {program.review_mode!r} is designed into the schema but "
            f"not built yet — v1 builds loan_level only (PRD §3)")
    check_overlay_fits_program(engagement, program)
    if not loans:
        raise ConfigError("engagement has no loans — nothing to review")

    wb = Workbook()
    wb.properties.creator = "Credit Review OS"
    wb.properties.created = _EPOCH
    wb.properties.modified = _EPOCH

    cover = wb.active
    cover.title = "Cover"
    build_cover(cover, program, engagement, loans)

    # _config is written before the linesheets because its threshold cells and
    # rating-scale-map block are the [POL]/MAP() formula targets; it is moved
    # to the back of the tab order below.
    refs = write_config_sheet(wb.create_sheet("_config"), program, engagement)

    grades = tuple(sorted(engagement.rating_scale_map, key=lambda g: (len(g), g)))
    exceptions_per_loan, cells_per_loan = [], []
    for loan in loans:
        ctx = BuildContext(pol_registry=refs.pol_registry, map_range=refs.map_range,
                           asof_cell=refs.asof_cell, values=dict(loan.values),
                           rating_grades=grades)
        builder = LinesheetBuilder(
            wb.create_sheet(loan.sheet_name),
            program.sections,
            title=program.title,
            subtitle=(f"{engagement.client_name}  ·  {engagement.engagement_id}"
                      f"  ·  Loan {loan.loan_id}"),
            ctx=ctx)
        row_cells = builder.build()
        exceptions_per_loan.append((loan, builder.exceptions))
        cells_per_loan.append((loan, row_cells))

    ws_master = wb.create_sheet("Master")
    ws_mart = wb.create_sheet("Data Mart")
    ws_findings = wb.create_sheet("Findings")
    findings_refs = build_findings(ws_findings, engagement, exceptions_per_loan)
    master_refs = build_master(ws_master, engagement, cells_per_loan,
                               refs.map_range, refs.framework_range, findings_refs)
    add_asset_quality(ws_findings, findings_refs, master_refs)
    build_mart(ws_mart, program, engagement, loans, master_refs)
    build_methodology(wb.create_sheet("_methodology"), program, engagement,
                      loans, master_refs)

    # Hidden structure map for the re-ingest pass: row ids and cell addresses
    # only — never borrower data. See credit_review.ingest.
    structure = {
        "version": 1,
        "engagement_id": engagement.engagement_id,
        "lob": program.lob,
        "loans": [
            {"sheet": loan.sheet_name,
             "mart_id": mart_id(engagement, i),
             "master_row": master_refs.first_row + i - 1,
             "cells": cells,
             "exceptions": [
                 {"id": exc.row_id, "class": exc.exc_class, "subtype": exc.subtype,
                  "severity": exc.severity, "flag": exc.flag_cell,
                  "status": exc.status_cell}
                 for exc in exceptions]}
            for i, ((loan, cells), (_, exceptions)) in enumerate(
                zip(cells_per_loan, exceptions_per_loan), start=1)],
    }
    ws_map = wb.create_sheet("_map")
    ws_map["A1"] = json.dumps(structure, separators=(",", ":"), sort_keys=True)
    ws_map.sheet_state = "hidden"

    build_readme(wb.create_sheet("_readme"), program, engagement)

    wb.move_sheet("_config", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_config"))
    wb.move_sheet("_readme", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_readme"))
    wb.move_sheet("_map", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("_map"))
    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    """Serialize deterministically: same workbook content -> identical bytes.

    ``Workbook.save`` stamps each zip member with the current local time and
    overwrites ``dcterms:modified`` in ``docProps/core.xml`` with the save-time
    clock, so two otherwise-identical saves differ. Re-pack every member (same
    order, same compression) with a fixed timestamp, and pin ``dcterms:modified``
    back to the fixed epoch.
    """
    raw = io.BytesIO()
    wb.save(raw)
    fixed_modified = (b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                      + _EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
                      + b"</dcterms:modified>")
    out = io.BytesIO()
    with zipfile.ZipFile(raw) as src, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for name in src.namelist():
            data = src.read(name)
            if name == "docProps/core.xml":
                data = re.sub(rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                              fixed_modified, data)
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(info, data)
    return out.getvalue()


def build_demo_workbook() -> tuple[Workbook, Program, Engagement, list[Loan]]:
    """Load the synthetic demo engagement shipped with the package and build it."""
    program = load_program(PROGRAMS_DIR / "c_and_i.yaml")
    engagement = load_engagement(ENGAGEMENTS_DIR / "demo_engagement.yaml")
    loans = load_loans(engagement.loans_path)
    return build_engagement_workbook(program, engagement, loans), program, engagement, loans
