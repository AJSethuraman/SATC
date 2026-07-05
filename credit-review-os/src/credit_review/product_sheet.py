"""``PS_<product>`` — the Mode B product sheet: conformance sample grid +
URCCP pool-classification panel + per-test analytics.

Grid: one row per sampled file — **loan number and segment tag only** (person
names never enter Mode B), keyed numeric attributes, attestation columns
(pass/fail/na dropdowns), computed test columns (live formulas over the row's
attributes vs ``_config`` policy knobs), a FAILS helper, and a computed
FRINGE flag (attribute within the overlay's fringe band of its policy limit,
on the approved side — the buy-box edge).

Analytics: one row per conformance test — applicable n, fails, exception
rate, tolerance (a ``_config`` knob), and the finding flag: policy and
documentation tests flag when rate exceeds tolerance; **compliance tests flag
on any fail, regardless of rate**. These analytics rows are the product's
findings and carry the standard tracking lane.

Pool panel: keyed delinquency buckets classify live per URCCP
(65 FR 36903; OCC Bulletin 2000-20; FDIC FIL-40-2000): closed-end Substandard
>= 90 cumulative DPD / Loss >= 120; open-end Substandard >= 90 / Loss >= 180;
residential-secured Substandard = keyed qualifying >=90-DPD balances (LTV
above the overlay's qualifier line) and Loss = keyed fair-value writedowns
(the 180-day writedown itself is an attestation, never a computation). An
overlay may **tighten** the classification clock via ``classification_
overrides`` — never loosen it past the URCCP floor (validated here).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import ConfigError, Engagement
from credit_review.config_mode_b import ATTESTATION_VALUES, BUCKETS, ProductSample
from credit_review.linesheet import (
    CLEAR_FLAG,
    FMT_DATE,
    FMT_NUM,
    FMT_RATIO,
    FMT_TEXT,
    FMT_USD,
    OPEN_FLAG,
    LinesheetError,
)

FMT_PCT1 = "0.0%"

_INPUT_FILL = PatternFill("solid", fgColor=KB.CANVAS)
_hair = Side(style="thin", color=KB.MIST)
_BORDER = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

_ATTR_FMT = {"num": FMT_NUM, "ratio": FMT_RATIO, "pct": FMT_PCT1, "usd": FMT_USD,
             "text": FMT_TEXT}

BUCKET_LABELS = {
    "current": "Current", "dpd_30_59": "30-59 DPD", "dpd_60_89": "60-89 DPD",
    "dpd_90_119": "90-119 DPD", "dpd_120_179": "120-179 DPD",
    "dpd_180_plus": "180+ DPD",
}

# URCCP floors: the delinquency clock may be tightened by overlay override,
# never loosened past these (65 FR 36903).
_BUCKET_START_DPD = {"current": 0, "dpd_30_59": 30, "dpd_60_89": 60,
                     "dpd_90_119": 90, "dpd_120_179": 120, "dpd_180_plus": 180}
URCCP_FLOORS = {
    "closed_end": {"substandard_from_dpd": 90, "loss_from_dpd": 120},
    "open_end": {"substandard_from_dpd": 90, "loss_from_dpd": 180},
}


def _classification_dpd(product: dict, overlay_block: dict) -> dict[str, int]:
    """Resolve the classification clock: URCCP floor, tightened by overrides.

    Raises ConfigError if an override would loosen past the floor.
    """
    ctype = product["classification_type"]
    floors = URCCP_FLOORS[ctype]
    overrides = overlay_block.get("classification_overrides", {}) or {}
    resolved = {}
    for key, floor in floors.items():
        value = overrides.get(key, floor)
        if value not in _BUCKET_START_DPD.values():
            raise ConfigError(
                f"classification_overrides.{key}: {value!r} must be a bucket "
                f"boundary ({sorted(set(_BUCKET_START_DPD.values()))})")
        if value > floor:
            raise ConfigError(
                f"classification_overrides.{key}={value} would LOOSEN the URCCP "
                f"floor of {floor} days for {ctype} — bank policy may only "
                f"tighten the interagency standard (65 FR 36903)")
        resolved[key] = value
    return resolved


@dataclass(frozen=True)
class TestRef:
    """One analytics row — the Findings register consumes these."""

    test_id: str
    label: str
    test_class: str
    severity: str
    flag_cell: str
    status_cell: str
    owner_cell: str
    due_cell: str
    cleared_cell: str
    rate_cell: str
    fails_cell: str


@dataclass(frozen=True)
class ProductRefs:
    """Where everything on one PS_ sheet landed."""

    product_id: str
    sheet_name: str
    tests: tuple[TestRef, ...]
    grid_first: int
    grid_last: int
    segment_col: str
    fails_col: str
    fringe_col: str
    substandard_count: str      # unqualified cell addrs on this sheet
    substandard_dollars: str
    loss_count: str
    loss_dollars: str
    sample_n: int
    fringe_rate: str = ""       # fringe-vs-core norms block
    core_rate: str = ""
    fringe_gap: str = ""
    fringe_gap_flag: str = ""
    segment_rows: dict = field(default_factory=dict)   # segment id -> analytics row


def _resolve(expr: str, attr_cells: dict[str, str], pol: dict[str, str]) -> str:
    def sub_attr(m: re.Match) -> str:
        aid = m.group(1)
        if aid not in attr_cells:
            raise LinesheetError(f"computed test references unknown attribute {{{aid}}}")
        return attr_cells[aid]

    text = re.sub(r"\{([a-zA-Z0-9_]+)\}", sub_attr, expr)

    def sub_pol(m: re.Match) -> str:
        key = m.group(1)
        if key not in pol:
            raise LinesheetError(f"unknown policy threshold [POL {key}]")
        return pol[key]

    return re.sub(r"\[POL\s+([a-z0-9_]+)\]", sub_pol, text)


def build_product_sheet(ws: Worksheet, product: dict, engagement: Engagement,
                        sample: ProductSample, pol: dict[str, str]) -> ProductRefs:
    KB.hide_gridlines(ws)
    attributes = list(product["attributes"])
    tests = list(product["tests"])

    # -- grid geometry: Loan# | Segment | attrs... | tests... | FAILS | FRINGE
    col_loan, col_seg = 1, 2
    attr_col = {a["id"]: col_seg + 1 + i for i, a in enumerate(attributes)}
    test_col = {t["id"]: col_seg + len(attributes) + 1 + i for i, t in enumerate(tests)}
    col_fails = col_seg + len(attributes) + len(tests) + 1
    col_fringe = col_fails + 1
    last_col = col_fringe

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 11
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    row = KB.brand_banner(
        ws, 1, last_col, product["label"],
        f"{engagement.client_name}  ·  {engagement.engagement_id}"
        f"  ·  conformance sample + URCCP pool")
    headers = (["Loan #", "Segment"] + [a["label"] for a in attributes]
               + [t["label"] for t in tests] + ["Fails", "Fringe"])
    row = KB.header_row(ws, row, headers, center_cols=tuple(range(1, last_col)))
    KB.freeze_below(ws, row - 1)

    # -- sample grid ---------------------------------------------------------
    plan = engagement.overlay_products[product["id"]]["sample_plan"]
    segment_ids = [s["id"] for s in plan["segments"]]
    seg_dv = DataValidation(type="list", formula1='"' + ",".join(segment_ids) + '"',
                            allow_blank=False, showErrorMessage=True,
                            error="Segment must come from the sample plan (see Cover).")
    att_dv = DataValidation(type="list", formula1='"' + ",".join(ATTESTATION_VALUES) + '"',
                            allow_blank=True, showErrorMessage=True,
                            error="pass, fail, or na.")
    ws.add_data_validation(seg_dv)
    ws.add_data_validation(att_dv)

    grid_first = row
    for f in sample.files:
        attr_cells = {aid: f"{get_column_letter(c)}{row}" for aid, c in attr_col.items()}
        cell = ws.cell(row, col_loan, f["loan_number"])
        cell.font = KB.DATA_FONT
        cell.number_format = FMT_TEXT
        seg = ws.cell(row, col_seg, f["segment"])
        seg.font = KB.DATA_FONT
        seg.fill = _INPUT_FILL
        seg.border = _BORDER
        seg.alignment = _CENTER
        seg_dv.add(seg)
        for a in attributes:
            c = ws.cell(row, attr_col[a["id"]], f[a["id"]])
            c.font = KB.DATA_FONT
            c.fill = _INPUT_FILL
            c.border = _BORDER
            c.alignment = _RIGHT
            c.number_format = _ATTR_FMT.get(a.get("fmt", "num"), FMT_NUM)
        for t in tests:
            c = ws.cell(row, test_col[t["id"]])
            if t["kind"] == "attestation":
                c.value = f.get(t["id"])
                c.fill = _INPUT_FILL
                att_dv.add(c)
            else:
                when = _resolve(t["when"], attr_cells, pol)
                c.value = f'=IFERROR(IF({when},"fail","pass"),"")'
            c.font = KB.DATA_FONT
            c.border = _BORDER
            c.alignment = _CENTER
            c.number_format = FMT_TEXT
        first_t = get_column_letter(min(test_col.values()))
        last_t = get_column_letter(max(test_col.values()))
        ws.cell(row, col_fails,
                f'=COUNTIF({first_t}{row}:{last_t}{row},"fail")').font = KB.DATA_FONT
        ws.cell(row, col_fails).alignment = _CENTER
        fringe_terms = []
        for rule in product["fringe_rules"]:
            attr = attr_cells[rule["attribute"]]
            limit, band = pol[rule["limit_key"]], pol[rule["band_key"]]
            if rule["direction"] == "floor":
                fringe_terms.append(f"AND({attr}>={limit},{attr}<=({limit}+{band}))")
            else:
                fringe_terms.append(f"AND({attr}<={limit},{attr}>=({limit}-{band}))")
        fr = ws.cell(row, col_fringe,
                     f'=IF(OR({",".join(fringe_terms)}),"FRINGE","")'
                     if fringe_terms else "")
        fr.font = KB.DATA_FONT
        fr.alignment = _CENTER
        row += 1
    grid_last = row - 1

    for t in tests:
        letter = get_column_letter(test_col[t["id"]])
        rng = f"{letter}{grid_first}:{letter}{grid_last}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'{letter}{grid_first}="fail"'],
            fill=PatternFill("solid", fgColor=KB.ALERT_FG), font=KB.ALERT_FONT))
    row += 1

    # -- per-test analytics (these rows ARE the product's findings) ----------
    def band(r: int, label: str) -> int:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        c = ws.cell(r, 1, label)
        c.fill = KB.HDR_FILL
        c.font = KB.WHITE_BOLD
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        return r + 1

    row = band(row, "Test analytics — exception rate vs tolerance "
                    "(compliance: any fail is a finding)")
    acols = ["Test", "Class", "Severity", "n", "Fails", "Rate", "Tolerance",
             "Flag", "Status", "Owner", "Due", "Cleared"]
    row = KB.header_row(ws, row, acols, center_cols=(3, 4, 5, 6, 7, 8))
    status_dv = DataValidation(type="list", formula1='"open,cleared,waived"',
                               allow_blank=True, showErrorMessage=True,
                               error="open, cleared, or waived.")
    ws.add_data_validation(status_dv)

    refs: list[TestRef] = []
    for t in tests:
        letter = get_column_letter(test_col[t["id"]])
        rng = f"{letter}{grid_first}:{letter}{grid_last}"
        ws.cell(row, 1, t["label"]).font = KB.DATA_FONT
        ws.cell(row, 2, t["class"]).font = KB.DATA_FONT
        ws.cell(row, 3, t["severity"]).font = KB.DATA_FONT
        n = ws.cell(row, 4, f'=COUNTIF({rng},"pass")+COUNTIF({rng},"fail")')
        fails = ws.cell(row, 5, f'=COUNTIF({rng},"fail")')
        rate = ws.cell(row, 6, f'=IFERROR($E{row}/$D{row},"")')
        rate.number_format = FMT_PCT1
        tol_key = f"tol_{t['class']}"
        if tol_key not in pol:
            raise ConfigError(f"overlay tolerances missing class {t['class']!r}")
        tol = ws.cell(row, 7, f"={pol[tol_key]}")
        tol.number_format = FMT_PCT1
        if t["class"] == "compliance":
            flag = ws.cell(row, 8, f'=IF($E{row}>0,"{OPEN_FLAG}","{CLEAR_FLAG}")')
        else:
            # n=0 (nothing applicable) must stay blank: a "" rate compares
            # ABOVE any numeric tolerance in Excel's text-vs-number ordering.
            flag = ws.cell(row, 8,
                           f'=IF($D{row}=0,"",'
                           f'IF($F{row}>$G{row},"{OPEN_FLAG}","{CLEAR_FLAG}"))')
        for c in (n, fails, rate, tol, flag):
            c.font = KB.DATA_FONT
            c.alignment = _CENTER
        ws.conditional_formatting.add(f"H{row}", FormulaRule(
            formula=[f'$H${row}="{OPEN_FLAG}"'],
            fill=PatternFill("solid", fgColor=KB.ALERT_FG), font=KB.ALERT_FONT))
        status = ws.cell(row, 9, "open")
        status.fill = _INPUT_FILL
        status.border = _BORDER
        status.alignment = _CENTER
        status.font = KB.DATA_FONT
        status_dv.add(status)
        for col, fmt in ((10, FMT_TEXT), (11, FMT_DATE), (12, FMT_DATE)):
            c = ws.cell(row, col)
            c.fill = _INPUT_FILL
            c.border = _BORDER
            c.font = KB.DATA_FONT
            c.number_format = fmt
        refs.append(TestRef(
            test_id=t["id"], label=t["label"], test_class=t["class"],
            severity=t["severity"], flag_cell=f"H{row}", status_cell=f"I{row}",
            owner_cell=f"J{row}", due_cell=f"K{row}", cleared_cell=f"L{row}",
            rate_cell=f"F{row}", fails_cell=f"E{row}"))
        row += 1
    row += 1

    # -- fringe-vs-core norms block -------------------------------------------
    fr_col = get_column_letter(col_fringe)
    fl_col = get_column_letter(col_fails)
    fr_rng = f"${fr_col}${grid_first}:${fr_col}${grid_last}"
    fl_rng = f"${fl_col}${grid_first}:${fl_col}${grid_last}"
    row = band(row, "Buy-box fringe vs core — files failing >=1 test "
                    "(fringe = within the overlay band of a policy limit)")
    fringe_cells: dict[str, str] = {}

    def norm(label: str, formula: str, fmt: str | None = FMT_PCT1) -> str:
        nonlocal row
        ws.cell(row, 1, label).font = KB.SECONDARY
        c = ws.cell(row, 2, formula)
        c.font = Font(name="Arial", bold=True, size=10, color=KB.INK)
        c.alignment = _RIGHT
        if fmt:
            c.number_format = fmt
        addr = f"B{row}"
        row += 1
        return addr

    n_fringe = norm("Fringe files (n)", f'=COUNTIF({fr_rng},"FRINGE")', FMT_NUM)
    fringe_cells["fringe_rate"] = norm(
        "Fringe fail-any rate",
        f'=IFERROR(SUMPRODUCT(({fr_rng}="FRINGE")*({fl_rng}>0))/{n_fringe},"")')
    n_core = norm("Core files (n)",
                  f'={grid_last - grid_first + 1}-{n_fringe}', FMT_NUM)
    fringe_cells["core_rate"] = norm(
        "Core fail-any rate",
        f'=IFERROR(SUMPRODUCT(({fr_rng}<>"FRINGE")*({fl_rng}>0))/{n_core},"")')
    fringe_cells["gap"] = norm(
        "Gap (fringe - core)",
        f'=IFERROR({fringe_cells["fringe_rate"]}-{fringe_cells["core_rate"]},"")')
    margin_ref = pol.get("fringe_margin", "1")   # no margin knob -> never flags
    fringe_cells["flag"] = norm(
        "Fringe exceeds core by more than the margin?",
        f'=IFERROR(IF({fringe_cells["gap"]}>{margin_ref},'
        f'"{OPEN_FLAG}","{CLEAR_FLAG}"),"")', None)
    flag_addr = fringe_cells["flag"]
    ws.conditional_formatting.add(flag_addr, FormulaRule(
        formula=[f'${flag_addr[0]}${flag_addr[1:]}="{OPEN_FLAG}"'],
        fill=PatternFill("solid", fgColor=KB.ALERT_FG), font=KB.ALERT_FONT))
    row += 1

    # -- per-segment / stratum analytics ---------------------------------------
    seg_col = "B"
    sg_rng = f"${seg_col}${grid_first}:${seg_col}${grid_last}"
    row = band(row, "Sample segments — coverage and fail-any rate by "
                    "stratum (random + judgmental)")
    row = KB.header_row(ws, row, ["Segment", "Method / stratum", "Planned",
                                  "Sampled", "Fail-any", "Rate"],
                        center_cols=(2, 3, 4, 5))
    segment_rows: dict[str, int] = {}
    for seg in plan["segments"]:
        stratum = ", ".join(f"{k}: {v}" for k, v in (seg.get("stratum") or {}).items())
        ws.cell(row, 1, seg["id"]).font = KB.DATA_FONT
        ws.cell(row, 2, seg["method"] + (f" ({stratum})" if stratum else "")) \
            .font = KB.DATA_FONT
        ws.cell(row, 3, seg["size"]).font = KB.DATA_FONT
        n_cell = ws.cell(row, 4, f'=COUNTIF({sg_rng},"{seg["id"]}")')
        f_cell = ws.cell(row, 5,
                         f'=SUMPRODUCT(({sg_rng}="{seg["id"]}")*({fl_rng}>0))')
        r_cell = ws.cell(row, 6, f'=IFERROR($E{row}/$D{row},"")')
        r_cell.number_format = FMT_PCT1
        for c in (n_cell, f_cell, r_cell):
            c.font = KB.DATA_FONT
            c.alignment = _CENTER
        segment_rows[seg["id"]] = row
        row += 1
    row += 1

    # -- URCCP pool panel -------------------------------------------------------
    ctype = product["classification_type"]
    overlay_block = engagement.overlay_products[product["id"]]
    bold_ink = Font(name="Arial", bold=True, size=10, color=KB.INK)
    bold_crimson = Font(name="Arial", bold=True, size=10, color=KB.CRIMSON)

    def styled(cell, font):
        cell.font = font
        cell.alignment = _RIGHT
        cell.number_format = FMT_USD if cell.column == 3 else FMT_NUM

    if ctype in ("closed_end", "open_end"):
        dpd = _classification_dpd(product, overlay_block)
        sub_from, loss_from = dpd["substandard_from_dpd"], dpd["loss_from_dpd"]
        row = band(row, f"Pool classification — URCCP ({ctype.replace('_', '-')}: "
                        f"Substandard >={sub_from} cumulative DPD, Loss >="
                        f"{loss_from}) · 65 FR 36903; OCC 2000-20; FDIC FIL-40-2000")
        row = KB.header_row(ws, row, ["Bucket", "Count", "Dollars"], right_from=1)
        bucket_row = {}
        for bucket in BUCKETS:
            ws.cell(row, 1, BUCKET_LABELS[bucket]).font = KB.DATA_FONT
            for col, key, fmt in ((2, "count", FMT_NUM), (3, "dollars", FMT_USD)):
                c = ws.cell(row, col, sample.pool[bucket][key])
                c.fill = _INPUT_FILL
                c.border = _BORDER
                c.font = KB.DATA_FONT
                c.alignment = _RIGHT
                c.number_format = fmt
            bucket_row[bucket] = row
            row += 1
        last_bucket = bucket_row["dpd_180_plus"]
        sub_start = next(bucket_row[b] for b in BUCKETS
                         if _BUCKET_START_DPD[b] == sub_from)
        loss_start = next(bucket_row[b] for b in BUCKETS
                          if _BUCKET_START_DPD[b] == loss_from)
        ws.cell(row, 1, f"Substandard (>={sub_from} DPD)").font = bold_ink
        styled(ws.cell(row, 2, f"=SUM($B${sub_start}:$B${last_bucket})"), bold_ink)
        styled(ws.cell(row, 3, f"=SUM($C${sub_start}:$C${last_bucket})"), bold_ink)
        substandard = (f"B{row}", f"C{row}")
        row += 1
        ws.cell(row, 1, f"Loss (>={loss_from} DPD)").font = bold_crimson
        styled(ws.cell(row, 2, f"=SUM($B${loss_start}:$B${last_bucket})"), bold_ink)
        styled(ws.cell(row, 3, f"=SUM($C${loss_start}:$C${last_bucket})"), bold_ink)
        loss = (f"B{row}", f"C{row}")
    else:   # residential_secured
        row = band(row, "Pool classification — URCCP (residential-secured: "
                        "Substandard = >=90 DPD with LTV above the _config "
                        "qualifier; Loss = fair-value writedowns taken — the "
                        "180-day writedown is attested, not computed) · "
                        "65 FR 36903; OCC 2000-20; FDIC FIL-40-2000")
        row = KB.header_row(ws, row, ["Bucket", "Count", "Dollars"], right_from=1)
        for bucket in BUCKETS:
            ws.cell(row, 1, BUCKET_LABELS[bucket]).font = KB.DATA_FONT
            for col, key, fmt in ((2, "count", FMT_NUM), (3, "dollars", FMT_USD)):
                c = ws.cell(row, col, sample.pool[bucket][key])
                c.fill = _INPUT_FILL
                c.border = _BORDER
                c.font = KB.DATA_FONT
                c.alignment = _RIGHT
                c.number_format = fmt
            row += 1
        ws.cell(row, 1, ">=90 DPD with LTV above qualifier (keyed)") \
            .font = KB.DATA_FONT
        qual_row = row
        for col, key in ((2, "count"), (3, "dollars")):
            c = ws.cell(row, col, sample.pool["qualifying_90_ltv60"][key])
            c.fill = _INPUT_FILL
            c.border = _BORDER
            styled(c, KB.DATA_FONT)
        row += 1
        ws.cell(row, 1, "Fair-value writedowns taken (keyed)").font = KB.DATA_FONT
        wd_row = row
        for col, key in ((2, "count"), (3, "dollars")):
            c = ws.cell(row, col, sample.pool["writedown_loss"][key])
            c.fill = _INPUT_FILL
            c.border = _BORDER
            styled(c, KB.DATA_FONT)
        row += 1
        ws.cell(row, 1, "Substandard (qualifying >=90 DPD)").font = bold_ink
        styled(ws.cell(row, 2, f"=$B${qual_row}"), bold_ink)
        styled(ws.cell(row, 3, f"=$C${qual_row}"), bold_ink)
        substandard = (f"B{row}", f"C{row}")
        row += 1
        ws.cell(row, 1, "Loss (writedowns)").font = bold_crimson
        styled(ws.cell(row, 2, f"=$B${wd_row}"), bold_ink)
        styled(ws.cell(row, 3, f"=$C${wd_row}"), bold_ink)
        loss = (f"B{row}", f"C{row}")

    return ProductRefs(
        product_id=product["id"], sheet_name=ws.title, tests=tuple(refs),
        grid_first=grid_first, grid_last=grid_last,
        segment_col="B", fails_col=fl_col, fringe_col=fr_col,
        substandard_count=substandard[0], substandard_dollars=substandard[1],
        loss_count=loss[0], loss_dollars=loss[1],
        sample_n=len(sample.files),
        fringe_rate=fringe_cells["fringe_rate"], core_rate=fringe_cells["core_rate"],
        fringe_gap=fringe_cells["gap"], fringe_gap_flag=fringe_cells["flag"],
        segment_rows=segment_rows)
