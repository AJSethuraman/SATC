"""``Findings`` — the exception register and its aggregates.

One row per (loan, exception rule). The flag/status/owner/due/cleared columns
are **live formulas referencing the loan's linesheet** — the linesheet stays
the single source of truth; this sheet only collects and counts. Aggregates
(by class, by severity, by status, by loan) are SUMPRODUCT formulas over the
register, so they move the moment the reviewer keys a value or clears an
exception.

An exception is counted as *open* when its linesheet flag says OPEN **and**
its status is still ``open`` (a ``waived`` or ``cleared`` status drops it from
the open counts while staying visible in the register).
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import (
    EXCEPTION_CLASSES,
    EXCEPTION_STATUSES,
    Engagement,
    Loan,
)
from credit_review.linesheet import ExceptionRef, FMT_DATE, OPEN_FLAG

SEVERITY_ORDER = ("high", "medium", "low")

_COLS = ["Loan", "Finding", "Class", "Subtype", "Severity",
         "Flag", "Status", "Owner", "Due", "Cleared"]
# Register column letters (col A = Loan ... col J = Cleared).
_C_LOAN, _C_CLASS, _C_SEV, _C_FLAG, _C_STATUS = "A", "C", "E", "F", "G"


@dataclass(frozen=True)
class FindingsRefs:
    """Where the register landed — Master and later sheets aggregate against it."""

    first_row: int
    last_row: int
    next_row: int = 0   # first free row after the aggregates (asset-quality block)

    def open_count_formula(self, *criteria: tuple[str, str]) -> str:
        """SUMPRODUCT counting register rows that are open, with extra
        ``(column_letter, value)`` equality criteria."""
        f, l = self.first_row, self.last_row
        terms = [f'(Findings!${_C_FLAG}${f}:${_C_FLAG}${l}="{OPEN_FLAG}")',
                 f'(Findings!${_C_STATUS}${f}:${_C_STATUS}${l}="open")']
        terms += [f'(Findings!${col}${f}:${col}${l}={value!r})'.replace("'", '"')
                  for col, value in criteria]
        return "=SUMPRODUCT(" + "*".join(terms) + ")"


def build_findings(ws: Worksheet, engagement: Engagement,
                   per_loan: list[tuple[Loan, list[ExceptionRef]]]) -> FindingsRefs:
    KB.hide_gridlines(ws)
    widths = {"A": 12, "B": 44, "C": 14, "D": 20, "E": 10,
              "F": 8, "G": 10, "H": 14, "I": 12, "J": 12}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    row = KB.brand_banner(
        ws, 1, len(_COLS), "Findings & Exceptions",
        f"{engagement.client_name}  ·  {engagement.engagement_id}")
    row = KB.header_row(ws, row, _COLS, center_cols=(4, 5, 6))
    KB.freeze_below(ws, row - 1)

    first = row
    for loan, exceptions in per_loan:
        for exc in exceptions:
            sheet = f"'{loan.sheet_name}'"
            static = [loan.loan_id, exc.label, exc.exc_class, exc.subtype or "",
                      exc.severity]
            for col, value in enumerate(static, start=1):
                cell = ws.cell(row, col, value)
                cell.font = KB.DATA_FONT
            refs = [(6, exc.flag_cell, None), (7, exc.status_cell, None),
                    (8, exc.owner_cell, None), (9, exc.due_cell, FMT_DATE),
                    (10, exc.cleared_cell, FMT_DATE)]
            for col, addr, fmt in refs:
                # blanks stay blank (a bare =ref shows an empty cell as 0)
                cell = ws.cell(row, col,
                               f'=IF({sheet}!{addr}="","",{sheet}!{addr})')
                cell.font = KB.DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if fmt:
                    cell.number_format = fmt
            row += 1
    last = row - 1
    refs = FindingsRefs(first_row=first, last_row=last)

    if last >= first:
        flag_range = f"{_C_FLAG}{first}:{_C_FLAG}{last}"
        ws.conditional_formatting.add(flag_range, FormulaRule(
            formula=[f'${_C_FLAG}{first}="{OPEN_FLAG}"'],
            fill=PatternFill("solid", fgColor=KB.ALERT_FG), font=KB.ALERT_FONT))

    def _band(r: int, label: str) -> int:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws.cell(r, 1, label)
        cell.fill = KB.HDR_FILL
        cell.font = KB.WHITE_BOLD
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        return r + 1

    def _agg(r: int, label: str, formula: str) -> int:
        lab = ws.cell(r, 1, label)
        lab.font = KB.SECONDARY
        val = ws.cell(r, 2, formula)
        val.font = Font(name="Arial", bold=True, size=11, color=KB.INK)
        val.alignment = Alignment(horizontal="left", vertical="center")
        return r + 1

    row += 1
    row = _band(row, "Open exceptions by class")
    for cls in EXCEPTION_CLASSES:
        row = _agg(row, cls, refs.open_count_formula((_C_CLASS, cls)))
    row += 1
    row = _band(row, "Open exceptions by severity")
    for sev in SEVERITY_ORDER:
        row = _agg(row, sev, refs.open_count_formula((_C_SEV, sev)))
    row += 1
    row = _band(row, "Fired exceptions by status")
    f, l = refs.first_row, refs.last_row
    for status in EXCEPTION_STATUSES:
        formula = (f'=SUMPRODUCT((Findings!${_C_FLAG}${f}:${_C_FLAG}${l}="{OPEN_FLAG}")'
                   f'*(Findings!${_C_STATUS}${f}:${_C_STATUS}${l}="{status}"))')
        row = _agg(row, status, formula)
    row += 1
    row = _band(row, "Open exceptions by loan")
    for loan, _ in per_loan:
        row = _agg(row, loan.loan_id, refs.open_count_formula((_C_LOAN, loan.loan_id)))
    row = _agg(row, "TOTAL open", refs.open_count_formula())

    return FindingsRefs(first_row=refs.first_row, last_row=refs.last_row, next_row=row + 1)


def add_asset_quality(ws: Worksheet, refs: FindingsRefs, master) -> None:
    """Append the criticized/classified asset-quality summary, computed by
    formula from the Master roll-up (criticized includes Special Mention;
    classified does not — the interagency boundary)."""
    f, l = master.first_row, master.last_row
    out = f"Master!${master.outstanding_col}${f}:${master.outstanding_col}${l}"

    def _sum(flag_col: str) -> str:
        rng = f"Master!${flag_col}${f}:${flag_col}${l}"
        return f"=SUMPRODUCT(({rng}=TRUE)*{out})"

    def _count(flag_col: str) -> str:
        rng = f"Master!${flag_col}${f}:${flag_col}${l}"
        return f"=SUMPRODUCT(({rng}=TRUE)*1)"

    row = refs.next_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    band = ws.cell(row, 1, "Asset quality")
    band.fill = KB.HDR_FILL
    band.font = KB.WHITE_BOLD
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1
    for label, formula, fmt in (
            ("Criticized $ (outstanding)", _sum(master.criticized_col), '$#,##0'),
            ("Criticized count", _count(master.criticized_col), None),
            ("Classified $ (outstanding)", _sum(master.classified_col), '$#,##0'),
            ("Classified count", _count(master.classified_col), None)):
        lab = ws.cell(row, 1, label)
        lab.font = KB.SECONDARY
        val = ws.cell(row, 2, formula)
        val.font = Font(name="Arial", bold=True, size=11, color=KB.INK)
        val.alignment = Alignment(horizontal="left", vertical="center")
        if fmt:
            val.number_format = fmt
        row += 1
