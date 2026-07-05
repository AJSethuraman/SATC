"""``Data Mart`` — the de-identified flat grid (the cross-engagement seam).

One row per loan, keyed by an **engagement-scoped id** (``<ENG>-L01`` ...);
no borrower name, no loan/obligor number, no TIN in any form. Every data cell
is a live formula into the Master roll-up, so the mart tracks the reviewer's
keyed assessments without duplicating them. Adapts the ``satc_system``
``mart_sheets`` flat-grid pattern: header row, then data — machine-consumable.
"""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import Engagement, Loan, Program
from credit_review.linesheet import FMT_USD
from credit_review.master import MasterRefs

MART_COLUMNS = ["mart_id", "lob", "committed", "outstanding", "orig_grade",
                "rev_grade", "bucket", "criticized", "classified", "open_exceptions"]

# Master register columns the mart mirrors (see master._COLS).
_MASTER_SRC = {"committed": "D", "outstanding": "E", "orig_grade": "F",
               "rev_grade": "G", "bucket": "H", "criticized": "I",
               "classified": "J", "open_exceptions": "K"}


def mart_id(engagement: Engagement, index: int) -> str:
    """Engagement-scoped de-identified loan key (1-based index)."""
    return f"{engagement.engagement_id}-L{index:02d}"


def build_mart(ws: Worksheet, program: Program, engagement: Engagement,
               loans: list[Loan], master: MasterRefs) -> None:
    KB.hide_gridlines(ws)
    ws.column_dimensions["A"].width = 20
    for letter in "BCDEFGHIJ":
        ws.column_dimensions[letter].width = 14

    row = KB.header_row(ws, 1, [c for c in MART_COLUMNS], right_from=2)
    for i, _loan in enumerate(loans, start=1):
        master_row = master.first_row + i - 1
        ws.cell(row, 1, mart_id(engagement, i)).font = KB.DATA_FONT
        ws.cell(row, 2, program.lob).font = KB.DATA_FONT
        for col, name in enumerate(MART_COLUMNS[2:], start=3):
            cell = ws.cell(row, col, f"=Master!${_MASTER_SRC[name]}${master_row}")
            cell.font = KB.DATA_FONT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if name in ("committed", "outstanding"):
                cell.number_format = FMT_USD
        row += 1
