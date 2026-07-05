"""Config-driven loan linesheet builder (one ``LS_<loan_id>`` sheet per loan).

Adapts — does not import — the ``satc_system`` ``LineSheetBuilder`` token
grammar, restyled to the KeyBank house style. Formula tokens are resolved at
build time so the workbook ships live Excel formulas, never hardcoded results:

  ``{row_id}``     -> the value cell of another row on this linesheet
  ``[POL key]``    -> the engagement policy-threshold cell on ``_config``
  ``[ASOF]``       -> the engagement review_as_of date cell on ``_config``
  ``MAP(expr)``    -> VLOOKUP of an internal grade against the ``_config``
                      rating-scale-map block (grade -> regulatory bucket)

Row ``kind`` vocabulary: input / input_num / input_text / rating_input /
computed / exception / subhead / note / spacer. An ``exception`` row renders a
live flag formula (``OPEN`` when its ``when`` condition holds, ``—`` when
tested clean, blank while inputs are missing) plus the reviewer-managed
tracking fields (status / owner / due / cleared). The ``evidence`` kind is
schema-reserved for the staleness engine and refused loudly.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import EXCEPTION_STATUSES, RESERVED_KINDS, mask_tin_last4

# Column layout: A margin, B label, C value/flag, D basis/note, then the
# exception-tracking lane: E status, F owner, G due date, H cleared date.
COL_LABEL = 2
COL_VALUE = 3
COL_NOTE = 4
COL_STATUS = 5
COL_OWNER = 6
COL_DUE = 7
COL_CLEARED = 8
LAST_COL = 8
VALUE_LETTER = "C"

# Number formats (currency shows parens for negatives, dash for zero).
FMT_USD = '$#,##0;($#,##0);"-"'
FMT_NUM = '#,##0;(#,##0);"-"'
FMT_PCT = "0.00%"
FMT_RATIO = "0.00"
FMT_DATE = "mm/dd/yyyy"
FMT_TEXT = "@"
_FMT = {"usd": FMT_USD, "num": FMT_NUM, "text": FMT_TEXT, "pct": FMT_PCT,
        "ratio": FMT_RATIO, "date": FMT_DATE}

# House-styled cell vocabulary. Inputs sit on the CANVAS tint so reviewer-entry
# cells read differently from computed ones; computed values are plain INK.
_INPUT_FILL = PatternFill("solid", fgColor=KB.CANVAS)
_hair = Side(style="thin", color=KB.MIST)
_CELL_BORDER = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")

OPEN_FLAG = "OPEN"
CLEAR_FLAG = "—"


class LinesheetError(Exception):
    """Raised when a linesheet row or formula token cannot be resolved."""


@dataclass
class BuildContext:
    """Everything a linesheet needs to resolve cross-references and prefills."""

    pol_registry: dict[str, str] = field(default_factory=dict)  # threshold key -> '_config!$B$n'
    map_range: str = ""                                         # rating_scale_map block on _config
    asof_cell: str = ""                                         # review_as_of cell on _config
    values: dict[str, Any] = field(default_factory=dict)        # row_id -> prefilled value
    rating_grades: tuple[str, ...] = ()                         # overlay's internal grades


@dataclass(frozen=True)
class ExceptionRef:
    """Where one exception row landed — the Findings sheet aggregates these."""

    row_id: str
    label: str
    exc_class: str
    subtype: str | None
    severity: str
    flag_cell: str      # unqualified, e.g. 'C14'
    status_cell: str
    owner_cell: str
    due_cell: str
    cleared_cell: str


class LinesheetBuilder:
    """Renders one program's linesheet config onto a worksheet for one loan."""

    def __init__(self, ws: Worksheet, program_sections: tuple[dict, ...],
                 title: str, subtitle: str, ctx: BuildContext) -> None:
        self.ws = ws
        self.sections = program_sections
        self.title = title
        self.subtitle = subtitle
        self.ctx = ctx
        self.row_cells: dict[str, str] = {}   # row_id -> unqualified same-sheet cell, e.g. 'C9'
        self.exceptions: list[ExceptionRef] = []
        self._rating_dv: DataValidation | None = None
        self._status_dv: DataValidation | None = None

    # -- formula resolution ------------------------------------------------
    def _resolve_tokens(self, formula: str) -> str:
        text = formula

        def sub_row(m: re.Match) -> str:
            rid = m.group(1)
            if rid not in self.row_cells:
                raise LinesheetError(f"Unknown row id referenced: {{{rid}}}")
            return self.row_cells[rid]

        text = re.sub(r"\{([a-zA-Z0-9_]+)\}", sub_row, text)

        def sub_pol(m: re.Match) -> str:
            key = m.group(1)
            if key not in self.ctx.pol_registry:
                raise LinesheetError(f"Unknown policy threshold [POL {key}] — not in the "
                                     f"engagement overlay's thresholds")
            return self.ctx.pol_registry[key]

        text = re.sub(r"\[POL\s+([a-z0-9_]+)\]", sub_pol, text)

        if "[ASOF]" in text:
            if not self.ctx.asof_cell:
                raise LinesheetError("[ASOF] used but no review_as_of cell is available")
            text = text.replace("[ASOF]", self.ctx.asof_cell)

        def sub_map(m: re.Match) -> str:
            if not self.ctx.map_range:
                raise LinesheetError("MAP() used but no rating-scale map range is available")
            return f"VLOOKUP({m.group(1)},{self.ctx.map_range},2,FALSE)"

        text = re.sub(r"MAP\(([^()]+)\)", sub_map, text)
        return text

    # -- row writers -------------------------------------------------------
    def _write_value_row(self, row: int, item: dict, *, formula: str | None,
                         font: Font, fill: PatternFill | None,
                         align: Alignment, fmt: str) -> None:
        ws = self.ws
        label = ws.cell(row, COL_LABEL, item.get("label", ""))
        label.font = KB.DATA_FONT
        cell = ws.cell(row, COL_VALUE)
        if formula is not None:
            cell.value = formula
        else:
            rid = item.get("id", "")
            prefill = self.ctx.values.get(rid, None)
            if "tin" in rid:
                prefill = mask_tin_last4(prefill)   # last-4 only, everywhere
            # Text-comparing formulas (MAP grade lookups, ="yes" checks) need
            # text: normalize YAML's native scalars (rating_originator: 4 ->
            # int; flood_zone_collateral: yes -> bool) to the strings the
            # formulas expect, instead of silently blanking a lookup.
            kind = item.get("kind", "input")
            if prefill is not None:
                if kind == "rating_input":
                    prefill = str(prefill)
                elif kind == "input_text":
                    if isinstance(prefill, bool):
                        prefill = "yes" if prefill else "no"
                    elif not isinstance(prefill, str):
                        prefill = str(prefill)
            cell.value = prefill
        cell.font = font
        if fill is not None:
            cell.fill = fill
        cell.border = _CELL_BORDER
        cell.alignment = align
        cell.number_format = fmt
        if item.get("note"):
            note = ws.cell(row, COL_NOTE, item["note"])
            note.font = KB.NOTE_FONT
            note.alignment = Alignment(wrap_text=True, vertical="center")
        rid = item.get("id")
        if rid:
            self.row_cells[rid] = f"{VALUE_LETTER}{row}"

    def _tracking_input(self, row: int, col: int, value, fmt: str,
                        align: Alignment = _CENTER):
        cell = self.ws.cell(row, col, value)
        cell.font = KB.DATA_FONT
        cell.fill = _INPUT_FILL
        cell.border = _CELL_BORDER
        cell.alignment = align
        cell.number_format = fmt
        return cell

    def _render_exception(self, row: int, item: dict) -> int:
        ws = self.ws
        label = ws.cell(row, COL_LABEL, item["label"])
        label.font = KB.DATA_FONT
        when = self._resolve_tokens(item["when"])
        # Blank while inputs are missing (lookups error), OPEN when the tested
        # condition holds, an em-dash when tested and clean.
        flag = ws.cell(row, COL_VALUE,
                       f'=IFERROR(IF({when},"{OPEN_FLAG}","{CLEAR_FLAG}"),"")')
        flag.font = KB.DATA_FONT
        flag.border = _CELL_BORDER
        flag.alignment = _CENTER
        flag.number_format = FMT_TEXT
        # Red leads only where something demands attention — the OPEN flag.
        addr = f"{VALUE_LETTER}{row}"
        ws.conditional_formatting.add(addr, FormulaRule(
            formula=[f'${VALUE_LETTER}${row}="{OPEN_FLAG}"'],
            fill=PatternFill("solid", fgColor=KB.ALERT_FG), font=KB.ALERT_FONT))

        exc_class = item["class"]
        subtype = item.get("subtype")
        severity = item["severity"]
        meta = ws.cell(row, COL_NOTE,
                       " / ".join([exc_class] + ([subtype] if subtype else []) + [severity]))
        meta.font = KB.NOTE_FONT
        meta.alignment = _LEFT

        if self._status_dv is None:
            self._status_dv = DataValidation(
                type="list", formula1='"' + ",".join(EXCEPTION_STATUSES) + '"',
                allow_blank=True, showErrorMessage=True,
                error="Exception status: open, cleared, or waived.")
            ws.add_data_validation(self._status_dv)
        status = self._tracking_input(row, COL_STATUS, "open", FMT_TEXT)
        self._status_dv.add(status)
        self._tracking_input(row, COL_OWNER, None, FMT_TEXT, _LEFT)
        self._tracking_input(row, COL_DUE, None, FMT_DATE)
        self._tracking_input(row, COL_CLEARED, None, FMT_DATE)

        rid = item.get("id")
        if rid:
            self.row_cells[rid] = addr
        self.exceptions.append(ExceptionRef(
            row_id=rid or "", label=item["label"], exc_class=exc_class,
            subtype=subtype, severity=severity, flag_cell=addr,
            status_cell=f"{get_column_letter(COL_STATUS)}{row}",
            owner_cell=f"{get_column_letter(COL_OWNER)}{row}",
            due_cell=f"{get_column_letter(COL_DUE)}{row}",
            cleared_cell=f"{get_column_letter(COL_CLEARED)}{row}"))
        return row + 1

    def _render_row(self, row: int, item: dict) -> int:
        kind = item.get("kind", "input")
        ws = self.ws

        if kind == "spacer":
            ws.row_dimensions[row].height = 6
            return row + 1
        if kind == "subhead":
            cell = ws.cell(row, COL_LABEL, item["label"])
            cell.font = Font(name="Arial", bold=True, size=10, color=KB.SLATE)
            return row + 1
        if kind == "note":
            cell = ws.cell(row, COL_LABEL, item["label"])
            cell.font = KB.NOTE_FONT
            return row + 1

        if kind in ("input", "input_num", "input_text"):
            fmt = _FMT[item.get("fmt", {"input": "usd", "input_num": "num",
                                        "input_text": "text"}[kind])]
            align = _LEFT if kind == "input_text" else _RIGHT
            self._write_value_row(row, item, formula=None, font=KB.DATA_FONT,
                                  fill=_INPUT_FILL, align=align, fmt=fmt)
            return row + 1

        if kind == "rating_input":
            self._write_value_row(row, item, formula=None, font=KB.DATA_FONT,
                                  fill=_INPUT_FILL, align=_CENTER, fmt=FMT_TEXT)
            if self.ctx.rating_grades:
                if self._rating_dv is None:
                    grades = ",".join(self.ctx.rating_grades)
                    self._rating_dv = DataValidation(
                        type="list", formula1=f'"{grades}"', allow_blank=True,
                        showErrorMessage=True,
                        error="Pick one of the bank's internal grades (see _config).")
                    self.ws.add_data_validation(self._rating_dv)
                self._rating_dv.add(self.ws.cell(row, COL_VALUE))
            return row + 1

        if kind == "computed":
            formula = "=" + self._resolve_tokens(item["formula"]).lstrip("=")
            self._write_value_row(row, item, formula=formula, font=KB.DATA_FONT,
                                  fill=None, align=_RIGHT,
                                  fmt=_FMT[item.get("fmt", "usd")])
            return row + 1

        if kind == "exception":
            return self._render_exception(row, item)

        if kind in RESERVED_KINDS:
            raise LinesheetError(
                f"Row kind {kind!r} is schema-reserved and not built yet "
                f"(evidence staleness arrives with the exception engine, issue #62)")
        raise LinesheetError(f"Unknown row kind: {kind!r}")

    # -- public ------------------------------------------------------------
    def build(self) -> dict[str, str]:
        ws = self.ws
        KB.hide_gridlines(ws)
        widths = {"A": 2, "B": 42, "C": 18, "D": 34, "E": 10, "F": 14, "G": 12, "H": 12}
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width

        row = KB.brand_banner(ws, 1, LAST_COL, self.title, self.subtitle)
        row = KB.header_row(
            ws, row, ["", "Line", "Value / flag", "Basis / note",
                      "Status", "Owner", "Due", "Cleared"],
            center_cols=(2, 4, 6, 7))
        KB.freeze_below(ws, row - 1)

        for section in self.sections:
            # Ink section band with the white Arial-bold label — the content-tab
            # section treatment (section_band's onyx is for system tabs).
            last = get_column_letter(LAST_COL)
            ws.merge_cells(f"B{row}:{last}{row}")
            band = ws.cell(row, COL_LABEL, section["title"])
            band.fill = KB.HDR_FILL
            band.font = KB.WHITE_BOLD
            band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1
            for item in section.get("rows", []):
                row = self._render_row(row, item)
            row += 1  # gap between sections

        return dict(self.row_cells)
