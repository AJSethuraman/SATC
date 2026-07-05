"""Re-ingest: read a filled engagement workbook back into de-identified data.

``ingest_workbook(path) -> (DeIdentifiedMart, PortfolioFindings)``

The workbook is the source of truth — the reviewer keyed assessments into it
and its formulas computed the flags and roll-ups. Re-ingest re-evaluates the
whole workbook with the deterministic ``formulas`` engine (so it works whether
or not Excel cached values on save) and extracts:

- the **de-identified mart**: one record per loan keyed by the engagement-
  scoped mart id — no borrower name, no loan number, no TIN in any form; and
- **portfolio findings**: open-exception counts by class / severity / status
  and by (de-identified) loan, plus the criticized/classified totals.

Cell geometry comes from the hidden ``_map`` sheet the builder stamps (a JSON
structure map — row ids and addresses only, never borrower data), so ingest
needs no access to the original configs.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path

from credit_review.linesheet import OPEN_FLAG

MAP_SHEET = "_map"


@dataclass(frozen=True)
class DeIdentifiedMart:
    engagement_id: str
    lob: str
    rows: tuple[dict, ...]   # one per loan, keyed by mart_id — de-identified


@dataclass(frozen=True)
class PortfolioFindings:
    open_by_class: dict[str, int]
    open_by_severity: dict[str, int]
    fired_by_status: dict[str, int]
    open_by_loan: dict[str, int]          # keyed by de-identified mart_id
    total_open: int
    criticized_dollars: float
    criticized_count: int
    classified_dollars: float
    classified_count: int
    records: tuple[dict, ...] = field(default=())   # de-identified exception rows


class _Solution:
    """Cell lookup over a `formulas` calculation of the whole workbook."""

    def __init__(self, path: str | Path) -> None:
        # Imported lazily so environments that only *build* (e.g. the ASCII
        # bundle's target machine) need just openpyxl + PyYAML.
        import formulas

        model = formulas.ExcelModel().loads(str(path)).finish()
        sol = model.calculate()
        self._cells: dict[str, object] = {}
        for key, ranges in sol.items():
            # keys look like "'[book.xlsx]SHEET'!A1"
            if "]" not in key or "!" not in key:
                continue
            sheet = key.split("]", 1)[1].split("'", 1)[0]
            addr = key.rsplit("!", 1)[1]
            try:
                self._cells[f"{sheet}!{addr}"] = ranges.value[0, 0]
            except Exception:
                continue

    def get(self, sheet: str, addr: str, default=None):
        return self._cells.get(f"{sheet.upper()}!{addr.upper().replace('$', '')}", default)


def _load_map(path: str | Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    try:
        if MAP_SHEET not in wb.sheetnames:
            raise ValueError(f"not a Credit Review OS workbook: no {MAP_SHEET} sheet")
        return json.loads(wb[MAP_SHEET]["A1"].value)
    finally:
        wb.close()


def ingest_workbook(path: str | Path) -> tuple[DeIdentifiedMart, PortfolioFindings]:
    structure = _load_map(path)
    sol = _Solution(path)
    if structure.get("mode") == "product_conformance":
        return _ingest_mode_b(structure, sol)

    mart_rows = []
    open_by_loan: dict[str, int] = {}
    records = []
    open_by_class: dict[str, int] = {}
    open_by_severity: dict[str, int] = {}
    fired_by_status: dict[str, int] = {}
    total_open = 0

    for loan in structure["loans"]:
        mid, sheet, master_row = loan["mart_id"], loan["sheet"], loan["master_row"]
        mart_rows.append({
            "mart_id": mid,
            "lob": structure["lob"],
            "committed": sol.get("Master", f"D{master_row}"),
            "outstanding": sol.get("Master", f"E{master_row}"),
            "orig_grade": sol.get("Master", f"F{master_row}"),
            "rev_grade": sol.get("Master", f"G{master_row}"),
            "bucket": sol.get("Master", f"H{master_row}"),
            "criticized": bool(sol.get("Master", f"I{master_row}")),
            "classified": bool(sol.get("Master", f"J{master_row}")),
            "open_exceptions": int(sol.get("Master", f"K{master_row}") or 0),
        })
        open_by_loan[mid] = 0
        for exc in loan["exceptions"]:
            flag = sol.get(sheet, exc["flag"])
            status = sol.get(sheet, exc["status"]) or ""
            fired = flag == OPEN_FLAG
            is_open = fired and status == "open"
            if fired:
                fired_by_status[status] = fired_by_status.get(status, 0) + 1
            if is_open:
                open_by_class[exc["class"]] = open_by_class.get(exc["class"], 0) + 1
                open_by_severity[exc["severity"]] = \
                    open_by_severity.get(exc["severity"], 0) + 1
                open_by_loan[mid] += 1
                total_open += 1
            records.append({
                "mart_id": mid, "rule": exc["id"], "class": exc["class"],
                "subtype": exc.get("subtype") or "", "severity": exc["severity"],
                "fired": fired, "status": status})

    criticized = [r for r in mart_rows if r["criticized"]]
    classified = [r for r in mart_rows if r["classified"]]
    mart = DeIdentifiedMart(engagement_id=structure["engagement_id"],
                            lob=structure["lob"], rows=tuple(mart_rows))
    findings = PortfolioFindings(
        open_by_class=open_by_class, open_by_severity=open_by_severity,
        fired_by_status=fired_by_status, open_by_loan=open_by_loan,
        total_open=total_open,
        criticized_dollars=float(sum(r["outstanding"] or 0 for r in criticized)),
        criticized_count=len(criticized),
        classified_dollars=float(sum(r["outstanding"] or 0 for r in classified)),
        classified_count=len(classified),
        records=tuple(records))
    return mart, findings


def _ingest_mode_b(structure: dict,
                   sol: "_Solution") -> tuple[DeIdentifiedMart, PortfolioFindings]:
    """Mode B: product-level outputs only — no loan numbers, no person data.

    Retail has no Special Mention bucket, so criticized == classified ==
    Substandard + Loss (URCCP).
    """
    mart_rows = []
    records = []
    open_by_class: dict[str, int] = {}
    open_by_severity: dict[str, int] = {}
    fired_by_status: dict[str, int] = {}
    open_by_loan: dict[str, int] = {}
    total_open = 0
    substandard = loss = 0.0

    for i, product in enumerate(structure["products"], start=1):
        pid, sheet = product["id"], product["sheet"]
        prow = product["products_row"]
        mid = f"{structure['engagement_id']}-P{i:02d}"
        sub_d = float(sol.get(sheet, product["panel"]["substandard_dollars"]) or 0)
        loss_d = float(sol.get(sheet, product["panel"]["loss_dollars"]) or 0)
        substandard += sub_d
        loss += loss_d
        mart_rows.append({
            "mart_id": mid,
            "product": pid,
            "population_count": sol.get("Products", f"B{prow}"),
            "population_dollars": sol.get("Products", f"C{prow}"),
            "sample_n": sol.get("Products", f"D{prow}"),
            "coverage": sol.get("Products", f"E{prow}"),
            "open_findings": int(sol.get("Products", f"F{prow}") or 0),
            "substandard_dollars": sub_d,
            "loss_dollars": loss_d,
            "sample_dollars": sol.get("Products", f"J{prow}"),
            "coverage_dollars": sol.get("Products", f"K{prow}"),
        })
        open_by_loan[mid] = 0
        for test in product["tests"]:
            flag = sol.get(sheet, test["flag"])
            status = sol.get(sheet, test["status"]) or ""
            fired = flag == OPEN_FLAG
            is_open = fired and status == "open"
            if fired:
                fired_by_status[status] = fired_by_status.get(status, 0) + 1
            if is_open:
                open_by_class[test["class"]] = open_by_class.get(test["class"], 0) + 1
                open_by_severity[test["severity"]] = \
                    open_by_severity.get(test["severity"], 0) + 1
                open_by_loan[mid] += 1
                total_open += 1
            records.append({
                "mart_id": mid, "rule": test["id"], "class": test["class"],
                "subtype": "", "severity": test["severity"], "fired": fired,
                "status": status,
                "rate": sol.get(sheet, test["rate"]),
                "fails": int(sol.get(sheet, test["fails"]) or 0)})

    classified_dollars = substandard + loss   # URCCP: no Special Mention bucket
    classified_count = sum(1 for r in mart_rows
                           if (r["substandard_dollars"] or r["loss_dollars"]))
    mart = DeIdentifiedMart(engagement_id=structure["engagement_id"],
                            lob=structure["lob"], rows=tuple(mart_rows))
    findings = PortfolioFindings(
        open_by_class=open_by_class, open_by_severity=open_by_severity,
        fired_by_status=fired_by_status, open_by_loan=open_by_loan,
        total_open=total_open,
        criticized_dollars=classified_dollars,
        criticized_count=classified_count,
        classified_dollars=classified_dollars,
        classified_count=classified_count,
        records=tuple(records))
    return mart, findings
