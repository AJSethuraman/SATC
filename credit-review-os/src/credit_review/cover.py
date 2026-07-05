"""Cover sheet: engagement identity, scope, and reviewer independence."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import Engagement, Loan, Program

_LAST_COL = 4
_LABEL_COL = 2
_VALUE_COL = 3


def _kv(ws: Worksheet, row: int, label: str, value) -> int:
    lab = ws.cell(row, _LABEL_COL, label)
    lab.font = KB.SECONDARY
    val = ws.cell(row, _VALUE_COL, value)
    val.font = KB.DATA_FONT
    val.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _section(ws: Worksheet, row: int, label: str) -> int:
    cell = ws.cell(row, _LABEL_COL, label)
    cell.font = Font(name="Arial", bold=True, size=11, color=KB.INK)
    ws.row_dimensions[row].height = 20
    return row + 1


def build_cover(ws: Worksheet, program: Program, engagement: Engagement,
                loans: list[Loan]) -> None:
    KB.hide_gridlines(ws)
    for letter, width in {"A": 2, "B": 34, "C": 52, "D": 10}.items():
        ws.column_dimensions[letter].width = width

    row = KB.brand_banner(
        ws, 1, _LAST_COL,
        f"Loan Review — {engagement.client_name}",
        f"{program.title}  ·  Engagement {engagement.engagement_id}")
    row += 1

    row = _section(ws, row, "Engagement")
    row = _kv(ws, row, "Client bank", engagement.client_name)
    row = _kv(ws, row, "Engagement ID", engagement.engagement_id)
    row = _kv(ws, row, "Line of business", program.lob)
    row = _kv(ws, row, "Review mode", program.review_mode)
    row += 1

    row = _section(ws, row, "Scope")
    scope = engagement.scope
    row = _kv(ws, row, "Portfolio ($)", scope.get("portfolio_dollars"))
    ws.cell(row - 1, _VALUE_COL).number_format = '$#,##0'
    row = _kv(ws, row, "Portfolio (count)", scope.get("portfolio_count"))
    row = _kv(ws, row, "Sampling basis", scope.get("sample_basis"))
    row = _kv(ws, row, "Loans in this workbook", len(loans))
    row += 1

    row = _section(ws, row, "Reviewer & independence")
    reviewer = engagement.reviewer
    row = _kv(ws, row, "Reviewer", reviewer.get("name"))
    row = _kv(ws, row, "Independent", "Yes" if reviewer.get("independent") else "No")
    if reviewer.get("independence_note"):
        row = _kv(ws, row, "Independence basis", reviewer["independence_note"])
    row += 1

    note = ws.cell(row, _LABEL_COL,
                   "Borrower TINs appear as last-4 only throughout this workpaper. "
                   "Linesheets identify credits by name and loan/obligor number.")
    note.font = KB.NOTE_FONT
    ws.merge_cells(start_row=row, start_column=_LABEL_COL, end_row=row, end_column=_LAST_COL)
    note.alignment = Alignment(wrap_text=True, vertical="top")
