"""``_methodology`` + ``_readme`` — the "prove the method" tabs.

``_methodology`` renders the program's regulatory crosswalk (every program
element -> the interagency requirement it satisfies, with citation) and the
guidance's own engagement-level asks captured from the overlay: coverage/scope
(reviewed $ and count computed **live** from the Master roll-up), reviewer
independence, and the findings-to-board reporting trail — so the deliverable
is self-proving to an examiner or board.

``_readme`` carries the working notes: how to use the workbook, the pin-cite
verification caveat, and the PII rules.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import Engagement, Loan, Program
from credit_review.master import MasterRefs

PIN_CITE_CAVEAT = (
    "Citation caveat: regulator quotes in the crosswalk were transcribed from the "
    "canonical primary documents, but exact page/paragraph pin-cites are "
    "unconfirmed — confirm them against the live PDFs before quoting in a filed "
    "workpaper.")

PII_NOTES = (
    "PII rules (binding): borrower name and loan/obligor number appear only in "
    "this engagement workbook (the bank's own data, returned to the bank). "
    "Taxpayer identification numbers are last-4 only, everywhere. The Data Mart "
    "and any re-ingest export are de-identified (engagement-scoped loan ids; no "
    "name, no loan number, no TIN). Real engagement workbooks are encrypted at "
    "rest; every fixture in the repository is synthetic.")

_WRAP = Alignment(wrap_text=True, vertical="top")


def _band(ws: Worksheet, row: int, label: str, ncols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row, 1, label)
    cell.fill = KB.HDR_FILL
    cell.font = KB.WHITE_BOLD
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def _kv(ws: Worksheet, row: int, label: str, value, fmt: str | None = None) -> int:
    ws.cell(row, 1, label).font = KB.SECONDARY
    cell = ws.cell(row, 2, value)
    cell.font = KB.DATA_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if fmt:
        cell.number_format = fmt
    return row + 1


def build_methodology(ws: Worksheet, program: Program, engagement: Engagement,
                      loans: list[Loan], master: MasterRefs) -> None:
    KB.hide_gridlines(ws)
    for letter, width in {"A": 30, "B": 52, "C": 52, "D": 46}.items():
        ws.column_dimensions[letter].width = width

    row = KB.brand_banner(
        ws, 1, 4, "Methodology — Regulatory Crosswalk",
        f"{program.title}  ·  {engagement.client_name}  ·  {engagement.engagement_id}")
    row += 1

    row = _band(ws, row, "Program element -> interagency requirement (with citation)", 4)
    hdr = ["Element", "How this program implements it",
           "Interagency requirement satisfied", "Citation"]
    row = KB.header_row(ws, row, hdr)
    for entry in program.crosswalk:
        values = [entry["element"], entry.get("implementation", ""),
                  entry["requirement"], entry["source"]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.font = KB.DATA_FONT
            cell.alignment = _WRAP
        row += 1
    row += 1

    row = _band(ws, row, "Coverage / scope (2020 guidance: documented review scope)", 4)
    scope = engagement.scope
    f, l = master.first_row, master.last_row
    out_range = f"Master!${master.outstanding_col}${f}:${master.outstanding_col}${l}"
    row = _kv(ws, row, "Portfolio (dollars)", scope.get("portfolio_dollars"), "$#,##0")
    portfolio_cell = f"$B${row - 1}"
    row = _kv(ws, row, "Reviewed (dollars, live)", f"=SUM({out_range})", "$#,##0")
    reviewed_cell = f"$B${row - 1}"
    row = _kv(ws, row, "Coverage (% of portfolio $)",
              f"=IFERROR({reviewed_cell}/{portfolio_cell},\"\")", "0.0%")
    row = _kv(ws, row, "Portfolio (count)", scope.get("portfolio_count"))
    row = _kv(ws, row, "Reviewed (count)", len(loans))
    row = _kv(ws, row, "Sampling basis", scope.get("sample_basis"))
    row = _kv(ws, row, "Review as-of date", engagement.review_as_of, "mm/dd/yyyy")
    row += 1

    row = _band(ws, row, "Reviewer independence (2020 guidance: independent assessment)", 4)
    reviewer = engagement.reviewer
    row = _kv(ws, row, "Reviewer", reviewer.get("name"))
    row = _kv(ws, row, "Independent", "Yes" if reviewer.get("independent") else "No")
    row = _kv(ws, row, "Basis", reviewer.get("independence_note", ""))
    row += 1

    row = _band(ws, row, "Findings-to-board reporting", 4)
    note = ws.cell(row, 1,
                   "Exceptions, rating findings, and the criticized/classified "
                   "asset-quality summary aggregate on the Findings sheet for "
                   "presentation to the credit committee / board. Severity tiers "
                   "correspond to materiality (OCC Loan Portfolio Management); "
                   "high-severity items warrant senior-management / board attention.")
    note.font = KB.DATA_FONT
    note.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 44
    row += 2

    caveat = ws.cell(row, 1, PIN_CITE_CAVEAT)
    caveat.font = KB.NOTE_FONT
    caveat.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30


PII_NOTES_MODE_B = (
    "PII rules (binding — stricter than the loan-level mode): Mode B artifacts "
    "contain no individual-person names, ever. Sampled files are identified by "
    "loan/account number only, and that number stays inside this "
    "encrypted-at-rest workbook — the Data Mart and every re-ingest export are "
    "product/segment level. No TIN in any form. Every repository fixture is "
    "synthetic.")


def build_methodology_mode_b(ws: Worksheet, program: Program,
                             engagement: Engagement,
                             product_rows: dict[str, int]) -> None:
    """Mode B methodology: cited crosswalk + sampling design + coverage."""
    KB.hide_gridlines(ws)
    for letter, width in {"A": 30, "B": 52, "C": 52, "D": 46}.items():
        ws.column_dimensions[letter].width = width

    row = KB.brand_banner(
        ws, 1, 4, "Methodology — Regulatory Crosswalk",
        f"{program.title}  ·  {engagement.client_name}  ·  {engagement.engagement_id}")
    row += 1

    row = _band(ws, row, "Program element -> requirement satisfied (with citation)", 4)
    row = KB.header_row(ws, row, ["Element", "How this program implements it",
                                  "Requirement satisfied", "Citation"])
    for entry in program.crosswalk:
        values = [entry["element"], entry.get("implementation", ""),
                  entry["requirement"], entry["source"]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.font = KB.DATA_FONT
            cell.alignment = _WRAP
        row += 1
    row += 1

    row = _band(ws, row, "Sampling design — stratified random + judgmental "
                         "segments (documented basis; coverage live)", 4)
    row = KB.header_row(ws, row, ["Product / segment", "Method / stratum",
                                  "Size / basis", "Coverage (live)"])
    for pid, block in engagement.overlay_products.items():
        pop = block["population"]
        prow = product_rows[pid]
        cell = ws.cell(row, 1, pid)
        cell.font = Font(name="Arial", bold=True, size=10, color=KB.INK)
        ws.cell(row, 2, f"population {pop['count']:,} / ${pop['dollars']:,}") \
            .font = KB.DATA_FONT
        cov = ws.cell(row, 4, f"=Products!$E${prow}")
        cov.font = KB.DATA_FONT
        cov.number_format = "0.00%"
        row += 1
        for seg in block["sample_plan"]["segments"]:
            stratum = ", ".join(f"{k}: {v}"
                                for k, v in (seg.get("stratum") or {}).items())
            ws.cell(row, 1, f"    {seg['id']}").font = KB.DATA_FONT
            ws.cell(row, 2, seg["method"] + (f" — {stratum}" if stratum else "")) \
                .font = KB.DATA_FONT
            c = ws.cell(row, 3, f"n={seg['size']} — {seg['basis']}")
            c.font = KB.DATA_FONT
            c.alignment = _WRAP
            row += 1
    row += 1

    row = _band(ws, row, "Reviewer independence", 4)
    reviewer = engagement.reviewer
    row = _kv(ws, row, "Reviewer", reviewer.get("name"))
    row = _kv(ws, row, "Independent", "Yes" if reviewer.get("independent") else "No")
    row = _kv(ws, row, "Basis", reviewer.get("independence_note", ""))
    row += 1

    row = _band(ws, row, "Findings-to-board reporting", 4)
    note = ws.cell(row, 1,
                   "Conformance findings (rate vs tolerance; compliance "
                   "per-occurrence), the buy-box fringe-vs-core comparison, and "
                   "the URCCP asset-quality totals aggregate on the Findings and "
                   "Products sheets for committee/board presentation.")
    note.font = KB.DATA_FONT
    note.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 44
    row += 2

    caveat = ws.cell(row, 1, PIN_CITE_CAVEAT)
    caveat.font = KB.NOTE_FONT
    caveat.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30


def build_readme(ws: Worksheet, program: Program, engagement: Engagement,
                 pii_notes: str = PII_NOTES) -> None:
    KB.hide_gridlines(ws)
    ws.column_dimensions["A"].width = 110

    lines = [
        ("Credit Review OS — engagement workbook", True),
        (f"{program.title} · {engagement.client_name} · {engagement.engagement_id}", False),
        ("", False),
        ("HOW TO USE", True),
        ("Key your assessments into the cream input cells on each LS_ linesheet: "
         "figures, the independent reviewer rating, evidence as-of dates, and "
         "exception status/owner/due/cleared. Everything else computes live — "
         "exception flags, the rating-disagreement finding, the Master roll-up, "
         "the criticized/classified totals, and the Findings aggregates.", False),
        ("Policy thresholds, the rating-scale map, and the review as-of date live "
         "on _config (the knob panel). Edit a knob and the workbook recomputes; "
         "no code change needed.", False),
        ("", False),
        ("METHODOLOGY", True),
        ("The _methodology tab maps every program element to the interagency "
         "requirement it satisfies, with citation, and records the engagement's "
         "coverage, reviewer independence, and findings-to-board reporting trail.", False),
        (PIN_CITE_CAVEAT, False),
        ("", False),
        ("PII", True),
        (pii_notes, False),
    ]
    row = 1
    for text, bold in lines:
        cell = ws.cell(row, 1, text)
        if bold:
            cell.fill = KB.SECT_FILL
            cell.font = KB.SECTION_FONT
        else:
            cell.font = KB.DATA_FONT
            cell.alignment = _WRAP
            if len(text) > 90:
                ws.row_dimensions[row].height = 15 * (len(text) // 90 + 1)
        row += 1
