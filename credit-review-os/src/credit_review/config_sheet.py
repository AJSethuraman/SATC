"""``_config`` — the knob panel: engagement settings, policy thresholds, and
the rating-scale map, written as plain key/value rows with onyx section bands
(the quiet system-tab treatment every template in this repo uses).

The sheet doubles as the formula-resolution target for the linesheet tokens:
``[POL key]`` points at a threshold cell, and ``MAP(...)`` VLOOKUPs against the
rating-scale-map block — edit a knob, the linesheets recompute live.
``write_config_sheet`` returns those addresses as a :class:`ConfigRefs`.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from credit_review import keybank_style as KB
from credit_review.config import Engagement, Program

_INK_BOLD = Font(name="Arial", bold=True, size=11, color=KB.INK)


@dataclass(frozen=True)
class ConfigRefs:
    """Absolute ``_config`` addresses the rest of the workbook resolves against."""

    pol_registry: dict[str, str]   # threshold key -> '_config!$B$n'
    map_range: str                 # rating_scale_map grade->bucket block, 'A:B' absolute
    framework_range: str           # bucket->criticized/classified block, 'A:C' absolute
    asof_cell: str                 # the engagement review_as_of date cell ([ASOF] target)


def write_config_sheet(ws: Worksheet, program: Program,
                       engagement: Engagement) -> ConfigRefs:
    """Populate ``_config`` and return the formula-resolution addresses."""
    KB.hide_gridlines(ws)
    rows: list[list] = [
        ["Credit Review OS -- CONFIG (the knob panel). Thresholds here drive "
         "the linesheet formulas; edit values, no code change needed."],
        [],
        ["[ENGAGEMENT]"],
        ["key", "value", "help"],
        ["client_name", engagement.client_name, "Client bank under review."],
        ["engagement_id", engagement.engagement_id, "Unique engagement identifier."],
        ["review_as_of", engagement.review_as_of,
         "Evidence currency is measured against this date ([ASOF] formula target)."],
        ["lob", program.lob, "Line of business this program covers."],
        ["review_mode", program.review_mode,
         "loan_level (v1) | product_conformance (reserved for consumer/resi)."],
        [],
        ["[THRESHOLDS]"],
        ["key", "value", "help"],
    ]
    threshold_row_of: dict[str, int] = {}
    for key in sorted(engagement.thresholds):
        rows.append([key, engagement.thresholds[key],
                     "Engagement policy threshold (overlay); [POL] formulas point here."])
        threshold_row_of[key] = len(rows)

    rows += [[], ["[RATING_SCALE_MAP]"], ["internal grade", "regulatory bucket"]]
    map_first = len(rows) + 1
    for grade in sorted(engagement.rating_scale_map, key=lambda g: (len(g), g)):
        rows.append([grade, engagement.rating_scale_map[grade]])
    map_last = len(rows)
    rows += [[], ["[RATING_FRAMEWORK]"], ["bucket", "criticized", "classified"]]
    fw_first = len(rows) + 1
    for bucket in program.buckets:
        rows.append([bucket,
                     "TRUE" if bucket in program.criticized else "FALSE",
                     "TRUE" if bucket in program.classified else "FALSE"])
    fw_last = len(rows)

    asof_row = None
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if row and row[0] == "review_as_of":
            asof_row = i
            ws.cell(i, 2).number_format = "mm/dd/yyyy"
        if row and isinstance(row[0], str) and row[0].startswith("["):
            band = ws.cell(i, 1)
            band.font = KB.SECTION_FONT
            band.fill = KB.SECT_FILL
    ws["A1"].font = _INK_BOLD
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72

    return ConfigRefs(
        pol_registry={key: f"_config!$B${r}" for key, r in threshold_row_of.items()},
        map_range=f"_config!$A${map_first}:$B${map_last}",
        framework_range=f"_config!$A${fw_first}:$C${fw_last}",
        asof_cell=f"_config!$B${asof_row}",
    )
